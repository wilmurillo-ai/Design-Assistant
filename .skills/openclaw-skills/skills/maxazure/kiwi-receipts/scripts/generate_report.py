#!/usr/bin/env python3
"""Generate IRD-ready GST report from receipt data."""

import argparse
import csv
import json
import sys
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
except ImportError:
    print("Error: openpyxl is required but not installed.", file=sys.stderr)
    print("Install it with:  pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# NZ Tax constants (Task 10)
# ---------------------------------------------------------------------------
NZ_TAX_BRACKETS = [
    (15600, 0.105),
    (53500, 0.175),
    (78100, 0.30),
    (180000, 0.33),
    (float("inf"), 0.39),
]
ACC_LEVY_RATE = 0.0167
ACC_MAX_EARNINGS = 152790

# ---------------------------------------------------------------------------
# Xero category map (Task 11)
# ---------------------------------------------------------------------------
XERO_CATEGORY_MAP = {
    "materials": "Cost of Goods Sold - Materials",
    "tools": "Tools and Equipment",
    "fuel": "Motor Vehicle Expenses - Fuel",
    "vehicle": "Motor Vehicle Expenses",
    "safety": "Health and Safety",
    "subcontractor": "Subcontractor Expenses",
    "office": "Office Expenses",
    "other": "General Expenses",
}


def get_gst_period(dt: date) -> tuple[int, int, int, int]:
    """Return (start_month, end_month, year, due_month) for 2-monthly GST period."""
    m = dt.month
    start = m if m % 2 == 1 else m - 1
    end = start + 1
    year = dt.year
    due_month = end + 2
    due_year = year
    if due_month > 12:
        due_month -= 12
        due_year += 1
    return start, end, year, due_month


def period_label(start_month: int, end_month: int, year: int) -> str:
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{months[start_month-1]}-{months[end_month-1]} {year}"


def filter_by_period(receipts: list, period: str) -> list:
    """Filter receipts by period. 'current' = current 2-month period, or 'YYYY-MM'."""
    if period == "current":
        today = date.today()
        start_m, end_m, year, _ = get_gst_period(today)
        start_date = date(year, start_m, 1)
        if end_m == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, end_m + 1, 1)
    elif period in ("all", "annual"):
        return receipts
    else:
        # YYYY-MM format — find the 2-month period containing this month
        parts = period.split("-")
        year, month = int(parts[0]), int(parts[1])
        start_m = month if month % 2 == 1 else month - 1
        end_m = start_m + 1
        start_date = date(year, start_m, 1)
        if end_m == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, end_m + 1, 1)

    filtered = []
    for r in receipts:
        try:
            rd = date.fromisoformat(r["date"])
            if start_date <= rd < end_date:
                filtered.append(r)
        except (ValueError, KeyError):
            continue
    return filtered


HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
CURRENCY_FMT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def add_summary_sheet(wb: Workbook, receipts: list, period_str: str, business_name: str = "", gst_number: str = ""):
    ws = wb.active
    ws.title = "GST Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    total_incl = sum(r.get("total", 0) for r in receipts)
    total_gst = sum(r.get("gst", 0) for r in receipts)
    total_excl = total_incl - total_gst

    rows = [
        ("Business Name", business_name or "(not set)"),
        ("IRD / GST Number", gst_number or "(not set)"),
        ("", ""),
        ("GST Period", period_str),
        ("Number of Receipts", len(receipts)),
        ("", ""),
        ("Total Purchases (incl GST)", total_incl),
        ("Total GST on Purchases", total_gst),
        ("Total Purchases (excl GST)", total_excl),
    ]

    ws.cell(row=1, column=1, value="GST Return Summary").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, color="666666")

    for i, (label, value) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=value)
        if isinstance(value, (int, float)) and label:
            cell.number_format = CURRENCY_FMT
        cell.border = THIN_BORDER


def add_receipts_sheet(wb: Workbook, receipts: list):
    ws = wb.create_sheet("All Receipts")
    headers = ["Date", "Merchant", "Category", "Items", "Amount (excl GST)", "GST", "Total"]
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    sorted_receipts = sorted(receipts, key=lambda r: r.get("date", ""))
    for i, r in enumerate(sorted_receipts, start=2):
        total = r.get("total", 0)
        gst = r.get("gst", 0)
        excl = round(total - gst, 2)
        items_desc = ", ".join(
            item.get("description", "") for item in r.get("items", [])
        ) or r.get("merchant", "")

        ws.cell(row=i, column=1, value=r.get("date", ""))
        ws.cell(row=i, column=2, value=r.get("merchant", ""))
        ws.cell(row=i, column=3, value=r.get("category", "other"))
        ws.cell(row=i, column=4, value=items_desc)
        ws.cell(row=i, column=5, value=excl).number_format = CURRENCY_FMT
        ws.cell(row=i, column=6, value=gst).number_format = CURRENCY_FMT
        ws.cell(row=i, column=7, value=total).number_format = CURRENCY_FMT

        for col in range(1, 8):
            ws.cell(row=i, column=col).border = THIN_BORDER


def add_category_sheet(wb: Workbook, receipts: list):
    ws = wb.create_sheet("By Category")
    headers = ["Category", "Count", "Total (excl GST)", "GST", "Total (incl GST)"]
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    categories: dict[str, dict] = {}
    for r in receipts:
        cat = r.get("category", "other")
        if cat not in categories:
            categories[cat] = {"count": 0, "total": 0, "gst": 0}
        categories[cat]["count"] += 1
        categories[cat]["total"] += r.get("total", 0)
        categories[cat]["gst"] += r.get("gst", 0)

    for i, (cat, data) in enumerate(sorted(categories.items()), start=2):
        excl = round(data["total"] - data["gst"], 2)
        ws.cell(row=i, column=1, value=cat.title())
        ws.cell(row=i, column=2, value=data["count"])
        ws.cell(row=i, column=3, value=excl).number_format = CURRENCY_FMT
        ws.cell(row=i, column=4, value=round(data["gst"], 2)).number_format = CURRENCY_FMT
        ws.cell(row=i, column=5, value=round(data["total"], 2)).number_format = CURRENCY_FMT
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = THIN_BORDER


def add_ird_sheet(wb: Workbook, receipts: list, period_str: str, income: list | None = None):
    ws = wb.create_sheet("IRD GST101A")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20

    # Box 11: total purchases incl GST
    box_11 = round(sum(r.get("total", 0) for r in receipts), 2)
    # Box 12: GST on purchases = Box 11 x 3/23 (official IRD formula)
    box_12 = round(box_11 * 3 / 23, 2)
    # Box 14: Total GST credit = Box 12 + Box 13 (Box 13 = 0 by default)
    box_14 = box_12

    # Compute income-side boxes if income data is available
    if income:
        box_5 = round(sum(r.get("amount_incl_gst", 0) for r in income), 2)
        box_6 = 0.00
        box_7 = round(box_5 - box_6, 2)
        box_8 = round(box_7 * 3 / 23, 2)
        box_9 = 0.00
        box_10 = round(box_8 + box_9, 2)
        box_15 = round(box_10 - box_14, 2)
    else:
        box_5 = "— enter from accounts —"
        box_6 = 0.00
        box_7 = "— calculated —"
        box_8 = "— calculated —"
        box_9 = 0.00
        box_10 = "— calculated —"
        box_15 = "— calculated —"

    ws.cell(row=1, column=1, value="IRD GST Return (GST101A) Reference").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Period: {period_str}").font = Font(italic=True)
    if not income:
        ws.cell(row=3, column=1, value="Note: Box 5 (sales) must be entered from your accounting records").font = Font(italic=True, color="CC0000")
    else:
        ws.cell(row=3, column=1, value="Income data loaded — sales boxes calculated automatically").font = Font(italic=True, color="007A33")

    headers = ["Box", "Description", "Amount"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    style_header_row(ws, 5, 3)

    ird_rows = [
        ("5", "Total sales and income for the period (incl GST and zero-rated supplies)", box_5),
        ("6", "Zero-rated supplies included in Box 5", box_6),
        ("7", "Box 5 minus Box 6", box_7),
        ("8", "Multiply Box 7 by three and divide by twenty-three (x 3/23)", box_8),
        ("9", "Adjustments from your calculation sheet", box_9),
        ("10", "Total GST collected on sales and income (Box 8 + Box 9)", box_10),
        ("", "", ""),
        ("11", "Total purchases and expenses (incl GST), excl imported goods", box_11),
        ("12", "Multiply Box 11 by three and divide by twenty-three (x 3/23)", box_12),
        ("13", "Credit adjustments from your calculation sheet", 0.00),
        ("14", "Total GST credit for purchases and expenses (Box 12 + Box 13)", box_14),
        ("", "", ""),
        ("15", "Difference (Box 10 minus Box 14). Positive = pay, Negative = refund", box_15),
    ]

    for i, (box, desc, amt) in enumerate(ird_rows, start=6):
        ws.cell(row=i, column=1, value=box).font = Font(bold=True)
        ws.cell(row=i, column=2, value=desc)
        cell = ws.cell(row=i, column=3, value=amt)
        if isinstance(amt, (int, float)):
            cell.number_format = CURRENCY_FMT
            cell.font = Font(bold=True)
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = THIN_BORDER


# ---------------------------------------------------------------------------
# Task 8: Income sheet
# ---------------------------------------------------------------------------

def add_income_sheet(wb: Workbook, income: list):
    ws = wb.create_sheet("Income")
    headers = ["Date", "Client", "Description", "Invoice #", "Amount (excl GST)", "GST", "Total (incl GST)"]
    widths = [12, 25, 35, 14, 20, 14, 20]
    col_letters = ["A", "B", "C", "D", "E", "F", "G"]
    for letter, w in zip(col_letters, widths):
        ws.column_dimensions[letter].width = w

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    sorted_income = sorted(income, key=lambda r: r.get("date", ""))
    for i, r in enumerate(sorted_income, start=2):
        ws.cell(row=i, column=1, value=r.get("date", ""))
        ws.cell(row=i, column=2, value=r.get("client", ""))
        ws.cell(row=i, column=3, value=r.get("description", ""))
        ws.cell(row=i, column=4, value=r.get("invoice_number", ""))
        ws.cell(row=i, column=5, value=r.get("amount_excl_gst", 0)).number_format = CURRENCY_FMT
        ws.cell(row=i, column=6, value=r.get("gst", 0)).number_format = CURRENCY_FMT
        ws.cell(row=i, column=7, value=r.get("amount_incl_gst", 0)).number_format = CURRENCY_FMT
        for col in range(1, 8):
            ws.cell(row=i, column=col).border = THIN_BORDER


# ---------------------------------------------------------------------------
# Task 9: Depreciation
# ---------------------------------------------------------------------------

def calculate_depreciation(asset: dict, tax_year_end: date) -> dict:
    """Calculate depreciation for one asset for the tax year ending on tax_year_end."""
    purchase_date = date.fromisoformat(asset["purchase_date"])
    cost = asset["cost"]
    method = asset.get("method", "DV")
    dv_rate = asset.get("dv_rate", 0)
    sl_rate = asset.get("sl_rate", 0)
    business_percent = asset.get("business_percent", 100) / 100.0

    # NZ tax year: 1 April to 31 March (only March balance date supported)
    tax_year_start = date(tax_year_end.year - 1, 4, 1)

    # For DV method, accumulate year by year from purchase
    if method == "DV":
        book_value = cost
        # Walk through complete prior years
        # First partial year
        if purchase_date < tax_year_start:
            # Calculate depreciation for the purchase year
            first_year_end = _next_tax_year_end(purchase_date, tax_year_end.month, tax_year_end.day)
            months_first = _months_owned(purchase_date, first_year_end)
            dep_first = book_value * dv_rate * (months_first / 12)
            book_value -= dep_first

            # Full years between first year end and current tax year start
            year_cursor = first_year_end
            while True:
                next_ye = date(year_cursor.year + 1, year_cursor.month, year_cursor.day)
                if next_ye > tax_year_start:
                    break
                dep_full = book_value * dv_rate
                book_value -= dep_full
                year_cursor = next_ye

        book_value_start = round(book_value, 2)

        # Current year depreciation
        if purchase_date >= tax_year_start:
            # Purchased during current tax year — pro-rate
            months = _months_owned(purchase_date, tax_year_end)
            depreciation_full = cost * dv_rate * (months / 12)
            book_value_start = cost
        else:
            depreciation_full = book_value * dv_rate

        depreciation_business = depreciation_full * business_percent
        book_value_end = book_value_start - depreciation_full

    else:  # SL
        book_value = cost
        if purchase_date < tax_year_start:
            first_year_end = _next_tax_year_end(purchase_date, tax_year_end.month, tax_year_end.day)
            months_first = _months_owned(purchase_date, first_year_end)
            dep_first = cost * sl_rate * (months_first / 12)
            book_value -= dep_first

            year_cursor = first_year_end
            while True:
                next_ye = date(year_cursor.year + 1, year_cursor.month, year_cursor.day)
                if next_ye > tax_year_start:
                    break
                dep_full = cost * sl_rate
                book_value -= dep_full
                year_cursor = next_ye

        book_value_start = round(max(book_value, 0), 2)

        if purchase_date >= tax_year_start:
            months = _months_owned(purchase_date, tax_year_end)
            depreciation_full = cost * sl_rate * (months / 12)
            book_value_start = cost
        else:
            depreciation_full = cost * sl_rate

        # Ensure we don't depreciate below zero
        depreciation_full = min(depreciation_full, book_value_start)
        depreciation_business = depreciation_full * business_percent
        book_value_end = book_value_start - depreciation_full

    return {
        "depreciation_full": round(depreciation_full, 2),
        "depreciation_business": round(depreciation_business, 2),
        "book_value_start": round(book_value_start, 2),
        "book_value_end": round(max(book_value_end, 0), 2),
    }


def _next_tax_year_end(from_date: date, end_month: int, end_day: int) -> date:
    """Return the next tax year end date after from_date."""
    candidate = date(from_date.year, end_month, end_day)
    if candidate <= from_date:
        candidate = date(from_date.year + 1, end_month, end_day)
    return candidate


def _months_owned(start: date, end: date) -> int:
    """Count the number of months (rounded up) between start and end."""
    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day >= start.day:
        months += 1
    return max(months, 1)


def add_depreciation_sheet(wb: Workbook, assets: list, tax_year_end: date) -> float:
    """Add depreciation schedule sheet. Returns total business depreciation."""
    ws = wb.create_sheet("Depreciation")
    headers = ["Asset", "Category", "Purchase Date", "Cost", "Method", "Rate",
               "Business %", "Opening Value", "Depreciation", "Closing Value"]
    widths = [30, 22, 14, 14, 10, 10, 12, 16, 16, 16]
    col_letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    for letter, w in zip(col_letters, widths):
        ws.column_dimensions[letter].width = w

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    total_depreciation = 0.0
    row = 2
    for asset in sorted(assets, key=lambda a: a.get("name", "")):
        if asset.get("disposed", False):
            continue
        dep = calculate_depreciation(asset, tax_year_end)
        method = asset.get("method", "DV")
        rate = asset.get("dv_rate", 0) if method == "DV" else asset.get("sl_rate", 0)

        ws.cell(row=row, column=1, value=asset.get("name", ""))
        ws.cell(row=row, column=2, value=asset.get("category", ""))
        ws.cell(row=row, column=3, value=asset.get("purchase_date", ""))
        ws.cell(row=row, column=4, value=asset.get("cost", 0)).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5, value=method)
        ws.cell(row=row, column=6, value=f"{rate:.0%}")
        ws.cell(row=row, column=7, value=f"{asset.get('business_percent', 100)}%")
        ws.cell(row=row, column=8, value=dep["book_value_start"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=9, value=dep["depreciation_business"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=10, value=dep["book_value_end"]).number_format = CURRENCY_FMT

        for col in range(1, 11):
            ws.cell(row=row, column=col).border = THIN_BORDER

        total_depreciation += dep["depreciation_business"]
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=9, value=round(total_depreciation, 2)).number_format = CURRENCY_FMT
    ws.cell(row=row, column=9).font = Font(bold=True)
    for col in range(1, 11):
        ws.cell(row=row, column=col).border = THIN_BORDER

    return round(total_depreciation, 2)


# ---------------------------------------------------------------------------
# Task 10: IR3 Annual Tax
# ---------------------------------------------------------------------------

def calculate_income_tax(taxable_income: float) -> tuple[float, list]:
    """Progressive NZ income tax calculation. Returns (total_tax, bracket_breakdown)."""
    if taxable_income <= 0:
        return 0.0, []
    remaining = taxable_income
    total_tax = 0.0
    breakdown = []
    prev_threshold = 0
    for threshold, rate in NZ_TAX_BRACKETS:
        band = min(remaining, threshold - prev_threshold)
        if band <= 0:
            break
        tax = band * rate
        total_tax += tax
        breakdown.append({
            "from": prev_threshold,
            "to": prev_threshold + band,
            "rate": rate,
            "amount": round(band, 2),
            "tax": round(tax, 2),
        })
        remaining -= band
        prev_threshold = threshold
    return round(total_tax, 2), breakdown


def calculate_acc_levy(gross_income: float) -> float:
    """Calculate ACC earners' levy."""
    liable = min(gross_income, ACC_MAX_EARNINGS)
    return round(max(liable, 0) * ACC_LEVY_RATE, 2)


def add_ir3_sheet(wb: Workbook, receipts: list, income: list | None,
                  total_depreciation: float, tax_year: str,
                  business_name: str, gst_number: str,
                  tax_history: dict | None):
    ws = wb.create_sheet("IR3 Annual Tax")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    row = 1
    ws.cell(row=row, column=1, value=f"IR3 Individual Tax Return — {tax_year}").font = TITLE_FONT
    row += 1
    if business_name:
        ws.cell(row=row, column=1, value=f"Business: {business_name}").font = Font(italic=True)
        row += 1
    if gst_number:
        ws.cell(row=row, column=1, value=f"IRD Number: {gst_number}").font = Font(italic=True)
        row += 1
    row += 1

    # --- INCOME section ---
    ws.cell(row=row, column=1, value="INCOME").font = Font(bold=True, size=12)
    row += 1

    gross_incl = round(sum(r.get("amount_incl_gst", 0) for r in (income or [])), 2)
    gross_excl = round(sum(r.get("amount_excl_gst", 0) for r in (income or [])), 2)

    ws.cell(row=row, column=1, value="Gross Income (incl GST)")
    ws.cell(row=row, column=2, value=gross_incl).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1
    ws.cell(row=row, column=1, value="Gross Income (excl GST)")
    ws.cell(row=row, column=2, value=gross_excl).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 2

    # --- EXPENSES section ---
    ws.cell(row=row, column=1, value="EXPENSES").font = Font(bold=True, size=12)
    row += 1

    categories: dict[str, float] = {}
    for r in receipts:
        cat = r.get("category", "other")
        total = r.get("total", 0)
        gst = r.get("gst", 0)
        excl = total - gst
        categories[cat] = categories.get(cat, 0) + excl

    total_expenses = 0.0
    for cat in sorted(categories.keys()):
        amt = round(categories[cat], 2)
        ws.cell(row=row, column=1, value=f"  {cat.title()}")
        ws.cell(row=row, column=2, value=amt).number_format = CURRENCY_FMT
        ws.cell(row=row, column=2).border = THIN_BORDER
        total_expenses += amt
        row += 1

    total_expenses = round(total_expenses, 2)
    ws.cell(row=row, column=1, value="Total Expenses (excl GST)").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_expenses).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 2

    # --- DEPRECIATION section ---
    ws.cell(row=row, column=1, value="DEPRECIATION").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Total Depreciation (business portion)")
    ws.cell(row=row, column=2, value=total_depreciation).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 2

    # --- TAX CALCULATION section ---
    ws.cell(row=row, column=1, value="TAX CALCULATION").font = Font(bold=True, size=12)
    row += 1

    taxable_income = round(gross_excl - total_expenses - total_depreciation, 2)

    ws.cell(row=row, column=1, value="Gross Income (excl GST)")
    ws.cell(row=row, column=2, value=gross_excl).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1
    ws.cell(row=row, column=1, value="Less: Total Expenses")
    ws.cell(row=row, column=2, value=-total_expenses).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1
    ws.cell(row=row, column=1, value="Less: Depreciation")
    ws.cell(row=row, column=2, value=-total_depreciation).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1
    ws.cell(row=row, column=1, value="Taxable Income").font = Font(bold=True)
    ws.cell(row=row, column=2, value=taxable_income).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 2

    # Bracket breakdown
    income_tax, breakdown = calculate_income_tax(taxable_income)
    ws.cell(row=row, column=1, value="Tax Bracket Breakdown").font = Font(bold=True, size=11)
    row += 1
    for b in breakdown:
        label = f"  ${b['from']:,.0f} - ${b['to']:,.0f} @ {b['rate']:.1%}"
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=b["tax"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    ws.cell(row=row, column=1, value="Income Tax").font = Font(bold=True)
    ws.cell(row=row, column=2, value=income_tax).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1

    acc_levy = calculate_acc_levy(gross_excl)
    ws.cell(row=row, column=1, value=f"ACC Earners' Levy ({ACC_LEVY_RATE:.2%})")
    ws.cell(row=row, column=2, value=acc_levy).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1

    total_tax = round(income_tax + acc_levy, 2)
    ws.cell(row=row, column=1, value="Total Tax + ACC").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_tax).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1

    # Provisional tax from history
    provisional_paid = 0.0
    if tax_history and "years" in tax_history:
        # Use previous year's RIT as provisional tax basis
        years_data = tax_history["years"]
        # tax_year is "YYYY-YYYY" format; first year is the start year
        start_year = tax_year.split("-")[0] if "-" in tax_year else tax_year
        prev_year = str(int(start_year))
        if prev_year in years_data:
            provisional_paid = years_data[prev_year].get("provisional_tax_paid", 0)

    ws.cell(row=row, column=1, value="Less: Provisional Tax Paid")
    ws.cell(row=row, column=2, value=-provisional_paid).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).border = THIN_BORDER
    row += 1

    terminal_tax = round(total_tax - provisional_paid, 2)
    ws.cell(row=row, column=1, value="Terminal Tax (to pay / refund)").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=terminal_tax).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).font = Font(bold=True, size=12)
    ws.cell(row=row, column=2).border = THIN_BORDER


# ---------------------------------------------------------------------------
# Task 11: Xero CSV Export
# ---------------------------------------------------------------------------

def generate_xero_csv(receipts: list, income: list | None, output_path: Path):
    """Generate Xero-compatible CSV import file."""
    rows = []

    for r in receipts:
        dt = r.get("date", "")
        try:
            d = date.fromisoformat(dt)
            formatted_date = d.strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            formatted_date = dt

        total = r.get("total", 0)
        category = r.get("category", "other")
        xero_cat = XERO_CATEGORY_MAP.get(category, "General Expenses")
        items_desc = ", ".join(
            item.get("description", "") for item in r.get("items", [])
        ) or ""

        rows.append({
            "Date": formatted_date,
            "Amount": round(-total, 2),  # Negative for expenses
            "Payee": r.get("merchant", ""),
            "Description": items_desc,
            "Reference": r.get("id", ""),
            "Category": xero_cat,
            "_sort_date": dt,
        })

    if income:
        for r in income:
            dt = r.get("date", "")
            try:
                d = date.fromisoformat(dt)
                formatted_date = d.strftime("%d/%m/%Y")
            except (ValueError, TypeError):
                formatted_date = dt

            rows.append({
                "Date": formatted_date,
                "Amount": round(r.get("amount_incl_gst", 0), 2),  # Positive for income
                "Payee": r.get("client", ""),
                "Description": r.get("description", ""),
                "Reference": r.get("invoice_number", ""),
                "Category": "Sales",
                "_sort_date": dt,
            })

    # Sort by date
    rows.sort(key=lambda x: x.get("_sort_date", ""))

    # Write CSV
    fieldnames = ["Date", "Amount", "Payee", "Description", "Reference", "Category"]
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    # Print JSON summary
    total_expenses = sum(r["Amount"] for r in rows if r["Amount"] < 0)
    total_income = sum(r["Amount"] for r in rows if r["Amount"] > 0)
    summary = {
        "format": "xero-csv",
        "output": str(output_path),
        "total_rows": len(rows),
        "expense_rows": sum(1 for r in rows if r["Amount"] < 0),
        "income_rows": sum(1 for r in rows if r["Amount"] > 0),
        "total_expenses": round(total_expenses, 2),
        "total_income": round(total_income, 2),
    }
    print(json.dumps(summary))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate IRD GST report")
    parser.add_argument("--data", required=True, help="Path to receipts.json")
    parser.add_argument("--output", required=True, help="Output XLSX path")
    parser.add_argument("--period", default="current", help="Period: 'current', 'all', or 'YYYY-MM'")
    parser.add_argument("--business-name", default="", help="Business name")
    parser.add_argument("--gst-number", default="", help="GST/IRD number")
    parser.add_argument("--income", default=None, help="Path to income.json")
    parser.add_argument("--assets", default=None, help="Path to assets.json")
    parser.add_argument("--tax-history", default=None, help="Path to tax-history.json")
    parser.add_argument("--format", default="xlsx", choices=["xlsx", "xero-csv"], help="Output format")
    args = parser.parse_args()

    data_path = Path(args.data)
    if not data_path.exists():
        print(f"No receipt data found at {data_path}", file=sys.stderr)
        sys.exit(1)

    def load_json(path: Path, label: str):
        try:
            with open(path) as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            print(f"Error: {label} is not valid JSON: {e}", file=sys.stderr)
            sys.exit(1)

    all_receipts = load_json(data_path, "receipts.json")
    receipts = filter_by_period(all_receipts, args.period)

    # Load income records
    income_records = None
    if args.income:
        income_path = Path(args.income)
        if income_path.exists():
            all_income = load_json(income_path, "income.json")
            income_records = filter_by_period(all_income, args.period)

    # Load assets (no period filter -- all active assets)
    assets = None
    if args.assets:
        assets_path = Path(args.assets)
        if assets_path.exists():
            assets = load_json(assets_path, "assets.json")

    # Load tax history
    tax_history = None
    if args.tax_history:
        th_path = Path(args.tax_history)
        if th_path.exists():
            tax_history = load_json(th_path, "tax-history.json")

    if not receipts and not income_records and not assets:
        print(f"No data found for period: {args.period}", file=sys.stderr)
        sys.exit(1)
    if not receipts:
        receipts = []

    # Determine period label
    if args.period == "current":
        today = date.today()
        s, e, y, _ = get_gst_period(today)
        p_label = period_label(s, e, y)
    elif args.period in ("all", "annual"):
        p_label = "All Periods"
    else:
        parts = args.period.split("-")
        year, month = int(parts[0]), int(parts[1])
        s = month if month % 2 == 1 else month - 1
        p_label = period_label(s, s + 1, year)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # --- Xero CSV branch ---
    if args.format == "xero-csv":
        generate_xero_csv(receipts, income_records, output_path)
        return

    # --- XLSX branch ---
    wb = Workbook()
    add_summary_sheet(wb, receipts, p_label, args.business_name, args.gst_number)
    add_receipts_sheet(wb, receipts)
    add_category_sheet(wb, receipts)
    add_ird_sheet(wb, receipts, p_label, income=income_records)

    # Income sheet (if income data available)
    if income_records:
        add_income_sheet(wb, income_records)

    # Depreciation and IR3 (for annual / "all" period)
    total_depreciation = 0.0
    if assets and args.period in ("all", "annual"):
        # Determine tax year end (default: March 31 of latest receipt year)
        years_in_data = set()
        for r in receipts:
            try:
                years_in_data.add(date.fromisoformat(r["date"]).year)
            except (ValueError, KeyError):
                pass
        if income_records:
            for r in income_records:
                try:
                    years_in_data.add(date.fromisoformat(r["date"]).year)
                except (ValueError, KeyError):
                    pass
        max_year = max(years_in_data) if years_in_data else date.today().year
        tax_year_end = date(max_year, 3, 31)
        total_depreciation = add_depreciation_sheet(wb, assets, tax_year_end)

    if args.period in ("all", "annual"):
        # Tax year ends on March 31 of the latest year in data
        # Reuse max_year from depreciation block, or compute it
        if not assets or args.period != "all":
            years_in_data = set()
            for r in receipts + (income_records or []):
                try:
                    years_in_data.add(date.fromisoformat(r["date"]).year)
                except (ValueError, KeyError):
                    pass
            max_year = max(years_in_data) if years_in_data else date.today().year
        tax_year = f"{max_year - 1}-{max_year}"

        add_ir3_sheet(wb, receipts, income_records, total_depreciation,
                      tax_year, args.business_name, args.gst_number, tax_history)

    wb.save(str(output_path))

    # JSON output
    total_incl = sum(r.get("total", 0) for r in receipts)
    total_gst = sum(r.get("gst", 0) for r in receipts)
    output_data = {
        "period": p_label,
        "receipt_count": len(receipts),
        "total_incl_gst": round(total_incl, 2),
        "total_gst": round(total_gst, 2),
        "total_excl_gst": round(total_incl - total_gst, 2),
        "output": str(output_path),
    }

    if income_records:
        income_total_incl = round(sum(r.get("amount_incl_gst", 0) for r in income_records), 2)
        income_total_excl = round(sum(r.get("amount_excl_gst", 0) for r in income_records), 2)
        income_total_gst = round(sum(r.get("gst", 0) for r in income_records), 2)
        output_data["income_count"] = len(income_records)
        output_data["income_total_incl_gst"] = income_total_incl
        output_data["income_total_excl_gst"] = income_total_excl
        output_data["income_total_gst"] = income_total_gst

    print(json.dumps(output_data))


if __name__ == "__main__":
    main()
