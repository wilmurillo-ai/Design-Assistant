#!/usr/bin/env python3
"""
PansExcel Pro Max - 全功能智能 Excel 处理专家
功能：数据分析 | 图表可视化 | 数据处理 | 报表模板 | AI智能 | 数据验证 | PDF导出 | 模板库
"""

import argparse, json, sys, os, re, math, csv, io
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle, Protection
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule, CellIsRule, FormulaRule
    from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, RadarChart, ScatterChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ========== 专业配色 ==========
C = {
    "primary":     "2E5BFF", "primary_l":  "5B8DEF", "primary_d":  "1A3DB8",
    "success":     "27AE60", "warning":    "F39C12", "danger":     "E74C3C",
    "purple":      "9B59B6", "teal":       "16A085",
    "white":       "FFFFFF", "gray_l":     "F8FAFC", "gray":       "F1F5F9",
    "gray_d":      "E2E8F0", "text_d":     "1E293B", "text_m":     "475569",
    "text_l":      "94A3B8", "gold":       "F1C40F",
}
CHART = ["2E5BFF","5B8DEF","27AE60","F39C12","E74C3C","9B59B6","16A085","FF6B81","1ABC9C","E67E22"]

def S(color=C["gray_d"]):
    return Side(style='thin', color=color)

def B(color=C["gray_d"]):
    s=S(color); return Border(left=s,right=s,top=s,bottom=s)

def BH():
    s=Side(style='medium',color=C["white"]); return Border(left=s,right=s,top=Side(style='thin',color=C["primary_d"]),bottom=Side(style='thin',color=C["primary_d"]))

def fill(hex_c): return PatternFill(start_color=hex_c,end_color=hex_c,fill_type="solid")

def fnt(bold=False,size=11,color=C["text_d"],italic=False):
    return Font(bold=bold,size=size,color=color,italic=italic,name="微软雅黑")

def aln(h="center",v="center",wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

def style(cell,bg=None,ft=True,sz=11,fc=C["text_d"],bld=False,ha="center",va="center",brd=True,num_fmt=None):
    if bg: cell.fill=fill(bg)
    cell.font=fnt(bold=bld,size=sz,color=fc)
    cell.alignment=aln(ha,va)
    if brd: cell.border=B()
    if num_fmt: cell.number_format=num_fmt

# ========== 样式预设 ==========
ST = {
    "h":    {"bg":C["primary"],    "fc":C["white"],   "bld":True,  "sz":11},
    "h2":   {"bg":C["primary_l"],  "fc":C["white"],   "bld":True,  "sz":11},
    "title":{"sz":20,"fc":C["primary"],"bld":True},
    "sub":  {"sz":14,"fc":C["text_d"],"bld":True,"ha":"left"},
    "kpi":  {"sz":28,"fc":C["primary"],"bld":True},
    "num":  {"ha":"right","num_fmt":"#,##0"},
    "money":{"ha":"right","num_fmt":"¥#,##0.00"},
    "pct":  {"ha":"right","num_fmt":"0.00%"},
    "date": {"ha":"center","num_fmt":"YYYY-MM-DD"},
}

def ap(cell,s):
    bg=s.get("bg"); ft=s.get("ft",True); sz=s.get("sz",11); fc=s.get("fc",C["text_d"])
    bld=s.get("bld",False); ha=s.get("ha","center"); va=s.get("va","center")
    brd=s.get("brd",True); num=s.get("num_fmt")
    style(cell,bg=bg,ft=ft,sz=sz,fc=fc,bld=bld,ha=ha,va=va,brd=brd,num_fmt=num)

# ========== 工具类 ==========
class X:
    """Excel 核心工具"""

    def __init__(self):
        self.wb=None; self.ws=None; self.path=None

    # --- 文件操作 ---
    def open(self,path):
        self.wb=load_workbook(path); self.ws=self.wb.active; self.path=path; return self

    def save(self,path=None):
        p=path or self.path; self.wb.save(p); return p

    def new(self,name="Sheet1"):
        self.wb=Workbook(); self.ws=self.wb.active; self.ws.title=name; return self

    def from_dict(self,data:Dict,sheet="数据")->'X':
        self.new(sheet)
        headers=list(data.keys())
        nr=max((len(v) for v in data.values()),default=0)
        # 标题
        self.ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        c=self.ws["A1"]; c.value=f"📊 {sheet}"; ap(c,ST["title"]); c.fill=fill(C["gray_l"])
        # 表头
        for col,h in enumerate(headers,1):
            cell=self.ws.cell(row=3,column=col,value=h); ap(cell,ST["h"])
        # 数据
        for r in range(nr):
            bg=C["gray_l"] if r%2==0 else C["white"]
            for col,h in enumerate(headers,1):
                val=data[h][r] if r<len(data[h]) else ""
                cell=self.ws.cell(row=4+r,column=col,value=val)
                cell.fill=fill(bg); ap(cell,{"sz":10,"ha":"center"})
                # 金额列
                if any(kw in h for kw in ["金额","销售额","利润","成本","收入","预算"]):
                    cell.number_format="¥#,##0.00"; cell.alignment=aln("right")
                # 百分比列
                elif any(kw in h for kw in ["率","占比","比例"]):
                    cell.number_format="0.00%"
        self._w()
        return self

    def _w(self,min_w=10,max_w=40):
        for col in self.ws.columns:
            mx=max((len(str(c.value or "")) for c in col),default=0)
            self.ws.column_dimensions[get_column_letter(col[0].column)].width=min(max_w,max(min_w,mx+3))

    # --- 样式 ---
    def beautify(self):
        """一键美化"""
        ws=self.ws
        # 表头
        for cell in ws[ws.min_row]:
            if cell.value is not None: ap(cell,ST["h"])
        # 交替行
        for row in ws.iter_rows(min_row=ws.min_row+1):
            bg=C["gray_l"] if (row[0].row-1)%2==0 else C["white"]
            for cell in row:
                if not cell.fill or cell.fill.fill_type=="none":
                    cell.fill=fill(bg)
                if not cell.font: cell.font=fnt(sz=10)
                if not cell.border: cell.border=B()
        self._w()
        return self

    def set_header(self,row_num=None,bg=C["primary"]):
        row_num=row_num or self.ws.min_row
        for cell in self.ws[row_num]:
            cell.fill=fill(bg); ap(cell,ST["h"])
        return self

    def add_title(self,text,row=None,cols=None):
        row=row or self.ws.min_row
        cols=cols or self.ws.max_column
        self.ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=cols)
        c=self.ws.cell(row=row,column=1,value=text)
        ap(c,ST["title"]); c.fill=fill(C["gray_l"])
        self.ws.row_dimensions[row].height=40
        return self

    def autofilter(self):
        self.ws.auto_filter.ref=f"A{self.ws.min_row}:{get_column_letter(self.ws.max_column)}{self.ws.max_row}"
        return self

    def freeze(self,cell="A2"):
        self.ws.freeze_panes=cell; return self

    # --- 图表 ---
    def add_bar(self,pos="F2",title="",w=18,h=12,combo_col=None):
        c=BarChart(type="col",style=10); c.title=title; c.width=w; c.height=h
        c.title_font=Font(bold=True,size=13,color=C["text_d"],name="微软雅黑")
        nr=self.ws.max_row; nc=self.ws.max_column
        if nr<2 or nc<2: return c
        dr=Reference(self.ws,min_col=2,min_row=1,max_col=nc,max_row=nr)
        cr=Reference(self.ws,min_col=1,min_row=2,max_row=nr)
        c.add_data(dr,titles_from_data=True); c.set_categories(cr)
        for i,s in enumerate(c.series):
            s.graphicalProperties.solidFill=CHART[i%len(CHART)]
        self.ws.add_chart(c,pos); return c

    def add_line(self,pos="F2",title="",w=18,h=12):
        c=LineChart(style=12); c.title=title; c.width=w; c.height=h
        c.title_font=Font(bold=True,size=13,color=C["text_d"],name="微软雅黑")
        nr=self.ws.max_row; nc=self.ws.max_column
        if nr<2 or nc<2: return c
        dr=Reference(self.ws,min_col=2,min_row=1,max_col=nc,max_row=nr)
        cr=Reference(self.ws,min_col=1,min_row=2,max_row=nr)
        c.add_data(dr,titles_from_data=True); c.set_categories(cr)
        for i,s in enumerate(c.series):
            s.graphicalProperties.line.solidFill=CHART[i%len(CHART)]; s.smooth=True
            s.line.width=20000
        self.ws.add_chart(c,pos); return c

    def add_pie(self,pos="F2",title="",limit=8):
        c=PieChart(); c.title=title; c.width=14; c.height=10
        c.title_font=Font(bold=True,size=13,color=C["text_d"],name="微软雅黑")
        nr=self.ws.max_row-1
        if nr<1: return c
        dr=Reference(self.ws,min_col=2,min_row=2,max_row=min(nr+1,limit+1))
        cr=Reference(self.ws,min_col=1,min_row=2,max_row=min(nr+1,limit+1))
        c.add_data(dr); c.set_categories(cr)
        for i,pt in enumerate(c.series[0].data_points):
            pt.graphicalProperties.solidFill=CHART[i%len(CHART)]
        self.ws.add_chart(c,pos); return c

    def add_area(self,pos="F2",title="",w=18,h=12):
        c=AreaChart(type="standard",style=10); c.title=title; c.width=w; c.height=h
        nr=self.ws.max_row; nc=self.ws.max_column
        if nr<2 or nc<2: return c
        dr=Reference(self.ws,min_col=2,min_row=1,max_col=nc,max_row=nr)
        cr=Reference(self.ws,min_col=1,min_row=2,max_row=nr)
        c.add_data(dr,titles_from_data=True); c.set_categories(cr)
        self.ws.add_chart(c,pos); return c

    def add_radar(self,pos="F2",title=""):
        c=RadarChart(style=10); c.title=title
        nr=self.ws.max_row
        if nr<2: return c
        dr=Reference(self.ws,min_col=2,min_row=1,max_row=nr)
        c.add_data(dr,titles_from_data=True)
        self.ws.add_chart(c,pos); return c

    def add_combo(self,bar_cols=None,line_cols=None,pos="F2",title="",w=18,h=12):
        c=BarChart(type="col",style=10); c.title=title; c.width=w; c.height=h
        c.grouping="clustered"
        nr=self.ws.max_row; nc=self.ws.max_column
        for col in (bar_cols or []):
            dr=Reference(self.ws,min_col=col,min_row=1,max_row=nr)
            c.add_data(dr,titles_from_data=True)
            s=c.series[-1]; s.graphicalProperties.solidFill=C["primary"]; s.type="col"
        for col in (line_cols or []):
            dr=Reference(self.ws,min_col=col,min_row=1,max_row=nr)
            c.add_data(dr,titles_from_data=True)
            s=c.series[-1]; s.graphicalProperties.line.solidFill=C["warning"]; s.type="line"; s.smooth=True
        cr=Reference(self.ws,min_col=1,min_row=2,max_row=nr); c.set_categories(cr)
        self.ws.add_chart(c,pos); return c

    # --- 条件格式 ---
    def cf_scale(self,rng,mid_pct=50):
        self.ws.conditional_formatting.add(rng,ColorScaleRule(
            start_type='min',start_color=C["success"],
            mid_type='percentile',mid_value=mid_pct,mid_color=C["warning"],
            end_type='max',end_color=C["danger"]))

    def cf_bar(self,rng,color=C["primary"]):
        self.ws.conditional_formatting.add(rng,DataBarRule(start_type='min',end_type='max',color=color,showValue=True))

    def cf_icon(self,rng,icon="3TrafficLights1"):
        self.ws.conditional_formatting.add(rng,IconSetRule(icon_set=icon,type="number",values="1,2,3"))

    def cf_top(self,rng,n=10):
        self.ws.conditional_formatting.add(rng,CellIsRule(
            operator='top',formula=[str(n)],
            fill=PatternFill(start_color=C["primary"],end_color=C["primary"],fill_type="solid"),
            font=Font(bold=True,color=C["white"],name="微软雅黑")))

    def cf_avg(self,rng):
        self.ws.conditional_formatting.add(rng,CellIsRule(
            operator='aboveAverage',
            fill=PatternFill(start_color="D4EDDA",end_color="D4EDDA",fill_type="solid"),
            font=Font(bold=True,color=C["success"],name="微软雅黑")))

    def cf_gt(self,rng,val=0):
        self.ws.conditional_formatting.add(rng,CellIsRule(
            operator='greaterThan',formula=[str(val)],
            fill=PatternFill(start_color="D4EDDA",end_color="D4EDDA",fill_type="solid"),
            font=Font(bold=True,color=C["success"],name="微软雅黑")))

    def cf_lt(self,rng,val=0):
        self.ws.conditional_formatting.add(rng,CellIsRule(
            operator='lessThan',formula=[str(val)],
            fill=PatternFill(start_color="F8D7DA",end_color="F8D7DA",fill_type="solid"),
            font=Font(bold=True,color=C["danger"],name="微软雅黑")))

    # --- 数据验证 ---
    def dv_list(self,col,options:List,prompt="请选择："):
        """下拉菜单"""
        dv=DataValidation(type="list",formula1='"'+','.join(options)+'"',allow_blank=True)
        dv.prompt=prompt; dv.promptTitle="数据验证"
        self.ws.add_data_validation(dv)
        col_l=get_column_letter(col)
        dv.add(f"{col_l}{self.ws.min_row+1}:{col_l}{self.ws.max_row}")

    def dv_number(self,col,min_v=None,max_v=None):
        """数字范围"""
        dv=DataValidation(type="whole",operator="between",
                          formula1=str(min_v or 0),formula2=str(max_v or 999999999),allow_blank=True)
        self.ws.add_data_validation(dv)
        col_l=get_column_letter(col)
        dv.add(f"{col_l}{self.ws.min_row+1}:{col_l}{self.ws.max_row}")

    def dv_date(self,col):
        """日期验证"""
        dv=DataValidation(type="date",allow_blank=True)
        self.ws.add_data_validation(dv)
        col_l=get_column_letter(col)
        dv.add(f"{col_l}{self.ws.min_row+1}:{col_l}{self.ws.max_row}")

    def dv_length(self,col,min_l=0,max_l=255):
        """文本长度"""
        dv=DataValidation(type="textLength",operator="between",
                          formula1=str(min_l),formula2=str(max_l),allow_blank=True)
        self.ws.add_data_validation(dv)
        col_l=get_column_letter(col)
        dv.add(f"{col_l}{self.ws.min_row+1}:{col_l}{self.ws.max_row}")

    # --- 数据清洗 ---
    def deduplicate(self,cols=None):
        seen=set(); rows=[]
        for row in self.ws.iter_rows(min_row=self.ws.min_row):
            key=tuple(row[c-1].value for c in (cols or range(1,self.ws.max_column+1)))
            if key not in seen: seen.add(key); rows.append(row)
        if len(rows)<self.ws.max_row-self.ws.min_row+1:
            # 删除重复行
            existing=set(id(r) for r in rows)
            for row in self.ws.iter_rows(min_row=self.ws.min_row+1):
                if id(row) not in existing:
                    self.ws.delete_rows(row[0].row)
        return len(self.ws.max_row - len(rows))

    def fill_blanks(self,value=""):
        n=0
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value is None or str(cell.value).strip()=="":
                    cell.value=value; n+=1
        return n

    def clean_pipeline(self,steps:List[str]):
        """清洗管道: dedup | blanks | trim | upper | lower | number | date"""
        for step in steps:
            if step=="dedup": self.deduplicate()
            elif step=="blanks": self.fill_blanks("")
            elif step=="trim": self._trim()
            elif step=="upper": self._case("upper")
            elif step=="lower": self._case("lower")
            elif step=="number": self._to_number()
            elif step=="date": self._to_date()

    def _trim(self):
        for row in self.ws.iter_rows():
            for cell in row:
                if isinstance(cell.value,str): cell.value=cell.value.strip()

    def _case(self,mode):
        for row in self.ws.iter_rows():
            for cell in row:
                if isinstance(cell.value,str):
                    cell.value=cell.value.upper() if mode=="upper" else cell.value.lower()

    def _to_number(self):
        for row in self.ws.iter_rows():
            for cell in row:
                if isinstance(cell.value,str):
                    try:
                        cell.value=float(cell.value.replace(",",""))
                    except Exception:
                        pass

    def _to_date(self):
        from datetime import datetime as dt
        for row in self.ws.iter_rows():
            for cell in row:
                if isinstance(cell.value,str) and re.match(r'\d{4}[-/]\d{2}[-/]\d{2}',cell.value):
                    try:
                        cell.value=dt.strptime(cell.value[:10],cell.value[4]).date()
                    except Exception:
                        pass

    # --- 分列 ---
    def split_col(self,col:int,delimiter:str=","):
        """分列"""
        max_r=self.ws.max_row
        new_data={}
        for r in range(self.ws.min_row,max_r+1):
            val=self.ws.cell(row=r,column=col).value
            if val is None: parts=[]
            elif isinstance(val,str): parts=val.split(delimiter)
            else: parts=[str(val)]
            for i,p in enumerate(parts):
                key=f"{get_column_letter(col)}_{i+1}"
                if key not in new_data: new_data[key]=[]
                new_data[key].append(p.strip() if isinstance(p,str) else p)
        # 填充不足的列
        for key in new_data:
            while len(new_data[key])<max_r-self.ws.min_row+1: new_data[key].append("")
        # 插入新列
        for i,key in enumerate(sorted(new_data.keys())):
            for r,val in enumerate(new_data[key],start=self.ws.min_row):
                self.ws.cell(row=r,column=self.ws.max_column+i+1,value=val)
        return list(new_data.keys())

    # --- 合并文件 ---
    def merge_files(self,files:List[str],key_col=1,how="left")->str:
        """合并多个文件"""
        wb=Workbook(); ws=wb.active; ws.title="合并数据"
        written=False
        for f in files:
            wb2=load_workbook(f); ws2=wb2.active
            h_row=list(ws2.iter_rows(min_row=ws2.min_row,max_row=ws2.min_row,values_only=True))[0]
            if not written:
                for col,h in enumerate(h_row,1):
                    cell=ws.cell(row=1,column=col,value=h); ap(cell,ST["h"])
                written=True
            for row in ws2.iter_rows(min_row=ws2.min_row+1,values_only=True):
                if not any(v is not None for v in row): continue
                new_row=list(row)
                # 找匹配行（简单 left join）
                for r2 in ws.iter_rows(min_row=2,values_only=True):
                    if r2[key_col-1]==row[key_col-1]:
                        for col,h in enumerate(h_row,1):
                            if col>len(new_row): new_row.append(r2[col-1])
                ws.append(new_row)
        wb.save("merged_output.xlsx")
        return "merged_output.xlsx"

    # --- 读取数据为 dict ---
    def read_data(self)->Dict:
        data={}
        rows=list(self.ws.iter_rows(values_only=True))
        if not rows: return {}
        for i,row in enumerate(rows):
            for col_idx,val in enumerate(row):
                col_name=get_column_letter(col_idx+1)
                if col_name not in data: data[col_name]=[]
                data[col_name].append(val)
        return data


# ========== 导入导出 ==========
class IO:
    """CSV/JSON/PDF 导入导出"""

    @staticmethod
    def read_csv(path:str,encoding="utf-8")->Dict:
        with open(path,encoding=encoding,newline="") as f:
            reader=csv.DictReader(f)
            rows=list(reader)
        if not rows: return {}
        headers=list(rows[0].keys())
        data={h:[] for h in headers}
        for row in rows:
            for h in headers: data[h].append(row.get(h,""))
        return data

    @staticmethod
    def write_csv(data:Dict,path:str,encoding="utf-8"):
        headers=list(data.keys())
        with open(path,"w",encoding=encoding,newline="") as f:
            writer=csv.DictWriter(f,fieldnames=headers)
            writer.writeheader()
            nr=max((len(v) for v in data.values()),default=0)
            for r in range(nr):
                writer.writerow({h:data[h][r] if r<len(data[h]) else "" for h in headers})

    @staticmethod
    def read_json(path:str)->Dict:
        with open(path,encoding="utf-8") as f: return json.load(f)

    @staticmethod
    def write_json(data:Any,path:str):
        with open(path,"w",encoding="utf-8") as f: json.dump(data,f,ensure_ascii=False,indent=2)

    @staticmethod
    def excel_to_json(xlsx_path:str,json_path:str):
        x=X().open(xlsx_path)
        data=x.read_data()
        IO.write_json(data,json_path)

    @staticmethod
    def excel_to_csv(xlsx_path:str,csv_path:str):
        x=X().open(xlsx_path)
        data=x.read_data()
        IO.write_csv(data,csv_path)

    @staticmethod
    def csv_to_excel(csv_path:str,xlsx_path:str,sheet="数据"):
        data=IO.read_csv(csv_path)
        X().from_dict(data,sheet).save(xlsx_path)

    @staticmethod
    def json_to_excel(json_path:str,xlsx_path:str,sheet="数据"):
        data=IO.read_json(json_path)
        X().from_dict(data,sheet).save(xlsx_path)

    @staticmethod
    def export_pdf(xlsx_path:str,pdf_path:str=None):
        """导出 PDF（依赖 LibreOffice 或 wkhtmltopdf）"""
        pdf_path=pdf_path or xlsx_path.replace(".xlsx",".pdf")
        import subprocess
        # 尝试 LibreOffice
        result=subprocess.run(
            ["libreoffice","--headless","--convert-to","pdf","--outdir",os.path.dirname(pdf_path),xlsx_path],
            capture_output=True,text=True,timeout=60
        )
        if result.returncode!=0:
            return {"status":"error","message":"请安装 LibreOffice: brew install --cask libreoffice","hint":"或手动在 Excel 中另存为 PDF"}
        return {"status":"success","file":pdf_path}


# ========== AI 智能 ==========
class AI:
    """AI 数据分析"""

    @staticmethod
    def describe(data:Dict)->str:
        if not data: return "数据为空"
        headers=list(data.keys())
        nr=max(len(v) for v in data.values())
        lines=[f"📊 数据概览  共 {nr} 行 × {len(headers)} 列"]
        for h in headers:
            vals=[v for v in data[h] if v is not None and v!=""]
            if not vals: continue
            nums=[float(v) for v in vals if isinstance(v,(int,float))]
            if len(nums)>len(vals)*0.7 and nums:
                lines.append(f"\n📈 「{h}」（数值型）")
                lines.append(f"  总计: {sum(nums):,.2f}")
                lines.append(f"  平均: {sum(nums)/len(nums):,.2f}")
                lines.append(f"  最大: {max(nums):,.2f}  最小: {min(nums):,.2f}")
            else:
                lines.append(f"\n📋 「{h}」（文本型）唯一值: {len(set(str(v) for v in vals))} 个")
        return "\n".join(lines)

    @staticmethod
    def suggest_chart(data:Dict)->Dict:
        headers=list(data.keys())
        kw_map={"time":any(k in h for k in ["日期","时间","月","年","季度"] for h in headers),
                "pct":any(k in h for k in ["占比","比例","率"] for h in headers),
                "multi":len(headers)>=3}
        if kw_map["pct"]: return {"type":"pie","reason":"含占比字段，适合饼图"}
        if kw_map["time"]: return {"type":"line","reason":"含时间序列，适合折线图"}
        if kw_map["multi"]: return {"type":"bar","reason":"多维度对比，适合柱状图"}
        return {"type":"bar","reason":"适合柱状图展示"}

    @staticmethod
    def insights(data:Dict)->List[str]:
        insights=[]
        for h in list(data.keys()):
            vals=[v for v in data[h] if isinstance(v,(int,float)) and v!=0]
            if not vals: continue
            avg=sum(vals)/len(vals); mx=max(vals); mn=min(vals)
            mx_i=data[h].index(mx); labels=data.get(list(data.keys())[0],[])
            lbl=labels[mx_i] if mx_i<len(labels) else ""
            if mx>avg*2: insights.append(f"⚠️ 「{h}」最大值 {mx:,.0f} 明显高于均值 {avg:,.0f}（{lbl}）")
            if mn<0: insights.append(f"🔴 「{h}」存在负值 {mn:,.0f}，需检查")
            if len(set(vals))==1: insights.append(f"ℹ️ 「{h}」为常量 {vals[0]}")
        if not insights: insights=["✅ 数据未发现明显异常"]
        return insights

    @staticmethod
    def formula(desc:str,cols:List[str])->Dict:
        d=desc.lower()
        patterns=[
            (r'毛利率|利润率',lambda c:f'=IF({c[1] if len(c)>1 else c[0]}=0,0,{c[0]}/{c[1] if len(c)>1 else c[0]})'),
            (r'统计.*大于(\d+)',lambda c,_,m:f'=COUNTIF({c[0]}:{c[0]},">{m.group(1)}")'),
            (r'统计.*小于(\d+)',lambda c,_,m:f'=COUNTIF({c[0]}:{c[0]},"<{m.group(1)}")'),
            (r'求和|总和|sum',lambda c:f'=SUM({c[0]}:{c[0]})'),
            (r'平均|均值|average',lambda c:f'=AVERAGE({c[0]}:{c[0]})'),
            (r'最大|max',lambda c:f'=MAX({c[0]}:{c[0]})'),
            (r'最小|min',lambda c:f'=MIN({c[0]}:{c[0]})'),
            (r'排名|rank',lambda c:f'=RANK({c[0]}2,{c[0]}:{c[0]},0)'),
            (r'查找|vlookup|lookup',lambda c:f'=VLOOKUP(A2,{c[0]}:{c[-1]},{len(c)},0)' if len(c)>1 else '=INDEX(...)'),
            (r'同比|年增',lambda c,oc,_:f'=IF({c[1]}=0,0,({c[0]}-{c[1]})/{c[1]})' if len(c)>1 else '=IF(prev=0,0,(curr-prev)/prev)'),
            (r'环比|月增',lambda c,oc,_:f'=IF({c[1]}=0,0,({c[0]}-{c[1]})/{c[1]})' if len(c)>1 else '=IF(prev=0,0,(curr-prev)/prev)'),
            (r'如果|if.*大于|if.*小于',lambda c:f'=IF({c[0]}>100,"高","低")'),
            (r'年龄|age',lambda c:f'=DATEDIF({c[0]},TODAY(),"Y")'),
            (r'工作日|networkdays',lambda c:f'=NETWORKDAYS({c[0]},{c[1] if len(c)>1 else c[0]})'),
        ]
        for pat,fn in patterns:
            m=re.search(pat,d)
            if m: return {"status":"success","input":desc,"columns":cols,"formulas":[{"formula":fn(cols or ["A"]),"type":"Excel","description":f"根据「{desc}」生成","note":"复制到单元格并修改列引用"}]}
        return {"status":"success","input":desc,"columns":cols,"formulas":[{"formula":"=SUM(A:A)","type":"Excel","description":"示例：求和","note":"请更详细描述需求"}]}


# ========== 报表生成器 ==========
class R:
    """报表模板"""

    @staticmethod
    def sales_dashboard(data:Dict,output="销售看板.xlsx")->str:
        wb=Workbook(); ws=wb.active; ws.title="销售看板"
        headers=list(data.keys()); nr=max(len(v) for v in data.values())

        # 大标题
        ws.merge_cells(f"A1:{get_column_letter(max(4,len(headers)+4))}1")
        c=ws["A1"]; c.value=f"📊 销售数据看板  {datetime.now().strftime('%Y年%m月%d日')}"
        ap(c,ST["title"]); c.fill=fill(C["gray_l"]); ws.row_dimensions[1].height=50

        # KPI 卡片
        kpis=[]
        for h in headers:
            vals=[v for v in data[h] if isinstance(v,(int,float))]
            if vals: kpis.append((h,sum(vals)))
        kpi_cols=max(4,len(kpis))
        for i,(name,val) in enumerate(kpis[:6]):
            col=i*2+1
            ws.merge_cells(start_row=3,start_column=col,end_row=3,end_column=col+1)
            lc=ws.cell(row=3,column=col,value=name); ap(lc,{"sz":9,"fc":C["text_m"],"ha":"center"})
            ws.merge_cells(start_row=4,start_column=col,end_row=4,end_column=col+1)
            vc=ws.cell(row=4,column=col,value=val)
            ap(vc,ST["kpi"]); vc.number_format="¥#,##0"
            ws.row_dimensions[5].height=8

        # 数据表
        dr=7
        for col,h in enumerate(headers,1):
            cell=ws.cell(row=dr,column=col,value=h); ap(cell,ST["h"])
        for r in range(nr):
            bg=C["gray_l"] if r%2==0 else C["white"]
            for col,h in enumerate(headers,1):
                val=data[h][r] if r<len(data[h]) else ""
                cell=ws.cell(row=dr+r+1,column=col,value=val)
                cell.fill=fill(bg); ap(cell,{"sz":10})
                if any(kw in h for kw in ["金额","销售","利润","成本","收入","预算","实际"]):
                    cell.number_format="¥#,##0.00"; cell.alignment=aln("right")

        # 柱状图
        if nr>0 and len(headers)>1:
            nc=len(headers)
            dc=Reference(ws,min_col=2,min_row=dr,max_col=nc,max_row=dr+nr)
            cr=Reference(ws,min_col=1,min_row=dr+1,max_row=dr+nr)
            chart=BarChart(type="col",style=10); chart.title="各部门对比"
            chart.width=16; chart.height=10
            chart.add_data(dc,titles_from_data=True); chart.set_categories(cr)
            for i,s in enumerate(chart.series):
                s.graphicalProperties.solidFill=CHART[i%len(CHART)]
            ws.add_chart(chart,f"{get_column_letter(nc+1)}3")

            # 饼图
            pie=PieChart(); pie.title="占比分布"; pie.width=12; pie.height=10
            pdr=Reference(ws,min_col=2,min_row=dr+1,max_row=dr+min(nr,6))
            pcr=Reference(ws,min_col=1,min_row=dr+1,max_row=dr+min(nr,6))
            pie.add_data(pdr); pie.set_categories(pcr)
            for i,pt in enumerate(pie.series[0].data_points):
                pt.graphicalProperties.solidFill=CHART[i%len(CHART)]
            ws.add_chart(pie,f"{get_column_letter(nc+1)}17")

        ws.column_dimensions["A"].width=15
        for col in range(2,len(headers)+1): ws.column_dimensions[get_column_letter(col)].width=14
        wb.save(output); return output

    @staticmethod
    def financial(data:Dict,output="财务报表.xlsx")->str:
        wb=Workbook()
        ws=wb.active; ws.title="利润表"
        R._fin_sheet(ws,data,"利润表（损益表）")
        ws2=wb.create_sheet("资产负债表")
        R._fin_sheet(ws2,data,"资产负债表")
        wb.save(output); return output

    @staticmethod
    def _fin_sheet(ws,data:Dict,title:str):
        headers=list(data.keys()); nr=max(len(v) for v in data.values())
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        c=ws["A1"]; c.value=title; ap(c,ST["title"]); c.fill=fill(C["gray_l"])
        for col,h in enumerate(headers,1):
            cell=ws.cell(row=3,column=col,value=h); ap(cell,ST["h"])
        for r in range(nr):
            bg=C["gray_l"] if r%2==0 else C["white"]
            for col,h in enumerate(headers,1):
                val=data[h][r] if r<len(data[h]) else ""
                cell=ws.cell(row=4+r,column=col,value=val)
                cell.fill=fill(bg); ap(cell,{"sz":10})
                if any(kw in h for kw in ["金额","收入","利润","成本"]):
                    cell.number_format="¥#,##0.00"; cell.alignment=aln("right")
        ws.column_dimensions["A"].width=20
        for col in range(2,len(headers)+1): ws.column_dimensions[get_column_letter(col)].width=14

    @staticmethod
    def gantt(data:Dict,output="项目进度.xlsx")->str:
        wb=Workbook(); ws=wb.active; ws.title="项目进度"
        headers=["项目名称","负责人","开始日期","结束日期","进度(%)","状态","优先级","预算","实际"]
        items=list(data.values())[0] if data else []

        ws.merge_cells(f"A1:I1")
        c=ws["A1"]; c.value="📅 项目进度看板"; ap(c,ST["title"]); c.fill=fill(C["gray_l"]); ws.row_dimensions[1].height=45

        for col,h in enumerate(headers,1):
            cell=ws.cell(row=3,column=col,value=h); ap(cell,ST["h"])

        for i,item in enumerate(items if isinstance(items,list) else []):
            bg=C["gray_l"] if i%2==0 else C["white"]
            for col in range(1,len(headers)+1):
                val=item[col-1] if col-1<len(item) else ""
                cell=ws.cell(row=4+i,column=col,value=val)
                cell.fill=fill(bg); ap(cell,{"sz":10})

            # 进度条
            try: pg=float(str(item[4] if len(item)>4 else 50).replace("%",""))
            except: pg=50
            pc=ws.cell(row=4+i,column=9,value=f"{pg:.0%}")
            pc.fill=fill(C["success"] if pg>=80 else C["warning"] if pg>=50 else C["danger"])
            pc.font=Font(bold=True,size=9,color=C["white"],name="微软雅黑"); pc.alignment=aln(); pc.border=B()

        ws.column_dimensions["A"].width=16
        for col in ["B","C","D","E","F","G","H","I"]: ws.column_dimensions[col].width=12
        wb.save(output); return output

    @staticmethod
    def reconciliation(data:Dict,output="对账清单.xlsx")->str:
        wb=Workbook(); ws=wb.active; ws.title="对账"
        headers=list(data.keys()); nr=max(len(v) for v in data.values())

        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        c=ws["A1"]; c.value=f"🔍 对账清单  {datetime.now().strftime('%Y-%m-%d')}"
        ap(c,ST["title"]); c.fill=fill(C["gray_l"])

        for col,h in enumerate(headers,1):
            cell=ws.cell(row=3,column=col,value=h); ap(cell,ST["h"])

        diff_col=None
        for col,h in enumerate(headers,1):
            if "差异" in str(h) or "差额" in str(h): diff_col=col; break

        for r in range(nr):
            bg=C["gray_l"] if r%2==0 else C["white"]
            for col,h in enumerate(headers,1):
                val=data[h][r] if r<len(data[h]) else ""
                cell=ws.cell(row=4+r,column=col,value=val)
                cell.fill=fill(bg); ap(cell,{"sz":10})

            # 差异自动标色
            if diff_col:
                dc=ws.cell(row=4+r,column=diff_col)
                try:
                    dv=float(str(dc.value or "0").replace(",","").replace("¥","").replace("%",""))
                    if dv>0: dc.fill=fill("D4EDDA"); dc.font=Font(bold=True,color=C["success"],name="微软雅黑")
                    elif dv<0: dc.fill=fill("F8D7DA"); dc.font=Font(bold=True,color=C["danger"],name="微软雅黑")
                except Exception:
                    pass

        # 汇总
        sr=4+nr+2
        ws.cell(row=sr,column=1,value="汇总").font=Font(bold=True,size=12,color=C["text_d"],name="微软雅黑")
        ws.cell(row=sr+1,column=1,value=f"总行数: {nr}").font=Font(size=10,color=C["text_m"],name="微软雅黑")
        ws.cell(row=sr+2,column=1,value=f"生成: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font=Font(size=10,color=C["text_m"],name="微软雅黑")

        wb.save(output); return output

    @staticmethod
    def daily_report(data:Dict,output="日报.xlsx")->str:
        """每日工作日报"""
        wb=Workbook(); ws=wb.active; ws.title="日报"
        headers=list(data.keys()) if data else ["事项","负责人","状态","备注"]

        ws.merge_cells(f"A1:{get_column_letter(max(4,len(headers)))}1")
        c=ws["A1"]; c.value=f"📋 工作日报  {datetime.now().strftime('%Y年%m月%d日 %A')}"
        ap(c,ST["title"]); c.fill=fill(C["gray_l"]); ws.row_dimensions[1].height=45

        # KPI
        for col,h in enumerate(headers,1):
            cell=ws.cell(row=3,column=col,value=h); ap(cell,ST["h"])

        if data:
            nr=max(len(v) for v in data.values())
            for r in range(nr):
                bg=C["gray_l"] if r%2==0 else C["white"]
                for col,h in enumerate(headers,1):
                    val=data[h][r] if r<len(data[h]) else ""
                    cell=ws.cell(row=4+r,column=col,value=val)
                    cell.fill=fill(bg); ap(cell,{"sz":10})
                    if h=="状态":
                        if val=="完成": cell.fill=fill("D4EDDA"); cell.font=Font(bold=True,color=C["success"],name="微软雅黑")
                        elif val in ["进行中","处理中"]: cell.fill=fill("FFF3CD"); cell.font=Font(bold=True,color=C["warning"],name="微软雅黑")
                        elif val=="待处理": cell.fill=fill("F8D7DA"); cell.font=Font(bold=True,color=C["danger"],name="微软雅黑")

        ws.column_dimensions["A"].width=20
        for col in range(2,len(headers)+1): ws.column_dimensions[get_column_letter(col)].width=14

        # 明日计划
        ws2=wb.create_sheet("明日计划")
        ws2["A1"]="📅 明日计划"; ws2["A1"].font=Font(bold=True,size=16,color=C["primary"],name="微软雅黑")
        ws2["A1"].fill=fill(C["gray_l"]); ws2.merge_cells("A1:D1")
        for col,h in enumerate(["事项","负责人","预计时间","备注"],1):
            cell=ws2.cell(row=3,column=col,value=h); ap(cell,ST["h"])

        wb.save(output); return output

    @staticmethod
    def weekly_report(data:Dict,output="周报.xlsx")->str:
        """周报"""
        wb=Workbook(); ws=wb.active; ws.title="本周总结"
        headers=list(data.keys()) if data else ["日期","完成任务","进行中","下周计划"]

        ws.merge_cells(f"A1:{get_column_letter(max(4,len(headers)))}1")
        c=ws["A1"]; c.value=f"📊 周报  {datetime.now().strftime('%Y年 第%W周')}"
        ap(c,ST["title"]); c.fill=fill(C["gray_l"]); ws.row_dimensions[1].height=45

        for col,h in enumerate(headers,1):
            cell=ws.cell(row=3,column=col,value=h); ap(cell,ST["h"])

        if data:
            nr=max(len(v) for v in data.values())
            for r in range(nr):
                bg=C["gray_l"] if r%2==0 else C["white"]
                for col,h in enumerate(headers,1):
                    val=data[h][r] if r<len(data[h]) else ""
                    cell=ws.cell(row=4+r,column=col,value=val)
                    cell.fill=fill(bg); ap(cell,{"sz":10,"ha":"left","brd":True})

        ws.column_dimensions["A"].width=15
        for col in range(2,len(headers)+1): ws.column_dimensions[get_column_letter(col)].width=20

        # KPI 汇总
        ws2=wb.create_sheet("KPI汇总")
        ws2["A1"]="📈 KPI 汇总"; ws2["A1"].font=Font(bold=True,size=16,color=C["primary"],name="微软雅黑")
        ws2["A1"].fill=fill(C["gray_l"]); ws2.merge_cells("A1:C1")

        for col,h in enumerate(["指标","本周","环比"],1):
            cell=ws2.cell(row=3,column=col,value=h); ap(cell,ST["h"])

        wb.save(output); return output


# ========== 主程序 ==========
def main():
    p=argparse.ArgumentParser(description="PansExcel Pro Max",formatter_class=argparse.RawDescriptionHelpFormatter)
    sub=p.add_subparsers(dest="cmd")

    def add(name,**kwargs): return sub.add_parser(name,**kwargs)

    # create
    sp=add("create"); sp.add_argument("-d","--data",required=True); sp.add_argument("-o","--output",default="out.xlsx"); sp.add_argument("-s","--sheet",default="数据")

    # report
    sp=add("report"); sp.add_argument("-d","--data",required=True); sp.add_argument("-t","--template",required=True,
        choices=["sales","financial","gantt","recon","daily","weekly","sales-dashboard"])
    sp.add_argument("-o","--output",default="report.xlsx")

    # analyze
    sp=add("analyze"); sp.add_argument("-f","--file",required=True)

    # formula
    sp=add("formula"); sp.add_argument("--描述",required=True); sp.add_argument("-c","--columns")

    # chart
    sp=add("chart"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-t","--type",default="bar",choices=["bar","line","pie","area","radar","combo","scatter"])
    sp.add_argument("-p","--pos",default="F2"); sp.add_argument("-T","--title",default="")

    # format
    sp=add("format"); sp.add_argument("-f","--file",required=True); sp.add_argument("-o","--output")

    # cond
    sp=add("cond"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-r","--range",default="B2:B100")
    sp.add_argument("-t","--type",default="scale",choices=["scale","bar","icon","top","avg","gt","lt"])

    # dv (data validation)
    sp=add("dv"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-c","--col",type=int,required=True)
    sp.add_argument("-t","--type",default="list",choices=["list","number","date","length"])
    sp.add_argument("-o","--options",help="列表选项，逗号分隔")
    sp.add_argument("--min",type=float); sp.add_argument("--max",type=float)
    sp.add_argument("--min-len",type=int,default=0); sp.add_argument("--max-len",type=int,default=255)

    # clean
    sp=add("clean"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-s","--steps",default="dedup,blanks,trim",help="逗号分隔: dedup,blanks,trim,upper,lower,number,date")
    sp.add_argument("-o","--output")

    # split
    sp=add("split"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-c","--col",type=int,required=True)
    sp.add_argument("-d","--delimiter",default=",")
    sp.add_argument("-o","--output")

    # import/export
    sp=add("import"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-o","--output"); sp.add_argument("-t","--type",choices=["xlsx","csv","json"],default="xlsx")
    sp=add("export"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-o","--output")
    sp.add_argument("-t","--type",choices=["csv","json","pdf"],default="csv")
    sp=add("pdf"); sp.add_argument("-f","--file",required=True); sp.add_argument("-o","--output")

    # merge
    sp=add("merge"); sp.add_argument("-f","--files",nargs="+",required=True)
    sp.add_argument("-o","--output",default="merged.xlsx")
    sp.add_argument("-k","--key",type=int,default=1)

    # insights
    sp=add("insights"); sp.add_argument("-f","--file")

    args=p.parse_args()

    if not HAS_OPENPYXL:
        print(json.dumps({"status":"error","message":"需要安装: pip install openpyxl"},ensure_ascii=False,indent=2)); return

    def load_data(s):
        try: return json.loads(s)
        except: 
            try: return eval(s)
            except: return {}

    def get_cols(s):
        return [c.strip() for c in s.split(",")] if s else []

    x=X(); ai=AI()

    try:
        if args.cmd=="create":
            data=load_data(args.data)
            path=X().from_dict(data,args.sheet).save(args.output)
            print(json.dumps({"status":"success","file":path},ensure_ascii=False,indent=2))

        elif args.cmd=="report":
            data=load_data(args.data)
            if args.template=="sales": path=R.sales_dashboard(data,args.output)
            elif args.template=="sales-dashboard": path=R.sales_dashboard(data,args.output)
            elif args.template=="financial": path=R.financial(data,args.output)
            elif args.template=="gantt": path=R.gantt(data,args.output)
            elif args.template=="recon": path=R.reconciliation(data,args.output)
            elif args.template=="daily": path=R.daily_report(data,args.output)
            elif args.template=="weekly": path=R.weekly_report(data,args.output)
            else: path=args.output
            print(json.dumps({"status":"success","file":path},ensure_ascii=False,indent=2))

        elif args.cmd=="analyze":
            x.open(args.file); data=x.read_data()
            print(json.dumps({"status":"success","description":ai.describe(data),
                             "chart":ai.suggest_chart(data),"insights":ai.insights(data)},ensure_ascii=False,indent=2))

        elif args.cmd=="formula":
            cols=get_cols(args.columns)
            print(json.dumps(ai.formula(args.描述,cols),ensure_ascii=False,indent=2))

        elif args.cmd=="chart":
            x.open(args.file)
            t=args.type
            if t=="bar": x.add_bar(pos=args.pos,title=args.title)
            elif t=="line": x.add_line(pos=args.pos,title=args.title)
            elif t=="pie": x.add_pie(pos=args.pos,title=args.title)
            elif t=="area": x.add_area(pos=args.pos,title=args.title)
            elif t=="radar": x.add_radar(pos=args.pos,title=args.title)
            elif t=="combo": x.add_combo(pos=args.pos,title=args.title)
            x.save(args.file)
            print(json.dumps({"status":"success","message":"图表已添加"},ensure_ascii=False,indent=2))

        elif args.cmd=="format":
            x.open(args.file); x.美化(); x.add_title("数据报表"); x.autofilter(); x.freeze()
            x.save(args.output or args.file)
            print(json.dumps({"status":"success","file":x.save(args.output or args.file)},ensure_ascii=False,indent=2))

        elif args.cmd=="cond":
            x.open(args.file)
            if args.type=="scale": x.cf_scale(args.range)
            elif args.type=="bar": x.cf_bar(args.range)
            elif args.type=="icon": x.cf_icon(args.range)
            elif args.type=="top": x.cf_top(args.range)
            elif args.type=="avg": x.cf_avg(args.range)
            elif args.type=="gt": x.cf_gt(args.range)
            elif args.type=="lt": x.cf_lt(args.range)
            x.save(args.file)
            print(json.dumps({"status":"success","message":"条件格式已应用"},ensure_ascii=False,indent=2))

        elif args.cmd=="dv":
            x.open(args.file)
            if args.type=="list":
                opts=[c.strip() for c in (args.options or "").split(",") if c.strip()]
                x.dv_list(args.col,opts)
            elif args.type=="number": x.dv_number(args.col,args.min,args.max)
            elif args.type=="date": x.dv_date(args.col)
            elif args.type=="length": x.dv_length(args.col,args.min_len,args.max_len)
            x.save(args.file)
            print(json.dumps({"status":"success","message":"数据验证已添加"},ensure_ascii=False,indent=2))

        elif args.cmd=="clean":
            x.open(args.file); steps=[s.strip() for s in args.steps.split(",")]
            removed=x.deduplicate(); filled=x.fill_blanks()
            x.clean_pipeline(steps)
            path=x.save(args.output or args.file)
            print(json.dumps({"status":"success","file":path,"removed_duplicates":removed,"filled_blanks":filled},ensure_ascii=False,indent=2))

        elif args.cmd=="split":
            x.open(args.file); new_cols=x.split_col(args.col,args.delimiter)
            x.save(args.output or args.file)
            print(json.dumps({"status":"success","new_columns":new_cols},ensure_ascii=False,indent=2))

        elif args.cmd=="import":
            ext=os.path.splitext(args.file)[1].lower()
            if ext==".csv":
                data=IO.read_csv(args.file); output=args.output or args.file.replace(".csv",".xlsx")
                X().from_dict(data).save(output)
            elif ext==".json":
                data=IO.read_json(args.file); output=args.output or args.file.replace(".json",".xlsx")
                X().from_dict(data).save(output)
            else:
                print(json.dumps({"status":"error","message":"仅支持 .csv 和 .json 文件"},ensure_ascii=False,indent=2)); return
            print(json.dumps({"status":"success","file":output},ensure_ascii=False,indent=2))

        elif args.cmd=="export":
            if args.type=="csv":
                x.open(args.file); data=x.read_data(); output=args.output or args.file.replace(".xlsx",".csv")
                IO.write_csv(data,output)
            elif args.type=="json":
                x.open(args.file); data=x.read_data(); output=args.output or args.file.replace(".xlsx",".json")
                IO.write_json(data,output)
            elif args.type=="pdf":
                result=IO.export_pdf(args.file,args.output); print(json.dumps(result,ensure_ascii=False,indent=2)); return
            print(json.dumps({"status":"success","file":output},ensure_ascii=False,indent=2))

        elif args.cmd=="pdf":
            print(json.dumps(IO.export_pdf(args.file,args.output),ensure_ascii=False,indent=2))

        elif args.cmd=="merge":
            path=x.merge_files(args.files,key_col=args.key,how="left")
            print(json.dumps({"status":"success","file":path},ensure_ascii=False,indent=2))

        elif args.cmd=="insights":
            x.open(args.file); data=x.read_data()
            print(json.dumps({"status":"success","insights":ai.insights(data)},ensure_ascii=False,indent=2))

        else:
            print("""
PansExcel Pro Max - 全功能智能 Excel

  create  python3 excel.py create -d '{"列":["值1","值2"]}' -o out.xlsx
  report  python3 excel.py report -d '{...}' -t sales -o report.xlsx
  chart   python3 excel.py chart -f data.xlsx -t bar -T "标题" -p F2
  format  python3 excel.py format -f data.xlsx -o 美化.xlsx
  cond    python3 excel.py cond -f data.xlsx -r B2:B100 -t scale
  dv      python3 excel.py dv -f data.xlsx -c 3 -t list -o "选项1,选项2,选项3"
  clean   python3 excel.py clean -f data.xlsx -s "dedup,blanks,trim"
  split   python3 excel.py split -f data.xlsx -c 1 -d ","
  import  python3 excel.py import -f data.csv -o out.xlsx
  export  python3 excel.py export -f data.xlsx -t csv
  pdf     python3 excel.py pdf -f data.xlsx
  merge   python3 excel.py merge -f a.xlsx b.xlsx -o merged.xlsx
  insights python3 excel.py insights -f data.xlsx
""")
    except Exception as e:
        print(json.dumps({"status":"error","message":str(e)},ensure_ascii=False,indent=2))


# ============================================================
# 第二阶段升级：瀑布图 | 直方图 | 迷你图 | AI图表 | 数据透视 | 仪表板
# ============================================================

class ChartAdvanced:
    """高级图表"""

    @staticmethod
    def waterfall(ws, pos="H2", title="瀑布图", data_col=2, label_col=1):
        """瀑布图（用堆叠柱状图模拟）"""
        nr = ws.max_row
        if nr < 3:
            return None

        # 读取数据，计算瀑布
        labels = []
        values = []
        running = 0
        for r in range(2, nr + 1):
            lbl = ws.cell(row=r, column=label_col).value or f"项{r-1}"
            val = ws.cell(row=r, column=data_col).value
            if val is None:
                continue
            try:
                val = float(val)
            except:
                continue
            labels.append(str(lbl))
            if r == 2:
                running = val
                values.append(val)
            else:
                if val > running:
                    values.append(val - running)
                else:
                    values.append(running - val)
                running = val

        if not labels:
            return None

        # 创建辅助列（堆叠数据）
        ws.cell(row=1, column=ws.max_column + 1, value="瀑布基准")
        ws.cell(row=1, column=ws.max_column + 1, value="瀑布变化")
        base_col = ws.max_column - 1
        change_col = ws.max_column

        running2 = 0
        for i, val in enumerate(values):
            if i == 0:
                ws.cell(row=i + 2, column=base_col, value=0)
                ws.cell(row=i + 2, column=change_col, value=val)
                running2 = val
            else:
                ws.cell(row=i + 2, column=base_col, value=running2)
                ws.cell(row=i + 2, column=change_col, value=val)
                running2 += val

        # 创建堆叠柱状图
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.style = 10
        chart.title = title
        chart.width = 18
        chart.height = 12
        chart.title_font = Font(bold=True, size=13, color=C["text_d"], name="微软雅黑")

        n = len(values)
        base_ref = Reference(ws, min_col=base_col, min_row=2, max_row=n + 1)
        change_ref = Reference(ws, min_col=change_col, min_row=2, max_row=n + 1)
        cats = Reference(ws, min_col=label_col, min_row=2, max_row=n + 1)

        chart.add_data(base_ref, titles_from_data=False)
        chart.add_data(change_ref, titles_from_data=False)
        chart.set_categories(cats)

        # 基准列透明
        series = chart.series[0]
        series.graphicalProperties.noFill = True
        series.graphicalProperties.line.solidFill = "FFFFFF"

        # 变化列颜色
        for i, val in enumerate(values[1:], 1):
            change = values[i]
            if i == 1:
                series2 = chart.series[1]
                series2.graphicalProperties.solidFill = C["primary"]

        ws.add_chart(chart, pos)
        return chart

    @staticmethod
    def histogram(ws, data_col=2, label_col=1, bins=10, pos="H2", title="分布直方图"):
        """直方图"""
        # 收集数值
        vals = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=data_col).value
            if isinstance(v, (int, float)) and v is not None:
                vals.append(float(v))

        if not vals:
            return None

        mn, mx = min(vals), max(vals)
        if mx == mn:
            return None

        bin_width = (mx - mn) / bins
        hist = [0] * bins
        for v in vals:
            idx = min(int((v - mn) / bin_width), bins - 1)
            hist[idx] += 1

        # 写入直方图数据
        ws.cell(row=1, column=ws.max_column + 1, value="区间")
        ws.cell(row=1, column=ws.max_column + 2, value="频数")
        c1 = ws.max_column - 1
        c2 = ws.max_column

        for i in range(bins):
            ws.cell(row=i + 2, column=c1, value=f"{mn + i*bin_width:.1f}~{mn + (i+1)*bin_width:.1f}")
            ws.cell(row=i + 2, column=c2, value=hist[i])

        # 柱状图
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        chart.width = 16
        chart.height = 10
        chart.title_font = Font(bold=True, size=13, color=C["text_d"], name="微软雅黑")
        dr = Reference(ws, min_col=c2, min_row=1, max_row=bins + 1)
        cr = Reference(ws, min_col=c1, min_row=2, max_row=bins + 1)
        chart.add_data(dr, titles_from_data=True)
        chart.set_categories(cr)
        for i, s in enumerate(chart.series):
            s.graphicalProperties.solidFill = CHART[i % len(CHART)]
        ws.add_chart(chart, pos)
        return chart

    @staticmethod
    def add_sparkline(ws, pos_cell, data_range, chart_type="line"):
        """迷你图"""
        from openpyxl.cell.text import Cell
        # openpyxl 迷你图支持（3.1+）
        try:
            from openpyxl.chart.sparkline import Sparkline
            sp = Sparkline()
            if chart_type == "line":
                sp.type = "line"
            elif chart_type == "bar":
                sp.type = "bar"
            elif chart_type == "win":
                sp.type = "winLoss"
            sp.ref = data_range
            ws[pos_cell].value = ""
            ws.add_sparkline(ws[pos_cell], sp)
        except Exception:
            pass

    @staticmethod
    def dynamic_chart(ws, title="动态图表", pos="H2"):
        """动态图表（带数据标签）"""
        chart = LineChart()
        chart.style = 12
        chart.title = title
        chart.width = 18
        chart.height = 12
        chart.title_font = Font(bold=True, size=14, color=C["text_d"], name="微软雅黑")
        nr = ws.max_row
        nc = ws.max_column
        if nr > 1 and nc > 1:
            dr = Reference(ws, min_col=2, min_row=1, max_col=nc, max_row=nr)
            cr = Reference(ws, min_col=1, min_row=2, max_row=nr)
            chart.add_data(dr, titles_from_data=True)
            chart.set_categories(cr)
            for i, s in enumerate(chart.series):
                s.graphicalProperties.line.solidFill = CHART[i % len(CHART)]
                s.smooth = True
                s.line.width = 25000
                s.dLbls = DataLabelList()
                s.dLbls.showVal = True
                s.dLbls.showCatName = False
                s.dLbls.showLegendKey = False
        ws.add_chart(chart, pos)
        return chart


class PivotTable:
    """简易数据透视（模拟）"""

    @staticmethod
    def create(ws_source, row_field=1, col_field=None, val_field=2, agg="sum"):
        """创建透视结果"""
        # 读取数据
        rows = list(ws_source.iter_rows(values_only=True))
        headers = rows[0] if rows else []
        data_rows = rows[1:]

        # 聚合
        agg_map = {
            "sum": lambda vals: sum(v for v in vals if isinstance(v, (int, float)) and v is not None),
            "avg": lambda vals: sum(v for v in vals if isinstance(v, (int, float))) / max(1, len([v for v in vals if isinstance(v, (int, float))])) if any(isinstance(v, (int, float)) for v in vals) else 0,
            "count": lambda vals: len([v for v in vals if v is not None]),
            "max": lambda vals: max((v for v in vals if isinstance(v, (int, float))), default=0),
            "min": lambda vals: min((v for v in vals if isinstance(v, (int, float))), default=0),
        }
        fn = agg_map.get(agg, agg_map["sum"])

        # 构建透视
        from collections import defaultdict
        pivot = defaultdict(lambda: defaultdict(list))

        for row in data_rows:
            row_key = str(row[row_field - 1] or "")
            if col_field:
                col_key = str(row[col_field - 1] or "")
                val = row[val_field - 1] if val_field <= len(row) else None
                pivot[row_key][col_key].append(val)
            else:
                val = row[val_field - 1] if val_field <= len(row) else None
                pivot[row_key]["_total"].append(val)

        # 创建透视表
        wb = Workbook()
        ws = wb.active
        ws.title = "透视表"

        # 表头
        col_keys = sorted({ck for rv in pivot.values() for ck in rv.keys()})
        ws.cell(row=1, column=1, value=str(headers[row_field - 1]) if row_field <= len(headers) else "行标签")
        for i, ck in enumerate(col_keys, 2):
            cell = ws.cell(row=1, column=i, value=ck)
            ap(cell, ST["h"])

        # 数据
        for ri, (rk, cols) in enumerate(sorted(pivot.items())):
            bg = C["gray_l"] if ri % 2 == 0 else C["white"]
            cell = ws.cell(row=ri + 2, column=1, value=rk)
            cell.fill = fill(bg); ap(cell, {"sz": 10, "ha": "left"})
            for ci, ck in enumerate(col_keys, 2):
                vals = cols.get(ck, [])
                result = fn(vals)
                cell = ws.cell(row=ri + 2, column=ci, value=result if vals else "")
                cell.fill = fill(bg)
                ap(cell, {"sz": 10, "num_fmt": "#,##0.00"})

        # 格式
        ws.column_dimensions["A"].width = 20
        for col in range(2, len(col_keys) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 14

        # 汇总行
        last_r = len(pivot) + 2
        ws.cell(row=last_r, column=1, value="合计").font = Font(bold=True, size=11, color=C["text_d"], name="微软雅黑")
        for ci in range(2, len(col_keys) + 2):
            total = sum(ws.cell(row=r, column=ci).value or 0 for r in range(2, last_r))
            cell = ws.cell(row=last_r, column=ci, value=total)
            ap(cell, ST["h"])

        return wb


class AIGraph:
    """AI 驱动的图表推荐与生成"""

    CHART_PATTERNS = {
        "占比|比例|份额": "pie",
        "趋势|增长|下降|随时间": "line",
        "对比|排名|排序|分布": "bar",
        "堆积|累计|组成": "area",
        "多维度|能力|综合": "radar",
        "对比+趋势|双轴": "combo",
        "瀑布|分解|增减": "waterfall",
        "直方|分布|频次": "histogram",
    }

    @staticmethod
    def describe(data: Dict) -> str:
        """生成数据描述"""
        if not data:
            return "数据为空"
        headers = list(data.keys())
        nr = max(len(v) for v in data.values())
        lines = [f"📊 数据概览：{nr} 行 × {len(headers)} 列"]

        for h in headers:
            vals = [v for v in data[h] if v is not None and v != ""]
            if not vals:
                continue
            nums = [float(v) for v in vals if isinstance(v, (int, float))]
            if nums and len(nums) > len(vals) * 0.7:
                lines.append(f"\n📈 「{h}」数值型")
                lines.append(f"  总计: {sum(nums):,.2f}  均值: {sum(nums)/len(nums):,.2f}")
                lines.append(f"  范围: [{min(nums):,.2f} ~ {max(nums):,.2f}]")
            else:
                lines.append(f"\n📋 「{h}」文本型（{len(set(str(v) for v in vals))} 个唯一值）")
        return "\n".join(lines)

    @staticmethod
    def suggest(data: Dict) -> Dict:
        """智能推荐图表"""
        headers = list(data.keys())
        # 时间特征
        time_kw = ["日期", "时间", "月", "年", "季度", "周", "week", "month", "year", "date"]
        has_time = any(any(k in h.lower() for k in time_kw) for h in headers)
        # 占比特征
        pct_kw = ["占比", "比例", "份额", "rate", "ratio", "%"]
        has_pct = any(any(k in h for k in pct_kw) for h in headers)
        # 分解特征
        has_diff = any(k in h for h in headers for k in ["差异", "变化", "delta", "diff"])

        if has_pct and len(headers) >= 2:
            return {"type": "pie", "reason": "含占比数据，适合饼图", "columns": headers[:2]}
        if has_time and len(headers) >= 2:
            return {"type": "line", "reason": "含时间维度，适合折线图", "columns": headers}
        if has_diff:
            return {"type": "waterfall", "reason": "含差异数据，适合瀑布图", "columns": headers}
        if len(headers) >= 3:
            return {"type": "bar", "reason": "多维度对比，适合柱状图", "columns": headers}
        return {"type": "bar", "reason": "默认柱状图", "columns": headers}

    @staticmethod
    def chart_from_nl(data: Dict, description: str) -> Dict:
        """自然语言 → 图表生成"""
        desc = description.lower()
        # 匹配关键词
        for kw, chart_type in AIGraph.CHART_PATTERNS.items():
            if any(k in desc for k in kw.split("|")):
                return {
                    "status": "success",
                    "input": description,
                    "chart_type": chart_type,
                    "reason": f"根据「{description}」推荐 {chart_type} 图",
                    "tip": f"使用: python3 excel.py chart -f <file> -t {chart_type} -T '{description}'"
                }
        # 默认推荐
        return {
            "status": "success",
            "input": description,
            "chart_type": "bar",
            "reason": "默认推荐柱状图",
            "tip": f"使用: python3 excel.py chart -f <file> -t bar -T '{description}'"
        }

    @staticmethod
    def generate_insights(data: Dict) -> List[str]:
        """生成深度洞察"""
        insights = []
        headers = list(data.keys())

        for h in headers:
            vals = [v for v in data[h] if isinstance(v, (int, float)) and v is not None]
            if not vals or len(vals) < 2:
                continue

            avg = sum(vals) / len(vals)
            std = math.sqrt(sum((v - avg) ** 2 for v in vals) / len(vals)) if len(vals) > 1 else 0
            mx = max(vals)
            mn = min(vals)
            mx_i = data[h].index(mx)
            mn_i = data[h].index(mn)
            labels = data.get(list(data.keys())[0], [])
            mx_lbl = labels[mx_i] if mx_i < len(labels) else ""
            mn_lbl = labels[mn_i] if mn_i < len(labels) else ""

            # 离群值检测（超过2个标准差）
            outliers = [v for v in vals if abs(v - avg) > 2 * std]
            if outliers:
                insights.append(f"⚠️ 「{h}」存在离群值: {[f'{o:.1f}' for o in outliers]}")

            if mx > avg * 2:
                insights.append(f"📈 「{h}」最大值 {mx:,.0f} 是均值的 {mx/avg:.1f} 倍（{mx_lbl}）")
            if mn < 0:
                insights.append(f"🔴 「{h}」存在负值 {mn:,.0f}，需检查数据")
            if mx == mn:
                insights.append(f"ℹ️ 「{h}」所有值相同，为常量 {mx}")
            if std / max(abs(avg), 1) < 0.05:
                insights.append(f"📊 「{h}」数据波动极小（变异系数<5%），分布均匀")

            # 相关性（如果有多个数值列）
            nums_self = vals  # vals 就是数值列表
            nums_cols = [(h2, [float(v) for v in data[h2] if isinstance(v, (int, float))]) for h2 in headers if h2 != h]
            for h2, nums2 in nums_cols:
                if len(nums_self) == len(nums2) and len(nums_self) > 2:
                    # 简单相关系数
                    mean1 = sum(nums_self) / len(nums_self)
                    mean2 = sum(nums2) / len(nums2)
                    cov = sum((a - mean1) * (b - mean2) for a, b in zip(nums_self, nums2)) / len(nums_self)
                    std1 = math.sqrt(sum((a - mean1) ** 2 for a in nums_self) / len(nums_self))
                    std2 = math.sqrt(sum((b - mean2) ** 2 for b in nums2) / len(nums2))
                    if std1 > 0 and std2 > 0:
                        corr = cov / (std1 * std2)
                        if abs(corr) > 0.8:
                            direction = "正" if corr > 0 else "负"
                            insights.append(f"🔗 「{h}」与「{h2}」强{direction}相关（r={corr:.2f}）")

        if not insights:
            insights = ["✅ 数据未发现明显异常，可进行可视化分析"]
        return insights


class DashboardBuilder:
    """仪表板构建器"""

    @staticmethod
    def create(data: Dict, output: str = "仪表板.xlsx") -> str:
        """一键生成专业仪表板"""
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "仪表板"

        headers = list(data.keys())
        nr = max(len(v) for v in data.values())
        nc = len(headers)

        # ==================== 标题 ====================
        ws_main.merge_cells(f"A1:{get_column_letter(nc + 4)}1")
        c = ws_main["A1"]
        c.value = f"📊 数据仪表板  {datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
        ap(c, ST["title"])
        c.fill = fill(C["gray_l"])
        ws_main.row_dimensions[1].height = 55

        # ==================== KPI 卡片行 ====================
        kpi_headers = [h for h in headers if any(kw in h for kw in ["金额", "销售", "利润", "成本", "收入", "预算", "实际", "率", "数"])]
        kpi_vals = []
        for h in kpi_headers:
            vals = [float(v) for v in data[h] if isinstance(v, (int, float))]
            if vals:
                kpi_vals.append((h, sum(vals), max(vals), sum(vals)/len(vals)))

        kpi_per_row = min(4, len(kpi_vals))
        for i, (name, total, mx, avg) in enumerate(kpi_vals[:4]):
            col = i * 3 + 1
            ws_main.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
            lc = ws_main.cell(row=3, column=col, value=name)
            lc.font = Font(size=9, color=C["text_m"], name="微软雅黑")
            lc.alignment = aln("center")

            ws_main.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 2)
            vc = ws_main.cell(row=4, column=col, value=total)
            vc.font = Font(bold=True, size=22, color=C["primary"], name="微软雅黑")
            vc.number_format = "¥#,##0"
            vc.alignment = aln("center")

            ws_main.cell(row=5, column=col, value=f"均值 {avg:,.0f}  |  最大 {mx:,.0f}").font = Font(size=8, color=C["text_l"], name="微软雅黑")
            ws_main.cell(row=5, column=col).alignment = aln("center")
            ws_main.row_dimensions[6].height = 10

        # ==================== 数据表格 ====================
        data_row = 8
        for col, h in enumerate(headers, 1):
            cell = ws_main.cell(row=data_row, column=col, value=h)
            ap(cell, ST["h"])

        for r in range(nr):
            bg = C["gray_l"] if r % 2 == 0 else C["white"]
            for col, h in enumerate(headers, 1):
                val = data[h][r] if r < len(data[h]) else ""
                cell = ws_main.cell(row=data_row + r + 1, column=col, value=val)
                cell.fill = fill(bg)
                ap(cell, {"sz": 10, "ha": "center"})
                if any(kw in h for kw in ["金额", "销售", "利润", "成本", "收入"]):
                    cell.number_format = "¥#,##0.00"
                    cell.alignment = aln("right")
                elif any(kw in h for kw in ["率", "占比", "比例"]):
                    cell.number_format = "0.00%"
                    cell.alignment = aln("right")

        # ==================== 图表区 ====================
        chart_area_col = nc + 2

        # 柱状图
        if nr > 0 and len(headers) > 1:
            chart1 = BarChart(type="col", style=10)
            chart1.title = "📊 数据对比"
            chart1.width = 16
            chart1.height = 10
            chart1.title_font = Font(bold=True, size=13, color=C["text_d"], name="微软雅黑")
            dr = Reference(ws_main, min_col=2, min_row=data_row, max_col=min(nc, 3), max_row=data_row + nr)
            cr = Reference(ws_main, min_col=1, min_row=data_row + 1, max_row=data_row + nr)
            chart1.add_data(dr, titles_from_data=True)
            chart1.set_categories(cr)
            for i, s in enumerate(chart1.series):
                s.graphicalProperties.solidFill = CHART[i % len(CHART)]
            ws_main.add_chart(chart1, f"{get_column_letter(chart_area_col)}3")

        # 饼图
        if nr > 0:
            pie = PieChart()
            pie.title = "🥧 占比分布"
            pie.width = 14
            pie.height = 10
            pie.title_font = Font(bold=True, size=13, color=C["text_d"], name="微软雅黑")
            pdr = Reference(ws_main, min_col=2, min_row=data_row + 1, max_row=data_row + min(nr, 7))
            pcr = Reference(ws_main, min_col=1, min_row=data_row + 1, max_row=data_row + min(nr, 7))
            pie.add_data(pdr)
            pie.set_categories(pcr)
            for i, pt in enumerate(pie.series[0].data_points):
                pt.graphicalProperties.solidFill = CHART[i % len(CHART)]
            ws_main.add_chart(pie, f"{get_column_letter(chart_area_col)}17")

        # 折线图
        if nr > 2:
            line = LineChart()
            line.title = "📈 趋势"
            line.width = 16
            line.height = 10
            line.title_font = Font(bold=True, size=13, color=C["text_d"], name="微软雅黑")
            lr = Reference(ws_main, min_col=2, min_row=data_row, max_col=min(nc, 3), max_row=data_row + nr)
            lcr = Reference(ws_main, min_col=1, min_row=data_row + 1, max_row=data_row + nr)
            line.add_data(lr, titles_from_data=True)
            line.set_categories(lcr)
            for i, s in enumerate(line.series):
                s.graphicalProperties.line.solidFill = CHART[i % len(CHART)]
                s.smooth = True
                s.dLbls = DataLabelList()
                s.dLbls.showVal = False
            ws_main.add_chart(line, f"{get_column_letter(chart_area_col)}31")

        # ==================== 数据洞察 ====================
        insights_row = data_row + nr + 3
        ai = AIGraph()
        insights = ai.generate_insights(data)

        ws_main.cell(row=insights_row, column=1, value="💡 数据洞察").font = Font(bold=True, size=13, color=C["primary"], name="微软雅黑")
        for i, ins in enumerate(insights[:5]):
            ws_main.cell(row=insights_row + i + 1, column=1, value=ins).font = Font(size=10, color=C["text_m"], name="微软雅黑")

        # 列宽
        ws_main.column_dimensions["A"].width = 18
        for col in range(2, nc + 1):
            ws_main.column_dimensions[get_column_letter(col)].width = 14

        # ==================== 数据表 ====================
        ws_data = wb.create_sheet("原始数据")
        for col, h in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=h)
            ap(cell, ST["h"])
        for r in range(nr):
            bg = C["gray_l"] if r % 2 == 0 else C["white"]
            for col, h in enumerate(headers, 1):
                val = data[h][r] if r < len(data[h]) else ""
                cell = ws_data.cell(row=r + 2, column=col, value=val)
                cell.fill = fill(bg)
                ap(cell, {"sz": 10})
        ws_data.column_dimensions["A"].width = 20
        for col in range(2, nc + 1):
            ws_data.column_dimensions[get_column_letter(col)].width = 14

        wb.save(output)
        return output


class ColumnAnalyzer:
    """列类型自动检测"""

    TYPES = ["number", "money", "percent", "date", "text"]

    @staticmethod
    def detect(data: Dict) -> Dict[str, str]:
        """自动检测每列的类型"""
        result = {}
        for h, vals in data.items():
            clean_vals = [v for v in vals if v is not None and v != ""]
            if not clean_vals:
                result[h] = "text"
                continue

            # 数值检测
            nums = [float(v) for v in clean_vals if isinstance(v, (int, float))]
            text_nums = [float(v) for v in clean_vals
                        if isinstance(v, str) and v.replace(",", "").replace(".", "").replace("-", "").isdigit()]

            total_numeric = len(nums) + len(text_nums)

            if total_numeric >= len(clean_vals) * 0.8:
                # 金额检测（常见金额关键词）
                if any(k in str(h) for k in ["金额", "价格", "销售额", "利润", "成本", "收入", "预算"]):
                    result[h] = "money"
                # 百分比检测
                elif any(k in str(h) for k in ["率", "占比", "比例", "%", "rate"]):
                    result[h] = "percent"
                else:
                    result[h] = "number"
            # 日期检测
            elif any(ColumnAnalyzer._is_date(str(v)) for v in clean_vals[:min(3, len(clean_vals))]):
                result[h] = "date"
            else:
                result[h] = "text"

        return result

    @staticmethod
    def _is_date(s: str) -> bool:
        patterns = [r"\d{4}-\d{2}-\d{2}", r"\d{4}/\d{2}/\d{2}", r"\d{2}[-/]\d{2}[-/]\d{4}"]
        return any(re.match(p, str(s)) for p in patterns)

    @staticmethod
    def apply_format(ws, col: int, col_type: str):
        """根据类型自动格式化列"""
        for r in range(ws.min_row + 1, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            if col_type == "money":
                cell.number_format = "¥#,##0.00"
                cell.alignment = aln("right")
            elif col_type == "percent":
                if isinstance(cell.value, str) and "%" in cell.value:
                    try:
                        cell.value = float(cell.value.replace("%", "")) / 100
                    except:
                        pass
                cell.number_format = "0.00%"
                cell.alignment = aln("right")
            elif col_type == "date":
                cell.number_format = "YYYY-MM-DD"
                cell.alignment = aln("center")


class FormulaBuilder:
    """公式构建器"""

    FORMULAS = {
        "毛利率": lambda c: f'=IF({c[1] if len(c)>1 else "A"}=0,0,{c[0]}/{c[1] if len(c)>1 else "A"})',
        "净利率": lambda c: f'=IF({c[1] if len(c)>1 else "A"}=0,0,{c[0]}/{c[1] if len(c)>1 else "A"})',
        "同比": lambda c: f'=IF({c[1] if len(c)>1 else "A"}=0,0,({c[0]}-{c[1] if len(c)>1 else "A"})/{c[1] if len(c)>1 else "A"})',
        "环比": lambda c: f'=IF({c[1] if len(c)>1 else "A"}=0,0,({c[0]}-{c[1] if len(c)>1 else "A"})/{c[1] if len(c)>1 else "A"})',
        "求和": lambda c: f'=SUM({c[0]}:{c[0]})',
        "平均": lambda c: f'=AVERAGE({c[0]}:{c[0]})',
        "最大": lambda c: f'=MAX({c[0]}:{c[0]})',
        "最小": lambda c: f'=MIN({c[0]}:{c[0]})',
        "计数": lambda c: f'=COUNTA({c[0]}:{c[0]})',
        "排名": lambda c: f'=RANK({c[0]}2,{c[0]}:{c[0]},0)',
        "条件求和": lambda c: f'=SUMIF({c[0]}:{c[0]},">0",{c[1] if len(c)>1 else c[0]}:{c[1] if len(c)>1 else c[0]})',
        "条件计数": lambda c: f'=COUNTIF({c[0]}:{c[0]},">0")',
        "VLOOKUP": lambda c: f'=VLOOKUP(A2,{c[0]}:{c[-1]},{len(c)},0)' if len(c) > 1 else '=VLOOKUP(...)',
        "INDEX_MATCH": lambda c: f'=INDEX({c[-1]}:{c[-1]},MATCH(A2,{c[0]}:{c[0]},0))' if len(c) > 1 else '=INDEX(MATCH(...))',
        "IF条件": lambda c: f'=IF({c[0]}>0,"达标","未达标")',
        "IF大于": lambda c: f'=IF({c[0]}>100,"高","低")',
        "IF排名": lambda c: f'=IF(RANK({c[0]}2,{c[0]}:{c[0]},0)<=3,"Top3","其他")',
        "LAMBDA": lambda c: f'=LAMBDA(x,{c[0]}*2)(A2)' if c else '=LAMBDA(x,x*2)(A2)',
        "LET": lambda c: f'=LET(x,{c[0]}:{c[0]},AVERAGE(x))' if c else '=LET(x,range,AVERAGE(x))',
    }

    @staticmethod
    def build(description: str, columns: List[str] = None) -> Dict:
        """根据描述构建公式"""
        desc = description.lower()
        cols = columns or ["A"]
        results = []

        for kw, fn in FormulaBuilder.FORMULAS.items():
            if kw.lower() in desc:
                formula = fn(cols)
                results.append({
                    "formula": formula,
                    "type": "Excel",
                    "description": f"「{kw}」公式",
                    "note": "复制到 Excel 单元格中使用"
                })

        if not results:
            # 通用搜索
            import re as _re
            m = _re.search(r'(?:大于|小于|等于|>=|<=|>|<)([\d.]+)', desc)
            if m:
                op_map = {">": ">", "<": "<", "=": "=", ">=": ">=", "<=": "<="}
                for op, excel_op in op_map.items():
                    if op in desc:
                        formula = f'=COUNTIF({cols[0]}:{cols[0]},"{excel_op}{m.group(1)}")'
                        results.append({"formula": formula, "type": "Excel", "description": f"条件计数（{excel_op}{m.group(1)}）", "note": ""})
                        break

        if not results:
            results.append({"formula": "=SUM(A:A)", "type": "Excel", "description": "求和示例", "note": "请更详细描述需求"})

        return {"status": "success", "input": description, "columns": columns, "formulas": results}


# ============================================================
# 主程序 v3 — 整合所有新命令
# ============================================================
def main_v3():
    import argparse as _ap
    p = _ap.ArgumentParser(description="PansExcel Pro Max v3", formatter_class=_ap.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="cmd")

    def add(name, **kw): return sub.add_parser(name, **kw)

    # 所有命令...
    sp = add("create"); sp.add_argument("-d","--data",required=True); sp.add_argument("-o","--output",default="out.xlsx"); sp.add_argument("-s","--sheet",default="数据")
    sp = add("report"); sp.add_argument("-d","--data",required=True); sp.add_argument("-t","--template",required=True,
        choices=["sales","financial","gantt","recon","daily","weekly","sales-dashboard"])
    sp.add_argument("-o","--output",default="report.xlsx")

    sp = add("analyze"); sp.add_argument("-f","--file",required=True); sp.add_argument("-o","--output")
    sp = add("formula"); sp.add_argument("--描述",required=True); sp.add_argument("-c","--columns")
    sp = add("chart"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-t","--type",default="bar",choices=["bar","line","pie","area","radar","combo","scatter","waterfall","histogram","sparkline"])
    sp.add_argument("-p","--pos",default="F2"); sp.add_argument("-T","--title",default="")
    sp.add_argument("-c","--col",type=int,default=2); sp.add_argument("--bins",type=int,default=10)

    sp = add("format"); sp.add_argument("-f","--file",required=True); sp.add_argument("-o","--output")
    sp = add("cond"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-r","--range",default="B2:B100")
    sp.add_argument("-t","--type",default="scale",choices=["scale","bar","icon","top","avg","gt","lt"])
    sp.add_argument("-v","--val",type=float,default=0)

    sp = add("dv"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-c","--col",type=int,required=True)
    sp.add_argument("-t","--type",default="list",choices=["list","number","date","length"])
    sp.add_argument("-o","--options"); sp.add_argument("--min",type=float); sp.add_argument("--max",type=float)
    sp.add_argument("--min-len",type=int,default=0); sp.add_argument("--max-len",type=int,default=255)

    sp = add("clean"); sp.add_argument("-f","--file",required=True); sp.add_argument("-s","--steps",default="dedup,blanks,trim"); sp.add_argument("-o","--output")
    sp = add("split"); sp.add_argument("-f","--file",required=True); sp.add_argument("-c","--col",type=int,required=True); sp.add_argument("-d","--delimiter",default=","); sp.add_argument("-o","--output")
    sp = add("import"); sp.add_argument("-f","--file",required=True); sp.add_argument("-o","--output"); sp.add_argument("-t","--type",choices=["xlsx","csv","json"],default="xlsx")
    sp = add("export"); sp.add_argument("-f","--file",required=True); sp.add_argument("-o","--output"); sp.add_argument("-t","--type",choices=["csv","json","pdf"],default="csv")
    sp = add("pdf"); sp.add_argument("-f","--file",required=True); sp.add_argument("-o","--output")
    sp = add("merge"); sp.add_argument("-f","--files",nargs="+",required=True); sp.add_argument("-o","--output",default="merged.xlsx"); sp.add_argument("-k","--key",type=int,default=1)
    sp = add("insights"); sp.add_argument("-f","--file")

    # 新命令
    sp = add("dashboard"); sp.add_argument("-d","--data",required=True); sp.add_argument("-o","--output",default="仪表板.xlsx")
    sp = add("pivot"); sp.add_argument("-f","--file",required=True)
    sp.add_argument("-r","--row",type=int,default=1); sp.add_argument("-c","--col",type=int); sp.add_argument("-v","--val",type=int,default=2)
    sp.add_argument("-a","--agg",default="sum",choices=["sum","avg","count","max","min"]); sp.add_argument("-o","--output",default="透视表.xlsx")

    sp = add("autofmt"); sp.add_argument("-f","--file",required=True)
    sp = add("describe"); sp.add_argument("-f","--file",required=True)

    args = p.parse_args()

    if not HAS_OPENPYXL:
        print(json.dumps({"status":"error","message":"pip install openpyxl"})); return

    def ld(s):
        try: return json.loads(s)
        except:
            try: return eval(s)
            except: return {}

    def gc(s):
        return [c.strip() for c in s.split(",")] if s else []

    x = X()
    ai = AIGraph()

    try:
        if args.cmd == "create":
            path = X().from_dict(ld(args.data), args.sheet).save(args.output)
            print(json.dumps({"status":"success","file":path}))

        elif args.cmd == "report":
            d = ld(args.data)
            if args.template == "sales": path = R.sales_dashboard(d, args.output)
            elif args.template == "sales-dashboard": path = R.sales_dashboard(d, args.output)
            elif args.template == "financial": path = R.financial(d, args.output)
            elif args.template == "gantt": path = R.gantt(d, args.output)
            elif args.template == "recon": path = R.reconciliation(d, args.output)
            elif args.template == "daily": path = R.daily_report(d, args.output)
            elif args.template == "weekly": path = R.weekly_report(d, args.output)
            else: path = args.output
            print(json.dumps({"status":"success","file":path}))

        elif args.cmd == "analyze":
            x.open(args.file)
            data = x.read_data()
            col_types = ColumnAnalyzer.detect(data)
            result = {
                "status": "success",
                "description": ai.describe(data),
                "column_types": col_types,
                "chart_suggestion": ai.suggest(data),
                "insights": AIGraph.generate_insights(data)
            }
            if args.output:
                with open(args.output, "w") as f:
                    json.dump(result, f, ensure_ascii=False, indent=2)
            print(json.dumps(result, ensure_ascii=False, indent=2))

        elif args.cmd == "describe":
            x.open(args.file)
            data = x.read_data()
            print(json.dumps({"status":"success","description":ai.describe(data)}, ensure_ascii=False, indent=2))

        elif args.cmd == "formula":
            cols = gc(args.columns)
            result = FormulaBuilder.build(args.描述, cols)
            print(json.dumps(result, ensure_ascii=False, indent=2))

        elif args.cmd == "chart":
            x.open(args.file)
            t = args.type
            if t == "bar": x.add_bar(pos=args.pos, title=args.title)
            elif t == "line": x.add_line(pos=args.pos, title=args.title)
            elif t == "pie": x.add_pie(pos=args.pos, title=args.title)
            elif t == "area": x.add_area(pos=args.pos, title=args.title)
            elif t == "radar": x.add_radar(pos=args.pos, title=args.title)
            elif t == "combo": x.add_combo(pos=args.pos, title=args.title)
            elif t == "waterfall":
                ChartAdvanced.waterfall(x.ws, pos=args.pos, title=args.title, data_col=args.col)
            elif t == "histogram":
                ChartAdvanced.histogram(x.ws, data_col=args.col, bins=args.bins, pos=args.pos, title=args.title)
            elif t == "sparkline":
                last_r = x.ws.max_row
                ChartAdvanced.add_sparkline(x.ws, f"{get_column_letter(args.col)}{last_r + 2}",
                                          f"{get_column_letter(args.col)}2:{get_column_letter(args.col)}{last_r}")
            else:
                x.add_bar(pos=args.pos, title=args.title)
            x.save(args.file)
            print(json.dumps({"status":"success","message":f"{t}图表已添加"}))

        elif args.cmd == "format":
            x.open(args.file); x.beautify(); x.add_title("数据报表"); x.autofilter(); x.freeze()
            x.save(args.output or args.file)
            print(json.dumps({"status":"success","file":x.save(args.output or args.file)}))

        elif args.cmd == "cond":
            x.open(args.file)
            if args.type == "scale": x.cf_scale(args.range)
            elif args.type == "bar": x.cf_bar(args.range)
            elif args.type == "icon": x.cf_icon(args.range)
            elif args.type == "top": x.cf_top(args.range)
            elif args.type == "avg": x.cf_avg(args.range)
            elif args.type == "gt": x.cf_gt(args.range)
            elif args.type == "lt": x.cf_lt(args.range)
            x.save(args.file)
            print(json.dumps({"status":"success","message":"条件格式已应用"}))

        elif args.cmd == "dv":
            x.open(args.file)
            if args.type == "list":
                opts = [c.strip() for c in (args.options or "").split(",") if c.strip()]
                x.dv_list(args.col, opts)
            elif args.type == "number": x.dv_number(args.col, args.min, args.max)
            elif args.type == "date": x.dv_date(args.col)
            elif args.type == "length": x.dv_length(args.col, args.min_len, args.max_len)
            x.save(args.file)
            print(json.dumps({"status":"success","message":"数据验证已添加"}))

        elif args.cmd == "clean":
            x.open(args.file); x.deduplicate(); x.fill_blanks(); x.clean_pipeline([s.strip() for s in args.steps.split(",")])
            x.save(args.output or args.file)
            print(json.dumps({"status":"success","file":x.save(args.output or args.file)}))

        elif args.cmd == "split":
            x.open(args.file); new_cols = x.split_col(args.col, args.delimiter)
            x.save(args.output or args.file)
            print(json.dumps({"status":"success","new_columns":new_cols}))

        elif args.cmd == "import":
            ext = os.path.splitext(args.file)[1].lower()
            if ext == ".csv":
                d = IO.read_csv(args.file); out = args.output or args.file.replace(".csv",".xlsx")
                X().from_dict(d).save(out)
            elif ext == ".json":
                d = IO.read_json(args.file); out = args.output or args.file.replace(".json",".xlsx")
                X().from_dict(d).save(out)
            else:
                print(json.dumps({"status":"error","message":"仅支持 .csv 和 .json"})); return
            print(json.dumps({"status":"success","file":out}))

        elif args.cmd == "export":
            if args.type == "csv":
                x.open(args.file); d = x.read_data(); out = args.output or args.file.replace(".xlsx",".csv")
                IO.write_csv(d, out)
            elif args.type == "json":
                x.open(args.file); d = x.read_data(); out = args.output or args.file.replace(".xlsx",".json")
                IO.write_json(d, out)
            elif args.type == "pdf":
                print(json.dumps(IO.export_pdf(args.file, args.output))); return
            print(json.dumps({"status":"success","file":out}))

        elif args.cmd == "pdf":
            print(json.dumps(IO.export_pdf(args.file, args.output)))

        elif args.cmd == "merge":
            path = x.merge_files(args.files, key_col=args.key)
            print(json.dumps({"status":"success","file":path}))

        elif args.cmd == "insights":
            x.open(args.file); d = x.read_data()
            print(json.dumps({"status":"success","insights":AIGraph.generate_insights(d)}, ensure_ascii=False, indent=2))

        elif args.cmd == "dashboard":
            d = ld(args.data)
            path = DashboardBuilder.create(d, args.output)
            print(json.dumps({"status":"success","file":path}))

        elif args.cmd == "pivot":
            x.open(args.file)
            wb_out = PivotTable.create(x.ws, row_field=args.row, col_field=args.col, val_field=args.val, agg=args.agg)
            wb_out.save(args.output)
            print(json.dumps({"status":"success","file":args.output}))

        elif args.cmd == "autofmt":
            x.open(args.file)
            data = x.read_data()
            col_types = ColumnAnalyzer.detect(data)
            # 反向映射列名→列号
            name_to_col = {}
            for row in x.ws.iter_rows(min_row=x.ws.min_row, max_row=x.ws.min_row, values_only=True):
                for col_idx, name in enumerate(row, 1):
                    name_to_col[str(name)] = col_idx
            # 应用格式
            for h, t in col_types.items():
                if h in name_to_col:
                    ColumnAnalyzer.apply_format(x.ws, name_to_col[h], t)
            x.beautify()
            x.save(args.file)
            print(json.dumps({"status":"success","column_types":col_types}))

        else:
            print("""
PansExcel Pro Max v3

  dashboard  python3 excel.py dashboard -d '{...}' -o out.xlsx
  pivot     python3 excel.py pivot -f data.xlsx -r 1 -v 2 -a sum -o out.xlsx
  autofmt   python3 excel.py autofmt -f data.xlsx   # 自动检测并格式化列类型
  chart     python3 excel.py chart -f data.xlsx -t waterfall -p H2  # 瀑布图
  chart     python3 excel.py chart -f data.xlsx -t histogram -c 2 --bins 10  # 直方图
  analyze   python3 excel.py analyze -f data.xlsx -o result.json
""")
    except Exception as e:
        print(json.dumps({"status":"error","message":str(e)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main_v3()
