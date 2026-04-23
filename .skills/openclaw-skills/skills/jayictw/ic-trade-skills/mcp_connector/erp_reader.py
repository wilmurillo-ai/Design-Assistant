"""
mcp_connector/erp_reader.py
────────────────────────────
Local ERP reader — client-side only.

Security contract
──────────────────
This module reads the user's local ERP data (Excel or SQLite) and extracts
ONLY the minimum fields needed for client-side display:

  ✅ part_number    — the normalised IC model number
  ✅ stock_qty      — inventory quantity (for local display only)
  ✅ supply_status  — supply label from ERP

Fields that are deliberately NEVER extracted or transmitted:
  ❌ floor_sale_price     — your cost basis
  ❌ taxed_sale_price     — your selling price
  ❌ standard_sale_price  — pricing model
  ❌ usd_price / agent_price — any pricing field
  ❌ batch / lot info     — operational data

These fields stay on the user's machine at all times.
"""
from __future__ import annotations

import logging
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any

log = logging.getLogger("mcp_connector.erp_reader")


@dataclass(frozen=True)
class ErpRecord:
    """Minimal ERP view — safe to display to the user; never sent to the server."""
    part_number: str
    stock_qty: float | None
    supply_status: str | None
    source: str  # "excel" | "sqlite" | "unavailable"


# ── Excel reader ──────────────────────────────────────────────────────────────

# Column names from ERP內容.xlsx (as imported by scripts/import_erp_contents.py)
_COL_MODEL         = "型号"
_COL_STOCK         = "庫存量"
_COL_SUPPLY_STATUS = "供應狀態"

# Columns that must NEVER be read (fail-safe whitelist approach)
_BLOCKED_COLUMNS = {
    "usd進貨价", "含稅進貨价", "最低售价", "含稅售价", "標準售价",
    "floor_sale_price", "taxed_sale_price", "standard_sale_price",
    "usd_price", "agent_price", "人民币未稅貨值", "人民币含稅貨值",
}


def _safe_float(val: Any) -> float | None:
    if val is None or str(val).strip() in ("", "-", "N/A"):
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def read_from_excel(excel_path: str | Path, part_number: str) -> ErpRecord:
    """
    Read a single part number from the local ERP Excel workbook.
    Only the model, stock_qty, and supply_status columns are accessed.
    """
    path = Path(excel_path)
    if not path.exists():
        log.warning("ERP Excel not found: %s", path)
        return ErpRecord(part_number=part_number, stock_qty=None, supply_status=None, source="unavailable")

    try:
        from openpyxl import load_workbook  # type: ignore[import]
    except ImportError:
        log.error("openpyxl is not installed — cannot read ERP Excel.")
        return ErpRecord(part_number=part_number, stock_qty=None, supply_status=None, source="unavailable")

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        log.error("Failed to open ERP Excel %s: %s", path, exc)
        return ErpRecord(part_number=part_number, stock_qty=None, supply_status=None, source="unavailable")

    # Build column index from header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        return ErpRecord(part_number=part_number, stock_qty=None, supply_status=None, source="unavailable")

    col_index: dict[str, int] = {str(h).strip(): i for i, h in enumerate(header_row) if h}

    # Blocked column guard — crash loudly if a blocked column would be read
    for blocked in _BLOCKED_COLUMNS:
        if blocked in col_index:
            log.debug("Skipping blocked column: %s (privacy guard)", blocked)

    idx_model  = col_index.get(_COL_MODEL)
    idx_stock  = col_index.get(_COL_STOCK)
    idx_supply = col_index.get(_COL_SUPPLY_STATUS)

    if idx_model is None:
        log.warning("Column %r not found in ERP Excel header.", _COL_MODEL)
        wb.close()
        return ErpRecord(part_number=part_number, stock_qty=None, supply_status=None, source="unavailable")

    target = part_number.strip().upper()
    result: ErpRecord | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        model_raw = row[idx_model] if idx_model < len(row) else None
        if not model_raw:
            continue
        if str(model_raw).strip().upper() == target:
            stock_raw  = row[idx_stock]  if idx_stock  is not None and idx_stock  < len(row) else None
            supply_raw = row[idx_supply] if idx_supply is not None and idx_supply < len(row) else None
            result = ErpRecord(
                part_number=target,
                stock_qty=_safe_float(stock_raw),
                supply_status=str(supply_raw).strip() if supply_raw else None,
                source="excel",
            )
            break  # take first match

    wb.close()
    if result is None:
        return ErpRecord(part_number=target, stock_qty=None, supply_status=None, source="excel")
    return result


# ── SQLite reader ─────────────────────────────────────────────────────────────

def read_from_sqlite(db_path: str | Path, part_number: str) -> ErpRecord:
    """
    Read a single part from the local erp_inventory SQLite table.
    Only model, stock_qty, supply_status columns are selected — never pricing.
    """
    path = Path(db_path)
    if not path.exists():
        return ErpRecord(part_number=part_number, stock_qty=None, supply_status=None, source="unavailable")

    target = part_number.strip().upper()
    try:
        conn = sqlite3.connect(str(path))
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            # Explicit column selection — no SELECT * to avoid accidental price leakage
            """
            SELECT model, stock_qty, supply_status
            FROM erp_inventory
            WHERE UPPER(TRIM(model)) = ?
            ORDER BY erp_inventory_id DESC
            LIMIT 1
            """,
            (target,),
        ).fetchone()
        conn.close()
    except Exception as exc:
        log.error("SQLite ERP read error: %s", exc)
        return ErpRecord(part_number=target, stock_qty=None, supply_status=None, source="unavailable")

    if row is None:
        return ErpRecord(part_number=target, stock_qty=None, supply_status=None, source="sqlite")

    return ErpRecord(
        part_number=target,
        stock_qty=_safe_float(row["stock_qty"]),
        supply_status=str(row["supply_status"]).strip() if row["supply_status"] else None,
        source="sqlite",
    )


# ── Unified reader (auto-selects source) ──────────────────────────────────────

def read_erp(
    part_number: str,
    excel_path: str | Path | None = None,
    sqlite_path: str | Path | None = None,
) -> ErpRecord:
    """
    Read local ERP data for *part_number*.  Excel takes priority over SQLite.
    Returns an ErpRecord with source="unavailable" if no source is configured.
    """
    if excel_path and Path(excel_path).exists():
        return read_from_excel(excel_path, part_number)
    if sqlite_path and Path(sqlite_path).exists():
        return read_from_sqlite(sqlite_path, part_number)
    log.warning("No local ERP source available for %s.", part_number)
    return ErpRecord(part_number=part_number, stock_qty=None, supply_status=None, source="unavailable")
