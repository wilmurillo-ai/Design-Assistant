"""
mcp_connector/erp_reader_generic.py
─────────────────────────────────────
Generic ERP Reader — reads ANY standard Excel or CSV inventory file.

Design goals
─────────────
1. Works with any column layout — user declares which columns map to what.
2. Extracts ONLY the minimum safe fields: part_number + stock_qty + status.
3. Pricing columns are NEVER read, even if present (whitelist approach).
4. No assumption about file format — auto-detects .xlsx / .xls / .csv.
5. Zero company-specific logic — suitable for any IC trading operation.

Privacy contract
─────────────────
This reader will NEVER extract: purchase price, sale price, cost, margin,
floor price, or any monetary value. These stay on the user's machine.
Only part number and quantity are eligible for the combined view merger.

Usage
──────
    from mcp_connector.erp_reader_generic import GenericErpReader, ColumnMap

    # Declare which Excel columns hold what data
    col_map = ColumnMap(
        part_number="Part Number",   # column header name in your Excel
        stock_qty="QTY Available",
        supply_status="Status",      # optional
    )

    reader = GenericErpReader(col_map)
    record = reader.lookup("STM32L412CBU6", "path/to/inventory.xlsx")
    print(record.stock_qty)   # e.g. 5000.0
    print(record.supply_status)  # e.g. "In Stock"
"""
from __future__ import annotations

import csv
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

log = logging.getLogger("mcp_connector.erp_reader_generic")

# ── Column map — user-configurable ────────────────────────────────────────────

@dataclass
class ColumnMap:
    """
    Maps your Excel column headers to the standard field names this reader uses.
    All fields except `part_number` are optional.

    Examples for common ERP / BOM formats:
        SAP export:      part_number="Material", stock_qty="Unrestricted"
        Standard BOM:    part_number="MPN",       stock_qty="Stock Qty"
        Simple CSV:      part_number="PN",         stock_qty="QTY"
    """
    part_number: str   = "Part Number"    # REQUIRED — header name in your file
    stock_qty: str     = "Stock Qty"      # header for available quantity
    supply_status: str = "Status"         # header for supply / lifecycle status
    # Add more mappings here as needed — they are display-only, never priced.
    package: str       = "Package"        # optional: component package
    date_code: str     = "Date Code"      # optional: manufacturing date code

    # Safety: columns that must NEVER be read regardless of mapping
    _PRICE_KEYWORDS: list[str] = field(default_factory=lambda: [
        "price", "cost", "floor", "margin", "sale", "purchase",
        "unit price", "买入", "卖出", "底价", "售价", "进价", "含税",
        "taxed", "tax", "usd价", "cny价", "单价", "成本",
    ])

    def is_price_column(self, header: str) -> bool:
        """Return True if this header looks like a pricing column."""
        low = header.lower().strip()
        return any(kw in low for kw in self._PRICE_KEYWORDS)


# ── Result type ───────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class ErpLookupResult:
    """
    Minimal ERP record — safe to display and merge with remote API data.
    No pricing fields. Only quantity and status signals.
    """
    part_number: str
    stock_qty: float | None
    supply_status: str | None
    package: str | None
    date_code: str | None
    source_file: str
    found: bool

    @property
    def has_stock(self) -> bool:
        return bool(self.stock_qty and self.stock_qty > 0)

    def as_dict(self) -> dict[str, Any]:
        return {
            "part_number":   self.part_number,
            "stock_qty":     self.stock_qty,
            "has_stock":     self.has_stock,
            "supply_status": self.supply_status,
            "package":       self.package,
            "date_code":     self.date_code,
            "source_file":   self.source_file,
        }


# ── Generic reader ────────────────────────────────────────────────────────────

class GenericErpReader:
    """
    Reads any standard Excel (.xlsx/.xls) or CSV file to look up a part number.

    The reader is intentionally column-agnostic: you tell it which column
    headers to use via ColumnMap, and it does the rest.

    Thread safety: instances are stateless between calls; safe to reuse.
    """

    def __init__(self, col_map: ColumnMap | None = None):
        self.col_map = col_map or ColumnMap()

    def lookup(
        self,
        part_number: str,
        file_path: str | Path,
        sheet_name: str | int = 0,
    ) -> ErpLookupResult:
        """
        Look up *part_number* in *file_path*.

        Args:
            part_number: IC model number to search for (case-insensitive).
            file_path:   Path to .xlsx, .xls, or .csv file.
            sheet_name:  Sheet name or 0-indexed sheet number (Excel only).

        Returns:
            ErpLookupResult with found=False if part not found or file missing.
        """
        path = Path(file_path)
        target = part_number.strip().upper()

        if not path.exists():
            log.warning("ERP file not found: %s", path)
            return self._not_found(target, str(path))

        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            return self._read_excel(target, path, sheet_name)
        if suffix == ".csv":
            return self._read_csv(target, path)

        log.warning("Unsupported file type: %s (expected .xlsx, .xls, .csv)", suffix)
        return self._not_found(target, str(path))

    def _not_found(self, part_number: str, source: str) -> ErpLookupResult:
        return ErpLookupResult(
            part_number=part_number,
            stock_qty=None,
            supply_status=None,
            package=None,
            date_code=None,
            source_file=source,
            found=False,
        )

    # ── Excel ──────────────────────────────────────────────────────────────

    def _read_excel(
        self,
        target: str,
        path: Path,
        sheet_name: str | int,
    ) -> ErpLookupResult:
        try:
            from openpyxl import load_workbook  # type: ignore[import]
        except ImportError:
            log.error("openpyxl is required: pip install openpyxl")
            return self._not_found(target, str(path))

        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            log.error("Cannot open Excel %s: %s", path, exc)
            return self._not_found(target, str(path))

        # Select sheet
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name] if sheet_name < len(wb.worksheets) else wb.active
        else:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            wb.close()
            return self._not_found(target, str(path))

        idx = self._build_index(list(header_row))
        result = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            pn_idx = idx.get("part_number")
            if pn_idx is None or pn_idx >= len(row):
                continue
            cell_val = row[pn_idx]
            if cell_val and str(cell_val).strip().upper() == target:
                result = self._extract_row(row, idx, target, str(path))
                break

        wb.close()
        return result or self._not_found(target, str(path))

    # ── CSV ────────────────────────────────────────────────────────────────

    def _read_csv(self, target: str, path: Path) -> ErpLookupResult:
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return self._not_found(target, str(path))

                idx = self._build_index(list(reader.fieldnames))
                pn_col = reader.fieldnames[idx.get("part_number", -1)] if idx.get("part_number") is not None else None
                if not pn_col:
                    log.warning("Part number column %r not found in CSV headers.", self.col_map.part_number)
                    return self._not_found(target, str(path))

                for row_dict in reader:
                    if str(row_dict.get(pn_col, "")).strip().upper() == target:
                        # Convert dict row to positional list for _extract_row
                        fieldnames = list(reader.fieldnames)
                        row_list = [row_dict.get(h) for h in fieldnames]
                        return self._extract_row(row_list, idx, target, str(path))

        except Exception as exc:
            log.error("CSV read error %s: %s", path, exc)

        return self._not_found(target, str(path))

    # ── Helpers ────────────────────────────────────────────────────────────

    def _build_index(self, headers: list[Any]) -> dict[str, int]:
        """
        Map logical field names to column indices.
        Skips any column whose header matches a price keyword (safety guard).
        """
        index: dict[str, int] = {}
        cm = self.col_map

        for i, h in enumerate(headers):
            if not h:
                continue
            h_str = str(h).strip()

            # Safety guard — never index pricing columns
            if cm.is_price_column(h_str):
                log.debug("Skipping price column: %r", h_str)
                continue

            if h_str == cm.part_number:
                index["part_number"] = i
            elif h_str == cm.stock_qty:
                index["stock_qty"] = i
            elif h_str == cm.supply_status:
                index["supply_status"] = i
            elif h_str == cm.package:
                index["package"] = i
            elif h_str == cm.date_code:
                index["date_code"] = i

        return index

    def _extract_row(
        self,
        row: list[Any],
        idx: dict[str, int],
        part_number: str,
        source: str,
    ) -> ErpLookupResult:
        def get(key: str) -> Any:
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        return ErpLookupResult(
            part_number=part_number,
            stock_qty=self._to_float(get("stock_qty")),
            supply_status=self._to_str(get("supply_status")),
            package=self._to_str(get("package")),
            date_code=self._to_str(get("date_code")),
            source_file=source,
            found=True,
        )

    def read_all(
        self,
        file_path: str | Path,
        sheet_name: str | int = 0,
    ) -> list[ErpLookupResult]:
        """
        Read ALL rows from the file and return them as a list of ErpLookupResult.

        Used for batch quoting — iterates every part in the inventory file
        without requiring individual lookup calls.

        Args:
            file_path:   Path to .xlsx, .xls, or .csv file.
            sheet_name:  Sheet name or 0-indexed sheet number (Excel only).

        Returns:
            List of ErpLookupResult (one per data row).  Empty rows are skipped.
        """
        path = Path(file_path)
        if not path.exists():
            log.warning("ERP file not found: %s", path)
            return []

        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            return self._read_excel_all(path, sheet_name)
        if suffix == ".csv":
            return self._read_csv_all(path)

        log.warning("Unsupported file type: %s (expected .xlsx, .xls, .csv)", suffix)
        return []

    def _read_excel_all(self, path: Path, sheet_name: str | int) -> list[ErpLookupResult]:
        try:
            from openpyxl import load_workbook  # type: ignore[import]
        except ImportError:
            log.error("openpyxl is required: pip install openpyxl")
            return []

        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            log.error("Cannot open Excel %s: %s", path, exc)
            return []

        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name] if sheet_name < len(wb.worksheets) else wb.active
        else:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            wb.close()
            return []

        idx     = self._build_index(list(header_row))
        pn_idx  = idx.get("part_number")
        results = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if pn_idx is None or pn_idx >= len(row):
                continue
            pn_val = row[pn_idx]
            if not pn_val or not str(pn_val).strip():
                continue
            part = str(pn_val).strip().upper()
            results.append(self._extract_row(list(row), idx, part, str(path)))

        wb.close()
        return results

    def _read_csv_all(self, path: Path) -> list[ErpLookupResult]:
        results = []
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return []

                fieldnames = list(reader.fieldnames)
                idx    = self._build_index(fieldnames)
                pn_col = fieldnames[idx["part_number"]] if "part_number" in idx else None
                if not pn_col:
                    log.warning("Part number column %r not found.", self.col_map.part_number)
                    return []

                for row_dict in reader:
                    pn_val = str(row_dict.get(pn_col, "")).strip()
                    if not pn_val:
                        continue
                    part     = pn_val.upper()
                    row_list = [row_dict.get(h) for h in fieldnames]
                    results.append(self._extract_row(row_list, idx, part, str(path)))

        except Exception as exc:
            log.error("CSV read_all error %s: %s", path, exc)

        return results

    @staticmethod
    def _to_float(val: Any) -> float | None:
        if val is None or str(val).strip() in ("", "-", "N/A", "NA"):
            return None
        try:
            return float(str(val).replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_str(val: Any) -> str | None:
        if val is None or str(val).strip() in ("", "-"):
            return None
        return str(val).strip()


# ── Convenience function ──────────────────────────────────────────────────────

def lookup_part(
    part_number: str,
    file_path: str | Path,
    col_map: ColumnMap | None = None,
) -> ErpLookupResult:
    """
    One-shot convenience wrapper for single lookups.

    Example:
        result = lookup_part("STM32L412CBU6", "~/Desktop/inventory.xlsx")
        if result.has_stock:
            print(f"Local stock: {result.stock_qty:,.0f} units")
    """
    return GenericErpReader(col_map).lookup(part_number, file_path)


# ── Example column maps for common ERP formats ────────────────────────────────

EXAMPLE_COLUMN_MAPS = {
    # Generic BOM / inventory spreadsheet
    "generic": ColumnMap(
        part_number="Part Number",
        stock_qty="Stock Qty",
        supply_status="Status",
        package="Package",
        date_code="Date Code",
    ),
    # SAP MM report
    "sap_mm": ColumnMap(
        part_number="Material",
        stock_qty="Unrestricted",
        supply_status="MRP Type",
        package="Base Unit",
        date_code="",
    ),
    # Arrow / Avnet distributor export
    "distributor_generic": ColumnMap(
        part_number="MPN",
        stock_qty="QOH",
        supply_status="Lifecycle",
        package="Package",
        date_code="DC",
    ),
    # Simple CSV / manual BOM
    "simple_csv": ColumnMap(
        part_number="PN",
        stock_qty="QTY",
        supply_status="",
        package="PKG",
        date_code="",
    ),
}


# ── CLI for quick testing ─────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse, json, sys

    parser = argparse.ArgumentParser(description="Generic ERP Reader — test lookup")
    parser.add_argument("file", help="Path to Excel or CSV file")
    parser.add_argument("part_number", help="Part number to look up")
    parser.add_argument("--pn-col",     default="Part Number", help="Column header for part number")
    parser.add_argument("--qty-col",    default="Stock Qty",   help="Column header for stock quantity")
    parser.add_argument("--status-col", default="Status",      help="Column header for supply status")
    parser.add_argument("--sheet",      default=0,             help="Sheet name or index (Excel only)")
    args = parser.parse_args()

    cm = ColumnMap(
        part_number=args.pn_col,
        stock_qty=args.qty_col,
        supply_status=args.status_col,
    )
    sheet = args.sheet if args.sheet == 0 else (int(args.sheet) if args.sheet.isdigit() else args.sheet)
    result = GenericErpReader(cm).lookup(args.part_number, args.file, sheet_name=sheet)
    print(json.dumps(result.as_dict(), ensure_ascii=False, indent=2))
    sys.exit(0 if result.found else 1)
