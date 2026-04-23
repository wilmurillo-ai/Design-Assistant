"""
微信文章抓取核心脚本 v5
修复版：参考开源项目 wechat-article-exporter 的策略
- 保留完整 HTML 结构（<!DOCTYPE><html><head><body>）
- 下载 CSS/图片到 assets/，用时间戳文件名
- 替换所有资源 URL 为本地路径
- 移除 #js_content 的 visibility:hidden 等内联样式
- 添加 minimal CSS 保底样式
"""
import re, sys, os, json, time, hashlib, urllib.request, urllib.parse
from datetime import datetime
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/117.0.0.0 Safari/537.36 WAE/1.0"
)

VALID_FORMATS = ("markdown", "html", "text", "json", "excel")
_file_cache = {}
_ASSET_COUNTER = 0
_MAX_PARALLEL = 8

# ---------------------------------------------------------------------------
# 下载工具
# ---------------------------------------------------------------------------

def _robust_urlopen(url: str, headers: dict = None, timeout: int = 12) -> bytes:
    hdrs = dict(headers or {})
    hdrs["User-Agent"] = USER_AGENT
    hdrs["Referer"] = "https://mp.weixin.qq.com/"
    for attempt in range(2):
        try:
            req = urllib.request.Request(url, headers=hdrs)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                data = resp.read()
                return data[:2 * 1024 * 1024]
        except Exception:
            if attempt == 0:
                time.sleep(0.5)
    return b""


def _download(url: str, headers: dict = None) -> tuple[bytes, str]:
    global _file_cache
    if not url or url in _file_cache:
        return _file_cache.get(url, (b"", ""))
    content = _robust_urlopen(url, headers)
    mime = ""
    _file_cache[url] = (content, mime)
    return content, mime


def _parallel_download(urls: list) -> dict:
    results = {}
    with ThreadPoolExecutor(max_workers=_MAX_PARALLEL) as pool:
        futures = {pool.submit(_download, u): u for u in urls if u}
        for future in as_completed(futures, timeout=20):
            u = futures[future]
            try:
                results[u] = future.result()
            except Exception:
                results[u] = (b"", "")
    return results


# ---------------------------------------------------------------------------
# 资源文件名生成（时间戳格式，与参考文件一致）
# ---------------------------------------------------------------------------

def _detect_content_ext(content: bytes, url: str = "") -> str:
    """根据文件内容魔数判断真实类型，返回扩展名（含点）"""
    if not content:
        return ""
    # 图片魔数（UTF-8 decode-safe: only use byte comparisons）
    b0, b1 = content[0], content[1]
    if b0 == 0xff and b1 in (0xd8, 0xd9):          # JPEG (SOI or EOI)
        return ".jpg"
    if content[:8] == b"\x89PNG\r\n\x1a\n":    # PNG
        return ".png"
    if content[:6] in (b"GIF87a", b"GIF89a"):      # GIF
        return ".gif"
    if content[:4] == b"RIFF" and b"WEBP" in content[:30]:  # WebP
        return ".webp"
    if b0 == 0x42 and b1 == 0x4d:                   # BMP
        return ".bmp"
    # SVG
    if content[:4] == b"<svg" or content[:5] == b"<?xml":
        return ".svg"
    # CSS / JS（文本）：尝试 decode
    try:
        text = content[:300].decode("ascii", errors="ignore")
        if any(c in text for c in ("{}", "@media", "@charset", ".wx-", ".rich_", ".app-")):
            return ".css"
        if "function" in text and ("$" in text or "document" in text):
            return ".js"
    except Exception:
        pass
    # 从 URL 猜扩展名（兜底）
    path = urllib.parse.urlparse(url).path.lower()
    ext = os.path.splitext(path)[1].lower()
    if ext in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".css", ".js"):
        return ext
    return ""


# ---------------------------------------------------------------------------
# Minimal CSS（保底样式，确保正文可读）
# ---------------------------------------------------------------------------

MINIMAL_CSS = """
.rich_media_area_primary {
    max-width: 667px;
    margin: 0 auto;
    padding: 24px 24px 60px;
    background: #fff;
    min-height: 100vh;
}
.rich_media_content {
    font-family: -apple-system, BlinkMacSystemFont, 'Microsoft YaHei',
                 'Helvetica Neue', Helvetica, Arial, sans-serif;
    font-size: 17px;
    line-height: 1.9;
    color: #333;
    word-wrap: break-word;
}
.rich_media_content p {
    margin: 1em 0;
    min-height: 1.6em;
}
.rich_media_content h2,
.rich_media_content h3 {
    font-weight: 700;
    margin: 1.4em 0 0.7em;
    color: #1a1a1a;
}
.rich_media_content img {
    max-width: 100%;
    height: auto;
    display: block;
    margin: 1em auto;
    border-radius: 4px;
}
blockquote {
    border-left: 4px solid #ccc;
    padding: 8px 16px;
    margin: 16px 0;
    background: #f9f9f9;
    border-radius: 0 4px 4px 0;
}
pre {
    background: #f5f5f5;
    padding: 16px;
    border-radius: 8px;
    overflow-x: auto;
    font-size: 14px;
}
.rich_media_meta {
    font-size: 14px;
    color: #999;
    margin: 8px 0;
}
#activity-name {
    font-size: 22px;
    font-weight: 700;
    color: #1a1a1a;
    margin-bottom: 12px;
    line-height: 1.4;
}
"""


# ---------------------------------------------------------------------------
# 核心：生成独立 HTML
# ---------------------------------------------------------------------------

def _build_standalone_html(raw_html: str, title: str, author: str,
                             date: str, url: str) -> tuple[str, list]:
    global _ASSET_COUNTER
    """
    参考开源项目 wechat-article-exporter 的 normalizeHtml 策略：
    1. 解析完整 HTML 文档（保留 <html><head><body>）
    2. 下载所有 CSS link、图片、背景图到 assets/
    3. 用本地时间戳文件名替换所有资源 URL
    4. 移除 #js_content 的 visibility:hidden / display:none 内联样式
    5. 删除无用标签（script, qr_code, ad 等）
    6. 构建完整的 <!DOCTYPE html><html> 文档
    返回: (完整HTML字符串, assets列表)
    """
    soup = BeautifulSoup(raw_html, "html5lib")

    # ── 1. 收集所有资源 URL ──────────────────────────────────────────
    css_urls = []
    for link in soup.find_all("link", rel="stylesheet"):
        href = link.get("href", "") or link.get("data-href", "") or ""
        if href and not href.startswith("data:"):
            css_urls.append(urllib.parse.urljoin(url, href))

    img_urls = []
    for img in soup.find_all("img"):
        for attr in ("data-src", "src"):
            src = img.get(attr, "")
            if src and not src.startswith("data:"):
                img_urls.append(urllib.parse.urljoin(url, src))

    # 背景图
    bg_urls = set()
    for el in soup.find_all(style=True):
        for m in re.findall(r'url\(["\']?([^"\'()]+)["\']?\)', el.get("style", "")):
            if not m.startswith("data:") and not m.startswith("blob:"):
                bg_urls.add(urllib.parse.urljoin(url, m))

    all_urls = list(set(css_urls + img_urls + list(bg_urls)))

    # ── 2. 并发下载 ─────────────────────────────────────────────────
    global _ASSET_COUNTER
    _ASSET_COUNTER = int(time.time() * 1000) % 1000000
    download_results = {}
    if all_urls:
        download_results = _parallel_download(all_urls)

    # ── 3. 建立 URL → (本地文件名, 内容) 映射 ──────────────────────
    # 用内容 hash 去重：相同内容的资源只存一个文件
    url_to_asset = {}   # url -> (local_filename, bytes)
    assets = []         # [(filename, bytes)]
    content_hashes = {} # content_hash -> filename（已分配的文件名）

    for res_url in all_urls:
        content, _ = download_results.get(res_url, (b"", ""))
        if not content:
            continue

        # 用内容 hash 去重
        content_hash = hashlib.md5(content[:100000]).hexdigest()[:16]
        ext_from_url = os.path.splitext(urllib.parse.urlparse(res_url).path.lower())[1]
        if ext_from_url not in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".css", ".js"):
            ext_from_url = ""
        # 内容检测比 URL 更可靠
        real_ext = _detect_content_ext(content, res_url) if not ext_from_url else ext_from_url
        if not real_ext:
            real_ext = ext_from_url

        # 同一内容用同一文件名（去重）
        if content_hash in content_hashes:
            fname = content_hashes[content_hash]
        else:
            _ASSET_COUNTER += 1
            ts = int(time.time() * 1000)
            rnd = str(time.time()).split(".")[-1][:6]
            fname = f"{ts}.{rnd}{real_ext}"
            content_hashes[content_hash] = fname
            assets.append((fname, content))

        url_to_asset[res_url] = (fname, content)

    # ── 4. 替换 HTML 中的资源 URL ──────────────────────────────────

    # 4a. <link href="..."> → <link href="./assets/{fname}">
    for link in soup.find_all("link"):
        href = link.get("href", "") or link.get("data-href", "") or ""
        if not href or href.startswith("data:"):
            continue
        abs_href = urllib.parse.urljoin(url, href)
        if abs_href in url_to_asset:
            fname = url_to_asset[abs_href][0]
            link["href"] = f"./assets/{fname}"
        elif href.startswith("http") or href.startswith("//"):
            # 没下载成功的外部链接，移除
            link.decompose()

    # 4b. <img data-src="..."> / <img src="...">
    for img in soup.find_all("img"):
        # 清理没有 src 也没有 data-src 的空 img（视频/音频占位符等）
        data_src = img.get("data-src", "") or ""
        src_attr = img.get("src", "") or ""
        if not data_src and not src_attr:
            # 无任何 URL 的空 img，移除
            img.decompose()
            continue
        # 优先用 data-src（懒加载），同步到 src
        if data_src and not data_src.startswith("data:") and not data_src.startswith("blob:"):
            abs_src = urllib.parse.urljoin(url, data_src)
            if abs_src in url_to_asset:
                fname = url_to_asset[abs_src][0]
                img["data-src"] = f"./assets/{fname}"
                img["src"] = f"./assets/{fname}"
            elif data_src.startswith("http"):
                img["src"] = data_src  # 保留远程 URL（降级）
        elif src_attr and not src_attr.startswith("data:") and not src_attr.startswith("blob:"):
            abs_src = urllib.parse.urljoin(url, src_attr)
            if abs_src in url_to_asset:
                fname = url_to_asset[abs_src][0]
                img["src"] = f"./assets/{fname}"
        # 清理可能导致问题的属性
        for bad_attr in ("onclick", "data-onclick"):
            if img.get(bad_attr):
                del img[bad_attr]

    # 4c. style 属性中的 background-image URL（处理相对路径如 .././ 或 ../../）
    def _replace_bg(m):
        raw = m.group(1).strip()
        if raw.startswith("data:") or raw.startswith("blob:"):
            return m.group(0)
        # 解析相对 URL：去除 ../ 和 ./，保留有效路径
        abs_url_raw = urllib.parse.urljoin(url, raw)
        abs_url = abs_url_raw
        if abs_url in url_to_asset:
            fname = url_to_asset[abs_url][0]
            return f'url("./assets/{fname}")'
        return m.group(0)  # 保留原始 URL（降级）

    for el in soup.find_all(style=True):
        new_style = re.sub(r'url\(["\']?([^"\'()]+)["\']?\)', _replace_bg, el.get("style", ""))
        el["style"] = new_style

    # ── 5. 清理无用标签（但先保存 bottom_bar） ───────────────────────
    # BeautifulSoup 把 #js_article_bottom_bar 放在 #js_article 外面，
    # 所以先提取其 HTML，再删除
    bottom_bar_tag = soup.find(id="js_article_bottom_bar")
    bottom_bar_html = str(bottom_bar_tag) if bottom_bar_tag else ""

    for tag in soup.find_all(["script", "noscript"]):
        tag.decompose()
    for sel in ["#js_pc_qr_code", "#js_top_ad_area", "#js_tags_preview_toast",
                 "#content_bottom_area", "#wx_stream_article_slide_tip",
                 "#js_article_bottom_bar"]:
        for t in soup.select(sel):
            t.decompose()

    # ── 6. 移除 #js_content 的内联样式（这是导致空白页的核心原因）──
    js_content = soup.find(id="js_content")
    if js_content:
        if "style" in js_content.attrs:
            del js_content["style"]
    rich_media = soup.find(class_="rich_media_area_primary")
    if rich_media and "style" in rich_media.attrs:
        del rich_media["style"]

    # ── 7. 提取 #js_article 完整结构（含底部栏） ───────────────────
    js_article = soup.find(id="js_article")
    if js_article:
        article_html = str(js_article)
        if bottom_bar_html:
            article_html += "\n" + bottom_bar_html
    else:
        article_html = str(soup.find(id="js_content") or soup)

    # ── 8. 组装完整 HTML 文档 ────────────────────────────────────────
    # 参考开源项目：构建干净的 <!DOCTYPE><html><head>...<body> 结构
    # 收集所有 CSS link（已替换为本地路径 ./assets/*.css）
    local_css_links = []
    seen_css_paths = set()
    for link in soup.find_all("link"):
        href = link.get("href", "") or ""
        if href and href not in seen_css_paths and (
            href.endswith(".css") or "./assets/" in href
        ):
            # 跳过不存在的 app_wx.css（MINIMAL_CSS 已内联）
            if "app_wx.css" not in href:
                local_css_links.append(
                    f'<link rel="stylesheet" href="{href}">'
                )
                seen_css_paths.add(href)

    body_cls = soup.find("body")
    body_class = body_cls.get("class", []) if body_cls else []
    body_cls_str = " ".join(body_class) if body_class else ""

    html_doc = f'''<!DOCTYPE html>
<html lang="zh_CN">
<head>
    <meta charset="utf-8">
    <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
    <meta http-equiv="X-UA-Compatible" content="IE=edge">
    <meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=0,viewport-fit=cover">
    <title>{title}</title>
{chr(10).join(local_css_links)}
    <style>
{MINIMAL_CSS}
    </style>
</head>
<body class="{body_cls_str}">
{article_html}
</body>
</html>'''

    # ── 9. 更新 title ───────────────────────────────────────────────
    html_doc = re.sub(r"<title>[^<]*</title>", f"<title>{title}</title>", html_doc, count=1)

    return html_doc, assets


# ---------------------------------------------------------------------------
# 主抓取函数
# ---------------------------------------------------------------------------

def fetch_article(url: str, fmt: str = "markdown") -> dict:
    if fmt not in VALID_FORMATS:
        raise ValueError(f"不支持的格式: {fmt}")
    if not _is_valid_mp_url(url):
        raise ValueError(f"非法的微信文章URL: {url}")

    raw_bytes = _robust_urlopen(url, {
        "Referer": "https://mp.weixin.qq.com/",
        "Origin": "https://mp.weixin.qq.com",
    })
    if not raw_bytes:
        raise ValueError(f"无法获取页面: {url}")
    raw_html = raw_bytes.decode("utf-8", errors="replace")
    soup = BeautifulSoup(raw_html, "html5lib")

    title_tag = (soup.find("h1", id="activity-name")
                 or soup.find("meta", property="og:title")
                 or soup.find("title"))
    title = ((title_tag.get("content", "") or title_tag.get_text(strip=True))
             if title_tag and hasattr(title_tag, 'get')
             else (title_tag.get_text(strip=True) if title_tag else "未命名文章"))

    author_tag = soup.find("meta", property="og:article:author")
    author = author_tag.get("content", "") if author_tag else ""

    date_tag = soup.find("em", id="publish_time")
    date_str = date_tag.get_text(strip=True) if date_tag else ""

    cover_tag = soup.find("meta", property="og:image")
    cover_url = cover_tag.get("content", "") if cover_tag else ""

    desc_tag = soup.find("meta", property="og:description")
    summary = desc_tag.get("content", "") if desc_tag else ""

    orig_tag = soup.find("span", id="copyright_logo")
    is_original = "是" if orig_tag else "否"

    topic_tag = soup.find("meta", property="og:article:tag")
    topic = topic_tag.get("content", "") if topic_tag else ""

    # 阅读量/评论数：从页面 JS 变量中提取（需先发请求到微信 API 获取真实数据）
    read_num = _extract_read_count(raw_html, url)
    comment_count = _extract_comment_count(raw_html)

    article_id = _extract_article_id(url)
    article = (soup.find("div", id="js_content")
               or soup.find("div", class_="rich_media_content"))

    if fmt == "html":
        html_str, assets = _build_standalone_html(raw_html, title, author, date_str, url)
        content_obj = {"html": html_str, "assets": assets}
    elif fmt == "markdown":
        article_html = str(article) if article else ""
        content_obj = _to_markdown(article_html, title, author, date_str, url)
    elif fmt == "text":
        article_html = str(article) if article else ""
        content_obj = _to_text(article_html)
    elif fmt == "json":
        article_html = str(article) if article else ""
        raw_text = _to_text(article_html)
        content_obj = json.dumps({
            "标题": title, "作者": author, "日期": date_str,
            "URL": url, "文章ID": article_id, "摘要": summary,
            "封面图": cover_url, "话题": topic, "是否原创": is_original,
            "正文": raw_text,
        }, ensure_ascii=False, indent=2)
    elif fmt == "excel":
        article_html = str(article) if article else ""
        content_obj = _to_text(article_html)

    return {
        "title": title,
        "content": content_obj,
        "format": fmt,
        "url": url,
        "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "article_id": article_id,
        "author": author,
        "date": date_str,
        "cover_url": cover_url,
        "summary": summary,
        "topic": topic,
        "is_original": is_original,
        "read_count": read_num,
        "comment_count": comment_count,
    }


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def _is_valid_mp_url(url: str) -> bool:
    try:
        return urllib.parse.urlparse(url).hostname == "mp.weixin.qq.com"
    except Exception:
        return False


def _extract_article_id(url: str) -> str:
    m = re.search(r'/s/([a-zA-Z0-9_-]+)', url)
    return m.group(1) if m else hashlib.md5(url.encode()).hexdigest()[:12]


def _extract_read_count(raw_html: str, article_url: str) -> str:
    """
    从微信文章页面的 JS 变量中提取 mid/sn，再请求微信阅读量 API。
    API: https://mp.weixin.qq.com/mp/vdapgetdatacount
    成功返回 JSON: {"list":[{"read_count":xxx,"like_count":xxx,"在看数":xxx}]}
    """
    mid_m = re.search(r'var\s+mid\s*=\s*["\']?(\d+)', raw_html)
    sn_m = re.search(r'var\s+sn\s*=\s*["\']?([a-zA-Z0-9]+)', raw_html)
    biz_m = re.search(r'var\s+biz\s*=\s*["\']?([^"\'\s;]+)', raw_html)
    idx_m = re.search(r'var\s+idx\s*=\s*["\']?(\d+)', raw_html)

    if not mid_m or not sn_m or not biz_m:
        return ""
    mid = mid_m.group(1)
    sn = sn_m.group(1)
    biz = biz_m.group(1)
    idx = idx_m.group(1) if idx_m else "1"

    api_url = (
        f"https://mp.weixin.qq.com/mp/vdapgetdatacount"
        f"?__biz={biz}&mid={mid}&idx={idx}&sn={sn}"
        f"&count_type=1&scene=0&is_ok=1&appid=wx8b0013715a442d37"
        f"&devicetype=Windows&version=17010000&clientversion=17010000"
        f"&lang=zh_CN&pass_ticket=&wx_header=1"
    )
    data = _robust_urlopen(api_url, headers={
        "Referer": article_url,
        "User-Agent": USER_AGENT,
    }, timeout=8)
    if not data:
        return ""
    try:
        import json as _json
        obj = _json.loads(data.decode("utf-8", errors="replace"))
        read_count = obj.get("list", [{}])[0].get("read_count", "")
        return str(read_count) if read_count else ""
    except Exception:
        return ""


def _extract_comment_count(raw_html: str) -> str:
    """从页面 HTML 中提取评论数（JS 变量形式：window.__comment_count = N）"""
    # 方式1：直接正则匹配 JS 变量
    for pat in (
        r'window\.__comment_count\s*=\s*(\d+)',
        r'comment_count\s*[=:]\s*["\']?(\d+)',
        r'"comment_count"\s*:\s*(\d+)',
        r'var\s+comment_count\s*=\s*(\d+)',
    ):
        m = re.search(pat, raw_html)
        if m:
            return m.group(1)
    return ""


def _to_markdown(html: str, title: str, author: str, date: str, url: str) -> str:
    import markdownify
    md = markdownify.markdownify(html, heading_style="ATX", bullets="-",
                                   links=True, images=True)
    md = re.sub(r'\n{3,}', '\n\n', md)
    meta = " | ".join(filter(None, [author, date]))
    return f"# {title}\n{meta}\n\n> 来源: [微信文章]({url})\n\n{md}"


def _to_text(html: str) -> str:
    soup = BeautifulSoup(html, "html5lib")
    lines = [l for l in soup.get_text(separator="\n", strip=True).split("\n") if l.strip()]
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 保存函数
# ---------------------------------------------------------------------------

def _save_standalone_html(article: dict, output_dir: str, safe_title: str) -> str:
    obj = article["content"]
    html_str = obj.get("html", "") if isinstance(obj, dict) else str(obj)
    assets = obj.get("assets", []) if isinstance(obj, dict) else []

    article_id = article.get("article_id", "")
    base_name = f"{safe_title}_{article_id}"
    article_dir = os.path.join(output_dir, base_name)
    assets_dir = os.path.join(article_dir, "assets")
    os.makedirs(assets_dir, exist_ok=True)

    # 保存 HTML
    html_path = os.path.join(article_dir, "index.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_str)

    # 保存 assets（只保存真正下载到内容的文件）
    saved_count = 0
    for fname, data in assets:
        if data:  # 只保存有内容的
            fp = os.path.join(assets_dir, fname)
            with open(fp, "wb") as f:
                f.write(data)
            saved_count += 1

    return html_path


def _save_excel(article: dict, output_dir: str, safe_title: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center")
    val_font = Font(name="Microsoft YaHei", size=9)
    val_align = Alignment(vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "文章信息"

    cols = [
        ("A", "来源", 12), ("B", "ID", 28), ("C", "链接", 45),
        ("D", "标题", 40), ("E", "封面图", 45), ("F", "摘要", 60),
        ("G", "抓取时间", 20), ("H", "发布时间", 20),
        ("I", "话题", 15), ("J", "是否原创", 12),
        ("K", "正文", 80), ("L", "作者", 15),
    ]

    for col_letter, col_name, col_width in cols:
        c = ws[f"{col_letter}1"]
        c.value = col_name
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border
        c.alignment = hdr_align
        ws.column_dimensions[col_letter].width = col_width

    row_data = [
        "微信公众号",
        article.get("article_id", ""),
        article.get("url", ""),
        article.get("title", ""),
        article.get("cover_url", ""),
        article.get("summary", ""),
        article.get("fetched_at", ""),
        article.get("date", ""),
        article.get("topic", ""),
        article.get("is_original", ""),
        article.get("content", ""),
        article.get("author", ""),
    ]

    for i, (col_letter, _, _) in enumerate(cols):
        c = ws[f"{col_letter}2"]
        c.value = row_data[i] if i < len(row_data) else ""
        c.font = val_font
        c.border = border
        c.alignment = val_align
        c.fill = PatternFill("solid",
                             fgColor="EBF3FB" if i % 2 == 0 else "FFFFFF")

    ws.row_dimensions[2].height = 300
    filename = f"{safe_title}_{article.get('article_id', '')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


def save_article(article: dict, output_dir: str = None) -> str:
    if output_dir is None:
        output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    safe_title = re.sub(r'[<>:"/\\|?*]', '_', article["title"])[:80]
    fmt = article["format"]

    if fmt == "excel":
        return _save_excel(article, output_dir, safe_title)
    elif fmt == "html":
        return _save_standalone_html(article, output_dir, safe_title)
    elif fmt == "markdown":
        p = os.path.join(output_dir, f"{safe_title}_{article['article_id']}.md")
        with open(p, "w", encoding="utf-8") as f:
            f.write(article["content"])
        return p
    elif fmt == "text":
        p = os.path.join(output_dir, f"{safe_title}_{article['article_id']}.txt")
        with open(p, "w", encoding="utf-8") as f:
            f.write(article["content"])
        return p
    elif fmt == "json":
        p = os.path.join(output_dir, f"{safe_title}_{article['article_id']}.json")
        with open(p, "w", encoding="utf-8") as f:
            f.write(article["content"])
        return p
    return ""


# ---------------------------------------------------------------------------
# CLI 入口
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python fetch_article.py <url> [format] [output_dir]")
        sys.exit(1)
    url = sys.argv[1]
    fmt = sys.argv[2] if len(sys.argv) > 2 else "markdown"
    out_dir = sys.argv[3] if len(sys.argv) > 3 else None
    print(f"抓取中: {url}", flush=True)
    try:
        article = fetch_article(url, fmt)
        fp = save_article(article, out_dir)
        print(f"完成: {article['title']}", flush=True)
        print(f"保存: {fp}", flush=True)
        if fmt == "html":
            obj = article["content"]
            html_sz = len(obj.get("html", "")) if isinstance(obj, dict) else 0
            asset_count = len(obj.get("assets", [])) if isinstance(obj, dict) else 0
            print(f"格式: html | HTML: {html_sz//1024}KB | 资源文件: {asset_count}个",
                  flush=True)
        else:
            c = article["content"]
            sz = len(c) if isinstance(c, str) else 0
            print(f"格式: {fmt} | 字符数: {sz}", flush=True)
    except Exception as e:
        print(f"错误: {e}", flush=True)
        import traceback
        traceback.print_exc()
        sys.exit(1)
