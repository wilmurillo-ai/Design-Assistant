import requests, shutil, sys, json, subprocess, datetime
from openpyxl import load_workbook 
from concurrent.futures import ThreadPoolExecutor, as_completed

TEMPLATE_PATH = "template/ScanResult.xlsx"

NEW_FILE = f'ScanResult_{datetime.date.today()}.xlsx'

def getCVEversion(name, version, ecosystem):
    url = 'https://api.osv.dev/v1/query'
    payload = {
        "version": version,
        "package": {
            "name": name,
            "ecosystem": ecosystem
        }
    }
    res = requests.post(url, json=payload).json()
    
    if 'vulns' in res:
        return res['vulns']
    return None

if sys.argv.__len__() == 1:
    print('Please add the packages language !\n\033[31mUsage\033[0m:\n\t\033[33mpython3 ./scanner.py [eg: npm, python, os]\033[0m')
    exit()
program = sys.argv[1].lower()
lines = []
all_rows = []

if program == 'centos':
    with open('packages/packages-lin.txt', 'r') as file:
        lines = file.readlines()
    eco = 'Red Hat'
elif program == 'os':
    result = subprocess.run(
        "lsb_release -a | grep 'Distributor ID' | awk '{print $3}'",
        shell=True,
        capture_output=True,
        text=True
    )
    eco = result.stdout.strip()
    operatingSystems = ['debian', 'ubuntu', 'kali']
    if eco.lower() in operatingSystems:
        if eco == 'kali':
            eco = 'Debian'
        result = subprocess.run(
            "dpkg-query -W -f='${binary:Package} ${Version}\n'",
            shell=True,
            capture_output=True,
            text=True
        )
        lines = result.stdout.strip().split('\n')
    else:
        print(f'\033[31mError\033[0m:\n\tYour os \033[35m{eco}\033[0m not in the specified list')
        exit()
elif program == 'npm':
    with open('packages/packages.json', 'r') as file:
        jsonLines = file.readlines()
    line = "".join(jsonLines)
    resp = json.loads(line)
    packages = resp['dependencies']
    for name, version in packages.items():
        lines.append(f'{name}=={version.split('^')[1]}')
    eco = 'npm'
elif program == 'python':
    with open('packages/packages.txt', 'r') as file:
        lines = file.readlines()
    eco = 'PyPI'
else:
    print('\033[31mError\033[0m:\n\tPlease choose either \033[30mnpm/python/os\033[0m')
    exit()

def check_package(pkg, eco):
    pkg = pkg.strip()
    if '==' in pkg:
        name, version = pkg.split('==', 1)
    else:
        name, version = pkg.split()

    vulns = getCVEversion(name, version, eco)
    return name, version, vulns

def print_package_result(name, version, vulns):
    print(f'\033[34mPackage\033[0m: {name} {version}')
    if not vulns:
        print('\033[34m\tStatus\033[0m: Package is up to date')
    else:
        rows = []
        for vuln in vulns:
            # if eco == 'centos':
            upstream = vuln.get('id')
            # cves = [a for a in upstream if a.startswith("CVE-")]
            # if eco == 'PyPI':
            #     aliases = vuln.get('aliases', [])
            #     cves = [a for a in aliases if a.startswith("CVE-")]
        if upstream:
            print(f'\033[34m\tCVE\033[0m: \033[31m⚠ {upstream} \033[0m')
            rows.append({
                'MachineA': f'{name} - {version} - {upstream}'
            })
            return rows
        return None

max_workers = 20

with ThreadPoolExecutor(max_workers=max_workers) as executor:
    future_to_pkg = {
        executor.submit(check_package, line, eco): line 
        for line in lines
    }
    
    for future in as_completed(future_to_pkg):
        try:
            name, version, vulns = future.result()
            row = print_package_result(name, version, vulns)
            if row:
                all_rows.extend(row)
        except Exception as e:
            print(f'\033[31mError\033[0m: {e}')

if all_rows:
    shutil.copy(TEMPLATE_PATH, NEW_FILE)

    wb = load_workbook(NEW_FILE)
    ws = wb.active
    
    machine_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "MachineA":
            machine_col = col
            break
    
    if machine_col is None:
        raise Exception("Header 'MachineA' not found in template")
    
    row_index = 2
    while ws.cell(row=row_index, column=machine_col).value is not None:
        row_index += 1
    
    for row_data in all_rows:
        ws.cell(row=row_index, column=machine_col, value=row_data['MachineA'])
        row_index += 1

    wb.save(NEW_FILE)