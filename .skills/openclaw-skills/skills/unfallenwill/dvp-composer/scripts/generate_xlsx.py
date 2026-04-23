#!/usr/bin/env python3
"""Generate DVP Excel file from JSON input."""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
MODULE_FILLS = {
    "default": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "alt": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
}
SEVERITY_FILLS = {
    "Critical": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "Major": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    "Minor": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
}
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _apply_cell(cell, fill=None):
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill


def create_check_list_sheet(ws, checks):
    headers = [
        "Check ID", "Module", "Category", "Check Description", "Logic Rule",
        "Applicable Scope", "Trigger Condition", "Query Wording",
        "Severity", "Execution Method", "Status", "Notes",
    ]
    for col, header in enumerate(headers, 1):
        _apply_header(ws.cell(row=1, column=col, value=header))

    current_module = None
    fill_key = "default"
    for row_idx, check in enumerate(checks, 2):
        mod = check.get("module", "")
        if mod != current_module:
            current_module = mod
            fill_key = "alt" if fill_key == "default" else "default"
        row_fill = MODULE_FILLS[fill_key]

        values = [
            check.get("check_id", ""), mod, check.get("category", ""),
            check.get("description", ""), check.get("logic", ""),
            check.get("scope", ""), check.get("trigger", ""),
            check.get("query_wording", ""), check.get("severity", ""),
            check.get("method", ""), check.get("status", "Active"),
            check.get("notes", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            _apply_cell(cell, row_fill)
            if col == 9 and value in SEVERITY_FILLS:
                cell.fill = SEVERITY_FILLS[value]
                if value == "Critical":
                    cell.font = Font(bold=True, color="FFFFFF")

    widths = [12, 10, 15, 35, 30, 20, 20, 35, 10, 15, 8, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def create_summary_sheet(ws, summary):
    sections = [
        ("Document Title", f"Data Validation Plan - {summary.get('protocol_number', '')}"),
        ("Protocol Number", summary.get("protocol_number", "")),
        ("Study Phase", summary.get("study_phase", "")),
        ("Indication", summary.get("indication", "")),
        ("Sponsor", summary.get("sponsor", "")),
        ("Version", summary.get("version", "1.0")),
        ("Date", summary.get("date", str(date.today()))),
        ("Author", summary.get("author", "")),
        ("", ""),
        ("Scope", summary.get("scope", "")),
        ("Key Data", summary.get("key_data", "")),
        ("Risk Summary", summary.get("risk_summary", "")),
        ("Validation Strategy", summary.get("validation_strategy", "")),
    ]
    for idx, (label, value) in enumerate(sections, 1):
        if label:
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=idx, column=2, value=value).alignment = Alignment(wrap_text=True)

    roles_start = len(sections) + 2
    ws.cell(row=roles_start, column=1, value="Roles & Responsibilities").font = Font(bold=True, size=12)
    hr = roles_start + 1
    for col, h in enumerate(["Role", "Responsibility"], 1):
        _apply_header(ws.cell(row=hr, column=col, value=h))
    for off, role in enumerate(summary.get("roles", [])):
        r = hr + off + 1
        ws.cell(row=r, column=1, value=role.get("role", "")).border = THIN_BORDER
        ws.cell(row=r, column=2, value=role.get("responsibility", "")).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60


def create_revision_history_sheet(ws, revisions):
    for col, h in enumerate(["Version", "Date", "Author", "Description of Changes", "Reviewer", "Approval Status"], 1):
        _apply_header(ws.cell(row=1, column=col, value=h))
    for row_idx, rev in enumerate(revisions, 2):
        for col, key in enumerate(["version", "date", "author", "description", "reviewer", "status"], 1):
            cell = ws.cell(row=row_idx, column=col, value=rev.get(key, ""))
            _apply_cell(cell)
    for i, w in enumerate([10, 15, 15, 40, 15, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_ext_data_recon_sheet(ws, items):
    headers = [
        "Recon ID", "Data Source", "Provider", "Key Fields",
        "Recon Method", "Frequency", "Discrepancy Handling", "Responsible Party", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        _apply_header(ws.cell(row=1, column=col, value=h))
    for row_idx, item in enumerate(items, 2):
        for col, key in enumerate(
            ["recon_id", "data_source", "provider", "key_fields",
             "method", "frequency", "discrepancy_handling", "responsible_party", "notes"], 1
        ):
            _apply_cell(ws.cell(row=row_idx, column=col, value=item.get(key, "")))
    for i, w in enumerate([12, 15, 20, 25, 15, 15, 30, 15, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_from_template(input_data, template_path, output_path):
    wb = openpyxl.load_workbook(template_path)
    check_sheet = None
    for name in wb.sheetnames:
        lower = name.lower()
        if any(kw in lower for kw in ["check", "list"]):
            check_sheet = wb[name]
            break
    if check_sheet is None:
        check_sheet = wb.active

    headers = [cell.value for cell in check_sheet[1]]
    field_mapping = {
        "check id": "check_id", "module": "module", "category": "category",
        "description": "description", "check description": "description",
        "logic": "logic", "logic rule": "logic", "scope": "scope",
        "applicable scope": "scope", "trigger": "trigger",
        "trigger condition": "trigger", "query": "query_wording",
        "query wording": "query_wording", "severity": "severity",
        "priority": "severity", "method": "method",
        "execution method": "method", "status": "status", "notes": "notes",
    }
    header_to_field = {}
    for idx, header in enumerate(headers):
        if header:
            header_to_field[idx] = field_mapping.get(header.lower().strip(), header.lower().strip())

    for row_idx, check in enumerate(input_data.get("checks", []), 2):
        for col_idx, field_name in header_to_field.items():
            check_sheet.cell(row=row_idx, column=col_idx + 1, value=check.get(field_name, ""))

    wb.save(output_path)
    print(f"DVP generated using template: {output_path}")


def generate_default(input_data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Check List"
    create_check_list_sheet(ws, input_data.get("checks", []))

    ws_s = wb.create_sheet("Summary")
    create_summary_sheet(ws_s, input_data.get("summary", {}))

    ws_r = wb.create_sheet("Revision History")
    create_revision_history_sheet(ws_r, input_data.get("revisions", [
        {"version": "1.0", "date": str(date.today()), "author": "", "description": "Initial draft"}
    ]))

    ws_e = wb.create_sheet("Ext Data Recon")
    create_ext_data_recon_sheet(ws_e, input_data.get("reconciliation", []))

    wb.save(output_path)
    print(f"DVP generated: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate DVP Excel file")
    parser.add_argument("--input", required=True, help="Path to JSON input file")
    parser.add_argument("--output", required=True, help="Path to output Excel file")
    parser.add_argument("--template", help="Path to Excel template (optional)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    with open(input_path, "r", encoding="utf-8") as f:
        input_data = json.load(f)

    if args.template:
        template_path = Path(args.template)
        if not template_path.exists():
            print(f"Error: Template file not found: {template_path}", file=sys.stderr)
            sys.exit(1)
        generate_from_template(input_data, str(template_path), args.output)
    else:
        generate_default(input_data, args.output)


if __name__ == "__main__":
    main()
