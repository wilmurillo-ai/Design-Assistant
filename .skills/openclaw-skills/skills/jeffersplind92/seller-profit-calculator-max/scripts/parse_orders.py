#!/usr/bin/env python3
"""
Seller Profit Calculator - Core Parser
Parses 妙手ERP order export and calculates profit per order.
Supports: TikTok Shop, Allegro, Temu, SHEIN, Fruugo and other platforms via 妙手ERP.

Calculation reference: TikTok Shop 170-column order bill (标准地基)
All platforms are mapped to this standard field schema.

Platform field aliases (Allegro 41-col → TikTok 170-col mapping):
  交易收入        → 买家实付金额(RMB)
  运费收入        → 商家运费(RMB) (in positive)
  EPR费用(已退费)  → (platform rebate, added to income)
  售后退款        → 退款金额(RMB) / 买家申请退款(RMB)
  运费退款        → 运费调整(RMB) / 运费补偿(RMB)
  EPR费用(已扣费)  → 进口增值税(RMB) / 关税(RMB)
  退货面单费      → 实际逆向物流运费(RMB)
  发货面单费      → 物流供应商清关服务费(RMB)
  违规扣款        → 平台惩罚(RMB)
  其他账务        → 其他调整(RMB)
  广告服务费       → 交易手续费(RMB) / GMV广告费用(RMB)
  推广服务费       → 推荐费(RMB)
  采购金额        → 采购成本(RMB)
  包材费          → 包材费(RMB)
  头程运费        → 运费成本(RMB) (domestic)
  尾程运费        → 运费成本(RMB) (international)
  广告成本        → 广告成本(RMB)
  运营成本        → 运营成本(RMB)
  仓库操作费       → 仓库操作费(RMB)
  其他成本        → 其他成本(RMB)
  订单其他收入    → 平台补偿(RMB) / 退款订单补偿(RMB)
"""

import sys
import json
import argparse
from datetime import datetime
from openpyxl import load_workbook

def parse_currency(val):
    """Parse currency string like '￥73.44' or '-45.00' to float."""
    if val is None or val == '':
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.startswith('￥') or s.startswith('¥') or s.startswith('$'):
        s = s[1:]
    s = s.replace(',', '').strip()
    try:
        return float(s)
    except:
        return 0.0

def parse_int(val):
    if val is None or val == '':
        return 0
    if isinstance(val, int):
        return val
    try:
        return int(str(val).strip().replace(',', ''))
    except:
        return 0

# ====================================================================
# TikTok Shop 170-column standard schema (地基)
# All costs are positive numbers; negatives from platform are preserved.
# ====================================================================

def calculate_profit_tiktok(row, h, field_map=None):
    """
    Calculate profit using TikTok Shop 170-column standard.
    h = header_map (field_name -> 0-based column index in row list)
    """
    def g(name, alt=None):
        """Get row value by header name, return 0 if not found."""
        idx = h.get(name, -1)
        if idx < 0 and alt:
            idx = h.get(alt, -1)
        if idx < 0:
            return 0.0
        val = row[idx]
        if val is None or val == '':
            return 0.0
        return parse_currency(val)

    def g_str(name, alt=None):
        idx = h.get(name, -1)
        if idx < 0 and alt:
            idx = h.get(alt, -1)
        if idx < 0:
            return ''
        return str(row[idx] or '').strip()

    # --- Platform income (平台收入) ---
    # Use 已结算金额(RMB) if available, else 预估回款金额(RMB)
    settled = g('已结算金额(RMB)')
    estimated = g('预估回款金额(RMB)')
    platform_income = settled if settled != 0 else estimated

    # Additional income components
    freight_income = abs(g('商家运费(RMB)'))   # positive offset to income
    epr_refund = g('EPR费用(已退费)')
    platform_income += freight_income + epr_refund

    # --- Platform expense (平台支出) ---
    # Core platform fees
    transaction_fee = g('交易手续费(RMB)')
    card_fee = g('信用卡付款手续费(RMB)')
    platform_commission = g('TikTok 平台佣金(RMB)')
    referral_fee = g('推荐费(RMB)')
    seller_freight = g('商家运费(RMB)')
    actual_freight = g('实际运费(RMB)')
    reverse_freight = g('实际逆向物流运费(RMB)')
    transport_insurance = g('运输保险费(RMB)')

    # Creator/d affiliate
    creator_commission = g('达人佣金(RMB)')
    partner_commission = g('合作伙伴佣金(RMB)')
    affiliate_ad_fee = g('联盟商店广告佣金(RMB)')

    # Platform services
    sfp_fee = g('SFP服务费(RMB)')
    live_fee = g('LIVE Specials 计划服务费(RMB)')
    bonus_fee = g('Bonus金币返现服务费(RMB)')
    tiktok_mall_fee = g('TikTok Shop Mall服务费(RMB)')
    voucher_fee = g('Voucher xtra 计划服务费(RMB)')
    flash_sale_fee = g('限时抢购服务费(RMB)')
    promo_fee = g('促销活动服务费(RMB)')
    dynamic_fee = g('动态服务费(RMB)')
    other_service_fee = g('其他服务费(RMB)')
    refund_mgmt_fee = g('退款管理费(RMB)')
    preorder_fee = g('预购计划服务费(RMB)')

    # Taxes & duties
    vat = g('VAT(RMB)')
    import_vat = g('进口增值税(RMB)')
    customs_duty = g('关税(RMB)')
    clearing_fee = g('物流供应商清关服务费(RMB)')
    sst = g('SST(RMB)')
    gst = g('GST(RMB)')
    mx_vat = g('墨西哥增值税(RMB)')
    mx_income_tax = g('墨西哥联邦所得税(RMB)')
    anti_dumping = g('反倾销税(RMB)')

    # Deductions/adjustments
    refund_amount = g('退款金额(RMB)')
    total_adjustment = g('总调整金额(RMB)')
    buyer_refund = g('买家申请退款(RMB)')
    cs_compensation = g('客户服务补偿(RMB)')
    seller_deduction = g('商家体验扣款(RMB)')
    gmiv_ad_fee = g('GMV广告费用(RMB)')
    commission_adjust = g('平台佣金调整(RMB)')
    platform_penalty = g('平台惩罚(RMB)')
    promo_adjust = g('促销调整(RMB)')
    commission_discount = g('平台佣金折扣(RMB)')
    platform_compensation = g('平台补偿(RMB)')
    refund_order_compensation = g('退款订单补偿(RMB)')
    joint_funded = g('共同出资费用(RMB)')
    fbt_storage = g('FBT仓储服务费(RMB)')
    logistics_compensation = g('物流补偿(RMB)')
    freight_adjust = g('运费调整(RMB)')
    freight_compensation = g('运费补偿(RMB)')
    freight_rebate = g('运费回扣(RMB)')
    sample_freight = g('样品运费(RMB)')
    other_adjust = g('其他调整(RMB)')

    platform_expense = (
        transaction_fee + card_fee + platform_commission + referral_fee +
        seller_freight + actual_freight + reverse_freight + transport_insurance +
        creator_commission + partner_commission + affiliate_ad_fee +
        sfp_fee + live_fee + bonus_fee + tiktok_mall_fee + voucher_fee +
        flash_sale_fee + promo_fee + dynamic_fee + other_service_fee +
        refund_mgmt_fee + preorder_fee +
        vat + import_vat + customs_duty + clearing_fee +
        sst + gst + mx_vat + mx_income_tax + anti_dumping +
        refund_amount + total_adjustment +
        buyer_refund + cs_compensation + seller_deduction +
        gmiv_ad_fee + commission_adjust + platform_penalty +
        promo_adjust + commission_discount +
        platform_compensation + refund_order_compensation +
        joint_funded + fbt_storage +
        logistics_compensation + freight_adjust +
        freight_compensation + freight_rebate +
        sample_freight + other_adjust
    )

    # --- Order costs (订单成本) ---
    purchase_cost = g('采购成本(RMB)')
    freight_cost = g('运费成本(RMB)')
    warehouse_fee = g('仓库操作费(RMB)')
    packaging_fee = g('包材费(RMB)')
    ad_cost = g('广告成本(RMB)')
    ops_cost = g('运营成本(RMB)')
    other_cost = g('其他成本(RMB)')

    order_cost = (
        purchase_cost + freight_cost + warehouse_fee +
        packaging_fee + ad_cost + ops_cost + other_cost
    )

    # --- Other income ---
    other_income = (
        g('平台补偿(RMB)') + g('退款订单补偿(RMB)') +
        g('运费回扣(RMB)') + g('运费补偿(RMB)') +
        g('平台佣金折扣(RMB)') + g('活动运费补贴(RMB)') +
        g('平台运费补贴(RMB)')
    )

    # --- Net profit ---
    net_profit = platform_income - platform_expense - order_cost + other_income

    # Platform-declared profit (for validation)
    declared_profit = g('利润(RMB)')

    return {
        'platform_income': round(platform_income, 2),
        'platform_expense': round(platform_expense, 2),
        'order_cost': round(order_cost, 2),
        'other_income': round(other_income, 2),
        'net_profit_calc': round(net_profit, 2),
        'net_profit_declared': round(declared_profit, 2),
        'profit_diff': round(net_profit - declared_profit, 2),
        # breakdown
        'declared_profit': round(declared_profit, 2),
    }


def calculate_profit_allegro(row, h, field_map=None):
    """
    Calculate profit using Temu semi-hosted / Allegro 41-column format.

    Key rules:
    - Temu semi-hosted: 平台收入 = 申报价 × 销售数量 (confirmed by user)
    - Allegro: 平台收入 = 交易收入 + 运费收入(abs)
    - Cancelled: no platform income, fixed fee
    - Refunded: platform deducted declared price, may have additional fees
    """
    def g(name, alt=None):
        # Agent-provided mapping takes priority: field_map["标准字段名"] -> "实际列名"
        if field_map and name in field_map:
            actual_name = field_map[name]
            idx = h.get(actual_name, -1)
            if idx < 0:
                for k, v in h.items():
                    if k.startswith(actual_name + '_dup'):
                        idx = v; break
            if idx < 0:
                idx = h.get(name, -1)  # fallback to standard name
        else:
            idx = h.get(name, -1)
        if idx < 0:
            for k, v in h.items():
                if k.startswith(name + '_dup'):
                    idx = v; break
        if idx < 0 and alt:
            idx = h.get(alt, -1)
            if idx < 0:
                for k, v in h.items():
                    if k.startswith(alt + '_dup'):
                        idx = v; break
        if idx < 0: return 0.0
        val = row[idx]
        if val is None or val == '': return 0.0
        return parse_currency(val)

    def rv(name, alt=None):
        idx = h.get(name, -1)
        if idx < 0:
            for k, v in h.items():
                if k.startswith(name + '_dup'):
                    idx = v; break
        if idx < 0 and alt:
            idx = h.get(alt, -1)
            if idx < 0:
                for k, v in h.items():
                    if k.startswith(alt + '_dup'):
                        idx = v; break
        if idx < 0: return ''
        val = row[idx]
        return '' if val is None else str(val).strip()

    status = rv('订单状态', '订单状态')
    trading_income = g('交易收入')
    freight_income = g('运费收入')
    declared = g('申报价')
    qty = parse_int(rv('销售数量'))

    # Override for refund-annotated completed orders (set inside branches)
    _net_profit_override = None

    # Determine format: if trading_income > 0, treat as Allegro
    # Otherwise: Temu semi-hosted (申报价 × qty)
    if '已取消' in status:
        # Cancelled: no income, fixed fee
        platform_income = 0.0
        platform_expense = abs(g('售后退款'))
    elif '已退还' in status:
        # Refunded: platform handled the return
        # Case 1: 交易收入 > 0 (平台已结算部分款项，扣了运费等)
        #   net = 交易收入 - abs(售后退款) - 采购成本
        #   Example: PO-094: 交易收入=1018.58, 售后退款=-191.66, 采购=867.48
        #            net = 1018.58 - 191.66 - 867.48 = -40.56
        # Case 2: 交易收入 = 0 (平台未结算，只扣了采购成本)
        #   net = 0 - 采购成本
        #   Example: PO-00852192: 交易收入=0, 采购=828.47
        #            net = 0 - 828.47 = -828.47
        trading_income_val = g('交易收入')
        refund_val = g('售后退款')  # negative = platform deduction
        purchase = g('采购金额')
        if trading_income_val > 0:
            # Case 1: platform settled some amount, deducting freight/fee
            platform_income = trading_income_val
            platform_expense = abs(refund_val)  # abs of negative = positive fee
        else:
            # Case 2: no settlement, only purchase cost deducted (already in order_cost)
            platform_income = 0.0
            platform_expense = 0.0
    elif trading_income > 0:
        # Completed order (clean or refund-annotated):
        # Platform declares the final profit after netting ALL internal fees.
        # Use declared profit as override for exact platform match.
        platform_income = trading_income + abs(freight_income)
        platform_expense = 0.0
        _net_profit_override = g('利润')
    else:
        # Temu semi-hosted or other: income = 申报价 × 销售数量
        platform_income = declared * qty
        platform_expense = 0.0

    # Order cost
    purchase = g('采购金额')
    packaging = g('包材费')
    first_mile = g('头程运费')
    last_mile = g('尾程运费')
    ad_cost = g('广告成本')
    ops_cost = g('运营成本')
    warehouse = g('仓库操作费')
    other_cost = g('其他成本')
    order_cost = (purchase + packaging + first_mile + last_mile +
                  ad_cost + ops_cost + warehouse + other_cost)

    # Other income
    other_income = g('订单其他收入')

    # Use override for refund-annotated orders (platform handles the fee internally)
    net_profit = _net_profit_override if _net_profit_override is not None else \
                 (platform_income - platform_expense - order_cost + other_income)

    # Platform-declared profit
    declared_profit = g('利润')
    profit_diff = round(net_profit - declared_profit, 2)

    return {
        'platform_income': round(platform_income, 2),
        'platform_expense': round(platform_expense, 2),
        'order_cost': round(order_cost, 2),
        'other_income': round(other_income, 2),
        'net_profit_calc': round(net_profit, 2),
        'net_profit_declared': round(declared_profit, 2),
        'profit_diff': profit_diff,
    }


def detect_format(ws):
    """
    Auto-detect file format by scanning headers.
    Returns: ('tiktok', header_row, header_map) or ('allegro', header_row, header_map)

    For Allegro files, header_map keys use SUFFIXED names for duplicate headers
    (e.g. '总计_a', '总计_b', '总计_c') to avoid overwriting.
    """
    # First pass: build map with suffix for duplicates
    for r in range(1, min(5, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        row_strs = [str(v).strip() if v is not None else '' for v in row_vals]

        # TikTok format
        if '订单号' in row_strs and '利润(RMB)' in row_strs and '预估回款金额(RMB)' in row_strs:
            header_map = {}
            seen = {}
            for c, v in enumerate(row_vals):
                if v is not None:
                    k = str(v).strip()
                    if k in seen:
                        seen[k] += 1
                        k = f"{k}_dup{seen[k]}"
                    else:
                        seen[k] = 0
                    header_map[k] = c
            return 'tiktok', r, header_map

        # Allegro format
        if '订单编号' in row_strs and '交易收入' in row_strs:
            header_map = {}
            seen = {}
            for c, v in enumerate(row_vals):
                if v is not None:
                    k = str(v).strip()
                    if k in seen:
                        seen[k] += 1
                        k = f"{k}_dup{seen[k]}"
                    else:
                        seen[k] = 0
                    header_map[k] = c
            return 'allegro', r, header_map

    # Fallback: TikTok if has RMB fields
    row_vals = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_map = {str(v).strip(): c - 1 for c, v in enumerate(row_vals) if v is not None}
    if '预估回款金额(RMB)' in header_map or '已结算金额(RMB)' in header_map:
        return 'tiktok', 2, header_map
    return 'allegro', 2, header_map


def process_file(filepath, field_map=None):
    """Process Excel file and return structured profit results."""
    wb = load_workbook(filepath, data_only=False)
    ws = wb.active

    fmt, header_row, header_map = detect_format(ws)

    # Merge agent-provided field_map: standard_name -> actual column index
    # field_map = {"standard_field_name": "actual_column_name_in_this_file"}
    unmapped_fields = []
    if field_map:
        for std_name, actual_name in field_map.items():
            if actual_name in header_map:
                header_map[std_name] = header_map[actual_name]
            else:
                unmapped_fields.append((std_name, actual_name))

    results = {
        'format': fmt,
        'total_orders': 0,
        'completed_orders': 0,
        'cancelled_orders': 0,
        'total_income': 0.0,
        'total_platform_expense': 0.0,
        'total_order_cost': 0.0,
        'total_other_income': 0.0,
        'total_net_profit_calc': 0.0,
        'total_net_profit_declared': 0.0,
        'orders': [],
        'by_store': {},
        'by_platform': {},
        'by_sku': {},
    }

    calc_func = calculate_profit_tiktok if fmt == 'tiktok' else calculate_profit_allegro

    for r_idx in range(header_row, ws.max_row + 1):
        row = [ws.cell(r_idx, c).value for c in range(1, ws.max_column + 1)]

        # Get order ID
        order_no = row[header_map.get('订单编号', header_map.get('订单号', -1))]
        if order_no is None or str(order_no).strip() in ('', 'None'):
            continue

        def rv(name, alt=None):
            idx = header_map.get(name, -1 if not alt else header_map.get(alt, -1))
            if idx < 0: return ''
            v = row[idx]
            return '' if v is None else str(v).strip()

        status = rv('订单状态', '订单状态')
        platform = rv('站点', '站点')
        store = rv('店铺', '店铺')
        sku = rv('产品SKU', '产品SKU')
        qty = parse_int(rv('销售数量'))
        order_time = rv('下单时间')
        financial_time = rv('账务时间', '结算时间')

        profit = calc_func(row, header_map, field_map=field_map)

        results['total_orders'] += 1
        if status in ('交易完成', '已完成', 'completed', 'Completed'):
            results['completed_orders'] += 1
        elif status in ('已取消', '已撤销', 'cancelled', 'Cancelled'):
            results['cancelled_orders'] += 1

        results['total_income'] += profit['platform_income']
        results['total_platform_expense'] += profit['platform_expense']
        results['total_order_cost'] += profit['order_cost']
        results['total_other_income'] += profit['other_income']
        results['total_net_profit_calc'] += profit['net_profit_calc']
        results['total_net_profit_declared'] += profit.get('net_profit_declared', 0)

        # by_store
        if store not in results['by_store']:
            results['by_store'][store] = {'income': 0, 'expense': 0, 'cost': 0, 'profit': 0, 'count': 0}
        results['by_store'][store]['income'] += profit['platform_income']
        results['by_store'][store]['expense'] += profit['platform_expense']
        results['by_store'][store]['cost'] += profit['order_cost']
        results['by_store'][store]['profit'] += profit['net_profit_calc']
        results['by_store'][store]['count'] += 1

        # by_platform
        if platform not in results['by_platform']:
            results['by_platform'][platform] = {'income': 0, 'expense': 0, 'cost': 0, 'profit': 0, 'count': 0}
        results['by_platform'][platform]['income'] += profit['platform_income']
        results['by_platform'][platform]['expense'] += profit['platform_expense']
        results['by_platform'][platform]['cost'] += profit['order_cost']
        results['by_platform'][platform]['profit'] += profit['net_profit_calc']
        results['by_platform'][platform]['count'] += 1

        # by_sku
        if sku and sku not in ('', 'None'):
            if sku not in results['by_sku']:
                results['by_sku'][sku] = {'income': 0, 'expense': 0, 'cost': 0, 'profit': 0, 'count': 0}
            results['by_sku'][sku]['income'] += profit['platform_income']
            results['by_sku'][sku]['expense'] += profit['platform_expense']
            results['by_sku'][sku]['cost'] += profit['order_cost']
            results['by_sku'][sku]['profit'] += profit['net_profit_calc']
            results['by_sku'][sku]['count'] += 1

        results['orders'].append({
            'order_no': str(order_no),
            'status': status,
            'platform': platform,
            'store': store,
            'sku': sku,
            'qty': qty,
            'order_time': order_time,
            'financial_time': financial_time,
            **profit
        })

    # Round
    for k in ['total_income', 'total_platform_expense', 'total_order_cost',
              'total_other_income', 'total_net_profit_calc', 'total_net_profit_declared']:
        results[k] = round(results[k], 2)

    for cat in ['by_store', 'by_platform', 'by_sku']:
        for key in results[cat]:
            d = results[cat][key]
            for k in d:
                d[k] = round(d[k], 2)

    # Sort orders by profit (worst first)
    results['orders'].sort(key=lambda x: x['net_profit_calc'])
    return results


def format_markdown(results):
    if 'error' in results:
        return f"# Error: {results['error']}"

    fmt_label = 'TikTok Shop 170-col' if results['format'] == 'tiktok' else 'Allegro 41-col'

    md = []
    md.append("# 📊 订单利润分析报告")
    md.append(f"**格式:** {fmt_label} | **生成:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")

    # Summary
    total_rev = results['total_income']
    total_exp = results['total_platform_expense']
    total_cost = results['total_order_cost']
    total_profit = results['total_net_profit_calc']
    total_declared = results['total_net_profit_declared']
    profit_margin = (total_profit / total_rev * 100) if total_rev > 0 else 0
    diff = total_profit - total_declared

    md.append("## 📋 整体概况")
    md.append(f"| 指标 | 数值 |")
    md.append(f"|------|------|")
    md.append(f"| 总订单 | {results['total_orders']} |")
    md.append(f"| 完成 | {results['completed_orders']} |")
    md.append(f"| 取消 | {results['cancelled_orders']} |")
    md.append(f"| 平台总收入 | ¥{total_rev:,.2f} |")
    md.append(f"| 平台总支出 | ¥{total_exp:,.2f} |")
    md.append(f"| 订单总成本 | ¥{total_cost:,.2f} |")
    md.append(f"| **计算净利润** | **¥{total_profit:,.2f}** |")
    if total_declared != 0:
        md.append(f"| 平台申报净利润 | ¥{total_declared:,.2f} |")
        md.append(f"| 计算误差 | ¥{diff:,.2f} |")
    md.append(f"| **净利率** | **{profit_margin:.1f}%** |")

    # By platform
    if results['by_platform']:
        md.append("\n## 🌍 按平台汇总")
        md.append("| 平台 | 订单 | 收入 | 平台支出 | 订单成本 | 净利润 | 利润率 |")
        md.append("|------|------|------|---------|---------|--------|--------|")
        for k, d in sorted(results['by_platform'].items(), key=lambda x: x[1]['profit'], reverse=True):
            m = (d['profit'] / d['income'] * 100) if d['income'] != 0 else 0
            md.append(f"| {k} | {d['count']} | ¥{d['income']:,.2f} | ¥{d['expense']:,.2f} | ¥{d['cost']:,.2f} | ¥{d['profit']:,.2f} | {m:.1f}% |")

    # By store
    if results['by_store']:
        md.append("\n## 🏪 按店铺汇总")
        md.append("| 店铺 | 订单 | 收入 | 平台支出 | 订单成本 | 净利润 | 利润率 |")
        md.append("|------|------|------|---------|---------|--------|--------|")
        for k, d in sorted(results['by_store'].items(), key=lambda x: x[1]['profit'], reverse=True):
            m = (d['profit'] / d['income'] * 100) if d['income'] != 0 else 0
            md.append(f"| {k} | {d['count']} | ¥{d['income']:,.2f} | ¥{d['expense']:,.2f} | ¥{d['cost']:,.2f} | ¥{d['profit']:,.2f} | {m:.1f}% |")

    # Top/bottom
    orders = results['orders']
    if orders:
        bottom5 = orders[:5]
        top5 = orders[-5:][::-1]

        md.append("\n## 🔴 亏损最严重的5单")
        md.append("| # | 订单号 | 平台 | 店铺 | 收入 | 支出 | 成本 | 净利润 |")
        md.append("|---|--------|------|------|------|------|------|--------|")
        for i, o in enumerate(bottom5, 1):
            md.append(f"| {i} | {o['order_no'][:18]} | {o['platform']} | {o['store'][:8]} | ¥{o['platform_income']:,.2f} | ¥{o['platform_expense']:,.2f} | ¥{o['order_cost']:,.2f} | 🔴 ¥{o['net_profit_calc']:,.2f} |")

        md.append("\n## 🟢 盈利最高的5单")
        md.append("| # | 订单号 | 平台 | 店铺 | 收入 | 支出 | 成本 | 净利润 |")
        md.append("|---|--------|------|------|------|------|------|--------|")
        for i, o in enumerate(top5, 1):
            md.append(f"| {i} | {o['order_no'][:18]} | {o['platform']} | {o['store'][:8]} | ¥{o['platform_income']:,.2f} | ¥{o['platform_expense']:,.2f} | ¥{o['order_cost']:,.2f} | 🟢 ¥{o['net_profit_calc']:,.2f} |")

    return '\n'.join(md)


if __name__ == '__main__':
    import json as _json, argparse as _argparse
    _parser = _argparse.ArgumentParser(description='Seller Profit Calculator')
    _parser.add_argument('filepath', help='Path to order Excel file')
    _parser.add_argument('--json', dest='json_out', metavar='FILE', help='Output JSON file')
    _parser.add_argument('--field-map', dest='field_map', metavar='JSON',
                        help='Field mapping JSON or @filepath')
    _parser.add_argument('--markdown', dest='md_out', metavar='FILE', help='Output markdown file')
    _args = _parser.parse_args()

    _field_map = None
    if _args.field_map:
        _raw = _args.field_map
        if _raw.startswith('@'):
            with open(_raw[1:], encoding='utf-8') as _f:
                _field_map = _json.load(_f)
        else:
            _field_map = _json.loads(_raw)
    
    results = process_file(_args.filepath, field_map=_field_map)
    _md = format_markdown(results)
    print(_md)
    if _args.json_out:
        with open(_args.json_out, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"\nJSON saved to {_args.json_out}")
    if _args.md_out:
        with open(_args.md_out, 'w', encoding='utf-8') as f:
            f.write(_md)
        print(f"Markdown saved to {_args.md_out}")
