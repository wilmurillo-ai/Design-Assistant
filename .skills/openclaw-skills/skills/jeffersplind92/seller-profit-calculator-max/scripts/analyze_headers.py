#!/usr/bin/env python3
"""
analyze_headers.py - Analyze Excel file headers and sample rows for field mapping.

Usage:
    python3 analyze_headers.py <file.xlsx>                    # print analysis prompt
    python3 analyze_headers.py <file.xlsx> --show-headers      # print headers + sample rows only

This script does NOT call LLM. It prepares structured data for the Agent to analyze.
The Agent uses its own reasoning to produce the field_map JSON.
"""
import sys
import json
import argparse
from openpyxl import load_workbook


STANDARD_FIELDS = {
    # 基础信息
    "订单编号": "order_no",
    "订单号": "order_no",
    "订单状态": "status",
    "销售数量": "qty",
    "退货数量": "return_qty",
    "履约类型": "fulfillment_type",
    "店铺": "store",
    "站点": "platform",
    # 时间
    "下单时间": "order_time",
    "付款时间": "pay_time",
    "发货时间": "ship_time",
    "账务时间": "finance_time",
    "结算时间": "finance_time",
    # 收入/利润 (Allegro/Temu)
    "利润": "profit_declared",
    "成本利润率": "cost_profit_rate",
    "销售利润率": "sales_profit_rate",
    "总计": "total_income",
    "交易收入": "trading_income",
    "运费收入": "freight_income",
    "EPR费用(已退费)": "epr_refund",
    "售后退款": "refund",
    "运费退款": "freight_refund",
    "EPR费用(已扣费)": "epr_fee",
    "退货面单费": "return_shipping_fee",
    "发货面单费": "shipping_fee",
    "违规扣款": "violation_deduction",
    "广告服务费": "ad_service_fee",
    "推广服务费": "promotion_fee",
    "其他账务": "other_finance",
    "总计.1": "total_platform_expense",
    # 成本 (Allegro/Temu)
    "采购金额": "purchase_cost",
    "包材费": "packaging_fee",
    "头程运费": "first_mile",
    "尾程运费": "last_mile",
    "广告成本": "ad_cost",
    "运营成本": "ops_cost",
    "仓库操作费": "warehouse_fee",
    "其他成本": "other_cost",
    "总计.2": "total_order_cost",
    "订单其他收入": "other_order_income",
    # 申报价 (Temu)
    "申报价": "declared_price",
    # TikTok 170-col fields
    "已结算金额(RMB)": "settled_amount",
    "预估回款金额(RMB)": "estimated_amount",
    "买家实付金额(RMB)": "buyer_paid",
    "商家运费(RMB)": "seller_freight",
    "退款金额(RMB)": "refund_tiktok",
    "采购成本(RMB)": "purchase_cost_tiktok",
    "运费成本(RMB)": "freight_cost",
    "交易手续费(RMB)": "transaction_fee",
    "TikTok平台佣金(RMB)": "platform_commission",
    "VAT(RMB)": "vat",
    "进口增值税(RMB)": "import_vat",
    "关税(RMB)": "customs_duty",
    "平台惩罚(RMB)": "platform_penalty",
    "平台补偿(RMB)": "platform_compensation",
    "退款订单补偿(RMB)": "refund_compensation",
    "FBT仓储服务费(RMB)": "fbt_fee",
    "物流供应商清关服务费(RMB)": "clearing_fee",
    "实际逆向物流运费(RMB)": "actual_reverse_freight",
    "实际运费(RMB)": "actual_freight",
    "运输保险费(RMB)": "transport_insurance",
    "达人佣金(RMB)": "creator_commission",
    "推荐费(RMB)": "referral_fee",
    "信用卡付款手续费(RMB)": "card_fee",
    "买家申请退款(RMB)": "buyer_refund",
    "客户服务补偿(RMB)": "cs_compensation",
    "商家体验扣款(RMB)": "seller_deduction",
    "GMV广告费用(RMB)": "gmv_ad_fee",
    "平台佣金调整(RMB)": "commission_adjust",
    "促销调整(RMB)": "promo_adjust",
    "平台佣金折扣(RMB)": "commission_discount",
    "共同出资费用(RMB)": "joint_funded",
    "物流补偿(RMB)": "logistics_compensation",
    "运费调整(RMB)": "freight_adjust",
    "运费补偿(RMB)": "freight_compensation",
    "运费回扣(RMB)": "freight_rebate",
    "样品运费(RMB)": "sample_freight",
    "其他调整(RMB)": "other_adjust",
    "总调整金额(RMB)": "total_adjustment",
    "SFP服务费(RMB)": "sfp_fee",
    "LIVE Specials计划服务费(RMB)": "live_fee",
    "Bonus金币返现服务费(RMB)": "bonus_fee",
    "TikTok Shop Mall服务费(RMB)": "mall_fee",
    "Voucher Xtra计划服务费(RMB)": "voucher_fee",
    "限时抢购服务费(RMB)": "flash_sale_fee",
    "促销活动服务费(RMB)": "promo_service_fee",
    "动态服务费(RMB)": "dynamic_fee",
    "其他服务费(RMB)": "other_service_fee",
    "退款管理费(RMB)": "refund_mgmt_fee",
    "预购计划服务费(RMB)": "preorder_fee",
    "SST(RMB)": "sst",
    "GST(RMB)": "gst",
    "墨西哥增值税(RMB)": "mx_vat",
    "墨西哥联邦所得税(RMB)": "mx_income_tax",
    "反倾销税(RMB)": "anti_dumping",
    "联盟商店广告佣金(RMB)": "affiliate_ad_fee",
    "合作伙伴佣金(RMB)": "partner_commission",
    "活动运费补贴(RMB)": "activity_freight_subsidy",
    "平台运费补贴(RMB)": "platform_freight_subsidy",
    "产品SKU": "sku",
}


def analyze_file(filepath, show_headers_only=False):
    """Read file and return structured header + sample data for LLM analysis."""
    wb = load_workbook(filepath, data_only=False)
    ws = wb.active

    # Detect header row (usually row 1 or 2)
    header_row = 1
    for r in range(1, min(5, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        strs = [str(v).strip() if v is not None else '' for v in vals]
        if any(k in strs for k in ['订单编号', '订单号', '订单号']):
            header_row = r
            break

    # Read headers
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        headers.append(str(v).strip() if v is not None else f'_col_{c}')

    # Read 3 sample rows
    samples = []
    for r in range(header_row + 1, min(header_row + 4, ws.max_row + 1)):
        row_data = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            row_data.append(str(v).strip() if v is not None else '')
        # Skip empty rows
        if any(v for v in row_data):
            samples.append(row_data)

    result = {
        'file': filepath,
        'header_row': header_row,
        'total_rows': ws.max_row - header_row,
        'total_columns': ws.max_column,
        'headers': headers,
        'sample_rows': samples,
    }

    if show_headers_only:
        return result

    # Build matched analysis (which standard fields are present)
    matched = []
    unmatched = []
    for h in headers:
        if h in STANDARD_FIELDS:
            matched.append({'column': h, 'standard': STANDARD_FIELDS[h]})
        elif h not in ('', f'_col_{headers.index(h)+1}'):
            unmatched.append(h)

    result['matched_standard_fields'] = matched
    result['unmatched_columns'] = unmatched

    return result


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Analyze Excel headers for field mapping')
    parser.add_argument('filepath', help='Path to Excel file')
    parser.add_argument('--show-headers', action='store_true',
                        help='Only show headers and samples, no standard field matching')
    parser.add_argument('--json', metavar='FILE', help='Save output as JSON')
    args = parser.parse_args()

    result = analyze_file(args.filepath, show_headers_only=args.show_headers)
    output = json.dumps(result, ensure_ascii=False, indent=2)

    if args.json:
        with open(args.json, 'w', encoding='utf-8') as f:
            f.write(output)
        print(f'Saved to {args.json}')
    else:
        print(output)
