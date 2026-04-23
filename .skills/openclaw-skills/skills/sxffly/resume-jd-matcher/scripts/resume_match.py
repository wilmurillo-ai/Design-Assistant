# -*- coding: utf-8 -*-
"""
简历与 JD 智能匹配 - V2.0
支持双模式：subagent（OpenClaw） / api（外部 API）
"""

import os
import json
import yaml
import glob
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import requests
import docx

# 读取配置文件
def load_config():
    config_path = r"C:\Users\Administrator\.openclaw\workspace\config_resume_match.yaml"
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

config = load_config()

# ========== 运行模式检查 ==========
MODE = config.get('mode', 'subagent')

if MODE == 'api':
    # API 模式 - 验证配置
    if not config.get('api') or not config['api'].get('active_provider'):
        raise Exception("mode='api' 但未配置 API 信息，请检查 config_resume_match.yaml")
    
    # API 配置 - 支持多供应商切换
    def get_active_api_config():
        """获取当前激活的 API 供应商配置"""
        api_config = config['api']
        active_provider = api_config.get('active_provider', 'tencent')
        providers = api_config.get('api_providers', {})
        
        if active_provider not in providers:
            print(f"[警告] 未找到供应商 '{active_provider}'，使用默认供应商 'tencent'")
            active_provider = 'tencent'
        
        provider_config = providers[active_provider]
        return {
            'provider': active_provider,
            'name': provider_config.get('name', active_provider),
            'api_key': provider_config.get('api_key', ''),
            'api_url': provider_config.get('api_url', ''),
            'model': provider_config.get('model', ''),
            'description': provider_config.get('description', '')
        }

    API_CONFIG = get_active_api_config()
    API_KEY = API_CONFIG['api_key']
    API_URL = API_CONFIG['api_url']
    MODEL = API_CONFIG['model']
    MAX_TOKENS = config['api'].get('max_tokens', 3000)
    TIMEOUT = config['api'].get('timeout', 120)
else:
    # Subagent 模式 - 无需 API 配置
    API_CONFIG = None
    API_KEY = None
    API_URL = None
    MODEL = None
    MAX_TOKENS = None
    TIMEOUT = 60

# 路径配置
JD_FOLDER = config['paths']['jd_folder']
JL_FOLDER = config['paths']['jl_folder']
OUTPUT_FOLDER = config['paths']['output_folder']

# 简历最大字符数
MAX_CHARS = config['resume']['max_chars']  # 0 表示不限制

# 测试参数
TEST_LIMIT = 0  # 0 表示不限制，处理所有简历

# 日志配置
LOG_FOLDER = OUTPUT_FOLDER  # 日志输出到 D:\jg
LOG_FILE = None  # 在 main() 中初始化

def init_log_file():
    """初始化日志文件名（带时间戳，到分钟级别）"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    log_file = os.path.join(LOG_FOLDER, f"resume_match_{timestamp}.log")
    return log_file

def log_message(message, log_file=None):
    """写入日志文件"""
    global LOG_FILE
    if log_file:
        LOG_FILE = log_file
    if not LOG_FILE:
        LOG_FILE = init_log_file()
    
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_line = f"[{timestamp}] {message}\n"
    print(log_line.strip())  # 同时打印到控制台
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(log_line)
    except Exception as e:
        print(f"写入日志失败：{e}")

def get_latest_excel():
    """获取最新的结果文件路径"""
    excel_files = glob.glob(os.path.join(OUTPUT_FOLDER, "AI_Resume_All_*.xlsx"))
    if excel_files:
        return max(excel_files, key=os.path.getmtime)
    return None

def load_existing_results():
    """从最新 Excel 加载已有结果"""
    results = []
    latest_file = get_latest_excel()
    if not latest_file:
        return results, None
    
    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True)
        ws = wb.worksheets[0]  # 读取第一个工作表（简历匹配结果）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                result = {
                    'resume_name': row[0],
                    'position': row[1],
                    'requirement': row[2],
                    'match_level': row[3] if row[3] else '无相关信息',
                    'ai_result': row[4] if row[4] else '',
                    'process_time': row[5] if row[5] else ''
                }
                results.append(result)
        wb.close()
        print(f"    已加载 {len(results)} 条历史记录")
        return results, latest_file
    except Exception as e:
        print(f"    加载历史记录失败：{e}")
        return results, None

def get_processed_records():
    """获取已处理的记录（用于增量处理）"""
    processed = set()
    latest_file = get_latest_excel()
    if not latest_file:
        return processed
    
    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True)
        ws = wb.worksheets[0]  # 读取第一个工作表（简历匹配结果）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                # 规范化应聘者名称：去除首尾空格
                resume_name = str(row[0]).strip()
                position = str(row[1]).strip()
                requirement = str(row[2]).strip()
                key = (resume_name, position, requirement[:50])
                processed.add(key)
        wb.close()
        print(f"    已从 Excel 加载 {len(processed)} 条已处理记录")
    except Exception as e:
        print(f"    读取历史记录失败：{e}")
    
    return processed

def read_docx(file_path):
    try:
        doc = docx.Document(file_path)
        return '\n'.join([para.text for para in doc.paragraphs])
    except Exception as e:
        print(f"读取 Word 失败 {file_path}: {e}")
        return ""

def read_pdf(file_path):
    """
    PDF 解析 - 使用 pdfplumber 提取文本层
    适用于文本型 PDF（大部分电子简历）
    扫描版/图片型 PDF 需要 OCR 支持（当前不支持）
    """
    try:
        import pdfplumber
        text = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                extracted = page.extract_text() or ''
                if extracted.strip():
                    text.append(extracted)
        
        full_text = '\n'.join(text)
        
        # 检测是否为扫描版 PDF（文本层内容过少）
        if len(full_text) < 50:
            print(f"  警告：{os.path.basename(file_path)} 可能是扫描版 PDF，提取文本较少 ({len(full_text)} 字符)")
            print(f"  建议：提供文本型 PDF 或 Word 格式简历以获得更好效果")
        
        return full_text
    except Exception as e:
        print(f"PDF 解析失败 {file_path}: {e}")
        return ""

def save_as_json(file_path, text_content):
    base_dir = os.path.dirname(file_path)
    parsed_dir = os.path.join(base_dir, 'parsed')
    os.makedirs(parsed_dir, exist_ok=True)
    
    file_name = os.path.basename(file_path) + '.json'
    parsed_json = os.path.join(parsed_dir, file_name)
    
    try:
        data = {
            'file_name': os.path.basename(file_path),
            'folder': os.path.basename(base_dir),
            'full_text': text_content,
            'parse_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        with open(parsed_json, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"保存 JSON 失败：{e}")
        return False

def extract_text_from_file(file_path):
    base_dir = os.path.dirname(file_path)
    parsed_dir = os.path.join(base_dir, 'parsed')
    file_name = os.path.basename(file_path) + '.json'
    parsed_json = os.path.join(parsed_dir, file_name)
    
    if os.path.exists(parsed_json):
        try:
            with open(parsed_json, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, dict):
                    for key in ['full_text', 'text', 'content']:
                        if key in data and data[key]:
                            return data[key]
        except Exception as e:
            print(f"读取 JSON 失败：{e}")
    
    ext = os.path.splitext(file_path)[1].lower()
    text_content = ""
    if ext == '.docx':
        text_content = read_docx(file_path)
    elif ext == '.pdf':
        text_content = read_pdf(file_path)
    
    if text_content and len(text_content) > 10:
        save_as_json(file_path, text_content)
    
    return text_content

def get_jd_content(jd_folder):
    jd_files = glob.glob(os.path.join(jd_folder, '*.docx'))
    jd_contents = {}
    for jd_file in jd_files:
        if not os.path.basename(jd_file).startswith('~$'):
            jd_contents[os.path.basename(jd_file)] = extract_text_from_file(jd_file)
    return jd_contents

def find_matching_jd(folder_name, jd_contents):
    folder_lower = folder_name.lower()
    
    if 'investment' in folder_lower or '投资' in folder_name:
        for jd_name, jd_content in jd_contents.items():
            if '投资' in jd_content or 'Investment' in jd_content:
                return [(folder_name, extract_requirements(jd_content))]
    
    if '合规' in folder_name or 'compliance' in folder_lower:
        results = []
        for jd_name, jd_content in jd_contents.items():
            if '合规' in jd_content:
                domestic = extract_requirements_by_position(jd_content, '境内')
                if domestic:
                    results.append(('合规岗-境内', domestic))
                overseas = extract_requirements_by_position(jd_content, '境外')
                if overseas:
                    results.append(('合规岗-境外', overseas))
        return results if results else None
    
    return None

def extract_requirements(jd_content):
    lines = jd_content.split('\n')
    in_req = False
    requirements = []
    for line in lines:
        line = line.strip()
        if any(kw in line for kw in ['任职要求', '岗位要求', 'Qualifications']):
            in_req = True
            continue
        if in_req and line:
            if '（' in line and '）' in line and ('境内' in line or '境外' in line):
                break
            requirements.append(line)
    return requirements if requirements else [jd_content[:2000]]

def extract_requirements_by_position(jd_content, position_type):
    lines = jd_content.split('\n')
    in_position = False
    in_req = False
    requirements = []
    marker = f'合规岗（{position_type}）'
    
    for line in lines:
        line = line.strip()
        if marker in line:
            in_position = True
            in_req = False
            continue
        if in_position:
            if '任职要求' in line:
                in_req = True
                continue
            if line and '（' in line and '）' in line and '合规岗' in line:
                break
            if in_req and line and any(kw in line for kw in ['岗位职责', '投递邮箱']):
                break
            if in_req and line:
                requirements.append(line)
    
    return requirements if requirements else None

def call_subagent(resume_text, job_requirement, position_name, resume_name=""):
    """
    调用 OpenClaw 子 Agent 分析简历匹配度（mode="subagent"）
    """
    from sessions_spawn import sessions_spawn
    from subagents import subagents
    from sessions_history import sessions_history
    
    prompt = f"""请分析以下简历与该条任职要求的匹配程度：

## 简历内容：
{resume_text}

## 任职要求：
{job_requirement}

## 岗位名称：{position_name}

请针对这条任职要求进行分析，直接返回 JSON 格式：
{{"分析": "针对该条任职要求的匹配分析", "匹配等级": "完全匹配"}}

重要要求：
1. **请统一使用中文返回分析结果**
2. "分析"字段：只针对该条任职要求进行匹配分析，不要分析其他维度
3. "匹配等级"必须从以下选项中选择：完全匹配、高度匹配、匹配、部分匹配、不匹配、无相关信息
4. 即使简历或任职要求中有英文内容，分析内容也必须用中文
5. 如果简历中有英文内容（如英文公司名、英文职位等），请在分析中保留原文并附上中文说明
"""
    
    try:
        # 1. 调用子 Agent
        spawn_result = sessions_spawn(
            task=prompt,
            runtime="subagent",
            mode="run",
            timeoutSeconds=TIMEOUT,
            label=f"resume-{resume_name[:20] if resume_name else 'analysis'}"
        )
        
        child_key = spawn_result.get('childSessionKey')
        if not child_key:
            print(f"[错误] 子 Agent 调用失败：未返回 childSessionKey")
            return None
        
        # 2. 等待子 Agent 完成（轮询状态，间隔 2 秒）
        max_wait = TIMEOUT
        waited = 0
        while waited < max_wait:
            agents = subagents(action="list")
            for agent in agents.get('recent', []):
                if agent.get('sessionKey') == child_key:
                    if agent.get('status') == 'done':
                        # 3. 获取结果
                        history = sessions_history(sessionKey=child_key, limit=10, includeTools=False)
                        if history.get('messages'):
                            last_msg = history['messages'][-1]
                            if last_msg.get('role') == 'assistant':
                                content = last_msg.get('content', [])
                                if content:
                                    # 提取 text 字段
                                    for item in content:
                                        if item.get('type') == 'text':
                                            return item.get('text', '')
            time.sleep(2)
            waited += 2
        
        print(f"[警告] 子 Agent 超时（{TIMEOUT}秒）")
        return None
        
    except Exception as e:
        print(f"[错误] 子 Agent 调用异常：{e}")
        return None

def call_ai_api(resume_text, job_requirement, position_name, requirement_index, resume_name=""):
    """
    AI 分析入口 - 根据模式路由
    """
    if MODE == 'subagent':
        return call_subagent(resume_text, job_requirement, position_name, resume_name)
    else:
        return call_external_api(resume_text, job_requirement, position_name, requirement_index, resume_name)

def call_external_api(resume_text, job_requirement, position_name, requirement_index, resume_name=""):
    """
    调用外部 API 分析简历与 JD 匹配度（mode="api"）
    兼容多种 API 格式：
    - 腾讯/阿里等：使用 content 字段
    - 中国移动 Qwen3：使用 reasoning_content 字段
    """
    # 不限制字符数，发送完整简历
    prompt = f"""请分析以下简历与该条任职要求的匹配程度：

## 简历内容：
{resume_text}

## 任职要求：
{job_requirement}

## 岗位名称：{position_name}

请针对这条任职要求进行分析，直接返回 JSON 格式：
{{"分析": "针对该条任职要求的匹配分析", "匹配等级": "完全匹配"}}

重要要求：
1. **请统一使用中文返回分析结果**
2. "分析"字段：只针对该条任职要求进行匹配分析，不要分析其他维度
3. "匹配等级"必须从以下选项中选择：完全匹配、高度匹配、匹配、部分匹配、不匹配、无相关信息
4. 即使简历或任职要求中有英文内容，分析内容也必须用中文
5. 如果简历中有英文内容（如英文公司名、英文职位等），请在分析中保留原文并附上中文说明

例如：如果任职要求是学历相关，就只分析学历匹配情况；如果是工作经验相关，就只分析工作经验匹配情况。
"""

    # 构建请求头 - 支持自定义头部（中国移动需要 x-gateway-apikey）
    if API_CONFIG.get('provider') == 'cmhk':
        # 中国移动使用 x-gateway-apikey 头部
        headers = {
            "x-gateway-apikey": API_KEY,
            "Content-Type": "application/json"
        }
        # 中国移动不需要 model 参数（URL 已包含）
        data = {
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": MAX_TOKENS
        }
    else:
        # 其他 API 使用标准 Authorization 头部
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }
        data = {
            "model": MODEL,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": MAX_TOKENS
        }
    
    try:
        response = requests.post(API_URL, headers=headers, json=data, timeout=TIMEOUT)
        if response.status_code == 200:
            result = response.json()
            if 'choices' in result:
                message = result['choices'][0]['message']
                # 兼容不同 API 的返回格式
                # 优先使用 content 字段（腾讯、阿里等）
                # 如果 content 为空，使用 reasoning_content（中国移动 Qwen3）
                ai_response = message.get('content') or message.get('reasoning_content', '')
                return ai_response
            elif 'content' in result:
                content = result['content']
                if isinstance(content, list):
                    texts = [item.get('text', '') for item in content if item.get('type') == 'text']
                    return '\n'.join(texts)
                return str(content)
            return str(result)
        else:
            print(f"API 失败：{response.status_code}")
            print(f"响应：{response.text[:200]}")
            return None
    except Exception as e:
        print(f"API 异常：{e}")
        return None

def parse_ai_response(ai_response):
    try:
        cleaned = ai_response
        if '```json' in cleaned:
            cleaned = cleaned.split('```json')[1].split('```')[0]
        elif '```' in cleaned:
            cleaned = cleaned.split('```')[1].split('```')[0]
        result = json.loads(cleaned.strip())
        if '匹配等级' not in result:
            result['匹配等级'] = '无相关信息'
        if '分析' not in result:
            result['分析'] = ai_response[:500]
        return result
    except:
        levels = ['完全匹配', '高度匹配', '匹配', '部分匹配', '不匹配', '无相关信息']
        for level in levels:
            if level in ai_response:
                return {"分析": ai_response[:500], "匹配等级": level}
        return {"分析": ai_response[:500], "匹配等级": "无相关信息"}

def get_resume_files(folder_path):
    resume_files = []
    for ext in ['*.docx', '*.pdf']:
        for f in glob.glob(os.path.join(folder_path, ext)):
            if not os.path.basename(f).startswith('~$'):
                resume_files.append(f)
    # 按文件名排序，确保每次遍历顺序一致
    resume_files.sort(key=lambda x: os.path.basename(x).lower())
    return resume_files

def create_output_excel(results, output_path):
    from openpyxl.styles import Border, Side
    
    log_message(f"开始生成 Excel 文件：{output_path}")
    print(f"    正在生成 Excel...")
    
    score_map = {'完全匹配': 100, '高度匹配': 90, '匹配': 80, '部分匹配': 50, '不匹配': 0, '无相关信息': 60}
    
    resume_scores = {}
    for r in results:
        key = (r['position'], r['resume_name'])
        if key not in resume_scores:
            resume_scores[key] = []
        score = score_map.get(r.get('match_level', '无相关信息'), 60)
        resume_scores[key].append(score)
    
    candidate_summary = []
    for (position, name), scores in resume_scores.items():
        avg_score = sum(scores) / len(scores) if scores else 0
        candidate_summary.append({'position': position, 'name': name, 'score': round(avg_score, 2)})
    # 按应聘岗位升序、得分降序排序
    candidate_summary.sort(key=lambda x: (x['position'], -x['score']))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "简历匹配结果"
    
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="微软雅黑")
    normal_font = Font(name="微软雅黑")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['应聘者名称', '应聘岗位', '任职要求', '匹配度', 'AI 分析详情', '处理时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    color1 = openpyxl.styles.PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    color2 = openpyxl.styles.PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    
    prev_key = None
    use_color1 = True
    
    for row_idx, result in enumerate(results, 2):
        current_key = (result['resume_name'], result['position'])
        if prev_key is not None and current_key != prev_key:
            use_color1 = not use_color1
        prev_key = current_key
        row_fill = color1 if use_color1 else color2
        
        ai_detail = ""
        try:
            ai_data = json.loads(result['ai_result'])
            # 只显示"分析"字段
            ai_detail = ai_data.get('分析', result['ai_result'])
        except:
            ai_detail = result['ai_result']
        
        data = [result['resume_name'], result['position'], result['requirement'], result.get('match_level', '无相关信息'), ai_detail, result['process_time']]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 18
    
    for row_idx in range(2, len(results) + 2):
        ws.row_dimensions[row_idx].height = 80
    ws.freeze_panes = 'A2'
    
    ws2 = wb.create_sheet(title="应聘者总体评估表")
    for col, header in enumerate(['应聘岗位', '应聘者名称', '得分'], 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    for row_idx, candidate in enumerate(candidate_summary, 2):
        row_fill = color1 if row_idx % 2 == 0 else color2
        for col_idx, value in enumerate([candidate['position'], candidate['name'], candidate['score']], 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 15
    
    try:
        wb.save(output_path)
        print(f"结果已保存到：{output_path}")
        log_message(f"结果已保存：{output_path}")
    except Exception as e:
        print(f"[ERROR] 保存文件失败：{e}")
        log_message(f"[ERROR] 保存 Excel 失败：{e}")
        import traceback
        log_message(traceback.format_exc())
        raise

def main():
    # 初始化日志文件
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    log_file = init_log_file()
    log_message(f"日志文件：{log_file}", log_file)
    
    log_message("=" * 50, log_file)
    log_message("简历与 JD 智能匹配程序 V2.0 启动", log_file)
    log_message("=" * 50, log_file)
    print("=" * 50)
    print("简历与 JD 智能匹配程序 V2.0")
    print("=" * 50)
    
    # 显示当前运行模式
    if MODE == 'subagent':
        print(f"\n[模式] OpenClaw Subagent 模式")
        print(f"    使用用户配置的大模型（无需 API 配置）")
        log_message(f"运行模式：subagent（OpenClaw）")
    else:
        print(f"\n[模式] 外部 API 模式")
        print(f"    API 供应商：{API_CONFIG['name']} ({API_CONFIG['provider']})")
        print(f"    模型：{MODEL}")
        print(f"    说明：{API_CONFIG['description']}")
        log_message(f"运行模式：api - {API_CONFIG['name']} - {MODEL}")
        
        # 显示可用的供应商列表
        available_providers = list(config['api'].get('api_providers', {}).keys())
        print(f"    可用供应商：{', '.join(available_providers)}")
        print(f"    切换方法：修改 config_resume_match.yaml 中 api.active_provider")
    
    # 加载已有结果
    existing_results, latest_file = load_existing_results()
    if latest_file:
        log_message(f"加载已有结果：{latest_file}，共 {len(existing_results)} 条记录", log_file)
    
    # 获取已处理的记录（增量处理）
    processed_records = get_processed_records()
    log_message(f"已处理记录数：{len(processed_records)} 条", log_file)
    print(f"    已处理 {len(processed_records)} 条记录")
    
    print("\n[1] 读取 JD 文件...")
    jd_contents = get_jd_content(JD_FOLDER)
    log_message(f"找到 JD 文件：{len(jd_contents)} 个", log_file)
    print(f"    找到 {len(jd_contents)} 个 JD 文件")
    
    print("\n[2] 扫描简历目录...")
    jl_folders = [f for f in os.listdir(JL_FOLDER) if os.path.isdir(os.path.join(JL_FOLDER, f))]
    log_message(f"找到简历文件夹：{len(jl_folders)} 个", log_file)
    print(f"    找到 {len(jl_folders)} 个简历文件夹")
    
    # 统计需要处理的简历总数
    total_resumes_to_process = 0
    total_resumes_actual = 0
    
    new_results = []
    processed_resumes = set()
    
    for folder_name in jl_folders:
        if TEST_LIMIT > 0 and len(processed_resumes) >= TEST_LIMIT:
            log_message(f"已达到测试限制 {TEST_LIMIT} 份简历")
            print(f"\n已达到测试限制 {TEST_LIMIT} 份简历")
            break
        
        folder_path = os.path.join(JL_FOLDER, folder_name)
        log_message(f"处理文件夹：{folder_name}")
        print(f"\n[3] 处理文件夹：{folder_name}")
        
        jd_requirements = find_matching_jd(folder_name, jd_contents)
        if jd_requirements is None:
            log_message(f"警告：文件夹 {folder_name} 无法找到匹配的 JD")
            print("    警告：无法找到匹配的 JD")
            continue
        
        resume_files = get_resume_files(folder_path)
        log_message(f"文件夹 {folder_name}: 找到 {len(resume_files)} 份简历，匹配岗位：{[item[0] for item in jd_requirements]}", log_file)
        print(f"    找到 {len(resume_files)} 个简历文件")
        print(f"    匹配岗位：{[item[0] for item in jd_requirements]}")
        
        # 统计需要处理的简历总数
        total_resumes_to_process += len(resume_files)
        
        for resume_file in resume_files:
            if TEST_LIMIT > 0 and len(processed_resumes) >= TEST_LIMIT:
                break
            
            resume_name = os.path.basename(resume_file)
            resume_key = resume_name.replace('.docx', '').replace('.pdf', '').strip()
            
            # 检查这份简历的所有任职要求是否都已处理
            all_done = True
            for (position_name, req_list) in jd_requirements:
                for requirement in req_list:
                    record_key = (resume_key, position_name, requirement[:50])
                    if record_key not in processed_records:
                        all_done = False
                        break
                if not all_done:
                    break
            
            if all_done:
                # 统计已处理的记录数
                skip_count = 0
                for (position_name, req_list) in jd_requirements:
                    skip_count += len(req_list)
                log_message(f"跳过简历 {resume_name} - {skip_count} 条任职要求已处理", log_file)
                print(f"    简历 {resume_name} [已处理，跳过] - {skip_count} 条任职要求")
                continue  # 跳过的不计入限制
            
            # 实际处理的简历
            total_resumes_actual += 1
            
            log_message(f"处理简历：{resume_name}", log_file)
            print(f"    处理简历：{resume_name}")
            resume_content = extract_text_from_file(resume_file)
            
            if not resume_content.strip():
                log_message(f"警告：简历 {resume_name} 内容为空")
                print("    警告：简历内容为空")
                continue
            
            resume_has_new = False
            
            for (position_name, req_list) in jd_requirements:
                for req_idx, requirement in enumerate(req_list):
                    if TEST_LIMIT > 0 and len(processed_resumes) >= TEST_LIMIT:
                        break
                    
                    # 检查该条记录是否已处理
                    record_key = (resume_key, position_name, requirement[:50])
                    if record_key in processed_records:
                        print(f"      任职要求 {req_idx+1}: {requirement[:30]}... [已跳过]")
                        continue
                    
                    print(f"      {position_name} - 任职要求 {req_idx+1}: {requirement[:30]}...")
                    
                    try:
                        ai_response = call_ai_api(resume_content, requirement, position_name, req_idx+1, resume_name)
                    except Exception as e:
                        log_message(f"    [ERROR] AI 调用失败：{e}")
                        log_message(f"    跳过此条要求，继续处理")
                        ai_response = None  # 继续处理下一条
                    
                    if ai_response:
                        parsed = parse_ai_response(ai_response)
                        result = {
                            'resume_name': resume_key,
                            'position': position_name,
                            'requirement': requirement,
                            'ai_result': json.dumps(parsed, ensure_ascii=False, indent=2),
                            'process_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'match_level': parsed.get('匹配等级', '无相关信息')
                        }
                        new_results.append(result)
                        resume_has_new = True
                        log_message(f"      -> {result['match_level']}")
                        print(f"      -> {result['match_level']}")
                    else:
                        log_message("      -> AI 分析失败")
                        print("      -> AI 分析失败")
                    
                    time.sleep(1)
            
            if resume_has_new:
                processed_resumes.add(resume_key)
    
    # 合并已有结果和新结果
    if new_results:
        all_results = existing_results + new_results
        
        # 排序
        all_results.sort(key=lambda x: (x['position'], x['resume_name'], x['process_time']))
        
        # 生成输出文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(OUTPUT_FOLDER, f"AI_Resume_All_{timestamp}.xlsx")
        
        try:
            create_output_excel(all_results, output_file)
            log_message(f"匹配完成！新增 {len(new_results)} 条记录，总计 {len(all_results)} 条，处理 {len(processed_resumes)} 份简历", log_file)
            print(f"\n匹配完成！新增 {len(new_results)} 条记录，总计 {len(all_results)} 条")
            print(f"本次处理 {len(processed_resumes)} 份简历")
        except Exception as e:
            log_message(f"[ERROR] 保存 Excel 失败：{e}", log_file)
            print(f"\n[ERROR] 保存 Excel 失败：{e}")
            import traceback
            log_message(traceback.format_exc(), log_file)
    else:
        log_message("没有找到需要处理的新记录", log_file)
        print("\n没有找到需要处理的新记录")
    
    # 执行总结
    log_message("=" * 50, log_file)
    log_message("本次执行总结", log_file)
    log_message("=" * 50, log_file)
    log_message(f"需要处理简历：{total_resumes_to_process} 份", log_file)
    log_message(f"实际处理简历：{total_resumes_actual} 份", log_file)
    log_message(f"跳过简历：{total_resumes_to_process - total_resumes_actual} 份", log_file)
    log_message(f"新增记录：{len(new_results)} 条", log_file)
    log_message(f"总计记录：{len(existing_results) + len(new_results)} 条", log_file)
    log_message("=" * 50, log_file)
    log_message("程序执行完毕", log_file)
    log_message("=" * 50, log_file)

if __name__ == "__main__":
    main()
