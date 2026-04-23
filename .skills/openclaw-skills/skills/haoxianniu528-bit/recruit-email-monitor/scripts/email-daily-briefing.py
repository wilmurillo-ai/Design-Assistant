#!/usr/bin/env python3
"""
每日招聘邮件简报
每天早上 9:00 运行，汇总表格中的信息并发送
"""

import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import json

# 表格路径
EXCEL_PATH = '/home/erhao/shared/招聘邮件汇总.xlsx'

# 简报输出路径
BRIEFING_PATH = '/home/erhao/shared/招聘邮件每日简报.txt'

def load_pending_emails():
    """加载待处理的邮件（状态不是已完成的）"""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        
        emails = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 日期列
                status = row[4] if len(row) > 4 else ''
                # 待处理：状态不是"已完成"
                if status and '✅' not in status and '完成' not in status:
                    emails.append({
                        'date': row[0],
                        'account': row[1],
                        'subject': row[2],
                        'from': row[3],
                        'status': status,
                        'type': row[5] if len(row) > 5 else '',
                        'link': row[6] if len(row) > 6 else '',
                        'deadline': row[7] if len(row) > 7 else ''
                    })
        
        return emails
    
    except Exception as e:
        print(f"❌ 读取表格失败：{e}")
        return []

def generate_briefing(emails):
    """生成简报内容"""
    if not emails:
        return f"""
═══════════════════════════════════════════════════
📧 招聘邮件每日简报
日期：{datetime.now().strftime('%Y年%m月%d日 %A')}
═══════════════════════════════════════════════════

✅ 所有邮件都已处理完毕！

祝你有愉快的一天！✨
"""
    
    # 按类型分组
    by_type = defaultdict(list)
    for email in emails:
        by_type[email['type']].append(email)
    
    # 检查是否有即将截止的
    today = datetime.now()
    urgent = []
    for email in emails:
        if email.get('deadline'):
            try:
                deadline_str = str(email['deadline'])[:10]
                deadline = datetime.strptime(deadline_str, '%Y-%m-%d')
                days_left = (deadline - today).days
                if days_left <= 3:
                    urgent.append((email, days_left))
            except:
                pass
    
    # 生成简报
    briefing = f"""
═══════════════════════════════════════════════════
📧 招聘邮件每日简报
日期：{datetime.now().strftime('%Y年%m月%d日 %A')}
═══════════════════════════════════════════════════

📊 待处理概览
───────────────────────────────────────────────────
待处理邮件：{len(emails)} 封
"""
    
    # 类型统计
    for email_type, type_emails in sorted(by_type.items()):
        briefing += f"  • {email_type}: {len(type_emails)} 封\n"
    
    if urgent:
        briefing += f"\n⚠️ 即将截止：{len(urgent)} 封（3 天内）\n"
    
    briefing += f"""
───────────────────────────────────────────────────
📋 待处理详情
───────────────────────────────────────────────────
"""
    
    # 先列出紧急的
    if urgent:
        briefing += "\n🔴 即将截止（3 天内）\n"
        briefing += "─" * 40 + "\n"
        for email, days_left in urgent:
            deadline_str = str(email.get('deadline', '未知'))[:10]
            briefing += f"""
• {email['subject']}
  状态：{email['status']}
  截止：{deadline_str}（还剩{days_left}天）
  邮箱：{email['account']}
"""
            if email.get('link'):
                briefing += f"  链接：{email['link'][:100]}...\n"
    
    # 按类型列出详情
    for email_type, type_emails in sorted(by_type.items()):
        emoji = {
            '笔试/测评': '✍️',
            '面试': '🎤',
            'Offer/录用': '🎉',
            '宣讲会': '📢',
            '投递确认': '✅',
            '其他': '📧'
        }.get(email_type, '📧')
        
        briefing += f"\n{emoji} {email_type} ({len(type_emails)}封)\n"
        briefing += "─" * 40 + "\n"
        
        for i, email in enumerate(type_emails, 1):
            # 跳过已在紧急列表中的
            if any(e is email for e, _ in urgent):
                continue
            
            briefing += f"""
{i}. {email['subject']}
   状态：{email['status']}
   邮箱：{email['account']}
   发件人：{email['from']}
   时间：{email['date']}
"""
            if email.get('deadline'):
                briefing += f"   截止：{str(email['deadline'])[:10]}\n"
            if email.get('link'):
                briefing += f"   链接：{email['link']}\n"
    
    briefing += f"""
═══════════════════════════════════════════════════
💡 提醒
───────────────────────────────────────────────────
"""
    
    # 智能提醒
    reminders = []
    
    if urgent:
        reminders.append("• ⚠️ 有邮件即将截止，请优先处理！")
    
    if '笔试/测评' in by_type:
        reminders.append("• 注意查看笔试/测评邮件，通常有截止时间")
    
    if '面试' in by_type:
        reminders.append("• 面试邮件请及时回复确认")
    
    if '宣讲会' in by_type:
        reminders.append("• 宣讲会通常需要提前报名")
    
    if 'Offer/录用' in by_type:
        reminders.append("• 🎉 恭喜！Offer 邮件请仔细阅读条款")
    
    if reminders:
        briefing += "\n".join(reminders) + "\n"
    else:
        briefing += "• 保持关注，继续加油！💪\n"
    
    briefing += f"""
═══════════════════════════════════════════════════
生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
═══════════════════════════════════════════════════
"""
    
    return briefing

def send_briefing(briefing):
    """发送简报（通过 Feishu）"""
    import subprocess
    
    # 保存简报到文件
    try:
        with open(BRIEFING_PATH, 'w', encoding='utf-8') as f:
            f.write(briefing)
        print(f"✅ 简报已保存到：{BRIEFING_PATH}")
    except Exception as e:
        print(f"❌ 保存简报失败：{e}")
    
    # 打印简报内容
    print("\n" + briefing)
    
    # 通过 OpenClaw CLI 发送 Feishu 消息
    print("\n📤 正在发送 Feishu 消息...")
    try:
        # 发送到主人的 Feishu (ou_8de02604ccd510eeb4897ffd70d96c1d)
        cmd = [
            'openclaw', 'message', 'send',
            '--channel', 'feishu',
            '--target', 'user:ou_8de02604ccd510eeb4897ffd70d96c1d',
            '--message', briefing.strip()
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        
        if result.returncode == 0:
            print("✅ Feishu 消息发送成功！")
        else:
            print(f"❌ 发送失败：{result.stderr}")
    except subprocess.TimeoutExpired:
        print("❌ 发送超时")
    except Exception as e:
        print(f"❌ 发送异常：{e}")

def main():
    print('🌅 每日招聘邮件简报\n')
    print(f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print()
    
    # 加载待处理的邮件
    emails = load_pending_emails()
    
    # 生成简报
    briefing = generate_briefing(emails)
    
    # 发送简报
    send_briefing(briefing)
    
    print("\n✅ 每日简报完成")

if __name__ == '__main__':
    main()
