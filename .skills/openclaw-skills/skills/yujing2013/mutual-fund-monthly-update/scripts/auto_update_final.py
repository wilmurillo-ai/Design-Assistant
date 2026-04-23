#!/usr/bin/env python3
"""
互认基金月度数据自动更新脚本 - 完整版
支持ZIP打包上传
"""

import re
import json
import openpyxl
from pathlib import Path
import pdfplumber
import zipfile
import tempfile
import shutil


class FundUpdater:
    """基金数据更新器"""
    
    def __init__(self, skill_path):
        """初始化"""
        template_path = Path(skill_path) / 'references/extraction_templates.json'
        with open(template_path, 'r', encoding='utf-8') as f:
            self.templates = json.load(f)
        
        self.pdf_to_sheet = {
            '华夏': '华夏', '南方东英': '南方东英', '博时': '博时', '弘收': '弘收',
            '高腾': '高腾亚洲', '摩根亚洲债券': '摩根亚洲', '摩根国际': '摩根国际',
            '易方达': '易方达', '汇丰亚洲债券': '汇丰亚洲债券', '汇丰亚洲高收益': '汇丰亚洲高收益',
            '汇丰亚洲高入息': '汇丰亚洲高入息', '汇丰亚洲多元资产': '汇丰亚洲多元资产高入息',
            '海通': '海通亚洲', '东亚联丰亚洲债券': '东亚联丰亚洲债券及货币',
            '东亚联丰亚洲策略': '东亚联丰亚洲策略', '中银香港全天候中国': '中银香港全天候中国高息',
            '中银香港全天候亚洲': '中银香港全天候亚洲',
        }
    
    def process_zip(self, zip_path, output_dir):
        """处理ZIP文件"""
        # 创建临时目录
        temp_dir = tempfile.mkdtemp()
        
        try:
            # 解压ZIP
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # 查找Excel文件
            excel_files = [f for f in Path(temp_dir).rglob('*.xlsx') if not f.name.startswith('._')]
            if not excel_files:
                return None, "未找到Excel文件"
            
            excel_path = excel_files[0]
            
            # 查找PDF文件
            pdf_files = [f for f in Path(temp_dir).rglob('*.pdf') if not f.name.startswith('._')]
            if not pdf_files:
                return None, "未找到PDF文件"
            
            # 生成输出路径
            output_name = excel_path.stem + '_已更新.xlsx'
            output_path = Path(output_dir) / output_name
            
            # 执行更新
            count = self.update_excel(str(excel_path), str(temp_dir), str(output_path))
            
            return output_path, f"成功更新 {count} 个基金"
        
        finally:
            # 清理临时目录
            shutil.rmtree(temp_dir, ignore_errors=True)
    
    def update_excel(self, excel_path, pdf_dir, output_path):
        """更新Excel"""
        wb = openpyxl.load_workbook(excel_path)
        pdf_dir = Path(pdf_dir)
        
        # PDF内容到Sheet的映射
        pdf_content_to_sheet = [
            ('华夏精选固定收益', '华夏'),
            ('華夏精選固定收益', '华夏'),  # 繁体中文
            ('南方东英', '南方东英'),
            ('博時精選新興市場', '博时'),
            ('弘收', '弘收'),
            ('高騰亞洲收益', '高腾亚洲'),
            ('摩根亚洲债券', '摩根亚洲'),
            ('摩根国际', '摩根国际'),
            ('易方达', '易方达'),
            ('汇丰亚洲债券', '汇丰亚洲债券'),
            ('滙豐亞洲高收益', '汇丰亚洲高收益'),
            ('滙豐亞洲高入息', '汇丰亚洲高入息'),
            ('滙豐亞洲多元資產', '汇丰亚洲多元资产高入息'),
            ('海通亞洲', '海通亚洲'),
            ('东亚联丰亚洲债券', '东亚联丰亚洲债券及货币'),
            ('東亞聯豐亞洲策略', '东亚联丰亚洲策略'),
            ('中銀香港全天候中國', '中银香港全天候中国高息'),
            ('中銀香港全天候亞洲', '中银香港全天候亚洲'),
        ]
        
        seen = set()
        new_month_date = None  # 新的月度日期（YYYYMM格式）
        
        # 查找所有PDF文件（排除macOS隐藏文件）
        pdf_files = [f for f in pdf_dir.rglob('*.pdf') if not f.name.startswith('._')]
        
        for pdf_path in sorted(pdf_files):
            # 从PDF内容识别基金名称
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
            
            # 匹配Sheet
            sheet_name = None
            for keyword, name in pdf_content_to_sheet:
                if keyword in text:
                    sheet_name = name
                    break
            
            if not sheet_name or sheet_name in seen or sheet_name not in wb.sheetnames:
                continue
            
            seen.add(sheet_name)
            ws = wb[sheet_name]
            
            # === 0. 更新日期字段 ===
            # 提取新的月度日期
            if not new_month_date:
                match = re.search(r'(\d{2})(\d{2})', pdf_path.name)
                if match:
                    new_month_date = f"{2000 + int(match.group(1))}{int(match.group(2)):02d}"
            
            if new_month_date:
                self._update_dates(ws, new_month_date)
            
            # === 1. 数据左移 ===
            self._left_shift_data(ws)
            
            # === 2. 提取新数据 ===
            date, duration, ytm = self._extract_from_pdf(pdf_path, sheet_name, text)
            
            # === 3. 填入新数据 ===
            if duration:
                ws.cell(row=12, column=6).value = duration
            
            if ytm:
                ws.cell(row=13, column=6).value = ytm
        
        wb.save(output_path)
        wb.close()
        
        return len(seen)
    
    def _update_dates(self, ws, new_month_date):
        """更新日期字段：上月日期=原本月日期，本月日期=新月度日期"""
        # 行7列3: 上月日期
        # 行8列3: 本月日期
        
        current_month_date = ws.cell(row=8, column=3).value
        
        if current_month_date:
            # 将原本月日期更新到上月日期
            ws.cell(row=7, column=3).value = current_month_date
        
        # 将新的月度日期更新到本月日期
        ws.cell(row=8, column=3).value = new_month_date
    
    def _left_shift_data(self, ws):
        """数据左移：列6→列4（只修改蓝色数据格，不修改绿色标题格）"""
        
        # 核心指标（行12-14）- 只移动数据，不修改标题
        for row in [12, 13, 14]:
            val = ws.cell(row=row, column=6).value
            if val:
                ws.cell(row=row, column=4).value = val
        
        # 行业分布（行18-31）- 移动名称和数据
        for row in range(18, 32):
            name = ws.cell(row=row, column=5).value
            val = ws.cell(row=row, column=6).value
            if name:
                ws.cell(row=row, column=3).value = name
            if val:
                ws.cell(row=row, column=4).value = val
        
        # 区域分布（行36-49）
        for row in range(36, 50):
            name = ws.cell(row=row, column=5).value
            val = ws.cell(row=row, column=6).value
            if name:
                ws.cell(row=row, column=3).value = name
            if val:
                ws.cell(row=row, column=4).value = val
        
        # 信用评级（行52-65）
        for row in range(52, 66):
            name = ws.cell(row=row, column=5).value
            val = ws.cell(row=row, column=6).value
            if name:
                ws.cell(row=row, column=3).value = name
            if val:
                ws.cell(row=row, column=4).value = val
    
    def _extract_from_pdf(self, pdf_path, sheet_name, text=None):
        """从PDF提取数据"""
        # 提取日期
        match = re.search(r'(\d{2})(\d{2})', pdf_path.name)
        date = f"{2000 + int(match.group(1))}{int(match.group(2)):02d}" if match else None
        
        # 匹配模板
        template = next((t for k, t in self.templates.items() if k in sheet_name or sheet_name in k), None)
        if not template:
            return date, None, None
        
        # 如果没有提供text，读取PDF
        if text is None:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
        
        # 提取久期
        duration = None
        duration_config = template.get('duration', {})
        if duration_config.get('pattern'):
            match = re.search(duration_config['pattern'], text)
            if match:
                duration = float(match.group(1))
        
        # 提取YTM
        ytm = None
        ytm_config = template.get('ytm', {})
        if ytm_config.get('pattern'):
            match = re.search(ytm_config['pattern'], text, re.DOTALL)
            if match:
                ytm = float(match.group(1)) / 100
        
        return date, duration, ytm


def main():
    """主函数"""
    import sys
    
    if len(sys.argv) < 3:
        print("用法: python auto_update_final.py <zip_path> <output_dir>")
        print("   或: python auto_update_final.py <excel_path> <pdf_dir> <output_path>")
        sys.exit(1)
    
    skill_path = Path(__file__).parent.parent
    updater = FundUpdater(skill_path)
    
    if len(sys.argv) == 3:
        # ZIP模式
        output_path, message = updater.process_zip(sys.argv[1], sys.argv[2])
        print(f"{message}")
        if output_path:
            print(f"文件: {output_path}")
    else:
        # 独立文件模式
        count = updater.update_excel(sys.argv[1], sys.argv[2], sys.argv[3])
        print(f"✅ 已更新 {count} 个基金")


if __name__ == '__main__':
    main()
