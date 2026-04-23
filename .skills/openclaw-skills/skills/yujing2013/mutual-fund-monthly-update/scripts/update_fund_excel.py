#!/usr/bin/env python3
"""
互认基金月度Excel更新脚本
功能：解析PDF月报，更新Excel中对应基金Sheet的数据

依赖：
- openpyxl>=3.0.0
- pdfplumber>=0.7.0

请确保已安装依赖：pip install openpyxl pdfplumber
"""

import os
import sys
import re
import shutil
from pathlib import Path
from datetime import datetime

# 检查依赖，不自动安装
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 缺少依赖：openpyxl")
    print("请运行：pip install openpyxl")
    sys.exit(1)

try:
    import pdfplumber
except ImportError:
    print("❌ 缺少依赖：pdfplumber")
    print("请运行：pip install pdfplumber")
    sys.exit(1)

# ==================== 基金匹配规则 ====================
FUND_MAPPING = {
    "华夏": "华夏",
    "南方东英": "南方东英",
    "易方达": "易方达",
    "汇丰亚洲高收益": "汇丰亚洲高收益",
    "汇丰亚洲高入息": "汇丰亚洲多元资产高入息",
    "汇丰亚洲多元资产": "汇丰亚洲多元资产高入息",
    "汇丰亚洲债券": "汇丰亚洲债券",
    "博时": "博时",
    "弘收": "弘收",
    "高腾": "高腾",
    "海通": "海通亚洲",
    "东亚联丰亚洲债券": "东亚联丰亚洲债券及货币",
    "东亚联丰亚洲策略": "东亚联丰亚洲策略",
    "摩根亚洲债券": "摩根亚洲",
    "摩根国际": "摩根国际",
    "中银香港全天候中国": "中银香港全天候中国高息",
    "中银香港全天候亚洲": "中银香港全天候亚洲",
}

# ==================== PDF数据提取器 ====================

def extract_huaxia_pdf(pdf_path, manual_data=None):
    """
    提取华夏基金PDF数据
    
    注意：PDF中的行业/区域分布以图表形式呈现，无法通过文本提取。
    需要手动提供图表数据或从图片中提取。
    """
    data = {
        "date": None,
        "duration": None,
        "ytm": None,
        "industry": {},
        "region": {},
        "credit": {},
    }
    
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += page.extract_text() or ""
        
        # 移除换行符，合并文本
        merged_text = full_text.replace('\n', ' ').replace('\r', ' ')
        # 移除多余空格
        merged_text = re.sub(r'\s+', ' ', merged_text)
        
        # 提取截至日期
        date_match = re.search(r'截至\s*(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', merged_text)
        if date_match:
            year, month, day = date_match.groups()
            data["date"] = f"{year}.{int(month):02d}.{int(day):02d}"
        
        # 提取久期（PDF两列并排导致文字混合，直接匹配数字+Years）
        # 模式：数字.Years 或 数字 Yea (Years被截断)
        duration_match = re.search(r'([\d.]+)\s*Yea(?:rs?)?', merged_text, re.IGNORECASE)
        if duration_match:
            try:
                val = float(duration_match.group(1))
                # 验证范围（久期通常在0-30年）
                if 0 < val < 30:
                    data["duration"] = val
            except:
                pass
        
        # 提取到期收益率
        ytm_match = re.search(r'加權平均最差收益率\d*\s*([\d.]+)\s*%', merged_text)
        if ytm_match:
            data["ytm"] = float(ytm_match.group(1)) / 100
    
    # 如果提供了手动数据，合并
    if manual_data:
        if "industry" in manual_data:
            data["industry"] = manual_data["industry"]
        if "region" in manual_data:
            data["region"] = manual_data["region"]
        if "credit" in manual_data:
            data["credit"] = manual_data["credit"]
    
    return data


def parse_filename(filename):
    """解析Excel文件名，提取年月信息"""
    # 支持带UUID后缀的文件名
    # 格式：互认基金月度更新_YYYYMMvsYYYYMM---UUID.xlsx
    match = re.search(r'_(\d{6})vs(\d{6})(?:---[a-f0-9-]+)?\.xlsx$', filename)
    if match:
        last_last_month = match.group(1)
        last_month = match.group(2)
        year = int(last_month[:4])
        month = int(last_month[4:6])
        if month == 12:
            this_month = f"{year + 1}01"
        else:
            this_month = f"{year}{month + 1:02d}"
        return last_last_month, last_month, this_month
    return None, None, None


def match_pdf_to_sheet(pdf_filename):
    """根据PDF文件名匹配对应的Excel Sheet名"""
    pdf_name = Path(pdf_filename).stem
    
    for keyword, sheet_name in FUND_MAPPING.items():
        if keyword in pdf_name:
            return sheet_name
    return None


def update_sheet(ws, fund_data, last_month, this_month):
    """
    更新Sheet数据
    
    三原则：
    1. 名称 & 顺序完全沿用 Excel 原有格式
    2. 数值精准对应 PDF
    3. 本月没有的就空着
    """
    # 更新日期
    ws.cell(row=7, column=3).value = last_month
    ws.cell(row=8, column=3).value = this_month
    
    # 核心指标左移
    old_duration = ws.cell(row=12, column=6).value
    old_ytm = ws.cell(row=13, column=6).value
    
    if old_duration is not None:
        ws.cell(row=12, column=4).value = old_duration
    if old_ytm is not None:
        ws.cell(row=13, column=4).value = old_ytm
    
    # 填充新的本月数据
    if fund_data.get("duration") is not None:
        ws.cell(row=12, column=6).value = fund_data["duration"]
    if fund_data.get("ytm") is not None:
        ws.cell(row=13, column=6).value = fund_data["ytm"]
    
    # 行业分布：左移 + 按原有名称填充PDF数据
    for row_idx in range(18, 29):
        excel_name = ws.cell(row=row_idx, column=5).value
        
        # 左移
        ws.cell(row=row_idx, column=3).value = excel_name
        ws.cell(row=row_idx, column=4).value = ws.cell(row=row_idx, column=6).value
        
        # 原则2 & 3：PDF有数据就填，没有就空着
        if excel_name and excel_name in fund_data.get("industry", {}):
            ws.cell(row=row_idx, column=5).value = excel_name
            ws.cell(row=row_idx, column=6).value = fund_data["industry"][excel_name]
        elif excel_name:
            ws.cell(row=row_idx, column=5).value = excel_name
            ws.cell(row=row_idx, column=6).value = None
    
    # 区域分布：左移 + 按原有名称填充PDF数据
    for row_idx in range(32, 44):
        excel_name = ws.cell(row=row_idx, column=5).value
        
        if not excel_name:
            continue
        
        # 左移
        ws.cell(row=row_idx, column=3).value = excel_name
        ws.cell(row=row_idx, column=4).value = ws.cell(row=row_idx, column=6).value
        
        # 原则2 & 3
        if excel_name in fund_data.get("region", {}):
            ws.cell(row=row_idx, column=5).value = excel_name
            ws.cell(row=row_idx, column=6).value = fund_data["region"][excel_name]
        else:
            ws.cell(row=row_idx, column=5).value = excel_name
            ws.cell(row=row_idx, column=6).value = None
    
    return True


def main(excel_path, pdf_paths, manual_data_dict=None):
    """
    主函数
    
    Args:
        excel_path: Excel文件路径
        pdf_paths: PDF文件路径列表
        manual_data_dict: 手动数据字典 {sheet_name: data}
            用于提供PDF图表中无法提取的数据
    """
    print("=" * 60)
    print("互认基金月度Excel更新工具")
    print("=" * 60)
    
    # 解析文件名
    excel_name = Path(excel_path).name
    last_last_month, last_month, this_month = parse_filename(excel_name)
    
    if not last_month:
        print(f"❌ 无法解析Excel文件名: {excel_name}")
        return None
    
    print(f"\n📂 输入Excel: {excel_name}")
    print(f"   上上月: {last_last_month}")
    print(f"   上月: {last_month}")
    print(f"   本月: {this_month}")
    
    # 复制Excel到新文件
    output_name = f"互认基金月度更新_{last_month}vs{this_month}.xlsx"
    output_path = Path(excel_path).parent / output_name
    shutil.copy(excel_path, output_path)
    print(f"\n📄 输出Excel: {output_name}")
    
    # 打开工作簿
    wb = openpyxl.load_workbook(output_path)
    
    # 处理每个PDF
    matched_funds = []
    unmatched_pdfs = []
    
    for pdf_path in pdf_paths:
        pdf_name = Path(pdf_path).name
        sheet_name = match_pdf_to_sheet(pdf_name)
        
        if not sheet_name:
            unmatched_pdfs.append(pdf_name)
            print(f"\n⚠️ 未匹配: {pdf_name}")
            continue
        
        if sheet_name not in wb.sheetnames:
            unmatched_pdfs.append(f"{pdf_name} (Sheet不存在: {sheet_name})")
            print(f"\n⚠️ Sheet不存在: {sheet_name}")
            continue
        
        print(f"\n✅ 匹配成功: {pdf_name} → {sheet_name}")
        matched_funds.append(f"{Path(pdf_name).stem}→{sheet_name}")
        
        # 获取手动数据（如果有）
        manual_data = None
        if manual_data_dict and sheet_name in manual_data_dict:
            manual_data = manual_data_dict[sheet_name]
        
        # 提取PDF数据
        fund_data = extract_huaxia_pdf(pdf_path, manual_data)
        print(f"   提取数据:")
        print(f"     日期: {fund_data.get('date')}")
        print(f"     久期: {fund_data.get('duration')}")
        print(f"     YTM: {fund_data.get('ytm')}")
        print(f"     行业数: {len(fund_data.get('industry', {}))}")
        print(f"     区域数: {len(fund_data.get('region', {}))}")
        
        # 更新Sheet
        ws = wb[sheet_name]
        update_sheet(ws, fund_data, last_month, this_month)
        print(f"   ✅ Sheet已更新")
    
    # 保存工作簿
    wb.save(output_path)
    wb.close()
    
    # 生成报告
    print("\n" + "=" * 60)
    print("📊 更新报告")
    print("=" * 60)
    print(f"输入Excel文件名: {excel_name}")
    print(f"解析结果: 上月年月={last_month}, 本月年月={this_month}")
    print(f"输出Excel文件名: {output_name}")
    print(f"上传PDF数量: {len(pdf_paths)}个")
    print(f"匹配成功的基金: {len(matched_funds)}个")
    for i, fund in enumerate(matched_funds, 1):
        print(f"   {i}. {fund}")
    
    if unmatched_pdfs:
        print(f"未匹配的PDF: {len(unmatched_pdfs)}个")
        for pdf in unmatched_pdfs:
            print(f"   - {pdf}")
    
    print(f"\n✅ 完成！输出文件: {output_path}")
    
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python update_fund_excel.py <excel_file> <pdf_file1> [pdf_file2] ...")
        print("\n注意: PDF中的行业/区域分布图表数据无法自动提取")
        print("需要在代码中手动提供 manual_data_dict 参数")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    pdf_files = sys.argv[2:]
    
    main(excel_file, pdf_files)
