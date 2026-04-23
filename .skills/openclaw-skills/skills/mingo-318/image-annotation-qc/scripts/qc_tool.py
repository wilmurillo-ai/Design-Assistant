#!/usr/bin/env python3
"""
图像标注质检脚本 (增强版)
支持 COCO/YOLO/VOC/LabelMe 格式的标注质量检查
- 自动保存报告 (JSON/TXT)
- 可视化错误标注 (PNG)
- 多场景质检 (general/road/industrial/security)
- 自动检测格式
"""

import json
import os
import sys
import argparse
import random
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
from dataclasses import dataclass, asdict
from collections import defaultdict
from datetime import datetime

# 可视化依赖
try:
    from PIL import Image, ImageDraw, ImageFont
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# Excel 支持
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


@dataclass
class QCResult:
    total_images: int
    total_annotations: int
    error_count: int
    accuracy: float
    error_rate_by_type: Dict[str, int]
    quality_score: float
    errors: List[Dict]
    
    def to_dict(self) -> dict:
        return asdict(self)


class AnnotationQC:
    """图像标注质检器"""
    
    SUPPORTED_FORMATS = ['coco', 'yolo', 'voc', 'labelme', 'auto']
    DOMAINS = ['general', 'road', 'industrial', 'security']
    
    ERROR_WEIGHTS = {
        '漏标': 1.0,
        '错标': 0.8,
        '框偏': 0.5,
        '框大': 0.3,
        '框小': 0.5,
        '重复': 0.5,
        '标签错': 0.8,
        '模糊': 0.2
    }
    
    def __init__(self, image_dir: str, annotation_path: str, format: str = 'auto', 
                 domain: str = 'general', output_dir: str = None):
        self.image_dir = Path(image_dir)
        self.annotation_path = Path(annotation_path)
        self.format = format.lower()
        self.domain = domain.lower()
        self.errors = defaultdict(list)
        self.error_details = []
        
        # 自动创建输出目录
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            # 默认保存到标注目录的 qc_report 子目录
            if self.annotation_path.is_file():
                self.output_dir = self.annotation_path.parent / 'qc_report'
            else:
                self.output_dir = self.annotation_path / 'qc_report'
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.image_sizes = {}
        self.image_map = {}
        
    def detect_format(self) -> str:
        """自动检测标注格式"""
        ann_path = Path(self.annotation_path)
        
        # 目录优先检查
        if ann_path.is_dir():
            # LabelMe: *.json 带 shapes 字段
            json_files = list(ann_path.glob('*.json'))
            for json_file in json_files[:3]:  # 检查前3个
                try:
                    with open(json_file) as f:
                        data = json.load(f)
                        if 'shapes' in data:
                            return 'labelme'
                except:
                    pass
            
            # YOLO: *.txt 文件
            if list(ann_path.glob('*.txt')):
                return 'yolo'
            
            # VOC: *.xml 文件
            if list(ann_path.glob('*.xml')):
                return 'voc'
        
        # 文件检查
        if ann_path.is_file():
            if ann_path.suffix == '.json':
                try:
                    with open(ann_path) as f:
                        data = json.load(f)
                        if 'images' in data and 'annotations' in data:
                            return 'coco'
                        if 'shapes' in data:
                            return 'labelme'
                except:
                    pass
            elif ann_path.suffix == '.xml':
                return 'voc'
        
        return 'labelme'  # 默认（更常见）
    
    def load_image_sizes(self):
        """加载图像尺寸"""
        if not HAS_PIL:
            return
            
        for img_file in self.image_dir.glob('*'):
            if img_file.suffix.lower() in ['.jpg', '.jpeg', '.png', '.bmp']:
                try:
                    with Image.open(img_file) as img:
                        self.image_sizes[img_file.stem] = img.size
                except:
                    pass
    
    def load_annotations(self) -> Dict:
        """加载标注数据"""
        if self.format == 'auto':
            self.format = self.detect_format()
            print(f"🔍 自动检测格式: {self.format}")
        
        if self.format == 'coco':
            return self._load_coco()
        elif self.format == 'yolo':
            return self._load_yolo()
        elif self.format == 'voc':
            return self._load_voc()
        elif self.format == 'labelme':
            return self._load_labelme()
        else:
            raise ValueError(f"不支持的格式: {self.format}")
    
    def _load_coco(self) -> Dict:
        with open(self.annotation_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        self.image_map = {img['id']: img['file_name'] for img in data.get('images', [])}
        return data
    
    def _load_yolo(self) -> Dict:
        annotations = {'images': [], 'annotations': []}
        ann_id = 1
        
        for txt_file in self.annotation_path.glob('*.txt'):
            image_file = self.image_dir / (txt_file.stem + '.jpg')
            if not image_file.exists():
                image_file = self.image_dir / (txt_file.stem + '.png')
            
            annotations['images'].append({
                'id': txt_file.stem,
                'file_name': image_file.name,
                'width': 1920,
                'height': 1080
            })
            
            if image_file.exists() and HAS_PIL:
                try:
                    with Image.open(image_file) as img:
                        w, h = img.size
                        annotations['images'][-1]['width'] = w
                        annotations['images'][-1]['height'] = h
                except:
                    pass
            
            with open(txt_file, 'r') as f:
                for line in f:
                    parts = line.strip().split()
                    if len(parts) >= 5:
                        class_id, x, y, w, h = map(float, parts)
                        img_w = annotations['images'][-1]['width']
                        img_h = annotations['images'][-1]['height']
                        bbox = [x * img_w - w * img_w / 2, 
                                y * img_h - h * img_h / 2,
                                w * img_w, h * img_h]
                        
                        annotations['annotations'].append({
                            'id': ann_id,
                            'image_id': txt_file.stem,
                            'category_id': int(class_id),
                            'bbox': bbox,
                            'area': bbox[2] * bbox[3]
                        })
                        ann_id += 1
        
        self.image_map = {img['id']: img['file_name'] for img in annotations['images']}
        return annotations
    
    def _load_voc(self) -> Dict:
        import xml.etree.ElementTree as ET
        annotations = {'images': [], 'annotations': []}
        ann_id = 1
        
        for xml_file in self.annotation_path.glob('*.xml'):
            try:
                tree = ET.parse(xml_file)
                root = tree.getroot()
                
                size = root.find('size')
                w = int(size.find('width').text)
                h = int(size.find('height').text)
                
                img_id = xml_file.stem
                annotations['images'].append({
                    'id': img_id,
                    'file_name': img_id + '.jpg',
                    'width': w,
                    'height': h
                })
                
                for obj in root.findall('object'):
                    bbox = obj.find('bndbox')
                    xmin = float(bbox.find('xmin').text)
                    ymin = float(bbox.find('ymin').text)
                    xmax = float(bbox.find('xmax').text)
                    ymax = float(bbox.find('ymax').text)
                    
                    annotations['annotations'].append({
                        'id': ann_id,
                        'image_id': img_id,
                        'category_id': obj.find('name').text,
                        'bbox': [xmin, ymin, xmax - xmin, ymax - ymin],
                        'area': (xmax - xmin) * (ymax - ymin)
                    })
                    ann_id += 1
            except:
                continue
        
        self.image_map = {img['id']: img['file_name'] for img in annotations['images']}
        return annotations
    
    def _load_labelme(self) -> Dict:
        annotations = {'images': [], 'annotations': []}
        
        # 支持单个文件或目录
        json_files = []
        if self.annotation_path.is_file():
            json_files = [self.annotation_path]
        else:
            json_files = self.annotation_path.glob('*.json')
        
        for json_file in json_files:
            try:
                with open(json_file, 'r') as f:
                    data = json.load(f)
                
                image_id = json_file.stem
                img_w = data.get('imageWidth', 1920)
                img_h = data.get('imageHeight', 1080)
                
                annotations['images'].append({
                    'id': image_id,
                    'file_name': data.get('imagePath', image_id + '.png'),
                    'width': img_w,
                    'height': img_h
                })
                
                for idx, shape in enumerate(data.get('shapes', [])):
                    points = shape.get('points', [])
                    if len(points) >= 2:
                        x_coords = [p[0] for p in points]
                        y_coords = [p[1] for p in points]
                        bbox = [min(x_coords), min(y_coords), 
                                max(x_coords) - min(x_coords),
                                max(y_coords) - min(y_coords)]
                        
                        area = (max(x_coords) - min(x_coords)) * (max(y_coords) - min(y_coords))
                        
                        annotations['annotations'].append({
                            'id': len(annotations['annotations']) + 1,
                            'image_id': image_id,
                            'category_id': shape.get('label', 'unknown'),
                            'bbox': bbox,
                            'area': area,
                            'shape_type': shape.get('shape_type', 'polygon'),
                            'points': points
                        })
            except:
                continue
        
        self.image_map = {img['id']: img['file_name'] for img in annotations['images']}
        return annotations
    
    def check_annotations(self, data: Dict, sample_size: int = None) -> QCResult:
        """执行质检检查"""
        images = data.get('images', [])
        annotations = data.get('annotations', [])
        
        total_images = len(images)
        total_annotations = len(annotations)
        
        self.load_image_sizes()
        
        if sample_size and sample_size < total_images:
            images = random.sample(images, sample_size)
        
        self.errors = defaultdict(list)
        self.error_details = []
        
        self._check_bbox_validity(annotations, images)
        self._check_polygon_validity(annotations, images)
        self._check_label_validity(annotations, images)
        self._check_duplicates(annotations)
        self._check_category_consistency(annotations)
        self._check_domain_rules(annotations, images)
        
        error_count = sum(len(errs) for errs in self.errors.values())
        
        if total_annotations > 0:
            accuracy = ((total_annotations - error_count) / total_annotations) * 100
            error_rate_by_type = {k: len(v) for k, v in self.errors.items()}
            
            total_weight = sum(
                len(v) * self.ERROR_WEIGHTS.get(k, 0.5) 
                for k, v in self.errors.items()
            )
            quality_score = max(0, 100 - (total_weight / total_annotations * 100))
        else:
            accuracy = 100.0
            error_rate_by_type = {}
            quality_score = 100.0
        
        return QCResult(
            total_images=total_images,
            total_annotations=total_annotations,
            error_count=error_count,
            accuracy=round(accuracy, 2),
            error_rate_by_type=error_rate_by_type,
            quality_score=round(quality_score, 2),
            errors=self.error_details
        )
    
    def _check_bbox_validity(self, annotations: List[Dict], images: List[Dict]):
        """检查边界框有效性"""
        img_sizes = {img['id']: (img.get('width', 2560), img.get('height', 1440)) 
                     for img in images}
        
        for ann in annotations:
            bbox = ann.get('bbox', [])
            if len(bbox) < 4:
                self._add_error('错标', ann, '边界框不完整')
                continue
            
            x, y, w, h = bbox
            img_id = ann.get('image_id')
            img_w, img_h = img_sizes.get(img_id, (2560, 1440))
            
            if self.format == 'yolo':
                if w <= 0 or h <= 0 or w > 1 or h > 1:
                    self._add_error('框小', ann, '归一化坐标超出范围 [0,1]')
            else:
                if w <= 0 or h <= 0:
                    self._add_error('框小', ann, f'宽度或高度为负: {w}x{h}')
                elif w > img_w * 1.5 or h > img_h * 1.5:
                    self._add_error('框大', ann, f'框过大超过图像: {w}x{h} > {img_w}x{img_h}')
                
                if x < 0 or y < 0:
                    self._add_error('框偏', ann, f'坐标为负: x={x}, y={y}')
                
                if x + w > img_w * 1.1 or y + h > img_h * 1.1:
                    self._add_error('框大', ann, f'框超出图像边界')
    
    def _check_polygon_validity(self, annotations: List[Dict], images: List[Dict]):
        """检查分割标注（polygon）的有效性"""
        img_sizes = {img['id']: (img.get('width', 2560), img.get('height', 1440)) 
                     for img in images}
        
        for ann in annotations:
            points = ann.get('points', [])
            if not points:
                continue
            
            img_id = ann.get('image_id')
            img_w, img_h = img_sizes.get(img_id, (2560, 1440))
            
            # 1. 顶点数量检查
            if len(points) < 3:
                self._add_error('错标', ann, f'多边形顶点数不足: {len(points)} < 3')
                continue
            
            # 2. 坐标范围检查
            x_coords = [p[0] for p in points]
            y_coords = [p[1] for p in points]
            
            min_x, max_x = min(x_coords), max(x_coords)
            min_y, max_y = min(y_coords), max(y_coords)
            
            # 检查负坐标
            if min_x < 0 or min_y < 0:
                self._add_error('框偏', ann, f'多边形坐标为负: x={min_x}, y={min_y}')
            
            # 检查超出图像边界
            if max_x > img_w * 1.1 or max_y > img_h * 1.1:
                self._add_error('框大', ann, f'多边形超出图像边界')
            
            # 3. 面积合理性检查
            area = self._polygon_area(points)
            if area <= 0:
                self._add_error('错标', ann, '多边形面积为零（可能自相交）')
            elif area < 100:  # 最小面积阈值
                self._add_error('框小', ann, f'多边形面积过小: {area:.1f}px²')
            elif area > img_w * img_h * 0.95:
                self._add_error('框大', ann, f'多边形面积过大: {area/img_w/img_h*100:.1f}%')
            
            # 4. 检查点是否共线（退化多边形）
            if len(points) >= 3:
                if self._is_degenerate_polygon(points):
                    self._add_error('错标', ann, '多边形顶点共线，无有效区域')
    
    def _polygon_area(self, points: List) -> float:
        """计算多边形面积（ Shoelace 算法）"""
        if len(points) < 3:
            return 0
        
        n = len(points)
        area = 0
        for i in range(n):
            j = (i + 1) % n
            area += points[i][0] * points[j][1]
            area -= points[j][0] * points[i][1]
        
        return abs(area) / 2
    
    def _is_degenerate_polygon(self, points: List, threshold: float = 1.0) -> bool:
        """检查多边形是否退化（所有点近似共线）"""
        if len(points) < 3:
            return True
        
        # 计算所有点到平均线的距离
        n = len(points)
        sum_x = sum(p[0] for p in points)
        sum_y = sum(p[1] for p in points)
        avg_x = sum_x / n
        avg_y = sum_y / n
        
        # 计算斜率
        dx = points[-1][0] - points[0][0]
        dy = points[-1][1] - points[0][1]
        
        if abs(dx) < 0.01 and abs(dy) < 0.01:
            return True
        
        # 计算每个点到直线的距离
        max_dist = 0
        for i in range(n - 1):
            # 点到线段的距离
            x0, y0 = points[i]
            x1, y1 = points[(i + 1) % n]
            
            # 线段长度
            line_len = ((x1 - x0)**2 + (y1 - y0)**2) ** 0.5
            if line_len < 0.01:
                continue
            
            # 点到线段距离
            t = max(0, min(1, ((avg_x - x0) * (x1 - x0) + (avg_y - y0) * (y1 - y0)) / (line_len ** 2)))
            proj_x = x0 + t * (x1 - x0)
            proj_y = y0 + t * (y1 - y0)
            dist = ((avg_x - proj_x)**2 + (avg_y - proj_y)**2) ** 0.5
            max_dist = max(max_dist, dist)
        
        return max_dist < threshold
    
    def _check_duplicates(self, annotations: List[Dict]):
        """检查重复标注"""
        seen = {}
        for ann in annotations:
            img_id = ann.get('image_id')
            cat_id = ann.get('category_id')
            bbox = ann.get('bbox', [])
            
            if not bbox:
                continue
            
            key = (img_id, cat_id)
            if key in seen:
                if self._iou(seen[key]['bbox'], bbox) > 0.8:
                    self._add_error('重复', ann, '与已有标注高度重叠')
            else:
                seen[key] = ann
    
    def _iou(self, bbox1, bbox2) -> float:
        """计算 IoU"""
        x1 = max(bbox1[0], bbox2[0])
        y1 = max(bbox1[1], bbox2[1])
        x2 = min(bbox1[0] + bbox1[2], bbox2[0] + bbox2[2])
        y2 = min(bbox1[1] + bbox1[3], bbox2[1] + bbox2[3])
        
        if x2 <= x1 or y2 <= y1:
            return 0
        
        inter = (x2 - x1) * (y2 - y1)
        area1 = bbox1[2] * bbox1[3]
        area2 = bbox2[2] * bbox2[3]
        union = area1 + area2 - inter
        
        return inter / union if union > 0 else 0
    
    def _check_category_consistency(self, annotations: List[Dict]):
        """检查类别一致性"""
        category_map = defaultdict(list)
        for ann in annotations:
            cat_id = ann.get('category_id')
            category_map[cat_id].append(ann.get('image_id'))
    
    def _check_label_validity(self, annotations: List[Dict], images: List[Dict]):
        """检查标签有效性"""
        # 收集所有标签
        all_labels = set()
        label_stats = defaultdict(list)
        
        for ann in annotations:
            label = ann.get('category_id', '')
            
            # 处理数字类型标签（YOLO 用数字）
            if isinstance(label, int):
                all_labels.add(f"class_{label}")
                continue
            
            if not label:
                self._add_error('标签错', ann, '标签为空')
                continue
            
            label_str = str(label).strip()
            all_labels.add(label_str)
            label_stats[label_str].append(ann.get('image_id'))
            
            # 检查标签格式（只能包含字母、数字、下划线、连字符）
            import re
            if not re.match(r'^[\w\-]+$', label_str):
                self._add_error('标签错', ann, f'标签包含非法字符: {label_str}')
            
            # 检查标签长度
            if len(label_str) > 50:
                self._add_error('标签错', ann, f'标签过长: {len(label_str)} > 50')
        
        # 标签统计摘要
        if all_labels:
            print(f"   标签类别: {', '.join(sorted(all_labels))}")
    
    def _check_domain_rules(self, annotations: List[Dict], images: List[Dict]):
        """根据场景执行特定规则"""
        img_sizes = {img['id']: (img.get('width', 1920), img.get('height', 1080)) 
                     for img in images}
        
        for ann in annotations:
            img_id = ann.get('image_id')
            img_w, img_h = img_sizes.get(img_id, (1920, 1080))
            bbox = ann.get('bbox', [])
            w, h = bbox[2], bbox[3]
            
            if self.domain == 'road':
                # 交通场景：标记小目标
                if w < 32 or h < 32:
                    pass  # 小目标不计入错误，仅提示
                    
            elif self.domain == 'industrial':
                # 工业场景：检查微小缺陷
                area_ratio = (w * h) / (img_w * img_h)
                if area_ratio < 0.001:
                    self._add_error('框小', ann, '标注区域过小，可能是微小缺陷')
    
    def _add_error(self, error_type: str, ann: Dict, message: str):
        """添加错误记录"""
        self.errors[error_type].append(ann)
        self.error_details.append({
            'type': error_type,
            'message': message,
            'annotation': ann
        })
    
    def visualize_errors(self, result: QCResult):
        """生成可视化错误标注图片"""
        if not HAS_PIL:
            print("⚠️ 需要安装 Pillow 才能生成可视化图片: pip install pillow")
            return
        
        if not result.errors:
            print("✅ 没有错误需要可视化")
            return
        
        vis_dir = self.output_dir / 'visual'
        vis_dir.mkdir(exist_ok=True)
        
        # 按图像分组错误
        errors_by_image = defaultdict(list)
        for err in result.errors:
            img_id = err['annotation'].get('image_id')
            errors_by_image[img_id].append(err)
        
        # 加载标注数据
        annotations = self.load_annotations()
        anns_by_image = defaultdict(list)
        for ann in annotations.get('annotations', []):
            anns_by_image[ann.get('image_id')].append(ann)
        
        # 绘制每个有错误的图像
        for img_id, errors in errors_by_image.items():
            img_info = next((img for img in annotations.get('images', []) if img['id'] == img_id), None)
            if not img_info:
                continue
            
            img_path = self.image_dir / img_info['file_name']
            if not img_path.exists():
                continue
            
            try:
                with Image.open(img_path) as img:
                    draw = ImageDraw.Draw(img)
                    
                    # 绘制所有标注（绿色）
                    for ann in anns_by_image.get(img_id, []):
                        bbox = ann.get('bbox', [])
                        if bbox and len(bbox) >= 4:
                            x, y, w, h = bbox[:4]
                            if w > 0 and h > 0:  # 确保有效
                                draw.rectangle([x, y, x+w, y+h], outline='green', width=2)
                    
                    # 绘制错误标注（红色）
                    for err in errors:
                        bbox = err['annotation'].get('bbox', [])
                        if bbox and len(bbox) >= 4:
                            x, y, w, h = bbox[:4]
                            if w > 0 and h > 0:  # 确保有效
                                draw.rectangle([x, y, x+w, y+h], outline='red', width=4)
                    
                    output_path = vis_dir / f"{img_id}_qc.png"
                    img.save(output_path)
                    print(f"📊 已生成可视化: {output_path}")
            except Exception as e:
                print(f"⚠️ 可视化失败 {img_id}: {e}")
    
    def generate_report(self, result: QCResult, output_path: str = None, 
                       formats: List[str] = None) -> Dict[str, str]:
        """生成质检报告（支持多种格式）"""
        if formats is None:
            formats = ['txt', 'json']
        
        outputs = {}
        
        # 1. TXT 报告
        if 'txt' in formats:
            txt_report = self._generate_txt_report(result)
            txt_path = output_path or str(self.output_dir / 'qc_report.txt')
            with open(txt_path, 'w', encoding='utf-8') as f:
                f.write(txt_report)
            outputs['txt'] = txt_path
            print(f"📄 TXT报告已保存: {txt_path}")
        
        # 2. JSON 报告（详细错误列表）
        if 'json' in formats:
            json_path = str(self.output_dir / 'qc_report.json')
            report_data = {
                'summary': {
                    'total_images': result.total_images,
                    'total_annotations': result.total_annotations,
                    'error_count': result.error_count,
                    'accuracy': result.accuracy,
                    'quality_score': result.quality_score,
                    'error_rate_by_type': result.error_rate_by_type,
                    'timestamp': datetime.now().isoformat()
                },
                'errors': result.errors
            }
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, ensure_ascii=False, indent=2)
            outputs['json'] = json_path
            print(f"📄 JSON报告已保存: {json_path}")
        
        # 3. Excel 报告（如果可用）
        if 'xlsx' in formats and HAS_XLSX:
            xlsx_path = str(self.output_dir / 'qc_report.xlsx')
            self._generate_xlsx_report(result, xlsx_path)
            outputs['xlsx'] = xlsx_path
            print(f"📄 Excel报告已保存: {xlsx_path}")
        
        return outputs
    
    def _generate_txt_report(self, result: QCResult) -> str:
        """生成文本报告"""
        quality_level = "优秀" if result.quality_score >= 90 else \
                       "良好" if result.quality_score >= 80 else \
                       "合格" if result.quality_score >= 70 else "不合格"
        
        report = f"""
{'='*50}
        图像标注质检报告
{'='*50}

📊 基本统计
{'-'*30}
总图像数: {result.total_images}
总标注数: {result.total_annotations}
错误总数: {result.error_count}

📈 质量指标
{'-'*30}
准确率: {result.accuracy}%
质量分: {result.quality_score} / 100

📋 错误分类统计
{'-'*30}
"""
        if result.error_rate_by_type:
            for error_type, count in sorted(result.error_rate_by_type.items()):
                report += f"  {error_type}: {count} 个\n"
        else:
            report += "  无错误\n"
        
        # 详细错误列表
        if result.errors:
            report += f"""
⚠️ 错误详情
{'-'*30}
"""
            for i, err in enumerate(result.errors[:20], 1):  # 最多显示20条
                ann = err['annotation']
                img_id = ann.get('image_id', 'unknown')
                report += f"  {i}. [{err['type']}] {img_id} - {err['message']}\n"
            
            if len(result.errors) > 20:
                report += f"  ... 还有 {len(result.errors) - 20} 个错误\n"
        
        report += f"""
{'='*50}
质量等级: {quality_level}
{'='*50}
报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
        return report
    
    def _generate_xlsx_report(self, result: QCResult, xlsx_path: str):
        """生成 Excel 报告"""
        wb = openpyxl.Workbook()
        
        # Sheet 1: 汇总
        ws_summary = wb.active
        ws_summary.title = "汇总"
        
        headers = ["指标", "数值"]
        for col, header in enumerate(headers, 1):
            ws_summary.cell(1, col, header)
        
        data = [
            ("总图像数", result.total_images),
            ("总标注数", result.total_annotations),
            ("错误总数", result.error_count),
            ("准确率", f"{result.accuracy}%"),
            ("质量分", result.quality_score),
        ]
        for row, (key, value) in enumerate(data, 2):
            ws_summary.cell(row, 1, key)
            ws_summary.cell(row, 2, value)
        
        # Sheet 2: 错误详情
        ws_errors = wb.create_sheet("错误详情")
        err_headers = ["序号", "错误类型", "图像ID", "标注ID", "类别", "错误描述"]
        for col, header in enumerate(err_headers, 1):
            ws_errors.cell(1, col, header)
        
        for row, err in enumerate(result.errors, 2):
            ann = err['annotation']
            ws_errors.cell(row, 1, row - 1)
            ws_errors.cell(row, 2, err['type'])
            ws_errors.cell(row, 3, ann.get('image_id', ''))
            ws_errors.cell(row, 4, ann.get('id', ''))
            ws_errors.cell(row, 5, str(ann.get('category_id', '')))
            ws_errors.cell(row, 6, err['message'])
        
        wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(
        description='图像标注质检工具 - 支持 COCO/YOLO/VOC/LabelMe 格式',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 自动检测格式
  python3 qc_tool.py -i ./images -a ./annotations
  
  # 指定格式
  python3 qc_tool.py -i ./images -a ./annotations.json -f coco
  
  # 指定场景（工业质检）
  python3 qc_tool.py -i ./images -a ./annotations -f labelme -d industrial
  
  # 抽样检查
  python3 qc_tool.py -i ./images -a ./annotations -s 100
  
  # 指定输出目录
  python3 qc_tool.py -i ./images -a ./annotations -o ./my_report
        """
    )
    
    parser.add_argument('--image-dir', '-i', required=True, 
                       help='图像目录路径')
    parser.add_argument('--annotation', '-a', required=True, 
                       help='标注文件路径（JSON/XML）或目录（YOLO/LabelMe）')
    parser.add_argument('--format', '-f', 
                       choices=AnnotationQC.SUPPORTED_FORMATS,
                       default='auto', 
                       help='标注格式 (默认: auto 自动检测)')
    parser.add_argument('--domain', '-d',
                       choices=AnnotationQC.DOMAINS,
                       default='general',
                       help='应用场景 (默认: general)')
    parser.add_argument('--sample', '-s', type=int, 
                       help='抽样检查数量（可选）')
    parser.add_argument('--output', '-o', 
                       help='报告输出目录 (默认: 标注目录/qc_report)')
    parser.add_argument('--formats',
                       nargs='+',
                       choices=['txt', 'json', 'xlsx'],
                       default=['txt', 'json'],
                       help='报告格式 (默认: txt json)')
    parser.add_argument('--no-visual', action='store_true',
                       help='禁止生成可视化图片')
    
    args = parser.parse_args()
    
    print(f"""
╔{'='*50}╗
║       🔍 图像标注质检工具 v2.0                ║
╚{'='*50}╝
    """)
    
    print(f"📂 图像目录: {args.image_dir}")
    print(f"📂 标注路径: {args.annotation}")
    print(f"🏷️  格式: {args.format} | 场景: {args.domain}")
    print()
    
    # 创建质检器
    qc = AnnotationQC(
        args.image_dir, 
        args.annotation, 
        format=args.format,
        domain=args.domain,
        output_dir=args.output
    )
    
    # 加载标注
    print("📂 加载标注数据...")
    data = qc.load_annotations()
    print(f"   加载完成: {len(data.get('images', []))} 图像, {len(data.get('annotations', []))} 标注")
    
    # 执行质检
    print("✅ 开始质检...")
    result = qc.check_annotations(data, args.sample)
    
    # 生成报告
    print("📄 生成报告...")
    outputs = qc.generate_report(result, formats=args.formats)
    
    # 生成可视化
    if not args.no_visual and HAS_PIL:
        print("🎨 生成可视化图片...")
        qc.visualize_errors(result)
    
    # 输出摘要
    print(f"""
╔{'='*50}╗
║                    质检结果                    ║
╠{'='*50}╣
║  图像数: {result.total_images:<6}  标注数: {result.total_annotations:<6}    ║
║  错误数: {result.error_count:<6}  质量分: {result.quality_score:<6}    ║
║  准确率: {result.accuracy:<6}%                         ║
╚{'='*50}╝
    """)
    
    # 返回状态码
    sys.exit(0 if result.quality_score >= 70 else 1)


if __name__ == '__main__':
    main()
