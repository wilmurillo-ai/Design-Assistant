#!/usr/bin/env python3
"""Excel 创建工具 - openpyxl 支持图表"""
import argparse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import datetime


def style_header(ws, headers, header_fill='4472C4', font_color='FFFFFF'):
    header_font = Font(name='微软雅黑', bold=True, color=font_color, size=11)
    fill = PatternFill(start_color=header_fill, end_color=header_fill, fill_type='solid')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def style_data(ws, rows, start_row=2):
    data_font = Font(name='微软雅黑', size=10)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ri, row in enumerate(rows, start_row):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def add_bar_chart(ws, title, data_range, categories_range):
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = '数值'
    chart.x_axis.title = '类别'
    data = Reference(ws, range_string=data_range)
    cats = Reference(ws, range_string=categories_range)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "E5")


def create_workbook(title, headers, rows, output_path, charts=None):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    # 写入标题行
    ws.append(headers)
    style_header(ws, headers)
    # 写入数据
    for row in rows:
        ws.append(row)
    style_data(ws, rows)
    auto_width(ws)
    # 添加图表
    if charts:
        for chart in charts:
            if chart['type'] == 'bar':
                add_bar_chart(ws, chart['title'], chart['data'], chart['categories'])
    wb.save(output_path)
    print(f"[OK] Excel 已保存: {output_path}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='创建 Excel 工作簿')
    parser.add_argument('--title', '-t', default='工作表1', help='工作表标题')
    parser.add_argument('--headers', default='[]', help='表头 JSON')
    parser.add_argument('--rows', default='[]', help='数据行 JSON')
    parser.add_argument('--output', '-o', default='output.xlsx', help='输出路径')
    args = parser.parse_args()
    import json
    create_workbook(args.title, json.loads(args.headers), json.loads(args.rows), args.output)
