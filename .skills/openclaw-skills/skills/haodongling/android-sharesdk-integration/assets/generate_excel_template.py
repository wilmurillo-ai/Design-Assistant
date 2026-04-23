#!/usr/bin/env python3
"""
生成 ShareSDK 配置模板 Excel 文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_sheet_with_header(wb, sheet_name, headers, data):
    """创建带表头的 Sheet"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40

    # 冻结首行
    ws.freeze_panes = 'A2'

    return ws


def create_platform_urls_sheet(wb):
    """创建平台申请地址 Sheet"""
    headers = ["平台", "开放平台地址", "说明"]
    data = [
        ["微信/微信朋友圈/微信收藏", "https://open.weixin.qq.com", "微信开放平台，需企业资质"],
        ["QQ / QZone", "https://connect.qq.com", "QQ互联平台"],
        ["新浪微博", "https://open.weibo.com", "微博开放平台"],
        ["Facebook", "https://developers.facebook.com", "Facebook for Developers"],
        ["Twitter / X", "https://developer.twitter.com", "Twitter Developer Portal"],
        ["Instagram", "https://developers.instagram.com", "Instagram Basic Display API"],
        ["Google+", "https://console.cloud.google.com/apis/credentials", "Google Cloud Console"],
        ["LinkedIn", "https://www.linkedin.com/developers/apps", "LinkedIn Developer Portal"],
        ["Line", "https://developers.line.biz/console/", "LINE Developers Console"],
        ["KakaoTalk / KakaoStory", "https://developers.kakao.com/console/app", "Kakao Developers"],
        ["支付宝", "https://open.alipay.com/platform/home.htm", "支付宝开放平台"],
        ["MobTech", "https://www.mob.com/", "用于获取 appKey/appSecret"],
    ]
    return create_sheet_with_header(wb, "平台申请地址", headers, data)


def create_basic_info_sheet(wb):
    """创建基础信息 Sheet"""
    headers = ["配置项", "说明", "您的信息（必填）"]
    data = [
        ["appKey", "MobSDK应用Key，从MobTech官网(https://www.mob.com/)注册应用获取", ""],
        ["appSecret", "MobSDK应用密钥，与appKey一同获取", ""],
    ]
    ws = create_sheet_with_header(wb, "基础信息", headers, data)

    # 标记必填项
    for row in range(2, 4):
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True, color="FF0000")

    return ws


def create_wechat_sheet(wb):
    """创建微信系列 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["Wechat.appId", "微信应用ID，从微信开放平台获取", ""],
        ["Wechat.appSecret", "微信应用密钥", ""],
        ["Wechat.userName", "微信小程序原始ID（分享小程序时需要，格式：gh_xxxxxxxxxxx）", ""],
        ["Wechat.path", "小程序页面路径（如：pages/index/index.html）", ""],
        ["Wechat.withShareTicket", "是否带shareTicket（true/false）", ""],
        ["Wechat.miniprogramType", "小程序类型：0正式版、1开发版、2体验版", ""],
        ["Wechat.bypassApproval", "是否绕过审核（仅分享文字图片时可用，默认false）", ""],
        ["", "", ""],
        ["WechatMoments.appId", "微信朋友圈：与微信相同appId", ""],
        ["WechatMoments.appSecret", "微信朋友圈：与微信相同appSecret", ""],
        ["WechatMoments.bypassApproval", "是否绕过审核（默认false）", ""],
        ["", "", ""],
        ["WechatFavorite.appId", "微信收藏：与微信相同appId", ""],
        ["WechatFavorite.appSecret", "微信收藏：与微信相同appSecret", ""],
    ]
    return create_sheet_with_header(wb, "微信系列", headers, data)


def create_qq_sheet(wb):
    """创建QQ系列 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["QQ.appId", "QQ互联应用ID", ""],
        ["QQ.appKey", "QQ互联应用Key", ""],
        ["QQ.shareByAppClient", "是否使用客户端分享（默认true）", ""],
        ["QQ.bypassApproval", "是否绕过审核（默认false）", ""],
        ["", "", ""],
        ["QZone.appId", "QQ空间：与QQ相同appId", ""],
        ["QZone.appKey", "QQ空间：与QQ相同appKey", ""],
        ["QZone.shareByAppClient", "是否使用客户端分享（默认true）", ""],
        ["QZone.bypassApproval", "是否绕过审核（默认false）", ""],
    ]
    return create_sheet_with_header(wb, "QQ系列", headers, data)


def create_weibo_sheet(wb):
    """创建微博 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["SinaWeibo.appKey", "新浪微博 App Key", ""],
        ["SinaWeibo.appSecret", "新浪微博 App Secret", ""],
        ["SinaWeibo.redirectUri", "回调地址（默认：https://api.weibo.com/oauth2/default.html）", "https://api.weibo.com/oauth2/default.html"],
        ["SinaWeibo.shareByAppClient", "是否使用客户端分享（默认true）", ""],
    ]
    return create_sheet_with_header(wb, "微博", headers, data)


def create_overseas_sheet(wb):
    """创建海外平台 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["Facebook.appKey", "Facebook App ID", ""],
        ["Facebook.appSecret", "Facebook App Secret", ""],
        ["Facebook.callbackUri", "回调地址", ""],
        ["Facebook.faceBookLoginProtocolScheme", "登录协议Scheme", ""],
        ["Facebook.faceBookAppType", "App类型（默认game）", ""],
        ["Facebook.shareByAppClient", "是否使用客户端分享", ""],
        ["", "", ""],
        ["Twitter.appKey", "Twitter App Key", ""],
        ["Twitter.appSecret", "Twitter App Secret", ""],
        ["Twitter.callbackUri", "回调地址", ""],
        ["Twitter.IsUseV2", "是否使用API v2（默认true）", ""],
        ["Twitter.shareByAppClient", "是否使用客户端分享", ""],
        ["", "", ""],
        ["Instagram.appId", "Instagram App ID", ""],
        ["Instagram.appSecret", "Instagram App Secret", ""],
        ["Instagram.callbackUri", "回调地址", ""],
        ["", "", ""],
        ["GooglePlus.appId", "Google Client ID", ""],
        ["GooglePlus.appSecret", "Google Client Secret", ""],
        ["GooglePlus.callbackUri", "回调地址", ""],
        ["GooglePlus.shareByAppClient", "是否使用客户端分享", ""],
    ]
    return create_sheet_with_header(wb, "海外平台", headers, data)


def create_other_sheet(wb):
    """创建其他平台 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["LinkedIn.appKey", "LinkedIn App Key", ""],
        ["LinkedIn.appSecret", "LinkedIn App Secret", ""],
        ["LinkedIn.callbackUri", "回调地址", ""],
        ["LinkedIn.shareByAppClient", "是否使用客户端分享", ""],
        ["", "", ""],
        ["Line.appId", "Line App ID", ""],
        ["Line.appSecret", "Line App Secret", ""],
        ["Line.callbackUri", "回调地址", ""],
        ["Line.callbackscheme", "回调协议", ""],
        ["", "", ""],
        ["KakaoTalk.appKey", "KakaoTalk App Key", ""],
        ["KakaoTalk.shareByAppClient", "是否使用客户端分享", ""],
        ["", "", ""],
        ["KakaoStory.appKey", "KakaoStory App Key", ""],
        ["", "", ""],
        ["Alipay.appId", "支付宝App ID", ""],
        ["", "", ""],
        ["AlipayMoments.appId", "支付宝朋友圈：与支付宝相同appId", ""],
        ["", "", ""],
        ["短信", "无需配置，直接可用", "-"],
        ["邮件", "无需配置，直接可用", "-"],
        ["WhatsApp", "无需配置，直接可用", "-"],
        ["复制链接", "无需配置，直接可用", "-"],
    ]
    return create_sheet_with_header(wb, "其他平台", headers, data)


def create_instructions_sheet(wb):
    """创建填写说明 Sheet"""
    headers = ["说明项", "详细内容"]
    data = [
        ["填写步骤", "1. 先在MobTech官网(https://www.mob.com/)注册应用，获取appKey和appSecret\n2. 根据需要的分享平台，前往对应开放平台申请应用（地址见'平台申请地址'Sheet）\n3. 获取各平台的appId/appKey等参数，填写到对应Sheet中\n4. 不需要的平台可以留空，Agent会自动忽略"],
        ["必填项", "基础信息Sheet中的appKey和appSecret是必填项，否则ShareSDK无法正常工作"],
        ["微信配置", "微信、微信朋友圈、微信收藏使用相同的appId和appSecret\n如需分享小程序，需要额外配置userName、path等参数"],
        ["QQ配置", "QQ和QQ空间使用相同的appId和appKey"],
        ["分享类型", "文字/图片分享：所有平台都支持\n链接分享：微信、QQ、微博等支持\n小程序分享：仅微信支持"],
        ["常见问题", "微信分享失败：检查签名MD5是否与微信开放平台配置一致\nQQ分享失败：检查是否在QQ互联配置了正确的包名和签名"],
    ]

    ws = wb.create_sheet(title="填写说明")

    # 设置表头
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 设置列宽和行高
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 60

    return ws


def main():
    wb = Workbook()

    # 删除默认创建的 Sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 按顺序创建各个 Sheet
    create_basic_info_sheet(wb)
    create_wechat_sheet(wb)
    create_qq_sheet(wb)
    create_weibo_sheet(wb)
    create_overseas_sheet(wb)
    create_other_sheet(wb)
    create_platform_urls_sheet(wb)
    create_instructions_sheet(wb)

    # 调整 Sheet 顺序（基础信息放第一个）
    sheets = wb.sheetnames
    target_order = [
        "基础信息",
        "平台申请地址",
        "填写说明",
        "微信系列",
        "QQ系列",
        "微博",
        "海外平台",
        "其他平台",
    ]

    for idx, name in enumerate(target_order):
        if name in sheets:
            wb.move_sheet(name, offset=-wb.sheetnames.index(name) + idx)

    # 保存文件
    output_path = "/Users/haodongling/.openclaw/workspace/skills/android-sharesdk-integration/assets/ShareSDK_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
