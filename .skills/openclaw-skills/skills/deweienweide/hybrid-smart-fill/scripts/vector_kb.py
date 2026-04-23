# -*- coding: utf-8 -*-
"""
向量知识库构建系统 - 基于TF-IDF + BM25的混合检索
"""
import json
import re
import os
from pathlib import Path
from collections import defaultdict, Counter
import math

# 基础分词函数（无需jieba，使用简单的中文分词）
def simple_tokenize(text):
    """简单的中文分词，按字符和空格分割"""
    if not text:
        return []
    # 移除特殊字符，保留中文、英文、数字
    text = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9\s]', ' ', str(text))
    # 按空格和字符分割
    tokens = []
    for word in text.split():
        if len(word) > 1:
            tokens.append(word.lower())
        else:
            tokens.extend(list(word.lower()))
    return tokens

class BM25Retriever:
    """BM25检索器"""
    def __init__(self, k1=1.5, b=0.75):
        self.k1 = k1
        self.b = b
        self.corpus = []
        self.doc_freqs = []
        self.idf = {}
        self.avgdl = 0
        self.N = 0
        
    def fit(self, documents):
        """训练BM25模型"""
        self.corpus = documents
        self.N = len(documents)
        self.doc_freqs = []
        total_length = 0
        
        # 计算文档频率
        freq = defaultdict(int)
        for doc in documents:
            tokens = simple_tokenize(doc)
            self.doc_freqs.append(Counter(tokens))
            total_length += len(tokens)
            for token in set(tokens):
                freq[token] += 1
        
        self.avgdl = total_length / self.N if self.N > 0 else 0
        
        # 计算IDF
        for token, freq in freq.items():
            self.idf[token] = math.log((self.N - freq + 0.5) / (freq + 0.5) + 1)
    
    def search(self, query, top_k=5):
        """搜索最相关的文档"""
        query_tokens = simple_tokenize(query)
        scores = []
        
        for idx, doc_freq in enumerate(self.doc_freqs):
            score = 0
            doc_length = sum(doc_freq.values())
            for token in query_tokens:
                if token in doc_freq:
                    tf = doc_freq[token]
                    idf = self.idf.get(token, 0)
                    # BM25公式
                    numerator = tf * (self.k1 + 1)
                    denominator = tf + self.k1 * (1 - self.b + self.b * doc_length / self.avgdl)
                    score += idf * (numerator / denominator)
            scores.append((idx, score))
        
        # 按分数排序
        scores.sort(key=lambda x: x[1], reverse=True)
        return scores[:top_k]

class TFIDFRetriever:
    """TF-IDF检索器"""
    def __init__(self):
        self.vocabulary = {}
        self.idf = {}
        self.doc_vectors = []
        
    def fit(self, documents):
        """训练TF-IDF模型"""
        # 构建词汇表
        all_tokens = set()
        doc_token_lists = []
        for doc in documents:
            tokens = simple_tokenize(doc)
            doc_token_lists.append(tokens)
            all_tokens.update(tokens)
        
        self.vocabulary = {token: idx for idx, token in enumerate(sorted(all_tokens))}
        
        # 计算IDF
        doc_freq = defaultdict(int)
        for tokens in doc_token_lists:
            for token in set(tokens):
                doc_freq[token] += 1
        
        N = len(documents)
        for token in self.vocabulary:
            self.idf[token] = math.log(N / (doc_freq.get(token, 0) + 1)) + 1
        
        # 计算文档向量
        self.doc_vectors = []
        for tokens in doc_token_lists:
            vector = [0] * len(self.vocabulary)
            token_count = Counter(tokens)
            doc_len = sum(token_count.values())
            
            for token, count in token_count.items():
                if token in self.vocabulary:
                    tf = count / doc_len
                    tfidf = tf * self.idf[token]
                    vector[self.vocabulary[token]] = tfidf
            
            self.doc_vectors.append(vector)
    
    def search(self, query, top_k=5):
        """搜索最相关的文档"""
        query_tokens = simple_tokenize(query)
        query_vector = [0] * len(self.vocabulary)
        
        if not query_tokens:
            return []
        
        token_count = Counter(query_tokens)
        query_len = sum(token_count.values())
        
        for token, count in token_count.items():
            if token in self.vocabulary:
                tf = count / query_len
                tfidf = tf * self.idf.get(token, 0)
                query_vector[self.vocabulary[token]] = tfidf
        
        # 计算余弦相似度
        scores = []
        query_norm = math.sqrt(sum(x**2 for x in query_vector))
        
        if query_norm == 0:
            return []
        
        for idx, doc_vector in enumerate(self.doc_vectors):
            dot_product = sum(q * d for q, d in zip(query_vector, doc_vector))
            doc_norm = math.sqrt(sum(x**2 for x in doc_vector))
            
            if doc_norm > 0:
                similarity = dot_product / (query_norm * doc_norm)
            else:
                similarity = 0
            
            scores.append((idx, similarity))
        
        scores.sort(key=lambda x: x[1], reverse=True)
        return scores[:top_k]

class HybridRetriever:
    """混合检索器 - 结合BM25和TF-IDF"""
    def __init__(self, bm25_weight=0.5, tfidf_weight=0.5):
        self.bm25_weight = bm25_weight
        self.tfidf_weight = tfidf_weight
        self.bm25 = BM25Retriever()
        self.tfidf = TFIDFRetriever()
        self.documents = []
        self.metadata = []
        
    def fit(self, documents, metadata=None):
        """训练混合检索模型"""
        self.documents = documents
        self.metadata = metadata if metadata else [{} for _ in documents]
        self.bm25.fit(documents)
        self.tfidf.fit(documents)
        
        # 归一化权重
        total = self.bm25_weight + self.tfidf_weight
        self.bm25_weight /= total
        self.tfidf_weight /= total
    
    def search(self, query, top_k=5):
        """混合搜索"""
        # 获取两种检索结果
        bm25_results = self.bm25.search(query, top_k * 2)
        tfidf_results = self.tfidf.search(query, top_k * 2)
        
        # 合并分数
        combined_scores = defaultdict(float)
        
        for idx, score in bm25_results:
            combined_scores[idx] += self.bm25_weight * score
        
        for idx, score in tfidf_results:
            combined_scores[idx] += self.tfidf_weight * score
        
        # 排序并返回
        results = [(idx, score, self.documents[idx], self.metadata[idx]) 
                   for idx, score in combined_scores.items()]
        results.sort(key=lambda x: x[1], reverse=True)
        
        return results[:top_k]

class VectorKnowledgeBase:
    """向量知识库"""
    def __init__(self):
        self.retriever = HybridRetriever()
        self.documents = []
        self.metadata = []
        self.structured_data = {}  # 存储结构化数据
        
    def load_knowledge_base(self, kb_path):
        """加载知识库文件（JSON格式，按文件组织）"""
        print(f"正在加载知识库: {kb_path}")
        
        # 读取JSON知识库
        with open(kb_path, 'r', encoding='utf-8') as f:
            kb_data = json.load(f)
        
        self.structured_data = kb_data
        
        # 构建文档集合用于检索
        documents = []
        metadata = []
        
        # 遍历每个文件的数据
        for file_key, file_data in kb_data.items():
            content = file_data.get('content', {})
            filename = file_data.get('filename', file_key)
            file_type = file_data.get('type', '')
            
            print(f"  处理文件: {filename} ({file_type})")
            
            # 处理Excel文件内容
            if file_type == 'xlsx':
                if isinstance(content, list) and len(content) > 0:
                    # 假设第一行是表头
                    headers = content[0]
                    rows = content[1:]
                    
                    for row in rows:
                        doc_parts = []
                        for col_idx, (header, value) in enumerate(zip(headers, row)):
                            if value is not None and str(value).strip():
                                doc_parts.append(f"{header} {value}")
                        
                        doc = ' '.join(doc_parts)
                        if doc:
                            documents.append(doc)
                            metadata.append({
                                'type': 'Excel数据',
                                'source': filename,
                                'data': dict(zip(headers, row))
                            })
            
            # 处理文档内容
            elif file_type in ['doc', 'docx']:
                if isinstance(content, str):
                    # 按段落分割
                    paragraphs = [p.strip() for p in content.split('\n') if p.strip()]
                    for para in paragraphs:
                        if para:
                            documents.append(para)
                            metadata.append({
                                'type': '文档段落',
                                'source': filename
                            })
            
            # 处理字典格式的结构化数据
            elif isinstance(content, dict):
                # 将字典展开为键值对文档
                for key, value in content.items():
                    if value is not None:
                        doc = f"{key} {value}"
                        documents.append(doc)
                        metadata.append({
                            'type': '结构化数据',
                            'source': filename,
                            'field': key
                        })
        
        self.documents = documents
        self.metadata = metadata
        
        print(f"知识库加载完成，共 {len(documents)} 条文档")
        return self
    
    def load_xlsx_data(self, xlsx_path, sheet_name=None):
        """加载Excel数据"""
        print(f"正在加载Excel数据: {xlsx_path}")
        
        from openpyxl import load_workbook
        
        wb = load_workbook(xlsx_path, data_only=True)
        sheets = wb.sheetnames
        
        documents = []
        metadata = []
        
        for sheet in sheets:
            if sheet_name and sheet != sheet_name:
                continue
                
            ws = wb[sheet]
            data = []
            headers = []
            
            # 读取数据
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(cell) if cell is not None else '' for cell in row]
                    continue
                row_data = [cell if cell is not None else '' for cell in row]
                data.append(row_data)
            
            # 为每行创建文档
            for row_idx, row in enumerate(data):
                doc_parts = []
                for col_idx, (header, value) in enumerate(zip(headers, row)):
                    if value:
                        doc_parts.append(f"{header} {value}")
                
                doc = ' '.join(doc_parts)
                if doc:
                    documents.append(doc)
                    metadata.append({
                        'type': 'Excel数据',
                        'source': xlsx_path,
                        'sheet': sheet,
                        'row': row_idx + 1,
                        'data': dict(zip(headers, row))
                    })
        
        self.documents.extend(documents)
        self.metadata.extend(metadata)
        
        print(f"Excel数据加载完成，新增 {len(documents)} 条文档")
        return self
    
    def build_index(self):
        """构建检索索引"""
        print("正在构建混合检索索引...")
        self.retriever.fit(self.documents, self.metadata)
        print("索引构建完成")
        return self
    
    def search(self, query, top_k=5):
        """搜索"""
        results = self.retriever.search(query, top_k)
        
        output = []
        for idx, score, doc, meta in results:
            output.append({
                'score': score,
                'document': doc,
                'metadata': meta
            })
        
        return output
    
    def get_structured_value(self, key_path):
        """获取结构化数据值（支持点号路径）"""
        keys = key_path.split('.')
        value = self.structured_data
        
        for key in keys:
            if isinstance(value, dict):
                value = value.get(key)
            else:
                return None
        
        return value

# 测试代码
if __name__ == '__main__':
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    
    kb = VectorKnowledgeBase()
    
    # 加载知识库
    kb_path = r'c:\需求\机构外部渠道需求\机构渠道外部机构表单填写需求\知识库.json'
    kb.load_knowledge_base(kb_path)
    
    # 构建索引
    kb.build_index()
    
    # 测试搜索
    test_queries = [
        '法人代表',
        '联系电话',
        '注册资本',
        '地址',
        '股东信息',
        '2024年',
        '债券基金规模'
    ]
    
    for query in test_queries:
        print(f"\n搜索: {query}")
        results = kb.search(query, top_k=3)
        for i, result in enumerate(results):
            doc_preview = result['document'][:80] if len(result['document']) > 80 else result['document']
            print(f"  [{i+1}] 分数: {result['score']:.4f}")
            print(f"      文档: {doc_preview}...")
            print(f"      类型: {result['metadata'].get('type', '')}")
