"""
Markdown folder to Excel converter.

Extracts HTML tables from OCR-generated Markdown page files and writes them into
one formatted `.xlsx` workbook.

Usage:
    python md_to_excel.py /path/to/markdowns /path/to/output.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path
import re
from typing import Optional, Tuple, Dict
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# For LaTeX to human-readable conversion
try:
    import unicodeit

    HAS_UNICODEIT = True
except ImportError:
    HAS_UNICODEIT = False


def convert_latex_to_readable(text: str) -> str:
    """
    Convert LaTeX expressions in text to human-readable Unicode math formulas.

    Handles both display math ($$...$$) and inline math ($...$).

    Args:
        text: Text that may contain LaTeX expressions

    Returns:
        Text with LaTeX expressions converted to readable Unicode
    """
    if not text or not isinstance(text, str):
        return text

    # Check if there's any LaTeX to convert
    if '$' not in text and '\\' not in text:
        return text

    def convert_single_latex(latex_content: str) -> str:
        """Convert a single LaTeX expression to readable text."""

        # Try unicodeit first if available (best results for scientific notation)
        if HAS_UNICODEIT:
            try:
                result = unicodeit.replace(latex_content)
                # Clean up any remaining LaTeX artifacts
                result = _cleanup_latex_artifacts(result)
                return result.strip()
            except Exception:
                pass

        # Fallback to comprehensive manual conversion
        return _manual_latex_convert(latex_content)

    # Pattern to match display math ($$...$$ and \[...\]) and inline math ($...$ or \(...\))
    # Process display math first ($$...$$ and \[...\])
    display_math_pattern = r'\$\$([\s\S]*?)\$\$'
    bracket_display_pattern = r'\\\[\s*([\s\S]*?)\s*\\\]'

    def replace_display_math(match):
        latex_content = match.group(1)
        converted = convert_single_latex(latex_content)
        return f' {converted} '  # Add spaces for readability

    result = re.sub(display_math_pattern, replace_display_math, text, flags=re.DOTALL)
    result = re.sub(bracket_display_pattern, replace_display_math, result, flags=re.DOTALL)

    # Process inline math ($...$ and \(...\))
    inline_math_pattern = r'\$([^$]+)\$'
    paren_inline_pattern = r'\\\(\s*([^)]*?)\s*\\\)'

    def replace_inline_math(match):
        latex_content = match.group(1)
        return convert_single_latex(latex_content)

    result = re.sub(inline_math_pattern, replace_inline_math, result)
    result = re.sub(paren_inline_pattern, replace_inline_math, result)

    # Normalize superscript/subscript runs for Excel display
    result = _normalize_supsub_for_excel(result)

    # Clean up multiple spaces
    result = re.sub(r' +', ' ', result)
    result = result.strip()

    return result


def _cleanup_latex_artifacts(text: str) -> str:
    """Clean up remaining LaTeX artifacts after conversion."""
    result = text

    # Remove \left and \right
    result = re.sub(r'\\left\s*\.?', '', result)
    result = re.sub(r'\\right\s*\.?', '', result)

    # Remove text formatting commands but keep content
    result = re.sub(r'\\(?:text|mathrm|mathbf|mathit|textbf|textit)\s*\{([^{}]*)\}', r'\1', result)

    # Clean up spacing commands
    result = result.replace(r'\quad', '  ')
    result = result.replace(r'\qquad', '   ')
    result = re.sub(r'\\[,;:!]', ' ', result)

    # Remove remaining unknown commands
    result = re.sub(r'\\[a-zA-Z]+(?:\s*\{[^{}]*\})*', '', result)

    # Clean up braces
    result = result.replace('{', '').replace('}', '')

    return result


def _normalize_supsub_for_excel(text: str) -> str:
    """
    Excel sometimes fails to render sub/superscript letters. Keep numeric runs
    in Unicode (they render well), but rewrite any letter-containing runs to
    ASCII fallbacks like _(...) or ^(...).
    """
    if not text:
        return text

    sub_map = {
        '₀': '0', '₁': '1', '₂': '2', '₃': '3', '₄': '4',
        '₅': '5', '₆': '6', '₇': '7', '₈': '8', '₉': '9',
        '₊': '+', '₋': '-', '₌': '=', '₍': '(', '₎': ')',
        'ₐ': 'a', 'ₑ': 'e', 'ₕ': 'h', 'ᵢ': 'i', 'ⱼ': 'j', 'ₖ': 'k',
        'ₗ': 'l', 'ₘ': 'm', 'ₙ': 'n', 'ₒ': 'o', 'ₚ': 'p', 'ᵣ': 'r',
        'ₛ': 's', 'ₜ': 't', 'ᵤ': 'u', 'ᵥ': 'v', 'ₓ': 'x',
    }
    sup_map = {
        '⁰': '0', '¹': '1', '²': '2', '³': '3', '⁴': '4',
        '⁵': '5', '⁶': '6', '⁷': '7', '⁸': '8', '⁹': '9',
        '⁺': '+', '⁻': '-', '⁼': '=', '⁽': '(', '⁾': ')',
        'ᵃ': 'a', 'ᵇ': 'b', 'ᶜ': 'c', 'ᵈ': 'd', 'ᵉ': 'e', 'ᶠ': 'f',
        'ᵍ': 'g', 'ʰ': 'h', 'ⁱ': 'i', 'ʲ': 'j', 'ᵏ': 'k', 'ˡ': 'l',
        'ᵐ': 'm', 'ⁿ': 'n', 'ᵒ': 'o', 'ᵖ': 'p', 'ʳ': 'r', 'ˢ': 's',
        'ᵗ': 't', 'ᵘ': 'u', 'ᵛ': 'v', 'ʷ': 'w', 'ˣ': 'x', 'ʸ': 'y',
        'ᶻ': 'z', 'ᴬ': 'A', 'ᴮ': 'B', 'ᴰ': 'D', 'ᴱ': 'E', 'ᴳ': 'G',
        'ᴴ': 'H', 'ᴵ': 'I', 'ᴶ': 'J', 'ᴷ': 'K', 'ᴸ': 'L', 'ᴹ': 'M',
        'ᴺ': 'N', 'ᴼ': 'O', 'ᴾ': 'P', 'ᴿ': 'R', 'ᵀ': 'T', 'ᵁ': 'U',
        'ⱽ': 'V', 'ᵂ': 'W',
    }

    sub_letters = set(ch for ch in sub_map if ch.isalpha())
    sup_letters = set(ch for ch in sup_map if ch.isalpha())

    def replace_sub(match):
        run = match.group(1)
        has_letter = any(ch in sub_letters for ch in run)
        if not has_letter:
            return run  # purely numeric/symbolic run renders fine
        plain = ''.join(sub_map.get(ch, '') or ch for ch in run)
        return f'_({plain})' if len(plain) > 1 else f'_{plain}'

    def replace_sup(match):
        run = match.group(1)
        has_letter = any(ch in sup_letters for ch in run)
        if not has_letter:
            return run  # keep numeric-only superscripts
        plain = ''.join(sup_map.get(ch, '') or ch for ch in run)
        return f'^({plain})' if len(plain) > 1 else f'^{plain}'

    # Replace letter-containing sub/superscript runs
    text = re.sub(r'([₀-₉₊₋₌₍₎ₐₑₕᵢⱼₖₗₘₙₒₚᵣₛₜᵤᵥₓ]+)', replace_sub, text)
    text = re.sub(r'([⁰¹²³⁴⁵⁶⁷⁸⁹⁺⁻⁼⁽⁾ᵃᵇᶜᵈᵉᶠᵍʰⁱʲᵏˡᵐⁿᵒᵖʳˢᵗᵘᵛʷˣʸᶻᴬᴮᴰᴱᴳᴴᴵᴶᴷᴸᴹᴺᴼᴾᴿᵀᵁⱽᵂ]+)', replace_sup, text)
    return text


def _manual_latex_convert(latex_content: str) -> str:
    """Comprehensive manual LaTeX to Unicode conversion."""
    result = latex_content

    # ==================== Greek Letters ====================
    greek_lower = {
        r'\alpha': 'α', r'\beta': 'β', r'\gamma': 'γ', r'\delta': 'δ',
        r'\epsilon': 'ε', r'\varepsilon': 'ε', r'\zeta': 'ζ', r'\eta': 'η',
        r'\theta': 'θ', r'\vartheta': 'ϑ', r'\iota': 'ι', r'\kappa': 'κ',
        r'\lambda': 'λ', r'\mu': 'μ', r'\nu': 'ν', r'\xi': 'ξ',
        r'\omicron': 'ο', r'\pi': 'π', r'\varpi': 'ϖ', r'\rho': 'ρ',
        r'\varrho': 'ϱ', r'\sigma': 'σ', r'\varsigma': 'ς', r'\tau': 'τ',
        r'\upsilon': 'υ', r'\phi': 'φ', r'\varphi': 'ϕ', r'\chi': 'χ',
        r'\psi': 'ψ', r'\omega': 'ω',
    }
    greek_upper = {
        r'\Gamma': 'Γ', r'\Delta': 'Δ', r'\Theta': 'Θ', r'\Lambda': 'Λ',
        r'\Xi': 'Ξ', r'\Pi': 'Π', r'\Sigma': 'Σ', r'\Upsilon': 'Υ',
        r'\Phi': 'Φ', r'\Psi': 'Ψ', r'\Omega': 'Ω',
    }

    # ==================== Math Operators ====================
    operators = {
        r'\times': '×', r'\div': '÷', r'\pm': '±', r'\mp': '∓',
        r'\cdot': '·', r'\ast': '∗', r'\star': '⋆', r'\circ': '∘',
        r'\bullet': '•', r'\oplus': '⊕', r'\ominus': '⊖', r'\otimes': '⊗',
    }

    # ==================== Relations ====================
    relations = {
        r'\leq': '≤', r'\le': '≤', r'\geq': '≥', r'\ge': '≥',
        r'\neq': '≠', r'\ne': '≠', r'\approx': '≈', r'\cong': '≅',
        r'\equiv': '≡', r'\sim': '∼', r'\simeq': '≃', r'\propto': '∝',
        r'\ll': '≪', r'\gg': '≫', r'\prec': '≺', r'\succ': '≻',
        r'\preceq': '⪯', r'\succeq': '⪰',
    }

    # ==================== Set Theory ====================
    sets = {
        r'\subset': '⊂', r'\supset': '⊃', r'\subseteq': '⊆', r'\supseteq': '⊇',
        r'\in': '∈', r'\ni': '∋', r'\notin': '∉', r'\cup': '∪', r'\cap': '∩',
        r'\emptyset': '∅', r'\varnothing': '∅', r'\setminus': '∖',
    }

    # ==================== Logic & Quantifiers ====================
    logic = {
        r'\forall': '∀', r'\exists': '∃', r'\nexists': '∄',
        r'\neg': '¬', r'\lnot': '¬', r'\land': '∧', r'\lor': '∨',
        r'\wedge': '∧', r'\vee': '∨', r'\implies': '⟹', r'\iff': '⟺',
    }

    # ==================== Arrows ====================
    arrows = {
        r'\rightarrow': '→', r'\to': '→', r'\leftarrow': '←',
        r'\leftrightarrow': '↔', r'\Rightarrow': '⇒', r'\Leftarrow': '⇐',
        r'\Leftrightarrow': '⇔', r'\mapsto': '↦', r'\uparrow': '↑',
        r'\downarrow': '↓', r'\nearrow': '↗', r'\searrow': '↘',
        r'\longrightarrow': '⟶', r'\longleftarrow': '⟵',
    }

    # ==================== Calculus & Analysis ====================
    calculus = {
        r'\infty': '∞', r'\partial': '∂', r'\nabla': '∇',
        r'\sum': '∑', r'\prod': '∏', r'\coprod': '∐',
        r'\int': '∫', r'\iint': '∬', r'\iiint': '∭', r'\oint': '∮',
        r'\prime': '′', r'\dprime': '″',
    }

    # ==================== Delimiters ====================
    # Use placeholders for literal braces to protect them during cleanup
    LBRACE_PLACEHOLDER = '\x00LBRACE\x00'
    RBRACE_PLACEHOLDER = '\x00RBRACE\x00'

    delimiters = {
        r'\langle': '⟨', r'\rangle': '⟩', r'\lfloor': '⌊', r'\rfloor': '⌋',
        r'\lceil': '⌈', r'\rceil': '⌉', r'\lvert': '|', r'\rvert': '|',
        r'\lVert': '‖', r'\rVert': '‖',
        r'\{': LBRACE_PLACEHOLDER, r'\}': RBRACE_PLACEHOLDER,
        r'\lbrace': LBRACE_PLACEHOLDER, r'\rbrace': RBRACE_PLACEHOLDER,
        r'\lbrack': '[', r'\rbrack': ']',
    }

    # ==================== Miscellaneous ====================
    misc = {
        r'\ldots': '…', r'\cdots': '⋯', r'\vdots': '⋮', r'\ddots': '⋱',
        r'\degree': '°', r'\angle': '∠', r'\measuredangle': '∡',
        r'\perp': '⊥', r'\parallel': '∥', r'\triangle': '△',
        r'\square': '□', r'\diamond': '◇', r'\clubsuit': '♣',
        r'\spadesuit': '♠', r'\heartsuit': '♡', r'\diamondsuit': '♢',
        r'\aleph': 'ℵ', r'\hbar': 'ℏ', r'\ell': 'ℓ', r'\wp': '℘',
        r'\Re': 'ℜ', r'\Im': 'ℑ',
    }

    # ==================== Spacing ====================
    spacing = {
        r'\quad': '  ', r'\qquad': '    ', r'\,': ' ', r'\;': ' ',
        r'\:': ' ', r'\!': '', r'\ ': ' ', r'~': ' ',
    }

    # Apply all symbol replacements (order matters - longer patterns first)
    all_symbols = {}
    all_symbols.update(greek_lower)
    all_symbols.update(greek_upper)
    all_symbols.update(operators)
    all_symbols.update(relations)
    all_symbols.update(sets)
    all_symbols.update(logic)
    all_symbols.update(arrows)
    all_symbols.update(calculus)
    all_symbols.update(delimiters)
    all_symbols.update(misc)
    all_symbols.update(spacing)

    # Sort by length (longest first) to avoid partial replacements
    for latex, unicode_char in sorted(all_symbols.items(), key=lambda x: -len(x[0])):
        result = result.replace(latex, unicode_char)

    # ==================== Handle Complex Structures ====================

    def extract_braced_content(s: str, start: int) -> Tuple[str, int]:
        """Extract content within braces, handling nested braces.
        Returns (content, end_position)."""
        if start >= len(s) or s[start] != '{':
            return '', start
        depth = 0
        content_start = start + 1
        i = start
        while i < len(s):
            if s[i] == '{':
                depth += 1
            elif s[i] == '}':
                depth -= 1
                if depth == 0:
                    return s[content_start:i], i + 1
            i += 1
        return s[content_start:], len(s)

    # Handle fractions: \frac{a}{b} -> (a)/(b)
    def process_fractions(s: str) -> str:
        """Process all \frac commands in the string."""
        result_str = s
        while r'\frac' in result_str:
            idx = result_str.find(r'\frac')
            if idx == -1:
                break

            # Find start of first brace
            brace_start = idx + 5
            while brace_start < len(result_str) and result_str[brace_start] in ' \t':
                brace_start += 1

            if brace_start >= len(result_str) or result_str[brace_start] != '{':
                break

            # Extract numerator
            num, num_end = extract_braced_content(result_str, brace_start)

            # Find start of second brace
            den_start = num_end
            while den_start < len(result_str) and result_str[den_start] in ' \t':
                den_start += 1

            if den_start >= len(result_str) or result_str[den_start] != '{':
                break

            # Extract denominator
            den, den_end = extract_braced_content(result_str, den_start)

            # Recursively process numerator and denominator
            num_converted = process_fractions(num) if r'\frac' in num else num
            den_converted = process_fractions(den) if r'\frac' in den else den

            # Format the fraction
            if len(num_converted) > 1 and not num_converted.replace('.', '').replace('-', '').isalnum():
                num_converted = f'({num_converted})'
            if len(den_converted) > 1 and not den_converted.replace('.', '').replace('-', '').isalnum():
                den_converted = f'({den_converted})'

            fraction_str = f'{num_converted}/{den_converted}'

            # Replace in result
            result_str = result_str[:idx] + fraction_str + result_str[den_end:]

        return result_str

    result = process_fractions(result)

    # Handle square roots: \sqrt{x} -> √(x), \sqrt[n]{x} -> ⁿ√(x)
    def process_sqrt(s: str) -> str:
        """Process all \\sqrt commands."""
        result_str = s
        # Handle \sqrt[n]{x} first
        while r'\sqrt[' in result_str:
            match = re.search(r'\\sqrt\s*\[([^\]]*)\]\s*\{', result_str)
            if not match:
                break
            n_val = match.group(1)
            brace_start = match.end() - 1
            content, end_pos = extract_braced_content(result_str, brace_start)
            replacement = f'{n_val}√({content})'
            result_str = result_str[:match.start()] + replacement + result_str[end_pos:]

        # Handle \sqrt{x}
        while r'\sqrt' in result_str:
            match = re.search(r'\\sqrt\s*\{', result_str)
            if not match:
                # Try \sqrt x (without braces)
                match2 = re.search(r'\\sqrt\s+(\w)', result_str)
                if match2:
                    result_str = result_str[:match2.start()] + f'√{match2.group(1)}' + result_str[match2.end():]
                else:
                    break
                continue
            brace_start = match.end() - 1
            content, end_pos = extract_braced_content(result_str, brace_start)
            replacement = f'√({content})'
            result_str = result_str[:match.start()] + replacement + result_str[end_pos:]

        return result_str

    result = process_sqrt(result)

    # ==================== Superscripts ====================
    superscript_map = {
        '0': '⁰', '1': '¹', '2': '²', '3': '³', '4': '⁴',
        '5': '⁵', '6': '⁶', '7': '⁷', '8': '⁸', '9': '⁹',
        '+': '⁺', '-': '⁻', '=': '⁼', '(': '⁽', ')': '⁾',
        'a': 'ᵃ', 'b': 'ᵇ', 'c': 'ᶜ', 'd': 'ᵈ', 'e': 'ᵉ',
        'f': 'ᶠ', 'g': 'ᵍ', 'h': 'ʰ', 'i': 'ⁱ', 'j': 'ʲ',
        'k': 'ᵏ', 'l': 'ˡ', 'm': 'ᵐ', 'n': 'ⁿ', 'o': 'ᵒ',
        'p': 'ᵖ', 'r': 'ʳ', 's': 'ˢ', 't': 'ᵗ', 'u': 'ᵘ',
        'v': 'ᵛ', 'w': 'ʷ', 'x': 'ˣ', 'y': 'ʸ', 'z': 'ᶻ',
        'A': 'ᴬ', 'B': 'ᴮ', 'D': 'ᴰ', 'E': 'ᴱ', 'G': 'ᴳ',
        'H': 'ᴴ', 'I': 'ᴵ', 'J': 'ᴶ', 'K': 'ᴷ', 'L': 'ᴸ',
        'M': 'ᴹ', 'N': 'ᴺ', 'O': 'ᴼ', 'P': 'ᴾ', 'R': 'ᴿ',
        'T': 'ᵀ', 'U': 'ᵁ', 'V': 'ⱽ', 'W': 'ᵂ',
    }

    def convert_superscript(content: str) -> str:
        """Convert content to superscript Unicode if possible."""
        # First, process any nested ^n patterns within the content
        processed = content
        # Handle simple nested superscripts like q^2 -> q²
        for char, sup in superscript_map.items():
            if len(char) == 1 and char.isdigit():
                processed = re.sub(rf'\^{char}(?![0-9])', sup, processed)

        converted = []
        i = 0
        while i < len(processed):
            char = processed[i]
            if char in superscript_map:
                converted.append(superscript_map[char])
            elif char == ' ':
                converted.append(' ')
            elif char == '/':
                converted.append('/')  # Keep division in superscript
            elif char in '²³⁴⁵⁶⁷⁸⁹⁰¹⁺⁻':
                converted.append(char)  # Already converted superscript
            else:
                return f'^({processed})'  # Can't fully convert, use fallback
            i += 1
        return ''.join(converted)

    # Handle ^{...} and ^x patterns
    def replace_superscript(match):
        content = match.group(1) if match.group(1) else match.group(2)
        return convert_superscript(content)

    result = re.sub(r'\^{([^{}]*)}|\^(\w)', replace_superscript, result)

    # ==================== Subscripts ====================
    subscript_map = {
        '0': '₀', '1': '₁', '2': '₂', '3': '₃', '4': '₄',
        '5': '₅', '6': '₆', '7': '₇', '8': '₈', '9': '₉',
        '+': '₊', '-': '₋', '=': '₌', '(': '₍', ')': '₎',
    }

    def convert_subscript(content: str) -> str:
        """Convert content to subscript Unicode where broadly supported; otherwise keep ASCII."""
        # If the subscript contains letters, stick to ASCII so Excel never drops glyphs
        contains_letters = any(ch.isalpha() for ch in content)
        if contains_letters:
            return f'_({content})' if len(content) > 1 else f'_{content}'

        converted = []
        for char in content:
            if char in subscript_map:
                converted.append(subscript_map[char])
            else:
                # Unknown symbol: fall back to readable ASCII
                return f'_({content})' if len(content) > 1 else f'_{content}'

        return ''.join(converted)

    # Handle _{...} and _x patterns
    def replace_subscript(match):
        content = match.group(1) if match.group(1) else match.group(2)
        return convert_subscript(content)

    result = re.sub(r'_{([^{}]*)}|_(\w)', replace_subscript, result)

    # ==================== Math Functions ====================
    math_functions = [
        'sin', 'cos', 'tan', 'cot', 'sec', 'csc',
        'arcsin', 'arccos', 'arctan', 'arccot',
        'sinh', 'cosh', 'tanh', 'coth',
        'log', 'ln', 'lg', 'exp',
        'lim', 'sup', 'inf', 'max', 'min',
        'det', 'dim', 'ker', 'deg', 'arg', 'gcd', 'lcm',
        'mod', 'bmod', 'pmod',
    ]
    for func in math_functions:
        result = re.sub(rf'\\{func}\b', func, result)

    # ==================== Accents/Modifiers ====================
    # Use postfix modifier characters instead of combining diacritics for Excel compatibility
    # Combining characters don't display well in Excel cells
    accent_patterns = [
        # Hat/circumflex: use modifier letter circumflex (ˆ U+02C6) as postfix
        (r'\\hat\s*\{([^{}]+)\}', r'\1ˆ'),
        (r'\\hat\s*(\w)', r'\1ˆ'),
        # Bar/overline: use macron (¯ U+00AF) as postfix
        (r'\\bar\s*\{([^{}]+)\}', r'\1¯'),
        (r'\\bar\s*(\w)', r'\1¯'),
        (r'\\overline\s*\{([^{}]*)\}', r'\1¯'),
        # Tilde: use small tilde (˜ U+02DC) as postfix
        (r'\\tilde\s*\{([^{}]+)\}', r'\1˜'),
        (r'\\tilde\s*(\w)', r'\1˜'),
        # Dot: use middle dot (· U+00B7) as superscript-like postfix
        (r'\\dot\s*\{([^{}]+)\}', r'\1̇'),
        (r'\\dot\s*(\w)', r'\1̇'),
        # Double dot: use diaeresis (¨ U+00A8) as postfix
        (r'\\ddot\s*\{([^{}]+)\}', r'\1¨'),
        (r'\\ddot\s*(\w)', r'\1¨'),
        # Vector: use arrow (→) as postfix
        (r'\\vec\s*\{([^{}]+)\}', r'\1→'),
        (r'\\vec\s*(\w)', r'\1→'),
        # Underline: use low line
        (r'\\underline\s*\{([^{}]*)\}', r'_\1_'),
    ]
    for pattern, replacement in accent_patterns:
        result = re.sub(pattern, replacement, result)

    # ==================== Text Formatting ====================
    result = re.sub(r'\\(?:text|mathrm|mathbf|mathit|textbf|textit|mbox|hbox)\s*\{([^{}]*)\}', r'\1', result)
    result = re.sub(r'\\mathbb\s*\{([^{}]*)\}', r'\1', result)  # Could map to double-struck if needed
    result = re.sub(r'\\mathcal\s*\{([^{}]*)\}', r'\1', result)
    result = re.sub(r'\\mathfrak\s*\{([^{}]*)\}', r'\1', result)

    # ==================== Delimiters ====================
    result = re.sub(r'\\left\s*([(\[{|.])', r'\1', result)
    result = re.sub(r'\\right\s*([)\]}|.])', r'\1', result)
    result = result.replace(r'\left.', '').replace(r'\right.', '')
    result = result.replace(r'\bigl', '').replace(r'\bigr', '')
    result = result.replace(r'\Bigl', '').replace(r'\Bigr', '')
    result = result.replace(r'\biggl', '').replace(r'\biggr', '')
    result = result.replace(r'\Biggl', '').replace(r'\Biggr', '')

    # ==================== Clean Up ====================
    # Remove any remaining backslash commands
    result = re.sub(r'\\[a-zA-Z]+\s*', '', result)

    # Clean up structural braces (not literal braces)
    result = result.replace('{', '').replace('}', '')

    # Restore literal braces from placeholders
    result = result.replace(LBRACE_PLACEHOLDER, '{').replace(RBRACE_PLACEHOLDER, '}')

    # Clean up multiple spaces
    result = re.sub(r' +', ' ', result)

    return result.strip()


def contains_math_unicode(text: str) -> bool:
    """
    Check if text contains Unicode mathematical symbols that need special font support.

    Args:
        text: Text to check

    Returns:
        True if text contains special Unicode math characters
    """
    if not text or not isinstance(text, str):
        return False

    # Check for combining diacritical marks (like hat, bar, tilde accents)
    # Range U+0300 to U+036F
    for char in text:
        code = ord(char)
        # Latin-1 Supplement superscripts: ¹²³ (U+00B2, U+00B3, U+00B9)
        if char in '¹²³':
            return True
        # Latin-1 Supplement: ± × ÷ ¯ ¨ (operators and modifiers)
        if char in '±×÷¯¨':
            return True
        # Modifier letters: ˆ ˜ (U+02C6, U+02DC) used for hat and tilde
        if char in 'ˆ˜':
            return True
        # Combining diacritical marks
        if 0x0300 <= code <= 0x036F:
            return True
        # Greek letters (U+0370-U+03FF)
        if 0x0370 <= code <= 0x03FF:
            return True
        # Superscripts and subscripts (U+2070-U+209F)
        if 0x2070 <= code <= 0x209F:
            return True
        # General punctuation (includes prime ′)
        if 0x2032 <= code <= 0x2037:
            return True
        # Mathematical operators (U+2200-U+22FF)
        if 0x2200 <= code <= 0x22FF:
            return True
        # Miscellaneous mathematical symbols (U+2A00-U+2AFF)
        if 0x2A00 <= code <= 0x2AFF:
            return True
        # Letterlike symbols (U+2100-U+214F)
        if 0x2100 <= code <= 0x214F:
            return True
        # Arrows (U+2190-U+21FF)
        if 0x2190 <= code <= 0x21FF:
            return True
        # Modifier letters (superscript/subscript letters)
        if 0x1D00 <= code <= 0x1DBF:
            return True
        # Latin Extended-D (more subscript chars)
        if 0xA720 <= code <= 0xA7FF:
            return True

    return False


def contains_latex_markup(text: str) -> bool:
    """
    Lightweight check for LaTeX-style markup so we can switch to a math-friendly font
    even if the conversion produces mostly ASCII characters.
    """
    if not text or not isinstance(text, str):
        return False
    return bool(re.search(r'(\$[^$]*\$)|\\[a-zA-Z]+', text, flags=re.DOTALL))


def markdown_folder_to_excel(markdown_folder: Path, excel_path: Path) -> None:
    """
    Extract HTML tables from markdown files and export to a formatted Excel file.

    Args:
        markdown_folder: Path to folder containing .md files named like XXXXX_00001.md
        excel_path: Path for the output Excel file
    """

    # ==================== Helper Functions ====================

    def parse_filename(filepath: Path) -> Tuple[Optional[str], Optional[int]]:
        """Extract report name and page number from filename like 'Report_Name_00001.md'."""
        match = re.match(r"(.+)_(\d+)\.md$", filepath.name)
        if match:
            return match.group(1), int(match.group(2))
        return None, None

    def convert_to_number(value: str):
        """
        Convert string to int or float if it represents a number.
        Float numbers are rounded to 2 decimal places.
        """
        if not isinstance(value, str):
            return value

        original = value.strip()
        if not original:
            return original

        # Remove common formatting: currency symbols, commas, spaces, parentheses for negatives
        cleaned = original

        # Handle negative numbers in parentheses: (123.45) -> -123.45
        is_negative_parens = cleaned.startswith('(') and cleaned.endswith(')')
        if is_negative_parens:
            cleaned = cleaned[1:-1]

        # Remove currency symbols and thousand separators
        cleaned = re.sub(r'[$€£¥₹,\s]', '', cleaned)

        # Handle percentage
        is_percent = cleaned.endswith('%')
        if is_percent:
            cleaned = cleaned[:-1]

        # Handle trailing negative sign
        if cleaned.endswith('-'):
            cleaned = '-' + cleaned[:-1]

        try:
            if '.' in cleaned:
                num = float(cleaned)
            else:
                # Check if it's actually an integer
                num = int(cleaned)
                if is_negative_parens:
                    num = -num
                if is_percent:
                    return round(num / 100, 2)
                return num

            if is_negative_parens:
                num = -num
            if is_percent:
                num = num / 100

            # Round floats to 2 decimal places
            return round(num, 2)
        except ValueError:
            return original

    def estimate_text_width(value) -> float:
        """Estimate the display width of a cell value."""
        if value is None:
            return 0
        text = str(value)
        # Rough estimation: normal chars = 1, wide chars = 1.5
        width = 0
        for char in text:
            if ord(char) > 255:  # Non-ASCII (likely CJK)
                width += 1.8
            else:
                width += 1
        return width

    # ==================== Get and Sort Files ====================

    md_files = list(markdown_folder.glob("*.md"))

    files_with_info = []
    for f in md_files:
        name, page_num = parse_filename(f)
        if page_num is not None:
            files_with_info.append((f, name, page_num))

    # Sort by page number
    files_with_info.sort(key=lambda x: x[2])

    if not files_with_info:
        raise ValueError(f"No valid markdown files found in {markdown_folder}")

    # ==================== Setup Workbook and Styles ====================

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Tables"

    # Color definitions (Business Dark Blue theme)
    BUSINESS_DARK_BLUE = "1F4E79"
    WHITE = "FFFFFF"
    HIGHLIGHT_GOLD = "FFD966"
    GRAY_TEXT = "666666"
    LIGHT_GRAY_BORDER = "BFBFBF"

    # Style objects
    # Use a wide-coverage font for mathematical Unicode symbols
    # Arial Unicode MS covers most sub/superscripts better than Cambria Math in cells
    MATH_FONT_NAME = "Arial Unicode MS"

    header_fill = PatternFill(start_color=BUSINESS_DARK_BLUE, end_color=BUSINESS_DARK_BLUE, fill_type="solid")
    header_font = Font(color=WHITE, bold=True, size=10)
    header_font_math = Font(name=MATH_FONT_NAME, color=WHITE, bold=True, size=10)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    page_header_fill = PatternFill(start_color=HIGHLIGHT_GOLD, end_color=HIGHLIGHT_GOLD, fill_type="solid")
    page_header_font = Font(bold=True, size=11, color="000000")

    no_table_font = Font(italic=True, color=GRAY_TEXT, size=10)

    # Font for cells containing mathematical Unicode
    math_font = Font(name=MATH_FONT_NAME, size=10)
    math_font_bold = Font(name=MATH_FONT_NAME, bold=True, size=10)

    cell_border = Border(
        left=Side(style='thin', color=LIGHT_GRAY_BORDER),
        right=Side(style='thin', color=LIGHT_GRAY_BORDER),
        top=Side(style='thin', color=LIGHT_GRAY_BORDER),
        bottom=Side(style='thin', color=LIGHT_GRAY_BORDER)
    )

    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # Track column widths across all tables
    col_max_widths: Dict[int, float] = {}

    current_row = 1

    # ==================== Table Writing Function ====================

    def write_table(table_html: str, start_row: int) -> Tuple[int, int]:
        """
        Parse an HTML table and write it to the worksheet.

        Returns:
            Tuple of (next_available_row, max_column_used)
        """
        soup = BeautifulSoup(table_html, 'html.parser')
        table = soup.find('table')
        if not table:
            return start_row, 1

        all_rows = table.find_all('tr')
        if not all_rows:
            return start_row, 1

        # Identify header rows (rows in <thead> or rows containing only <th> elements)
        thead = table.find('thead')
        thead_trs = set(thead.find_all('tr')) if thead else set()

        # Grid to track cells occupied by rowspan
        # Key: (local_row_offset, column) -> True if occupied
        occupied: Dict[Tuple[int, int], bool] = {}

        # Track merge operations to apply styling to merged cells later
        merge_info = []  # List of (start_row, start_col, end_row, end_col, is_header, value)

        local_row = 0
        max_col_used = 1

        for row_idx, tr in enumerate(all_rows):
            cells = tr.find_all(['th', 'td'])

            # Determine if this is a header row
            is_header_row = (
                    tr in thead_trs or
                    all(cell.name == 'th' for cell in cells)
            )

            col = 1

            for cell_elem in cells:
                # Skip columns occupied by previous rowspans
                while (local_row, col) in occupied:
                    col += 1

                # Get span attributes
                colspan = int(cell_elem.get('colspan', 1))
                rowspan = int(cell_elem.get('rowspan', 1))

                # Extract and convert cell value
                raw_text = cell_elem.get_text(strip=True)
                had_latex_markup = contains_latex_markup(raw_text)
                # Convert LaTeX expressions to human-readable format
                raw_text = convert_latex_to_readable(raw_text)
                value = convert_to_number(raw_text)

                # Determine if this specific cell should be styled as header
                is_header_cell = is_header_row or cell_elem.name == 'th'

                # Calculate Excel positions
                excel_row = start_row + local_row
                excel_col = col

                # Write the cell value
                cell = ws.cell(row=excel_row, column=excel_col, value=value)

                # Check if value contains math Unicode characters
                has_math = contains_math_unicode(str(value)) if value else False
                if not has_math and had_latex_markup:
                    # Even if conversion is mostly ASCII, ensure math font for better fidelity
                    has_math = True

                # Apply base styling
                cell.border = cell_border

                if is_header_cell:
                    cell.fill = header_fill
                    cell.font = header_font_math if has_math else header_font
                    cell.alignment = header_alignment
                else:
                    # Use math font for cells with mathematical symbols
                    if has_math:
                        cell.font = math_font
                    if isinstance(value, (int, float)):
                        cell.alignment = right_align
                    else:
                        cell.alignment = left_align

                # Track column width
                text_width = estimate_text_width(value)
                if excel_col not in col_max_widths:
                    col_max_widths[excel_col] = 0
                col_max_widths[excel_col] = max(col_max_widths[excel_col], text_width)

                # Handle cell merging
                if colspan > 1 or rowspan > 1:
                    end_excel_row = excel_row + rowspan - 1
                    end_excel_col = excel_col + colspan - 1

                    # Store merge info for styling
                    merge_info.append((
                        excel_row, excel_col,
                        end_excel_row, end_excel_col,
                        is_header_cell
                    ))

                    # Merge the cells
                    ws.merge_cells(
                        start_row=excel_row,
                        start_column=excel_col,
                        end_row=end_excel_row,
                        end_column=end_excel_col
                    )

                    # Mark cells as occupied for rowspan tracking
                    for r_offset in range(rowspan):
                        for c_offset in range(colspan):
                            if r_offset > 0 or c_offset > 0:
                                occupied[(local_row + r_offset, col + c_offset)] = True

                max_col_used = max(max_col_used, col + colspan - 1)
                col += colspan

            local_row += 1

        # Apply borders to all cells in merged ranges
        for (sr, sc, er, ec, is_hdr) in merge_info:
            for r in range(sr, er + 1):
                for c in range(sc, ec + 1):
                    cell = ws.cell(row=r, column=c)
                    if not isinstance(cell, MergedCell):
                        cell.border = cell_border
                        if is_hdr:
                            cell.fill = header_fill

        return start_row + local_row, max_col_used

    # ==================== Process Each File ====================

    total_files = len(files_with_info)

    for file_idx, (filepath, report_name, page_num) in enumerate(files_with_info):
        # Read file content
        try:
            content = filepath.read_text(encoding='utf-8')
        except UnicodeDecodeError:
            try:
                content = filepath.read_text(encoding='latin-1')
            except Exception as e:
                print(f"Warning: Could not read {filepath}: {e}")
                continue
        except Exception as e:
            print(f"Warning: Could not read {filepath}: {e}")
            continue

        # Find all HTML tables using regex (handles nested tags)
        table_pattern = r'<table[^>]*>.*?</table>'
        tables = re.findall(table_pattern, content, re.DOTALL | re.IGNORECASE)

        # Write page header with highlighting
        page_header_cell = ws.cell(
            row=current_row,
            column=1,
            value=f"Tables from Page {page_num}"
        )
        page_header_cell.fill = page_header_fill
        page_header_cell.font = page_header_font
        page_header_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        if not tables:
            # No tables found on this page
            no_table_cell = ws.cell(
                row=current_row,
                column=1,
                value="(This page does not have explicit tables)"
            )
            no_table_cell.font = no_table_font
            current_row += 1
        else:
            # Process each table on the page
            for table_idx, table_html in enumerate(tables):
                current_row, _ = write_table(table_html, current_row)

                # Add 2 empty lines between tables (not after the last table)
                if table_idx < len(tables) - 1:
                    current_row += 2

        # Add 3 blank lines between pages (not after the last page)
        if file_idx < total_files - 1:
            current_row += 3

    # ==================== Post-Processing ====================

    max_row = ws.max_row
    max_col = ws.max_column

    # Set white background for ALL cells to hide grid lines
    for row in range(1, max_row + 50):  # Add some buffer rows
        for col in range(1, max(max_col + 10, 20)):  # Add some buffer columns
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                current_fill = cell.fill
                # Only apply white fill if cell doesn't have a custom fill
                if (current_fill.patternType is None or
                        current_fill.fgColor.rgb in ('00000000', None)):
                    cell.fill = white_fill

    # Adjust column widths (elegant but not too wide)
    MAX_COL_WIDTH = 45
    MIN_COL_WIDTH = 8
    DEFAULT_WIDTH = 12

    for col in range(1, max_col + 1):
        if col in col_max_widths:
            # Add padding and apply constraints
            width = col_max_widths[col] + 3
            width = max(MIN_COL_WIDTH, min(width, MAX_COL_WIDTH))
        else:
            width = DEFAULT_WIDTH

        ws.column_dimensions[get_column_letter(col)].width = width

    # Set elegant row heights
    DEFAULT_ROW_HEIGHT = 18
    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = DEFAULT_ROW_HEIGHT

    # ==================== Save Workbook ====================

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)

    print(f"Successfully extracted tables to: {excel_path}")
    print(f"  - Processed {total_files} page(s)")
    print(f"  - Output has {max_row} rows and {max_col} columns")


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert OCR Markdown page files with HTML tables into one Excel workbook.",
    )
    parser.add_argument(
        "markdown_folder",
        type=Path,
        help="Directory containing per-page Markdown files such as report_00001.md.",
    )
    parser.add_argument(
        "output",
        type=Path,
        help="Output .xlsx file path.",
    )
    return parser


if __name__ == "__main__":
    args = _build_parser().parse_args()
    markdown_folder_to_excel(args.markdown_folder, args.output)
