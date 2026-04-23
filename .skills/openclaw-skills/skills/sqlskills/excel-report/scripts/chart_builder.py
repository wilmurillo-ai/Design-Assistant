"""
chart_builder.py — Excel 图表生成器
excel-report Skill 专用模块
基于 openpyxl 内置图表类型构建多样化图表
"""

from __future__ import annotations

import math
from typing import Any, Callable, Optional, Sequence

from openpyxl import Workbook
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    RadarChart,
    ScatterChart,
    AreaChart,
    DoughnutChart,
)
from openpyxl.chart.axis import TextAxis, NumericAxis, DateAxis
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.units import cm_to_EMU, EMU_to_pixels
from openpyxl.worksheet.drawing import Drawing


# ---------------------------------------------------------------------------
# 默认调色板
# ---------------------------------------------------------------------------
DEFAULT_COLORS = [
    "4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5",
    "70AD47", "264478", "9E480E", "636363", "997300",
    "7030A0", "00B0F0", "FF6699", "00B050", "C00000",
]

# KPI 仪表盘配色
KPI_COLORS = {
    "good":      "00B050",   # 绿色
    "warning":   "FFC000",   # 橙色
    "bad":       "FF0000",   # 红色
    "neutral":   "4472C4",   # 蓝色
}


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def _color_hex(hex_str: str) -> str:
    """标准化为 6 位大写 HEX，不带 #"""
    h = hex_str.lstrip("#").upper()
    return h if len(h) == 6 else h.zfill(6)


def _ensure_serie_ref(chart, values, titles):
    """
    绑定数据到图表系列。
    支持两种写法:
      - values=list-of-lists / values=list-of-tuples (2D 矩阵)
      - values=list / values=range (一维，自动补0扩展)
    """
    from openpyxl.chart import Reference
    if not isinstance(values, (list, tuple)):
        values = [values]
    if values and isinstance(values[0], (int, float, str)):
        # 一维转二维
        values = [values]

    refs = []
    n_rows = len(values)
    n_cols = len(values[0]) if values else 0

    for col_idx, col_data in enumerate(zip(*values)):
        col_letter = get_column_letter(col_idx + 2)   # 从第2列开始（col1 是分类轴）
        row_start = 2
        row_end = row_start + n_rows - 1
        refs.append(Reference(chart.data, min_col=col_idx + 2,
                               min_row=row_start, max_row=row_end))

    # 分类轴
    cat_ref = Reference(chart.data, min_col=1,
                         min_row=row_start, max_row=row_end)

    # 标题行（第一行）
    if titles:
        for i, title in enumerate(titles):
            if i < len(refs):
                refs[i].title = title

    return refs, cat_ref


def _apply_series_colors(chart, colors: list[str]) -> None:
    """给每个系列分配颜色"""
    for i, sp in enumerate(chart.series):
        c = colors[i % len(colors)]
        if sp.graphicalProperties is None:
            sp.graphicalProperties = GraphicalProperties()
        sp.graphicalProperties.solidFill = c


# ---------------------------------------------------------------------------
# ChartConfig — 图表配置对象
# ---------------------------------------------------------------------------

class ChartConfig:
    """
    标准图表配置
    Attributes:
        title      图表标题
        x          X 轴标签 / 分类轴名称
        y          Y 轴标签 / 数值轴名称
        colors     系列颜色列表
        style      openpyxl 内置样式编号 (1-48)
        width      图表宽度  (cm)
        height     图表高度  (cm)
        legend_pos 图例位置  'r'|'l'|'t'|'b'|None
        show_grid  显示网格线
        show_value 显示数值标签
    """

    def __init__(
        self,
        title: str = "",
        x: str = "",
        y: str = "",
        colors: Optional[list[str]] = None,
        style: int = 10,
        width: float = 14,
        height: float = 8,
        legend_pos: Optional[str] = "r",
        show_grid: bool = True,
        show_value: bool = False,
        **kwargs,
    ):
        self.title = title or ""
        self.x = x or ""
        self.y = y or ""
        self.colors = colors or DEFAULT_COLORS[:]
        self.style = style
        self.width = width
        self.height = height
        self.legend_pos = legend_pos
        self.show_grid = show_grid
        self.show_value = show_value
        for k, v in kwargs.items():
            setattr(self, k, v)


# ---------------------------------------------------------------------------
# 单图表构建器
# ---------------------------------------------------------------------------

class ChartBuilder:
    """
    统一图表构建入口。
    用法示例:
        wb = Workbook()
        ws = wb.active
        # 填入数据...
        chart = ChartBuilder.bar(ws, config, categories, [series1, series2])
        ws.add_chart(chart, "A1")
        wb.save("output.xlsx")
    """

    def __init__(self, ws, config: Optional[ChartConfig] = None, **kwargs):
        self.ws = ws
        self.config = config or ChartConfig(**kwargs)

    # ---- 工厂方法 ------------------------------------------------------------

    @classmethod
    def bar(
        cls,
        ws,
        config: Optional[ChartConfig] = None,
        categories: Optional[Sequence] = None,
        series: Optional[Sequence[Sequence]] = None,
        series_labels: Optional[Sequence[str]] = None,
        stacked: bool = False,
        direction: str = "col",   # 'col'=列柱状, 'bar'=条形
        **kwargs,
    ) -> BarChart:
        """柱状图 / 条形图"""
        cfg = config or ChartConfig(**kwargs)
        chart = BarChart()
        chart.type = "col" if direction == "col" else "bar"
        chart.grouping = "stacked" if stacked else "clustered"

        _setup_chart(chart, cfg, categories, series, series_labels)

        if direction == "bar":
            chart.x_axis.majorGridlines = None
            chart.y_axis.majorGridlines = None
        return chart

    @classmethod
    def line(
        cls,
        ws,
        config: Optional[ChartConfig] = None,
        categories: Optional[Sequence] = None,
        series: Optional[Sequence[Sequence]] = None,
        series_labels: Optional[Sequence[str]] = None,
        smooth: bool = False,
        markers: bool = True,
        **kwargs,
    ) -> LineChart:
        """折线图"""
        cfg = config or ChartConfig(**kwargs)
        chart = LineChart()
        chart.smooth = smooth
        _setup_chart(chart, cfg, categories, series, series_labels)

        for sp in chart.series:
            sp.marker = Marker("circle") if markers else None

        return chart

    @classmethod
    def pie(
        cls,
        ws,
        config: Optional[ChartConfig] = None,
        categories: Optional[Sequence] = None,
        values: Optional[Sequence] = None,
        series_label: Optional[str] = None,
        **kwargs,
    ) -> PieChart:
        """饼图"""
        cfg = config or ChartConfig(**kwargs)
        chart = PieChart()
        _setup_chart(chart, cfg, categories, [values] if values else None, [series_label] if series_label else None)

        # 饼图加百分比标签
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = False

        _apply_series_colors(chart, cfg.colors)
        return chart

    @classmethod
    def doughnut(
        cls,
        ws,
        config: Optional[ChartConfig] = None,
        categories: Optional[Sequence] = None,
        values: Optional[Sequence] = None,
        series_label: Optional[str] = None,
        hole_size: float = 50,
        **kwargs,
    ) -> DoughnutChart:
        """环形图"""
        cfg = config or ChartConfig(**kwargs)
        chart = DoughnutChart()
        chart.holeSize = hole_size
        _setup_chart(chart, cfg, categories, [values] if values else None, [series_label] if series_label else None)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        _apply_series_colors(chart, cfg.colors)
        return chart

    @classmethod
    def radar(
        cls,
        ws,
        config: Optional[ChartConfig] = None,
        categories: Optional[Sequence] = None,
        series: Optional[Sequence[Sequence]] = None,
        series_labels: Optional[Sequence[str]] = None,
        radar_style: str = "filled",
        **kwargs,
    ) -> RadarChart:
        """雷达图"""
        cfg = config or ChartConfig(**kwargs)
        chart = RadarChart()
        chart.type = radar_style   # 'filled' | 'marker' | 'standard'
        _setup_chart(chart, cfg, categories, series, series_labels)
        _apply_series_colors(chart, cfg.colors)
        return chart

    @classmethod
    def area(
        cls,
        ws,
        config: Optional[ChartConfig] = None,
        categories: Optional[Sequence] = None,
        series: Optional[Sequence[Sequence]] = None,
        series_labels: Optional[Sequence[str]] = None,
        stacked: bool = False,
        **kwargs,
    ) -> AreaChart:
        """面积图"""
        cfg = config or ChartConfig(**kwargs)
        chart = AreaChart()
        chart.grouping = "stacked" if stacked else "standard"
        _setup_chart(chart, cfg, categories, series, series_labels)
        _apply_series_colors(chart, cfg.colors)
        return chart

    @classmethod
    def scatter(
        cls,
        ws,
        config: Optional[ChartConfig] = None,
        x_values: Optional[Sequence] = None,
        y_series: Optional[Sequence[Sequence]] = None,
        series_labels: Optional[Sequence[str]] = None,
        bubble: bool = False,
        **kwargs,
    ) -> ScatterChart:
        """散点图 / 气泡图"""
        cfg = config or ChartConfig(**kwargs)
        chart = ScatterChart()
        chart.scatter = ScatterChart()

        if bubble:
            chart.bubblesize = 10

        # 分类轴设为数值
        chart.x_axis = NumericAxis()
        chart.y_axis = NumericAxis()

        # 绑定数据
        from openpyxl.chart import Reference
        n = len(x_values) if x_values else 0

        if x_values and y_series:
            for i, (y_vals, label) in enumerate(zip(y_series, series_labels or [])):
                x_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
                y_ref = Reference(ws, min_col=i + 2, min_row=2, max_row=n + 1)
                chart.series.append(_make_scatter_series(x_ref, y_ref, label,
                                                          cfg.colors[i % len(cfg.colors)]))
        return chart

    # ---- KPI 仪表盘 -----------------------------------------------------------

    @classmethod
    def kpi_card(
        cls,
        ws,
        cell: str,
        value: float | int,
        title: str = "",
        unit: str = "",
        subtitle: str = "",
        status: str = "neutral",   # good | warning | bad | neutral
        color: Optional[str] = None,
        icon_char: str = "",
    ) -> None:
        """
        在指定单元格写入一个 KPI 卡片（纯单元格方案，无 Drawing）。
        返回 None。KPI 以合并单元格 + 条件字体/填充呈现。

        参数:
            cell      左上角单元格坐标，如 'A1'
            value     数值
            title     KPI 名称
            unit      单位，如 '%' 或 '万元'
            subtitle  副标题/说明
            status    good=绿, warning=橙, bad=红, neutral=蓝
            color     手动指定颜色（覆盖 status）
            icon_char 图标字符，如 '📈'
        """
        row, col = cls._parse_cell(cell)
        accent = color or KPI_COLORS.get(status, KPI_COLORS["neutral"])

        # 背景色
        fill = PatternFill("solid", fgColor=accent + "22")   # 12% 透明度近似
        accent_fill = PatternFill("solid", fgColor=accent)

        thin = Side(style="thin", color=accent)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 标题行
        title_cell = ws.cell(row=row, column=col, value=title)
        title_cell.font = Font(bold=True, color="FFFFFF", size=9)
        title_cell.fill = accent_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # KPI 数值行（跨列合并）
        val_cell = ws.cell(row=row + 1, column=col,
                            value=f"{icon_char} {value}{unit}")
        val_cell.font = Font(bold=True, color=accent, size=18)
        val_cell.fill = fill
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border = border

        # 副标题行
        if subtitle:
            sub_cell = ws.cell(row=row + 2, column=col, value=subtitle)
            sub_cell.font = Font(color="595959", size=8)
            sub_cell.fill = fill
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")
            sub_cell.border = border

    # ---- 内部工具 ------------------------------------------------------------

    @staticmethod
    def _parse_cell(cell: str):
        """'A1' → (1, 1)"""
        import re
        m = re.match(r"([A-Z]+)(\d+)", cell.upper())
        if not m:
            raise ValueError(f"Invalid cell: {cell}")
        return int(m.group(2)), column_index_from_string(m.group(1))

    # ---- 公共渲染方法 --------------------------------------------------------

    def apply(self, chart) -> None:
        """把当前 config 的样式应用到已构建的 chart"""
        _apply_style(chart, self.config)

    def embed_data(
        self,
        ws,
        headers: Sequence[str],
        categories: Sequence,
        *series_data: Sequence,
    ) -> None:
        """
        把数据写入工作表，供图表引用。
        第一列=分类轴，其余列为数据系列。
        返回写入数据的起始行号(1-based)。
        """
        # 写标题行
        for c_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=c_idx, value=h)
        # 写分类列
        for r_idx, cat in enumerate(categories, 2):
            ws.cell(row=r_idx, column=1, value=cat)
        # 写数据列
        for s_idx, s_data in enumerate(series_data, 2):
            for r_idx, v in enumerate(s_data, 2):
                ws.cell(row=r_idx, column=s_idx, value=v)


# ---------------------------------------------------------------------------
# 内部辅助
# ---------------------------------------------------------------------------

def _setup_chart(
    chart,
    cfg: ChartConfig,
    categories: Optional[Sequence],
    series_data: Optional[Sequence[Sequence]],
    series_labels: Optional[Sequence[str]],
) -> None:
    """把数据绑定 + 样式应用到任意图表对象"""
    from openpyxl.chart import Reference

    # 标题
    chart.title = cfg.title

    # 尺寸
    chart.width = cfg.width
    chart.height = cfg.height

    # 样式
    if cfg.style and cfg.style > 0:
        chart.style = cfg.style

    # 图例
    if cfg.legend_pos is not None:
        chart.legend = Legend()
        chart.legend.position = cfg.legend_pos

    # 数据绑定
    if categories and series_data:
        _bind_data(chart, categories, series_data, series_labels)
    elif categories:
        _bind_single(chart, categories)

    # 网格线
    if not cfg.show_grid:
        chart.x_axis.majorGridlines = None
        chart.y_axis.majorGridlines = None

    # 轴标题
    if cfg.x:
        chart.x_axis.title = cfg.x
        chart.x_axis.titleLayout = None
    if cfg.y:
        chart.y_axis.title = cfg.y
        chart.y_axis.titleLayout = None

    # 数值标签
    if cfg.show_value:
        for sp in chart.series:
            sp.dLbls = DataLabelList()
            sp.dLbls.showVal = True

    # 颜色
    _apply_series_colors(chart, cfg.colors)


def _bind_data(chart, categories, series_data, series_labels):
    """多系列绑定"""
    from openpyxl.chart import Reference

    n = len(categories)
    # 分类轴: 假设在 sheet 的 col=1, row_start=2
    # 这里 chart.data 还没设置；返回 (refs, cat_ref)
    # 调用方负责引用 ws
    pass   # 见下面的 _bind_chart_data


def _bind_chart_data(
    chart,
    ws,
    categories: Sequence,
    series_data: Sequence[Sequence],
    series_labels: Optional[Sequence[str]] = None,
) -> None:
    """
    完整绑定：把数据写入临时 ws（如果 chart 已有 data），
    然后绑定 series_ref + cat_ref 到图表。
    """
    from openpyxl.chart import Reference

    n_rows = len(categories)
    n_series = len(series_data) if series_data else 0

    # 把数据写到 chart.data 引用的 sheet（默认就是 ws）
    ws.cell(row=1, column=1, value="分类")
    for i, c in enumerate(categories, 2):
        ws.cell(row=i, column=1, value=c)

    labels = series_labels or [f"系列{i+1}" for i in range(n_series)]
    for s_idx, (s_data, label) in enumerate(zip(series_data, labels)):
        col = s_idx + 2
        ws.cell(row=1, column=col, value=label)
        for r_idx, v in enumerate(s_data, 2):
            ws.cell(row=r_idx, column=col, value=v)

    # Reference
    cat_ref = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
    chart.set_categories(cat_ref)

    for s_idx in range(n_series):
        col = s_idx + 2
        vals_ref = Reference(ws, min_col=col, min_row=1,
                              max_row=n_rows + 1)
        chart.add_data(vals_ref, titles_from_data=True)


def _bind_single(chart, categories):
    """单系列绑定（只有饼图常用）"""
    from openpyxl.chart import Reference
    n = len(categories)
    cat_ref = Reference(ws=None, min_col=1, min_row=2, max_row=n + 1)
    # 占位，实际由外部 ws 指定
    pass


def _make_scatter_series(x_ref, y_ref, title, color):
    """构造散点图系列"""
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.marker import Marker
    series = ScatterChart().series
    # openpyxl ScatterChart.series 是 list，手动添加
    return _ScatterSeries(x_ref, y_ref, title, color)


class _ScatterSeries:
    def __init__(self, x_ref, y_ref, title, color):
        self.xVal = x_ref
        self.yVal = y_ref
        self.title = title
        c = _color_hex(color)
        self.graphicalProperties = GraphicalProperties()
        self.graphicalProperties.line.solidFill = c
        self.graphicalProperties.solidFill = c


def _apply_style(chart, cfg: ChartConfig) -> None:
    """把 ChartConfig 的通用属性应用到图表"""
    chart.title = cfg.title
    chart.width = cfg.width
    chart.height = cfg.height
    if cfg.legend_pos is not None:
        chart.legend = Legend()
        chart.legend.position = cfg.legend_pos
    if cfg.style and cfg.style > 0:
        chart.style = cfg.style
    if not cfg.show_grid:
        chart.x_axis.majorGridlines = None
        chart.y_axis.majorGridlines = None
    _apply_series_colors(chart, cfg.colors)


# ---------------------------------------------------------------------------
# 多图表布局管理器
# ---------------------------------------------------------------------------

class MultiChartLayout:
    """
    多图表网格布局。

    用法:
        layout = MultiChartLayout(ws, cols=2)
        layout.add(chart_a, title="图A", row=0, col=0)
        layout.add(chart_b, title="图B", row=0, col=1)
        layout.render()          # 写入所有图表和标题
        wb.save(...)
    """

    def __init__(
        self,
        ws,
        cols: int = 2,
        chart_width: float = 14,
        chart_height: float = 8,
        gap_h: float = 0.3,
        gap_v: float = 0.3,
        start_row: int = 1,
        start_col: int = 1,
    ):
        self.ws = ws
        self.cols = cols
        self.chart_width = chart_width
        self.chart_height = chart_height
        self.gap_h = gap_h
        self.gap_v = gap_v
        self.start_row = start_row
        self.start_col = start_col
        self.charts: list[dict] = []
        self._current_row = start_row

    def add(
        self,
        chart,
        title: str = "",
        subtitle: str = "",
        row: Optional[int] = None,
        col: Optional[int] = None,
        width: Optional[float] = None,
        height: Optional[float] = None,
        kpi_values: Optional[list[dict]] = None,
    ) -> "MultiChartLayout":
        """
        注册一个图表。

        kpi_values: 在图表下方渲染 mini-KPI 卡片
            [{"value": 85, "label": "完成率", "unit": "%", "status": "good"}, ...]
        """
        if row is None:
            row = len(self.charts) // self.cols
        if col is None:
            col = len(self.charts) % self.cols

        w = width or self.chart_width
        h = height or self.chart_height

        anchor = self._calc_anchor(row, col, w, h)

        self.charts.append({
            "chart": chart,
            "title": title,
            "subtitle": subtitle,
            "row": row,
            "col": col,
            "width": w,
            "height": h,
            "anchor": anchor,
            "kpi_values": kpi_values or [],
        })
        return self

    def _calc_anchor(self, row: int, col: int, w: float, h: float) -> str:
        """返回 anchor 字符串，如 'A1' 或 'C10'"""
        from openpyxl.utils import get_column_letter
        # 每个图表宽度用 cm 换算（1 cm ≈ 360000 EMU，openpyxl 默认单位）
        col_letter = get_column_letter(self.start_col + col * (int(self.chart_width // 4) + 1))
        row_num = self.start_row + row * (int(self.chart_height // 3) + 2)
        return f"{col_letter}{row_num}"

    def render(self) -> None:
        """把标题、图表、KPI 全部写入工作表"""
        # 先渲染标题
        for item in self.charts:
            if item["title"]:
                r, c = self._title_pos(item["row"], item["col"])
                cell = self.ws.cell(row=r, column=c, value=item["title"])
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4472C4")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                self.ws.merge_cells(
                    start_row=r, start_column=c,
                    end_row=r, end_column=c + 1
                )

        # 再渲染 KPI
        for item in self.charts:
            if item["kpi_values"]:
                kpi_start_row = self._kpi_start_row(item["row"], item["height"])
                for i, kpi in enumerate(item["kpi_values"]):
                    kr = kpi_start_row + i
                    kc = self._anchor_col(item["col"])
                    self.ws.cell(row=kr, column=kc,
                                   value=kpi.get("label", "")).font = Font(size=9)
                    self.ws.cell(row=kr, column=kc + 1,
                                   value=f"{kpi.get('value', 0)}{kpi.get('unit', '')}").font = Font(
                                   bold=True,
                                   color=KPI_COLORS.get(kpi.get("status", "neutral"))
                                 )

        # 最后渲染图表
        for item in self.charts:
            chart = item["chart"]
            chart.width = item["width"]
            chart.height = item["height"]
            self.ws.add_chart(chart, item["anchor"])

    def _title_pos(self, row: int, col: int) -> tuple[int, int]:
        from openpyxl.utils import column_index_from_string
        anchor = self._calc_anchor(row, col, self.chart_width, self.chart_height)
        import re
        m = re.match(r"([A-Z]+)(\d+)", anchor)
        col_idx = column_index_from_string(m.group(1))
        return int(m.group(2)) - 1, col_idx

    def _kpi_start_row(self, row: int, height: float) -> int:
        anchor = self._calc_anchor(row, 0, self.chart_width, self.chart_height)
        import re
        m = re.match(r"([A-Z]+)(\d+)", anchor)
        return int(m.group(2)) + int(height // 3) + 1

    def _anchor_col(self, col: int) -> int:
        from openpyxl.utils import column_index_from_string
        anchor = self._calc_anchor(0, col, self.chart_width, self.chart_height)
        import re
        m = re.match(r"([A-Z]+)(\d+)", anchor)
        return column_index_from_string(m.group(1))


# ---------------------------------------------------------------------------
# 便捷 API — 给定数据直接返回图表对象（不写 ws）
# ---------------------------------------------------------------------------

def build_chart(
    ws,
    chart_type: str,
    categories: Sequence,
    *series_data: Sequence,
    series_labels: Optional[Sequence[str]] = None,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
    colors: Optional[list[str]] = None,
    style: int = 10,
    width: float = 14,
    height: float = 8,
    **kwargs,
):
    """
    一行调用构建任意类型图表。
    内部自动写数据到 ws → 绑定引用 → 返回 Chart 对象。

    chart_type: 'bar' | 'line' | 'pie' | 'doughnut' | 'radar' | 'area' | 'scatter'
    """
    type_map = {
        "bar":      ChartBuilder.bar,
        "line":     ChartBuilder.line,
        "pie":      ChartBuilder.pie,
        "doughnut": ChartBuilder.doughnut,
        "radar":    ChartBuilder.radar,
        "area":     ChartBuilder.area,
        "scatter":  ChartBuilder.scatter,
    }
    factory = type_map.get(chart_type.lower())
    if not factory:
        raise ValueError(f"Unsupported chart_type: {chart_type}. "
                         f"Choose from: {list(type_map.keys())}")

    cfg = ChartConfig(
        title=title,
        x=x_axis,
        y=y_axis,
        colors=colors or DEFAULT_COLORS,
        style=style,
        width=width,
        height=height,
        **kwargs,
    )

    # 写数据到 ws（假设第1列=分类，其余列=数据，1-based）
    _write_series_data(ws, categories, series_labels or [], *series_data)

    # 构造图表
    if chart_type.lower() in ("pie", "doughnut"):
        chart = factory(ws, cfg, categories, series_data[0] if series_data else None,
                         series_labels=series_labels, **kwargs)
    else:
        chart = factory(ws, cfg, categories, series_data,
                         series_labels=series_labels, **kwargs)

    # 绑定数据引用
    _finalize_chart(chart, ws, categories, series_data)

    return chart


def _write_series_data(ws, categories, series_labels, *series_data):
    """写入图表数据到工作表"""
    # 分类
    ws.cell(row=1, column=1, value="分类")
    for r, cat in enumerate(categories, 2):
        ws.cell(row=r, column=1, value=cat)
    # 系列
    labels = list(series_labels) if series_labels else [f"系列{i+1}" for i in range(len(series_data))]
    for s_idx, (s_label, s_vals) in enumerate(zip(labels, series_data), 2):
        ws.cell(row=1, column=s_idx, value=s_label)
        for r, v in enumerate(s_vals, 2):
            ws.cell(row=r, column=s_idx, value=v)


def _finalize_chart(chart, ws, categories, series_data):
    """把 Reference 绑定到图表"""
    from openpyxl.chart import Reference

    n = len(categories)
    cat_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    chart.set_categories(cat_ref)

    for s_idx in range(len(series_data)):
        col = s_idx + 2
        vals_ref = Reference(ws, min_col=col, min_row=1, max_row=n + 1)
        chart.add_data(vals_ref, titles_from_data=True)


# ---------------------------------------------------------------------------
# KPI Gauge — 模拟仪表盘进度条（纯单元格 + 渐变填充 + 形状估算）
# ---------------------------------------------------------------------------

def kpi_gauge(
    ws,
    cell: str,
    value: float,
    max_val: float = 100,
    label: str = "",
    unit: str = "",
    status: str = "neutral",
    bar_color: Optional[str] = None,
    track_color: str = "E0E0E0",
    bar_length: int = 10,
    show_value: bool = True,
) -> None:
    """
    渲染一个单格仪表盘（模拟进度条）。

    使用 Unicode 方块字符模拟进度条段:
        ▰▰▰▰▱▱▱▱▱▱  85%
    """
    import re
    r, c = ChartBuilder._parse_cell(cell)

    bar = bar_color or KPI_COLORS.get(status, KPI_COLORS["neutral"])

    # 计算填充比例
    ratio = min(max(value / max_val, 0), 1)
    filled = round(ratio * bar_length)
    empty = bar_length - filled

    bar_text = "▰" * filled + "▱" * empty
    pct_text = f"{value:.1f}{unit}" if show_value else ""

    cell_val = f"{label} {bar_text} {pct_text}".strip()
    cell_obj = ws.cell(row=r, column=c, value=cell_val)
    cell_obj.font = Font(name="Consolas", size=10, color=bar)

    # track 背景（同一单元格背景色）
    track_fill = PatternFill("solid", fgColor=track_color)
    # 不覆盖 bar 字符颜色，只设置背景


def kpi_gauge_row(
    ws,
    start_cell: str,
    gauges: list[dict],
    row_height: int = 20,
) -> None:
    """
    渲染一行多个 KPI gauge（横向排列）。

    gauges: [{"value": 85, "max": 100, "label": "完成率",
               "unit": "%", "status": "good"}, ...]
    """
    r, start_c = ChartBuilder._parse_cell(start_cell)

    for i, g in enumerate(gauges):
        gauge_cell = get_column_letter(start_c + i * 2)
        kpi_gauge(
            ws,
            f"{gauge_cell}{r}",
            value=g.get("value", 0),
            max_val=g.get("max", 100),
            label=g.get("label", ""),
            unit=g.get("unit", ""),
            status=g.get("status", "neutral"),
        )
