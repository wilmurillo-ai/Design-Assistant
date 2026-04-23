#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Excel可视化报表生成器 v4.0 PRO - 大厂生产级"""

from __future__ import annotations
import argparse, json, os, sys, re, smtplib
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

SCRIPT_DIR = Path(__file__).parent.resolve()
SKILL_DIR = SCRIPT_DIR.parent
TEMPLATES_DIR = SKILL_DIR / "templates"
STYLES_DIR = SKILL_DIR / "styles"
OUTPUT_DIR = SKILL_DIR / "output"
SAMPLE_DATA_DIR = SKILL_DIR / "sample_data"

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import pandas as pd
except ImportError as e:
    print(f"[ERROR] Missing dependency: {e}")
    sys.exit(1)


class Colors:
    HEADER_BG = "1F3864"
    HEADER_FG = "FFFFFF"
    KPI_BG = "DEEAF1"
    KPI_VALUE = "1F3864"
    KPI_LABEL = "595959"
    ROW_ODD = "FFFFFF"
    ROW_EVEN = "F2F2F2"


class Borders:
    @staticmethod
    def thin():
        s = Side(style='thin', color='BFBFBF')
        return Border(left=s, right=s, top=s, bottom=s)
    @staticmethod
    def header():
        top = Side(style='medium', color='1F3864')
        bottom = Side(style='double', color='1F3864')
        left = Side(style='thin', color='BDD7EE')
        right = Side(style='thin', color='BDD7EE')
        return Border(left=left, right=right, top=top, bottom=bottom)
    @staticmethod
    def kpi_card():
        s = Side(style='medium', color='1F3864')
        return Border(left=s, right=s, top=s, bottom=s)
    @staticmethod
    def bottom_only():
        s = Side(style='thin', color='BFBFBF')
        return Border(bottom=s)


class Formatters:
    NUMBER = '#,##0'
    NUMBER_2 = '#,##0.00'
    CURRENCY_CN = '¥#,##0.00'
    PERCENT = '0.0%'
    TEXT = '@'


class StyleApplier:
    @staticmethod
    def apply_header_style(cell, is_kpi=False):
        cell.font = Font(name='Arial', bold=True, color=Colors.HEADER_FG, size=10 if not is_kpi else 9)
        cell.fill = PatternFill(fill_type='solid', start_color=Colors.HEADER_BG, end_color=Colors.HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Borders.header()
    
    @staticmethod
    def apply_data_style(cell, row_idx, is_numeric=False):
        bg_color = Colors.ROW_ODD if row_idx % 2 == 1 else Colors.ROW_EVEN
        cell.fill = PatternFill(fill_type='solid', start_color=bg_color, end_color=bg_color)
        if is_numeric:
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = Formatters.NUMBER_2
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.font = Font(name='Arial', size=10)
        cell.border = Borders.bottom_only()
    
    @staticmethod
    def apply_kpi_label(cell):
        cell.font = Font(name='Arial', bold=False, color=Colors.KPI_LABEL, size=9)
        cell.fill = PatternFill(fill_type='solid', start_color=Colors.KPI_BG, end_color=Colors.KPI_BG)
        cell.alignment = Alignment(horizontal='left', vertical='bottom')
        cell.border = Borders.kpi_card()
    
    @staticmethod
    def apply_kpi_value(cell):
        cell.font = Font(name='Arial', bold=True, color=Colors.KPI_VALUE, size=16)
        cell.fill = PatternFill(fill_type='solid', start_color=Colors.KPI_BG, end_color=Colors.KPI_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Borders.kpi_card()


class DataParser:
    def __init__(self):
        self.data = None
        self.mapping = {}
        self.sheet_name = "SrcData"
    
    def load(self, path):
        suffix = path.suffix.lower()
        try:
            if suffix == ".xlsx":
                self.data = pd.read_excel(path, engine="openpyxl")
            elif suffix == ".csv":
                for enc in ["utf-8-sig", "utf-8", "gbk", "gb2312", "gb18030"]:
                    try:
                        self.data = pd.read_csv(path, encoding=enc)
                        break
                    except:
                        continue
            elif suffix == ".json":
                self.data = pd.read_json(path)
            if self.data is None:
                return False
            self.data.columns = self.data.columns.str.strip()
            return True
        except Exception as e:
            print(f"[ERROR] Load failed: {e}")
            return False
    
    def map_columns(self, template_cols):
        mapping = {}
        actual_cols = {str(c).lower().strip(): c for c in self.data.columns}
        for tc in template_cols:
            field = tc.get("field", "")
            aliases = [tc.get("label", field), field] + tc.get("aliases", [])
            for alias in aliases:
                al = str(alias).lower().strip()
                if al in actual_cols:
                    mapping[field] = actual_cols[al]
                    break
        self.mapping = mapping
        return mapping
    
    def get_col_letter(self, field):
        if field not in self.mapping:
            return None
        actual_name = self.mapping[field]
        if actual_name not in self.data.columns:
            return None
        idx = self.data.columns.get_loc(actual_name) + 1
        return get_column_letter(idx)
    
    def get_data_rows(self):
        return len(self.data) if self.data is not None else 0


class FormulaBuilder:
    """SQL公式转Excel公式"""
    
    def __init__(self, parser):
        self.parser = parser
    
    def build(self, formula, sheet, n_rows):
        """将SQL公式转换为Excel公式"""
        last_row = n_rows + 1
        quoted = f"'{sheet}'"
        
        # 1. 处理 LAG_SUM(field, n) - 取前n行的值，用MAX确保索引>=1
        m = re.match(r'LAG_SUM\(([^,]+),\s*(\d+)\s*\)', formula, re.IGNORECASE)
        if m:
            field = m.group(1).strip()
            n = int(m.group(2))
            col = self.parser.get_col_letter(field)
            if col:
                rng = f'{quoted}!{col}2:{col}{last_row}'
                cnt = f'COUNTA({rng})'
                idx = f'MAX(1,{cnt}-{n})'
                return f'=INDEX({rng},{idx})'
            return '=0'
        
        # 2. 处理 LAG(field, n) - 取前n行的值
        m = re.match(r'LAG\(([^,]+),\s*(\d+)\s*\)', formula, re.IGNORECASE)
        if m:
            field = m.group(1).strip()
            n = int(m.group(2))
            col = self.parser.get_col_letter(field)
            if col:
                rng = f'{quoted}!{col}2:{col}{last_row}'
                cnt = f'COUNTA({rng})'
                idx = f'MAX(1,{cnt}-{n})'
                return f'=INDEX({rng},{idx})'
            return '=0'
        
        # 3. 处理 LAST(field) - 取最后一个值
        m = re.match(r'LAST\(([^)]+)\)', formula, re.IGNORECASE)
        if m:
            field = m.group(1).strip()
            col = self.parser.get_col_letter(field)
            if col:
                rng = f'{quoted}!{col}2:{col}{last_row}'
                return f'=OFFSET({rng},COUNTA({rng})-1,0,1,1)'
            return '=0'
        
        # 4. 处理 COUNT_IF(field, condition)
        m = re.match(r'COUNT_IF\(([^,]+),\s*(.+?)\s*\)', formula, re.IGNORECASE)
        if m:
            field = m.group(1).strip()
            cond = m.group(2).strip()
            col = self.parser.get_col_letter(field)
            
            if not col and field == 'achievement_rate':
                actual_col = self.parser.get_col_letter('actual')
                target_col = self.parser.get_col_letter('target')
                if actual_col and target_col:
                    rng_a = f'{quoted}!{actual_col}2:{actual_col}{last_row}'
                    rng_t = f'{quoted}!{target_col}2:{target_col}{last_row}'
                    return f'=SUMPRODUCT(--({rng_a}>={rng_t}))'
            
            if col:
                rng = f'{quoted}!{col}2:{col}{last_row}'
                return f'=COUNTIF({rng},"{cond}")'
            return '=0'
        
        # 5. 处理复合比率公式: (FUNC(a) - FUNC(b)) / FUNC(c)
        # 支持 FUNC(field) 或 FUNC(field, n) 格式
        m = re.match(r'\((\w+)\(([^)]+)\)\s*-\s*(\w+)\(([^)]+)\)\s*\)\s*/\s*(\w+)\(([^)]+)\)', formula, re.IGNORECASE)
        if m:
            f1, a1, f2, a2, f3, a3 = m.groups()
            
            def parse_arg(arg):
                """解析参数，返回 (field, n)"""
                parts = arg.split(',')
                field = parts[0].strip()
                n = int(parts[1].strip()) if len(parts) > 1 else None
                return field, n
            
            field1, n1 = parse_arg(a1)
            field2, n2 = parse_arg(a2)
            field3, n3 = parse_arg(a3)
            
            col1 = self.parser.get_col_letter(field1)
            col2 = self.parser.get_col_letter(field2)
            col3 = self.parser.get_col_letter(field3)
            
            def build_expr(func, col, n, quoted, last_row):
                if not col:
                    return None
                rng = f'{quoted}!{col}2:{col}{last_row}'
                fu = func.upper()
                if fu == 'LAG_SUM' or fu == 'LAG':
                    lag_n = n if n else 12
                    return f'INDEX({rng},MAX(1,COUNTA({rng})-{lag_n}))'
                elif fu in ('SUM', 'AVG', 'MAX', 'MIN', 'COUNT'):
                    return f'{fu}({rng})'
                return f'{fu}({rng})'
            
            expr1 = build_expr(f1, col1, n1, quoted, last_row)
            expr2 = build_expr(f2, col2, n2, quoted, last_row)
            expr3 = build_expr(f3, col3, n3, quoted, last_row)
            
            if expr1 and expr2 and expr3:
                return f'=IF({expr3}=0,0,({expr1}-{expr2})/{expr3})'
        
        # 6. 处理 SUM(a)/SUM(b) 比率
        m = re.match(r'(\w+)\(([^)]+)\)\s*/\s*(\w+)\(([^)]+)\)', formula, re.IGNORECASE)
        if m:
            func1, arg1 = m.group(1), m.group(2)
            func2, arg2 = m.group(3), m.group(4)
            
            col1 = self.parser.get_col_letter(arg1)
            col2 = self.parser.get_col_letter(arg2)
            
            if col1 and col2:
                r1 = f'{quoted}!{col1}2:{col1}{last_row}'
                r2 = f'{quoted}!{col2}2:{col2}{last_row}'
                f1 = func1.upper() if func1.upper() in ('SUM','AVG','COUNT','MAX','MIN') else func1
                f2 = func2.upper() if func2.upper() in ('SUM','AVG','COUNT','MAX','MIN') else func2
                return f'=IF({f2}({r2})=0,0,{f1}({r1})/{f2}({r2}))'
        
        # 7. 处理简单聚合函数
        m = re.match(r'(\w+)\(([^)]+)\)', formula, re.IGNORECASE)
        if m:
            func, field = m.group(1), m.group(2)
            col = self.parser.get_col_letter(field)
            
            if col:
                rng = f'{quoted}!{col}2:{col}{last_row}'
                func_upper = func.upper()
                
                if func_upper == 'AVG':
                    return f'=AVERAGE({rng})'
                elif func_upper == 'COUNT_DISTINCT':
                    return f'=SUMPRODUCT(1/COUNTIF({rng},{rng}))'
                elif func_upper == 'MODE':
                    return f'=MODE.SNGL({rng})'
                elif func_upper == 'SUM':
                    return f'=SUM({rng})'
                elif func_upper == 'COUNT':
                    return f'=COUNTA({rng})'
                elif func_upper == 'MAX':
                    return f'=MAX({rng})'
                elif func_upper == 'MIN':
                    return f'=MIN({rng})'
        
        return f'={formula}'


class ReportGeneratorPRO:
    SHEET_NAMES = {"data_input": "SrcData", "dashboard": "Dashboard", "detail": "Detail", "ranking": "Ranking"}
    
    def __init__(self, template_path):
        self.template_path = template_path
        self.template = self._load_template()
        self.themes = self._load_themes()
        self.wb = None
        self.parser = DataParser()
        self.formula_builder = FormulaBuilder(self.parser)
        self._chart_col_start = 50
    
    def _load_template(self):
        with open(self.template_path, "r", encoding="utf-8") as f:
            return json.load(f)
    
    def _load_themes(self):
        themes_path = STYLES_DIR / "themes.json"
        if themes_path.exists():
            with open(themes_path, "r", encoding="utf-8") as f:
                return json.load(f)
        return {}
    
    def load_data(self, data_path):
        print(f"[DATA] Loading: {data_path}")
        if not self.parser.load(data_path):
            return False
        template_cols = []
        for sheet in self.template.get("sheets", []):
            if sheet.get("type") == "data_input":
                template_cols.extend(sheet.get("columns", []))
        self.parser.map_columns(template_cols)
        print(f"[OK] {self.parser.get_data_rows()} rows")
        return True
    
    def generate(self, output_path=None):
        if self.parser.data is None:
            raise ValueError("No data loaded")
        print("[WORK] Generating...")
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        industry = self.template.get("industry", "general")
        theme = self.themes.get(industry, self.themes.get("general", {}))
        for sheet_cfg in self.template.get("sheets", []):
            self._create_sheet(sheet_cfg, theme)
        if output_path is None:
            output_path = OUTPUT_DIR / self._filename()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)
        print(f"[OK] {output_path}")
        return output_path
    
    def _filename(self):
        p = self.template.get("output_filename", "report.xlsx")
        d = datetime.now().strftime("%Y%m%d")
        return p.replace("{date}", d).replace("{month}", d[:7])
    
    def _create_sheet(self, cfg, theme):
        stype = cfg.get("type", "data_input")
        print(f"  [SHEET] {cfg.get('name', stype)}")
        ws = self.wb.create_sheet(self.SHEET_NAMES.get(stype, "Sheet"))
        if stype == "data_input":
            self._create_data_sheet(ws, cfg, theme)
        elif stype == "dashboard":
            self._create_dashboard(ws, cfg, theme)
        elif stype in ("detail", "ranking"):
            self._create_detail(ws, cfg, theme)
    
    def _create_data_sheet(self, ws, cfg, theme):
        cols = cfg.get("columns", [])
        data = self.parser.data
        self.parser.sheet_name = self.SHEET_NAMES["data_input"]
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=col.get("label", col.get("field")))
            StyleApplier.apply_header_style(cell)
        if data is not None:
            for ri, row in enumerate(data.itertuples(index=False), 2):
                for ci, col_cfg in enumerate(cols):
                    cell = ws.cell(row=ri, column=ci+1)
                    if ci < len(row):
                        val = row[ci]
                        is_numeric = isinstance(val, (int, float)) and not isinstance(val, bool)
                        cell.value = val
                        StyleApplier.apply_data_style(cell, ri, is_numeric)
        for i in range(1, len(cols)+1):
            ws.column_dimensions[get_column_letter(i)].width = 14
        ws.freeze_panes = "A2"
        if data is not None:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(data)+1}"
    
    def _create_dashboard(self, ws, cfg, theme):
        ws.merge_cells("A1:N1")
        ws["A1"] = self.template.get("name", "Dashboard")
        ws["A1"].font = Font(name='Arial', bold=True, size=18, color=Colors.HEADER_BG)
        ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 36
        for idx, kpi in enumerate(cfg.get("kpis", [])):
            self._create_kpi(ws, kpi, idx)
        for chart_cfg in cfg.get("charts", []):
            self._create_chart(ws, chart_cfg, theme)
    
    def _create_kpi(self, ws, kpi, idx):
        col = idx * 4 + 1
        label = kpi.get("label", "KPI")
        formula = kpi.get("formula", "")
        fmt = kpi.get("format", "number")
        
        # 标签
        ws.cell(row=3, column=col, value=label)
        StyleApplier.apply_kpi_label(ws.cell(row=3, column=col))
        
        # 合并值单元格
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+2)
        value_cell = ws.cell(row=4, column=col)
        StyleApplier.apply_kpi_value(value_cell)
        
        # 公式
        if formula:
            excel_formula = self.formula_builder.build(formula, self.parser.sheet_name, self.parser.get_data_rows())
            value_cell.value = excel_formula
            value_cell.number_format = self._detect_format(label, fmt)
        else:
            value_cell.value = "--"
    
    def _detect_format(self, label, default_fmt):
        label_lower = label.lower()
        if any(k in label_lower for k in ['率', '达成', '增长', '完成', '占比']):
            return Formatters.PERCENT
        if any(k in label_lower for k in ['额', '营收', '销售', '成本', '金额', '资产']):
            return Formatters.CURRENCY_CN
        if any(k in label_lower for k in ['时间', '分钟', '次数', '数量']):
            return Formatters.NUMBER
        if default_fmt == "percent":
            return Formatters.PERCENT
        elif default_fmt == "currency":
            return Formatters.CURRENCY_CN
        return Formatters.NUMBER_2
    
    def _create_chart(self, ws, chart_cfg, theme):
        ctype = chart_cfg.get("type", "bar")
        title = chart_cfg.get("title", "Chart")
        pos = chart_cfg.get("position", {})
        row, col = pos.get("row", 10), pos.get("col", 1)
        w, h = pos.get("width", 12), pos.get("height", 8)
        x_field = chart_cfg.get("x", "")
        y_fields = chart_cfg.get("y", [])
        dim = chart_cfg.get("dimension", "")
        met = chart_cfg.get("metric", "")
        print(f"    [CHART] {title}")
        
        if self.parser.data is None:
            return
        try:
            df = self.parser.data
            if ctype == "pie" and dim and met:
                g = df.groupby(dim)[met].sum().reset_index()
                self._add_pie(ws, title, g[dim].astype(str).tolist(), g[met].tolist(), row, col, w, h)
            elif ctype in ("bar", "bar_stacked"):
                x_col = self.parser.get_col_letter(x_field) if x_field else None
                y_col = self.parser.get_col_letter(y_fields[0]) if y_fields else None
                if x_col and y_col:
                    self._add_bar(ws, title, x_col, y_col, row, col, w, h)
            elif ctype == "line":
                x_col = self.parser.get_col_letter(x_field) if x_field else None
                y_col = self.parser.get_col_letter(y_fields[0]) if y_fields else None
                if x_col and y_col:
                    self._add_line(ws, title, x_col, y_col, row, col, w, h)
        except Exception as e:
            print(f"      [WARN] {e}")
    
    def _add_bar(self, ws, title, x_col, y_col, row, col, w, h):
        n_rows = self.parser.get_data_rows()
        dc = self._chart_col_start
        qs = f"'{self.parser.sheet_name}'"
        for i in range(min(n_rows, 15)):
            ws.cell(row=row+i+1, column=dc, value=f"={qs}!{x_col}{i+2}")
            ws.cell(row=row+i+1, column=dc+1, value=f"={qs}!{y_col}{i+2}")
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        chart.width, chart.height = w, h
        chart.grouping = "clustered"
        chart.style = 10
        n = min(n_rows, 15)
        chart.add_data(Reference(ws, min_col=dc+1, min_row=row, max_row=row+n), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=dc, min_row=row+1, max_row=row+n))
        ws.add_chart(chart, f"{get_column_letter(col)}{row}")
        self._chart_col_start += 3
    
    def _add_line(self, ws, title, x_col, y_col, row, col, w, h):
        n_rows = self.parser.get_data_rows()
        dc = self._chart_col_start
        qs = f"'{self.parser.sheet_name}'"
        for i in range(min(n_rows, 15)):
            ws.cell(row=row+i+1, column=dc, value=f"={qs}!{x_col}{i+2}")
            ws.cell(row=row+i+1, column=dc+1, value=f"={qs}!{y_col}{i+2}")
        chart = LineChart()
        chart.title = title
        chart.width, chart.height = w, h
        chart.smooth = True
        chart.style = 10
        n = min(n_rows, 15)
        chart.add_data(Reference(ws, min_col=dc+1, min_row=row+1, max_row=row+n))
        chart.set_categories(Reference(ws, min_col=dc, min_row=row+1, max_row=row+n))
        ws.add_chart(chart, f"{get_column_letter(col)}{row}")
        self._chart_col_start += 3
    
    def _add_pie(self, ws, title, cats, vals, row, col, w, h):
        dc = self._chart_col_start
        for i, (c, v) in enumerate(zip(cats[:10], vals[:10])):
            ws.cell(row=row+i+1, column=dc, value=c)
            ws.cell(row=row+i+1, column=dc+1, value=v)
        chart = PieChart()
        chart.title = title
        chart.width, chart.height = w, h
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        n = min(len(cats), 10)
        chart.add_data(Reference(ws, min_col=dc+1, min_row=row, max_row=row+n))
        chart.set_categories(Reference(ws, min_col=dc, min_row=row+1, max_row=row+n))
        ws.add_chart(chart, f"{get_column_letter(col)}{row}")
        self._chart_col_start += 3
    
    def _create_detail(self, ws, cfg, theme):
        stype = cfg.get("type", "detail")
        ws.title = self.SHEET_NAMES.get(stype, "Detail")
        data = self.parser.data
        if data is not None:
            for ci, col in enumerate(data.columns, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                StyleApplier.apply_header_style(cell)
            for ri, row in enumerate(data.itertuples(index=False), 2):
                for ci in range(len(row)):
                    cell = ws.cell(row=ri, column=ci+1, value=row[ci])
                    is_num = isinstance(row[ci], (int, float)) and not isinstance(row[ci], bool)
                    StyleApplier.apply_data_style(cell, ri, is_num)
            for i in range(1, len(data.columns)+1):
                ws.column_dimensions[get_column_letter(i)].width = 14
            ws.freeze_panes = "A2"


def list_templates():
    ind_names = {"manufacturing": "制造业", "retail": "零售业", "finance": "金融业",
                 "internet": "互联网", "medical": "医疗", "general": "通用"}
    print("\n" + "=" * 60)
    print("Excel Report Generator v4.0 PRO")
    print("=" * 60)
    for ind_dir in sorted(TEMPLATES_DIR.iterdir()):
        if ind_dir.is_dir():
            print(f"\n[{ind_names.get(ind_dir.name, ind_dir.name)}]")
            for f in sorted(ind_dir.glob("*.json")):
                with open(f, "r", encoding="utf-8") as fp:
                    tpl = json.load(fp)
                    print(f"  - {tpl.get('name', f.stem)} ({f.stem})")
    print("\n" + "=" * 60 + "\n")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--list", action="store_true")
    parser.add_argument("--template", "-t")
    parser.add_argument("--data", "-d")
    parser.add_argument("--output", "-o")
    parser.add_argument("--batch", action="store_true")
    parser.add_argument("--data-dir", default="sample_data")
    args = parser.parse_args()
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    if args.list:
        list_templates()
        return
    
    tpl_id = args.template
    tpl_path = None
    for ind_dir in TEMPLATES_DIR.iterdir():
        if ind_dir.is_dir():
            for f in ind_dir.glob("*.json"):
                if f.stem == tpl_id:
                    tpl_path = f
                    break
    
    if not tpl_path:
        print(f"[ERROR] Template not found: {tpl_id}")
        sys.exit(1)
    
    data_path = Path(args.data)
    if not data_path.exists():
        data_path = SAMPLE_DATA_DIR / args.data
        if not data_path.exists():
            print(f"[ERROR] Data file not found: {args.data}")
            sys.exit(1)
    
    output_path = Path(args.output) if args.output else None
    gen = ReportGeneratorPRO(tpl_path)
    if gen.load_data(data_path):
        result = gen.generate(output_path)
        import subprocess
        subprocess.Popen(["start", "", str(result)], shell=True)
        print(f"\n[DONE] {result}")


if __name__ == "__main__":
    main()
