#!/usr/bin/env python3
"""
build_xlsx.py — Expense Claim Skill
Generates a formatted expense claim spreadsheet from a structured data file.

Usage:
    python3 build_xlsx.py --data expenses.json --out /mnt/user-data/outputs/expense_claim.xlsx

expenses.json format:
[
  {
    "date": "13-Feb-26",
    "vendor": "Shiromaru Ramen",
    "location": "Troy, MI, USA",
    "category": "Travel Expenses - Meal",
    "currency": "USD",
    "amount_orig": 39.22,
    "amount_inr": 39.22,
    "bill_no": "30",
    "notes": "Pre-departure dinner"
  },
  ...
]

Config (optional, passed as --config config.json):
{
  "employee_name": "Arun Prasad Vailoppilly",
  "company": "Tata Consultancy Services",
  "swon": "1085692",
  "travel_period": "Feb 13 - Feb 24, 2026",
  "fx_usd_to_inr": 86.5,
  "fx_eur_to_inr": 90.0,
  "fx_eur_to_usd": 1.065
}
"""

import argparse
import json
import os
import sys
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Run: pip install openpyxl --break-system-packages")
    sys.exit(1)

# ── Styles ────────────────────────────────────────────────────────────────────

FILLS = {
    "header":    PatternFill("solid", start_color="1F3864"),
    "subheader": PatternFill("solid", start_color="2E75B6"),
    "travel":    PatternFill("solid", start_color="D9E1F2"),
    "meal":      PatternFill("solid", start_color="E2EFDA"),
    "hotel":     PatternFill("solid", start_color="FCE4D6"),
    "other":     PatternFill("solid", start_color="F5F5F5"),
    "total":     PatternFill("solid", start_color="FFF2CC"),
}

CATEGORY_FILL = {
    "Travel Expenses - Conveyance":        "travel",
    "Ticket Expenses - Air Tickets":       "travel",
    "Ticket Expenses - Train Tickets":     "travel",
    "Ticket Expenses - Bus Tickets":       "travel",
    "Travel Expenses - Meal":              "meal",
    "Client Entertainment - Travel":       "meal",
    "Hotel Accommodation":                 "hotel",
    "Car Expenses":                        "other",
    "Communication Expenses":              "other",
    "Conference and Training Courses":     "other",
}

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def sc(cell, bold=False, white=False, fill_key=None, align="left", num_format=None):
    color = "FFFFFF" if white else "000000"
    cell.font = Font(name="Arial", size=10, bold=bold, color=color)
    if fill_key:
        cell.fill = FILLS[fill_key]
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER
    if num_format:
        cell.number_format = num_format


def build_xlsx(expenses: list, config: dict, out_path: str):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "All Expenses"
    ws2 = wb.create_sheet("Summary")

    emp = config.get("employee_name", "—")
    company = config.get("company", "—")
    period = config.get("travel_period", "—")

    # ── Sheet 1: All Expenses ─────────────────────────────────────────────────

    ws1.merge_cells("A1:I1")
    ws1["A1"] = f"EXPENSE CLAIM — {emp}"
    sc(ws1["A1"], bold=True, white=True, fill_key="header", align="center")
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:I2")
    ws1["A2"] = f"{company}  |  Travel Period: {period}"
    sc(ws1["A2"], white=True, fill_key="subheader", align="center")
    ws1.row_dimensions[2].height = 18

    ws1.row_dimensions[3].height = 6

    headers = ["#", "Date", "Vendor / Description", "Location",
               "Category", "Currency", "Amount (Original)", "Amount (INR equiv.)", "Notes"]
    ws1.append([""] * 9)
    ws1.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col)
        cell.value = h
        sc(cell, bold=True, white=True, fill_key="subheader", align="center")
    ws1.row_dimensions[4].height = 20

    for i, exp in enumerate(expenses):
        row = 5 + i
        fk = CATEGORY_FILL.get(exp.get("category", ""), "other")
        vals = [
            i + 1,
            exp.get("date", ""),
            exp.get("vendor", ""),
            exp.get("location", ""),
            exp.get("category", ""),
            exp.get("currency", ""),
            exp.get("amount_orig", 0),
            exp.get("amount_inr", 0),
            exp.get("notes", ""),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col)
            cell.value = val
            center = col in (1, 2, 5, 6, 7, 8)
            sc(cell, fill_key=fk, align="center" if center else "left")
            if col in (7, 8):
                cell.number_format = "#,##0.00"
        ws1.row_dimensions[row].height = 30

    col_widths = [4, 12, 32, 26, 28, 10, 20, 20, 38]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A5"

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────

    ws2.merge_cells("A1:E1")
    ws2["A1"] = "EXPENSE SUMMARY BY CATEGORY"
    sc(ws2["A1"], bold=True, white=True, fill_key="header", align="center")
    ws2.row_dimensions[1].height = 28

    sum_headers = ["Category", "# Items", "Original Amounts", "INR Equivalent", "USD Equivalent"]
    ws2.append([""] * 5)
    ws2.append(sum_headers)
    for col, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=3, column=col)
        cell.value = h
        sc(cell, bold=True, white=True, fill_key="subheader", align="center")

    cat_totals = defaultdict(lambda: {"count": 0, "inr": 0.0, "orig_parts": []})
    fx = config.get("fx_usd_to_inr", 86.5)
    for exp in expenses:
        cat = exp.get("category", "Other")
        cat_totals[cat]["count"] += 1
        cat_totals[cat]["inr"] += exp.get("amount_inr", 0)
        curr = exp.get("currency", "")
        orig = exp.get("amount_orig", 0)
        cat_totals[cat]["orig_parts"].append(f"{curr} {orig:,.2f}")

    ordered_cats = [c for c in [
        "Travel Expenses - Conveyance",
        "Ticket Expenses - Air Tickets",
        "Ticket Expenses - Train Tickets",
        "Ticket Expenses - Bus Tickets",
        "Hotel Accommodation",
        "Travel Expenses - Meal",
        "Client Entertainment - Travel",
        "Car Expenses",
        "Communication Expenses",
        "Conference and Training Courses",
    ] if c in cat_totals]

    grand_inr = 0.0
    srow = 4
    for cat in ordered_cats:
        d = cat_totals[cat]
        fk = CATEGORY_FILL.get(cat, "other")
        usd_eq = d["inr"] / fx
        vals = [cat, d["count"], "; ".join(d["orig_parts"]), d["inr"], usd_eq]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=srow, column=col)
            cell.value = val
            sc(cell, fill_key=fk, align="center" if col > 1 else "left")
            if col in (4, 5):
                cell.number_format = "#,##0.00"
        ws2.row_dimensions[srow].height = 22
        grand_inr += d["inr"]
        srow += 1

    # Grand total
    grand_usd = grand_inr / fx
    for col, val in enumerate(["GRAND TOTAL", len(expenses), "—", grand_inr, grand_usd], 1):
        cell = ws2.cell(row=srow, column=col)
        cell.value = val
        sc(cell, bold=True, fill_key="total", align="center" if col > 1 else "left")
        if col in (4, 5):
            cell.number_format = "#,##0.00"
    ws2.row_dimensions[srow].height = 22

    srow += 2
    ws2.merge_cells(f"A{srow}:E{srow}")
    note = (f"FX rates used: ₹{config.get('fx_usd_to_inr', 86.5)}/USD, "
            f"₹{config.get('fx_eur_to_inr', 90.0)}/EUR "
            f"(approx. {period}). Verify with finance team.")
    ws2[f"A{srow}"] = note
    ws2[f"A{srow}"].font = Font(name="Arial", size=9, italic=True, color="595959")

    for i, w in enumerate([32, 10, 30, 18, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    print(f"✓ Saved: {out_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--data", required=True, help="Path to expenses.json")
    parser.add_argument("--config", default=None, help="Path to config.json (optional)")
    parser.add_argument("--out", default="/mnt/user-data/outputs/expense_claim.xlsx")
    args = parser.parse_args()

    with open(args.data) as f:
        expenses = json.load(f)

    config = {}
    if args.config and os.path.exists(args.config):
        with open(args.config) as f:
            config = json.load(f)

    build_xlsx(expenses, config, args.out)


if __name__ == "__main__":
    main()
