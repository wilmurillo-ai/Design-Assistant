#!/usr/bin/env python3
"""
生成 FlyVerify 配置模板 Excel 文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_sheet_with_header(wb, sheet_name, headers, data):
    """创建带表头的 Sheet"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40

    # 冻结首行
    ws.freeze_panes = 'A2'

    return ws


def create_basic_info_sheet(wb):
    """创建基础信息 Sheet"""
    headers = ["配置项", "说明", "您的信息（必填）"]
    data = [
        ["appKey", "FlyVerify应用Key，从MobTech官网(https://www.mob.com/)注册应用获取", ""],
        ["appSecret", "FlyVerify应用密钥，与appKey一同获取", ""],
        ["包名", "Android应用包名，如：com.example.app", ""],
        ["签名MD5", "Android签名MD5（去掉冒号并转小写），如：a1b2c3d4e5f6...", ""],
        ["秒验审核", "是否已在MobTech后台提交秒验审核并通过（填写是或否）", "否"],
    ]
    ws = create_sheet_with_header(wb, "基础信息", headers, data)

    # 标记必填项
    for row in range(2, 7):
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True, color="FF0000")

    return ws


def create_instructions_sheet(wb):
    """创建填写说明 Sheet"""
    headers = ["说明项", "详细内容"]
    data = [
        ["填写步骤", "1. 先在 MobTech 官网(https://www.mob.com/) 注册应用，获取 appKey 和 appSecret\n2. 在开发者后台提交秒验审核并等待通过\n3. 填写基础信息（必填）\n4. 填写完成后告诉我\"填好了\"，我将继续下一步"],
        ["必填项", "基础信息 Sheet 中的 appKey、appSecret、包名是必填项\n秒验审核必须通过才能正常使用"],
        ["重要提醒", "秒验审核通过后才能正常使用一键验证功能\n签名 MD5 用于在 MobTech 后台配置应用"],
        ["网络要求", "秒验必须在手机开启移动蜂窝网络的前提下才能成功取号\n仅支持中国移动、联通、电信三大运营商\nWiFi 环境下可能无法取号"],
        ["常见问题", "取号失败：检查签名MD5、包名是否与后台一致\n秒验审核未通过：需在 MobTech 后台提交秒验审核"],
    ]

    ws = wb.create_sheet(title="填写说明")

    # 设置表头
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 设置列宽和行高
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 60

    return ws


def create_privacy_sheet(wb):
    """创建隐私合规 Sheet"""
    headers = ["说明项", "详细内容"]
    data = [
        ["合规要求", "根据 MobTech 隐私合规要求和中国区 App 上架规范，使用 FlyVerify 需要在用户同意隐私政策后才能初始化 SDK"],
        ["代码位置", "在用户点击隐私政策\"同意\"按钮的回调中添加授权代码"],
        ["调用时机", "必须先展示隐私政策弹窗，用户点击\"同意\"后才能调用"],
        ["授权代码", "FlySDK.submitPolicyGrantResult(true, null);\n\n参数说明：\n- true: 用户同意隐私政策\n- null: 使用默认的隐私控制器"],
        ["隐私政策内容", "在 App 隐私政策中应包含：\n- 使用了第三方 MobTech 秒验服务\n- 用于一键验证/登录功能\n- 涉及运营商网关认证和手机号验证\n- 用户授权后可能收集相关个人信息"],
        ["合规指南", "https://www.mob.com/wiki/detailed?wiki=421&id=717", ""],
    ]

    ws = wb.create_sheet(title="隐私合规")

    # 设置表头
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 设置列宽和行高
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 50

    return ws


def main():
    wb = Workbook()

    # 删除默认创建的 Sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 按顺序创建各个 Sheet
    create_basic_info_sheet(wb)
    create_privacy_sheet(wb)
    create_instructions_sheet(wb)

    # 调整 Sheet 顺序
    sheets = wb.sheetnames
    target_order = [
        "基础信息",
        "隐私合规",
        "填写说明",
    ]

    for idx, name in enumerate(target_order):
        if name in sheets:
            wb.move_sheet(name, offset=-wb.sheetnames.index(name) + idx)

    # 保存文件
    output_path = "/Users/haodongling/.openclaw/workspace/skills/andorid-flyverify-integration/assets/FlyVerify_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
