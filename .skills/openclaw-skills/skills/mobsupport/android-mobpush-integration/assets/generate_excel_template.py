#!/usr/bin/env python3
"""
生成 MobPush 配置模板 Excel 文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_sheet_with_header(wb, sheet_name, headers, data):
    """创建带表头的 Sheet"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40

    # 冻结首行
    ws.freeze_panes = 'A2'

    return ws


def create_basic_info_sheet(wb):
    """创建基础信息 Sheet"""
    headers = ["配置项", "说明", "您的信息（必填）"]
    data = [
        ["appKey", "MobPush应用Key，从MobTech官网(https://www.mob.com/)注册应用获取", ""],
        ["appSecret", "MobPush应用密钥，与appKey一同获取", ""],
        ["包名", "Android应用包名，如：com.example.app", ""],
        ["签名MD5", "Android签名MD5（去掉冒号并转小写），如：a1b2c3d4e5f6...", ""],
    ]
    ws = create_sheet_with_header(wb, "基础信息", headers, data)

    # 标记必填项
    for row in range(2, 6):
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True, color="FF0000")

    return ws


def create_huawei_sheet(wb):
    """创建华为通道 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["启用", "是否启用华为通道（填写是或否）", "否"],
        ["HUAWEI.appId", "华为 App ID，从华为开发者联盟获取", ""],
        ["HUAWEI.appSecret", "华为 Client Secret", ""],
        ["回执状态", "是否在华为后台开启回执状态（填写是或否）", "否"],
        ["", "", ""],
        ["申请地址", "https://developer.huawei.com/consumer/cn/", ""],
    ]
    return create_sheet_with_header(wb, "华为通道", headers, data)


def create_xiaomi_sheet(wb):
    """创建小米通道 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["启用", "是否启用小米通道（填写是或否）", "否"],
        ["XIAOMI.appId", "小米 AppID", ""],
        ["XIAOMI.appKey", "小米 AppKey", ""],
        ["XIAOMI.appSecret", "小米 AppSecret", ""],
        ["", "", ""],
        ["申请地址", "https://dev.mi.com/console/", ""],
    ]
    return create_sheet_with_header(wb, "小米通道", headers, data)


def create_oppo_sheet(wb):
    """创建OPPO通道 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["启用", "是否启用OPPO通道（填写是或否）", "否"],
        ["OPPO.appId", "OPPO AppID", ""],
        ["OPPO.appKey", "OPPO AppKey", ""],
        ["OPPO.appSecret", "OPPO AppSecret", ""],
        ["OPPO.masterSecret", "OPPO MasterSecret", ""],
        ["通知渠道", "Android 8.0+ 是否适配通知渠道（填写是或否）", "否"],
        ["", "", ""],
        ["申请地址", "https://open.oppomobile.com/", ""],
    ]
    return create_sheet_with_header(wb, "OPPO通道", headers, data)


def create_vivo_sheet(wb):
    """创建vivo通道 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["启用", "是否启用vivo通道（填写是或否）", "否"],
        ["VIVO.appId", "vivo AppID", ""],
        ["VIVO.appKey", "vivo AppKey", ""],
        ["VIVO.appSecret", "vivo AppSecret", ""],
        ["", "", ""],
        ["申请地址", "https://vivo.open.com.cn/", ""],
    ]
    return create_sheet_with_header(wb, "vivo通道", headers, data)


def create_meizu_sheet(wb):
    """创建魅族通道 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["启用", "是否启用魅族通道（填写是或否）", "否"],
        ["MEIZU.appId", "魅族 AppID", ""],
        ["MEIZU.appKey", "魅族 AppKey", ""],
        ["MEIZU.appSecret", "魅族 AppSecret", ""],
        ["", "", ""],
        ["申请地址", "https://open.flyme.cn/", ""],
    ]
    return create_sheet_with_header(wb, "魅族通道", headers, data)


def create_honor_sheet(wb):
    """创建荣耀通道 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["启用", "是否启用荣耀通道（填写是或否）", "否"],
        ["HONOR.appId", "荣耀 APP ID", ""],
        ["HONOR.appSecret", "荣耀 APP Secret", ""],
        ["HONOR.clientId", "荣耀 Client ID", ""],
        ["HONOR.clientSecret", "荣耀 Client Secret", ""],
        ["回执状态", "是否在荣耀后台开启回执状态（填写是或否）", "否"],
        ["", "", ""],
        ["注意事项", "荣耀通道仅支持 Magic UI 4.0 及以上版本", ""],
        ["申请地址", "https://developer.hihonor.com/", ""],
    ]
    return create_sheet_with_header(wb, "荣耀通道", headers, data)


def create_fcm_sheet(wb):
    """创建FCM通道 Sheet"""
    headers = ["配置项", "说明", "您的信息"]
    data = [
        ["启用", "是否启用FCM通道（填写是或否）", "否"],
        ["google-services.json", "需放入 app 模块根目录", "已放入/待放入"],
        ["私钥证书", "在 MobTech 后台上传 Firebase 服务账号私钥证书", "待上传"],
        ["", "", ""],
        ["申请地址", "https://console.firebase.google.com/", ""],
    ]
    return create_sheet_with_header(wb, "FCM通道", headers, data)


def create_instructions_sheet(wb):
    """创建填写说明 Sheet"""
    headers = ["说明项", "详细内容"]
    data = [
        ["填写步骤", "1. 先在 MobTech 官网(https://www.mob.com/) 注册应用，获取 appKey 和 appSecret\n2. 填写基础信息（必填）\n3. 根据需要的厂商通道，在对应 Sheet 中填写参数\n4. 不需要的厂商通道可留空或填写\"否\"\n5. 填写完成后告诉我\"填好了\"，我将继续下一步"],
        ["必填项", "基础信息 Sheet 中的 appKey、appSecret、包名是必填项\n签名 MD5 用于在 MobTech 后台配置应用"],
        ["厂商通道说明", "厂商通道可提升推送到达率：\n- 华为：适合华为手机用户\n- 小米：适合小米/红米手机用户\n- OPPO：适合 OPPO/realme/一加手机用户\n- vivo：适合 vivo 手机用户\n- 魅族：适合魅族手机用户\n- 荣耀：适合荣耀/华为旧机型升级用户\n- FCM：适合海外用户或使用 Google Play 的应用"],
        ["隐私合规", "必须在用户同意隐私政策后才能初始化 SDK\n详见\"隐私合规\"Sheet"],
        ["常见问题", "收不到推送：检查签名MD5、包名是否与后台一致\n厂商推送失败：检查厂商参数是否正确\nAndroid 13+ 无通知：需动态申请 POST_NOTIFICATIONS 权限"],
    ]

    ws = wb.create_sheet(title="填写说明")

    # 设置表头
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 设置列宽和行高
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 60

    return ws


def create_privacy_sheet(wb):
    """创建隐私合规 Sheet"""
    headers = ["说明项", "详细内容"]
    data = [
        ["合规要求", "根据 MobTech 隐私合规要求和中国区 App 上架规范，使用 MobPush 需要在用户同意隐私政策后才能初始化 SDK"],
        ["代码位置", "在用户点击隐私政策\"同意\"按钮的回调中添加授权代码"],
        ["调用时机", "必须先展示隐私政策弹窗，用户点击\"同意\"后才能调用"],
        ["授权代码", "MobSDK.submitPolicyGrantResult(true);\n\n如使用主动控制器方案：\nMobSDK.submitPolicyGrantResult(customController, true);"],
        ["隐私政策内容", "在 App 隐私政策中应包含：\n- 使用了第三方 MobTech MobPush 服务\n- 用于消息推送、通知提醒等功能\n- 用户授权后可能收集相关个人信息"],
        ["合规指南", "https://www.mob.com/wiki/detailed?wiki=421&id=717", ""],
    ]

    ws = wb.create_sheet(title="隐私合规")

    # 设置表头
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 设置列宽和行高
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 50

    return ws


def main():
    wb = Workbook()

    # 删除默认创建的 Sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 按顺序创建各个 Sheet
    create_basic_info_sheet(wb)
    create_privacy_sheet(wb)
    create_huawei_sheet(wb)
    create_xiaomi_sheet(wb)
    create_oppo_sheet(wb)
    create_vivo_sheet(wb)
    create_meizu_sheet(wb)
    create_honor_sheet(wb)
    create_fcm_sheet(wb)
    create_instructions_sheet(wb)

    # 调整 Sheet 顺序
    sheets = wb.sheetnames
    target_order = [
        "基础信息",
        "隐私合规",
        "填写说明",
        "华为通道",
        "小米通道",
        "OPPO通道",
        "vivo通道",
        "魅族通道",
        "荣耀通道",
        "FCM通道",
    ]

    for idx, name in enumerate(target_order):
        if name in sheets:
            wb.move_sheet(name, offset=-wb.sheetnames.index(name) + idx)

    # 保存文件
    output_path = "/Users/haodongling/.openclaw/workspace/skills/android-mobpush-integration/assets/MobPush_Config_Template.xlsx"
    wb.save(output_path)
    print(f"Excel 模板已生成: {output_path}")


if __name__ == "__main__":
    main()
