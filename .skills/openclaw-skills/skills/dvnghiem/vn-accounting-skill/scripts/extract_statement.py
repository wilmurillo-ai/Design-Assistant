# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl>=3.1.0",
#     "pdfplumber>=0.10.0",
#     "Pillow>=10.0.0",
#     "pytesseract>=0.3.10",
#     "pdf2image>=1.17.0",
# ]
# ///
"""Extract transaction data from bank statement PDF/image files and output to Excel.

Usage:
  uv run extract_statement.py /path/to/statement.pdf
  uv run extract_statement.py statement.pdf -o ~/accounting/vcb_march.xlsx
  uv run extract_statement.py statement.pdf --dry-run
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from ocr_utils import clean_amount, eprint, extract_from_file, parse_date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass
class Transaction:
    transaction_date: str = ""
    description: str = ""
    transaction_type: str = ""  # CREDIT or DEBIT
    amount: float = 0.0
    balance: float = 0.0
    reference: str = ""


@dataclass
class BankStatement:
    bank_name: str = ""
    account_number: str = ""
    account_holder: str = ""
    statement_period: str = ""
    opening_balance: float = 0.0
    closing_balance: float = 0.0
    transactions: list[Transaction] = field(default_factory=list)
    ocr_confidence: float = 0.0
    extraction_confidence: float = 0.0
    ocr_method: str = ""
    raw_text: str = ""
    source_file: str = ""


def parse_statement_header(text: str) -> dict:
    info = {}

    bank_patterns = [
        r"(Vietcombank|VCB|Techcombank|TCB|BIDV|Agribank|MB\s*Bank|MBBank|VPBank|ACB|Sacombank|TPBank|HDBank|VietinBank|CTG)",
        r"(?:Ngân hàng|Bank)[\s.:]*(.+?)(?:\n|$)",
    ]
    for pat in bank_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            info["bank_name"] = m.group(1).strip()
            break

    acct_patterns = [
        r"(?:Account\s*(?:No|Number|#)|Số tài khoản|STK)[\s.:]*(\d[\d\s-]+\d)",
        r"(?:Acct)[\s.:]*(\d{6,})",
    ]
    for pat in acct_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            info["account_number"] = re.sub(r"[\s-]", "", m.group(1).strip())
            break

    holder_patterns = [
        r"(?:Account\s*(?:Name|Holder)|Chủ tài khoản|Tên TK)[\s.:]*(.+?)(?:\n|$)",
    ]
    for pat in holder_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            info["account_holder"] = m.group(1).strip()
            break

    period_patterns = [
        r"(?:Period|Statement\s*Period|Từ ngày|Kỳ sao kê)[\s.:]*(.+?)(?:\n|$)",
        r"(\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4})\s*[-–to đến]\s*(\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4})",
    ]
    for pat in period_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            info["statement_period"] = m.group(0).strip() if len(m.groups()) > 1 else m.group(1).strip()
            break

    return info


def parse_transactions_from_tables(tables: list[list]) -> list[Transaction]:
    transactions = []
    for table in tables:
        if not table or len(table) < 2:
            continue
        header = None
        header_idx = -1
        for i, row in enumerate(table):
            if row and any(
                h and re.search(r"date|ngày|debit|credit|nợ|có|amount|số tiền|balance|số dư", str(h), re.IGNORECASE)
                for h in row
            ):
                header = [str(h or "").strip().lower() for h in row]
                header_idx = i
                break

        if header is None:
            continue

        col_map = {}
        for idx, h in enumerate(header):
            if re.search(r"date|ngày", h):
                col_map.setdefault("date", idx)
            elif re.search(r"description|diễn giải|nội dung|mô tả", h):
                col_map.setdefault("description", idx)
            elif re.search(r"debit|nợ|chi", h):
                col_map.setdefault("debit", idx)
            elif re.search(r"credit|có|thu", h):
                col_map.setdefault("credit", idx)
            elif re.search(r"amount|số tiền", h):
                col_map.setdefault("amount", idx)
            elif re.search(r"balance|số dư", h):
                col_map.setdefault("balance", idx)
            elif re.search(r"ref|reference|mã giao dịch|số (?:tham )?chiếu", h):
                col_map.setdefault("reference", idx)

        for row in table[header_idx + 1:]:
            if not row or all(not cell for cell in row):
                continue
            txn = Transaction()

            if "date" in col_map and col_map["date"] < len(row):
                txn.transaction_date = parse_date(str(row[col_map["date"]] or "").strip())

            if "description" in col_map and col_map["description"] < len(row):
                txn.description = str(row[col_map["description"]] or "").strip()

            if "debit" in col_map and col_map["debit"] < len(row):
                debit_val = clean_amount(str(row[col_map["debit"]] or ""))
                if debit_val > 0:
                    txn.amount = debit_val
                    txn.transaction_type = "DEBIT"

            if "credit" in col_map and col_map["credit"] < len(row):
                credit_val = clean_amount(str(row[col_map["credit"]] or ""))
                if credit_val > 0:
                    txn.amount = credit_val
                    txn.transaction_type = "CREDIT"

            if not txn.transaction_type and "amount" in col_map and col_map["amount"] < len(row):
                amt = clean_amount(str(row[col_map["amount"]] or ""))
                txn.amount = abs(amt)
                txn.transaction_type = "DEBIT" if amt < 0 else "CREDIT"

            if "balance" in col_map and col_map["balance"] < len(row):
                txn.balance = clean_amount(str(row[col_map["balance"]] or ""))

            if "reference" in col_map and col_map["reference"] < len(row):
                txn.reference = str(row[col_map["reference"]] or "").strip()

            if txn.amount > 0 or txn.transaction_date:
                transactions.append(txn)

    return transactions


def parse_transactions_from_text(text: str) -> list[Transaction]:
    """Fallback: parse from raw text."""
    transactions = []
    line_pattern = re.compile(
        r"(\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4})\s+(.+?)\s+([\d.,]+)\s*$",
        re.MULTILINE,
    )
    for m in line_pattern.finditer(text):
        txn = Transaction()
        txn.transaction_date = parse_date(m.group(1))
        txn.description = m.group(2).strip()
        txn.amount = clean_amount(m.group(3))
        txn.transaction_type = "DEBIT"  # Default; agent should review
        if txn.amount > 0:
            transactions.append(txn)
    return transactions


def validate_statement(statement: BankStatement) -> list[str]:
    warnings = []

    if not statement.bank_name:
        warnings.append("Missing bank name")
    if not statement.transactions:
        warnings.append("No transactions extracted")
        return warnings

    # Balance continuity
    prev_balance = statement.opening_balance
    for i, txn in enumerate(statement.transactions):
        if txn.balance > 0 and prev_balance > 0:
            if txn.transaction_type == "CREDIT":
                expected = prev_balance + txn.amount
            else:
                expected = prev_balance - txn.amount
            tolerance = max(abs(expected) * 0.01, 1.0)
            if abs(expected - txn.balance) > tolerance:
                warnings.append(
                    f"Balance gap at row {i + 1}: expected {expected:,.0f}, got {txn.balance:,.0f}"
                )
            prev_balance = txn.balance

    if statement.closing_balance > 0 and statement.transactions:
        last_balance = statement.transactions[-1].balance
        if last_balance > 0 and abs(last_balance - statement.closing_balance) > 1.0:
            warnings.append(f"Closing balance mismatch: last={last_balance:,.0f}, stated={statement.closing_balance:,.0f}")

    if statement.ocr_confidence < 85:
        warnings.append(f"Low OCR confidence: {statement.ocr_confidence}% (method: {statement.ocr_method})")

    return warnings


HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADERS = ["Date", "Description", "Type", "Amount", "Balance", "Reference"]
COL_WIDTHS = [12, 45, 10, 18, 18, 20]


def write_statement_excel(statement: BankStatement, excel_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    info_rows = [
        ("Bank:", statement.bank_name),
        ("Account:", statement.account_number),
        ("Holder:", statement.account_holder),
        ("Period:", statement.statement_period),
        ("Opening Balance:", statement.opening_balance),
        ("OCR Method:", statement.ocr_method),
        ("OCR Confidence:", f"{statement.ocr_confidence}%"),
    ]
    for row_idx, (label, value) in enumerate(info_rows, 1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    header_row = len(info_rows) + 2
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    for i, txn in enumerate(statement.transactions):
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=txn.transaction_date)
        ws.cell(row=row, column=2, value=txn.description)
        ws.cell(row=row, column=3, value=txn.transaction_type)
        ws.cell(row=row, column=4, value=txn.amount).number_format = "#,##0"
        ws.cell(row=row, column=5, value=txn.balance).number_format = "#,##0"
        ws.cell(row=row, column=6, value=txn.reference)

    close_row = header_row + len(statement.transactions) + 2
    ws.cell(row=close_row, column=1, value="Closing Balance:").font = Font(bold=True)
    ws.cell(row=close_row, column=5, value=statement.closing_balance).number_format = "#,##0"

    total_credits = sum(t.amount for t in statement.transactions if t.transaction_type == "CREDIT")
    total_debits = sum(t.amount for t in statement.transactions if t.transaction_type == "DEBIT")
    ws.cell(row=close_row + 1, column=1, value="Total Credits:").font = Font(bold=True)
    ws.cell(row=close_row + 1, column=4, value=total_credits).number_format = "#,##0"
    ws.cell(row=close_row + 2, column=1, value="Total Debits:").font = Font(bold=True)
    ws.cell(row=close_row + 2, column=4, value=total_debits).number_format = "#,##0"

    ws.auto_filter.ref = f"A{header_row}:{openpyxl.utils.get_column_letter(len(HEADERS))}{header_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(excel_path)
    wb.close()


def save_json_backup(statement: BankStatement, output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    safe_name = re.sub(r"[^\w-]", "_", statement.bank_name or "unknown")
    json_path = os.path.join(output_dir, f"statement_{safe_name}_{datetime.now():%Y%m%d_%H%M%S}.json")
    data = asdict(statement)
    data.pop("raw_text", None)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return json_path


def main():
    parser = argparse.ArgumentParser(description="Extract bank statement data from PDF/image files")
    parser.add_argument("file", help="Path to the bank statement file (PDF, JPG, PNG)")
    parser.add_argument("--output", "-o", default=None, help="Output Excel file path")
    parser.add_argument("--json-dir", default=None, help="Directory for JSON backup")
    parser.add_argument("--format", choices=["excel", "json", "both"], default="both")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate only")
    args = parser.parse_args()

    file_path = os.path.abspath(args.file)
    if not os.path.exists(file_path):
        eprint(f"File not found: {file_path}")
        sys.exit(1)

    eprint(f"Extracting from: {file_path}")
    ocr = extract_from_file(file_path)
    eprint(f"OCR method: {ocr.method} | Confidence: {ocr.confidence}% | Text length: {len(ocr.text)}")

    if not ocr.text.strip():
        eprint("ERROR: No text extracted. Check OCR dependencies.")
        sys.exit(1)

    header_info = parse_statement_header(ocr.text)
    statement = BankStatement(
        bank_name=header_info.get("bank_name", ""),
        account_number=header_info.get("account_number", ""),
        account_holder=header_info.get("account_holder", ""),
        statement_period=header_info.get("statement_period", ""),
        raw_text=ocr.text,
        source_file=file_path,
        ocr_confidence=ocr.confidence,
        ocr_method=ocr.method,
    )

    # Parse transactions (prefer table extraction, fallback to text)
    if ocr.tables:
        statement.transactions = parse_transactions_from_tables(ocr.tables)
    if not statement.transactions:
        statement.transactions = parse_transactions_from_text(ocr.text)

    for pat in [r"(?:Opening|Số dư đầu kỳ)[\s.:]*([0-9.,]+)"]:
        m = re.search(pat, ocr.text, re.IGNORECASE)
        if m:
            statement.opening_balance = clean_amount(m.group(1))
            break
    for pat in [r"(?:Closing|Số dư cuối kỳ)[\s.:]*([0-9.,]+)"]:
        m = re.search(pat, ocr.text, re.IGNORECASE)
        if m:
            statement.closing_balance = clean_amount(m.group(1))
            break

    filled = sum(1 for v in [
        statement.bank_name, statement.account_number,
        statement.transactions, statement.opening_balance,
    ] if v)
    statement.extraction_confidence = round(filled / 4 * 100, 1)

    eprint(f"Parsed {len(statement.transactions)} transactions from {statement.bank_name or 'unknown bank'}")

    warnings = validate_statement(statement)
    for w in warnings:
        eprint(f"  ⚠ {w}")

    if args.dry_run:
        data = asdict(statement)
        data.pop("raw_text", None)
        data["warnings"] = warnings
        print(json.dumps(data, ensure_ascii=False, indent=2))
        return

    bank_slug = re.sub(r"[^\w]", "", statement.bank_name or "unknown").lower()
    date_slug = datetime.now().strftime("%Y%m%d")
    default_name = f"statement_{bank_slug}_{date_slug}.xlsx"
    excel_path = os.path.abspath(args.output or default_name)

    if args.format in ("excel", "both"):
        write_statement_excel(statement, excel_path)
        eprint(f"Excel output: {excel_path}")

    if args.format in ("json", "both"):
        json_dir = args.json_dir or os.path.dirname(excel_path) or "."
        json_path = save_json_backup(statement, json_dir)
        eprint(f"JSON backup: {json_path}")

    result = {
        "bank": statement.bank_name,
        "account": statement.account_number,
        "transactions": len(statement.transactions),
        "ocr_confidence": statement.ocr_confidence,
        "ocr_method": statement.ocr_method,
        "output": excel_path,
        "warnings": warnings,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
