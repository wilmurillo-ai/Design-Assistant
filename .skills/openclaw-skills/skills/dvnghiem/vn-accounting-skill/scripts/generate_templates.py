# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl>=3.1.0",
# ]
# ///
"""Generate empty Excel templates for accounting tracking sheets.

Usage:
  uv run generate_templates.py all --output-dir ~/accounting/
  uv run generate_templates.py invoice
  uv run generate_templates.py po
  uv run generate_templates.py statement
"""

from __future__ import annotations

import argparse
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _write_headers(ws, headers, widths, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


def create_invoice_template(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Tracking"
    _write_headers(ws, [
        "Invoice#", "Date", "Vendor", "TaxCode", "Subtotal",
        "VAT", "Total", "Status", "OCR%", "Extract%", "OCR Method", "FilePath", "ProcessedAt",
    ], [15, 12, 30, 15, 15, 15, 15, 12, 8, 8, 10, 40, 20], "4472C4")
    wb.save(path)
    print(f"Created: {path}")


def create_po_template(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Tracking"
    _write_headers(ws, [
        "PO#", "Date", "Vendor", "Delivery", "Total",
        "Status", "DaysLeft", "PaymentTerms", "OCR%", "Extract%", "OCR Method", "FilePath", "ProcessedAt",
    ], [15, 12, 30, 12, 18, 12, 10, 25, 8, 8, 10, 40, 20], "548235")
    wb.save(path)
    print(f"Created: {path}")


def create_statement_template(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"
    for row, label in enumerate(["Bank:", "Account:", "Holder:", "Period:", "Opening Balance:", "OCR Method:", "OCR Confidence:"], 1):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
    header_row = 9
    fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    headers = ["Date", "Description", "Type", "Amount", "Balance", "Reference"]
    widths = [12, 45, 10, 18, 18, 20]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.auto_filter.ref = f"A{header_row}:{openpyxl.utils.get_column_letter(len(headers))}{header_row}"
    wb.save(path)
    print(f"Created: {path}")


def main():
    parser = argparse.ArgumentParser(description="Generate Excel templates for accounting")
    parser.add_argument("template", choices=["invoice", "po", "statement", "all"])
    parser.add_argument("--output-dir", "-o", default=".")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    if args.template in ("invoice", "all"):
        create_invoice_template(os.path.join(args.output_dir, "invoice_tracking.xlsx"))
    if args.template in ("po", "all"):
        create_po_template(os.path.join(args.output_dir, "po_tracking.xlsx"))
    if args.template in ("statement", "all"):
        create_statement_template(os.path.join(args.output_dir, "statement_template.xlsx"))


if __name__ == "__main__":
    main()
