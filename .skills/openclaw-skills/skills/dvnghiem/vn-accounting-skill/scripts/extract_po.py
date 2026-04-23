# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl>=3.1.0",
#     "pdfplumber>=0.10.0",
#     "Pillow>=10.0.0",
#     "pytesseract>=0.3.10",
#     "pdf2image>=1.17.0",
# ]
# ///
"""Extract structured data from Purchase Order PDF/image files and append to Excel.

Usage:
  uv run extract_po.py /path/to/po.pdf
  uv run extract_po.py po.jpg --output ~/accounting/po_tracking.xlsx
  uv run extract_po.py po.pdf --dry-run
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from ocr_utils import OCRResult, clean_amount, eprint, extract_from_file, parse_date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass
class POItem:
    description: str
    quantity: float
    unit_price: float
    total: float


@dataclass
class PurchaseOrder:
    po_number: str = ""
    po_date: str = ""
    vendor_name: str = ""
    delivery_date: str = ""
    items: list[POItem] = field(default_factory=list)
    total_amount: float = 0.0
    payment_terms: str = ""
    ocr_confidence: float = 0.0
    extraction_confidence: float = 0.0
    ocr_method: str = ""
    raw_text: str = ""
    source_file: str = ""
    status: str = "extracted"
    days_left: int | None = None


def parse_po_text(ocr: OCRResult) -> PurchaseOrder:
    text = ocr.text
    po = PurchaseOrder(
        raw_text=text, source_file=ocr.source_file,
        ocr_confidence=ocr.confidence, ocr_method=ocr.method,
    )

    po_patterns = [
        r"(?:P\.?O\.?\s*(?:No|Number|#|Num)|Purchase\s*Order\s*(?:No|#))[\s.:]*\s*([A-Z0-9/-]+)",
        r"(?:Đơn\s*(?:đặt hàng|mua hàng)\s*(?:số|#))[\s.:]*\s*([A-Z0-9/-]+)",
        r"(?:Số\s*(?:PO|ĐĐH))[\s.:]*\s*([A-Z0-9/-]+)",
    ]
    for pat in po_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            po.po_number = m.group(1).strip()
            break

    date_patterns = [
        r"(?:P\.?O\.?\s*Date|Order\s*Date|Ngày\s*(?:đặt hàng|đơn hàng)|Date)[\s.:]*(\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4})",
    ]
    for pat in date_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            po.po_date = parse_date(m.group(1))
            break
    if not po.po_date:
        m = re.search(r"(\d{1,2}[/.-]\d{1,2}[/.-]\d{4})", text)
        if m:
            po.po_date = parse_date(m.group(1))

    vendor_patterns = [
        r"(?:Vendor|Supplier|Nhà cung cấp|Đơn vị cung cấp|To)[\s.:]*(.+?)(?:\n|$)",
    ]
    for pat in vendor_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            po.vendor_name = m.group(1).strip()
            break

    delivery_patterns = [
        r"(?:Deliver(?:y|ed)\s*(?:Date|By)|Ship\s*(?:Date|By)|Ngày\s*giao\s*hàng|Hạn\s*giao|Expected\s*Delivery)[\s.:]*(\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4})",
    ]
    for pat in delivery_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            po.delivery_date = parse_date(m.group(1))
            break

    total_patterns = [
        r"(?:Total\s*Amount|Grand\s*Total|Tổng\s*(?:cộng|giá trị))[\s.:]*([0-9.,]+)",
    ]
    for pat in total_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            po.total_amount = clean_amount(m.group(1))
            break

    pay_patterns = [
        r"(?:Payment\s*Terms?|Điều khoản thanh toán|Hình thức thanh toán)[\s.:]*(.+?)(?:\n|$)",
    ]
    for pat in pay_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            po.payment_terms = m.group(1).strip()
            break

    if po.delivery_date:
        try:
            delivery = datetime.strptime(po.delivery_date, "%Y-%m-%d")
            po.days_left = (delivery - datetime.now()).days
        except ValueError:
            pass

    filled = sum(1 for v in [
        po.po_number, po.po_date, po.vendor_name,
        po.delivery_date, po.total_amount,
    ] if v)
    po.extraction_confidence = round(filled / 5 * 100, 1)

    return po


def validate_po(po: PurchaseOrder) -> list[str]:
    warnings = []
    if not po.po_number:
        warnings.append("Missing PO number")
    if not po.po_date:
        warnings.append("Missing PO date")
    if not po.vendor_name:
        warnings.append("Missing vendor name")
    if not po.total_amount:
        warnings.append("Missing total amount")

    if po.days_left is not None and po.days_left < 0:
        warnings.append(f"OVERDUE: delivery was {abs(po.days_left)} days ago ({po.delivery_date})")
    elif po.days_left is not None and po.days_left <= 3:
        warnings.append(f"URGENT: delivery in {po.days_left} days ({po.delivery_date})")

    if po.items and po.total_amount > 0:
        items_total = sum(item.total for item in po.items)
        tolerance = max(po.total_amount * 0.01, 1.0)
        if abs(items_total - po.total_amount) > tolerance:
            warnings.append(f"Total mismatch: items sum={items_total:,.0f}, total={po.total_amount:,.0f}")

    if po.ocr_confidence < 85:
        warnings.append(f"Low OCR confidence: {po.ocr_confidence}% (method: {po.ocr_method})")
    if po.extraction_confidence < 85:
        warnings.append(f"Low extraction confidence: {po.extraction_confidence}%")

    return warnings


def check_duplicate(po: PurchaseOrder, excel_path: str) -> bool:
    if not os.path.exists(excel_path):
        return False
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 5:
            if (
                str(row[0] or "") == po.po_number
                and str(row[2] or "") == po.vendor_name
                and abs(float(row[4] or 0) - po.total_amount) < 1.0
            ):
                wb.close()
                return True
    wb.close()
    return False


HEADER_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADERS = [
    "PO#", "Date", "Vendor", "Delivery", "Total",
    "Status", "DaysLeft", "PaymentTerms", "OCR%", "Extract%", "OCR Method", "FilePath", "ProcessedAt",
]
COL_WIDTHS = [15, 12, 30, 12, 18, 12, 10, 25, 8, 8, 10, 40, 20]


def init_excel(excel_path: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Tracking"
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HEADERS))}1"
    ws.freeze_panes = "A2"
    wb.save(excel_path)
    return wb


def append_to_excel(po: PurchaseOrder, excel_path: str, status: str = "extracted"):
    if not os.path.exists(excel_path):
        wb = init_excel(excel_path)
    else:
        wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    next_row = ws.max_row + 1
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_data = [
        po.po_number, po.po_date, po.vendor_name, po.delivery_date,
        po.total_amount, status, po.days_left, po.payment_terms,
        po.ocr_confidence, po.extraction_confidence, po.ocr_method,
        po.source_file, now,
    ]
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=next_row, column=col_idx, value=value)
    ws.cell(row=next_row, column=5).number_format = "#,##0"
    wb.save(excel_path)
    wb.close()


def save_json_backup(po: PurchaseOrder, output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    safe_name = re.sub(r"[^\w-]", "_", po.po_number or "unknown")
    json_path = os.path.join(output_dir, f"po_{safe_name}_{datetime.now():%Y%m%d_%H%M%S}.json")
    data = asdict(po)
    data.pop("raw_text", None)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return json_path


def main():
    parser = argparse.ArgumentParser(description="Extract PO data from PDF/image files")
    parser.add_argument("file", help="Path to the PO file (PDF, JPG, PNG)")
    parser.add_argument("--output", "-o", default="po_tracking.xlsx")
    parser.add_argument("--json-dir", default=None)
    parser.add_argument("--format", choices=["excel", "json", "both"], default="both")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    file_path = os.path.abspath(args.file)
    if not os.path.exists(file_path):
        eprint(f"File not found: {file_path}")
        sys.exit(1)

    eprint(f"Extracting text from: {file_path}")
    ocr = extract_from_file(file_path)
    eprint(f"OCR method: {ocr.method} | Confidence: {ocr.confidence}% | Text length: {len(ocr.text)}")

    if not ocr.text.strip():
        eprint("ERROR: No text extracted. Check OCR dependencies.")
        sys.exit(1)

    po = parse_po_text(ocr)
    eprint(f"Parsed PO: {po.po_number or '(unknown)'} from {po.vendor_name or '(unknown)'}")

    warnings = validate_po(po)
    if warnings:
        for w in warnings:
            eprint(f"  ⚠ {w}")
        if po.extraction_confidence < 85 or po.ocr_confidence < 85:
            po.status = "needs_review"

    excel_path = os.path.abspath(args.output)
    if not args.dry_run and check_duplicate(po, excel_path):
        eprint(f"Duplicate detected: {po.po_number} from {po.vendor_name}")
        po.status = "duplicate"

    if args.dry_run:
        data = asdict(po)
        data.pop("raw_text", None)
        data["warnings"] = warnings
        print(json.dumps(data, ensure_ascii=False, indent=2))
        return

    if args.format in ("excel", "both"):
        append_to_excel(po, excel_path, po.status)
        eprint(f"Appended to: {excel_path}")

    if args.format in ("json", "both"):
        json_dir = args.json_dir or os.path.dirname(excel_path) or "."
        json_path = save_json_backup(po, json_dir)
        eprint(f"JSON backup: {json_path}")

    result = {
        "po_number": po.po_number,
        "vendor": po.vendor_name,
        "total": po.total_amount,
        "delivery_date": po.delivery_date,
        "days_left": po.days_left,
        "status": po.status,
        "ocr_confidence": po.ocr_confidence,
        "ocr_method": po.ocr_method,
        "extraction_confidence": po.extraction_confidence,
        "warnings": warnings,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
