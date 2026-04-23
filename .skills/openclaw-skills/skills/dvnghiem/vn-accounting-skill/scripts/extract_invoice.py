# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl>=3.1.0",
#     "pdfplumber>=0.10.0",
#     "Pillow>=10.0.0",
#     "pytesseract>=0.3.10",
#     "pdf2image>=1.17.0",
# ]
# ///
"""Extract structured data from invoice PDF/image files and append to Excel tracking sheet.

Usage:
  uv run extract_invoice.py /path/to/invoice.pdf
  uv run extract_invoice.py invoice.jpg --output ~/accounting/invoice_tracking.xlsx
  uv run extract_invoice.py invoice.pdf --dry-run
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path

# Add scripts dir to path so we can import sibling modules
sys.path.insert(0, str(Path(__file__).resolve().parent))
from ocr_utils import OCRResult, clean_amount, eprint, extract_from_file, parse_date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass
class InvoiceItem:
    description: str
    quantity: float
    unit_price: float
    total: float


@dataclass
class Invoice:
    invoice_number: str = ""
    invoice_date: str = ""
    vendor_name: str = ""
    vendor_tax_code: str = ""
    items: list[InvoiceItem] = field(default_factory=list)
    subtotal: float = 0.0
    vat_rate: float = 0.0
    vat_amount: float = 0.0
    total_amount: float = 0.0
    payment_terms: str = ""
    ocr_confidence: float = 0.0
    extraction_confidence: float = 0.0
    ocr_method: str = ""
    raw_text: str = ""
    source_file: str = ""
    status: str = "extracted"


def parse_invoice_text(ocr: OCRResult) -> Invoice:
    """Parse OCR result into Invoice structure using regex patterns."""
    text = ocr.text
    invoice = Invoice(
        raw_text=text,
        source_file=ocr.source_file,
        ocr_confidence=ocr.confidence,
        ocr_method=ocr.method,
    )

    # Invoice number patterns
    inv_patterns = [
        r"(?:Invoice\s*(?:No|Number|#|Num)[\s.:]*)\s*([A-Z0-9/-]+)",
        r"(?:Số\s*(?:hóa đơn|HĐ|hoá đơn)[\s.:]*)\s*([A-Z0-9/-]+)",
        r"(?:Mã\s*(?:hóa đơn|HĐ)[\s.:]*)\s*([A-Z0-9/-]+)",
        r"(?:Ký hiệu[\s.:]*\s*[A-Z0-9]+\s*Số[\s.:]*)\s*(\d+)",
        r"(?:Số[\s.:]*)\s*(\d{7,})",
    ]
    for pat in inv_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            invoice.invoice_number = m.group(1).strip()
            break

    # Date patterns
    date_patterns = [
        r"(?:Invoice\s*Date|Ngày\s*\w*\s*tháng\s*\w*\s*năm|Ngày|Date)[\s.:]*(\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4})",
        r"[Nn]gày\s*(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})",
    ]
    for pat in date_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            if len(m.groups()) == 3:
                invoice.invoice_date = f"{int(m.group(3)):04d}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
            else:
                invoice.invoice_date = parse_date(m.group(1))
            break
    if not invoice.invoice_date:
        m = re.search(r"(\d{1,2}[/.-]\d{1,2}[/.-]\d{4})", text)
        if m:
            invoice.invoice_date = parse_date(m.group(1))

    # Vendor name
    vendor_patterns = [
        r"(?:Đơn vị bán hàng|Vendor|Supplier|Nhà cung cấp|Tên đơn vị)[\s.:]*(.+?)(?:\n|$)",
        r"(?:Company|Công ty)[\s.:]*(.+?)(?:\n|$)",
    ]
    for pat in vendor_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            invoice.vendor_name = m.group(1).strip()
            break

    # Tax code
    tax_patterns = [
        r"(?:Tax\s*(?:Code|ID|No)|Mã số thuế|MST)[\s.:]*(\d[\d-]+)",
    ]
    for pat in tax_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            invoice.vendor_tax_code = m.group(1).strip()
            break

    # Amounts
    amount_patterns = {
        "subtotal": [
            r"(?:Sub\s*total|Cộng tiền hàng|Thành tiền chưa thuế)[\s.:]*([0-9.,]+)",
        ],
        "vat_amount": [
            r"(?:Tiền thuế GTGT|VAT\s*Amount|Thuế GTGT|Tax Amount)[\s.:]*([0-9.,]+)",
        ],
        "total_amount": [
            r"(?:Tổng cộng tiền thanh toán|Total\s*Amount|Grand\s*Total|Tổng\s*(?:cộng|thanh toán))[\s.:]*([0-9.,]+)",
        ],
    }
    for field_name, patterns in amount_patterns.items():
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                setattr(invoice, field_name, clean_amount(m.group(1)))
                break

    # VAT rate
    vat_rate_patterns = [
        r"(?:Thuế suất GTGT|VAT|Thuế suất)[\s.:]*(\d+)\s*%",
    ]
    for pat in vat_rate_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            invoice.vat_rate = float(m.group(1))
            break

    # Payment terms
    pay_patterns = [
        r"(?:Payment\s*Terms?|Điều khoản thanh toán|Hình thức thanh toán)[\s.:]*(.+?)(?:\n|$)",
    ]
    for pat in pay_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            invoice.payment_terms = m.group(1).strip()
            break

    # Extraction confidence (how many fields we got)
    filled = sum(1 for v in [
        invoice.invoice_number, invoice.invoice_date, invoice.vendor_name,
        invoice.vendor_tax_code, invoice.subtotal, invoice.total_amount,
    ] if v)
    invoice.extraction_confidence = round(filled / 6 * 100, 1)

    return invoice


def validate_invoice(invoice: Invoice) -> list[str]:
    """Validate extracted invoice data. Returns list of warnings."""
    warnings = []

    if not invoice.invoice_number:
        warnings.append("Missing invoice number")
    if not invoice.invoice_date:
        warnings.append("Missing invoice date")
    if not invoice.vendor_name:
        warnings.append("Missing vendor name")
    if not invoice.total_amount:
        warnings.append("Missing total amount")

    # Verify subtotal + VAT = total
    if invoice.subtotal > 0 and invoice.vat_amount >= 0 and invoice.total_amount > 0:
        expected_total = invoice.subtotal + invoice.vat_amount
        tolerance = max(invoice.total_amount * 0.01, 1.0)
        if abs(expected_total - invoice.total_amount) > tolerance:
            warnings.append(
                f"Total mismatch: subtotal({invoice.subtotal:,.0f}) + VAT({invoice.vat_amount:,.0f}) "
                f"= {expected_total:,.0f}, but total = {invoice.total_amount:,.0f}"
            )

    # Verify VAT calculation
    if invoice.subtotal > 0 and invoice.vat_rate > 0:
        expected_vat = invoice.subtotal * invoice.vat_rate / 100
        tolerance = max(expected_vat * 0.01, 1.0)
        if invoice.vat_amount > 0 and abs(expected_vat - invoice.vat_amount) > tolerance:
            warnings.append(
                f"VAT mismatch: {invoice.subtotal:,.0f} × {invoice.vat_rate}% = {expected_vat:,.0f}, "
                f"but VAT amount = {invoice.vat_amount:,.0f}"
            )

    # OCR confidence
    if invoice.ocr_confidence < 85:
        warnings.append(f"Low OCR confidence: {invoice.ocr_confidence}% (method: {invoice.ocr_method})")

    # Extraction confidence
    if invoice.extraction_confidence < 85:
        warnings.append(f"Low extraction confidence: {invoice.extraction_confidence}%")

    return warnings


def check_duplicate(invoice: Invoice, excel_path: str) -> bool:
    """Check if invoice already exists in the tracking file."""
    if not os.path.exists(excel_path):
        return False
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 7:
            existing_inv_no = str(row[0] or "")
            existing_vendor = str(row[2] or "")
            existing_total = float(row[6] or 0)
            if (
                existing_inv_no == invoice.invoice_number
                and existing_vendor == invoice.vendor_name
                and abs(existing_total - invoice.total_amount) < 1.0
            ):
                wb.close()
                return True
    wb.close()
    return False


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADERS = [
    "Invoice#", "Date", "Vendor", "TaxCode", "Subtotal",
    "VAT", "Total", "Status", "OCR%", "Extract%", "OCR Method", "FilePath", "ProcessedAt",
]
COL_WIDTHS = [15, 12, 30, 15, 15, 15, 15, 12, 8, 8, 10, 40, 20]


def init_excel(excel_path: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Tracking"
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HEADERS))}1"
    ws.freeze_panes = "A2"
    wb.save(excel_path)
    return wb


def append_to_excel(invoice: Invoice, excel_path: str, status: str = "extracted"):
    if not os.path.exists(excel_path):
        wb = init_excel(excel_path)
    else:
        wb = openpyxl.load_workbook(excel_path)

    ws = wb.active
    next_row = ws.max_row + 1
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_data = [
        invoice.invoice_number, invoice.invoice_date, invoice.vendor_name,
        invoice.vendor_tax_code, invoice.subtotal, invoice.vat_amount,
        invoice.total_amount, status, invoice.ocr_confidence,
        invoice.extraction_confidence, invoice.ocr_method, invoice.source_file, now,
    ]
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=next_row, column=col_idx, value=value)
    for col_idx in (5, 6, 7):
        ws.cell(row=next_row, column=col_idx).number_format = "#,##0"

    wb.save(excel_path)
    wb.close()


def save_json_backup(invoice: Invoice, output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    safe_name = re.sub(r"[^\w-]", "_", invoice.invoice_number or "unknown")
    json_path = os.path.join(output_dir, f"invoice_{safe_name}_{datetime.now():%Y%m%d_%H%M%S}.json")
    data = asdict(invoice)
    data.pop("raw_text", None)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return json_path


def main():
    parser = argparse.ArgumentParser(description="Extract invoice data from PDF/image files")
    parser.add_argument("file", help="Path to the invoice file (PDF, JPG, PNG)")
    parser.add_argument("--output", "-o", default="invoice_tracking.xlsx",
                        help="Path to the output Excel tracking file")
    parser.add_argument("--json-dir", default=None, help="Directory for JSON backup")
    parser.add_argument("--format", choices=["excel", "json", "both"], default="both")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate only")
    args = parser.parse_args()

    file_path = os.path.abspath(args.file)
    if not os.path.exists(file_path):
        eprint(f"File not found: {file_path}")
        sys.exit(1)

    # OCR / Extract text
    eprint(f"Extracting text from: {file_path}")
    ocr = extract_from_file(file_path)
    eprint(f"OCR method: {ocr.method} | Confidence: {ocr.confidence}% | Text length: {len(ocr.text)}")

    if not ocr.text.strip():
        eprint("ERROR: No text extracted from file. Check OCR dependencies (tesseract, poppler).")
        sys.exit(1)

    # Parse
    invoice = parse_invoice_text(ocr)
    eprint(f"Parsed invoice: {invoice.invoice_number or '(unknown)'} from {invoice.vendor_name or '(unknown)'}")

    # Validate
    warnings = validate_invoice(invoice)
    if warnings:
        for w in warnings:
            eprint(f"  ⚠ {w}")
        if invoice.extraction_confidence < 85 or invoice.ocr_confidence < 85:
            invoice.status = "needs_review"

    # Check duplicate
    excel_path = os.path.abspath(args.output)
    if not args.dry_run and check_duplicate(invoice, excel_path):
        eprint(f"Duplicate detected: {invoice.invoice_number} from {invoice.vendor_name}")
        invoice.status = "duplicate"

    if args.dry_run:
        data = asdict(invoice)
        data.pop("raw_text", None)
        data["warnings"] = warnings
        print(json.dumps(data, ensure_ascii=False, indent=2))
        return

    if args.format in ("excel", "both"):
        append_to_excel(invoice, excel_path, invoice.status)
        eprint(f"Appended to: {excel_path}")

    if args.format in ("json", "both"):
        json_dir = args.json_dir or os.path.dirname(excel_path) or "."
        json_path = save_json_backup(invoice, json_dir)
        eprint(f"JSON backup: {json_path}")

    result = {
        "invoice_number": invoice.invoice_number,
        "vendor": invoice.vendor_name,
        "total": invoice.total_amount,
        "status": invoice.status,
        "ocr_confidence": invoice.ocr_confidence,
        "ocr_method": invoice.ocr_method,
        "extraction_confidence": invoice.extraction_confidence,
        "warnings": warnings,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
