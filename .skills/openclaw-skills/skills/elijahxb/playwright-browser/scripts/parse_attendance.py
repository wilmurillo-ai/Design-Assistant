#!/usr/bin/env python3
"""
Parse attendance records and save to Excel.
Each source file becomes a sheet named after the file.
"""

import os
import re
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Paths
INPUT_DIR = os.path.join(os.path.expanduser("~"), "Desktop", "考勤")
OUTPUT_FILE = os.path.join(INPUT_DIR, "考勤汇总.xlsx")

# Attendance type patterns (Chinese)
PATTERNS = {
    "请假": re.compile(r"请假\s*(\d+\.?\d*)\s*天"),
    "年假": re.compile(r"年假\s*(\d+\.?\d*)\s*天"),
    "加班": re.compile(r"加班\s*(\d+\.?\d*)\s*天"),
    "事假": re.compile(r"事假\s*(\d+\.?\d*)\s*天"),
    "病假": re.compile(r"病假\s*(\d+\.?\d*)\s*天"),
    "调休": re.compile(r"调休\s*(\d+\.?\d*)\s*天"),
    "迟到": re.compile(r"迟到\s*(\d+\.?\d*)\s*[天次]"),
    "早退": re.compile(r"早退\s*(\d+\.?\d*)\s*[天次]"),
    "旷工": re.compile(r"旷工\s*(\d+\.?\d*)\s*天"),
}


def parse_line(line):
    """Parse a single attendance record line"""
    line = line.strip()
    if not line:
        return None

    # Split name and details: name is separated by space(s) from the rest
    parts = re.split(r'\s+', line, maxsplit=1)
    if len(parts) < 2:
        return {"姓名": parts[0], "原始记录": line, "满勤": "是"}

    name = parts[0]
    detail = parts[1]

    record = {
        "姓名": name,
        "原始记录": detail,
        "满勤": "",
    }

    # Check for 满勤
    if "满勤" in detail:
        record["满勤"] = "是"

    # Extract each attendance type
    for att_type, pattern in PATTERNS.items():
        match = pattern.search(detail)
        if match:
            record[att_type] = float(match.group(1))
            if record["满勤"] == "":
                record["满勤"] = "否"

    if record["满勤"] == "":
        record["满勤"] = "否"

    return record


def parse_file(filepath):
    """Parse an entire attendance file"""
    # Try multiple encodings
    for encoding in ['utf-8', 'utf-8-sig', 'gbk', 'gb2312', 'gb18030']:
        try:
            with open(filepath, 'r', encoding=encoding) as f:
                text = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        print(f"  [!] Cannot decode file: {filepath}")
        return []

    records = []
    for line in text.strip().splitlines():
        record = parse_line(line)
        if record:
            records.append(record)

    return records


def create_excel(all_data, output_path):
    """Create Excel workbook with one sheet per source file"""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="微软雅黑", size=10)
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_left = Alignment(horizontal="left", vertical="center")

    full_mark_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    absence_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Red
    overtime_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")    # Blue

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Column definitions
    columns = ["序号", "姓名", "满勤", "请假", "年假", "事假", "病假", "加班", "调休", "迟到", "早退", "旷工", "原始记录"]

    for sheet_name, records in all_data.items():
        ws = wb.create_sheet(title=sheet_name)

        # Write title row
        title_text = f"考勤明细 - {sheet_name}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # Write headers (row 2)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[2].height = 25

        # Write data rows
        for row_idx, record in enumerate(records, 1):
            data_row = row_idx + 2  # offset for title + header

            values = [
                row_idx,                        # 序号
                record.get("姓名", ""),          # 姓名
                record.get("满勤", ""),           # 满勤
                record.get("请假", ""),           # 请假
                record.get("年假", ""),           # 年假
                record.get("事假", ""),           # 事假
                record.get("病假", ""),           # 病假
                record.get("加班", ""),           # 加班
                record.get("调休", ""),           # 调休
                record.get("迟到", ""),           # 迟到
                record.get("早退", ""),           # 早退
                record.get("旷工", ""),           # 旷工
                record.get("原始记录", ""),        # 原始记录
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=data_row, column=col_idx, value=value if value != "" else None)
                cell.font = data_font
                cell.border = thin_border

                if col_idx <= 3 or (col_idx >= 4 and col_idx <= 12):
                    cell.alignment = data_align_center
                else:
                    cell.alignment = data_align_left

                # Conditional coloring
                if col_idx == 3:  # 满勤 column
                    if value == "是":
                        cell.fill = full_mark_fill
                    elif value == "否":
                        cell.fill = absence_fill

                # Highlight overtime
                if col_idx == 8 and value and value != "":  # 加班
                    cell.fill = overtime_fill

                # Highlight numeric absence values > 0
                if col_idx in [4, 5, 6, 7, 10, 11, 12] and isinstance(value, (int, float)) and value > 0:
                    cell.fill = absence_fill

        # Auto-adjust column widths
        col_widths = {
            1: 6,    # 序号
            2: 10,   # 姓名
            3: 8,    # 满勤
            4: 8,    # 请假
            5: 8,    # 年假
            6: 8,    # 事假
            7: 8,    # 病假
            8: 8,    # 加班
            9: 8,    # 调休
            10: 8,   # 迟到
            11: 8,   # 早退
            12: 8,   # 旷工
            13: 30,  # 原始记录
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = width

        # Freeze panes (freeze header)
        ws.freeze_panes = "A3"

        # Summary row
        summary_row = len(records) + 3
        ws.cell(row=summary_row, column=1, value="").border = thin_border
        summary_cell = ws.cell(row=summary_row, column=2, value="合计")
        summary_cell.font = Font(name="微软雅黑", bold=True, size=10)
        summary_cell.alignment = data_align_center
        summary_cell.border = thin_border

        # Count 满勤
        full_count = sum(1 for r in records if r.get("满勤") == "是")
        ws.cell(row=summary_row, column=3, value=f"{full_count}/{len(records)}人").font = Font(name="微软雅黑", bold=True, size=10)
        ws.cell(row=summary_row, column=3).alignment = data_align_center
        ws.cell(row=summary_row, column=3).border = thin_border

        # Sum numeric columns
        for col_idx, key in [(4, "请假"), (5, "年假"), (6, "事假"), (7, "病假"), (8, "加班"), (9, "调休"), (10, "迟到"), (11, "早退"), (12, "旷工")]:
            total = sum(r.get(key, 0) for r in records if isinstance(r.get(key), (int, float)))
            cell = ws.cell(row=summary_row, column=col_idx, value=total if total > 0 else None)
            cell.font = Font(name="微软雅黑", bold=True, size=10)
            cell.alignment = data_align_center
            cell.border = thin_border

        ws.cell(row=summary_row, column=13, value="").border = thin_border

        print(f"  [Sheet] {sheet_name}: {len(records)} records")

    wb.save(output_path)
    print(f"\n[*] Excel saved: {output_path}")


def main():
    print(f"[*] Scanning: {INPUT_DIR}")

    # Find all text files
    files = sorted(glob.glob(os.path.join(INPUT_DIR, "*.txt")))

    if not files:
        print("[!] No .txt files found!")
        return

    print(f"[*] Found {len(files)} file(s):\n")

    all_data = {}
    for filepath in files:
        filename = os.path.splitext(os.path.basename(filepath))[0]
        print(f"  Processing: {filename}")

        records = parse_file(filepath)
        if records:
            all_data[filename] = records
            for r in records:
                print(f"    {r['姓名']:>4s} | 满勤:{r.get('满勤','?'):>1s} | ", end="")
                details = []
                for key in ["请假", "年假", "事假", "病假", "加班", "调休", "迟到", "早退", "旷工"]:
                    val = r.get(key)
                    if isinstance(val, (int, float)) and val > 0:
                        details.append(f"{key}:{val}天")
                print(", ".join(details) if details else "满勤")
        print()

    if all_data:
        create_excel(all_data, OUTPUT_FILE)
    else:
        print("[!] No valid records found.")


if __name__ == "__main__":
    main()
