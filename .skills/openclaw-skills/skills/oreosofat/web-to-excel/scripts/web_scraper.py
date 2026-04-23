#!/usr/bin/env python3
"""
web_scraper.py — 从网页抓取参数并写入 Excel（通用版）

用法:
  python3 web_scraper.py <excel_path> <sheet_name> <start_row> <end_row> <url> [col_index]

参数:
  excel_path   Excel 文件路径（支持 ~ 展开）
  sheet_name   工作表名称
  start_row    起始行（整数）
  end_row      结束行（整数）
  url          目标 URL（参数配置页）
  col_index    网页列索引（0=第1列车型，1=第2列...），默认 0

字段映射由用户在使用时通过对话提供，追加到 FIELD_MAP 即可。
"""
import sys, os, importlib, subprocess, time

# ── 自动安装依赖 ──────────────────────────────────────────────────────────
for pkg in ['websockets', 'openpyxl']:
    if importlib.util.find_spec(pkg) is None:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])

# ── 通用字段映射（用户可在使用时追加/覆盖）──────────────────────────────────
# 用法：在调用本脚本前，在对话中确认映射关系，然后我来执行写入。
# 下方为最常见的摩托范网站字段，仅作参考示例：

FIELD_MAP = {
    # 动力系统
    '电机最大(峰值)功率kW)': 'AI',
    '最大功率':              'AI',
    '电机额定功率(kW)':      'AJ',
    '额定功率':              'AJ',
    '额定电压(V)':           'AK',
    '电压':                  'AK',
    '电池能量(kWh/度)':     'AL',
    '电池容量':              'AL',
    '电机最大扭矩(N·m)':     'AN',
    '最大扭矩':              'AN',
    # 续航
    '官方续航里程(km)':      'AP',
    '续航':                  'AP',
    # 传动/车架
    '传动方式':              'AS',
    '驱动形式':              'AS',
    '电机布局':              'AS',
    '车架型式':              'AT',
    '车架类型':              'AT',
    # 悬挂
    '前悬挂系统':            'AV',
    '前悬挂':                'AV',
    '后悬挂系统':            'AW',
    '后悬挂':                'AW',
    # 制动
    '前制动系统':            'AZ',
    '前制动':                'AZ',
    '后制动系统':            'BA',
    '后制动':                'BA',
    # 车轮/轮胎
    '前轮规格':              'BE',
    '前轮胎':                'BE',
    '后轮规格':              'BF',
    '后轮胎':                'BF',
    '轴距(mm)':              'BB',
    '轴距':                  'BB',
    # 质量/速度
    '整备质量(kg)':          'BI',
    '整备质量':              'BI',
    '最大允许总质量(kg)':    'BJ',
    '最大有效载荷(kg)':      'BM',
    '载荷':                  'BM',
    '最高车速(km/h)':       'BK',
    '最高车速':              'BK',
    # 价格
    '参考价':                'BO',
    '价格':                  'BO',
}


def col_to_num(letter: str) -> int:
    """列字母 → 列号（A=1, Z=26, AA=27...）"""
    num = 0
    for c in letter.upper():
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num


def clean_val(v) -> any:
    """清洗值：去除¥符号、逗号、空格，暂无报价→跳过"""
    if v is None: return None
    s = str(v).strip().replace('¥', '').replace(',', '').replace(' ', '')
    if s in ('-', '暂无报价', '', 'None', '—'): return None
    try: return float(s)
    except: return s.strip() or None


def parse_params(text: str) -> tuple:
    """
    解析网页参数文本（制表符分隔）→ (field_data, column_names)

    field_data    : {字段名: [col0值, col1值, ...]}
    column_names  : [第1列车型名, 第2列车型名, ...]
    """
    lines = [l for l in text.strip().split('\n') if l.strip() and '\t' in l]
    if not lines:
        return {}, []
    header = lines[0].split('\t')
    column_names = header[1:]              # 跳过第一列（字段名列）
    field_data = {}
    for line in lines[1:]:
        parts = line.split('\t')
        if len(parts) < 2: continue
        field = parts[0].strip()
        values = parts[1:]
        if field not in field_data:
            field_data[field] = values
    return field_data, column_names


def write_row(sheet, row_num: int, field_data: dict,
              col_idx: int, field_map: dict) -> list:
    """向指定行写入数据，返回更新列表"""
    updates = []
    for web_field, excel_col in field_map.items():
        if web_field not in field_data: continue
        vals = field_data[web_field]
        if col_idx >= len(vals): continue
        val = clean_val(vals[col_idx])
        if val is None: continue
        col_num = col_to_num(excel_col)
        old = sheet.cell(row_num, col_num).value
        if old is None:                  # 只填空值，不覆盖已有
            sheet.cell(row_num, col_num).value = val
            # 找对应参考价（若存在）
            ref_price = ''
            if '参考价' in field_data and col_idx < len(field_data['参考价']):
                ref_price = field_data['参考价'][col_idx]
            updates.append((row_num, excel_col, col_num, ref_price, None, val))
    return updates


# ── 主程序入口 ─────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if len(sys.argv) < 6:
        print("用法: python3 web_scraper.py <excel> <sheet> <start_row> <end_row> <url> [col_index=0]")
        sys.exit(1)

    excel_path  = os.path.expanduser(sys.argv[1])
    sheet_name  = sys.argv[2]
    start_row   = int(sys.argv[3])
    end_row     = int(sys.argv[4])
    url         = sys.argv[5]
    col_index   = int(sys.argv[6]) if len(sys.argv) > 6 else 0

    print(f'\n{"="*60}')
    print(f'目标文件 : {excel_path}')
    print(f'Sheet    : {sheet_name}')
    print(f'填写行   : {start_row}～{end_row}')
    print(f'目标URL  : {url}')
    print(f'列索引   : {col_index}（0=第1列）')
    print(f'{"="*60}\n')

    # ① 连接浏览器
    sys.path.insert(0,
        os.path.expanduser('~/Library/Application Support/QClaw/openclaw/config/skills/browser-cdp/scripts'))
    from cdp_client import CDPClient

    client = CDPClient('http://127.0.0.1:9334')
    client.connect()
    tabs = client.list_tabs()

    target_tab = None
    for t in tabs:
        if url[:40] in t['url']:
            target_tab = t
            print(f'  → 复用已有标签: {t["id"]}')
            break

    if not target_tab:
        client.create_tab(url)
        time.sleep(3)
        tabs = client.list_tabs()
        for t in tabs:
            if url[:40] in t['url']:
                target_tab = t
                print(f'  → 新建标签: {t["id"]}')
                break

    if not target_tab:
        print('❌ 无法打开目标页面'); sys.exit(1)

    client.attach(target_tab['id'])
    time.sleep(2)

    # ② 抓取参数
    result = client.send("Runtime.evaluate", {
        "expression": "(function(){"
            "var t=document.body.innerText;"
            "var i=t.indexOf('参考价');"
            "return i>=0 ? t.substring(i,i+25000) : 'NO_DATA';"
        "})()"
    })
    text = result.get('result', {}).get('value', '')
    if not text or text == 'NO_DATA':
        print('❌ 未找到参数数据，请检查页面是否已加载或URL是否正确')
        sys.exit(1)
    print(f'  → 获取 {len(text)} chars 参数数据\n')

    # ③ 解析
    field_data, column_names = parse_params(text)
    print(f'  网页列头: {column_names}')
    print(f'  字段数量: {len(field_data)} 个\n')

    # ④ 写入 Excel
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]

    all_updates = []
    for row_num in range(start_row, end_row + 1):
        name = sheet.cell(row_num, 10).value   # J列=产品/车型名
        ups = write_row(sheet, row_num, field_data, col_index, FIELD_MAP)
        all_updates.extend(ups)
        if ups:
            print(f'  ✅ 行{row_num} {str(name):<30} 更新了 {len(ups)} 个字段')

    if not all_updates:
        print('  ⚠️  无新数据写入（目标单元格可能已有数据，或字段名不匹配）')

    wb.save(excel_path)
    print(f'\n✅ 已保存: {excel_path}')

    # ⑤ 输出摘要
    print(f'\n{"="*60}')
    print(f'📊 数据填充完成报告')
    print(f'文件   : {excel_path}')
    print(f'Sheet  : {sheet_name}')
    print(f'填写范围: 第 {start_row}～{end_row} 行')
    print(f'网页列索引: {col_index}')
    if column_names:
        print(f'网页列名 : {column_names}')
    print(f'{"="*60}')
    if all_updates:
        print(f'{"行":<5} {"列":<6} {"字段":<30} {"新值"}')
        print('-' * 60)
        for row, col, col_num, ref_price, old, new in all_updates:
            print(f'  行{row:<3} {col:<6} {new}')
    else:
        print('（无更新）')
