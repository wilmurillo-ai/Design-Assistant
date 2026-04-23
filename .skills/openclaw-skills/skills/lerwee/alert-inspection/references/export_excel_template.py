#!/usr/bin/env python3

"""
运行时导出模板

使用方式：
1. agent 在实际执行巡检时，先把本次数据写入临时 JSON 文件
2. 再基于本模板生成一个临时脚本，例如 /tmp/alert_inspection_export_20260307_160500.py
3. 将下方占位符替换为本次真实文件路径和环境名称
4. 执行临时脚本导出 Excel

必须替换的占位符：
- {{HOSTS_JSON}}
- {{PROBLEMS_JSON}}
- {{OUTPUT_XLSX}}
- {{ENVIRONMENT_NAME}}
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


HOSTS_JSON = Path("{{HOSTS_JSON}}")
PROBLEMS_JSON = Path("{{PROBLEMS_JSON}}")
OUTPUT_XLSX = Path("{{OUTPUT_XLSX}}")
ENVIRONMENT_NAME = "{{ENVIRONMENT_NAME}}"
CURRENT_TIME = "{{CURRENT_TIME}}"

PRIORITY_LABELS = {
    5: "紧急(P5)",
    4: "严重(P4)",
    3: "次要(P3)",
    2: "警告(P2)",
    1: "信息(P1)",
}


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def adjust_widths(ws):
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 10), 40)


def to_int(value, default=0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def main():
    hosts = read_json(HOSTS_JSON)
    problems = read_json(PROBLEMS_JSON)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "巡检概览"

    total = len(problems)
    counts = {level: 0 for level in PRIORITY_LABELS}
    for item in problems:
        level = int(item.get("priority", 0) or 0)
        if level in counts:
            counts[level] += 1

    normal_host_count = len([h for h in hosts if to_int(h.get("active_status"), -1) == 0])

    overview_rows = [
        [f"🚨 设备健康巡检报告 · {ENVIRONMENT_NAME}"],
        ["巡检时间：", CURRENT_TIME],
        ["报告生成:", "lerwee运维智能体"],
        [],
        ["📊 告警概览"],
        ["告警等级", "数量", "占比"],
        ["紧急(P5)", counts[5], f"{(counts[5] / total * 100):.2f}%" if total else "0.00%"],
        ["严重(P4)", counts[4], f"{(counts[4] / total * 100):.2f}%" if total else "0.00%"],
        ["次要(P3)", counts[3], f"{(counts[3] / total * 100):.2f}%" if total else "0.00%"],
        ["警告(P2)", counts[2], f"{(counts[2] / total * 100):.2f}%" if total else "0.00%"],
        ["信息(P1)", counts[1], f"{(counts[1] / total * 100):.2f}%" if total else "0.00%"],
        ["合计", total, "100.00%" if total else "0.00%"],
        [],
        ["📌 巡检结论"],
        [f"● 🔴 {counts[5] + counts[4]}条高危告警需立即处理"],
        [f"● ⚠️ {counts[3] + counts[2]}条告警需关注"],
        [f"● ✅ {normal_host_count}台主机无告警"],
    ]
    for row in overview_rows:
        ws1.append(row)

    ws2 = wb.create_sheet("正常主机")
    ws2.append(["主机名", "IP", "监控类型", "监控状态", "采集状态"])
    for host in hosts:
        if to_int(host.get("active_status"), -1) != 0:
            continue
        ws2.append([
            host.get("name", ""),
            host.get("ip", ""),
            host.get("classification", ""),
            host.get("active_status_label") or host.get("monitor_status", ""),
            host.get("power_label") or host.get("collect_status", ""),
        ])

    ws3 = wb.create_sheet("异常主机")
    ws3.append(["主机名", "IP", "监控类型", "监控状态", "采集状态"])
    for host in hosts:
        if to_int(host.get("active_status"), -1) == 0:
            continue
        ws3.append([
            host.get("name", ""),
            host.get("ip", ""),
            host.get("classification", ""),
            host.get("active_status_label") or host.get("monitor_status", ""),
            host.get("power_label") or host.get("collect_status", ""),
        ])

    ws4 = wb.create_sheet("异常详细清单")
    ws4.append(["主机名", "IP", "告警描述", "告警等级", "告警时间", "持续时长"])
    for item in problems:
        ws4.append([
            item.get("name", ""),
            item.get("ip", ""),
            item.get("description", ""),
            PRIORITY_LABELS.get(int(item.get("priority", 0) or 0), item.get("priority", "")),
            item.get("clock", ""),
            item.get("duration", ""),
        ])

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        adjust_widths(ws)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(str(OUTPUT_XLSX))


if __name__ == "__main__":
    main()
