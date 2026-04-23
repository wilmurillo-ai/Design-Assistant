"""
Excel English-to-Chinese Translator
=====================================
Translates text in an Excel (.xlsx) file from English to Chinese.
- Sends one sheet at a time to the LLM (batched by sheet).
- Chinese translation is appended below the original English text in the SAME cell.
- All formatting, merged cells, images, and styles are preserved via openpyxl.

Usage:
    python3 translate.py <input_file.xlsx>
    python3 translate.py <input_file.xlsx> -o <output_file.xlsx>

Credentials:
    Read from ~/.openai/secret (KEY=VALUE format), fallback to environment variables.
    OPENAI_API_KEY   - Your API key (required)
    OPENAI_BASE_URL  - API base URL (optional)
"""

import os
import re
import argparse
import openpyxl
from openpyxl.styles import Alignment
from openai import OpenAI

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
MODEL_NAME = "qwen3.5-flash"
MAX_TOKENS_PER_BATCH = 80000   # stay well under 128K context limit

# ---------------------------------------------------------------------------
# Read credentials from ~/.openai/secret (KEY=VALUE format), fallback to env
# ---------------------------------------------------------------------------
def _load_secret(path: str = "~/.openai/secret") -> dict:
    secrets = {}
    expanded = os.path.expanduser(path)
    if os.path.exists(expanded):
        with open(expanded) as f:
            for line in f:
                line = line.strip()
                if line and "=" in line and not line.startswith("#"):
                    k, _, v = line.partition("=")
                    secrets[k.strip()] = v.strip()
    return secrets

_secrets = _load_secret()
_api_key = _secrets.get("OPENAI_API_KEY") or os.environ.get("OPENAI_API_KEY", "")
_base_url = _secrets.get("OPENAI_BASE_URL") or os.environ.get("OPENAI_BASE_URL", "https://api.openai.com/v1")

client = OpenAI(api_key=_api_key, base_url=_base_url)

# ---------------------------------------------------------------------------
# Batch translate: send all cells from one sheet in a single request
# ---------------------------------------------------------------------------

def translate_batch(texts: list[str]) -> dict[int, str]:
    """Send a numbered list of texts to the LLM, return {index: translated}."""
    numbered = "\n".join(f"{i+1}. {t}" for i, t in enumerate(texts))
    prompt = (
        "You are a professional translator. Translate each numbered English item into Chinese.\n"
        "Rules:\n"
        "- Output ONLY the translations, one per line, keeping the same number prefix.\n"
        "- Format: '1. 翻译内容'\n"
        "- Do NOT add explanations, notes, or extra lines.\n"
        "- If an item is already Chinese or doesn't need translation, output it as-is.\n\n"
        f"{numbered}"
    )
    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
        )
        raw = response.choices[0].message.content.strip()
    except Exception as exc:
        print(f"  [warn] Batch API error: {exc}")
        return {}

    results = {}
    for line in raw.splitlines():
        m = re.match(r"^(\d+)\.\s*(.*)", line.strip())
        if m:
            idx = int(m.group(1)) - 1
            results[idx] = m.group(2).strip()
    return results


# ---------------------------------------------------------------------------
# Core processing
# ---------------------------------------------------------------------------

def translate_sheet(wb, sheet_name: str) -> int:
    """Translate all text cells in a sheet. Returns number of cells translated."""
    sheet = wb[sheet_name]

    # Collect translatable cells
    cells = []  # list of (coordinate, original_text)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                text = cell.value.strip()
                if text and not text.isdigit() and len(text) >= 2:
                    cells.append((cell.coordinate, cell.value))

    if not cells:
        print(f"  Sheet '{sheet_name}': no translatable cells, skipping.")
        return 0

    print(f"  Sheet '{sheet_name}': {len(cells)} cells to translate...")

    # Split into batches to avoid hitting context/timeout limits
    BATCH_SIZE = 100
    translated_count = 0

    for batch_start in range(0, len(cells), BATCH_SIZE):
        batch = cells[batch_start: batch_start + BATCH_SIZE]
        texts = [t for _, t in batch]

        results = translate_batch(texts)

        for local_idx, (coord, original) in enumerate(batch):
            translated = results.get(local_idx, "")
            if not translated or translated == original:
                continue
            # Skip if already translated (idempotent re-run)
            if translated in original:
                continue
            cell = sheet[coord]
            cell.value = f"{original}\n{translated}"
            existing = cell.alignment
            cell.alignment = Alignment(
                horizontal=existing.horizontal,
                vertical=existing.vertical,
                text_rotation=existing.text_rotation,
                wrap_text=True,
                shrink_to_fit=existing.shrink_to_fit,
                indent=existing.indent,
            )
            translated_count += 1

    print(f"  Sheet '{sheet_name}': {translated_count} cells written.")
    return translated_count


def translate_excel(input_path: str, output_path: str) -> None:
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path}")

    print(f"Loading: {input_path}")
    wb = openpyxl.load_workbook(input_path)

    total = 0
    for sheet_name in wb.sheetnames:
        total += translate_sheet(wb, sheet_name)

    print(f"Saving: {output_path}  (total {total} cells translated)")
    wb.save(output_path)
    print("Done!")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Translate an Excel file from English to Chinese (sheet-by-sheet batch)."
    )
    parser.add_argument("input", help="Path to the input .xlsx file.")
    parser.add_argument("-o", "--output",
                        help="Output file path. Defaults to <input>_translated.xlsx.")
    args = parser.parse_args()

    if not args.output:
        name, ext = os.path.splitext(args.input)
        args.output = f"{name}_translated{ext}"

    translate_excel(args.input, args.output)


if __name__ == "__main__":
    main()
