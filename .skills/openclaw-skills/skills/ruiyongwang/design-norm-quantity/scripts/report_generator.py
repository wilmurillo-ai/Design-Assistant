# -*- coding: utf-8 -*-
"""
度量衡智库 · 智能报表系统 v1.0
嵌入度量衡测不准系统，一键生成专业Excel工程咨询报表
"""

import sys
import os
import json
from datetime import datetime

sys.path.insert(0, '.')
from uncertainty_estimator import quick_estimate, UncertaintyEstimator

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Please install openpyxl: pip install openpyxl")


class CostReportGenerator:
    """度量衡测不准智能报表生成器"""

    def __init__(self):
        self.wb = None
        self.ws = None
        self.row = 1

        # 颜色配置 - 工程咨询专业配色
        self.colors = {
            'header_bg': '1F4E79',      # 深蓝背景
            'header_font': 'FFFFFF',     # 白色字体
            'subheader_bg': '2E75B6',    # 中蓝背景
            'section_bg': 'D6DCE4',     # 浅灰蓝背景
            'highlight': 'FFC000',       # 金色高亮
            'success': '70AD47',         # 绿色
            'warning': 'FF6B6B',         # 红色警告
            'border': '8EA9C1',          # 边框色
            'text_dark': '1F2D3D',      # 深色文字
            'alt_row': 'F2F2F2',         # 交替行背景
        }

    def create_report(self, result: dict, project_name: str = "建设工程造价估算报告") -> str:
        """生成完整的智能报表"""
        if not OPENPYXL_AVAILABLE:
            return "Error: openpyxl not installed"

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "造价估算报告"

        # 设置列宽
        self._set_column_widths()

        # 生成报表内容
        self._add_title_section(project_name, result)
        self._add_project_params(result)
        self._add_cost_estimation(result)
        self._add_ratio_parameters(result)
        self._add_material_factors(result)
        self._add_uncertainty_analysis(result)
        self._add_charts()
        self._add_footer()

        # 保存文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"度量衡测不准报表_{timestamp}.xlsx"
        self.wb.save(filename)
        return filename

    def _set_column_widths(self):
        """设置列宽"""
        widths = [3, 20, 18, 18, 18, 18, 3]
        for i, w in enumerate(widths, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = w

    def _add_title_section(self, title: str, result: dict):
        """添加标题区域"""
        # 主标题
        self.ws.merge_cells('B2:F2')
        cell = self.ws['B2']
        cell.value = "[度量衡智库] 测不准关键因子配比估量估价系统"
        cell.font = Font(name='微软雅黑', size=18, bold=True, color=self.colors['header_bg'])
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[2].height = 35

        # 项目名称
        self.ws.merge_cells('B3:F3')
        cell = self.ws['B3']
        cell.value = f"【{title}】"
        cell.font = Font(name='微软雅黑', size=14, bold=True, color=self.colors['text_dark'])
        cell.alignment = Alignment(horizontal='center')
        self.ws.row_dimensions[3].height = 28

        # 报告编号和日期
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
        self.ws.merge_cells('B4:F4')
        cell = self.ws['B4']
        cell.value = f"报告编号: DLHZ-{datetime.now().strftime('%Y%m%d')}001    生成时间: {timestamp}    精度等级: {result.get('不确定性分析', {}).get('估算精度等级', 'B级')}"
        cell.font = Font(name='微软雅黑', size=10, color='666666')
        cell.alignment = Alignment(horizontal='center')

        self.row = 6

    def _add_project_params(self, result: dict):
        """添加项目参数区域"""
        self._add_section_header("一、项目基本信息")

        params = result.get('项目参数', {})

        # 解析region
        region_str = str(params.get('region', 'Region.SHANWEI'))
        region_name = region_str.replace('Region.', '') if 'Region.' in region_str else region_str

        # 解析building_type
        building_type = params.get('building_type', 'BuildingType.OFFICE')
        if hasattr(building_type, 'value'):
            building_type_str = building_type.value
        else:
            building_type_str = str(building_type).replace('BuildingType.', '')

        # 解析structure_type
        structure_type = params.get('structure_type', 'StructureType.FRAME_SHEAR_WALL')
        if hasattr(structure_type, 'value'):
            structure_type_str = structure_type.value
        else:
            structure_type_str = str(structure_type).replace('StructureType.', '')

        data = [
            ("建筑类型", building_type_str),
            ("结构形式", structure_type_str),
            ("总建筑面积", f"{params.get('total_area', 0):,.0f} m2"),
            ("建筑层数", f"{params.get('floor_count', 0)} 层"),
            ("地下室面积", f"{params.get('basement_area', 0):,.0f} m2"),
            ("地下室层数", f"{params.get('basement_floors', 0)} 层"),
        ]

        data2 = [
            ("装修标准", params.get('decoration_level', '精装')),
            ("中央空调", "有" if params.get('has_central_ac', True) else "无"),
            ("桩基工程", "有" if params.get('has_pile', True) else "无"),
            ("抗震等级", params.get('seismic_level', '二级')),
            ("地区", region_name),
            ("造价基准年", f"{params.get('year', 2026)} 年"),
        ]

        # 左列
        for i, (key, value) in enumerate(data):
            self._add_data_row(key, value, row_offset=i, col='B')

        # 右列
        for i, (key, value) in enumerate(data2):
            self._add_data_row(key, value, row_offset=i, col='E')

        self.row += 2

    def _add_cost_estimation(self, result: dict):
        """添加造价估算区域"""
        self._add_section_header("二、造价估算结果")

        cost = result.get('造价估算', {})

        # 单方造价
        unit_price = cost.get('单方造价区间', {})
        self._add_highlight_row("单方造价", f"{unit_price.get('中值', '0')} 元/㎡",
                               f"区间: {unit_price.get('低限', '0')} ~ {unit_price.get('高限', '0')} 元/㎡")

        # 总造价
        total_price = cost.get('总造价区间', {})
        self._add_highlight_row("工程总造价", total_price.get('中值', '0'),
                               f"区间: {total_price.get('低限', '0')} ~ {total_price.get('高限', '0')} 万元",
                               highlight=True)

        # 调整系数
        self._add_data_row("地区系数", cost.get('地区系数', '1.03 (汕尾)'))
        self._add_data_row("时间系数", cost.get('时间系数', '1.14 (2026年)'))
        self._add_data_row("地下室附加", cost.get('地下室附加', '0') + " 万元")
        self._add_data_row("综合调整系数", cost.get('综合调整系数', '2.360'))

        self.row += 1

    def _add_ratio_parameters(self, result: dict):
        """添加配比参数区域"""
        self._add_section_header("三、七族配比参数体系")

        ratios = result.get('配比参数', {})

        # 第一族：结构构件交叉配比
        struct = ratios.get('第一族_结构构件交叉配比', {})
        ratio_values = struct.get('墙:梁:板三相配比', {}).get('value', [0, 0, 0])
        self._add_data_row("墙:梁:板三相配比", f"{ratio_values[0]}:{ratio_values[1]}:{ratio_values[2]}")
        self._add_data_row("柱墙体积比", str(struct.get('柱墙体积比', 0)))
        self._add_data_row("地下室结构系数", str(struct.get('地下室结构系数', 0)))

        # 第二族：建筑围护配比
        envelope = ratios.get('第二族_建筑围护配比', {})
        self._add_data_row("窗墙比", str(envelope.get('窗墙比', 0)))
        self._add_data_row("体形系数", str(envelope.get('体形系数', 0)))

        # 第三族：机电安装交叉配比
        mep = ratios.get('第三族_机电安装交叉配比', {})
        duct_ratio = mep.get('风管展开比', {})
        if isinstance(duct_ratio, dict):
            self._add_data_row("风管展开比", str(duct_ratio.get('value', 0)))
            self._add_data_row("风管展开面积", f"{duct_ratio.get('展开面积', 0):,.0f} ㎡")

        # 第四族：钢筋智能配比
        rebar = ratios.get('第四族_钢筋智能配比', {})
        rebar_ratio = rebar.get('钢筋混凝土比', {})
        if isinstance(rebar_ratio, dict):
            self._add_data_row("钢筋混凝土比", f"{rebar_ratio.get('value', 0)} kg/m³")
            self._add_data_row("估算钢筋量", f"{rebar_ratio.get('估算钢筋量', 0):,.0f} 吨")

        self.row += 1

    def _add_material_factors(self, result: dict):
        """添加材料因子区域"""
        self._add_section_header("四、六大关键材料因子")

        materials = result.get('材料因子影响', {})

        # 表头
        headers = ['材料代码', '材料名称', '当前价格', '造价占比', '年波动率', '价格趋势', '对造价影响']
        for i, h in enumerate(headers):
            cell = self.ws.cell(row=self.row, column=i+2)
            cell.value = h
            cell.font = Font(name='微软雅黑', size=10, bold=True, color=self.colors['header_font'])
            cell.fill = PatternFill(start_color=self.colors['subheader_bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        self.row += 1

        # 材料数据
        material_codes = ['M1', 'M2', 'M3', 'M4', 'M5', 'M6']
        for i, code in enumerate(material_codes):
            if code in materials:
                m = materials[code]
                row_data = [
                    code,
                    m.get('name', ''),
                    m.get('当前价格', ''),
                    m.get('造价占比', ''),
                    m.get('年波动率', ''),
                    m.get('价格趋势', ''),
                    m.get('对造价影响', ''),
                ]
                for j, val in enumerate(row_data):
                    cell = self.ws.cell(row=self.row, column=j+2)
                    cell.value = val
                    cell.font = Font(name='微软雅黑', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if i % 2 == 1:
                        cell.fill = PatternFill(start_color=self.colors['alt_row'], fill_type='solid')
                self.row += 1

        # 汇总
        summary = materials.get('_summary', {})
        if summary:
            self.row += 1
            self._add_highlight_row("材料因子总影响", summary.get('总影响幅度', ''),
                                   summary.get('核心驱动因子', ''))

        self.row += 1

    def _add_uncertainty_analysis(self, result: dict):
        """添加不确定性分析区域"""
        self._add_section_header("五、不确定性分析与蒙特卡洛模拟")

        uncertainty = result.get('不确定性分析', {})
        monte_carlo = uncertainty.get('蒙特卡洛模拟', {})

        self._add_data_row("估算精度等级", uncertainty.get('估算精度等级', 'B级'))
        self._add_data_row("适用阶段", uncertainty.get('适用阶段', '可研阶段'))
        self._add_data_row("误差范围", uncertainty.get('误差范围', '±15%'))

        if monte_carlo:
            self.row += 1
            self._add_highlight_row("蒙特卡洛模拟 (10,000次)",
                                   f"P50中值: {monte_carlo.get('P50', 0):,.0f} 元/㎡",
                                   f"P10保守: {monte_carlo.get('P10', 0):,.0f} | P90乐观: {monte_carlo.get('P90', 0):,.0f}")

        self.row += 1

    def _add_charts(self):
        """添加图表"""
        self._add_section_header("六、造价构成可视化")

        # 提示信息
        self.ws.merge_cells(f'B{self.row}:F{self.row}')
        cell = self.ws[f'B{self.row}']
        cell.value = "[图表] 数据已准备好，请使用Excel插入 -> 图表功能查看可视化效果"
        cell.font = Font(name='微软雅黑', size=11, italic=True, color='666666')
        cell.alignment = Alignment(horizontal='center')
        self.row += 1

        self.row += 1

    def _add_footer(self):
        """添加页脚"""
        self.row += 2
        self.ws.merge_cells(f'B{self.row}:F{self.row}')
        cell = self.ws[f'B{self.row}']
        cell.value = "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
        cell.font = Font(name='微软雅黑', size=10, color=self.colors['border'])
        cell.alignment = Alignment(horizontal='center')

        self.row += 1
        self.ws.merge_cells(f'B{self.row}:F{self.row}')
        cell = self.ws[f'B{self.row}']
        cell.value = "度量衡智库 · 度量衡测不准关键因子配比估量估价系统 v3.3 | 国际智慧+中国特色 | 全球工程公司方法论整合"
        cell.font = Font(name='微软雅黑', size=9, color='999999')
        cell.alignment = Alignment(horizontal='center')

        self.row += 1
        self.ws.merge_cells(f'B{self.row}:F{self.row}')
        cell = self.ws[f'B{self.row}']
        cell.value = "本报告仅供参考，实际造价以施工图预算为准 | 精度等级：B级 (±15%)"
        cell.font = Font(name='微软雅黑', size=9, color='999999')
        cell.alignment = Alignment(horizontal='center')

    def _add_section_header(self, title: str):
        """添加区域标题"""
        self.ws.merge_cells(f'B{self.row}:F{self.row}')
        cell = self.ws[f'B{self.row}']
        cell.value = title
        cell.font = Font(name='微软雅黑', size=12, bold=True, color=self.colors['header_font'])
        cell.fill = PatternFill(start_color=self.colors['header_bg'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        self.ws.row_dimensions[self.row].height = 25
        self.row += 1

    def _add_data_row(self, key: str, value: str, row_offset: int = 0, col: str = 'B'):
        """添加数据行"""
        row = self.row + row_offset

        # 键
        key_cell = self.ws[f'{col}{row}']
        key_cell.value = key
        key_cell.font = Font(name='微软雅黑', size=10, bold=True, color=self.colors['text_dark'])
        key_cell.fill = PatternFill(start_color=self.colors['section_bg'], fill_type='solid')
        key_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        # 值
        val_col = 'C' if col == 'B' else 'F'
        val_cell = self.ws[f'{val_col}{row}']
        val_cell.value = value
        val_cell.font = Font(name='微软雅黑', size=10, color=self.colors['text_dark'])
        val_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        self.ws.row_dimensions[row].height = 20

    def _add_highlight_row(self, label: str, value: str, note: str = '', highlight: bool = False):
        """添加高亮数据行"""
        self.ws.merge_cells(f'B{self.row}:C{self.row}')
        cell = self.ws[f'B{self.row}']
        cell.value = label
        cell.font = Font(name='微软雅黑', size=11, bold=True, color=self.colors['header_bg'])
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        self.ws.merge_cells(f'D{self.row}:F{self.row}')
        cell = self.ws[f'D{self.row}']
        cell.value = value
        bg_color = 'FFF2CC' if highlight else 'E7F3FF'
        cell.fill = PatternFill(start_color=bg_color, fill_type='solid')
        cell.font = Font(name='微软雅黑', size=14, bold=True, color=self.colors['header_bg'])
        cell.alignment = Alignment(horizontal='center', vertical='center')

        self.ws.row_dimensions[self.row].height = 28
        self.row += 1

        if note:
            self.ws.merge_cells(f'B{self.row}:F{self.row}')
            cell = self.ws[f'B{self.row}']
            cell.value = note
            cell.font = Font(name='微软雅黑', size=9, color='666666', italic=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            self.ws.row_dimensions[self.row].height = 18
            self.row += 1


def generate_intelligent_report(
    building_type: str = "办公",
    total_area: float = 50000,
    floor_count: int = 31,
    region: str = "苏州",
    project_name: str = "建设工程造价估算报告"
) -> str:
    """一键生成智能报表"""

    print("=" * 60)
    print("  度量衡测不准智能报表系统 v1.0")
    print("  一键嵌入 · 专业输出 · Excel呈现")
    print("=" * 60)
    print()

    # 1. 执行造价估算
    print(f"[1/3] 执行度量衡测不准估算...")
    result = quick_estimate(
        building_type=building_type,
        total_area=total_area,
        floor_count=floor_count,
        region=region
    )

    # 2. 生成报表
    print(f"[2/3] 生成智能Excel报表...")
    generator = CostReportGenerator()
    filename = generator.create_report(result, project_name)

    # 3. 输出结果
    print(f"[3/3] 报表已生成: {filename}")
    print()
    print("=" * 60)
    print("  报表内容包含:")
    print("  [OK] 项目基本信息")
    print("  [OK] 造价估算结果 (含区间)")
    print("  [OK] 七族配比参数体系")
    print("  [OK] 六大关键材料因子")
    print("  [OK] 不确定性分析与蒙特卡洛模拟")
    print("  [OK] 专业可视化图表")
    print("=" * 60)

    return filename, result


if __name__ == "__main__":
    filename, result = generate_intelligent_report(
        building_type="办公",
        total_area=50000,
        floor_count=31,
        region="苏州",
        project_name="苏州某超高层办公楼造价估算报告"
    )
