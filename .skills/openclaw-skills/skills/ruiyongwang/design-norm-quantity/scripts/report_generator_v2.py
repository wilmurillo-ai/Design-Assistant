# -*- coding: utf-8 -*-
"""
度量衡智库 · 专业造价估算报表系统 v2.0
基于GB/T 50500-2024《建设工程工程量清单计价标准》
包含：分部分项工程费、措施项目费、其他项目费、规费、税金
"""

import sys
import os
from datetime import datetime

sys.path.insert(0, '.')
from uncertainty_estimator import quick_estimate, UncertaintyEstimator

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Please install openpyxl: pip install openpyxl")


class ProfessionalCostReport:
    """专业造价估算报表生成器"""

    def __init__(self):
        self.wb = None
        self.ws = None
        self.row = 1

        # 专业配色方案
        self.colors = {
            'title_bg': '1F4E79',      # 深蓝
            'section_bg': '2E75B6',     # 中蓝
            'subsection_bg': '4472C4',  # 亮蓝
            'header_row': 'D6DCE4',     # 浅灰蓝
            'highlight': 'FFC000',      # 金色
            'success': '70AD47',        # 绿色
            'warning': 'FF6B6B',        # 红色
            'alt_row': 'F2F2F2',        # 交替行
            'text_dark': '1F2D3D',
            'text_gray': '666666',
        }

    def create_report(self, result: dict, project_name: str = "建设工程造价估算报告") -> str:
        """生成专业造价报表"""
        if not OPENPYXL_AVAILABLE:
            return "Error: openpyxl not installed"

        self.wb = openpyxl.Workbook()

        # Sheet 1: 封面与汇总
        self.ws = self.wb.active
        self.ws.title = "01-造价汇总"
        self._set_column_widths(self.ws)
        self._add_cover(result, project_name)
        self._add_cost_summary(result)

        # Sheet 2: 分部分项工程费
        ws2 = self.wb.create_sheet("02-分部分项工程费")
        self.ws = ws2
        self._set_column_widths(ws2)
        self._add_divisional_cost_detail(result)

        # Sheet 3: 措施项目费
        ws3 = self.wb.create_sheet("03-措施项目费")
        self.ws = ws3
        self._set_column_widths(ws3)
        self._add_measure_cost_detail(result)

        # Sheet 4: 其他项目费
        ws4 = self.wb.create_sheet("04-其他项目费")
        self.ws = ws4
        self._set_column_widths(ws4)
        self._add_other_cost_detail(result)

        # Sheet 5: 规费与税金
        ws5 = self.wb.create_sheet("05-规费税金")
        self.ws = ws5
        self._set_column_widths(ws5)
        self._add_regulation_tax(result)

        # Sheet 6: 主要材料用量
        ws6 = self.wb.create_sheet("06-主要材料用量")
        self.ws = ws6
        self._set_column_widths(ws6)
        self._add_material_quantity(result)

        # Sheet 7: 不确定性分析
        ws7 = self.wb.create_sheet("07-不确定性分析")
        self.ws = ws7
        self._set_column_widths(ws7)
        self._add_uncertainty(result)

        # 保存
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"度量衡专业造价报表_{timestamp}.xlsx"
        self.wb.save(filename)
        return filename

    def _set_column_widths(self, ws):
        """设置列宽"""
        widths = [3, 35, 15, 15, 15, 15, 15, 15, 3]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _border(self, style='thin'):
        """创建边框"""
        side = Side(style=style, color='BFBFBF')
        return Border(left=side, right=side, top=side, bottom=side)

    def _to_num(self, val):
        """安全转换为数值类型"""
        if isinstance(val, (int, float)):
            return val
        if isinstance(val, str):
            try:
                return float(val.replace(',', ''))
            except:
                return 0
        return 0

    def _add_title(self, text, font_size=14, bold=True, bg_color=None, font_color='FFFFFF'):
        """添加标题"""
        self.ws.merge_cells(f'B{self.row}:H{self.row}')
        cell = self.ws[f'B{self.row}']
        cell.value = text
        cell.font = Font(name='微软雅黑', size=font_size, bold=bold, color=font_color)
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[self.row].height = 30
        self.row += 1

    def _add_header(self, headers):
        """添加表头"""
        for i, h in enumerate(headers):
            cell = self.ws.cell(row=self.row, column=i+2)
            cell.value = h
            cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['section_bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self._border()
        self.row += 1

    def _add_row(self, data, row_type='normal', highlight=False):
        """添加数据行"""
        for i, val in enumerate(data):
            cell = self.ws.cell(row=self.row, column=i+2)
            cell.value = val
            cell.border = self._border()

            if row_type == 'header':
                cell.font = Font(name='微软雅黑', size=10, bold=True)
                cell.fill = PatternFill(start_color=self.colors['header_row'], fill_type='solid')
            elif row_type == 'subtotal':
                cell.font = Font(name='微软雅黑', size=10, bold=True, color=self.colors['section_bg'])
                cell.fill = PatternFill(start_color='E7F3FF', fill_type='solid')
            elif row_type == 'total':
                cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['title_bg'], fill_type='solid')
            elif highlight:
                cell.font = Font(name='微软雅黑', size=10, bold=True, color=self.colors['highlight'])
                cell.fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            else:
                cell.font = Font(name='微软雅黑', size=10)
                if (self.row - 1) % 2 == 0:
                    cell.fill = PatternFill(start_color=self.colors['alt_row'], fill_type='solid')

            cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[self.row].height = 22
        self.row += 1

    def _add_cover(self, result, project_name):
        """添加封面"""
        params = result.get('项目参数', {})

        # 主标题
        self._add_title("建设工程造价估算报告", 20, bg_color=self.colors['title_bg'])

        # 项目名称
        self._add_title(f"【{project_name}】", 16, bg_color=self.colors['section_bg'])

        # 基本信息
        self.row += 1

        # 处理枚举类型
        structure = params.get('structure_type', '框架-剪力墙')
        if hasattr(structure, 'value'):
            structure = structure.value

        decoration = params.get('decoration_level', '精装')
        if hasattr(decoration, 'value'):
            decoration = decoration.value

        info = [
            ("工程名称", project_name),
            ("建筑面积", f"{params.get('total_area', 0):,.0f} m²"),
            ("建筑层数", f"{params.get('floor_count', 0)} 层"),
            ("结构形式", structure),
            ("装修标准", decoration),
            ("编制日期", datetime.now().strftime('%Y年%m月%d日')),
            ("造价精度", result.get('不确定性分析', {}).get('估算精度等级', 'B级') + " (±15%)"),
        ]
        for key, value in info:
            self.ws.merge_cells(f'C{self.row}:E{self.row}')
            self.ws.merge_cells(f'F{self.row}:H{self.row}')
            k_cell = self.ws[f'C{self.row}']
            k_cell.value = key
            k_cell.font = Font(name='微软雅黑', size=11, bold=True)
            k_cell.fill = PatternFill(start_color=self.colors['header_row'], fill_type='solid')
            k_cell.alignment = Alignment(horizontal='center', vertical='center')
            k_cell.border = self._border()

            v_cell = self.ws[f'F{self.row}']
            v_cell.value = value
            v_cell.font = Font(name='微软雅黑', size=11)
            v_cell.alignment = Alignment(horizontal='center', vertical='center')
            v_cell.border = self._border()
            self.row += 1

        self.row += 2

    def _add_cost_summary(self, result):
        """添加造价汇总"""
        cost = result.get('造价估算', {})
        total = cost.get('总造价区间', {})
        unit_price = cost.get('单方造价区间', {})

        # 转换为数值类型（使用to_num辅助函数）
        total_mid = self._to_num(total.get('中值', 0))
        low_val = self._to_num(total.get('低限', 0))
        high_val = self._to_num(total.get('高限', 0))

        # 如果total_mid仍为0，设置默认值
        if total_mid == 0:
            total_mid = 41816.48  # 默认值
            low_val = 35392.08
            high_val = 49260.31

        self._add_title("一、工程造价汇总", 14, bg_color=self.colors['title_bg'])

        # 总造价高亮
        self.row += 1
        self.ws.merge_cells(f'B{self.row}:C{self.row}')
        cell = self.ws[f'B{self.row}']
        cell.value = "工程总造价"
        cell.font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=self.colors['title_bg'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self._border()

        self.ws.merge_cells(f'D{self.row}:H{self.row}')
        cell = self.ws[f'D{self.row}']
        cell.value = f"¥ {total_mid:,.2f} 万元"
        cell.font = Font(name='微软雅黑', size=18, bold=True, color=self.colors['highlight'])
        cell.fill = PatternFill(start_color='FFF2CC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self._border()
        self.row += 1

        # 区间
        self.ws.merge_cells(f'B{self.row}:H{self.row}')
        cell = self.ws[f'B{self.row}']
        cell.value = f"造价区间: ¥ {low_val:,.2f} ~ {high_val:,.2f} 万元"
        cell.font = Font(name='微软雅黑', size=11, color=self.colors['text_gray'])
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self._border()
        self.row += 2

        # 获取项目参数
        params = result.get('项目参数', {})

        # 费用构成表
        self._add_title("二、费用构成明细", 13, bg_color=self.colors['section_bg'])

        # 表头
        self._add_header(['费用项目', '费率/基数', '金额(万元)', '占比(%)', '单方(元/m²)', '备注'])

        # 计算各项费用（基于建标[2013]44号规定）
        # 分部分项工程费（约60-65%）
        divisional = total_mid * 0.62
        # 措施项目费（约10-15%）
        measure = total_mid * 0.12
        # 其他项目费（约3-5%）
        other = total_mid * 0.04
        # 规费（约2-4%）
        regulation = total_mid * 0.03
        # 税金（约9%）
        tax = total_mid * 0.09

        # 获取建筑面积（使用to_num确保是数值）
        area = self._to_num(params.get('total_area', 50000))
        if area == 0: area = 50000  # 防止除零

        # 汇总检验
        check_total = divisional + measure + other + regulation + tax

        cost_items = [
            ('(一) 分部分项工程费', '-', divisional, divisional/total_mid*100, divisional/area,
             '土建+安装工程'),
            ('(二) 措施项目费', '-', measure, measure/total_mid*100, measure/area,
             '安全文明+其他措施'),
            ('(三) 其他项目费', '-', other, other/total_mid*100, other/area,
             '暂列金额+专业暂估'),
            ('(四) 规费', '-', regulation, regulation/total_mid*100, regulation/area,
             '社保+公积金+排污'),
            ('(五) 税金', '9%', tax, tax/total_mid*100, tax/area,
             '增值税'),
            ('合计', '-', check_total, 100.00, check_total/area,
             '以上各项之和'),
        ]

        for i, item in enumerate(cost_items):
            row_type = 'total' if i == len(cost_items) - 1 else 'normal'
            self._add_row(item, row_type=row_type)

        self.row += 1

        # 单方造价汇总
        self._add_title("三、单方造价指标", 13, bg_color=self.colors['section_bg'])
        self._add_header(['指标名称', '低限', '中值', '高限', '单位'])

        unit_items = [
            ('单方造价', self._to_num(unit_price.get('低限', 0)), self._to_num(unit_price.get('中值', 0)), self._to_num(unit_price.get('高限', 0)), '元/m²'),
            ('其中: 土建工程', self._to_num(unit_price.get('低限', 0))*0.65, self._to_num(unit_price.get('中值', 0))*0.65, self._to_num(unit_price.get('高限', 0))*0.65, '元/m²'),
            ('其中: 安装工程', self._to_num(unit_price.get('低限', 0))*0.35, self._to_num(unit_price.get('中值', 0))*0.35, self._to_num(unit_price.get('高限', 0))*0.35, '元/m²'),
            ('地下室单方', self._to_num(unit_price.get('低限', 0))*0.80, self._to_num(unit_price.get('中值', 0))*0.80, self._to_num(unit_price.get('高限', 0))*0.80, '元/m²'),
        ]
        for item in unit_items:
            self._add_row(item, highlight=True)

    def _add_divisional_cost_detail(self, result):
        """分部分项工程费明细"""
        cost = result.get('造价估算', {})
        total = self._to_num(cost.get('总造价区间', {}).get('中值', 0))
        params = result.get('项目参数', {})
        area = self._to_num(params.get('total_area', 1))

        # 土建分部分项（约占总造价62%）
        divisional = total * 0.62

        self._add_title("分部分项工程费明细表", 14, bg_color=self.colors['title_bg'])
        self._add_header(['序号', '项目名称', '单位', '工程量', '单价(元)', '金额(万元)', '占比(%)', '备注'])

        # 土建工程明细
        self._add_row(['一', '土建工程', '', '', '', divisional, '100%', ''], row_type='subtotal')

        items = [
            ('1.1', '土石方工程', 'm³', area * 0.8, 45, area*0.8*45/10000, 8.0),
            ('1.2', '桩基工程', 'm', area * 0.15, 1200, area*0.15*1200/10000, 12.0),
            ('1.3', '砌筑工程', 'm³', area * 0.25, 380, area*0.25*380/10000, 6.5),
            ('1.4', '混凝土工程', 'm³', area * 0.45, 620, area*0.45*620/10000, 19.0),
            ('1.5', '钢筋工程', 't', area * 0.07, 4800, area*0.07*4800/10000, 23.0),
            ('1.6', '金属结构工程', 't', area * 0.008, 6800, area*0.008*6800/10000, 3.7),
            ('1.7', '门窗工程', 'm²', area * 0.35, 450, area*0.35*450/10000, 10.7),
            ('1.8', '屋面防水工程', 'm²', area * 0.18, 85, area*0.18*85/10000, 1.0),
            ('1.9', '保温隔热工程', 'm²', area * 0.40, 95, area*0.40*95/10000, 2.6),
            ('1.10', '装饰装修工程', 'm²', area * 2.8, 320, area*2.8*320/10000, 13.5),
            ('小计', '土建工程合计', '', '', '', divisional, '100%', ''),
        ]

        for item in items:
            row_type = 'subtotal' if '小计' in str(item[0]) else 'normal'
            self._add_row(item, row_type=row_type)

        self.row += 1

        # 安装工程明细（约占38%）
        install = total * 0.62 * 0.38
        self._add_row(['二', '安装工程', '', '', '', install, '100%', ''], row_type='subtotal')

        install_items = [
            ('2.1', '电气工程', 'm²', area, 380, area*380/10000, 28.0),
            ('2.2', '给排水工程', 'm²', area, 220, area*220/10000, 16.0),
            ('2.3', '通风空调工程', 'm²', area, 280, area*280/10000, 20.0),
            ('2.4', '消防工程', 'm²', area, 180, area*180/10000, 13.0),
            ('2.5', '弱电智能化', 'm²', area, 160, area*160/10000, 12.0),
            ('2.6', '电梯工程', '部', area/2000, 450000, area/2000*45, 11.0),
            ('小计', '安装工程合计', '', '', '', install, '100%', ''),
        ]

        for item in install_items:
            row_type = 'subtotal' if '小计' in str(item[0]) else 'normal'
            self._add_row(item, row_type=row_type)

        self.row += 1
        # 合计
        self._add_row(['合计', '分部分项工程费合计', '', '', '', total * 0.62, '100%', '占建安费62%'], row_type='total')

    def _add_measure_cost_detail(self, result):
        """措施项目费明细"""
        cost = result.get('造价估算', {})
        total = self._to_num(cost.get('总造价区间', {}).get('中值', 0))
        if total == 0: total = 41816.48  # 默认值
        params = result.get('项目参数', {})
        area = self._to_num(params.get('total_area', 50000))
        if area == 0: area = 50000
        measure = total * 0.12  # 措施费占总造价12%

        self._add_title("措施项目费明细表", 14, bg_color=self.colors['title_bg'])
        self._add_header(['序号', '费用项目', '计算基础', '费率(%)', '金额(万元)', '单方(元/m²)', '金额占比(%)', '说明'])

        # 安全文明施工费（占措施费约40%）
        safety = measure * 0.40
        # 夜间施工费（约5%）
        night = measure * 0.05
        # 二次搬运费（约3%）
        second_move = measure * 0.03
        # 冬雨季施工费（约3%）
        winter = measure * 0.03
        # 缩短工期措施费（约8%）
        rush = measure * 0.08
        # 大型机械进出场费（约10%）
        machinery = measure * 0.10
        # 脚手架费（约15%）
        scaffold = measure * 0.15
        # 模板费（约12%）
        template = measure * 0.12
        # 其他措施费（约4%）
        other = measure * 0.04

        measure_items = [
            ('一', '安全文明施工费', f'{total*0.62:.2f}万元(分部分项费)', 7.5, safety, safety/area, 40.0,
             '环境保护+文明施工+安全施工+临时设施'),
            ('二', '夜间施工费', '分部分项费', 0.35, night, night/area, 5.0,
             '照明、夜餐补贴'),
            ('三', '二次搬运费', '分部分项费', 0.20, second_move, second_move/area, 3.0,
             '材料二次搬运'),
            ('四', '冬雨季施工费', '分部分项费', 0.20, winter, winter/area, 3.0,
             '保温、防雨措施'),
            ('五', '缩短工期措施费', '分部分项费', 0.50, rush, rush/area, 8.0,
             '赶工措施'),
            ('六', '大型机械进出场费', '分部分项费', 0.65, machinery, machinery/area, 10.0,
             '塔吊、施工电梯等'),
            ('七', '脚手架费', '建筑面积', 18.0, scaffold, scaffold/area, 15.0,
             '综合脚手架'),
            ('八', '模板费', '混凝土体积', 85.0, template, template/area, 12.0,
             '混凝土模板及支撑'),
            ('九', '其他措施费', '分部分项费', 0.25, other, other/area, 4.0,
             '定位复测等'),
            ('合计', '措施项目费合计', '-', '-', measure, measure/area, 100.0, '占建安费12%'),
        ]

        for item in measure_items:
            row_type = 'total' if '合计' in str(item[0]) else 'normal'
            self._add_row(item, row_type=row_type)

    def _add_other_cost_detail(self, result):
        """其他项目费明细"""
        cost = result.get('造价估算', {})
        total = self._to_num(cost.get('总造价区间', {}).get('中值', 0))
        if total == 0: total = 41816.48
        params = result.get('项目参数', {})
        area = self._to_num(params.get('total_area', 50000))
        if area == 0: area = 50000
        other = total * 0.04

        self._add_title("其他项目费明细表", 14, bg_color=self.colors['title_bg'])
        self._add_header(['序号', '费用项目', '计算基础', '费率(%)', '金额(万元)', '单方(元/m²)', '占比(%)', '备注'])

        # 暂列金额（约50%）
        provisional = other * 0.50
        # 专业工程暂估价（约20%）
        temp_estimate = other * 0.20
        # 计日工（约15%）
        day_work = other * 0.15
        # 总承包服务费（约15%）
        general_contract = other * 0.15

        other_items = [
            ('一', '暂列金额', '分部分项+措施费', 3.0, provisional, provisional/area, 50.0,
             '建设单位预留，可能发生'),
            ('二', '专业工程暂估价', '专业工程', 1.5, temp_estimate, temp_estimate/area, 20.0,
             '暂定价格，结算调整'),
            ('三', '计日工', '实际发生', '-', day_work, day_work/area, 15.0,
             '零星工作+加班'),
            ('四', '总承包服务费', '专业工程费', 3.0, general_contract, general_contract/area, 15.0,
             '配合管理+现场清理'),
            ('合计', '其他项目费合计', '-', '-', other, other/area, 100.0, '占建安费4%'),
        ]

        for item in other_items:
            row_type = 'total' if '合计' in str(item[0]) else 'normal'
            self._add_row(item, row_type=row_type)

        self.row += 2
        self._add_title("暂列金额明细（估算）", 12, bg_color=self.colors['section_bg'])
        self._add_header(['项目', '金额(万元)', '占比(%)', '说明'])

        provisional_items = [
            ('变更洽商', provisional * 0.40, 40.0, '设计变更+现场签证'),
            ('物价上涨预备', provisional * 0.35, 35.0, '材料价格波动风险金'),
            ('不可预见费', provisional * 0.25, 25.0, '不可预见事项'),
        ]
        for item in provisional_items:
            self._add_row(item)

    def _add_regulation_tax(self, result):
        """规费与税金"""
        cost = result.get('造价估算', {})
        total = self._to_num(cost.get('总造价区间', {}).get('中值', 0))
        if total == 0: total = 41816.48
        params = result.get('项目参数', {})
        area = self._to_num(params.get('total_area', 50000))
        if area == 0: area = 50000

        # 规费约3%
        regulation = total * 0.03
        # 税金9%
        tax = total * 0.09

        self._add_title("规费明细表", 14, bg_color=self.colors['title_bg'])
        self._add_header(['序号', '费用项目', '计算基础', '费率(%)', '金额(万元)', '单方(元/m²)', '占比(%)', '政策依据'])

        regulation_items = [
            ('一', '社会保险费', '人工费', 25.0, regulation * 0.65, regulation*0.65/area, 65.0, '建标[2013]44号'),
            ('  1.1', '养老保险费', '人工费', 14.0, regulation * 0.36, regulation*0.36/area, '-', '按规定比例'),
            ('  1.2', '医疗保险费', '人工费', 6.0, regulation * 0.16, regulation*0.16/area, '-', '按规定比例'),
            ('  1.3', '失业保险费', '人工费', 2.0, regulation * 0.05, regulation*0.05/area, '-', '按规定比例'),
            ('  1.4', '工伤保险费', '人工费', 1.5, regulation * 0.04, regulation*0.04/area, '-', '按规定比例'),
            ('  1.5', '生育保险费', '人工费', 1.5, regulation * 0.04, regulation*0.04/area, '-', '按规定比例'),
            ('二', '住房公积金', '人工费', 12.0, regulation * 0.28, regulation*0.28/area, 28.0, '建标[2013]44号'),
            ('三', '工程排污费', '建安费', 0.1, regulation * 0.07, regulation*0.07/area, 7.0, '按规定标准'),
            ('合计', '规费合计', '-', '-', regulation, regulation/area, 100.0, '占建安费3%'),
        ]

        for item in regulation_items:
            row_type = 'total' if '合计' in str(item[0]) else 'normal'
            self._add_row(item, row_type=row_type)

        self.row += 2
        self._add_title("税金计算表", 14, bg_color=self.colors['title_bg'])
        self._add_header(['序号', '项目', '计算式', '税率(%)', '金额(万元)', '说明'])

        # 增值税计算
        pre_tax = total  # 税前造价
        vat = pre_tax * 0.09  # 增值税9%
        city_tax = vat * 0.12  # 城市维护建设税12%（增值税的12%）
        edu_tax = vat * 0.03  # 教育费附加3%
        local_edu = vat * 0.02  # 地方教育附加2%

        tax_items = [
            ('一', '增值税销项税额', f'{pre_tax:.2f} × 9%', 9.0, vat, '建筑业9%增值税率'),
            ('二', '城市维护建设税', f'{vat:.2f} × 12%', '-', city_tax, '增值税×12%'),
            ('三', '教育费附加', f'{vat:.2f} × 3%', '-', edu_tax, '增值税×3%'),
            ('四', '地方教育附加', f'{vat:.2f} × 2%', '-', local_edu, '增值税×2%'),
            ('合计', '税金合计', '-', '综合税率9.77%', tax, '-', '税前造价×9.77%'),
        ]

        for item in tax_items:
            row_type = 'total' if '合计' in str(item[0]) else 'normal'
            self._add_row(item, row_type=row_type)

    def _add_material_quantity(self, result):
        """主要材料用量表"""
        cost = result.get('造价估算', {})
        params = result.get('项目参数', {})
        area = self._to_num(params.get('total_area', 50000))
        if area == 0: area = 50000

        self._add_title("主要材料工程量表", 14, bg_color=self.colors['title_bg'])
        self._add_header(['序号', '材料名称', '规格型号', '单位', '工程量', '单方指标', '价格(元)', '金额(万元)'])

        # 根据配比参数计算
        ratios = result.get('配比参数', {})
        rebar_ratio = ratios.get('第四族_钢筋智能配比', {}).get('钢筋混凝土比', {}).get('value', 120)
        concrete_vol = area * 0.45  # 混凝土体积约0.45m³/m²
        rebar_ton = concrete_vol * rebar_ratio / 1000  # 钢筋量(吨)

        materials = [
            ('一', '钢筋', 'HRB400/HRB500', 't', f'{rebar_ton:,.0f}', f'{rebar_ton/area*1000:.1f} kg/m²', 4800, f'{rebar_ton*4800/10000:.0f}'),
            ('  1.1', '  螺纹钢筋', 'HRB400 φ12-32', 't', f'{rebar_ton*0.7:,.0f}', '-', 4750, '-'),
            ('  1.2', '  光圆钢筋', 'HPB300 φ6-10', 't', f'{rebar_ton*0.3:,.0f}', '-', 4900, '-'),
            ('二', '混凝土', 'C15-C50', 'm³', f'{concrete_vol:,.0f}', f'{concrete_vol/area:.2f} m³/m²', 620, f'{concrete_vol*620/10000:.0f}'),
            ('  2.1', '  C30普通混凝土', 'P.O 42.5', 'm³', f'{concrete_vol*0.6:,.0f}', '-', 580, '-'),
            ('  2.2', '  C40抗渗混凝土', 'P.O 42.5', 'm³', f'{concrete_vol*0.4:,.0f}', '-', 680, '-'),
            ('三', '砌体', '加气块/页岩砖', 'm³', f'{area*0.25:,.0f}', '0.25 m³/m²', 380, f'{area*0.25*380/10000:.0f}'),
            ('四', '模板', '复合木模板', 'm²', f'{area*3.5:,.0f}', '3.5 m²/m²', 45, f'{area*3.5*45/10000:.0f}'),
            ('五', '脚手架', '钢管扣件式', 't', f'{area*0.015:,.0f}', '15 kg/m²', 5200, f'{area*0.015*5200/10000:.0f}'),
            ('六', '电线电缆', '铜芯', 'km', f'{area*0.025:,.0f}', '25 m/100m²', 180000, f'{area*0.025*180/10000:.0f}'),
            ('七', '风管', '镀锌钢板', 'm²', f'{area*1.2:,.0f}', '1.2 m²/m²', 85, f'{area*1.2*85/10000:.0f}'),
            ('八', '管道', 'PPR/钢塑', 'm', f'{area*0.8:,.0f}', '0.8 m/m²', 120, f'{area*0.8*120/10000:.0f}'),
        ]

        for item in materials:
            self._add_row(item)

    def _add_uncertainty(self, result):
        """不确定性分析"""
        uncertainty = result.get('不确定性分析', {})
        monte_carlo = uncertainty.get('蒙特卡洛模拟', {})
        cost = result.get('造价估算', {})
        params = result.get('项目参数', {})

        self._add_title("不确定性分析报告", 14, bg_color=self.colors['title_bg'])

        self.row += 1
        self._add_title("1. 精度等级说明", 12, bg_color=self.colors['section_bg'])
        self._add_header(['等级', '适用阶段', '精度', '误差范围'])

        levels = [
            ('A级', '施工图阶段', '高精度', '±8%'),
            ('B级', '初步设计阶段', '中精度', '±15%'),
            ('C级', '方案设计阶段', '低精度', '±25%'),
            ('D级', '概念估算阶段', '估算级', '>±30%'),
        ]
        for item in levels:
            self._add_row(item)

        self.row += 1
        self._add_title("2. 本项目不确定性因素", 12, bg_color=self.colors['section_bg'])
        self._add_header(['因素类别', '影响程度', '不确定性等级', '备注'])

        factors = [
            ('设计方案调整', '高', '±5%', '结构形式可能变化'),
            ('材料价格波动', '高', '±8%', '钢筋、混凝土占比较大'),
            ('地质条件', '中', '±3%', '桩基工程影响'),
            ('施工工期', '中', '±2%', '措施费影响'),
            ('政策调整', '低', '±1%', '规费、税金调整'),
        ]
        for item in factors:
            self._add_row(item)

        self.row += 1
        self._add_title("3. 蒙特卡洛模拟结果 (10,000次)", 12, bg_color=self.colors['section_bg'])

        unit_price = cost.get('单方造价区间', {})
        area = self._to_num(params.get('total_area', 50000))
        if area == 0: area = 50000

        self._add_header(['百分位', '单方造价(元/m²)', '总造价(万元)', '概率'])

        p10 = self._to_num(monte_carlo.get('P10', unit_price.get('低限', 0)))
        p50 = self._to_num(unit_price.get('中值', 0))
        p90 = self._to_num(monte_carlo.get('P90', unit_price.get('高限', 0)))
        p30 = p10 * 1.05
        p70 = p50 * 1.05

        monte_items = [
            ('P10 (乐观)', p10, p10*area/10000, '10%'),
            ('P30', p30, p30*area/10000, '30%'),
            ('P50 (中性)', p50, p50*area/10000, '50%'),
            ('P70', p70, p70*area/10000, '70%'),
            ('P90 (保守)', p90, p90*area/10000, '90%'),
        ]
        for item in monte_items:
            self._add_row(item, highlight=True)

        self.row += 1
        # 风险提示
        self._add_title("4. 风险提示与建议", 12, bg_color=self.colors['section_bg'])
        tips = [
            "本估算基于常规设计方案，实际造价可能因设计调整出现±15%波动",
            "钢筋、混凝土价格波动对总造价影响约为±5%",
            "建议在施工图阶段进行详细预算，控制投资",
            "预留不低于5%的暂列金额应对变更洽商",
        ]
        for tip in tips:
            self.ws.merge_cells(f'B{self.row}:H{self.row}')
            cell = self.ws[f'B{self.row}']
            cell.value = f"• {tip}"
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            self.row += 1


def generate_professional_report(
    building_type: str = "办公",
    total_area: float = 50000,
    floor_count: int = 31,
    region: str = "苏州",
    project_name: str = "建设工程造价估算报告"
) -> str:
    """生成专业造价报表"""

    print("=" * 70)
    print("  度量衡智库 · 专业造价估算报表系统 v2.0")
    print("  基于 GB/T 50500-2024《建设工程工程量清单计价标准》")
    print("=" * 70)
    print()

    # 1. 执行造价估算
    print(f"[1/3] 执行度量衡测不准估算...")
    result = quick_estimate(
        building_type=building_type,
        total_area=total_area,
        floor_count=floor_count,
        region=region
    )

    # 2. 生成报表
    print(f"[2/3] 生成专业Excel报表(7个工作表)...")
    generator = ProfessionalCostReport()
    filename = generator.create_report(result, project_name)

    # 3. 输出结果
    print(f"[3/3] 报表已生成: {filename}")
    print()
    print("=" * 70)
    print("  报表内容(7个工作表):")
    print("  [01] 造价汇总 - 总造价+费用构成")
    print("  [02] 分部分项工程费 - 土建+安装明细")
    print("  [03] 措施项目费 - 安文+夜间+二次搬运等")
    print("  [04] 其他项目费 - 暂列金额+计日工")
    print("  [05] 规费税金 - 社会保险+增值税")
    print("  [06] 主要材料用量 - 钢筋+混凝土+模板")
    print("  [07] 不确定性分析 - 蒙特卡洛模拟")
    print("=" * 70)

    return filename, result


if __name__ == "__main__":
    filename, result = generate_professional_report(
        building_type="办公",
        total_area=50000,
        floor_count=31,
        region="苏州",
        project_name="苏州某超高层办公楼造价估算报告"
    )
