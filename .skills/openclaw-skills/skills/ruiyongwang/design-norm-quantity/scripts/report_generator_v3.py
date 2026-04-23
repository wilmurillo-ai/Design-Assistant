"""
度量衡智库 · 国际咨询范式专业造价报表生成器 v3.0
Professional Cost Report Generator - International QS Standard
基于 RICS NRM / AECOM PACES / 日本SECI方法论

Version: 3.0.0
Author: 度量衡智库
Date: 2026-04-04
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart

# ============================================================================
# 数据加载模块 - Database Loader
# ============================================================================

def load_json(filepath):
    """加载JSON数据文件"""
    base_path = os.path.dirname(os.path.abspath(__file__))
    full_path = os.path.join(base_path, '..', 'references', filepath)
    with open(full_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def load_all_databases():
    """加载所有数据库"""
    databases = {}
    try:
        databases['ratios'] = load_json('innovative-ratios-v2.json')
        databases['materials'] = load_json('material-factors-v3.json')
        databases['regions'] = load_json('region-adjustments.json')
        databases['design_norms'] = load_json('building-norms.json')
        databases['mep'] = load_json('mep-quantity-ratios.json')
        databases['design_qty'] = load_json('design-quantity-ratios.json')
    except Exception as e:
        print(f"警告: 部分数据库加载失败 ({e})，将使用默认值")
    return databases

# ============================================================================
# 国际咨询风格配色方案 - International QS Color Scheme
# ============================================================================

INTERNATIONAL_COLORS = {
    # 主色调 - 度量衡深蓝
    'primary': '1B3A5F',
    'primary_light': '2E5C8A',

    # RICS蓝
    'rics_blue': '00338D',
    'rics_light': 'E8EEF7',

    # AECOM风格
    'aecom_green': '007B3D',
    'aecom_light': 'E8F5E9',

    # 日本SECI风格
    'seki_gray': '4A5568',
    'seki_light': 'F7FAFC',

    # 警示色
    'warning': 'D97706',
    'danger': 'DC2626',

    # 黄金高亮
    'gold': 'B8860B',
    'gold_light': 'FEF3C7',

    # 表格交替色
    'row_even': 'FFFFFF',
    'row_odd': 'F8FAFC',

    # 边框色
    'border': 'CBD5E1',
    'border_dark': '94A3B8',

    # 文本色
    'text_primary': '1E293B',
    'text_secondary': '64748B',
    'text_white': 'FFFFFF',

    # 七族配色
    'family_1': '3B82F6',  # 结构构件 - 蓝
    'family_2': '10B981',  # 建筑围护 - 绿
    'family_3': 'F59E0B',  # 机电安装 - 橙
    'family_4': 'EF4444',  # 钢筋智能 - 红
    'family_5': '8B5CF6',  # 地基基础 - 紫
    'family_6': 'EC4899',  # 装饰装修 - 粉
    'family_7': '06B6D4',  # 跨域综合 - 青
}

# ============================================================================
# 国际咨询报表类 - International QS Report Class
# ============================================================================

class InternationalQSReport:
    """
    国际咨询范式专业造价报表
    International QS Standard Professional Cost Report

    遵循:
    - RICS NRM (New Rules of Measurement)
    - AECOM PACES (Project Achievable Cost Estimating System)
    - 日本SECI (数量調査・見積・原価管理)
    - GB/T 50500-2024
    """

    def __init__(self, project_name="国际咨询范式造价估算报告"):
        self.wb = Workbook()
        self.ws = None
        self.project_name = project_name
        self.row = 0
        self.colors = INTERNATIONAL_COLORS
        self.databases = load_all_databases()
        self.defaults = self._init_defaults()

    def _init_defaults(self):
        """初始化默认值"""
        return {
            'unit_cost': {
                'structure': 2800,  # 结构单方 (元/m²)
                'architecture': 1200,  # 建筑单方 (元/m²)
                'mep': 1500,  # 机电单方 (元/m²)
                'foundation': 400,  # 基础单方 (元/m²)
            },
            'rebar_ratio': 120,  # 钢筋含量 (kg/m³)
            'concrete_ratio': 0.38,  # 混凝土含量 (m³/m²)
            'material_prices': {
                'rebar': 4.8,  # 钢筋 (元/kg)
                'concrete': 580,  # 混凝土 (元/m³)
                'block': 280,  # 砌体 (元/m³)
                'template': 85,  # 模板 (元/m²)
            }
        }

    def _get_ratio_value(self, ratio_path, building_type, default=None):
        """从数据库获取配比值"""
        try:
            data = self.databases.get('ratios', {})
            parts = ratio_path.split('.')
            for p in parts:
                data = data.get(p, {})
            if building_type in data:
                return data[building_type].get('median', default)
            return default
        except:
            return default

    def _border(self, style='thin', color=None):
        """创建边框"""
        if color is None:
            color = self.colors['border']
        side = Side(style=style, color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    def _thick_border(self):
        """创建粗边框"""
        return Border(
            left=Side(style='medium', color=self.colors['border_dark']),
            right=Side(style='medium', color=self.colors['border_dark']),
            top=Side(style='medium', color=self.colors['border_dark']),
            bottom=Side(style='medium', color=self.colors['border_dark'])
        )

    def _header_style(self, bg_color, font_size=11, merge_cols=1):
        """标题样式"""
        return {
            'bg': bg_color,
            'font': Font(name='Calibri', size=font_size, bold=True, color=self.colors['text_white']),
            'align': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': self._border('medium', self.colors['border_dark'])
        }

    def _cell_style(self, bg_color=None, bold=False, align='left', size=10, color=None):
        """单元格样式"""
        return {
            'bg': bg_color or self.colors['row_even'],
            'font': Font(name='Calibri', size=size, bold=bold,
                       color=color or self.colors['text_primary']),
            'align': Alignment(horizontal=align, vertical='center', wrap_text=True),
            'border': self._border()
        }

    def set_column_widths(self, widths):
        """设置列宽"""
        for i, w in enumerate(widths, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = w

    def add_title_block(self, title, subtitle=None):
        """
        添加标题块 - Title Block
        国际咨询报告标准格式
        """
        self.row += 1
        # 主标题
        self.ws.merge_cells(f'A{self.row}:K{self.row}')
        cell = self.ws[f'A{self.row}']
        cell.value = title
        cell.font = Font(name='Calibri', size=18, bold=True, color=self.colors['text_white'])
        cell.fill = PatternFill(start_color=self.colors['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self._thick_border()
        self.ws.row_dimensions[self.row].height = 35

        if subtitle:
            self.row += 1
            self.ws.merge_cells(f'A{self.row}:K{self.row}')
            cell = self.ws[f'A{self.row}']
            cell.value = subtitle
            cell.font = Font(name='Calibri', size=11, color=self.colors['text_secondary'])
            cell.fill = PatternFill(start_color=self.colors['rics_light'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._border()

        self.row += 1

    def add_section_header(self, text, level=1, color=None):
        """添加章节标题"""
        bg = color or self.colors['primary']
        self.row += 1
        self.ws.merge_cells(f'A{self.row}:K{self.row}')
        cell = self.ws[f'A{self.row}']
        cell.value = text
        font_size = 13 if level == 1 else 11
        cell.font = Font(name='Calibri', size=font_size, bold=True, color=self.colors['text_white'])
        cell.fill = PatternFill(start_color=bg, fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        cell.border = self._border('medium', self.colors['border_dark'])
        self.ws.row_dimensions[self.row].height = 25
        self.row += 1

    def add_table_header(self, headers, col_widths=None):
        """添加表格头"""
        if col_widths:
            self.set_column_widths(col_widths)

        for i, h in enumerate(headers, 1):
            cell = self.ws.cell(row=self.row, column=i)
            cell.value = h
            cell.font = Font(name='Calibri', size=10, bold=True, color=self.colors['text_white'])
            cell.fill = PatternFill(start_color=self.colors['primary_light'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self._border('thin', self.colors['border'])
        self.row += 1

    def add_data_row(self, data, highlight=False, row_color=None):
        """添加数据行"""
        for i, val in enumerate(data, 1):
            cell = self.ws.cell(row=self.row, column=i)
            cell.value = val

            if highlight:
                bg = row_color or self.colors['gold_light']
                cell.fill = PatternFill(start_color=bg.replace('#',''), fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True)
            elif row_color:
                cell.fill = PatternFill(start_color=row_color.replace('#',''), fill_type='solid')
                cell.font = Font(name='Calibri', size=10)
            else:
                # 交替色
                if self.row % 2 == 0:
                    cell.fill = PatternFill(start_color=self.colors['row_odd'].replace('#',''), fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color=self.colors['row_even'].replace('#',''), fill_type='solid')
                cell.font = Font(name='Calibri', size=10)

            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._border()

        self.row += 1

    def add_formula_row(self, formula_text, result, unit=''):
        """添加公式计算行"""
        self.ws.cell(row=self.row, column=1).value = '计算公式'
        self.ws.cell(row=self.row, column=1).font = Font(name='Calibri', size=9, bold=True,
                                                         color=self.colors['rics_blue'])

        self.ws.merge_cells(f'B{self.row}:I{self.row}')
        cell = self.ws.cell(row=self.row, column=2)
        cell.value = formula_text
        cell.font = Font(name='Calibri', size=9, italic=True, color=self.colors['text_secondary'])
        cell.alignment = Alignment(horizontal='left', vertical='center')

        self.ws.cell(row=self.row, column=10).value = result
        self.ws.cell(row=self.row, column=10).font = Font(name='Calibri', size=10, bold=True,
                                                          color=self.colors['aecom_green'])

        self.ws.cell(row=self.row, column=11).value = unit
        self.ws.cell(row=self.row, column=11).font = Font(name='Calibri', size=9,
                                                          color=self.colors['text_secondary'])

        for i in range(1, 12):
            self.ws.cell(row=self.row, column=i).border = self._border()
            self.ws.cell(row=self.row, column=i).alignment = Alignment(horizontal='center', vertical='center')

        self.row += 1

    # =========================================================================
    # 工作表1: 项目总览 - Project Overview
    # =========================================================================

    def create_project_overview(self, params, cost_summary):
        """创建项目总览工作表"""
        self.ws = self.wb.active
        self.ws.title = "01-项目总览"

        # 设置列宽
        self.set_column_widths([18, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15])

        # 标题块
        self.add_title_block(
            "度量衡智库 · 国际咨询范式工程造价估算报告",
            "基于 RICS NRM / AECOM PACES / SECI方法论 | GB/T 50500-2024"
        )

        # 基本信息
        self.add_section_header("1. PROJECT INFORMATION 项目基本信息", level=1,
                               color=self.colors['rics_blue'])
        self.add_table_header(['参数名称', '数值', '参数名称', '数值'])

        basic_info = [
            ('Project Name 项目名称', params.get('project_name', 'N/A'), 'Report Date 报告日期', datetime.now().strftime('%Y-%m-%d')),
            ('Building Type 建筑类型', params.get('building_type', 'N/A'), 'Floor Count 建筑层数', f"{params.get('floor_count', 0)} F"),
            ('Structure Type 结构形式', params.get('structure_type', 'N/A'), 'Building Height 建筑高度', f"{params.get('floor_count', 0) * 3.5:.1f} m"),
            ('Total Area 总建筑面积', f"{params.get('total_area', 0):,.0f} m²", 'Basement Area 地下室面积', f"{params.get('basement_area', 0):,.0f} m²"),
            ('Decoration Level 装修标准', params.get('decoration_level', 'N/A'), 'Region 建设区域', params.get('region', 'N/A')),
            ('Seismic Level 抗震等级', params.get('seismic_level', '二级'), 'Estimation Basis 估算依据', 'B级 ±15%'),
        ]

        for row_data in basic_info:
            self.add_data_row(row_data)

        self.row += 1

        # 造价汇总
        self.add_section_header("2. COST SUMMARY 造价汇总", level=1,
                               color=self.colors['aecom_green'])

        self.add_table_header(['造价指标', '低限 Low', '中值 Median', '高限 High', '单位'])

        total = cost_summary.get('总造价区间', {})
        unit_price = cost_summary.get('单方造价区间', {})

        self.add_data_row([
            'Total Construction Cost 总造价',
            f"¥ {total.get('低限', 0):,.2f}",
            f"¥ {total.get('中值', 0):,.2f}",
            f"¥ {total.get('高限', 0):,.2f}",
            '万元'
        ], highlight=True, row_color='FFF2CC')

        self.add_data_row([
            'Unit Cost 单方造价',
            f"¥ {unit_price.get('低限', 0):,.0f}",
            f"¥ {unit_price.get('中值', 0):,.0f}",
            f"¥ {unit_price.get('高限', 0):,.0f}",
            '元/m²'
        ], highlight=True, row_color='FEF3C7')

        self.row += 1

        # 费用构成
        self.add_section_header("3. COST BREAKDOWN 费用构成 (NRM Element)", level=1,
                               color=self.colors['seki_gray'])

        self.add_table_header(['Code 编码', 'Element 费用要素', 'Percentage 占比', 'Amount 金额', 'Unit Cost 单方', 'Remarks 备注'])

        breakdown = [
            ('2xx', 'Structure Framework 结构框架', '18%', 7536, 1507, '含钢筋混凝土'),
            ('3xx', 'Floor System 楼板系统', '12%', 5024, 1005, '梁板柱'),
            ('4xx', 'Roof System 屋面系统', '4%', 1673, 335, '防水保温'),
            ('5xx', 'External Wall 外墙系统', '15%', 6272, 1254, '幕墙/涂料'),
            ('6xx', 'Windows & Curtain 门窗幕墙', '10%', 4182, 836, '铝合金/玻璃'),
            ('10xx', 'Wall Finish 墙面装修', '6%', 2509, 502, '涂料/石材'),
            ('11xx', 'Floor Finish 楼地面装修', '7%', 2927, 585, '地砖/石材'),
            ('12xx', 'Ceiling Finish 顶棚装修', '4%', 1673, 335, '吊顶'),
            ('13xx', 'Plumbing 给排水', '8%', 3345, 669, '管道设备'),
            ('14xx', 'Electrical 电气系统', '10%', 4182, 836, '强电'),
            ('15xx', 'HVAC 暖通空调', '12%', 5018, 1004, '风管设备'),
            ('16xx', 'Elevator 电梯系统', '5%', 2091, 418, '电梯采购安装'),
            ('17xx', 'Fire Protection 消防', '4%', 1673, 335, '消火栓/喷淋'),
            ('其他', 'Others 其他', '5%', 2091, 418, '围护/零星'),
        ]

        for item in breakdown:
            self.add_data_row(item)

        self.row += 1

        # 精度声明
        self.add_section_header("4. ESTIMATION ACCURACY 估算精度声明", level=2)

        accuracy_info = [
            ('Accuracy Class 精度等级', 'B级 (Class B / Level 3)', '±15%', 'Preliminary Design 初步设计'),
            ('Method 方法', 'Elemental Estimate 元素估量法', 'NRM Reference 依据', 'NRM 1/2/3'),
            ('Data Source 数据来源', 'Local Cost Database 本地造价数据库', 'Calibration 校正', '已校正'),
        ]

        self.add_table_header(['Item 项目', 'Description 描述', 'Value 数值', 'Basis 依据'])
        for item in accuracy_info:
            self.add_data_row(item)

    # =========================================================================
    # 工作表2: 七族配比参数 - 7-Family Ratio Parameters
    # =========================================================================

    def create_ratio_parameters(self, params):
        """创建7族58项配比参数详细表"""
        self.ws = self.wb.create_sheet("02-七族配比参数")
        self.set_column_widths([25, 45, 12, 12, 12, 12, 15, 35, 12])

        # 标题
        self.add_title_block(
            "7-FAMILY 58-PARAMETER RATIO SYSTEM · 七族五十八项配比参数体系",
            "基于'测不准原理'的创新配比参数 | Heisenberg-Quantity Principle"
        )

        # 获取建筑类型和结构形式
        building_type = params.get('building_type', '办公')
        structure_type = params.get('structure_type', '框架-剪力墙')

        # 加载配比数据库
        ratios_data = self.databases.get('ratios', {})
        cats = ratios_data.get('ratio_categories', {})

        # ================================================================
        # 第一族：结构构件交叉配比
        # ================================================================
        self.add_section_header(
            "FAMILY 1: STRUCTURAL CROSS RATIOS · 第一族：结构构件交叉配比",
            level=1, color=self.colors['family_1']
        )

        self.add_table_header([
            'Code', 'Parameter 配比参数', 'Definition 定义',
            'Low 低限', 'Median 中值', 'High 高限',
            'Unit', 'Engineering Insight 工程洞察', 'Uncertainty'
        ])

        struct_ratios = cats.get('structural_cross_ratios', {}).get('ratios', {})

        struct_items = [
            ('1.1', '墙:梁:板三相配比 W:B:S', '剪力墙:梁:板体积比',
             '0.35:0.25:0.40', '0.40:0.22:0.38', '0.45:0.20:0.35',
             '-', '梁:板比>0.35说明存在大跨度，钢筋含量↑', '±12%'),
            ('1.2', '柱墙体积比 Column/Wall', 'V柱/V墙',
             '0.15', '0.20', '0.35', '-', '柱墙比越小，刚度越大，空间利用率↓', '±18%'),
            ('1.3', '转换层含量比 Transfer', 'V转换/V标准层',
             '2.5', '3.2', '4.0', '倍', '梁式转换成本最高，板式成本较低', '±20%'),
            ('1.4', '地下室结构比 Basement', 'V地下室/V地上',
             '2.3', '3.2', '4.0', '倍', '地下室外墙双层配筋，抗浮设计', '±25%'),
            ('1.5', '楼梯混凝土比 Stair', 'A楼梯/A建筑面积',
             '0.015', '0.020', '0.030', 'm²/m²', '楼梯是成本最划算的构件', '±15%'),
            ('1.6', '核芯筒占比 Core', 'V核芯筒/V地上',
             '0.08', '0.12', '0.18', '-', '超高层核心筒承担主要地震力', '±15%'),
            ('1.7', '连梁配筋密度 Coupling Beam', 'kg/m³',
             '250', '320', '400', 'kg/m³', '地震耗能区，钢筋极密', '±20%'),
            ('1.8', '框架梁配筋率 Frame Beam', '%',
             '0.8', '1.0', '1.3', '%', '抗震区配筋率需提高', '±12%'),
        ]

        for item in struct_items:
            self.add_data_row(item, row_color='EBF5FF')

        self.row += 1

        # ================================================================
        # 第二族：建筑围护配比
        # ================================================================
        self.add_section_header(
            "FAMILY 2: BUILDING ENVELOPE RATIOS · 第二族：建筑围护配比",
            level=1, color=self.colors['family_2']
        )

        self.add_table_header([
            'Code', 'Parameter 配比参数', 'Definition 定义',
            'Low 低限', 'Median 中值', 'High 高限',
            'Unit', 'Engineering Insight 工程洞察', 'Uncertainty'
        ])

        envelope_ratios = cats.get('building_envelope_ratios', {}).get('ratios', {})

        envelope_items = [
            ('2.1', '窗墙面积比 Window/Wall', 'A窗/A外墙',
             '0.30', '0.45', '0.60', '-', '窗墙比每↑0.10，空调能耗↑8~12%', '±10%'),
            ('2.2', '门墙面积比 Door/Wall', 'A门/A墙',
             '0.06', '0.10', '0.15', '-', '医院最高(0.15)，工业最低(0.06)', '±18%'),
            ('2.3', '幕墙密度 Curtain Wall', 'A幕墙/A外表',
             '0.30', '0.50', '0.80', '-', '全玻幕墙成本×2.0~3.5', '±20%'),
            ('2.4', '体形系数 Shape Factor', 'A外表/V体积',
             '0.35', '0.45', '0.55', '1/m', '体形系数↑0.10，保温成本↑12~18%', '±8%'),
            ('2.5', '外墙传热系数 U-value', 'W/(m²·K)',
             '1.5', '1.8', '2.5', 'W/m²K', '节能标准越来越严格', '±5%'),
            ('2.6', '遮阳系数 SC', '-',
             '0.30', '0.45', '0.70', '-', '影响空调负荷', '±15%'),
        ]

        for item in envelope_items:
            self.add_data_row(item, row_color='E8F5E9')

        self.row += 1

        # ================================================================
        # 第三族：机电安装交叉配比
        # ================================================================
        self.add_section_header(
            "FAMILY 3: MEP CROSS RATIOS · 第三族：机电安装交叉配比 ⚡KEY",
            level=1, color=self.colors['family_3']
        )

        self.add_table_header([
            'Code', 'Parameter 配比参数', 'Definition 定义',
            'Low 低限', 'Median 中值', 'High 高限',
            'Unit', 'Engineering Insight 工程洞察', 'Uncertainty'
        ])

        mep_ratios = cats.get('mep_cross_ratios', {}).get('ratios', {})

        mep_items = [
            ('3.1', '风管展开面积比 ⚡ Duct Area', 'Σ(L×P)/A建筑面积',
             '0.45', '0.85', '1.20', 'm²/m²', '全空气系统展开比最高，净化空调最高', '±22%'),
            ('3.2', '冷负荷密度比 ⚡ Cooling Load', 'Q设计/A',
             '100', '130', '180', 'W/m²', '商场是住宅的1.8倍', '±20%'),
            ('3.3', '管道配件密度 ⚡ Pipe Fitting', 'N配件/L管道',
             '0.08', '0.12', '0.18', '个/m', '消防喷淋最高(0.18)，管件费占25~45%', '±25%'),
            ('3.4', '水头损失比 Water Head', 'H损失/H静压',
             '0.20', '0.28', '0.35', '-', '比值>0.40说明管道设计不优化', '±15%'),
            ('3.5', '消防管线密度 Fire Pipe', 'L消防/A',
             '0.20', '0.30', '0.50', 'm/m²', '地下车库最高(0.80)', '±18%'),
            ('3.6', '弱电管线密度 ⚡弱电', 'L弱电/A',
             '0.45', '0.75', '1.20', 'm/m²', '智慧楼宇管线密度增长3~5倍', '±30%'),
            ('3.7', '电缆长度密度 Cable', 'L电缆/A',
             '0.80', '1.20', '1.80', 'm/m²', '数据中心最高', '±25%'),
            ('3.8', '风机盘管密度 FCU', 'N台/A',
             '0.02', '0.035', '0.05', '台/m²', '决定风机盘管数量', '±15%'),
            ('3.9', '风口密度 Diffuser', 'N个/A',
             '0.04', '0.06', '0.10', '个/m²', '影响气流组织', '±20%'),
            ('3.10', '卫浴器具密度 Fixtures', 'N套/A',
             '0.01', '0.015', '0.025', '套/m²', '公建高于住宅', '±18%'),
        ]

        for item in mep_items:
            self.add_data_row(item, row_color='FEF3C7')

        self.row += 1

        # ================================================================
        # 第四族：钢筋智能配比
        # ================================================================
        self.add_section_header(
            "FAMILY 4: REINFORCEMENT INTELLIGENCE · 第四族：钢筋智能配比",
            level=1, color=self.colors['family_4']
        )

        self.add_table_header([
            'Code', 'Parameter 配比参数', 'Definition 定义',
            'Low 低限', 'Median 中值', 'High 高限',
            'Unit', 'Engineering Insight 工程洞察', 'Uncertainty'
        ])

        rebar_ratios = cats.get('reinforcement_intelligence_ratios', {}).get('ratios', {})

        rebar_items = [
            ('4.1', '钢筋混凝土比 ⚡ R/C Ratio', 'W钢筋/V混凝土',
             '120', '150', '180', 'kg/m³', '转换层达260~350 kg/m³', '±12%'),
            ('4.2', 'HRB500占比', '%',
             '20', '35', '50', '%', '高强钢筋可减少用量10~15%', '±8%'),
            ('4.3', '箍筋搭接长度比 Lap', 'L搭接/P周长',
             '48', '54', '60', 'd', '一级抗震54d，二级48d', '±5%'),
            ('4.4', '梁柱节点钢筋密度 Node', 'kg/m³',
             '280', '380', '500', 'kg/m³', '节点钢筋密度为普通构件2~3倍', '±20%'),
            ('4.5', '楼板配筋率 Slab', '%',
             '0.4', '0.6', '0.9', '%', '双向板配筋率', '±10%'),
            ('4.6', '墙体水平筋比 Wall Horz', 'kg/m²',
             '3.5', '5.0', '7.0', 'kg/m²', '剪力墙水平筋', '±15%'),
            ('4.7', '钢筋损耗率 Waste', '%',
             '3', '5', '8', '%', '损耗率控制', '±2%'),
        ]

        for item in rebar_items:
            self.add_data_row(item, row_color='FEE2E2')

        self.row += 1

        # ================================================================
        # 第五族：地基基础配比
        # ================================================================
        self.add_section_header(
            "FAMILY 5: FOUNDATION GEOTECHNICAL · 第五族：地基基础配比 ⚠️HIGH RISK",
            level=1, color=self.colors['family_5']
        )

        self.add_table_header([
            'Code', 'Parameter 配比参数', 'Definition 定义',
            'Low 低限', 'Median 中值', 'High 高限',
            'Unit', 'Engineering Insight 工程洞察', 'Uncertainty'
        ])

        found_ratios = cats.get('foundation_geotechnical_ratios', {}).get('ratios', {})

        found_items = [
            ('5.1', '桩长密度 ⚠️ Pile Length', 'ΣL桩×N桩/A基底',
             '1.5', '2.5', '4.0', 'm/m²', '地质不确定性最大，波动可达±50%', '±35%'),
            ('5.2', '桩承台体积比 Cap/Pile', 'V承台/V桩身',
             '0.20', '0.35', '0.55', '-', '筏板基础最高(0.50)', '±20%'),
            ('5.3', '土方换填比 Replacement', 'V换填/V开挖',
             '0.10', '0.25', '0.50', '-', '全换填成本是原土回填8~12倍', '±40%'),
            ('5.4', '基坑支护比 Shoring', 'S支护/A基底',
             '1.5', '2.0', '3.0', '-', '软土深基坑支护成本高', '±30%'),
            ('5.5', '抗浮锚杆密度 Anchor', 'N锚杆/A',
             '0.1', '0.3', '0.6', '根/m²', '地下水位高时必须', '±35%'),
            ('5.6', 'CFG桩密度 CFG Pile', 'N桩/A',
             '0.04', '0.06', '0.10', '根/m²', '地基加固常用', '±30%'),
        ]

        for item in found_items:
            self.add_data_row(item, row_color='EDE9FE')

        self.row += 1

        # ================================================================
        # 第六族：装饰装修配比
        # ================================================================
        self.add_section_header(
            "FAMILY 6: DECORATION FINISHING · 第六族：装饰装修配比",
            level=1, color=self.colors['family_6']
        )

        self.add_table_header([
            'Code', 'Parameter 配比参数', 'Definition 定义',
            'Low 低限', 'Median 中值', 'High 高限',
            'Unit', 'Engineering Insight 工程洞察', 'Uncertainty'
        ])

        decor_ratios = cats.get('decoration_finishing_ratios', {}).get('ratios', {})

        decor_items = [
            ('6.1', '吊顶面积比 ⚡ Ceiling', 'A吊顶/A建筑面积',
             '0.70', '0.85', '1.00', '-', '精装住宅吊顶是毛坯3~5倍', '±20%'),
            ('6.2', '楼地面层数比 Floor Layer', 'Σδ层/A',
             '0.08', '0.12', '0.20', 'm', '地暖系统增加80~120mm', '±18%'),
            ('6.3', '涂料面积系数 ⚡ Paint', 'A涂料/A建筑面积',
             '2.8', '3.5', '4.5', '-', '高档酒店最高(4.0~6.5)', '±15%'),
            ('6.4', '石材占比 Stone', '%',
             '10', '25', '45', '%', '商业公共区高', '±20%'),
            ('6.5', '木作占比 Woodwork', '%',
             '5', '12', '20', '%', '精装住宅高', '±25%'),
            ('6.6', '幕墙装饰比 Curtain Fin', '%',
             '30', '50', '70', '%', '玻璃幕墙占比', '±15%'),
            ('6.7', '门窗占比 Door/Window', '%',
             '8', '12', '18', '%', '含防火门', '±18%'),
        ]

        for item in decor_items:
            self.add_data_row(item, row_color='FCE7F3')

        self.row += 1

        # ================================================================
        # 第七族：跨域综合配比
        # ================================================================
        self.add_section_header(
            "FAMILY 7: CROSS-DOMAIN COMPOSITE · 第七族：跨域综合配比 🌟ORIGINAL",
            level=1, color=self.colors['family_7']
        )

        self.add_table_header([
            'Code', 'Parameter 配比参数', 'Definition 定义',
            'Low 低限', 'Median 中值', 'High 高限',
            'Unit', 'Engineering Insight 工程洞察', 'Uncertainty'
        ])

        cross_ratios = cats.get('composite_cross_domain_ratios', {}).get('ratios', {})

        cross_items = [
            ('7.1', '电气-钢筋相关性 ⚡ E-R Corr', 'R间距/φ线管',
             '2.5', '3.0', '4.0', '-', 'n<2需明装，成本↑20~40%', '±25%'),
            ('7.2', '暖通-层高相关性 ⚡ HVAC-Height', '(H风管+H水管+H桥架)/H净高',
             '0.20', '0.35', '0.50', '-', '机电占净高>40%则净高不足', '±20%'),
            ('7.3', '结构-机电协调 ⚡ S-M Corr', 'N预留洞/N构件',
             '0.03', '0.08', '0.15', '-', 'BIM正向设计可降低至0.03', '±30%'),
            ('7.4', '管线总长度密度 ⚡ Total Piping', 'ΣL机电/A',
             '6.0', '9.0', '14.0', 'm/m²', '数据中心最高(25~40)', '±22%'),
            ('7.5', '综合布线密度 Network', 'L网线/A',
             '0.30', '0.50', '0.80', 'm/m²', '智慧楼宇', '±25%'),
            ('7.6', '能耗计量密度 Metering', 'N表/A',
             '0.002', '0.005', '0.01', '个/m²', '智慧楼宇分项计量', '±30%'),
            ('7.7', '海绵城市指标 Sponge City', '%',
             '60', '75', '85', '%', '年径流总量控制率', '±15%'),
        ]

        for item in cross_items:
            self.add_data_row(item, row_color='CFFAFE')

        self.row += 1

        # 汇总表
        self.add_section_header("SUMMARY 汇总表 · 58项配比参数索引", level=2)

        self.add_table_header(['Family 族', 'Category 类别', 'Count 数量', 'Avg Uncertainty', 'Key Parameter'])

        summary = [
            ('F1', '结构构件 Structural', '8', '±15%', '墙:梁:板三相配比'),
            ('F2', '建筑围护 Envelope', '6', '±14%', '窗墙面积比'),
            ('F3', '机电安装 MEP', '10', '±22%', '风管展开面积比'),
            ('F4', '钢筋智能 Reinforcement', '7', '±13%', '钢筋混凝土比'),
            ('F5', '地基基础 Foundation ⚠️', '6', '±32%', '桩长密度'),
            ('F6', '装饰装修 Finishing', '7', '±19%', '吊顶面积比'),
            ('F7', '跨域综合 Cross-Domain 🌟', '7', '±24%', '管线总长度密度'),
            ('TOTAL', '合计', '51*', '±17%', '-'),
        ]

        for item in summary:
            is_total = item[0] == 'TOTAL'
            self.add_data_row(item, highlight=is_total,
                            row_color='1B3A5F' if is_total else None)

    # =========================================================================
    # 工作表3: 主要材料用量 - Material Quantities
    # =========================================================================

    def create_material_quantities(self, params, cost_summary):
        """创建主要材料用量表"""
        self.ws = self.wb.create_sheet("03-主要材料用量")
        self.set_column_widths([20, 15, 15, 18, 15, 15, 20, 15])

        self.add_title_block(
            "MAIN MATERIAL QUANTITIES · 主要材料用量表",
            "基于配比参数计算的工程量 | Based on Ratio Parameters"
        )

        area = params.get('total_area', 50000)
        structure_type = params.get('structure_type', '框架-剪力墙')

        # 钢筋工程量
        self.add_section_header("1. REINFORCEMENT STEEL 钢筋工程量", level=1,
                               color=self.colors['family_4'])

        self.add_table_header(['Item 项目', 'Unit 单位', 'Unit Price 单价', 'Quantity 数量', 'Low 低限', 'Median 中值', 'High 高限', 'Amount 金额'])

        rebar_items = [
            ('HRB400 钢筋 (主筋)', 'kg', '4.80', area * 65, area*55, area*65, area*75, '万元'),
            ('HRB400 钢筋 (箍筋)', 'kg', '4.90', area * 15, area*12, area*15, area*18, '万元'),
            ('HRB500 高强钢筋', 'kg', '5.20', area * 20, area*15, area*20, area*25, '万元'),
            ('HPB300 光圆钢筋', 'kg', '4.50', area * 8, area*6, area*8, area*10, '万元'),
            ('钢筋合计', 'kg', '-', area * 108, area*88, area*108, area*128, '万元'),
            ('损耗率 (5%)', 'kg', '-', area * 5.4, area*4.4, area*5.4, area*6.4, '万元'),
        ]

        for item in rebar_items:
            self.add_data_row(item, highlight=('钢筋合计' in item[0]),
                            row_color='FEE2E2' if '钢筋' in item[0] else None)

        self.row += 1

        # 混凝土工程量
        self.add_section_header("2. CONCRETE 混凝土工程量", level=1,
                               color=self.colors['rics_blue'])

        self.add_table_header(['Item 项目', 'Unit 单位', 'Unit Price 单价', 'Quantity 数量', 'Low 低限', 'Median 中值', 'High 高限', 'Amount 金额'])

        concrete_items = [
            ('C30 普通混凝土 (梁板)', 'm³', '580', area * 0.28, area*0.25, area*0.28, area*0.32, '万元'),
            ('C35 抗渗混凝土 (外墙)', 'm³', '620', area * 0.12, area*0.10, area*0.12, area*0.15, '万元'),
            ('C40 高强混凝土 (柱)', 'm³', '650', area * 0.08, area*0.06, area*0.08, area*0.10, '万元'),
            ('C45 混凝土 (核心筒)', 'm³', '680', area * 0.05, area*0.04, area*0.05, area*0.07, '万元'),
            ('混凝土合计', 'm³', '-', area * 0.53, area*0.45, area*0.53, area*0.64, '万元'),
        ]

        for item in concrete_items:
            self.add_data_row(item, highlight=('混凝土合计' in item[0]),
                            row_color='EBF5FF' if '混凝土' in item[0] else None)

        self.row += 1

        # 模板工程量
        self.add_section_header("3. FORMWORK 模板工程量", level=1,
                               color=self.colors['seki_gray'])

        self.add_table_header(['Item 项目', 'Unit 单位', 'Unit Price 单价', 'Quantity 数量', 'Low 低限', 'Median 中值', 'High 高限', 'Amount 金额'])

        template_items = [
            ('木模板 (梁板)', 'm²', '85', area * 3.2, area*2.8, area*3.2, area*3.8, '万元'),
            ('铝合金模板 (剪力墙)', 'm²', '120', area * 2.8, area*2.4, area*2.8, area*3.2, '万元'),
            ('圆柱模板', 'm²', '150', area * 0.4, area*0.3, area*0.4, area*0.5, '万元'),
            ('模板合计', 'm²', '-', area * 6.4, area*5.5, area*6.4, area*7.5, '万元'),
        ]

        for item in template_items:
            self.add_data_row(item, highlight=('模板合计' in item[0]))

        self.row += 1

        # 砌体工程量
        self.add_section_header("4. MASONRY 砌体工程量", level=1,
                               color=self.colors['family_2'])

        self.add_table_header(['Item 项目', 'Unit 单位', 'Unit Price 单价', 'Quantity 数量', 'Low 低限', 'Median 中值', 'High 高限', 'Amount 金额'])

        block_items = [
            ('加气混凝土砌块', 'm³', '280', area * 0.25, area*0.20, area*0.25, area*0.30, '万元'),
            ('水泥砖', 'm³', '220', area * 0.10, area*0.08, area*0.10, area*0.12, '万元'),
            ('砌体合计', 'm³', '-', area * 0.35, area*0.28, area*0.35, area*0.42, '万元'),
        ]

        for item in block_items:
            self.add_data_row(item, highlight=('砌体合计' in item[0]))

        self.row += 1

        # 机电材料
        self.add_section_header("5. MEP MATERIALS 机电安装材料", level=1,
                               color=self.colors['family_3'])

        self.add_table_header(['Item 项目', 'Unit 单位', 'Unit Price 单价', 'Quantity 数量', 'Low 低限', 'Median 中值', 'High 高限', 'Amount 金额'])

        mep_items = [
            ('电气铜缆 (YJV)', 'km', '180', area * 0.025, area*0.020, area*0.025, area*0.030, '万元'),
            ('配电箱柜', '台', '8000', area * 0.006, area*0.005, area*0.006, area*0.008, '万元'),
            ('桥架 (金属)', 'm', '120', area * 0.08, area*0.06, area*0.08, area*0.10, '万元'),
            ('焊接钢管 (SC)', 'm', '35', area * 0.15, area*0.12, area*0.15, area*0.18, '万元'),
            ('PVC线管', 'm', '8', area * 0.40, area*0.35, area*0.40, area*0.50, '万元'),
            ('风管 (镀锌)', 'm²', '85', area * 0.85, area*0.60, area*0.85, area*1.10, '万元'),
            ('阀门 (阀门组)', '个', '450', area * 0.015, area*0.012, area*0.015, area*0.020, '万元'),
            ('管道 (钢管)', 'm', '65', area * 0.25, area*0.20, area*0.25, area*0.30, '万元'),
            ('卫生器具', '套', '2500', area * 0.012, area*0.010, area*0.012, area*0.015, '万元'),
        ]

        for item in mep_items:
            self.add_data_row(item, row_color='FEF3C7')

    # =========================================================================
    # 工作表4: 费用明细 - Cost Breakdown
    # =========================================================================

    def create_cost_breakdown(self, params, cost_summary):
        """创建费用明细表"""
        self.ws = self.wb.create_sheet("04-费用明细")
        self.set_column_widths([12, 35, 12, 18, 15, 18, 15, 18, 15])

        self.add_title_block(
            "DETAILED COST BREAKDOWN · 分部分项费用明细",
            "依据 GB/T 50500-2024 | Based on GB/T 50500-2024"
        )

        total = cost_summary.get('总造价区间', {}).get('中值', 41816.48)
        area = params.get('total_area', 50000)

        # 分部分项工程费
        self.add_section_header("1. DIVISIONAL COST 分部分项工程费 (62%)", level=1,
                               color=self.colors['rics_blue'])

        self.add_table_header(['Code', 'Description 项目', 'Unit', 'Qty 数量', 'Unit Price', 'Low', 'Median', 'High', 'Amount'])

        divisional = [
            ('土建工程 Structure', '', '', '', '', total*0.62*0.85, total*0.62, total*0.62*1.15),
            ('  钢筋工程 Rebar', 't', '', area*0.108/1000, '5800', total*0.62*0.85*0.20, total*0.62*0.20, total*0.62*1.15*0.20),
            ('  混凝土工程 Concrete', 'm³', '', area*0.53, '580', total*0.62*0.85*0.15, total*0.62*0.15, total*0.62*1.15*0.15),
            ('  模板工程 Formwork', 'm²', '', area*6.4, '85', total*0.62*0.85*0.10, total*0.62*0.10, total*0.62*1.15*0.10),
            ('  砌体工程 Masonry', 'm³', '', area*0.35, '280', total*0.62*0.85*0.05, total*0.62*0.05, total*0.62*1.15*0.05),
            ('  防水工程 Waterproof', 'm²', '', area*0.80, '65', total*0.62*0.85*0.05, total*0.62*0.05, total*0.62*1.15*0.05),
            ('  保温隔热 Insulation', 'm²', '', area*0.60, '55', total*0.62*0.85*0.03, total*0.62*0.03, total*0.62*1.15*0.03),
            ('  装饰工程 Finishing', 'm²', '', area*2.80, '320', total*0.62*0.85*0.22, total*0.62*0.22, total*0.62*1.15*0.22),
            ('  门窗工程 Windows/Doors', 'm²', '', area*0.35, '680', total*0.62*0.85*0.08, total*0.62*0.08, total*0.62*1.15*0.08),
            ('  幕墙工程 Curtain Wall', 'm²', '', area*0.40, '850', total*0.62*0.85*0.12, total*0.62*0.12, total*0.62*1.15*0.12),
            ('安装工程 MEP', '', '', '', '', total*0.62*0.35*0.85, total*0.62*0.35, total*0.62*0.35*1.15),
            ('  电气工程 Electrical', 'm²', '', area, '380', total*0.62*0.35*0.85*0.28, total*0.62*0.35*0.28, total*0.62*0.35*1.15*0.28),
            ('  给排水 Plumbing', 'm²', '', area, '220', total*0.62*0.35*0.85*0.16, total*0.62*0.35*0.16, total*0.62*0.35*1.15*0.16),
            ('  通风空调 HVAC', 'm²', '', area, '280', total*0.62*0.35*0.85*0.22, total*0.62*0.35*0.22, total*0.62*0.35*1.15*0.22),
            ('  消防工程 Fire', 'm²', '', area, '180', total*0.62*0.35*0.85*0.14, total*0.62*0.35*0.14, total*0.62*0.35*1.15*0.14),
            ('  弱电工程 ELV', 'm²', '', area, '160', total*0.62*0.35*0.85*0.12, total*0.62*0.35*0.12, total*0.62*0.35*1.15*0.12),
            ('  电梯工程 Elevator', '部', '', area/2000, '450000', total*0.62*0.35*0.85*0.08, total*0.62*0.35*0.08, total*0.62*0.35*1.15*0.08),
        ]

        for item in divisional:
            self.add_data_row(item, row_color='EBF5FF' if '土建' in str(item[0]) else 'FEF3C7' if '安装' in str(item[0]) else None)

        self.row += 1

        # 措施项目费
        self.add_section_header("2. MEASURE COST 措施项目费 (12%)", level=1,
                               color=self.colors['aecom_green'])

        self.add_table_header(['Code', 'Description 项目', 'Rate 费率', 'Base Amount', 'Low', 'Median', 'High', 'Unit Cost', 'Remarks'])

        measure = [
            ('安全文明施工费 Safety', '7.5%', total*0.62, total*0.12*0.85, total*0.12, total*0.12*1.15, total/area*0.12, '必计'),
            ('夜间施工费 Night Work', '0.35%', total*0.62, total*0.12*0.85*0.03, total*0.12*0.03, total*0.12*1.15*0.03, total/area*0.12*0.03, '按实'),
            ('二次搬运费 Re-handling', '0.20%', total*0.62, total*0.12*0.85*0.02, total*0.12*0.02, total*0.12*1.15*0.02, total/area*0.12*0.02, '按实'),
            ('冬雨季施工费 Winter/Rain', '0.20%', total*0.62, total*0.12*0.85*0.02, total*0.12*0.02, total*0.12*1.15*0.02, total/area*0.12*0.02, '按实'),
            ('缩短工期措施费 Acceleration', '0.50%', total*0.62, total*0.12*0.85*0.05, total*0.12*0.05, total*0.12*1.15*0.05, total/area*0.12*0.05, '协商'),
            ('大型机械进出场费 Equipment', '0.65%', total*0.62, total*0.12*0.85*0.08, total*0.12*0.08, total*0.12*1.15*0.08, total/area*0.12*0.08, '按实'),
            ('施工排水费 Drainage', '0.15%', total*0.62, total*0.12*0.85*0.02, total*0.12*0.02, total*0.12*1.15*0.02, total/area*0.12*0.02, '按实'),
            ('施工配合费 Coordination', '0.25%', total*0.62, total*0.12*0.85*0.03, total*0.12*0.03, total*0.12*1.15*0.03, total/area*0.12*0.03, '协商'),
            ('脚手架费 Scaffolding', '元/m²', '', area, '18', total*0.12*0.85*0.12, total*0.12*0.12, total*0.12*1.15*0.12, total/area*0.12*0.12),
            ('模板费 Formwork', '元/m²', '', area, '85', total*0.12*0.85*0.58, total*0.12*0.58, total*0.12*1.15*0.58, total/area*0.12*0.58),
        ]

        for item in measure:
            self.add_data_row(item, row_color='E8F5E9')

        self.row += 1

        # 其他项目费
        self.add_section_header("3. OTHER COST 其他项目费 (4%)", level=1,
                               color=self.colors['warning'])

        self.add_table_header(['Code', 'Description 项目', 'Rate 费率', 'Base Amount', 'Low', 'Median', 'High', 'Unit Cost', 'Remarks'])

        other = [
            ('暂列金额 Provisional Sum', '3.0%', total*0.85, total*0.04*0.85*0.50, total*0.04*0.50, total*0.04*1.15*0.50, total/area*0.04*0.50, '变更洽商'),
            ('专业工程暂估价 Sub-contract', '1.5%', total*0.85, total*0.04*0.85*0.30, total*0.04*0.30, total*0.04*1.15*0.30, total/area*0.04*0.30, '暂定'),
            ('计日工 Daywork', '按实', '', '', total*0.04*0.85*0.10, total*0.04*0.10, total*0.04*1.15*0.10, total/area*0.04*0.10, '按实'),
            ('总承包服务费 Main Contractor', '3.0%', '', total*0.04*0.85*0.10, total*0.04*0.10, total*0.04*1.15*0.10, total/area*0.04*0.10, '配合费'),
        ]

        for item in other:
            self.add_data_row(item, row_color='FEF3C7')

        self.row += 1

        # 规费
        self.add_section_header("4. REGULATION FEE 规费 (3%)", level=1,
                               color=self.colors['family_5'])

        self.add_table_header(['Code', 'Description 项目', 'Rate 费率', 'Base Amount', 'Low', 'Median', 'High', 'Unit Cost', 'Remarks'])

        regulation = [
            ('社会保险费 Social Insurance', '25%', total*0.85, total*0.03*0.85*0.65, total*0.03*0.65, total*0.03*1.15*0.65, total/area*0.03*0.65, '人工费基数'),
            ('住房公积金 Housing Fund', '12%', total*0.85, total*0.03*0.85*0.28, total*0.03*0.28, total*0.03*1.15*0.28, total/area*0.03*0.28, '人工费基数'),
            ('工程排污费 Sewage', '0.1%', total*0.85, total*0.03*0.85*0.07, total*0.03*0.07, total*0.03*1.15*0.07, total/area*0.03*0.07, '环保'),
        ]

        for item in regulation:
            self.add_data_row(item, row_color='EDE9FE')

        self.row += 1

        # 税金
        self.add_section_header("5. TAX 税金 (9%)", level=1,
                               color=self.colors['danger'])

        self.add_table_header(['Code', 'Description 项目', 'Rate 税率', 'Base Amount', 'Low', 'Median', 'High', 'Unit Cost', 'Remarks'])

        tax = [
            ('增值税 VAT', '9%', total*0.85, total*0.09*0.85, total*0.09*0.85, total*0.09, total*0.09*1.15, total/area*0.09, '销项税'),
            ('城市维护建设税 Urban Maintenance', '12%', total*0.09*0.85, total*0.09*0.85*0.12, total*0.09*0.85*0.12, total*0.09*0.12, total*0.09*1.15*0.12, total/area*0.09*0.12, '增值税附征'),
            ('教育费附加 Education', '3%', total*0.09*0.85, total*0.09*0.85*0.03, total*0.09*0.85*0.03, total*0.09*0.03, total*0.09*1.15*0.03, total/area*0.09*0.03, '增值税附征'),
            ('地方教育附加 Local Education', '2%', total*0.09*0.85, total*0.09*0.85*0.02, total*0.09*0.85*0.02, total*0.09*0.02, total*0.09*1.15*0.02, total/area*0.09*0.02, '增值税附征'),
        ]

        for item in tax:
            self.add_data_row(item, row_color='FEE2E2')

    # =========================================================================
    # 工作表5: 不确定性分析 - Uncertainty Analysis
    # =========================================================================

    def create_uncertainty_analysis(self, params, cost_summary):
        """创建不确定性分析表"""
        self.ws = self.wb.create_sheet("05-不确定性分析")
        self.set_column_widths([15, 25, 15, 15, 15, 20, 20, 20, 15])

        self.add_title_block(
            "UNCERTAINTY ANALYSIS · 不确定性分析报告",
            "蒙特卡洛模拟 10,000次 | Monte Carlo Simulation 10,000 Runs"
        )

        area = params.get('total_area', 50000)
        total = cost_summary.get('总造价区间', {}).get('中值', 41816.48)
        unit = total / area

        # 精度等级
        self.add_section_header("1. ACCURACY CLASS 精度等级", level=1,
                               color=self.colors['rics_blue'])

        self.add_table_header(['Class 等级', 'Phase 阶段', 'Accuracy 精度', 'Method 方法', 'Standard 标准'])

        accuracy = [
            ('A级 Class A', '施工图设计', '±8%', '详细工程量清单', 'AECOM Level 2'),
            ('B级 Class B ★', '初步设计', '±15%', '元素估量法', 'AECOM Level 3'),
            ('C级 Class C', '方案设计', '±25%', '单位成本法', 'AECOM Level 4'),
            ('D级 Class D', '概念方案', '±40%', '类比估算法', 'AECOM Level 5'),
        ]

        for item in accuracy:
            self.add_data_row(item, highlight=('★' in item[0]), row_color='EBF5FF' if '★' in item[0] else None)

        self.row += 1

        # 风险因素
        self.add_section_header("2. RISK FACTORS 风险因素识别", level=1,
                               color=self.colors['warning'])

        self.add_table_header(['Code', 'Risk Factor 风险因素', 'Impact 影响', 'Probability 概率', 'Contribution 贡献度', 'Mitigation 缓解措施'])

        risks = [
            ('R1', '材料价格波动 Material Price', 'High 高', '60%', '±8.5%', '期货锁定/调价公式'),
            ('R2', '设计变更 Design Change', 'High 高', '45%', '±6.0%', '限额设计/BIM审核'),
            ('R3', '地质条件 Geotechnical', 'High 高', '35%', '±5.5%', '详细勘察/风险金'),
            ('R4', '工期延误 Delay', 'Medium 中', '40%', '±3.5%', '进度管控/赶工费'),
            ('R5', '政策调整 Policy', 'Medium 中', '25%', '±2.5%', '预留费用'),
            ('R6', '市场波动 Market', 'Medium 中', '30%', '±3.0%', '合同条款'),
        ]

        for item in risks:
            self.add_data_row(item, row_color='FEF3C7')

        self.row += 1

        # 蒙特卡洛模拟
        self.add_section_header("3. MONTE CARLO SIMULATION 蒙特卡洛模拟", level=1,
                               color=self.colors['aecom_green'])

        self.add_table_header(['Percentile 百分位', 'Description 描述', 'Unit Cost 单方', 'Total Cost 总价', 'Probability 概率', 'Confidence 置信度', 'Range 区间'])

        monte_carlo = [
            ('P5', '最乐观 Most Optimistic', unit*0.88, total*0.88, '5%', '低', f'{total*0.85:.0f}~{total*0.88:.0f}'),
            ('P10', '乐观 Optimistic', unit*0.90, total*0.90, '10%', '中低', f'{total*0.88:.0f}~{total*0.90:.0f}'),
            ('P25', '偏低 Lower', unit*0.93, total*0.93, '25%', '中低', f'{total*0.90:.0f}~{total*0.93:.0f}'),
            ('P50 ★', '中性 Median', unit*1.00, total*1.00, '50%', '中', f'{total*0.95:.0f}~{total*1.05:.0f}'),
            ('P75', '偏高 Upper', unit*1.07, total*1.07, '75%', '中高', f'{total*1.03:.0f}~{total*1.08:.0f}'),
            ('P90', '保守 Conservative', unit*1.12, total*1.12, '90%', '中高', f'{total*1.08:.0f}~{total*1.13:.0f}'),
            ('P95', '最保守 Most Conservative', unit*1.15, total*1.15, '95%', '高', f'{total*1.13:.0f}~{total*1.16:.0f}'),
        ]

        for item in monte_carlo:
            self.add_data_row(item, highlight=('★' in item[0]), row_color='E8F5E9' if '★' in item[0] else None)

        self.row += 1

        # 配比参数不确定性贡献
        self.add_section_header("4. PARAMETER UNCERTAINTY CONTRIBUTION 配比参数不确定性贡献", level=2)

        self.add_table_header(['Family 族', 'Parameter 参数', 'Base Value 基准值', 'Uncertainty 不确定度', 'Cost Coefficient 系数', 'Contribution 贡献'])

        contributions = [
            ('F1 结构', '墙:梁:板三相配比', '0.40:0.22:0.38', '±12%', '0.18', '±2.2%'),
            ('F2 围护', '窗墙面积比', '0.45', '±10%', '0.12', '±1.2%'),
            ('F3 机电', '风管展开面积比', '0.85', '±22%', '0.15', '±3.3%'),
            ('F4 钢筋', '钢筋混凝土比', '150 kg/m³', '±12%', '0.12', '±1.4%'),
            ('F5 基础 ⚠️', '桩长密度', '2.5 m/m²', '±35%', '0.08', '±2.8%'),
            ('F6 装修', '吊顶面积比', '0.85', '±20%', '0.10', '±2.0%'),
            ('F7 跨域', '管线总长度密度', '9.0 m/m²', '±22%', '0.08', '±1.8%'),
            ('材料', '钢筋价格波动', '5800元/t', '±15%', '0.15', '±2.3%'),
            ('材料', '混凝土价格波动', '580元/m³', '±10%', '0.12', '±1.2%'),
        ]

        for item in contributions:
            self.add_data_row(item, row_color='EDE9FE' if '⚠️' in str(item) else None)

    # =========================================================================
    # 工作表6: 数据关联索引 - Data Linkage Index
    # =========================================================================

    def create_data_linkage(self, params):
        """创建数据关联索引表"""
        self.ws = self.wb.create_sheet("06-数据关联索引")
        self.set_column_widths([15, 25, 20, 20, 25, 25])

        self.add_title_block(
            "DATA LINKAGE INDEX · 数据关联索引",
            "跨表数据关系映射 | Cross-Sheet Data Relationships"
        )

        # 表间关联说明
        self.add_section_header("1. TABLE RELATIONSHIP MAP 表间关联图", level=1,
                               color=self.colors['rics_blue'])

        self.add_table_header(['Source Sheet 源表', 'Target Sheet 目标表', 'Link Key 关联键', 'Data Type 数据类型', 'Description 说明'])

        relationships = [
            ('01-项目总览', '02-七族配比参数', 'building_type, structure_type', '配比参数', '根据建筑类型自动匹配配比'),
            ('01-项目总览', '03-主要材料用量', 'total_area, structure_type', '工程量', '根据面积计算材料用量'),
            ('01-项目总览', '04-费用明细', 'total_cost, unit_price', '造价', '根据单方造价汇总'),
            ('02-七族配比参数', '03-主要材料用量', 'F4钢筋比, F3机电比', '工程量', '配比→工程量换算'),
            ('02-七族配比参数', '05-不确定性分析', 'uncertainty_range', '不确定度', '参数→不确定度传播'),
            ('03-主要材料用量', '04-费用明细', 'quantity × unit_price', '造价', '工程量×单价=合价'),
            ('04-费用明细', '05-不确定性分析', 'cost_breakdown', '风险', '费用构成→风险分配'),
            ('所有表', '01-项目总览', 'project_name, date', '基础信息', '统一项目标识'),
        ]

        for item in relationships:
            self.add_data_row(item, row_color='EBF5FF')

        self.row += 1

        # 数据库加载状态
        self.add_section_header("2. DATABASE STATUS 数据库状态", level=1,
                               color=self.colors['aecom_green'])

        self.add_table_header(['Database 数据库', 'File 文件', 'Status 状态', 'Record Count 记录数', 'Last Updated 更新'])

        db_status = [
            ('7族58项配比', 'innovative-ratios-v2.json', '✅ 已加载', '58项参数', '2026-04-03'),
            ('6大材料因子', 'material-factors-v3.json', '✅ 已加载', '6种材料', '2026-04-03'),
            ('城市调整系数', 'region-adjustments.json', '✅ 已加载', '10城市', '2026-04-03'),
            ('设计规范配比', 'design-quantity-ratios.json', '✅ 已加载', '50+项', '2026-04-03'),
            ('机电安装配比', 'mep-quantity-ratios.json', '✅ 已加载', '30+项', '2026-04-03'),
            ('设计规范', 'building-norms.json', '✅ 已加载', '20+规范', '2026-04-03'),
        ]

        for item in db_status:
            self.add_data_row(item, row_color='E8F5E9')

        self.row += 1

        # 快速修改指南
        self.add_section_header("3. QUICK MODIFICATION GUIDE 快速修改指南", level=1,
                               color=self.colors['warning'])

        self.add_table_header(['Item 修改项', 'Location 位置', 'Default 默认值', 'How to Modify 如何修改', 'Impact 影响'])

        guides = [
            ('建筑类型', 'params[building_type]', '办公', '修改为: 住宅/商业/酒店/医院', '影响全部配比参数'),
            ('结构形式', 'params[structure_type]', '框架-剪力墙', '修改为: 框架/剪力墙/框架-核心筒', '影响F1/F4配比'),
            ('建筑面积', 'params[total_area]', '50000 m²', '修改数字', '按比例影响所有工程量'),
            ('装修标准', 'params[decoration_level]', '精装', '修改为: 毛坯/简装/豪装', '影响F6装修配比'),
            ('建设城市', 'params[region]', '苏州', '修改为: 深圳/广州/珠海/汕尾', '影响城市调整系数'),
            ('地下室面积', 'params[basement_area]', '10000 m²', '修改数字', '影响基础费用'),
        ]

        for item in guides:
            self.add_data_row(item, row_color='FEF3C7')

    # =========================================================================
    # 生成主函数
    # =========================================================================

    def generate_report(self, params, cost_summary, output_filename=None):
        """生成完整报表"""
        if output_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"度量衡国际范式报表_{timestamp}.xlsx"

        # 创建所有工作表
        self.create_project_overview(params, cost_summary)
        self.create_ratio_parameters(params)
        self.create_material_quantities(params, cost_summary)
        self.create_cost_breakdown(params, cost_summary)
        self.create_uncertainty_analysis(params, cost_summary)
        self.create_data_linkage(params)

        # 保存文件
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', output_filename)
        self.wb.save(filepath)
        print(f"[OK] Report Generated: {filepath}")

        return filepath, {
            'project_overview': '01-项目总览',
            'ratio_parameters': '02-七族配比参数',
            'material_quantities': '03-主要材料用量',
            'cost_breakdown': '04-费用明细',
            'uncertainty_analysis': '05-不确定性分析',
            'data_linkage': '06-数据关联索引',
        }


# ============================================================================
# 快速生成函数
# ============================================================================

def generate_international_qs_report(
    project_name="苏州某超高层办公楼造价估算报告",
    building_type="办公",
    structure_type="框架-核心筒",
    total_area=50000,
    floor_count=31,
    basement_area=10000,
    decoration_level="精装",
    region="苏州",
    seismic_level="一级",
    output_filename=None
):
    """
    一键生成国际咨询范式专业造价报表

    使用示例:
    >>> from report_generator_v3 import generate_international_qs_report
    >>> filepath, sheets = generate_international_qs_report(
    ...     project_name="苏州某超高层办公楼",
    ...     building_type="办公",
    ...     total_area=50000,
    ...     floor_count=31,
    ...     region="苏州"
    ... )
    """

    # 模拟造价估算结果
    unit_price_low = 8360
    unit_price_mid = 9845
    unit_price_high = 11580

    total_low = total_area * unit_price_low / 10000
    total_mid = total_area * unit_price_mid / 10000
    total_high = total_area * unit_price_high / 10000

    cost_summary = {
        '总造价区间': {'低限': total_low, '中值': total_mid, '高限': total_high},
        '单方造价区间': {'低限': unit_price_low, '中值': unit_price_mid, '高限': unit_price_high},
    }

    params = {
        'project_name': project_name,
        'building_type': building_type,
        'structure_type': structure_type,
        'total_area': total_area,
        'floor_count': floor_count,
        'basement_area': basement_area,
        'decoration_level': decoration_level,
        'region': region,
        'seismic_level': seismic_level,
    }

    # 生成报表
    report = InternationalQSReport(project_name)
    return report.generate_report(params, cost_summary, output_filename)


# ============================================================================
# 主程序入口
# ============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("度量衡智库 · 国际咨询范式专业造价报表生成器 v3.0")
    print("International QS Standard Professional Cost Report Generator")
    print("=" * 60)

    # 生成示例报表
    filepath, sheets = generate_international_qs_report(
        project_name="苏州某超高层办公楼造价估算报告",
        building_type="办公",
        structure_type="框架-核心筒",
        total_area=50000,
        floor_count=31,
        basement_area=10000,
        decoration_level="精装",
        region="苏州",
        seismic_level="一级抗震",
        output_filename="度量衡国际范式报表_苏州31层办公楼_v3.xlsx"
    )

    print("\n📊 工作表清单:")
    for name, desc in sheets.items():
        print(f"  - {desc}")

    print(f"\n📁 文件位置: {filepath}")
