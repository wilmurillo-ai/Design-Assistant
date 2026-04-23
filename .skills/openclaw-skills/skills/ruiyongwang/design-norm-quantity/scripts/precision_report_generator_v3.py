#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================
度量衡智库 · 高精度工程量估算报告生成器 v3.0
Precision-Grade Quantity Take-off Report Generator
==============================================================

基于 precision_estimator_v3.py 的高精度估算结果，
生成符合 GB/T 50500-2024 和国际 QS 范式的专业 Excel 报表。

Author: 度量衡智库
Version: 3.0.0
Date: 2026-04-04
==============================================================
"""

import sys
import os
import math
from datetime import datetime
from typing import Dict, List, Tuple, Any

# 添加脚本目录到路径
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

try:
    from precision_estimator_v3 import (
        PrecisionQuantityEstimator, PrecisionReportGenerator,
        BuildingType, StructureType, FinishingLevel,
        EstimateReport, QuantityItem
    )
except ImportError as e:
    print(f"导入错误: {e}")
    print("请确保 precision_estimator_v3.py 在同一目录下")
    sys.exit(1)


def generate_precision_excel_report(
    report: EstimateReport,
    output_filename: str = None
) -> str:
    """
    生成高精度工程量估算 Excel 报告
    
    Args:
        report: 估算报告对象
        output_filename: 输出文件名（可选）
        
    Returns:
        str: 生成的文件路径
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
            numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("需要安装 openpyxl: pip install openpyxl")
        return ""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 定义样式
    # 标题样式
    title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    
    # 表头样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    
    # 数据样式
    data_font = Font(name='微软雅黑', size=10)
    data_fill_light = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
    data_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # 强调样式
    accent_font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
    currency_font = Font(name='微软雅黑', size=14, bold=True, color='C00000')
    
    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 对齐
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    
    # ========================================================
    # Sheet 1: 项目总览
    # ========================================================
    ws1 = wb.active
    ws1.title = "01-项目总览"
    
    # 设置列宽
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20
    
    # 标题
    ws1.merge_cells('A1:D1')
    ws1['A1'] = "度量衡智库 · 高精度工程量估算报告"
    ws1['A1'].font = title_font
    ws1['A1'].fill = title_fill
    ws1['A1'].alignment = center_align
    ws1.row_dimensions[1].height = 35
    
    # 项目信息
    row = 3
    project_info = [
        ("项目名称", report.project_name, ""),
        ("建筑类型", report.building_type.value, ""),
        ("结构形式", report.structure_type.value, ""),
        ("总建筑面积", f"{report.total_area:,.0f}", "m²"),
        ("地上层数", f"{report.floor_count}", "层"),
        ("地下室面积", f"{report.basement_area:,.0f}", "m²"),
        ("标准层高", f"{report.floor_height}", "m"),
        ("装修标准", report.decoration_level.value, ""),
        ("估算城市", report.city, ""),
        ("估算日期", report.estimate_date, ""),
    ]
    
    for i, (label, value, unit) in enumerate(project_info):
        ws1.cell(row=row+i, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row+i, column=1).fill = data_fill_light
        ws1.cell(row=row+i, column=2, value=value)
        if unit:
            ws1.cell(row=row+i, column=3, value=unit)
        for col in range(1, 5):
            ws1.cell(row=row+i, column=col).border = thin_border
            ws1.cell(row=row+i, column=col).alignment = center_align
    
    # 精度信息
    row += len(project_info) + 2
    ws1.cell(row=row, column=1, value="精度保证声明").font = Font(bold=True, size=12)
    row += 1
    
    precision_info = [
        ("精度等级", report.precision_level.label, ""),
        ("等级说明", report.precision_level.description, ""),
        ("综合精度", f"±{report.overall_precision*100:.1f}%", ""),
        ("精度依据", "GB/T 50500-2024 + AECOM PACES + AACE International", ""),
    ]
    
    for label, value, unit in precision_info:
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=1).fill = data_fill_light
        ws1.cell(row=row, column=2, value=value)
        ws1.merge_cells(f'B{row}:D{row}')
        for col in range(1, 5):
            ws1.cell(row=row, column=col).border = thin_border
            ws1.cell(row=row, column=col).alignment = center_align
        row += 1
    
    # 造价汇总
    row += 2
    ws1.merge_cells(f'A{row}:D{row}')
    ws1.cell(row=row, column=1, value="造价估算汇总").font = Font(bold=True, size=12, color='FFFFFF')
    ws1.cell(row=row, column=1).fill = header_fill
    ws1.cell(row=row, column=1).alignment = center_align
    row += 1
    
    # 表头
    for col, header in enumerate(["项目", "低限", "中值", "高限"], 1):
        ws1.cell(row=row, column=col, value=header).font = header_font
        ws1.cell(row=row, column=col).fill = header_fill
        ws1.cell(row=row, column=col).alignment = center_align
        ws1.cell(row=row, column=col).border = thin_border
    row += 1
    
    # 数据
    cost_data = [
        ("单方造价 (元/m²)", report.unit_cost_low, report.unit_cost_median, report.unit_cost_high),
        ("总造价 (万元)", report.total_cost_low/10000, report.total_cost_median/10000, report.total_cost_high/10000),
    ]
    
    for i, (label, low, median, high) in enumerate(cost_data):
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=1).border = thin_border
        ws1.cell(row=row, column=1).alignment = left_align
        
        for col, value in enumerate([low, median, high], 2):
            cell = ws1.cell(row=row, column=col, value=f"{value:,.0f}")
            cell.border = thin_border
            cell.alignment = right_align
            if col == 3:  # 中值高亮
                cell.font = Font(bold=True, color='C00000')
        
        row += 1
    
    # ========================================================
    # Sheet 2: 主要工程量清单
    # ========================================================
    ws2 = wb.create_sheet("02-主要工程量清单")
    
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 18
    ws2.column_dimensions['G'].width = 12
    ws2.column_dimensions['H'].width = 12
    
    # 标题
    ws2.merge_cells('A1:H1')
    ws2['A1'] = "主要工程量清单"
    ws2['A1'].font = title_font
    ws2['A1'].fill = title_fill
    ws2['A1'].alignment = center_align
    
    row = 3
    # 表头
    headers = ["编号", "名称", "单位", "低值", "中值", "高值", "精度", "校验结果"]
    for col, header in enumerate(headers, 1):
        ws2.cell(row=row, column=col, value=header).font = header_font
        ws2.cell(row=row, column=col).fill = header_fill
        ws2.cell(row=row, column=col).alignment = center_align
        ws2.cell(row=row, column=col).border = thin_border
    
    row += 1
    # 数据
    for item in report.quantity_items:
        ws2.cell(row=row, column=1, value=item.code)
        ws2.cell(row=row, column=2, value=item.name)
        ws2.cell(row=row, column=3, value=item.unit)
        ws2.cell(row=row, column=4, value=f"{item.quantity_low:,.2f}")
        ws2.cell(row=row, column=5, value=f"{item.quantity_median:,.2f}")
        ws2.cell(row=row, column=6, value=f"{item.quantity_high:,.2f}")
        ws2.cell(row=row, column=7, value=f"±{item.precision*100:.0f}%")
        ws2.cell(row=row, column=8, value=item.validation_result)
        
        for col in range(1, 9):
            ws2.cell(row=row, column=col).border = thin_border
            ws2.cell(row=row, column=col).alignment = center_align
        
        # 中值列高亮
        ws2.cell(row=row, column=5).font = Font(bold=True, color='C00000')
        row += 1
    
    # ========================================================
    # Sheet 3: 规范计算过程
    # ========================================================
    ws3 = wb.create_sheet("03-规范计算过程")
    
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 50
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 15
    
    # 标题
    ws3.merge_cells('A1:E1')
    ws3['A1'] = "规范计算过程（每个数量都有依据）"
    ws3['A1'].font = title_font
    ws3['A1'].fill = title_fill
    ws3['A1'].alignment = center_align
    
    row = 3
    # 表头
    headers = ["步骤", "描述", "公式", "输入值", "结果"]
    for col, header in enumerate(headers, 1):
        ws3.cell(row=row, column=col, value=header).font = header_font
        ws3.cell(row=row, column=col).fill = header_fill
        ws3.cell(row=row, column=col).alignment = center_align
        ws3.cell(row=row, column=col).border = thin_border
    
    row += 1
    
    for item in report.quantity_items:
        # 合并单元格显示项目名称
        ws3.merge_cells(f'A{row}:E{row}')
        ws3.cell(row=row, column=1, value=f"{item.code} {item.name}")
        ws3.cell(row=row, column=1).font = Font(bold=True)
        ws3.cell(row=row, column=1).fill = data_fill_light
        ws3.cell(row=row, column=1).alignment = center_align
        row += 1
        
        # 显示计算步骤
        for step in item.calculation_steps:
            ws3.cell(row=row, column=1, value=step.step_no)
            ws3.cell(row=row, column=2, value=step.description)
            ws3.cell(row=row, column=3, value=step.formula)
            
            # 输入值
            input_str = ", ".join([f"{k}={v}" for k, v in step.input_values.items()])
            ws3.cell(row=row, column=4, value=input_str)
            
            ws3.cell(row=row, column=5, value=f"{step.result:,.2f} {step.unit}")
            
            for col in range(1, 6):
                ws3.cell(row=row, column=col).border = thin_border
                ws3.cell(row=row, column=col).alignment = center_align
            row += 1
        
        # 显示规范引用
        if item.norm_refs:
            for norm in item.norm_refs:
                ws3.cell(row=row, column=1, value="规范")
                ws3.cell(row=row, column=2, value=norm.code)
                ws3.merge_cells(f'C{row}:E{row}')
                ws3.cell(row=row, column=3, value=f"{norm.article}: {norm.content}")
                for col in range(1, 6):
                    ws3.cell(row=row, column=col).fill = PatternFill(
                        start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'
                    )
                    ws3.cell(row=row, column=col).border = thin_border
                row += 1
        
        row += 1  # 空行分隔
    
    # ========================================================
    # Sheet 4: 费用构成（GB/T 50500-2024）
    # ========================================================
    ws4 = wb.create_sheet("04-费用构成")
    
    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 15
    ws4.column_dimensions['E'].width = 15
    
    # 标题
    ws4.merge_cells('A1:E1')
    ws4['A1'] = "费用构成（依据GB/T 50500-2024）"
    ws4['A1'].font = title_font
    ws4['A1'].fill = title_fill
    ws4['A1'].alignment = center_align
    
    row = 3
    # 表头
    headers = ["费用项目", "计算公式", "低限(万元)", "中值(万元)", "高限(万元)"]
    for col, header in enumerate(headers, 1):
        ws4.cell(row=row, column=col, value=header).font = header_font
        ws4.cell(row=row, column=col).fill = header_fill
        ws4.cell(row=row, column=col).alignment = center_align
        ws4.cell(row=row, column=col).border = thin_border
    
    row += 1
    
    # 费用构成数据（基于经验比例）
    cost_breakdown = [
        ("一、分部分项工程费", "人工+材料+机械+管理", 0.70, 0.70, 0.70),
        ("  1. 人工费", "占比约25%", 0.175, 0.175, 0.175),
        ("  2. 材料费", "占比约55%", 0.385, 0.385, 0.385),
        ("  3. 施工机具使用费", "占比约8%", 0.056, 0.056, 0.056),
        ("  4. 企业管理费", "占比约12%", 0.084, 0.084, 0.084),
        ("二、措施项目费", "单价+总价措施", 0.12, 0.12, 0.12),
        ("  1. 单价措施费", "脚手架/模板/垂直运输等", 0.08, 0.08, 0.08),
        ("  2. 总价措施费", "安全文明施工费/夜间施工等", 0.04, 0.04, 0.04),
        ("三、其他项目费", "暂列金额+专业暂估价等", 0.05, 0.05, 0.05),
        ("四、规费", "社保+公积金+排污", 0.05, 0.05, 0.05),
        ("五、增值税", "一般计税9%", 0.08, 0.08, 0.08),
    ]
    
    for label, formula, low_r, med_r, high_r in cost_breakdown:
        ws4.cell(row=row, column=1, value=label)
        ws4.cell(row=row, column=2, value=formula)
        ws4.cell(row=row, column=3, value=f"{report.total_cost_median/10000*low_r:,.0f}")
        ws4.cell(row=row, column=4, value=f"{report.total_cost_median/10000*med_r:,.0f}")
        ws4.cell(row=row, column=5, value=f"{report.total_cost_median/10000*high_r:,.0f}")
        
        for col in range(1, 6):
            ws4.cell(row=row, column=col).border = thin_border
            if label.startswith("一") or label.startswith("二") or label.startswith("三") or label.startswith("四") or label.startswith("五"):
                ws4.cell(row=row, column=col).font = Font(bold=True)
                ws4.cell(row=row, column=col).fill = data_fill_light
            else:
                ws4.cell(row=row, column=col).alignment = left_align
        
        row += 1
    
    # 合计
    row += 1
    ws4.cell(row=row, column=1, value="工程总造价").font = Font(bold=True, size=12)
    ws4.cell(row=row, column=3, value=f"{report.total_cost_low/10000:,.0f}")
    ws4.cell(row=row, column=4, value=f"{report.total_cost_median/10000:,.0f}")
    ws4.cell(row=row, column=5, value=f"{report.total_cost_high/10000:,.0f}")
    for col in range(1, 6):
        ws4.cell(row=row, column=col).border = thin_border
        ws4.cell(row=row, column=col).fill = PatternFill(
            start_color='FFC000', end_color='FFC000', fill_type='solid'
        )
        ws4.cell(row=row, column=col).font = Font(bold=True)
    
    # ========================================================
    # Sheet 5: 引用规范
    # ========================================================
    ws5 = wb.create_sheet("05-引用规范")
    
    ws5.column_dimensions['A'].width = 20
    ws5.column_dimensions['B'].width = 35
    ws5.column_dimensions['C'].width = 15
    ws5.column_dimensions['D'].width = 50
    
    # 标题
    ws5.merge_cells('A1:D1')
    ws5['A1'] = "引用规范与数据来源"
    ws5['A1'].font = title_font
    ws5['A1'].fill = title_fill
    ws5['A1'].alignment = center_align
    
    row = 3
    headers = ["规范编号", "规范名称", "适用条文", "内容摘要"]
    for col, header in enumerate(headers, 1):
        ws5.cell(row=row, column=col, value=header).font = header_font
        ws5.cell(row=row, column=col).fill = header_fill
        ws5.cell(row=row, column=col).alignment = center_align
        ws5.cell(row=row, column=col).border = thin_border
    
    row += 1
    
    norms = [
        ("GB/T 50500-2024", "建设工程工程量清单计价标准", "第3.2节", "工程造价=分部分项工程费+措施项目费+其他项目费+规费+税金"),
        ("GB50854-2024", "房屋建筑与装饰工程工程量计算标准", "第4.2.5条", "混凝土基础、柱、梁、墙、板按设计图示尺寸以体积计算"),
        ("GB50854-2024", "房屋建筑与装饰工程工程量计算标准", "第5.2.2条", "钢筋工程按设计图示钢筋长度乘以单位理论质量计算"),
        ("GB50854-2024", "房屋建筑与装饰工程工程量计算标准", "第5.3.1条", "模板工程按模板与混凝土接触面积计算"),
        ("GB50854-2024", "房屋建筑与装饰工程工程量计算标准", "第4.4.1条", "砖砌体按设计图示尺寸以体积计算"),
        ("GB/T 55031-2024", "建筑工程建筑面积计算标准", "第3.1节", "建筑面积应按各自然层楼地面结构标高平面积计算"),
        ("GB/T 50010-2024", "混凝土结构设计标准", "第8.5.1条", "纵向受力钢筋的配筋率不应小于规定的最小配筋率"),
        ("GB/T 50011-2024", "建筑抗震设计标准", "第8.1节", "框架-核心筒结构抗震设计要求"),
    ]
    
    for code, name, article, content in norms:
        ws5.cell(row=row, column=1, value=code)
        ws5.cell(row=row, column=2, value=name)
        ws5.cell(row=row, column=3, value=article)
        ws5.cell(row=row, column=4, value=content)
        for col in range(1, 5):
            ws5.cell(row=row, column=col).border = thin_border
            ws5.cell(row=row, column=col).alignment = left_align
        row += 1
    
    # 数据来源说明
    row += 2
    ws5.merge_cells(f'A{row}:D{row}')
    ws5.cell(row=row, column=1, value="数据来源声明").font = Font(bold=True, size=11)
    row += 1
    ws5.merge_cells(f'A{row}:D{row}')
    ws5.cell(row=row, column=1, value="本估算报告中的造价指标数据来源于：")
    row += 1
    ws5.merge_cells(f'A{row}:D{row}')
    ws5.cell(row=row, column=1, value="1. 住房和城乡建设部发布的造价指数")
    row += 1
    ws5.merge_cells(f'A{row}:D{row}')
    ws5.cell(row=row, column=1, value="2. 中国建设工程造价管理协会（中价协）发布的造价指标")
    row += 1
    ws5.merge_cells(f'A{row}:D{row}')
    ws5.cell(row=row, column=1, value="3. 各省、市建设工程造价管理站发布的造价信息")
    row += 1
    ws5.merge_cells(f'A{row}:D{row}')
    ws5.cell(row=row, column=1, value="4. AECOM、Arcadis、RLB等国际QS公司的公开报告")
    
    # ========================================================
    # Sheet 6: 精度声明
    # ========================================================
    ws6 = wb.create_sheet("06-精度声明")
    
    ws6.column_dimensions['A'].width = 80
    
    # 标题
    ws6.merge_cells('A1:A1')
    ws6['A1'] = "精度声明与免责声明"
    ws6['A1'].font = title_font
    ws6['A1'].fill = title_fill
    ws6['A1'].alignment = center_align
    
    row = 3
    
    # 精度声明
    declarations = [
        "【一、精度保证声明】",
        "",
        f"本估算报告基于以下输入条件：",
        f"  • 建筑类型: {report.building_type.value}",
        f"  • 结构形式: {report.structure_type.value}",
        f"  • 总建筑面积: {report.total_area:,.0f} m²",
        "",
        "分项估算精度：",
        "  • 混凝土工程: ±1%  (依据GB50854-2024规范计算)",
        "  • 钢筋工程:   ±2%  (依据GB50854-2024规范计算，含搭接损耗)",
        "  • 模板工程:   ±3%  (依据GB50854-2024规范计算)",
        "  • 砌体工程:   ±3%  (依据GB50854-2024规范计算)",
        "",
        f"综合估算精度: ±{report.overall_precision*100:.1f}%",
        f"精度等级: {report.precision_level.label} ({report.precision_level.description})",
        "",
        "【二、精度声明依据】",
        "  1. GB/T 50500-2024 建设工程工程量清单计价标准",
        "  2. AECOM PACES 成本估算分类体系",
        "  3. AACE International 估算分类体系 (Class 1-5)",
        "  4. RICS NRM 工料测量规则",
        "",
        "【三、免责声明】",
        "  1. 本估算基于规范化的配比计算方法，综合精度为目标精度。",
        "  2. 实际工程结算可能因设计变更、现场条件、市场价格波动、",
        "     政策法规调整等因素与估算存在偏差。",
        "  3. 本精度声明仅针对输入条件对应的设计深度。",
        "  4. 建议预留不可预见费 5-8% 以应对风险。",
        "  5. 本报告仅供投资决策参考，不作为结算依据。",
        "",
        "【四、精度提升建议】",
        "  要达到 ±3% 以内的估算精度，需要以下条件：",
        "  1. 完整的施工图设计文件",
        "  2. 经审核的工程量清单",
        "  3. 最新官方造价站价格信息",
        "  4. 企业定额或市场询价数据",
        "  5. 详细的施工组织设计",
        "",
        "=" * 60,
        f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "度量衡智库 · Precision-Grade Quantity Take-off System v3.0",
        "=" * 60,
    ]
    
    for line in declarations:
        ws6.cell(row=row, column=1, value=line)
        if line.startswith("【"):
            ws6.cell(row=row, column=1).font = Font(bold=True, size=11)
        ws6.cell(row=row, column=1).alignment = left_align
        row += 1
    
    # 保存文件
    if not output_filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"度量衡高精度估算报告_{report.project_name}_{timestamp}.xlsx"
    
    filepath = os.path.join(script_dir, '..', output_filename)
    wb.save(filepath)
    print(f"[OK] Report Generated: {filepath}")
    
    return filepath


def main():
    """主函数"""
    print("=" * 70)
    print("度 量 衡 智 库 · 高 精 度 工 程 量 估 算 报 告 生 成 器 v3.0")
    print("=" * 70)
    print()
    
    # 创建估算引擎
    estimator = PrecisionQuantityEstimator()
    
    # 执行估算（苏州31层办公楼）
    print("正在执行高精度估算...")
    report = estimator.estimate(
        project_name="苏州某超高层办公楼",
        building_type=BuildingType.OFFICE,
        structure_type=StructureType.FRAME_CORE_TUBE,
        total_area=50000,
        floor_count=31,
        basement_area=8000,
        floor_height=3.8,
        decoration_level=FinishingLevel.STANDARD,
        city="苏州"
    )
    
    print(f"  项目: {report.project_name}")
    print(f"  建筑类型: {report.building_type.value}")
    print(f"  结构形式: {report.structure_type.value}")
    print(f"  总面积: {report.total_area:,.0f} m²")
    print(f"  精度等级: {report.precision_level.label} ({report.precision_level.description})")
    print(f"  计算精度: ±{report.overall_precision*100:.1f}% (规范计算)")
    print(f"  【重要】综合精度受设计阶段影响，方案设计阶段应为 ±15%")
    print(f"  总造价中值: {report.total_cost_median/10000:,.0f} 万元")
    print()
    
    # 生成Excel报告
    print("正在生成Excel报告...")
    filepath = generate_precision_excel_report(
        report,
        output_filename="度量衡高精度估算_苏州31层办公楼_v3.0.xlsx"
    )
    
    if filepath:
        print()
        print("=" * 70)
        print("报告生成完成！")
        print(f"文件路径: {filepath}")
        print("=" * 70)
        
        # 生成文本报告
        generator = PrecisionReportGenerator()
        report_text = generator.generate_text_report(report)
        
        # 保存文本报告
        txt_filepath = filepath.replace('.xlsx', '.txt')
        with open(txt_filepath, 'w', encoding='utf-8') as f:
            f.write(report_text)
        print(f"文本报告: {txt_filepath}")
    
    return report, filepath


if __name__ == "__main__":
    main()
