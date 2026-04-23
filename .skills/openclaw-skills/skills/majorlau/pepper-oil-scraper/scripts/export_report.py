#!/usr/bin/env python3
"""
export_report.py — 将爬取数据导出为 Excel 报告
每个维度一个 sheet，带格式化表头和数据源链接
"""

import json
import glob
import argparse
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def load_data(input_dir: str) -> list[dict]:
    """加载目录下所有 JSON 数据文件"""
    all_records = []
    json_files = glob.glob(str(Path(input_dir) / "*.json"))

    for fp in sorted(json_files):
        try:
            with open(fp, "r", encoding="utf-8") as f:
                data = json.load(f)
            # merged file
            if "records" in data and isinstance(data["records"], list):
                all_records.extend(data["records"])
            # single adapter file
            elif "records" in data:
                for rec in data["records"]:
                    rec["_source_adapter"] = data.get("adapter_id", "")
                    rec["_source_category"] = data.get("category", "")
                    all_records.extend(data["records"])
        except Exception as e:
            print(f"  跳过 {fp}: {e}")

    return all_records


def categorize(records: list[dict]) -> dict[str, list[dict]]:
    """分类到各维度"""
    cats = {
        "1_市场规模": [],
        "2_原料价格": [],
        "3_产量面积": [],
        "4_企业财务": [],
        "5_进出口": [],
        "6_渠道占比": [],
        "7_政策标准": [],
        "8_全部数据": records,
    }

    for rec in records:
        metric = (rec.get("metric", "") or "").lower()
        if any(k in metric for k in ["市场规模", "market", "cagr", "全球"]):
            cats["1_市场规模"].append(rec)
        elif any(k in metric for k in ["价格", "报价", "指数", "price", "批发"]):
            cats["2_原料价格"].append(rec)
        elif any(k in metric for k in ["面积", "产量", "种植", "消费量"]):
            cats["3_产量面积"].append(rec)
        elif any(k in metric for k in ["营业", "净利", "毛利", "产能", "营收"]):
            cats["4_企业财务"].append(rec)
        elif any(k in metric for k in ["进出口", "出口", "进口", "海关", "hs编码"]):
            cats["5_进出口"].append(rec)
        elif any(k in metric for k in ["渠道", "占比", "份额"]):
            cats["6_渠道占比"].append(rec)
        elif any(k in metric for k in ["标准", "gb", "db", "政策"]):
            cats["7_政策标准"].append(rec)

    return cats


def records_to_df(records: list[dict]) -> "pd.DataFrame":
    """转换记录列表为 DataFrame"""
    rows = []
    for rec in records:
        rows.append({
            "指标": rec.get("metric", ""),
            "数值": rec.get("value", ""),
            "单位": rec.get("unit", ""),
            "标准化数值": rec.get("value_standardized", rec.get("value", "")),
            "标准化单位": rec.get("unit_standardized", rec.get("unit", "")),
            "时间": rec.get("period", ""),
            "地区": rec.get("region", ""),
            "置信度": rec.get("confidence", ""),
            "数据来源": rec.get("source_url", ""),
            "原文": (rec.get("original_text", "") or "")[:200],
            "采集站点": rec.get("_source_adapter", ""),
        })
    return pd.DataFrame(rows)


def export_excel(input_dir: str, output_path: str):
    """导出 Excel 报告"""
    if not HAS_OPENPYXL:
        print("错误：需要安装 openpyxl 和 pandas")
        print("  pip install pandas openpyxl --break-system-packages")
        return

    print(f"加载数据: {input_dir}")
    records = load_data(input_dir)
    print(f"  总记录数: {len(records)}")

    if not records:
        print("  无数据可导出")
        return

    cats = categorize(records)

    wb = Workbook()
    wb.remove(wb.active)

    # 样式
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    for sheet_name, recs in cats.items():
        if not recs:
            continue

        df = records_to_df(recs)
        ws = wb.create_sheet(title=sheet_name[:31])

        # 写入表头
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 写入数据
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                # 如果是 URL 列，加超链接
                if headers[col_idx - 1] == "数据来源" and isinstance(val, str) and val.startswith("http"):
                    cell.font = link_font
                    try:
                        cell.hyperlink = val
                    except Exception:
                        pass

        # 列宽
        col_widths = {
            "指标": 25, "数值": 12, "单位": 10,
            "标准化数值": 12, "标准化单位": 10,
            "时间": 12, "地区": 12, "置信度": 8,
            "数据来源": 50, "原文": 60, "采集站点": 15,
        }
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(header, 15)

        # 冻结首行
        ws.freeze_panes = "A2"
        # 自动筛选
        ws.auto_filter.ref = ws.dimensions

        print(f"  Sheet '{sheet_name}': {len(recs)} 行")

    # 添加概览 sheet
    ws_overview = wb.create_sheet(title="概览", index=0)
    ws_overview["A1"] = "花椒油/藤椒油产业数据采集报告"
    ws_overview["A1"].font = Font(name="微软雅黑", bold=True, size=16, color="C0392B")
    ws_overview["A3"] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_overview["A4"] = f"总记录数: {len(records)}"
    ws_overview["A6"] = "各维度数据量:"
    for i, (name, recs) in enumerate(cats.items()):
        ws_overview.cell(row=7 + i, column=1, value=name)
        ws_overview.cell(row=7 + i, column=2, value=len(recs))
    ws_overview.column_dimensions["A"].width = 30
    ws_overview.column_dimensions["B"].width = 15

    # 保存
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n✅ Excel 报告已保存: {output_path}")


def export_markdown(input_dir: str, output_path: str):
    """导出 Markdown 报告（备选）"""
    records = load_data(input_dir)
    cats = categorize(records)

    lines = ["# 花椒油/藤椒油产业数据采集报告\n"]
    lines.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    lines.append(f"总记录数: {len(records)}\n")

    for sheet_name, recs in cats.items():
        if not recs or sheet_name == "8_全部数据":
            continue
        lines.append(f"\n## {sheet_name}\n")
        lines.append(f"| 指标 | 数值 | 单位 | 时间 | 地区 | 来源 |")
        lines.append(f"|------|------|------|------|------|------|")
        for rec in recs[:50]:  # 限制行数
            metric = rec.get("metric", "")
            value = rec.get("value", "")
            unit = rec.get("unit", "")
            period = rec.get("period", "")
            region = rec.get("region", "")
            url = rec.get("source_url", "")
            if url:
                source = f"[链接]({url})"
            else:
                source = ""
            lines.append(f"| {metric} | {value} | {unit} | {period} | {region} | {source} |")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"\n✅ Markdown 报告已保存: {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="导出爬取数据报告")
    parser.add_argument("--input", "-i", required=True, help="数据目录（含 JSON 文件）")
    parser.add_argument("--output", "-o", required=True, help="输出文件路径（.xlsx 或 .md）")
    parser.add_argument("--format", "-f", choices=["xlsx", "md"], default=None, help="输出格式")
    args = parser.parse_args()

    fmt = args.format or ("md" if args.output.endswith(".md") else "xlsx")

    if fmt == "xlsx":
        export_excel(args.input, args.output)
    else:
        export_markdown(args.input, args.output)
