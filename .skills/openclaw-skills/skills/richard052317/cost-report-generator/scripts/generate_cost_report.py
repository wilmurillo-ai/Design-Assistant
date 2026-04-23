import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sys
import os

def generate_cost_report(source_file, output_dir=None):
    """
    从项目费用 Excel 文件生成成本分析报告
    
    Args:
        source_file: 源 Excel 文件路径（如 EVP 项目费用.xlsx）
        output_dir: 输出目录，默认为源文件所在目录
    
    Returns:
        输出文件路径
    """
    # 读取源文件（第 4 行作为表头，索引从 1 开始）
    # 尝试读取第一个工作表，表头在第 2 行（索引 1）
    # HP 项目文件格式：表头在第 2 行（索引 1）
    df = pd.read_excel(source_file, sheet_name=0, header=1)
    print(f"读取数据：{len(df)}行，{len(df.columns)}列")
    
    # HP 项目文件：第 0 行是数据，不需要跳过
    # 提取基本信息
    customers = [str(df.iloc[i]['客户大类']) if pd.notna(df.iloc[i]['客户大类']) else '' for i in range(len(df))]
    projects = [str(df.iloc[i]['项目大类']) if pd.notna(df.iloc[i]['项目大类']) else '' for i in range(len(df))]
    order_ids = [str(df.iloc[i]['内部订单号']) if pd.notna(df.iloc[i]['内部订单号']) else '' for i in range(len(df))]
    
    # 筛选有效客户分类
    valid_customers = ['帝国', 'BAT', 'MT', 'ITMS']
    valid_indices = [i for i in range(len(df)) if customers[i] in valid_customers]
    
    # 按客户分类分组
    customer_groups = {}
    for cust_type in valid_customers:
        customer_groups[cust_type] = [i for i in valid_indices if customers[i] == cust_type]
    
    print(f"客户分布：{ {k: len(v) for k, v in customer_groups.items()} }")
    
    # 列分类
    def find_cols(keyword):
        return [col for col in df.columns if keyword in str(col)]
    
    hr_cols = find_cols('人工费用')
    sample_cols = find_cols('打样费用')
    manufacturing_cols = find_cols('制造费用')
    material_cols = find_cols('物料耗用')
    test_cols = find_cols('测试费用')
    cert_cols = find_cols('外发认证费用')
    mold_cols = find_cols('开模费用')
    modify_cols = find_cols('改模费用')
    other_cols = find_cols('IDC 费用') + find_cols('品烟费用') + find_cols('其它费用') + find_cols('手板费用')
    
    def extract_month(col_str):
        if '1-' in col_str:
            return '1-2 月'
        if '年' in col_str and '月' in col_str:
            year_pos = col_str.find('年')
            month_pos = col_str.find('月')
            if year_pos >= 0 and month_pos > year_pos:
                between = col_str[year_pos+1:month_pos]
                if between.isdigit():
                    return f"{between} 月"
        return None
    
    def get_data(row_idx, cat_cols, year):
        year_short = str(year)[-2:]
        months_order = ['1-2 月', '1 月', '2 月', '3 月', '4 月', '5 月', '6 月', '7 月', '8 月', '9 月', '10 月', '11 月', '12 月'] if year == 2026 else ['1 月', '2 月', '3 月', '4 月', '5 月', '6 月', '7 月', '8 月', '9 月', '10 月', '11 月', '12 月']
        
        yearly_total = 0
        monthly_data = {m: 0.0 for m in months_order}
        actual_months_count = 0
        
        for col in cat_cols:
            col_str = str(col)
            if year_short not in col_str or '年' not in col_str:
                continue
            val = df.iloc[row_idx][col]
            try:
                num = float(val) if pd.notna(val) else 0.0
            except:
                num = 0.0
            
            if '月' not in col_str and '1-' not in col_str:
                yearly_total += num
            else:
                month = extract_month(col_str)
                if month and month in monthly_data:
                    monthly_data[month] += num
                    if num != 0:
                        actual_months_count += 1
        
        if yearly_total == 0:
            yearly_total = sum(v for v in monthly_data.values())
        
        monthly_sum = sum(v for v in monthly_data.values())
        max_months = 2 if year == 2026 else 12
        divisor = min(actual_months_count, max_months) if actual_months_count > 0 else max_months
        monthly_avg = monthly_sum / divisor if divisor > 0 else 0
        
        return round(yearly_total, 2), {m: (round(v, 2) if v != 0 else None) for m, v in monthly_data.items()}, round(monthly_avg, 2)
    
    # 创建工作簿
    wb = Workbook()
    del wb['Sheet']
    
    # 样式定义
    header_font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    subheader_font = Font(name='微软雅黑', bold=True, size=9, color='333333')
    normal_font = Font(name='微软雅黑', size=9)
    bold_font = Font(name='微软雅黑', bold=True, size=9)
    header_fill_blue = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
    header_fill_light = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    num_format = '#,##0.00'
    thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    
    def create_sheet(name, year):
        ws = wb.create_sheet(name)
        months = ['1-2 月', '1 月', '2 月', '3 月', '4 月', '5 月', '6 月', '7 月', '8 月', '9 月', '10 月', '11 月', '12 月'] if year == 2026 else ['1 月', '2 月', '3 月', '4 月', '5 月', '6 月', '7 月', '8 月', '9 月', '10 月', '11 月', '12 月']
        categories = ['HR 费用', '试产/打样费用', '测试费用', '开发投资', '其他费用']
        num_months = len(months)
        
        col = 1
        ws.merge_cells(start_row=1, start_column=1, end_column=3, end_row=1)
        cell = ws.cell(row=1, column=1, value="基本信息")
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill_blue, center_align, thin_border
        
        col = 4
        for cat in categories:
            end_col = col + num_months + 1
            ws.merge_cells(start_row=1, start_column=col, end_column=end_col, end_row=1)
            cell = ws.cell(row=1, column=col, value=cat)
            cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill_blue, center_align, thin_border
            col = end_col + 1
        
        ws.merge_cells(start_row=1, start_column=col, end_column=col+1, end_row=1)
        cell = ws.cell(row=1, column=col, value="总计")
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill_blue, center_align, thin_border
        
        for c, val in enumerate(["客户大类", "项目大类", "内部订单号"], 1):
            cell = ws.cell(row=2, column=c, value=val)
            cell.font = subheader_font
            cell.fill = header_fill_light
            cell.alignment = center_align
            cell.border = thin_border
        
        col = 4
        for cat in categories:
            cell = ws.cell(row=2, column=col, value="汇总")
            cell.font = subheader_font
            cell.fill = header_fill_light
            cell.alignment = center_align
            cell.border = thin_border
            col += 1
            cell = ws.cell(row=2, column=col, value="月均")
            cell.font = subheader_font
            cell.fill = header_fill_light
            cell.alignment = center_align
            cell.border = thin_border
            col += 1
            for month in months:
                cell = ws.cell(row=2, column=col, value=month)
                cell.font = normal_font
                cell.fill = header_fill_light
                cell.alignment = center_align
                cell.border = thin_border
                col += 1
        
        ws.cell(row=2, column=col, value="汇总").font = subheader_font
        ws.cell(row=2, column=col).fill = header_fill_light
        ws.cell(row=2, column=col).alignment = center_align
        ws.cell(row=2, column=col).border = thin_border
        col += 1
        ws.cell(row=2, column=col, value="月度合计").font = subheader_font
        ws.cell(row=2, column=col).fill = header_fill_light
        ws.cell(row=2, column=col).alignment = center_align
        ws.cell(row=2, column=col).border = thin_border
        
        return ws, months, thin_border
    
    def add_data(ws, indices, start_row, year, border):
        months = ['1-2 月', '1 月', '2 月', '3 月', '4 月', '5 月', '6 月', '7 月', '8 月', '9 月', '10 月', '11 月', '12 月'] if year == 2026 else ['1 月', '2 月', '3 月', '4 月', '5 月', '6 月', '7 月', '8 月', '9 月', '10 月', '11 月', '12 月']
        cats = [('HR', hr_cols), ('Sample', sample_cols + manufacturing_cols + material_cols), ('Test', test_cols + cert_cols), ('Mold', mold_cols + modify_cols), ('Other', other_cols)]
        
        row = start_row
        use_alt_fill = False
        for idx in indices:
            if customers[idx] == '':
                continue
            row_fill = alt_fill if use_alt_fill else None
            
            for c, val in enumerate([customers[idx], projects[idx], order_ids[idx]], 1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.border = border
                cell.alignment = left_align
                cell.font = normal_font
                if row_fill:
                    cell.fill = row_fill
            
            col = 4
            grand_total = 0.0
            monthly_sum = 0.0
            
            for _, cat_cols in cats:
                yearly, monthly, monthly_avg = get_data(idx, cat_cols, year)
                cell = ws.cell(row=row, column=col, value=yearly if yearly != 0 else None)
                cell.border = border
                cell.alignment = right_align
                cell.font = bold_font if yearly != 0 else normal_font
                cell.number_format = num_format
                if row_fill:
                    cell.fill = row_fill
                col += 1
                
                cell = ws.cell(row=row, column=col, value=monthly_avg if monthly_avg != 0 else None)
                cell.border = border
                cell.alignment = right_align
                cell.font = normal_font
                cell.number_format = num_format
                if row_fill:
                    cell.fill = row_fill
                col += 1
                
                for m in months:
                    v = monthly.get(m)
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.border = border
                    cell.alignment = right_align
                    cell.font = normal_font
                    cell.number_format = num_format
                    if row_fill:
                        cell.fill = row_fill
                    if v is not None:
                        monthly_sum += v
                    col += 1
                
                grand_total += yearly
            
            cell = ws.cell(row=row, column=col, value=round(grand_total, 2))
            cell.font = bold_font
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            cell.alignment = right_align
            cell.border = border
            cell.number_format = num_format
            col += 1
            
            cell = ws.cell(row=row, column=col, value=round(monthly_sum, 2))
            cell.font = bold_font
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.alignment = right_align
            cell.border = border
            cell.number_format = num_format
            
            row += 1
            use_alt_fill = not use_alt_fill
        return row
    
    # 生成 2025 年工作表
    print("生成 2025 年成本分析...")
    ws25, _, border = create_sheet("2025 年成本分析", 2025)
    row = 3
    for cust_type in valid_customers:
        indices = customer_groups[cust_type]
        if indices:
            ws25.cell(row=row-1, column=1, value=f"【{cust_type}客户】").font = Font(name='微软雅黑', bold=True, size=11, color='2E74B5')
            row = add_data(ws25, indices, row, 2025, border)
            row += 1
    
    # 生成 2026 年工作表
    print("生成 2026 年成本分析...")
    ws26, _, border = create_sheet("2026 年成本分析", 2026)
    row = 3
    for cust_type in valid_customers:
        indices = customer_groups[cust_type]
        if indices:
            ws26.cell(row=row-1, column=1, value=f"【{cust_type}客户】").font = Font(name='微软雅黑', bold=True, size=11, color='2E74B5')
            row = add_data(ws26, indices, row, 2026, border)
            row += 1
    
    # 设置列宽
    for ws in [ws25, ws26]:
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        for i in range(6, 72):
            col_letter = get_column_letter(i)
            if col_letter not in ['D', 'E']:
                ws.column_dimensions[col_letter].width = 10
        ws.column_dimensions['BT'].width = 14
    
    # 确定输出路径
    if output_dir is None:
        output_dir = os.path.dirname(source_file)
    
    timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M')
    source_basename = os.path.basename(source_file).replace('.xlsx', '')
    output_file = os.path.join(output_dir, f"{source_basename}_成本分析报告_{timestamp}.xlsx")
    
    wb.save(output_file)
    print(f"✅ 报告已生成：{output_file}")
    
    return output_file

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法：python generate_cost_report.py <源 Excel 文件路径> [输出目录]")
        sys.exit(1)
    
    source_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else None
    
    generate_cost_report(source_file, output_dir)
