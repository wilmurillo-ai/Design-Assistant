#!/usr/bin/env python3
"""
Voice Excel Editor CLI.

职责：
- 调用 SenseAudio ASR 完成语音指令转写
- 对转写文本做规则化规范
- 校验并执行结构化 Excel 操作计划
- 生成修改后工作簿和执行日志

说明：
- 规划本身由上层 Agent 结合 references/ 中的协议文件完成
- 本脚本负责稳定的 I/O、协议校验、坐标解析和 Excel 实际修改
"""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import importlib
import json
import os
import re
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


def ensure_python_package(import_name: str, pip_name: Optional[str] = None) -> None:
    pip_name = pip_name or import_name
    try:
        importlib.import_module(import_name)
        return
    except ImportError:
        print(f"缺少 {pip_name}，正在自动安装...", file=sys.stderr)

    result = subprocess.run(
        [sys.executable, "-m", "pip", "install", pip_name],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"自动安装 {pip_name} 失败，请手动执行: {sys.executable} -m pip install {pip_name}\n"
            f"{result.stderr.strip()}"
        )

    importlib.import_module(import_name)


ensure_python_package("requests")
ensure_python_package("openpyxl")
import requests
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_API_BASE = "https://api.senseaudio.cn"
ASR_API_PATH = "/v1/audio/transcriptions"
DEFAULT_ASR_MODEL = "sense-asr-pro"
MAX_AUDIO_SIZE_MB = 10
DEFAULT_LANGUAGE = "zh"

SKILL_DIR = Path(__file__).resolve().parent.parent
WORKSPACE_DIR = SKILL_DIR.parent
OUTPUT_DIR = WORKSPACE_DIR / "outputs"

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SUPPORTED_ACTIONS = {
    "merge_cells",
    "align_center",
    "bold_font",
    "fill_color",
    "set_border",
    "set_column_width",
    "set_row_height",
    "wrap_text",
    "write_value",
    "clear_cells",
    "copy_range",
    "write_formula",
    "calculate_average",
    "calculate_sum",
    "calculate_max",
    "calculate_min",
    "calculate_count",
    "insert_rows",
    "delete_rows",
    "insert_columns",
    "delete_columns",
    "create_sheet",
    "rename_sheet",
}
CALCULATION_FORMULAS = {
    "calculate_average": "AVERAGE",
    "calculate_sum": "SUM",
    "calculate_max": "MAX",
    "calculate_min": "MIN",
    "calculate_count": "COUNT",
}


class APIError(Exception):
    pass


@dataclass
class TranscriptionResult:
    output_dir: Path
    json_path: Path
    transcript_txt_path: Path
    response_json: Dict[str, Any]


@dataclass
class NormalizeResult:
    output_path: Path
    normalized_text: str


@dataclass
class ExecutionResult:
    output_dir: Path
    workbook_path: Path
    execution_log_path: Path
    feedback_path: Path
    media_ref: Optional[str]
    executed_operations: int
    modified_ranges: List[str]
    log_entries: List[Dict[str, Any]]


class SenseAudioClient:
    def __init__(self, api_key: Optional[str] = None, api_base: Optional[str] = None):
        self.api_key = (api_key or os.getenv("SENSEAUDIO_API_KEY", "")).strip()
        self.api_base = (api_base or os.getenv("SENSEAUDIO_API_BASE", DEFAULT_API_BASE)).rstrip("/")

    @property
    def configured(self) -> bool:
        return bool(self.api_key)

    @property
    def headers_auth_only(self) -> Dict[str, str]:
        if not self.api_key:
            raise APIError(missing_key_message())
        return {"Authorization": f"Bearer {self.api_key}"}

    def auth_check(self) -> Dict[str, Any]:
        if not self.configured:
            raise APIError(missing_key_message())
        return {
            "configured": True,
            "api_base": self.api_base,
            "message": "已检测到 `SENSEAUDIO_API_KEY`。当前 auth-check 只做本地配置检查。",
        }

    def transcribe(
        self,
        *,
        audio_path: Path,
        language: Optional[str],
        output_dir: Optional[Path] = None,
    ) -> TranscriptionResult:
        if not audio_path.exists():
            raise APIError(f"音频文件不存在: {audio_path}")
        if audio_path.stat().st_size > MAX_AUDIO_SIZE_MB * 1024 * 1024:
            raise APIError(
                f"音频文件大于 {MAX_AUDIO_SIZE_MB}MB，当前 Skill 不自动切片，请先拆分后再处理。"
            )

        data: List[Tuple[str, str]] = [
            ("model", DEFAULT_ASR_MODEL),
            ("response_format", "verbose_json"),
            ("enable_itn", "true"),
            ("enable_punctuation", "true"),
        ]
        if language:
            data.append(("language", language))

        with audio_path.open("rb") as audio_file:
            files = {"file": (audio_path.name, audio_file, guess_mime_type(audio_path))}
            response = requests.post(
                f"{self.api_base}{ASR_API_PATH}",
                headers=self.headers_auth_only,
                data=data,
                files=files,
                timeout=600,
            )

        parsed = handle_json_response(response)
        out_dir = prepare_audio_output_dir(audio_path, output_dir)
        json_path = out_dir / "asr_verbose.json"
        transcript_txt_path = out_dir / "transcript_raw.txt"
        json_path.write_text(json.dumps(parsed, ensure_ascii=False, indent=2), encoding="utf-8")
        transcript_txt_path.write_text((parsed.get("text") or "").strip() + "\n", encoding="utf-8")
        return TranscriptionResult(
            output_dir=out_dir,
            json_path=json_path,
            transcript_txt_path=transcript_txt_path,
            response_json=parsed,
        )


def missing_key_message() -> str:
    return (
        "未检测到 `SENSEAUDIO_API_KEY`。\n"
        "请先配置：\n"
        'export SENSEAUDIO_API_KEY="YOUR_API_KEY"\n'
        'export SENSEAUDIO_API_BASE="https://api.senseaudio.cn"'
    )


def guess_mime_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".wav":
        return "audio/wav"
    if suffix == ".mp3":
        return "audio/mpeg"
    if suffix in {".m4a", ".mp4"}:
        return "audio/mp4"
    if suffix == ".aac":
        return "audio/aac"
    return "application/octet-stream"


def handle_json_response(response: requests.Response) -> Dict[str, Any]:
    if response.status_code in (401, 403):
        raise APIError("认证失败：`SENSEAUDIO_API_KEY` 无效、过期或无权限。")
    if response.status_code == 429:
        raise APIError("请求频率过高，接口返回 HTTP 429。")
    if response.status_code >= 400:
        raise APIError(f"HTTP {response.status_code}: {response.text[:500]}")
    try:
        parsed = response.json()
    except ValueError as exc:
        raise APIError(f"接口返回了非 JSON 内容: {response.text[:300]}") from exc
    base_resp = parsed.get("base_resp")
    if isinstance(base_resp, dict) and base_resp.get("status_code", 0) not in (0, None):
        raise APIError(f"{base_resp.get('status_code')}: {base_resp.get('status_msg')}")
    if isinstance(parsed, dict) and parsed.get("code") and parsed.get("message"):
        raise APIError(f"{parsed['code']}: {parsed['message']}")
    return parsed


def prepare_audio_output_dir(audio_path: Path, output_dir: Optional[Path] = None) -> Path:
    ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    base = output_dir or OUTPUT_DIR / f"{audio_path.stem}-{ts}"
    base.mkdir(parents=True, exist_ok=True)
    return base


def prepare_excel_output_dir(excel_path: Path, output_dir: Optional[Path] = None) -> Path:
    ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    base = output_dir or OUTPUT_DIR / f"{excel_path.stem}-{ts}"
    base.mkdir(parents=True, exist_ok=True)
    return base


def build_media_reference(path: Path) -> Optional[str]:
    try:
        relative = path.resolve().relative_to(WORKSPACE_DIR.resolve())
    except ValueError:
        return None
    return f"MEDIA:./{relative.as_posix()}"


def normalize_instruction_text(text: str) -> str:
    normalized = text.strip()
    if not normalized:
        raise APIError("规范化失败：输入文本为空。")
    replacements = [
        (r"[，,]+", "，"),
        (r"[。；;]+", "。"),
        (r"\s+", ""),
        (r"格子", "单元格"),
        (r"单元", "单元格"),
        (r"前(\d+)个格子", r"前\1个单元格"),
        (r"第(\d+)行前(\d+)个单元格", r"第\1行的前\2个单元格"),
        (r"第(\d+)列前(\d+)个数", r"第\1列的前\2个数"),
        (r"保存到", "写入到"),
        (r"放到", "写入到"),
    ]
    for pattern, replacement in replacements:
        normalized = re.sub(pattern, replacement, normalized)
    normalized = replace_chinese_ordinals(normalized)
    normalized = restore_spacing_after_numbers(normalized)
    normalized = normalized.replace("，然后", "，然后").replace("然后", "然后")
    if not normalized.endswith(("。", "！", "？")):
        normalized += "。"
    return normalized


def replace_chinese_ordinals(text: str) -> str:
    number_map = {
        "一": 1,
        "二": 2,
        "三": 3,
        "四": 4,
        "五": 5,
        "六": 6,
        "七": 7,
        "八": 8,
        "九": 9,
        "十": 10,
    }

    def parse_simple_cn_number(token: str) -> Optional[int]:
        if token.isdigit():
            return int(token)
        if token == "十":
            return 10
        if "十" in token:
            parts = token.split("十")
            left = number_map.get(parts[0], 1) if parts[0] else 1
            right = number_map.get(parts[1], 0) if len(parts) > 1 and parts[1] else 0
            return left * 10 + right
        if token in number_map:
            return number_map[token]
        return None

    def repl(match: re.Match[str]) -> str:
        prefix = match.group(1)
        token = match.group(2)
        suffix = match.group(3)
        value = parse_simple_cn_number(token)
        return f"{prefix}{value if value is not None else token}{suffix}"

    return re.sub(r"(第|前)([一二三四五六七八九十\d]+)(行|列|个|项|个单元格|个数)", repl, text)


def restore_spacing_after_numbers(text: str) -> str:
    return text


def load_json_file(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise APIError(f"文件不存在: {path}")
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except ValueError as exc:
        raise APIError(f"JSON 文件格式错误: {path}") from exc
    if not isinstance(data, dict):
        raise APIError("计划文件顶层必须是 JSON 对象。")
    return data


def validate_excel_path(path: Path) -> None:
    if not path.exists():
        raise APIError(f"Excel 文件不存在: {path}")
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise APIError(f"暂不支持该 Excel 格式: {path.suffix}")


def validate_plan(plan: Dict[str, Any]) -> None:
    if "operations" not in plan or not isinstance(plan["operations"], list):
        raise APIError("计划文件缺少 `operations` 数组。")
    ambiguities = plan.get("ambiguities", [])
    if not isinstance(ambiguities, list):
        raise APIError("`ambiguities` 必须是数组。")
    if ambiguities:
        raise APIError(f"计划存在未消解歧义，拒绝执行: {json.dumps(ambiguities, ensure_ascii=False)}")
    sheet_selector = plan.get("sheet_selector") or {"mode": "active"}
    if not isinstance(sheet_selector, dict):
        raise APIError("`sheet_selector` 必须是对象。")
    mode = sheet_selector.get("mode", "active")
    if mode not in {"active", "by_name"}:
        raise APIError("`sheet_selector.mode` 只支持 `active` 或 `by_name`。")
    if mode == "by_name" and not str(sheet_selector.get("sheet_name") or "").strip():
        raise APIError("当 `sheet_selector.mode=by_name` 时必须提供 `sheet_name`。")
    for index, operation in enumerate(plan["operations"], start=1):
        if not isinstance(operation, dict):
            raise APIError(f"第 {index} 个操作不是对象。")
        action = str(operation.get("action") or "").strip()
        if action not in SUPPORTED_ACTIONS:
            raise APIError(f"第 {index} 个操作动作不受支持: {action}")


def select_worksheet(workbook, plan: Dict[str, Any]) -> Worksheet:
    selector = plan.get("sheet_selector") or {"mode": "active"}
    mode = selector.get("mode", "active")
    if mode == "active":
        return workbook.active
    sheet_name = str(selector.get("sheet_name") or "").strip()
    if sheet_name not in workbook.sheetnames:
        raise APIError(f"工作表不存在: {sheet_name}")
    return workbook[sheet_name]


def get_used_range(ws: Worksheet) -> Optional[str]:
    if ws.max_row < 1 or ws.max_column < 1:
        return None
    non_empty = [cell.coordinate for row in ws.iter_rows() for cell in row if cell.value is not None]
    if not non_empty:
        return None
    rows = [ws[cell].row for cell in non_empty]
    cols = [ws[cell].column for cell in non_empty]
    return f"{get_column_letter(min(cols))}{min(rows)}:{get_column_letter(max(cols))}{max(rows)}"


def normalize_range_reference(ws: Worksheet, ref: str) -> str:
    ref = ref.strip()
    if ref == "used_range":
        used = get_used_range(ws)
        if not used:
            raise APIError("当前工作表没有已使用区域，无法执行 `used_range` 操作。")
        return used
    return ref


def iter_cells(ws: Worksheet, ref: str) -> Iterable[Cell]:
    normalized = normalize_range_reference(ws, ref)
    if re.fullmatch(r"[A-Z]+:[A-Z]+", normalized):
        start_col, end_col = normalized.split(":")
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        for col_idx in range(start_idx, end_idx + 1):
            for row_idx in range(1, ws.max_row + 1):
                yield ws.cell(row=row_idx, column=col_idx)
        return
    if re.fullmatch(r"\d+:\d+", normalized):
        start_row, end_row = [int(x) for x in normalized.split(":")]
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                yield ws.cell(row=row_idx, column=col_idx)
        return
    if ":" in normalized:
        min_col, min_row, max_col, max_row = range_boundaries(normalized)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                yield cell
        return
    yield ws[normalized]


def copy_cell_style_and_value(source: Cell, target: Cell) -> None:
    target.value = source.value
    if source.has_style:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)


def apply_operation(workbook, ws: Worksheet, operation: Dict[str, Any]) -> Dict[str, Any]:
    action = operation["action"]
    detail: Dict[str, Any] = {"action": action, "status": "success", "modified_ranges": []}

    if action == "merge_cells":
        ref = normalize_range_reference(ws, require_str(operation, "range"))
        ws.merge_cells(ref)
        detail["modified_ranges"] = [ref]
        return detail

    if action == "align_center":
        ref = normalize_range_reference(ws, require_str(operation, "range"))
        vertical_enabled = bool(operation.get("vertical", True))
        for cell in iter_cells(ws, ref):
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center" if vertical_enabled else cell.alignment.vertical,
                wrap_text=cell.alignment.wrap_text,
            )
        detail["modified_ranges"] = [ref]
        return detail

    if action == "bold_font":
        ref = normalize_range_reference(ws, require_str(operation, "range"))
        enabled = bool(operation.get("enabled", True))
        for cell in iter_cells(ws, ref):
            cell.font = copy.copy(cell.font)
            cell.font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=enabled,
                italic=cell.font.italic,
                vertAlign=cell.font.vertAlign,
                underline=cell.font.underline,
                strike=cell.font.strike,
                color=cell.font.color,
                charset=cell.font.charset,
                outline=cell.font.outline,
                shadow=cell.font.shadow,
                condense=cell.font.condense,
                extend=cell.font.extend,
                family=cell.font.family,
                scheme=cell.font.scheme,
            )
        detail["modified_ranges"] = [ref]
        return detail

    if action == "fill_color":
        ref = normalize_range_reference(ws, require_str(operation, "range"))
        color = require_str(operation, "color")
        for cell in iter_cells(ws, ref):
            cell.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
        detail["modified_ranges"] = [ref]
        return detail

    if action == "set_border":
        ref = normalize_range_reference(ws, require_str(operation, "range"))
        style_raw = require_str(operation, "style").lower()
        style = "thick" if style_raw == "bold" else style_raw
        if style not in {"thin", "medium", "thick"}:
            raise APIError(f"不支持的边框样式: {style_raw}")
        side = Side(border_style=style, color="000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        for cell in iter_cells(ws, ref):
            cell.border = border
        detail["modified_ranges"] = [ref]
        return detail

    if action == "set_column_width":
        ref = normalize_range_reference(ws, require_str(operation, "range"))
        width = float(operation["width"])
        start_col, end_col = ref.split(":")
        for col_idx in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        detail["modified_ranges"] = [ref]
        return detail

    if action == "set_row_height":
        ref = normalize_range_reference(ws, require_str(operation, "range"))
        height = float(operation["height"])
        start_row, end_row = [int(x) for x in ref.split(":")]
        for row_idx in range(start_row, end_row + 1):
            ws.row_dimensions[row_idx].height = height
        detail["modified_ranges"] = [ref]
        return detail

    if action == "wrap_text":
        ref = normalize_range_reference(ws, require_str(operation, "range"))
        enabled = bool(operation.get("enabled", True))
        for cell in iter_cells(ws, ref):
            cell.alignment = copy.copy(cell.alignment).copy(wrap_text=enabled)
        detail["modified_ranges"] = [ref]
        return detail

    if action == "write_value":
        target = require_str(operation, "target_cell")
        value = operation.get("value")
        ws[target].value = value
        detail["modified_ranges"] = [target]
        return detail

    if action == "clear_cells":
        ref = normalize_range_reference(ws, require_str(operation, "range"))
        for cell in iter_cells(ws, ref):
            cell.value = None
        detail["modified_ranges"] = [ref]
        return detail

    if action == "copy_range":
        source_ref = normalize_range_reference(ws, require_str(operation, "source_range"))
        target_cell = ws[require_str(operation, "target_cell")]
        min_col, min_row, max_col, max_row = range_boundaries(source_ref)
        start_target_row = target_cell.row
        start_target_col = target_cell.column
        for row_offset, row_idx in enumerate(range(min_row, max_row + 1)):
            for col_offset, col_idx in enumerate(range(min_col, max_col + 1)):
                source = ws.cell(row=row_idx, column=col_idx)
                target = ws.cell(row=start_target_row + row_offset, column=start_target_col + col_offset)
                copy_cell_style_and_value(source, target)
        end_target = ws.cell(
            row=start_target_row + (max_row - min_row),
            column=start_target_col + (max_col - min_col),
        )
        detail["modified_ranges"] = [source_ref, f"{target_cell.coordinate}:{end_target.coordinate}"]
        return detail

    if action == "write_formula":
        target = require_str(operation, "target_cell")
        formula = require_str(operation, "formula")
        ws[target].value = formula
        detail["modified_ranges"] = [target]
        return detail

    if action in CALCULATION_FORMULAS:
        source_ref = normalize_range_reference(ws, require_str(operation, "source_range"))
        target = require_str(operation, "target_cell")
        formula_name = CALCULATION_FORMULAS[action]
        ws[target].value = f"={formula_name}({source_ref})"
        detail["modified_ranges"] = [source_ref, target]
        return detail

    if action == "insert_rows":
        row = int(operation["row"])
        amount = int(operation.get("amount", 1))
        ws.insert_rows(row, amount)
        detail["modified_ranges"] = [f"{row}:{row + amount - 1}"]
        return detail

    if action == "delete_rows":
        row = int(operation["row"])
        amount = int(operation.get("amount", 1))
        ws.delete_rows(row, amount)
        detail["modified_ranges"] = [f"{row}:{row + amount - 1}"]
        return detail

    if action == "insert_columns":
        column = int(operation["column"])
        amount = int(operation.get("amount", 1))
        ws.insert_cols(column, amount)
        end_column = get_column_letter(column + amount - 1)
        detail["modified_ranges"] = [f"{get_column_letter(column)}:{end_column}"]
        return detail

    if action == "delete_columns":
        column = int(operation["column"])
        amount = int(operation.get("amount", 1))
        ws.delete_cols(column, amount)
        end_column = get_column_letter(column + amount - 1)
        detail["modified_ranges"] = [f"{get_column_letter(column)}:{end_column}"]
        return detail

    if action == "create_sheet":
        title = require_str(operation, "title")
        workbook.create_sheet(title=title)
        detail["modified_ranges"] = [title]
        return detail

    if action == "rename_sheet":
        new_title = require_str(operation, "new_title")
        old_title = ws.title
        ws.title = new_title
        detail["modified_ranges"] = [f"{old_title}->{new_title}"]
        return detail

    raise APIError(f"未实现的动作: {action}")


def require_str(operation: Dict[str, Any], key: str) -> str:
    value = str(operation.get(key) or "").strip()
    if not value:
        raise APIError(f"动作 `{operation.get('action')}` 缺少字段 `{key}`。")
    return value


def build_feedback(log_entries: Sequence[Dict[str, Any]]) -> str:
    success_entries = [entry for entry in log_entries if entry.get("status") == "success"]
    failed_entries = [entry for entry in log_entries if entry.get("status") == "failed"]
    lines = [f"已完成 {len(success_entries)} 项操作。"]
    for entry in success_entries:
        ranges = ", ".join(entry.get("modified_ranges") or []) or "无"
        lines.append(f"- 已执行 `{entry['action']}`，影响区域：{ranges}")
    if failed_entries:
        lines.append(f"失败 {len(failed_entries)} 项。")
        for entry in failed_entries:
            lines.append(f"- `{entry['action']}` 失败：{entry.get('error', '未知错误')}")
    return "\n".join(lines) + "\n"


def execute_plan(
    *,
    excel_path: Path,
    plan_path: Path,
    output_dir: Optional[Path] = None,
) -> ExecutionResult:
    validate_excel_path(excel_path)
    plan = load_json_file(plan_path)
    validate_plan(plan)

    workbook = load_workbook(excel_path)
    ws = select_worksheet(workbook, plan)
    out_dir = prepare_excel_output_dir(excel_path, output_dir)
    execution_log_path = out_dir / "execution_log.json"
    feedback_path = out_dir / "user_feedback.txt"
    workbook_path = out_dir / "modified.xlsx"

    log_entries: List[Dict[str, Any]] = []
    modified_ranges: List[str] = []
    for operation in plan["operations"]:
        try:
            entry = apply_operation(workbook, ws, operation)
        except Exception as exc:
            entry = {
                "action": operation.get("action"),
                "status": "failed",
                "error": str(exc),
                "operation": operation,
                "modified_ranges": [],
            }
            log_entries.append(entry)
            execution_log_path.write_text(
                json.dumps({"status": "failed", "log_entries": log_entries}, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            raise APIError(f"执行在 `{operation.get('action')}` 处失败: {exc}") from exc
        entry["operation"] = operation
        log_entries.append(entry)
        modified_ranges.extend(entry.get("modified_ranges") or [])

    workbook.save(workbook_path)
    feedback = build_feedback(log_entries)
    feedback_path.write_text(feedback, encoding="utf-8")
    execution_log_path.write_text(
        json.dumps(
            {
                "status": "success",
                "executed_operations": len(log_entries),
                "modified_ranges": modified_ranges,
                "log_entries": log_entries,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return ExecutionResult(
        output_dir=out_dir,
        workbook_path=workbook_path,
        execution_log_path=execution_log_path,
        feedback_path=feedback_path,
        media_ref=build_media_reference(workbook_path),
        executed_operations=len(log_entries),
        modified_ranges=modified_ranges,
        log_entries=log_entries,
    )


def command_auth_check(args: argparse.Namespace) -> int:
    client = SenseAudioClient()
    print(json.dumps(client.auth_check(), ensure_ascii=False, indent=2))
    return 0


def command_transcribe(args: argparse.Namespace) -> int:
    client = SenseAudioClient()
    result = client.transcribe(
        audio_path=Path(args.audio).expanduser().resolve(),
        language=args.language,
        output_dir=Path(args.output_dir).expanduser().resolve() if args.output_dir else None,
    )
    print(json.dumps(
        {
            "output_dir": str(result.output_dir),
            "json_path": str(result.json_path),
            "transcript_raw_path": str(result.transcript_txt_path),
        },
        ensure_ascii=False,
        indent=2,
    ))
    return 0


def command_normalize(args: argparse.Namespace) -> int:
    if args.text:
        raw_text = args.text
    elif args.text_file:
        raw_text = Path(args.text_file).expanduser().resolve().read_text(encoding="utf-8")
    else:
        raise APIError("请通过 `--text` 或 `--text-file` 提供待规范化文本。")
    normalized = normalize_instruction_text(raw_text)
    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    elif args.text_file:
        output_path = Path(args.text_file).expanduser().resolve().with_name("transcript_normalized.txt")
    else:
        out_dir = OUTPUT_DIR / f"normalized-{dt.datetime.now().strftime('%Y%m%d-%H%M%S')}"
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / "transcript_normalized.txt"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(normalized + "\n", encoding="utf-8")
    print(json.dumps({"normalized_path": str(output_path), "text": normalized}, ensure_ascii=False, indent=2))
    return 0


def command_execute(args: argparse.Namespace) -> int:
    result = execute_plan(
        excel_path=Path(args.excel).expanduser().resolve(),
        plan_path=Path(args.plan_file).expanduser().resolve(),
        output_dir=Path(args.output_dir).expanduser().resolve() if args.output_dir else None,
    )
    print(json.dumps(
        {
            "status": "success",
            "output_dir": str(result.output_dir),
            "modified_workbook": str(result.workbook_path),
            "execution_log": str(result.execution_log_path),
            "user_feedback": str(result.feedback_path),
            "executed_operations": result.executed_operations,
            "modified_ranges": result.modified_ranges,
        },
        ensure_ascii=False,
        indent=2,
    ))
    if result.media_ref:
        print(result.media_ref)
    return 0


def command_run(args: argparse.Namespace) -> int:
    client = SenseAudioClient()
    transcription = client.transcribe(
        audio_path=Path(args.audio).expanduser().resolve(),
        language=args.language,
        output_dir=Path(args.output_dir).expanduser().resolve() if args.output_dir else None,
    )
    normalized_text = normalize_instruction_text(transcription.transcript_txt_path.read_text(encoding="utf-8"))
    normalized_path = transcription.output_dir / "transcript_normalized.txt"
    normalized_path.write_text(normalized_text + "\n", encoding="utf-8")

    payload: Dict[str, Any] = {
        "output_dir": str(transcription.output_dir),
        "asr_verbose_json": str(transcription.json_path),
        "transcript_raw_path": str(transcription.transcript_txt_path),
        "transcript_normalized_path": str(normalized_path),
    }

    if args.plan_file:
        result = execute_plan(
            excel_path=Path(args.excel).expanduser().resolve(),
            plan_path=Path(args.plan_file).expanduser().resolve(),
            output_dir=transcription.output_dir,
        )
        payload.update(
            {
                "status": "success",
                "modified_workbook": str(result.workbook_path),
                "execution_log": str(result.execution_log_path),
                "user_feedback": str(result.feedback_path),
                "media_ref": result.media_ref,
                "executed_operations": result.executed_operations,
                "modified_ranges": result.modified_ranges,
            }
        )
    else:
        payload.update(
            {
                "status": "needs_plan",
                "message": (
                    "已完成转写与规范化。请读取 references/planning_prompt.md 和 "
                    "references/operation_schema.md 生成 operation_plan.json，再运行 execute。"
                ),
            }
        )

    print(json.dumps(payload, ensure_ascii=False, indent=2))
    if args.plan_file and result.media_ref:
        print(result.media_ref)
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Voice Excel Editor CLI")
    subparsers = parser.add_subparsers(dest="command", required=True)

    auth_check_parser = subparsers.add_parser("auth-check", help="检查 SenseAudio 配置")
    auth_check_parser.set_defaults(func=command_auth_check)

    transcribe_parser = subparsers.add_parser("transcribe", help="转写语音指令")
    transcribe_parser.add_argument("--audio", required=True, help="语音文件路径")
    transcribe_parser.add_argument("--language", default=DEFAULT_LANGUAGE, help="识别语言，默认 zh")
    transcribe_parser.add_argument("--output-dir", help="输出目录")
    transcribe_parser.set_defaults(func=command_transcribe)

    normalize_parser = subparsers.add_parser("normalize", help="规范化转写文本")
    normalize_parser.add_argument("--text", help="直接输入文本")
    normalize_parser.add_argument("--text-file", help="文本文件路径")
    normalize_parser.add_argument("--output", help="规范化文本输出路径")
    normalize_parser.set_defaults(func=command_normalize)

    execute_parser = subparsers.add_parser("execute", help="执行 Excel 操作计划")
    execute_parser.add_argument("--excel", required=True, help="Excel 文件路径")
    execute_parser.add_argument("--plan-file", required=True, help="结构化计划 JSON 路径")
    execute_parser.add_argument("--output-dir", help="输出目录")
    execute_parser.set_defaults(func=command_execute)

    run_parser = subparsers.add_parser("run", help="串联转写、规范化与执行")
    run_parser.add_argument("--excel", required=True, help="Excel 文件路径")
    run_parser.add_argument("--audio", required=True, help="语音文件路径")
    run_parser.add_argument("--language", default=DEFAULT_LANGUAGE, help="识别语言，默认 zh")
    run_parser.add_argument("--plan-file", help="结构化计划 JSON 路径")
    run_parser.add_argument("--output-dir", help="输出目录")
    run_parser.set_defaults(func=command_run)

    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return int(args.func(args))
    except APIError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    except KeyboardInterrupt:
        print("ERROR: 用户中断执行。", file=sys.stderr)
        return 130


if __name__ == "__main__":
    raise SystemExit(main())
