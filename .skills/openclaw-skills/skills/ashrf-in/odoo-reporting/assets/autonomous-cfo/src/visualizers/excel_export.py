"""
Excel export for financial reports.
Generates professional Excel workbooks with formatting.
"""
from typing import Dict, Any, List, Optional
from io import BytesIO
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Export financial reports to Excel.
    
    Usage:
        exporter = ExcelExporter()
        excel_bytes = exporter.export_balance_sheet(data)
        
        # Save to file
        with open("balance_sheet.xlsx", "wb") as f:
            f.write(excel_bytes)
    """
    
    # Styles
    HEADER_FONT = Font(bold=True, size=14, color="1A2E3A")
    SUBHEADER_FONT = Font(bold=True, size=11, color="333333")
    AMOUNT_FONT = Font(size=10)
    TOTAL_FONT = Font(bold=True, size=11)
    
    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    TOTAL_BORDER = Border(
        top=Side(style='double', color='CD7F32'),
        bottom=Side(style='thin', color='CD7F32')
    )
    
    HEADER_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="FFFAF0", end_color="FFFAF0", fill_type="solid")
    
    def __init__(self):
        if Workbook is None:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    def export_balance_sheet(self, data: Dict[str, Any]) -> bytes:
        """
        Export balance sheet to Excel.
        
        Args:
            data: Balance sheet data dict
        
        Returns:
            Excel file bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        row = 1
        
        # Title
        ws.cell(row=row, column=1, value="Statement of Financial Position")
        ws.cell(row=row, column=1).font = self.HEADER_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        # Company and date
        ws.cell(row=row, column=1, value=data.get("company_name", ""))
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        row += 1
        
        ws.cell(row=row, column=1, value=f"As at {data.get('as_of_date', '')}")
        row += 2
        
        # Assets section
        ws.cell(row=row, column=1, value="ASSETS")
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        row += 1
        
        # Non-current assets
        if data.get("assets", {}).get("non_current"):
            ws.cell(row=row, column=1, value="Non-Current Assets")
            ws.cell(row=row, column=1).font = Font(bold=True, size=10)
            row += 1
            
            for name, amount in data["assets"]["non_current"].items():
                ws.cell(row=row, column=1, value=f"  {name}")
                ws.cell(row=row, column=3, value=amount)
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                row += 1
            
            nc_total = sum(data["assets"]["non_current"].values())
            ws.cell(row=row, column=1, value="Total Non-Current Assets")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=3, value=nc_total)
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=3).font = Font(bold=True)
            row += 2
        
        # Current assets
        if data.get("assets", {}).get("current"):
            ws.cell(row=row, column=1, value="Current Assets")
            ws.cell(row=row, column=1).font = Font(bold=True, size=10)
            row += 1
            
            for name, amount in data["assets"]["current"].items():
                ws.cell(row=row, column=1, value=f"  {name}")
                ws.cell(row=row, column=3, value=amount)
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                row += 1
            
            c_total = sum(data["assets"]["current"].values())
            ws.cell(row=row, column=1, value="Total Current Assets")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=3, value=c_total)
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=3).font = Font(bold=True)
            row += 2
        
        # Total assets
        total_assets = data.get("assets", {}).get("total", 0)
        ws.cell(row=row, column=1, value="TOTAL ASSETS")
        ws.cell(row=row, column=1).font = self.TOTAL_FONT
        ws.cell(row=row, column=3, value=total_assets)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=3).font = self.TOTAL_FONT
        ws.cell(row=row, column=3).fill = self.TOTAL_FILL
        row += 3
        
        # Liabilities section
        ws.cell(row=row, column=1, value="LIABILITIES AND EQUITY")
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        row += 1
        
        # Equity
        if data.get("equity"):
            ws.cell(row=row, column=1, value="Equity")
            ws.cell(row=row, column=1).font = Font(bold=True, size=10)
            row += 1
            
            for name, amount in data["equity"].items():
                if isinstance(amount, (int, float)):
                    ws.cell(row=row, column=1, value=f"  {name}")
                    ws.cell(row=row, column=3, value=amount)
                    ws.cell(row=row, column=3).number_format = '#,##0.00'
                    row += 1
            row += 1
        
        # Liabilities
        if data.get("liabilities"):
            ws.cell(row=row, column=1, value="Liabilities")
            ws.cell(row=row, column=1).font = Font(bold=True, size=10)
            row += 1
            
            if data["liabilities"].get("non_current"):
                for name, amount in data["liabilities"]["non_current"].items():
                    ws.cell(row=row, column=1, value=f"  {name}")
                    ws.cell(row=row, column=3, value=amount)
                    ws.cell(row=row, column=3).number_format = '#,##0.00'
                    row += 1
            
            if data["liabilities"].get("current"):
                for name, amount in data["liabilities"]["current"].items():
                    ws.cell(row=row, column=1, value=f"  {name}")
                    ws.cell(row=row, column=3, value=amount)
                    ws.cell(row=row, column=3).number_format = '#,##0.00'
                    row += 1
            row += 1
        
        # Total liabilities + equity
        total_le = data.get("liabilities", {}).get("total", 0) + data.get("equity", {}).get("total", 0)
        ws.cell(row=row, column=1, value="TOTAL LIABILITIES AND EQUITY")
        ws.cell(row=row, column=1).font = self.TOTAL_FONT
        ws.cell(row=row, column=3, value=total_le)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=3).font = self.TOTAL_FONT
        ws.cell(row=row, column=3).fill = self.TOTAL_FILL
        
        # Column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 15
        
        # Footer
        row += 3
        ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=row, column=1).font = Font(italic=True, size=8, color="888888")
        
        # Export
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_profit_loss(self, data: Dict[str, Any]) -> bytes:
        """
        Export P&L statement to Excel.
        
        Args:
            data: P&L data dict
        
        Returns:
            Excel file bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit & Loss"
        
        row = 1
        
        # Title
        ws.cell(row=row, column=1, value="Statement of Profit or Loss")
        ws.cell(row=row, column=1).font = self.HEADER_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        # Company and period
        ws.cell(row=row, column=1, value=data.get("company_name", ""))
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        row += 1
        
        ws.cell(row=row, column=1, value=f"For the period {data.get('period', '')}")
        row += 2
        
        # Revenue
        ws.cell(row=row, column=1, value="REVENUE")
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        row += 1
        
        if data.get("revenue"):
            for name, amount in data["revenue"].items():
                ws.cell(row=row, column=1, value=f"  {name}")
                ws.cell(row=row, column=3, value=amount)
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                row += 1
        
        total_revenue = data.get("totals", {}).get("revenue", 0)
        ws.cell(row=row, column=1, value="Total Revenue")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_revenue)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=3).font = Font(bold=True)
        row += 2
        
        # Expenses
        ws.cell(row=row, column=1, value="EXPENSES")
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        row += 1
        
        if data.get("expenses"):
            for name, amount in data["expenses"].items():
                ws.cell(row=row, column=1, value=f"  {name}")
                ws.cell(row=row, column=3, value=-abs(amount))  # Show as negative
                ws.cell(row=row, column=3).number_format = '(#,##0.00)'
                row += 1
        
        total_expenses = data.get("totals", {}).get("total_expenses", 0)
        ws.cell(row=row, column=1, value="Total Expenses")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=-abs(total_expenses))
        ws.cell(row=row, column=3).number_format = '(#,##0.00)'
        ws.cell(row=row, column=3).font = Font(bold=True)
        row += 2
        
        # Net profit
        net_profit = data.get("totals", {}).get("net_profit", 0) or data.get("totals", {}).get("profit_for_period", 0)
        ws.cell(row=row, column=1, value="NET PROFIT")
        ws.cell(row=row, column=1).font = self.TOTAL_FONT
        ws.cell(row=row, column=3, value=net_profit)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=3).font = self.TOTAL_FONT
        ws.cell(row=row, column=3).fill = self.TOTAL_FILL
        
        # Column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 15
        
        # Footer
        row += 3
        ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=row, column=1).font = Font(italic=True, size=8, color="888888")
        
        # Export
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_cash_flow(self, data: Dict[str, Any]) -> bytes:
        """Export cash flow statement to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow"
        
        row = 1
        
        # Title
        ws.cell(row=row, column=1, value="Statement of Cash Flows")
        ws.cell(row=row, column=1).font = self.HEADER_FONT
        row += 1
        
        ws.cell(row=row, column=1, value=data.get("company_name", ""))
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        row += 1
        
        ws.cell(row=row, column=1, value=f"For the period {data.get('period', '')}")
        row += 2
        
        # Operating activities
        ws.cell(row=row, column=1, value="Cash Flows from Operating Activities")
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        row += 1
        
        if data.get("operating"):
            for name, amount in data["operating"].items():
                ws.cell(row=row, column=1, value=f"  {name}")
                ws.cell(row=row, column=3, value=amount)
                ws.cell(row=row, column=3).number_format = '#,##0.00;(#,##0.00)'
                row += 1
            
            ws.cell(row=row, column=1, value="Net Cash from Operating Activities")
            ws.cell(row=row, column=3, value=data["operating"].get("net", 0))
            ws.cell(row=row, column=3).number_format = '#,##0.00;(#,##0.00)'
            ws.cell(row=row, column=3).font = Font(bold=True)
            row += 2
        
        # Summary
        if data.get("summary"):
            ws.cell(row=row, column=1, value=f"Net Change in Cash: {data['summary'].get('net_change', 0):,.2f}")
            row += 1
            ws.cell(row=row, column=1, value=f"Opening Cash: {data['summary'].get('opening', 0):,.2f}")
            row += 1
            ws.cell(row=row, column=1, value=f"Closing Cash: {data['summary'].get('closing', 0):,.2f}")
        
        # Column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 15
        
        # Export
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_aging(self, data: Dict[str, Any]) -> bytes:
        """Export AR/AP aging report to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Aging Report"
        
        row = 1
        
        # Title
        ws.cell(row=row, column=1, value="AR/AP Aging Report")
        ws.cell(row=row, column=1).font = self.HEADER_FONT
        row += 1
        
        ws.cell(row=row, column=1, value=f"As at {data.get('as_of_date', '')}")
        row += 2
        
        # Headers
        headers = ["Partner", "Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = self.HEADER_FILL
        row += 1
        
        # Data
        if data.get("partners"):
            for partner in data["partners"]:
                ws.cell(row=row, column=1, value=partner.get("name", ""))
                ws.cell(row=row, column=2, value=partner.get("current", 0))
                ws.cell(row=row, column=3, value=partner.get("days_30", 0))
                ws.cell(row=row, column=4, value=partner.get("days_60", 0))
                ws.cell(row=row, column=5, value=partner.get("days_90", 0))
                ws.cell(row=row, column=6, value=partner.get("over_90", 0))
                ws.cell(row=row, column=7, value=partner.get("total", 0))
                
                for col in range(2, 8):
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
                row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Export
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
