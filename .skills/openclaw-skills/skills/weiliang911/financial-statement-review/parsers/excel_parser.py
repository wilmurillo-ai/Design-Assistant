"""
Excel文件解析器

支持解析 .xlsx 和 .xls 格式的财务文件
"""
import re
from typing import Dict, List, Any, Optional
import logging

from .base_parser import BaseParser, ParseResult, FinancialData, normalize_account_name, extract_number
from .file_identifier import FileIdentifier

# 配置日志
logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """
    Excel文件解析器
    
    支持解析：
    - 资产负债表
    - 利润表
    - 现金流量表
    - 科目余额表
    """
    
    name = "excel_parser"
    description = "Excel文件解析器 (.xlsx/.xls)"
    supported_extensions = ['.xlsx', '.xls', '.xlsm']
    supported_software = ['GENERIC', 'KINGDEE', 'UFIDA', 'SAP']
    
    # 常见工作表名称映射
    SHEET_NAME_PATTERNS = {
        'balance_sheet': [
            '资产负债表', 'balance sheet', 'balance_sheet',
            '資產負債表', '资产', '负债'
        ],
        'income_statement': [
            '利润表', '损益表', 'income statement', 'income_statement',
            '損益表', '利润', '损益'
        ],
        'cash_flow': [
            '现金流量表', 'cash flow', 'cash_flow',
            '現金流量表', '现金流'
        ],
        'trial_balance': [
            '科目余额表', '试算平衡表', 'trial balance', '科目汇总表',
            'accvouch', 'ledger'
        ],
    }
    
    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)
        self.header_row = self.config.get('header_row', 0)
        self.data_start_row = self.config.get('data_start_row', 1)
    
    def can_parse(self, file_path: str) -> bool:
        """检查是否能解析该文件"""
        if not self.validate_file_exists(file_path):
            return False
        
        ext = self.get_file_extension(file_path)
        return ext in self.supported_extensions
    
    def parse(self, file_path: str) -> ParseResult:
        """解析Excel文件"""
        if not self.can_parse(file_path):
            return self.create_error_result(file_path, "无法解析该文件格式")
        
        try:
            # 尝试导入 openpyxl
            try:
                import openpyxl
            except ImportError:
                return self.create_error_result(
                    file_path, 
                    "缺少依赖库：请安装 openpyxl (pip install openpyxl)"
                )
            
            # 加载工作簿
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            result = self.create_success_result(file_path)
            result.file_type = 'EXCEL'
            result.data = FinancialData()
            result.data.source_file = file_path
            result.data.source_software = 'EXCEL'
            
            # 识别并解析各个报表
            sheet_names = wb.sheetnames
            
            # 尝试识别标准财务报表
            bs_sheet = self._find_sheet(sheet_names, self.SHEET_NAME_PATTERNS['balance_sheet'])
            is_sheet = self._find_sheet(sheet_names, self.SHEET_NAME_PATTERNS['income_statement'])
            cf_sheet = self._find_sheet(sheet_names, self.SHEET_NAME_PATTERNS['cash_flow'])
            
            # 解析资产负债表
            if bs_sheet:
                result.data.balance_sheet = self._parse_balance_sheet(wb[bs_sheet])
                result.add_warning(f"找到并解析了资产负债表: {bs_sheet}")
            
            # 解析利润表
            if is_sheet:
                result.data.income_statement = self._parse_income_statement(wb[is_sheet])
                result.add_warning(f"找到并解析了利润表: {is_sheet}")
            
            # 解析现金流量表
            if cf_sheet:
                result.data.cash_flow = self._parse_cash_flow(wb[cf_sheet])
                result.add_warning(f"找到并解析了现金流量表: {cf_sheet}")
            
            # 如果没有找到标准报表，尝试智能识别
            if not bs_sheet and not is_sheet and not cf_sheet:
                self._auto_detect_and_parse(wb, result)
            
            wb.close()
            
            # 提取公司信息
            self._extract_company_info(result.data)
            
            return result
            
        except Exception as e:
            logger.error(f"解析Excel文件失败: {e}")
            return self.create_error_result(file_path, f"解析失败: {str(e)}")
    
    def _find_sheet(self, sheet_names: List[str], patterns: List[str]) -> Optional[str]:
        """根据模式查找工作表"""
        for name in sheet_names:
            name_lower = name.lower()
            for pattern in patterns:
                if pattern.lower() in name_lower:
                    return name
        return None
    
    def _parse_balance_sheet(self, sheet) -> Dict[str, float]:
        """解析资产负债表"""
        data = {}
        
        for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
            if not row or len(row) < 2:
                continue
            
            # 第一列通常是科目名称
            account_name = str(row[0]) if row[0] else ""
            if not account_name:
                continue
            
            # 标准化科目名称
            standard_name = normalize_account_name(account_name)
            
            # 查找数值（通常是期末余额，在后面的列）
            for cell_value in reversed(row[1:]):
                value = extract_number(cell_value)
                if value is not None:
                    data[standard_name] = value
                    break
        
        return data
    
    def _parse_income_statement(self, sheet) -> Dict[str, float]:
        """解析利润表"""
        data = {}
        
        for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
            if not row or len(row) < 2:
                continue
            
            account_name = str(row[0]) if row[0] else ""
            if not account_name:
                continue
            
            standard_name = normalize_account_name(account_name)
            
            # 查找数值（通常是本期金额）
            for cell_value in row[1:]:
                value = extract_number(cell_value)
                if value is not None:
                    data[standard_name] = value
                    break
        
        return data
    
    def _parse_cash_flow(self, sheet) -> Dict[str, float]:
        """解析现金流量表"""
        data = {}
        
        for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
            if not row or len(row) < 2:
                continue
            
            account_name = str(row[0]) if row[0] else ""
            if not account_name:
                continue
            
            standard_name = normalize_account_name(account_name)
            
            for cell_value in row[1:]:
                value = extract_number(cell_value)
                if value is not None:
                    data[standard_name] = value
                    break
        
        return data
    
    def _auto_detect_and_parse(self, workbook, result: ParseResult):
        """自动检测并解析报表"""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 读取前几行进行内容分析
            sample_data = []
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
                sample_data.append(row)
            
            # 检测报表类型
            sheet_type = self._detect_sheet_type(sample_data)
            
            if sheet_type == 'balance_sheet':
                result.data.balance_sheet = self._parse_balance_sheet(sheet)
                result.add_warning(f"自动识别为资产负债表: {sheet_name}")
            elif sheet_type == 'income_statement':
                result.data.income_statement = self._parse_income_statement(sheet)
                result.add_warning(f"自动识别为利润表: {sheet_name}")
            elif sheet_type == 'cash_flow':
                result.data.cash_flow = self._parse_cash_flow(sheet)
                result.add_warning(f"自动识别为现金流量表: {sheet_name}")
    
    def _detect_sheet_type(self, sample_data: List[tuple]) -> str:
        """根据样本数据检测工作表类型"""
        text_content = ' '.join([
            str(cell) for row in sample_data for cell in row if cell
        ]).lower()
        
        # 关键字匹配
        bs_keywords = ['资产', '负债', '流动资产', '流动负债', '所有者权益', 'asset', 'liability']
        is_keywords = ['收入', '成本', '利润', '营业', 'revenue', 'profit', 'income']
        cf_keywords = ['现金', '流量', '经营', '投资', '筹资', 'cash', 'flow']
        
        bs_score = sum(1 for kw in bs_keywords if kw in text_content)
        is_score = sum(1 for kw in is_keywords if kw in text_content)
        cf_score = sum(1 for kw in cf_keywords if kw in text_content)
        
        scores = {
            'balance_sheet': bs_score,
            'income_statement': is_score,
            'cash_flow': cf_score
        }
        
        detected_type = max(scores, key=scores.get)
        
        # 如果最高分小于2，可能不是财务报表
        if scores[detected_type] < 2:
            return 'unknown'
        
        return detected_type
    
    def _extract_company_info(self, data: FinancialData):
        """从解析的数据中提取公司信息"""
        # 可以从文件名或报表标题中提取
        # 这里可以根据实际情况扩展
        pass


# 兼容 .xls 格式的解析器（使用 xlrd）
class XLSParser(ExcelParser):
    """XLS格式解析器（旧版Excel）"""
    
    name = "xls_parser"
    description = "XLS文件解析器（旧版Excel）"
    supported_extensions = ['.xls']
    
    def parse(self, file_path: str) -> ParseResult:
        """解析XLS文件"""
        try:
            try:
                import xlrd
            except ImportError:
                return self.create_error_result(
                    file_path,
                    "缺少依赖库：请安装 xlrd (pip install xlrd)"
                )
            
            workbook = xlrd.open_workbook(file_path)
            
            result = self.create_success_result(file_path)
            result.file_type = 'XLS'
            result.data = FinancialData()
            result.data.source_file = file_path
            result.data.source_software = 'EXCEL'
            
            # 解析逻辑与Excel类似，使用xlrd API
            sheet_names = workbook.sheet_names()
            
            # 识别标准财务报表
            bs_sheet = self._find_sheet(sheet_names, self.SHEET_NAME_PATTERNS['balance_sheet'])
            is_sheet = self._find_sheet(sheet_names, self.SHEET_NAME_PATTERNS['income_statement'])
            cf_sheet = self._find_sheet(sheet_names, self.SHEET_NAME_PATTERNS['cash_flow'])
            
            if bs_sheet:
                sheet_idx = sheet_names.index(bs_sheet)
                result.data.balance_sheet = self._parse_sheet_xlrd(workbook.sheet_by_index(sheet_idx))
                result.add_warning(f"找到并解析了资产负债表: {bs_sheet}")
            
            if is_sheet:
                sheet_idx = sheet_names.index(is_sheet)
                result.data.income_statement = self._parse_sheet_xlrd(workbook.sheet_by_index(sheet_idx))
                result.add_warning(f"找到并解析了利润表: {is_sheet}")
            
            if cf_sheet:
                sheet_idx = sheet_names.index(cf_sheet)
                result.data.cash_flow = self._parse_sheet_xlrd(workbook.sheet_by_index(sheet_idx))
                result.add_warning(f"找到并解析了现金流量表: {cf_sheet}")
            
            return result
            
        except Exception as e:
            return self.create_error_result(file_path, f"解析失败: {str(e)}")
    
    def _parse_sheet_xlrd(self, sheet) -> Dict[str, float]:
        """使用xlrd解析工作表"""
        data = {}
        
        for row_idx in range(min(sheet.nrows, 100)):
            row = sheet.row_values(row_idx)
            
            if not row or len(row) < 2:
                continue
            
            account_name = str(row[0]) if row[0] else ""
            if not account_name:
                continue
            
            standard_name = normalize_account_name(account_name)
            
            # 查找数值
            for cell_value in reversed(row[1:]):
                value = extract_number(cell_value)
                if value is not None:
                    data[standard_name] = value
                    break
        
        return data


__all__ = ['ExcelParser', 'XLSParser']
