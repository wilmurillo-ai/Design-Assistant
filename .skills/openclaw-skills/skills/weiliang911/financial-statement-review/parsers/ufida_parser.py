"""
用友财务软件文件解析器

支持解析用友软件导出的各种格式文件
- 用友U8/NC/YonSuite/畅捷通等版本
- 支持标准报表格式和凭证数据
"""
import re
from typing import Dict, List, Any, Optional
from pathlib import Path

from .base_parser import BaseParser, ParseResult, FinancialData, normalize_account_name, extract_number


class UfidaParser(BaseParser):
    """
    用友财务软件文件解析器
    
    支持版本：U8、NC、YonSuite、畅捷通T3/T6/T+等
    """
    
    name = "ufida_parser"
    description = "用友财务软件文件解析器"
    supported_extensions = ['.xlsx', '.xls', '.csv', '.txt']
    supported_software = ['UFIDA']
    
    # 用友特有标识
    UFIDA_MARKERS = [
        '用友', 'UFIDA', 'U8', 'NC', 'YonSuite',
        '畅捷通', 'T3', 'T6', 'T+', '好会计', '好生意'
    ]
    
    # 用友标准报表格式
    UFIDA_REPORT_FORMATS = {
        'balance_sheet': {
            'names': ['资产负债表', 'Balance Sheet'],
            'headers': ['项目', '行次', '期末余额', '年初余额'],
        },
        'income_statement': {
            'names': ['利润表', '损益表', 'Income Statement'],
            'headers': ['项目', '行次', '本期金额', '上期金额'],
        },
        'cash_flow': {
            'names': ['现金流量表', 'Cash Flow'],
            'headers': ['项目', '行次', '本期金额'],
        },
        'voucher': {
            'names': ['凭证', '记账凭证', 'Voucher', 'GL_accvouch'],
            'headers': ['日期', '凭证号', '摘要', '科目', '借方', '贷方'],
        },
        'ledger': {
            'names': ['科目余额表', '总账', '明细账', 'Ledger'],
            'headers': ['科目编码', '科目名称', '期初余额', '本期发生', '期末余额'],
        },
    }
    
    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)
        self.version = self.config.get('version', 'auto')
    
    def can_parse(self, file_path: str) -> bool:
        """检查是否能解析该文件（用友格式）"""
        if not self.validate_file_exists(file_path):
            return False
        
        ext = self.get_file_extension(file_path)
        if ext not in self.supported_extensions:
            return False
        
        # 检查文件名是否包含用友标识
        file_name = Path(file_path).name.lower()
        for marker in [m.lower() for m in self.UFIDA_MARKERS]:
            if marker in file_name:
                return True
        
        # 对于Excel文件，尝试读取内容识别
        if ext in ['.xlsx', '.xls']:
            return self._check_excel_content(file_path)
        
        # 对于CSV文件，尝试识别用友格式
        if ext == '.csv':
            return self._check_csv_content(file_path)
        
        return False
    
    def _check_excel_content(self, file_path: str) -> bool:
        """检查Excel文件内容是否为用友格式"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # 检查工作表名称
            for sheet_name in wb.sheetnames:
                sheet_lower = sheet_name.lower()
                # 用友常用工作表名称
                ufida_sheets = ['gl_accvouch', 'code', 'customer', 'vendor', 'department']
                if any(s in sheet_lower for s in ufida_sheets):
                    wb.close()
                    return True
                
                for marker in self.UFIDA_MARKERS:
                    if marker.lower() in sheet_lower:
                        wb.close()
                        return True
            
            # 检查第一个工作表的内容
            if wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
                for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                    for cell in row:
                        if cell:
                            cell_str = str(cell).lower()
                            # 用友特有的列名
                            if any(kw in cell_str for kw in ['ufdata', '用友', 'cicode', 'ccuscode']):
                                wb.close()
                                return True
            
            wb.close()
        except:
            pass
        
        return False
    
    def _check_csv_content(self, file_path: str) -> bool:
        """检查CSV文件内容是否为用友格式"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                header = f.read(2048).lower()
                
                # 用友CSV特有的列名
                ufida_columns = ['ccode', 'ccuscode', 'csocode', 'cvouchtype', 'gl_accvouch']
                if any(col in header for col in ufida_columns):
                    return True
                
                for marker in [m.lower() for m in self.UFIDA_MARKERS]:
                    if marker in header:
                        return True
        except:
            pass
        
        return False
    
    def parse(self, file_path: str) -> ParseResult:
        """解析用友格式文件"""
        if not self.can_parse(file_path):
            return self.create_error_result(file_path, "无法识别为用友格式文件")
        
        ext = self.get_file_extension(file_path)
        
        try:
            if ext in ['.xlsx', '.xls']:
                return self._parse_excel(file_path)
            elif ext == '.csv':
                return self._parse_csv(file_path)
            elif ext == '.txt':
                return self._parse_txt(file_path)
            else:
                return self.create_error_result(file_path, f"不支持的文件格式: {ext}")
                
        except Exception as e:
            return self.create_error_result(file_path, f"解析失败: {str(e)}")
    
    def _parse_excel(self, file_path: str) -> ParseResult:
        """解析用友Excel文件"""
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        result = self.create_success_result(file_path)
        result.file_type = 'EXCEL'
        result.software_type = 'UFIDA'
        result.data = FinancialData()
        result.data.source_file = file_path
        result.data.source_software = '用友'
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # 识别工作表类型
            sheet_type = self._identify_sheet_type(sheet_name, sheet)
            
            if sheet_type == 'balance_sheet':
                result.data.balance_sheet = self._parse_ufida_sheet(
                    sheet, 'balance_sheet'
                )
                result.add_warning(f"工作表 '{sheet_name}' 识别为资产负债表")
            
            elif sheet_type == 'income_statement':
                result.data.income_statement = self._parse_ufida_sheet(
                    sheet, 'income_statement'
                )
                result.add_warning(f"工作表 '{sheet_name}' 识别为利润表")
            
            elif sheet_type == 'cash_flow':
                result.data.cash_flow = self._parse_ufida_sheet(
                    sheet, 'cash_flow'
                )
                result.add_warning(f"工作表 '{sheet_name}' 识别为现金流量表")
            
            elif sheet_type == 'voucher':
                # 凭证数据，可以汇总生成财务报表
                self._parse_voucher_sheet(sheet, result.data)
                result.add_warning(f"工作表 '{sheet_name}' 识别为记账凭证")
            
            elif sheet_type == 'ledger':
                # 科目余额表
                self._parse_ledger_sheet(sheet, result.data)
                result.add_warning(f"工作表 '{sheet_name}' 识别为科目余额表")
        
        wb.close()
        
        # 提取公司信息
        self._extract_company_info(file_path, result.data)
        
        return result
    
    def _parse_csv(self, file_path: str) -> ParseResult:
        """解析用友CSV文件"""
        import csv
        
        result = self.create_success_result(file_path)
        result.file_type = 'CSV'
        result.software_type = 'UFIDA'
        result.data = FinancialData()
        result.data.source_file = file_path
        result.data.source_software = '用友'
        
        # 检测编码
        encoding = self._detect_encoding(file_path)
        
        with open(file_path, 'r', encoding=encoding, errors='replace') as f:
            reader = csv.reader(f)
            rows = list(reader)
        
        if not rows:
            result.add_error("CSV文件为空")
            return result
        
        # 检测文件类型（凭证、科目余额表等）
        headers = [h.lower() for h in rows[0]]
        
        if any(col in headers for col in ['ccode', 'ccode_name', '科目']):
            # 科目余额表格式
            self._parse_csv_ledger(rows, result.data)
        elif any(col in headers for col in ['ino_id', 'cbookid', '凭证号']):
            # 凭证格式
            self._parse_csv_voucher(rows, result.data)
        else:
            # 通用解析
            result.data.raw_data['csv_content'] = rows
        
        return result
    
    def _parse_txt(self, file_path: str) -> ParseResult:
        """解析用友TXT文件（某些旧版本用友导出格式）"""
        result = self.create_success_result(file_path)
        result.file_type = 'TXT'
        result.software_type = 'UFIDA'
        result.data = FinancialData()
        result.data.source_file = file_path
        result.data.source_software = '用友'
        
        # 检测编码
        encoding = self._detect_encoding(file_path)
        
        with open(file_path, 'r', encoding=encoding, errors='replace') as f:
            content = f.read()
        
        # 用友TXT格式通常是固定宽度或特定分隔符
        # 这里需要根据实际格式进行调整
        result.data.raw_data['text_content'] = content
        
        return result
    
    def _detect_encoding(self, file_path: str) -> str:
        """检测文件编码"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'latin-1']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    f.read(1024)
                return encoding
            except:
                continue
        
        return 'utf-8'
    
    def _identify_sheet_type(self, sheet_name: str, sheet) -> str:
        """识别用友工作表类型"""
        sheet_name_lower = sheet_name.lower()
        
        # 用友特有的工作表命名
        if 'gl_accvouch' in sheet_name_lower or '凭证' in sheet_name:
            return 'voucher'
        
        if 'code' in sheet_name_lower and '科目' in sheet_name:
            return 'ledger'
        
        for report_type, config in self.UFIDA_REPORT_FORMATS.items():
            for name_pattern in config['names']:
                if name_pattern.lower() in sheet_name_lower:
                    return report_type
        
        # 内容分析
        return self._analyze_sheet_content(sheet)
    
    def _analyze_sheet_content(self, sheet) -> str:
        """分析工作表内容"""
        sample_text = ""
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            sample_text += ' '.join([str(cell) for cell in row if cell]) + ' '
        
        sample_text = sample_text.lower()
        
        # 用友特有的列名标识
        if any(kw in sample_text for kw in ['ino_id', 'dbill_date', 'cdigest']):
            return 'voucher'
        
        if any(kw in sample_text for kw in ['ccode', 'ccode_name', 'mb', 'md', 'mc']):
            return 'ledger'
        
        if any(kw in sample_text for kw in ['资产', '负债']):
            return 'balance_sheet'
        
        if any(kw in sample_text for kw in ['收入', '成本', '利润']):
            return 'income_statement'
        
        return 'unknown'
    
    def _parse_ufida_sheet(self, sheet, report_type: str) -> Dict[str, float]:
        """按用友格式解析工作表"""
        data = {}
        
        # 查找数据起始行（跳过标题）
        data_start_row = 2  # 默认从第3行开始
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
            if row and any(kw in str(cell) for cell in row for kw in ['项目', '科目', '行次']):
                data_start_row = row_idx + 2
                break
        
        # 解析数据
        for row in sheet.iter_rows(min_row=data_start_row, max_row=200, values_only=True):
            if not row or len(row) < 3:
                continue
            
            # 科目名称通常在第二列（第一列是行次）
            account_name = str(row[1]) if row[1] else ""
            if not account_name:
                continue
            
            # 跳过标题行和空行
            if any(kw in account_name for kw in ['项目', '编制单位', '行次', '附注']):
                continue
            
            standard_name = normalize_account_name(account_name)
            
            # 期末余额通常在第三列或第四列
            for cell_value in row[2:5]:
                value = extract_number(cell_value)
                if value is not None:
                    data[standard_name] = value
                    break
        
        return data
    
    def _parse_voucher_sheet(self, sheet, data: FinancialData):
        """解析用友凭证数据"""
        # 用友凭证表通常包含：日期、凭证号、摘要、科目、借方、贷方
        
        # 这里可以实现凭证汇总逻辑
        # 将凭证数据按科目汇总，生成试算平衡表
        
        vouchers = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 6:
                continue
            
            voucher = {
                'date': row[0],
                'voucher_no': row[1],
                'summary': row[2],
                'account': row[3],
                'debit': extract_number(row[4]) or 0,
                'credit': extract_number(row[5]) or 0,
            }
            
            vouchers.append(voucher)
        
        data.raw_data['vouchers'] = vouchers
    
    def _parse_ledger_sheet(self, sheet, data: FinancialData):
        """解析用友科目余额表"""
        balance_sheet = {}
        income_statement = {}
        
        for row in sheet.iter_rows(min_row=2, max_row=500, values_only=True):
            if not row or len(row) < 4:
                continue
            
            # 科目名称
            account_name = str(row[1]) if row[1] else ""
            if not account_name:
                continue
            
            standard_name = normalize_account_name(account_name)
            
            # 期末余额
            value = extract_number(row[-1]) if len(row) >= 4 else None
            
            if value is not None:
                # 分类
                if any(prefix in standard_name for prefix in ['收入', '成本', '费用', '利润']):
                    income_statement[standard_name] = value
                else:
                    balance_sheet[standard_name] = value
        
        data.balance_sheet = balance_sheet
        data.income_statement = income_statement
    
    def _parse_csv_ledger(self, rows: List[List], data: FinancialData):
        """解析CSV格式的科目余额表"""
        if not rows:
            return
        
        headers = rows[0]
        
        # 查找列索引
        name_col = self._find_column(headers, ['ccode_name', '科目名称', 'name'])
        end_balance_col = self._find_column(headers, ['me', '期末余额', 'ending_balance'])
        
        if name_col is None:
            name_col = 1
        
        balance_sheet = {}
        income_statement = {}
        
        for row in rows[1:]:
            if len(row) <= name_col:
                continue
            
            account_name = row[name_col].strip() if row[name_col] else ""
            if not account_name:
                continue
            
            standard_name = normalize_account_name(account_name)
            
            value = None
            if end_balance_col is not None and len(row) > end_balance_col:
                value = extract_number(row[end_balance_col])
            
            if value is not None:
                if any(prefix in standard_name for prefix in ['收入', '成本', '费用', '利润']):
                    income_statement[standard_name] = value
                else:
                    balance_sheet[standard_name] = value
        
        data.balance_sheet = balance_sheet
        data.income_statement = income_statement
    
    def _parse_csv_voucher(self, rows: List[List], data: FinancialData):
        """解析CSV格式的凭证"""
        # 实现凭证汇总逻辑
        data.raw_data['vouchers'] = rows[1:]  # 跳过标题行
    
    def _find_column(self, headers: List[str], keywords: List[str]) -> Optional[int]:
        """查找列索引"""
        headers_lower = [h.lower() for h in headers]
        
        for i, header in enumerate(headers_lower):
            for kw in keywords:
                if kw.lower() in header:
                    return i
        return None
    
    def _extract_company_info(self, file_path: str, data: FinancialData):
        """提取用友文件中的公司信息"""
        file_name = Path(file_path).stem
        
        # 尝试匹配公司名称
        patterns = [
            r'([^_]+)_.*资产负债表',
            r'([^_]+)_.*利润表',
            r'用友(.*?)账套',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, file_name)
            if match:
                data.company_name = match.group(1)
                break


__all__ = ['UfidaParser']
