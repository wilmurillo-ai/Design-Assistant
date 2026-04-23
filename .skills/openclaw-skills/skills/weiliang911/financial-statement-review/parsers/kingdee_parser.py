"""
金蝶财务软件文件解析器

支持解析金蝶软件导出的各种格式文件
- 金蝶KIS/K3/EAS/云星空等版本
- 支持标准报表格式和科目余额表
"""
import re
from typing import Dict, List, Any, Optional
from pathlib import Path

from .base_parser import BaseParser, ParseResult, FinancialData, normalize_account_name, extract_number
from .excel_parser import ExcelParser
from .csv_parser import CSVParser


class KingdeeParser(BaseParser):
    """
    金蝶财务软件文件解析器
    
    支持版本：KIS、K3、EAS、云星空、云星辰等
    """
    
    name = "kingdee_parser"
    description = "金蝶财务软件文件解析器"
    supported_extensions = ['.xlsx', '.xls', '.csv', '.xml']
    supported_software = ['KINGDEE']
    
    # 金蝶特有标识
    KINGDEE_MARKERS = [
        '金蝶', 'Kingdee', 'KIS', 'K3', 'EAS',
        '云星空', '云星辰', '精斗云'
    ]
    
    # 金蝶标准报表格式
    KINGDEE_REPORT_FORMATS = {
        'balance_sheet': {
            'names': ['资产负债表', 'Balance Sheet', ' balancesheet'],
            'account_col': 0,  # 科目名称列
            'period_col': -1,  # 期末数列（默认最后一列）
        },
        'income_statement': {
            'names': ['利润表', '损益表', 'Income Statement', 'Profit'],
            'account_col': 0,
            'period_col': -1,
        },
        'cash_flow': {
            'names': ['现金流量表', 'Cash Flow'],
            'account_col': 0,
            'period_col': -1,
        },
        'trial_balance': {
            'names': ['科目余额表', '试算平衡表', '科目汇总表', 'Trial Balance'],
            'account_col': 1,  # 金蝶科目余额表通常第二列是科目名称
            'period_col': -2,  # 期末余额列
        },
    }
    
    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)
        self.version = self.config.get('version', 'auto')  # auto, KIS, K3, EAS
    
    def can_parse(self, file_path: str) -> bool:
        """检查是否能解析该文件（金蝶格式）"""
        if not self.validate_file_exists(file_path):
            return False
        
        ext = self.get_file_extension(file_path)
        if ext not in self.supported_extensions:
            return False
        
        # 检查文件名是否包含金蝶标识
        file_name = Path(file_path).name.lower()
        for marker in [m.lower() for m in self.KINGDEE_MARKERS]:
            if marker in file_name:
                return True
        
        # 对于Excel文件，尝试读取内容识别
        if ext in ['.xlsx', '.xls']:
            return self._check_excel_content(file_path)
        
        return False
    
    def _check_excel_content(self, file_path: str) -> bool:
        """检查Excel文件内容是否为金蝶格式"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # 检查工作表名称
            for sheet_name in wb.sheetnames:
                sheet_lower = sheet_name.lower()
                for marker in self.KINGDEE_MARKERS:
                    if marker.lower() in sheet_lower:
                        wb.close()
                        return True
            
            # 检查第一个工作表的内容
            if wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
                for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                    for cell in row:
                        if cell:
                            cell_str = str(cell).lower()
                            for marker in [m.lower() for m in self.KINGDEE_MARKERS]:
                                if marker in cell_str:
                                    wb.close()
                                    return True
            
            wb.close()
        except:
            pass
        
        return False
    
    def parse(self, file_path: str) -> ParseResult:
        """解析金蝶格式文件"""
        if not self.can_parse(file_path):
            return self.create_error_result(file_path, "无法识别为金蝶格式文件")
        
        ext = self.get_file_extension(file_path)
        
        try:
            if ext in ['.xlsx', '.xls']:
                return self._parse_excel(file_path)
            elif ext == '.csv':
                return self._parse_csv(file_path)
            elif ext == '.xml':
                return self._parse_xml(file_path)
            else:
                return self.create_error_result(file_path, f"不支持的文件格式: {ext}")
                
        except Exception as e:
            return self.create_error_result(file_path, f"解析失败: {str(e)}")
    
    def _parse_excel(self, file_path: str) -> ParseResult:
        """解析金蝶Excel文件"""
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        result = self.create_success_result(file_path)
        result.file_type = 'EXCEL'
        result.software_type = 'KINGDEE'
        result.data = FinancialData()
        result.data.source_file = file_path
        result.data.source_software = '金蝶'
        
        # 识别并解析各个报表
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # 识别报表类型
            report_type = self._identify_report_type(sheet_name, sheet)
            
            if report_type == 'balance_sheet':
                result.data.balance_sheet = self._parse_kingdee_sheet(
                    sheet, self.KINGDEE_REPORT_FORMATS['balance_sheet']
                )
                result.add_warning(f"工作表 '{sheet_name}' 识别为资产负债表")
            
            elif report_type == 'income_statement':
                result.data.income_statement = self._parse_kingdee_sheet(
                    sheet, self.KINGDEE_REPORT_FORMATS['income_statement']
                )
                result.add_warning(f"工作表 '{sheet_name}' 识别为利润表")
            
            elif report_type == 'cash_flow':
                result.data.cash_flow = self._parse_kingdee_sheet(
                    sheet, self.KINGDEE_REPORT_FORMATS['cash_flow']
                )
                result.add_warning(f"工作表 '{sheet_name}' 识别为现金流量表")
            
            elif report_type == 'trial_balance':
                # 科目余额表，提取多个报表数据
                self._parse_trial_balance_sheet(sheet, result.data)
                result.add_warning(f"工作表 '{sheet_name}' 识别为科目余额表")
        
        wb.close()
        
        # 提取公司信息
        self._extract_company_info(file_path, result.data)
        
        return result
    
    def _parse_csv(self, file_path: str) -> ParseResult:
        """解析金蝶CSV文件"""
        # 使用通用CSV解析器，但添加金蝶特定的后处理
        csv_parser = CSVParser()
        result = csv_parser.parse(file_path)
        
        if result.success:
            result.software_type = 'KINGDEE'
            result.data.source_software = '金蝶'
        
        return result
    
    def _parse_xml(self, file_path: str) -> ParseResult:
        """解析金蝶XML文件"""
        # 金蝶某些版本支持XML格式导出
        result = self.create_success_result(file_path)
        result.file_type = 'XML'
        result.software_type = 'KINGDEE'
        result.data = FinancialData()
        result.data.source_file = file_path
        result.data.source_software = '金蝶'
        
        try:
            import xml.etree.ElementTree as ET
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # 解析XML结构（需要根据实际金蝶XML格式调整）
            # 这里提供基本框架
            for elem in root.iter():
                tag = elem.tag.lower()
                
                if 'balancesheet' in tag or '资产负债表' in tag:
                    result.data.balance_sheet = self._parse_xml_element(elem)
                elif 'incomestatement' in tag or '利润表' in tag:
                    result.data.income_statement = self._parse_xml_element(elem)
                elif 'cashflow' in tag or '现金流量' in tag:
                    result.data.cash_flow = self._parse_xml_element(elem)
            
        except Exception as e:
            result.add_error(f"XML解析失败: {str(e)}")
        
        return result
    
    def _parse_xml_element(self, element) -> Dict[str, float]:
        """解析XML元素"""
        data = {}
        
        for child in element:
            account_name = child.get('name', '') or child.tag
            value_str = child.get('value', '') or child.text
            
            if account_name and value_str:
                standard_name = normalize_account_name(account_name)
                value = extract_number(value_str)
                if value is not None:
                    data[standard_name] = value
        
        return data
    
    def _identify_report_type(self, sheet_name: str, sheet) -> str:
        """识别金蝶报表类型"""
        sheet_name_lower = sheet_name.lower()
        
        for report_type, config in self.KINGDEE_REPORT_FORMATS.items():
            for name_pattern in config['names']:
                if name_pattern.lower() in sheet_name_lower:
                    return report_type
        
        # 如果名称无法识别，尝试分析内容
        return self._analyze_sheet_content(sheet)
    
    def _analyze_sheet_content(self, sheet) -> str:
        """分析工作表内容识别报表类型"""
        sample_text = ""
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            sample_text += ' '.join([str(cell) for cell in row if cell]) + ' '
        
        sample_text = sample_text.lower()
        
        if any(kw in sample_text for kw in ['资产', '负债', '所有者权益']):
            return 'balance_sheet'
        elif any(kw in sample_text for kw in ['营业收入', '营业成本', '利润总额']):
            return 'income_statement'
        elif any(kw in sample_text for kw in ['经营活动', '投资活动', '筹资活动']):
            return 'cash_flow'
        elif any(kw in sample_text for kw in ['科目代码', '科目名称', '期初余额']):
            return 'trial_balance'
        
        return 'unknown'
    
    def _parse_kingdee_sheet(self, sheet, format_config: Dict) -> Dict[str, float]:
        """按金蝶格式解析工作表"""
        data = {}
        
        account_col = format_config['account_col']
        period_col = format_config['period_col']
        
        for row in sheet.iter_rows(min_row=1, max_row=200, values_only=True):
            if not row or len(row) <= max(abs(account_col), abs(period_col)):
                continue
            
            # 获取科目名称
            if account_col >= 0:
                account_name = row[account_col] if account_col < len(row) else None
            else:
                account_name = row[account_col] if abs(account_col) <= len(row) else None
            
            if not account_name:
                continue
            
            account_name = str(account_name).strip()
            
            # 跳过标题行
            if any(kw in account_name for kw in ['项目', '科目', '编制单位']):
                continue
            
            standard_name = normalize_account_name(account_name)
            
            # 获取期末余额
            if period_col >= 0:
                value_cell = row[period_col] if period_col < len(row) else None
            else:
                value_cell = row[period_col] if abs(period_col) <= len(row) else None
            
            value = extract_number(value_cell)
            if value is not None:
                data[standard_name] = value
        
        return data
    
    def _parse_trial_balance_sheet(self, sheet, data: FinancialData):
        """解析金蝶科目余额表"""
        balance_sheet = {}
        income_statement = {}
        
        # 查找标题行
        header_row = 0
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
            if row and any(kw in str(cell) for cell in row for kw in ['科目名称', '科目']):
                header_row = i
                break
        
        # 解析数据行
        for row in sheet.iter_rows(min_row=header_row + 2, max_row=500, values_only=True):
            if not row or len(row) < 4:
                continue
            
            # 科目名称通常在第二列
            account_name = str(row[1]) if row[1] else ""
            if not account_name:
                continue
            
            standard_name = normalize_account_name(account_name)
            
            # 期末余额通常在倒数第二列
            value = extract_number(row[-2]) if len(row) >= 2 else None
            
            if value is not None:
                # 分类存储
                if any(prefix in standard_name for prefix in ['收入', '成本', '费用', '营业外', '利润']):
                    income_statement[standard_name] = value
                else:
                    balance_sheet[standard_name] = value
        
        data.balance_sheet = balance_sheet
        data.income_statement = income_statement
    
    def _extract_company_info(self, file_path: str, data: FinancialData):
        """提取金蝶文件中的公司信息"""
        # 从文件名中提取
        file_name = Path(file_path).stem
        
        # 尝试匹配公司名称模式
        # 例如: "XX公司_2024年资产负债表.xlsx"
        patterns = [
            r'([^_]+)_.*资产负债表',
            r'([^_]+)_.*利润表',
            r'金蝶(.*?)20',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, file_name)
            if match:
                data.company_name = match.group(1)
                break


__all__ = ['KingdeeParser']
