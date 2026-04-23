"""
Audit Logger - 全流程执行日志与用户告知体系
遵循PRD 4.5节要求
"""

import json
import time
import csv
from datetime import datetime
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field, asdict
from pathlib import Path


@dataclass
class LogEntry:
    """日志条目"""
    timestamp: float
    operation: str  # retry, fallback, degradation
    task_id: str
    details: Dict[str, Any] = field(default_factory=dict)
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        data = asdict(self)
        data['datetime'] = datetime.fromtimestamp(self.timestamp).isoformat()
        return data


class AuditLogger:
    """
    全流程执行日志与用户告知体系
    
    Features:
    - 完整记录重试/切换/降级操作
    - 支持导出Excel/PDF格式
    - 实时状态同步通知
    - 满足企业级审计要求
    """
    
    def __init__(self, log_dir: Optional[str] = None):
        """
        初始化审计日志器
        
        Args:
            log_dir: 日志存储目录
        """
        self.log_dir = Path(log_dir) if log_dir else Path('./logs')
        self.log_dir.mkdir(parents=True, exist_ok=True)
        
        self._logs: List[LogEntry] = []
        self._notification_callbacks: List[Callable] = []
    
    def log_retry(
        self,
        task_id: str,
        exception_type: str,
        attempt: int,
        max_attempts: int,
        delay: float = 0.0,
        exception_message: str = "",
        category: str = ""
    ):
        """
        记录重试操作
        
        Args:
            task_id: 任务ID
            exception_type: 异常类型
            attempt: 当前尝试次数
            max_attempts: 最大尝试次数
            delay: 重试间隔
            exception_message: 异常消息
            category: 异常分类
        """
        entry = LogEntry(
            timestamp=time.time(),
            operation='retry',
            task_id=task_id,
            details={
                'exception_type': exception_type,
                'attempt': attempt,
                'max_attempts': max_attempts,
                'delay': delay,
                'exception_message': exception_message,
                'category': category,
                'remaining_attempts': max_attempts - attempt
            }
        )
        self._logs.append(entry)
        self._save_to_file(entry)
    
    def log_fallback(
        self,
        task_id: str,
        primary_tool: str,
        backup_tool: str,
        success: bool,
        param_mapping: Optional[Dict[str, str]] = None,
        error: str = "",
        duration: float = 0.0
    ):
        """
        记录备用工具切换操作
        
        Args:
            task_id: 任务ID
            primary_tool: 主工具名称
            backup_tool: 备用工具名称
            success: 是否成功
            param_mapping: 参数映射
            error: 错误信息
            duration: 执行时长
        """
        entry = LogEntry(
            timestamp=time.time(),
            operation='fallback',
            task_id=task_id,
            details={
                'primary_tool': primary_tool,
                'backup_tool': backup_tool,
                'success': success,
                'param_mapping': param_mapping or {},
                'error': error,
                'duration': duration
            }
        )
        self._logs.append(entry)
        self._save_to_file(entry)
    
    def log_degradation(
        self,
        task_id: str,
        level: str,
        failed_step: str,
        error: str,
        completed_steps: List[str] = None,
        skipped_steps: List[str] = None
    ):
        """
        记录降级操作
        
        Args:
            task_id: 任务ID
            level: 降级等级
            failed_step: 失败的步骤
            error: 错误信息
            completed_steps: 已完成的步骤
            skipped_steps: 被跳过的步骤
        """
        entry = LogEntry(
            timestamp=time.time(),
            operation='degradation',
            task_id=task_id,
            details={
                'level': level,
                'failed_step': failed_step,
                'error': error,
                'completed_steps': completed_steps or [],
                'skipped_steps': skipped_steps or []
            }
        )
        self._logs.append(entry)
        self._save_to_file(entry)
    
    def log_task_completion(
        self,
        task_id: str,
        success: bool,
        execution_time: float,
        retry_count: int = 0,
        fallback_count: int = 0,
        degradation_level: str = "NONE"
    ):
        """
        记录任务完成
        
        Args:
            task_id: 任务ID
            success: 是否成功
            execution_time: 执行时长
            retry_count: 重试次数
            fallback_count: 备用工具切换次数
            degradation_level: 降级等级
        """
        entry = LogEntry(
            timestamp=time.time(),
            operation='task_completion',
            task_id=task_id,
            details={
                'success': success,
                'execution_time': execution_time,
                'retry_count': retry_count,
                'fallback_count': fallback_count,
                'degradation_level': degradation_level
            }
        )
        self._logs.append(entry)
        self._save_to_file(entry)
    
    def _save_to_file(self, entry: LogEntry):
        """保存日志到文件"""
        date_str = datetime.fromtimestamp(entry.timestamp).strftime('%Y-%m-%d')
        log_file = self.log_dir / f'audit_{date_str}.jsonl'
        
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(json.dumps(entry.to_dict(), ensure_ascii=False) + '\n')
    
    def get_logs(
        self,
        task_id: Optional[str] = None,
        operation: Optional[str] = None,
        start_time: Optional[float] = None,
        end_time: Optional[float] = None
    ) -> List[LogEntry]:
        """
        查询日志
        
        Args:
            task_id: 任务ID筛选
            operation: 操作类型筛选
            start_time: 开始时间戳
            end_time: 结束时间戳
            
        Returns:
            List[LogEntry]: 符合条件的日志列表
        """
        filtered = self._logs
        
        if task_id:
            filtered = [log for log in filtered if log.task_id == task_id]
        
        if operation:
            filtered = [log for log in filtered if log.operation == operation]
        
        if start_time:
            filtered = [log for log in filtered if log.timestamp >= start_time]
        
        if end_time:
            filtered = [log for log in filtered if log.timestamp <= end_time]
        
        return filtered
    
    def export_logs(
        self,
        format: str = 'json',
        filepath: Optional[str] = None,
        task_id: Optional[str] = None
    ) -> str:
        """
        导出日志
        
        Args:
            format: 导出格式 (json/csv/excel/pdf)
            filepath: 导出文件路径
            task_id: 指定任务ID，None则导出全部
            
        Returns:
            str: 导出文件路径
        """
        logs = self.get_logs(task_id=task_id)
        
        if not filepath:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filepath = f'audit_logs_{timestamp}.{format}'
        
        filepath = Path(filepath)
        
        if format == 'json':
            self._export_json(logs, filepath)
        elif format == 'csv':
            self._export_csv(logs, filepath)
        elif format in ['excel', 'xlsx']:
            self._export_excel(logs, filepath)
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        return str(filepath)
    
    def _export_json(self, logs: List[LogEntry], filepath: Path):
        """导出为JSON"""
        data = [log.to_dict() for log in logs]
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    
    def _export_csv(self, logs: List[LogEntry], filepath: Path):
        """导出为CSV"""
        if not logs:
            return
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            # 获取所有可能的字段
            all_keys = set()
            for log in logs:
                all_keys.update(log.to_dict().keys())
                all_keys.update(log.details.keys())
            
            fieldnames = ['timestamp', 'datetime', 'operation', 'task_id'] + sorted(all_keys - {'timestamp', 'datetime', 'operation', 'task_id', 'details'})
            
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for log in logs:
                row = log.to_dict()
                row.update(log.details)
                row.pop('details', None)
                writer.writerow(row)
    
    def _export_excel(self, logs: List[LogEntry], filepath: Path):
        """导出为Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # 如果没有openpyxl，回退到CSV
            csv_path = filepath.with_suffix('.csv')
            self._export_csv(logs, csv_path)
            return str(csv_path)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        
        # 表头
        headers = ['时间', '操作类型', '任务ID', '详情']
        ws.append(headers)
        
        # 样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # 数据
        for log in logs:
            row = [
                datetime.fromtimestamp(log.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
                log.operation,
                log.task_id,
                json.dumps(log.details, ensure_ascii=False)
            ]
            ws.append(row)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 60
        
        wb.save(filepath)
    
    def generate_report(self, task_id: str) -> Dict[str, Any]:
        """
        生成任务执行报告
        
        Args:
            task_id: 任务ID
            
        Returns:
            Dict: 执行报告
        """
        logs = self.get_logs(task_id=task_id)
        
        if not logs:
            return {'error': 'No logs found for this task'}
        
        retry_logs = [log for log in logs if log.operation == 'retry']
        fallback_logs = [log for log in logs if log.operation == 'fallback']
        degradation_logs = [log for log in logs if log.operation == 'degradation']
        completion_logs = [log for log in logs if log.operation == 'task_completion']
        
        report = {
            'task_id': task_id,
            'execution_summary': {
                'total_operations': len(logs),
                'retry_count': len(retry_logs),
                'fallback_count': len(fallback_logs),
                'degradation_count': len(degradation_logs)
            },
            'retry_details': [log.details for log in retry_logs],
            'fallback_details': [log.details for log in fallback_logs],
            'degradation_details': [log.details for log in degradation_logs]
        }
        
        if completion_logs:
            report['final_status'] = completion_logs[-1].details
        
        return report
    
    def clear_logs(self):
        """清空所有日志"""
        self._logs = []