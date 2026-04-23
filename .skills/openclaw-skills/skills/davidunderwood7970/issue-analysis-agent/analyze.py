#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
客服问题分析脚本 - 核心分析模块
技能名称：issue-analysis-agent
版本：v2.0.0
更新：2026-03-23

功能：
1. 读取 Excel 客服问题数据
2. 统计分析（总数、解决率、TOP5 等）
3. 未解问题人统计（解决者中未解决问题最多的）
4. 输出结构化数据 JSON

团队协作：找茬 🐛 负责
"""

from openpyxl import load_workbook
from collections import defaultdict, Counter
import json
import sys
from pathlib import Path

def analyze_excel(file_path):
    """分析 Excel 文件，返回统计结果"""
    
    print(f"📂 读取文件：{file_path}")
    
    # 读取 Excel
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 读取表头
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"📋 表头字段：{headers[:10]}...")
    
    # 读取所有数据
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 有问题描述
            data.append(dict(zip(headers, row)))
    
    total = len(data)
    print(f"✅ 读取 {total} 条数据")
    
    if total == 0:
        print("❌ 数据为空，请检查 Excel 文件")
        return None
    
    # 统计解决状态
    resolved = sum(1 for d in data if '已解决' in str(d.get('解决状态', '')))
    unresolved = total - resolved
    rate = f"{resolved/total*100:.1f}%"
    
    print(f"📊 解决状态：已解决 {resolved} ({rate}), 未解决 {unresolved}")
    
    # 问题类型统计
    types = Counter(d.get('类型', '其他') for d in data)
    
    # 平台统计
    platforms = Counter(d.get('所属平台', '其他') for d in data)
    
    # 模块统计
    modules = Counter(d.get('模块', '其他') for d in data)
    
    # 反馈人统计
    reporters = Counter(d.get('反馈人', '其他') for d in data)
    
    # 解决人统计（只统计已解决的）
    resolvers = Counter(
        d.get('解决者', '其他') 
        for d in data 
        if '已解决' in str(d.get('解决状态', ''))
    )
    
    # 未解问题人统计（重点！解决者中还有未解决问题的）
    unresolved_resolvers = Counter(
        d.get('解决者', '其他')
        for d in data
        if d.get('解决者') and '已解决' not in str(d.get('解决状态', ''))
    )
    
    print(f"⚠️ 未解问题人：{len(unresolved_resolvers)} 人有未解决问题")
    
    # 未解决模块统计
    unresolved_modules = Counter(
        d.get('模块', '其他')
        for d in data
        if '已解决' not in str(d.get('解决状态', ''))
    )
    
    # 周次统计（如果有周次字段）
    weekly = Counter()
    for d in data:
        week = d.get('所属周') or d.get('周次')
        if week:
            weekly[f"第{week}周"] += 1
    
    # 构建输出数据
    output = {
        'summary': {
            'total': total,
            'resolved': resolved,
            'unresolved': unresolved,
            'rate': rate
        },
        'types': dict(types.most_common(10)),
        'platforms': dict(platforms.most_common(10)),
        'modules': dict(modules.most_common(10)),
        'reporters': dict(reporters.most_common(10)),
        'resolvers': dict(resolvers.most_common(10)),
        'unresolved_resolvers': dict(unresolved_resolvers.most_common(5)),  # 未解问题人 TOP5
        'unresolved_modules': dict(unresolved_modules.most_common(10)),
        'weekly': dict(sorted(weekly.items(), key=lambda x: int(x[0].replace('第', '').replace('周', '')))) if weekly else {}
    }
    
    return output

def save_data(data, output_path):
    """保存分析结果到 JSON 文件"""
    
    output_dir = Path(output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"✅ 数据已保存：{output_path}")

def main():
    """主函数"""
    
    if len(sys.argv) < 2:
        print("用法：python3 analyze.py <excel 文件路径>")
        print("示例：python3 analyze.py /path/to/issue_data.xlsx")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_json = Path(excel_file).parent / 'analysis_data_latest.json'
    
    # 分析
    data = analyze_excel(excel_file)
    
    if data:
        # 保存
        save_data(data, output_json)
        
        # 打印摘要
        print(f"\n📊 数据摘要:")
        print(f"  问题总数：{data['summary']['total']}")
        print(f"  已解决：{data['summary']['resolved']} ({data['summary']['rate']})")
        print(f"  未解决：{data['summary']['unresolved']}")
        print(f"\n🏷️ 问题类型 TOP3:")
        for i, (t, c) in enumerate(list(data['types'].items())[:3], 1):
            print(f"  {i}. {t}: {c}")
        print(f"\n🖥️ 平台 TOP3:")
        for i, (p, c) in enumerate(list(data['platforms'].items())[:3], 1):
            print(f"  {i}. {p}: {c}")
        print(f"\n⚠️ 未解问题人 TOP5:")
        for i, (name, c) in enumerate(list(data['unresolved_resolvers'].items())[:5], 1):
            print(f"  {i}. {name}: {c}")
        
        print(f"\n✅ 分析完成！")

if __name__ == '__main__':
    main()
