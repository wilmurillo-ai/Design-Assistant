#!/usr/bin/env python3
"""
获取点点数据所有渠道的上架榜和下架榜
并导出 Excel 发送到钉钉群
"""

import asyncio
import sys
import yaml
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from notifiers.dingtalk import DingTalkNotifier
from diandian_manual_fetch_fixed import DiandianManualFetcher

import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger('FetchAllChannels')


def load_config():
    """加载配置"""
    config_path = Path(__file__).parent.parent / "config" / "dingtalk.yaml"
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def export_to_excel(all_results: dict, date_str: str):
    """导出所有数据到 Excel"""
    wb = Workbook()
    # 删除默认工作表
    default_ws = wb.active
    wb.remove(default_ws)
    
    # 样式
    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    channels = {
        "huawei": "华为",
        "xiaomi": "小米",
        "tencent": "应用宝",
        "oppo": "OPPO",
        "vivo": "vivo",
        "baidu": "百度",
        "qihoo360": "360",
        "wandoujia": "豌豆荚",
    }
    
    # 创建上架榜工作表
    ws_new = wb.create_sheet(title="新上架应用")
    ws_new.append(["渠道", "应用名称", "开发者", "上架日期"])
    for cell in ws_new[1]:
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    # 创建下架榜工作表
    ws_offline = wb.create_sheet(title="下架应用")
    ws_offline.append(["渠道", "应用名称", "开发者", "下架日期"])
    for cell in ws_offline[1]:
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    # 填充数据
    total_new = 0
    total_offline = 0
    
    for channel_key, channel_name in channels.items():
        result = all_results.get(channel_key, {})
        new_apps = result.get("new_apps", [])
        offline_apps = result.get("offline_apps", [])
        
        # 填充上架数据
        for app in new_apps:
            ws_new.append([
                channel_name,
                app.app_name,
                app.developer,
                app.release_date
            ])
            total_new += 1
        
        # 填充下架数据
        for app in offline_apps:
            ws_offline.append([
                channel_name,
                app.app_name,
                app.developer,
                app.offline_date
            ])
            total_offline += 1
    
    # 调整列宽
    for ws in [ws_new, ws_offline]:
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
    
    # 保存文件
    output_dir = Path(__file__).parent.parent / "reports" / "diandian_exports"
    output_dir.mkdir(exist_ok=True)
    
    filename = f"点点数据_上架下架榜_{date_str}.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)
    
    logger.info(f"✅ Excel 已保存：{output_path}")
    logger.info(f"📊 总计：新上架 {total_new} 个，下架 {total_offline} 个")
    
    return {
        "path": output_path,
        "filename": filename,
        "total_new": total_new,
        "total_offline": total_offline
    }


async def send_to_dingtalk(export_info, date_str):
    """发送到钉钉群"""
    config = load_config()
    
    notifier = DingTalkNotifier(config)
    
    text = f"# 📊 点点数据每日榜单 ({date_str})\n\n"
    text += f"- ✅ 已获取所有 8 个安卓渠道数据\n"
    text += f"- 🆕 新上架应用：**{export_info['total_new']}** 个\n"
    text += f"- 📉 下架应用：**{export_info['total_offline']}** 个\n\n"
    text += "数据文件见附件。"
    
    try:
        result = await notifier.send_file(text, str(export_info['path']), export_info['filename'])
        logger.info("✅ 已发送到钉钉群")
        return True
    except Exception as e:
        logger.error(f"❌ 发送到钉钉失败：{e}")
        import traceback
        traceback.print_exc()
        return False


async def main():
    """主函数"""
    date_str = datetime.now().strftime('%Y%m%d')
    logger.info(f"🚀 开始获取点点数据所有渠道榜单 - {date_str}")
    
    fetcher = DiandianManualFetcher()
    all_results = await fetcher.fetch_all_channels(limit=100)
    
    # 统计
    total_new = sum(len(r.get('new_apps', [])) for r in all_results.values())
    total_offline = sum(len(r.get('offline_apps', [])) for r in all_results.values())
    logger.info(f"📊 全部渠道获取完成：新上架 {total_new}，下架 {total_offline}")
    
    # 导出 Excel
    export_info = export_to_excel(all_results, date_str)
    
    # 发送钉钉
    success = await send_to_dingtalk(export_info, date_str)
    
    if success:
        print("\n✅ 任务完成！数据已发送到钉钉群")
    else:
        print("\n⚠️ 数据已保存但发送失败")
    
    return success


if __name__ == "__main__":
    success = asyncio.run(main())
    sys.exit(0 if success else 1)
