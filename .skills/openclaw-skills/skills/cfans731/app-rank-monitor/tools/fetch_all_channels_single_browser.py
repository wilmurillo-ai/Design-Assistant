#!/usr/bin/env python3
"""
获取点点数据所有渠道的上架榜和下架榜
使用单次浏览器启动，避免 SingletonLock 冲突
并导出 Excel 发送到钉钉群
"""

import asyncio
import sys
import yaml
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from playwright.async_api import async_playwright
from rankers.base import NewApp, OfflineApp
from notifiers.dingtalk import DingTalkNotifier

import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger('FetchAllChannelsSingleBrowser')


class DiandianFetcher:
    """点点数据爬虫 - 单次浏览器版"""
    
    CHANNELS = {
        "huawei": "华为",
        "xiaomi": "小米",
        "tencent": "应用宝",
        "oppo": "OPPO",
        "vivo": "vivo",
        "baidu": "百度",
        "qihoo360": "360",
        "wandoujia": "豌豆荚",
    }
    
    def __init__(self):
        self.config = self._load_config()
    
    def _load_config(self) -> dict:
        config_path = Path(__file__).parent.parent / "config" / "diandian.yaml"
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        return {"token": ""}
    
    async def _navigate_to_page(self, page):
        """导航到市场情报页面 - 直接访问 URL"""
        # 直接访问上架监控页面
        await page.goto("https://app.diandian.com/rank/android-new", wait_until="domcontentloaded", timeout=60000)
        await asyncio.sleep(5)
        
        # 检查登录
        page_content = await page.content()
        is_logged = '退出' in page_content or '账号设置' in page_content
        if not is_logged:
            logger.error("❌ 未登录！")
            return False
        
        logger.info("✅ 已登录，已直达国内安卓新上架榜单页面")
        return True
    
    async def _extract_apps(self, page, limit: int = 100, app_type: str = 'new'):
        """提取应用数据"""
        try:
            # 等待表格
            try:
                await page.wait_for_selector('table, [class*="dd-data-table"], [class*="list"]', timeout=5000)
                await asyncio.sleep(2)
            except:
                logger.warning("⚠️ 未找到表格，尝试直接提取")
            
            # 滚动加载
            for i in range(5):
                await page.evaluate("window.scrollBy(0, 1200)")
                await asyncio.sleep(2)
            
            # JS 提取 - 作为 Python 字符串
            js_code = r"""
function() {
    const apps = [];
    let rows = document.querySelectorAll('table tr');
    if (rows.length < 5) {
        rows = document.querySelectorAll('[class*="row"], [class*="item"], [class*="dd-data-table-row"], .list-item');
    }
    console.log('找到 ' + rows.length + ' 行');
    for (let i = 0; i < rows.length; i++) {
        const row = rows[i];
        const text = row.textContent.trim();
        if (!text || text.length < 5 || text.length > 300) continue;
        if (text.includes('排名') || text.includes('应用') || text.includes('开发者') || text.includes('分类')) continue;
        
        const cells = row.querySelectorAll('td, div, span');
        let rank = '';
        let name = '';
        let developer = '';
        let package_name = '';
        let category = '';
        let date = '';
        
        if (cells.length >= 3) {
            rank = cells[0] ? cells[0].textContent.trim() : '';
            name = cells[1] ? cells[1].textContent.trim() : '';
            developer = cells[2] ? cells[2].textContent.trim() : '';
            if (cells.length >= 4) category = cells[3] ? cells[3].textContent.trim() : '';
            if (cells.length >= 5) date = cells[4] ? cells[4].textContent.trim() : '';
        } else {
            const lines = text.split('\n').filter(l => l.trim());
            if (lines.length >= 2) {
                if (/^\d+$/.test(lines[0])) {
                    rank = lines[0];
                    name = lines[1];
                    developer = lines[2] || '';
                } else {
                    name = lines[0];
                    developer = lines[1] || '';
                }
            }
        }
        
        name = name.replace(/[\(\（].*[\)\）]/, '').trim();
        if (name && name.length > 1 && name.length < 100 && !/^[\d\s]+$/.test(name)) {
            if (!apps.some(a => a.name === name)) {
                apps.push({
                    rank: rank,
                    name: name,
                    developer: developer,
                    package: package_name,
                    category: category,
                    date: date || new Date().toISOString().split('T')[0]
                });
            }
        }
        if (apps.length >= 100) break;
    }
    return apps.slice(0, 100);
}
            """
            apps_data = await page.evaluate(js_code)
            
            # 转换
            result = []
            for app in apps_data:
                if app.get('name'):
                    if app_type == 'new':
                        result.append(NewApp(
                            app_name=app['name'],
                            package_name=app.get('package', ''),
                            developer=app.get('developer', ''),
                            category=app.get('category', ''),
                            release_date=app.get('date', datetime.now().strftime('%Y-%m-%d'))
                        ))
                    else:
                        result.append(OfflineApp(
                            app_name=app['name'],
                            package_name=app.get('package', ''),
                            developer=app.get('developer', ''),
                            category=app.get('category', ''),
                            offline_date=app.get('date', datetime.now().strftime('%Y-%m-%d'))
                        ))
            
            return result
            
        except Exception as e:
            logger.error(f"❌ 提取失败：{e}")
            import traceback
            traceback.print_exc()
            return []
    
    async def fetch_channel(self, page, channel_key: str, channel_name: str, app_type: str, limit: int = 100):
        """获取单个渠道数据"""
        monitor_text = "新上架" if app_type == "new" else "下架"
        logger.info(f"📱 开始获取 {channel_name} 渠道 {monitor_text} 数据...")
        
        try:
            # 如果是下架，导航到下架页面
            if app_type == 'offline':
                await page.goto("https://app.diandian.com/rank/android-offline", wait_until="domcontentloaded", timeout=60000)
                await asyncio.sleep(5)
            else:
                # 确保在上架页面
                current_url = page.url
                if 'android-new' not in current_url:
                    await page.goto("https://app.diandian.com/rank/android-new", wait_until="domcontentloaded", timeout=60000)
                    await asyncio.sleep(5)
            
            # 选择渠道
            await asyncio.sleep(2)
            js_code = f"""function() {{
                const btns = Array.from(document.querySelectorAll('*')).filter(el => 
                    el.textContent.trim() === '{channel_name}' &&
                    el.tagName === 'DIV'
                );
                if (btns.length > 0) {{
                    btns[0].click();
                    console.log('已点击 {channel_name}');
                }}
            }}"""
            await page.evaluate(js_code)
            await asyncio.sleep(3)
            logger.info(f"✅ 已选择 {channel_name} (JS)")
            
            # 点击列表模式
            try:
                await page.wait_for_selector('text="列表模式"', timeout=5000)
                js_list_mode = """function() {
                    const btn = Array.from(document.querySelectorAll('*')).find(el => 
                        el.textContent.includes('列表模式')
                    );
                    if (btn) {
                        btn.click();
                        console.log('已点击 列表模式');
                    }
                }"""
                await page.evaluate(js_list_mode)
                await asyncio.sleep(5)
                logger.info("✅ 已点击列表模式 (JS)")
            except Exception as e:
                logger.warning(f"⚠️ 未找到列表模式按钮：{e}")
            
            # 截图
            screenshot_path = Path(__file__).parent.parent / f"debug_{channel_key}_{app_type}_single.png"
            await page.screenshot(path=str(screenshot_path), full_page=True)
            
            # 提取数据
            apps = await self._extract_apps(page, limit, app_type)
            logger.info(f"✅ {channel_name} {monitor_text}：提取到 {len(apps)} 个应用")
            
            return apps
            
        except Exception as e:
            logger.error(f"❌ {channel_name} {monitor_text} 获取失败：{e}")
            import traceback
            traceback.print_exc()
            return []
    
    async def fetch_all_channels(self, limit: int = 100) -> dict:
        """获取所有渠道数据 - 单次浏览器启动"""
        user_data_dir = Path(__file__).parent.parent / ".browser_data" / "diandian"
        user_data_dir.mkdir(parents=True, exist_ok=True)
        
        results = {}
        
        async with async_playwright() as p:
            context = await p.chromium.launch_persistent_context(
                user_data_dir=str(user_data_dir),
                headless=False,
                channel='chrome',
                args=['--disable-blink-features=AutomationControlled', '--no-sandbox']
            )
            
            page = context.pages[0] if context.pages else await context.new_page()
            await page.add_init_script("Object.defineProperty(navigator, {webdriver: {get: () => undefined}})")
            
            # 导航到目标页面
            success = await self._navigate_to_page(page)
            if not success:
                await context.close()
                return {}
            
            # 获取每个渠道
            for channel_key, channel_name in self.CHANNELS.items():
                new_apps = await self.fetch_channel(page, channel_key, channel_name, 'new', limit)
                offline_apps = await self.fetch_channel(page, channel_key, channel_name, 'offline', limit)
                
                results[channel_key] = {
                    "new_apps": new_apps,
                    "offline_apps": offline_apps,
                    "channel_name": channel_name
                }
                
                logger.info(f"✅ {channel_name} 完成：新上架 {len(new_apps)} 个，下架 {len(offline_apps)} 个")
                await asyncio.sleep(3)
            
            await context.close()
        
        return results


def load_config():
    """加载钉钉配置"""
    config_path = Path(__file__).parent.parent / "config" / "dingtalk.yaml"
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def export_to_excel(all_results: dict, date_str: str):
    """导出所有数据到 Excel"""
    wb = Workbook()
    # 删除默认工作表
    default_ws = wb.active
    wb.remove(default_ws)
    
    # 样式
    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 创建上架榜工作表
    ws_new = wb.create_sheet(title="新上架应用")
    ws_new.append(["渠道", "应用名称", "开发者", "上架日期"])
    for cell in ws_new[1]:
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    # 创建下架榜工作表
    ws_offline = wb.create_sheet(title="下架应用")
    ws_offline.append(["渠道", "应用名称", "开发者", "下架日期"])
    for cell in ws_offline[1]:
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    # 填充数据
    total_new = 0
    total_offline = 0
    
    for channel_key, data in all_results.items():
        channel_name = data.get("channel_name", channel_key)
        new_apps = data.get("new_apps", [])
        offline_apps = data.get("offline_apps", [])
        
        # 填充上架数据
        for app in new_apps:
            ws_new.append([
                channel_name,
                app.app_name,
                app.developer,
                app.release_date
            ])
            total_new += 1
        
        # 填充下架数据
        for app in offline_apps:
            ws_offline.append([
                channel_name,
                app.app_name,
                app.developer,
                app.offline_date
            ])
            total_offline += 1
    
    # 调整列宽
    for ws in [ws_new, ws_offline]:
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
    
    # 保存文件
    output_dir = Path(__file__).parent.parent / "reports" / "diandian_exports"
    output_dir.mkdir(exist_ok=True)
    
    filename = f"点点数据_上架下架榜_{date_str}.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)
    
    logger.info(f"✅ Excel 已保存：{output_path}")
    logger.info(f"📊 总计：新上架 {total_new} 个，下架 {total_offline} 个")
    
    return {
        "path": output_path,
        "filename": filename,
        "total_new": total_new,
        "total_offline": total_offline
    }


async def send_to_dingtalk(export_info, date_str):
    """发送到钉钉群"""
    notifier = DingTalkNotifier()
    
    text = f"# 📊 点点数据每日榜单 ({date_str})\n\n"
    text += f"- ✅ 已获取所有 8 个安卓渠道数据\n"
    text += f"- 🆕 新上架应用：**{export_info['total_new']}** 个\n"
    text += f"- 📉 下架应用：**{export_info['total_offline']}** 个\n\n"
    text += "数据文件见附件。"
    
    try:
        result = await notifier.send_file(text, str(export_info['path']), export_info['filename'])
        logger.info("✅ 已发送到钉钉群")
        return True
    except Exception as e:
        logger.error(f"❌ 发送到钉钉失败：{e}")
        import traceback
        traceback.print_exc()
        return False


async def main():
    """主函数"""
    date_str = datetime.now().strftime('%Y%m%d')
    logger.info(f"🚀 开始获取点点数据所有渠道榜单 - {date_str}")
    
    fetcher = DiandianFetcher()
    all_results = await fetcher.fetch_all_channels(limit=100)
    
    if not all_results:
        logger.error("❌ 获取失败，未获取到任何数据")
        return False
    
    # 统计
    total_new = sum(len(r.get('new_apps', [])) for r in all_results.values())
    total_offline = sum(len(r.get('offline_apps', [])) for r in all_results.values())
    logger.info(f"📊 全部渠道获取完成：新上架 {total_new}，下架 {total_offline}")
    
    # 导出 Excel
    export_info = export_to_excel(all_results, date_str)
    
    # 发送钉钉
    success = await send_to_dingtalk(export_info, date_str)
    
    if success:
        print(f"\n✅ 任务完成！数据已发送到钉钉群")
        print(f"📄 文件：{export_info['path']}")
        print(f"📊 统计：新上架 {export_info['total_new']}，下架 {export_info['total_offline']}")
    else:
        print(f"\n⚠️ 数据已保存但发送失败")
        print(f"📄 文件：{export_info['path']}")
    
    return success


if __name__ == "__main__":
    success = asyncio.run(main())
    sys.exit(0 if success else 1)
