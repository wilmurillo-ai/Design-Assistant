#!/usr/bin/env python3
"""
iOS 榜单监控 - 增强版
参考原 iOS-app-rank-monitor 的实现方式

功能:
- 免费榜、付费榜、新上架榜
- Markdown 消息通知（TOP 20）
- Excel 完整榜单（100 个应用）
- 钉钉文件上传
"""

import httpx
import yaml
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger('IOSRankMonitor')


def load_config() -> dict:
    """加载配置"""
    # 尝试多个路径
    possible_paths = [
        Path(__file__).parent.parent / "config" / "ios_rank.yaml",
        Path(__file__).parent / "config" / "ios_rank.yaml",
        Path(__file__).parent.parent.parent / "config" / "ios_rank.yaml",
    ]
    
    for config_path in possible_paths:
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
    
    raise FileNotFoundError("配置文件不存在")


def fetch_ios_rank(ranking_type: str, limit: int = 100, api_version: str = "v2") -> list:
    """
    获取 iOS 榜单数据
    
    Args:
        ranking_type: 榜单类型 (top-free, top-paid, newapplications)
        limit: 获取数量 (默认 100)
        api_version: API 版本 (v1 或 v2)
    
    Returns:
        应用列表
    """
    # 构建 URL
    if api_version == "v2":
        # 新 API: https://rss.marketingtools.apple.com/api/v2/cn/apps/{type}/{limit}/apps.json
        url = f"https://rss.marketingtools.apple.com/api/v2/cn/apps/{ranking_type}/{limit}/apps.json"
    else:
        # 旧 API: https://itunes.apple.com/cn/rss/{type}/limit={limit}/json
        url = f"https://itunes.apple.com/cn/rss/{ranking_type}/limit={limit}/json"
    
    logger.info(f"📊 请求：{url}")
    
    try:
        with httpx.Client(timeout=30) as client:
            response = client.get(url)
            response.raise_for_status()
            data = response.json()
            
            # 解析数据
            apps = []
            
            if api_version == "v2":
                # v2 API 数据结构
                results = data.get('results', [])
                if not results:
                    results = data.get('feed', {}).get('results', [])
                if not results:
                    results = data.get('feed', {}).get('entry', [])
                
                for idx, item in enumerate(results, 1):
                    try:
                        if 'url' in item and 'name' in item:
                            # v2 API 新结构
                            app = {
                                'name': item.get('name', ''),
                                'artist': item.get('artistName', ''),
                                'category': item.get('genreNames', [''])[0] if item.get('genreNames') else '',
                                'price': '',
                                'rank': str(idx),
                                'icon': item.get('artworkUrl100', ''),
                                'link': item.get('url', ''),
                                'id': item.get('id', '')
                            }
                        elif 'attributes' in item:
                            # 旧结构
                            attributes = item.get('attributes', {})
                            app = {
                                'name': attributes.get('name', ''),
                                'artist': attributes.get('artistName', ''),
                                'category': attributes.get('genreNames', [''])[0] if attributes.get('genreNames') else '',
                                'price': attributes.get('price', {}).get('amount', '') if isinstance(attributes.get('price'), dict) else '',
                                'rank': str(idx),
                                'icon': attributes.get('artwork', {}).get('url', '').replace('{w}x{h}bb', '100x100bb') if isinstance(attributes.get('artwork'), dict) else '',
                                'link': attributes.get('url', ''),
                                'id': item.get('id', '')
                            }
                        else:
                            continue
                        
                        apps.append(app)
                    except Exception as e:
                        logger.debug(f"解析失败：{e}")
                        continue
            else:
                # v1 API 数据结构
                entries = data.get('feed', {}).get('entry', [])
                if not entries:
                    entries = data.get('entry', [])
                
                for entry in entries:
                    try:
                        if isinstance(entry.get('im:name'), dict):
                            name = entry.get('im:name', {}).get('label', '')
                        else:
                            name = entry.get('im:name', '')
                        
                        if isinstance(entry.get('im:artist'), dict):
                            artist = entry.get('im:artist', {}).get('label', '')
                        else:
                            artist = entry.get('im:artist', '')
                        
                        app = {
                            'name': name,
                            'artist': artist,
                            'category': entry.get('category', {}).get('attributes', {}).get('label', '') if isinstance(entry.get('category'), dict) else '',
                            'price': entry.get('im:price', {}).get('label', '') if isinstance(entry.get('im:price'), dict) else '',
                            'rank': entry.get('im:rank', {}).get('label', '') if isinstance(entry.get('im:rank'), dict) else str(len(apps) + 1),
                            'icon': entry.get('im:image', [{}])[-1].get('label', '') if isinstance(entry.get('im:image'), list) else '',
                            'link': entry.get('link', [{}])[0].get('attributes', {}).get('href', '') if isinstance(entry.get('link'), list) else '',
                            'id': entry.get('id', {}).get('label', '').split('/')[-2] if isinstance(entry.get('id'), dict) else ''
                        }
                        apps.append(app)
                    except Exception as e:
                        logger.debug(f"解析 v1 失败：{e}")
                        continue
            
            logger.info(f"✅ 获取到 {len(apps)} 个应用")
            return apps
            
    except Exception as e:
        logger.error(f"❌ 获取失败：{e}")
        return []


def format_report(apps: list, report_type: str) -> str:
    """
    格式化报告消息（Markdown 格式）
    
    Args:
        apps: 应用列表
        report_type: 报告类型
    
    Returns:
        Markdown 格式的报告
    """
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    message = f"【iOS 榜单监控】{report_type}\n\n"
    message += f"**数据源**: 苹果官方 RSS\n"
    message += f"**更新时间**: {timestamp}\n"
    message += f"**数量**: {len(apps)}个\n\n"
    
    if apps:
        is_new_apps = '新上架' in report_type
        
        # 只显示 TOP 20
        display_limit = 20
        
        if is_new_apps:
            # 新上架榜单
            message += "| 排名 | 应用名 | 开发者 | 详情 |\n"
            message += "|------|--------|--------|------|\n"
            
            for app in apps[:display_limit]:
                name = app['name']
                artist = app['artist'][:20] + '...' if len(app['artist']) > 20 else app['artist']
                rank = app['rank'] or str(apps.index(app) + 1)
                app_link = app['link']
                
                if isinstance(app_link, dict):
                    app_link = app_link.get('attributes', {}).get('href', '')
                if not app_link or not app_link.startswith('http'):
                    app_link = '#'
                
                message += f"| {rank} | {name} | {artist} | [查看]({app_link}) |\n"
        else:
            # 免费榜/付费榜
            message += "| 排名 | 应用名 | 开发者 | 分类 | 价格 | 详情 |\n"
            message += "|------|--------|--------|------|------|------|\n"
            
            for app in apps[:display_limit]:
                name = app['name']
                artist = app['artist'][:20] + '...' if len(app['artist']) > 20 else app['artist']
                category = app['category']
                price = app['price']
                rank = app['rank']
                app_link = app['link']
                
                if isinstance(app_link, dict):
                    app_link = app_link.get('attributes', {}).get('href', '')
                if not app_link or not app_link.startswith('http'):
                    app_link = '#'
                
                message += f"| {rank} | {name} | {artist} | {category} | {price} | [查看]({app_link}) |\n"
        
        if len(apps) > display_limit:
            message += f"\n... 还有 {len(apps) - display_limit} 个应用\n\n"
            message += f"**📊 完整榜单 Excel 文件将单独发送，支持钉钉在线预览，请注意查收**\n"
    else:
        message += "暂无数据\n"
    
    return message


def save_full_report(apps: list, report_type: str, filename_prefix: str) -> str:
    """保存完整的应用列表到 Excel 文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        logger.error("❌ openpyxl 未安装，请运行：pip install openpyxl")
        return ""
    
    reports_dir = Path(__file__).parent.parent / "reports"
    reports_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{filename_prefix}_{timestamp}.xlsx"
    file_path = reports_dir / file_name
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = report_type[:31]
    
    # 定义样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 100
    
    # 写入表头
    headers = ['排名', '应用名', '开发者', '分类', '价格', '详情链接']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for idx, app in enumerate(apps, 2):
        rank = app.get('rank', str(idx - 1))
        name = app['name']
        artist = app['artist']
        category = app.get('category', '')
        price = app.get('price', '')
        
        app_link = app.get('link', '')
        if isinstance(app_link, dict):
            app_link = app_link.get('attributes', {}).get('href', '')
        
        ws.cell(row=idx, column=1, value=rank).alignment = cell_alignment
        ws.cell(row=idx, column=2, value=name).alignment = cell_alignment
        ws.cell(row=idx, column=3, value=artist).alignment = cell_alignment
        ws.cell(row=idx, column=4, value=category).alignment = cell_alignment
        ws.cell(row=idx, column=5, value=price).alignment = cell_alignment
        
        link_cell = ws.cell(row=idx, column=6, value="点击查看")
        if app_link:
            link_cell.hyperlink = app_link
            link_cell.style = "Hyperlink"
        link_cell.alignment = cell_alignment
    
    # 添加统计信息
    stats_row = len(apps) + 3
    ws.cell(row=stats_row, column=1, value="生成时间:")
    ws.cell(row=stats_row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    ws.cell(row=stats_row + 1, column=1, value="应用总数:")
    ws.cell(row=stats_row + 1, column=2, value=len(apps))
    ws.cell(row=stats_row + 2, column=1, value="数据源:")
    ws.cell(row=stats_row + 2, column=2, value="苹果官方 RSS API")
    
    # 自动调整行高
    for row in ws.iter_rows(min_row=1, max_row=len(apps) + 1):
        ws.row_dimensions[row[0].row].height = 25
    
    wb.save(file_path)
    logger.info(f"📊 Excel 报告已保存：{file_path}")
    
    return str(file_path)


def load_dingtalk_config() -> dict:
    """加载钉钉配置"""
    # 尝试多个路径
    possible_paths = [
        Path(__file__).parent.parent.parent / "config" / "dingtalk.yaml",
        Path(__file__).parent.parent / "config" / "dingtalk.yaml",
        Path(__file__).parent / "config" / "dingtalk.yaml",
    ]
    
    for config_path in possible_paths:
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
                # 配置在 dingtalk 命名空间下
                return config.get('dingtalk', {})
    return {}


def get_dingtalk_access_token(client_id: str, client_secret: str) -> str:
    """获取钉钉 access_token"""
    try:
        url = "https://oapi.dingtalk.com/gettoken"
        params = {"appkey": client_id, "appsecret": client_secret}
        
        with httpx.Client(timeout=30) as client:
            response = client.get(url, params=params)
            result = response.json()
            
            if result.get('errcode') == 0:
                return result.get('access_token', '')
            else:
                logger.error(f"❌ 获取 token 失败：{result}")
                return ""
    except Exception as e:
        logger.error(f"❌ 获取 token 异常：{e}")
        return ""


def upload_file_to_dingtalk(file_path: str, client_id: str, client_secret: str) -> str:
    """上传文件到钉钉"""
    try:
        access_token = get_dingtalk_access_token(client_id, client_secret)
        if not access_token:
            return ""
        
        upload_url = "https://oapi.dingtalk.com/media/upload"
        
        with open(file_path, 'rb') as f:
            files = {'media': f}
            params = {'access_token': access_token, 'type': 'file'}
            
            with httpx.Client(timeout=30) as client:
                response = client.post(upload_url, params=params, files=files)
                result = response.json()
                
                if result.get('errcode') == 0:
                    media_id = result.get('media_id', '')
                    logger.info(f"✅ 文件上传成功，media_id: {media_id}")
                    return media_id
                else:
                    logger.error(f"❌ 文件上传失败：{result}")
                    return ""
    except Exception as e:
        logger.error(f"❌ 上传异常：{e}")
        return ""


def send_file_to_chat(media_id: str, file_name: str, chat_id: str, access_token: str) -> bool:
    """发送文件到钉钉群聊"""
    try:
        send_url = "https://oapi.dingtalk.com/chat/send"
        
        data = {
            "chatid": chat_id,
            "sender_id": "system",
            "msgtype": "file",
            "file": {"media_id": media_id}
        }
        
        params = {"access_token": access_token}
        
        with httpx.Client(timeout=30) as client:
            response = client.post(send_url, params=params, json=data)
            result = response.json()
            
            if result.get('errcode') == 0:
                logger.info(f"✅ 文件已发送到群聊：{file_name}")
                return True
            else:
                logger.error(f"❌ 发送失败：{result}")
                return False
    except Exception as e:
        logger.error(f"❌ 发送异常：{e}")
        return False


def send_markdown_to_dingtalk(message: str, webhook: str) -> bool:
    """发送 Markdown 消息到钉钉"""
    payload = {
        "msgtype": "markdown",
        "markdown": {
            "title": "iOS 榜单监控",
            "text": message
        },
        "at": {"isAtAll": True}
    }
    
    try:
        with httpx.Client(timeout=30) as client:
            response = client.post(webhook, json=payload)
            result = response.json()
            
            if result.get('errcode') == 0:
                logger.info("✅ 钉钉消息发送成功")
                return True
            else:
                logger.error(f"❌ 钉钉发送失败：{result}")
                return False
    except Exception as e:
        logger.error(f"❌ 发送异常：{e}")
        return False


def fetch_and_send_all():
    """获取所有榜单并发送"""
    # 加载配置
    config = load_config()
    dingtalk_config = load_dingtalk_config()
    
    webhook = config.get('dingtalk_webhook', '')
    client_id = dingtalk_config.get('client_id', '')
    client_secret = dingtalk_config.get('client_secret', '')
    chat_id = dingtalk_config.get('chat_id', '')
    
    if not webhook:
        logger.error("❌ 钉钉 Webhook 未配置")
        return False
    
    rankings = config.get('rankings', [])
    if not rankings:
        logger.error("❌ 未配置榜单")
        return False
    
    success_count = 0
    
    for ranking in rankings:
        ranking_type = ranking.get('type', '')
        name = ranking.get('name', '')
        limit = ranking.get('limit', 100)
        api_version = ranking.get('api_version', 'v2')
        
        if not ranking_type:
            continue
        
        logger.info(f"\n📊 获取 {name}...")
        
        # 获取数据
        apps = fetch_ios_rank(ranking_type, limit, api_version)
        
        if apps:
            # 保存完整报告
            report_type = ranking.get('report_name', name)
            filename_prefix = f"ios_{ranking_type}"
            full_report_path = save_full_report(apps, report_type, filename_prefix)
            
            # 格式化报告
            message = format_report(apps, report_type)
            
            # Step 1: 发送榜单消息
            logger.info(f"📤 发送 {name} 报告...")
            if send_markdown_to_dingtalk(message, webhook):
                success_count += 1
            
            # Step 2: 上传并发送 Excel 文件
            if client_id and client_secret and chat_id and full_report_path:
                logger.info(f"📤 上传 {name} Excel 文件...")
                logger.info(f"   文件路径：{full_report_path}")
                logger.info(f"   client_id: {client_id[:10]}...")
                logger.info(f"   chat_id: {chat_id}")
                file_name = Path(full_report_path).name
                media_id = upload_file_to_dingtalk(full_report_path, client_id, client_secret)
                
                if media_id:
                    logger.info(f"   ✅ 上传成功，media_id: {media_id[:20]}...")
                    access_token = get_dingtalk_access_token(client_id, client_secret)
                    if access_token:
                        logger.info(f"   ✅ 获取 token 成功")
                        success = send_file_to_chat(media_id, file_name, chat_id, access_token)
                        logger.info(f"   {'✅' if success else '❌'} 文件发送{'成功' if success else '失败'}")
                    else:
                        logger.error(f"   ❌ 获取 token 失败")
                else:
                    logger.error(f"   ❌ 文件上传失败")
                
                import time
                time.sleep(2)  # 避免发送太快
            else:
                logger.warning(f"⚠️ 跳过文件上传：client_id={bool(client_id)}, chat_id={bool(chat_id)}, file={bool(full_report_path)}")
        else:
            logger.warning(f"⚠️ {name} 无数据")
    
    logger.info(f"\n✅ 成功发送 {success_count}/{len(rankings)} 个报告")
    return success_count == len(rankings)


if __name__ == "__main__":
    success = fetch_and_send_all()
    exit(0 if success else 1)
