#!/usr/bin/env python3
"""
获取苹果 iOS 榜单数据并发送到钉钉
- 使用苹果 RSS API 获取 3 个榜单（免费榜、付费榜、新上架）
- 生成 Excel 文件
- 发送 TOP20 消息 + Excel 文件到钉钉群
"""

import asyncio
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import List, Dict
import httpx
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import yaml

# 添加 src 目录到 Python 路径
sys.path.insert(0, str(Path(__file__).parent / "src"))

from notifiers.dingtalk import DingTalkNotifier
from notifiers.upload_to_dingtalk import upload_file, send_file_to_chat
from utils.logger import setup_logger

logger = setup_logger()

# 苹果 iTunes RSS API 配置 (经典 API，仍然可用)
# 格式：https://itunes.apple.com/{region}/rss/topapplications/limit={limit}/explicit=true/json
APPLE_RSS_BASE = "https://itunes.apple.com"
REGION = "cn"      # 中国区
LIMIT = 100        # 获取 100 个应用

# 榜单类型映射
CHART_TYPES = {
    "top_free": {
        "name": "免费榜",
        "emoji": "🆓",
        "rss_path": "topfreeapplications"
    },
    "top_paid": {
        "name": "付费榜",
        "emoji": "💰",
        "rss_path": "toppaidapplications"
    },
    "new_free": {
        "name": "新上架",
        "emoji": "🆕",
        "rss_path": "newapplications"
    }
}


def parse_apple_rss(json_content: str) -> List[Dict]:
    """解析苹果 RSS JSON 数据"""
    apps = []
    
    try:
        import json
        data = json.loads(json_content)
        
        # iTunes RSS JSON 格式：feed.entry 是应用列表
        entries = data.get('feed', {}).get('entry', [])
        
        for idx, entry in enumerate(entries):
            app_data = {}
            
            # 应用 ID (从 id 标签获取)
            id_elem = entry.get('id', {})
            app_id = id_elem.get('attributes', {}).get('im:id', '')
            app_data['app_id'] = app_id
            
            # 应用名称
            name_elem = entry.get('im:name', {})
            app_data['app_name'] = name_elem.get('label', '')
            
            # 开发者
            artist_elem = entry.get('im:artist', {})
            app_data['developer'] = artist_elem.get('label', '')
            
            # 价格
            price_elem = entry.get('im:price', {})
            app_data['price'] = price_elem.get('label', 'Free')
            
            # 分类
            category_elem = entry.get('category', {})
            app_data['category'] = category_elem.get('attributes', {}).get('label', '')
            
            # 图标 URL (获取高分辨率)
            images = entry.get('im:image', [])
            if images:
                # 最后一个通常是最高分辨率的
                app_data['icon_url'] = images[-1].get('label', '')
            
            # 排名 (从索引获取)
            app_data['rank'] = idx + 1
            
            # 只有当有应用名称时才添加
            if app_data.get('app_name'):
                apps.append(app_data)
        
        return apps
        
    except json.JSONDecodeError as e:
        logger.error(f"JSON 解析错误：{e}")
        return []
    except Exception as e:
        logger.error(f"解析 RSS 数据异常：{e}")
        import traceback
        traceback.print_exc()
        return []


async def fetch_chart(chart_type: str) -> List[Dict]:
    """获取单个榜单数据"""
    config = CHART_TYPES[chart_type]
    rss_path = config['rss_path']
    
    # 构建 RSS URL (使用 JSON 格式)
    url = f"{APPLE_RSS_BASE}/{REGION}/rss/{rss_path}/limit={LIMIT}/explicit=true/json"
    
    logger.info(f"📡 获取 {config['name']} ({url})")
    
    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            response = await client.get(url)
            
            if response.status_code == 200:
                apps = parse_apple_rss(response.text)
                # 验证解析结果
                if apps:
                    logger.info(f"✅ {config['name']} 获取成功：{len(apps)} 个应用")
                    return apps
                else:
                    logger.warning(f"⚠️ {config['name']} 解析结果为空")
                    return []
                logger.info(f"✅ {config['name']} 获取成功：{len(apps)} 个应用")
                return apps
            else:
                logger.error(f"❌ {config['name']} 获取失败：HTTP {response.status_code}")
                return []
                
    except Exception as e:
        logger.error(f"❌ {config['name']} 获取异常：{e}")
        return []


def create_excel_report(all_charts: Dict[str, List[Dict]], output_path: str) -> str:
    """创建 Excel 报告"""
    wb = openpyxl.Workbook()
    
    # 删除默认 sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 样式定义
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for chart_type, apps in all_charts.items():
        config = CHART_TYPES[chart_type]
        sheet = wb.create_sheet(title=config['name'])
        
        # 表头
        headers = ['排名', '应用名称', '开发者', '分类', '价格', 'App ID']
        sheet.append(headers)
        
        # 设置表头样式
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置列宽
        column_widths = [8, 35, 25, 15, 12, 20]
        for col_num, width in enumerate(column_widths, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # 填充数据
        for idx, app in enumerate(apps, 2):
            row_data = [
                app.get('rank', idx - 1),
                app.get('app_name', ''),
                app.get('developer', ''),
                app.get('category', ''),
                app.get('price', 'Free'),
                app.get('app_id', '')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=idx, column=col_num)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
    
    # 保存文件
    wb.save(output_path)
    logger.info(f"✅ Excel 报告已保存：{output_path}")
    return output_path


def build_top20_markdown(chart_type: str, apps: List[Dict]) -> str:
    """构建 TOP20 Markdown 表格"""
    config = CHART_TYPES[chart_type]
    
    markdown = f"### {config['emoji']} {config['name']} TOP 20\n\n"
    markdown += "| 排名 | 应用名称 | 开发者 | 分类 |\n"
    markdown += "|------|----------|--------|------|\n"
    
    for app in apps[:20]:
        rank = app.get('rank', '-')
        app_name = app.get('app_name', '未知')
        developer = app.get('developer', '未知')[:20]  # 截断长开发者名
        category = app.get('category', '-')
        
        markdown += f"| {rank} | {app_name} | {developer} | {category} |\n"
    
    markdown += "\n"
    return markdown


async def send_to_dingtalk(all_charts: Dict[str, List[Dict]], excel_path: str, report_time: str):
    """发送报告到钉钉"""
    notifier = DingTalkNotifier()
    
    # 构建消息
    markdown_text = f"""## 🍎 iOS 榜单报告 ({report_time})

**数据源**: Apple RSS API
**获取时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
**榜单数量**: 3 个 (各 100 个应用)

---

"""
    
    # 添加各榜单 TOP20
    for chart_type in ["top_free", "top_paid", "new_free"]:
        apps = all_charts.get(chart_type, [])
        markdown_text += build_top20_markdown(chart_type, apps)
    
    markdown_text += """---
📊 完整数据请查看附件 Excel 文件
"""
    
    content = {
        "msgtype": "markdown",
        "markdown": {
            "title": f"🍎 iOS 榜单报告 {report_time}",
            "text": markdown_text
        },
        "at": {
            "isAtAll": True
        }
    }
    
    # 发送 Markdown 消息
    logger.info("📤 发送 TOP20 消息到钉钉...")
    await notifier._send(content)
    
    # 上传并发送 Excel 文件
    logger.info("📎 上传 Excel 文件到钉钉...")
    media_id = upload_file(excel_path)
    
    if media_id:
        if send_file_to_chat(media_id, Path(excel_path).name):
            logger.info("✅ Excel 文件已发送到钉钉群")
        else:
            logger.error("❌ 发送 Excel 文件失败")
    else:
        logger.error("❌ 上传 Excel 文件失败")


async def main():
    """主函数"""
    logger.info("=" * 60)
    logger.info("🚀 开始获取 iOS 榜单数据")
    logger.info("=" * 60)
    
    # 获取所有榜单
    all_charts = {}
    for chart_type in ["top_free", "top_paid", "new_free"]:
        apps = await fetch_chart(chart_type)
        all_charts[chart_type] = apps
        await asyncio.sleep(1)  # 避免请求过快
    
    # 检查是否有数据
    total_apps = sum(len(apps) for apps in all_charts.values())
    if total_apps == 0:
        logger.error("❌ 未获取到任何榜单数据")
        return False
    
    logger.info(f"📊 共获取 {total_apps} 个应用数据")
    
    # 生成报告文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    reports_dir = Path(__file__).parent / "src" / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    excel_filename = f"ios_rank_report_{timestamp}.xlsx"
    excel_path = reports_dir / excel_filename
    
    # 创建 Excel 报告
    create_excel_report(all_charts, str(excel_path))
    
    # 发送到钉钉
    report_time = datetime.now().strftime('%Y-%m-%d %H:%M')
    await send_to_dingtalk(all_charts, str(excel_path), report_time)
    
    logger.info("=" * 60)
    logger.info("✅ iOS 榜单报告完成")
    logger.info("=" * 60)
    
    return True


if __name__ == "__main__":
    import logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    result = asyncio.run(main())
    print(f"\n执行结果：{'成功' if result else '失败'}")
    exit(0 if result else 1)
