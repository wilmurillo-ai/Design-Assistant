import time
import os
import requests
from urllib.parse import urljoin
import sys
import re
from datetime import datetime, timedelta
import json
from bs4 import BeautifulSoup
import openpyxl as op
from openpyxl import Workbook, load_workbook
from ocr_utils import process_pdf, save_to_combined_excel

# ==========================================
#              爬取速度设置
# ==========================================
REQUEST_DELAY = 1.5
# ==========================================

# 文件名清洗
def sanitize_filename(name, default_ext='.pdf'):
    name = (name or '').strip()
    invalid = '\\/:*?"<>|'
    for ch in invalid:
        name = name.replace(ch, '_')
    if default_ext and not name.lower().endswith(default_ext):
        name = name + default_ext
    return name

def normalize_title(s):
    s = (s or '').strip().lower()
    s = re.sub(r'\.pdf$', '', s)
    s = re.sub(r'[\s\-_]+', '', s)
    return s

def calculate_default_date():
    today = datetime.now().date()
    if today.weekday() == 0:
        return today - timedelta(days=3)
    return today - timedelta(days=1)

def get_default_date_str():
    return calculate_default_date().strftime("%Y-%m-%d")

def contains_chinese(s):
    return bool(re.search(r'[\u4e00-\u9fa5]', s or ''))

def is_english_title(s):
    s = s or ''
    return (not contains_chinese(s)) and bool(re.search(r'[a-zA-Z]', s))

def is_garbled(s):
    s = s or ''
    if '�' in s:
        return True
    if re.search(r'(锟斤拷|Ã|Â)', s):
        return True
    return False

def is_generic_result_announcement(s):
    return normalize_title(s) == normalize_title('转让结果公告')

def should_use_attachment_only(title, raw_title):
    if normalize_title(raw_title) == normalize_title(title):
        return True
    if is_garbled(title) or is_english_title(title) or is_generic_result_announcement(title):
        return True
    return False

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if getattr(sys, 'frozen', False):
    # 如果是打包后的exe运行，使用exe所在的目录
    _SCRIPT_DIR = os.path.dirname(sys.executable)

# 全局Session，保持会话
session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
})

# 定义爬取源配置
CRAWL_SOURCES = {
    "announcement": {
        "name": "银登公告",
        "url": "https://www.yindeng.com.cn/xxpl/xxpl_bldkzr/bldkzr_zrgg/index.html",
        "folder_suffix": "银登公告"
    },
    "result": {
        "name": "银登结果",
        "url": "https://www.yindeng.com.cn/xxpl/xxpl_bldkzr/bldkzr_zrjggg/index.html",
        "folder_suffix": "银登公告结果"
    }
}

time_start = None

# 全局变量定义，用于 run_crawl 内部
wb = None
ws = None
row_count = 1
all_recognition_data = []
last_run_added_rows = 0
last_run_no_update = False

def download_pdf(pdf_url, file_name, save_dir, referer=None, base_url=None, log_func=print):
    try:
        headers = {
            "Referer": referer or base_url,
            "Accept": "application/pdf,application/octet-stream;q=0.9,*/*;q=0.8",
        }
        
        safe_name = sanitize_filename(file_name)
        file_path = os.path.join(save_dir, safe_name)
        if os.path.exists(file_path):
            base, ext = os.path.splitext(safe_name)
            idx = 2
            while True:
                candidate = f"{base}({idx}){ext}"
                candidate_path = os.path.join(save_dir, candidate)
                if not os.path.exists(candidate_path):
                    file_path = candidate_path
                    break
                idx += 1
        
        response = session.get(pdf_url, headers=headers, stream=True, allow_redirects=True, timeout=30)
        if response.status_code != 200:
            log_func(f"下载失败，状态码: {response.status_code}")
            return None
            
        content_type = (response.headers.get("Content-Type") or "").lower()
        first_bytes = b""
        with open(file_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    if not first_bytes:
                        first_bytes = chunk[:5]
                    f.write(chunk)
                    
        if ("pdf" not in content_type) and not first_bytes.startswith(b"%PDF"):
            log_func("响应内容不是PDF，可能需要会话或权限")
            try:
                os.remove(file_path)
            except Exception:
                pass
            return None
            
        log_func(f"成功下载: {safe_name}")
        return file_path
    except Exception as e:
        log_func(f"下载出错: {e}")
        return None

# 获取当前页面的所有结果链接
def get_announcement_links(page_url, log_func=print):
    try:
        resp = session.get(page_url, timeout=15)
        if resp.status_code != 200:
            log_func(f"页面请求失败: {resp.status_code}")
            return [], None
        
        # 尝试检测编码，防止乱码
        resp.encoding = resp.apparent_encoding or 'utf-8'
        
        soup = BeautifulSoup(resp.text, 'lxml')
        
        # 获取总条数/总页数信息
        total_items = 0
        try:
            total_span = soup.find('span', class_='pageTotal', id='pageTotal')
            if total_span:
                match = re.search(r'\d+', total_span.get_text())
                if match:
                    total_items = int(match.group())
        except Exception:
            pass
            
        # 解析列表
        links = []
        list_box = soup.find('ul', class_='list-box')
        if list_box:
            rows = list_box.find_all('div', class_='rightListContent list-item')
            log_func(f"找到 {len(rows)} 个结果行")
            for row in rows:
                try:
                    # 标题和链接
                    link_element = row.find('span', class_='rightListImport unlock list-content').find('a')
                    title = link_element.get_text(strip=True)
                    href = link_element.get('href')
                    if href and not href.startswith('http'):
                        href = urljoin(page_url, href)
                    
                    # 日期
                    date_element = row.find('span', class_='rightListTime data-time')
                    date = date_element.get_text(strip=True)
                    
                    if href and title:
                        links.append({"title": title, "date": date, "href": href})
                except Exception as e:
                    log_func(f"处理结果行时出错: {e}")
                    continue
                    
        return links, total_items
    except Exception as e:
        log_func(f"获取结果链接出错: {e}")
        return [], None

# 检查页面是否包含PDF链接或内容
def process_announcement_page(page_url, title, log_func=print):
    try:
        resp = session.get(page_url, timeout=15)
        if resp.status_code != 200:
            return []
        resp.encoding = resp.apparent_encoding or 'utf-8'
        soup = BeautifulSoup(resp.text, 'lxml')
        
        pdf_links = []
        
        # 1. 查找页面中的PDF链接
        # 优先查找附件区域
        attach_area = soup.select('div.attach_main.attachment div.attach_file a[href$=".pdf"]')
        if not attach_area:
             attach_area = soup.select('div.attach_file a[href$=".pdf"]')
        
        # 如果还没找到，查找所有PDF链接
        if not attach_area:
             attach_area = soup.select('a[href$=".pdf"]')
             
        for element in attach_area:
            pdf_url = element.get('href')
            if pdf_url:
                if not pdf_url.startswith('http'):
                    pdf_url = urljoin(page_url, pdf_url)
                
                raw_title = element.get('title') or element.get_text(strip=True) or pdf_url.split('/')[-1]
                # 简单过滤，避免获取到无关的PDF
                # 统一使用公告标题作为文件名
                pdf_name = sanitize_filename(title)
                pdf_links.append({"url": pdf_url, "name": pdf_name})
        
        # 2. 查找iframe
        if not pdf_links:
            iframes = soup.find_all('iframe')
            for iframe in iframes:
                src = iframe.get('src')
                if src and ".pdf" in src.lower():
                    pdf_url = src
                    if not pdf_url.startswith('http'):
                        pdf_url = urljoin(page_url, pdf_url)
                    # 统一使用公告标题作为文件名
                    pdf_name = sanitize_filename(title)
                    pdf_links.append({"url": pdf_url, "name": pdf_name})
                    
        # 3. 如果仍然没有找到PDF，标记为文本
        if not pdf_links:
            # 简单检查是否有内容区域
            content_div = soup.find('div', class_='article-box')
            if content_div:
                 pdf_links.append({"url": page_url, "name": f"{title}.txt", "is_text": True})
                 
        return pdf_links
    except Exception as e:
        log_func(f"处理结果页面出错: {e}")
        return []

def parse_date(s):
    try:
        return datetime.strptime(s.strip(), "%Y-%m-%d").date()
    except Exception:
        return None

# 这里我们把 run_crawl 改造成一个通用函数，可以指定 source_type
def run_crawl_task(source_type, start_date_str=None, log_cb=None):
    # global wb, ws, row_count, time_start, last_run_added_rows, last_run_no_update, all_recognition_data, base_url
    
    config = CRAWL_SOURCES.get(source_type)
    if not config:
        if log_cb: log_cb(f"未知的爬取源: {source_type}")
        return None, None
        
    base_url = config['url']
    source_name = config['name']
    
    def log(msg):
        if log_cb:
            log_cb(f"[{source_name}] {str(msg)}")
        else:
            print(f"[{source_name}] {str(msg)}")
            
    time_start = time.time()
    all_recognition_data = [] # Reset for each run
    
    # 确定保存目录和文件路径
    date_prefix = start_date_str if start_date_str else get_default_date_str()
    date_prefix = date_prefix.replace(":", "-").replace("/", "-")
    folder_name = f"{date_prefix}{config['folder_suffix']}"
    
    current_pdf_dir = os.path.join(_SCRIPT_DIR, folder_name)
    if not os.path.exists(current_pdf_dir):
        os.makedirs(current_pdf_dir)
        log(f"创建下载目录: {current_pdf_dir}")
    else:
        log(f"使用下载目录: {current_pdf_dir}")
        
    # Excel路径也放在该目录下
    current_excel_path = os.path.join(current_pdf_dir, f"{date_prefix}{source_name}爬取记录.xlsx")
    current_combined_output_path = os.path.join(current_pdf_dir, f"{date_prefix}债权转让证明信息汇总.xlsx")
    
    # 初始化Excel
    if os.path.exists(current_excel_path):
        try:
            wb_loaded = load_workbook(current_excel_path)
            wb = wb_loaded
            ws = wb.active
            row_count = ws.max_row
            log(f"继续写入现有Excel，当前 {row_count} 行")
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.append(["标题", "发布日期", "详情链接", "PDF文件名", "PDF原始链接", "识别结果", "保存路径"])
            row_count = 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["标题", "发布日期", "详情链接", "PDF文件名", "PDF原始链接", "识别结果", "保存路径"])
        row_count = 1
        
    # 获取第一页数据并计算总页数
    log("正在获取首页数据并计算总页数...")
    first_page_links, total_items = get_announcement_links(base_url, log_func=log)
    
    total_pages = 1
    if total_items:
        items_per_page = 10
        total_pages = (total_items + items_per_page - 1) // items_per_page
        log(f"根据总条目数({total_items})计算得到总页数: {total_pages}")
    else:
        log("未能获取总条目数，将默认爬取直到无数据")
        total_pages = 1000 # 临时最大值
        
    pages_to_crawl = total_pages
        
    log(f"计划爬取 {pages_to_crawl} 页")
    
    start_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str.strip(), "%Y-%m-%d").date()
        except Exception:
            start_date = None
            
    crawled_pages = 0
    starting_rows = row_count
    
    # 循环爬取
    for page in range(1, pages_to_crawl + 1):
        # 银登网翻页逻辑
        current_page_index = total_pages - page + 1 if page > 1 else 1
        page_url = base_url if page == 1 else f"{base_url.replace('/index.html', '')}/index_{current_page_index}.html"
        
        log(f"正在处理第 {page} 页: {page_url}")
        
        if page == 1:
            announcements = first_page_links
        else:
            # 翻页前等待，控制速度
            time.sleep(REQUEST_DELAY)
            announcements, _ = get_announcement_links(page_url, log_func=log)
            
        if not announcements:
            log(f"第 {page} 页未获取到数据，跳过")
            continue
            
        log(f"找到 {len(announcements)} 个结果")
        
        # 时间过滤检查
        if start_date:
            parsed = [parse_date(a['date']) for a in announcements]
            has_target = any((d is not None) and (d >= start_date) for d in parsed)
            if not has_target:
                log(f"第 {page} 页全部早于 {start_date}，提前结束")
                break
                
        for announcement in announcements:
            if start_date:
                d = parse_date(announcement['date'])
                if d and d < start_date:
                    continue
            
            # 处理详情页前等待，控制速度
            time.sleep(REQUEST_DELAY)
            
            log(f"处理结果: {announcement['title']} ({announcement['date']})")
            pdf_links = process_announcement_page(announcement['href'], announcement['title'], log_func=log)
            
            if pdf_links:
                for pdf_link in pdf_links:
                    rec_json = ""
                    saved_file_path = ""
                    if not pdf_link.get('is_text', False):
                        saved_file_path = download_pdf(pdf_link['url'], pdf_link['name'], current_pdf_dir, referer=announcement['href'], base_url=base_url, log_func=log)
                        try:
                            # 只有结果公告需要进行OCR识别并汇总
                            if source_type == "result" and saved_file_path:
                                data = process_pdf(saved_file_path)
                                if data:
                                    rec_json = json.dumps(data, ensure_ascii=False)
                                    all_recognition_data.append(data)
                        except Exception as e:
                            log(f"PDF处理异常: {e}")
                            rec_json = ""
                    
                    row_count += 1
                    ws.cell(row_count, 1).value = announcement['title']
                    ws.cell(row_count, 2).value = announcement['date']
                    ws.cell(row_count, 3).value = announcement['href']
                    ws.cell(row_count, 4).value = pdf_link['name']
                    ws.cell(row_count, 5).value = pdf_link['url']
                    ws.cell(row_count, 6).value = rec_json
                    ws.cell(row_count, 7).value = saved_file_path
            else:
                log(f"未找到PDF或内容: {announcement['title']}")
                row_count += 1
                ws.cell(row_count, 1).value = announcement['title']
                ws.cell(row_count, 2).value = announcement['date']
                ws.cell(row_count, 3).value = announcement['href']
                ws.cell(row_count, 4).value = "未找到PDF或内容"
                ws.cell(row_count, 5).value = ""
                ws.cell(row_count, 6).value = ""
                ws.cell(row_count, 7).value = ""
                
        try:
            wb.save(current_excel_path)
            log(f"Excel记录已保存: {current_excel_path}")
        except Exception as e:
            log(f"保存Excel出错: {e}")

        # 只有结果公告需要生成汇总表
        if source_type == "result":
            try:
                save_to_combined_excel(all_recognition_data, current_combined_output_path)
            except Exception:
                pass
        log(f"第 {page} 页处理完成")
        crawled_pages += 1
        
    time_end = time.time()
    log(f"爬取完成，总用时: {time_end - time_start:.2f} 秒")
    last_run_added_rows = row_count - starting_rows
    last_run_no_update = (start_date is not None and last_run_added_rows <= 0)
    if last_run_no_update:
        log(f"{start_date}暂无更新")
        
    return current_pdf_dir, current_excel_path, last_run_added_rows, last_run_no_update

# 为了兼容旧接口，提供 run_crawl 默认调用 result
def run_crawl(start_date_str=None, on_complete=None, log_cb=None):
    pdf_dir, excel_path, added, no_update = run_crawl_task("result", start_date_str, log_cb)
    # 更新全局变量状态以供 main.py 检查
    global last_run_added_rows, last_run_no_update
    last_run_added_rows = added
    last_run_no_update = no_update
    
    if on_complete:
        on_complete()
    return pdf_dir, excel_path
