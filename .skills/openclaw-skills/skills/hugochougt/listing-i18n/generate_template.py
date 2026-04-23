#!/usr/bin/env python3
"""
listing-i18n template — 生成中文产品信息 Excel 输入模板
用法: python3 generate_template.py [output_path]
"""

import sys
import os


def check_deps():
    try:
        import openpyxl
    except ImportError:
        print("缺少依赖 openpyxl，请先安装：python3 -m pip install openpyxl")
        sys.exit(1)


def generate_template(output_path="product_template.xlsx"):
    check_deps()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── Sheet 1: Products ──
    ws = wb.active
    ws.title = "Products"

    headers = [
        ("product_id", "产品编号", 12),
        ("brand", "品牌名", 15),
        ("product_name", "中文产品名", 30),
        ("category", "品类", 20),
        ("specs", "核心规格（分号分隔）", 40),
        ("selling_points", "核心卖点（分号分隔）", 50),
        ("keywords_cn", "中文关键词（逗号分隔）", 30),
        ("package_includes", "包装清单", 30),
        ("custom_attributes", "自定义属性（key:value | key:value）", 40),
        ("images_note", "图片描述（可选）", 25),
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写表头：第一行英文字段名，第二行中文说明
    for col_idx, (field, desc, width) in enumerate(headers, start=1):
        # 英文字段名
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        # 中文说明
        desc_cell = ws.cell(row=2, column=col_idx, value=desc)
        desc_cell.font = Font(italic=True, color="666666", size=10)
        desc_cell.alignment = Alignment(wrap_text=True)
        # 列宽
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # 示例数据
    sample = [
        "SP-001",
        "SoundPulse",
        "SP-Pro 主动降噪真无线蓝牙耳机",
        "电子产品 > 耳机",
        "蓝牙5.3; 续航40小时; 13mm动圈; IPX5防水; USB-C快充; 支持LDAC; 重量58g",
        "混合主动降噪深度达35dB; 13mm大动圈+LDAC还原Hi-Res音质; 单次充电续航40小时; IPX5防水运动无忧; 6麦通话降噪; 10分钟快充听3小时",
        "降噪耳机, 蓝牙耳机, 无线耳机, 运动耳机, 长续航耳机",
        "耳机x1, 充电仓x1, USB-C线x1, 硅胶耳塞(S/M/L)x3, 说明书x1",
        "color:黑色 | connectivity:蓝牙5.3 | driver_size:13mm",
        "黑色外观，哑光质感",
    ]
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=3, column=col_idx, value=val)

    ws.freeze_panes = "A3"

    # ── Sheet 2: Instructions ──
    ins = wb.create_sheet("Instructions")
    ins.column_dimensions["A"].width = 20
    ins.column_dimensions["B"].width = 60
    ins.column_dimensions["C"].width = 50

    title_font = Font(bold=True, size=13)
    section_font = Font(bold=True, size=11, color="4472C4")

    ins.cell(row=1, column=1, value="产品信息模板填写说明").font = title_font
    ins.merge_cells("A1:C1")

    instructions = [
        ("", "", ""),
        ("字段", "说明", "示例"),
        ("product_id", "产品唯一编号，自定义即可", "SP-001"),
        ("brand", "品牌英文名（不翻译）", "SoundPulse"),
        ("product_name", "产品中文全称", "SP-Pro 主动降噪真无线蓝牙耳机"),
        ("category", "产品品类，用 > 分隔层级", "电子产品 > 耳机"),
        ("specs", "核心规格参数，用分号 ; 分隔每项", "蓝牙5.3; 续航40小时; IPX5防水"),
        ("selling_points", "产品核心卖点(中文)，用分号 ; 分隔。\n建议5-6个卖点，每个卖点包含具体数据。", "降噪深度35dB; 续航40小时; IPX5防水"),
        ("keywords_cn", "中文搜索关键词，用逗号 , 分隔", "降噪耳机, 蓝牙耳机, 无线耳机"),
        ("package_includes", "包装清单，列出所有配件", "耳机x1, 充电仓x1, 说明书x1"),
        ("custom_attributes", "品类特有属性，格式: key:value | key:value\n属性名用英文，值用中文。", "color:黑色 | material:铝合金 | wattage:15W"),
        ("images_note", "（可选）主图描述，帮助生成更准确的文案", "产品正面45度角，白色背景"),
        ("", "", ""),
        ("注意事项", "", ""),
        ("1. 每行一个产品", "", ""),
        ("2. 第1行为字段名（请勿修改），第2行为中文说明，第3行起填写产品数据", "", ""),
        ("3. selling_points 越详细，翻译质量越高。建议包含具体数字和对比数据", "", ""),
        ("4. custom_attributes 的 key 请用英文，便于系统识别", "", ""),
        ("5. 建议每次不超过 20 个产品，确保翻译质量", "", ""),
    ]

    for row_idx, (a, b, c) in enumerate(instructions, start=2):
        ins.cell(row=row_idx, column=1, value=a)
        ins.cell(row=row_idx, column=2, value=b)
        ins.cell(row=row_idx, column=3, value=c)

    # 表头行样式
    for col in range(1, 4):
        ins.cell(row=3, column=col).font = section_font
    ins.cell(row=15, column=1).font = section_font

    wb.save(output_path)
    print(f"✅ 模板已生成: {output_path}")
    print(f"   请在 Products sheet 第3行起填写产品信息")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "product_template.xlsx"
    generate_template(path)
