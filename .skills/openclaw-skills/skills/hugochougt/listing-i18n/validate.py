#!/usr/bin/env python3
"""
listing-i18n validate — 检查翻译输出文件的字符数/字节数限制（支持 Amazon + Shopify）
用法: python3 validate.py <output_file.xlsx>
"""

import sys
import os
import math
import re


def check_deps():
    try:
        import openpyxl
    except ImportError:
        print("缺少依赖 openpyxl，请先安装：python3 -m pip install openpyxl")
        sys.exit(1)


AMAZON_LIMITS = {
    "title": {"max": 200, "warn": 150, "unit": "chars"},
    "bullet_point": {"max": 500, "warn_min": 150, "warn_max": 250, "unit": "chars"},
    "backend_keywords": {"max": 249, "unit": "bytes"},
    "description": {"warn_min_words": 150, "warn_max_words": 300},
}

SHOPIFY_LIMITS = {
    "title": {"max": 100, "unit": "chars"},
    "seo_title": {"max": 70, "unit": "chars"},
    "seo_description": {"max": 320, "unit": "chars"},
    "description_html": {"warn_min_words": 100, "warn_max_words": 400},
    "tags": {"warn_min": 8, "warn_max": 15},
}


TAG_RE = re.compile(r"<[^>]+>")
CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\u3040-\u30ff\uff66-\uff9f]")


def approximate_word_count(text):
    plain_text = TAG_RE.sub(" ", str(text or ""))
    words = [part for part in plain_text.split() if part]
    cjk_chars = len(CJK_RE.findall(plain_text))
    cjk_units = math.ceil(cjk_chars / 2) if cjk_chars else 0
    return max(len(words), cjk_units)


def validate_amazon_row(row, row_idx, lang):
    """校验一行 Amazon 数据，返回 issues 列表"""
    issues = []
    product_id = row[0] or f"Row{row_idx}"

    # A: product_id, B: brand, C: title, D-H: bullet 1-5, I: description, J: backend_keywords
    title = row[2] or ""
    bullets = [row[i] or "" for i in range(3, 8)]
    description = row[8] or ""
    backend_kw = row[9] or ""

    # Title
    if len(title) > AMAZON_LIMITS["title"]["max"]:
        issues.append(("error", f"[{lang}] {product_id}: Title {len(title)} chars > {AMAZON_LIMITS['title']['max']} 上限"))
    elif len(title) > AMAZON_LIMITS["title"]["warn"]:
        issues.append(("warn", f"[{lang}] {product_id}: Title {len(title)} chars > {AMAZON_LIMITS['title']['warn']} 建议值"))

    # Bullet Points
    for bp_idx, bp in enumerate(bullets, start=1):
        if len(bp) > AMAZON_LIMITS["bullet_point"]["max"]:
            issues.append(("error", f"[{lang}] {product_id}: Bullet {bp_idx} {len(bp)} chars > {AMAZON_LIMITS['bullet_point']['max']} 上限"))
        elif len(bp) < AMAZON_LIMITS["bullet_point"]["warn_min"]:
            issues.append(("warn", f"[{lang}] {product_id}: Bullet {bp_idx} {len(bp)} chars < {AMAZON_LIMITS['bullet_point']['warn_min']} 建议下限"))
        elif len(bp) > AMAZON_LIMITS["bullet_point"]["warn_max"]:
            issues.append(("warn", f"[{lang}] {product_id}: Bullet {bp_idx} {len(bp)} chars > {AMAZON_LIMITS['bullet_point']['warn_max']} 建议上限"))

    # Backend Keywords (bytes)
    kw_bytes = len(backend_kw.encode("utf-8"))
    if kw_bytes > AMAZON_LIMITS["backend_keywords"]["max"]:
        issues.append(("error", f"[{lang}] {product_id}: Backend Keywords {kw_bytes} bytes > {AMAZON_LIMITS['backend_keywords']['max']} 上限"))

    # Description word count
    desc_words = approximate_word_count(description)
    if desc_words < AMAZON_LIMITS["description"]["warn_min_words"]:
        issues.append(("warn", f"[{lang}] {product_id}: Description 过短 ({desc_words} 词 < {AMAZON_LIMITS['description']['warn_min_words']})"))
    elif desc_words > AMAZON_LIMITS["description"]["warn_max_words"]:
        issues.append(("warn", f"[{lang}] {product_id}: Description 过长 ({desc_words} 词 > {AMAZON_LIMITS['description']['warn_max_words']})"))

    return issues


def validate_shopify_row(row, row_idx, lang):
    """校验一行 Shopify 数据，返回 issues 列表"""
    issues = []
    product_id = row[0] or f"Row{row_idx}"

    # A: product_id, B: brand, C: title, D: description_html, E: seo_title,
    # F: seo_description, G: tags, H: product_type
    description_html = row[3] or ""
    title = row[2] or ""
    seo_title = row[4] or ""
    seo_description = row[5] or ""
    tags = row[6] or ""
    product_type = row[7] or ""

    if len(title) > SHOPIFY_LIMITS["title"]["max"]:
        issues.append(("error", f"[{lang}] {product_id}: Title {len(title)} chars > {SHOPIFY_LIMITS['title']['max']} 上限"))

    # SEO Title
    if len(seo_title) > SHOPIFY_LIMITS["seo_title"]["max"]:
        issues.append(("error", f"[{lang}] {product_id}: SEO Title {len(seo_title)} chars > {SHOPIFY_LIMITS['seo_title']['max']} 上限"))

    # SEO Description
    if len(seo_description) > SHOPIFY_LIMITS["seo_description"]["max"]:
        issues.append(("error", f"[{lang}] {product_id}: SEO Description {len(seo_description)} chars > {SHOPIFY_LIMITS['seo_description']['max']} 上限"))

    # Description HTML non-empty and word count
    if not description_html.strip():
        issues.append(("error", f"[{lang}] {product_id}: Description HTML 为空"))
    else:
        desc_words = approximate_word_count(description_html)
        if desc_words < SHOPIFY_LIMITS["description_html"]["warn_min_words"]:
            issues.append(("warn", f"[{lang}] {product_id}: Description HTML 过短 ({desc_words} 词 < {SHOPIFY_LIMITS['description_html']['warn_min_words']})"))
        elif desc_words > SHOPIFY_LIMITS["description_html"]["warn_max_words"]:
            issues.append(("warn", f"[{lang}] {product_id}: Description HTML 过长 ({desc_words} 词 > {SHOPIFY_LIMITS['description_html']['warn_max_words']})"))

    # Tags check
    if not tags.strip():
        issues.append(("warn", f"[{lang}] {product_id}: Tags 为空"))
    else:
        tag_count = len([tag for tag in tags.split(",") if tag.strip()])
        if tag_count < SHOPIFY_LIMITS["tags"]["warn_min"]:
            issues.append(("warn", f"[{lang}] {product_id}: Tags 数量过少 ({tag_count} < {SHOPIFY_LIMITS['tags']['warn_min']})"))
        elif tag_count > SHOPIFY_LIMITS["tags"]["warn_max"]:
            issues.append(("warn", f"[{lang}] {product_id}: Tags 数量过多 ({tag_count} > {SHOPIFY_LIMITS['tags']['warn_max']})"))

    if not str(product_type).strip():
        issues.append(("warn", f"[{lang}] {product_id}: Product Type 为空"))

    return issues


def validate(filepath):
    check_deps()
    import openpyxl

    wb = openpyxl.load_workbook(filepath)
    all_issues = []
    total_products = 0
    sheets_checked = []

    for sheet_name in wb.sheetnames:
        is_amazon = sheet_name.startswith("Amazon_")
        is_shopify = sheet_name.startswith("Shopify_")

        if not is_amazon and not is_shopify:
            continue

        ws = wb[sheet_name]
        lang = sheet_name.split("_", 1)[1] if "_" in sheet_name else sheet_name
        platform = "Amazon" if is_amazon else "Shopify"
        sheets_checked.append(sheet_name)
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for row_idx, row in enumerate(rows, start=2):
            if not row or not row[0]:
                continue
            total_products += 1

            if is_amazon:
                all_issues.extend(validate_amazon_row(row, row_idx, f"{platform}/{lang}"))
            else:
                all_issues.extend(validate_shopify_row(row, row_idx, f"{platform}/{lang}"))

    # Report
    errors = [i for i in all_issues if i[0] == "error"]
    warnings = [i for i in all_issues if i[0] == "warn"]

    print(f"\n{'='*60}")
    print(f"Listing I18n 验证报告")
    print(f"{'='*60}")
    print(f"文件: {filepath}")
    print(f"检查 Sheet: {', '.join(sheets_checked) or '无'}")
    print(f"检查产品数: {total_products}")
    print(f"严重问题: {len(errors)}")
    print(f"警告提醒: {len(warnings)}")
    print(f"{'='*60}\n")

    if not sheets_checked:
        print("  ⚠️  未找到 Amazon_* 或 Shopify_* 格式的 sheet\n")
        return False

    if errors:
        print("❌ 严重问题（必须修复）:\n")
        for _, msg in errors:
            print(f"  ❌ {msg}")
        print()

    if warnings:
        print("⚠️  警告提醒（建议优化）:\n")
        for _, msg in warnings:
            print(f"  ⚠️  {msg}")
        print()

    if not errors and not warnings:
        print("  ✅ 全部通过，没有发现问题。\n")

    return len(errors) == 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 validate.py <output_file.xlsx>")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        sys.exit(1)

    success = validate(filepath)
    sys.exit(0 if success else 1)
