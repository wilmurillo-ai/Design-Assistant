#!/usr/bin/env python3
"""
listing-i18n inspect — 检查输入产品文件并输出字段映射建议
用法:
  python3 inspect_input.py <input_file>
  python3 inspect_input.py <input_file> --sheet Products
"""

import csv
import os
import sys

EXPECTED_FIELDS = [
    "product_id",
    "brand",
    "product_name",
    "category",
    "specs",
    "selling_points",
    "keywords_cn",
    "package_includes",
    "custom_attributes",
    "images_note",
]

REQUIRED_FIELDS = EXPECTED_FIELDS[:8]


def check_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("缺少依赖 openpyxl，请先安装：python3 -m pip install openpyxl")
        sys.exit(1)


def normalize(value):
    return str(value or "").strip()


def detect_header_row(rows):
    best_index = 0
    best_score = -1

    for idx, row in enumerate(rows[:5]):
        normalized = [normalize(cell) for cell in row]
        score = sum(1 for cell in normalized if cell in EXPECTED_FIELDS)
        if score > best_score:
            best_index = idx
            best_score = score

    return best_index


def data_start_index(rows, header_index):
    start = header_index + 1
    if start >= len(rows):
        return start

    first = normalize(rows[start][0]) if len(rows[start]) > 0 else ""
    second = normalize(rows[start][1]) if len(rows[start]) > 1 else ""
    if first == "产品编号" and second == "品牌名":
        return start + 1

    return start


def load_csv(filepath):
    with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))

    header_index = detect_header_row(rows)
    header = [normalize(cell) for cell in rows[header_index]]
    start = data_start_index(rows, header_index)
    data_rows = rows[start: start + 3]
    total_rows = max(len(rows) - start, 0)

    return {
        "source": os.path.basename(filepath),
        "sheet": None,
        "header_row": header_index + 1,
        "header": header,
        "samples": data_rows,
        "total_rows": total_rows,
    }


def load_excel(filepath, requested_sheet=None):
    check_openpyxl()
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    if requested_sheet:
        if requested_sheet not in wb.sheetnames:
            print(f"Sheet 不存在: {requested_sheet}")
            print(f"可用 sheet: {', '.join(wb.sheetnames)}")
            sys.exit(1)
        ws = wb[requested_sheet]
    else:
        ws = wb["Products"] if "Products" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header_index = detect_header_row(rows)
    header = [normalize(cell) for cell in rows[header_index]]
    start = data_start_index(rows, header_index)
    data_rows = rows[start: start + 3]
    total_rows = max(len(rows) - start, 0)

    return {
        "source": os.path.basename(filepath),
        "sheet": ws.title,
        "header_row": header_index + 1,
        "header": header,
        "samples": data_rows,
        "total_rows": total_rows,
        "all_sheets": wb.sheetnames,
    }


def analyze_field_mapping(header):
    mapping = {}
    missing_required = []
    missing_optional = []

    for field in REQUIRED_FIELDS:
        if field in header:
            mapping[field] = field
        else:
            missing_required.append(field)

    for field in EXPECTED_FIELDS[8:]:
        if field in header:
            mapping[field] = field
        else:
            missing_optional.append(field)

    return mapping, missing_required, missing_optional


def print_report(report):
    mapping, missing_required, missing_optional = analyze_field_mapping(report["header"])

    print("=" * 60)
    print("Listing I18n 输入检查")
    print("=" * 60)
    print(f"文件: {report['source']}")
    if report.get("all_sheets"):
        print(f"工作表: {', '.join(report['all_sheets'])}")
    if report["sheet"]:
        print(f"使用 Sheet: {report['sheet']}")
    print(f"检测到表头行: 第 {report['header_row']} 行")
    print(f"数据行数: {report['total_rows']}")
    print()

    print("字段列表:")
    for idx, field in enumerate(report["header"], start=1):
        print(f"  {idx:02d}. {field}")
    print()

    print("字段映射建议:")
    for field in EXPECTED_FIELDS:
        target = mapping.get(field, "未匹配")
        print(f"  {field:<18} -> {target}")
    print()

    if missing_required:
        print("缺少必填字段:")
        for field in missing_required:
            print(f"  - {field}")
        print()

    if missing_optional:
        print("缺少可选字段:")
        for field in missing_optional:
            print(f"  - {field}")
        print()

    print("前 3 行示例数据:")
    if not report["samples"]:
        print("  无数据")
    else:
        for idx, row in enumerate(report["samples"], start=1):
            values = [normalize(cell) for cell in row[: len(report["header"])]]
            print(f"  Row {idx}: {values}")

    print()
    if missing_required:
        print("结论: 输入文件缺少必填字段，不能直接进入翻译。")
        return 1

    print("结论: 输入文件结构可用于翻译，请继续确认平台和目标市场。")
    return 0


def parse_args(argv):
    if len(argv) < 2:
        print(__doc__.strip())
        sys.exit(1)

    filepath = argv[1]
    sheet = None

    if len(argv) >= 4 and argv[2] == "--sheet":
        sheet = argv[3]
    elif len(argv) > 2:
        print(__doc__.strip())
        sys.exit(1)

    return filepath, sheet


def main():
    filepath, sheet = parse_args(sys.argv)
    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        sys.exit(1)

    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        report = load_csv(filepath)
    else:
        report = load_excel(filepath, requested_sheet=sheet)

    sys.exit(print_report(report))


if __name__ == "__main__":
    main()
