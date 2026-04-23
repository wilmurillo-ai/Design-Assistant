#!/usr/bin/env python3
"""
listing-i18n build-output — 根据 JSON 描述生成最终 Excel 文件
用法: python3 build_output.py <translations.json> [output_file.xlsx]
"""

import json
import os
import sys

AMAZON_HEADERS = [
    "product_id",
    "brand",
    "title",
    "bullet_point_1",
    "bullet_point_2",
    "bullet_point_3",
    "bullet_point_4",
    "bullet_point_5",
    "description",
    "backend_keywords",
    "custom_attributes",
    "source_product_name",
]

SHOPIFY_HEADERS = [
    "product_id",
    "brand",
    "title",
    "description_html",
    "seo_title",
    "seo_description",
    "tags",
    "product_type",
    "custom_attributes",
    "source_product_name",
]


def check_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("缺少依赖 openpyxl，请先安装：python3 -m pip install openpyxl")
        sys.exit(1)


def load_payload(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        payload = json.load(f)

    if not isinstance(payload, dict):
        raise ValueError("顶层 JSON 必须是对象")

    sheets = payload.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ValueError("JSON 必须包含非空 sheets 数组")

    return payload


def expected_headers(sheet_name):
    if sheet_name.startswith("Amazon_"):
        return AMAZON_HEADERS
    if sheet_name.startswith("Shopify_"):
        return SHOPIFY_HEADERS
    if sheet_name == "Source_CN":
        return None
    raise ValueError(f"不支持的 sheet 名: {sheet_name}")


def validate_sheet(sheet):
    if not isinstance(sheet, dict):
        raise ValueError("每个 sheet 项必须是对象")

    name = sheet.get("name")
    headers = sheet.get("headers")
    rows = sheet.get("rows")

    if not name or not isinstance(name, str):
        raise ValueError("sheet.name 必须是非空字符串")
    if not isinstance(headers, list) or not headers:
        raise ValueError(f"{name}: headers 必须是非空数组")
    if not isinstance(rows, list):
        raise ValueError(f"{name}: rows 必须是数组")

    required_headers = expected_headers(name)
    if required_headers and headers != required_headers:
        raise ValueError(
            f"{name}: headers 不匹配，期望 {required_headers}，实际 {headers}"
        )

    width = len(headers)
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, list):
            raise ValueError(f"{name}: 第 {idx} 行必须是数组")
        if len(row) != width:
            raise ValueError(
                f"{name}: 第 {idx} 行列数错误，期望 {width}，实际 {len(row)}"
            )


def autosize_columns(ws):
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)


def build_workbook(payload, output_path):
    check_openpyxl()
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet in payload["sheets"]:
        validate_sheet(sheet)
        ws = wb.create_sheet(sheet["name"])
        ws.append(sheet["headers"])
        for row in sheet["rows"]:
            ws.append(row)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"
        autosize_columns(ws)

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print(__doc__.strip())
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "products_listing_i18n.xlsx"

    if not os.path.exists(input_path):
        print(f"文件不存在: {input_path}")
        sys.exit(1)

    try:
        payload = load_payload(input_path)
        build_workbook(payload, output_path)
    except Exception as exc:
        print(f"生成失败: {exc}")
        sys.exit(1)

    print(f"✅ 输出文件已生成: {output_path}")


if __name__ == "__main__":
    main()
