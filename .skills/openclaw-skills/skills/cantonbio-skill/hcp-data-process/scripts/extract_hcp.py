#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


GROUP_PREFIX = "Group:"
STOP_PREFIXES = ("Group Column", "Group Summaries", "~End", "Original Filename:")
HEADER_FILL = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
BODY_FILL = PatternFill(fill_type="solid", start_color="E2F0D9", end_color="E2F0D9")


@dataclass
class GroupSection:
    title: str
    mean_header: str
    cv_header: str
    records: list[tuple[object, object, object]]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def canonical_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_text(value).lower())


def first_nonblank(row: list[object]) -> str:
    for value in row:
        text = normalize_text(value)
        if text:
            return text
    return ""


def parse_tsv_scalar(value: str) -> object:
    text = value.strip()
    if not text:
        return None
    if text.upper() == "TRUE":
        return True
    if text.upper() == "FALSE":
        return False
    if re.fullmatch(r"[+-]?[0-9]+", text):
        if len(text.lstrip("+-")) > 1 and text.lstrip("+-").startswith("0"):
            return text
        try:
            return int(text)
        except ValueError:
            return text
    if re.fullmatch(r"[+-]?(?:[0-9]*\.[0-9]+|[0-9]+\.[0-9]*)(?:[Ee][+-]?[0-9]+)?", text):
        try:
            return float(text)
        except ValueError:
            return text
    return text


def load_raw_rows(path: Path) -> tuple[list[list[object]], Workbook]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        read_wb = load_workbook(path, data_only=True)
        write_wb = load_workbook(path)
        ws = read_wb.worksheets[0]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        return rows, write_wb

    if suffix == ".xls":
        with path.open("r", encoding="utf-16") as handle:
            reader = csv.reader(handle, delimiter="\t")
            rows = [[parse_tsv_scalar(cell) for cell in row] for row in reader]

        wb = Workbook()
        ws = wb.active
        ws.title = safe_sheet_title(path.stem)
        for row_index, row in enumerate(rows, start=1):
            for col_index, value in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=value)
        return rows, wb

    raise ValueError(f"Unsupported file type: {path}")


def safe_sheet_title(name: str) -> str:
    cleaned = re.sub(r"[:\\\\/?*\\[\\]]", "_", name).strip() or "Sheet"
    return cleaned[:31]


def build_header_map(header_row: list[object]) -> dict[str, tuple[int, str]]:
    mapping: dict[str, tuple[int, str]] = {}
    for index, value in enumerate(header_row):
        label = normalize_text(value)
        if not label:
            continue
        mapping[canonical_header(label)] = (index, label)
    return mapping


def parse_group_sections(rows: list[list[object]]) -> list[GroupSection]:
    groups: list[GroupSection] = []
    row_count = len(rows)
    row_index = 0

    while row_index < row_count:
        row = rows[row_index]
        first_value = first_nonblank(row)
        if not first_value.startswith(GROUP_PREFIX):
            row_index += 1
            continue

        title = first_value
        header_index = row_index + 1
        while header_index < row_count and not first_nonblank(rows[header_index]):
            header_index += 1
        if header_index >= row_count:
            break

        header_map = build_header_map(rows[header_index])
        sample_meta = header_map.get("sample")
        mean_meta = header_map.get("meanresult")
        cv_meta = header_map.get("cv") or header_map.get("cvpercent")

        if not sample_meta or not mean_meta or not cv_meta:
            row_index = header_index + 1
            continue

        sample_index, _ = sample_meta
        mean_index, mean_header = mean_meta
        cv_index, cv_header = cv_meta
        records: list[tuple[object, object, object]] = []

        data_index = header_index + 1
        while data_index < row_count:
            current_first = first_nonblank(rows[data_index])
            if not current_first:
                data_index += 1
                continue
            if current_first.startswith(GROUP_PREFIX) or current_first.startswith(STOP_PREFIXES):
                break

            row_values = rows[data_index]
            sample_value = row_values[sample_index] if sample_index < len(row_values) else None
            sample_text = normalize_text(sample_value)
            if sample_text:
                mean_value = row_values[mean_index] if mean_index < len(row_values) else None
                cv_value = row_values[cv_index] if cv_index < len(row_values) else None
                records.append((sample_value, mean_value, cv_value))
            data_index += 1

        if records:
            groups.append(
                GroupSection(
                    title=title,
                    mean_header=mean_header,
                    cv_header=cv_header,
                    records=records,
                )
            )

        row_index = data_index

    return groups


def reset_or_create_sheet(wb: Workbook, sheet_name: str):
    if sheet_name in wb.sheetnames:
        index = wb.sheetnames.index(sheet_name)
        ws = wb[sheet_name]
        wb.remove(ws)
        return wb.create_sheet(title=sheet_name, index=index)
    return wb.create_sheet(title=sheet_name)


def apply_layout(ws) -> None:
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15.66
    ws.column_dimensions["C"].width = 13


def fill_row(ws, row: int, start_col: int, end_col: int, fill: PatternFill) -> None:
    for column in range(start_col, end_col + 1):
        ws.cell(row=row, column=column).fill = fill


def populate_summary_sheet(ws, groups: list[GroupSection]) -> None:
    start_row = 1
    start_col = 1
    end_col = start_col + 2
    row = start_row

    for group in groups:
        ws.cell(row=row, column=start_col, value=group.title)
        fill_row(ws, row, start_col, end_col, HEADER_FILL)
        row += 1
        ws.cell(row=row, column=start_col, value="Sample")
        ws.cell(row=row, column=start_col + 1, value=group.mean_header)
        ws.cell(row=row, column=start_col + 2, value=group.cv_header)
        fill_row(ws, row, start_col, end_col, HEADER_FILL)
        row += 1

        for sample_value, mean_value, cv_value in group.records:
            ws.cell(row=row, column=start_col, value=sample_value)
            ws.cell(row=row, column=start_col + 1, value=mean_value)
            ws.cell(row=row, column=start_col + 2, value=cv_value)
            fill_row(ws, row, start_col, end_col, BODY_FILL)
            row += 1

        fill_row(ws, row, start_col, end_col, HEADER_FILL)
        row += 2

    apply_layout(ws)


def iter_input_files(paths: Iterable[Path]) -> Iterable[Path]:
    for path in paths:
        if path.is_dir():
            for child in sorted(path.iterdir()):
                if child.suffix.lower() in {".xlsx", ".xlsm", ".xls"} and not child.name.startswith("~$"):
                    yield child
        else:
            yield path


def build_output_path(input_path: Path, output_dir: Path | None) -> Path:
    base_dir = output_dir or input_path.parent
    return base_dir / f"{input_path.stem}_extracted.xlsx"


def process_file(input_path: Path, output_dir: Path | None, sheet_name: str, overwrite: bool) -> Path:
    rows, workbook = load_raw_rows(input_path)
    groups = parse_group_sections(rows)
    if not groups:
        raise ValueError(f"No extractable group sections found in {input_path.name}")

    summary_sheet = reset_or_create_sheet(workbook, sheet_name)
    populate_summary_sheet(summary_sheet, groups)

    output_path = build_output_path(input_path, output_dir)
    if output_path.exists() and not overwrite:
        raise FileExistsError(f"Output already exists: {output_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract HCP LQC summary tables from instrument export files."
    )
    parser.add_argument("inputs", nargs="+", help="Input file(s) or directory/directories.")
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="Directory for processed workbooks. Defaults to the source file directory.",
    )
    parser.add_argument(
        "--sheet-name",
        default="Sheet1",
        help="Name of the generated summary sheet. Default: Sheet1",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing output files.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_paths = [Path(item).expanduser().resolve() for item in args.inputs]
    failures = 0

    for input_path in iter_input_files(input_paths):
        try:
            output_path = process_file(
                input_path=input_path,
                output_dir=args.output_dir,
                sheet_name=args.sheet_name,
                overwrite=args.overwrite,
            )
            print(f"OK  {input_path.name} -> {output_path}")
        except Exception as exc:
            failures += 1
            print(f"ERR {input_path}: {exc}", file=sys.stderr)

    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
