#!/usr/bin/env python3
from __future__ import annotations
import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook

TARGET_COLS = {
    "B": "order_no",
    "D": "ship_method",
    "F": "product_name_ko",
    "G": "price",
    "K": "option",
    "L": "qty",
    "N": "n_value",
}


def find_start_row(ws, min_row: int = 11) -> int:
    last_row = min_row - 1
    for row in range(min_row, ws.max_row + 1):
        v = ws[f"B{row}"].value
        if isinstance(v, int):
            last_row = row
        elif isinstance(v, str) and v.strip().isdigit():
            last_row = row
    return last_row + 1


def write_rows(workbook_path: Path, records: List[Dict[str, Any]], sheet_name: Optional[str], out_path: Path) -> int:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    row = find_start_row(ws)
    first_row = row
    for rec in records:
        for col, key in TARGET_COLS.items():
            ws[f"{col}{row}"] = rec.get(key, "")
        row += 1
    wb.save(out_path)
    return first_row


def main() -> None:
    parser = argparse.ArgumentParser(description="Write normalized BUYMA records into workbook")
    parser.add_argument("--base", required=True, help="Base workbook path")
    parser.add_argument("--records", required=True, help="Normalized JSON records path")
    parser.add_argument("--out", required=True, help="Output workbook path")
    parser.add_argument("--sheet", help="Optional target sheet name")
    args = parser.parse_args()

    records = json.loads(Path(args.records).expanduser().resolve().read_text(encoding="utf-8"))
    first_row = write_rows(
        Path(args.base).expanduser().resolve(),
        records,
        args.sheet,
        Path(args.out).expanduser().resolve(),
    )
    print(json.dumps({"first_written_row": first_row, "count": len(records)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
