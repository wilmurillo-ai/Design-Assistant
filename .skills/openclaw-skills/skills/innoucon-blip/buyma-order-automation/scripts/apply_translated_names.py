#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font

def normalize_order_no(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return s[:6] if len(s) >= 6 else s


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply translated Korean product names to workbook by order number")
    parser.add_argument("--workbook", required=True, help="Workbook path")
    parser.add_argument("--json", required=True, help="JSON path with order_no/product_name_ko")
    parser.add_argument("--start-row", type=int, required=True, help="Start row of newly written rows")
    parser.add_argument("--end-row", type=int, required=True, help="End row of newly written rows")
    parser.add_argument("--sheet", help="Optional sheet name")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    json_path = Path(args.json).expanduser().resolve()

    items = json.loads(json_path.read_text(encoding="utf-8"))

    mapping = defaultdict(list)
    for item in items:
        order_no = normalize_order_no(item.get("order_no"))
        product_name_ko = str(item.get("product_name_ko", "")).strip()
        if order_no and product_name_ko:
            mapping[order_no].append(product_name_ko)

    wb = load_workbook(workbook_path)
    RED_FONT = Font(color="FF0000")
    ws = wb[args.sheet] if args.sheet else wb.active

    updated = 0
    missing = 0
    overrun = 0
    used_index = defaultdict(int)

    for row in range(args.start_row, args.end_row + 1):
        order_no = normalize_order_no(ws[f"B{row}"].value)
        if not order_no:
            continue

        names = mapping.get(order_no, [])
        idx = used_index[order_no]

        if not names:
            missing += 1
            continue

        if idx >= len(names):
            overrun += 1
            continue

        cell = ws[f"F{row}"]
        cell.value = names[idx]

        if len(names) > 1:
            cell.font = RED_FONT

        used_index[order_no] += 1
        updated += 1

    wb.save(workbook_path)
    print(json.dumps({
        "updated_rows": updated,
        "missing_rows": missing,
        "overrun_rows": overrun,
        "start_row": args.start_row,
        "end_row": args.end_row
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
