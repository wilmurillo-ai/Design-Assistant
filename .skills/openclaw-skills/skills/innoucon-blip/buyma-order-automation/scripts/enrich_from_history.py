#!/usr/bin/env python3
from __future__ import annotations
import argparse
import json
from pathlib import Path
from typing import Dict, List, Set, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Font

RED = "00FF0000"


def normalize(value) -> str:
    return str(value).strip() if value is not None else ""


def map_m_value(value: str) -> str:
    if value in {"9부산", "5신강"}:
        return value
    if value in {"6노원", "7마산", "8울산"}:
        return "화라"
    return ""


def collect_history(ws, until_row: int) -> Dict[str, Dict[str, Set[str]]]:
    hist: Dict[str, Dict[str, Set[str]]] = {}
    for row in range(2, until_row):
        f = normalize(ws[f"F{row}"].value)
        if not f:
            continue
        item = hist.setdefault(f, {"I": set(), "J": set(), "M": set()})
        i_val = normalize(ws[f"I{row}"].value)
        j_val = normalize(ws[f"J{row}"].value)
        m_val = map_m_value(normalize(ws[f"M{row}"].value))
        if i_val:
            item["I"].add(i_val)
        if j_val:
            item["J"].add(j_val)
        if m_val:
            item["M"].add(m_val)
    return hist


def set_joined(cell, values: Set[str]) -> None:
    if not values:
        return
    sorted_vals = sorted(values)
    cell.value = "\n".join(sorted_vals)
    if len(sorted_vals) > 1:
        cell.font = Font(color=RED)


def enrich(workbook_path: Path, start_row: int, end_row: int, sheet_name: str | None) -> Dict[str, int]:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    history = collect_history(ws, start_row)
    updated = 0
    multi = 0
    for row in range(start_row, end_row + 1):
        f = normalize(ws[f"F{row}"].value)
        if not f or f not in history:
            continue
        item = history[f]
        before = (ws[f"I{row}"].value, ws[f"J{row}"].value, ws[f"M{row}"].value)
        set_joined(ws[f"I{row}"], item["I"])
        set_joined(ws[f"J{row}"], item["J"])
        set_joined(ws[f"M{row}"], item["M"])
        after = (ws[f"I{row}"].value, ws[f"J{row}"].value, ws[f"M{row}"].value)
        if after != before:
            updated += 1
        if any(len(item[k]) > 1 for k in ("I", "J", "M")):
            multi += 1
    wb.save(workbook_path)
    return {"updated_rows": updated, "multi_match_rows": multi}


def main() -> None:
    parser = argparse.ArgumentParser(description="Enrich current workbook rows from historical rows")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--start-row", type=int, required=True)
    parser.add_argument("--end-row", type=int, required=True)
    parser.add_argument("--sheet")
    args = parser.parse_args()
    result = enrich(Path(args.workbook).expanduser().resolve(), args.start_row, args.end_row, args.sheet)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
