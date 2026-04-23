#!/usr/bin/env python3
from __future__ import annotations
import argparse
import json
from pathlib import Path
from typing import Dict, List
from openpyxl import load_workbook

REQUIRED_COLS = ["B", "D", "F", "G", "K", "L"]


def validate(workbook_path: Path, start_row: int, end_row: int, sheet_name: str | None) -> Dict[str, object]:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    blanks: List[Dict[str, object]] = []
    for row in range(start_row, end_row + 1):
        missing = [col for col in REQUIRED_COLS if ws[f"{col}{row}"].value in (None, "")]
        if missing:
            blanks.append({"row": row, "missing": missing})
    return {
        "ok": len(blanks) == 0,
        "checked_rows": max(0, end_row - start_row + 1),
        "blank_rows": blanks,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate required workbook cells")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--start-row", type=int, required=True)
    parser.add_argument("--end-row", type=int, required=True)
    parser.add_argument("--sheet")
    args = parser.parse_args()
    result = validate(Path(args.workbook).expanduser().resolve(), args.start_row, args.end_row, args.sheet)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
