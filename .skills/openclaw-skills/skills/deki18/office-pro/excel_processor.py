"""
Office Pro - Enterprise Document Automation Suite
Excel Processor Module

Enterprise-grade Excel processing based on openpyxl with template support
"""

from __future__ import annotations

import io
import json
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, Tuple, Type, Generator

from core import (
    DocumentProcessor, require_document,
    DependencyError, DocumentNotLoadedError, TemplateNotFoundError,
    TemplateRenderError, ParameterError
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.workbook import Workbook as WorkbookType
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import (
        Font, Fill, PatternFill, Border, Side, Alignment, Protection,
        NamedStyle
    )
    from openpyxl.utils import get_column_letter, coordinate_to_tuple
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, ScatterChart, Reference,
        Series
    )
    from openpyxl.chart.label import DataLabelList
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.comments import Comment
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.chart._chart import ChartBase as Chart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    WorkbookType = None
    Chart = None

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ChartFactory:
    """
    Chart factory for creating different chart types
    
    Follows Open/Closed Principle - new chart types can be registered
    without modifying existing code
    """
    
    _charts: Dict[str, Type[Chart]] = {}
    
    @classmethod
    def register(cls, chart_type: str, chart_class: Type[Chart]) -> None:
        """Register a chart type"""
        cls._charts[chart_type.lower()] = chart_class
    
    @classmethod
    def create(cls, chart_type: str) -> Chart:
        """Create a chart by type"""
        chart_type_lower = chart_type.lower()
        if chart_type_lower not in cls._charts:
            raise ParameterError(
                f"Unknown chart type: {chart_type}. "
                f"Available types: {list(cls._charts.keys())}"
            )
        return cls._charts[chart_type_lower]()
    
    @classmethod
    def available_types(cls) -> List[str]:
        """Get list of available chart types"""
        return list(cls._charts.keys())


if OPENPYXL_AVAILABLE:
    ChartFactory.register('bar', BarChart)
    ChartFactory.register('line', LineChart)
    ChartFactory.register('pie', PieChart)
    ChartFactory.register('scatter', ScatterChart)


class XlsxTemplateEngine:
    """
    xlsx-template style Excel template engine
    
    Supports placeholders in Excel templates: ${variable}, ${table:data}, ${image:logo}
    """
    
    PLACEHOLDER_PATTERN = re.compile(r'\$\{([^}]+)\}')
    
    def __init__(self, workbook: WorkbookType):
        """
        Initialize template engine
        
        Args:
            workbook: openpyxl Workbook object
        """
        self.workbook = workbook
        self.substitutions: Dict[str, Any] = {}
    
    def substitute(self, data: Dict[str, Any]) -> None:
        """
        Execute data substitution
        
        Args:
            data: Substitution data dictionary
        """
        self.substitutions = data
        
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            self._process_worksheet(worksheet)
    
    def _process_worksheet(self, worksheet: Worksheet) -> None:
        """Process all cells in worksheet"""
        cells_to_update = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if '${' in cell.value:
                        new_value = self._replace_placeholders(cell.value)
                        if new_value != cell.value:
                            cells_to_update.append((cell, new_value))
        
        for cell, new_value in cells_to_update:
            cell.value = new_value
    
    def _replace_placeholders(self, text: str) -> Any:
        """Replace placeholders in text"""
        if not isinstance(text, str):
            return text
        
        def replace_match(match):
            placeholder = match.group(1).strip() if match.group(1) else ""
            
            if not placeholder:
                return match.group(0)
            
            if ':' in placeholder:
                prefix, _, name = placeholder.partition(':')
                if not name.strip():
                    return match.group(0)
            
            if placeholder.startswith('table:'):
                table_name = placeholder[6:].strip()
                if not table_name:
                    return match.group(0)
                return self._handle_table_placeholder(table_name)
            
            if placeholder.startswith('image:'):
                image_name = placeholder[6:].strip()
                if not image_name:
                    return match.group(0)
                return self._handle_image_placeholder(image_name)
            
            if placeholder in self.substitutions:
                value = self.substitutions[placeholder]
                return self._format_value(value)
            
            if '.' in placeholder:
                value = self._get_nested_value(self.substitutions, placeholder)
                if value is not None:
                    return self._format_value(value)
            
            return match.group(0)
        
        return self.PLACEHOLDER_PATTERN.sub(replace_match, text)
    
    def _handle_table_placeholder(self, path: str) -> str:
        """Handle table placeholder"""
        value = self._get_nested_value(self.substitutions, path)
        return self._format_value(value) if value is not None else f"${{table:{path}}}"
    
    def _handle_image_placeholder(self, path: str) -> str:
        """Handle image placeholder"""
        return f"${{image:{path}}}"
    
    def _get_nested_value(self, data: Dict, path: str) -> Any:
        """Get nested dictionary value"""
        keys = path.split('.')
        value = data
        for key in keys:
            if isinstance(value, dict) and key in value:
                value = value[key]
            else:
                return None
        return value
    
    def _format_value(self, value: Any) -> str:
        """Format value to string"""
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")
        return str(value)


class ExcelProcessor(DocumentProcessor):
    """
    Enterprise-grade Excel processor
    
    Features:
    - Workbook creation and editing
    - Template-driven data substitution
    - Chart generation with factory pattern
    - Data import/export
    """
    
    _document_type_name = "Excel workbook"
    
    def __init__(self, template_dir: Optional[str] = None):
        """
        Initialize Excel processor
        
        Args:
            template_dir: Template directory path
        """
        if not OPENPYXL_AVAILABLE:
            raise DependencyError(
                "openpyxl is required. Install with: pip install openpyxl"
            )
        
        self._template_engine: Optional[XlsxTemplateEngine] = None
        super().__init__(template_dir)
    
    def _get_default_template_dir(self) -> str:
        """Get default template directory"""
        skill_root = Path(__file__).parent
        templates_dir = skill_root / "assets" / "templates" / "excel"
        return str(templates_dir)
    
    @require_document
    def save(self, path: str) -> None:
        """
        Save workbook
        
        Args:
            path: Save path
        """
        self._ensure_output_dir(path)
        self._document.save(path)
    
    def create_document(self) -> WorkbookType:
        """
        Create new workbook
        
        Returns:
            Workbook object
        """
        self._document = Workbook()
        self._template_engine = None
        return self._document
    
    def load_document(self, path: str) -> WorkbookType:
        """
        Load existing workbook
        
        Args:
            path: File path
            
        Returns:
            Workbook object
        """
        self._document = load_workbook(path, data_only=False)
        self._template_engine = None
        return self._document
    
    def load_workbook(self, path: str, data_only: bool = False) -> WorkbookType:
        """
        Load existing workbook (alias for load_document)
        
        Args:
            path: File path
            data_only: Whether to read only data (not formulas)
            
        Returns:
            Workbook object
        """
        self._document = load_workbook(path, data_only=data_only)
        self._template_engine = None
        return self._document
    
    def load_template(self, template_name: str) -> WorkbookType:
        """
        Load template file and initialize template engine
        
        Args:
            template_name: Template filename
            
        Returns:
            Workbook object
        """
        template_path = self._validate_template_path(template_name)
        self.load_workbook(str(template_path))
        self._template_engine = XlsxTemplateEngine(self._document)
        return self._document
    
    def render_template(self, data: Dict[str, Any]) -> WorkbookType:
        """
        Render template with data
        
        Args:
            data: Substitution data dictionary
            
        Returns:
            Rendered Workbook object
        """
        if not self._template_engine:
            raise DocumentNotLoadedError("template")
        
        try:
            self._template_engine.substitute(data)
            return self._document
        except Exception as e:
            raise TemplateRenderError(f"Template rendering failed: {e}")
    
    def get_sheet(self, name: Optional[str] = None) -> Worksheet:
        """
        Get worksheet
        
        Args:
            name: Worksheet name, defaults to active sheet
            
        Returns:
            Worksheet object
        """
        if not self._document:
            raise DocumentNotLoadedError("workbook")
        
        if name:
            return self._document[name]
        return self._document.active
    
    @require_document
    def create_sheet(self, title: str, index: Optional[int] = None) -> Worksheet:
        """
        Create worksheet
        
        Args:
            title: Worksheet title
            index: Insert position
            
        Returns:
            Worksheet object
        """
        return self._document.create_sheet(title=title, index=index)
    
    @require_document
    def remove_sheet(self, name: str) -> None:
        """
        Remove worksheet
        
        Args:
            name: Worksheet name
        """
        sheet = self._document[name]
        self._document.remove(sheet)
    
    def write_cell(self, cell: str, value: Any, sheet: Optional[str] = None) -> None:
        """
        Write cell
        
        Args:
            cell: Cell coordinate (e.g. 'A1')
            value: Value
            sheet: Worksheet name
        """
        ws = self.get_sheet(sheet)
        ws[cell] = value
    
    def read_cell(self, cell: str, sheet: Optional[str] = None) -> Any:
        """
        Read cell
        
        Args:
            cell: Cell coordinate
            sheet: Worksheet name
            
        Returns:
            Cell value
        """
        ws = self.get_sheet(sheet)
        return ws[cell].value
    
    def write_range(self, start_cell: str, data: List[List[Any]], 
                    sheet: Optional[str] = None) -> None:
        """
        Write data range
        
        Args:
            start_cell: Starting cell
            data: 2D data list
            sheet: Worksheet name
        """
        ws = self.get_sheet(sheet)
        start_row, start_col = coordinate_to_tuple(start_cell)
        
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                ws.cell(
                    row=start_row + row_idx,
                    column=start_col + col_idx,
                    value=value
                )
    
    def set_cell_style(self, cell: str, 
                       font: Optional[Dict] = None,
                       fill: Optional[Dict] = None,
                       border: Optional[Dict] = None,
                       alignment: Optional[Dict] = None,
                       number_format: Optional[str] = None,
                       sheet: Optional[str] = None) -> None:
        """
        Set cell style
        
        Args:
            cell: Cell coordinate
            font: Font settings {'name': 'Arial', 'size': 12, 'bold': True}
            fill: Fill settings {'color': 'FFFF00', 'pattern': 'solid'}
            border: Border settings {'style': 'thin', 'color': '000000'}
            alignment: Alignment settings {'horizontal': 'center', 'vertical': 'center'}
            number_format: Number format '#,##0.00'
            sheet: Worksheet name
        """
        ws = self.get_sheet(sheet)
        cell_obj = ws[cell]
        
        if font:
            cell_obj.font = Font(**font)
        if fill:
            cell_obj.fill = PatternFill(**fill)
        if border:
            side = Side(**border)
            cell_obj.border = Border(left=side, right=side, top=side, bottom=side)
        if alignment:
            cell_obj.alignment = Alignment(**alignment)
        if number_format:
            cell_obj.number_format = number_format
    
    def set_column_width(self, column: str, width: float, 
                         sheet: Optional[str] = None) -> None:
        """
        Set column width
        
        Args:
            column: Column letter (e.g. 'A')
            width: Width
            sheet: Worksheet name
        """
        ws = self.get_sheet(sheet)
        ws.column_dimensions[column].width = width
    
    def set_row_height(self, row: int, height: float, 
                       sheet: Optional[str] = None) -> None:
        """
        Set row height
        
        Args:
            row: Row number (1-based)
            height: Height
            sheet: Worksheet name
        """
        ws = self.get_sheet(sheet)
        ws.row_dimensions[row].height = height
    
    def merge_cells(self, range_str: str, sheet: Optional[str] = None) -> None:
        """
        Merge cells
        
        Args:
            range_str: Range (e.g. 'A1:B2')
            sheet: Worksheet name
        """
        ws = self.get_sheet(sheet)
        ws.merge_cells(range_str)
    
    def add_chart(self, chart_type: str, data_range: str, 
                  title: Optional[str] = None,
                  position: Optional[str] = None,
                  sheet: Optional[str] = None) -> Chart:
        """
        Add chart using factory pattern
        
        Args:
            chart_type: Chart type (bar/line/pie/scatter)
            data_range: Data range
            title: Chart title
            position: Chart position (cell coordinate)
            sheet: Worksheet name
            
        Returns:
            Chart object
        """
        ws = self.get_sheet(sheet)
        
        chart = ChartFactory.create(chart_type)
        chart.add_data(Reference(ws, data_range))
        
        if title:
            chart.title = title
        
        ws.add_chart(chart, position or 'E5')
        
        return chart
    
    def read_dataframe(self, sheet: Optional[str] = None, 
                       header: int = 0) -> Any:
        """
        Read data to pandas DataFrame
        
        Args:
            sheet: Worksheet name
            header: Header row number
            
        Returns:
            pandas DataFrame
        """
        if not PANDAS_AVAILABLE:
            raise DependencyError("pandas is required. Install with: pip install pandas")
        
        ws = self.get_sheet(sheet)
        
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        if not data:
            return pd.DataFrame()
        
        return pd.DataFrame(data[header+1:], columns=data[header])
    
    def read_dataframe_chunked(self, sheet: Optional[str] = None, 
                               chunk_size: int = 1000) -> Generator[Any, None, None]:
        """
        Read data in chunks for large files
        
        Args:
            sheet: Worksheet name
            chunk_size: Number of rows per chunk
            
        Yields:
            pandas DataFrame for each chunk
        """
        if not PANDAS_AVAILABLE:
            raise DependencyError("pandas is required. Install with: pip install pandas")
        
        ws = self.get_sheet(sheet)
        
        data = []
        header = None
        
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1:
                header = row
                continue
            
            data.append(row)
            
            if len(data) >= chunk_size:
                yield pd.DataFrame(data, columns=header)
                data = []
        
        if data:
            yield pd.DataFrame(data, columns=header)
    
    def write_dataframe(self, df: Any, start_cell: str = 'A1', 
                        sheet: Optional[str] = None,
                        include_header: bool = True,
                        index: bool = False) -> None:
        """
        Write pandas DataFrame to Excel
        
        Args:
            df: pandas DataFrame
            start_cell: Starting cell
            sheet: Worksheet name
            include_header: Whether to include header
            index: Whether to include index
        """
        if not PANDAS_AVAILABLE:
            raise DependencyError("pandas is required. Install with: pip install pandas")
        
        ws = self.get_sheet(sheet)
        start_row, start_col = coordinate_to_tuple(start_cell)
        
        if index:
            for i, idx in enumerate(df.index):
                ws.cell(row=start_row + i + (1 if include_header else 0), 
                       column=start_col, value=idx)
            start_col += 1
        
        if include_header:
            for col_idx, col_name in enumerate(df.columns):
                ws.cell(row=start_row, column=start_col + col_idx, value=col_name)
            start_row += 1
        
        for row_idx, row in enumerate(df.itertuples(index=False)):
            for col_idx, value in enumerate(row):
                ws.cell(row=start_row + row_idx, 
                       column=start_col + col_idx, 
                       value=value)
    
    def import_csv(self, csv_path: str, sheet: Optional[str] = None, 
                   delimiter: str = ',',
                   encoding: str = 'utf-8') -> None:
        """
        Import data from CSV file
        
        Args:
            csv_path: CSV file path
            sheet: Target worksheet name
            delimiter: Delimiter
            encoding: Encoding
        """
        import csv
        
        ws = self.get_sheet(sheet)
        
        with open(csv_path, 'r', encoding=encoding) as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
    
    def export_csv(self, csv_path: str, sheet: Optional[str] = None,
                   delimiter: str = ',',
                   encoding: str = 'utf-8') -> None:
        """
        Export to CSV file
        
        Args:
            csv_path: Output CSV file path
            sheet: Source worksheet name
            delimiter: Delimiter
            encoding: Encoding
        """
        import csv
        
        ws = self.get_sheet(sheet)
        
        with open(csv_path, 'w', encoding=encoding, newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
    
    @property
    def workbook(self) -> Optional[WorkbookType]:
        """Get current workbook"""
        return self._document
    
    @property
    def sheetnames(self) -> List[str]:
        """Get all worksheet names"""
        if not self._document:
            return []
        return self._document.sheetnames
