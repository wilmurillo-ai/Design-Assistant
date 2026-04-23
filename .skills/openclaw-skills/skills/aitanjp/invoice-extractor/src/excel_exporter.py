"""
Excel导出模块
将发票信息导出到Excel文件
"""

import pandas as pd
from pathlib import Path
from typing import List, Optional
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from invoice_model import InvoiceInfo


class ExcelExporter:
    """发票信息Excel导出器"""

    def __init__(self):
        self.default_columns = [
            "发票代码",
            "发票号码",
            "开票日期",
            "发票类型",
            "购买方名称",
            "购买方税号",
            "购买方地址电话",
            "购买方开户行",
            "销售方名称",
            "销售方税号",
            "销售方地址电话",
            "销售方开户行",
            "合计金额",
            "合计税额",
            "价税合计",
            "商品明细摘要",
            "备注",
            "复核人",
            "收款人",
            "开票人",
            "源文件",
            "提取时间",
        ]

    def export_to_excel(
        self,
        invoices: List[InvoiceInfo],
        output_path: str,
        sheet_name: str = "发票信息",
        include_items: bool = False
    ) -> bool:
        """
        将发票信息导出到Excel

        Args:
            invoices: 发票信息列表
            output_path: 输出Excel文件路径
            sheet_name: 工作表名称
            include_items: 是否包含商品明细

        Returns:
            导出成功返回True
        """
        if not invoices:
            print("[FAIL] 没有发票信息可导出")
            return False

        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # 准备数据
            data = []
            for invoice in invoices:
                row = invoice.to_dict()
                # 添加商品明细摘要
                row["商品明细摘要"] = invoice.get_items_summary()
                data.append(row)

            # 创建DataFrame
            df = pd.DataFrame(data)

            # 确保列顺序
            columns = [col for col in self.default_columns if col in df.columns]
            df = df[columns]

            # 导出到Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # 格式化Excel
                self._format_worksheet(worksheet, len(df))

                # 如果包含商品明细，添加明细工作表
                if include_items:
                    self._add_items_sheet(writer, invoices)

            print(f"[OK] 成功导出 {len(invoices)} 条发票信息到: {output_path}")
            return True

        except Exception as e:
            print(f"[FAIL] Excel导出失败: {e}")
            return False

    def _format_worksheet(self, worksheet, row_count: int):
        """格式化工作表"""
        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        cell_font = Font(name='微软雅黑', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        number_alignment = Alignment(horizontal='right', vertical='center')

        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 格式化表头
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 格式化数据行
        for row in worksheet.iter_rows(min_row=2, max_row=row_count + 1):
            for cell in row:
                cell.font = cell_font
                cell.border = border

                # 根据列名设置对齐方式
                col_name = worksheet.cell(1, cell.column).value
                if col_name in ['合计金额', '合计税额', '价税合计']:
                    cell.alignment = number_alignment
                    # 格式化为货币格式（使用Excel内置的货币格式）
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = cell_alignment

        # 调整列宽
        self._adjust_column_widths(worksheet)

        # 冻结首行
        worksheet.freeze_panes = 'A2'

    def _adjust_column_widths(self, worksheet):
        """自动调整列宽"""
        column_widths = {
            "发票代码": 15,
            "发票号码": 15,
            "开票日期": 12,
            "发票类型": 12,
            "购买方名称": 30,
            "购买方税号": 20,
            "购买方地址电话": 35,
            "购买方开户行": 35,
            "销售方名称": 30,
            "销售方税号": 20,
            "销售方地址电话": 35,
            "销售方开户行": 35,
            "合计金额": 12,
            "合计税额": 12,
            "价税合计": 12,
            "商品明细摘要": 40,
            "备注": 30,
            "复核人": 10,
            "收款人": 10,
            "开票人": 10,
            "源文件": 40,
            "提取时间": 18,
        }

        for col_idx, cell in enumerate(worksheet[1], 1):
            col_name = cell.value
            if col_name in column_widths:
                worksheet.column_dimensions[cell.column_letter].width = column_widths[col_name]
            else:
                # 默认宽度
                worksheet.column_dimensions[cell.column_letter].width = 15

    def _add_items_sheet(self, writer, invoices: List[InvoiceInfo]):
        """添加商品明细工作表"""
        items_data = []

        for invoice in invoices:
            for item in invoice.items:
                items_data.append({
                    "发票代码": invoice.invoice_code,
                    "发票号码": invoice.invoice_number,
                    "开票日期": invoice.invoice_date,
                    "商品名称": item.name,
                    "规格型号": item.specification,
                    "单位": item.unit,
                    "数量": item.quantity,
                    "单价": item.unit_price,
                    "金额": item.amount,
                    "税率": item.tax_rate,
                    "税额": item.tax_amount,
                })

        if items_data:
            df_items = pd.DataFrame(items_data)
            df_items.to_excel(writer, sheet_name="商品明细", index=False)

            # 格式化商品明细表
            worksheet = writer.sheets["商品明细"]
            self._format_worksheet(worksheet, len(df_items))

    def export_summary(
        self,
        invoices: List[InvoiceInfo],
        output_path: str
    ) -> bool:
        """
        导出发票汇总统计

        Args:
            invoices: 发票信息列表
            output_path: 输出文件路径

        Returns:
            导出成功返回True
        """
        try:
            # 计算统计数据
            total_count = len(invoices)
            total_amount = sum(inv.total_amount for inv in invoices)
            total_tax = sum(inv.total_tax_amount for inv in invoices)
            total_with_tax = sum(inv.total_amount_with_tax for inv in invoices)

            # 按销售方统计
            seller_stats = {}
            for inv in invoices:
                seller = inv.seller_name or "未知"
                if seller not in seller_stats:
                    seller_stats[seller] = {
                        "发票数量": 0,
                        "合计金额": 0.0,
                        "合计税额": 0.0,
                        "价税合计": 0.0,
                    }
                seller_stats[seller]["发票数量"] += 1
                seller_stats[seller]["合计金额"] += inv.total_amount
                seller_stats[seller]["合计税额"] += inv.total_tax_amount
                seller_stats[seller]["价税合计"] += inv.total_amount_with_tax

            # 创建汇总数据
            summary_data = [
                ["统计项目", "数值"],
                ["发票总数", total_count],
                ["合计金额", total_amount],
                ["合计税额", total_tax],
                ["价税合计", total_with_tax],
                ["", ""],
                ["销售方", "发票数量", "合计金额", "合计税额", "价税合计"],
            ]

            for seller, stats in seller_stats.items():
                summary_data.append([
                    seller,
                    stats["发票数量"],
                    stats["合计金额"],
                    stats["合计税额"],
                    stats["价税合计"],
                ])

            # 导出到Excel
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(output_path, sheet_name="汇总统计", index=False, header=False)

            # 格式化
            workbook = load_workbook(output_path)
            worksheet = workbook["汇总统计"]

            # 设置标题样式
            title_font = Font(name='微软雅黑', size=12, bold=True)
            title_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')

            for cell in worksheet[1]:
                cell.font = title_font
                cell.fill = title_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 设置数值格式
            for row in worksheet.iter_rows(min_row=2, max_row=5):
                if len(row) > 1:
                    row[1].number_format = '#,##0.00'

            workbook.save(output_path)

            print(f"[OK] 成功导出汇总统计到: {output_path}")
            return True

        except Exception as e:
            print(f"[FAIL] 汇总统计导出失败: {e}")
            return False


def export_invoices(
    invoices: List[InvoiceInfo],
    output_dir: str = "output",
    filename_prefix: str = "发票信息"
) -> bool:
    """
    便捷函数：导出发票信息到Excel

    Args:
        invoices: 发票信息列表
        output_dir: 输出目录
        filename_prefix: 文件名前缀

    Returns:
        导出成功返回True
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 导出主表
    main_file = output_dir / f"{filename_prefix}_{timestamp}.xlsx"
    exporter = ExcelExporter()

    success = exporter.export_to_excel(
        invoices,
        str(main_file),
        include_items=True
    )

    if success:
        print(f"\n导出文件列表:")
        print(f"  - 主表: {main_file}")

    return success
