"""
export_quotation.py - 报价导出脚本
生成与 wood_quotation_template.xlsx 格式一致的报价 Excel 文件
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl 未安装，执行: pip install openpyxl")
    sys.exit(1)


def create_quotation_workbook(quotation: dict, output_path: str):
    """
    创建报价 Excel 工作簿
    对齐 wood_quotation_template.xlsx 的格式和列结构
    """
    wb = openpyxl.Workbook()

    # ---------- Sheet1: 报价明细 ----------
    ws = wb.active
    ws.title = "报价明细"

    # 样式定义
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")  # 蓝色表头
    subheader_fill = PatternFill("solid", fgColor="4472C4")
    money_font = Font(bold=True, size=14, color="C00000")  # 红色金额
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def style_header(cell, text):
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    def style_cell_simple(row, col, value, bold=False, align="left"):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.font = Font(bold=bold)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = thin_border
        return cell

    # 第1行：标题
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "黄杰木工厂 · 定制木作报价单"
    title_cell.font = Font(bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="D9E2F3")
    ws.row_dimensions[1].height = 36

    # 第2行：图纸尺寸
    ws.merge_cells("A2:F2")
    dim_cell = ws["A2"]
    dim_cell.value = f"图纸尺寸：{quotation.get('图纸尺寸', 'N/A')}（宽{quotation['宽_w']} × 深{quotation['深_d']} × 高{quotation['高_h']}mm）"
    dim_cell.font = Font(size=11)
    dim_cell.alignment = Alignment(horizontal="center")
    dim_cell.fill = PatternFill("solid", fgColor="EAF2FF")
    ws.row_dimensions[2].height = 24

    # 第3行：空行
    ws.row_dimensions[3].height = 6

    # ---------- 材料成本 ----------
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    style_header(ws.cell(row=row, column=1), "一、材料成本")
    row += 1

    # 材料表头
    headers = ["材料名称", "价格档位", "单价(元)", "用量", "金额(元)", "备注"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=row, column=col), h)
    row += 1

    for m in quotation.get("材料明细", []):
        style_cell_simple(row, 1, m["材料"])
        style_cell_simple(row, 2, m.get("规格", m.get("价格档位", "")))
        style_cell_simple(row, 3, m["单价"], align="right")
        style_cell_simple(row, 4, m["用量"], align="right")
        style_cell_simple(row, 5, m["金额(元)"], align="right")
        style_cell_simple(row, 6, "")
        row += 1

    # 材料合计
    ws.merge_cells(f"A{row}:D{row}")
    style_cell_simple(row, 1, "材料成本合计", bold=True)
    ws.cell(row=row, column=5).value = quotation.get("材料成本", 0)
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=6).border = thin_border
    row += 2

    # ---------- 工艺成本 ----------
    ws.merge_cells(f"A{row}:F{row}")
    style_header(ws.cell(row=row, column=1), "二、工艺成本")
    row += 1

    headers2 = ["工艺名称", "单价(元)", "备注", "", "", ""]
    for col, h in enumerate(headers2[:3], 1):
        style_header(ws.cell(row=row, column=col), h)
    row += 1

    for p in quotation.get("工艺明细", []):
        style_cell_simple(row, 1, p["工艺"])
        style_cell_simple(row, 2, p["单价(元)"], align="right")
        style_cell_simple(row, 3, p.get("备注", "含损耗"))
        row += 1

    ws.merge_cells(f"A{row}:B{row}")
    style_cell_simple(row, 1, "工艺成本合计", bold=True)
    ws.cell(row=row, column=3).value = quotation.get("工艺成本", 0)
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=3).border = thin_border
    row += 2

    # ---------- 损耗 ----------
    ws.merge_cells(f"A{row}:F{row}")
    style_header(ws.cell(row=row, column=1), "三、损耗核算")
    row += 1

    style_cell_simple(row, 1, "综合损耗率")
    ws.cell(row=row, column=2).value = quotation["损耗"]["综合损耗率"]
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=2).border = thin_border
    row += 1

    style_cell_simple(row, 1, "损耗金额")
    ws.cell(row=row, column=2).value = quotation["损耗"]["损耗金额(元)"]
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=2).border = thin_border
    row += 2

    # ---------- 报价汇总 ----------
    ws.merge_cells(f"A{row}:F{row}")
    style_header(ws.cell(row=row, column=1), "四、报价汇总")
    row += 1

    summary = [
        ("板材用量", f"{quotation['板材用量']['含损耗张数']}张（利用率 {quotation['板材用量']['利用率']}）"),
        ("总成本", f"{quotation['总成本']}元"),
        ("利润率", quotation["利润率"]),
    ]
    for label, val in summary:
        style_cell_simple(row, 1, label)
        ws.merge_cells(f"B{row}:C{row}")
        ws.cell(row=row, column=2).value = val
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=2).border = thin_border
        row += 1

    # 报价金额（重点高亮）
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    price_cell = ws.cell(row=row, column=1, value="💰 报价金额（含利润）")
    price_cell.font = Font(bold=True, size=14, color="FFFFFF")
    price_cell.fill = PatternFill("solid", fgColor="C00000")
    price_cell.alignment = Alignment(horizontal="right", vertical="center")
    price_cell.border = thin_border

    ws.merge_cells(f"D{row}:F{row}")
    ws.cell(row=row, column=4).value = f"{quotation['报价金额']}元"
    ws.cell(row=row, column=4).font = Font(bold=True, size=16, color="C00000")
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="FCE4D6")
    ws.cell(row=row, column=4).border = thin_border
    ws.row_dimensions[row].height = 32
    row += 2

    # 底部说明
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1).value = "⚠️ 本报价单仅供内部参考，采购数据不对外泄露。如有疑问请联系：黄杰（0592-7153916）"
    ws.cell(row=row, column=1).font = Font(size=9, color="808080", italic=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

    # 列宽
    col_widths = [20, 16, 14, 10, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- Sheet2: 数据源 ----------
    ws2 = wb.create_sheet("数据源")
    ws2["A1"] = "来源"
    ws2["B1"] = "内容"
    ws2["A2"] = "核算模板"
    ws2["B2"] = "./data/wood_quotation_template.xlsx"
    ws2["A3"] = "采购数据"
    ws2["B3"] = "./data/purchase_cost_data.xlsx"

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python export_quotation.py <报价JSON> <输出路径.xlsx>")
        print("示例: python export_quotation.py '{\"status\":\"success\",...}' ./output.xlsx")
        sys.exit(1)

    quotation = json.loads(sys.argv[1])
    output_path = sys.argv[2]

    if quotation.get("status") == "error":
        print(f"❌ 导出失败：{quotation.get('message')}")
        sys.exit(1)

    path = create_quotation_workbook(quotation, output_path)
    print(f"✅ 报价单已导出：{path}")
