#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 数据导入主脚本

功能：根据配置文件从源数据导入到目标模板
支持：单文件/多文件/目录批量、.xlsx/.csv、多层合并表头、自动表头检测、
      字段映射、数据转换、验证、增量更新、合并单元格保护、自动备份
"""

import argparse
import copy
import csv
import glob
import json
import sys
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Any, List, Dict, Optional

try:
    import python_calamine as pc

    CALAMINE_AVAILABLE = True
except ImportError:
    CALAMINE_AVAILABLE = False

# 添加脚本目录到 Python 路径
script_dir = Path(__file__).parent
sys.path.insert(0, str(script_dir))

from config_parser import parse_config, validate_config  # pylint: disable=import-error
from data_mapper import apply_transforms, validate_field  # pylint: disable=import-error
from data_matcher import find_matching_row  # pylint: disable=import-error
from error_handler import (  # pylint: disable=import-error
    ErrorLog,
    backup_file,
    log_error,
)
from merged_cells_handler import (  # pylint: disable=import-error
    get_merged_ranges,
    handle_merged_cells,
    is_cell_merged,
)
from openpyxl import load_workbook, Workbook
from validator import ValidationReport  # pylint: disable=import-error
from worksheet_analyzer import (  # pylint: disable=import-error
    analyze_worksheet,
    detect_header_row,
    find_data_range,
    read_headers_with_merged,
)


class ExcelImporter:
    """Excel 数据导入器"""

    def __init__(
        self, config_path, dry_run=False, verbose=False, no_backup=False
    ) -> None:
        self.config_path = Path(config_path)
        self.config = parse_config(config_path)
        validate_config(self.config)
        self.dry_run = dry_run
        self.verbose = verbose
        self.no_backup = no_backup

        self._create_directories()

        self.source_wb = None
        self.target_wb = None
        self.source_ws = None
        self.target_ws = None
        self.source_is_csv = False  # CSV 模式标记
        self.csv_data: Optional[List[List[str]]] = None  # CSV 数据缓存
        self.error_log = ErrorLog()
        self.validation_report = ValidationReport()
        self.stats = {"added": 0, "updated": 0, "skipped": 0}
        self.legacy_merged_ranges: List[str] = []  # .xls 文件的合并单元格信息

    def _create_directories(self) -> None:
        output_path = Path(self.config["target"]["output_path"]).parent
        output_path.mkdir(parents=True, exist_ok=True)

        if self.config["error_handling"].get("backup", False):
            backup_path = Path(
                self.config["error_handling"].get("backup_path", "backup/")
            )
            backup_path.mkdir(parents=True, exist_ok=True)

        log_path = Path(
            self.config["error_handling"].get("error_log_path", "logs/")
        ).parent
        log_path.mkdir(parents=True, exist_ok=True)

    def import_data(self) -> Any:
        """执行数据导入（支持单文件/多文件/目录批量）"""
        if self.dry_run:
            print(f"[DRY RUN] 预览模式：只分析，不写入")
        print(f"开始执行导入任务: {self.config['task_name']}")
        print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("-" * 60)

        try:
            # 展开源文件列表
            source_files = self._expand_sources()
            if not source_files:
                print("❌ 没有找到有效的源文件")
                return False

            # 加载目标
            if (
                self.config["error_handling"].get("backup", False)
                and not self.no_backup
            ):
                self._backup_target_file()

            self._load_workbooks()

            source_structure = self._analyze_source()
            target_structure = self._analyze_target()

            merged_ranges = self._handle_merged_cells()

            # 多文件模式：合并所有源数据后一次导入
            all_source_data = self._read_all_sources(source_files)
            self.validation_report.total = len(all_source_data)

            print(
                f"✓ 共读取 {len(all_source_data)} 条有效源数据"
                f"（来自 {len(source_files)} 个文件）"
            )

            self._import_records(
                all_source_data, source_structure, target_structure, merged_ranges
            )

            if not self.dry_run:
                self._save_result()
                self._generate_report()

            print("-" * 60)
            if self.dry_run:
                print("[DRY RUN] 预览导入结果:")
                print(f"  新增: {self.stats['added']} 条")
                print(f"  更新: {self.stats['updated']} 条")
                print(f"  跳过: {self.stats['skipped']} 条")
                if self.validation_report.failed > 0:
                    fail_reasons = [e.get("message", str(e)) for e in self.error_log.errors[:3]]
                    reasons_str = "; ".join(fail_reasons) if fail_reasons else "未知"
                    print(
                        f"  失败: {self.validation_report.failed} 条 (原因: {reasons_str})"
                    )
                print("[DRY RUN] 未执行实际写入")
            else:
                print("✅ 导入完成！")
                print(f"总记录数: {self.validation_report.total}")
                print(f"成功: {self.validation_report.successful}")
                print(f"失败: {self.validation_report.failed}")
                print(f"  ├ 新增: {self.stats['added']}")
                print(f"  ├ 更新: {self.stats['updated']}")
                print(f"  └ 跳过: {self.stats['skipped']}")

            return True

        except Exception as e:
            print(f"❌ 导入失败: {str(e)}")
            log_error(str(e), self.error_log)
            return False

    def _expand_sources(self) -> List[Dict[str, str]]:
        """展开源配置，返回文件列表（支持单文件/目录/多源）

        Returns:
            [{"file_path": ..., "sheet_name": ..., "header_row": ..., "key_field": ...}, ...]
        """
        source_config = self.config["source"]
        base_dir = self.config_path.parent
        sources = []

        if "directory" in source_config or source_config.get("type") == "directory":
            # 目录批量模式
            dir_path = source_config.get("directory", "")
            if not Path(dir_path).is_absolute():
                dir_path = str(base_dir / dir_path)
            pattern = source_config.get("pattern", "*.xlsx")
            files = sorted(glob.glob(str(Path(dir_path) / pattern)))
            # 也支持 CSV
            if "*.xlsx" in pattern:
                csv_files = sorted(
                    glob.glob(str(Path(dir_path) / pattern.replace(".xlsx", ".csv")))
                )
                files.extend(csv_files)
            print(f"📁 目录模式: 在 {dir_path} 中找到 {len(files)} 个文件")

            for f in files:
                sources.append(
                    {
                        "file_path": f,
                        "sheet_name": source_config.get("sheet_name", ""),
                        "header_row": source_config.get("header_row", 1),
                        "key_field": source_config.get("key_field", ""),
                    }
                )

        elif "sources" in self.config:
            # 多源显式配置
            for src in self.config["sources"]:
                fp = src.get("file_path", "")
                if not Path(fp).is_absolute():
                    fp = str(base_dir / fp)
                if Path(fp).exists():
                    sources.append(
                        {
                            "file_path": fp,
                            "sheet_name": src.get("sheet_name", ""),
                            "header_row": src.get("header_row", 1),
                            "key_field": src.get("key_field", ""),
                        }
                    )
                else:
                    print(f"⚠ 跳过不存在的文件: {fp}")

        else:
            # 单文件模式（默认）
            fp = source_config.get("file_path", "")
            if not Path(fp).is_absolute():
                fp = str(base_dir / fp)
            sources.append(
                {
                    "file_path": fp,
                    "sheet_name": source_config.get("sheet_name", ""),
                    "header_row": source_config.get("header_row", 1),
                    "key_field": source_config.get("key_field", ""),
                }
            )

        return sources

    def _read_all_sources(self, source_files: List[Dict]) -> List[Dict]:
        """从所有源文件读取数据，合并去重

        Returns:
            [{"row_num": int, "data": {field: value}}, ...]
        """
        base_source_config = self.config["source"]
        key_field = base_source_config.get("key_field", "")

        # 收集所有 field_mappings（多源模式下每个源可能有自己的映射）
        global_mappings = self.config.get("field_mappings", [])

        all_data: List[Dict] = []
        processed_keys = set()

        for idx, src in enumerate(source_files, 1):
            fp = src["file_path"]
            print(f"  [{idx}/{len(source_files)}] 读取: {Path(fp).name}")

            try:
                is_csv = fp.lower().endswith(".csv")
                is_legacy_xls = fp.lower().endswith((".xls", ".xlsm")) and not fp.lower().endswith(".xlsx")

                if is_csv:
                    headers, rows = self._load_csv(fp, src.get("header_row", 1))
                elif is_legacy_xls:
                    if not CALAMINE_AVAILABLE:
                        print(f"    ⚠️ 跳过 .xls/.xlsm 文件: python-calamine 未安装")
                        print(f"       请运行: pip install python-calamine")
                        continue
                    sheet_name = src.get("sheet_name") or base_source_config.get(
                        "sheet_name", ""
                    )
                    headers, rows = self._load_legacy_xls(
                        fp, sheet_name, src.get("header_row", 1)
                    )
                else:
                    wb = load_workbook(fp)
                    sheet_name = src.get("sheet_name") or base_source_config.get(
                        "sheet_name", ""
                    )
                    if not sheet_name and wb.sheetnames:
                        sheet_name = wb.sheetnames[0]
                    ws = wb[sheet_name]

                    # 自动检测表头
                    header_row = src.get("header_row", 1)
                    if header_row == "auto" or base_source_config.get(
                        "auto_detect_header"
                    ):
                        det = detect_header_row(ws)
                        header_row = det["header_row"]
                        print(
                            f"    自动检测表头: 第{header_row}行 (置信度{det['confidence']:.0%})"
                        )

                    first_row = self._detect_first_header_row(
                        ws, header_row, sheet_name
                    )
                    struct = analyze_worksheet(ws, header_row, first_row)
                    headers = struct["headers"]
                    rows = self._read_xlsx_rows(ws, struct)

                # 使用全局映射（多源模式）或默认映射
                mappings = global_mappings
                if "sources" in self.config:
                    src_idx = idx - 1
                    if src_idx < len(self.config["sources"]):
                        src_mappings = self.config["sources"][src_idx].get(
                            "field_mappings", []
                        )
                        if src_mappings:
                            mappings = src_mappings

                count = 0
                for row_data in rows:
                    if key_field and key_field not in row_data:
                        continue
                    key_val = row_data.get(key_field)
                    if not key_val or str(key_val).strip() == "":
                        self.error_log.add_warning(f"关键字段为空，跳过")
                        self.stats["skipped"] += 1
                        continue

                    key_str = str(key_val).strip()
                    if key_str in processed_keys:
                        continue  # 去重
                    processed_keys.add(key_str)

                    all_data.append({"row_num": count, "data": row_data, "source": fp})
                    count += 1

                print(f"    读取 {count} 条有效数据")
                if not is_csv and not is_legacy_xls:
                    wb.close()

            except Exception as e:
                print(f"    ❌ 读取失败: {e}")
                continue

        return all_data

    def _load_csv(self, file_path: str, header_row: int = 1) -> tuple:
        """加载 CSV 文件

        Returns:
            (headers_dict: {列名: 列号}, rows: [{列名: 值}, ...])
        """
        # 尝试多种编码
        encodings = ["utf-8", "utf-8-sig", "gbk", "gb2312", "latin-1"]
        content = None
        used_encoding = "utf-8"
        for enc in encodings:
            try:
                with open(file_path, "r", encoding=enc) as f:
                    content = f.read()
                used_encoding = enc
                break
            except (UnicodeDecodeError, UnicodeError):
                continue

        if content is None:
            raise ValueError(f"无法解码文件: {file_path}")

        # 自动检测分隔符
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(content[:2048])
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","

        reader = csv.reader(StringIO(content), delimiter=delimiter)
        rows_raw = list(reader)

        if header_row < 1 or header_row > len(rows_raw):
            header_row = 1

        # 表头（清除 BOM 和首尾空白）
        header_line = rows_raw[header_row - 1]
        headers = {}
        for col_idx, h in enumerate(header_line):
            h_clean = h.strip().lstrip("\ufeff").lstrip("\ufeff")
            if h_clean:
                headers[h_clean] = col_idx

        # 数据行
        data_rows = []
        for row in rows_raw[header_row:]:
            row_data = {}
            for name, col_idx in headers.items():
                if col_idx < len(row):
                    row_data[name] = row[col_idx].strip() if row[col_idx] else None
                else:
                    row_data[name] = None
            # 跳过全空行
            if any(v for v in row_data.values()):
                data_rows.append(row_data)

        print(
            f"    CSV: {len(headers)}列, {len(data_rows)}行数据, 分隔符='{delimiter}'"
        )
        return headers, data_rows

    def _load_legacy_xls(
        self, file_path: str, sheet_name: str, header_row: int = 1
    ) -> tuple:
        """加载 .xls/.xlsm 旧格式文件 (BIFF格式)

        Returns:
            (headers_dict: {列名: 列号}, rows: [{列名: 值}, ...])
        """
        if not CALAMINE_AVAILABLE:
            raise ImportError(
                "python-calamine 未安装，无法读取 .xls/.xlsm 文件。"
                "请运行: pip install python-calamine"
            )

        import python_calamine as pc

        wb = pc.load_workbook(file_path)

        if sheet_name and sheet_name in wb.sheet_names:
            ws = wb.get_sheet_by_name(sheet_name)
        else:
            ws = wb.get_sheet_by_index(0)

        self.legacy_merged_ranges = ws.merged_cell_ranges

        all_rows = ws.to_python()
        if not all_rows or header_row < 1 or header_row > len(all_rows):
            return {}, []

        header_row_data = all_rows[header_row - 1]
        headers = {}
        for col_idx, val in enumerate(header_row_data):
            if val is not None:
                h_clean = str(val).strip()
                if h_clean:
                    headers[h_clean] = col_idx

        data_rows = []
        for row_data_raw in all_rows[header_row:]:
            row_data = {}
            for name, col_idx in headers.items():
                if col_idx < len(row_data_raw):
                    cell_value = row_data_raw[col_idx]
                    if cell_value is not None:
                        if isinstance(cell_value, float) and cell_value == int(cell_value):
                            cell_value = int(cell_value)
                        row_data[name] = cell_value
                    else:
                        row_data[name] = None
                else:
                    row_data[name] = None

            if any(v is not None and str(v).strip() != "" for v in row_data.values()):
                data_rows.append(row_data)

        print(
            f"    .xls: {len(headers)}列, {len(data_rows)}行数据, 合并单元格:{len(self.legacy_merged_ranges)}个"
        )
        return headers, data_rows

    def _read_xlsx_rows(self, ws, structure: Dict) -> List[Dict]:
        """从 xlsx 工作表读取数据行"""
        header_row = structure["header_row"]
        max_col = structure["max_col"]
        headers = structure["headers"]

        rows = []
        actual_last = self._find_actual_last_data_row(ws, header_row)

        for row in range(header_row + 1, actual_last + 1):
            row_data = {}
            for name, col in headers.items():
                cell = ws.cell(row=row, column=col)
                row_data[name] = cell.value if cell is not None else None
            if any(v is not None and str(v).strip() != "" for v in row_data.values()):
                rows.append(row_data)

        return rows

    def _backup_target_file(self) -> None:
        """备份目标文件"""
        target_path = Path(self.config["target"]["file_path"])
        backup_path = Path(self.config["error_handling"].get("backup_path", "backup/"))

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{target_path.stem}_备份_{timestamp}{target_path.suffix}"
        backup_file_path = backup_path / backup_filename

        backup_file(target_path, backup_file_path)
        print(f"✓ 已备份目标文件: {backup_file_path}")

    def _load_workbooks(self) -> None:
        """加载目标工作簿"""
        target_path = Path(self.config["target"]["file_path"])
        if not target_path.exists():
            self._create_target_template(target_path)
            self.target_wb.save(target_path)
            self.target_wb = load_workbook(target_path)
            self.target_ws = self.target_wb[self.config["target"]["sheet_name"]]
        else:
            self.target_wb = load_workbook(target_path)
            self.target_ws = self.target_wb[self.config["target"]["sheet_name"]]
        print(f"✓ 已加载目标模板: {target_path}")

    def _create_target_template(self, target_path: Path) -> None:
        """目标文件不存在时自动创建模板"""
        from openpyxl import Workbook

        self.target_wb = Workbook()

        # 重命名默认 sheet 为目标 sheet 名
        target_sheet_name = self.config["target"]["sheet_name"]
        self.target_wb.active.title = target_sheet_name
        self.target_ws = self.target_wb.active

        # 根据 field_mappings 创建表头
        target_header_row = self.config["target"].get("header_row", 1)
        field_mappings = self.config.get("field_mappings", [])
        for col_idx, mapping in enumerate(field_mappings, 1):
            self.target_ws.cell(
                row=target_header_row,
                column=col_idx,
                value=mapping.get("target", mapping.get("source", "")),
            )

        # 额外目标列
        extra_headers = self.config["target"].get("extra_headers", [])
        for col_idx, header in enumerate(extra_headers, len(field_mappings) + 1):
            self.target_ws.cell(row=target_header_row, column=col_idx, value=header)

        if "data_start_row" not in self.config["target"]:
            self.config["target"]["data_start_row"] = target_header_row + 1

        print(f"✓ 目标文件不存在，已自动创建模板: {target_path}")

    def _analyze_target(self) -> Any:
        """分析目标模板表结构"""
        header_row = self.config["target"]["header_row"]
        # 自动检测
        if header_row == "auto" or self.config["target"].get("auto_detect_header"):
            print("🔍 自动检测目标表头位置...")
            det = detect_header_row(self.target_ws)
            header_row = det["header_row"]
            self.config["target"]["header_row"] = header_row
            self.config["target"]["data_start_row"] = det["data_start_row"]
            print(f"✓ 自动检测: 表头第{header_row}行, 置信度{det['confidence']:.0%}")
        first_header_row = self._detect_first_header_row(
            self.target_ws, header_row, self.config["target"].get("sheet_name", "")
        )
        structure = analyze_worksheet(self.target_ws, header_row, first_header_row)
        print(f"✓ 目标模板: {structure['total_rows']}行, {structure['max_col']}列")
        return structure

    def _analyze_source(self) -> Any:
        """分析源数据表结构（仅用于单文件模式，多文件/目录/CSV模式下跳过）"""
        source_config = self.config["source"]
        is_multi = "sources" in self.config
        is_dir = (
            source_config.get("type") == "directory" or "directory" in source_config
        )

        fp = source_config.get("file_path", "")
        is_csv = fp.lower().endswith(".csv")

        if is_multi or is_dir or is_csv:
            # 多源/目录/CSV模式：使用默认结构，实际分析在 _read_all_sources 中进行
            fm = self.config.get("field_mappings", [])
            return {
                "total_rows": 0,
                "max_col": len(fm) if fm else 1,
                "header_row": source_config.get("header_row", 1),
                "headers": {m["source"]: i + 1 for i, m in enumerate(fm)},
                "first_header_row": source_config.get("header_row", 1),
            }

        header_row = source_config.get("header_row", 1)
        if header_row == "auto" or source_config.get("auto_detect_header"):
            if fp.lower().endswith(".csv"):
                # CSV auto header: 始终第1行
                header_row = 1
            else:
                wb = load_workbook(fp)
                sn = source_config.get("sheet_name", wb.sheetnames[0])
                det = detect_header_row(wb[sn])
                header_row = det["header_row"]
                wb.close()

        source_headers = {}
        max_col = 1

        if fp.lower().endswith(".csv"):
            headers_dict, _ = self._load_csv(fp, header_row)
            source_headers = {name: col + 1 for name, col in headers_dict.items()}
            max_col = len(source_headers)
        else:
            wb = load_workbook(fp)
            sn = source_config.get("sheet_name", wb.sheetnames[0])
            ws = wb[sn]
            first = self._detect_first_header_row(ws, header_row, sn)
            struct = analyze_worksheet(ws, header_row, first)
            source_headers = struct["headers"]
            max_col = struct["max_col"]
            wb.close()

        print(f"✓ 源数据: {len(source_headers)}列, 表头行{header_row}")
        return {
            "total_rows": 0,
            "max_col": max_col,
            "header_row": header_row,
            "headers": source_headers,
            "first_header_row": header_row,
        }

    def _detect_first_header_row(self, ws, header_row, sheet_name="") -> int:
        """自动检测第一行表头（多层表头时向上查找）

        Args:
            ws: 工作表
            header_row: 已知的主表头行号
            sheet_name: 工作表名（仅用于日志）

        Returns:
            int: 第一行表头的行号
        """
        # 检查上方行是否有合并单元格延伸到 header_row
        for mr in ws.merged_cells.ranges:
            if mr.max_row == header_row and mr.min_row < header_row:
                # 向上延伸的第一行
                return mr.min_row
        return header_row

    def _handle_merged_cells(self) -> Any:
        """处理合并单元格"""
        merged_ranges = get_merged_ranges(self.target_ws)
        print(f"✓ 发现 {len(merged_ranges)} 个合并单元格区域")
        return handle_merged_cells(self.target_ws, merged_ranges)

    def _find_actual_last_data_row(self, ws, header_row) -> int:
        """找到工作表中实际有数据的最后一行（跳过尾部空行）

        Args:
            ws: 工作表
            header_row: 表头行号

        Returns:
            int: 最后一行行号
        """
        actual_last = header_row
        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip() != "":
                    actual_last = row
                    break
        return actual_last

    def _is_empty_row(self, ws, row, max_col) -> bool:
        """检查一行是否为空"""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip() != "":
                return False
        return True

    def _find_next_available_row(self, ws, start_row, max_col) -> int:
        """从 start_row 开始找到第一个空行（跳过已有数据的行）

        Args:
            ws: 工作表
            start_row: 搜索起始行
            max_col: 最大列数

        Returns:
            int: 第一个空行的行号
        """
        row = start_row
        while row <= ws.max_row + 100:  # 安全上限
            if self._is_empty_row(ws, row, max_col):
                return row
            row += 1
        return row

    def _import_records(
        self, all_source_data, source_structure, target_structure, merged_ranges
    ) -> None:
        """将所有源数据导入到目标模板"""
        print("-" * 60)
        print("开始导入数据...")

        target_header_row = self.config["target"]["header_row"]
        target_data_start_row = self.config["target"].get(
            "data_start_row", target_header_row + 1
        )
        key_field = self.config["source"].get("key_field", "")
        field_mappings = self.config.get("field_mappings", [])

        if not field_mappings:
            print("⚠ 警告: field_mappings 为空，无法导入")
            return

        target_headers = read_headers_with_merged(
            self.target_ws,
            target_header_row,
            target_structure["max_col"],
            target_structure.get("first_header_row", target_header_row),
        )

        # 构建列映射 (key=source_field_name, value=mapping_dict)
        column_mapping = {}
        target_key_field = key_field
        for mapping in field_mappings:
            source_field = mapping["source"]
            target_field = mapping["target"]
            if target_field in target_headers:
                column_mapping[source_field] = {
                    "target_col": target_headers[target_field],
                    "required": mapping.get("required", False),
                    "default": mapping.get("default", None),
                    "transform": mapping.get("transform", None),
                    "transform_params": mapping.get("transform_params", {}),
                    "validate": mapping.get("validate", None),
                    "validate_params": mapping.get("validate_params", {}),
                }
                if source_field == key_field:
                    target_key_field = target_field

        if key_field and target_key_field not in target_headers:
            print(f"⚠ 关键字段 '{key_field}' 在目标表头中未找到，以追加模式导入")

        actual_target_last = self._find_actual_last_data_row(
            self.target_ws, target_header_row
        )
        search_end_row = max(actual_target_last, target_data_start_row)

        for idx, source_record in enumerate(all_source_data, 1):
            try:
                row_data = source_record["data"]
                target_row = None
                if key_field and target_key_field in target_headers:
                    key_value = row_data.get(key_field)
                    if key_value:
                        target_row = find_matching_row(
                            key_value,
                            self.target_ws,
                            target_headers,
                            target_key_field,
                            target_data_start_row,
                            search_end_row,
                        )

                if target_row is not None:
                    self.stats["updated"] += 1
                else:
                    target_row = self._find_next_available_row(
                        self.target_ws,
                        target_data_start_row,
                        target_structure["max_col"],
                    )
                    self.stats["added"] += 1
                    if target_row > search_end_row:
                        search_end_row = target_row

                self._write_record(row_data, column_mapping, target_row, merged_ranges)
                self.validation_report.successful += 1

                if idx % 10 == 0:
                    print(
                        f"  已处理: {idx}/{len(all_source_data)} (新增:{self.stats['added']} 更新:{self.stats['updated']})"
                    )

            except Exception as e:
                self.validation_report.failed += 1
                src_info = source_record.get("source", "")
                error_msg = f"[{Path(src_info).name if src_info else '?'}] {str(e)}"
                log_error(error_msg, self.error_log)
                if self.config["error_handling"].get("stop_on_error", False):
                    raise

        print(f"✓ 数据导入完成: 共处理 {len(all_source_data)} 条记录")

    def _write_record(
        self, row_data: Dict, column_mapping: Dict, target_row: int, merged_ranges
    ) -> None:
        """将一条记录写入目标行（先全部验证，再一次性写入，失败则不写入任何字段）"""
        write_plan = []
        for source_field, mapping in column_mapping.items():
            source_value = row_data.get(source_field)

            if source_value is None and mapping["required"]:
                if mapping["default"] is not None:
                    source_value = mapping["default"]
                else:
                    raise ValueError(f"必填字段 '{source_field}' 为空且无默认值")

            if source_value is not None and mapping["transform"]:
                source_value = apply_transforms(
                    source_value,
                    mapping["transform"],
                    mapping.get("transform_params", {}),
                )

            if mapping.get("validate"):
                is_valid, error_msg = validate_field(
                    source_value,
                    mapping["validate"],
                    mapping.get("validate_params", {}),
                )
                if not is_valid:
                    raise ValueError(f"字段 '{source_field}' 验证失败: {error_msg}")

            write_plan.append((mapping["target_col"], source_value))

        for col, value in write_plan:
            target_cell = self.target_ws.cell(row=target_row, column=col)
            if is_cell_merged(target_cell, self.target_ws.merged_cells.ranges):
                continue
            # 保留原有样式
            cached_styles = {
                "number_format": target_cell.number_format,
                "font": copy.copy(target_cell.font) if target_cell.font else None,
                "fill": copy.copy(target_cell.fill) if target_cell.fill else None,
                "border": copy.copy(target_cell.border) if target_cell.border else None,
                "alignment": copy.copy(target_cell.alignment) if target_cell.alignment else None,
            }
            target_cell.value = value
            for attr, val in cached_styles.items():
                if val is not None:
                    setattr(target_cell, attr, val)

    def _save_result(self) -> None:
        """保存结果文件"""
        output_path = Path(self.config["target"]["output_path"])
        if self.target_wb is not None:
            self.target_wb.save(output_path)
        print(f"✓ 结果已保存: {output_path}")

    def _generate_report(self) -> None:
        """生成报告"""
        if self.error_log.has_errors():
            log_path = Path(
                self.config["error_handling"].get(
                    "error_log_path", "logs/import_errors.log"
                )
            )
            log_path.parent.mkdir(parents=True, exist_ok=True)
            self.error_log.save(log_path)
            print(f"✓ 错误日志已保存: {log_path}")

        report_path = (
            Path(self.config["target"]["output_path"]).parent / "import_report.json"
        )
        report_data = {
            "task_name": self.config["task_name"],
            "timestamp": datetime.now().isoformat(),
            "summary": {
                "total_records": self.validation_report.total,
                "successful": self.validation_report.successful,
                "failed": self.validation_report.failed,
                "skipped": self.stats["skipped"],
            },
            "details": {
                "added": self.stats["added"],
                "updated": self.stats["updated"],
            },
            "errors": self.error_log.errors[:20],
            "warnings": self.error_log.warnings[:20],
        }

        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)

        print(f"✓ 导入报告已保存: {report_path}")


def main() -> None:
    """主函数"""
    parser = argparse.ArgumentParser(
        description="Excel 数据导入工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python excel_import.py config.yaml
  python excel_import.py config.yaml --dry-run
  python excel_import.py config.yaml --verbose --no-backup
        """,
    )
    parser.add_argument("config_path", help="配置文件路径")
    parser.add_argument(
        "--dry-run", action="store_true", help="只分析不写入，输出预览报告"
    )
    parser.add_argument("--verbose", action="store_true", help="输出每条记录的处理详情")
    parser.add_argument("--no-backup", action="store_true", help="跳过备份步骤")

    args = parser.parse_args()

    config_path = args.config_path

    if not Path(config_path).exists():
        print(f"错误: 配置文件不存在: {config_path}")
        sys.exit(1)

    try:
        importer = ExcelImporter(
            config_path,
            dry_run=args.dry_run,
            verbose=args.verbose,
            no_backup=args.no_backup,
        )
        success = importer.import_data()
        sys.exit(0 if success else 1)

    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback

        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
