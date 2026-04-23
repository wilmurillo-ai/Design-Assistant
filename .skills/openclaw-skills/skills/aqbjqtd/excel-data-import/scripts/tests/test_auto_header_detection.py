#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动表头检测功能测试脚本

测试场景：
1. 简单表头 - 第1行就是列名
2. 有标题行 - 第1行标题，第3行列名
3. 多层说明 - 前3行说明，第4行列名
4. 合并单元格标题 - 第1-2行合并，第3行列名
5. 边缘情况 - 所有行都是数据
"""

import sys
import os
import openpyxl
from openpyxl.styles import Alignment

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from worksheet_analyzer import detect_header_row


def create_test_case_1():
    """测试用例1: 简单表头（第1行就是列名）"""
    print("\n" + "="*80)
    print("测试用例1: 简单表头 - 第1行就是列名")
    print("="*80)

    wb = openpyxl.Workbook()
    ws = wb.active

    # 第1行：直接是表头
    ws.append(['姓名', '学号', '身份证号', '手机号', '班级'])
    # 第2-5行：数据
    ws.append(['张三', '2024001', '110101200001011234', '13800138000', '1班'])
    ws.append(['李四', '2024002', '110101200001022345', '13800138001', '2班'])
    ws.append(['王五', '2024003', '110101200001033456', '13800138002', '1班'])

    result = detect_header_row(ws)

    print(f"检测结果: 第 {result['header_row']} 行")
    print(f"置信度: {result['confidence']*100:.1f}%")
    print(f"原因: {result['reason']}")

    assert result['header_row'] == 1, f"期望第1行，实际第{result['header_row']}行"
    print("✅ 测试通过")

    wb.close()
    return True


def create_test_case_2():
    """测试用例2: 有标题行（第1行标题，第3行列名）"""
    print("\n" + "="*80)
    print("测试用例2: 有标题行 - 第1行标题，第3行列名")
    print("="*80)

    wb = openpyxl.Workbook()
    ws = wb.active

    # 第1行：大标题
    ws.append(['学生信息表'])
    # 第2行：空行
    ws.append([])
    # 第3行：表头
    ws.append(['姓名', '学号', '身份证号', '手机号', '班级'])
    # 第4行：数据
    ws.append(['张三', '2024001', '110101200001011234', '13800138000', '1班'])

    result = detect_header_row(ws)

    print(f"检测结果: 第 {result['header_row']} 行")
    print(f"置信度: {result['confidence']*100:.1f}%")
    print(f"原因: {result['reason']}")

    assert result['header_row'] == 3, f"期望第3行，实际第{result['header_row']}行"
    print("✅ 测试通过")

    wb.close()
    return True


def create_test_case_3():
    """测试用例3: 多层说明（前3行各种说明，第4行列名）"""
    print("\n" + "="*80)
    print("测试用例3: 多层说明 - 前3行说明，第4行列名")
    print("="*80)

    wb = openpyxl.Workbook()
    ws = wb.active

    # 第1行：标题
    ws.append(['2024年度学生成绩表'])
    # 第2行：说明
    ws.append(['制表部门: 教务处', '制表日期: 2024-01-15'])
    # 第3行：更多说明
    ws.append(['说明: 本表包含所有班级的学生信息'])
    # 第4行：表头
    ws.append(['姓名', '学号', '身份证号', '手机号', '班级'])
    # 第5行：数据
    ws.append(['张三', '2024001', '110101200001011234', '13800138000', '1班'])

    result = detect_header_row(ws)

    print(f"检测结果: 第 {result['header_row']} 行")
    print(f"置信度: {result['confidence']*100:.1f}%")
    print(f"原因: {result['reason']}")

    assert result['header_row'] == 4, f"期望第4行，实际第{result['header_row']}行"
    print("✅ 测试通过")

    wb.close()
    return True


def create_test_case_4():
    """测试用例4: 合并单元格标题（第1-2行合并单元格，第3行列名）"""
    print("\n" + "="*80)
    print("测试用例4: 合并单元格标题 - 第1-2行合并，第3行列名")
    print("="*80)

    wb = openpyxl.Workbook()
    ws = wb.active

    # 第1行：合并单元格标题
    ws.merge_cells('A1:E1')
    ws['A1'] = '学生信息管理系统'
    ws['A1'].alignment = Alignment(horizontal='center')

    # 第2行：合并单元格副标题
    ws.merge_cells('A2:E2')
    ws['A2'] = '2024-2025学年第一学期'
    ws['A2'].alignment = Alignment(horizontal='center')

    # 第3行：表头
    ws.append(['姓名', '学号', '身份证号', '手机号', '班级'])
    # 第4行：数据
    ws.append(['张三', '2024001', '110101200001011234', '13800138000', '1班'])

    result = detect_header_row(ws)

    print(f"检测结果: 第 {result['header_row']} 行")
    print(f"置信度: {result['confidence']*100:.1f}%")
    print(f"原因: {result['reason']}")

    assert result['header_row'] == 3, f"期望第3行，实际第{result['header_row']}行"
    print("✅ 测试通过")

    wb.close()
    return True


def create_test_case_5():
    """测试用例5: 边缘情况（所有行看起来都像数据）"""
    print("\n" + "="*80)
    print("测试用例5: 边缘情况 - 所有行都是数据")
    print("="*80)

    wb = openpyxl.Workbook()
    ws = wb.active

    # 所有行都是纯数据（数字）
    for i in range(1, 11):
        ws.append([i, i*100, i*1000, i*10000, i*100000])

    result = detect_header_row(ws)

    print(f"检测结果: 第 {result['header_row']} 行")
    print(f"置信度: {result['confidence']*100:.1f}%")
    print(f"原因: {result['reason']}")

    # 这种情况应该选择第1行，但置信度会很低
    assert result['header_row'] == 1, f"期望第1行，实际第{result['header_row']}行"
    assert result['confidence'] < 0.5, "置信度应该较低"
    print("✅ 测试通过（置信度较低符合预期）")

    wb.close()
    return True


def main():
    """运行所有测试用例"""
    print("\n" + "="*80)
    print("自动表头检测功能测试")
    print("="*80)

    test_cases = [
        create_test_case_1,
        create_test_case_2,
        create_test_case_3,
        create_test_case_4,
        create_test_case_5,
    ]

    passed = 0
    failed = 0

    for test_case in test_cases:
        try:
            if test_case():
                passed += 1
        except AssertionError as e:
            print(f"❌ 测试失败: {e}")
            failed += 1
        except Exception as e:
            print(f"❌ 测试出错: {e}")
            failed += 1

    print("\n" + "="*80)
    print(f"测试完成: {passed} 通过, {failed} 失败")
    print("="*80)

    return failed == 0


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
