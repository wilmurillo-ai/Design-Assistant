#!/usr/bin/env python3
"""
高德POI医美机构数据采集脚本
用法: python3 poi_fetch.py <城市名> [--key <高德KEY>]
      python3 poi_fetch.py <城市名> --skip-excel
"""
import json, time, urllib.request, urllib.parse, os, sys, glob

# ========== 配置 ==========
DEFAULT_KEY = "0c166a2bf61c1e4e6c96e3b645233e54"  # 默认key（主人已创建）
KEY = os.environ.get("AMAP_KEY", DEFAULT_KEY)
SKIP_EXCEL = "--skip-excel" in sys.argv

def fetch_json(url):
    with urllib.request.urlopen(url, timeout=10) as r:
        return json.loads(r.read().decode("utf-8"))

def get_districts(city_name):
    """通过高德行政区划API获取城市的区级adcode列表"""
    url = f"https://restapi.amap.com/v3/config/district?keywords={urllib.parse.quote(city_name)}&subdistrict=1&key={KEY}"
    data = fetch_json(url)
    districts = []
    for d in data.get("districts", []):
        province_list = d.get("districts", [])
        for city in province_list:
            if city.get("level") in ("city", "district"):
                districts.append({
                    "name": city.get("name",""),
                    "adcode": city.get("adcode",""),
                    "center": city.get("center",""),
                    "level": city.get("level","")
                })
    return districts

def fetch_keyword(keyword, adcode, name):
    """抓取某区某关键词全部页（最多200条）"""
    kw_enc = urllib.parse.quote(keyword)
    url_base = (f"https://restapi.amap.com/v3/place/text?key={KEY}"
                f"&keywords={kw_enc}&city={adcode}&citylimit=true&offset=20&extensions=all")
    all_pois = []
    page = 1

    while page <= 10:
        url = f"{url_base}&page={page}"
        try:
            data = fetch_json(url)
        except Exception as e:
            print(f"    ⚠ p{page}: {e}")
            break

        pois = data.get("pois") or []
        all_pois.extend(pois)
        total = int(data.get("count", 0))
        pages = max((total + 19) // 20, 1)
        print(f"  {name}[{keyword}] p{page}/{pages}: +{len(pois)} (累计{len(all_pois)}/{total})")
        if not pois or page >= pages:
            break
        page += 1
        time.sleep(0.8)

    return all_pois

def save_json(path, pois):
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"count": len(pois), "pois": pois, "status": "1"}, f, ensure_ascii=False)

def export_excel(outdir, city_name):
    """将JSON数据导出为Excel（3个Sheet）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("⚠ openpyxl未安装，跳过Excel导出")
        return

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()

    def safe(v):
        if v is None: return ""
        if isinstance(v, (list, dict)): return str(v) if v else ""
        return str(v)

    def safe_float(v):
        try:
            f = float(v)
            return f if f > 0 else ""
        except: return ""

    # --- Sheet 1: 汇总 ---
    ws0 = wb.active
    ws0.title = "汇总"
    ws0.sheet_view.showGridLines = False
    ws0.merge_cells("A1:D1")
    ws0["A1"] = f"{city_name} 医疗美容+生活美容 机构分布汇总"
    ws0["A1"].font = Font(name="微软雅黑", bold=True, size=15, color="2F75B6")
    ws0["A1"].alignment = center
    ws0.row_dimensions[1].height = 34

    ws0.merge_cells("A2:D2")
    ws0["A2"] = "⚠️ 高德API硬上限：每词每区最多返回200条 | 数据来源: 高德地图POI"
    ws0["A2"].font = Font(name="微软雅黑", size=9, color="CC6600")

    for col, h in enumerate(["区县", "医疗美容(家)", "生活美容(家)", "小计(家)"], 1):
        c = ws0.cell(row=4, column=col, value=h)
        c.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor="2F75B6")
        c.alignment = center
        c.border = border
    ws0.row_dimensions[4].height = 26

    ym_fill = PatternFill("solid", fgColor="FDECEA")
    sh_fill = PatternFill("solid", fgColor="E8F8F0")
    ym_total = sh_total = 0
    row = 5

    for f in sorted(glob.glob(f"{outdir}/*_医疗美容.json")):
        name = os.path.basename(f).replace("_医疗美容.json","")
        ym = len(json.load(open(f)).get("pois", []))
        sh_file = f.replace("医疗美容","生活美容")
        sh = len(json.load(open(sh_file)).get("pois", [])) if os.path.exists(sh_file) else 0
        ym_total += ym; sh_total += sh

        for col, v in enumerate([name, ym, sh, ym+sh], 1):
            c = ws0.cell(row=row, column=col, value=v)
            c.font = Font(name="微软雅黑", size=10)
            c.border = border
            c.alignment = center if col > 1 else left_align
            if col == 2: c.fill = ym_fill
            elif col == 3: c.fill = sh_fill
        ws0.row_dimensions[row].height = 22
        row += 1

    for col, v in enumerate(["合计", ym_total, sh_total, ym_total+sh_total], 1):
        c = ws0.cell(row=row, column=col, value=v)
        c.font = Font(name="微软雅黑", bold=True, size=11)
        c.fill = PatternFill("solid", fgColor="D6E4F0")
        c.border = border
        c.alignment = center
    ws0.row_dimensions[row].height = 26

    ws0.column_dimensions["A"].width = 14
    ws0.column_dimensions["B"].width = 16
    ws0.column_dimensions["C"].width = 16
    ws0.column_dimensions["D"].width = 14

    # --- 明细表函数 ---
    def write_detail(keyword, fill_color, label):
        ws2 = wb.create_sheet(title=f"{label}明细")
        ws2.sheet_view.showGridLines = False
        headers = ["区县", "机构名称", "类型标签", "电话", "地址", "评分",
                   "人均(元)", "营业时间", "商圈", "经度", "纬度"]
        for col, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.alignment = center
            c.border = border
        ws2.row_dimensions[1].height = 26

        row = 2
        for f in sorted(glob.glob(f"{outdir}/*{keyword}.json")):
            district = os.path.basename(f).replace(f"_{keyword}.json","")
            try:
                pois = json.load(open(f)).get("pois") or []
                for poi in pois:
                    loc = safe(poi.get("location",""))
                    lng = lat = ""
                    if "," in loc:
                        lng = safe_float(loc.split(",")[0])
                        lat = safe_float(loc.split(",")[1])
                    biz_ext = poi.get("biz_ext") or {}

                    row_data = [
                        district,
                        safe(poi.get("name","")),
                        safe(poi.get("keytag","")),
                        safe(poi.get("tel","")),
                        safe(poi.get("address","")),
                        safe_float(biz_ext.get("rating","")),
                        safe_float(biz_ext.get("cost","")),
                        safe(biz_ext.get("opentime2","")),
                        safe(poi.get("business_area","")),
                        lng, lat
                    ]
                    for col, val in enumerate(row_data, 1):
                        c = ws2.cell(row=row, column=col, value=val)
                        c.font = Font(name="微软雅黑", size=9)
                        c.border = border
                        c.alignment = left_align if col in [2,5,8] else center
                    ws2.row_dimensions[row].height = 18
                    row += 1
            except Exception as e:
                print(f"  ⚠ {district}: {e}")

        widths = [10, 34, 14, 22, 42, 8, 12, 32, 12, 13, 13]
        for col, w in enumerate(widths, 1):
            ws2.column_dimensions[chr(64+col)].width = w
        print(f"  {label}明细: {row-2} 条")

    print("  导出Excel...")
    write_detail("医疗美容", "C0392B", "医疗美容")
    write_detail("生活美容", "1E8449", "生活美容")

    date_str = time.strftime("%Y%m%d")
    outpath = f"{outdir}/{city_name}医美生活美容数据_{date_str}.xlsx"
    wb.save(outpath)
    return outpath

def main(city_name):
    key_arg_idx = None
    for i, a in enumerate(sys.argv):
        if a == "--key" and i+1 < len(sys.argv):
            global KEY
            KEY = sys.argv[i+1]
        if a == "--city" and i+1 < len(sys.argv):
            city_name = sys.argv[i+1]

    print(f"=" * 50)
    print(f"高德POI数据采集 — {city_name}")
    print(f"=" * 50)

    # 1. 获取区划
    print(f"\n[1/4] 获取城市区划...")
    districts = get_districts(city_name)
    if not districts:
        print(f"❌ 未找到城市: {city_name}")
        return
    print(f"  → 找到 {len(districts)} 个区/县")

    # 2. 建立目录
    safe_name = city_name.replace("市","").replace("省","").replace("区","").replace("县","")
    outdir = f"~/.openclaw/workspace/data/{safe_name}_poi"
    outdir = os.path.expanduser(outdir)
    os.makedirs(outdir, exist_ok=True)
    print(f"  → 数据目录: {outdir}")

    # 3. 采集数据
    print(f"\n[2/4] 开始采集数据（每区翻页，最多200条/词）...")
    ym_total = sh_total = 0
    for district in districts:
        name = district["name"]
        adcode = district["adcode"]
        ym_file = f"{outdir}/{name}_医疗美容.json"
        sh_file = f"{outdir}/{name}_生活美容.json"

        if os.path.exists(ym_file) and os.path.getsize(ym_file) > 200:
            ym_pois = json.load(open(ym_file)).get("pois", [])
            print(f"  ⏭ {name}: 已有 {len(ym_pois)} 条医疗美容，跳过")
        else:
            print(f"\n  [{name}] 医疗美容...")
            ym_pois = fetch_keyword("医疗美容", adcode, name)
            save_json(ym_file, ym_pois)
            time.sleep(0.5)

        if os.path.exists(sh_file) and os.path.getsize(sh_file) > 200:
            sh_pois = json.load(open(sh_file)).get("pois", [])
            print(f"  ⏭ {name}: 已有 {len(sh_pois)} 条生活美容，跳过")
        else:
            print(f"  [{name}] 生活美容...")
            sh_pois = fetch_keyword("生活美容", adcode, name)
            save_json(sh_file, sh_pois)
            time.sleep(0.5)

        ym_total += len(ym_pois)
        sh_total += len(sh_pois)

    print(f"\n  → 医疗美容: {ym_total} 条")
    print(f"  → 生活美容: {sh_total} 条")

    # 4. 导出Excel
    if not SKIP_EXCEL:
        print(f"\n[3/4] 生成Excel...")
        excel_path = export_excel(outdir, safe_name)
        print(f"  ✅ {excel_path}")
    else:
        print(f"\n[3/4] 跳过Excel导出（--skip-excel）")

    # 5. 汇总
    print(f"\n[4/4] 完成！")
    print(f"  数据目录: {outdir}")
    print(f"  医疗美容: {ym_total} 条")
    print(f"  生活美容: {sh_total} 条")
    print(f"  合计: {ym_total+sh_total} 条")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1])
