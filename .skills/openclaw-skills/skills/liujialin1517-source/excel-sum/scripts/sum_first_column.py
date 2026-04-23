#!/usr/bin/env python3
"""
Excel 第一列求和脚本
用于计算 Excel 文件第一列所有数值的总和
"""

import sys
import os

try:
    from openpyxl import load_workbook
except ImportError:
    print("错误：缺少 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)


def sum_first_column(file_path):
    """
    读取 Excel 文件并计算第一列的总和
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        tuple: (总和, 计数, 跳过的行数, 错误信息)
    """
    if not os.path.exists(file_path):
        return None, 0, 0, f"文件不存在: {file_path}"
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        total = 0
        count = 0
        skipped = 0
        
        # 遍历第一列（从第一行开始）
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            value = row[0]
            
            # 跳过空值
            if value is None:
                skipped += 1
                continue
            
            # 尝试转换为数值
            try:
                # 如果是字符串，去除空格后转换
                if isinstance(value, str):
                    value = value.strip()
                    if value == '':
                        skipped += 1
                        continue
                    total += float(value)
                else:
                    total += float(value)
                count += 1
            except (ValueError, TypeError):
                # 无法转换为数值的跳过
                skipped += 1
                continue
        
        wb.close()
        return total, count, skipped, None
        
    except Exception as e:
        return None, 0, 0, f"读取文件错误: {str(e)}"


def main():
    if len(sys.argv) < 2:
        print("用法: python sum_first_column.py <excel文件路径>")
        print("示例: python sum_first_column.py data.xlsx")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    total, count, skipped, error = sum_first_column(file_path)
    
    if error:
        print(f"错误: {error}")
        sys.exit(1)
    
    # 输出结果
    print("=" * 40)
    print(f"文件: {os.path.basename(file_path)}")
    print("=" * 40)
    print(f"第一列数值总和: {total}")
    print(f"求和单元格数量: {count}")
    print(f"跳过的单元格数: {skipped}")
    print("=" * 40)


if __name__ == "__main__":
    main()
