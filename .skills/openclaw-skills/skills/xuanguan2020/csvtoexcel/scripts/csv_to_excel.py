#!/usr/bin/env python3
"""
CSV to Excel Converter with Advanced Formatting

Converts one or multiple CSV files to Excel format with proper formatting,
Chinese character support, and professional styling.

Usage:
    # Single CSV to Excel
    python csv_to_excel.py input.csv output.xlsx
    
    # Multiple CSVs to one Excel (each CSV becomes a sheet)
    python csv_to_excel.py file1.csv file2.csv file3.csv --output combined.xlsx
    
    # With custom sheet names
    python csv_to_excel.py data1.csv data2.csv --output report.xlsx --sheet-names "Sales" "Inventory"
"""

import sys
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def detect_csv_encoding(file_path):
    """
    Detect CSV file encoding to handle Chinese characters properly.
    Tries common encodings: UTF-8, GBK, GB2312, UTF-8-SIG.
    """
    encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig', 'latin1']
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                f.read()
            return encoding
        except (UnicodeDecodeError, UnicodeError):
            continue
    
    # Default to utf-8 if nothing works
    return 'utf-8'


def read_csv_with_encoding(file_path):
    """
    Read CSV file with proper encoding detection.
    Returns list of rows.
    """
    encoding = detect_csv_encoding(file_path)
    
    with open(file_path, 'r', encoding=encoding, newline='') as f:
        # Try to detect delimiter
        sample = f.read(1024)
        f.seek(0)
        
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(sample)
            reader = csv.reader(f, dialect)
        except:
            # Fallback to comma delimiter
            reader = csv.reader(f)
        
        rows = list(reader)
    
    return rows


def format_worksheet(ws, has_header=True):
    """
    Apply professional formatting to worksheet:
    - Bold header row with background color
    - Auto-adjust column widths
    - Add borders
    - Center align headers
    - Freeze header row
    """
    # Header styling
    if has_header and ws.max_row > 0:
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Freeze header row
        ws.freeze_panes = ws['A2']
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    # Handle Chinese characters (count as 2 width units)
                    cell_value = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    max_length = max(max_length, length)
            except:
                pass
        
        # Set column width with padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long content
        ws.column_dimensions[column_letter].width = adjusted_width


def csv_to_excel(csv_files, output_file, sheet_names=None):
    """
    Convert CSV file(s) to Excel with formatting.
    
    Args:
        csv_files: List of CSV file paths
        output_file: Output Excel file path
        sheet_names: Optional list of sheet names (defaults to CSV filenames)
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    for idx, csv_file in enumerate(csv_files):
        csv_path = Path(csv_file)
        
        if not csv_path.exists():
            print(f"⚠️  Warning: File not found: {csv_file}")
            continue
        
        # Determine sheet name
        if sheet_names and idx < len(sheet_names):
            sheet_name = sheet_names[idx]
        else:
            sheet_name = csv_path.stem
        
        # Excel sheet names have max 31 characters
        sheet_name = sheet_name[:31]
        
        # Read CSV data
        print(f"📖 Reading: {csv_file}")
        rows = read_csv_with_encoding(csv_file)
        
        if not rows:
            print(f"⚠️  Warning: Empty file: {csv_file}")
            continue
        
        # Create worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Write data to worksheet
        for row_data in rows:
            ws.append(row_data)
        
        # Apply formatting
        format_worksheet(ws, has_header=True)
        
        print(f"✅ Added sheet: {sheet_name} ({len(rows)} rows, {len(rows[0]) if rows else 0} columns)")
    
    # Save workbook
    wb.save(output_file)
    print(f"\n🎉 Excel file created: {output_file}")


def main():
    if len(sys.argv) < 3:
        print("Usage:")
        print("  Single CSV:   python csv_to_excel.py input.csv output.xlsx")
        print("  Multiple CSVs: python csv_to_excel.py file1.csv file2.csv --output combined.xlsx")
        print("  With sheet names: python csv_to_excel.py file1.csv file2.csv --output report.xlsx --sheet-names 'Sales' 'Inventory'")
        sys.exit(1)
    
    # Parse arguments
    args = sys.argv[1:]
    csv_files = []
    output_file = None
    sheet_names = None
    
    i = 0
    while i < len(args):
        if args[i] == '--output':
            output_file = args[i + 1]
            i += 2
        elif args[i] == '--sheet-names':
            sheet_names = []
            i += 1
            while i < len(args) and not args[i].startswith('--'):
                sheet_names.append(args[i])
                i += 1
        else:
            csv_files.append(args[i])
            i += 1
    
    # Handle simple case: input.csv output.xlsx
    if len(csv_files) == 2 and output_file is None and not csv_files[1].endswith('.csv'):
        output_file = csv_files[1]
        csv_files = [csv_files[0]]
    
    if not csv_files:
        print("❌ Error: No CSV files specified")
        sys.exit(1)
    
    if not output_file:
        print("❌ Error: No output file specified")
        sys.exit(1)
    
    # Convert
    csv_to_excel(csv_files, output_file, sheet_names)


if __name__ == "__main__":
    main()
