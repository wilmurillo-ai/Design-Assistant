#!/usr/bin/env python3
"""
仓配网络规划优化 - 核心脚本 v2.7

更新日志:
- v2.7: 参数引导流程，支持必选/禁止仓库约束
- v2.6: 智能单位检测，自动转换时效单位

用法: python3 run_optimization.py --store-file <门店Excel> [--rdc-count <数量>] [--required <必选仓库>] [--forbidden <禁止仓库>] [--output <输出文件>]
"""

import json
import pandas as pd
import pulp
from collections import defaultdict, deque
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import argparse
import os


def load_trunk_data(file_path='trunk_lines.json'):
    """加载干线数据"""
    with open(file_path, 'r', encoding='utf-8') as f:
        trunk_data = json.load(f)
    trunk_lines = trunk_data['trunk_lines']
    
    trunk_graph = defaultdict(list)
    trunk_shipping_days = {}
    for line in trunk_lines:
        trunk_graph[line['from_house_id']].append({
            'to': line['to_house_id'],
            'efficiency': line['efficiency']
        })
        trunk_shipping_days[(line['from_house_id'], line['to_house_id'])] = \
            [int(x) for x in line['shipping_days'].split(',')]
    
    warehouse_name_map = {}
    for line in trunk_lines:
        warehouse_name_map[line['from_house_id']] = line['from_house_name']
        warehouse_name_map[line['to_house_id']] = line['to_house_name']
    
    warehouse_list = [{'id': k, 'name': v} for k, v in warehouse_name_map.items()]
    
    return trunk_graph, warehouse_list, warehouse_name_map, trunk_shipping_days


def load_distribution_data(file_path='distribution_cycles.json'):
    """加载共配数据"""
    with open(file_path, 'r', encoding='utf-8') as f:
        dist_data = json.load(f)
    
    dist_efficiency = defaultdict(dict)
    dist_shipping_days = defaultdict(dict)
    
    for d in dist_data['distribution_cycles']:
        key_city = (d['province_name'], d['city_name'])
        house = d['start_house_name']
        
        if key_city not in dist_efficiency[house]:
            dist_efficiency[house][key_city] = d['efficiency']
        else:
            dist_efficiency[house][key_city] = min(dist_efficiency[house][key_city], d['efficiency'])
        
        if key_city not in dist_shipping_days[house]:
            dist_shipping_days[house][key_city] = [int(x) for x in d['shipping_date'].split(',')]
    
    province_map = defaultdict(lambda: defaultdict(list))
    for house, cities in dist_efficiency.items():
        for (prov, city), eff in cities.items():
            province_map[house][prov].append(city)
    
    return dist_efficiency, dist_shipping_days, province_map


def find_best_distribution(province, city, dist_efficiency, province_map):
    """查找最佳配送基地"""
    for house, cities in dist_efficiency.items():
        for (prov, c), eff in cities.items():
            if province in prov and city in c:
                return house, (prov, c), eff
    
    for house, provinces in province_map.items():
        for prov, cities in provinces.items():
            if province in prov:
                return house, (prov, cities[0]), 2
    
    return None, None, None


def get_next_shipping_day(arrival_day, shipping_days):
    """计算等待天数"""
    if arrival_day in shipping_days:
        return 0
    for i in range(1, 8):
        next_day = ((arrival_day + i - 1) % 7) + 1
        if next_day in shipping_days:
            return i
    return 7


def calculate_full_efficiency(rdc_id, fsl_name, province, city, trunk_graph, 
                            warehouse_list, dist_efficiency, dist_shipping_days, trunk_shipping_days):
    """计算完整时效（含等待时间）"""
    fsl_id = None
    for w in warehouse_list:
        if w['name'] == fsl_name:
            fsl_id = w['id']
            break
    if not fsl_id:
        return None

    if rdc_id == fsl_id:
        dist_eff = dist_efficiency.get(fsl_name, {}).get((province, city), 2)
        dist_days = dist_shipping_days.get(fsl_name, {}).get((province, city), [1,2,3,4,5,6,7])
        wait = get_next_shipping_day(1, dist_days)
        return {'total_days': wait + dist_eff, 'wait_days': wait, 'trunk_days': 0, 'dist_days': dist_eff}

    queue = deque([(rdc_id, [rdc_id], 0, 0, 0)])
    visited = {rdc_id}

    while queue:
        current, path, transit, wait_days, hops = queue.popleft()
        if hops > 2:
            continue
        for next_hop in trunk_graph.get(current, []):
            next_id = next_hop['to']
            line_eff = next_hop['efficiency']
            arrival_day = ((1 + int(transit)) % 7) + 1
            key = (current, next_id)
            shipping = trunk_shipping_days.get(key, [1,2,3,4,5,6,7])
            wait = get_next_shipping_day(arrival_day, shipping)
            new_transit = transit + wait + line_eff
            new_wait = wait_days + wait
            new_path = path + [next_id]

            if next_id == fsl_id:
                dist_eff = dist_efficiency.get(fsl_name, {}).get((province, city), 2)
                dist_days = dist_shipping_days.get(fsl_name, {}).get((province, city), [1,2,3,4,5,6,7])
                final_arrival = ((1 + int(new_transit)) % 7) + 1
                dist_wait = get_next_shipping_day(final_arrival, dist_days)
                return {'total_days': new_transit + dist_wait + dist_eff, 
                       'wait_days': new_wait + dist_wait,
                       'trunk_days': transit + line_eff, 'dist_days': dist_eff, 'path': new_path}

            if next_id not in visited and hops < 2:
                visited.add(next_id)
                queue.append((next_id, new_path, new_transit, new_wait, hops + 1))

    return None


def process_store_data(df):
    """处理门店数据"""
    city_to_province = {
        '贵阳市': '贵州省', '成都市': '四川省', '武汉市': '湖北省',
        '南昌市': '江西省', '郑州市': '河南省', '西安市': '陕西省',
        '重庆市': '重庆市', '昆明市': '云南省', '长沙市': '湖南省',
        '漯河市': '河南省', '开封市': '河南省', '景德镇市': '江西省',
        '荆州市': '湖北省', '襄阳市': '湖北省', '新乡市': '河南省',
        '宜昌市': '湖北省', '宜宾市': '四川省', '南阳市': '河南省',
    }
    
    records = []
    for idx, row in df.iterrows():
        province = str(row['省']) if pd.notna(row.get('省')) else None
        city = str(row['市']) if pd.notna(row.get('市')) else None
        district = str(row['区']) if pd.notna(row.get('区')) else '全区域'
        
        if not city:
            continue
        
        if not province or province == 'nan':
            province = city_to_province.get(city, city + '市')
        
        records.append({
            '省份': province,
            '城市': city,
            '区县': district,
        })
    
    return pd.DataFrame(records)


def optimize_and_generate_excel(store_file, num_rdc=3, output_file='result.xlsx', 
                                trunk_file='trunk_lines.json', dist_file='distribution_cycles.json',
                                required_warehouses=None, forbidden_warehouses=None):
    """执行优化并生成Excel"""
    print('=== 仓配网络规划优化 v2.7 ===')
    
    # 参数引导
    print('\n【Step 3: 参数配置】')
    print(f'  RDC数量上限: {num_rdc}')
    print(f'  必选仓库: {required_warehouses or "无"}')
    print(f'  禁止仓库: {forbidden_warehouses or "无"}')
    print()
    
    # 加载门店数据
    df = pd.read_excel(store_file)
    print(f'原始数据: {len(df)} 行')
    df = process_store_data(df)
    area_stats = df.groupby(['省份', '城市', '区县']).size().reset_index(name='门店数量')
    print(f'有效区域: {len(area_stats)} 个')
    
    # 加载干线数据
    trunk_graph, warehouse_list, warehouse_name_map, trunk_shipping_days = load_trunk_data(trunk_file)
    print(f'仓库: {len(warehouse_list)} 个')
    print(f'干线: {len(trunk_graph)} 条')
    
    # 加载共配数据
    dist_efficiency, dist_shipping_days, province_map = load_distribution_data(dist_file)
    print(f'共配记录: {sum(len(v) for v in dist_efficiency.values())} 条')
    
    # 优化
    print(f'\n=== 运行优化 (RDC={num_rdc}) ===')
    prob = pulp.LpProblem('WarehouseNetwork', pulp.LpMinimize)
    warehouses_ids = [w['id'] for w in warehouse_list]
    
    provinces = df['省份'].unique().tolist()
    
    y = pulp.LpVariable.dicts('y', warehouses_ids, cat='Binary')
    x = pulp.LpVariable.dicts('x', [(w['id'], p) for w in warehouse_list for p in provinces], cat='Binary')
    
    cost_terms = []
    for w in warehouse_list:
        for p in provinces:
            best_days = 5
            for fsl in warehouse_list:
                fsl_name = fsl['name']
                if fsl_name in dist_efficiency:
                    result = calculate_full_efficiency(
                        w['id'], fsl_name, p, p, trunk_graph,
                        warehouse_list, dist_efficiency, dist_shipping_days, trunk_shipping_days
                    )
                    if result:
                        best_days = min(best_days, result['total_days'])
            cost_terms.append(x[(w['id'], p)] * 1 * best_days)
    
    prob += pulp.lpSum(cost_terms)
    for p in provinces:
        prob += pulp.lpSum([x[(w['id'], p)] for w in warehouse_list]) == 1
    for w in warehouse_list:
        for p in provinces:
            prob += x[(w['id'], p)] <= y[w['id']]
    prob += pulp.lpSum([y[i] for i in warehouses_ids]) <= num_rdc  # RDC数量上限
    
    # 必选仓库约束
    if required_warehouses:
        for rw in required_warehouses:
            for w in warehouse_list:
                if w['name'] == rw:
                    prob += y[w['id']] == 1
                    print(f'  强制包含: {rw}')
    
    # 禁止仓库约束
    if forbidden_warehouses:
        for fw in forbidden_warehouses:
            for w in warehouse_list:
                if w['name'] == fw:
                    prob += y[w['id']] == 0
                    print(f'  强制排除: {fw}')
    
    prob.solve(pulp.PULP_CBC_CMD(msg=0))
    print(f'求解状态: {pulp.LpStatus[prob.status]}')
    
    selected = [w for w in warehouse_list if y[w['id']].value() > 0.5]
    print(f'选中的RDC: {[w["name"] for w in selected]}')
    
    # 生成Excel
    print('\n=== 生成Excel ===')
    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    def style_header(ws, cols):
        for col in cols:
            ws[f'{col}1'].fill = header_fill
            ws[f'{col}1'].font = header_font
    
    # 计算加权平均时效
    supply_records = []  # 收集供应关系数据用于计算
    
    # 执行摘要
    ws1 = wb.active
    ws1.title = "执行摘要"
    ws1['A1'] = "仓配网络规划分析报告 v2.6"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A3'] = "RDC数量"
    ws1['B3'] = len(selected)
    ws1['A4'] = "覆盖区域"
    ws1['B4'] = len(area_stats)
    ws1['A5'] = "总门店数"
    ws1['B5'] = len(df)
    ws1['A6'] = "时效单位"
    ws1['B6'] = "天"
    ws1['A7'] = "加权平均时效"
    ws1['B7'] = "待计算"  # 先占位，后面填入
    
    # RDC选址
    ws2 = wb.create_sheet("RDC选址")
    ws2['A1'] = "RDC仓库"
    ws2['B1'] = "仓库ID"
    style_header(ws2, ['A', 'B'])
    for i, w in enumerate(selected, 2):
        ws2[f'A{i}'] = w['name']
        ws2[f'B{i}'] = w['id']
    
    # 供应关系
    ws3 = wb.create_sheet("供应关系")
    headers = ["RDC仓库", "中转仓", "配送基地", "省份", "城市", "区县", 
               "门店数量", "链路类型", "等待时间", "干线时效(天)", "配送时效", "总时效(天)"]
    for i, h in enumerate(headers, 1):
        ws3.cell(row=1, column=i, value=h)
    style_header(ws3, [get_column_letter(i) for i in range(1, len(headers)+1)])
    
    transfer_records = []
    supply_records = []  # 收集供应关系用于计算加权平均
    row = 2
    for area in area_stats.itertuples():
        province, city, district = area.省份, area.城市, area.区县
        store_count = area.门店数量
        
        fsl_name, fsl_key, _ = find_best_distribution(province, city, dist_efficiency, province_map)
        
        if fsl_name:
            best_rdc, best_result = None, None
            for w in selected:
                result = calculate_full_efficiency(
                    w['id'], fsl_name, fsl_key[0], fsl_key[1], trunk_graph,
                    warehouse_list, dist_efficiency, dist_shipping_days, trunk_shipping_days
                )
                if result and (best_result is None or result['total_days'] < best_result['total_days']):
                    best_rdc = w
                    best_result = result
            
            if best_rdc and best_result:
                tc_name = ''
                if best_rdc['name'] == fsl_name:
                    link_type = '直达'
                elif len(best_result.get('path', [])) == 2:
                    link_type = '调拨一次'
                else:
                    link_type = '中转'
                    if best_result.get('path') and len(best_result['path']) > 1:
                        tc_name = warehouse_name_map.get(best_result['path'][1], '')
                        transfer_records.append({
                            'RDC': best_rdc['name'], '中转仓': tc_name, '目的仓': fsl_name,
                            '链路': ' -> '.join([warehouse_name_map.get(p, '') for p in best_result['path']]),
                            '时效': round(best_result['total_days'], 2)
                        })
                
                ws3[f'A{row}'] = best_rdc['name']
                ws3[f'B{row}'] = tc_name
                ws3[f'C{row}'] = fsl_name
                ws3[f'D{row}'] = province
                ws3[f'E{row}'] = city
                ws3[f'F{row}'] = district
                ws3[f'G{row}'] = store_count
                ws3[f'H{row}'] = link_type
                ws3[f'I{row}'] = best_result['wait_days']
                ws3[f'J{row}'] = round(best_result['trunk_days'], 2)
                ws3[f'K{row}'] = best_result['dist_days']
                ws3[f'L{row}'] = round(best_result['total_days'], 2)
                
                # 收集数据用于计算加权平均
                supply_records.append({
                    'store_count': store_count,
                    'total_days': round(best_result['total_days'], 2)
                })
                
                row += 1
    
    # 调拨中转仓
    ws4 = wb.create_sheet("调拨中转仓")
    for i, h in enumerate(["序号", "RDC仓库", "中转仓", "目的仓", "完整链路", "总时效(天)"], 1):
        ws4.cell(row=1, column=i, value=h)
    style_header(ws4, ['A', 'B', 'C', 'D', 'E', 'F'])
    
    seen, unique_transfers = set(), []
    for t in transfer_records:
        key = (t['RDC'], t['中转仓'], t['目的仓'])
        if key not in seen:
            seen.add(key)
            unique_transfers.append(t)
    for i, t in enumerate(unique_transfers, 2):
        ws4[f'A{i}'] = i - 1
        ws4[f'B{i}'] = t['RDC']
        ws4[f'C{i}'] = t['中转仓']
        ws4[f'D{i}'] = t['目的仓']
        ws4[f'E{i}'] = t['链路']
        ws4[f'F{i}'] = t['时效']
    
    # 省份分布
    ws5 = wb.create_sheet("省份分布")
    for i, h in enumerate(["省份", "区域数", "RDC归属"], 1):
        ws5.cell(row=1, column=i, value=h)
    style_header(ws5, ['A', 'B', 'C'])
    
    province_assign = {}
    for p in provinces:
        for w in selected:
            if x[(w['id'], p)].value() > 0.5:
                province_assign[p] = w['name']
                break
    for i, p in enumerate(sorted(province_assign.keys()), 2):
        ws5[f'A{i}'] = p
        ws5[f'B{i}'] = len(area_stats[area_stats['省份']==p])
        ws5[f'C{i}'] = province_assign.get(p, '')
    
    for ws in [ws3, ws4, ws5]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    # 计算加权平均时效
    total_stores = sum(r['store_count'] for r in supply_records)
    weighted_sum = sum(r['store_count'] * r['total_days'] for r in supply_records)
    weighted_avg_efficiency = weighted_sum / total_stores if total_stores > 0 else 0
    
    # 更新执行摘要
    ws1['B7'] = round(weighted_avg_efficiency, 2)
    print(f'加权平均时效: {round(weighted_avg_efficiency, 2)} 天')
    
    wb.save(output_file)
    print(f'✅ Excel报告已生成: {output_file}')
    print(f'供应关系记录: {row - 2} 条')


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='仓配网络规划优化 v2.7')
    parser.add_argument('--store-file', required=True, help='门店数据Excel文件')
    parser.add_argument('--rdc-count', type=int, default=3, help='RDC数量上限（默认3）')
    parser.add_argument('--required', nargs='*', default=[], help='必选仓库（空格分隔，如：成都仓 广州仓')
    parser.add_argument('--forbidden', nargs='*', default=[], help='禁止仓库（空格分隔，如：北京仓 上海仓')
    parser.add_argument('--output', default='result.xlsx', help='输出文件路径')
    parser.add_argument('--trunk-data', default='trunk_lines.json', help='干线数据文件')
    parser.add_argument('--dist-data', default='distribution_cycles.json', help='共配数据文件')
    args = parser.parse_args()
    
    optimize_and_generate_excel(
        args.store_file, 
        args.rdc_count, 
        args.output,
        args.trunk_data,
        args.dist_data,
        required_warehouses=args.required if args.required else None,
        forbidden_warehouses=args.forbidden if args.forbidden else None
    )
