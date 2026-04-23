"""
HR 智能体 - 员工管理模块单元测试
使用内存 Excel（不依赖实际文件）测试所有功能
"""

import sys
import os
import unittest
from datetime import datetime
from decimal import Decimal

# 确保能导入被测模块
sys.path.insert(0, os.path.dirname(__file__))

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("跳过测试：openpyxl 未安装")

if OPENPYXL_AVAILABLE:
    from employee_manager import (
        Employee, Department, EmployeeManager,
        FIELD_ALIASES, REQUIRED_FIELDS, RECOMMENDED_FIELDS
    )

def create_test_excel(file_path: str, employees: list = None, headers: list = None):
    """创建测试用 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "员工花名册"
    
    if headers is None:
        headers = [
            "工号", "姓名", "部门编码", "部门名称", "岗位", "性别",
            "手机号", "邮箱", "入职日期", "转正日期", "在职状态",
            "基本工资", "社保城市", "汇报对象"
        ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入数据
    if employees:
        for row_idx, emp_data in enumerate(employees, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col, value=emp_data.get(header))
    
    wb.save(file_path)
    return file_path


def create_org_excel(file_path: str, departments: list = None):
    """创建测试用组织架构 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "组织架构"
    
    headers = ["部门编码", "部门名称", "上级部门", "负责人"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    if departments:
        for row_idx, dept_data in enumerate(departments, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col, value=dept_data.get(header))
    
    wb.save(file_path)
    return file_path


class TestEmployee(unittest.TestCase):
    """Employee 数据模型测试"""
    
    def test_create_basic_employee(self):
        """测试创建基本员工"""
        emp = Employee(empNo="E001", name="张三")
        self.assertEqual(emp.empNo, "E001")
        self.assertEqual(emp.name, "张三")
        self.assertEqual(emp.status, "在职")  # 默认状态
    
    def test_employee_validation_empty_fields(self):
        """测试空字段验证"""
        emp = Employee(empNo="", name="")
        valid, errors = emp.validate()
        self.assertFalse(valid)
        self.assertIn("工号不能为空", errors)
        self.assertIn("姓名不能为空", errors)
    
    def test_employee_validation_phone(self):
        """测试手机号验证"""
        emp = Employee(empNo="E001", name="张三", phone="13800138000")
        valid, errors = emp.validate()
        self.assertTrue(valid)
        
        emp.phone = "123"
        valid, errors = emp.validate()
        self.assertFalse(valid)
        self.assertTrue(any("手机号" in e for e in errors))
    
    def test_employee_validation_email(self):
        """测试邮箱验证"""
        emp = Employee(empNo="E001", name="张三", email="test@example.com")
        valid, errors = emp.validate()
        self.assertTrue(valid)
        
        emp.email = "invalid-email"
        valid, errors = emp.validate()
        self.assertFalse(valid)
    
    def test_employee_validation_leave_requires_date(self):
        """测试离职状态需要日期"""
        emp = Employee(empNo="E001", name="张三", status="离职")
        valid, errors = emp.validate()
        self.assertFalse(valid)
        self.assertTrue(any("离职日期" in e for e in errors))
        
        emp.leaveDate = "2026-03-01"
        valid, errors = emp.validate()
        self.assertTrue(valid)
    
    def test_to_dict(self):
        """测试转字典"""
        emp = Employee(empNo="E001", name="张三", deptCode="D001")
        d = emp.toDict()
        self.assertEqual(d["empNo"], "E001")
        self.assertEqual(d["name"], "张三")
        self.assertNotIn("_rowIndex", d)
        self.assertNotIn("_isNew", d)
    
    def test_normalize_status(self):
        """测试状态标准化"""
        self.assertEqual(EmployeeManager._normalizeStatus("在职"), "在职")
        self.assertEqual(EmployeeManager._normalizeStatus("试用期"), "试用期")
        self.assertEqual(EmployeeManager._normalizeStatus("已离职"), "离职")
        self.assertEqual(EmployeeManager._normalizeStatus("inactive"), "离职")
        self.assertEqual(EmployeeManager._normalizeStatus("Probation"), "试用期")


class TestEmployeeManagerCRUD(unittest.TestCase):
    """EmployeeManager CRUD 操作测试"""
    
    @classmethod
    def setUpClass(cls):
        """创建测试 Excel 文件"""
        cls.test_dir = os.path.join(os.path.dirname(__file__), "test_data")
        os.makedirs(cls.test_dir, exist_ok=True)
        
        cls.excel_path = os.path.join(cls.test_dir, "test_employees.xlsx")
        
        test_employees = [
            {
                "工号": "E001", "姓名": "张三", "部门编码": "D001",
                "部门名称": "技术部", "岗位": "高级工程师", "性别": "男",
                "手机号": "13800138001", "邮箱": "zhangsan@company.com",
                "入职日期": "2023-03-15", "转正日期": "2023-06-15",
                "在职状态": "在职", "基本工资": 20000, "社保城市": "北京",
                "汇报对象": "E003",
            },
            {
                "工号": "E002", "姓名": "李四", "部门编码": "D001",
                "部门名称": "技术部", "岗位": "工程师", "性别": "男",
                "手机号": "13800138002", "邮箱": "lisi@company.com",
                "入职日期": "2024-01-10", "转正日期": "2024-04-10",
                "在职状态": "在职", "基本工资": 15000, "社保城市": "北京",
                "汇报对象": "E001",
            },
            {
                "工号": "E003", "姓名": "王五", "部门编码": "D001",
                "部门名称": "技术部", "岗位": "技术总监", "性别": "男",
                "手机号": "13800138003", "邮箱": "wangwu@company.com",
                "入职日期": "2022-06-01", "转正日期": "2022-09-01",
                "在职状态": "在职", "基本工资": 35000, "社保城市": "北京",
                "汇报对象": "",
            },
            {
                "工号": "E004", "姓名": "赵六", "部门编码": "D002",
                "部门名称": "产品部", "岗位": "产品经理", "性别": "女",
                "手机号": "13800138004", "邮箱": "zhaoliu@company.com",
                "入职日期": "2024-06-01", "转正日期": "",
                "在职状态": "试用期", "基本工资": 18000, "社保城市": "北京",
                "汇报对象": "E005",
            },
            {
                "工号": "E005", "姓名": "钱七", "部门编码": "D002",
                "部门名称": "产品部", "岗位": "产品总监", "性别": "女",
                "手机号": "13800138005", "邮箱": "qianqi@company.com",
                "入职日期": "2021-09-01", "转正日期": "2021-12-01",
                "在职状态": "在职", "基本工资": 30000, "社保城市": "北京",
                "汇报对象": "",
            },
            {
                "工号": "E006", "姓名": "孙八", "部门编码": "D001",
                "部门名称": "技术部", "岗位": "实习工程师", "性别": "男",
                "手机号": "13800138006", "邮箱": "",
                "入职日期": "2026-01-10", "转正日期": "",
                "在职状态": "试用期", "基本工资": 5000, "社保城市": "北京",
                "汇报对象": "E001",
            },
            {
                "工号": "E007", "姓名": "周九", "部门编码": "D003",
                "部门名称": "市场部", "岗位": "市场专员", "性别": "女",
                "手机号": "13800138007", "邮箱": "zhoujiu@company.com",
                "入职日期": "2023-05-01", "离职日期": "2025-12-31",
                "在职状态": "离职", "基本工资": 12000, "社保城市": "北京",
                "汇报对象": "",
            },
        ]
        
        create_test_excel(cls.excel_path, test_employees)
    
    def setUp(self):
        """每个测试前加载管理器"""
        self.manager = EmployeeManager(self.excel_path)
    
    def test_load_employees(self):
        """测试加载员工数据"""
        self.assertEqual(len(self.manager.employees), 7)
        self.assertIn("E001", self.manager.employees)
        self.assertEqual(self.manager.employees["E001"].name, "张三")
    
    def test_get_employee(self):
        """测试获取单个员工"""
        emp = self.manager.getEmployee("E001")
        self.assertIsNotNone(emp)
        self.assertEqual(emp.name, "张三")
        self.assertEqual(emp.deptName, "技术部")
        self.assertEqual(emp.baseSalary, 20000)
    
    def test_get_nonexistent_employee(self):
        """测试获取不存在的员工"""
        emp = self.manager.getEmployee("E999")
        self.assertIsNone(emp)
    
    def test_list_all_active(self):
        """测试列出所有在职员工"""
        employees = self.manager.listEmployees()
        self.assertEqual(len(employees), 6)  # 不含离职的 E007
    
    def test_list_all_include_inactive(self):
        """测试列出包含离职员工"""
        employees = self.manager.listEmployees(includeInactive=True)
        self.assertEqual(len(employees), 7)
    
    def test_list_by_department(self):
        """测试按部门筛选"""
        employees = self.manager.listEmployees(deptCode="D001")
        names = [e.name for e in employees]
        self.assertIn("张三", names)
        self.assertIn("李四", names)
        self.assertIn("孙八", names)
        self.assertNotIn("赵六", names)
    
    def test_list_by_status(self):
        """测试按状态筛选"""
        employees = self.manager.listEmployees(status="试用期")
        names = [e.name for e in employees]
        self.assertIn("赵六", names)
        self.assertIn("孙八", names)
        self.assertNotIn("张三", names)
    
    def test_list_by_keyword(self):
        """测试关键词搜索"""
        # 搜索手机号
        employees = self.manager.listEmployees(keyword="13800138004")
        self.assertEqual(len(employees), 1)
        self.assertEqual(employees[0].name, "赵六")
        
        # 搜索姓名
        employees = self.manager.listEmployees(keyword="王")
        self.assertEqual(len(employees), 1)
        self.assertEqual(employees[0].name, "王五")
    
    def test_add_employee(self):
        """测试添加员工"""
        new_emp = Employee(
            empNo="E010", name="新员工", deptCode="D001",
            deptName="技术部", position="工程师",
            hireDate="2026-04-01", phone="13800138010"
        )
        ok, msg = self.manager.addEmployee(new_emp)
        self.assertTrue(ok)
        self.assertEqual(len(self.manager.employees), 8)
        self.assertIn("E010", self.manager.employees)
    
    def test_add_duplicate_employee(self):
        """测试添加重复工号"""
        new_emp = Employee(empNo="E001", name="另一个张三")
        ok, msg = self.manager.addEmployee(new_emp)
        self.assertFalse(ok)
        self.assertIn("已存在", msg)
    
    def test_add_employee_with_invalid_phone(self):
        """测试添加手机号不正确的员工"""
        new_emp = Employee(empNo="E011", name="测试", phone="123")
        ok, msg = self.manager.addEmployee(new_emp)
        self.assertFalse(ok)
    
    def test_update_employee(self):
        """测试更新员工"""
        ok, msg = self.manager.updateEmployee("E002", {
            "position": "高级工程师",
            "baseSalary": 18000,
        })
        self.assertTrue(ok)
        emp = self.manager.getEmployee("E002")
        self.assertEqual(emp.position, "高级工程师")
        self.assertEqual(emp.baseSalary, 18000)
    
    def test_update_nonexistent_employee(self):
        """测试更新不存在的员工"""
        ok, msg = self.manager.updateEmployee("E999", {"name": "不存在"})
        self.assertFalse(ok)
    
    def test_delete_soft(self):
        """测试软删除"""
        ok, msg = self.manager.deleteEmployee("E006", soft=True)
        self.assertTrue(ok)
        self.assertIn("离职", msg)
        
        emp = self.manager.getEmployee("E006")
        self.assertEqual(emp.status, "离职")
        self.assertIsNotNone(emp.leaveDate)
    
    def test_delete_hard(self):
        """测试物理删除"""
        ok, msg = self.manager.deleteEmployee("E007", soft=False)
        self.assertTrue(ok)
        self.assertIsNone(self.manager.getEmployee("E007"))
    
    def test_batch_update_status(self):
        """测试批量更新状态"""
        result = self.manager.batchUpdateStatus(
            ["E004", "E006"], "在职", "2026-04-01"
        )
        self.assertEqual(result["successCount"], 2)
        self.assertEqual(self.manager.employees["E004"].status, "在职")
    
    def test_get_statistics(self):
        """测试统计信息"""
        stats = self.manager.getStatistics()
        self.assertEqual(stats["total"], 7)
        self.assertEqual(stats["active"], 6)
        self.assertIn("在职", stats["byStatus"])
        self.assertIn("试用期", stats["byStatus"])
        self.assertIn("离职", stats["byStatus"])
        self.assertIn("技术部", stats["byDepartment"])
    
    def test_employee_count_by_dept(self):
        """测试部门人数统计"""
        count = self.manager.getEmployeeCount(deptCode="D001")
        self.assertEqual(count, 4)  # 张三、李四、王五、孙八


class TestEmployeeManagerOrg(unittest.TestCase):
    """组织架构关联测试"""
    
    @classmethod
    def setUpClass(cls):
        cls.test_dir = os.path.join(os.path.dirname(__file__), "test_data")
        os.makedirs(cls.test_dir, exist_ok=True)
        
        # 员工表
        cls.emp_path = os.path.join(cls.test_dir, "test_org_employees.xlsx")
        create_test_excel(cls.emp_path, [
            {
                "工号": "E001", "姓名": "张三", "部门编码": "DEV",
                "岗位": "开发", "在职状态": "在职", "汇报对象": "E003",
            },
            {
                "工号": "E002", "姓名": "李四", "部门编码": "DEV-FE",
                "岗位": "前端", "在职状态": "在职", "汇报对象": "E001",
            },
            {
                "工号": "E003", "姓名": "王五", "部门编码": "DEV",
                "岗位": "技术总监", "在职状态": "在职", "汇报对象": "",
            },
            {
                "工号": "E004", "姓名": "赵六", "部门编码": "PM",
                "岗位": "产品", "在职状态": "在职", "汇报对象": "E005",
            },
            {
                "工号": "E005", "姓名": "钱七", "部门编码": "PM",
                "岗位": "产品总监", "在职状态": "在职", "汇报对象": "",
            },
        ])
        
        # 组织架构
        cls.org_path = os.path.join(cls.test_dir, "test_org_structure.xlsx")
        create_org_excel(cls.org_path, [
            {"部门编码": "TECH", "部门名称": "技术中心", "上级部门": "", "负责人": "E003"},
            {"部门编码": "DEV", "部门名称": "开发部", "上级部门": "TECH", "负责人": "E003"},
            {"部门编码": "DEV-FE", "部门名称": "前端组", "上级部门": "DEV", "负责人": "E001"},
            {"部门编码": "DEV-BE", "部门名称": "后端组", "上级部门": "DEV", "负责人": "E002"},
            {"部门编码": "PM", "部门名称": "产品部", "上级部门": "", "负责人": "E005"},
        ])
    
    def setUp(self):
        self.manager = EmployeeManager(self.emp_path)
        self.manager.loadDepartments(self.org_path)
    
    def test_load_departments(self):
        """测试加载组织架构"""
        self.assertEqual(len(self.manager.departments), 5)
        self.assertIn("TECH", self.manager.departments)
        self.assertEqual(self.manager.departments["DEV-FE"].parentCode, "DEV")
    
    def test_get_dept_tree(self):
        """测试部门树"""
        tree = self.manager.getDeptTree()
        self.assertEqual(len(tree), 2)  # TECH 和 PM 两个根节点
        
        tech = next(t for t in tree if t["deptCode"] == "TECH")
        self.assertEqual(len(tech["children"]), 1)  # DEV
        
        dev = tech["children"][0]
        self.assertEqual(len(dev["children"]), 2)  # DEV-FE, DEV-BE
    
    def test_dept_tree_employee_count(self):
        """测试部门树人数统计"""
        tree = self.manager.getDeptTree()
        tech = next(t for t in tree if t["deptCode"] == "TECH")
        self.assertEqual(tech["employeeCount"], 3)  # DEV + DEV-FE 的员工
    
    def test_get_reporting_chain(self):
        """测试汇报链"""
        chain = self.manager.getReportingChain("E002")
        self.assertEqual(len(chain), 3)
        self.assertEqual(chain[0]["name"], "李四")
        self.assertEqual(chain[1]["name"], "张三")
        self.assertEqual(chain[2]["name"], "王五")
    
    def test_get_reporting_chain_top_level(self):
        """测试顶层员工汇报链"""
        chain = self.manager.getReportingChain("E003")
        self.assertEqual(len(chain), 1)
        self.assertEqual(chain[0]["name"], "王五")
    
    def test_validate_data(self):
        """测试数据校验"""
        result = self.manager.validateData()
        self.assertEqual(result["errorCount"], 0)
        self.assertIn("共检查 5 名员工", result["info"])
    
    def test_list_by_parent_dept(self):
        """测试按父部门筛选（包含子部门）"""
        # 筛选 TECH 应包含 DEV 和 DEV-FE 的员工
        employees = self.manager.listEmployees(deptCode="DEV")
        empNos = [e.empNo for e in employees]
        self.assertIn("E001", empNos)  # DEV
        self.assertIn("E002", empNos)  # DEV-FE
        self.assertIn("E003", empNos)  # DEV
        self.assertNotIn("E004", empNos)  # PM


class TestEmployeeManagerValidation(unittest.TestCase):
    """数据校验测试"""
    
    @classmethod
    def setUpClass(cls):
        cls.test_dir = os.path.join(os.path.dirname(__file__), "test_data")
        os.makedirs(cls.test_dir, exist_ok=True)
        
        cls.excel_path = os.path.join(cls.test_dir, "test_validation.xlsx")
        create_test_excel(cls.excel_path, [
            {
                "工号": "E001", "姓名": "张三", "部门编码": "D001",
                "在职状态": "在职", "合同开始日期": "2024-01-01",
                "合同结束日期": "2026-12-31",
            },
            {
                "工号": "E002", "姓名": "李四", "部门编码": "INVALID_DEPT",
                "在职状态": "在职", "汇报对象": "NONEXISTENT",
            },
            {
                "工号": "E003", "姓名": "",  # 姓名为空
                "在职状态": "在职",
            },
        ])
    
    def setUp(self):
        self.manager = EmployeeManager(self.excel_path)
    
    def test_validate_missing_name(self):
        """测试姓名缺失检查"""
        result = self.manager.validateData()
        self.assertTrue(any("姓名为空" in e for e in result["errors"]))
    
    def test_validate_dept_not_in_org(self):
        """测试部门不在组织架构中"""
        # 先创建组织架构并加载
        org_path = os.path.join(self.test_dir, "test_val_org.xlsx")
        create_org_excel(org_path, [
            {"部门编码": "D001", "部门名称": "技术部", "上级部门": "", "负责人": ""},
        ])
        self.manager.loadDepartments(org_path)
        result = self.manager.validateData()
        self.assertTrue(any("INVALID_DEPT" in w for w in result["warnings"]))


class TestColumnMapping(unittest.TestCase):
    """列映射测试"""
    
    def test_field_aliases_completeness(self):
        """测试别名覆盖完整性"""
        required_fields = ["empNo", "name", "hireDate", "status"]
        for field in required_fields:
            self.assertIn(field, FIELD_ALIASES)
            self.assertTrue(len(FIELD_ALIASES[field]) >= 2)
    
    def test_required_fields_defined(self):
        """测试必填字段定义"""
        self.assertIn("empNo", REQUIRED_FIELDS)
        self.assertIn("name", REQUIRED_FIELDS)


class TestMappingQuality(unittest.TestCase):
    """映射质量测试"""
    
    @classmethod
    def setUpClass(cls):
        cls.test_dir = os.path.join(os.path.dirname(__file__), "test_data")
        os.makedirs(cls.test_dir, exist_ok=True)
        
        cls.excel_path = os.path.join(cls.test_dir, "test_mapping.xlsx")
        create_test_excel(cls.excel_path, [
            {
                "工号": "E001", "姓名": "张三", "部门": "技术部",
                "入职时间": "2023-01-01", "状态": "在职",
            }
        ])
    
    def test_mapping_quality(self):
        """测试映射质量报告"""
        manager = EmployeeManager(self.excel_path)
        quality = manager.getMappingQuality()
        
        # 工号和姓名应该能映射
        self.assertTrue(quality["requiredMapped"])
        self.assertTrue(quality["details"]["empNo"]["mapped"])
        self.assertTrue(quality["details"]["name"]["mapped"])
    
    def test_export_mapping(self):
        """测试导出列映射"""
        manager = EmployeeManager(self.excel_path)
        mapping = manager.exportColumnMapping()
        self.assertIn("empNo", mapping)
        self.assertIn("name", mapping)


if __name__ == "__main__":
    if not OPENPYXL_AVAILABLE:
        print("❌ 跳过测试：openpyxl 未安装")
        print("请运行: pip install openpyxl")
        sys.exit(1)
    
    # 运行测试
    unittest.main(verbosity=2)
