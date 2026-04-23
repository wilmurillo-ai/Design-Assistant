"""
HR 智能体 - Excel 数据适配器
处理 Excel 文件的读取、写入和结构识别
支持 .xls（通过 xlrd）和 .xlsx/.xlsm（通过 openpyxl）格式
"""

import os
import re
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass
import json

# 注意：实际使用时需要安装 openpyxl 和 xlrd
# pip install openpyxl xlrd
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("警告：openpyxl 未安装，.xlsx 文件将无法处理")

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    print("警告：xlrd 未安装，.xls 文件将无法处理")


@dataclass
class ColumnInfo:
    """列信息"""
    index: int              # 列索引（从0开始）
    letter: str             # 列字母（A, B, C...）
    name: str               # 列名
    dataType: str           # 推断的数据类型
    sampleValues: List[Any] # 样例值（前5个非空值）


@dataclass
class TableStructure:
    """表格结构"""
    sheetName: str
    rowCount: int
    columnCount: int
    columns: List[ColumnInfo]
    hasHeader: bool


class ExcelAdapter:
    """Excel 适配器"""
    
    # 列名别名映射库
    COLUMN_ALIASES = {
        # 组织架构
        "deptCode": ["部门编码", "部门ID", "部门代码", "dept_id", "dept_code", "部门编号", "部门号"],
        "deptName": ["部门名称", "部门", "部门名", "dept_name", "department", "部门名称"],
        "parentCode": ["上级部门", "父部门", "上级编码", "parent_id", "parent_code", "上级部门编码", "父部门编码"],
        "manager": ["部门负责人", "负责人", "经理", "主管", "manager", "head", "部门经理", "部门主管"],
        "level": ["部门层级", "层级", "级别", "level", "grade", "部门级别"],
        
        # 员工
        "empNo": ["工号", "员工编号", "员工ID", "编号", "emp_id", "employee_id", "员工号", "职工号"],
        "name": ["姓名", "员工姓名", "名字", "name", "employee_name", "员工名字"],
        "position": ["岗位", "职位", "职务", "岗位名称", "position", "job_title", "职级"],
        "hireDate": ["入职日期", "入职时间", "入职日", "hire_date", "join_date", "入职年月", "到岗日期"],
        "status": ["在职状态", "状态", "员工状态", "status", "employment_status", "是否在职"],
        "phone": ["手机号", "电话", "联系电话", "手机", "phone", "mobile", "手机号码"],
        "email": ["邮箱", "电子邮件", "邮箱地址", "email", "邮件"],
        "idCard": ["身份证号", "身份证", "证件号", "id_card", "id_number", "身份证号码"],
        
        # 薪资
        "baseSalary": ["基本工资", "底薪", "基础工资", "base_salary", "basic_salary", "基本薪资"],
        "performanceBonus": ["绩效奖金", "绩效工资", "奖金", "performance_bonus", "绩效", "绩效工资"],
        "positionAllowance": ["岗位津贴", "津贴", "岗位补贴", "position_allowance", "职务津贴"],
        "transportAllowance": ["交通补贴", "交通津贴", "车补", "transport_allowance", "交通补助"],
        "mealAllowance": ["餐补", "餐饮补贴", "伙食补贴", "meal_allowance", "餐费补贴"],
        "communicationAllowance": ["通讯补贴", "通讯津贴", "communication_allowance", "电话补贴"],
        "socialInsuranceBase": ["社保基数", "缴费基数", "社保缴费基数"],
        "housingFundBase": ["公积金基数", "公积金缴费基数", "住房公积金额度"],
        "preTaxDeductions": ["税前扣减", "税前扣除", "考勤扣款", "迟到扣款", "事假扣款", "病假扣款"],
        "specialDeduction": ["专项附加扣除", "专项扣除", "附加扣除", "专项附加", "special_deduction"],
        "grossPay": ["应发合计", "应发工资", "应发", "gross_pay", "税前工资", "应发薪资"],
        "totalDeductions": ["扣款合计", "扣款总计", "total_deductions", "扣除合计"],
        "netPay": ["实发工资", "实发", "net_pay", "税后工资", "实发薪资", "到手工资"],
        
        # 考勤
        "year": ["年份", "年度", "年"],
        "month": ["月份", "月"],
        "shouldAttendDays": ["应出勤天数", "应出勤", "出勤标准", "应到天数"],
        "actualAttendDays": ["实际出勤天数", "实际出勤", "出勤天数", "实到天数"],
        "lateCount": ["迟到次数", "迟到", "迟到天数"],
        "earlyLeaveCount": ["早退次数", "早退", "早退天数"],
        "personalLeaveDays": ["事假天数", "事假", "事假(天)"],
        "sickLeaveDays": ["病假天数", "病假", "病假(天)"],
        "absentDays": ["旷工天数", "旷工", "旷工(天)"],
        "overtimeHours": ["加班小时", "加班时长", "加班小时数", "加班(h)"],
    }
    
    def __init__(self, filePath: str):
        self.filePath = filePath
        self.workbook = None
        self._xlsBook = None  # xlrd 用于 .xls 格式
        
        if not os.path.exists(filePath):
            raise FileNotFoundError(f"文件不存在: {filePath}")
        
        # 根据文件扩展名判断格式
        _, ext = os.path.splitext(filePath)
        ext = ext.lower()
        if ext == '.xls':
            if not XLRD_AVAILABLE:
                raise ImportError("xlrd 未安装，无法处理 .xls 文件。请执行: pip install xlrd")
            self._fileFormat = 'xls'
        elif ext in ('.xlsx', '.xlsm'):
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl 未安装，无法处理 .xlsx/.xlsm 文件。请执行: pip install openpyxl")
            self._fileFormat = 'xlsx'
        else:
            # 默认尝试 openpyxl（.xlsx），失败则提示
            if OPENPYXL_AVAILABLE:
                self._fileFormat = 'xlsx'
            elif XLRD_AVAILABLE:
                self._fileFormat = 'xls'
            else:
                raise ImportError("openpyxl 和 xlrd 均未安装，无法处理 Excel 文件")
    
    @property
    def fileFormat(self) -> str:
        """返回当前文件格式: 'xls' 或 'xlsx'"""
        return getattr(self, '_fileFormat', 'xlsx')
    
    def _loadWorkbook(self):
        """加载工作簿"""
        if self.fileFormat == 'xls':
            return self._loadXlsWorkbook()
        else:
            return self._loadXlsxWorkbook()
    
    def _loadXlsxWorkbook(self):
        """加载 .xlsx/.xlsm 工作簿（openpyxl）"""
        if self.workbook is None:
            self.workbook = load_workbook(self.filePath, data_only=True)
        return self.workbook
    
    def _loadXlsWorkbook(self):
        """加载 .xls 工作簿（xlrd）"""
        if self._xlsBook is None:
            self._xlsBook = xlrd.open_workbook(self.filePath)
        return self._xlsBook
    
    def getSheetNames(self) -> List[str]:
        """获取所有工作表名称"""
        if self.fileFormat == 'xls':
            book = self._loadXlsWorkbook()
            return book.sheet_names()
        else:
            wb = self._loadXlsxWorkbook()
            return wb.sheetnames
    
    def analyzeStructure(self, sheetName: Optional[str] = None) -> TableStructure:
        """
        分析表格结构
        
        Args:
            sheetName: 工作表名称，None 则使用第一个
        
        Returns:
            表格结构信息
        """
        if self.fileFormat == 'xls':
            return self._analyzeStructureXls(sheetName)
        else:
            return self._analyzeStructureXlsx(sheetName)
    
    def _analyzeStructureXlsx(self, sheetName: Optional[str] = None) -> TableStructure:
        """分析 .xlsx 表格结构"""
        wb = self._loadXlsxWorkbook()
        
        if sheetName is None:
            sheetName = wb.sheetnames[0]
        
        ws = wb[sheetName]
        
        # 获取数据范围
        maxRow = ws.max_row
        maxCol = ws.max_column
        
        if maxRow == 0 or maxCol == 0:
            raise ValueError("表格为空")
        
        # 检测是否有表头（第一行是否为列名）
        hasHeader = self._detectHeader(ws, maxCol)
        
        # 分析列信息
        columns = []
        headerRow = 1 if hasHeader else 0
        
        for colIdx in range(1, maxCol + 1):
            colLetter = get_column_letter(colIdx)
            
            # 获取列名
            if hasHeader:
                cellValue = ws.cell(row=1, column=colIdx).value
                colName = str(cellValue) if cellValue is not None else f"Column{colIdx}"
            else:
                colName = f"Column{colIdx}"
            
            # 获取样例值（前5个非空值）
            sampleValues = []
            for rowIdx in range(headerRow + 1, min(headerRow + 6, maxRow + 1)):
                value = ws.cell(row=rowIdx, column=colIdx).value
                if value is not None:
                    sampleValues.append(value)
            
            # 推断数据类型
            dataType = self._inferDataType(sampleValues)
            
            columns.append(ColumnInfo(
                index=colIdx - 1,
                letter=colLetter,
                name=colName,
                dataType=dataType,
                sampleValues=sampleValues[:5]
            ))
        
        return TableStructure(
            sheetName=sheetName,
            rowCount=maxRow - (1 if hasHeader else 0),
            columnCount=maxCol,
            columns=columns,
            hasHeader=hasHeader
        )
    
    def _analyzeStructureXls(self, sheetName: Optional[str] = None) -> TableStructure:
        """分析 .xls 表格结构（xlrd）"""
        book = self._loadXlsWorkbook()
        
        if sheetName is None:
            sheetName = book.sheet_names()[0]
        
        sheet = book.sheet_by_name(sheetName)
        
        maxRow = sheet.nrows
        maxCol = sheet.ncols
        
        if maxRow == 0 or maxCol == 0:
            raise ValueError("表格为空")
        
        # 检测是否有表头
        hasHeader = self._detectHeaderXls(sheet, maxCol)
        
        # 分析列信息
        columns = []
        headerRow = 1 if hasHeader else 0
        
        for colIdx in range(maxCol):
            colLetter = get_column_letter(colIdx + 1)
            
            # 获取列名
            if hasHeader:
                cellValue = sheet.cell_value(0, colIdx)
                colName = str(cellValue) if cellValue else f"Column{colIdx + 1}"
            else:
                colName = f"Column{colIdx + 1}"
            
            # 获取样例值（前5个非空值）
            sampleValues = []
            for rowIdx in range(headerRow, min(headerRow + 5, maxRow)):
                value = sheet.cell_value(rowIdx, colIdx)
                if value is not None and value != '':
                    # xlrd 日期处理：将 xlrd.xldate 转为字符串日期
                    if sheet.cell_type(rowIdx, colIdx) == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate_as_tuple(value, book.datemode)
                            value = f"{value[0]}-{value[1]:02d}-{value[2]:02d}"
                        except Exception:
                            pass
                    sampleValues.append(value)
            
            # 推断数据类型
            dataType = self._inferDataType(sampleValues)
            
            columns.append(ColumnInfo(
                index=colIdx,
                letter=colLetter,
                name=colName,
                dataType=dataType,
                sampleValues=sampleValues[:5]
            ))
        
        return TableStructure(
            sheetName=sheetName,
            rowCount=maxRow - (1 if hasHeader else 0),
            columnCount=maxCol,
            columns=columns,
            hasHeader=hasHeader
        )
    
    def _detectHeaderXls(self, sheet, maxCol: int) -> bool:
        """检测 .xls 表格第一行是否为表头"""
        if sheet.nrows < 2:
            return True
        
        headerStringCount = 0
        
        for colIdx in range(maxCol):
            firstRowValue = str(sheet.cell_value(0, colIdx)) if sheet.cell_value(0, colIdx) != '' else None
            secondRowValue = sheet.cell_value(1, colIdx)
            
            if firstRowValue and isinstance(firstRowValue, str) and len(firstRowValue) < 50:
                headerStringCount += 1
                
                # 第二行是数字 → 典型表头特征
                if isinstance(secondRowValue, (int, float)) and secondRowValue != '':
                    return True
                
                if isinstance(secondRowValue, str):
                    firstHasChinese = bool(re.search(r'[\u4e00-\u9fff]', firstRowValue))
                    if firstHasChinese and not re.search(r'[\u4e00-\u9fff]', secondRowValue):
                        return True
                    
                    if re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', firstRowValue):
                        return True
        
        return headerStringCount >= maxCol * 0.6
    
    def _detectHeader(self, ws, maxCol: int) -> bool:
        """
        检测第一行是否为表头

        判断逻辑：
        1. 第一行是字符串类型
        2. 后续行包含与第一行不同类型的值（数字、日期等）
        3. 或者第一行值与后续行值重复度低（说明第一行是唯一标识而非数据）
        """
        if ws.max_row < 2:
            return True  # 只有一行，当作表头
        
        # 统计第一行中字符串类型的列
        headerStringCount = 0
        totalCols = maxCol
        
        for colIdx in range(1, maxCol + 1):
            firstRowValue = ws.cell(row=1, column=colIdx).value
            secondRowValue = ws.cell(row=2, column=colIdx).value
            
            # 第一行是短字符串（像列名）
            if isinstance(firstRowValue, str) and len(firstRowValue) < 50:
                headerStringCount += 1
                
                # 第二行是不同类型的值 → 典型表头特征
                if isinstance(secondRowValue, (int, float)):
                    return True
                
                # 第二行也是字符串，检查是否看起来像数据（非列名）
                if isinstance(secondRowValue, str):
                    # 如果第一行像中文列名（纯中文/含中文），第二行像数据
                    firstHasChinese = bool(re.search(r'[\u4e00-\u9fff]', firstRowValue))
                    if firstHasChinese and not re.search(r'[\u4e00-\u9fff]', secondRowValue):
                        return True
                    
                    # 第一行像编码字段名（含下划线、驼峰），第二行是实际值
                    if re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', firstRowValue):
                        return True
        
        # 如果第一行大多数是短字符串（>60%），也认为是表头
        return headerStringCount >= totalCols * 0.6
    
    def _inferDataType(self, values: List[Any]) -> str:
        """推断数据类型"""
        if not values:
            return "string"
        
        # 检查是否为日期
        datePattern = re.compile(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$')
        dateCount = sum(1 for v in values if isinstance(v, str) and datePattern.match(v))
        if dateCount >= len(values) * 0.8:
            return "date"
        
        # 检查是否为数字
        numericCount = sum(1 for v in values if isinstance(v, (int, float)))
        if numericCount >= len(values) * 0.8:
            return "number"
        
        # 检查是否为枚举（值种类少）
        uniqueValues = set(str(v) for v in values if v is not None)
        if len(uniqueValues) <= 5 and len(values) >= 5:
            return "enum"
        
        return "string"
    
    def matchColumns(self, tableType: str, structure: TableStructure) -> Dict[str, Any]:
        """
        匹配标准字段与表格列
        
        Args:
            tableType: organization, employee, salary
            structure: 表格结构
        
        Returns:
            匹配结果
        """
        # 根据表格类型获取字段定义
        if tableType == "organization":
            fieldDefs = {k: v for k, v in self.COLUMN_ALIASES.items() 
                        if k in ["deptCode", "deptName", "parentCode", "manager", "level"]}
        elif tableType == "employee":
            fieldDefs = {k: v for k, v in self.COLUMN_ALIASES.items() 
                        if k in ["empNo", "name", "deptCode", "position", "hireDate", 
                                "status", "phone", "email", "idCard"]}
        elif tableType == "salary":
            fieldDefs = {k: v for k, v in self.COLUMN_ALIASES.items() 
                        if k in ["empNo", "baseSalary", "performanceBonus", "positionAllowance",
                                "transportAllowance", "mealAllowance", "communicationAllowance",
                                "socialInsuranceBase", "housingFundBase", "preTaxDeductions",
                                "grossPay", "totalDeductions", "netPay"]}
        else:
            raise ValueError(f"未知的表格类型: {tableType}")
        
        matches = []
        
        for fieldKey, aliases in fieldDefs.items():
            bestMatch = None
            bestScore = 0
            
            for col in structure.columns:
                score = self._calculateMatchScore(col.name, aliases)
                if score > bestScore:
                    bestScore = score
                    bestMatch = col
            
            matchInfo = {
                "field": fieldKey,
                "fieldName": aliases[0],
                "detectedColumn": bestMatch.name if bestMatch else None,
                "columnIndex": bestMatch.index if bestMatch else None,
                "confidence": bestScore,
                "isRequired": fieldKey in ["deptCode", "deptName", "empNo", "name", "baseSalary"],
                "dataType": bestMatch.dataType if bestMatch else "string"
            }
            matches.append(matchInfo)
        
        # 按置信度排序
        matches.sort(key=lambda x: x["confidence"], reverse=True)
        
        return {
            "tableType": tableType,
            "totalFields": len(fieldDefs),
            "matchedFields": sum(1 for m in matches if m["confidence"] > 0.5),
            "matches": matches
        }
    
    def _calculateMatchScore(self, columnName: str, aliases: List[str]) -> float:
        """计算列名匹配分数"""
        columnName = columnName.lower().strip()
        
        for alias in aliases:
            alias = alias.lower().strip()
            
            # 完全匹配
            if columnName == alias:
                return 1.0
            
            # 包含关系
            if alias in columnName or columnName in alias:
                return 0.8
            
            # 编辑距离相似（简化版）
            if self._similarity(columnName, alias) > 0.7:
                return 0.6
        
        return 0.0
    
    def _similarity(self, s1: str, s2: str) -> float:
        """计算字符串相似度（简化版）"""
        if not s1 or not s2:
            return 0.0
        
        # 使用 Jaccard 相似度
        set1 = set(s1)
        set2 = set(s2)
        intersection = len(set1 & set2)
        union = len(set1 | set2)
        
        return intersection / union if union > 0 else 0.0
    
    def readData(self, sheetName: Optional[str] = None, 
                 startRow: int = 2, maxRows: Optional[int] = None) -> List[Dict]:
        """
        读取表格数据
        
        Args:
            sheetName: 工作表名称
            startRow: 开始行（从1开始，xlsx模式；从0开始，xls模式会自动调整）
            maxRows: 最大读取行数
        
        Returns:
            数据列表，每条记录是一个字典
        """
        if self.fileFormat == 'xls':
            return self._readDataXls(sheetName, startRow, maxRows)
        else:
            return self._readDataXlsx(sheetName, startRow, maxRows)
    
    def _readDataXlsx(self, sheetName: Optional[str] = None, 
                       startRow: int = 2, maxRows: Optional[int] = None) -> List[Dict]:
        """读取 .xlsx 表格数据"""
        wb = self._loadXlsxWorkbook()
        
        if sheetName is None:
            sheetName = wb.sheetnames[0]
        
        ws = wb[sheetName]
        
        # 获取表头
        headers = []
        for colIdx in range(1, ws.max_column + 1):
            headerValue = ws.cell(row=1, column=colIdx).value
            headers.append(str(headerValue) if headerValue else f"Column{colIdx}")
        
        # 读取数据
        data = []
        endRow = ws.max_row + 1
        if maxRows:
            endRow = min(endRow, startRow + maxRows)
        
        for rowIdx in range(startRow, endRow):
            rowData = {}
            for colIdx, header in enumerate(headers, 1):
                value = ws.cell(row=rowIdx, column=colIdx).value
                rowData[header] = value
            
            # 跳过空行
            if any(v is not None for v in rowData.values()):
                data.append(rowData)
        
        return data
    
    def _readDataXls(self, sheetName: Optional[str] = None,
                      startRow: int = 2, maxRows: Optional[int] = None) -> List[Dict]:
        """读取 .xls 表格数据（xlrd）"""
        book = self._loadXlsWorkbook()
        
        if sheetName is None:
            sheetName = book.sheet_names()[0]
        
        sheet = book.sheet_by_name(sheetName)
        
        # 获取表头（xldr 行从 0 开始，第一行即表头）
        headers = []
        for colIdx in range(sheet.ncols):
            headerValue = sheet.cell_value(0, colIdx)
            headers.append(str(headerValue) if headerValue and headerValue != '' else f"Column{colIdx + 1}")
        
        # startRow 从 1 开始（第 2 行），xlrd 从 0 开始所以直接用 startRow-1 作为起始
        dataStart = startRow - 1  # 转为 0-based
        endRow = sheet.nrows
        if maxRows:
            endRow = min(endRow, dataStart + maxRows)
        
        data = []
        for rowIdx in range(dataStart, endRow):
            rowData = {}
            for colIdx, header in enumerate(headers):
                value = sheet.cell_value(rowIdx, colIdx)
                # xlrd 日期处理
                if sheet.cell_type(rowIdx, colIdx) == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_tuple(value, book.datemode)
                        value = f"{value[0]}-{value[1]:02d}-{value[2]:02d}"
                    except Exception:
                        pass
                elif value == '':
                    value = None
                rowData[header] = value
            
            # 跳过空行
            if any(v is not None for v in rowData.values()):
                data.append(rowData)
        
        return data
    
    def writeData(self, data: List[Dict], sheetName: str = "Sheet1", 
                  headers: Optional[List[str]] = None):
        """
        写入表格数据（仅支持 .xlsx 格式）
        
        Args:
            data: 数据列表
            sheetName: 工作表名称
            headers: 表头列表，None 则使用数据字典的 key
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 未安装，无法写入 Excel 文件")
        
        if not data:
            return
        
        # 确定表头
        if headers is None:
            headers = list(data[0].keys())
        
        # 创建新工作簿或加载现有工作簿
        if os.path.exists(self.filePath):
            wb = load_workbook(self.filePath)
        else:
            wb = Workbook()
        
        # 创建或清空工作表
        if sheetName in wb.sheetnames:
            ws = wb[sheetName]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheetName)
        
        # 写入表头
        for colIdx, header in enumerate(headers, 1):
            ws.cell(row=1, column=colIdx, value=header)
        
        # 写入数据
        for rowIdx, rowData in enumerate(data, 2):
            for colIdx, header in enumerate(headers, 1):
                value = rowData.get(header)
                ws.cell(row=rowIdx, column=colIdx, value=value)
        
        # 保存
        wb.save(self.filePath)


# 工具函数
def get_excel_support_status() -> Dict[str, bool]:
    """获取 Excel 格式支持状态"""
    return {
        "xlsx": OPENPYXL_AVAILABLE,
        "xls": XLRD_AVAILABLE,
    }


def analyze_excel_structure(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """分析 Excel 表格结构"""
    try:
        adapter = ExcelAdapter(file_path)
        structure = adapter.analyzeStructure(sheet_name)
        
        return {
            "success": True,
            "filePath": file_path,
            "sheetName": structure.sheetName,
            "rowCount": structure.rowCount,
            "columnCount": structure.columnCount,
            "hasHeader": structure.hasHeader,
            "columns": [
                {
                    "index": col.index,
                    "letter": col.letter,
                    "name": col.name,
                    "dataType": col.dataType,
                    "sampleValues": col.sampleValues[:3]  # 只返回前3个样例
                }
                for col in structure.columns
            ]
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


def match_columns(file_path: str, table_type: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """匹配表格列与标准字段"""
    try:
        adapter = ExcelAdapter(file_path)
        structure = adapter.analyzeStructure(sheet_name)
        result = adapter.matchColumns(table_type, structure)
        return {
            "success": True,
            **result
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


def read_excel_data(file_path: str, sheet_name: Optional[str] = None, 
                    max_rows: int = 100) -> Dict[str, Any]:
    """读取 Excel 数据"""
    try:
        adapter = ExcelAdapter(file_path)
        data = adapter.readData(sheet_name, maxRows=max_rows)
        return {
            "success": True,
            "filePath": file_path,
            "sheetName": sheet_name,
            "rowCount": len(data),
            "data": data
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }
