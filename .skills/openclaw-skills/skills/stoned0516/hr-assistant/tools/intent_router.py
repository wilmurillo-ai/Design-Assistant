"""
HR 智能体 - 意图路由引擎
将用户的自然语言意图路由到具体的 Tool 调用

架构：
  用户输入 → IntentClassifier（意图识别）→ ParameterExtractor（参数提取）→ ToolExecutor（执行）→ ResponseFormatter（响应格式化）
  
支持：
  - 关键词匹配（精确模式）
  - 正则表达式匹配（灵活模式）
  - 多轮对话上下文（上下文跟踪）
  - 参数缺失追问（智能补全）
"""

import re
import os
import json
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from enum import Enum
from datetime import datetime
from decimal import Decimal


# ============================================================
# 数据模型
# ============================================================

class IntentType(Enum):
    """意图类型"""
    # 初始化引导
    START_ONBOARDING = "start_onboarding"
    SET_STORAGE_TYPE = "set_storage_type"
    BIND_TABLE = "bind_table"
    CHECK_CONFIG = "check_config"
    RESET_CONFIG = "reset_config"
    
    # 员工查询
    READ_EMPLOYEE = "read_employee"
    LIST_EMPLOYEES = "list_employees"
    SEARCH_EMPLOYEE = "search_employee"
    
    # 员工操作
    ADD_EMPLOYEE = "add_employee"
    UPDATE_EMPLOYEE = "update_employee"
    DELETE_EMPLOYEE = "delete_employee"
    BATCH_UPDATE_STATUS = "batch_update_status"
    
    # 组织架构
    QUERY_DEPARTMENT = "query_department"
    GET_DEPT_TREE = "get_department_tree"
    GET_REPORTING_CHAIN = "get_reporting_chain"
    
    # 薪资
    CALCULATE_PAYROLL = "calculate_payroll"
    CALCULATE_BONUS = "calculate_bonus"
    CALCULATE_INSURANCE = "calculate_insurance"
    GET_TAX_BRACKET = "get_tax_bracket"
    
    # 考勤
    BIND_ATTENDANCE = "bind_attendance"
    VIEW_ATTENDANCE = "view_attendance"
    ATTENDANCE_SUMMARY = "attendance_summary"
    ATTENDANCE_RULES = "attendance_rules"
    
    # 报表
    EMPLOYEE_STATISTICS = "employee_statistics"
    DATA_VALIDATION = "data_validation"
    EXPORT_REPORT = "export_report"
    
    # 其他
    HELP = "help"
    UNKNOWN = "unknown"


@dataclass
class IntentResult:
    """意图识别结果"""
    intent: IntentType
    confidence: float          # 置信度 0-1
    matchedPatterns: List[str]  # 匹配到的模式
    entities: Dict[str, Any] = field(default_factory=dict)  # 提取到的实体
    missingParams: List[str] = field(default_factory=list)   # 缺失的必要参数
    needsFollowUp: bool = False  # 是否需要追问


@dataclass
class ConversationContext:
    """对话上下文"""
    sessionId: str = ""
    turns: List[Dict] = field(default_factory=list)  # 对话历史
    lastIntent: Optional[IntentType] = None           # 上一个意图
    pendingAction: Optional[Dict] = None              # 待完成的操作
    lastResults: Optional[Dict] = None                # 上一次操作结果
    lastQuery: str = ""                               # 上一次查询
    
    def addTurn(self, role: str, content: str, intent: Optional[IntentType] = None):
        """添加对话轮次"""
        self.turns.append({
            "role": role,
            "content": content,
            "intent": intent.value if intent else None,
            "timestamp": datetime.now().isoformat()
        })
        # 保留最近 20 轮
        if len(self.turns) > 20:
            self.turns = self.turns[-20:]


# ============================================================
# 意图分类器
# ============================================================

class IntentClassifier:
    """
    意图分类器
    三层匹配策略：精确关键词 → 正则表达式 → 模糊匹配
    """
    
    # 意图模式定义
    INTENT_PATTERNS = {
        # === 初始化引导 ===
        IntentType.START_ONBOARDING: {
            "keywords": ["开始初始化", "初始化", "首次使用", "绑定表格", "设置数据", "开始设置", "首次绑定"],
            "regex": [
                r"(?:我要|帮我|想要|准备)?(?:开始|进行)?初始化",
                r"首次使用",
                r"(?:绑定|关联|设置).*(?:表格|数据)",
            ],
            "excludeKeywords": ["重新", "重置", "再次", "修改"],
            "examples": ["开始初始化", "我要初始化", "首次使用", "帮我绑定表格"]
        },
        IntentType.SET_STORAGE_TYPE: {
            "keywords": ["存储方式", "数据源", "excel", "存储类型"],
            "regex": [
                r"(?:选择|设置|使用)(?:Excel|excel)",
                r"数据(?:存储|来源|源)",
            ],
            "examples": ["用Excel存储", "选择Excel", "设置存储方式"]
        },
        IntentType.BIND_TABLE: {
            "keywords": ["绑定组织架构", "绑定花名册", "绑定薪资", "上传表格", "指定文件"],
            "regex": [
                r"绑定(?:组织架构|部门|花名册|员工|薪资|工资)",
                r"(?:上传|指定|加载).*(?:组织架构|花名册|薪资|Excel|文件)",
            ],
            "examples": ["绑定组织架构表", "上传花名册", "指定薪资表"]
        },
        IntentType.CHECK_CONFIG: {
            "keywords": ["查看配置", "当前配置", "配置状态", "绑定了哪些", "初始化状态"],
            "regex": [
                r"(?:查看|看看|检查|显示)(?:配置|设置|状态)",
                r"绑定了哪些",
                r"当前.*(?:配置|设置)",
            ],
            "examples": ["查看配置", "绑定了哪些表格", "配置状态"]
        },
        IntentType.RESET_CONFIG: {
            "keywords": ["重新初始化", "重置配置", "重新开始", "修改配置", "清除配置"],
            "regex": [
                r"(?:重新|再次)(?:初始化|设置|开始)",
                r"(?:重置|清除|删除|重写)(?:配置|设置)",
            ],
            "examples": ["重新初始化", "重置配置"]
        },
        
        # === 员工查询 ===
        IntentType.READ_EMPLOYEE: {
            "keywords": ["查员工", "员工信息", "工号", "员工详情"],
            "regex": [
                r"(?:查|看|查看|显示)(?:员工?|人员?).*?(?:信息|详情|资料)",
                r"(?:员工|人员?)\s*(?:E\d{3,}|编号)\s*(?:的信息|详情)?",  # E001的信息
                r"(?:工号|编号)\s*(?:是|=|：)\s*(\w+)",
                r"(.{2,4})\s*(?:的)?(?:信息|详情|资料|情况)$",  # 张伟的信息、王芳详情
            ],
            "excludeKeywords": ["离职", "辞职", "删除", "移除"],
            "examples": ["查看E001的信息", "工号E002的详情", "员工E003"]
        },
        IntentType.LIST_EMPLOYEES: {
            "keywords": ["员工列表", "花名册", "所有员工", "员工名单", "人员列表"],
            "regex": [
                r"(?:查看|列出|显示|看看).*(?:所有|全部)?(?:员工|人员|花名册|名单)",
                r"(?:有哪些|列出|罗列)(?:员工|人员|人)",
                r"(?:显示|看看)(?:花名册|员工列表)",
            ],
            "examples": ["列出所有员工", "看看花名册", "显示员工列表"]
        },
        IntentType.SEARCH_EMPLOYEE: {
            "keywords": ["搜索员工", "查找员工", "找员工", "筛选员工", "查询员工"],
            "regex": [
                r"(?:搜索|查找|找|筛选|查询).*(?:员工|人员|人)",
                r"(?:部门|技术部|人事部).*(?:员工|人)",
                r"(?:试用期|离职|在职|正式).*(?:员工|人)",
                r"^(?!.*(?:添加|新增|录入|入职|多少|几个|有多少|统计)).+?(?:部|中心|组|团队)(?:的)?(?:员工|人员|人)$",  # 前端开发部员工（排除更具体的意图）
            ],
            "excludeKeywords": ["添加", "新增", "录入", "入职", "多少", "几个", "统计"],
            "examples": ["查找技术部员工", "试用期员工有哪些", "搜索张伟"]
        },
        
        # === 员工操作 ===
        IntentType.ADD_EMPLOYEE: {
            "keywords": ["添加员工", "新增员工", "入职登记", "新人入职", "录入员工", "加个员工"],
            "regex": [
                r"(?:添加|新增|录入|登记|加入).*(?:员工|人员|新人|人)",
                r"(?:新人|新员工|新来).*(?:入职|登记|加入)",
            ],
            "examples": ["添加新员工", "新人入职登记", "录入一个员工"]
        },
        IntentType.UPDATE_EMPLOYEE: {
            "keywords": ["修改员工", "更新信息", "调岗", "调薪", "转正", "改信息"],
            "regex": [
                r"(?:修改|更新|变更|调整).*(?:员工|信息|薪资|岗位)",
                r"调(?:岗|薪|部门|职位)",
                r"(.{2,4})\s*(?:转正|转正了)",
                r"(?:改|修改|更新)(?:员工?.*)(?:信息|数据|资料)",
            ],
            "examples": ["张伟转正", "调薪E001", "修改员工信息"]
        },
        IntentType.DELETE_EMPLOYEE: {
            "keywords": ["删除员工", "移除员工", "员工离职", "办离职", "移除人员"],
            "regex": [
                r"(?:删除|移除).*(?:员工|人员)",
                r"(?:办|办理|处理|走).*(?:离职|辞职)",
                r"(.{2,4})\s*(?:离职|辞职|走了|不干了)",
                r"[Ee]\d{2,4}\s*(?:离职|辞职|走人|离开)",
            ],
            "examples": ["员工E003离职", "办离职", "删除E005"]
        },
        IntentType.BATCH_UPDATE_STATUS: {
            "keywords": ["批量转正", "批量离职", "批量更新", "一起转正", "一起离职", "批量操作"],
            "regex": [
                r"批量.+(?:转正|离职|更新|修改)",
                r"(?:全部|所有|一起|同时).+(?:转正|离职)",
                r"把这(?:些|几|批).+(?:转正|离职)",
            ],
            "examples": ["批量转正", "这3个人一起离职", "批量更新状态"]
        },
        
        # === 组织架构 ===
        IntentType.QUERY_DEPARTMENT: {
            "keywords": ["部门信息", "部门详情", "哪个部门", "部门查询"],
            "regex": [
                r"(?:查|看|查看|显示).*(?:部门信息|部门详情)",
                r"(?:哪个|什么)部门",
            ],
            "examples": ["查看部门信息", "技术部详情"]
        },
        IntentType.GET_DEPT_TREE: {
            "keywords": ["组织架构", "部门结构", "部门树", "部门列表", "组织结构", "架构图"],
            "regex": [
                r"(?:查看|显示|看看).*(?:组织架构|部门结构|组织结构|架构图)",
                r"(?:有哪些|列出|显示)(?:部门)",
                r"组织架构",
            ],
            "examples": ["查看组织架构", "部门结构", "有哪些部门"]
        },
        IntentType.GET_REPORTING_CHAIN: {
            "keywords": ["汇报链", "汇报关系", "上级", "直属上级", "汇报对象", "组织链"],
            "regex": [
                r"(?:查看|显示|看看).*(?:汇报链|汇报关系|汇报路线)",
                r"(.{2,4}).*(?:的)?(?:上级|汇报给谁|向谁汇报|直属上级)",
                r"(?:谁是|谁是其).*(?:上级|汇报对象|直属上级)",
            ],
            "examples": ["张伟的汇报链", "E003的上级是谁", "查看汇报关系"]
        },
        
        # === 薪资 ===
        IntentType.CALCULATE_PAYROLL: {
            "keywords": ["计算薪资", "算工资", "跑工资", "薪资计算", "本月工资", "工资条"],
            "regex": [
                r"(?:计算|算|核算|跑|生成).*(?:薪资|工资|薪酬)",
                r"(?:本月|这个月|当月).*(?:工资|薪资)",
                r"工资条",
                r"(?:给|帮|为).*(?:算|计算).*(?:工资|薪资)",
            ],
            "examples": ["计算本月薪资", "算工资", "帮E001算工资"]
        },
        IntentType.CALCULATE_BONUS: {
            "keywords": ["年终奖", "年终", "奖金计算", "十三薪", "年终奖个税"],
            "regex": [
                r"(?:计算|算|核算).*(?:年终奖|奖金|十三薪)",
                r"年终奖(?:个税|税|是多少)",
            ],
            "examples": ["计算年终奖", "年终奖个税是多少"]
        },
        IntentType.CALCULATE_INSURANCE: {
            "keywords": ["社保", "公积金", "五险一金", "社保计算", "公积金计算"],
            "regex": [
                r"(?:计算|算|查看|查).*(?:社保|公积金|五险一金)",
                r"(.+)(?:的)?(?:社保|公积金|五险一金)(?:是|多少|是多少)?",
            ],
            "examples": ["计算社保", "北京公积金多少", "五险一金"]
        },
        IntentType.GET_TAX_BRACKET: {
            "keywords": ["税率", "个税税率", "税级", "适用税率"],
            "regex": [
                r"(?:查看|查|适用|当前).*(?:税率|税级|个税税率)",
            ],
            "examples": ["查看税率", "适用税率是多少"]
        },
        
        # === 考勤 ===
        IntentType.BIND_ATTENDANCE: {
            "keywords": ["绑定考勤", "上传考勤", "考勤表", "考勤文件", "导入考勤"],
            "regex": [
                r"(?:绑定|上传|导入|加载).*(?:考勤|打卡|出勤)",
                r"考勤(?:表格|文件|表|数据)",
            ],
            "excludeKeywords": ["查看", "查看考勤", "统计", "汇总"],
            "examples": ["绑定考勤表", "上传考勤文件", "导入考勤数据"]
        },
        IntentType.VIEW_ATTENDANCE: {
            "keywords": ["查看考勤", "考勤记录", "出勤记录", "打卡记录", "考勤详情"],
            "regex": [
                r"(?:查看|看看|显示|查).*(?:考勤|出勤|打卡).*(?:记录|详情|数据)?",
                r"(?:考勤|出勤|打卡).*(?:记录|详情|数据)",
                r"(.{2,4}).*(?:的)?(?:考勤|出勤|打卡)",
            ],
            "excludeKeywords": ["绑定", "上传", "导入", "汇总", "统计"],
            "examples": ["查看考勤记录", "张伟的考勤", "本月出勤记录"]
        },
        IntentType.ATTENDANCE_SUMMARY: {
            "keywords": ["考勤汇总", "考勤统计", "出勤统计", "考勤月报", "考勤报表"],
            "regex": [
                r"(?:考勤|出勤).*(?:汇总|统计|月报|报表|概览)",
                r"(?:本月|这个月|当月).*(?:考勤|出勤).*(?:汇总|统计|情况)?",
                r"(?:汇总|统计|月报).*(?:考勤|出勤)",
            ],
            "examples": ["考勤汇总", "本月考勤统计", "出勤月报"]
        },
        IntentType.ATTENDANCE_RULES: {
            "keywords": ["考勤规则", "扣减规则", "迟到规则", "考勤制度", "扣款规则"],
            "regex": [
                r"(?:查看|设置|修改|调整).*(?:考勤规则|扣减规则|考勤制度|扣款标准)",
                r"(?:考勤规则|扣减规则|考勤制度|扣款标准)",
            ],
            "examples": ["查看考勤规则", "扣减规则是什么", "修改迟到扣款标准"]
        },
        
        # === 报表 ===
        IntentType.EMPLOYEE_STATISTICS: {
            "keywords": ["员工统计", "人数统计", "部门人数", "合同到期", "生日提醒", "数据统计", "人员统计", "人力统计"],
            "regex": [
                r"(?:查看|显示|看看).*(?:统计|人数|数据)",
                r"(?:有多少|几个|多少)(?:员工|人|在职|试用期)",
                r"合同(?:到期|续签)",
                r"(?:生日|节假日)提醒",
                r"(?:员工|人员|人力).*(?:统计|概况|概览)",
            ],
            "examples": ["员工统计", "有多少员工", "合同到期提醒"]
        },
        IntentType.DATA_VALIDATION: {
            "keywords": ["数据校验", "检查数据", "数据质量", "校验花名册", "数据检查", "校验数据", "检查花名册"],
            "regex": [
                r"(?:校验|检查|审查|验证).*(?:数据|花名册|信息)",
                r"数据(?:质量|完整性|准确性)",
                r"花名册.*(?:校验|检查|审查|验证)",
            ],
            "examples": ["校验花名册数据", "检查数据完整性"]
        },
        IntentType.EXPORT_REPORT: {
            "keywords": ["导出", "生成报表", "导出报表", "统计报告", "人力报表", "导出Excel", "导出工资条"],
            "regex": [
                r"(?:导出|生成|输出).*(?:报表|报告|Excel|工资条|数据)",
                r"生成.+(?:报表|报告)",
            ],
            "examples": ["导出员工报表", "生成薪资报表", "导出Excel"]
        },
        
        # === 帮助 ===
        IntentType.HELP: {
            "keywords": ["帮助", "帮助我", "怎么用", "使用说明", "功能", "你能做什么", "你会什么"],
            "regex": [
                r"(?:帮助|帮忙|help|怎么用|使用说明|功能列表)",
                r"你(?:能|会|可以)(?:做什么|干什么|帮忙)",
                r"有哪些功能",
            ],
            "examples": ["帮助", "你能做什么", "怎么用"]
        },
    }
    
    def classify(self, text: str, context: Optional[ConversationContext] = None) -> IntentResult:
        """
        分类用户输入的意图
        
        Args:
            text: 用户输入文本
            context: 对话上下文（用于多轮对话消歧）
        
        Returns:
            意图识别结果
        """
        text = text.strip()
        if not text:
            return IntentResult(
                intent=IntentType.UNKNOWN,
                confidence=0,
                matchedPatterns=[]
            )
        
        candidates = []
        
        # 第一层：关键词精确匹配（分两遍：精确全匹配 > 包含匹配）
        # 先收集所有匹配
        all_kw_matches = []
        for intent, patterns in self.INTENT_PATTERNS.items():
            exclude = patterns.get("excludeKeywords", [])
            if any(ex.lower() in text.lower() for ex in exclude):
                continue
            for kw in patterns["keywords"]:
                if kw.lower() in text.lower():
                    # 按匹配长度排序（更长的匹配更精确）
                    all_kw_matches.append((len(kw), intent, kw))
                    break
        
        # 如果多个意图匹配，选择关键词最长的（更精确）
        if all_kw_matches:
            all_kw_matches.sort(key=lambda x: -x[0])
            best_len, best_intent, best_kw = all_kw_matches[0]
            candidates.append((best_intent, 0.95, [best_kw]))
        
        # 第二层：正则表达式匹配
        for intent, patterns in self.INTENT_PATTERNS.items():
            exclude = patterns.get("excludeKeywords", [])
            if any(ex.lower() in text.lower() for ex in exclude):
                continue
            for regex in patterns.get("regex", []):
                match = re.search(regex, text, re.IGNORECASE)
                if match:
                    candidates.append((intent, 0.85, [regex]))
                    break  # 每个意图只取第一个匹配的正则
        
        # 第三层：模糊匹配（字符级相似度）
        if not candidates:
            for intent, patterns in self.INTENT_PATTERNS.items():
                for example in patterns.get("examples", []):
                    similarity = self._fuzzyMatch(text, example)
                    if similarity > 0.6:
                        candidates.append((intent, similarity * 0.8, [example]))
                        break
        
        # 如果有多个候选，选择置信度最高的
        if candidates:
            candidates.sort(key=lambda x: x[1], reverse=True)
            best_intent, best_conf, best_patterns = candidates[0]
            
            return IntentResult(
                intent=best_intent,
                confidence=best_conf,
                matchedPatterns=best_patterns
            )
        
        # 尝试从上下文推断
        if context and context.lastIntent:
            # 如果用户输入很短（如"确认"、"好的"、"是"、"E001"），可能是在回应上一次的追问
            short_responses = {"确认", "好的", "是", "对", "yes", "ok", "嗯", "可以", "行"}
            if text in short_responses and context.pendingAction:
                return IntentResult(
                    intent=context.lastIntent,
                    confidence=0.7,
                    matchedPatterns=["context_continuation"]
                )
            
            # 如果用户输入的是单个工号（如 E001），可能是补全参数
            if re.match(r'^[Ee]\d{3,}$', text) or re.match(r'^\d{4,}$', text):
                return IntentResult(
                    intent=context.lastIntent,
                    confidence=0.6,
                    matchedPatterns=["param_completion"]
                )
        
        return IntentResult(
            intent=IntentType.UNKNOWN,
            confidence=0.3,
            matchedPatterns=[]
        )
    
    @staticmethod
    def _fuzzyMatch(text: str, pattern: str) -> float:
        """简单的字符级模糊匹配（Jaccard 相似度）"""
        text_chars = set(text.lower())
        pattern_chars = set(pattern.lower())
        if not pattern_chars:
            return 0
        intersection = text_chars & pattern_chars
        union = text_chars | pattern_chars
        return len(intersection) / len(union)


# ============================================================
# 参数提取器
# ============================================================

class ParameterExtractor:
    """
    参数提取器
    从自然语言中提取 Tool 调用所需的参数
    """
    
    # 通用正则模式
    PATTERNS = {
        "empNo": [
            r'[Ee](\d{2,4})',
            r'(?:工号|编号|员工编号|empNo|emp_id)\s*(?:是|：|:|=)?\s*([Ee]?\d{2,4})',
        ],
        "empName": [
            r'(?:名字|姓名|员工|工号为\s*[Ee]?\d+\s*的|叫)\s*([^\s,，。、]{2,4})',
        ],
        "deptCode": [
            r'[Dd](\d{2,4})',
            r'(?:部门编码|部门ID)\s*(?:是|：|:|=)?\s*([Dd]?\d{2,4})',
        ],
        "deptName": [
            r'((?:技术|研发|产品|设计|运营|市场|销售|财务|人事|HR|行政|法务|客服|测试|前端|后端|移动|数据|战略|总经办|总裁办|人力资源)(?:部|中心|组|团队|事业部|部门)?)',
        ],
        "filePath": [
            r'(?:文件|路径|地址)\s*(?:是|：|:|=)?\s*([/\w.\-~]+\.(xlsx|xls|csv))',
            r'([/\w.\-~]+\.(xlsx|xls|csv))',
        ],
        "year": [
            r'(202[0-9]|19[0-9]{2})\s*年?',
        ],
        "month": [
            r'(1[0-2]|[1-9])\s*月?',
        ],
        "salary": [
            r'(?:工资|薪资|月薪|基本工资|底薪)\s*(?:是|：|:|=)?\s*(\d+(?:\.\d+)?)',
            r'(\d+(?:,\d{3})*(?:\.\d+)?)\s*(?:元|块|万)',
        ],
        "city": [
            r'(北京|上海|广州|深圳|杭州|成都|武汉|南京|重庆|天津|苏州|西安|长沙|沈阳|青岛|大连|宁波|厦门|福州|济南|合肥|昆明|郑州|哈尔滨|石家庄|太原|贵阳|南宁|拉萨|银川|呼和浩特|海口|兰州|长春|珠海)',
        ],
        "status": [
            r'(在职|正式|试用期|试用|离职|已离职)',
        ],
        "phone": [
            r'(1[3-9]\d{9})',
        ],
        "email": [
            r'([\w.+-]+@[\w-]+\.[\w.]+)',
        ],
    }
    
    # 数量表达
    AMOUNT_PATTERNS = [
        (r'(?:这|那|这(\d+)|那(\d+)|(\d+))(?:个|位|名|人|项|条)', "count"),
    ]
    
    def extract(self, text: str, intent: IntentResult, context: Optional[ConversationContext] = None) -> Dict[str, Any]:
        """
        从文本中提取参数
        
        Args:
            text: 用户输入
            intent: 已识别的意图
            context: 对话上下文
        
        Returns:
            参数字典
        """
        params = {}
        text_lower = text.lower()
        
        # 根据意图类型选择性地提取参数
        intent = intent.intent
        
        if intent in (
            IntentType.READ_EMPLOYEE, IntentType.UPDATE_EMPLOYEE, 
            IntentType.DELETE_EMPLOYEE, IntentType.GET_REPORTING_CHAIN
        ):
            params = self._extractEmployeeIdentifiers(text, params, context)
        
        if intent in (IntentType.SEARCH_EMPLOYEE, IntentType.LIST_EMPLOYEES):
            params = self._extractFilters(text, params)
        
        if intent == IntentType.ADD_EMPLOYEE:
            params = self._extractAddEmployeeParams(text, params)
        
        if intent == IntentType.UPDATE_EMPLOYEE:
            params = self._extractUpdateParams(text, params)
        
        if intent == IntentType.CALCULATE_PAYROLL:
            params = self._extractPayrollParams(text, params, context)
        
        if intent == IntentType.CALCULATE_BONUS:
            params = self._extractBonusParams(text, params)
        
        if intent == IntentType.CALCULATE_INSURANCE:
            params = self._extractInsuranceParams(text, params)
        
        if intent == IntentType.BIND_ATTENDANCE:
            params = self._extractBindAttendanceParams(text, params)
        
        if intent in (IntentType.VIEW_ATTENDANCE, IntentType.ATTENDANCE_SUMMARY):
            params = self._extractAttendanceParams(text, params)
        
        if intent == IntentType.ATTENDANCE_RULES:
            params = self._extractAttendanceRulesParams(text, params)
        
        if intent == IntentType.BIND_TABLE:
            params = self._extractBindParams(text, params)
        
        if intent == IntentType.BATCH_UPDATE_STATUS:
            params = self._extractBatchParams(text, params)
        
        if intent == IntentType.EXPORT_REPORT:
            params = self._extractExportParams(text, params)
        
        if intent == IntentType.EMPLOYEE_STATISTICS:
            params = self._extractStatsParams(text, params)
        
        # 通用参数提取（兜底）
        params = self._extractCommonParams(text, params)
        
        # 从上下文继承缺失参数
        if context and context.lastResults:
            for key, value in context.lastResults.get("extractedParams", {}).items():
                if key not in params:
                    params[key] = value
        
        return params
    
    def _extractEmployeeIdentifiers(self, text: str, params: Dict, context: Optional[ConversationContext]) -> Dict:
        """提取员工标识（工号或姓名）"""
        # 提取工号
        emp_match = re.search(r'[Ee](\d{2,4})', text)
        if emp_match:
            params["empNo"] = f"E{emp_match.group(1).zfill(3)}"
        
        # 提取姓名 - 先尝试从句首提取（如"张伟转正"、"王芳离职"）
        name = self._extractChineseName(text)
        if name:
            params["empName"] = name
        
        return params
    
    @staticmethod
    def _extractChineseName(text: str) -> Optional[str]:
        """从文本中提取中文姓名"""
        # 常见姓氏（覆盖98%+的中文姓氏）
        common_surnames = (
            "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜"
            "戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳酆鲍史唐"
            "费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元卜顾孟平黄"
            "和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊纪舒屈项祝董梁"
            "杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林刁钟徐邱骆高夏蔡田樊胡凌霍"
            "虞万支柯昝管卢莫经房裘缪干解应宗丁宣贲邓郁单杭洪包诸左石崔吉钮"
            "龚程嵇邢滑裴陆荣翁荀羊於惠甄麴家封芮羿储靳汲邴糜松井段富巫乌焦巴"
            "弓牧隗山谷车侯宓蓬全郗班仰秋仲伊宫宁仇栾暴甘钭厉戎祖武符刘景詹束龙叶幸司韶郜黎蓟薄印宿白怀蒲邰从鄂索咸籍赖卓蔺屠蒙池乔阴郁胥能苍双闻莘党翟谭贡劳逄姬申扶堵冉宰郦雍却璩桑桂濮牛寿通边扈燕冀郏浦尚农温别庄晏柴瞿阎充慕连茹习宦艾鱼容向古易慎戈廖庾终暨居衡步都耿满弘匡国文寇广禄阙东欧殳沃利蔚越夔隆师巩厍聂晁勾敖融冷訾辛阚那简饶空曾毋沙乜养鞠须丰巢关蒯相查后荆红游竺权逯盖益桓公"
        )
        
        # 姓名+操作后缀的模式列表
        action_suffixes = [
            r'(?:转正|离职|辞职|调岗|调薪|的汇报链|的上级|的汇报|的信息|的详情|的工资|的薪资|的数据)',
        ]
        
        # 模式1: "姓名+操作后缀"（如"张伟转正"、"赵雪的汇报链"）
        for suffix in action_suffixes:
            match = re.match(r'^([' + common_surnames + r'][\u4e00-\u9fff]{1,2})' + suffix, text)
            if match:
                name = match.group(1)
                if 2 <= len(name) <= 3:
                    return name
        
        # 模式2: "操作+姓名"（如"查看张伟"）
        match = re.search(r'(?:查|看|查看|显示|修改|更新|删除)\s*([' + common_surnames + r'][\u4e00-\u9fff]{1,2})', text)
        if match:
            name = match.group(1)
            if 2 <= len(name) <= 3:
                return name
        
        # 模式3: "帮/给/为 + 姓名"
        match = re.search(r'(?:帮|给|为)\s*([' + common_surnames + r'][\u4e00-\u9fff]{1,2})', text)
        if match:
            name = match.group(1)
            if 2 <= len(name) <= 3:
                return name
        
        return None
    
    def _extractEmployeeIdentifiers_legacy(self, text: str, params: Dict, context: Optional[ConversationContext]) -> Dict:
        """提取员工标识（工号或姓名）- 旧版备用"""
        # 提取工号
        emp_match = re.search(r'[Ee](\d{2,4})', text)
        if emp_match:
            params["empNo"] = f"E{emp_match.group(1).zfill(3)}"
        
        # 提取姓名（2-4个中文字符，排除常见词）
        name_patterns = [
            r'(?:查|看|查看|显示|修改|删除|更新|转正|离职|调岗|调薪|办|计算|算)\s*([^的]{2,4})\s*(?:的)?(?:信息|详情|工资|薪资|数据|员工)?',
            r'(?:帮|给|为)\s*([^的]{2,4})\s*',
            r'([^\s,，。、]{2,4})\s*(?:转正|离职|调岗|调薪|的上级|的汇报)',
        ]
        for pattern in name_patterns:
            match = re.search(pattern, text)
            if match:
                name = match.group(1)
                # 过滤掉常见非人名词
                non_names = {"员工", "人员", "信息", "配置", "薪资", "工资", "部门", "所有", "本月", "花名册", "数据", "帮助", "统计", "报表", "年终奖"}
                if name not in non_names and not re.match(r'^[\d\w]+$', name):
                    params["empName"] = name
                    break
        
        return params
    
    def _extractFilters(self, text: str, params: Dict) -> Dict:
        """提取筛选条件"""
        # 状态筛选
        for status_pattern in self.PATTERNS["status"]:
            match = re.search(status_pattern, text)
            if match:
                params["status"] = match.group(1)
                break
        
        # 部门筛选
        for dept_pattern in self.PATTERNS["deptName"]:
            match = re.search(dept_pattern, text)
            if match:
                params["deptName"] = match.group(1)
                break
        
        # 关键词（剩余文本去掉已知模式词）
        keywords = []
        noise_words = {"查看", "显示", "列出", "搜索", "查找", "看看", "的", "有哪些", "列表",
                       "员工", "人员", "花名册", "名单", "所有", "全部", "帮我", "请", "想要",
                       "一下", "一下", "吗", "呢", "？", "?", "请"}
        cleaned = text
        for w in noise_words:
            cleaned = cleaned.replace(w, "")
        cleaned = cleaned.strip()
        if cleaned and len(cleaned) >= 2:
            keywords.append(cleaned)
        if keywords:
            params["keyword"] = " ".join(keywords)
        
        return params
    
    def _extractAddEmployeeParams(self, text: str, params: Dict) -> Dict:
        """提取新增员工参数"""
        params = self._extractCommonParams(text, params)
        
        # 尝试解析结构化输入（如"姓名张三，工号E016，技术部，15000"）
        # 工号
        emp_match = re.search(r'[Ee](\d{2,4})', text)
        if emp_match:
            params["empNo"] = f"E{emp_match.group(1).zfill(3)}"
        
        # 姓名（复用姓氏提取逻辑）
        name = self._extractChineseName(text)
        if name:
            params["empName"] = name
        
        # 手机号
        phone_match = re.search(r'(1[3-9]\d{9})', text)
        if phone_match:
            params["phone"] = phone_match.group(1)
        
        # 部门
        for dept_pattern in self.PATTERNS["deptName"]:
            match = re.search(dept_pattern, text)
            if match:
                params["deptName"] = match.group(1)
                break
        
        # 职位
        position_match = re.search(r'(?:岗位|职位|职务|担任)\s*(?:是|：|:)?\s*([\u4e00-\u9fff\w]{2,8})', text)
        if position_match:
            params["position"] = position_match.group(1)
        
        # 基本工资
        salary_match = re.search(r'(?:基本工资|薪资|月薪|底薪|工资)\s*(?:是|：|:|=)?\s*(\d+(?:,\d{3})*(?:\.\d+)?)', text)
        if salary_match:
            raw = salary_match.group(1).replace(",", "")
            params["baseSalary"] = float(raw)
        
        # 入职日期
        date_match = re.search(r'(?:入职|到岗)(?:日期|时间)?\s*(?:是|：|:)?\s*(\d{4}[年/-]\d{1,2}[月/-]\d{1,2}[日]?)', text)
        if date_match:
            raw_date = date_match.group(1)
            # 标准化为 YYYY-MM-DD
            raw_date = raw_date.replace("年", "-").replace("月", "-").replace("日", "")
            params["hireDate"] = raw_date
        
        return params
    
    def _extractUpdateParams(self, text: str, params: Dict) -> Dict:
        """提取更新操作参数"""
        params = self._extractEmployeeIdentifiers(text, params, None)
        
        # 检测具体的更新类型
        if "转正" in text:
            params["updateType"] = "转正"
            params["newStatus"] = "在职"
        elif "调岗" in text:
            params["updateType"] = "调岗"
        elif "调薪" in text or "涨薪" in text or "加薪" in text:
            params["updateType"] = "调薪"
            # 尝试提取新薪资
            salary_match = re.search(r'(\d+(?:,\d{3})*(?:\.\d+)?)\s*(?:元|块)', text)
            if salary_match:
                raw = salary_match.group(1).replace(",", "")
                params["newSalary"] = float(raw)
        elif "离职" in text:
            params["updateType"] = "离职"
            params["newStatus"] = "离职"
        
        return params
    
    def _extractPayrollParams(self, text: str, params: Dict, context: Optional[ConversationContext]) -> Dict:
        """提取薪资计算参数"""
        params = self._extractEmployeeIdentifiers(text, params, context)
        
        # 年月
        now = datetime.now()
        year_match = re.search(r'(202[0-9])\s*年?', text)
        month_match = re.search(r'(1[0-2]|[1-9])\s*月?', text)
        
        if year_match:
            params["year"] = int(year_match.group(1))
        else:
            params["year"] = now.year
        
        if month_match:
            params["month"] = int(month_match.group(1))
        else:
            params["month"] = now.month
        
        # 是否批量（"所有人"、"全部"、"批量"）
        if any(w in text for w in ["所有人", "全部", "批量", "大家", "全体"]):
            params["batch"] = True
        
        return params
    
    def _extractBonusParams(self, text: str, params: Dict) -> Dict:
        """提取年终奖计算参数"""
        # 年终奖金额
        amount_match = re.search(r'(\d+(?:,\d{3})*(?:\.\d+)?)\s*(?:元|块|万)?', text)
        if amount_match:
            raw = amount_match.group(1).replace(",", "")
            if "万" in text:
                params["bonusAmount"] = float(raw) * 10000
            else:
                params["bonusAmount"] = float(raw)
        
        return params
    
    def _extractInsuranceParams(self, text: str, params: Dict) -> Dict:
        """提取社保计算参数"""
        # 工资基数（先尝试"城市+社保+数字"的模式）
        salary_match = re.search(r'(?:社保|公积金|五险一金)\s*(\d+(?:,\d{3})*(?:\.\d+)?)', text)
        if not salary_match:
            salary_match = re.search(r'(?:工资|薪资|月薪|基数|缴纳)\s*(?:是|：|:|=)?\s*(\d+(?:,\d{3})*(?:\.\d+)?)', text)
        if salary_match:
            raw = salary_match.group(1).replace(",", "")
            params["salary"] = float(raw)
        
        # 城市
        for city_pattern in self.PATTERNS["city"]:
            match = re.search(city_pattern, text)
            if match:
                params["city"] = match.group(1)
                break
        
        return params
    
    def _extractBindAttendanceParams(self, text: str, params: Dict) -> Dict:
        """提取考勤表绑定参数"""
        # 文件路径
        for fp_pattern in self.PATTERNS["filePath"]:
            match = re.search(fp_pattern, text)
            if match:
                params["filePath"] = match.group(1)
                break
        
        params["tableType"] = "attendance"
        return params
    
    def _extractAttendanceParams(self, text: str, params: Dict) -> Dict:
        """提取考勤查询参数"""
        # 员工标识
        params = self._extractEmployeeIdentifiers(text, params, None)
        
        # 年份
        year_match = re.search(self.PATTERNS["year"][0], text)
        if year_match:
            params["year"] = int(year_match.group(1))
        
        # 月份
        month_match = re.search(self.PATTERNS["month"][0], text)
        if month_match:
            params["month"] = int(month_match.group(1))
        
        return params
    
    def _extractAttendanceRulesParams(self, text: str, params: Dict) -> Dict:
        """提取考勤规则参数"""
        # 检测是否包含数字（用于设置扣款金额等）
        numbers = re.findall(r'(\d+(?:\.\d+)?)\s*(?:元|块)?', text)
        if numbers:
            params["amounts"] = [float(n) for n in numbers]
        return params
    
    def _extractBindParams(self, text: str, params: Dict) -> Dict:
        """提取表格绑定参数"""
        # 判断绑定哪种表
        if "组织架构" in text or "部门" in text:
            params["tableType"] = "organization"
        elif "花名册" in text or "员工" in text:
            params["tableType"] = "employee"
        elif "薪资" in text or "工资" in text:
            params["tableType"] = "salary"
        
        # 文件路径
        for fp_pattern in self.PATTERNS["filePath"]:
            match = re.search(fp_pattern, text)
            if match:
                params["filePath"] = match.group(1)
                break
        
        return params
    
    def _extractBatchParams(self, text: str, params: Dict) -> Dict:
        """提取批量操作参数"""
        # 操作类型
        if "转正" in text:
            params["newStatus"] = "在职"
        elif "离职" in text:
            params["newStatus"] = "离职"
        
        # 提取工号列表
        emp_matches = re.findall(r'[Ee](\d{2,4})', text)
        if emp_matches:
            params["empNos"] = [f"E{m.zfill(3)}" for m in emp_matches]
        
        return params
    
    def _extractExportParams(self, text: str, params: Dict) -> Dict:
        """提取导出参数"""
        if "薪资" in text or "工资条" in text:
            params["reportType"] = "payroll"
        elif "员工" in text or "人力" in text or "花名册" in text:
            params["reportType"] = "employee"
        elif "统计" in text:
            params["reportType"] = "statistics"
        
        return params
    
    def _extractStatsParams(self, text: str, params: Dict) -> Dict:
        """提取统计参数"""
        # 特定统计类型
        if "合同" in text and "到期" in text:
            params["statType"] = "contractExpiring"
        elif "生日" in text:
            params["statType"] = "birthday"
        elif "试用期" in text and ("到期" in text or "转正" in text):
            params["statType"] = "probation"
        
        # 部门筛选
        for dept_pattern in self.PATTERNS["deptName"]:
            match = re.search(dept_pattern, text)
            if match:
                params["deptName"] = match.group(1)
                break
        
        return params
    
    def _extractCommonParams(self, text: str, params: Dict) -> Dict:
        """提取通用参数"""
        # 城市
        if "city" not in params:
            for city_pattern in self.PATTERNS["city"]:
                match = re.search(city_pattern, text)
                if match:
                    params["city"] = match.group(1)
                    break
        
        # 手机号
        if "phone" not in params:
            phone_match = re.search(r'(1[3-9]\d{9})', text)
            if phone_match:
                params["phone"] = phone_match.group(1)
        
        # 邮箱
        if "email" not in params:
            email_match = re.search(r'([\w.+-]+@[\w-]+\.[\w.]+)', text)
            if email_match:
                params["email"] = email_match.group(1)
        
        return params
    
    def checkMissingParams(self, intent: IntentType, params: Dict) -> List[str]:
        """检查必要参数是否缺失"""
        required = {
            IntentType.READ_EMPLOYEE: ["empNo_or_empName"],
            IntentType.UPDATE_EMPLOYEE: ["empNo_or_empName"],
            IntentType.DELETE_EMPLOYEE: ["empNo_or_empName"],
            IntentType.GET_REPORTING_CHAIN: ["empNo_or_empName"],
            IntentType.CALCULATE_INSURANCE: ["city"],
            IntentType.CALCULATE_BONUS: ["bonusAmount"],
            IntentType.BIND_TABLE: ["tableType", "filePath"],
            IntentType.BATCH_UPDATE_STATUS: ["empNos", "newStatus"],
        }
        
        needed = required.get(intent, [])
        missing = []
        for p in needed:
            if p == "empNo_or_empName":
                if not params.get("empNo") and not params.get("empName"):
                    missing.append("empNo")
            elif p not in params or not params[p]:
                missing.append(p)
        return missing


# ============================================================
# Tool 执行器
# ============================================================

class ToolExecutor:
    """
    Tool 执行器
    统一调用各模块的工具函数，管理配置和生命周期
    自动记录审计日志和持久化薪资计算结果
    """
    
    def __init__(self, config: Optional[Dict] = None):
        """
        Args:
            config: 全局配置 {
                "filePath": "花名册路径",
                "orgFilePath": "组织架构路径",
                "salaryFilePath": "薪资表路径",
                "columnMapping": {},
                "orgColumnMapping": {},
                "dataDir": "数据目录（可选，默认 .hr-data/）",
            }
        """
        self.config = config or {}
        self._employeeManager = None
        self._payrollEngine = None
        self._store = None  # HRStore 延迟初始化
        self._attendanceManager = None  # AttendanceManager 延迟初始化
    
    @property
    def store(self):
        """延迟加载 HRStore"""
        if self._store is None:
            from hr_store import HRStore
            dataDir = self.config.get("dataDir")
            self._store = HRStore(dataDir)
        return self._store
    
    @property
    def employeeManager(self):
        """延迟加载 EmployeeManager"""
        if self._employeeManager is None and self.config.get("filePath"):
            from employee_manager import EmployeeManager
            # 构建 config 字典，包含 columnMapping 和 sheetName
            empConfig = dict(self.config.get("columnMapping") or {})
            if self.config.get("sheetName"):
                empConfig["sheetName"] = self.config["sheetName"]
            self._employeeManager = EmployeeManager(
                self.config["filePath"],
                empConfig
            )
            if self.config.get("orgFilePath"):
                # 传递 sheetName（Bug3 修复）
                orgConfig = dict(self.config.get("orgColumnMapping") or {})
                if self.config.get("orgSheetName"):
                    orgConfig["sheetName"] = self.config["orgSheetName"]
                self._employeeManager.loadDepartments(
                    self.config["orgFilePath"],
                    orgConfig if orgConfig else None
                )
        return self._employeeManager
    
    @property
    def payrollEngine(self):
        """延迟加载 PayrollEngine"""
        if self._payrollEngine is None:
            from payroll_engine import PayrollEngine
            self._payrollEngine = PayrollEngine()
        return self._payrollEngine
    
    @property
    def attendanceManager(self):
        """延迟加载 AttendanceManager"""
        if self._attendanceManager is None and self.config.get("attendanceFilePath"):
            from attendance_manager import AttendanceManager
            attConfig = dict(self.config.get("attendanceColumnMapping") or {})
            if self.config.get("attendanceSheetName"):
                attConfig["sheetName"] = self.config["attendanceSheetName"]
            self._attendanceManager = AttendanceManager(
                self.config["attendanceFilePath"],
                attConfig
            )
            self._attendanceManager.load()
        return self._attendanceManager
    
    def execute(self, intent: IntentType, params: Dict) -> Dict[str, Any]:
        """
        执行对应的 Tool 调用
        
        Args:
            intent: 意图类型
            params: 提取的参数
        
        Returns:
            执行结果 {success, message, data, ...}
        """
        executor_map = {
            # 初始化
            IntentType.START_ONBOARDING: self._exec_start_onboarding,
            IntentType.SET_STORAGE_TYPE: self._exec_set_storage_type,
            IntentType.BIND_TABLE: self._exec_bind_table,
            IntentType.CHECK_CONFIG: self._exec_check_config,
            IntentType.RESET_CONFIG: self._exec_reset_config,
            
            # 员工查询
            IntentType.READ_EMPLOYEE: self._exec_read_employee,
            IntentType.LIST_EMPLOYEES: self._exec_list_employees,
            IntentType.SEARCH_EMPLOYEE: self._exec_search_employee,
            
            # 员工操作
            IntentType.ADD_EMPLOYEE: self._exec_add_employee,
            IntentType.UPDATE_EMPLOYEE: self._exec_update_employee,
            IntentType.DELETE_EMPLOYEE: self._exec_delete_employee,
            IntentType.BATCH_UPDATE_STATUS: self._exec_batch_update_status,
            
            # 组织架构
            IntentType.QUERY_DEPARTMENT: self._exec_query_department,
            IntentType.GET_DEPT_TREE: self._exec_get_dept_tree,
            IntentType.GET_REPORTING_CHAIN: self._exec_get_reporting_chain,
            
            # 薪资
            IntentType.CALCULATE_PAYROLL: self._exec_calculate_payroll,
            IntentType.CALCULATE_BONUS: self._exec_calculate_bonus,
            IntentType.CALCULATE_INSURANCE: self._exec_calculate_insurance,
            IntentType.GET_TAX_BRACKET: self._exec_get_tax_bracket,
            
            # 考勤
            IntentType.BIND_ATTENDANCE: self._exec_bind_attendance,
            IntentType.VIEW_ATTENDANCE: self._exec_view_attendance,
            IntentType.ATTENDANCE_SUMMARY: self._exec_attendance_summary,
            IntentType.ATTENDANCE_RULES: self._exec_attendance_rules,
            
            # 报表
            IntentType.EMPLOYEE_STATISTICS: self._exec_employee_statistics,
            IntentType.DATA_VALIDATION: self._exec_data_validation,
            IntentType.EXPORT_REPORT: self._exec_export_report,
            
            # 帮助
            IntentType.HELP: self._exec_help,
        }
        
        executor = executor_map.get(intent)
        if executor:
            try:
                return executor(params)
            except Exception as e:
                return {
                    "success": False,
                    "message": f"执行出错: {str(e)}",
                    "intent": intent.value
                }
        
        return {
            "success": False,
            "message": "抱歉，我暂时无法处理这个请求。",
            "intent": intent.value
        }
    
    # --- 初始化引导 ---
    
    def _exec_start_onboarding(self, params: Dict) -> Dict:
        from onboarding import OnboardingManager
        mgr = OnboardingManager()
        status = mgr.getOnboardingStatus()
        
        if status["isFullyInitialized"]:
            return {
                "success": True,
                "message": "已完成初始化，可以直接使用。",
                "data": status
            }
        
        next_step = status["nextStep"]
        step_messages = {
            "set_storage_type": (
                "请选择数据存储方式：\n"
                "  1️⃣ Excel 本地文件（推荐）\n"
                "\n回复「用excel」选择。目前仅支持 Excel 本地存储，飞书多维表格规划中。"
            ),
            "bind_organization": (
                "📋 下一步：绑定组织架构表\n"
                "\n请直接上传你的组织架构 Excel 文件（.xls 或 .xlsx），我会自动分析表格结构并识别列映射。\n"
                "\n支持的列：部门编码、部门名称、上级部门、部门负责人等。"
            ),
            "bind_employee": (
                "👤 下一步：绑定员工花名册\n"
                "\n请直接上传你的员工花名册 Excel 文件（.xls 或 .xlsx），我会自动识别列映射。\n"
                "\n支持的列：工号、姓名、部门、职位、入职日期、在职状态、手机号等。"
            ),
            "bind_salary": (
                "💰 下一步：绑定薪资表\n"
                "\n请直接上传你的薪资 Excel 文件（.xls 或 .xlsx），我会自动识别列映射。\n"
                "\n支持的列：工号、基本工资、绩效奖金、津贴等。"
            ),
            "bind_attendance": (
                "📊 下一步：绑定考勤表（可选）\n"
                "\n考勤表绑定后，计算薪资时会自动扣除考勤相关款项。\n"
                "如果你暂时不需要考勤扣减功能，可以直接跳过此步，后续随时可以绑定。\n"
                "\n请上传考勤 Excel 文件（.xls 或 .xlsx），通常包含：工号、年份、月份、出勤天数、迟到次数、事假天数等。"
            ),
        }
        
        return {
            "success": True,
            "message": step_messages.get(next_step, "初始化状态检查完成"),
            "nextStep": next_step,
            "status": status,
            "needsInput": True
        }
    
    def _exec_set_storage_type(self, params: Dict) -> Dict:
        from onboarding import OnboardingManager
        mgr = OnboardingManager()
        storage = params.get("storageType", "excel")
        result = mgr.setStorageType(storage)
        # 记录审计日志
        if result.get("success"):
            self.store.appendAuditLog(
                action="set_storage_type",
                targetType="config",
                details={"storageType": storage}
            )
        return result
    
    def _exec_bind_table(self, params: Dict) -> Dict:
        from onboarding import OnboardingManager
        mgr = OnboardingManager()
        table_type = params.get("tableType", "")
        file_path = params.get("filePath", "")
        
        if not file_path:
            # 没有文件路径时，提示用户上传文件
            typeNames = {
                "organization": "组织架构表",
                "employee": "员工花名册",
                "salary": "薪资表",
                "attendance": "考勤表",
                "": "表格"
            }
            name = typeNames.get(table_type, "表格")
            return {
                "success": False,
                "needsInput": True,
                "message": f"请上传{name}的 Excel 文件（.xls 或 .xlsx），我会自动分析结构并完成绑定。"
            }
        
        # 转为绝对路径
        if not os.path.isabs(file_path):
            file_path = os.path.abspath(file_path)
        
        # 使用一站式 analyzeAndBind
        result = mgr.analyzeAndBind(table_type, file_path)
        
        # 记录审计日志
        if result.get("success"):
            self.store.appendAuditLog(
                action="bind_table",
                targetType="config",
                details={"tableType": table_type, "filePath": file_path}
            )
        return result
    
    def _exec_check_config(self, params: Dict) -> Dict:
        from onboarding import get_configuration_summary
        return get_configuration_summary()
    
    def _exec_reset_config(self, params: Dict) -> Dict:
        from onboarding import OnboardingManager
        mgr = OnboardingManager()
        result = mgr.resetConfiguration()
        # 记录审计日志
        if result.get("success"):
            self.store.appendAuditLog(
                action="reset_config",
                targetType="config",
                details={}
            )
        return result
    
    # --- 员工查询 ---
    
    def _exec_read_employee(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        emp_no = params.get("empNo")
        emp_name = params.get("empName")
        
        if not emp_no and emp_name and mgr.employees:
            # 按姓名查找
            for emp in mgr.employees.values():
                if emp.name == emp_name:
                    emp_no = emp.empNo
                    break
        
        if not emp_no:
            return {"success": False, "message": "请提供工号或姓名"}
        
        emp = mgr.getEmployee(emp_no)
        if not emp:
            return {"success": False, "message": f"员工不存在: {emp_no}"}
        
        return {
            "success": True,
            "message": f"已找到员工: {emp.name}（{emp.empNo}）",
            "data": emp.toDict()
        }
    
    def _exec_list_employees(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        employees = mgr.listEmployees(
            deptCode=params.get("deptCode"),
            status=params.get("status"),
            keyword=params.get("keyword"),
            includeInactive=params.get("includeInactive", False)
        )
        
        summary = f"共 {len(employees)} 名员工"
        if params.get("status"):
            summary += f"（{params['status']}）"
        if params.get("deptName"):
            summary += f"，{params['deptName']}"
        
        return {
            "success": True,
            "message": summary,
            "data": [e.toDict() for e in employees],
            "count": len(employees)
        }
    
    def _exec_search_employee(self, params: Dict) -> Dict:
        return self._exec_list_employees(params)
    
    # --- 员工操作 ---
    
    def _exec_add_employee(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        from employee_manager import Employee
        empNo = params.get("empNo", "")
        emp = Employee(
            empNo=empNo,
            name=params.get("empName", ""),
            deptCode=params.get("deptCode", ""),
            deptName=params.get("deptName", ""),
            position=params.get("position", ""),
            phone=params.get("phone", ""),
            email=params.get("email", ""),
            hireDate=params.get("hireDate", datetime.now().strftime("%Y-%m-%d")),
            status=params.get("status", "试用期"),
            baseSalary=float(params.get("baseSalary", 0)),
            socialInsuranceCity=params.get("socialInsuranceCity", ""),
            socialInsuranceBase=float(params.get("socialInsuranceBase", 0)),
            reportTo=params.get("reportTo", ""),
        )
        
        ok, msg = mgr.addEmployee(emp)
        if ok:
            mgr.save()
            # 记录审计日志
            self.store.appendAuditLog(
                action="add_employee",
                targetType="employee",
                targetId=empNo,
                details={"name": params.get("empName", ""), "dept": params.get("deptName", "")}
            )
        
        return {"success": ok, "message": msg}
    
    def _exec_update_employee(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        emp_no = params.get("empNo")
        emp_name = params.get("empName")
        
        if not emp_no and emp_name:
            for emp in mgr.employees.values():
                if emp.name == emp_name:
                    emp_no = emp.empNo
                    break
        
        if not emp_no:
            return {"success": False, "message": "请提供工号或姓名"}
        
        updates = {}
        if "newStatus" in params:
            updates["status"] = params["newStatus"]
        if "newSalary" in params:
            updates["baseSalary"] = params["newSalary"]
        if "deptName" in params:
            updates["deptName"] = params["deptName"]
        
        ok, msg = mgr.updateEmployee(emp_no, updates)
        if ok:
            mgr.save()
            # 记录审计日志
            self.store.appendAuditLog(
                action="update_employee",
                targetType="employee",
                targetId=emp_no,
                details={"updates": updates}
            )
        
        return {"success": ok, "message": msg}
    
    def _exec_delete_employee(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        emp_no = params.get("empNo")
        soft = params.get("soft", True)
        ok, msg = mgr.deleteEmployee(emp_no, soft)
        
        if ok:
            mgr.save()
            # 记录审计日志
            self.store.appendAuditLog(
                action="delete_employee",
                targetType="employee",
                targetId=emp_no,
                details={"soft": soft}
            )
        
        return {"success": ok, "message": msg}
    
    def _exec_batch_update_status(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        emp_nos = params.get("empNos", [])
        result = mgr.batchUpdateStatus(
            empNos=emp_nos,
            newStatus=params.get("newStatus", ""),
            leaveDate=params.get("leaveDate")
        )
        
        if result["successCount"] > 0:
            mgr.save()
            # 记录审计日志
            self.store.appendAuditLog(
                action="batch_update_status",
                targetType="employee",
                targetId=",".join(emp_nos),
                details={
                    "newStatus": params.get("newStatus", ""),
                    "successCount": result["successCount"],
                    "failedCount": result["failedCount"],
                }
            )
        
        status_name = params.get("newStatus", "")
        return {
            "success": True,
            "message": f"批量{status_name}完成：成功 {result['successCount']} 人，失败 {result['failedCount']} 人",
            **result
        }
    
    # --- 组织架构 ---
    
    def _exec_query_department(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr or not mgr.departments:
            return {"success": False, "message": "未加载组织架构，请先完成初始化"}
        
        dept_code = params.get("deptCode")
        dept_name = params.get("deptName")
        
        if dept_name and not dept_code:
            for code, dept in mgr.departments.items():
                if dept.deptName == dept_name:
                    dept_code = code
                    break
        
        if not dept_code:
            return {"success": False, "message": "请提供部门编码或名称"}
        
        dept = mgr.departments.get(dept_code)
        if not dept:
            return {"success": False, "message": f"部门不存在: {dept_code}"}
        
        from dataclasses import asdict
        return {
            "success": True,
            "message": f"部门: {dept.deptName}（{dept.deptCode}），{dept.employeeCount} 人",
            "data": asdict(dept)
        }
    
    def _exec_get_dept_tree(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr or not mgr.departments:
            return {"success": False, "message": "未加载组织架构，请先完成初始化"}
        
        tree = mgr.getDeptTree()
        total = mgr.getEmployeeCount()
        
        return {
            "success": True,
            "message": f"组织架构共 {len(mgr.departments)} 个部门，{total} 名员工",
            "data": tree,
            "totalDepts": len(mgr.departments),
            "totalEmployees": total
        }
    
    def _exec_get_reporting_chain(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        emp_no = params.get("empNo")
        emp_name = params.get("empName")
        
        if not emp_no and emp_name:
            for emp in mgr.employees.values():
                if emp.name == emp_name:
                    emp_no = emp.empNo
                    break
        
        if not emp_no:
            return {"success": False, "message": "请提供工号或姓名"}
        
        chain = mgr.getReportingChain(emp_no)
        if not chain:
            return {"success": False, "message": f"未找到员工: {emp_no}"}
        
        chain_str = " → ".join([f"{c['name']}({c['position'] or '无职位'})" for c in chain])
        return {
            "success": True,
            "message": f"汇报链: {chain_str}",
            "data": chain,
            "chainLength": len(chain)
        }
    
    # --- 薪资 ---
    
    def _loadCumulativeData(self, year: int, month: int) -> Dict[str, Dict]:
        """
        从 HRStore 加载上月薪资累计数据（用于累计预扣法）
        
        Args:
            year: 当前计算年份
            month: 当前计算月份
        
        Returns:
            {empNo: {cumulativeIncome, cumulativeInsurance, cumulativeTax}}
        """
        result = {}
        
        # 计算上月
        prev_month = month - 1
        prev_year = year
        if prev_month < 1:
            prev_month = 12
            prev_year -= 1
        
        # 1 月份没有上月数据，返回空
        if month == 1:
            return result
        
        payroll = self.store.loadPayrollResult(prev_year, prev_month)
        if not payroll:
            return result
        
        for record in payroll.get("results", []):
            empNo = record.get("empNo")
            if empNo:
                result[empNo] = {
                    "cumulativeIncome": Decimal(str(record.get("cumulativeIncome", 0))),
                    "cumulativeInsurance": Decimal(str(record.get("cumulativeInsurance", 0))),
                    "cumulativeTax": Decimal(str(record.get("cumulativeTax", 0))),
                }
        
        return result
    
    def _exec_calculate_payroll(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        engine = self.payrollEngine
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        year = params.get("year", datetime.now().year)
        month = params.get("month", datetime.now().month)
        emp_no = params.get("empNo")
        emp_name = params.get("empName")
        
        # 从 HRStore 加载上月累计数据（用于累计预扣法）
        cumulativeDataMap = self._loadCumulativeData(year, month)
        
        if params.get("batch"):
            # 批量计算
            results = []
            skipped_cities = []  # Bug6: 记录跳过的不支持城市
            
            # 考勤扣减集成：如果考勤表已绑定，批量计算考勤扣减
            attendance_deductions = {}
            att_mgr = self.attendanceManager
            if att_mgr and att_mgr._loaded:
                daily_salaries = {}
                for emp in mgr.employees.values():
                    if emp.status == "离职":
                        continue
                    if emp_no and emp.empNo != emp_no:
                        continue
                    if emp_name and emp.name != emp_name:
                        continue
                    # 日薪 = 基本工资 / 21.75
                    if emp.baseSalary and emp.baseSalary > 0:
                        daily_salaries[emp.empNo] = Decimal(str(emp.baseSalary)) / Decimal('21.75')
                attendance_deductions = att_mgr.getAllPreTaxDeductions(year, month, daily_salaries)
            
            for emp in mgr.employees.values():
                if emp.status == "离职":
                    continue
                if emp_no and emp.empNo != emp_no:
                    continue
                if emp_name and emp.name != emp_name:
                    continue
                
                city = emp.socialInsuranceCity or "北京"
                
                # Bug6 修复：跳过不支持的城市，不中断批量计算
                from payroll_engine import InsuranceCalculator
                if not InsuranceCalculator.getCityConfig(city):
                    skipped_cities.append((emp.empNo, emp.name, city))
                    continue
                
                cumData = cumulativeDataMap.get(emp.empNo, {})
                
                # 社保基数：优先取花名册中的 socialInsuranceBase，否则用 baseSalary
                insuranceBase = None
                if emp.socialInsuranceBase and emp.socialInsuranceBase > 0:
                    insuranceBase = Decimal(str(emp.socialInsuranceBase))
                
                # 公积金基数：优先取花名册中的 housingFundBase，否则与社保基数相同
                hfBase = None
                if emp.housingFundBase and emp.housingFundBase > 0:
                    hfBase = Decimal(str(emp.housingFundBase))
                
                # 公积金比例
                hfRate = None
                if emp.housingFundRate and emp.housingFundRate > 0:
                    hfRate = Decimal(str(emp.housingFundRate))
                
                # 专项附加扣除
                specialDed = Decimal(str(emp.specialDeduction)) if emp.specialDeduction and emp.specialDeduction > 0 else None
                
                result = engine.calculatePayroll(
                    employeeId=emp.empNo,
                    employeeName=emp.name,
                    year=year,
                    month=month,
                    baseSalary=Decimal(str(emp.baseSalary or 0)),
                    city=city,
                    cumulativeData=cumData,
                    insuranceBase=insuranceBase,
                    housingFundBase=hfBase,
                    housingFundRate=hfRate,
                    specialDeduction=specialDed,
                    preTaxDeductions=attendance_deductions.get(emp.empNo, Decimal('0')),
                )
                results.append(result)
            
            total_net = sum(float(r.netPay) for r in results)
            
            # 考勤扣减汇总
            attendance_total = sum(float(v) for v in attendance_deductions.values()) if attendance_deductions else 0
            
            # 持久化薪资计算结果
            payroll_data = [{
                "empNo": r.employeeId,
                "name": r.employeeName,
                "baseSalary": float(r.baseSalary),
                "preTaxDeductions": float(r.preTaxDeductions),
                "grossPay": float(r.grossPay),
                "socialInsurance": float(r.socialInsurance),
                "housingFund": float(r.housingFund),
                "totalDeductions": float(r.totalInsurance) + float(r.taxAmount),
                "totalInsurance": float(r.totalInsurance),
                "taxAmount": float(r.taxAmount),
                "netPay": float(r.netPay),
                "socialInsuranceBaseUsed": float(r.socialInsuranceBaseUsed),
                "housingFundBaseUsed": float(r.housingFundBaseUsed),
                "taxBracketLevel": r.taxBracketLevel,
                "taxBracketChanged": r.taxBracketChanged,
                # 累计数据，供下月计算使用
                "cumulativeIncome": float(r.cumulativeIncome),
                "cumulativeInsurance": float(r.cumulativeInsurance),
                "cumulativeTax": float(r.cumulativeTax),
            } for r in results]
            self.store.savePayrollResult(year, month, payroll_data)
            
            # 统计跳档人数
            bracket_changed_count = sum(1 for r in results if r.taxBracketChanged)
            
            # 记录审计日志
            self.store.appendAuditLog(
                action="calculate_payroll",
                targetType="payroll",
                targetId=f"{year}-{month:02d}",
                details={
                    "employeeCount": len(results),
                    "totalNetPay": total_net,
                }
            )
            
            msg = (
                f"已计算 {len(results)} 人 {year}年{month}月薪资\n"
                f"  应发合计: ¥{sum(float(r.grossPay) for r in results):,.2f}\n"
            )
            if attendance_total > 0:
                msg += f"  考勤扣减合计: ¥{attendance_total:,.2f}\n"
            msg += (
                f"  社保合计: ¥{sum(float(r.socialInsurance) for r in results):,.2f}\n"
                f"  公积金合计: ¥{sum(float(r.housingFund) for r in results):,.2f}\n"
                f"  个税合计: ¥{sum(float(r.taxAmount) for r in results):,.2f}\n"
                f"  实发合计: ¥{total_net:,.2f}"
            )
            if bracket_changed_count > 0:
                msg += f"\n  ⚠️ {bracket_changed_count} 人本月税率跳档（累计预扣法正常现象，全年个税总额不变）"
            
            # Bug6: 不支持城市的跳过警告
            if skipped_cities:
                skipped_msg = ", ".join(f"{name}({city})" for _, name, city in skipped_cities)
                msg += f"\n  ⚠️ {len(skipped_cities)} 人因社保城市不支持被跳过: {skipped_msg}"
            
            return {
                "success": True,
                "message": msg,
                "data": payroll_data,
                "count": len(results),
                "totalNetPay": total_net
            }
        
        # 单人计算
        if not emp_no and emp_name:
            for emp in mgr.employees.values():
                if emp.name == emp_name:
                    emp_no = emp.empNo
                    break
        
        if not emp_no:
            return {"success": False, "message": "请提供员工工号或姓名"}
        
        emp = mgr.getEmployee(emp_no)
        if not emp:
            return {"success": False, "message": f"员工不存在: {emp_no}"}
        
        city = emp.socialInsuranceCity or params.get("city", "北京")
        
        # Bug6 修复：检测不支持的城市
        from payroll_engine import InsuranceCalculator
        if not InsuranceCalculator.getCityConfig(city):
            return {
                "success": False,
                "message": (
                    f"暂不支持城市「{city}」的社保配置。"
                    f"\n当前支持的城市：{', '.join(InsuranceCalculator.CITY_CONFIGS.keys())}"
                ),
            }
        
        cumData = cumulativeDataMap.get(emp.empNo, {})
        
        # 社保基数
        insuranceBase = None
        if emp.socialInsuranceBase and emp.socialInsuranceBase > 0:
            insuranceBase = Decimal(str(emp.socialInsuranceBase))
        
        # 公积金基数
        hfBase = None
        if emp.housingFundBase and emp.housingFundBase > 0:
            hfBase = Decimal(str(emp.housingFundBase))
        
        # 公积金比例
        hfRate = None
        if emp.housingFundRate and emp.housingFundRate > 0:
            hfRate = Decimal(str(emp.housingFundRate))
        
        # 专项附加扣除
        specialDed = Decimal(str(emp.specialDeduction)) if emp.specialDeduction and emp.specialDeduction > 0 else None
        
        # 考勤扣减（如果考勤表已绑定）
        preTaxDed = Decimal('0')
        att_deduction_note = ""
        att_mgr = self.attendanceManager
        if att_mgr and att_mgr._loaded:
            if emp.baseSalary and emp.baseSalary > 0:
                daily_salary = Decimal(str(emp.baseSalary)) / Decimal('21.75')
                record = att_mgr.calculateDeductions(emp.empNo, year, month, daily_salary)
                preTaxDed = record.preTaxDeductions
                if preTaxDed > 0 and record.deductionDetails:
                    att_deduction_note = f"  考勤扣减: ¥{float(preTaxDed):,.2f}\n"
        
        result = engine.calculatePayroll(
            employeeId=emp.empNo,
            employeeName=emp.name,
            year=year,
            month=month,
            baseSalary=Decimal(str(emp.baseSalary or 0)),
            city=city,
            cumulativeData=cumData,
            insuranceBase=insuranceBase,
            housingFundBase=hfBase,
            housingFundRate=hfRate,
            specialDeduction=specialDed,
            preTaxDeductions=preTaxDed,
        )
        
        # 构建消息
        special_line = f"  专项附加扣除: ¥{float(emp.specialDeduction):,.2f}\n" if emp.specialDeduction and emp.specialDeduction > 0 else ""
        bracket_line = ""
        if result.taxBracketChanged:
            bracket_line = f"  ⚠️ 本月税率跳档至第{result.taxBracketLevel}级（{float(result.taxBracketRate)*100:.0f}%），累计预扣法正常现象\n"
        
        return {
            "success": True,
            "message": (
                f"{emp.name} {year}年{month}月薪资：\n"
                f"  基本工资: ¥{float(result.baseSalary):,.2f}\n"
                + att_deduction_note +
                f"  税前扣减合计: ¥{float(result.preTaxDeductions):,.2f}\n"
                + special_line +
                f"  应发: ¥{float(result.grossPay):,.2f}\n"
                f"  社保: ¥{float(result.socialInsurance):,.2f}（基数 ¥{float(result.socialInsuranceBaseUsed):,.0f}）\n"
                f"  公积金: ¥{float(result.housingFund):,.2f}（基数 ¥{float(result.housingFundBaseUsed):,.0f}）\n"
                f"  个税: ¥{float(result.taxAmount):,.2f}（第{result.taxBracketLevel}级 {float(result.taxBracketRate)*100:.0f}%）\n"
                + bracket_line +
                f"  实发: ¥{float(result.netPay):,.2f}"
            ),
            "data": {
                "employeeId": result.employeeId,
                "employeeName": result.employeeName,
                "baseSalary": float(result.baseSalary),
                "preTaxDeductions": float(result.preTaxDeductions),
                "grossPay": float(result.grossPay),
                "socialInsurance": float(result.socialInsurance),
                "housingFund": float(result.housingFund),
                "totalInsurance": float(result.totalInsurance),
                "taxAmount": float(result.taxAmount),
                "netPay": float(result.netPay),
                "details": {
                    "pensionInsurance": float(result.pensionInsurance),
                    "medicalInsurance": float(result.medicalInsurance),
                    "unemploymentInsurance": float(result.unemploymentInsurance),
                    "socialInsuranceBaseUsed": float(result.socialInsuranceBaseUsed),
                    "housingFundBaseUsed": float(result.housingFundBaseUsed),
                    "taxBracketLevel": result.taxBracketLevel,
                    "taxBracketRate": float(result.taxBracketRate),
                    "taxBracketChanged": result.taxBracketChanged,
                }
            }
        }
    
    def _exec_calculate_bonus(self, params: Dict) -> Dict:
        engine = self.payrollEngine
        amount = params.get("bonusAmount", 0)
        monthly_salary = params.get("monthlySalary", 0)
        
        if not amount:
            return {"success": False, "message": "请提供年终奖金额"}
        
        result = engine.calculateYearEndBonus(
            Decimal(str(amount)),
            Decimal(str(monthly_salary))
        )
        
        return {
            "success": True,
            "message": (
                f"年终奖 ¥{float(result.bonusAmount):,.2f}\n"
                f"  计税方式: {result.optimalMethod}\n"
                f"  个税: ¥{float(result.taxAmount):,.2f}\n"
                f"  税后: ¥{float(result.netBonus):,.2f}\n"
                f"  实际税率: {float(result.effectiveRate)}%"
            ),
            "data": {
                "bonusAmount": float(result.bonusAmount),
                "taxAmount": float(result.taxAmount),
                "netBonus": float(result.netBonus),
                "effectiveRate": float(result.effectiveRate),
                "optimalMethod": result.optimalMethod
            }
        }
    
    def _exec_calculate_insurance(self, params: Dict) -> Dict:
        from payroll_engine import InsuranceCalculator
        calculator = InsuranceCalculator()
        
        salary = params.get("salary", 0)
        city = params.get("city", "北京")
        
        if not salary:
            return {"success": False, "message": "请提供工资基数和缴纳城市"}
        
        result = calculator.calculateInsurance(Decimal(str(salary)), city)
        
        # Bug6 修复：检测不支持的城市
        if result.get("_unsupportedCity"):
            return {
                "success": False,
                "message": (
                    f"暂不支持城市「{city}」的社保配置。"
                    f"\n当前支持的城市：{', '.join(InsuranceCalculator.CITY_CONFIGS.keys())}"
                    f"\n请联系管理员添加该城市的社保配置。"
                ),
                "data": {},
            }
        
        total = float(result["totalInsurance"])
        return {
            "success": True,
            "message": (
                f"{city}社保公积金（基数 ¥{salary:,.0f}）：\n"
                f"  养老保险: ¥{float(result['pensionInsurance']):.2f}\n"
                f"  医疗保险: ¥{float(result['medicalInsurance']):.2f}\n"
                f"  失业保险: ¥{float(result['unemploymentInsurance']):.2f}\n"
                f"  公积金: ¥{float(result['housingFund']):.2f}\n"
                f"  个人合计: ¥{total:.2f}"
            ),
            "data": {k: float(v) for k, v in result.items()}
        }
    
    def _exec_get_tax_bracket(self, params: Dict) -> Dict:
        from payroll_engine import TaxCalculator
        taxable = params.get("taxableIncome", 0)
        bracket = TaxCalculator.getTaxBracket(Decimal(str(taxable)))
        
        if not bracket:
            return {"success": False, "message": "无法确定税率"}
        
        return {
            "success": True,
            "message": (
                f"适用税率: 第{bracket.level}级\n"
                f"  税率: {float(bracket.rate)*100:.0f}%\n"
                f"  速算扣除数: ¥{float(bracket.quickDeduction):,.0f}\n"
                f"  范围: ¥{float(bracket.minIncome):,.0f} ~ {f'¥{float(bracket.maxIncome):,.0f}' if bracket.maxIncome else '无上限'}"
            ),
            "data": {
                "level": bracket.level,
                "rate": float(bracket.rate),
                "quickDeduction": float(bracket.quickDeduction),
                "minIncome": float(bracket.minIncome),
                "maxIncome": float(bracket.maxIncome) if bracket.maxIncome else None,
            }
        }
    
    # --- 考勤 ---
    
    def _exec_bind_attendance(self, params: Dict) -> Dict:
        from onboarding import OnboardingManager
        mgr = OnboardingManager()
        file_path = params.get("filePath", "")
        
        if not file_path:
            return {
                "success": False,
                "needsInput": True,
                "message": "请上传考勤表的 Excel 文件（.xls 或 .xlsx），我会自动分析结构并完成绑定。\n\n考勤表通常包含：工号、年份、月份、应出勤天数、实际出勤、迟到次数、早退次数、事假天数、病假天数、旷工天数、加班小时等字段。"
            }
        
        if not os.path.isabs(file_path):
            file_path = os.path.abspath(file_path)
        
        result = mgr.analyzeAndBind("attendance", file_path)
        
        if result.get("success"):
            self.store.appendAuditLog(
                action="bind_table",
                targetType="config",
                details={"tableType": "attendance", "filePath": file_path}
            )
        
        return result
    
    def _exec_view_attendance(self, params: Dict) -> Dict:
        att_mgr = self.attendanceManager
        if not att_mgr or not att_mgr._loaded:
            return {"success": False, "message": "未绑定考勤表，请先上传考勤文件。"}
        
        year = params.get("year", datetime.now().year)
        month = params.get("month", datetime.now().month)
        emp_no = params.get("empNo")
        emp_name = params.get("empName")
        
        # 如果只有姓名，查找工号
        if not emp_no and emp_name and self.employeeManager:
            for emp in self.employeeManager.employees.values():
                if emp.name == emp_name:
                    emp_no = emp.empNo
                    break
        
        if emp_no:
            # 查看单人考勤
            record = att_mgr.getRecord(emp_no, year, month)
            if not record:
                return {"success": False, "message": f"未找到 {emp_no} {year}年{month}月的考勤记录"}
            
            parts = [
                f"📋 {record.empNo} {year}年{month}月考勤",
                f"  应出勤: {record.shouldAttendDays} 天",
                f"  实出勤: {record.actualAttendDays} 天",
                f"  迟到: {record.lateCount} 次",
                f"  早退: {record.earlyLeaveCount} 次",
                f"  事假: {record.personalLeaveDays} 天",
                f"  病假: {record.sickLeaveDays} 天",
                f"  旷工: {record.absentDays} 天",
                f"  加班: {record.overtimeHours} 小时",
            ]
            
            if record.preTaxDeductions and record.preTaxDeductions > 0:
                parts.append(f"  考勤扣减: ¥{float(record.preTaxDeductions):,.2f}")
                if record.deductionDetails:
                    for k, v in record.deductionDetails.items():
                        if k.endswith("明细"):
                            parts.append(f"    {k}: {v}")
                        elif isinstance(v, (int, float)):
                            parts.append(f"    {k}: ¥{v:,.2f}")
            
            return {
                "success": True,
                "message": "\n".join(parts),
                "data": {
                    "empNo": record.empNo,
                    "year": year,
                    "month": month,
                    "shouldAttendDays": float(record.shouldAttendDays),
                    "actualAttendDays": float(record.actualAttendDays),
                    "lateCount": float(record.lateCount),
                    "earlyLeaveCount": float(record.earlyLeaveCount),
                    "personalLeaveDays": float(record.personalLeaveDays),
                    "sickLeaveDays": float(record.sickLeaveDays),
                    "absentDays": float(record.absentDays),
                    "overtimeHours": float(record.overtimeHours),
                    "preTaxDeductions": float(record.preTaxDeductions),
                    "deductionDetails": record.deductionDetails,
                }
            }
        
        # 查看整月考勤
        records = att_mgr.getMonthlyRecords(year, month)
        if not records:
            return {"success": False, "message": f"{year}年{month}月暂无考勤记录"}
        
        parts = [
            f"📋 {year}年{month}月考勤记录（{len(records)} 人）",
            f"{'工号':<8} {'出勤':>4} {'迟到':>4} {'早退':>4} {'事假':>4} {'病假':>4} {'旷工':>4} {'加班h':>5} {'扣减':>8}",
            "-" * 55,
        ]
        for r in sorted(records, key=lambda x: x.empNo):
            parts.append(
                f"{r.empNo:<8} {r.actualAttendDays:>4.0f} {r.lateCount:>4.0f} "
                f"{r.earlyLeaveCount:>4.0f} {r.personalLeaveDays:>4.0f} "
                f"{r.sickLeaveDays:>4.0f} {r.absentDays:>4.0f} "
                f"{r.overtimeHours:>5.0f} {float(r.preTaxDeductions):>8.2f}"
            )
        
        return {
            "success": True,
            "message": "\n".join(parts),
            "count": len(records)
        }
    
    def _exec_attendance_summary(self, params: Dict) -> Dict:
        att_mgr = self.attendanceManager
        if not att_mgr or not att_mgr._loaded:
            return {"success": False, "message": "未绑定考勤表，请先上传考勤文件。"}
        
        year = params.get("year", datetime.now().year)
        month = params.get("month", datetime.now().month)
        
        summary = att_mgr.getDeductionSummary(year, month)
        
        parts = [
            f"📊 {year}年{month}月考勤汇总",
            f"  记录人数: {summary['recordCount']} 人",
            f"  考勤扣减总额: ¥{summary['totalDeductions']:,.2f}",
            "",
            "  详细统计:",
        ]
        for k, v in summary["summary"].items():
            parts.append(f"    {k}: {v}")
        
        return {
            "success": True,
            "message": "\n".join(parts),
            "data": summary
        }
    
    def _exec_attendance_rules(self, params: Dict) -> Dict:
        from attendance_manager import AttendanceRules
        rules = AttendanceRules.default()
        
        parts = [
            "📐 当前考勤扣减规则",
            "",
            "  ⏰ 迟到/早退:",
            f"    免费额度: 每月 {rules.lateFreeQuota} 次",
            f"    超额扣款: 每次 ¥{float(rules.lateDeductionPerTime):,.0f}",
            f"    早退规则: 同迟到（免费 {rules.earlyLeaveFreeQuota} 次，超额 ¥{float(rules.earlyLeaveDeductionPerTime):,.0f}/次）",
            "",
            "  📝 事假:",
            f"    扣薪比例: {float(rules.personalLeaveDeductionRate)*100:.0f}% 日薪（1天事假扣1天工资）",
            "",
            "  🏥 病假:",
            f"    扣薪比例: {float(rules.sickLeaveDeductionRate)*100:.0f}% 日薪",
            f"    免扣天数: 每月 {rules.sickLeaveFreeDays} 天",
            "",
            "  ⚠️ 旷工:",
            f"    扣薪倍数: {float(rules.absentDeductionMultiplier):.1f} 倍日薪",
            "",
            "  ⏳ 加班:",
            f"    加班费率: {float(rules.overtimePayRate):.1f} 倍（按日薪/8h 计算）",
            "",
            "  💡 日薪计算: 月薪 ÷ 21.75（法定月均工作日）",
            "",
            "  如需修改规则，可直接告诉我，如「迟到每次扣100元」。",
        ]
        
        return {
            "success": True,
            "message": "\n".join(parts),
            "data": {
                "lateFreeQuota": rules.lateFreeQuota,
                "lateDeductionPerTime": float(rules.lateDeductionPerTime),
                "earlyLeaveFreeQuota": rules.earlyLeaveFreeQuota,
                "earlyLeaveDeductionPerTime": float(rules.earlyLeaveDeductionPerTime),
                "personalLeaveDeductionRate": float(rules.personalLeaveDeductionRate),
                "sickLeaveDeductionRate": float(rules.sickLeaveDeductionRate),
                "sickLeaveFreeDays": rules.sickLeaveFreeDays,
                "absentDeductionMultiplier": float(rules.absentDeductionMultiplier),
                "overtimePayRate": float(rules.overtimePayRate),
            }
        }
    
    # --- 报表 ---
    
    def _exec_employee_statistics(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        stats = mgr.getStatistics()
        
        # 组织统计信息
        parts = [
            f"📊 员工统计",
            f"  总人数: {stats['total']} 人",
            f"  在职: {stats['active']} 人",
            f"",
            f"  按状态分布:",
        ]
        for status, count in stats.get("byStatus", {}).items():
            parts.append(f"    {status}: {count} 人")
        
        parts.append(f"")
        parts.append(f"  按部门分布:")
        for dept, count in sorted(stats.get("byDepartment", {}).items(), key=lambda x: -x[1]):
            parts.append(f"    {dept}: {count} 人")
        
        if stats.get("contractExpiring"):
            parts.append(f"")
            parts.append(f"  ⚠️ 30天内合同到期: {len(stats['contractExpiring'])} 人")
            for item in stats["contractExpiring"]:
                parts.append(f"    - {item['name']}（{item['deptName']}）还剩 {item['daysLeft']} 天")
        
        if stats.get("probationExpiring"):
            parts.append(f"")
            parts.append(f"  📋 30天内试用期到期: {len(stats['probationExpiring'])} 人")
            for item in stats["probationExpiring"]:
                parts.append(f"    - {item['name']}（{item['deptName']}）还剩 {item['daysLeft']} 天")
        
        if stats.get("birthdayReminder"):
            parts.append(f"")
            parts.append(f"  🎂 7天内生日: {len(stats['birthdayReminder'])} 人")
            for item in stats["birthdayReminder"]:
                parts.append(f"    - {item['name']}（{item['birthday']}）还剩 {item['daysLeft']} 天")
        
        return {
            "success": True,
            "message": "\n".join(parts),
            "data": stats
        }
    
    def _exec_data_validation(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        result = mgr.validateData()
        
        parts = [f"🔍 数据校验结果（共 {result['info'][0] if result['info'] else '?'}）"]
        
        if result["errors"]:
            parts.append(f"\n  ❌ 错误 ({len(result['errors'])}):")
            for err in result["errors"][:10]:
                parts.append(f"    - {err}")
        
        if result["warnings"]:
            parts.append(f"\n  ⚠️ 警告 ({len(result['warnings'])}):")
            for warn in result["warnings"][:10]:
                parts.append(f"    - {warn}")
        
        if not result["errors"] and not result["warnings"]:
            parts.append(f"\n  ✅ 数据质量良好，没有发现错误或警告")
        
        return {
            "success": result["errorCount"] == 0,
            "message": "\n".join(parts),
            "data": result
        }
    
    def _exec_export_report(self, params: Dict) -> Dict:
        mgr = self.employeeManager
        if not mgr:
            return {"success": False, "message": "未配置花名册文件，请先完成初始化"}
        
        report_type = params.get("reportType", "employee")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"hr_report_{report_type}_{timestamp}.xlsx"
        
        # 安全校验：确保导出目录是用户的花名册所在目录（已初始化的用户目录）
        source_dir = os.path.realpath(os.path.dirname(mgr.filePath))
        
        # 简单导出：把员工列表写入新 Excel
        try:
            from openpyxl import Workbook
            from employee_manager import Employee
            
            wb = Workbook()
            ws = wb.active
            
            if report_type == "employee":
                ws.title = "员工花名册"
                employees = mgr.listEmployees(includeInactive=True)
                headers = ["工号", "姓名", "部门", "岗位", "状态", "入职日期", "手机号", "基本工资"]
                for col, header in enumerate(headers, 1):
                    ws.cell(1, col, header)
                for row, emp in enumerate(employees, 2):
                    ws.cell(row, 1, emp.empNo)
                    ws.cell(row, 2, emp.name)
                    ws.cell(row, 3, emp.deptName)
                    ws.cell(row, 4, emp.position)
                    ws.cell(row, 5, emp.status)
                    ws.cell(row, 6, emp.hireDate)
                    ws.cell(row, 7, emp.phone)
                    ws.cell(row, 8, emp.baseSalary)
                
                output_path = os.path.join(source_dir, f"员工花名册_{timestamp}.xlsx")
            else:
                output_path = os.path.join(source_dir, default_filename)
            
            wb.save(output_path)
            
            return {
                "success": True,
                "message": f"已导出到: {output_path}",
                "filePath": output_path,
                "count": len(mgr.employees)
            }
        except Exception as e:
            return {"success": False, "message": f"导出失败: {str(e)}"}
    
    # --- 帮助 ---
    
    def _exec_help(self, params: Dict) -> Dict:
        return {
            "success": True,
            "message": (
                "👔 HR 智能体使用指南\n"
                "\n"
                "📋 员工管理:\n"
                "  • 「查看 E001」- 查询员工信息\n"
                "  • 「列出所有员工」- 查看花名册\n"
                "  • 「搜索技术部员工」- 按部门/状态/关键词筛选\n"
                "  • 「添加员工」- 新增员工\n"
                "  • 「张伟转正」- 修改员工状态\n"
                "  • 「E003离职」- 员工离职\n"
                "  • 「批量转正 E011,E012」- 批量操作\n"
                "\n"
                "🏢 组织架构:\n"
                "  • 「组织架构」- 查看部门树\n"
                "  • 「张伟的汇报链」- 查看汇报关系\n"
                "\n"
                "💰 薪资:\n"
                "  • 「计算本月薪资」- 批量计算\n"
                "  • 「帮E001算工资」- 单人计算\n"
                "  • 「年终奖36000」- 年终奖计算\n"
                "  • 「北京社保10000」- 社保计算\n"
                "\n"
                "📊 报表:\n"
                "  • 「员工统计」- 人数统计和提醒\n"
                "  • 「校验数据」- 数据质量检查\n"
                "  • 「导出报表」- 导出Excel\n"
                "\n"
                "⚙️ 设置:\n"
                "  • 「开始初始化」- 首次使用设置\n"
                "  • 「查看配置」- 查看当前配置\n"
            ),
        }


# ============================================================
# 响应格式化器
# ============================================================

class ResponseFormatter:
    """响应格式化器"""
    
    @staticmethod
    def format(result: Dict, intent: IntentResult) -> str:
        """格式化执行结果为用户可读的文本"""
        if result.get("success"):
            return result.get("message", "操作成功")
        else:
            return result.get("message", "操作失败，请重试")
    
    @staticmethod
    def askForParams(missing: List[str], intent: IntentType) -> str:
        """生成追问提示"""
        prompts = {
            "empNo": "请提供员工工号（如 E001）或姓名",
            "empName": "请提供员工姓名",
            "bonusAmount": "请提供年终奖金额",
            "city": "请提供缴纳城市（如 北京、上海、深圳）",
            "filePath": "请提供文件路径",
            "tableType": "请指定要绑定的表格类型（组织架构/花名册/薪资）",
            "empNos": "请提供要操作的员工工号列表",
            "newStatus": "请指定要变更的状态（在职/试用期/离职）",
        }
        
        parts = []
        for p in missing:
            parts.append(prompts.get(p, f"请提供 {p}"))
        
        return "请补充以下信息：\n" + "\n".join(f"  • {p}" for p in parts)


# ============================================================
# 意图路由器（统一入口）
# ============================================================

class IntentRouter:
    """
    意图路由器
    统一入口，串联 意图识别 → 参数提取 → Tool 执行 → 响应格式化
    """
    
    def __init__(self, config: Optional[Dict] = None):
        self.classifier = IntentClassifier()
        self.extractor = ParameterExtractor()
        self.executor = ToolExecutor(config)
        self.formatter = ResponseFormatter()
        self.context = ConversationContext()
    
    def process(self, text: str) -> Dict[str, Any]:
        """
        处理用户输入
        
        Args:
            text: 用户输入的文本
        
        Returns:
            处理结果 {
                "success": bool,
                "message": str,      # 格式化后的响应文本
                "intent": str,       # 识别的意图
                "confidence": float, # 置信度
                "data": dict,        # 原始数据（供进一步处理）
                "needsInput": bool,  # 是否需要用户进一步输入
            }
        """
        # 1. 意图识别
        intent_result = self.classifier.classify(text, self.context)
        
        # 2. 记录对话
        self.context.addTurn("user", text, intent_result.intent)
        self.context.lastQuery = text
        
        # 3. 未知意图
        if intent_result.intent == IntentType.UNKNOWN:
            response = self._handleUnknown(text)
            self.context.addTurn("assistant", response, IntentType.UNKNOWN)
            return {
                "success": False,
                "message": response,
                "intent": "unknown",
                "confidence": 0,
            }
        
        # 4. 参数提取
        params = self.extractor.extract(text, intent_result, self.context)
        
        # 5. 检查必要参数
        missing = self.extractor.checkMissingParams(intent_result.intent, params)
        
        if missing:
            # 设置待追问状态
            self.context.pendingAction = {
                "intent": intent_result.intent,
                "params": params,
                "missingParams": missing,
            }
            self.context.lastIntent = intent_result.intent
            
            followup = self.formatter.askForParams(missing, intent_result.intent)
            self.context.addTurn("assistant", followup, intent_result.intent)
            
            return {
                "success": False,
                "message": followup,
                "intent": intent_result.intent.value,
                "confidence": intent_result.confidence,
                "needsInput": True,
                "extractedParams": params,
            }
        
        # 6. 执行 Tool
        result = self.executor.execute(intent_result.intent, params)
        
        # 7. 格式化响应
        response = self.formatter.format(result, intent_result)
        
        # 8. 更新上下文
        self.context.lastIntent = intent_result.intent
        self.context.lastResults = {
            "result": result,
            "extractedParams": params,
        }
        self.context.pendingAction = None
        self.context.addTurn("assistant", response, intent_result.intent)
        
        return {
            "success": result.get("success", False),
            "message": response,
            "intent": intent_result.intent.value,
            "confidence": intent_result.confidence,
            "data": result.get("data"),
            "needsInput": result.get("needsInput", False),
        }
    
    def _handleUnknown(self, text: str) -> str:
        """处理未知意图"""
        # 检查是否是简短回答（上下文延续）
        if self.context.pendingAction:
            return self._handleContextContinuation(text)
        
        return (
            "抱歉，我没有理解您的意思。您可以试试：\n"
            "  • 「帮助」查看功能列表\n"
            "  • 「员工统计」查看人数\n"
            "  • 「计算本月薪资」计算工资\n"
            "  • 「查看组织架构」查看部门"
        )
    
    def _handleContextContinuation(self, text: str) -> str:
        """处理上下文延续（用户在回答追问）"""
        pending = self.context.pendingAction
        if not pending:
            return "请问有什么可以帮您？"
        
        intent = pending["intent"]
        params = pending["params"]
        missing = pending["missingParams"]
        
        # 尝试把用户的回答作为缺失参数
        if "empNo" in missing:
            # 用户可能输入了工号或姓名
            emp_match = re.search(r'[Ee](\d{2,4})', text)
            if emp_match:
                params["empNo"] = f"E{emp_match.group(1).zfill(3)}"
                missing.remove("empNo")
            else:
                params["empName"] = text.strip()
                params["empNo"] = text.strip()  # 先当姓名处理，executor 会尝试查找
                missing.remove("empNo")
        
        if "city" in missing:
            for city_pattern in ParameterExtractor.PATTERNS["city"]:
                match = re.search(city_pattern, text)
                if match:
                    params["city"] = match.group(1)
                    missing.remove("city")
                    break
        
        if "bonusAmount" in missing:
            amount_match = re.search(r'(\d+(?:,\d{3})*(?:\.\d+)?)', text)
            if amount_match:
                raw = amount_match.group(1).replace(",", "")
                params["bonusAmount"] = float(raw)
                missing.remove("bonusAmount")
        
        if missing:
            followup = self.formatter.askForParams(missing, intent)
            return followup
        
        # 参数补全，执行
        result = self.executor.execute(intent, params)
        response = self.formatter.format(result, IntentResult(intent=intent, confidence=0.7, matchedPatterns=[]))
        
        self.context.pendingAction = None
        self.context.lastResults = {"result": result, "extractedParams": params}
        
        return response


# ============================================================
# 便捷函数（供外部调用）
# ============================================================

def _load_onboarding_config() -> Dict:
    """
    从 OnboardingManager 的配置中构建 ToolExecutor 所需的 config 字典。
    自动在 process_user_input 调用时加载，无需手动传参。
    """
    try:
        from onboarding import OnboardingManager
        mgr = OnboardingManager()
        cfg = mgr.config

        result = {}

        # 花名册配置 → filePath / columnMapping / sheetName
        if cfg.employee and cfg.employee.isBound:
            result["filePath"] = cfg.employee.filePath
            result["columnMapping"] = cfg.employee.columnMapping
            if cfg.employee.sheetName:
                result["sheetName"] = cfg.employee.sheetName

        # 组织架构配置 → orgFilePath / orgColumnMapping / orgSheetName
        if cfg.organization and cfg.organization.isBound:
            result["orgFilePath"] = cfg.organization.filePath
            result["orgColumnMapping"] = cfg.organization.columnMapping
            if cfg.organization.sheetName:
                result["orgSheetName"] = cfg.organization.sheetName

        # 薪资表配置 → salaryFilePath / salaryColumnMapping
        if cfg.salary and cfg.salary.isBound:
            result["salaryFilePath"] = cfg.salary.filePath
            result["salaryColumnMapping"] = cfg.salary.columnMapping

        # 考勤表配置 → attendanceFilePath / attendanceColumnMapping / attendanceSheetName
        if cfg.attendance and cfg.attendance.isBound:
            result["attendanceFilePath"] = cfg.attendance.filePath
            result["attendanceColumnMapping"] = cfg.attendance.columnMapping
            if cfg.attendance.sheetName:
                result["attendanceSheetName"] = cfg.attendance.sheetName

        return result
    except Exception:
        return {}


def process_user_input(text: str, config: Optional[Dict] = None) -> Dict[str, Any]:
    """
    处理用户输入（便捷函数）
    
    Args:
        text: 用户输入文本
        config: HR 配置（如不传则自动从 OnboardingManager 加载）
    
    Returns:
        处理结果
    """
    if config is None:
        config = _load_onboarding_config()
    router = IntentRouter(config)
    return router.process(text)
