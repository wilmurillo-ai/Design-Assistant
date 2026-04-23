"""
HR 智能体 - 员工管理模块
基于 Excel 的员工增删改查、查询筛选、批量操作、组织架构关联
支持 .xls（通过 xlrd）和 .xlsx/.xlsm（通过 openpyxl）格式
"""

import os
import json
import re
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from decimal import Decimal
from copy import deepcopy

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("警告：openpyxl 未安装，.xlsx 文件功能将不可用")

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    print("警告：xlrd 未安装，.xls 文件功能将不可用")


# ============================================================
# 数据模型
# ============================================================

@dataclass
class Employee:
    """员工信息"""
    # 基本信息
    empNo: str                              # 工号（唯一标识）
    name: str                               # 姓名
    deptCode: str = ""                      # 部门编码
    deptName: str = ""                      # 部门名称
    position: str = ""                      # 岗位/职位
    jobLevel: str = ""                      # 职级
    
    # 个人信息
    gender: str = ""                        # 性别
    birthDate: str = ""                     # 出生日期
    idCard: str = ""                        # 身份证号
    phone: str = ""                         # 手机号
    email: str = ""                         # 邮箱
    address: str = ""                       # 住址
    
    # 职务信息
    hireDate: str = ""                      # 入职日期
    regularDate: str = ""                   # 转正日期
    leaveDate: str = ""                     # 离职日期
    status: str = "在职"                    # 在职状态（在职/试用期/离职）
    contractType: str = ""                  # 合同类型（固定期限/无固定期限/劳务派遣）
    contractStart: str = ""                 # 合同开始日期
    contractEnd: str = ""                   # 合同结束日期
    
    # 薪资信息
    baseSalary: float = 0                   # 基本工资
    socialInsuranceCity: str = ""           # 社保缴纳城市
    socialInsuranceBase: float = 0          # 社保缴费基数
    housingFundBase: float = 0              # 公积金缴费基数（0 表示与社保基数相同）
    housingFundRate: float = 0              # 公积金比例
    specialDeduction: float = 0             # 专项附加扣除（每月）
    
    # 汇报关系
    reportTo: str = ""                      # 汇报对象（上级工号）
    reportToName: str = ""                  # 汇报对象姓名
    
    # 扩展字段
    extraFields: Dict[str, Any] = field(default_factory=dict)
    
    # 内部字段（不写入 Excel）
    _rowIndex: int = -1                     # Excel 行号
    _isNew: bool = True                     # 是否新增（未写入文件）
    _isModified: bool = False               # 是否已修改
    
    def toDict(self) -> Dict:
        """转换为字典（排除内部字段）"""
        data = {}
        for key, value in asdict(self).items():
            if not key.startswith('_'):
                data[key] = value
        # 合并扩展字段
        data.update(self.extraFields)
        return data
    
    def validate(self) -> Tuple[bool, List[str]]:
        """验证员工信息的完整性"""
        errors = []
        
        if not self.empNo.strip():
            errors.append("工号不能为空")
        
        if not self.name.strip():
            errors.append("姓名不能为空")
        
        # 验证手机号格式
        if self.phone:
            phonePattern = re.compile(r'^1[3-9]\d{9}$')
            if not phonePattern.match(self.phone):
                errors.append(f"手机号格式不正确: {self.phone}")
        
        # 验证邮箱格式
        if self.email:
            emailPattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
            if not emailPattern.match(self.email):
                errors.append(f"邮箱格式不正确: {self.email}")
        
        # 验证身份证号格式（简单校验）
        if self.idCard:
            idPattern = re.compile(r'^\d{17}[\dXx]$')
            if not idPattern.match(self.idCard):
                errors.append(f"身份证号格式不正确: {self.idCard}")
        
        # 验证日期格式
        dateFields = [
            ("birthDate", "出生日期"),
            ("hireDate", "入职日期"),
            ("leaveDate", "离职日期"),
            ("contractStart", "合同开始日期"),
            ("contractEnd", "合同结束日期"),
        ]
        for fieldName, label in dateFields:
            value = getattr(self, fieldName)
            if value and not self._isValidDate(value):
                errors.append(f"{label}格式不正确: {value}")
        
        # 验证逻辑关系
        if self.status == "离职" and not self.leaveDate:
            errors.append("离职状态必须填写离职日期")
        
        return len(errors) == 0, errors
    
    @staticmethod
    def _isValidDate(dateStr: str) -> bool:
        """验证日期格式（支持多种格式）"""
        patterns = [
            r'^\d{4}-\d{2}-\d{2}$',
            r'^\d{4}/\d{2}/\d{2}$',
            r'^\d{4}\.\d{2}\.\d{2}$',
        ]
        return any(re.match(p, dateStr) for p in patterns)


@dataclass
class Department:
    """部门信息"""
    deptCode: str                           # 部门编码（唯一标识）
    deptName: str                           # 部门名称
    parentCode: str = ""                    # 上级部门编码
    manager: str = ""                       # 部门负责人工号
    managerName: str = ""                   # 部门负责人姓名
    level: int = 1                          # 部门层级
    sortOrder: int = 0                      # 排序号
    status: str = "正常"                    # 部门状态
    employeeCount: int = 0                  # 员工人数（自动计算）


# ============================================================
# 列映射配置
# ============================================================

# 标准字段 → 列名映射
STANDARD_EMPLOYEE_FIELDS = {
    "empNo": "工号",
    "name": "姓名",
    "deptCode": "部门编码",
    "deptName": "部门名称",
    "position": "岗位",
    "jobLevel": "职级",
    "gender": "性别",
    "birthDate": "出生日期",
    "idCard": "身份证号",
    "phone": "手机号",
    "email": "邮箱",
    "hireDate": "入职日期",
    "regularDate": "转正日期",
    "leaveDate": "离职日期",
    "status": "在职状态",
    "contractType": "合同类型",
    "contractStart": "合同开始日期",
    "contractEnd": "合同结束日期",
    "baseSalary": "基本工资",
    "socialInsuranceCity": "社保城市",
    "socialInsuranceBase": "社保基数",
    "housingFundBase": "公积金基数",
    "housingFundRate": "公积金比例",
    "specialDeduction": "专项附加扣除",
    "reportTo": "汇报对象",
}

# 列名 → 标准字段的反向别名映射
FIELD_ALIASES = {
    "empNo": ["工号", "员工编号", "员工ID", "编号", "emp_id", "employee_id", "员工号", "职工号", "工号编号"],
    "name": ["姓名", "员工姓名", "名字", "name", "employee_name", "员工名字"],
    "deptCode": ["部门编码", "部门ID", "部门代码", "dept_id", "dept_code", "部门编号"],
    "deptName": ["部门名称", "部门", "部门名", "dept_name", "department", "所属部门"],
    "position": ["岗位", "职位", "职务", "岗位名称", "position", "job_title", "职级", "岗位职级"],
    "jobLevel": ["职级", "级别", "level", "grade", "岗位级别", "P级"],
    "gender": ["性别", "gender", "sex"],
    "birthDate": ["出生日期", "生日", "出生年月", "birth_date", "birthday"],
    "idCard": ["身份证号", "身份证", "证件号", "id_card", "id_number", "身份证号码"],
    "phone": ["手机号", "电话", "联系电话", "手机", "phone", "mobile", "手机号码", "联系方式"],
    "email": ["邮箱", "电子邮件", "邮箱地址", "email", "邮件"],
    "hireDate": ["入职日期", "入职时间", "入职日", "hire_date", "join_date", "入职年月", "到岗日期"],
    "regularDate": ["转正日期", "转正时间", "regular_date"],
    "leaveDate": ["离职日期", "离职时间", "离职日", "leave_date", "exit_date"],
    "status": ["在职状态", "状态", "员工状态", "status", "employment_status", "是否在职", "人员状态"],
    "contractType": ["合同类型", "合同性质", "contract_type"],
    "contractStart": ["合同开始日期", "合同起始日", "合同生效日"],
    "contractEnd": ["合同结束日期", "合同到期日", "合同终止日"],
    "baseSalary": ["基本工资", "底薪", "基础工资", "base_salary", "basic_salary", "基本薪资", "月薪"],
    "socialInsuranceCity": ["社保城市", "缴纳城市", "社保缴纳地"],
    "socialInsuranceBase": ["社保基数", "缴费基数", "社保缴费基数"],
    "housingFundBase": ["公积金基数", "公积金缴费基数", "住房公积金额度"],
    "housingFundRate": ["公积金比例", "公积金缴存比例"],
    "specialDeduction": ["专项附加扣除", "专项扣除", "附加扣除", "专项附加", "special_deduction"],
    "reportTo": ["汇报对象", "直属上级", "上级", "汇报给", "manager"],
}

# 必填字段
REQUIRED_FIELDS = ["empNo", "name"]
# 核心字段（建议填写）
RECOMMENDED_FIELDS = ["empNo", "name", "deptCode", "deptName", "position", "hireDate", "status", "phone"]


# ============================================================
# 员工管理器
# ============================================================

class EmployeeManager:
    """
    员工管理器
    
    基于 Excel 文件的员工增删改查操作。
    所有操作都在内存中进行，需要显式调用 save() 写回文件。
    """
    
    def __init__(self, filePath: str, config: Optional[Dict] = None):
        """
        初始化员工管理器
        
        Args:
            filePath: 员工花名册 Excel 文件路径（支持 .xls 和 .xlsx）
            config: 列映射配置 {标准字段: 实际列名}，
                    也支持 sheetName 字段来指定工作表名称
        """
        # 根据文件扩展名判断需要哪个库
        _, ext = os.path.splitext(filePath)
        ext = ext.lower()
        
        if ext == '.xls':
            if not XLRD_AVAILABLE:
                raise ImportError("xlrd 未安装，无法处理 .xls 文件。请执行: pip install xlrd")
        elif ext in ('.xlsx', '.xlsm'):
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl 未安装，无法处理 .xlsx 文件。请执行: pip install openpyxl")
        else:
            # 未知扩展名，尝试 openpyxl
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl 未安装，无法处理 Excel 文件")
        
        self.filePath = filePath
        self.config = config or {}
        self.workbook: Optional[Workbook] = None
        self.sheetName: str = ""
        self.employees: Dict[str, Employee] = {}  # empNo -> Employee
        self.departments: Dict[str, Department] = {}  # deptCode -> Department
        self._columnMapping: Dict[str, str] = {}  # 标准字段 -> Excel列名
        self._columnIndex: Dict[str, int] = {}    # 标准字段 -> Excel列索引
        self._headerRow: List[str] = []            # Excel 表头
        self._dataStartRow: int = 2                # 数据开始行
        self._loaded: bool = False
        
        if os.path.exists(filePath):
            self._load()
    
    # ========================================================
    # 加载和保存
    # ========================================================
    
    def _load(self):
        """从 Excel 文件加载数据"""
        try:
            _, ext = os.path.splitext(self.filePath)
            ext = ext.lower()
            
            if ext == '.xls' and XLRD_AVAILABLE:
                self._loadXls()
            else:
                self._loadXlsx()
            
            self._loaded = True
            
        except Exception as e:
            raise RuntimeError(f"加载员工花名册失败: {e}")
    
    def _loadXlsx(self):
        """从 .xlsx 文件加载数据（openpyxl）"""
        self.workbook = load_workbook(self.filePath, data_only=True)

        # 优先使用 config 中指定的 sheetName（Bug3 修复）
        targetSheet = self.config.get("sheetName", "")
        if targetSheet and targetSheet in self.workbook.sheetnames:
            self.sheetName = targetSheet
        else:
            self.sheetName = self.workbook.sheetnames[0]

        ws = self.workbook[self.sheetName]
        
        # 读取表头
        self._headerRow = []
        for colIdx in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=colIdx).value
            self._headerRow.append(str(value) if value else f"Column{colIdx}")
        
        # 建立列映射（标准字段 → Excel列名/索引）
        self._buildColumnMapping()
        
        # 读取员工数据
        self._loadEmployees(ws)
    
    def _loadXls(self):
        """从 .xls 文件加载数据（xlrd）"""
        book = xlrd.open_workbook(self.filePath)

        # 优先使用 config 中指定的 sheetName（Bug3 修复）
        targetSheet = self.config.get("sheetName", "")
        if targetSheet and targetSheet in book.sheet_names():
            sheet = book.sheet_by_name(targetSheet)
            self.sheetName = targetSheet
        else:
            sheet = book.sheet_by_index(0)
            self.sheetName = book.sheet_names()[0]
        
        # 读取表头
        self._headerRow = []
        for colIdx in range(sheet.ncols):
            value = sheet.cell_value(0, colIdx)
            self._headerRow.append(str(value) if value and value != '' else f"Column{colIdx + 1}")
        
        # 建立列映射
        self._buildColumnMapping()
        
        # 读取员工数据（使用 sheet 对象）
        self._loadEmployeesXls(sheet, book)
    
    def _buildColumnMapping(self):
        """建立标准字段与 Excel 列的映射关系"""
        self._columnMapping = {}
        self._columnIndex = {}
        usedColumns = set()
        
        # 先使用配置中的映射
        if self.config:
            for standardField, actualColumnName in self.config.items():
                if actualColumnName in self._headerRow:
                    idx = self._headerRow.index(actualColumnName)
                    self._columnMapping[standardField] = actualColumnName
                    self._columnIndex[standardField] = idx
                    usedColumns.add(idx)
        
        # 第一遍：精确匹配（优先级最高）
        for standardField, aliases in FIELD_ALIASES.items():
            if standardField in self._columnMapping:
                continue
            
            for colIdx, headerName in enumerate(self._headerRow):
                if colIdx in usedColumns:
                    continue
                if self._exactMatchFieldName(headerName, aliases):
                    self._columnMapping[standardField] = headerName
                    self._columnIndex[standardField] = colIdx
                    usedColumns.add(colIdx)
                    break
        
        # 第二遍：包含匹配（已有列不再被抢占）
        for standardField, aliases in FIELD_ALIASES.items():
            if standardField in self._columnMapping:
                continue
            
            for colIdx, headerName in enumerate(self._headerRow):
                if colIdx in usedColumns:
                    continue
                if self._matchFieldName(headerName, aliases):
                    self._columnMapping[standardField] = headerName
                    self._columnIndex[standardField] = colIdx
                    usedColumns.add(colIdx)
                    break
    
    def _matchFieldName(self, columnName: str, aliases: List[str]) -> bool:
        """检查列名是否匹配别名列表（包含匹配）"""
        columnName = columnName.lower().strip()
        for alias in aliases:
            alias = alias.lower().strip()
            if alias in columnName or columnName in alias:
                return True
        return False
    
    def _exactMatchFieldName(self, columnName: str, aliases: List[str]) -> bool:
        """检查列名是否精确匹配别名列表"""
        columnName = columnName.lower().strip()
        for alias in aliases:
            alias = alias.lower().strip()
            if columnName == alias:
                return True
        return False
    
    def _loadEmployees(self, ws):
        """从工作表加载员工数据（.xlsx，openpyxl）"""
        self.employees = {}
        
        for rowIdx in range(self._dataStartRow, ws.max_row + 1):
            rowData = {}
            for colIdx in range(len(self._headerRow)):
                value = ws.cell(row=rowIdx, column=colIdx + 1).value
                rowData[self._headerRow[colIdx]] = value
            
            # 跳过空行
            if not any(v is not None for v in rowData.values()):
                continue
            
            # 将行数据转换为 Employee 对象
            emp = self._rowToEmployee(rowData)
            if emp and emp.empNo:
                emp._rowIndex = rowIdx
                emp._isNew = False
                self.employees[emp.empNo] = emp
    
    def _loadEmployeesXls(self, sheet, book):
        """从 .xls 工作表加载员工数据（xlrd）"""
        self.employees = {}
        
        for rowIdx in range(1, sheet.nrows):  # 从第 2 行开始（0-based，跳过表头）
            rowData = {}
            for colIdx in range(sheet.ncols):
                value = sheet.cell_value(rowIdx, colIdx)
                # xlrd 日期处理
                if sheet.cell_type(rowIdx, colIdx) == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_tuple(value, book.datemode)
                        value = f"{value[0]}-{value[1]:02d}-{value[2]:02d}"
                    except Exception:
                        pass
                elif value == '':
                    value = None
                rowData[self._headerRow[colIdx]] = value
            
            # 跳过空行
            if not any(v is not None for v in rowData.values()):
                continue
            
            emp = self._rowToEmployee(rowData)
            if emp and emp.empNo:
                emp._rowIndex = rowIdx + 1  # 1-based 行号
                emp._isNew = False
                self.employees[emp.empNo] = emp
    
    def _rowToEmployee(self, rowData: Dict) -> Optional[Employee]:
        """将 Excel 行数据转换为 Employee 对象"""
        # 获取工号和姓名
        empNo = self._getFieldValue(rowData, "empNo")
        name = self._getFieldValue(rowData, "name")
        
        if not empNo:
            return None
        
        empNo = str(empNo).strip()
        name = str(name).strip() if name else ""
        
        emp = Employee(empNo=empNo, name=name)
        
        # 填充其他字段
        fieldMapping = {
            "deptCode": "deptCode", "deptName": "deptName",
            "position": "position", "jobLevel": "jobLevel",
            "gender": "gender", "birthDate": "birthDate",
            "idCard": "idCard", "phone": "phone", "email": "email",
            "hireDate": "hireDate", "regularDate": "regularDate",
            "leaveDate": "leaveDate", "status": "status",
            "contractType": "contractType", "contractStart": "contractStart",
            "contractEnd": "contractEnd", "baseSalary": "baseSalary",
            "socialInsuranceCity": "socialInsuranceCity",
            "socialInsuranceBase": "socialInsuranceBase",
            "housingFundBase": "housingFundBase",
            "housingFundRate": "housingFundRate",
            "specialDeduction": "specialDeduction",
            "reportTo": "reportTo",
        }
        
        for attrName, standardField in fieldMapping.items():
            value = self._getFieldValue(rowData, standardField)
            if value is not None:
                setattr(emp, attrName, self._convertFieldValue(attrName, value))
        
        # 处理状态字段标准化
        if emp.status:
            emp.status = EmployeeManager._normalizeStatus(emp.status)
        
        # 收集未映射的字段到 extraFields
        mappedColumns = set(self._columnMapping.values())
        for colName, value in rowData.items():
            if colName not in mappedColumns and value is not None:
                # 跳过已经映射过的列
                isMapped = False
                for stdField, mappedCol in self._columnMapping.items():
                    if mappedCol == colName:
                        isMapped = True
                        break
                if not isMapped:
                    emp.extraFields[colName] = value
        
        return emp
    
    def _getFieldValue(self, rowData: Dict, standardField: str) -> Any:
        """根据标准字段名获取行数据中的值"""
        columnName = self._columnMapping.get(standardField)
        if columnName:
            return rowData.get(columnName)
        return None
    
    @staticmethod
    def _convertFieldValue(fieldName: str, value: Any) -> Any:
        """转换单元格值为合适的 Python 类型"""
        if value is None:
            return ""
        
        if isinstance(value, (int, float)):
            if fieldName in ("baseSalary", "socialInsuranceBase", "housingFundBase", "housingFundRate", "specialDeduction"):
                return float(value)
            return str(value)
        
        # 处理日期对象
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        
        return str(value).strip()
    
    @staticmethod
    def _normalizeStatus(status: str) -> str:
        """标准化在职状态"""
        status = str(status).strip().lower()
        
        if status in ("在职", "正式", "regular", "active"):
            return "在职"
        elif status in ("试用期", "试用", "probation", "intern"):
            return "试用期"
        elif status in ("离职", "已离职", "inactive", "resigned", "离开"):
            return "离职"
        
        return status
    
    def save(self):
        """保存所有修改到 Excel 文件

        注意：.xls 格式为只读（xlrd 不支持写入），尝试保存时会自动转换为 .xlsx。
        """
        # 检测文件格式，.xls 无法写入（xlrd 只读）
        _, ext = os.path.splitext(self.filePath)
        ext = ext.lower()
        if ext == '.xls':
            # 自动转换：将路径改为 .xlsx
            new_path = self.filePath.rsplit('.', 1)[0] + '.xlsx'
            import warnings
            warnings.warn(
                f".xls 格式不支持写入，将自动保存为 .xlsx: {os.path.basename(new_path)}",
                UserWarning,
                stacklevel=2
            )
            self.filePath = new_path
            # .xls 模式下 workbook 为 None（xlrd 不创建 openpyxl workbook），需要新建
            self.workbook = None

        if not self.workbook:
            # 新文件
            self.workbook = Workbook()
            ws = self.workbook.active
            ws.title = "员工花名册"
            self.sheetName = "员工花名册"

            # 写入表头
            headers = self._getOutputHeaders()
            for colIdx, header in enumerate(headers, 1):
                ws.cell(row=1, column=colIdx, value=header)
            self._headerRow = headers

            # 写入数据
            for rowIdx, emp in enumerate(self.employees.values(), 2):
                self._writeEmployeeRow(ws, rowIdx, emp)
        else:
            ws = self.workbook[self.sheetName]

            # 删除所有数据行（保留表头）
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)

            # 写入所有员工
            for rowIdx, emp in enumerate(self.employees.values(), 2):
                self._writeEmployeeRow(ws, rowIdx, emp)

        self.workbook.save(self.filePath)

        # 重置修改标记
        for emp in self.employees.values():
            emp._isNew = False
            emp._isModified = False
            emp._rowIndex = emp._rowIndex if emp._rowIndex > 0 else 0
    
    def _getOutputHeaders(self) -> List[str]:
        """获取输出表头列表"""
        if self._headerRow:
            return self._headerRow
        
        # 默认表头
        return [
            "工号", "姓名", "部门编码", "部门名称", "岗位", "职级",
            "性别", "出生日期", "身份证号", "手机号", "邮箱",
            "入职日期", "转正日期", "离职日期", "在职状态",
            "合同类型", "合同开始日期", "合同结束日期",
            "基本工资", "社保城市", "社保基数", "公积金基数", "公积金比例", "专项附加扣除",
            "汇报对象",
        ]
    
    def _writeEmployeeRow(self, ws, rowIdx: int, emp: Employee):
        """将员工数据写入 Excel 行"""
        fieldOrder = [
            ("empNo", "工号"), ("name", "姓名"),
            ("deptCode", "部门编码"), ("deptName", "部门名称"),
            ("position", "岗位"), ("jobLevel", "职级"),
            ("gender", "性别"), ("birthDate", "出生日期"),
            ("idCard", "身份证号"), ("phone", "手机号"),
            ("email", "邮箱"),
            ("hireDate", "入职日期"), ("regularDate", "转正日期"),
            ("leaveDate", "离职日期"), ("status", "在职状态"),
            ("contractType", "合同类型"),
            ("contractStart", "合同开始日期"),
            ("contractEnd", "合同结束日期"),
            ("baseSalary", "基本工资"),
            ("socialInsuranceCity", "社保城市"),
            ("socialInsuranceBase", "社保基数"),
            ("housingFundBase", "公积金基数"),
            ("housingFundRate", "公积金比例"),
            ("specialDeduction", "专项附加扣除"),
            ("reportTo", "汇报对象"),
        ]
        
        for attrName, defaultHeader in fieldOrder:
            # 查找对应的列
            mappedColumn = self._columnMapping.get(attrName)
            targetColumn = mappedColumn if mappedColumn else defaultHeader
            
            # 找到列索引
            if targetColumn in self._headerRow:
                colIdx = self._headerRow.index(targetColumn) + 1
            else:
                # 未找到列，使用默认顺序
                colIdx = fieldOrder.index((attrName, defaultHeader)) + 1
            
            value = getattr(emp, attrName, "")
            if isinstance(value, str) and not value:
                value = None
            elif isinstance(value, float) and value == 0:
                value = None
            
            ws.cell(row=rowIdx, column=colIdx, value=value)
        
        emp._rowIndex = rowIdx
    
    # ========================================================
    # CRUD 操作
    # ========================================================
    
    def getEmployee(self, empNo: str) -> Optional[Employee]:
        """根据工号获取员工"""
        return self.employees.get(empNo)
    
    def listEmployees(
        self,
        deptCode: Optional[str] = None,
        status: Optional[str] = None,
        keyword: Optional[str] = None,
        includeInactive: bool = False
    ) -> List[Employee]:
        """
        列出员工列表（支持筛选）
        
        Args:
            deptCode: 部门编码筛选
            status: 在职状态筛选
            keyword: 关键词搜索（匹配姓名、工号、手机号）
            includeInactive: 是否包含离职员工
        
        Returns:
            员工列表
        """
        result = list(self.employees.values())
        
        # 状态筛选
        if not includeInactive:
            result = [e for e in result if e.status != "离职"]
        
        if status:
            statusNormalized = EmployeeManager._normalizeStatus(status)
            result = [e for e in result if EmployeeManager._normalizeStatus(e.status) == statusNormalized]
        
        # 部门筛选（包含子部门）
        if deptCode:
            deptCodes = self._getDeptAndChildren(deptCode)
            result = [e for e in result if e.deptCode in deptCodes]
        
        # 关键词搜索
        if keyword:
            keyword = keyword.lower()
            result = [e for e in result if self._matchKeyword(e, keyword)]
        
        # 按部门、工号排序
        result.sort(key=lambda e: (e.deptCode or "", e.empNo))
        
        return result
    
    def _matchKeyword(self, emp: Employee, keyword: str) -> bool:
        """关键词匹配"""
        searchFields = [
            emp.empNo, emp.name, emp.phone, emp.email,
            emp.position, emp.deptName, emp.idCard
        ]
        return any(keyword in str(f).lower() for f in searchFields if f)
    
    def addEmployee(self, emp: Employee) -> Tuple[bool, str]:
        """
        添加新员工
        
        Args:
            emp: 员工对象
        
        Returns:
            (是否成功, 消息)
        """
        # 验证
        isValid, errors = emp.validate()
        if not isValid:
            return False, f"员工信息验证失败: {'; '.join(errors)}"
        
        # 检查工号唯一性
        if emp.empNo in self.employees:
            return False, f"工号已存在: {emp.empNo}（{self.employees[emp.empNo].name}）"
        
        # 设置默认状态
        if not emp.status:
            emp.status = "试用期"
        
        emp._isNew = True
        self.employees[emp.empNo] = emp
        
        return True, f"已添加员工: {emp.name}（{emp.empNo}）"
    
    def updateEmployee(self, empNo: str, updates: Dict[str, Any]) -> Tuple[bool, str]:
        """
        更新员工信息
        
        Args:
            empNo: 工号
            updates: 更新字段字典
        
        Returns:
            (是否成功, 消息)
        """
        emp = self.employees.get(empNo)
        if not emp:
            return False, f"员工不存在: {empNo}"
        
        # 应用更新
        for key, value in updates.items():
            if hasattr(emp, key) and not key.startswith('_'):
                setattr(emp, key, value)
            else:
                emp.extraFields[key] = value
        
        # 标准化状态
        if "status" in updates:
            emp.status = EmployeeManager._normalizeStatus(emp.status)
        
        # 验证
        isValid, errors = emp.validate()
        if not isValid:
            return False, f"更新后信息验证失败: {'; '.join(errors)}"
        
        emp._isModified = True
        
        return True, f"已更新员工: {emp.name}（{emp.empNo}）"
    
    def deleteEmployee(self, empNo: str, soft: bool = True) -> Tuple[bool, str]:
        """
        删除员工
        
        Args:
            empNo: 工号
            soft: 软删除（标记为离职）还是物理删除
        
        Returns:
            (是否成功, 消息)
        """
        emp = self.employees.get(empNo)
        if not emp:
            return False, f"员工不存在: {empNo}"
        
        if soft:
            emp.status = "离职"
            emp.leaveDate = datetime.now().strftime("%Y-%m-%d")
            emp._isModified = True
            return True, f"已将员工标记为离职: {emp.name}（{emp.empNo}）"
        else:
            del self.employees[empNo]
            return True, f"已删除员工: {emp.name}（{emp.empNo}）"
    
    # ========================================================
    # 批量操作
    # ========================================================
    
    def batchAddEmployees(self, employees: List[Employee]) -> Dict[str, Any]:
        """
        批量添加员工
        
        Args:
            employees: 员工列表
        
        Returns:
            操作结果 {success: int, failed: [{empNo, name, error}]}
        """
        successCount = 0
        failedList = []
        
        for emp in employees:
            ok, msg = self.addEmployee(emp)
            if ok:
                successCount += 1
            else:
                failedList.append({
                    "empNo": emp.empNo,
                    "name": emp.name,
                    "error": msg
                })
        
        return {
            "successCount": successCount,
            "failedCount": len(failedList),
            "failed": failedList
        }
    
    def batchUpdateStatus(
        self,
        empNos: List[str],
        newStatus: str,
        leaveDate: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        批量更新员工状态
        
        Args:
            empNos: 工号列表
            newStatus: 新状态
            leaveDate: 离职日期（离职时使用）
        
        Returns:
            操作结果
        """
        successCount = 0
        failedList = []
        
        for empNo in empNos:
            emp = self.employees.get(empNo)
            if not emp:
                failedList.append({"empNo": empNo, "error": "员工不存在"})
                continue
            
            emp.status = EmployeeManager._normalizeStatus(newStatus)
            if newStatus == "离职" or emp.status == "离职":
                emp.leaveDate = leaveDate or datetime.now().strftime("%Y-%m-%d")
            
            emp._isModified = True
            successCount += 1
        
        return {
            "successCount": successCount,
            "failedCount": len(failedList),
            "failed": failedList
        }
    
    def importEmployees(self, filePath: str, sheetName: Optional[str] = None) -> Dict[str, Any]:
        """
        从另一个 Excel 文件导入员工数据
        
        Args:
            filePath: 导入文件路径
            sheetName: 工作表名称
        
        Returns:
            导入结果
        """
        try:
            tempManager = EmployeeManager(filePath)
            imported = list(tempManager.employees.values())
            
            result = self.batchAddEmployees(imported)
            
            return {
                "success": True,
                "importedCount": result["successCount"],
                "failedCount": result["failedCount"],
                "failed": result["failed"],
                "sourceFile": filePath
            }
        except Exception as e:
            return {
                "success": False,
                "importedCount": 0,
                "error": str(e)
            }
    
    # ========================================================
    # 统计和报表
    # ========================================================
    
    def getStatistics(self) -> Dict[str, Any]:
        """
        获取员工统计信息
        
        Returns:
            统计数据
        """
        allEmps = list(self.employees.values())
        activeEmps = [e for e in allEmps if e.status != "离职"]
        
        # 按状态统计
        statusStats = {}
        for emp in allEmps:
            status = EmployeeManager._normalizeStatus(emp.status)
            statusStats[status] = statusStats.get(status, 0) + 1
        
        # 按部门统计
        deptStats = {}
        for emp in activeEmps:
            dept = emp.deptName or "未分配"
            deptStats[dept] = deptStats.get(dept, 0) + 1
        
        # 按月入职统计（最近12个月）
        monthlyHire = {}
        now = datetime.now()
        for emp in activeEmps:
            if emp.hireDate:
                try:
                    hireDate = datetime.strptime(emp.hireDate, "%Y-%m-%d")
                    monthKey = hireDate.strftime("%Y-%m")
                    if (now - hireDate).days <= 365:
                        monthlyHire[monthKey] = monthlyHire.get(monthKey, 0) + 1
                except ValueError:
                    pass
        
        # 合同到期提醒（30天内）
        contractExpiring = []
        for emp in activeEmps:
            if emp.contractEnd:
                try:
                    endDate = datetime.strptime(emp.contractEnd, "%Y-%m-%d")
                    daysLeft = (endDate - now).days
                    if 0 <= daysLeft <= 30:
                        contractExpiring.append({
                            "empNo": emp.empNo,
                            "name": emp.name,
                            "deptName": emp.deptName,
                            "contractEnd": emp.contractEnd,
                            "daysLeft": daysLeft
                        })
                except ValueError:
                    pass
        
        # 试用期到期提醒
        probationExpiring = []
        for emp in activeEmps:
            if emp.status == "试用期" and emp.regularDate:
                try:
                    regDate = datetime.strptime(emp.regularDate, "%Y-%m-%d")
                    daysLeft = (regDate - now).days
                    if 0 <= daysLeft <= 30:
                        probationExpiring.append({
                            "empNo": emp.empNo,
                            "name": emp.name,
                            "deptName": emp.deptName,
                            "regularDate": emp.regularDate,
                            "daysLeft": daysLeft
                        })
                except ValueError:
                    pass
        
        # 生日提醒（7天内）
        birthdayReminder = []
        for emp in activeEmps:
            if emp.birthDate:
                try:
                    birthDate = datetime.strptime(emp.birthDate, "%Y-%m-%d")
                    thisYearBirthday = birthDate.replace(year=now.year)
                    if thisYearBirthday < now:
                        thisYearBirthday = thisYearBirthday.replace(year=now.year + 1)
                    daysLeft = (thisYearBirthday - now).days
                    if daysLeft <= 7:
                        birthdayReminder.append({
                            "empNo": emp.empNo,
                            "name": emp.name,
                            "birthday": birthDate.strftime("%m-%d"),
                            "daysLeft": daysLeft
                        })
                except ValueError:
                    pass
        
        return {
            "total": len(allEmps),
            "active": len(activeEmps),
            "byStatus": statusStats,
            "byDepartment": deptStats,
            "monthlyHire": monthlyHire,
            "contractExpiring": contractExpiring,
            "probationExpiring": probationExpiring,
            "birthdayReminder": birthdayReminder,
        }
    
    def getEmployeeCount(self, deptCode: Optional[str] = None) -> int:
        """获取员工人数"""
        if deptCode:
            deptCodes = self._getDeptAndChildren(deptCode)
            return sum(1 for e in self.employees.values()
                      if e.status != "离职" and e.deptCode in deptCodes)
        return sum(1 for e in self.employees.values() if e.status != "离职")
    
    # ========================================================
    # 组织架构关联
    # ========================================================
    
    def loadDepartments(self, filePath: str, config: Optional[Dict] = None):
        """
        加载组织架构（支持 .xls 和 .xlsx）
        
        Args:
            filePath: 组织架构 Excel 文件路径
            config: 列映射配置，也支持 sheetName 字段
        """
        try:
            if not os.path.exists(filePath):
                raise FileNotFoundError(f"组织架构文件不存在: {filePath}")
            
            _, ext = os.path.splitext(filePath)
            ext = ext.lower()
            
            # 提取 sheetName（如有）
            targetSheet = ""
            if config and isinstance(config, dict):
                targetSheet = config.get("sheetName", "")
            
            if ext == '.xls' and XLRD_AVAILABLE:
                self._loadDepartmentsXls(filePath, targetSheet)
            else:
                self._loadDepartmentsXlsx(filePath, targetSheet)
            
        except Exception as e:
            raise RuntimeError(f"加载组织架构失败: {e}")

    def _loadDepartmentsXlsx(self, filePath: str, targetSheet: str = ""):
        """从 .xlsx 加载组织架构"""
        wb = load_workbook(filePath, data_only=True)

        # Bug3 修复：使用指定的 sheetName
        if targetSheet and targetSheet in wb.sheetnames:
            ws = wb[targetSheet]
        else:
            ws = wb[wb.sheetnames[0]]
        
        headers = []
        for colIdx in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=colIdx).value
            headers.append(str(value) if value else f"Column{colIdx}")
        
        colMapping = self._matchDeptColumns(headers)
        
        self.departments = {}
        for rowIdx in range(2, ws.max_row + 1):
            deptCode = ws.cell(row=rowIdx, column=colMapping.get("deptCode", 1) + 1).value
            deptName = ws.cell(row=rowIdx, column=colMapping.get("deptName", 2) + 1).value
            parentCode = ws.cell(row=rowIdx, column=colMapping.get("parentCode", 3) + 1).value
            manager = ws.cell(row=rowIdx, column=colMapping.get("manager", 4) + 1).value
            level = ws.cell(row=rowIdx, column=colMapping.get("level", 5) + 1).value
            
            if deptCode:
                dept = Department(
                    deptCode=str(deptCode).strip(),
                    deptName=str(deptName).strip() if deptName else "",
                    parentCode=str(parentCode).strip() if parentCode else "",
                    manager=str(manager).strip() if manager else "",
                    level=int(level) if level and isinstance(level, (int, float)) else 1,
                )
                self.departments[dept.deptCode] = dept
        
        self._finalizeDepartments()
    
    def _loadDepartmentsXls(self, filePath: str, targetSheet: str = ""):
        """从 .xls 加载组织架构（xlrd）"""
        book = xlrd.open_workbook(filePath)

        # Bug3 修复：使用指定的 sheetName
        if targetSheet and targetSheet in book.sheet_names():
            sheet = book.sheet_by_name(targetSheet)
        else:
            sheet = book.sheet_by_index(0)
        
        # 读取表头
        headers = []
        for colIdx in range(sheet.ncols):
            value = sheet.cell_value(0, colIdx)
            headers.append(str(value) if value and value != '' else f"Column{colIdx + 1}")
        
        colMapping = self._matchDeptColumns(headers)
        
        self.departments = {}
        for rowIdx in range(1, sheet.nrows):
            deptCode = sheet.cell_value(rowIdx, colMapping.get("deptCode", 0))
            deptName = sheet.cell_value(rowIdx, colMapping.get("deptName", 1))
            parentCode = sheet.cell_value(rowIdx, colMapping.get("parentCode", 2))
            manager = sheet.cell_value(rowIdx, colMapping.get("manager", 3))
            level = sheet.cell_value(rowIdx, colMapping.get("level", 4))
            
            if deptCode and deptCode != '':
                dept = Department(
                    deptCode=str(deptCode).strip(),
                    deptName=str(deptName).strip() if deptName and deptName != '' else "",
                    parentCode=str(parentCode).strip() if parentCode and parentCode != '' else "",
                    manager=str(manager).strip() if manager and manager != '' else "",
                    level=int(level) if level and isinstance(level, (int, float)) else 1,
                )
                self.departments[dept.deptCode] = dept
        
        self._finalizeDepartments()
    
    def _matchDeptColumns(self, headers: List[str]) -> Dict[str, int]:
        """匹配部门列名，返回 {标准字段: 列索引}"""
        deptAliases = {
            "deptCode": ["部门编码", "部门ID", "部门代码", "dept_id", "dept_code", "部门编号"],
            "deptName": ["部门名称", "部门", "部门名", "dept_name", "department"],
            "parentCode": ["上级部门", "父部门", "上级编码", "parent_id", "parent_code"],
            "manager": ["负责人", "经理", "主管", "manager", "head", "部门负责人"],
            "level": ["层级", "级别", "level", "grade", "部门层级"],
        }
        
        colMapping = {}
        usedCols = set()
        # 第一遍：精确匹配
        for fieldKey, aliases in deptAliases.items():
            for colIdx, header in enumerate(headers):
                if colIdx in usedCols:
                    continue
                if any(alias.lower() == header.lower() for alias in aliases):
                    colMapping[fieldKey] = colIdx
                    usedCols.add(colIdx)
                    break
        # 第二遍：包含匹配
        for fieldKey, aliases in deptAliases.items():
            if fieldKey in colMapping:
                continue
            for colIdx, header in enumerate(headers):
                if colIdx in usedCols:
                    continue
                if any(alias.lower() in header.lower() or header.lower() in alias.lower() for alias in aliases):
                    colMapping[fieldKey] = colIdx
                    usedCols.add(colIdx)
                    break
        
        return colMapping
    
    def _finalizeDepartments(self):
        """加载组织架构后的通用收尾逻辑"""
        self._syncDeptNames()
        
        for dept in self.departments.values():
            dept.employeeCount = self.getEmployeeCount(dept.deptCode)
        
        for dept in self.departments.values():
            if dept.manager:
                emp = self.employees.get(dept.manager)
                if emp:
                    dept.managerName = emp.name
    
    def _syncDeptNames(self):
        """同步员工的部门名称（从组织架构获取）"""
        for emp in self.employees.values():
            if emp.deptCode and not emp.deptName:
                dept = self.departments.get(emp.deptCode)
                if dept:
                    emp.deptName = dept.deptName
    
    def _getDeptAndChildren(self, deptCode: str) -> set:
        """获取部门及其所有子部门编码"""
        from collections import deque
        
        result = {deptCode}
        queue = deque([deptCode])
        
        while queue:
            current = queue.popleft()
            for dept in self.departments.values():
                if dept.parentCode == current and dept.deptCode not in result:
                    result.add(dept.deptCode)
                    queue.append(dept.deptCode)
        
        return result
    
    def getDeptTree(self) -> List[Dict]:
        """
        获取部门树结构
        
        Returns:
            部门树 JSON
        """
        # 找到根部门
        rootDepts = [d for d in self.departments.values() if not d.parentCode or d.parentCode not in self.departments]
        
        def buildTree(dept: Department) -> Dict:
            children = [d for d in self.departments.values() if d.parentCode == dept.deptCode]
            return {
                "deptCode": dept.deptCode,
                "deptName": dept.deptName,
                "manager": dept.manager,
                "managerName": dept.managerName,
                "employeeCount": self.getEmployeeCount(dept.deptCode),
                "level": dept.level,
                "children": [buildTree(c) for c in sorted(children, key=lambda x: x.sortOrder)]
            }
        
        return [buildTree(d) for d in sorted(rootDepts, key=lambda x: x.sortOrder)]
    
    def getReportingChain(self, empNo: str) -> List[Dict]:
        """
        获取员工的汇报链（从直属上级到顶层）
        
        Args:
            empNo: 员工工号
        
        Returns:
            汇报链列表 [{empNo, name, position, deptName}, ...]
        """
        chain = []
        visited = set()
        currentEmpNo = empNo
        
        while currentEmpNo and currentEmpNo not in visited:
            visited.add(currentEmpNo)
            emp = self.employees.get(currentEmpNo)
            if not emp:
                break
            
            chain.append({
                "empNo": emp.empNo,
                "name": emp.name,
                "position": emp.position,
                "deptName": emp.deptName,
                "reportTo": emp.reportTo,
            })
            
            currentEmpNo = emp.reportTo
        
        return chain
    
    # ========================================================
    # 数据校验
    # ========================================================
    
    def validateData(self) -> Dict[str, Any]:
        """
        全面校验花名册数据
        
        Returns:
            校验结果 {errors: [], warnings: [], info: []}
        """
        errors = []
        warnings = []
        info = []
        
        # 1. 必填字段检查
        for empNo, emp in self.employees.items():
            if not emp.name:
                errors.append(f"[{empNo}] 姓名为空")
            if not emp.hireDate:
                warnings.append(f"[{empNo}] {emp.name} 缺少入职日期")
        
        # 2. 工号唯一性检查（已在 addEmployee 时保证）
        
        # 3. 部门关联检查
        if self.departments:
            for empNo, emp in self.employees.items():
                if emp.status != "离职" and emp.deptCode:
                    if emp.deptCode not in self.departments:
                        warnings.append(f"[{empNo}] {emp.name} 的部门编码 {emp.deptCode} 在组织架构中不存在")
        
        # 4. 汇报关系检查
        for empNo, emp in self.employees.items():
            if emp.status != "离职" and emp.reportTo:
                if emp.reportTo not in self.employees:
                    warnings.append(f"[{empNo}] {emp.name} 的汇报对象 {emp.reportTo} 不存在于花名册")
                elif self.employees[emp.reportTo].status == "离职":
                    warnings.append(f"[{empNo}] {emp.name} 的汇报对象 {self.employees[emp.reportTo].name} 已离职")
        
        # 5. 日期逻辑检查
        for empNo, emp in self.employees.items():
            if emp.contractEnd and emp.contractStart:
                try:
                    end = datetime.strptime(emp.contractEnd, "%Y-%m-%d")
                    start = datetime.strptime(emp.contractStart, "%Y-%m-%d")
                    if end < start:
                        errors.append(f"[{empNo}] {emp.name} 的合同结束日期早于开始日期")
                except ValueError:
                    pass
            
            if emp.leaveDate and emp.hireDate:
                try:
                    leave = datetime.strptime(emp.leaveDate, "%Y-%m-%d")
                    hire = datetime.strptime(emp.hireDate, "%Y-%m-%d")
                    if leave < hire:
                        errors.append(f"[{empNo}] {emp.name} 的离职日期早于入职日期")
                except ValueError:
                    pass
        
        info.append(f"共检查 {len(self.employees)} 名员工")
        
        return {
            "errorCount": len(errors),
            "warningCount": len(warnings),
            "infoCount": len(info),
            "errors": errors,
            "warnings": warnings,
            "info": info,
        }
    
    # ========================================================
    # 配置导出
    # ========================================================
    
    def exportColumnMapping(self) -> Dict[str, str]:
        """导出当前列映射配置"""
        return {k: v for k, v in self._columnMapping.items()}
    
    def getMappingQuality(self) -> Dict[str, Any]:
        """获取列映射质量报告"""
        mapped = set(self._columnMapping.keys())
        
        quality = {}
        for field in RECOMMENDED_FIELDS:
            quality[field] = {
                "mapped": field in mapped,
                "columnName": self._columnMapping.get(field, ""),
            }
        
        return {
            "mappedFields": len(mapped),
            "totalStandardFields": len(STANDARD_EMPLOYEE_FIELDS),
            "recommendedMapped": len([f for f in RECOMMENDED_FIELDS if f in mapped]),
            "totalRecommended": len(RECOMMENDED_FIELDS),
            "requiredMapped": all(f in mapped for f in REQUIRED_FIELDS),
            "details": quality
        }


# ============================================================
# 工具函数（供 WorkBuddy Skill 调用）
# ============================================================

def read_employee(emp_no: str, config: Dict = None) -> Dict:
    """读取单个员工信息"""
    try:
        if not config or not config.get("filePath"):
            return {"success": False, "error": "未配置员工花名册文件路径"}
        
        manager = EmployeeManager(config["filePath"], config.get("columnMapping"))
        emp = manager.getEmployee(emp_no)
        
        if not emp:
            return {"success": False, "error": f"员工不存在: {emp_no}"}
        
        return {
            "success": True,
            "data": emp.toDict()
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def list_employees(
    config: Dict = None,
    dept_code: str = None,
    status: str = None,
    keyword: str = None,
    include_inactive: bool = False
) -> Dict:
    """列出员工列表"""
    try:
        if not config or not config.get("filePath"):
            return {"success": False, "error": "未配置员工花名册文件路径"}
        
        manager = EmployeeManager(config["filePath"], config.get("columnMapping"))
        
        if config.get("orgFilePath"):
            manager.loadDepartments(config["orgFilePath"], config.get("orgColumnMapping"))
        
        employees = manager.listEmployees(
            deptCode=dept_code,
            status=status,
            keyword=keyword,
            includeInactive=include_inactive
        )
        
        return {
            "success": True,
            "count": len(employees),
            "data": [e.toDict() for e in employees]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def add_employee(
    config: Dict = None,
    emp_data: Dict = None
) -> Dict:
    """添加新员工"""
    try:
        if not config or not config.get("filePath"):
            return {"success": False, "error": "未配置员工花名册文件路径"}
        if not emp_data:
            return {"success": False, "error": "员工数据不能为空"}
        
        manager = EmployeeManager(config["filePath"], config.get("columnMapping"))
        
        emp = Employee(
            empNo=emp_data.get("empNo", ""),
            name=emp_data.get("name", ""),
            deptCode=emp_data.get("deptCode", ""),
            deptName=emp_data.get("deptName", ""),
            position=emp_data.get("position", ""),
            gender=emp_data.get("gender", ""),
            phone=emp_data.get("phone", ""),
            email=emp_data.get("email", ""),
            hireDate=emp_data.get("hireDate", ""),
            status=emp_data.get("status", "试用期"),
            baseSalary=float(emp_data.get("baseSalary", 0)),
            reportTo=emp_data.get("reportTo", ""),
        )
        
        ok, msg = manager.addEmployee(emp)
        if ok:
            manager.save()
        
        return {"success": ok, "message": msg}
    except Exception as e:
        return {"success": False, "error": str(e)}


def update_employee(
    config: Dict = None,
    emp_no: str = None,
    updates: Dict = None
) -> Dict:
    """更新员工信息"""
    try:
        if not config or not config.get("filePath"):
            return {"success": False, "error": "未配置员工花名册文件路径"}
        if not emp_no:
            return {"success": False, "error": "工号不能为空"}
        if not updates:
            return {"success": False, "error": "更新数据不能为空"}
        
        manager = EmployeeManager(config["filePath"], config.get("columnMapping"))
        ok, msg = manager.updateEmployee(emp_no, updates)
        
        if ok:
            manager.save()
        
        return {"success": ok, "message": msg}
    except Exception as e:
        return {"success": False, "error": str(e)}


def batch_update_status(
    config: Dict = None,
    emp_nos: List[str] = None,
    new_status: str = None,
    leave_date: str = None
) -> Dict:
    """批量更新员工状态"""
    try:
        if not config or not config.get("filePath"):
            return {"success": False, "error": "未配置员工花名册文件路径"}
        if not emp_nos:
            return {"success": False, "error": "员工列表不能为空"}
        if not new_status:
            return {"success": False, "error": "新状态不能为空"}
        
        manager = EmployeeManager(config["filePath"], config.get("columnMapping"))
        result = manager.batchUpdateStatus(emp_nos, new_status, leave_date)
        
        if result["successCount"] > 0:
            manager.save()
        
        return {"success": True, **result}
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_employee_statistics(config: Dict = None) -> Dict:
    """获取员工统计信息"""
    try:
        if not config or not config.get("filePath"):
            return {"success": False, "error": "未配置员工花名册文件路径"}
        
        manager = EmployeeManager(config["filePath"], config.get("columnMapping"))
        stats = manager.getStatistics()
        
        return {"success": True, **stats}
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_department_tree(config: Dict = None) -> Dict:
    """获取部门树结构"""
    try:
        if not config or not config.get("filePath"):
            return {"success": False, "error": "未配置员工花名册文件路径"}
        if not config.get("orgFilePath"):
            return {"success": False, "error": "未配置组织架构文件路径"}
        
        manager = EmployeeManager(config["filePath"], config.get("columnMapping"))
        manager.loadDepartments(config["orgFilePath"], config.get("orgColumnMapping"))
        
        tree = manager.getDeptTree()
        return {"success": True, "tree": tree}
    except Exception as e:
        return {"success": False, "error": str(e)}


def validate_employee_data(config: Dict = None) -> Dict:
    """校验花名册数据"""
    try:
        if not config or not config.get("filePath"):
            return {"success": False, "error": "未配置员工花名册文件路径"}
        
        manager = EmployeeManager(config["filePath"], config.get("columnMapping"))
        
        if config.get("orgFilePath"):
            manager.loadDepartments(config["orgFilePath"], config.get("orgColumnMapping"))
        
        result = manager.validateData()
        return {"success": True, **result}
    except Exception as e:
        return {"success": False, "error": str(e)}
