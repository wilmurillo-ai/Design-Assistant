"""
HR 智能体 - 意图路由引擎测试
覆盖：意图识别、参数提取、Tool 执行、多轮对话、端到端
"""

import os
import sys
import unittest
import tempfile
import shutil
from decimal import Decimal

# 确保可以导入同目录的模块
sys.path.insert(0, os.path.dirname(__file__))

from intent_router import (
    IntentClassifier, IntentType, ParameterExtractor, ToolExecutor,
    IntentRouter, ConversationContext, IntentResult, ResponseFormatter,
    process_user_input
)


class TestIntentClassifier(unittest.TestCase):
    """意图分类器测试"""
    
    def setUp(self):
        self.classifier = IntentClassifier()
    
    # --- 初始化引导 ---
    
    def test_start_onboarding_keywords(self):
        result = self.classifier.classify("开始初始化")
        self.assertEqual(result.intent, IntentType.START_ONBOARDING)
        self.assertGreater(result.confidence, 0.8)
    
    def test_start_onboarding_variants(self):
        for text in ["我要初始化", "首次使用", "帮我绑定表格", "开始设置"]:
            result = self.classifier.classify(text)
            self.assertEqual(result.intent, IntentType.START_ONBOARDING, f"Failed for: {text}")
    
    def test_check_config(self):
        result = self.classifier.classify("查看配置")
        self.assertEqual(result.intent, IntentType.CHECK_CONFIG)
    
    def test_check_config_variants(self):
        for text in ["绑定了哪些", "当前配置", "配置状态"]:
            result = self.classifier.classify(text)
            self.assertEqual(result.intent, IntentType.CHECK_CONFIG, f"Failed for: {text}")
    
    def test_reset_config(self):
        result = self.classifier.classify("重新初始化")
        self.assertEqual(result.intent, IntentType.RESET_CONFIG)
    
    # --- 员工查询 ---
    
    def test_read_employee(self):
        result = self.classifier.classify("查看E001的信息")
        self.assertEqual(result.intent, IntentType.READ_EMPLOYEE)
    
    def test_read_employee_variants(self):
        for text in ["员工E002详情", "工号E003", "查看员工信息"]:
            result = self.classifier.classify(text)
            self.assertIn(result.intent, [IntentType.READ_EMPLOYEE, IntentType.SEARCH_EMPLOYEE], f"Failed for: {text}")
    
    def test_list_employees(self):
        result = self.classifier.classify("列出所有员工")
        self.assertIn(result.intent, [IntentType.LIST_EMPLOYEES, IntentType.SEARCH_EMPLOYEE])
    
    def test_search_employee(self):
        result = self.classifier.classify("搜索技术部员工")
        self.assertEqual(result.intent, IntentType.SEARCH_EMPLOYEE)
    
    def test_search_employee_by_status(self):
        result = self.classifier.classify("试用期员工有哪些")
        self.assertEqual(result.intent, IntentType.SEARCH_EMPLOYEE)
    
    # --- 员工操作 ---
    
    def test_add_employee(self):
        result = self.classifier.classify("添加新员工")
        self.assertEqual(result.intent, IntentType.ADD_EMPLOYEE)
    
    def test_add_employee_variants(self):
        for text in ["新增员工", "新人入职登记", "录入一个员工", "加个员工"]:
            result = self.classifier.classify(text)
            self.assertEqual(result.intent, IntentType.ADD_EMPLOYEE, f"Failed for: {text}")
    
    def test_update_employee_transfer(self):
        result = self.classifier.classify("张伟转正")
        self.assertEqual(result.intent, IntentType.UPDATE_EMPLOYEE)
    
    def test_update_employee_salary(self):
        result = self.classifier.classify("调薪E001")
        self.assertEqual(result.intent, IntentType.UPDATE_EMPLOYEE)
    
    def test_delete_employee(self):
        result = self.classifier.classify("员工E003离职")
        self.assertEqual(result.intent, IntentType.DELETE_EMPLOYEE)
    
    def test_delete_employee_variants(self):
        for text in ["办离职", "删除E005", "E003离职了"]:
            result = self.classifier.classify(text)
            self.assertEqual(result.intent, IntentType.DELETE_EMPLOYEE, f"Failed for: {text}")
    
    def test_batch_update(self):
        result = self.classifier.classify("批量转正")
        self.assertEqual(result.intent, IntentType.BATCH_UPDATE_STATUS)
    
    # --- 组织架构 ---
    
    def test_dept_tree(self):
        result = self.classifier.classify("查看组织架构")
        self.assertEqual(result.intent, IntentType.GET_DEPT_TREE)
    
    def test_dept_tree_variants(self):
        for text in ["部门结构", "有哪些部门", "组织结构"]:
            result = self.classifier.classify(text)
            self.assertEqual(result.intent, IntentType.GET_DEPT_TREE, f"Failed for: {text}")
    
    def test_reporting_chain(self):
        result = self.classifier.classify("张伟的汇报链")
        self.assertEqual(result.intent, IntentType.GET_REPORTING_CHAIN)
    
    def test_reporting_chain_who(self):
        result = self.classifier.classify("E003的上级是谁")
        self.assertEqual(result.intent, IntentType.GET_REPORTING_CHAIN)
    
    # --- 薪资 ---
    
    def test_calculate_payroll(self):
        result = self.classifier.classify("计算本月薪资")
        self.assertEqual(result.intent, IntentType.CALCULATE_PAYROLL)
    
    def test_calculate_payroll_variants(self):
        for text in ["算工资", "跑工资", "本月工资", "薪资计算"]:
            result = self.classifier.classify(text)
            self.assertEqual(result.intent, IntentType.CALCULATE_PAYROLL, f"Failed for: {text}")
    
    def test_calculate_bonus(self):
        result = self.classifier.classify("计算年终奖")
        self.assertEqual(result.intent, IntentType.CALCULATE_BONUS)
    
    def test_calculate_insurance(self):
        result = self.classifier.classify("计算社保")
        self.assertEqual(result.intent, IntentType.CALCULATE_INSURANCE)
    
    def test_calculate_insurance_city(self):
        result = self.classifier.classify("北京公积金多少")
        self.assertEqual(result.intent, IntentType.CALCULATE_INSURANCE)
    
    # --- 报表 ---
    
    def test_employee_statistics(self):
        result = self.classifier.classify("员工统计")
        self.assertEqual(result.intent, IntentType.EMPLOYEE_STATISTICS)
    
    def test_employee_count(self):
        result = self.classifier.classify("有多少员工")
        self.assertEqual(result.intent, IntentType.EMPLOYEE_STATISTICS)
    
    def test_data_validation(self):
        result = self.classifier.classify("校验花名册数据")
        self.assertEqual(result.intent, IntentType.DATA_VALIDATION)
    
    def test_export_report(self):
        result = self.classifier.classify("导出员工报表")
        self.assertEqual(result.intent, IntentType.EXPORT_REPORT)
    
    # --- 帮助 ---
    
    def test_help(self):
        result = self.classifier.classify("帮助")
        self.assertEqual(result.intent, IntentType.HELP)
    
    def test_help_variants(self):
        for text in ["你能做什么", "怎么用", "有哪些功能"]:
            result = self.classifier.classify(text)
            self.assertEqual(result.intent, IntentType.HELP, f"Failed for: {text}")
    
    # --- 未知意图 ---
    
    def test_unknown(self):
        result = self.classifier.classify("今天天气怎么样")
        self.assertEqual(result.intent, IntentType.UNKNOWN)
    
    def test_empty(self):
        result = self.classifier.classify("")
        self.assertEqual(result.intent, IntentType.UNKNOWN)


class TestParameterExtractor(unittest.TestCase):
    """参数提取器测试"""
    
    def setUp(self):
        self.extractor = ParameterExtractor()
    
    def test_extract_emp_no(self):
        text = "查看E001的信息"
        intent = IntentResult(intent=IntentType.READ_EMPLOYEE, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["empNo"], "E001")
    
    def test_extract_emp_no_with_leading_zero(self):
        text = "查看E003"
        intent = IntentResult(intent=IntentType.READ_EMPLOYEE, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["empNo"], "E003")
    
    def test_extract_name_from_action(self):
        text = "张伟转正"
        intent = IntentResult(intent=IntentType.UPDATE_EMPLOYEE, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["empName"], "张伟")
    
    def test_extract_update_transfer(self):
        text = "张伟转正"
        intent = IntentResult(intent=IntentType.UPDATE_EMPLOYEE, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["updateType"], "转正")
        self.assertEqual(params["newStatus"], "在职")
    
    def test_extract_update_salary(self):
        text = "调薪E001到20000元"
        intent = IntentResult(intent=IntentType.UPDATE_EMPLOYEE, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["empNo"], "E001")
        self.assertEqual(params["updateType"], "调薪")
        self.assertEqual(params["newSalary"], 20000)
    
    def test_extract_delete_with_name(self):
        text = "王芳离职"
        intent = IntentResult(intent=IntentType.DELETE_EMPLOYEE, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["empName"], "王芳")
    
    def test_extract_search_by_dept(self):
        text = "搜索技术部员工"
        intent = IntentResult(intent=IntentType.SEARCH_EMPLOYEE, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["deptName"], "技术部")
    
    def test_extract_search_by_status(self):
        text = "试用期员工有哪些"
        intent = IntentResult(intent=IntentType.SEARCH_EMPLOYEE, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["status"], "试用期")
    
    def test_extract_payroll_batch(self):
        text = "计算本月薪资，所有人"
        intent = IntentResult(intent=IntentType.CALCULATE_PAYROLL, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertTrue(params.get("batch"))
        self.assertIsNotNone(params.get("year"))
        self.assertIsNotNone(params.get("month"))
    
    def test_extract_payroll_single(self):
        text = "帮E001算工资"
        intent = IntentResult(intent=IntentType.CALCULATE_PAYROLL, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["empNo"], "E001")
    
    def test_extract_bonus(self):
        text = "年终奖36000"
        intent = IntentResult(intent=IntentType.CALCULATE_BONUS, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["bonusAmount"], 36000)
    
    def test_extract_bonus_wan(self):
        text = "年终奖3万"
        intent = IntentResult(intent=IntentType.CALCULATE_BONUS, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["bonusAmount"], 30000)
    
    def test_extract_insurance(self):
        text = "北京社保10000"
        intent = IntentResult(intent=IntentType.CALCULATE_INSURANCE, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["city"], "北京")
        self.assertEqual(params["salary"], 10000)
    
    def test_extract_city(self):
        text = "上海社保"
        intent = IntentResult(intent=IntentType.CALCULATE_INSURANCE, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["city"], "上海")
    
    def test_extract_batch_emp_nos(self):
        text = "批量转正E011,E012"
        intent = IntentResult(intent=IntentType.BATCH_UPDATE_STATUS, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["newStatus"], "在职")
        self.assertIn("E011", params["empNos"])
        self.assertIn("E012", params["empNos"])
    
    def test_extract_reporting_chain(self):
        text = "赵雪的汇报链"
        intent = IntentResult(intent=IntentType.GET_REPORTING_CHAIN, confidence=0.9, matchedPatterns=[])
        params = self.extractor.extract(text, intent)
        self.assertEqual(params["empName"], "赵雪")
    
    def test_check_missing_params(self):
        intent = IntentType.READ_EMPLOYEE
        params = {}
        missing = self.extractor.checkMissingParams(intent, params)
        self.assertIn("empNo", missing)
    
    def test_no_missing_params(self):
        intent = IntentType.READ_EMPLOYEE
        params = {"empNo": "E001"}
        missing = self.extractor.checkMissingParams(intent, params)
        self.assertEqual(missing, [])


class TestToolExecutor(unittest.TestCase):
    """Tool 执行器测试（需要测试数据）"""
    
    @classmethod
    def setUpClass(cls):
        """创建测试数据"""
        cls.test_dir = tempfile.mkdtemp()
        
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        # 创建组织架构
        org_wb = Workbook()
        org_ws = org_wb.active
        org_ws.title = "组织架构"
        org_headers = ["部门编码", "部门名称", "上级部门", "负责人", "层级"]
        for col, h in enumerate(org_headers, 1):
            org_ws.cell(1, col, h)
        
        org_data = [
            ["D001", "总公司", None, "E001", 1],
            ["D010", "技术中心", "D001", "E002", 2],
            ["D011", "前端开发部", "D010", "E003", 3],
            ["D012", "后端开发部", "D010", "E004", 3],
            ["D020", "人力资源部", "D001", "E005", 2],
        ]
        for row_idx, row_data in enumerate(org_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                org_ws.cell(row_idx, col_idx, value)
        
        cls.org_file = os.path.join(cls.test_dir, "组织架构.xlsx")
        org_wb.save(cls.org_file)
        
        # 创建员工花名册
        emp_wb = Workbook()
        emp_ws = emp_wb.active
        emp_ws.title = "员工花名册"
        emp_headers = ["工号", "姓名", "部门编码", "部门名称", "岗位", "在职状态", "入职日期",
                       "基本工资", "社保城市", "汇报对象", "手机号"]
        for col, h in enumerate(emp_headers, 1):
            emp_ws.cell(1, col, h)
        
        emp_data = [
            ["E001", "张伟", "D001", "总公司", "CEO", "在职", "2020-01-15", 50000, "北京", "", "13800000001"],
            ["E002", "李明", "D010", "技术中心", "CTO", "在职", "2020-03-01", 45000, "北京", "E001", "13800000002"],
            ["E003", "王芳", "D011", "前端开发部", "前端负责人", "在职", "2021-05-10", 35000, "北京", "E002", "13800000003"],
            ["E004", "刘强", "D012", "后端开发部", "后端负责人", "在职", "2021-06-15", 35000, "北京", "E002", "13800000004"],
            ["E005", "陈静", "D020", "人力资源部", "HR经理", "在职", "2021-01-20", 30000, "北京", "E001", "13800000005"],
            ["E006", "赵雪", "D011", "前端开发部", "前端工程师", "试用期", "2026-03-01", 25000, "北京", "E003", "13800000006"],
            ["E007", "孙磊", "D012", "后端开发部", "后端工程师", "在职", "2022-04-01", 28000, "北京", "E004", "13800000007"],
        ]
        for row_idx, row_data in enumerate(emp_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                emp_ws.cell(row_idx, col_idx, value)
        
        cls.emp_file = os.path.join(cls.test_dir, "员工花名册.xlsx")
        emp_wb.save(cls.emp_file)
    
    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.test_dir)
    
    def setUp(self):
        self.config = {
            "filePath": self.emp_file,
            "orgFilePath": self.org_file,
        }
        self.executor = ToolExecutor(self.config)
    
    def test_read_employee(self):
        result = self.executor.execute(IntentType.READ_EMPLOYEE, {"empNo": "E001"})
        self.assertTrue(result["success"])
        self.assertEqual(result["data"]["name"], "张伟")
        self.assertEqual(result["data"]["position"], "CEO")
    
    def test_read_employee_by_name(self):
        result = self.executor.execute(IntentType.READ_EMPLOYEE, {"empName": "李明"})
        self.assertTrue(result["success"])
        self.assertEqual(result["data"]["empNo"], "E002")
    
    def test_read_employee_not_found(self):
        result = self.executor.execute(IntentType.READ_EMPLOYEE, {"empNo": "E999"})
        self.assertFalse(result["success"])
    
    def test_list_all_employees(self):
        result = self.executor.execute(IntentType.LIST_EMPLOYEES, {})
        self.assertTrue(result["success"])
        self.assertEqual(result["count"], 7)  # E001-E007 (6在职 + 1试用期)
    
    def test_list_employees_by_dept(self):
        result = self.executor.execute(IntentType.SEARCH_EMPLOYEE, {"deptName": "前端开发部"})
        self.assertTrue(result["success"])
        self.assertGreaterEqual(result["count"], 2)
    
    def test_list_employees_by_status(self):
        result = self.executor.execute(IntentType.SEARCH_EMPLOYEE, {"status": "试用期"})
        self.assertTrue(result["success"])
        self.assertEqual(result["count"], 1)
    
    def test_search_by_keyword(self):
        result = self.executor.execute(IntentType.SEARCH_EMPLOYEE, {"keyword": "CEO"})
        self.assertTrue(result["success"])
        self.assertGreaterEqual(result["count"], 1)
    
    def test_dept_tree(self):
        result = self.executor.execute(IntentType.GET_DEPT_TREE, {})
        self.assertTrue(result["success"])
        self.assertEqual(result["totalDepts"], 5)
        self.assertEqual(result["totalEmployees"], 7)
    
    def test_reporting_chain(self):
        result = self.executor.execute(IntentType.GET_REPORTING_CHAIN, {"empNo": "E006"})
        self.assertTrue(result["success"])
        self.assertGreaterEqual(result["chainLength"], 3)
        # 赵雪 → 王芳 → 李明 → 张伟
        names = [c["name"] for c in result["data"]]
        self.assertEqual(names[0], "赵雪")
        self.assertEqual(names[1], "王芳")
    
    def test_calculate_payroll_single(self):
        result = self.executor.execute(IntentType.CALCULATE_PAYROLL, {
            "empNo": "E001",
            "year": 2026,
            "month": 4
        })
        self.assertTrue(result["success"])
        self.assertIn("实发", result["message"])
        data = result["data"]
        self.assertGreater(data["grossPay"], 0)
        self.assertGreater(data["totalInsurance"], 0)
        self.assertGreater(data["netPay"], 0)
    
    def test_calculate_payroll_batch(self):
        result = self.executor.execute(IntentType.CALCULATE_PAYROLL, {
            "batch": True,
            "year": 2026,
            "month": 4
        })
        self.assertTrue(result["success"])
        self.assertEqual(result["count"], 7)  # 7在职/试用期
        self.assertGreater(result["totalNetPay"], 0)
    
    def test_calculate_bonus(self):
        result = self.executor.execute(IntentType.CALCULATE_BONUS, {"bonusAmount": 36000})
        self.assertTrue(result["success"])
        self.assertIn("税后", result["message"])
        data = result["data"]
        self.assertEqual(data["bonusAmount"], 36000)
        self.assertGreater(data["netBonus"], 0)
    
    def test_calculate_insurance(self):
        result = self.executor.execute(IntentType.CALCULATE_INSURANCE, {
            "salary": 10000,
            "city": "北京"
        })
        self.assertTrue(result["success"])
        self.assertIn("养老保险", result["message"])
        data = result["data"]
        self.assertGreater(data["totalInsurance"], 0)
    
    def test_employee_statistics(self):
        result = self.executor.execute(IntentType.EMPLOYEE_STATISTICS, {})
        self.assertTrue(result["success"])
        self.assertIn("总人数", result["message"])
        self.assertIn("在职", result["message"])
    
    def test_data_validation(self):
        result = self.executor.execute(IntentType.DATA_VALIDATION, {})
        self.assertTrue(result["success"])
        self.assertIn("校验结果", result["message"])
    
    def test_export_report(self):
        # 导出报表（创建副本避免修改原始文件）
        import copy
        executor = ToolExecutor(self.config)
        result = executor.execute(IntentType.EXPORT_REPORT, {"reportType": "employee"})
        self.assertTrue(result["success"])
        self.assertIn("导出", result["message"])
        if os.path.exists(result.get("filePath", "")):
            os.remove(result["filePath"])
    
    def test_help(self):
        result = self.executor.execute(IntentType.HELP, {})
        self.assertTrue(result["success"])
        self.assertIn("使用指南", result["message"])
    
    def test_unknown_intent(self):
        result = self.executor.execute(IntentType.UNKNOWN, {})
        self.assertFalse(result["success"])
    
    def test_batch_transfer(self):
        # 直接调用底层方法，避免 save 修改共享测试文件
        mgr = self.executor.employeeManager
        result = mgr.batchUpdateStatus(["E006"], "在职")
        self.assertEqual(result["successCount"], 1)
    
    def test_delete_employee_soft(self):
        # 直接调用底层方法，避免 save 修改共享测试文件
        mgr = self.executor.employeeManager
        ok, msg = mgr.deleteEmployee("E006", soft=True)
        self.assertTrue(ok)
        self.assertIn("离职", msg)
        # 验证状态已变
        self.assertEqual(mgr.getEmployee("E006").status, "离职")


class TestConversationContext(unittest.TestCase):
    """对话上下文测试"""
    
    def test_add_turn(self):
        ctx = ConversationContext(sessionId="test-1")
        ctx.addTurn("user", "查看E001")
        self.assertEqual(len(ctx.turns), 1)
        self.assertEqual(ctx.turns[0]["role"], "user")
    
    def test_turn_limit(self):
        ctx = ConversationContext(sessionId="test-2")
        for i in range(25):
            ctx.addTurn("user", f"消息{i}")
        self.assertLessEqual(len(ctx.turns), 20)
    
    def test_context_continuation(self):
        """测试上下文延续识别"""
        classifier = IntentClassifier()
        ctx = ConversationContext(sessionId="test-3")
        ctx.lastIntent = IntentType.READ_EMPLOYEE
        ctx.pendingAction = {
            "intent": IntentType.READ_EMPLOYEE,
            "params": {},
            "missingParams": ["empNo"]
        }
        
        # 用户回答工号
        result = classifier.classify("E001", ctx)
        self.assertEqual(result.intent, IntentType.READ_EMPLOYEE)


class TestIntentRouterE2E(unittest.TestCase):
    """端到端意图路由测试（模拟真实对话场景）"""
    
    @classmethod
    def setUpClass(cls):
        cls.test_dir = tempfile.mkdtemp()
        
        from openpyxl import Workbook
        
        # 创建组织架构
        org_wb = Workbook()
        org_ws = org_wb.active
        org_ws.title = "组织架构"
        org_headers = ["部门编码", "部门名称", "上级部门", "负责人", "层级"]
        for col, h in enumerate(org_headers, 1):
            org_ws.cell(1, col, h)
        org_data = [
            ["D001", "总公司", None, "E001", 1],
            ["D010", "技术中心", "D001", "E002", 2],
            ["D011", "前端开发部", "D010", "E003", 3],
            ["D012", "后端开发部", "D010", "E004", 3],
            ["D020", "人力资源部", "D001", "E005", 2],
        ]
        for row_idx, row_data in enumerate(org_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                org_ws.cell(row_idx, col_idx, value)
        cls.org_file = os.path.join(cls.test_dir, "组织架构.xlsx")
        org_wb.save(cls.org_file)
        
        # 创建花名册
        emp_wb = Workbook()
        emp_ws = emp_wb.active
        emp_ws.title = "员工花名册"
        emp_headers = ["工号", "姓名", "部门编码", "部门名称", "岗位", "在职状态",
                       "入职日期", "基本工资", "社保城市", "汇报对象"]
        for col, h in enumerate(emp_headers, 1):
            emp_ws.cell(1, col, h)
        emp_data = [
            ["E001", "张伟", "D001", "总公司", "CEO", "在职", "2020-01-15", 50000, "北京", ""],
            ["E002", "李明", "D010", "技术中心", "CTO", "在职", "2020-03-01", 45000, "北京", "E001"],
            ["E003", "王芳", "D011", "前端开发部", "前端负责人", "在职", "2021-05-10", 35000, "北京", "E002"],
            ["E004", "刘强", "D012", "后端开发部", "后端负责人", "在职", "2021-06-15", 35000, "北京", "E002"],
            ["E005", "陈静", "D020", "人力资源部", "HR经理", "在职", "2021-01-20", 30000, "北京", "E001"],
            ["E006", "赵雪", "D011", "前端开发部", "前端工程师", "试用期", "2026-03-01", 25000, "北京", "E003"],
        ]
        for row_idx, row_data in enumerate(emp_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                emp_ws.cell(row_idx, col_idx, value)
        cls.emp_file = os.path.join(cls.test_dir, "员工花名册.xlsx")
        emp_wb.save(cls.emp_file)
    
    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.test_dir)
    
    def _makeRouter(self):
        config = {"filePath": self.emp_file, "orgFilePath": self.org_file}
        return IntentRouter(config)
    
    def test_e2e_query_employee(self):
        """场景1：查询员工信息"""
        router = self._makeRouter()
        result = router.process("查看E001的信息")
        self.assertTrue(result["success"])
        self.assertIn("张伟", result["message"])
    
    def test_e2e_query_employee_by_name(self):
        """场景2：按姓名查询"""
        router = self._makeRouter()
        result = router.process("王芳的信息")
        self.assertTrue(result["success"])
        self.assertIn("王芳", result["message"])
    
    def test_e2e_list_all(self):
        """场景3：列出所有员工"""
        router = self._makeRouter()
        result = router.process("列出所有员工")
        self.assertTrue(result["success"])
        self.assertIn("6", result["message"])  # 6人（不含离职）
    
    def test_e2e_search_by_dept(self):
        """场景4：按部门搜索"""
        router = self._makeRouter()
        result = router.process("前端开发部员工")
        self.assertTrue(result["success"])
    
    def test_e2e_search_probation(self):
        """场景5：试用期员工"""
        router = self._makeRouter()
        result = router.process("试用期员工有哪些")
        self.assertTrue(result["success"])
    
    def test_e2e_org_tree(self):
        """场景6：查看组织架构"""
        router = self._makeRouter()
        result = router.process("组织架构")
        self.assertTrue(result["success"])
        self.assertIn("5", result["message"])  # 5个部门
    
    def test_e2e_reporting_chain(self):
        """场景7：汇报链"""
        router = self._makeRouter()
        result = router.process("赵雪的汇报链")
        self.assertTrue(result["success"])
        self.assertIn("赵雪", result["message"])
    
    def test_e2e_payroll_single(self):
        """场景8：单人薪资计算"""
        router = self._makeRouter()
        result = router.process("帮E001算工资")
        self.assertTrue(result["success"])
        self.assertIn("实发", result["message"])
        self.assertIn("50,000", result["message"])  # 格式化后带千分位
    
    def test_e2e_payroll_batch(self):
        """场景9：批量薪资计算"""
        router = self._makeRouter()
        result = router.process("计算本月所有人薪资")
        self.assertTrue(result["success"])
        self.assertIn("6", result["message"])
    
    def test_e2e_bonus(self):
        """场景10：年终奖"""
        router = self._makeRouter()
        result = router.process("年终奖36000")
        self.assertTrue(result["success"])
        self.assertIn("税后", result["message"])
    
    def test_e2e_insurance(self):
        """场景11：社保计算"""
        router = self._makeRouter()
        result = router.process("北京社保10000")
        self.assertTrue(result["success"])
        self.assertIn("养老保险", result["message"])
    
    def test_e2e_statistics(self):
        """场景12：员工统计"""
        router = self._makeRouter()
        result = router.process("员工统计")
        self.assertTrue(result["success"])
        self.assertIn("总人数", result["message"])
        self.assertIn("6", result["message"])
    
    def test_e2e_data_validation(self):
        """场景13：数据校验"""
        router = self._makeRouter()
        result = router.process("校验数据")
        self.assertTrue(result["success"])
        self.assertIn("校验结果", result["message"])
    
    def test_e2e_help(self):
        """场景14：帮助"""
        router = self._makeRouter()
        result = router.process("帮助")
        self.assertTrue(result["success"])
        self.assertIn("使用指南", result["message"])
    
    def test_e2e_multi_turn_context(self):
        """场景15：多轮对话 - 追问后回答"""
        router = self._makeRouter()
        
        # 第一轮：触发需要参数的操作
        result1 = router.process("年终奖")
        # 年终奖缺少 bonusAmount，应该需要追问
        self.assertIn(result1["intent"], ["calculate_bonus", "calculate_bonus"])
        
        # 第二轮：回答追问
        result2 = router.process("36000")
        self.assertTrue(result2["success"])
        self.assertIn("税后", result2["message"])
    
    def test_e2e_unknown_fallback(self):
        """场景16：未知意图 -> 给出建议"""
        router = self._makeRouter()
        result = router.process("今天吃什么")
        self.assertFalse(result["success"])
        self.assertIn("帮助", result["message"])
    
    def test_e2e_transfer_flow(self):
        """场景17：转正操作 - 验证意图识别和参数提取"""
        router = self._makeRouter()
        # 只验证意图识别和参数提取，不实际执行（避免修改共享文件）
        intent = router.classifier.classify("赵雪转正")
        self.assertEqual(intent.intent, IntentType.UPDATE_EMPLOYEE)
        params = router.extractor.extract("赵雪转正", intent)
        self.assertEqual(params["empName"], "赵雪")
        self.assertEqual(params["newStatus"], "在职")
    
    def test_e2e_export(self):
        """场景18：导出报表 - 验证意图识别"""
        router = self._makeRouter()
        intent = router.classifier.classify("导出员工报表")
        self.assertEqual(intent.intent, IntentType.EXPORT_REPORT)
        params = router.extractor.extract("导出员工报表", intent)
        self.assertEqual(params["reportType"], "employee")


class TestResponseFormatter(unittest.TestCase):
    """响应格式化测试"""
    
    def test_success_format(self):
        result = {"success": True, "message": "操作成功"}
        formatted = ResponseFormatter.format(result, IntentResult(intent=IntentType.HELP, confidence=0.9, matchedPatterns=[]))
        self.assertEqual(formatted, "操作成功")
    
    def test_failure_format(self):
        result = {"success": False, "message": "操作失败"}
        formatted = ResponseFormatter.format(result, IntentResult(intent=IntentType.HELP, confidence=0.9, matchedPatterns=[]))
        self.assertEqual(formatted, "操作失败")
    
    def test_ask_for_params(self):
        prompt = ResponseFormatter.askForParams(["empNo", "city"], IntentType.READ_EMPLOYEE)
        self.assertIn("工号", prompt)
        self.assertIn("城市", prompt)


if __name__ == "__main__":
    unittest.main(verbosity=2)
