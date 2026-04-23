import os
import sys
import subprocess
import time

# 🔥 移除 Ollama 相关配置（交给 OpenClaw 处理）

def print_step(step_name):
    """打印当前步骤（中英双语）"""
    # 中文在前，英文在后，兼顾中外用户
    step_mapping = {
        "读取 TXT 文件": "Reading TXT file",
        "读取 Word 文档": "Reading Word document",
        "读取 PDF 文件": "Reading PDF file",
        "读取 Excel 文件": "Reading Excel file",
        "正在自动安装依赖库": "Automatically installing dependency library"
    }
    en_step = step_mapping.get(step_name, step_name)
    print(f"\n[Step] {step_name} / {en_step} ...")
    time.sleep(0.3)

def install_package(package_name):
    """自动安装Python包（中英双语提示）"""
    print_step(f"正在自动安装依赖库: {package_name}")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name], 
                              stdout=subprocess.DEVNULL, 
                              stderr=subprocess.DEVNULL)
        print(f"✅ 成功安装 {package_name} / Successfully installed {package_name}")
        return True
    except subprocess.CalledProcessError:
        err_msg = f"❌ 安装 {package_name} 失败，请手动运行: pip install {package_name} / Failed to install {package_name}, please run manually: pip install {package_name}"
        print(err_msg)
        return False

def read_txt(path):
    print_step("读取 TXT 文件")
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except:
        with open(path, "r", encoding="gbk") as f:
            return f.read()

def read_docx(path):
    print_step("读取 Word 文档")
    try:
        from docx import Document
        doc = Document(path)
        return "\n".join([p.text for p in doc.paragraphs])
    except ImportError:
        if install_package("python-docx"):
            return read_docx(path)  # 安装成功后重试
        else:
            err_msg = "❌ 读取Word失败：缺少 python-docx 库，且自动安装失败。 / Failed to read Word: Missing python-docx library, auto-install failed."
            return err_msg

def read_pdf(path):
    print_step("读取 PDF 文件")
    try:
        from pypdf import PdfReader
        reader = PdfReader(path)
        return "\n".join([page.extract_text() for page in reader.pages])
    except ImportError:
        if install_package("pypdf"):
            return read_pdf(path)  # 安装成功后重试
        else:
            err_msg = "❌ 读取PDF失败：缺少 pypdf 库，且自动安装失败。 / Failed to read PDF: Missing pypdf library, auto-install failed."
            return err_msg

def read_excel(path):
    """读取 Excel 文件 (.xlsx/.xls)，无需安装Microsoft Excel"""
    print_step("读取 Excel 文件")
    try:
        ext = path.lower().split(".")[-1]
        if ext == "xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            text = ""
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"=== 工作表: {sheet_name} / Worksheet: {sheet_name} ===\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = [str(cell) for cell in row if cell is not None]
                    if row_text:
                        text += "\t".join(row_text) + "\n"
            return text
        elif ext == "xls":
            import xlrd
            wb = xlrd.open_workbook(path)
            text = ""
            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                text += f"=== 工作表: {sheet.name} / Worksheet: {sheet.name} ===\n"
                for row_num in range(sheet.nrows):
                    row_text = []
                    for col_num in range(sheet.ncols):
                        val = sheet.cell_value(row_num, col_num)
                        if val is not None or val == 0:
                            row_text.append(str(val))
                    if row_text:
                        text += "\t".join(row_text) + "\n"
            return text
        else:
            err_msg = f"不支持的Excel格式：.{ext} / Unsupported Excel format: .{ext}"
            return err_msg
    except ImportError as e:
        lib_name = "openpyxl" if "openpyxl" in str(e) else "xlrd==1.2.0"
        if install_package(lib_name):
            return read_excel(path)  # 安装成功后重试
        else:
            err_msg = f"❌ 读取Excel失败：缺少 {lib_name} 库，且自动安装失败。 / Failed to read Excel: Missing {lib_name} library, auto-install failed."
            return err_msg
    except Exception as e:
        err_msg = f"❌ 读取Excel失败：{str(e)} / Failed to read Excel: {str(e)}"
        return err_msg

def read_file(path):
    ext = path.lower().split(".")[-1]
    if ext == "txt":
        return read_txt(path)
    elif ext == "docx":
        return read_docx(path)
    elif ext == "pdf":
        return read_pdf(path)
    elif ext in ["xlsx", "xls"]:
        return read_excel(path)
    else:
        supported = ["txt", "docx", "pdf", "xlsx", "xls"]
        err_msg = f"❌ 不支持该文件格式（支持：{', '.join(supported)}） / Unsupported file format (supported: {', '.join(supported)})"
        return err_msg

def main():
    if len(sys.argv) < 2:
        # 🔥 适配 OpenClaw：返回标准化提示（中英双语）
        err_msg = "❌ 使用方法：请传入需要读取的文件完整路径作为参数 / Usage: Please pass the full path of the file to read as a parameter"
        print(err_msg)
        return

    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        err_msg = f"❌ 文件不存在：{file_path} / File does not exist: {file_path}"
        print(err_msg)
        return

    # 读取文件内容
    content = read_file(file_path)
    
    # 🔥 核心适配：只输出纯内容（错误/正常），供 OpenClaw 捕获
    # 错误内容以 ❌ 开头，正常内容直接输出
    print(content)

if __name__ == "__main__":
    main()