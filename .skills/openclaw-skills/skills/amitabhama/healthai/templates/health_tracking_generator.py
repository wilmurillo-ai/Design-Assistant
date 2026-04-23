#!/usr/bin/env python3
"""
健康管理跟踪表生成脚本
自动生成12周健康管理打卡Excel表格

使用方法:
    python health_tracking_generator.py --output 输出路径.xlsx

依赖:
    pip install openpyxl
"""

import argparse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_workbook():
    """创建健康管理跟踪工作簿"""
    wb = Workbook()
    return wb


def style_header(cell, fill_color="4472C4"):
    """设置表头样式"""
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def style_data_cell(cell):
    """设置数据单元格样式"""
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def create_checkin_sheet(wb, user_info):
    """创建每周打卡记录表"""
    ws = wb.active
    ws.title = "每周打卡记录"

    # 用户信息
    ws.merge_cells('A1:H1')
    ws['A1'] = f"健康管理打卡记录 - 第{user_info.get('name', '用户')}（生成日期：{datetime.now().strftime('%Y-%m-%d')}）"
    ws['A1'].font = Font(bold=True, size=14)

    # 表头
    headers = ["周次", "日期", "体重(kg)", "腰围(cm)", "饮食打卡", "运动时长(min)", "饮酒情况", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        style_header(cell)

    # 数据行（12周）
    start_date = datetime.now()
    for week in range(1, 13):
        row = week + 3
        # 周次
        ws.cell(row=row, column=1, value=f"第{week}周")
        style_data_cell(ws.cell(row=row, column=1))
        # 日期（周一）
        week_date = start_date + timedelta(weeks=week-1)
        ws.cell(row=row, column=2, value=week_date.strftime("%Y-%m-%d"))
        style_data_cell(ws.cell(row=row, column=2))
        # 空数据列
        for col in range(3, 9):
            ws.cell(row=row, column=col, value="")
            style_data_cell(ws.cell(row=row, column=col))

    # 设置列宽
    column_widths = [8, 12, 10, 10, 20, 12, 10, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return ws


def create_diet_sheet(wb, diet_type="蛋奶素"):
    """创建饮食方案表"""
    ws = wb.create_sheet("饮食方案")

    # 标题
    ws.merge_cells('A1:E1')
    ws['A1'] = f"每日饮食方案（{diet_type}版本）"
    ws['A1'].font = Font(bold=True, size=14)

    # 表头
    headers = ["餐次", "时间", "推荐食物", "份量", "作用"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        style_header(cell)

    # 早餐数据
    breakfast_items = [
        ["早餐", "7:30-8:30", "鸡蛋+燕麦+牛奶+水果", "蛋白质+粗粮+钙", "补充蛋白质和碳水"],
    ]
    if diet_type == "杂食":
        breakfast_items[0][2] = "鸡蛋+全麦面包+牛奶+水果"

    for row_idx, item in enumerate(breakfast_items, 4):
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_data_cell(cell)

    # 午餐数据
    lunch_items = [
        ["午餐", "12:00-13:00", "杂粮饭+豆腐+菌菇+深绿蔬菜", "一拳蛋白+两捧蔬菜+半碗饭", "主食+蛋白+纤维"],
    ]
    if diet_type == "杂食":
        lunch_items[0][2] = "米饭+鱼/鸡胸肉+蔬菜+豆腐"

    for row_idx, item in enumerate(lunch_items, 5):
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_data_cell(cell)

    # 晚餐数据
    dinner_items = [
        ["晚餐", "18:30-19:30", "主食减半+豆制品+凉拌菜", "碳水减量，多蔬菜", "控制热量摄入"],
    ]
    if diet_type == "杂食":
        dinner_items[0][2] = "主食减半+鱼/鸡/豆制品+蔬菜汤"

    for row_idx, item in enumerate(dinner_items, 6):
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_data_cell(cell)

    # 加餐数据
    snack_items = [
        ["加餐(10:00/15:30)", "上午/下午", "坚果+水果", "坚果15g/水果1个", "健康零食，控制热量"],
    ]
    for row_idx, item in enumerate(snack_items, 7):
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_data_cell(cell)

    # 禁忌清单
    ws.merge_cells('A9:E9')
    ws['A9'] = "🚫 严格避免食物"
    ws['A9'].font = Font(bold=True, size=12)

    caution_items = [
        ["酒类", "直接伤肝，升高ALT"],
        ["动物内脏", "高胆固醇"],
        ["油炸食品", "高脂肪，易形成脂肪肝"],
        ["甜饮料/奶茶", "高糖"],
        ["夜宵(22点后)", "影响代谢"],
    ]

    for row_idx, item in enumerate(caution_items, 10):
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_data_cell(cell)

    # 设置列宽
    column_widths = [15, 12, 30, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return ws


def create_exercise_sheet(wb, exercise_time="中午"):
    """创建运动方案表"""
    ws = wb.create_sheet("运动方案")

    # 标题
    ws.merge_cells('A1:D1')
    ws['A1'] = f"每周运动计划（运动时间：{exercise_time}）"
    ws['A1'].font = Font(bold=True, size=14)

    # 说明
    ws.merge_cells('A2:D2')
    ws['A2'] = "目标：每周至少5次运动，每次30-60分钟，有氧为主+力量为辅"
    ws['A2'].font = Font(size=10, italic=True)

    # 表头
    headers = ["星期", "运动类型", "时长", "具体内容"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        style_header(cell)

    # 周运动计划
    workout_plan = [
        ["周一", "有氧", "45min", "跑步机/椭圆机/划船机，心率130左右"],
        ["周二", "力量", "50min", "胸+背（哑铃卧推/划船）"],
        ["周三", "有氧", "45min", "快走/慢跑/游泳"],
        ["周四", "力量", "50min", "腿+核心（深蹲/平板支撑）"],
        ["周五", "有氧", "45min", "骑行/椭圆机/跳绳"],
        ["周六", "综合", "60min", "球类运动（羽毛球/篮球）或爬山"],
        ["周日", "休息/拉伸", "30min", "瑜伽/泡沫轴拉伸"],
    ]

    for row_idx, item in enumerate(workout_plan, 5):
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_data_cell(cell)

    # 每日提醒
    ws.merge_cells('A13:D13')
    ws['A13'] = "💪 每日微运动提醒"
    ws['A13'].font = Font(bold=True, size=12)

    daily_reminders = [
        ["起床后", "靠墙俯卧撑", "10个 × 3组"],
        ["睡前", "平板支撑", "30秒 × 3组"],
        ["随时", "深蹲", "15个 × 2组"],
    ]

    for row_idx, item in enumerate(daily_reminders, 14):
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_data_cell(cell)

    # 心率说明
    ws.merge_cells('A18:D18')
    ws['A18'] = "📊 有效减脂心率区间计算：(220 - 年龄) × 60%~70%"
    ws['A18'].font = Font(size=10, italic=True)

    # 设置列宽
    column_widths = [10, 12, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return ws


def create_checkup_sheet(wb):
    """创建复查计划表"""
    ws = wb.create_sheet("复查计划")

    # 标题
    ws.merge_cells('A1:D1')
    ws['A1'] = "健康复查计划"
    ws['A1'].font = Font(bold=True, size=14)

    # 表头
    headers = ["检查项目", "建议时间", "目的", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        style_header(cell)

    # 复查项目
    checkup_items = [
        ["肝功能（ALT/AST）", "3个月后", "验证护肝效果", "空腹抽血"],
        ["血脂+血糖", "3-6个月后", "代谢评估", "空腹抽血"],
        ["肝脏弹性检测", "6个月后（必要时）", "排查脂肪肝程度", "B超或FibroScan"],
        ["全面体检", "12个月后", "年度复查", "全面评估"],
        ["叶酸+B12", "4周后（必要时）", "排查贫血原因", "如MCV偏高"],
    ]

    for row_idx, item in enumerate(checkup_items, 4):
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_data_cell(cell)

    # 营养补充建议
    ws.merge_cells('A10:D10')
    ws['A10'] = "💊 营养补充建议"
    ws['A10'].font = Font(bold=True, size=12)

    supplement_headers = ["补充剂", "剂量", "作用", "必要性"]
    for col, header in enumerate(supplement_headers, 1):
        cell = ws.cell(row=11, column=col)
        cell.value = header
        style_header(cell)

    supplements = [
        ["复合维生素B", "每日1片", "支持肝脏代谢", "建议"],
        ["深海鱼油（Omega-3）", "1000mg/日", "抗炎、护肝", "建议"],
        ["奶蓟草", "250mg/日", "护肝抗氧化", "可选"],
        ["维生素D", "1000IU/日", "骨骼+免疫", "建议（尤其冬季）"],
    ]

    for row_idx, item in enumerate(supplements, 12):
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_data_cell(cell)

    # 设置列宽
    column_widths = [20, 15, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return ws


def create_tracking_sheet(wb, goals):
    """创建目标追踪表"""
    ws = wb.create_sheet("目标追踪")

    # 标题
    ws.merge_cells('A1:F1')
    ws['A1'] = "健康目标追踪进度"
    ws['A1'].font = Font(bold=True, size=14)

    # 表头
    headers = ["目标项目", "起始值", "当前值", "目标值", "达成进度", "状态"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        style_header(cell)

    # 目标数据（示例）
    default_goals = [
        ["体重(kg)", "71", "", "65", "0%", "进行中"],
        ["BMI", "24.8", "", "<23", "0%", "进行中"],
        ["腰围(cm)", "", "", "<85", "0%", "进行中"],
        ["ALT(U/L)", "50.65", "", "<40", "0%", "进行中"],
        ["每周运动(分钟)", "0", "", "≥150", "0%", "进行中"],
    ]

    for row_idx, item in enumerate(default_goals, 4):
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            style_data_cell(cell)

    # 成功关键
    ws.merge_cells('A10:F10')
    ws['A10'] = "🔑 成功关键"
    ws['A10'].font = Font(bold=True, size=12)

    success_keys = [
        "1. 戒酒是第一位 — 不戒酒，指标很难降下来",
        "2. 坚持比强度重要 — 每天30分钟比一周一次2小时更有效",
        "3. 体重每周测一次 — 不要每天称（波动大易焦虑）",
        "4. 给自己3个月 — 习惯养成需要时间，不要急于看结果",
    ]

    for row_idx, key in enumerate(success_keys, 11):
        ws.merge_cells(f'A{row_idx}:F{row_idx}')
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=1).font = Font(size=10)

    # 设置列宽
    column_widths = [15, 10, 10, 10, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return ws


def generate_health_tracking_xlsx(output_path, user_info=None, diet_type="蛋奶素", exercise_time="中午"):
    """
    生成健康管理跟踪表Excel

    Args:
        output_path: 输出文件路径
        user_info: 用户信息字典 {'name': '姓名', 'age': 年龄, ...}
        diet_type: 饮食习惯 ('杂食', '蛋奶素', '全素')
        exercise_time: 运动偏好时间 ('早晨', '中午', '晚上')
    """
    if user_info is None:
        user_info = {}

    wb = create_workbook()

    # 创建各Sheet
    create_checkin_sheet(wb, user_info)
    create_diet_sheet(wb, diet_type)
    create_exercise_sheet(wb, exercise_time)
    create_checkup_sheet(wb)
    create_tracking_sheet(wb, user_info.get('goals', []))

    # 保存文件
    wb.save(output_path)
    print(f"✅ 健康管理跟踪表已生成：{output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="生成健康管理跟踪表")
    parser.add_argument("--output", "-o", default="健康管理跟踪表.xlsx", help="输出文件路径")
    parser.add_argument("--name", "-n", default="用户", help="用户名称")
    parser.add_argument("--diet", "-d", default="蛋奶素", choices=["杂食", "蛋奶素", "全素"], help="饮食习惯")
    parser.add_argument("--exercise", "-e", default="中午", choices=["早晨", "中午", "晚上"], help="运动偏好时间")

    args = parser.parse_args()

    user_info = {
        "name": args.name,
        "diet_type": args.diet,
        "exercise_time": args.exercise,
    }

    generate_health_tracking_xlsx(args.output, user_info, args.diet, args.exercise)
