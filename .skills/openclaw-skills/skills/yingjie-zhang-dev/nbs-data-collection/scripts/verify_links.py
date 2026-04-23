#!/usr/bin/env python3
"""
验证Excel中的数据来源链接有效性
Verify Data Source Links in Excel

用法:
    python3 verify_links.py excel_file.xlsx
    python3 verify_links.py "国民经济核算与价格指数_2003-2026.xlsx"
"""

import sys
import csv
import subprocess
from pathlib import Path
from typing import Dict, List, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要openpyxl库: pip install openpyxl")
    sys.exit(1)


def check_url_http(url: str, timeout: int = 8) -> Tuple[bool, str]:
    """检查URL是否返回200"""
    try:
        result = subprocess.run(
            ["curl", "-s", "-o", "/dev/null", "-w", "%{http_code}", 
             url, "--connect-timeout", str(timeout)],
            capture_output=True,
            text=True,
            timeout=timeout + 2
        )
        code = result.stdout.strip()
        return (code == "200", code)
    except Exception as e:
        return (False, str(e)[:50])


def extract_urls_from_excel(excel_path: str) -> Dict[str, List[Dict]]:
    """从Excel中提取所有链接"""
    wb = load_workbook(excel_path)
    all_links = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        links = []
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    links.append({
                        "cell": cell.coordinate,
                        "text": cell.value,
                        "url": cell.hyperlink.target,
                        "sheet": sheet_name
                    })
        
        if links:
            all_links[sheet_name] = links
    
    return all_links


def verify_all_links(excel_path: str, verbose: bool = True) -> Dict:
    """验证所有链接"""
    urls_by_sheet = extract_urls_from_excel(excel_path)
    
    summary = {
        "total": 0,
        "valid": 0,
        "invalid": 0,
        "errors": []
    }
    
    all_valid = True
    
    for sheet_name, links in urls_by_sheet.items():
        if verbose:
            print(f"\n=== {sheet_name} ===")
        
        for link in links:
            summary["total"] += 1
            url = link["url"]
            cell_ref = f"{sheet_name}!{link['cell']}"
            
            is_valid, code = check_url_http(url)
            
            if is_valid:
                summary["valid"] += 1
                if verbose:
                    print(f"  ✅ {link['cell']}: {link['text']}")
            else:
                summary["invalid"] += 1
                summary["errors"].append({
                    "cell": cell_ref,
                    "url": url,
                    "code": code
                })
                if verbose:
                    print(f"  ❌ {link['cell']}: HTTP {code}")
    
    return summary


def fix_invalid_links(excel_path: str, errors: List[Dict]):
    """修复无效链接（2004年鉴格式修复示例）"""
    from openpyxl.styles import Font
    
    # 2004年鉴特殊格式修复
    fix_map = {
        "https://www.stats.gov.cn/sj/ndsj/2004/indexch.htm": 
            "https://www.stats.gov.cn/sj/ndsj/yb2004-c/indexch.htm",
    }
    
    wb = load_workbook(excel_path)
    link_font = Font(color="0563C1", underline="single", size=10)
    
    fixed_count = 0
    
    for error in errors:
        url = error["url"]
        if url in fix_map:
            new_url = fix_map[url]
            cell_ref = error["cell"]
            sheet_name, cell_coord = cell_ref.split("!", 1)
            
            ws = wb[sheet_name]
            cell = ws[cell_coord]
            cell.hyperlink = new_url
            
            fixed_count += 1
            print(f"  🔧 修复: {cell_ref} -> {new_url}")
    
    if fixed_count > 0:
        wb.save(excel_path)
        print(f"\n已修复 {fixed_count} 个链接")
    
    return fixed_count


def main():
    if len(sys.argv) < 2:
        print("用法: python3 verify_links.py <excel_file> [--fix]")
        print("示例: python3 verify_links.py '国民经济核算与价格指数_2003-2026.xlsx'")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    do_fix = "--fix" in sys.argv
    
    if not Path(excel_path).exists():
        print(f"文件不存在: {excel_path}")
        sys.exit(1)
    
    print(f"检查文件: {excel_path}")
    
    # 验证所有链接
    summary = verify_all_links(excel_path)
    
    # 输出总结
    print(f"\n=== 总结 ===")
    print(f"总链接数: {summary['total']}")
    print(f"有效: {summary['valid']}")
    print(f"无效: {summary['invalid']}")
    
    if summary['invalid'] > 0:
        print(f"\n无效链接:")
        for err in summary['errors']:
            print(f"  - {err['cell']}: {err['url']} (HTTP {err['code']})")
        
        if do_fix:
            print("\n尝试修复...")
            fix_invalid_links(excel_path, summary['errors'])
    
    return 0 if summary['invalid'] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
