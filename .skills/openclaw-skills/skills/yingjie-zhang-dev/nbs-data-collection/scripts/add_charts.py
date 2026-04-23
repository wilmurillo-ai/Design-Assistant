#!/usr/bin/env python3
"""
为Excel添加折线图
Add Line Charts to Excel

用法:
    python3 add_charts.py excel_file.xlsx
    python3 add_charts.py "国民经济核算与价格指数_2003-2026.xlsx" --dry-run
"""

import sys
from pathlib import Path
from typing import Dict, List, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("需要openpyxl库: pip install openpyxl")
    sys.exit(1)


# 图表配置
# 格式: (sheet_name, data_column, chart_title, y_axis_title, chart_height)
CHART_CONFIGS = [
    ("实际GDP (季度)", 4, "实际GDP季度走势", "GDP(亿元)", 10),
    ("GDP增速", 2, "GDP同比增速走势", "增速(%)", 10),
    ("产出缺口", 4, "产出缺口走势", "产出缺口(亿元)", 10),
    ("产出缺口", 5, "产出缺口率走势", "产出缺口率(%)", 10),
    ("人均GDP", 4, "人均GDP走势", "人均GDP(元/季度)", 10),
    ("CPI (季度均值)", 2, "CPI季度均值走势", "CPI(上年同月=100)", 10),
    ("PPI (季度均值)", 2, "PPI季度均值走势", "PPI(上年同月=100)", 10),
    ("GDP平减指数", 2, "GDP平减指数走势", "GDP平减指数", 10),
]


def add_line_chart(ws, data_col: int, title: str, y_axis_title: str, 
                   height: float = 10, width: float = 20) -> LineChart:
    """
    为worksheet添加折线图
    
    参数:
        ws: Excel worksheet
        data_col: 数据列号（第几列）
        title: 图表标题
        y_axis_title: Y轴标题
        height: 图表高度（cm）
        width: 图表宽度（cm）
    
    返回:
        创建的图表对象
    """
    # 创建折线图
    chart = LineChart()
    chart.title = title
    chart.style = 10  # 使用内置样式10（蓝色主题）
    chart.height = height
    chart.width = width
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = "期间(季度)"
    
    # 引用数据范围
    # 假设第1行是表头，第2行开始是数据
    max_row = ws.max_row - 1  # 减1是因为最后一行可能是备注
    
    if max_row < 2:
        print(f"  ⚠️ 数据行不足，跳过")
        return None
    
    # 数据范围（从表头行获取标题）
    data = Reference(ws, min_col=data_col, min_row=1, max_row=max_row)
    # 类别轴（X轴，从第1列获取期间）
    periods = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(periods)
    
    # 设置图表位置（在数据右边）
    # 尝试找到合适的位置
    ws.add_chart(chart, f"H2")
    
    return chart


def add_charts_to_excel(excel_path: str, dry_run: bool = False) -> Dict:
    """
    为Excel添加工具图表
    
    参数:
        excel_path: Excel文件路径
        dry_run: 如果为True，只显示要添加的图表，不实际添加
    
    返回:
        摘要信息
    """
    path = Path(excel_path)
    if not path.exists():
        print(f"文件不存在: {excel_path}")
        return {"error": "File not found"}
    
    wb = load_workbook(excel_path)
    
    summary = {
        "total": 0,
        "added": 0,
        "skipped": 0,
        "errors": []
    }
    
    print(f"\n处理文件: {path.name}")
    
    for sheet_name, data_col, title, y_axis, height in CHART_CONFIGS:
        summary["total"] += 1
        
        if sheet_name not in wb.sheetnames:
            print(f"\n  ⚠️ Sheet不存在: {sheet_name}")
            summary["skipped"] += 1
            continue
        
        ws = wb[sheet_name]
        
        if dry_run:
            print(f"\n  [dry-run] 将添加图表: {title}")
            print(f"    - Sheet: {sheet_name}")
            print(f"    - 数据列: {data_col}")
            print(f"    - Y轴: {y_axis}")
            continue
        
        try:
            chart = add_line_chart(ws, data_col, title, y_axis, height)
            if chart:
                print(f"  ✅ 添加: {title} -> {sheet_name}")
                summary["added"] += 1
            else:
                summary["skipped"] += 1
        except Exception as e:
            print(f"  ❌ 错误: {title} - {str(e)}")
            summary["errors"].append({
                "sheet": sheet_name,
                "chart": title,
                "error": str(e)
            })
    
    if not dry_run:
        wb.save(excel_path)
        print(f"\n文件已保存: {excel_path}")
    
    return summary


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="为Excel添加折线图")
    parser.add_argument("excel_file", help="Excel文件路径")
    parser.add_argument("--dry-run", "-n", action="store_true", 
                        help="只显示要添加的图表，不实际添加")
    parser.add_argument("--config", "-c", help="图表配置文件(JSON)")
    
    args = parser.parse_args()
    
    print("=== Excel图表添加工具 ===")
    
    summary = add_charts_to_excel(args.excel_file, args.dry_run)
    
    if not args.dry_run:
        print(f"\n=== 总结 ===")
        print(f"计划添加: {summary['total']}")
        print(f"成功添加: {summary['added']}")
        print(f"跳过: {summary['skipped']}")
        if summary['errors']:
            print(f"错误: {len(summary['errors'])}")
    
    return 0 if not summary.get('errors') else 1


if __name__ == "__main__":
    sys.exit(main())
