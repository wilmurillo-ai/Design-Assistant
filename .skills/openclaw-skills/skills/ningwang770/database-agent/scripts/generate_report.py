#!/usr/bin/env python3
"""
Excel Report Generator
Generates Excel reports for database analysis results.
"""

import argparse
import json
from datetime import datetime
from typing import Dict, List


class ReportGenerator:
    """Generates Excel reports for database compliance and analysis."""
    
    def generate_compliance_report(self, results: List[Dict], output_file: str):
        """
        Generate Excel compliance report.
        
        Args:
            results: List of compliance check results
            output_file: Output Excel file path
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("Error: openpyxl library required. Install with: pip install openpyxl")
            return
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Headers
        headers = ['Table Name', 'Compliance Score', 'Violations', 'Warnings', 'Status']
        ws_summary.append(headers)
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        total_violations = 0
        total_warnings = 0
        
        for result in results:
            violations_count = len(result.get('violations', []))
            warnings_count = len(result.get('warnings', []))
            score = result.get('compliance_score', 0)
            
            total_violations += violations_count
            total_warnings += warnings_count
            
            status = 'PASS' if score >= 80 else 'FAIL'
            
            ws_summary.append([
                result.get('table_name', ''),
                score,
                violations_count,
                warnings_count,
                status
            ])
        
        # Add summary row
        ws_summary.append([])
        ws_summary.append(['Total', '', total_violations, total_warnings])
        
        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws_summary.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Details sheet
        ws_details = wb.create_sheet("Details")
        
        detail_headers = ['Table Name', 'Issue Type', 'Severity', 'Message', 'Suggestion']
        ws_details.append(detail_headers)
        
        # Style detail headers
        for col_num, header in enumerate(detail_headers, 1):
            cell = ws_details.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add violations and warnings
        for result in results:
            table_name = result.get('table_name', '')
            
            for violation in result.get('violations', []):
                ws_details.append([
                    table_name,
                    violation.get('type', ''),
                    violation.get('severity', ''),
                    violation.get('message', ''),
                    violation.get('suggestion', '')
                ])
            
            for warning in result.get('warnings', []):
                ws_details.append([
                    table_name,
                    warning.get('type', ''),
                    warning.get('severity', ''),
                    warning.get('message', ''),
                    warning.get('suggestion', '')
                ])
        
        # Adjust column widths
        for col_num in range(1, len(detail_headers) + 1):
            ws_details.column_dimensions[get_column_letter(col_num)].width = 30
        
        # Save workbook
        wb.save(output_file)
        print(f"Report generated: {output_file}")
    
    def generate_sql_analysis_report(self, results: List[Dict], output_file: str):
        """
        Generate Excel report for SQL analysis.
        
        Args:
            results: List of SQL analysis results
            output_file: Output Excel file path
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("Error: openpyxl library required. Install with: pip install openpyxl")
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SQL Analysis"
        
        # Headers
        headers = ['SQL', 'Severity', 'Issue Type', 'Issue Message', 'Recommendation']
        ws.append(headers)
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for result in results:
            sql = result.get('sql', '')
            severity = result.get('severity', '')
            
            for issue in result.get('issues', []):
                ws.append([
                    sql[:100] + '...' if len(sql) > 100 else sql,
                    issue.get('severity', ''),
                    issue.get('type', ''),
                    issue.get('message', ''),
                    ''  # Recommendations in separate column
                ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50
        
        # Save workbook
        wb.save(output_file)
        print(f"SQL analysis report generated: {output_file}")


def main():
    parser = argparse.ArgumentParser(description='Generate Excel reports')
    parser.add_argument('--input', required=True, help='JSON file containing analysis results')
    parser.add_argument('--output', required=True, help='Output Excel file path')
    parser.add_argument('--type', choices=['compliance', 'sql_analysis'], 
                       default='compliance', help='Report type')
    
    args = parser.parse_args()
    
    # Load results
    with open(args.input, 'r') as f:
        results = json.load(f)
    
    generator = ReportGenerator()
    
    if args.type == 'compliance':
        generator.generate_compliance_report(results, args.output)
    elif args.type == 'sql_analysis':
        generator.generate_sql_analysis_report(results, args.output)


if __name__ == '__main__':
    main()
