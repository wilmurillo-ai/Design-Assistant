#!/usr/bin/env python3
"""
LZ Create MCP 客户端 — CodeBuddy Skill 脚本

通过远程 MCP Server 执行腾讯云 Landing Zone 迁移全流程。
支持 9 个子命令，覆盖从问卷生成到 Terraform 代码的完整链路。

用法:
  python3 mcp_client.py <command> [options]

子命令:
  connect         测试 MCP 连接
  generate-survey 生成空白调研问卷
  complete-form   一键智能填充（凭据→扫描→AI→Excel）
  scan            扫描源云资源
  query-specs     查询腾讯云产品规格
  design-doc      AI 生成设计文档
  terraform       AI 生成 Terraform 代码
  list-files      列出工作区文件
  download        下载工作区文件
"""

import asyncio
import argparse
import base64
import re
import sys
import time
from pathlib import Path

MCP_DEFAULT = "http://159.75.221.23/mcp"


# =====================================================================
#  Helpers
# =====================================================================

async def get_client(url, timeout=3600):
    from fastmcp import Client
    return Client(url, timeout=timeout)


async def create_session(c):
    r = await c.call_tool("create_session", {})
    sid = re.search(r"session_id:\s*(\w+)", r.content[0].text).group(1)
    return sid


async def upload_excel(c, sid, excel_path):
    p = Path(excel_path)
    b64 = base64.b64encode(p.read_bytes()).decode("ascii")
    await c.call_tool("upload_file", {
        "session_id": sid,
        "filename": p.name,
        "content_base64": b64,
    })
    return p.name


def save_binary(text, output_path):
    """从 MCP download_file 响应保存二进制文件。"""
    if "[BASE64:" in text:
        b64 = text.strip().split("\n")[1]
        Path(output_path).write_bytes(base64.b64decode(b64))
    else:
        Path(output_path).write_text(text, encoding="utf-8")


# =====================================================================
#  Commands
# =====================================================================

async def cmd_connect(args):
    """测试 MCP Server 连接。"""
    async with await get_client(args.url, timeout=30) as c:
        tools = await c.list_tools()
        print(f"MCP Server 连接成功: {args.url}")
        print(f"可用 Tools ({len(tools)}):")
        for t in tools:
            print(f"  - {t.name}")


async def cmd_generate_survey(args):
    """生成空白问卷 Excel。"""
    async with await get_client(args.url) as c:
        sid = await create_session(c)
        print(f"Session: {sid}")

        r = await c.call_tool("generate_survey", {"session_id": sid})
        print(r.content[0].text)

        # 下载
        r = await c.call_tool("download_file", {
            "session_id": sid,
            "file_path": "uploads/TencentCloud_LZ_Migration_Survey.xlsx",
        })
        out = args.output or "TencentCloud_LZ_Migration_Survey.xlsx"
        save_binary(r.content[0].text, out)
        print(f"\n已保存: {out}")


async def cmd_complete_form(args):
    """一键智能填充：凭据→扫描→AI→填好的Excel。"""
    async with await get_client(args.url, timeout=600) as c:
        sid = args.session
        if not sid:
            sid = await create_session(c)
        print(f"Session: {sid}")

        # 如果用户通过 CLI 传入凭据，先生成问卷再注入凭据
        if args.ak and args.sk:
            # 生成空白问卷到 server
            await c.call_tool("generate_survey", {"session_id": sid})
            excel_path = "TencentCloud_LZ_Migration_Survey.xlsx"

            # 下载到本地临时文件，注入凭据，再上传
            import tempfile, shutil
            r_dl = await c.call_tool("download_file", {
                "session_id": sid,
                "file_path": f"uploads/{excel_path}",
            })
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name
                save_binary(r_dl.content[0].text, tmp_path)

            # 注入凭据
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            ws = wb["凭据清单"]
            hmap = {}
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=1, column=col).value
                if v:
                    hmap[str(v).strip()] = col

            # 找第一个空行
            tgt = None
            cc = hmap.get("云平台", 2)
            for rr in range(2, ws.max_row + 2):
                cv = ws.cell(row=rr, column=cc).value
                if not cv or str(cv).strip() in ("", "nan"):
                    tgt = rr
                    break
            if not tgt:
                tgt = ws.max_row + 1

            cred_type = "STS" if args.token else "AK/SK"
            fields = {
                "序号": str(tgt - 1),
                "云平台": args.cloud or "AWS",
                "账号名称": args.account_name or "scanned-account",
                "环境": args.cred_env or "生产+非生产(混合)",
                "Region": args.region or "ap-southeast-1",
                "凭据类型": cred_type,
                "AccessKey": args.ak,
                "SecretKey": args.sk,
                "SessionToken": args.token or "",
                "权限级别": "只读(推荐)",
            }
            if args.account_id:
                fields["账号ID"] = args.account_id
            for cn, val in fields.items():
                if cn in hmap and val:
                    ws.cell(row=tgt, column=hmap[cn], value=val)
            wb.save(tmp_path)
            print(f"凭据已注入 ({args.cloud or 'AWS'}/{args.account_name or 'scanned-account'})")

            # 上传注入后的 Excel
            b64 = base64.b64encode(Path(tmp_path).read_bytes()).decode("ascii")
            await c.call_tool("upload_file", {
                "session_id": sid,
                "filename": excel_path,
                "content_base64": b64,
            })
            Path(tmp_path).unlink(missing_ok=True)
            filename = excel_path
        else:
            # 上传用户提供的 Excel
            filename = await upload_excel(c, sid, args.excel)

        print(f"上传: {filename}")

        # 智能填充
        print(f"\n开始智能填充 (模型: {args.model})...\n")
        t0 = time.time()
        r = await c.call_tool("complete_form", {
            "session_id": sid,
            "excel_filename": filename,
            "model": args.model,
        })
        print(r.content[0].text)
        print(f"\n耗时: {time.time()-t0:.1f}s")

        # 将 FILLED 版本覆盖回原始文件名，确保同 session 后续步骤读到已填写的版本
        filled = f"FILLED_{filename}"
        try:
            r_filled = await c.call_tool("download_file", {
                "session_id": sid,
                "file_path": f"uploads/{filled}",
            })
            # 覆盖回原始文件名（后续 design-doc/terraform 默认读这个名字）
            filled_b64 = r_filled.content[0].text
            if "[BASE64:" in filled_b64:
                raw_b64 = filled_b64.strip().split("\n")[1]
            else:
                raw_b64 = base64.b64encode(filled_b64.encode("utf-8")).decode("ascii")
            await c.call_tool("upload_file", {
                "session_id": sid,
                "filename": filename,
                "content_base64": raw_b64,
            })
            print(f"\n已将 FILLED 版本覆盖回 {filename}（后续步骤自动使用已填写版本）")

            # 下载到本地
            out = args.output or filled
            save_binary(r_filled.content[0].text, out)
            print(f"已下载: {out}")
        except Exception as e:
            print(f"\n下载/覆盖失败: {e}")

        print(f"Session ID: {sid}（可用于后续 design-doc / terraform）")


async def cmd_scan(args):
    """扫描源云资源。"""
    async with await get_client(args.url) as c:
        sid = args.session
        if not sid:
            sid = await create_session(c)
            filename = await upload_excel(c, sid, args.excel)
            print(f"Session: {sid}, 上传: {filename}")
        else:
            filename = Path(args.excel).name if args.excel else "TencentCloud_LZ_Migration_Survey.xlsx"
            print(f"Session: {sid}")

        print(f"\n扫描中...\n")
        t0 = time.time()
        r = await c.call_tool("scan_resources", {
            "session_id": sid,
            "excel_filename": filename,
        })
        print(r.content[0].text)
        print(f"\n耗时: {time.time()-t0:.1f}s")

        # 下载扫描结果
        out_dir = Path(args.output or ".")
        out_dir.mkdir(parents=True, exist_ok=True)
        for fp, ln in [("scan_output/resources.md", "resources.md"), ("scan_output/resources.xlsx", "resources.xlsx")]:
            try:
                r = await c.call_tool("download_file", {"session_id": sid, "file_path": fp})
                save_binary(r.content[0].text, str(out_dir / ln))
                print(f"已下载: {out_dir / ln}")
            except Exception:
                pass

        print(f"\nSession ID: {sid}")


async def cmd_query_specs(args):
    """查询腾讯云产品规格。"""
    async with await get_client(args.url) as c:
        sid = args.session
        params = {"region": args.region}
        if args.products:
            params["products"] = args.products
        if sid:
            params["session_id"] = sid

        r = await c.call_tool("query_tc_specs", params)
        text = r.content[0].text

        if args.output:
            Path(args.output).write_text(text, encoding="utf-8")
            print(f"已保存: {args.output}")
        else:
            print(text[:3000])
            if len(text) > 3000:
                print(f"\n... (共 {len(text)} 字符，使用 --output 保存完整内容)")


async def cmd_design_doc(args):
    """AI 生成设计文档。"""
    async with await get_client(args.url) as c:
        sid = args.session
        if not sid:
            if not args.excel:
                print("错误: 无 --session 时必须提供 --excel")
                sys.exit(1)
            sid = await create_session(c)
            filename = await upload_excel(c, sid, args.excel)
            print(f"Session: {sid}, 上传: {filename}")

            # 如果还没扫描，先扫描
            if args.scan_first:
                print("\n先扫描资源...")
                await c.call_tool("scan_resources", {
                    "session_id": sid,
                    "excel_filename": filename,
                })
                print("扫描完成")
        else:
            # 有 session 时，如果用户又传了新 excel，上传覆盖
            if args.excel:
                filename = await upload_excel(c, sid, args.excel)
                print(f"Session: {sid}, 重新上传: {filename}")
            else:
                print(f"Session: {sid}（使用 session 中已有的问卷）")

        print(f"\n生成设计文档 (模型: {args.model})... 预计 5-15 分钟\n")
        t0 = time.time()
        r = await c.call_tool("generate_design_doc", {
            "session_id": sid,
            "model": args.model,
        })
        print(r.content[0].text)
        print(f"\n耗时: {(time.time()-t0)/60:.1f} 分钟")

        # 下载
        out_dir = Path(args.output or ".")
        out_dir.mkdir(parents=True, exist_ok=True)
        try:
            r = await c.call_tool("download_file", {
                "session_id": sid,
                "file_path": "design_output/LZ_Design_Doc.md",
            })
            out = out_dir / "LZ_Design_Doc.md"
            out.write_text(r.content[0].text, encoding="utf-8")
            print(f"已下载: {out}")
        except Exception as e:
            print(f"下载失败: {e}")

        print(f"\nSession ID: {sid}")


async def cmd_terraform(args):
    """AI 生成 Terraform 代码。"""
    if not args.session:
        print("错误: --session 必填（需要先执行 design-doc 获取 session_id）")
        sys.exit(1)

    async with await get_client(args.url) as c:
        sid = args.session
        print(f"Session: {sid}")

        params = {
            "session_id": sid,
            "env": args.env,
            "model": args.model,
        }
        if args.app:
            params["app_name"] = args.app

        print(f"\n生成 Terraform (env={args.env}, 模型={args.model})... 预计 10-30 分钟\n")
        t0 = time.time()
        r = await c.call_tool("generate_terraform", params)
        print(r.content[0].text)
        print(f"\n耗时: {(time.time()-t0)/60:.1f} 分钟")

        # 下载所有 tf 文件
        out_dir = Path(args.output or "./terraform")
        out_dir.mkdir(parents=True, exist_ok=True)
        r_list = await c.call_tool("list_workspace_files", {"session_id": sid})
        count = 0
        for line in r_list.content[0].text.split("\n"):
            line = line.strip()
            if "terraform_output" not in line:
                continue
            fp = line.split("(")[0].strip()
            if not fp:
                continue
            try:
                r_dl = await c.call_tool("download_file", {"session_id": sid, "file_path": fp})
                rel = fp.replace("terraform_output/", "")
                local = out_dir / rel
                local.parent.mkdir(parents=True, exist_ok=True)
                save_binary(r_dl.content[0].text, str(local))
                count += 1
            except Exception:
                pass

        print(f"\n已下载 {count} 个文件到 {out_dir}/")
        print(f"Session ID: {sid}")


async def cmd_list_files(args):
    """列出工作区文件。"""
    if not args.session:
        print("错误: --session 必填")
        sys.exit(1)
    async with await get_client(args.url) as c:
        r = await c.call_tool("list_workspace_files", {"session_id": args.session})
        print(r.content[0].text)


async def cmd_download(args):
    """下载工作区文件。"""
    if not args.session or not args.file:
        print("错误: --session 和 --file 必填")
        sys.exit(1)

    async with await get_client(args.url) as c:
        r = await c.call_tool("download_file", {
            "session_id": args.session,
            "file_path": args.file,
        })
        out = args.output or Path(args.file).name
        save_binary(r.content[0].text, out)
        print(f"已下载: {out}")


# =====================================================================
#  CLI
# =====================================================================

def main():
    parser = argparse.ArgumentParser(
        description="LZ Create MCP 客户端 — 腾讯云 Landing Zone 迁移全流程",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--url", default=MCP_DEFAULT, help=f"MCP Server URL (默认: {MCP_DEFAULT})")
    sub = parser.add_subparsers(dest="command", help="子命令")

    # connect
    sub.add_parser("connect", help="测试 MCP 连接")

    # generate-survey
    p = sub.add_parser("generate-survey", help="生成空白问卷")
    p.add_argument("--output", "-o", help="输出路径")

    # complete-form
    p = sub.add_parser("complete-form", help="一键智能填充（传凭据或传 Excel）")
    p.add_argument("--excel", "-e", help="Excel 问卷路径（已填凭据）。与 --ak/--sk 二选一")
    p.add_argument("--model", default="qwen3.5:397b-cloud", help="AI 模型")
    p.add_argument("--session", "-s", help="已有 Session ID")
    p.add_argument("--output", "-o", help="输出 Excel 路径")
    cg = p.add_argument_group("凭据（无需 Excel，直接传凭据即可）")
    cg.add_argument("--ak", help="AccessKey")
    cg.add_argument("--sk", help="SecretKey")
    cg.add_argument("--token", help="SessionToken（STS）")
    cg.add_argument("--cloud", default="AWS", help="云平台（默认 AWS）")
    cg.add_argument("--region", default="ap-southeast-1", help="源云 Region")
    cg.add_argument("--account-id", help="账号 ID")
    cg.add_argument("--account-name", default="scanned-account", help="账号名称")
    cg.add_argument("--cred-env", default="生产+非生产(混合)", help="环境")

    # scan
    p = sub.add_parser("scan", help="扫描源云资源")
    p.add_argument("--excel", "-e", help="Excel 问卷路径")
    p.add_argument("--session", "-s", help="已有 Session ID")
    p.add_argument("--output", "-o", default=".", help="输出目录")

    # query-specs
    p = sub.add_parser("query-specs", help="查询腾讯云产品规格")
    p.add_argument("--region", "-r", required=True, help="腾讯云地域 (如 ap-singapore)")
    p.add_argument("--products", "-p", help="产品列表 (如 cvm,mysql,redis)")
    p.add_argument("--session", "-s", help="Session ID")
    p.add_argument("--output", "-o", help="输出文件路径")

    # design-doc
    p = sub.add_parser("design-doc", help="AI 生成设计文档")
    p.add_argument("--excel", "-e", help="Excel 问卷路径")
    p.add_argument("--model", default="qwen3.5:397b-cloud", help="AI 模型")
    p.add_argument("--session", "-s", help="已有 Session ID")
    p.add_argument("--scan-first", action="store_true", help="先扫描再生成")
    p.add_argument("--output", "-o", default=".", help="输出目录")

    # terraform
    p = sub.add_parser("terraform", help="AI 生成 Terraform 代码")
    p.add_argument("--session", "-s", required=True, help="Session ID（必填）")
    p.add_argument("--env", default="nonprod", choices=["nonprod", "prod"], help="环境")
    p.add_argument("--app", help="应用名（留空自动检测）")
    p.add_argument("--model", default="qwen3.5:397b-cloud", help="AI 模型")
    p.add_argument("--output", "-o", default="./terraform", help="输出目录")

    # list-files
    p = sub.add_parser("list-files", help="列出工作区文件")
    p.add_argument("--session", "-s", required=True, help="Session ID")

    # download
    p = sub.add_parser("download", help="下载工作区文件")
    p.add_argument("--session", "-s", required=True, help="Session ID")
    p.add_argument("--file", "-f", required=True, help="工作区内文件路径")
    p.add_argument("--output", "-o", help="本地保存路径")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(0)

    cmd_map = {
        "connect": cmd_connect,
        "generate-survey": cmd_generate_survey,
        "complete-form": cmd_complete_form,
        "scan": cmd_scan,
        "query-specs": cmd_query_specs,
        "design-doc": cmd_design_doc,
        "terraform": cmd_terraform,
        "list-files": cmd_list_files,
        "download": cmd_download,
    }

    asyncio.run(cmd_map[args.command](args))


if __name__ == "__main__":
    main()
