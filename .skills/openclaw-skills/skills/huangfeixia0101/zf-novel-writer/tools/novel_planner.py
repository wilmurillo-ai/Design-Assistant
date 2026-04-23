#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
小说全局规划工具

根据用户输入的需求，自动生成4000章小说大纲，输出 story_outline.xlsx 格式。

功能：
1. 根据核心设定生成8个纪元规划
2. 生成每章的标题、主要事件、地点、女主等
3. 输出 Excel 格式，方便用户调整
4. 最终 story_outline.xlsx 作为 Writer Agent 的基本来源

用法：
    python novel_planner.py --generate "一个关于金融系统的小说，主角重生回到过去"
    python novel_planner.py --interactive
    python novel_planner.py --preview
"""
import json
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ============================================
# 核心设定（可以由用户输入或默认）
# ============================================

DEFAULT_SETTINGS = {
    "title": "{BOOK_NAME}",
    "subtitle": "{SUBTITLE}",
    "total_chapters": 4000,
    "eras": [
        {
            "name": "第一纪元：青铜期",
            "chapters": "1-500",
            "theme": "{ERA_1_THEME}",
            "system_level": "青铜级",
            "main_heroines": ["{HEROINE_1}"],
            "main_conflict": "{VILLAIN_1}",
            "protagonist_age_start": 23,
            "protagonist_age_end": 24,
            "shen_jin_status": "{RIVAL_STATUS_1}",
        },
        {
            "name": "第二纪元：白银期",
            "chapters": "501-1000",
            "theme": "{ERA_2_THEME}",
            "system_level": "白银级",
            "main_heroines": ["{HEROINE_2}"],
            "main_conflict": "{VILLAIN_2}",
            "protagonist_age_start": 24,
            "protagonist_age_end": 26,
            "shen_jin_status": "{RIVAL_STATUS_2}",
        },
        {
            "name": "第三纪元：黄金期",
            "chapters": "501-1000",
            "theme": "{ERA_3_THEME}",
            "system_level": "黄金级",
            "main_heroines": ["{HEROINE_3}"],
            "main_conflict": "{VILLAIN_3}",
            "protagonist_age_start": 26,
            "protagonist_age_end": 28,
            "shen_jin_status": "{RIVAL_STATUS_3}",
        },
        {
            "name": "第四纪元：铂金期",
            "chapters": "1501-2000",
            "theme": "{ERA_4_THEME}",
            "system_level": "铂金级",
            "main_heroines": ["{HEROINE_4}"],
            "main_conflict": "{VILLAIN_4}",
            "protagonist_age_start": 28,
            "protagonist_age_end": 30,
            "shen_jin_status": "{RIVAL_STATUS_4}",
        },
        {
            "name": "第五纪元：钻石期",
            "chapters": "2001-2500",
            "theme": "{ERA_5_THEME}",
            "system_level": "钻石级",
            "main_heroines": ["{HEROINE_5}", "{HEROINE_6}"],
            "main_conflict": "{VILLAIN_5}",
            "protagonist_age_start": 30,
            "protagonist_age_end": 32,
            "shen_jin_status": "{RIVAL_STATUS_5}",
        },
        {
            "name": "第六纪元：王者期",
            "chapters": "2501-3000",
            "theme": "{ERA_6_THEME}",
            "system_level": "王者级",
            "main_heroines": ["{HEROINE_7}"],
            "main_conflict": "{VILLAIN_6}",
            "protagonist_age_start": 32,
            "protagonist_age_end": 35,
            "shen_jin_status": "{RIVAL_STATUS_6}",
        },
        {
            "name": "第七纪元：至尊期",
            "chapters": "3001-3500",
            "theme": "{ERA_7_THEME}",
            "system_level": "至尊级",
            "main_heroines": ["{HEROINE_8}"],
            "main_conflict": "{VILLAIN_7}",
            "protagonist_age_start": 35,
            "protagonist_age_end": 38,
            "shen_jin_status": "{RIVAL_STATUS_7}",
        },
        {
            "name": "第八纪元：主宰期",
            "chapters": "3501-4000",
            "theme": "{ERA_8_THEME}",
            "system_level": "主宰级",
            "main_heroines": [],
            "main_conflict": "{VILLAIN_8}",
            "protagonist_age_start": 38,
            "protagonist_age_end": 40,
            "shen_jin_status": "{RIVAL_STATUS_8}",
        },
    ]
}


# ============================================
# 章节模式库（用于生成章节内容）
# ============================================

CHAPTER_PATTERNS = {
    # 第一阶段：资本积累（1-100）
    "capital_accumulation": {
        "patterns": [
            ("小试牛刀", "系统提示，试探性投资", "股市", "微赚", "初次验证系统"),
            ("意外收获", "投资获得意外收益", "各种场所", "小赚", "信心增强"),
            ("第一个教训", "投资失误，获得教训", "股市/房产", "小亏", "谨慎意识"),
            ("朋友相遇", "结识新的盟友或对手", "各种场所", "无", "人脉建立"),
            ("隐藏的危机", "发现潜在的威胁", "商业场合", "无", "危机伏笔"),
        ],
        "emotions": ["紧张", "兴奋", "谨慎", "期待"],
        "cool_points": ["赚钱", "打脸", "升级", "逆袭"],
    },
    # 感情线模式
    "romance": {
        "patterns": [
            ("初次相遇", "与女主初次相遇", "日常场景", "好感", "感情线开启"),
            ("共度难关", "一起面对困难", "危机场景", "信任", "关系加深"),
            ("心结解开", "解开误会", "对话场景", "感动", "情感升华"),
            ("浪漫时刻", "单独相处", "特殊场景", "心动", "甜蜜回忆"),
            ("矛盾冲突", "发生争执", "紧张场景", "波折", "情感考验"),
        ],
        "emotions": ["好奇", "信任", "心动", "纠结", "坚定"],
        "cool_points": ["英雄救美", "温柔守护", "默契配合"],
    },
    # 战斗/对抗模式
    "confrontation": {
        "patterns": [
            ("暗中较劲", "与对手进行隐晦对抗", "商业场合", "小胜", "展示实力"),
            ("正面冲突", "直接对抗", "关键场合", "激烈", "冲突升级"),
            ("以弱胜强", "用智慧战胜强敌", "各种场景", "大胜", "信心爆棚"),
            ("惨烈失败", "遭遇重大挫折", "危机场景", "惨败", "低谷时刻"),
            ("绝地反击", "在绝境中反击", "关键场景", "逆转", "高潮爽点"),
        ],
        "emotions": ["紧张", "愤怒", "冷静", "爆发"],
        "cool_points": ["打脸", "逆袭", "复仇", "碾压"],
    },
}


def generate_era_outline(era: Dict, era_index: int, chapters_per_era: int = 500) -> List[Dict]:
    """生成单个纪元的章节大纲"""
    era_name = era["name"]
    theme = era["theme"]
    system_level = era["system_level"]
    heroines = era["main_heroines"]
    main_conflict = era["main_conflict"]
    
    chapters = []
    
    # 计算章节起始编号（每个纪元500章）
    era_start = era_index * chapters_per_era + 1
    
    # 每个纪元分成5个阶段（每100章一个阶段）
    phase_length = chapters_per_era // 5
    
    for phase in range(5):
        for i in range(phase_length):
            chapter_num = era_start + phase * phase_length + i
            
            # 根据阶段选择模式
            if phase == 0:
                # 开局阶段：主要是资本积累
                pattern_type = "capital_accumulation"
            elif phase == 4:
                # 收尾阶段：可能是大高潮或为下一纪元铺垫
                pattern_type = "confrontation"
            else:
                # 中间阶段：混合各种模式
                import random
                pattern_type = random.choice(["capital_accumulation", "romance", "confrontation"])
            
            pattern_data = CHAPTER_PATTERNS.get(pattern_type, CHAPTER_PATTERNS["capital_accumulation"])
            patterns = pattern_data["patterns"]
            
            import random
            pattern = random.choice(patterns)
            
            # 生成章节数据
            chapter = {
                "chapter": chapter_num,
                "title": pattern[0],
                "main_event": pattern[1],
                "location": pattern[2],
                "heroine": random.choice(heroines) if heroines else "无",
                "conflict_type": pattern[4] if len(pattern) > 4 else main_conflict,
                "system_prompt": f"{system_level}系统提示",
                "system_level": system_level,
                "stage": f"第{phase + 1}阶段",
                "solved_events": "",
                "cool_point": pattern[3],
                "foreshadow_type": pattern[4] if len(pattern) > 4 else "伏笔",
                "suspense_hook": f"{pattern[0]}，故事继续发展",
                "notes": f"{era_name}，{theme}",
            }
            
            chapters.append(chapter)
    
    return chapters


def generate_full_outline(settings: Dict = None) -> List[Dict]:
    """生成完整的小说大纲"""
    if settings is None:
        settings = DEFAULT_SETTINGS
    
    all_chapters = []
    chapters_per_era = 500
    
    for era_index, era in enumerate(settings["eras"]):
        chapters = generate_era_outline(era, era_index, chapters_per_era)
        all_chapters.extend(chapters)
    
    return all_chapters


def save_to_excel(chapters: List[Dict], output_path: Path):
    """保存章节大纲到 Excel"""
    if not OPENPYXL_AVAILABLE:
        print("[ERROR] openpyxl not available. Please install: pip install openpyxl")
        return False
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "章节大纲"
    
    # 标题行
    headers = [
        "章节", "标题", "主要事件", "地点", "女主", 
        "冲突类型", "系统提示", "系统等级", "阶段",
        "已解决事件", "爽点", "伏笔类型", "悬念钩子", "备注"
    ]
    ws.append(headers)
    
    # 设置标题样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据行
    for chapter in chapters:
        row = [
            chapter.get("chapter", ""),
            chapter.get("title", ""),
            chapter.get("main_event", ""),
            chapter.get("location", ""),
            chapter.get("heroine", ""),
            chapter.get("conflict_type", ""),
            chapter.get("system_prompt", ""),
            chapter.get("system_level", ""),
            chapter.get("stage", ""),
            chapter.get("solved_events", ""),
            chapter.get("cool_point", ""),
            chapter.get("foreshadow_type", ""),
            chapter.get("suspense_hook", ""),
            chapter.get("notes", ""),
        ]
        ws.append(row)
    
    # 调整列宽
    column_widths = [8, 12, 30, 10, 12, 15, 15, 12, 10, 15, 12, 15, 25, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 保存
    wb.save(output_path)
    return True


def save_to_json(chapters: List[Dict], output_path: Path):
    """保存章节大纲到 JSON（作为备份）"""
    data = {
        "version": "1.0",
        "generated_at": datetime.now().isoformat(),
        "total_chapters": len(chapters),
        "settings": DEFAULT_SETTINGS,
        "chapters": chapters
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    return True


def preview_outline(chapters: List[Dict], limit: int = 30):
    """预览大纲（显示前N章）"""
    print(f"\n{'='*80}")
    print(f"小说大纲预览（共 {len(chapters)} 章）")
    print(f"{'='*80}\n")
    
    # 按纪元分组显示
    current_era = None
    for chapter in chapters[:limit]:
        era = chapter.get("notes", "").split("，")[0] if chapter.get("notes") else ""
        
        if era != current_era:
            print(f"\n--- {era} ---")
            current_era = era
        
        ch_num = chapter.get("chapter", "")
        title = chapter.get("title", "")
        main_event = chapter.get("main_event", "")[:30]
        heroine = chapter.get("heroine", "")
        
        print(f"  {ch_num:4d}. {title:10s} | {main_event:30s} | {heroine}")
    
    if len(chapters) > limit:
        print(f"\n... 还有 {len(chapters) - limit} 章未显示 ...")


def interactive_mode():
    """交互模式：让用户定制大纲"""
    print("\n" + "="*60)
    print("小说大纲生成器 - 交互模式")
    print("="*60)
    
    # 获取基本信息
    print("\n请回答以下问题（或直接回车使用默认值）：\n")
    
    # 书名
    book_title = input("书名 [{BOOK_NAME}]: ").strip()
    if not book_title:
        book_title = settings.get("title", "{BOOK_NAME}")
    
    # 总章节数
    total_input = input("总章节数 [4000]: ").strip()
    total_chapters = int(total_input) if total_input.isdigit() else 4000
    
    # 核心主题
    theme = input("核心主题 [金融系统/重生]: ").strip()
    if not theme:
        theme = "金融系统/重生"
    
    print(f"\n✅ 开始生成 {total_chapters} 章小说大纲...")
    print(f"   主题：{theme}")
    
    # 生成大纲
    chapters = generate_full_outline()
    
    # 预览
    preview_outline(chapters, limit=20)
    
    return chapters


def main():
    import argparse
    
    parser = argparse.ArgumentParser(
        description='小说全局规划工具 - 生成 story_outline.xlsx',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # 交互模式
  python novel_planner.py --interactive
  
  # 生成默认大纲
  python novel_planner.py --generate
  
  # 预览大纲（前30章）
  python novel_planner.py --preview
  
  # 输出到指定目录
  python novel_planner.py --output /path/to/output/
        """
    )
    
    parser.add_argument('--interactive', '-i', action='store_true', 
                        help='交互模式：回答问题定制大纲')
    parser.add_argument('--generate', '-g', action='store_true',
                        help='生成完整大纲')
    parser.add_argument('--preview', '-p', action='store_true',
                        help='预览大纲（前30章）')
    parser.add_argument('--output', '-o', type=str, default=None,
                        help='输出目录')
    parser.add_argument('--json', action='store_true',
                        help='同时输出 JSON 格式')
    
    args = parser.parse_args()
    
    if not OPENPYXL_AVAILABLE:
        print("[ERROR] openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)
    
    # 确定输出目录
    if args.output:
        output_dir = Path(args.output)
    else:
        output_dir = Path.cwd() / 'books' / settings.get("title", "{BOOK_NAME}")
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 交互模式
    if args.interactive:
        chapters = interactive_mode()
    else:
        # 生成默认大纲
        chapters = generate_full_outline()
    
    # 预览
    if args.preview or (not args.interactive and not args.generate):
        preview_outline(chapters, limit=30)
    
    # 保存
    if args.generate or args.interactive:
        xlsx_path = output_dir / 'story_outline.xlsx'
        print(f"\n[OK] 保存 Excel 到: {xlsx_path}")
        save_to_excel(chapters, xlsx_path)
        
        if args.json:
            json_path = output_dir / 'story_outline_backup.json'
            print(f"[OK] 保存 JSON 到: {json_path}")
            save_to_json(chapters, json_path)
        
        print(f"\n✅ 生成了 {len(chapters)} 章的大纲！")
        print(f"   请打开 {xlsx_path} 查看和调整。")


if __name__ == '__main__':
    main()
