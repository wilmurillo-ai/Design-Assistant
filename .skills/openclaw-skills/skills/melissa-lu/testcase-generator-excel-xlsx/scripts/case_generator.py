"""
用例生成工具

功能：
1. 生成新增用例 Excel 文件（12列）
2. 生成更新用例 Excel 文件（17列）
3. 严格遵循产品用例库格式规范
4. 支持多产品配置

用法:
    from case_generator import generate_new_cases_excel, generate_update_cases_excel
    
    new_cases = [{...}, {...}]
    generate_new_cases_excel(new_cases, '需求名称', product='uhost')
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_PRODUCT = 'uhost'

# 默认输出目录
DEFAULT_OUTPUT_DIR = os.path.expanduser('~/.openclaw/workspace/')

# 产品输出目录配置
PRODUCT_OUTPUT_DIRS = {
    'uhost': DEFAULT_OUTPUT_DIR
}


def get_output_dir(product=None):
    """获取产品输出目录"""
    if product and product in PRODUCT_OUTPUT_DIRS:
        return PRODUCT_OUTPUT_DIRS[product]
    return DEFAULT_OUTPUT_DIR


def generate_new_cases_excel(cases_data, requirement_name, output_dir=None, product=None):
    """
    生成新增用例 Excel 文件
    
    参数:
        cases_data: 用例数据列表，每个元素是一个字典，包含12个字段
        requirement_name: 需求名称（用于文件命名）
        output_dir: 输出目录（默认为产品配置中的输出目录或 ~/.openclaw/workspace/）
        product: 产品名称（用于读取配置）
    
    返回:
        生成的文件路径
    """
    if output_dir is None:
        output_dir = get_output_dir(product)
    
    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "新增用例"
    
    # 表头（12列）
    headers = [
        '用例ID', '用例名称', '所属模块', '标签', '前置条件', '备注',
        '步骤描述', '预期结果', '用例等级', '用例类型',
        '是否支持自动化', '是否支持拨测'
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    column_widths = {
        'A': 40, 'B': 35, 'C': 50, 'D': 15, 'E': 60, 'F': 20,
        'G': 80, 'H': 80, 'I': 10, 'J': 10, 'K': 15, 'L': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 写入用例数据
    for row_idx, case in enumerate(cases_data, 2):
        # 用例ID留空
        ws.cell(row_idx, 1, '')
        
        # 所属模块加 /ai 前缀
        module = case.get('所属模块', '')
        if module and not module.startswith('/ai'):
            module = '/ai' + module
        
        case_row = [
            '',  # 用例ID留空
            case.get('用例名称', ''),
            module,
            case.get('标签', ''),
            case.get('前置条件', ''),
            case.get('备注', ''),
            case.get('步骤描述', ''),
            case.get('预期结果', ''),
            case.get('用例等级', 'P2'),
            case.get('用例类型', '1'),
            case.get('是否支持自动化', '2'),
            case.get('是否支持拨测', '2')
        ]
        
        for col_idx, value in enumerate(case_row, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # 保存文件
    filename = f"{requirement_name}_新增用例.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    return filepath


def generate_update_cases_excel(updates_data, requirement_name, output_dir=None, product=None):
    """
    生成更新用例 Excel 文件
    
    参数:
        updates_data: 更新数据列表，每个元素是一个字典，包含17个字段
        requirement_name: 需求名称（用于文件命名）
        output_dir: 输出目录（默认为产品配置中的输出目录或 ~/.openclaw/workspace/）
        product: 产品名称（用于读取配置）
    
    返回:
        生成的文件路径
    """
    if output_dir is None:
        output_dir = get_output_dir(product)
    
    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "更新用例"
    
    # 表头（17列 = 原12列 + 新5列）
    headers = [
        '用例ID', '用例名称', '所属模块', '标签', '前置条件', '备注',
        '步骤描述', '预期结果', '用例等级', '用例类型',
        '是否支持自动化', '是否支持拨测',
        '更新类型', '更新位置', '原内容', '更新后内容', '更新原因'
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    column_widths = {
        'A': 40, 'B': 35, 'C': 50, 'D': 15, 'E': 60, 'F': 20,
        'G': 80, 'H': 80, 'I': 10, 'J': 10, 'K': 15, 'L': 15,
        'M': 15, 'N': 15, 'O': 50, 'P': 50, 'Q': 40
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 写入更新数据
    for row_idx, update in enumerate(updates_data, 2):
        update_row = [
            update.get('用例ID', ''),
            update.get('用例名称', ''),
            update.get('所属模块', ''),
            update.get('标签', ''),
            update.get('前置条件', ''),
            update.get('备注', ''),
            update.get('步骤描述', ''),
            update.get('预期结果', ''),
            update.get('用例等级', ''),
            update.get('用例类型', ''),
            update.get('是否支持自动化', ''),
            update.get('是否支持拨测', ''),
            update.get('更新类型', ''),
            update.get('更新位置', ''),
            update.get('原内容', ''),
            update.get('更新后内容', ''),
            update.get('更新原因', '')
        ]
        
        for col_idx, value in enumerate(update_row, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # 根据用例来源设置底色
            source = update.get('用例来源', '')
            if source == '镜像用例库':
                cell.fill = PatternFill(start_color="E7F5E7", end_color="E7F5E7", fill_type="solid")
            elif source == '主机用例库':
                cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    
    # 添加说明
    ws.cell(len(updates_data) + 3, 1, "说明：")
    ws.cell(len(updates_data) + 4, 1, "1. 绿色底色：镜像用例库的更新")
    ws.cell(len(updates_data) + 5, 1, "2. 黄色底色：主机用例库的更新")
    ws.cell(len(updates_data) + 6, 1, "3. 前12列为原用例完整信息，后5列为更新说明")
    
    # 保存文件
    filename = f"{requirement_name}_更新用例.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    return filepath


if __name__ == '__main__':
    # 示例用法
    
    # 新增用例示例
    new_cases = [
        {
            '用例名称': '测试用例示例',
            '所属模块': '/镜像常规/镜像常规/功能/自制镜像tab下/导入镜像',
            '标签': '镜像',
            '前置条件': '1.前置条件1\n2.前置条件2',
            '步骤描述': '【1】步骤1\n【2】步骤2',
            '预期结果': '【1】结果1\n【2】结果2',
            '用例等级': 'P2',
            '用例类型': '1',
            '是否支持自动化': '2',
            '是否支持拨测': '2'
        }
    ]
    
    filepath = generate_new_cases_excel(new_cases, '测试需求', product='uhost')
    print(f"✅ 新增用例文件已生成: {filepath}")