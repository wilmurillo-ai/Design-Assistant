"""
用例库版本管理工具

功能：
1. 检查当前用例库版本（用例数量、最后更新时间）
2. 接收增量更新文件，按 UUID 覆盖现有用例
3. 保存用例库元数据
4. 支持多产品配置

用法:
    python3 version_manager.py --product uhost
    python3 version_manager.py --product uphost
    python3 version_manager.py --list
"""

import os
import sys
import json
import argparse
from datetime import datetime
from openpyxl import load_workbook

METADATA_FILE = os.path.expanduser('~/.openclaw/workspace/.testcase_metadata.json')
SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_PRODUCT = 'uhost'

# 默认配置（向后兼容）
DEFAULT_CONFIG = {
    'uhost': {
        'display_name': 'UHost 主机',
        'case_libraries': {
            'cpu': {
                'path': '/Users/user/testdata/--用例同步库/uhost/UHost-cpu.xlsx',
                'name': 'UHost 主机用例库',
                'count': 306
            },
            'gpu': {
                'path': '/Users/user/testdata/--用例同步库/uhost/UHost-gpu.xlsx',
                'name': 'UHost GPU 特有用例库',
                'count': 69
            },
            'image': {
                'path': '/Users/user/testdata/--用例同步库/uhost/UHost-image.xlsx',
                'name': 'UHost 镜像用例库',
                'count': 56
            }
        },
        'total_cases': 431
    }
}

# 产品配置存储路径
PRODUCT_CONFIG_DIR = os.path.expanduser('~/.openclaw/products')


def load_product_config(product):
    """加载产品配置（优先从 products 目录加载，其次使用内置默认配置）"""
    
    # 首先尝试从 products 目录加载
    config_file = os.path.join(PRODUCT_CONFIG_DIR, product, 'config.json')
    if os.path.exists(config_file):
        with open(config_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    # 其次尝试 YAML 配置（如果存在）
    yaml_file = os.path.join(PRODUCT_CONFIG_DIR, product, 'config.yaml')
    if os.path.exists(yaml_file):
        # 如果没有 pyyaml，回退到内置默认配置
        try:
            import yaml
            with open(yaml_file, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        except ImportError:
            pass
    
    # 最后使用内置默认配置
    if product in DEFAULT_CONFIG:
        return DEFAULT_CONFIG[product]
    
    return None


def check_version(product=DEFAULT_PRODUCT):
    """检查用例库版本"""
    
    config = load_product_config(product)
    
    if config is None:
        print(f"❌ 未找到产品 '{product}' 的配置")
        return None
    
    result = {
        'timestamp': datetime.now().isoformat(),
        'product': product,
        'display_name': config.get('display_name', product),
        'libraries': {}
    }
    
    total_cases = 0
    
    for lib_type, lib_info in config.get('case_libraries', {}).items():
        lib_path = lib_info.get('path', '')
        
        try:
            if not os.path.exists(lib_path):
                result['libraries'][lib_type] = {
                    'path': lib_path,
                    'name': lib_info.get('name', lib_type),
                    'error': '文件不存在'
                }
                continue
            
            wb = load_workbook(lib_path, data_only=True)
            ws = wb.active
            case_count = ws.max_row - 1  # 减去表头
            
            # 获取文件修改时间
            mtime = os.path.getmtime(lib_path)
            mtime_str = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
            
            result['libraries'][lib_type] = {
                'path': lib_path,
                'name': lib_info.get('name', lib_type),
                'count': case_count,
                'last_modified': mtime_str
            }
            
            total_cases += case_count
            
        except Exception as e:
            result['libraries'][lib_type] = {
                'path': lib_path,
                'name': lib_info.get('name', lib_type),
                'error': str(e)
            }
    
    result['total'] = total_cases
    
    # 保存元数据
    os.makedirs(os.path.dirname(METADATA_FILE), exist_ok=True)
    with open(METADATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    return result


def update_library(update_file_path, product=DEFAULT_PRODUCT, lib_type='cpu'):
    """
    更新用例库（按UUID覆盖）
    
    参数:
        update_file_path: 增量更新文件路径
        product: 产品名称
        lib_type: 用例库类型（cpu/gpu/image）
    """
    config = load_product_config(product)
    
    if config is None:
        raise ValueError(f"未找到产品 '{product}' 的配置")
    
    if lib_type not in config.get('case_libraries', {}):
        raise ValueError(f"未知的用例库类型: {lib_type}")
    
    target_path = config['case_libraries'][lib_type]['path']
    
    # 读取目标用例库
    wb_target = load_workbook(target_path)
    ws_target = wb_target.active
    
    # 读取更新文件
    wb_update = load_workbook(update_file_path, data_only=True)
    ws_update = wb_update.active
    
    # 构建目标用例库的UUID映射（UUID -> row_index）
    uuid_map = {}
    for row in range(2, ws_target.max_row + 1):
        uuid = ws_target.cell(row, 1).value
        if uuid:
            uuid_map[uuid] = row
    
    # 遍历更新文件，按UUID覆盖
    updated_count = 0
    new_count = 0
    
    for row in range(2, ws_update.max_row + 1):
        uuid = ws_update.cell(row, 1).value
        
        if not uuid:
            continue
        
        # 如果UUID存在，覆盖
        if uuid in uuid_map:
            target_row = uuid_map[uuid]
            for col in range(1, ws_update.max_column + 1):
                ws_target.cell(target_row, col).value = ws_update.cell(row, col).value
            updated_count += 1
        else:
            # UUID不存在，新增
            new_row = ws_target.max_row + 1
            for col in range(1, ws_update.max_column + 1):
                ws_target.cell(new_row, col).value = ws_update.cell(row, col).value
            new_count += 1
    
    # 保存更新后的用例库
    wb_target.save(target_path)
    
    return {
        'updated': updated_count,
        'new': new_count,
        'total': updated_count + new_count
    }


def list_products():
    """列出所有已配置的产品"""
    if not os.path.exists(PRODUCT_CONFIG_DIR):
        return list(DEFAULT_CONFIG.keys())
    
    products = []
    for name in os.listdir(PRODUCT_CONFIG_DIR):
        product_dir = os.path.join(PRODUCT_CONFIG_DIR, name)
        if os.path.isdir(product_dir):
            config_file = os.path.join(product_dir, 'config.yaml')
            json_file = os.path.join(product_dir, 'config.json')
            if os.path.exists(config_file) or os.path.exists(json_file):
                products.append(name)
    
    # 如果 products 目录为空，返回默认配置中的产品
    if not products:
        products = list(DEFAULT_CONFIG.keys())
    
    return products


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='用例库版本管理工具')
    parser.add_argument('--product', '-p', default=DEFAULT_PRODUCT, help='产品名称 (默认: uhost)')
    parser.add_argument('--list', '-l', action='store_true', help='列出所有已配置的产品')
    parser.add_argument('--update', '-u', help='增量更新文件路径')
    parser.add_argument('--lib-type', '-t', default='cpu', help='要更新的用例库类型 (cpu/gpu/image)')
    
    args = parser.parse_args()
    
    if args.list:
        products = list_products()
        print("📦 已配置的产品:")
        if products:
            for p in products:
                config = load_product_config(p)
                display = config.get('display_name', p) if config else p
                print(f"  - {p} ({display})")
        else:
            print("  (暂无)")
        sys.exit(0)
    
    if args.update:
        result = update_library(args.update, args.product, args.lib_type)
        print(f"✅ 更新完成: 更新 {result['updated']} 条, 新增 {result['new']} 条")
    else:
        version_info = check_version(args.product)
        if version_info is None:
            sys.exit(1)
        print(f"📚 [{version_info['display_name']}] 用例库版本信息:")
        print(f"总计: {version_info['total']} 条")
        for lib_type, lib_info in version_info['libraries'].items():
            if 'error' in lib_info:
                print(f"❌ {lib_info['name']}: {lib_info['error']}")
            else:
                print(f"✅ {lib_info['name']}: {lib_info['count']} 条 (更新: {lib_info['last_modified']})")