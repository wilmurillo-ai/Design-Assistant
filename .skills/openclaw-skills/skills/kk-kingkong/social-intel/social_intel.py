#!/usr/bin/env python3
"""
Social Media Intelligence Hub v2
多平台爬虫 + 数据分析 + AI 洞察
用法: python social_intel.py -k "关键词" -p xhs --analyze --export csv
"""

import argparse
import asyncio
import json
import os
import re
import sys
import time
import hashlib
from collections import Counter
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional

# ── 依赖检查 ──────────────────────────────────────────────────
def _check_deps():
    missing = []
    try:
        import wordcloud
    except ImportError:
        missing.append("wordcloud")
    try:
        import matplotlib
    except ImportError:
        missing.append("matplotlib")
    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")
    if missing:
        print(f"[WARN] 缺少依赖: {', '.join(missing)}，部分功能不可用")

_check_deps()

import subprocess
import csv
import io

# ── 配置 ──────────────────────────────────────────────────────────
TIKHUB_MCP = "mcporter"
MEDIACRAWLER_DIR = "/Users/kk/openclaw-media"
CRAWL4AI_DIR = "/Users/kk/openclaw-crawl4ai"
OUTPUT_DIR = "/tmp/social_intel"
CACHE_DIR = "/tmp/social_intel_cache"
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(CACHE_DIR, exist_ok=True)

PLATFORM_MAP = {
    "xhs":        "tikhub-xiaohongshu",
    "douyin":     "tikhub-douyin",
    "dy":         "tikhub-douyin",
    "bilibili":   "tikhub-bilibili",
    "bili":       "tikhub-bilibili",
    "weibo":      "tikhub-weibo",
    "wb":         "tikhub-weibo",
    "kuaishou":   "tikhub-kuaishou",
    "ks":         "tikhub-kuaishou",
    "tiktok":     "tikhub-tiktok",
    "youtube":    "tikhub-youtube",
    "yt":         "tikhub-youtube",
}
ALL_PLATFORMS = list(PLATFORM_MAP.keys())


# ── 工具函数 ──────────────────────────────────────────────────────

def run_cmd(cmd, timeout=25):
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, shell=False)
        return r.stdout.strip()
    except Exception:
        return ""


def mcporter_call(tool_path, **kwargs):
    parts = tool_path.split(".")
    if len(parts) != 2:
        return None
    server, method = parts
    cmd = ["mcporter", "call", f"{server}.{method}"]
    for k, v in kwargs.items():
        cmd.append(f"{k}={v}")
    result = run_cmd(cmd, timeout=25)
    if not result:
        return None
    try:
        return json.loads(result)
    except json.JSONDecodeError:
        match = re.search(r'\{.*\}', result, re.DOTALL)
        if match:
            try:
                return json.loads(match.group())
            except Exception:
                pass
        return None


def get_cache_key(platform, keyword, page=1):
    raw = f"{platform}:{keyword}:{page}"
    return hashlib.md5(raw.encode()).hexdigest()


def get_from_cache(platform, keyword, page=1):
    key = get_cache_key(platform, keyword, page)
    path = os.path.join(CACHE_DIR, f"{platform}_{key}.json")
    if os.path.exists(path):
        age = time.time() - os.path.getmtime(path)
        if age < 86400:  # 24h 缓存
            with open(path) as f:
                return json.load(f)
    return None


def save_to_cache(platform, keyword, page, data):
    key = get_cache_key(platform, keyword, page)
    path = os.path.join(CACHE_DIR, f"{platform}_{key}.json")
    with open(path, "w") as f:
        json.dump(data, f)


def progress_bar(current, total, bar_len=30, prefix=""):
    filled = int(bar_len * current / total) if total > 0 else bar_len
    bar = "█" * filled + "░" * (bar_len - filled)
    pct = int(100 * current / total) if total > 0 else 100
    sys.stdout.write(f"\r{prefix}[{bar}] {pct}% ({current}/{total})")
    sys.stdout.flush()


# ── 解析器 ──────────────────────────────────────────────────────

def parse_xhs_notes(raw):
    items = raw.get("data", {}).get("data", {}).get("items", [])
    notes = []
    for it in items:
        n = it.get("note", {})
        desc = n.get("desc", "")
        notes.append({
            "id": n.get("id", ""),
            "platform": "xhs",
            "desc": desc,
            "liked": n.get("liked_count", 0),
            "collected": n.get("collected_count", 0),
            "comment_count": n.get("comment_count", 0),
            "share_count": n.get("share_count", 0),
            "timestamp": n.get("timestamp", 0),
            "date": datetime.fromtimestamp(n.get("timestamp", 0)).strftime("%m-%d") if n.get("timestamp") else "",
            "tags": re.findall(r"#(\S+)", desc),
            "images": [img.get("url", "") for img in n.get("images_list", [])[:2]],
            "url": f"https://www.xiaohongshu.com/explore/{n.get('id','')}",
            "author": n.get("user", {}).get("nickname", ""),
        })
    return notes


def parse_weibo_notes(raw):
    cards = raw.get("data", {}).get("data", {}).get("cards", [])
    notes = []
    for card in cards:
        mblog = card.get("mblog", {})
        if not mblog:
            continue
        text = re.sub(r"<[^>]+>", "", mblog.get("text", ""))
        notes.append({
            "id": mblog.get("id", ""),
            "platform": "weibo",
            "desc": text[:200],
            "liked": mblog.get("attitudes_count", 0),
            "collected": mblog.get("reposts_count", 0),
            "comment_count": mblog.get("comments_count", 0),
            "share_count": mblog.get("reposts_count", 0),
            "timestamp": mblog.get("created_timestamp", 0),
            "date": datetime.fromtimestamp(mblog.get("created_timestamp", 0)).strftime("%m-%d") if mblog.get("created_timestamp") else "",
            "tags": re.findall(r"#(\S+)#", mblog.get("text", "")),
            "url": f"https://weibo.com/detail/{mblog.get('id','')}",
            "author": mblog.get("user", {}).get("screen_name", ""),
        })
    return notes


def parse_bilibili_notes(raw):
    videos = raw.get("data", {}).get("data", {}).get("result", [])
    notes = []
    for v in videos:
        title = re.sub(r"<[^>]+>", "", v.get("title", ""))
        desc = re.sub(r"<[^>]+>", "", v.get("description", ""))
        notes.append({
            "id": v.get("bvid", ""),
            "platform": "bilibili",
            "desc": f"{title} {desc}".strip()[:200],
            "liked": v.get("like", 0),
            "view": v.get("view", 0),
            "danmaku": v.get("danmaku", 0),
            "comment_count": v.get("review", 0),
            "share_count": v.get("share", 0),
            "timestamp": 0,
            "date": v.get("pubdate", ""),
            "tags": v.get("tag", "").split(","),
            "url": f"https://www.bilibili.com/video/{v.get('bvid','')}",
            "author": v.get("author", ""),
        })
    return notes


def parse_douyin_notes(raw):
    items = raw.get("data", {}).get("data", {}).get("video_list", [])
    notes = []
    for v in items:
        aweme = v.get("aweme_info", {})
        desc = aweme.get("desc", "")
        notes.append({
            "id": aweme.get("aweme_id", ""),
            "platform": "douyin",
            "desc": desc,
            "liked": aweme.get("statistics", {}).get("digg_count", 0),
            "view": aweme.get("statistics", {}).get("play_count", 0),
            "comment_count": aweme.get("statistics", {}).get("comment_count", 0),
            "share_count": aweme.get("statistics", {}).get("share_count", 0),
            "timestamp": aweme.get("create_time", 0),
            "date": datetime.fromtimestamp(aweme.get("create_time", 0)).strftime("%m-%d") if aweme.get("create_time") else "",
            "tags": re.findall(r"#(\S+)", desc),
            "url": f"https://www.douyin.com/video/{aweme.get('aweme_id','')}",
            "author": aweme.get("author", {}).get("nickname", ""),
        })
    return notes


def parse_kuaishou_notes(raw):
    items = raw.get("data", {}).get("data", {}).get("feeds", [])
    notes = []
    for v in items:
        notes.append({
            "id": v.get("photo_id", ""),
            "platform": "kuaishou",
            "desc": v.get("caption", ""),
            "liked": v.get("like_count", 0),
            "view": v.get("view_count", 0),
            "comment_count": v.get("comment_count", 0),
            "share_count": v.get("share_count", 0),
            "timestamp": v.get("timestamp", 0),
            "date": v.get("photo_timestamp", ""),
            "tags": re.findall(r"#(\S+)", v.get("caption", "")),
            "url": v.get("share_url", ""),
            "author": v.get("user_name", ""),
        })
    return notes


# ── 搜索 ──────────────────────────────────────────────────────────

def search_tikhub(keyword, platform, page=1, use_cache=True):
    if use_cache:
        cached = get_from_cache(platform, keyword, page)
        if cached is not None:
            return cached, True

    server = PLATFORM_MAP.get(platform, "tikhub-xiaohongshu")
    result = None

    if platform in ("xhs",):
        result = mcporter_call(f"{server}.xiaohongshu_web_search_notes_v3",
                               keyword=keyword, page=str(page), sort="general")

    elif platform in ("weibo", "wb"):
        result = mcporter_call(f"{server}.weibo_web_search_notes",
                               keyword=keyword, page=str(page))
        if not result:
            result = mcporter_call(f"{server}.weibo_web_search",
                                   keyword=keyword, page=str(page))

    elif platform in ("bilibili", "bili"):
        result = mcporter_call(f"{server}.bilibili_web_search",
                               keyword=keyword, page=str(page), order="totalrank")

    elif platform in ("douyin", "dy"):
        result = mcporter_call(f"{server}.douyin_web_search_notes",
                               keyword=keyword, page=str(page), sort="0")

    elif platform in ("kuaishou", "ks"):
        result = mcporter_call(f"{server}.kuaishou_web_search_notes",
                               keyword=keyword, page=str(page))

    if result is None:
        return [], False

    notes = []
    if platform in ("xhs",):
        notes = parse_xhs_notes(result)
    elif platform in ("weibo", "wb"):
        notes = parse_weibo_notes(result)
    elif platform in ("bilibili", "bili"):
        notes = parse_bilibili_notes(result)
    elif platform in ("douyin", "dy"):
        notes = parse_douyin_notes(result)
    elif platform in ("kuaishou", "ks"):
        notes = parse_kuaishou_notes(result)

    if notes and use_cache:
        save_to_cache(platform, keyword, page, notes)

    return notes, False


def get_comments_xhs(note_id):
    """获取小红书笔记评论"""
    result = mcporter_call("tikhub-xiaohongshu.xiaohongshu_web_get_note_comments",
                           note_id=note_id)
    if not result:
        return []
    comments = result.get("data", {}).get("data", {}).get("comments", [])
    return [{"user": c.get("user_info", {}).get("nickname", ""),
             "content": c.get("content", ""),
             "liked": c.get("like_count", 0),
             "time": c.get("create_time", "")}
            for c in comments]


# ── 多平台并行搜索 ───────────────────────────────────────────────

def search_multiplatform(keyword, platforms, max_per_platform=20):
    """并行搜索多个平台"""
    results = {}
    total = len(platforms)

    def fetch(platform):
        notes = []
        for page in range(1, 3):
            if len(notes) >= max_per_platform:
                break
            n, cached = search_tikhub(keyword, platform, page)
            notes.extend(n)
            time.sleep(0.3)
        return platform, notes[:max_per_platform]

    print(f"🔍 并行搜索 {total} 个平台...")
    with ThreadPoolExecutor(max_workers=min(total, 4)) as executor:
        futures = {executor.submit(fetch, p): p for p in platforms}
        done = 0
        for future in as_completed(futures):
            done += 1
            platform, notes = future.result()
            results[platform] = notes
            progress_bar(done, total, prefix=f"  进度 ")

    print()
    return results


# ── 数据分析 ─────────────────────────────────────────────────────

def analyze_notes(notes, keyword, platform=""):
    if not notes:
        return {}

    all_tags = []
    all_text = []
    for n in notes:
        desc = n.get("desc", "")
        if desc:
            tags = re.findall(r"#(\S+)", desc)
            all_tags.extend(tags)
            clean = re.sub(r"#\S+|\s+", " ", desc).strip()
            if clean:
                all_text.append(clean[:120])
        for tag in n.get("tags", []):
            if isinstance(tag, str):
                all_tags.append(tag)

    tag_counts = Counter(all_tags)
    total_likes = sum(n.get("liked", 0) for n in notes)
    total_views = sum(n.get("view", 0) for n in notes)
    total_collected = sum(n.get("collected", 0) for n in notes)
    total_comments = sum(n.get("comment_count", 0) for n in notes)

    # 热度趋势（按日期聚合）
    date_counts = Counter()
    date_likes = {}
    for n in notes:
        d = n.get("date", "")
        if d:
            date_counts[d] += 1
            date_likes[d] = date_likes.get(d, 0) + n.get("liked", 0)

    # 高赞/高评论/高收藏 TOP3
    top_liked = sorted(notes, key=lambda x: x.get("liked", 0), reverse=True)[:3]
    top_commented = sorted(notes, key=lambda x: x.get("comment_count", 0), reverse=True)[:3]
    top_viewed = sorted(notes, key=lambda x: x.get("view", 0), reverse=True)[:3]

    return {
        "keyword": keyword,
        "platform": platform or (notes[0].get("platform", "")),
        "total_notes": len(notes),
        "total_likes": total_likes,
        "total_views": total_views,
        "total_collected": total_collected,
        "total_comments": total_comments,
        "avg_likes": round(total_likes / len(notes), 1) if notes else 0,
        "avg_views": round(total_views / len(notes), 1) if notes else 0,
        "top_tags": [{"tag": t, "count": c} for t, c in tag_counts.most_common(20)],
        "top_liked": top_liked,
        "top_commented": top_commented,
        "top_viewed": top_viewed,
        "sample_texts": all_text[:5],
        "date_trend": dict(sorted(date_likes.items())),
        "tag_diversity": len(tag_counts),
    }


def compare_keywords(keywords, platform, max_notes=20):
    """竞品对比分析"""
    results = {}
    for kw in keywords:
        print(f"  搜索: {kw}")
        notes = []
        for page in range(1, 3):
            if len(notes) >= max_notes:
                break
            n, _ = search_tikhub(kw, platform, page)
            notes.extend(n)
            time.sleep(0.5)
        results[kw] = {
            "notes": notes[:max_notes],
            "analysis": analyze_notes(notes[:max_notes], kw, platform),
        }
    return results


def analyze_trend(keyword, platform, days=7):
    """话题趋势追踪（多天）"""
    print(f"📈 趋势分析：{keyword} 近 {days} 天")
    all_notes = []

    for day_offset in range(days):
        date = (datetime.now() - timedelta(days=day_offset)).strftime("%Y-%m-%d")
        # 取最新两页作为当天快照
        day_notes = []
        for page in range(1, 3):
            n, _ = search_tikhub(keyword, platform, page)
            if not n:
                break
            day_notes.extend(n)
            time.sleep(0.4)
        total_likes = sum(x.get("liked", 0) for x in day_notes)
        all_notes.append({
            "date": date,
            "count": len(day_notes),
            "total_likes": total_likes,
            "avg_likes": round(total_likes / len(day_notes), 1) if day_notes else 0,
            "samples": day_notes[:3]
        })
        print(f"  {date}: {len(day_notes)} 条, 总赞 {total_likes}")

    return all_notes


# ── 导出 ──────────────────────────────────────────────────────────

def export_csv(notes, output_path):
    if not notes:
        return
    fieldnames = ["platform", "id", "author", "liked", "view", "collected",
                  "comment_count", "share_count", "date", "desc", "url"]
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(notes)
    print(f"💾 CSV 保存: {output_path}")


def export_excel(data_dict, output_path):
    """data_dict: {sheet_name: [notes_list]}"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet_name, notes in data_dict.items():
            if not notes:
                continue
            ws = wb.create_sheet(title=sheet_name[:31])
            fields = ["platform", "id", "author", "liked", "view", "collected",
                     "comment_count", "share_count", "date", "desc", "url"]
            headers = ["平台", "ID", "作者", "点赞", "浏览", "收藏",
                      "评论", "转发", "日期", "内容摘要", "链接"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for note in notes:
                row = [note.get(f, "") for f in fields]
                ws.append(row)
            # 设置列宽
            ws.column_dimensions["J"].width = 50
            ws.column_dimensions["K"].width = 40
        wb.save(output_path)
        print(f"💾 Excel 保存: {output_path}")
    except ImportError:
        print("[WARN] openpyxl 未安装，导出为 CSV")
        first_sheet = list(data_dict.keys())[0]
        export_csv(data_dict[first_sheet], output_path.replace(".xlsx", ".csv"))


# ── 词云图 ────────────────────────────────────────────────────────

def generate_wordcloud(notes, output_path, keyword=""):
    try:
        from wordcloud import WordCloud
        import matplotlib.pyplot as plt
    except ImportError:
        print("[WARN] wordcloud/matplotlib 未安装，跳过词云")
        return

    all_tags = []
    for n in notes:
        desc = n.get("desc", "")
        tags = re.findall(r"#(\S+)", desc)
        all_tags.extend(tags)

    if not all_tags:
        print("[WARN] 无标签数据，无法生成词云")
        return

    tag_str = " ".join(all_tags)
    wc = WordCloud(
        width=1200, height=600,
        background_color="white",
        font_path="/System/Library/Fonts/STHeiti Light.ttc",
        max_words=100,
        colormap="viridis",
    ).generate(tag_str)

    plt.figure(figsize=(12, 6))
    plt.imshow(wc, interpolation="bilinear")
    plt.axis("off")
    plt.title(f"{keyword or '话题'} 词云", fontsize=14)
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"💾 词云保存: {output_path}")


# ── AI 洞察 ───────────────────────────────────────────────────────

def ai_insight(analysis, notes):
    """调用 MiniMax LLM 生成洞察"""
    prompt = f"""你是社交媒体数据分析师。请根据以下数据，为关键词「{analysis['keyword']}」生成一份精简的洞察报告。

数据概览：
- 平台：{analysis['platform']}
- 样本量：{analysis['total_notes']} 条笔记
- 总点赞：{analysis['total_likes']:,}
- 总收藏：{analysis['total_collected']:,}
- 总浏览：{analysis['total_views']:,}
- 平均点赞：{analysis['avg_likes']:.1f}

高频话题标签 TOP10：
{chr(10).join([f"{i+1}. #{t['tag']} ({t['count']}次)" for i, t in enumerate(analysis.get('top_tags', [])[:10])])}

高赞内容示例：
{chr(10).join([f"- 赞{n.get('liked',0)}: {n.get('desc','')[:80]}" for n in analysis.get('top_liked',[])[:3]])}

请分析：
1. 内容的核心主题和用户关注点
2. 爆款内容的共性规律
3. 用户真实需求和痛点
4. 简要结论和建议

请用中文输出，简洁有条理，100字以内。"""

    return prompt  # 返回 prompt，供上层用 MiniMax LLM 处理


# ── 打印报告 ─────────────────────────────────────────────────────

def print_banner(platform, keyword, count):
    p_name = {"xhs": "小红书", "weibo": "微博", "bilibili": "B站",
              "douyin": "抖音", "kuaishou": "快手"}.get(platform, platform.upper())
    print(f"\n{'='*55}")
    print(f"📊 {p_name} 「{keyword}」数据分析报告  |  {count} 条笔记")
    print(f"{'='*55}")


def print_report(analysis):
    if not analysis:
        print("❌ 无数据")
        return

    print(f"\n📈 总体数据")
    metrics = [
        ("总点赞", analysis.get("total_likes", 0), ""),
        ("总收藏", analysis.get("total_collected", 0), ""),
        ("总浏览", analysis.get("total_views", 0), ""),
        ("总评论", analysis.get("total_comments", 0), ""),
        ("平均点赞", analysis.get("avg_likes", 0), ""),
        ("平均浏览", analysis.get("avg_views", 0), ""),
        ("标签多样性", analysis.get("tag_diversity", 0), "种标签"),
    ]
    for name, val, unit in metrics:
        if val > 0:
            val_str = f"{val:,}" if isinstance(val, int) else f"{val:.1f}"
            print(f"  {name}: {val_str}{unit}")

    tags = analysis.get("top_tags", [])
    if tags:
        print(f"\n🏷️ 高频话题 TOP15")
        for i, t in enumerate(tags[:15], 1):
            bar_len = min(t["count"], 20)
            bar = "▓" * bar_len + "░" * (20 - bar_len)
            print(f"  {i:2d}. #{t['tag']:<28} {t['count']:>3}  {bar}")

    if analysis.get("top_liked"):
        print(f"\n🔥 点赞 TOP3")
        for i, n in enumerate(analysis["top_liked"], 1):
            print(f"  {i}. 赞{n.get('liked', 0):>7,} | {n.get('desc','')[:62]}")
    if analysis.get("top_viewed") and analysis["top_viewed"][0].get("view", 0) > 0:
        print(f"\n👀 浏览 TOP3")
        for i, n in enumerate(analysis["top_viewed"], 1):
            print(f"  {i}. 浏{analysis['top_viewed'][0].get('view',0):>8,} | {n.get('desc','')[:60]}")

    # 日期趋势
    trend = analysis.get("date_trend", {})
    if len(trend) > 1:
        print(f"\n📅 内容趋势（按日期）")
        for d, likes in sorted(trend.items()):
            bar = "▓" * min(int(likes / 100), 40)
            print(f"  {d}: {likes:>6,}赞  {bar}")

    print()


def print_compare_report(results):
    """打印竞品对比报告"""
    print(f"\n{'='*55}")
    print(f"⚔️  竞品对比报告: {list(results.keys())}")
    print(f"{'='*55}")

    for kw, data in results.items():
        a = data["analysis"]
        p = a.get("platform", "").upper()
        print(f"\n【{kw}】({a.get('total_notes',0)}条 | 均赞{a.get('avg_likes',0):.0f})")
        for i, t in enumerate(a.get("top_tags", [])[:5], 1):
            print(f"  {i}. #{t['tag']}({t['count']})", end="  ")
        print()

    # 对比表格
    print(f"\n{'─'*55}")
    print(f"{'关键词':<12} {'笔记数':>8} {'总点赞':>10} {'均点赞':>8} {'标签多样性':>10}")
    print(f"{'─'*55}")
    for kw, data in results.items():
        a = data["analysis"]
        print(f"{kw:<12} {a.get('total_notes',0):>8} {a.get('total_likes',0):>10,} "
              f"{a.get('avg_likes',0):>8.0f} {a.get('tag_diversity',0):>10}")
    print()


def print_trend_report(trend_data, keyword):
    if not trend_data:
        return
    print(f"\n📈 「{keyword}」趋势追踪")
    print(f"{'─'*40}")
    dates = [d["date"] for d in trend_data]
    counts = [d["count"] for d in trend_data]
    avg_likes = [d["avg_likes"] for d in trend_data]

    for d in trend_data:
        bar_len = min(int(d["count"] / 2), 30)
        bar = "▓" * bar_len
        print(f"  {d['date']}  {d['count']:>3}条  均赞{d['avg_likes']:>5.0f}  {bar}")

    # 趋势判断
    if len(avg_likes) >= 3:
        recent = sum(avg_likes[:3]) / 3
        older = sum(avg_likes[-3:]) / 3 if len(avg_likes) > 3 else recent
        if recent > older * 1.2:
            trend = "📈 上升趋势"
        elif recent < older * 0.8:
            trend = "📉 下降趋势"
        else:
            trend = "➡️ 基本平稳"
        print(f"\n  趋势判断: {trend}")


# ── 主流程 ───────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Social Media Intelligence Hub v2")
    parser.add_argument("-k", "--keyword", help="搜索关键词")
    parser.add_argument("-p", "--platform", default="xhs",
                        help=f"平台: {', '.join(ALL_PLATFORMS)}，或用 all 表示全部")
    parser.add_argument("-m", "--max", type=int, default=20)
    parser.add_argument("--pages", type=int, default=1)
    parser.add_argument("--analyze", action="store_true")
    parser.add_argument("--export", choices=["csv", "excel", "both"], default=None,
                        help="导出格式")
    parser.add_argument("--output", "-o", default=None)
    parser.add_argument("--all-platforms", action="store_true",
                        help="搜索所有平台")
    parser.add_argument("--compare", nargs="+", metavar="KW",
                        help="竞品对比: --compare 关键词1 关键词2")
    parser.add_argument("--trend", action="store_true",
                        help="趋势分析（近7天）")
    parser.add_argument("--comments", metavar="NOTE_ID",
                        help="采集指定笔记的评论（需配合 -p 使用）")
    parser.add_argument("--no-cache", action="store_true")
    parser.add_argument("--wordcloud", action="store_true",
                        help="生成词云图")
    args = parser.parse_args()

    use_cache = not args.no_cache

    # 竞品对比模式
    if args.compare:
        kw_list = args.compare
        platform = args.platform or "xhs"
        print(f"⚔️ 竞品对比: {kw_list} @ {platform}")
        results = compare_keywords(kw_list, platform, max_notes=args.max)
        print_compare_report(results)
        if args.export:
            out = args.output or f"{OUTPUT_DIR}/compare_{'_'.join(kw_list)}.xlsx"
            export_excel({kw: r["notes"] for kw, r in results.items()}, out)
        return

    # 趋势分析模式
    if args.trend:
        keyword = args.keyword or "AI"
        platform = args.platform or "xhs"
        trend_data = analyze_trend(keyword, platform, days=7)
        print_trend_report(trend_data, keyword)
        return

    # 评论采集模式
    if args.comments:
        keyword = args.keyword or ""
        platform = args.platform or "xhs"
        print(f"💬 采集评论: note_id={args.comments}")
        comments = get_comments_xhs(args.comments) if platform == "xhs" else []
        if comments:
            print(f"  共 {len(comments)} 条评论:")
            for c in comments[:10]:
                print(f"  @{c['user']}: {c['content'][:60]}")
        else:
            print("  无评论数据")
        return

    # 关键词必填
    if not args.keyword:
        print("❌ 请指定 -k/--keyword")
        parser.print_help()
        return

    keyword = args.keyword
    platform = args.platform.lower()
    max_notes = args.max

    # 多平台模式
    if args.all_platforms or platform == "all":
        platforms = [p for p in ALL_PLATFORMS if p not in ("tiktok", "youtube", "yt")]
        all_results = search_multiplatform(keyword, platforms, max_notes)
        total = sum(len(v) for v in all_results.values())
        print(f"\n✅ 共获取 {total} 条笔记，来自 {len(all_results)} 个平台")

        if args.analyze:
            for p, notes in all_results.items():
                if notes:
                    a = analyze_notes(notes, keyword, p)
                    print_banner(p, keyword, len(notes))
                    print_report(a)
        else:
            print_banner("all", keyword, total)

        if args.export and all_results:
            out = args.output or f"{OUTPUT_DIR}/{keyword}_multiplatform.xlsx"
            export_excel(all_results, out)

        # 词云
        if args.wordcloud:
            all_notes = [n for notes in all_results.values() for n in notes]
            wc_path = args.output or f"{OUTPUT_DIR}/{keyword}_wordcloud.png"
            generate_wordcloud(all_notes, wc_path, keyword)

        return

    # 单平台搜索
    print(f"🔍 [{platform.upper()}] 搜索:「{keyword}」 最大: {max_notes} 条")
    all_notes = []
    pages = args.pages

    for page in range(1, pages + 1):
        if len(all_notes) >= max_notes:
            break
        progress_bar(page - 1, pages, prefix=f"  页{pages} ")
        notes, cached = search_tikhub(keyword, platform, page, use_cache)
        cache_tag = " [缓存]" if cached else ""
        if notes:
            print(f"\r  第 {page} 页 {cache_tag} → {len(notes)} 条")
            all_notes.extend(notes)
        else:
            print(f"\r  第 {page} 页 → 无数据")
        time.sleep(0.4)

    # 去重
    seen = set()
    unique = []
    for n in all_notes:
        if n["id"] not in seen:
            seen.add(n["id"])
            unique.append(n)
    unique = unique[:max_notes]
    print(f"\n✅ 去重后: {len(unique)} 条笔记")

    if not unique:
        print("❌ 未获取到数据")
        return

    # 分析
    analysis = analyze_notes(unique, keyword, platform)

    if args.analyze or True:
        print_banner(platform, keyword, len(unique))
        print_report(analysis)

    # 词云
    if args.wordcloud:
        wc_path = args.output or f"{OUTPUT_DIR}/{keyword}_{platform}_wordcloud.png"
        generate_wordcloud(unique, wc_path, keyword)

    # 导出
    if args.export:
        out = args.output or f"{OUTPUT_DIR}/{keyword}_{platform}.{'xlsx' if args.export=='excel' else 'csv'}"
        if args.export in ("csv", "both"):
            export_csv(unique, out if args.export == "csv" else out.replace(".xlsx", ".csv"))
        if args.export in ("excel", "both"):
            export_excel({platform.upper(): unique}, out.replace(".csv", ".xlsx") if args.export == "both" else out)

    # AI 洞察 prompt
    insight_prompt = ai_insight(analysis, unique)
    print(f"\n💡 AI 洞察 Prompt（复制到 LLM 使用）：")
    print(f"\n{insight_prompt}\n")


if __name__ == "__main__":
    main()
