#!/usr/bin/env python3
"""
Excel Generator - 生成作业统计Excel表格
"""

import os
import json
from typing import List, Dict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("请安装依赖: pip install openpyxl")
    exit(1)


def create_statistics(students_data: List[Dict], correct_answers: List[Dict]) -> Dict:
    """统计数据"""
    total_questions = len(correct_answers)
    
    # 统计每道题的错题人数
    question_stats = {}
    for i in range(1, total_questions + 1):
        question_stats[i] = {'incorrect_count': 0, 'total': len(students_data)}
    
    # 统计每个学生的正确情况
    student_stats = []
    for student in students_data:
        correct_count = 0
        incorrect_questions = []
        
        for answer in student.get('answers', []):
            q_num = answer['question']
            is_correct = answer.get('correct', False)
            
            if is_correct:
                correct_count += 1
            else:
                incorrect_questions.append(q_num)
                if q_num in question_stats:
                    question_stats[q_num]['incorrect_count'] += 1
        
        # 计算错误率
        error_rate = ((total_questions - correct_count) / total_questions * 100) if total_questions > 0 else 0
        
        student_stats.append({
            'name': student.get('name', student.get('index', '未知')),
            'correct_count': correct_count,
            'incorrect_count': total_questions - correct_count,
            'error_rate': round(error_rate, 1),
            'incorrect_questions': incorrect_questions
        })
    
    # 计算每道题的错误率
    for q_num in question_stats:
        stats = question_stats[q_num]
        stats['error_rate'] = round(stats['incorrect_count'] / stats['total'] * 100, 1) if stats['total'] > 0 else 0
    
    return {
        'students': student_stats,
        'questions': question_stats,
        'total_questions': total_questions,
        'total_students': len(students_data)
    }


def generate_excel(students_data: List[Dict], correct_answers: List[Dict], output_path: str, threshold: float = 40):
    """生成Excel统计表"""
    stats = create_statistics(students_data, correct_answers)
    
    wb = openpyxl.Workbook()
    
    # ===== Sheet 1: 学生统计 =====
    ws_students = wb.active
    ws_students.title = "学生统计"
    
    # 标题行
    headers = ["姓名/序号", "正确题数", "错误题数", "正确率", "错题列表"]
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws_students.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 填充数据
    for row, student in enumerate(stats['students'], 2):
        ws_students.cell(row=row, column=1, value=student['name'])
        ws_students.cell(row=row, column=2, value=student['correct_count'])
        ws_students.cell(row=row, column=3, value=student['incorrect_count'])
        ws_students.cell(row=row, column=4, value=f"{student['error_rate']}%")
        ws_students.cell(row=row, column=5, value=str(student['incorrect_questions']))
    
    # 设置列宽
    ws_students.column_dimensions['A'].width = 15
    ws_students.column_dimensions['B'].width = 12
    ws_students.column_dimensions['C'].width = 12
    ws_students.column_dimensions['D'].width = 10
    ws_students.column_dimensions['E'].width = 30
    
    # ===== Sheet 2: 错题统计 =====
    ws_questions = wb.create_sheet("错题统计")
    
    q_headers = ["题号", "错题人数", "错误率", "是否重点"]
    for col, header in enumerate(q_headers, 1):
        cell = ws_questions.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 重点标记样式
    highlight_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    
    for row, (q_num, q_stat) in enumerate(stats['questions'].items(), 2):
        ws_questions.cell(row=row, column=1, value=f"第{q_num}题")
        ws_questions.cell(row=row, column=2, value=q_stat['incorrect_count'])
        ws_questions.cell(row=row, column=3, value=f"{q_stat['error_rate']}%")
        
        is重点 = q_stat['error_rate'] >= threshold
        cell = ws_questions.cell(row=row, column=4, value="⭐ 重点" if is重点 else "")
        if is重点:
            cell.fill = highlight_fill
            cell.font = Font(bold=True, color="FFFFFF")
    
    ws_questions.column_dimensions['A'].width = 10
    ws_questions.column_dimensions['B'].width = 12
    ws_questions.column_dimensions['C'].width = 10
    ws_questions.column_dimensions['D'].width = 12
    
    # ===== Sheet 3: 班级概况 =====
    ws_summary = wb.create_sheet("班级概况")
    
    # 计算班级整体数据
    total_correct = sum(s['correct_count'] for s in stats['students'])
    total_possible = stats['total_questions'] * stats['total_students']
    avg_correct_rate = round(total_correct / total_possible * 100, 1) if total_possible > 0 else 0
    
    # 重点错题数量
    key_questions = [q for q, s in stats['questions'].items() if s['error_rate'] >= threshold]
    
    summary_data = [
        ("统计时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("班级人数", stats['total_students']),
        ("题目总数", stats['total_questions']),
        ("班级平均正确率", f"{avg_correct_rate}%"),
        ("重点错题数量", len(key_questions)),
        ("重点错题题号", ", ".join([f"第{q}题" for q in key_questions]) if key_questions else "无"),
    ]
    
    for row, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)
    
    ws_summary.column_dimensions['A'].width = 18
    ws_summary.column_dimensions['B'].width = 30
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel 统计表已生成: {output_path}")
    
    return stats


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 3:
        print("用法: python generate_excel.py <学生数据json> <正确答案json> [输出路径] [阈值]")
        sys.exit(1)
    
    student_file = sys.argv[1]
    correct_file = sys.argv[2]
    output_path = sys.argv[3] if len(sys.argv) > 3 else "作业统计表.xlsx"
    threshold = float(sys.argv[4]) if len(sys.argv) > 4 else 40
    
    with open(student_file, 'r', encoding='utf-8') as f:
        students_data = json.load(f)
    
    with open(correct_file, 'r', encoding='utf-8') as f:
        correct_answers = json.load(f)
    
    generate_excel(students_data, correct_answers, output_path, threshold)