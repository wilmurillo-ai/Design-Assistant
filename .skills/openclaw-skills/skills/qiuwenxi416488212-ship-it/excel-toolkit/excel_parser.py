#!/usr/bin/env python3
"""
Excel/XLSX 操作工具
Create, inspect, and edit Microsoft Excel workbooks
"""

import os
import warnings
from typing import Union, List, Dict, Any, Optional, Tuple

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ExcelParser:
    """Excel文件解析器"""
    
    def __init__(self, path: str = None):
        self.path = path
        self.wb = None
        self.ws = None
        
        if path and os.path.exists(path):
            self.load(path)
    
    def load(self, path: str, data_only: bool = False):
        """加载Excel文件"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError('openpyxl not installed')
        
        self.path = path
        self.wb = load_workbook(path, data_only=data_only)
        self.ws = self.wb.active
        return self
    
    def save(self, path: str = None):
        """保存Excel文件"""
        if self.wb is None:
            return
        
        save_path = path or self.path
        if save_path:
            self.wb.save(save_path)
    
    def close(self):
        """关闭工作簿"""
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws = None
    
    # ============== Sheet 操作 ==============
    def get_sheets(self) -> List[str]:
        """获取所有Sheet名称"""
        return self.wb.sheetnames if self.wb else []
    
    def select_sheet(self, name: str):
        """选择Sheet"""
        if self.wb:
            self.ws = self.wb[name]
    
    def create_sheet(self, name: str):
        """创建新Sheet"""
        if self.wb:
            return self.wb.create_sheet(name)
    
    def delete_sheet(self, name: str):
        """删除Sheet"""
        if self.wb and name in self.wb.sheetnames:
            del self.wb[name]
    
    # ============== 读取数据 ==============
    def read_cell(self, cell: str) -> Any:
        """读取单个单元格"""
        return self.ws[cell].value if self.ws else None
    
    def read_range(self, range_str: str) -> List[List[Any]]:
        """读取区域数据"""
        if not self.ws:
            return []
        
        data = []
        for row in self.ws[range_str]:
            data.append([cell.value for cell in row])
        return data
    
    def read_all(self, skip_empty: bool = True) -> List[List[Any]]:
        """读取所有数据"""
        if not self.ws:
            return []
        
        data = []
        for row in self.ws.iter_rows(values_only=True):
            if skip_empty and not any(row):
                continue
            data.append(list(row))
        return data
    
    def to_dataframe(self, sheet_name: str = None) -> 'pd.DataFrame':
        """转换为DataFrame"""
        if not PANDAS_AVAILABLE:
            raise ImportError('pandas not installed')
        
        if sheet_name:
            self.select_sheet(sheet_name)
        
        return pd.read_excel(self.path, sheet_name=sheet_name or 0)
    
    # ============== 写入数据 ==============
    def write_cell(self, cell: str, value: Any):
        """写入单个单元格"""
        if self.ws:
            self.ws[cell] = value
    
    def write_row(self, row: int, values: List[Any], start_col: int = 1):
        """写入一行"""
        if self.ws:
            for i, v in enumerate(values):
                self.ws.cell(row=row, column=start_col + i, value=v)
    
    def write_range(self, start_cell: str, data: List[List[Any]]):
        """写入区域"""
        if not self.ws:
            return
        
        row, col = self._parse_cell(start_cell)
        
        for i, r in enumerate(data):
            for j, v in enumerate(r):
                self.ws.cell(row=row + i, column=col + j, value=v)
    
    def append_row(self, values: List[Any]):
        """追加一行"""
        if self.ws:
            self.ws.append(values)
    
    # ============== 样式 ==============
    def set_style(self, range_str: str, **kwargs):
        """设置样式"""
        if not self.ws:
            return
        
        # 解析样式参数
        font = kwargs.get('font')
        fill = kwargs.get('fill')
        alignment = kwargs.get('alignment')
        
        for row in self.ws[range_str]:
            for cell in row:
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
    
    def set_column_width(self, col: int, width: float):
        """设置列宽"""
        if self.ws:
            self.ws.column_dimensions[get_column_letter(col)].width = width
    
    def autofilter(self, range_str: str):
        """添加自动筛选"""
        if self.ws:
            self.ws.auto_filter.ref = range_str
    
    # ============== 公式 ==============
    def write_formula(self, cell: str, formula: str):
        """写入公式"""
        if self.ws:
            self.ws[cell] = formula
    
    def get_formula(self, cell: str) -> str:
        """获取公式"""
        return self.ws[cell].value if self.ws else None
    
    # ============== 工具方法 ==============
    @staticmethod
    def _parse_cell(cell: str) -> Tuple[int, int]:
        """解析单元格坐标"""
        import re
        match = re.match(r'([A-Z]+)(\d+)', cell.upper())
        if match:
            col = column_index_from_string(match.group(1))
            row = int(match.group(2))
            return row, col
        return 1, 1
    
    @staticmethod
    def create_template(path: str, sheets: Dict[str, List[str]] = None):
        """创建Excel模板"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError('openpyxl not installed')
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # 默认Sheet
        if not sheets:
            sheets = {'Sheet1': []}
        
        for name, headers in sheets.items():
            ws = wb.create_sheet(name)
            if headers:
                ws.append(headers)
        
        wb.save(path)
        return path
    
    # ============== 批量操作 ==============
    @staticmethod
    def merge_files(file_paths: List[str], output_path: str, sheet_name: str = None):
        """合并多个Excel文件"""
        if not PANDAS_AVAILABLE:
            raise ImportError('pandas not installed')
        
        dfs = []
        for fp in file_paths:
            df = pd.read_excel(fp, sheet_name=sheet_name)
            dfs.append(df)
        
        merged = pd.concat(dfs, ignore_index=True)
        merged.to_excel(output_path, index=False)
        return output_path
    
    @staticmethod
    def split_by_column(input_path: str, output_dir: str, column: str):
        """按列拆分Excel"""
        if not PANDAS_AVAILABLE:
            raise ImportError('pandas not installed')
        
        df = pd.read_excel(input_path)
        
        os.makedirs(output_dir, exist_ok=True)
        
        for value, group in df.groupby(column):
            filename = os.path.join(output_dir, f'{value}.xlsx')
            group.to_excel(filename, index=False)
        
        return output_dir


# ============== 便捷函数 ==============
def read_excel(path: str, **kwargs) -> Union['pd.DataFrame', Dict]:
    """读取Excel"""
    if PANDAS_AVAILABLE:
        return pd.read_excel(path, **kwargs)
    raise ImportError('pandas not installed')


def write_excel(data: 'pd.DataFrame', path: str, sheet_name: str = 'Sheet1'):
    """写入Excel"""
    if PANDAS_AVAILABLE:
        data.to_excel(path, sheet_name=sheet_name, index=False)
        return path
    raise ImportError('pandas not installed')


def excel_to_csv(excel_path: str, csv_path: str = None, sheet: int = 0):
    """Excel转CSV"""
    if not PANDAS_AVAILABLE:
        raise ImportError('pandas not installed')
    
    df = pd.read_excel(excel_path, sheet_name=sheet)
    
    if csv_path is None:
        csv_path = os.path.splitext(excel_path)[0] + '.csv'
    
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    return csv_path


def csv_to_excel(csv_path: str, excel_path: str = None, sheet_name: str = 'Sheet1'):
    """CSV转Excel"""
    if not PANDAS_AVAILABLE:
        raise ImportError('pandas not installed')
    
    df = pd.read_csv(csv_path)
    
    if excel_path is None:
        excel_path = os.path.splitext(csv_path)[0] + '.xlsx'
    
    df.to_excel(excel_path, sheet_name=sheet_name, index=False)
    return excel_path


# 直接运行测试
if __name__ == '__main__':
    print('Excel toolkit loaded')
    print(f'openpyxl: {OPENPYXL_AVAILABLE}')
    print(f'pandas: {PANDAS_AVAILABLE}')