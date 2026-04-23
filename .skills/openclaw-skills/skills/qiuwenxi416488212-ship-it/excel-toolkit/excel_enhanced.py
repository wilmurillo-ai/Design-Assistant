#!/usr/bin/env python3
"""Excel增强功能"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference


class ExcelEnhanced:
    """Excel增强功能"""
    
    def __init__(self, path=None):
        self.path = path
        self.wb = load_workbook(path) if path else Workbook()
        self.ws = self.wb.active
    
    def set_style(self, cell, bold=False, color=None, bg_color=None):
        """设置单元格样式"""
        font = Font(bold=bold, color=color)
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid") if bg_color else None
        cell.font = font
        cell.fill = fill
    
    def set_column_width(self, col, width):
        """设置列宽"""
        self.ws.column_dimensions[col].width = width
    
    def set_row_height(self, row, height):
        """设置行高"""
        self.ws.row_dimensions[row].height = height
    
    def merge_cells(self, start, end, value=None):
        """合并单元格"""
        self.ws.merge_cells(start, end)
        if value:
            self.ws[start] = value
    
    def add_chart(self, chart_type, data_range, title):
        """添加图表"""
        if chart_type == "bar":
            chart = BarChart()
        else:
            chart = LineChart()
        chart.title = title
        self.ws.add_chart(chart, "E5")
        return chart
    
    def add_image(self, cell, image_path):
        """添加图片"""
        from openpyxl.drawing.image import Image
        img = Image(image_path)
        self.ws.add_image(img, cell)
    
    def freeze_panes(self, cell):
        """冻结窗格"""
        self.ws.freeze_panes = cell
    
    def add_filter(self, range):
        """添加筛选"""
        self.ws.auto_filter.ref = range
    
    def set_print_area(self, range):
        """设置打印区域"""
        self.ws.print_area = range
    
    def protect(self, password):
        """保护工作表"""
        self.ws.protection.set_password(password)
    
    def add_validation(self, cell, type="list", formula=None, showDropDown=False):
        """数据验证"""
        from openpyxl.worksheet.dat_valid import DataValidation
        dv = DataValidation(type=type, formula1=formula, showDropDown=showDropDown)
        self.ws.add_data_validation(dv)
        dv.add(cell)
        return dv
    
    def add_hyperlink(self, cell, url, display=None):
        """添加超链接"""
        from openpyxl.cell.cell import Hyperlink
        cell.hyperlink = Hyperlink(ref=cell, target=url, display=display)
    
    def write_formula(self, cell, formula):
        """写入公式"""
        self.ws[cell] = formula
    
    def group_rows(self, start, end, collapse=True):
        """分组"""
        self.ws.row_dimensions.group(start, end, hidden=collapse)
    
    def group_columns(self, start, end, collapse=True):
        """列分组"""
        self.ws.column_dimensions.group(start, end, hidden=collapse)


def create_excel_template(path, headers, data):
    """创建Excel模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    wb.save(path)
    return path


if __name__ == "__main__":
    print("Excel Enhanced loaded")