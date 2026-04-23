#!/usr/bin/env python3
import argparse
import csv
import json
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

CANONICAL_FIELDS = [
    "日期",
    "金额",
    "记账人",
    "一级类型",
    "二级类型",
    "备注",
    "收支类型",
    "支付平台",
    "导入来源",
    "流水号",
]

KEYWORD_RULES = [
    (["瑞幸", "库迪", "星巴克"], ("餐饮", "咖啡")),
    (["麦当劳", "吉吞吞", "shawerma", "拉面", "熊野", "潮汕小馆", "堂食", "点餐订单", "美团/大众点评点餐订单"], ("餐饮", "午饭")),
    (["美团外卖", "饿了么", "外卖"], ("餐饮", "外卖")),
    (["轻巧拿", "兆歌智能售货机", "自动售货机", "先购后付"], ("餐饮", "零食")),
    (["滴滴", "T3"], ("交通", "打车")),
    (["停车", "停车场", "停车费用", "西溪银泰城", "紫金广场停车场"], ("交通", "停车")),
    (["12306"], ("交通", "高铁")),
    (["航空", "机票", "携程机票"], ("交通", "机票")),
    (["盒马", "美团买菜", "叮咚买菜"], ("日用购物", "买菜")),
    (["京东", "天猫超市", "超市"], ("日用购物", "超市")),
    (["泡泡玛特", "盲盒", "毛绒挂件"], ("娱乐休闲", "兴趣消费")),
    (["apple.com/bill", "Apple", "腾讯云购买云服务", "腾讯云费用账户"], ("娱乐休闲", "兴趣消费")),
    (["发给", "转给"], ("餐饮", "午饭")),
    (["退款"], ("收入", "退款")),
]


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip().replace("\ufeff", "")


def normalize_amount(value):
    text = clean_text(value)
    if not text:
        return ""
    text = text.replace(",", "")
    text = text.replace("¥", "").replace("￥", "")
    m = re.search(r"-?\d+(?:\.\d+)?", text)
    if not m:
        return ""
    try:
        amount = abs(Decimal(m.group(0)))
        return format(amount.normalize(), 'f').rstrip('0').rstrip('.') if '.' in format(amount.normalize(), 'f') else format(amount.normalize(), 'f')
    except InvalidOperation:
        return ""


def infer_income_expense(text):
    text = clean_text(text)
    if text in ("收入", "入", "收入 "):
        return "收入"
    if text in ("支出", "支", "支出 "):
        return "支出"
    if text in ("不计收支",):
        return "不计收支"
    if any(k in text for k in ["收入", "退款", "到账"]):
        return "收入"
    if any(k in text for k in ["支出", "付款", "消费"]):
        return "支出"
    return ""


def infer_category(*texts):
    joined = " ".join(clean_text(t) for t in texts if clean_text(t))
    for keywords, category in KEYWORD_RULES:
        if any(k in joined for k in keywords):
            return category
    return ("其他", "暂未分类")


def build_note(*parts):
    vals = [clean_text(p) for p in parts if clean_text(p)]
    deduped = []
    for v in vals:
        if v not in deduped:
            deduped.append(v)
    return " | ".join(deduped)


def row_is_empty(row):
    return not any(clean_text(v) for v in row.values())


def is_zero_amount(row):
    amount = normalize_amount(row.get("金额"))
    return amount in ("0", "0.0", "0.00")


def dedupe_key(row):
    txn = clean_text(row.get("流水号"))
    if txn:
        return f"txn:{txn}"
    return "fallback:" + "|".join([
        clean_text(row.get("日期")),
        clean_text(row.get("金额")),
        clean_text(row.get("支付平台")),
        clean_text(row.get("备注")),
    ])


def normalize_wechat(row, default_bookkeeper=""):
    note = build_note(row.get("商品"), row.get("交易对方"), row.get("备注"))
    income_expense = infer_income_expense(row.get("收/支"))
    cat1, cat2 = infer_category(row.get("商品"), row.get("交易对方"), row.get("备注"))
    if income_expense == "收入" and cat1 == "其他" and cat2 == "暂未分类":
        cat1, cat2 = ("收入", "其他收入")
    return {
        "日期": clean_text(row.get("交易时间")),
        "金额": normalize_amount(row.get("金额(元)")),
        "记账人": default_bookkeeper,
        "一级类型": cat1,
        "二级类型": cat2,
        "备注": note,
        "收支类型": income_expense,
        "支付平台": "微信",
        "导入来源": "微信账单",
        "流水号": clean_text(row.get("交易单号")),
    }


def normalize_alipay(row, default_bookkeeper=""):
    note = build_note(row.get("商品名称") or row.get("商品说明"), row.get("交易对方"), row.get("备注"))
    income_expense = infer_income_expense(row.get("收/支"))
    cat1, cat2 = infer_category(row.get("商品名称") or row.get("商品说明"), row.get("交易对方"), row.get("备注"), row.get("交易分类"))

    if income_expense == "不计收支":
        return None
    if income_expense == "收入" and cat1 == "其他" and cat2 == "暂未分类":
        cat1, cat2 = ("收入", "其他收入")

    return {
        "日期": clean_text(row.get("交易付款时间") or row.get("交易创建时间") or row.get("交易时间")),
        "金额": normalize_amount(row.get("金额（元）") or row.get("金额(元)") or row.get("金额")),
        "记账人": default_bookkeeper,
        "一级类型": cat1,
        "二级类型": cat2,
        "备注": note,
        "收支类型": income_expense,
        "支付平台": "支付宝",
        "导入来源": "支付宝账单",
        "流水号": clean_text(row.get("交易订单号")),
    }


def detect_platform(fieldnames, platform_hint=None):
    if platform_hint:
        return platform_hint.lower()
    fields = set(clean_text(f) for f in fieldnames)
    if "交易单号" in fields and "交易时间" in fields:
        return "wechat"
    if "交易订单号" in fields or "交易付款时间" in fields:
        return "alipay"
    raise ValueError("Unable to detect platform from file headers. Use --platform wechat|alipay.")


def is_alipay_header_row(values):
    header = [clean_text(v) for v in values]
    expected = ["交易时间", "交易分类", "交易对方", "对方账号", "商品说明", "收/支", "金额", "收/付款方式", "交易状态", "交易订单号", "商家订单号", "备注"]
    return header[:len(expected)] == expected


def load_rows_from_csv(path):
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                raw_rows = list(csv.reader(f))

            header_idx = None
            for idx, row in enumerate(raw_rows):
                if is_wechat_header_row(row) or is_alipay_header_row(row):
                    header_idx = idx
                    break

            if header_idx is not None:
                fieldnames = [clean_text(v) for v in raw_rows[header_idx]]
                rows = []
                for row in raw_rows[header_idx + 1:]:
                    item = {}
                    for i, field in enumerate(fieldnames):
                        if not field:
                            continue
                        value = row[i] if i < len(row) else ""
                        item[field] = value
                    rows.append(item)
                return fieldnames, rows

            with open(path, "r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                return reader.fieldnames or [], rows
        except UnicodeDecodeError as e:
            last_error = e
            continue
    raise last_error or UnicodeDecodeError("decode", b"", 0, 1, "Unable to decode CSV with utf-8-sig/utf-8/gb18030")


def is_wechat_header_row(values):
    header = [clean_text(v) for v in values]
    expected = ["交易时间", "交易类型", "交易对方", "商品", "收/支", "金额(元)", "支付方式", "当前状态", "交易单号", "商户单号", "备注"]
    return header[:len(expected)] == expected


def load_rows_from_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for idx, row in enumerate(rows):
        if is_wechat_header_row(row):
            header_idx = idx
            break

    if header_idx is None:
        raise ValueError("Unable to locate supported header row in Excel file.")

    fieldnames = [clean_text(v) for v in rows[header_idx]]
    data_rows = []
    for row in rows[header_idx + 1:]:
        item = {}
        for i, field in enumerate(fieldnames):
            if not field:
                continue
            value = row[i] if i < len(row) else ""
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            item[field] = value
        data_rows.append(item)
    return fieldnames, data_rows


def load_rows(path):
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return load_rows_from_csv(path)
    if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return load_rows_from_xlsx(path)
    raise ValueError(f"Unsupported file type: {suffix}. Expected .csv or .xlsx")


def main():
    parser = argparse.ArgumentParser(description="Normalize WeChat/Alipay bill CSV into canonical family-bookkeeping rows.")
    parser.add_argument("input", help="Path to input CSV")
    parser.add_argument("--platform", choices=["wechat", "alipay"], help="Override platform detection")
    parser.add_argument("--bookkeeper", default="", help="Default 记账人")
    parser.add_argument("--format", choices=["json", "csv"], default="json", help="Output format")
    parser.add_argument("--output", help="Write result to file instead of stdout")
    args = parser.parse_args()

    fieldnames, rows = load_rows(args.input)
    platform = detect_platform(fieldnames, args.platform)

    normalized = []
    seen = set()

    for row in rows:
        if row_is_empty(row):
            continue
        if platform == "wechat":
            item = normalize_wechat(row, args.bookkeeper)
        else:
            item = normalize_alipay(row, args.bookkeeper)

        if not item:
            continue
        if not item["日期"] and not item["金额"] and not item["备注"]:
            continue
        if is_zero_amount(item):
            continue

        key = dedupe_key(item)
        if key in seen:
            continue
        seen.add(key)
        normalized.append(item)

    if args.format == "json":
        text = json.dumps(normalized, ensure_ascii=False, indent=2)
    else:
        from io import StringIO
        buf = StringIO()
        writer = csv.DictWriter(buf, fieldnames=CANONICAL_FIELDS)
        writer.writeheader()
        writer.writerows(normalized)
        text = buf.getvalue()

    if args.output:
        Path(args.output).write_text(text, encoding="utf-8")
    else:
        print(text)


if __name__ == "__main__":
    main()
