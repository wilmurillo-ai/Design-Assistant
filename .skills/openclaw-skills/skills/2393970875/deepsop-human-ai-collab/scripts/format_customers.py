#!/usr/bin/env python3
"""
format_customers.py
将 AiWa 返回的客户 JSON 数据生成 xlsx 文件
用法: python3 format_customers.py '<json_string>' '<output_path>'
输出: xlsx 文件路径
"""

import sys
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def main():
    if len(sys.argv) < 3:
        print('用法: python3 format_customers.py \'<json_string>\' \'<output_path>\'', file=sys.stderr)
        sys.exit(1)

    # 解析 JSON
    try:
        raw = json.loads(sys.argv[1])
        data = raw.get('data', raw) if isinstance(raw, dict) else raw
    except json.JSONDecodeError as e:
        print(f'JSON 解析失败: {e}', file=sys.stderr)
        sys.exit(1)

    if not isinstance(data, list) or len(data) == 0:
        print('无客户数据', file=sys.stderr)
        sys.exit(1)

    output_path = sys.argv[2]
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    # 表头定义
    headers = [
        ('序号',     'index',       8),
        ('联系人',   'personName',  14),
        ('职位',     'position',    14),
        ('公司名称', 'companyName', 40),
        ('行业',     'industry',    14),
        ('公司官网', 'companyUrl',  28),
        ('公司规模', 'companySize', 12),
        ('邮箱',     'email',       28),
        ('电话',     'phone',       16),
        ('WhatsApp', 'whatsapp',    18),
        ('LinkedIn', 'linkedin',    24),
        ('Facebook', 'facebook',    24),
        ('Instagram','instagram',   20),
        ('TikTok',   'tiktok',      16),
        ('Twitter',  'twitter',     16),
        ('YouTube',  'youtube',     16),
        ('Line',     'line',        14),
        ('任务ID',   'taskId',      38),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'AiWa客户数据'

    # 表头样式
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2B6CB0', end_color='2B6CB0', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (label, key, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # 数据行样式
    data_align = Alignment(vertical='center', wrap_text=False)
    alt_fill = PatternFill(start_color='EBF4FF', end_color='EBF4FF', fill_type='solid')

    for row_idx, item in enumerate(data, start=2):
        for col_idx, (label, key, width) in enumerate(headers, start=1):
            if key == 'index':
                value = row_idx - 1
            else:
                value = item.get(key)
                if value is None:
                    value = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 20

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(output_path)

if __name__ == '__main__':
    main()
