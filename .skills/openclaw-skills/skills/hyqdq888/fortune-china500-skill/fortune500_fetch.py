#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
财富中国 500 强数据获取脚本
支持按年份获取完整 500 强企业名单
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime
import re

def get_available_year(requested_year):
    """
    判断请求的年份数据是否可用
    
    规则：
    - 每年 7 月下旬发布新榜单
    - 当前月份 < 7 月：最新可用数据是前年的
    - 当前月份 >= 7 月：最新可用数据是去年的
    """
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    # 判断最新可用年份
    if current_month < 7:
        latest_available = current_year - 2
    else:
        latest_available = current_year - 1
    
    # 如果请求年份 > 最新可用年份，返回不可用
    if requested_year > latest_available:
        return {
            'status': 'unavailable',
            'message': f'{requested_year}年榜单尚未发布。财富中国 500 强于每年 7 月下旬发布，当前最新可用年份为{latest_available}年',
            'latest_available': latest_available
        }
    
    # 如果请求年份 < 2010，数据可能不可用
    if requested_year < 2010:
        return {
            'status': 'unavailable',
            'message': f'{requested_year}年数据过早，本技能支持 2010 年之后的数据',
            'latest_available': latest_available
        }
    
    return {
        'status': 'available',
        'year': requested_year,
        'latest_available': latest_available
    }

def fetch_fortune_china500(year, output_path=None):
    """
    获取财富中国 500 强企业名单
    
    Args:
        year: 年份（如 2025）
        output_path: 输出 Excel 文件路径
    
    Returns:
        dict: 包含状态、企业列表、文件路径等信息
    """
    # 检查年份是否可用
    year_check = get_available_year(year)
    if year_check['status'] == 'unavailable':
        return year_check
    
    # 构建 URL
    url = f"https://www.caifuzhongwen.com/fortune500/paiming/china500/{year}_%e4%b8%ad%e5%9b%bd500%e5%bc%ba.htm"
    
    try:
        # 获取网页内容
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        response = requests.get(url, headers=headers, timeout=30)
        response.encoding = 'utf-8'
        
        if response.status_code != 200:
            return {
                'status': 'error',
                'message': f'无法访问数据源，HTTP 状态码：{response.status_code}'
            }
        
        # 解析 HTML
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 提取企业数据（根据实际网页结构调整选择器）
        companies = []
        
        # 尝试不同的解析方式
        # 方式 1：查找表格
        tables = soup.find_all('table')
        if tables:
            for table in tables:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all('td')
                    if len(cells) >= 4:
                        try:
                            rank = int(cells[0].text.strip())
                            company = cells[1].text.strip()
                            industry = cells[2].text.strip()
                            revenue = cells[3].text.strip()
                            
                            companies.append({
                                'rank': rank,
                                'company': company,
                                'industry': industry,
                                'revenue': revenue
                            })
                        except (ValueError, IndexError):
                            continue
        
        # 方式 2：如果没有表格，尝试查找特定模式
        if not companies:
            text = soup.get_text()
            # 使用正则表达式匹配企业数据
            pattern = r'(\d+)\s+([^\d]+?)\s+([^\d]+?)\s+([\d,\.]+)'
            matches = re.findall(pattern, text)
            for match in matches:
                if len(match) == 4:
                    try:
                        rank = int(match[0])
                        if 1 <= rank <= 500:
                            companies.append({
                                'rank': rank,
                                'company': match[1].strip(),
                                'industry': match[2].strip(),
                                'revenue': match[3].strip()
                            })
                    except ValueError:
                        continue
        
        if not companies:
            return {
                'status': 'error',
                'message': '无法解析企业数据，网页格式可能已变更'
            }
        
        # 创建 Excel 文件
        if output_path is None:
            output_path = f'/tmp/财富中国 500 强_{year}年.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'{year}年中国 500 强'
        
        # 添加标题行
        headers = ['排名', '企业名称', '行业分类', '营业收入 (百万美元)', '数据源']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 添加企业数据
        for idx, company in enumerate(companies, 2):
            ws.cell(row=idx, column=1, value=company['rank'])
            ws.cell(row=idx, column=2, value=company['company'])
            ws.cell(row=idx, column=3, value=company['industry'])
            ws.cell(row=idx, column=4, value=company['revenue'])
            ws.cell(row=idx, column=5, value=f'{url} | 获取时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 60
        
        # 保存文件
        wb.save(output_path)
        
        return {
            'status': 'success',
            'year': year,
            'count': len(companies),
            'output_path': output_path,
            'url': url,
            'message': f'成功获取{year}年财富中国 500 强，共{len(companies)}家企业'
        }
        
    except Exception as e:
        return {
            'status': 'error',
            'message': f'获取失败：{str(e)}'
        }

if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print('用法：python fortune500_fetch.py <年份> [输出路径]')
        print('示例：python fortune500_fetch.py 2025')
        sys.exit(1)
    
    year = int(sys.argv[1])
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    result = fetch_fortune_china500(year, output_path)
    
    if result['status'] == 'success':
        print(f"✅ {result['message']}")
        print(f"📁 文件已保存：{result['output_path']}")
    else:
        print(f"❌ {result.get('message', '未知错误')}")
