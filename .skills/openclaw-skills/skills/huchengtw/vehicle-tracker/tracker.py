#!/usr/bin/env python3
import sys
import os
import json
import argparse
import datetime
import shutil
import re
from pathlib import Path

# Add google-workspace skill to path to reuse auth
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../google-workspace')))

SKILL_DIR = os.path.dirname(__file__)
CONFIG_PATH = os.path.join(SKILL_DIR, 'config.json')
TEMPLATE_PATH = os.path.join(SKILL_DIR, 'template.xlsx')
LOCALES_DIR = os.path.join(SKILL_DIR, 'locales')

# Try importing openpyxl
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def load_config():
    if not os.path.exists(CONFIG_PATH):
        return {"vehicles": {}, "category_defaults": {}}
    with open(CONFIG_PATH, 'r') as f:
        return json.load(f)

def load_locale(locale_code):
    """Load locale file, fallback to en-US if not found."""
    locale_path = os.path.join(LOCALES_DIR, f"{locale_code}.json")
    if not os.path.exists(locale_path):
        # Fallback to en-US
        locale_path = os.path.join(LOCALES_DIR, "en-US.json")
    if not os.path.exists(locale_path):
        # Return minimal defaults if no locale found
        return {
            "sheet_name": "Expenses",
            "columns": ["Mileage", "Category", "Cost", "Qty", "Unit", "Description", "Date", "Fuel Economy"],
            "photo_prefix": "; Photo: ",
            "messages": {}
        }
    with open(locale_path, 'r') as f:
        return json.load(f)

def msg(locale, key, **kwargs):
    """Get localized message with formatting."""
    messages = locale.get('messages', {})
    template = messages.get(key, key)
    try:
        return template.format(**kwargs)
    except KeyError:
        return template

# Global Config & Locale
_CONFIG = load_config()
_LOCALE = load_locale(_CONFIG.get('locale', 'en-US'))
_UNIT_SYSTEM = _CONFIG.get('unit_system', 'metric')  # 'metric' or 'imperial'

PHOTO_BASE_DIR = os.path.expanduser(_CONFIG.get('photo_base_dir', '~/vehicle_tracker'))
LOCAL_EXCEL_DIR = os.path.expanduser(_CONFIG.get('local_excel_dir', '~/vehicle_tracker'))
SHEET_NAME = _CONFIG.get('sheet_name') or _LOCALE.get('sheet_name', 'Expenses')
PHOTO_PREFIX = _LOCALE.get('photo_prefix', '; Photo: ')

# Select columns based on unit system
if _UNIT_SYSTEM == 'imperial':
    COLUMNS = _LOCALE.get('columns_imperial', _LOCALE.get('columns', ["Mileage", "Category", "Cost", "Qty", "Unit", "Description", "Date", "Fuel Economy"]))
    UNITS_INFO = _LOCALE.get('units_imperial', {})
    DEFAULT_CATEGORY_UNITS = _LOCALE.get('default_units_imperial', {})
else:
    COLUMNS = _LOCALE.get('columns_metric', _LOCALE.get('columns', ["Mileage", "Category", "Cost", "Qty", "Unit", "Description", "Date", "Fuel Economy"]))
    UNITS_INFO = _LOCALE.get('units_metric', {})
    DEFAULT_CATEGORY_UNITS = _LOCALE.get('default_units_metric', {})

def save_config(config):
    with open(CONFIG_PATH, 'w') as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

def sanitize_filename(name):
    """Replace illegal filename characters with _"""
    return re.sub(r'[\\/*?:"<>|]', '_', name)

def main():
    parser = argparse.ArgumentParser(description='Track vehicle expenses')
    parser.add_argument('--vehicle', help='Vehicle name', default=None)
    parser.add_argument('--mileage', required=True, type=float, help='Current mileage')
    parser.add_argument('--category', required=True, help='Expense category')
    parser.add_argument('--cost', required=True, type=str, help='Cost (e.g., 205)')
    parser.add_argument('--quantity', type=float, help='Quantity')
    parser.add_argument('--unit', help='Unit')
    parser.add_argument('--description', help='Description', default='')
    parser.add_argument('--date', help='Date (YYYY-MM-DD)', default=str(datetime.date.today()))
    parser.add_argument('--photos', nargs='*', help='List of photo file paths')
    
    parser.add_argument('--save-default', action='store_true', help='Save quantity as default for this category')
    parser.add_argument('--save-unit', action='store_true', help='Save unit for this category')
    parser.add_argument('--dry-run', action='store_true', help='Preview only, do not write')

    args = parser.parse_args()
    config = load_config()
    locale = _LOCALE

    # 1. Resolve Vehicle
    vehicle_name = args.vehicle
    
    aliases = config.get('aliases', {})
    if vehicle_name in aliases:
        vehicle_name = aliases[vehicle_name]
        
    spreadsheet_id = config['vehicles'].get(vehicle_name) if vehicle_name else None
    
    if not spreadsheet_id:
        vehicles = list(config['vehicles'].keys())
        if not vehicle_name:
            if len(vehicles) == 1:
                vehicle_name = vehicles[0]
                spreadsheet_id = config['vehicles'][vehicle_name]
            elif config.get('default_vehicle'):
                vehicle_name = config.get('default_vehicle')
                spreadsheet_id = config['vehicles'].get(vehicle_name)
        
    if not vehicle_name and args.vehicle:
        vehicle_name = args.vehicle

    if not vehicle_name:
        print(msg(locale, 'error_vehicle_not_specified'))
        sys.exit(1)

    # Determine Storage Mode
    mode = 'google_sheet'
    local_excel_path = None

    if not spreadsheet_id:
        mode = 'local_excel'
        safe_name = sanitize_filename(vehicle_name)
        local_excel_path = os.path.join(LOCAL_EXCEL_DIR, f"{safe_name}.xlsx")
        
        if not HAS_OPENPYXL:
            print(msg(locale, 'error_openpyxl_missing'))
            sys.exit(1)

    # 2. Resolve Category/Unit/Quantity
    category = args.category
    category_defaults = config.get('category_defaults', {})
    cat_config = category_defaults.get(category, {})
    
    # Unit Resolution
    unit = args.unit
    if not unit:
        unit = cat_config.get('unit')
    
    if not unit:
        print(msg(locale, 'error_unit_unknown', category=category))
        sys.exit(1)

    # Quantity Resolution
    quantity = args.quantity
    if quantity is None:
        quantity = cat_config.get('quantity')
    
    if quantity is None:
        print(msg(locale, 'error_quantity_missing', category=category))
        sys.exit(1)

    # Save to config if requested or new category
    if args.save_unit or args.save_default or (category not in category_defaults):
        if 'category_defaults' not in config:
            config['category_defaults'] = {}
        if category not in config['category_defaults']:
            config['category_defaults'][category] = {}
        
        if args.save_unit or 'unit' not in config['category_defaults'][category]:
            config['category_defaults'][category]['unit'] = unit
        if args.save_default:
            config['category_defaults'][category]['quantity'] = quantity
        
        save_config(config)

    # 3. Handle Photos
    photo_filenames = []
    if args.photos:
        target_dir = Path(PHOTO_BASE_DIR) / vehicle_name / category
        
        if not args.dry_run:
            target_dir.mkdir(parents=True, exist_ok=True)
        
        for i, photo_path in enumerate(args.photos):
            src = Path(photo_path)
            if not src.exists():
                print(msg(locale, 'warning_photo_not_found', path=str(src)))
                continue
            
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
            ext = src.suffix
            new_name = f"{timestamp}-{i+1:02d}{ext}"
            
            if not args.dry_run:
                dst = target_dir / new_name
                try:
                    shutil.copy2(src, dst)
                    print(msg(locale, 'success_photo_saved', path=str(dst)))
                    photo_filenames.append(new_name)
                except Exception as e:
                    print(f"Error saving photo {src}: {e}")
            else:
                photo_filenames.append(new_name)

    # Append photo filenames to description
    if photo_filenames:
        photo_names_str = PHOTO_PREFIX + " ".join(photo_filenames)
        if args.description:
            args.description = f"{args.description}{photo_names_str}"
        else:
            args.description = photo_names_str

    # 4. Write Data
    cost_str = str(args.cost)
    # Remove common currency symbols
    cost_val = re.sub(r'[NT$,¥€£₩]', '', cost_str).strip()
    try:
        final_cost = float(cost_val)
    except ValueError:
        final_cost = cost_str

    row_data = [
        args.mileage,
        category,
        final_cost,
        quantity,
        unit,
        args.description,
        args.date
    ]
    
    if args.dry_run:
        print(json.dumps({
            "mode": mode,
            "vehicle": vehicle_name,
            "target": spreadsheet_id if mode == 'google_sheet' else local_excel_path,
            "row_data": row_data,
            "photos": args.photos or []
        }, ensure_ascii=False, indent=2))
        return

    # Execute Write
    if mode == 'google_sheet':
        try:
            from google_auth import get_service
            sheets = get_service('sheets', 'v4')
            body = {'values': [row_data]}
            result = sheets.spreadsheets().values().append(
                spreadsheetId=spreadsheet_id,
                range=f"{SHEET_NAME}!A:H",
                valueInputOption='USER_ENTERED',
                body=body
            ).execute()
            print(msg(locale, 'success_row_added_sheet', vehicle=vehicle_name, row=str(row_data)))
        except Exception as e:
            print(msg(locale, 'error_write_sheet', error=str(e)))
            sys.exit(1)
            
    elif mode == 'local_excel':
        try:
            if not os.path.exists(local_excel_path):
                if os.path.exists(TEMPLATE_PATH):
                    shutil.copy2(TEMPLATE_PATH, local_excel_path)
                    print(msg(locale, 'success_excel_created_template', path=local_excel_path))
                else:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = SHEET_NAME
                    ws.append(COLUMNS)
                    wb.save(local_excel_path)
                    print(msg(locale, 'success_excel_created_blank', path=local_excel_path))

            wb = openpyxl.load_workbook(local_excel_path)
            
            if SHEET_NAME in wb.sheetnames:
                ws = wb[SHEET_NAME]
            else:
                ws = wb.active

            ws.append(row_data)
            wb.save(local_excel_path)
            print(msg(locale, 'success_row_added_excel', path=local_excel_path, row=str(row_data)))

        except Exception as e:
            print(msg(locale, 'error_write_excel', error=str(e)))
            sys.exit(1)

if __name__ == '__main__':
    main()
