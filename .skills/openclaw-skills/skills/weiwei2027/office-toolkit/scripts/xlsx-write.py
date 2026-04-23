#!/usr/bin/env python3
"""
Create XLSX file
Usage: xlsx-write.py <output.xlsx> [options]
"""
import sys
import argparse
import json
from pathlib import Path

def import_openpyxl():
    """Import openpyxl with helpful error message"""
    try:
        from openpyxl import Workbook
        return Workbook
    except ImportError:
        print("Error: openpyxl not installed.", file=sys.stderr)
        print("Install: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

def create_workbook(output_path, data=None, sheet_name="Sheet1"):
    """Create an Excel workbook"""
    Workbook = import_openpyxl()
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if data:
        # Data can be a list of lists (rows)
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    else:
        # Add sample data
        ws['A1'] = "Hello"
        ws['B1'] = "World"
        ws['A2'] = "office-toolkit"
        ws['B2'] = "Excel support"
    
    wb.save(output_path)
    return output_path

def main():
    parser = argparse.ArgumentParser(
        description='Create XLSX file',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  xlsx-write.py output.xlsx
  xlsx-write.py output.xlsx --sheet "Data"
  xlsx-write.py output.xlsx --data "[[1,2,3],[4,5,6]]"
        """
    )
    parser.add_argument('output', help='Output XLSX file')
    parser.add_argument('--sheet', default="Sheet1", help='Sheet name (default: Sheet1)')
    parser.add_argument('--data', help='JSON array of arrays for cell data')
    
    args = parser.parse_args()
    
    # Validate output path
    output_path = Path(args.output)
    if output_path.suffix.lower() not in ['.xlsx', '.xlsm']:
        print(f"Error: Output file must be .xlsx or .xlsm: {output_path}", file=sys.stderr)
        sys.exit(1)
    
    # Create parent directory
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Parse data
    data = None
    if args.data:
        try:
            data = json.loads(args.data)
            if not isinstance(data, list):
                raise ValueError("Data must be a list of lists")
        except json.JSONDecodeError as e:
            print(f"Error: Invalid JSON data: {e}", file=sys.stderr)
            sys.exit(1)
    
    try:
        create_workbook(
            output_path,
            data=data,
            sheet_name=args.sheet
        )
        print(f"Created: {output_path}")
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
