#!/usr/bin/env python3
"""
Read XLSX file and extract data
Usage: xlsx-read.py <file.xlsx> [--sheet SHEET] [--json]
"""
import sys
import argparse
import json
from pathlib import Path

def import_openpyxl():
    """Import openpyxl with helpful error message"""
    try:
        from openpyxl import load_workbook
        return load_workbook
    except ImportError:
        print("Error: openpyxl not installed.", file=sys.stderr)
        print("Install: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

def read_xlsx(file_path, sheet_name=None):
    """Read Excel file and extract data"""
    load_workbook = import_openpyxl()
    
    wb = load_workbook(file_path, data_only=True)
    
    data = {
        'file': str(file_path),
        'sheet_names': wb.sheetnames,
        'sheets': {}
    }
    
    sheets_to_read = [sheet_name] if sheet_name else wb.sheetnames
    
    for name in sheets_to_read:
        if name not in wb.sheetnames:
            print(f"Warning: Sheet '{name}' not found, skipping", file=sys.stderr)
            continue
            
        ws = wb[name]
        sheet_data = {
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'data': []
        }
        
        for row in ws.iter_rows(values_only=True):
            sheet_data['data'].append(row)
        
        data['sheets'][name] = sheet_data
    
    wb.close()
    return data

def main():
    parser = argparse.ArgumentParser(
        description='Read XLSX file and extract data',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  xlsx-read.py spreadsheet.xlsx
  xlsx-read.py spreadsheet.xlsx --json
  xlsx-read.py spreadsheet.xlsx --sheet "Sheet1"
        """
    )
    parser.add_argument('file', help='Input XLSX file')
    parser.add_argument('--json', action='store_true', help='Output as JSON')
    parser.add_argument('--sheet', help='Specific sheet to read')
    
    args = parser.parse_args()
    
    # Validate file
    file_path = Path(args.file)
    if not file_path.exists():
        print(f"Error: File not found: {file_path}", file=sys.stderr)
        sys.exit(1)
    if file_path.suffix.lower() not in ['.xlsx', '.xlsm']:
        print(f"Error: Not an Excel file: {file_path}", file=sys.stderr)
        sys.exit(1)
    
    try:
        content = read_xlsx(file_path, sheet_name=args.sheet)
        
        if args.json:
            print(json.dumps(content, indent=2, ensure_ascii=False, default=str))
        else:
            # Plain text output
            print(f"File: {content['file']}")
            print(f"Sheets: {', '.join(content['sheet_names'])}\n")
            
            for sheet_name, sheet_data in content['sheets'].items():
                print(f"=== Sheet: {sheet_name} ===")
                print(f"Dimensions: {sheet_data['max_row']} rows x {sheet_data['max_column']} cols")
                print()
                
                for row in sheet_data['data'][:20]:  # Limit output
                    print(" | ".join(str(cell) if cell is not None else "" for cell in row))
                
                if len(sheet_data['data']) > 20:
                    print(f"... ({len(sheet_data['data']) - 20} more rows)")
                print()
                        
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
