from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

DELIVERY_COLUMNS = [
    "序号",
    "评论用户名",
    "国家",
    "星级评分",
    "评论原文",
    "评论中文版",
    "评论概括",
    "情感倾向",
    "类别分类",
    "标签",
    "重点标记",
    "评论链接网址",
    "评论时间",
    "评论点赞数",
]

REQUIRED_ANALYSIS_COLUMNS = [
    "评论原文",
    "评论中文版",
    "评论概括",
    "情感倾向",
    "类别分类",
    "标签",
    "重点标记",
]

GUIDE_ROWS = [
    ("序号", "从 1 开始连续编号。"),
    ("评论用户名", "评论者的 Amazon 用户名；没有来源就留空。"),
    ("国家", "页面明确给出的国家/地区；没有来源就留空。"),
    ("星级评分", "仅在页面有显式星级时填写。"),
    ("评论原文", "保留原语言原文；若正文为空，可回退到标题。"),
    ("评论中文版", "忠实翻译，不额外润色。"),
    ("评论概括", "一句话概括评论核心观点。"),
    ("情感倾向", "只允许 Positive / Negative / Neutral。"),
    ("类别分类", "通用运营分类口径，默认 1-2 个高信号类别。"),
    ("标签", "短标题型标签，单个标签不超过 10 个字。"),
    ("重点标记", "优先标出性价比、质量、功能改进建议、竞品/型号等。"),
    ("评论链接网址", "可追溯到具体评论的原始链接。"),
    ("评论时间", "评论页面明确给出的时间；没有来源就留空。"),
    ("评论点赞数", "评论页面明确给出的 helpful/useful 数；没有来源就留空。"),
]

QUALITY_ROWS = [
    ("标签长度", "标签列中的单个标签不超过 10 个字。"),
    ("情感值", "必须是 Positive / Negative / Neutral。"),
    ("来源完整性", "国家、评分、时间、点赞数只在页面明确提供时填写。"),
    ("同类合并", "尽量把同义表达收敛为少量稳定标签。"),
    ("重点洞察", "优先标出性价比、质量、功能改进建议、竞品/型号。"),
]

FIELD_ALIASES = {
    "序号": "序号",
    "seq": "序号",
    "index": "序号",
    "id": "序号",
    "评论用户名": "评论用户名",
    "author": "评论用户名",
    "reviewer": "评论用户名",
    "reviewer_name": "评论用户名",
    "user_name": "评论用户名",
    "username": "评论用户名",
    "国家": "国家",
    "country": "国家",
    "country_cn": "国家",
    "region": "国家",
    "market": "国家",
    "星级评分": "星级评分",
    "rating": "星级评分",
    "star_rating": "星级评分",
    "rating_text": "星级评分",
    "评论原文": "评论原文",
    "review_text": "评论原文",
    "body": "评论原文",
    "title": "评论原文",
    "评论中文版": "评论中文版",
    "translated_text": "评论中文版",
    "translated_cn": "评论中文版",
    "评论概括": "评论概括",
    "summary": "评论概括",
    "summary_cn": "评论概括",
    "情感倾向": "情感倾向",
    "sentiment": "情感倾向",
    "类别分类": "类别分类",
    "category": "类别分类",
    "categories": "类别分类",
    "标签": "标签",
    "tag": "标签",
    "tags": "标签",
    "tags_cn": "标签",
    "重点标记": "重点标记",
    "focus_mark": "重点标记",
    "focus_marks": "重点标记",
    "priority_mark": "重点标记",
    "评论链接网址": "评论链接网址",
    "review_link": "评论链接网址",
    "评论时间": "评论时间",
    "review_time": "评论时间",
    "country_date": "评论时间",
    "评论点赞数": "评论点赞数",
    "helpful_votes": "评论点赞数",
    "helpful_vote": "评论点赞数",
    "helpful_vote_statement": "评论点赞数",
}

SENTIMENT_MAP = {
    "positive": "Positive",
    "pos": "Positive",
    "正向": "Positive",
    "积极": "Positive",
    "negative": "Negative",
    "neg": "Negative",
    "负向": "Negative",
    "消极": "Negative",
    "neutral": "Neutral",
    "neu": "Neutral",
    "中性": "Neutral",
    "一般": "Neutral",
}

ALLOWED_CATEGORIES = [
    "Praise on product",
    "Questions/Worrying",
    "Negative emotions",
    "Suggestion",
    "Nothing particular",
    "Being supportive to the brand",
    "Expecting",
    "Being Sarcastic",
    "Competitor comparison",
    "Status update",
]

CATEGORY_ALIASES = {
    "Priase on product": "Praise on product",
    "Product praise": "Praise on product",
    "Questions / Worrying": "Questions/Worrying",
    "Question": "Questions/Worrying",
    "Concern": "Questions/Worrying",
}

FOCUS_MARK_EXPANSIONS = {
    "品牌/售后": ["品牌", "售后"],
    "物流/包装": ["物流", "包装"],
    "竞品/型号": ["竞品", "型号"],
    "安装/适配": ["安装", "适配"],
    "软件/设置": ["软件", "设置"],
    "配件/兼容性": ["配件", "兼容性"],
}

FOCUS_MARK_MAP = {
    "value_for_money": "性价比",
    "quality": "质量",
    "feature_suggestion": "功能改进建议",
    "competitor_model": "竞品/型号",
    "software_setup": "软件/设置",
    "installation_fit": "安装/适配",
    "brand_after_sales": "品牌/售后",
    "logistics_packaging": "物流/包装",
    "accessory_compatibility": "配件/兼容性",
}

TAG_SHORTCUTS = {
    "用户整体满意": "整体满意",
    "对产品整体满意": "整体满意",
    "对品牌整体信任": "品牌信任",
    "产品质量不错": "质量不错",
    "产品质量一般": "质量一般",
    "产品质量较差": "质量较差",
    "建议改进功能": "功能建议",
    "软件设置复杂": "设置复杂",
    "包装到货完好": "包装完好",
}

TAG_SEPARATORS = ("；", ";", "、", "|", "，", ",")


def normalize_space(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


NESTED_RECORD_KEYS = (
    "machine",
    "delivery",
    "fields",
    "analysis",
    "semantic",
    "payload",
    "data",
    "row",
    "record",
)


def flatten_record(raw: Any) -> dict[str, Any]:
    if not isinstance(raw, dict):
        return {}

    flattened: dict[str, Any] = {}
    for nested_key in NESTED_RECORD_KEYS:
        nested = raw.get(nested_key)
        if isinstance(nested, dict):
            flattened.update(flatten_record(nested))

    for key, value in raw.items():
        if key in NESTED_RECORD_KEYS and isinstance(value, dict):
            continue
        flattened[key] = value

    return flattened


def normalize_key(key: Any) -> str:
    text = normalize_space(key)
    lowered = text.lower()
    return FIELD_ALIASES.get(text, FIELD_ALIASES.get(lowered, text))


def split_values(raw: Any) -> list[str]:
    value = normalize_space(raw)
    if not value:
        return []
    for separator in TAG_SEPARATORS[1:]:
        value = value.replace(separator, TAG_SEPARATORS[0])
    return [item.strip() for item in value.split(TAG_SEPARATORS[0]) if item.strip()]


def dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def shorten_tag(tag: str) -> str:
    candidate = normalize_space(tag).replace("“", "").replace("”", "")
    candidate = TAG_SHORTCUTS.get(candidate, candidate)
    if len(candidate) <= 10:
        return candidate
    for old, new in [
        ("用户", ""),
        ("产品", ""),
        ("设备", ""),
        ("整体", ""),
        ("非常", ""),
        ("明显", ""),
        ("需要", ""),
    ]:
        candidate = candidate.replace(old, new)
        if len(candidate) <= 10:
            return candidate
    return candidate[:10]


def normalize_sentiment(raw: Any) -> str:
    value = normalize_space(raw)
    if not value:
        return ""
    return SENTIMENT_MAP.get(
        value.lower(),
        value if value in {"Positive", "Negative", "Neutral"} else "Neutral",
    )


def normalize_categories(raw: Any) -> str:
    normalized: list[str] = []
    for item in split_values(str(raw).replace(" / ", "；").replace("/", "；")):
        candidate = CATEGORY_ALIASES.get(item, item)
        if candidate in ALLOWED_CATEGORIES and candidate not in normalized:
            normalized.append(candidate)
    if not normalized:
        return "Nothing particular"
    return " / ".join(normalized[:2])


def normalize_tags(raw: Any) -> str:
    cleaned: list[str] = []
    value = str(raw or "").replace("/", "；")
    for tag in split_values(value):
        lowered = tag.lower()
        if lowered in {"none", "null", "n/a", "na"} or tag in {"", "???", "无"}:
            continue
        cleaned.append(shorten_tag(tag))
    return "；".join(dedupe_preserve_order(cleaned))


def normalize_focus_marks(raw: Any) -> str:
    if not raw:
        return ""
    values = split_values(raw)
    expanded: list[str] = []
    for value in values:
        expanded.extend(FOCUS_MARK_EXPANSIONS.get(value, [value]))
    normalized: list[str] = []
    for value in expanded:
        lowered = value.lower()
        if lowered in {"none", "null", "n/a", "na"} or value in {"", "???", "无"}:
            continue
        mapped = FOCUS_MARK_MAP.get(lowered, value)
        normalized.append(mapped)
    return "；".join(dedupe_preserve_order(normalized))


MONTH_MAP = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def normalize_review_time(raw: Any) -> str:
    value = normalize_space(raw)
    if not value:
        return ""
    for separator in (" on ", " il ", " am ", " le ", " el ", " en ", " em "):
        if separator in value:
            value = value.rsplit(separator, 1)[-1].strip()

    # YYYY年M月D日
    match = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", value)
    if match:
        y, m, d = match.group(1), match.group(2).zfill(2), match.group(3).zfill(2)
        return f"{y}-{m}-{d}"

    # DD Month YYYY or D Month YY
    match = re.search(r"(\d{1,2})\s+([A-Za-zÀ-ÿ]+)\s+(\d{2,4})", value)
    if match:
        day = int(match.group(1))
        month_str = match.group(2).lower()
        year_str = match.group(3)
        year = int(year_str) if len(year_str) == 4 else (2000 + int(year_str))
        month = MONTH_MAP.get(month_str)
        if month:
            return f"{year}-{month:02d}-{day:02d}"

    # YYYY-MM-DD
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", value)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"

    # DD/MM/YYYY or DD.MM.YYYY
    match = re.search(r"(\d{1,2})[/.](\d{1,2})[/.](\d{4})", value)
    if match:
        d, m, y = match.group(1).zfill(2), match.group(2).zfill(2), match.group(3)
        return f"{y}-{m}-{d}"

    # Relative time: "X days/weeks/months ago"
    match = re.search(r"(\d+)\s+(day|week|month|year)s?\s+ago", value, re.I)
    if match:
        from datetime import datetime, timedelta

        amount = int(match.group(1))
        unit = match.group(2).lower()
        now = datetime.now()
        if unit == "day":
            delta = timedelta(days=amount)
        elif unit == "week":
            delta = timedelta(weeks=amount)
        elif unit == "month":
            delta = timedelta(days=amount * 30)
        else:
            delta = timedelta(days=amount * 365)
        target = now - delta
        return target.strftime("%Y-%m-%d")

    return ""


def normalize_helpful_votes(raw: Any) -> str:
    value = normalize_space(raw)
    if not value:
        return ""
    lowered = value.lower()
    if "one person" in lowered or "one customer" in lowered:
        return "1"
    match = re.search(r"(\d[\d,]*)", lowered)
    if match:
        return match.group(1).replace(",", "")
    return ""


def build_review_link(review_id: str, host: str = "", raw_link: Any = "") -> str:
    link = normalize_space(raw_link)
    if link:
        return link
    review_id = normalize_space(review_id)
    host = normalize_space(host)
    if not review_id or not host:
        return ""
    base = host if host.startswith("http") else f"https://{host}"
    return f"{base.rstrip('/')}/gp/customer-reviews/{review_id}/ref=cm_cr_arp_d_rvw_ttl?ie=UTF8"


def parse_number_value(value: Any) -> int | float | None:
    text = normalize_space(value).replace(",", "")
    if not text:
        return None
    try:
        numeric = float(text)
    except ValueError:
        return None
    if numeric.is_integer():
        return int(numeric)
    return numeric


def normalize_delivery_record(raw: dict[str, Any], index: int) -> dict[str, str]:
    raw = flatten_record(raw)
    record = {column: "" for column in DELIVERY_COLUMNS}
    host = normalize_space(raw.get("host") or raw.get("site") or raw.get("source_host"))
    primary_review_text = normalize_space(
        raw.get("评论原文")
        or raw.get("review_text")
        or raw.get("body")
        or raw.get("content")
        or raw.get("review_body")
        or raw.get("title")
        or raw.get("review_title")
    )

    for key, value in raw.items():
        mapped_key = normalize_key(key)
        if mapped_key not in record:
            continue
        if mapped_key == "序号":
            record[mapped_key] = normalize_space(value)
        elif mapped_key == "评论链接网址":
            record[mapped_key] = build_review_link(
                str(raw.get("review_id") or raw.get("reviewId") or ""),
                host,
                value,
            )
        elif mapped_key == "评论时间":
            record[mapped_key] = normalize_review_time(value)
        elif mapped_key == "评论点赞数":
            record[mapped_key] = normalize_helpful_votes(value)
        else:
            record[mapped_key] = normalize_space(value)

    if primary_review_text:
        record["评论原文"] = primary_review_text
    record["序号"] = record["序号"] or str(index)
    record["情感倾向"] = normalize_sentiment(record["情感倾向"])
    record["类别分类"] = normalize_categories(record["类别分类"])
    record["标签"] = normalize_tags(record["标签"])
    record["重点标记"] = normalize_focus_marks(record["重点标记"])
    record["评论时间"] = normalize_review_time(record["评论时间"])
    record["评论点赞数"] = normalize_helpful_votes(record["评论点赞数"])
    if not record["评论链接网址"]:
        record["评论链接网址"] = build_review_link(
            str(raw.get("review_id") or raw.get("reviewId") or ""),
            host,
            raw.get("review_link") or raw.get("评论链接网址") or "",
        )
    return record


def build_delivery_records(records: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [
        normalize_delivery_record(record, index)
        for index, record in enumerate(records, start=1)
    ]


def validate_delivery_records(
    records: list[dict[str, str]], *, strict: bool = False
) -> list[str]:
    warnings: list[str] = []
    for row_index, record in enumerate(records, start=1):
        for column in REQUIRED_ANALYSIS_COLUMNS:
            if strict and not record.get(column):
                warnings.append(f"第 {row_index} 行缺少必填字段：{column}")
        sentiment = record.get("情感倾向", "")
        if sentiment and sentiment not in {"Positive", "Negative", "Neutral"}:
            warnings.append(f"第 {row_index} 行情感倾向非法：{sentiment}")
        for field in ("标签",):
            for tag in split_values(record.get(field, "")):
                if len(tag) > 10:
                    warnings.append(f"第 {row_index} 行 {field} 超过 10 个字：{tag}")

    if strict and warnings:
        raise ValueError("\n".join(warnings))
    return warnings


def build_delivery_dataframe(records: list[dict[str, str]]) -> pd.DataFrame:
    return pd.DataFrame(records, columns=DELIVERY_COLUMNS)


def write_delivery_workbook(df: pd.DataFrame, output_xlsx: Path) -> None:
    guide_df = pd.DataFrame(GUIDE_ROWS, columns=["字段", "说明"])
    quality_df = pd.DataFrame(QUALITY_ROWS, columns=["检查项", "要求"])

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="comments")
        guide_df.to_excel(writer, index=False, sheet_name="guide")
        quality_df.to_excel(writer, index=False, sheet_name="quality")

    workbook = load_workbook(output_xlsx)
    for sheet_name in ("comments", "guide", "quality"):
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    widths = {
        "A": 8,  # 序号
        "B": 18,  # 评论用户名
        "C": 14,  # 国家
        "D": 10,  # 星级评分
        "E": 48,  # 评论原文
        "F": 48,  # 评论中文版
        "G": 24,  # 评论概括
        "H": 12,  # 情感倾向
        "I": 24,  # 类别分类
        "J": 18,  # 标签
        "K": 24,  # 重点标记
        "L": 40,  # 评论链接网址
        "M": 20,  # 评论时间
        "N": 16,  # 评论点赞数
    }
    comments_sheet = workbook["comments"]
    for column, width in widths.items():
        comments_sheet.column_dimensions[column].width = width

    workbook["guide"].column_dimensions["A"].width = 18
    workbook["guide"].column_dimensions["B"].width = 80
    workbook["quality"].column_dimensions["A"].width = 18
    workbook["quality"].column_dimensions["B"].width = 80
    workbook.save(output_xlsx)


def write_delivery_artifacts(
    records: list[dict[str, Any]],
    output_json: Path,
    output_xlsx: Path,
    output_csv: Path,
    *,
    strict: bool = False,
) -> dict[str, Any]:
    normalized = build_delivery_records(records)
    warnings = validate_delivery_records(normalized, strict=strict)

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    output_csv.parent.mkdir(parents=True, exist_ok=True)

    output_json.write_text(
        json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    dataframe = build_delivery_dataframe(normalized)
    write_delivery_workbook(dataframe, output_xlsx)
    dataframe.to_csv(output_csv, index=False, encoding="utf-8-sig")

    return {
        "rows": len(normalized),
        "warnings": warnings,
        "records": normalized,
    }
