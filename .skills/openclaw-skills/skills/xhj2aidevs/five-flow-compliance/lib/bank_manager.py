#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
银行流水管理模块
"""

import json
import uuid
import csv
from datetime import datetime
from pathlib import Path

class BankManager:
    """银行流水管理器"""
    
    def __init__(self, data_dir):
        self.data_dir = Path(data_dir)
        self.bank_dir = self.data_dir / 'bank'
        self.bank_dir.mkdir(exist_ok=True)
    
    def import_from_file(self, file_path):
        """从文件导入银行流水"""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        count = 0
        ext = file_path.suffix.lower()
        
        if ext == '.csv':
            count = self._import_csv(file_path)
        elif ext in ['.xlsx', '.xls']:
            count = self._import_excel(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")
        
        return count
    
    def _import_csv(self, file_path):
        """导入 CSV 格式"""
        count = 0
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                self._create_transaction(row)
                count += 1
        return count
    
    def _import_excel(self, file_path):
        """导入 Excel 格式"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 假设第一行是表头
            headers = [cell.value for cell in ws[1]]
            count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                self._create_transaction(row_data)
                count += 1
            
            return count
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
    
    def _create_transaction(self, row_data):
        """创建交易记录"""
        # 解析金额（支持收入/支出分开或正负数）
        amount = 0
        income_val = row_data.get('收入')
        expense_val = row_data.get('支出')
        
        if income_val and float(income_val) != 0:
            amount = float(income_val)
        elif expense_val and float(expense_val) != 0:
            amount = -float(expense_val)
        elif '金额' in row_data and row_data['金额']:
            amount = float(row_data['金额'])
        
        # 解析日期
        date_str = row_data.get('交易日期', row_data.get('日期', datetime.now().strftime('%Y-%m-%d')))
        if isinstance(date_str, datetime):
            date_str = date_str.strftime('%Y-%m-%d')
        
        transaction_id = f"BNK{date_str.replace('-', '')}{uuid.uuid4().hex[:6].upper()}"
        
        transaction = {
            "id": transaction_id,
            "date": date_str,
            "amount": amount,
            "balance": float(row_data.get('余额', 0)) if '余额' in row_data else None,
            "counterparty": row_data.get('对方户名', row_data.get('对方账户', '')),
            "account": row_data.get('本方账户', row_data.get('账户', '')),
            "remark": row_data.get('摘要', row_data.get('用途', '')),
            "reference": row_data.get('流水号', row_data.get('交易号', '')),
            "project_id": None,
            "contract_id": None,
            "invoice_id": None,
            "matched": False,
            "created_at": datetime.now().isoformat()
        }
        
        file_path = self.bank_dir / f"{transaction_id}.json"
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(transaction, f, ensure_ascii=False, indent=2)
        
        return transaction
    
    def list_transactions(self, start_date=None, end_date=None, limit=50):
        """列出交易记录"""
        transactions = []
        for file_path in self.bank_dir.glob('*.json'):
            with open(file_path, 'r', encoding='utf-8') as f:
                transaction = json.load(f)
                if start_date and transaction.get('date', '') < start_date:
                    continue
                if end_date and transaction.get('date', '') > end_date:
                    continue
                transactions.append(transaction)
        
        transactions.sort(key=lambda x: x['date'], reverse=True)
        return transactions[:limit]
    
    def get_transaction(self, transaction_id):
        """获取交易详情"""
        file_path = self.bank_dir / f"{transaction_id}.json"
        if file_path.exists():
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return None
    
    def match_transaction(self, transaction_id, project_id=None, contract_id=None, invoice_id=None):
        """匹配交易到业务"""
        transaction = self.get_transaction(transaction_id)
        if not transaction:
            return None
        
        if project_id:
            transaction['project_id'] = project_id
        if contract_id:
            transaction['contract_id'] = contract_id
        if invoice_id:
            transaction['invoice_id'] = invoice_id
        
        transaction['matched'] = True
        transaction['updated_at'] = datetime.now().isoformat()
        
        file_path = self.bank_dir / f"{transaction_id}.json"
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(transaction, f, ensure_ascii=False, indent=2)
        
        return transaction
    
    def get_statistics(self, start_date=None, end_date=None):
        """获取统计信息"""
        transactions = self.list_transactions(start_date=start_date, end_date=end_date, limit=10000)
        
        total_income = 0
        total_expense = 0
        matched_income = 0
        matched_expense = 0
        
        for t in transactions:
            if t['amount'] > 0:
                total_income += t['amount']
                if t.get('matched'):
                    matched_income += t['amount']
            else:
                total_expense += abs(t['amount'])
                if t.get('matched'):
                    matched_expense += abs(t['amount'])
        
        return {
            'total_transactions': len(transactions),
            'total_income': total_income,
            'total_expense': total_expense,
            'net_amount': total_income - total_expense,
            'matched_income': matched_income,
            'matched_expense': matched_expense,
            'unmatched_income': total_income - matched_income,
            'unmatched_expense': total_expense - matched_expense
        }
    
    def get_unmatched_transactions(self):
        """获取未匹配的流水"""
        transactions = []
        for file_path in self.bank_dir.glob('*.json'):
            with open(file_path, 'r', encoding='utf-8') as f:
                transaction = json.load(f)
                if not transaction.get('matched'):
                    transactions.append(transaction)
        return sorted(transactions, key=lambda x: x['date'], reverse=True)
