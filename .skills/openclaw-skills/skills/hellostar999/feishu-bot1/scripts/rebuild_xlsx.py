import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import datetime, os, json, lark_oapi as lark
from lark_oapi.api.im.v1 import *

# Read original file
filepath = 'C:/Users/10430/Desktop/采购数据.xlsx'
wb_src = openpyxl.load_workbook(filepath)
ws_src = wb_src.active
rows = list(ws_src.iter_rows(values_only=True))
print(f'Read {len(rows)} rows from original file')

# Create new clean xlsx
wb_new = openpyxl.Workbook()
ws_new = wb_new.active
ws_new.title = '采购数据'

# Styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
center = Alignment(horizontal='center', vertical='center')

# Write header
headers = ['采购时间', '部门', '项目', '采购人', '花费']
for col, h in enumerate(headers, 1):
    cell = ws_new.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

# Write data rows
for r_idx, row in enumerate(rows[1:], 2):  # skip header row
    for c_idx, val in enumerate(row, 1):
        if isinstance(val, datetime.datetime):
            ws_new.cell(row=r_idx, column=c_idx, value=val).number_format = 'YYYY-MM-DD'
        else:
            ws_new.cell(row=r_idx, column=c_idx, value=val)

# Auto-adjust column widths
for col in ws_new.columns:
    max_len = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws_new.column_dimensions[col_letter].width = min(max_len + 4, 30)

# Save locally
out_path = 'C:/Users/10430/Desktop/采购数据_干净版.xlsx'
wb_new.save(out_path)
print(f'Saved to: {out_path}')

# Upload via lark-oapi SDK
config = json.load(open('C:/Users/10430/.openclaw/workspace/skills/feishu-ops/scripts/config.json', encoding='utf-8'))
client = lark.Client.builder() \
    .app_id(config['app_id']) \
    .app_secret(config['app_secret']) \
    .log_level(lark.LogLevel.ERROR) \
    .build()

with open(out_path, 'rb') as f:
    file_data = f.read()

upload_resp = client.im.v1.file.create(
    CreateFileRequest.builder()
        .request_body(CreateFileRequestBody.builder()
            .file_type('xlsx')
            .file_name('采购数据_干净版.xlsx')
            .file(file_data)
            .build())
        .build()
)
print('Upload code:', upload_resp.code)
if upload_resp.code != 0:
    print('Upload failed:', upload_resp.msg)
    import sys
    sys.exit(1)

file_key = upload_resp.data.file_key
print('file_key:', file_key)

# Send to chat
msg_resp = client.im.v1.message.create(
    CreateMessageRequest.builder()
        .receive_id_type('chat_id')
        .request_body(CreateMessageRequestBody.builder()
            .receive_id('oc_2c6df8f6e06e88d34729baacc124b89e')
            .msg_type('file')
            .content(json.dumps({'file_key': file_key}))
            .build())
        .build()
)
print('Send code:', msg_resp.code, 'msg:', msg_resp.msg)
print('Done!')
