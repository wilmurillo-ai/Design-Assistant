"""
MSA Excel 报告生成器
支持 GR&R、偏倚、线性研究的 Excel 格式报告
"""

import io
from datetime import datetime
from typing import Dict, Any, Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, Color
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, LineChart, BarChart
from openpyxl.chart.label import DataLabelList


# ==================== 样式定义 ====================

class StyleManager:
    """Excel 样式管理器"""
    
    def __init__(self):
        # 字体
        self.title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        self.heading_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        self.subheading_font = Font(name='微软雅黑', size=11, bold=True)
        self.normal_font = Font(name='微软雅黑', size=10)
        self.small_font = Font(name='微软雅黑', size=9, color='666666')
        
        # 对齐
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
        
        # 边框
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 填充
        self.header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        self.alt_fill1 = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
        self.alt_fill2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        self.good_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
        self.warning_fill = PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid')
        self.bad_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')


# ==================== GR&R Excel 报告 ====================

def create_grr_excel(
    study_info: Dict[str, Any],
    grr_result: Dict[str, float],
    evaluation: Dict[str, Any],
    data_array: np.ndarray
) -> bytes:
    """创建 GR&R Excel 报告"""
    wb = Workbook()
    styles = StyleManager()
    
    # 1. 封面页
    ws_cover = wb.active
    ws_cover.title = "封面"
    _create_cover_sheet(ws_cover, study_info, styles)
    
    # 2. 结果摘要页
    ws_summary = wb.create_sheet("结果摘要")
    _create_grr_summary(ws_summary, grr_result, evaluation, styles)
    
    # 3. 原始数据页
    ws_data = wb.create_sheet("原始数据")
    _create_grr_data(ws_data, data_array, styles)
    
    # 4. 分析图表页
    ws_chart = wb.create_sheet("分析图表")
    _create_grr_charts(ws_chart, data_array, grr_result, evaluation, styles)
    
    # 5. 统计详情页
    ws_stats = wb.create_sheet("统计详情")
    _create_grr_statistics(ws_stats, data_array, grr_result, styles)
    
    # 保存为字节
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.read()


def _create_cover_sheet(ws, study_info: Dict[str, Any], styles):
    """创建封面页"""
    ws.merge_cells('A1:F1')
    ws['A1'] = "MSA 研究报告"
    ws['A1'].font = styles.title_font
    ws['A1'].alignment = styles.center_align
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"报告编号：MSA-{study_info.get('id', 'N/A')}-{datetime.now().strftime('%Y%m%d')}"
    ws['A2'].font = styles.normal_font
    ws['A2'].alignment = styles.center_align
    
    data = [
        ['研究类型:', study_info.get('study_type', 'N/A')],
        ['研究名称:', study_info.get('study_name', 'N/A')],
        ['质量特性:', study_info.get('quality_characteristic', 'N/A')],
        ['零件数量:', study_info.get('n_parts', 'N/A')],
        ['操作者数量:', study_info.get('n_operators', 'N/A')],
        ['试验次数:', study_info.get('n_trials', 'N/A')],
        ['生成日期:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]
    
    for i, (label, value) in enumerate(data, start=4):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = styles.subheading_font
        ws[f'A{i}'].alignment = styles.right_align
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = styles.normal_font
        ws[f'B{i}'].alignment = styles.left_align


def _create_grr_summary(ws, grr_result: Dict[str, float], evaluation: Dict[str, Any], styles):
    """创建 GR&R 结果摘要"""
    # 标题
    ws.merge_cells('A1:E1')
    ws['A1'] = "GR&R 分析结果摘要"
    ws['A1'].font = styles.heading_font
    ws['A1'].fill = styles.header_fill
    ws['A1'].alignment = styles.center_align
    
    # 变差分量表
    headers = ['变差分量', '数值', '百分比', '评估']
    data = [
        ['重复性 (EV)', grr_result['ev'], evaluation['percent_ev'], 
         '<10%' if evaluation['percent_ev'] < 10 else '>30%' if evaluation['percent_ev'] > 30 else '10-30%'],
        ['再现性 (AV)', grr_result['av'], evaluation['percent_av'],
         '<10%' if evaluation['percent_av'] < 10 else '>30%' if evaluation['percent_av'] > 30 else '10-30%'],
        ['GR&R (RR)', grr_result['rr'], evaluation['percent_grr'],
         '<10%' if evaluation['percent_grr'] < 10 else '>30%' if evaluation['percent_grr'] > 30 else '10-30%'],
        ['零件变差 (PV)', grr_result['pv'], evaluation['percent_pv'], ''],
        ['总变差 (TV)', grr_result['tv'], '100%', ''],
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = styles.subheading_font
        cell.fill = styles.header_fill
        cell.alignment = styles.center_align
        cell.border = styles.thin_border
    
    for i, row_data in enumerate(data, start=4):
        fill = styles.alt_fill1 if i % 2 == 0 else styles.alt_fill2
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = styles.normal_font
            cell.alignment = styles.center_align
            cell.border = styles.thin_border
            cell.fill = fill
            
            # 根据 %GRR 设置颜色
            if col == 4 and isinstance(value, str):
                if value == '<10%':
                    cell.fill = styles.good_fill
                elif value == '10-30%':
                    cell.fill = styles.warning_fill
                elif value == '>30%':
                    cell.fill = styles.bad_fill
    
    # 评估结论
    ws.merge_cells('A8:E8')
    ws['A8'] = f"评估结论：{evaluation.get('acceptance', 'N/A').upper()}"
    ws['A8'].font = styles.subheading_font
    ws['A8'].alignment = styles.left_align
    
    ws['A9'] = f"ndc (分级数): {evaluation.get('ndc', 'N/A')}"
    ws['A9'].font = styles.normal_font
    
    ws['A10'] = f"判定：{'✅ 合格 (ndc≥5)' if evaluation.get('ndc', 0) >= 5 else '❌ 建议改进 (ndc<5)'}"
    ws['A10'].font = styles.normal_font


def _create_grr_data(ws, data_array: np.ndarray, styles):
    """创建原始数据表"""
    n_parts, n_operators, n_trials = data_array.shape
    
    # 表头
    headers = ['零件', '操作者'] + [f'第{i+1}次' for i in range(n_trials)] + ['平均值', '极差', '标准差']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles.subheading_font
        cell.fill = styles.header_fill
        cell.alignment = styles.center_align
        cell.border = styles.thin_border
    
    # 数据行
    row_num = 2
    for i in range(n_parts):
        for j in range(n_operators):
            values = [data_array[i, j, k] for k in range(n_trials)]
            row = [f'零件{i+1}', f'操作者{j+1}'] + values + [
                np.mean(values),
                max(values) - min(values),
                np.std(values, ddof=1) if len(values) > 1 else 0
            ]
            
            fill = styles.alt_fill1 if row_num % 2 == 0 else styles.alt_fill2
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.font = styles.normal_font
                cell.alignment = styles.center_align
                cell.border = styles.thin_border
                cell.fill = fill
                if isinstance(value, float):
                    cell.number_format = '0.0000'
            
            row_num += 1
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12


def _create_grr_charts(ws, data_array: np.ndarray, grr_result: Dict, evaluation: Dict, styles):
    """创建分析图表（简化版，兼容移动端）"""
    n_parts, n_operators, n_trials = data_array.shape
    
    # 在 Sheet 中创建图表数据区域（从 K 列开始）
    # 零件平均值数据
    ws['K1'] = "零件平均值"
    ws['K1'].font = styles.subheading_font
    ws['L1'] = "平均值"
    ws['L1'].font = styles.subheading_font
    
    for i in range(n_parts):
        mean_val = float(np.mean(data_array[i, :, :]))
        ws.cell(row=i+2, column=11, value=f"零件{i+1}")  # K 列
        ws.cell(row=i+2, column=12, value=mean_val)       # L 列
    
    # %GRR 数据（从 N 列开始）
    ws['N1'] = "变差百分比"
    ws['N1'].font = styles.subheading_font
    ws['O1'] = "数值"
    ws['O1'].font = styles.subheading_font
    
    metrics = ['%EV', '%AV', '%GRR', '%PV']
    values = [
        float(evaluation['percent_ev']),
        float(evaluation['percent_av']),
        float(evaluation['percent_grr']),
        float(evaluation['percent_pv'])
    ]
    
    for i, (metric, val) in enumerate(zip(metrics, values)):
        ws.cell(row=i+2, column=14, value=metric)  # N 列
        ws.cell(row=i+2, column=15, value=val)     # O 列
    
    # 创建图表 1: 零件平均值图
    try:
        chart1 = LineChart()
        chart1.title = "零件平均值图"
        chart1.x_axis.title = "零件编号"
        chart1.y_axis.title = "平均值"
        
        # 数据范围：L2:L11
        data = Reference(ws, min_col=12, min_row=1, max_row=n_parts+1)
        # 分类范围：K2:K11
        cats = Reference(ws, min_col=11, min_row=2, max_row=n_parts+1)
        
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.style = 13
        chart1.shape = 1
        
        ws.add_chart(chart1, "A3")
    except Exception as e:
        ws['A3'] = f"零件平均值图 (图表渲染失败): {[round(float(np.mean(data_array[i, :, :])), 4) for i in range(n_parts)]}"
        ws['A3'].font = styles.normal_font
    
    # 创建图表 2: %GRR 柱状图
    try:
        chart2 = BarChart()
        chart2.title = "变差百分比"
        chart2.x_axis.title = "变差类型"
        chart2.y_axis.title = "百分比 (%)"
        
        # 数据范围：O2:O5
        data = Reference(ws, min_col=15, min_row=1, max_row=5)
        # 分类范围：N2:N5
        cats = Reference(ws, min_col=14, min_row=2, max_row=5)
        
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.style = 13
        chart2.shape = 1
        
        ws.add_chart(chart2, "A20")
    except Exception as e:
        ws['A20'] = f"变差百分比 (图表渲染失败): %EV={evaluation['percent_ev']:.1f}%, %AV={evaluation['percent_av']:.1f}%, %GRR={evaluation['percent_grr']:.1f}%, %PV={evaluation['percent_pv']:.1f}%"
        ws['A20'].font = styles.normal_font


def _create_grr_statistics(ws, data_array: np.ndarray, grr_result: Dict, styles):
    """创建统计详情表"""
    n_parts, n_operators, n_trials = data_array.shape
    
    # 基础统计
    headers = ['统计量', '值']
    stats = [
        ['总测量次数', n_parts * n_operators * n_trials],
        ['零件平均值', np.mean(data_array)],
        ['零件标准差', np.std(data_array, ddof=1)],
        ['零件内变差', np.mean([np.std(data_array[i, :, :], ddof=1) for i in range(n_parts)])],
        ['零件间变差', np.std([np.mean(data_array[i, :, :]) for i in range(n_parts)], ddof=1)],
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles.subheading_font
        cell.fill = styles.header_fill
        cell.alignment = styles.center_align
        cell.border = styles.thin_border
    
    for i, (stat, value) in enumerate(stats, start=2):
        ws.cell(row=i, column=1, value=stat).font = styles.normal_font
        ws.cell(row=i, column=2, value=value).font = styles.normal_font
        ws.cell(row=i, column=2).number_format = '0.0000'


# ==================== 偏倚 Excel 报告 ====================

def create_bias_excel(
    study_info: Dict[str, Any],
    bias_result: Dict[str, float],
    measurements: list
) -> bytes:
    """创建偏倚 Excel 报告"""
    wb = Workbook()
    styles = StyleManager()
    
    # 1. 封面页
    ws_cover = wb.active
    ws_cover.title = "封面"
    _create_cover_sheet(ws_cover, study_info, styles)
    
    # 2. 结果摘要页
    ws_summary = wb.create_sheet("结果摘要")
    _create_bias_summary(ws_summary, bias_result, styles)
    
    # 3. 原始数据页
    ws_data = wb.create_sheet("原始数据")
    _create_bias_data(ws_data, measurements, bias_result.get('reference_value', 0), styles)
    
    # 保存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _create_bias_summary(ws, bias_result: Dict[str, float], styles):
    """创建偏倚结果摘要"""
    ws.merge_cells('A1:D1')
    ws['A1'] = "偏倚分析结果摘要"
    ws['A1'].font = styles.heading_font
    ws['A1'].fill = styles.header_fill
    ws['A1'].alignment = styles.center_align
    
    data = [
        ['参考值', f"{bias_result.get('reference_value', 0):.4f}"],
        ['测量次数', str(bias_result.get('n_measurements', 0))],
        ['偏倚 (Bias)', f"{bias_result.get('bias', 0):.6f}"],
        ['偏倚百分比', f"{bias_result.get('bias_percent', 0):.2f}%"],
        ['t 统计量', f"{bias_result.get('t_statistic', 0):.3f}"],
        ['p 值', f"{bias_result.get('p_value', 0):.4f}"],
        ['显著性', '是' if bias_result.get('is_significant', False) else '否'],
    ]
    
    for i, (label, value) in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=label).font = styles.normal_font
        ws.cell(row=i, column=2, value=value).font = styles.normal_font
        ws.cell(row=i, column=1).border = styles.thin_border
        ws.cell(row=i, column=2).border = styles.thin_border


def _create_bias_data(ws, measurements: list, reference_value: float, styles):
    """创建偏倚原始数据"""
    headers = ['序号', '测量值', '偏倚 (测量值 - 参考值)']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles.subheading_font
        cell.fill = styles.header_fill
        cell.alignment = styles.center_align
        cell.border = styles.thin_border
    
    for i, val in enumerate(measurements, start=2):
        bias = val - reference_value
        ws.cell(row=i, column=1, value=i-1).font = styles.normal_font
        ws.cell(row=i, column=2, value=val).font = styles.normal_font
        ws.cell(row=i, column=3, value=bias).font = styles.normal_font
        ws.cell(row=i, column=2).number_format = '0.0000'
        ws.cell(row=i, column=3).number_format = '+0.0000;-0.0000'
        
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = styles.thin_border
    
    # 统计行
    row = len(measurements) + 2
    ws.cell(row=row, column=1, value='统计').font = styles.subheading_font
    ws.cell(row=row, column=2, value=f'平均值={np.mean(measurements):.4f}').font = styles.normal_font
    ws.cell(row=row, column=3, value=f'标准差={np.std(measurements, ddof=1):.4f}').font = styles.normal_font


# ==================== 线性 Excel 报告 ====================

def create_linearity_excel(
    study_info: Dict[str, Any],
    linearity_result: Dict[str, float],
    reference_values: np.ndarray,
    measurements: np.ndarray
) -> bytes:
    """创建线性 Excel 报告"""
    wb = Workbook()
    styles = StyleManager()
    
    # 1. 封面页
    ws_cover = wb.active
    ws_cover.title = "封面"
    _create_cover_sheet(ws_cover, study_info, styles)
    
    # 2. 结果摘要页
    ws_summary = wb.create_sheet("结果摘要")
    _create_linearity_summary(ws_summary, linearity_result, styles)
    
    # 3. 原始数据页
    ws_data = wb.create_sheet("原始数据")
    _create_linearity_data(ws_data, reference_values, measurements, styles)
    
    # 保存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _create_linearity_summary(ws, linearity_result: Dict[str, float], styles):
    """创建线性结果摘要"""
    ws.merge_cells('A1:D1')
    ws['A1'] = "线性分析结果摘要"
    ws['A1'].font = styles.heading_font
    ws['A1'].fill = styles.header_fill
    ws['A1'].alignment = styles.center_align
    
    data = [
        ['斜率 (Slope)', f"{linearity_result.get('slope', 0):.6f}"],
        ['截距 (Intercept)', f"{linearity_result.get('intercept', 0):.6f}"],
        ['R 平方', f"{linearity_result.get('r_squared', 0):.4f}"],
        ['斜率 p 值', f"{linearity_result.get('p_value_slope', 0):.4f}"],
        ['截距 p 值', f"{linearity_result.get('p_value_intercept', 0):.4f}"],
        ['线性度', f"{linearity_result.get('linearity', 0):.6f}"],
        ['线性接受', '是' if linearity_result.get('is_linear', False) else '否'],
    ]
    
    for i, (label, value) in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=label).font = styles.normal_font
        ws.cell(row=i, column=2, value=value).font = styles.normal_font
        ws.cell(row=i, column=1).border = styles.thin_border
        ws.cell(row=i, column=2).border = styles.thin_border


def _create_linearity_data(ws, reference_values: np.ndarray, measurements: np.ndarray, styles):
    """创建线性原始数据"""
    headers = ['零件编号', '参考值', '测量值', '偏倚 (测量值 - 参考值)']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles.subheading_font
        cell.fill = styles.header_fill
        cell.alignment = styles.center_align
        cell.border = styles.thin_border
    
    for i, (ref, meas) in enumerate(zip(reference_values, measurements), start=2):
        bias = meas - ref
        ws.cell(row=i, column=1, value=i-1).font = styles.normal_font
        ws.cell(row=i, column=2, value=ref).font = styles.normal_font
        ws.cell(row=i, column=3, value=meas).font = styles.normal_font
        ws.cell(row=i, column=4, value=bias).font = styles.normal_font
        
        for col in range(2, 5):
            ws.cell(row=i, column=col).number_format = '0.0000'
            ws.cell(row=i, column=col).border = styles.thin_border


# ==================== 主函数 ====================

def generate_msa_excel(
    study_type: str,
    study_info: Dict[str, Any],
    results: Dict[str, Any],
    chart_data: Optional[Dict[str, Any]] = None
) -> bytes:
    """生成 MSA Excel 报告"""
    
    if study_type == 'GRR':
        return create_grr_excel(
            study_info=study_info,
            grr_result=results.get('grr_error', {}),
            evaluation=results.get('evaluation', {}),
            data_array=chart_data.get('data_array', np.array([]))
        )
    
    elif study_type == 'BIAS':
        return create_bias_excel(
            study_info=study_info,
            bias_result=results.get('bias_result', {}),
            measurements=chart_data.get('measurements', [])
        )
    
    elif study_type == 'LINEARITY':
        return create_linearity_excel(
            study_info=study_info,
            linearity_result=results.get('linearity_result', {}),
            reference_values=chart_data.get('reference_values', np.array([])),
            measurements=chart_data.get('measurements', np.array([]))
        )
    
    else:
        raise ValueError(f"不支持的研究类型：{study_type}")
