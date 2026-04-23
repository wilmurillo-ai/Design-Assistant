import json
import yaml
import os
import shutil
import zlib
from datetime import datetime
from lake.lake_handle import MyParser, MyContext, remove_invalid_characters
from lake.lake_reader import unpack_lake_book_file

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ConversionReport:
    """转换报告类，收集转换过程中的详细信息"""

    def __init__(self, lakebook_name):
        self.lakebook_name = lakebook_name
        self.start_time = datetime.now()
        self.end_time = None

        # 文档统计
        self.total_docs = 0
        self.success_docs = []  # 成功转换的文档 [(title, type, output_path)]
        self.failed_docs = []  # 转换失败的文档 [(title, type, error)]
        self.skipped_docs = []  # 跳过的文档 [(title, reason)]

        # 按类型统计
        self.lake_count = 0
        self.laketable_count = 0
        self.lakesheet_count = 0

        # 资源下载统计
        self.images_downloaded = []  # 成功下载的图片 [(name, url, local_path)]
        self.images_failed = []  # 下载失败的图片 [(name, url, error)]
        self.files_downloaded = []  # 成功下载的附件 [(name, url, local_path)]
        self.files_failed = []  # 下载失败的附件 [(name, url, error)]
        self.expired_links = []  # 过期的链接 [(name, url)]

        # 特殊内容统计
        self.encrypted_content = []  # 加密内容 [(doc_title, count)]
        self.cross_doc_links = []  # 跨文档引用 [(doc_title, target)]

    def add_success(self, title, doc_type, output_path, record_count=None):
        self.success_docs.append(
            {
                "title": title,
                "type": doc_type,
                "path": output_path,
                "records": record_count,
            }
        )
        if doc_type == "lake":
            self.lake_count += 1
        elif doc_type == "laketable":
            self.laketable_count += 1
        elif doc_type == "lakesheet":
            self.lakesheet_count += 1

    def add_failure(self, title, doc_type, error):
        self.failed_docs.append({"title": title, "type": doc_type, "error": str(error)})

    def add_image_success(self, name, url, local_path):
        self.images_downloaded.append({"name": name, "url": url, "path": local_path})

    def add_image_failure(self, name, url, error=None):
        self.images_failed.append(
            {
                "name": name,
                "url": url,
                "error": str(error) if error else "Unknown error",
            }
        )

    def add_file_success(self, name, url, local_path):
        self.files_downloaded.append({"name": name, "url": url, "path": local_path})

    def add_file_failure(self, name, url, error=None):
        self.files_failed.append(
            {
                "name": name,
                "url": url,
                "error": str(error) if error else "Unknown error",
            }
        )

    def add_expired_link(self, name, url):
        self.expired_links.append({"name": name, "url": url})

    def add_encrypted(self, doc_title, count=1):
        self.encrypted_content.append({"doc": doc_title, "count": count})

    def finish(self):
        self.end_time = datetime.now()

    def generate_report(self):
        """生成 Markdown 格式的报告"""
        duration = (
            (self.end_time - self.start_time).total_seconds() if self.end_time else 0
        )

        lines = []
        lines.append(f"# 语雀文档转换报告")
        lines.append(f"")
        lines.append(f"**笔记本**: {self.lakebook_name}")
        lines.append(f"**转换时间**: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"**耗时**: {duration:.1f} 秒")
        lines.append(f"")

        # 总体统计
        lines.append(f"## 1. 总体统计")
        lines.append(f"")
        lines.append(f"| 项目 | 数量 |")
        lines.append(f"| --- | --- |")
        lines.append(f"| 总文档数 | {self.total_docs} |")
        lines.append(f"| 成功转换 | {len(self.success_docs)} |")
        lines.append(f"| 转换失败 | {len(self.failed_docs)} |")
        lines.append(f"| 富文本文档 (lake) | {self.lake_count} |")
        lines.append(f"| 数据库表格 (laketable) | {self.laketable_count} |")
        lines.append(f"| 电子表格 (lakesheet) | {self.lakesheet_count} |")
        lines.append(f"")

        # 成功转换的文档
        if self.success_docs:
            lines.append(f"## 2. 成功转换的文档")
            lines.append(f"")
            lines.append(f"| 文档标题 | 类型 | 输出路径 |")
            lines.append(f"| --- | --- | --- |")
            for doc in self.success_docs:
                title = doc["title"]
                doc_type = doc["type"]
                path = doc["path"]
                lines.append(f"| {title} | {doc_type} | {path} |")
            lines.append(f"")

        # 转换失败的文档
        if self.failed_docs:
            lines.append(f"## 3. 转换失败的文档 ⚠️")
            lines.append(f"")
            lines.append(f"以下文档转换失败，需要手动检查：")
            lines.append(f"")
            lines.append(f"| 文档标题 | 类型 | 错误信息 |")
            lines.append(f"| --- | --- | --- |")
            for doc in self.failed_docs:
                title = doc["title"]
                doc_type = doc["type"]
                error = doc["error"]
                lines.append(f"| {title} | {doc_type} | {error} |")
            lines.append(f"")

        # 图片下载统计
        total_images = len(self.images_downloaded) + len(self.images_failed)
        if total_images > 0:
            lines.append(f"## 4. 图片下载统计")
            lines.append(f"")
            lines.append(f"| 项目 | 数量 |")
            lines.append(f"| --- | --- |")
            lines.append(f"| 总图片数 | {total_images} |")
            lines.append(f"| 下载成功 | {len(self.images_downloaded)} |")
            lines.append(f"| 下载失败 | {len(self.images_failed)} |")
            lines.append(f"")

            if self.images_failed:
                lines.append(f"### 下载失败的图片 ⚠️")
                lines.append(f"")
                lines.append(f"以下图片下载失败，需要手动处理：")
                lines.append(f"")
                lines.append(f"| 图片名称 | URL | 错误原因 |")
                lines.append(f"| --- | --- | --- |")
                for img in self.images_failed[:20]:  # 最多显示20条
                    name = img["name"]
                    url = (
                        img["url"][:80] + "..." if len(img["url"]) > 80 else img["url"]
                    )
                    error = img["error"]
                    lines.append(f"| {name} | {url} | {error} |")
                if len(self.images_failed) > 20:
                    lines.append(
                        f"| ... | ... | 还有 {len(self.images_failed) - 20} 条 |"
                    )
                lines.append(f"")

        # 附件下载统计
        total_files = len(self.files_downloaded) + len(self.files_failed)
        if total_files > 0:
            lines.append(f"## 5. 附件下载统计")
            lines.append(f"")
            lines.append(f"| 项目 | 数量 |")
            lines.append(f"| --- | --- |")
            lines.append(f"| 总附件数 | {total_files} |")
            lines.append(f"| 下载成功 | {len(self.files_downloaded)} |")
            lines.append(f"| 下载失败 | {len(self.files_failed)} |")
            lines.append(f"")

            if self.files_failed:
                lines.append(f"### 下载失败的附件 ⚠️")
                lines.append(f"")
                lines.append(f"| 附件名称 | URL | 错误原因 |")
                lines.append(f"| --- | --- | --- |")
                for f in self.files_failed[:20]:
                    lines.append(f"| {f['name']} | {f['url'][:80]}... | {f['error']} |")
                lines.append(f"")

        # 过期链接
        if self.expired_links:
            lines.append(f"## 6. 过期链接 ⚠️")
            lines.append(f"")
            lines.append(f"以下链接已过期，无法下载：")
            lines.append(f"")
            lines.append(f"| 名称 | URL |")
            lines.append(f"| --- | --- |")
            for link in self.expired_links[:20]:
                lines.append(f"| {link['name']} | {link['url'][:80]}... |")
            lines.append(f"")

        # 加密内容
        if self.encrypted_content:
            lines.append(f"## 7. 加密内容")
            lines.append(f"")
            lines.append(f"以下文档包含加密内容（密钥未导出，无法解密）：")
            lines.append(f"")
            lines.append(f"| 文档标题 | 加密内容数量 |")
            lines.append(f"| --- | --- |")
            for item in self.encrypted_content:
                lines.append(f"| {item['doc']} | {item['count']} |")
            lines.append(f"")

        # 需要手动处理的内容汇总
        needs_manual = []
        if self.failed_docs:
            needs_manual.append(f"转换失败的文档 ({len(self.failed_docs)} 个)")
        if self.images_failed:
            needs_manual.append(f"下载失败的图片 ({len(self.images_failed)} 张)")
        if self.files_failed:
            needs_manual.append(f"下载失败的附件 ({len(self.files_failed)} 个)")
        if self.expired_links:
            needs_manual.append(f"过期的链接 ({len(self.expired_links)} 个)")
        if self.encrypted_content:
            needs_manual.append(
                f"加密内容 ({sum(x['count'] for x in self.encrypted_content)} 处)"
            )

        if needs_manual:
            lines.append(f"## 8. 需要手动处理的内容 📋")
            lines.append(f"")
            for item in needs_manual:
                lines.append(f"- {item}")
            lines.append(f"")

        # 报告结尾
        lines.append(f"---")
        lines.append(f"")
        lines.append(f"*报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*")

        return "\n".join(lines)


class GlobalContext:
    parent_id_and_child: dict = {}
    id_and_book = {}
    root_books = []
    file_count = 0
    all_file_count = 0
    failure_image_download_list = []
    file_total = 0
    total = 0
    download_image = True
    skip_existing = False
    root_path = ""
    report = None  # 转换报告对象


def load_meta_json(global_context: GlobalContext):
    full_path = "/".join([global_context.root_path, "$meta.json"])
    fp = open(full_path, "r+", encoding="utf-8")
    json_obj = json.load(fp)
    meta = json_obj["meta"]
    meta_obj = json.loads(meta)
    book_yml = meta_obj["book"]["tocYml"]
    books = yaml.load(book_yml, yaml.Loader)
    for book in books:
        if book.get("uuid"):
            global_context.id_and_book[book["uuid"]] = book
        if book["type"] == "META":
            continue
        if book["parent_uuid"] == "":
            global_context.root_books.append(book)
            continue
        parent_uuid = book["parent_uuid"]
        if global_context.parent_id_and_child.get(parent_uuid):
            global_context.parent_id_and_child[parent_uuid].append(book)
        else:
            global_context.parent_id_and_child[parent_uuid] = []
            global_context.parent_id_and_child[parent_uuid].append(book)
    global_context.file_total = len(global_context.id_and_book)
    global_context.total = len(global_context.id_and_book)


def create_tree_dir(global_context, parent_path, book):
    if book is None:
        return
    uuid = book["uuid"]
    name = book["title"].replace("/", "_")
    file_url = book["url"]
    if not os.path.exists(parent_path):
        parent_path = remove_invalid_characters(parent_path)
        os.makedirs(parent_path, exist_ok=True)
    book_children = global_context.parent_id_and_child.get(uuid)
    global_context.all_file_count += 1
    if file_url != "":
        ltm = LakeToMd(
            "{}/{}.json".format(global_context.root_path, file_url),
            target=os.path.join(parent_path, name),
        )
        ltm.to_md(global_context)
        global_context.failure_image_download_list += ltm.image_download_failure
        global_context.file_count += 1
        print(
            "\r转换进度: {}/{}/{}. ".format(
                global_context.file_count,
                global_context.all_file_count,
                global_context.file_total,
            ),
            end="",
        )
    if not book_children:
        return
    for child in book_children:
        child_title = child["title"].replace("/", "_")
        child_path = os.path.join(parent_path, child_title)
        create_tree_dir(global_context, child_path, child)


def format_cell_value(value, col_type="text"):
    if isinstance(value, dict):
        if "text" in value:
            return str(value["text"])
        if "class" in value and value["class"] == "formula":
            v = value.get("value", "")
            if v is not None and v != "":
                try:
                    fv = float(v)
                    return f"{fv:.2f}".rstrip("0").rstrip(".")
                except (ValueError, TypeError):
                    return str(v)
            return value.get("formula", "")
        return str(value.get("value", ""))
    if isinstance(value, str):
        try:
            fv = float(value)
            if fv == int(fv) and "." not in value.rstrip("0").rstrip("."):
                return str(int(fv))
            return f"{fv:.2f}".rstrip("0").rstrip(".")
        except (ValueError, TypeError):
            pass
    return str(value) if value is not None else ""


def convert_laketable_to_md(doc_data, report=None, doc_title=""):
    doc = doc_data["doc"]
    body_str = doc.get("body", "")
    if not body_str:
        return "", 0
    try:
        body = json.loads(body_str)
    except json.JSONDecodeError as e:
        if report:
            report.add_failure(doc_title, "laketable", f"JSON解析失败: {str(e)}")
        return "", 0
    sheets = body.get("sheet", [])
    if not sheets:
        return "", 0
    lines = []
    record_count = 0
    for sheet in sheets:
        columns = sheet.get("columns", [])
        if not columns:
            continue
        col_names = [c["name"] for c in columns]
        col_ids = [c["id"] for c in columns]
        col_types = {c["id"]: c.get("type", "text") for c in columns}
        # Build option ID -> display value mapping for select/multi_select columns
        col_options = {}
        for c in columns:
            if c.get("type") in ("select", "multi_select"):
                opts = c.get("options", [])
                col_options[c["id"]] = {o["id"]: o.get("value", o["id"]) for o in opts}
        lines.append("| " + " | ".join(col_names) + " |")
        lines.append("| " + " | ".join(["---"] * len(col_names)) + " |")
        records = doc.get("table_records", [])
        record_count = len(records)
        for rec in records:
            data = json.loads(rec["data"])
            row = []
            for cid in col_ids:
                cell = data.get(cid, {})
                val = cell.get("value", "") if isinstance(cell, dict) else cell
                # Resolve select/multi_select option IDs to display values
                if cid in col_options and val:
                    opt_map = col_options[cid]
                    if isinstance(val, list):
                        val = [opt_map.get(v, v) for v in val]
                    elif isinstance(val, str):
                        val = opt_map.get(val, val)
                row.append(format_cell_value(val, col_types.get(cid, "text")))
            lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines) + "\n" if lines else "", record_count


def convert_lakesheet_to_md(doc_data, report=None, doc_title=""):
    doc = doc_data["doc"]
    body_draft_str = doc.get("body_draft", "")
    if not body_draft_str:
        return "", 0
    try:
        bd = json.loads(body_draft_str)
    except json.JSONDecodeError as e:
        if report:
            report.add_failure(doc_title, "lakesheet", f"JSON解析失败: {str(e)}")
        return "", 0
    sheet_raw = bd.get("sheet", "")
    if not sheet_raw:
        return "", 0
    try:
        sheet_bytes = (
            sheet_raw.encode("latin-1") if isinstance(sheet_raw, str) else sheet_raw
        )
        decompressed = zlib.decompress(sheet_bytes)
        sheets = json.loads(decompressed)
    except Exception as e:
        if report:
            report.add_failure(doc_title, "lakesheet", f"zlib解压失败: {str(e)}")
        return "", 0
    if not sheets:
        return "", 0
    lines = []
    row_count = 0
    for sheet in sheets:
        data = sheet.get("data", {})
        if not data:
            continue
        max_row = 0
        max_col = 0
        for row_idx in data:
            ri = int(row_idx)
            for col_idx in data[row_idx]:
                ci = int(col_idx)
                cell = data[row_idx][col_idx]
                v = cell.get("v")
                if v and not (
                    isinstance(v, dict)
                    and v.get("class") == "formula"
                    and v.get("value") in (None, 0, "")
                ):
                    max_row = max(max_row, ri)
                    max_col = max(max_col, ci)
        row_count = max_row + 1
        for ri in range(max_row + 1):
            row = data.get(str(ri), {})
            cells = []
            for ci in range(max_col + 1):
                cell = row.get(str(ci), {})
                val = cell.get("v", "")
                cells.append(format_cell_value(val))
            lines.append("| " + " | ".join(cells) + " |")
            if ri == 0:
                lines.append("| " + " | ".join(["---"] * (max_col + 1)) + " |")
    return "\n".join(lines) + "\n" if lines else "", row_count


def convert_laketable_to_excel(doc_data, output_path, report=None, doc_title=""):
    """将laketable转换为Excel文件"""
    if not HAS_OPENPYXL:
        return False, "openpyxl未安装"

    doc = doc_data["doc"]
    body_str = doc.get("body", "")
    if not body_str:
        return False, "body为空"

    try:
        body = json.loads(body_str)
    except json.JSONDecodeError as e:
        return False, f"JSON解析失败: {str(e)}"

    sheets = body.get("sheet", [])
    if not sheets:
        return False, "没有sheet数据"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "数据"

        # 定义样式
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        record_count = 0
        for sheet in sheets:
            columns = sheet.get("columns", [])
            if not columns:
                continue

            # 获取列信息
            col_names = [c["name"] for c in columns]
            col_ids = [c["id"] for c in columns]
            col_types = {c["id"]: c.get("type", "text") for c in columns}

            # 构建select/multi_select选项映射
            col_options = {}
            for c in columns:
                if c.get("type") in ("select", "multi_select"):
                    opts = c.get("options", [])
                    col_options[c["id"]] = {
                        o["id"]: o.get("value", o["id"]) for o in opts
                    }

            # 写入表头
            for col_idx, col_name in enumerate(col_names, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            # 写入数据
            records = doc.get("table_records", [])
            record_count = len(records)

            for row_idx, rec in enumerate(records, 2):
                data = json.loads(rec["data"])
                for col_idx, cid in enumerate(col_ids, 1):
                    cell_data = data.get(cid, {})
                    val = (
                        cell_data.get("value", "")
                        if isinstance(cell_data, dict)
                        else cell_data
                    )

                    # 解析select/multi_select选项
                    if cid in col_options and val:
                        opt_map = col_options[cid]
                        if isinstance(val, list):
                            val = ", ".join([opt_map.get(v, v) for v in val])
                        elif isinstance(val, str):
                            val = opt_map.get(val, val)

                    # 处理日期类型
                    if col_types.get(cid) == "date" and isinstance(val, dict):
                        val = val.get("text", "")

                    # 处理数字类型
                    if col_types.get(cid) == "number" and val:
                        try:
                            val = float(val)
                        except (ValueError, TypeError):
                            pass

                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border

            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        # 保存文件
        wb.save(output_path)
        return True, record_count

    except Exception as e:
        return False, str(e)


class LakeToMd:
    body_html = None
    image_download_failure = []

    def __init__(self, filename, target):
        self.filename = filename
        self.target = target
        self.doc_title = os.path.basename(target)
        self.doc_type = "lake"
        self.record_count = None
        self.__body_html()

    def __body_html(self):
        fp = open(file=self.filename, mode="r+", encoding="utf-8")
        file_json = json.load(fp)
        doc = file_json.get("doc", {})
        doc_format = doc.get("format", "")
        doc_type = doc.get("type", "")
        if doc_format == "laketable" or doc_type == "Table":
            self.body_html = None
            self.table_data = file_json
            self.sheet_data = None
            self.doc_type = "laketable"
        elif doc_format == "lakesheet" or doc_type == "Sheet":
            self.body_html = None
            self.table_data = None
            self.sheet_data = file_json
            self.doc_type = "lakesheet"
        else:
            self.body_html = doc.get("body_draft_asl", "")
            if not self.body_html:
                self.body_html = doc.get("body_draft", "")
            self.table_data = None
            self.sheet_data = None
            self.doc_type = "lake"
        fp.close()

    def to_md(self, global_context):
        target = self.target.replace("/", os.path.sep).replace("\\", os.path.sep)
        target = remove_invalid_characters(target)
        os.makedirs(os.path.dirname(target), exist_ok=True)
        report = global_context.report
        output_files = []  # 记录生成的文件

        try:
            if self.table_data is not None:
                # laketable: 生成Markdown和Excel
                content, record_count = convert_laketable_to_md(
                    self.table_data, report, self.doc_title
                )
                self.record_count = record_count

                # 同时生成Excel文件
                excel_path = target + ".xlsx"
                success, result = convert_laketable_to_excel(
                    self.table_data, excel_path, report, self.doc_title
                )
                if success:
                    output_files.append(
                        os.path.relpath(excel_path, global_context.output_path)
                    )
                    if report:
                        report.add_file_success(
                            self.doc_title + ".xlsx",
                            "local",
                            os.path.relpath(excel_path, global_context.output_path),
                        )

            elif self.sheet_data is not None:
                content, row_count = convert_lakesheet_to_md(
                    self.sheet_data, report, self.doc_title
                )
                self.record_count = row_count

            elif self.body_html:
                name = os.path.basename(target)
                short_target = os.path.dirname(target)
                context = MyContext(
                    filename=name,
                    image_target=short_target,
                    download_image=global_context.download_image,
                    skip_existing=global_context.skip_existing,
                )
                mp = MyParser(self.body_html)
                content = mp.handle_descent(mp.soup, context)
                self.image_download_failure += context.failure_images

                # 记录图片下载信息
                if report:
                    for img_fail in context.failure_images:
                        # 格式: [name]url
                        if img_fail.startswith("[") and "]" in img_fail:
                            bracket_end = img_fail.index("]")
                            img_name = img_fail[1:bracket_end]
                            img_url = img_fail[bracket_end + 1 :]
                            report.add_image_failure(img_name, img_url)

                # 记录附件下载信息
                if report:
                    for f in context.downloaded_files:
                        report.add_file_success(f["name"], f["url"], f["path"])
                    for f in context.failed_files:
                        report.add_file_failure(f["name"], f["url"], f["error"])
                    for link in context.expired_links:
                        report.add_expired_link(link["name"], link["url"])

            else:
                content = ""

            # 写入Markdown文件
            with open(target + ".md", "w+", encoding="utf-8") as fp:
                fp.writelines(content)
                fp.flush()
            output_files.append(
                os.path.relpath(target + ".md", global_context.output_path)
            )

            # 记录成功
            if report:
                report.add_success(
                    self.doc_title,
                    self.doc_type,
                    ", ".join(output_files),
                    self.record_count,
                )

        except Exception as e:
            # 记录失败
            if report:
                report.add_failure(self.doc_title, self.doc_type, str(e))
            raise


def convert_to_md(global_context, file_path):
    output_path = file_path
    for root_book in global_context.root_books:
        title = root_book["title"].replace("/", "_")
        create_tree_dir(global_context, os.path.join(output_path, title), root_book)
    print("\n>>> Markdown 转换完成")


def start_convert(meta, lake_book, output, download_image_of_in, skip_existing=False):
    global_context = GlobalContext()
    temp_dir = "temp"

    # 获取lakebook文件名用于报告
    lakebook_name = os.path.basename(lake_book) if lake_book else "meta.json"

    if lake_book:
        global_context.root_path = unpack_lake_book_file(lake_book, temp_dir)
        print(">>> lake 文件抽取完成")
    else:
        global_context.root_path = meta
    if not global_context.root_path:
        print("参数错误！-i 或 -l 二者必须有一个")
        return
    try:
        load_meta_json(global_context)
        print(">>> meta.json 解析完成")

        # 初始化报告
        global_context.report = ConversionReport(lakebook_name)
        global_context.report.total_docs = global_context.file_total

        global_context.download_image = download_image_of_in
        global_context.skip_existing = skip_existing
        abspath = os.path.abspath(output)
        global_context.output_path = abspath  # 保存输出路径供报告使用

        print(">>> 开始进行 Markdown 转换")
        convert_to_md(global_context, abspath)

        print("共导出 %s 个文件" % global_context.file_count)
        print("图片下载失败列表:")
        print(" list: ", global_context.failure_image_download_list)

        # 生成报告
        global_context.report.finish()
        report_content = global_context.report.generate_report()
        report_path = os.path.join(abspath, "转换报告.md")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(report_content)
        print(f">>> 转换报告已生成: {report_path}")

        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
    except Exception as e:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        print(e)
