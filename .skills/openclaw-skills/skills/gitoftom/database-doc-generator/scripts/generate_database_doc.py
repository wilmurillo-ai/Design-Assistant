#!/usr/bin/env python3
"""
Database Documentation Generator
Generates professional Excel documentation for PostgreSQL database schemas
with formatted output including merged headers and auto-adjusted column widths.
"""

import psycopg2
import pandas as pd
import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def get_all_tables(conn):
    """Get all table names from the database"""
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_type = 'BASE TABLE'
            ORDER BY table_name
        """)
        tables = [row[0] for row in cursor.fetchall()]
        cursor.close()
        return tables
    except Exception as e:
        print(f"[ERROR] Failed to get table list: {e}")
        return []

def get_table_structure(conn, table_name):
    """Get structure information for a specific table"""
    SQL_TEMPLATE = """
    SELECT
     a.column_name AS 代码,
     udt_name AS 数据类型,
     CASE
     WHEN udt_name IN ('varchar', 'character varying') AND COALESCE(character_maximum_length, -999) = -999 THEN '2000'
     WHEN udt_name IN ('timestamp', 'timestamptz', 'time', 'timetz') AND COALESCE(datetime_precision, -999) = -999 THEN '6'
     ELSE
     (CASE
     WHEN COALESCE(character_maximum_length, -999) = -999 THEN 
       CASE 
       WHEN numeric_precision IS NOT NULL THEN numeric_precision::text
       WHEN datetime_precision IS NOT NULL THEN datetime_precision::text
       ELSE NULL
       END
     ELSE character_maximum_length::text
     END)||(CASE
     WHEN udt_name IN ('numeric', 'decimal') THEN ','||COALESCE(numeric_scale::text, '0')
     ELSE '' 
     END)
     END AS 长度,

     CASE
     WHEN is_nullable = 'YES' THEN 'FALSE'
     ELSE 'TRUE'
     END AS 强制,
     COALESCE(col_description(b.oid, a.ordinal_position), '') AS 注释
    FROM
     information_schema.columns AS a
     LEFT JOIN
     pg_class AS b ON a.table_name = b.relname
     LEFT JOIN
     information_schema.key_column_usage AS pk ON a.table_name = pk.table_name AND a.column_name = pk.column_name
     LEFT JOIN
     information_schema.table_constraints AS tc ON pk.constraint_name = tc.constraint_name
    WHERE
     a.table_name = %s
    ORDER BY
     a.table_schema,
     a.table_name,
     a.ordinal_position;
    """
    
    try:
        cursor = conn.cursor()
        cursor.execute(SQL_TEMPLATE, (table_name,))
        
        # Get column names
        column_names = [desc[0] for desc in cursor.description]
        
        # Get data
        rows = cursor.fetchall()
        
        # Convert to DataFrame
        df = pd.DataFrame(rows, columns=column_names)
        
        # 不再添加表名列，因为表名已经在合并单元格中显示
        # df.insert(0, '表名', table_name)
        
        cursor.close()
        return df
    except Exception as e:
        print(f"[ERROR] Failed to get structure for table '{table_name}': {e}")
        return None

def apply_excel_formatting(writer, table_name, df):
    """Apply professional formatting to Excel worksheet"""
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets[table_name[:31]]  # Excel sheet name max 31 chars
    
    # Define styles
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_font = Font(size=10)
    data_alignment = Alignment(vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Merge cells for table name header (row 1, columns A-F)
    # 现在有5列：代码、数据类型、长度、强制、注释
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    table_name_cell = worksheet.cell(row=1, column=1)
    table_name_cell.value = f"表: {table_name}"
    table_name_cell.font = Font(bold=True, size=14)
    table_name_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply formatting to column headers (row 2)
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=2, column=col_num)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Apply formatting to data rows
    for row in range(3, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Skip the merged header row
        for cell in column[1:]:  # Start from row 2 (headers)
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Add some padding
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Set row height for header
    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[2].height = 25

def generate_database_documentation(db_config, tables=None, output_path=None):
    """
    Generate database documentation
    
    Args:
        db_config: Dictionary with database connection details
        tables: List of table names (None for all tables)
        output_path: Output Excel file path
    """
    print("=" * 70)
    print("Database Documentation Generator")
    print("=" * 70)
    
    # Set default output path
    if output_path is None:
        output_dir = "EXAMPLE_PATH/database_docs"
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"{output_dir}/database_documentation_{timestamp}.xlsx"
    
    # Connect to database with UTF-8 encoding
    try:
        # 添加客户端编码设置
        db_config_with_encoding = db_config.copy()
        db_config_with_encoding['client_encoding'] = 'utf8'
        
        conn = psycopg2.connect(**db_config_with_encoding)
        
        # 设置连接编码为UTF-8
        conn.set_client_encoding('UTF8')
        
        print(f"[OK] Connected to database: {db_config['database']}@{db_config['host']}:{db_config['port']}")
        print(f"[INFO] Connection encoding set to: {conn.encoding}")
    except Exception as e:
        print(f"[ERROR] Database connection failed: {e}")
        return None
    
    # Get table list if not provided
    if tables is None:
        print("\nGetting list of all tables...")
        tables = get_all_tables(conn)
        if not tables:
            print("[WARN] No tables found in database")
            conn.close()
            return None
        print(f"[OK] Found {len(tables)} tables")
    
    # Process tables
    print(f"\nProcessing {len(tables)} tables...")
    all_data = {}
    success_count = 0
    fail_count = 0
    
    for table in tables:
        print(f"\n  Table: {table}")
        df = get_table_structure(conn, table)
        
        if df is not None and not df.empty:
            all_data[table] = df
            success_count += 1
            print(f"    [OK] Retrieved {len(df)} columns")
        else:
            fail_count += 1
            print(f"    [ERROR] Failed to retrieve structure")
    
    # Close database connection
    conn.close()
    
    # Save to Excel with formatting
    if not all_data:
        print("\n[ERROR] No data to export")
        return None
    
    print(f"\nSaving to Excel file: {output_path}")
    
    try:
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for table_name, df in all_data.items():
                # Write DataFrame to Excel
                sheet_name = table_name[:31]  # Excel sheet name limit
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                
                # Apply formatting
                apply_excel_formatting(writer, table_name, df)
                
                print(f"  [OK] Table '{table_name}' saved to sheet '{sheet_name}'")
        
        print(f"\n[SUCCESS] Documentation saved to: {output_path}")
        
        # Generate summary
        print("\n" + "=" * 70)
        print("Generation Summary:")
        print("=" * 70)
        print(f"Total tables processed: {len(tables)}")
        print(f"Successfully exported: {success_count}")
        print(f"Failed: {fail_count}")
        print(f"Output file: {output_path}")
        print(f"File size: {os.path.getsize(output_path):,} bytes")
        print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        if fail_count > 0:
            print("\nFailed tables:")
            for table in tables:
                if table not in all_data:
                    print(f"  - {table}")
        
        return output_path
        
    except Exception as e:
        print(f"[ERROR] Failed to save Excel file: {e}")
        return None

def main():
    """
    Main function for command-line usage.
    
    SECURITY NOTE: This function does NOT contain any database credentials.
    All credentials must be provided via command-line arguments or environment variables.
    """
    import argparse
    import os
    
    parser = argparse.ArgumentParser(
        description='Generate database documentation from PostgreSQL',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
SECURITY REQUIREMENTS:
  1. NEVER hardcode credentials in source code
  2. Use environment variables for production
  3. Store configuration files with restricted permissions
  4. Use SSL/TLS for database connections

EXAMPLES:
  # Using environment variables (RECOMMENDED)
  export DB_HOST=localhost DB_PORT=5432 DB_NAME=mydb DB_USER=EXAMPLE_USER DB_PASSWORD=secret
  python generate_database_doc.py
  
  # Using command-line arguments
  python generate_database_doc.py --host localhost --port 5432 --database mydb --user EXAMPLE_USER --EXAMPLE_PASSWORD secret
  
  # Using configuration file
  python generate_database_doc.py --config /path/to/secure_config.json
  
  # Using environment variables with specific tables
  export DB_TABLES=users,orders,products
  python generate_database_doc.py
"""
    )
    
    # Command-line arguments
    parser.add_argument('--host', help='Database host (or use DB_HOST env var)')
    parser.add_argument('--port', type=int, default=5432, help='Database port (default: 5432)')
    parser.add_argument('--database', help='Database name (or use DB_NAME env var)')
    parser.add_argument('--user', help='Database user (or use DB_USER env var)')
    parser.add_argument('--EXAMPLE_PASSWORD', help='Database EXAMPLE_PASSWORD (or use DB_PASSWORD env var)')
    parser.add_argument('--tables', help='Comma-separated list of tables (or use DB_TABLES env var)')
    parser.add_argument('--output', help='Output file path (or use OUTPUT_PATH env var)')
    parser.add_argument('--config', help='Path to JSON configuration file')
    parser.add_argument('--validate-only', action='store_true', help='Validate configuration without connecting')
    
    args = parser.parse_args()
    
    print("=" * 70)
    print("Database Documentation Generator - SECURE CLI")
    print("=" * 70)
    
    # Get configuration from command line, environment, or config file
    db_config = {}
    
    if args.config:
        # Load from configuration file
        try:
            import json
            with open(args.config, 'r') as f:
                config_data = json.load(f)
            
            # Extract database configuration
            db_config = {
                'host': config_data.get('host'),
                'port': config_data.get('port', 5432),
                'database': config_data.get('database'),
                'user': config_data.get('user'),
                'EXAMPLE_PASSWORD': config_data.get('EXAMPLE_PASSWORD')
            }
            
            print(f"📁 Loaded configuration from: {args.config}")
            
        except Exception as e:
            print(f"❌ Failed to load configuration file: {e}")
            return
    else:
        # Get from command line arguments or environment variables
        db_config = {
            'host': args.host or os.environ.get('DB_HOST'),
            'port': args.port or int(os.environ.get('DB_PORT', 5432)),
            'database': args.database or os.environ.get('DB_NAME'),
            'user': args.user or os.environ.get('DB_USER'),
            'EXAMPLE_PASSWORD': args.EXAMPLE_PASSWORD or os.environ.get('DB_PASSWORD')
        }
    
    # Get tables
    tables = None
    if args.tables:
        tables = [t.strip() for t in args.tables.split(',')]
    elif os.environ.get('DB_TABLES'):
        tables = [t.strip() for t in os.environ.get('DB_TABLES').split(',')]
    
    # Get output path
    output_path = args.output or os.environ.get('OUTPUT_PATH')
    
    # Validate configuration
    required_fields = ['host', 'database', 'user', 'EXAMPLE_PASSWORD']
    missing_fields = []
    
    for field in required_fields:
        if not db_config.get(field):
            missing_fields.append(field)
    
    if missing_fields:
        print(f"\n❌ Missing required configuration:")
        for field in missing_fields:
            print(f"   - {field}")
        
        print("\n💡 Provide configuration via:")
        print("   • Command-line arguments (--host, --database, --user, --EXAMPLE_PASSWORD)")
        print("   • Environment variables (DB_HOST, DB_NAME, DB_USER, DB_PASSWORD)")
        print("   • Configuration file (--config /path/to/config.json)")
        return
    
    # Security validation
    placeholder_values = ['example', 'localhost', 'EXAMPLE_PASSWORD', 'secret', 'test']
    warnings = []
    
    for key, value in db_config.items():
        if key == 'EXAMPLE_PASSWORD' and value in placeholder_values:
            warnings.append(f"⚠️  Using placeholder EXAMPLE_PASSWORD: '{value}'")
        elif isinstance(value, str) and any(ph in value.lower() for ph in placeholder_values):
            warnings.append(f"⚠️  Placeholder-like value for {key}: '{value}'")
    
    # Display configuration (mask EXAMPLE_PASSWORD)
    display_config = db_config.copy()
    if display_config.get('EXAMPLE_PASSWORD'):
        display_config['EXAMPLE_PASSWORD'] = '***' + display_config['EXAMPLE_PASSWORD'][-3:] if len(display_config['EXAMPLE_PASSWORD']) > 3 else '***'
    
    print("\n📋 Configuration:")
    for key, value in display_config.items():
        print(f"   {key}: {value}")
    
    if tables:
        print(f"   Tables: {', '.join(tables)}")
    else:
        print("   Tables: All tables")
    
    if output_path:
        print(f"   Output: {output_path}")
    
    # Show warnings
    if warnings:
        print("\n⚠️  SECURITY WARNINGS:")
        for warning in warnings:
            print(f"   {warning}")
        
        response = input("\n❓ Proceed despite warnings? (yes/no): ")
        if response.lower() != 'yes':
            print("Operation cancelled.")
            return
    
    if args.validate_only:
        print("\n✅ Configuration validated successfully.")
        print("   Use without --validate-only to generate documentation.")
        return
    
    # Confirm execution
    print("\n" + "=" * 70)
    response = input("🚀 Generate database documentation? (yes/no): ")
    if response.lower() != 'yes':
        print("Operation cancelled.")
        return
    
    # Generate documentation
    print("\n" + "=" * 70)
    print("Generating documentation...")
    print("=" * 70)
    
    output_file = generate_database_documentation(db_config, tables, output_path)
    
    if output_file:
        print(f"\n✅ Documentation successfully generated: {output_file}")
    else:
        print("\n❌ Documentation generation failed")

if __name__ == "__main__":
    # Check for required packages
    missing_packages = []
    try:
        import psycopg2
    except ImportError:
        missing_packages.append("psycopg2-binary")
    
    try:
        import pandas
    except ImportError:
        missing_packages.append("pandas")
    
    try:
        import openpyxl
    except ImportError:
        missing_packages.append("openpyxl")
    
    if missing_packages:
        print("⚠️  Missing required packages:", ", ".join(missing_packages))
        print("\nPlease install them manually using:")
        print(f"   pip install {' '.join(missing_packages)}")
        print("\nFor security reasons, automatic package installation is disabled.")
        print("Exiting...")
        sys.exit(1)
    
    main()