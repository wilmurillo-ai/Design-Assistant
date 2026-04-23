#!/usr/bin/env python3
"""
小红书数据提取 OCR - 主脚本
从小红书笔记图片提取结构化数据（租金/面积/单价等）
"""

import os
import sys
import argparse
import json
import re
from pathlib import Path

# 导入 OCR 模块
from vision_ocr import ocr_image, slice_image

def extract_rental_data(texts):
    """从 OCR 文本中提取租房数据"""
    prices = []
    unit_prices = []
    areas = []
    regions = []
    
    for text in texts:
        text = text.strip()
        if not text:
            continue
        
        # 提取价格
        if '元' in text and '元/平' not in text:
            match = re.search(r'(\d+) 元', text)
            if match:
                prices.append(match.group(1))
        
        # 提取单价
        if '元/平' in text:
            match = re.search(r'(\d+) 元/平', text)
            if match:
                unit_prices.append(match.group(1))
        
        # 提取面积
        if '平' in text and '元' not in text:
            match = re.search(r'(\d+) 平', text)
            if match:
                areas.append(match.group(1))
        
        # 提取区域
        for region in ['南山区', '福田区', '罗湖区', '龙岗区', '宝安区', '龙华区', '盐田区', '光明区', '坪山区']:
            if region in text:
                regions.append(region)
    
    # 去重
    prices = list(dict.fromkeys(prices))
    unit_prices = list(dict.fromkeys(unit_prices))
    areas = list(dict.fromkeys(areas))
    regions = list(dict.fromkeys(regions))
    
    # 配对数据
    min_len = min(len(prices), len(unit_prices), len(areas))
    records = []
    
    for i in range(min_len):
        record = {
            '月租金 (元)': int(prices[i]),
            '单价 (元/平)': int(unit_prices[i]),
            '面积 (平)': int(areas[i]),
            '区域': regions[i] if i < len(regions) else '',
        }
        records.append(record)
    
    return records


def export_to_excel(records, output_path):
    """导出为 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("❌ 错误：需要安装 openpyxl")
        print("   运行：pip3 install openpyxl")
        sys.exit(1)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "提取数据"
    
    # 表头
    headers = ['序号', '图片', '区域', '板块', '户型', '面积 (平)', '月租金 (元)', '单价 (元/平)', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF', size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # 数据
    for row, record in enumerate(records, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=record.get('图片', ''))
        ws.cell(row=row, column=3, value=record.get('区域', ''))
        ws.cell(row=row, column=4, value='')
        ws.cell(row=row, column=5, value='')
        ws.cell(row=row, column=6, value=record.get('面积 (平)', ''))
        ws.cell(row=row, column=7, value=record.get('月租金 (元)', ''))
        ws.cell(row=row, column=8, value=record.get('单价 (元/平)', ''))
        ws.cell(row=row, column=9, value='')
    
    # 列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30
    
    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(2, len(records) + 2):
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    # 保存
    wb.save(output_path)
    return len(records)


def main():
    parser = argparse.ArgumentParser(
        description='小红书数据提取 OCR - 从图片提取结构化数据',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  # 从 URL 提取
  python3 extract_data.py --url "https://www.xiaohongshu.com/explore/xxx" --output data.xlsx
  
  # 从本地图片提取
  python3 extract_data.py --images image1.jpg image2.jpg --output data.xlsx
  
  # 裁切长图后识别（9 等分）
  python3 extract_data.py --images long_image.jpg --slice 9 --output data.xlsx
        '''
    )
    
    parser.add_argument('--url', type=str, help='小红书笔记 URL')
    parser.add_argument('--images', nargs='+', type=str, help='本地图片路径（可多个）')
    parser.add_argument('--output', type=str, default='output.xlsx', help='输出文件路径')
    parser.add_argument('--slice', type=int, default=1, help='长图裁切份数（默认 1，即不裁切）')
    parser.add_argument('--languages', type=str, default='zh-Hans,en-US', help='OCR 语言（默认：zh-Hans,en-US）')
    parser.add_argument('--confidence', type=float, default=0.3, help='最低置信度（默认：0.3）')
    
    args = parser.parse_args()
    
    if not args.url and not args.images:
        parser.print_help()
        sys.exit(1)
    
    print("🔍 小红书数据提取 OCR")
    print("=" * 50)
    
    all_records = []
    
    # 处理 URL
    if args.url:
        print(f"\n📥 下载图片：{args.url}")
        # TODO: 实现 URL 下载
        print("   ⚠️  URL 下载功能待实现，请使用本地图片")
    
    # 处理本地图片
    if args.images:
        for img_path in args.images:
            print(f"\n📷 处理图片：{img_path}")
            
            if not os.path.exists(img_path):
                print(f"   ❌ 文件不存在：{img_path}")
                continue
            
            # 裁切图片（如果需要）
            if args.slice > 1:
                print(f"   ✂️  裁切成 {args.slice} 份...")
                pieces = slice_image(img_path, args.slice)
            else:
                pieces = [img_path]
            
            # OCR 识别
            for i, piece in enumerate(pieces):
                print(f"   🔍 识别部分 {i+1}/{len(pieces)}...")
                texts = ocr_image(piece, args.languages)
                
                if texts:
                    records = extract_rental_data(texts)
                    for r in records:
                        r['图片'] = f"{os.path.basename(img_path)}-{i+1}"
                    all_records.extend(records)
                    print(f"      ✅ 提取 {len(records)} 条记录")
                else:
                    print(f"      ⚠️  未检测到文字")
    
    # 导出
    if all_records:
        print(f"\n💾 导出 Excel: {args.output}")
        count = export_to_excel(all_records, args.output)
        print(f"   ✅ 导出成功！共 {count} 条记录")
        print(f"   📁 文件位置：{os.path.abspath(args.output)}")
    else:
        print("\n⚠️  未提取到任何数据")
        sys.exit(1)


if __name__ == "__main__":
    main()
