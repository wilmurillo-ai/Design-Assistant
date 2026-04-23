#!/usr/bin/env python3
"""Inspect spreadsheet-like import templates and summarize likely header structure."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
import xlrd


INSTRUCTION_KEYWORDS = (
    "填表说明",
    "字段说明",
    "模板说明",
    "数据字典",
    "说明",
    "示例",
    "example",
    "instruction",
)


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def score_headers(headers: list[str], sheet_name: str) -> int:
    non_empty = [item for item in headers if item]
    score = len(non_empty) * 2
    lowered_name = sheet_name.lower()
    if any(keyword.lower() in lowered_name for keyword in INSTRUCTION_KEYWORDS):
        score -= 8
    if any(any(keyword.lower() in header.lower() for keyword in INSTRUCTION_KEYWORDS) for header in non_empty):
        score -= 3
    if len(non_empty) >= 4:
        score += 4
    return score


def detect_header(rows: Iterable[list[str]]) -> tuple[int | None, list[str], list[list[str]]]:
    buffered = [row for row in rows]
    best_index = None
    best_headers: list[str] = []
    best_score = -10**9
    for index, row in enumerate(buffered[:15], start=1):
        score = score_headers(row, "")
        if score > best_score:
            best_score = score
            best_index = index
            best_headers = row
    if best_index is None:
        return None, [], []
    sample_rows = []
    for row in buffered[best_index : best_index + 3]:
        sample_rows.append(row[: len(best_headers)])
    return best_index, best_headers, sample_rows


def inspect_csv(path: Path, delimiter: str) -> dict[str, object]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        rows = [[normalize_cell(cell) for cell in row] for row in reader]
    header_row, headers, sample_rows = detect_header(rows)
    return {
        "file": str(path),
        "type": path.suffix.lower().lstrip("."),
        "sheets": [
            {
                "name": path.stem,
                "header_row": header_row,
                "headers": headers,
                "sample_rows": sample_rows,
                "score": score_headers(headers, path.stem),
            }
        ],
    }


def iter_ws_rows(worksheet) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        normalized = [normalize_cell(cell) for cell in row]
        if any(normalized):
            rows.append(normalized)
    return rows


def inspect_workbook(path: Path, requested_sheet: str | None) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    names = [requested_sheet] if requested_sheet else workbook.sheetnames
    for sheet_name in names:
        worksheet = workbook[sheet_name]
        rows = iter_ws_rows(worksheet)
        header_row, headers, sample_rows = detect_header(rows)
        sheets.append(
            {
                "name": sheet_name,
                "header_row": header_row,
                "headers": headers,
                "sample_rows": sample_rows,
                "score": score_headers(headers, sheet_name),
            }
        )
    sheets.sort(key=lambda item: item["score"], reverse=True)
    return {"file": str(path), "type": path.suffix.lower().lstrip("."), "sheets": sheets}


def inspect_legacy_xls(path: Path, requested_sheet: str | None) -> dict[str, object]:
    workbook = xlrd.open_workbook(str(path))
    sheets = []
    names = [requested_sheet] if requested_sheet else workbook.sheet_names()
    for sheet_name in names:
        worksheet = workbook.sheet_by_name(sheet_name)
        rows = []
        for row_index in range(worksheet.nrows):
            row = [normalize_cell(worksheet.cell_value(row_index, col_index)) for col_index in range(worksheet.ncols)]
            if any(row):
                rows.append(row)
        header_row, headers, sample_rows = detect_header(rows)
        sheets.append(
            {
                "name": sheet_name,
                "header_row": header_row,
                "headers": headers,
                "sample_rows": sample_rows,
                "score": score_headers(headers, sheet_name),
            }
        )
    sheets.sort(key=lambda item: item["score"], reverse=True)
    return {"file": str(path), "type": path.suffix.lower().lstrip("."), "sheets": sheets}


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect an import template and summarize headers.")
    parser.add_argument("--input", required=True, help="Path to .xlsx, .xlsm, .csv, or .tsv file")
    parser.add_argument("--sheet", help="Specific Excel sheet to inspect")
    parser.add_argument("--json", action="store_true", help="Emit compact JSON")
    args = parser.parse_args()

    path = Path(args.input).expanduser().resolve()
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        result = inspect_csv(path, "\t" if suffix == ".tsv" else ",")
    elif suffix == ".xls":
        result = inspect_legacy_xls(path, args.sheet)
    elif suffix in {".xlsx", ".xlsm"}:
        result = inspect_workbook(path, args.sheet)
    else:
        raise SystemExit(f"Unsupported file type: {suffix}")

    if args.json:
        print(json.dumps(result, ensure_ascii=False, separators=(",", ":")))
        return

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
