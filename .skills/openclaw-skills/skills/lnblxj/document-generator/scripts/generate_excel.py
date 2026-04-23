#!/usr/bin/env python3
"""
Excel Spreadsheet Generator
Generates .xlsx files from JSON configuration
"""

import json
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def hex_to_rgb(hex_color):
    """Convert hex color to RGB string for openpyxl"""
    return hex_color.lstrip('#')


def apply_cell_formatting(cell, formatting):
    """Apply formatting to a cell"""
    if not formatting:
        return
    
    # Font formatting
    if 'font' in formatting:
        font_config = formatting['font']
        font = Font(
            name=font_config.get('name', 'Calibri'),
            size=font_config.get('size', 11),
            bold=font_config.get('bold', False),
            italic=font_config.get('italic', False),
            underline='single' if font_config.get('underline', False) else None,
            color=hex_to_rgb(font_config['color']) if 'color' in font_config else None
        )
        cell.font = font
    
    # Background color
    if 'background' in formatting:
        fill = PatternFill(start_color=hex_to_rgb(formatting['background']),
                          end_color=hex_to_rgb(formatting['background']),
                          fill_type='solid')
        cell.fill = fill
    
    # Alignment
    if 'alignment' in formatting:
        align_config = formatting['alignment']
        alignment = Alignment(
            horizontal=align_config.get('horizontal', 'left'),
            vertical=align_config.get('vertical', 'top'),
            wrap_text=align_config.get('wrap_text', False)
        )
        cell.alignment = alignment
    
    # Borders
    if 'border' in formatting:
        border_config = formatting['border']
        side_style = border_config.get('style', 'thin')
        border = Border(
            left=Side(style=side_style) if border_config.get('left', True) else None,
            right=Side(style=side_style) if border_config.get('right', True) else None,
            top=Side(style=side_style) if border_config.get('top', True) else None,
            bottom=Side(style=side_style) if border_config.get('bottom', True) else None
        )
        cell.border = border
    
    # Number format
    if 'number_format' in formatting:
        cell.number_format = formatting['number_format']


def process_sheet(ws, sheet_config):
    """Process a single sheet configuration"""
    # Add data
    data = sheet_config.get('data', [])
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Handle formulas (strings starting with =)
            if isinstance(cell_value, str) and cell_value.startswith('='):
                cell.value = cell_value
            else:
                cell.value = cell_value
    
    # Apply formatting
    formatting = sheet_config.get('formatting', {})
    
    # Column widths
    if 'column_widths' in formatting:
        for col_idx, width in enumerate(formatting['column_widths'], start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Row heights
    if 'row_heights' in formatting:
        for row_idx, height in enumerate(formatting['row_heights'], start=1):
            ws.row_dimensions[row_idx].height = height
    
    # Header row formatting
    if 'header_row' in formatting:
        header_row_idx = formatting['header_row'] + 1  # Convert to 1-based
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            apply_cell_formatting(cell, {
                'font': {'bold': True, 'size': 11},
                'background': '#D3D3D3',
                'alignment': {'horizontal': 'center'}
            })
    
    # Cell-specific formatting
    if 'cell_formatting' in formatting:
        for cell_ref, cell_format in formatting['cell_formatting'].items():
            # Parse cell reference (e.g., "A1" or "1,1")
            if ',' in cell_ref:
                row_idx, col_idx = map(int, cell_ref.split(','))
                row_idx += 1  # Convert to 1-based
                col_idx += 1
            else:
                from openpyxl.utils import coordinate_to_tuple
                row_idx, col_idx = coordinate_to_tuple(cell_ref)
            
            cell = ws.cell(row=row_idx, column=col_idx)
            apply_cell_formatting(cell, cell_format)
    
    # Merged cells
    if 'merged_cells' in formatting:
        for merge_range in formatting['merged_cells']:
            ws.merge_cells(merge_range)
    
    # Freeze panes
    if 'freeze_panes' in formatting:
        ws.freeze_panes = formatting['freeze_panes']
    
    # Auto-filter
    if 'auto_filter' in formatting:
        ws.auto_filter.ref = formatting['auto_filter']


def generate_excel_document(config, output_path):
    """Generate Excel document from configuration"""
    # Create or load workbook
    if 'template' in config:
        wb = load_workbook(config['template'])
    else:
        wb = Workbook()
        # Remove default sheet if we're creating new sheets
        if 'sheets' in config and len(config['sheets']) > 0:
            wb.remove(wb.active)
    
    # Process sheets
    for sheet_config in config.get('sheets', []):
        sheet_name = sheet_config.get('name', 'Sheet1')
        
        # Create or get sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        process_sheet(ws, sheet_config)
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel document generated: {output_path}")


def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_excel.py <config.json> <output.xlsx>")
        sys.exit(1)
    
    config_path = sys.argv[1]
    output_path = sys.argv[2]
    
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        generate_excel_document(config, output_path)
    except Exception as e:
        print(f"Error generating Excel document: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
