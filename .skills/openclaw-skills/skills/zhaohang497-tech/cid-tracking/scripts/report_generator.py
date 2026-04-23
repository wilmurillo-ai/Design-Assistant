#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 报表生成器 - 生成专业的 CID 投流日报/周报
使用 openpyxl 创建美观、清晰、规整的 Excel 报表
"""

import json
import argparse
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Color
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


class StyleConfig:
    """样式配置"""
    
    # 字体
    FONT_TITLE = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
    FONT_HEADER = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    FONT_NORMAL = Font(name='微软雅黑', size=10)
    FONT_NUMBER = Font(name='Arial', size=10)
    
    # 填充色
    FILL_TITLE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')  # 深蓝
    FILL_HEADER = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')  # 蓝色
    FILL_SUBHEADER = PatternFill(start_color='85C1E9', end_color='85C1E9', fill_type='solid')  # 浅蓝
    FILL_ALT1 = PatternFill(start_color='F8F9F9', end_color='F8F9F9', fill_type='solid')  # 浅灰
    FILL_ALT2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白色
    FILL_POSITIVE = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')  # 浅绿
    FILL_NEGATIVE = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')  # 浅红
    
    # 边框
    THIN_BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 对齐
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')


class ReportGenerator:
    """Excel 报表生成器"""
    
    def __init__(self, config: Dict = None):
        self.config = config or {}
        self.wb = Workbook()
        self.styles = StyleConfig()
    
    def create_summary_sheet(self, data: List[Dict], date: str) -> None:
        """创建汇总看板工作表"""
        ws = self.wb.active
        ws.title = "汇总看板"
        
        # 标题
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"CID 投流日报 - {date}"
        title_cell.font = self.styles.FONT_TITLE
        title_cell.fill = self.styles.FILL_TITLE
        title_cell.alignment = self.styles.ALIGN_CENTER
        
        # 计算汇总数据
        total_cost = sum(item.get('cost', 0) for item in data)
        total_conversion = sum(item.get('conversion', 0) for item in data)
        total_impression = sum(item.get('impression', 0) for item in data)
        total_click = sum(item.get('click', 0) for item in data)
        
        avg_ctr = (total_click / total_impression * 100) if total_impression > 0 else 0
        avg_cvr = (total_conversion / total_click * 100) if total_click > 0 else 0
        avg_cpc = (total_cost / total_click) if total_click > 0 else 0
        avg_cpa = (total_cost / total_conversion) if total_conversion > 0 else 0
        estimated_gmv = total_conversion * 200  # 假设客单价 200
        roi = (estimated_gmv / total_cost) if total_cost > 0 else 0
        
        # 核心指标卡片
        metrics = [
            ('总消耗', f'¥{total_cost:,.2f}', ''),
            ('总转化', f'{total_conversion}', ''),
            ('总展示', f'{total_impression:,}', ''),
            ('总点击', f'{total_click:,}', ''),
            ('CTR', f'{avg_ctr:.2f}%', ''),
            ('CVR', f'{avg_cvr:.2f}%', ''),
            ('CPC', f'¥{avg_cpc:.2f}', ''),
            ('CPA', f'¥{avg_cpa:.2f}', ''),
            ('估算 GMV', f'¥{estimated_gmv:,.2f}', ''),
            ('ROI', f'{roi:.2f}', 'positive' if roi >= 1.5 else 'negative')
        ]
        
        # 写入指标
        start_row = 3
        for i, (name, value, status) in enumerate(metrics):
            row = start_row + i
            col_name = get_column_letter(1)
            col_value = get_column_letter(2)
            
            # 指标名称
            name_cell = ws[f'{col_name}{row}']
            name_cell.value = name
            name_cell.font = self.styles.FONT_NORMAL
            name_cell.fill = self.styles.FILL_SUBHEADER
            name_cell.alignment = self.styles.ALIGN_CENTER
            name_cell.border = self.styles.THIN_BORDER
            
            # 指标值
            value_cell = ws[f'{col_value}{row}']
            value_cell.value = value
            value_cell.font = self.styles.FONT_NUMBER
            value_cell.alignment = self.styles.ALIGN_RIGHT
            value_cell.border = self.styles.THIN_BORDER
            
            if status == 'positive':
                value_cell.fill = self.styles.FILL_POSITIVE
            elif status == 'negative':
                value_cell.fill = self.styles.FILL_NEGATIVE
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        
        # 添加备注
        note_row = start_row + len(metrics) + 2
        ws[f'A{note_row}'] = "备注：GMV 按转化数 × 客单价 (¥200) 估算"
        ws[f'A{note_row}'].font = Font(size=9, italic=True, color='666666')
    
    def create_platform_sheet(self, data: List[Dict]) -> None:
        """创建分平台数据工作表"""
        ws = self.wb.create_sheet("分平台数据")
        
        # 按平台分组
        platforms = {}
        for item in data:
            platform = item.get('platform', 'unknown')
            if platform not in platforms:
                platforms[platform] = {
                    'cost': 0, 'conversion': 0,
                    'impression': 0, 'click': 0
                }
            platforms[platform]['cost'] += item.get('cost', 0)
            platforms[platform]['conversion'] += item.get('conversion', 0)
            platforms[platform]['impression'] += item.get('impression', 0)
            platforms[platform]['click'] += item.get('click', 0)
        
        # 表头
        headers = ['平台', '消耗', '转化数', '展示量', '点击量', 'CTR', 'CVR', 'CPA', 'ROI']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.styles.FONT_HEADER
            cell.fill = self.styles.FILL_HEADER
            cell.alignment = self.styles.ALIGN_CENTER
            cell.border = self.styles.THIN_BORDER
        
        # 数据行
        platform_names = {'oceanengine': '抖音', 'kuaishou': '快手', 'tencent': '腾讯'}
        row = 2
        for platform, stats in platforms.items():
            name = platform_names.get(platform, platform)
            ctr = (stats['click'] / stats['impression'] * 100) if stats['impression'] > 0 else 0
            cvr = (stats['conversion'] / stats['click'] * 100) if stats['click'] > 0 else 0
            cpa = (stats['cost'] / stats['conversion']) if stats['conversion'] > 0 else 0
            gmv = stats['conversion'] * 200
            roi = (gmv / stats['cost']) if stats['cost'] > 0 else 0
            
            values = [
                name,
                f"¥{stats['cost']:,.2f}",
                stats['conversion'],
                f"{stats['impression']:,}",
                f"{stats['click']:,}",
                f"{ctr:.2f}%",
                f"{cvr:.2f}%",
                f"¥{cpa:.2f}",
                f"{roi:.2f}"
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.FONT_NUMBER if col > 1 else self.styles.FONT_NORMAL
                cell.alignment = self.styles.ALIGN_CENTER if col == 1 else self.styles.ALIGN_RIGHT
                cell.border = self.styles.THIN_BORDER
                
                # 隔行换色
                cell.fill = self.styles.FILL_ALT1 if row % 2 == 0 else self.styles.FILL_ALT2
            
            row += 1
        
        # 调整列宽
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def create_campaign_sheet(self, data: List[Dict]) -> None:
        """创建计划明细工作表"""
        ws = self.wb.create_sheet("计划明细")
        
        # 表头
        headers = ['平台', '计划 ID', '计划名称', '消耗', '转化', '展示', '点击', 'CTR', 'CVR', 'CPA', 'ROI']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.styles.FONT_HEADER
            cell.fill = self.styles.FILL_HEADER
            cell.alignment = self.styles.ALIGN_CENTER
            cell.border = self.styles.THIN_BORDER
        
        # 数据行
        platform_names = {'oceanengine': '抖音', 'kuaishou': '快手', 'tencent': '腾讯'}
        row = 2
        for item in sorted(data, key=lambda x: x.get('cost', 0), reverse=True):
            platform = platform_names.get(item.get('platform', ''), item.get('platform', ''))
            ctr = (item.get('click', 0) / item.get('impression', 1) * 100) if item.get('impression', 0) > 0 else 0
            cvr = (item.get('conversion', 0) / item.get('click', 1) * 100) if item.get('click', 0) > 0 else 0
            cpa = (item.get('cost', 0) / item.get('conversion', 1)) if item.get('conversion', 0) > 0 else 0
            gmv = item.get('conversion', 0) * 200
            roi = (gmv / item.get('cost', 1)) if item.get('cost', 0) > 0 else 0
            
            values = [
                platform,
                item.get('campaign_id', ''),
                item.get('campaign_name', ''),
                f"¥{item.get('cost', 0):,.2f}",
                item.get('conversion', 0),
                f"{item.get('impression', 0):,}",
                f"{item.get('click', 0):,}",
                f"{ctr:.2f}%",
                f"{cvr:.2f}%",
                f"¥{cpa:.2f}",
                f"{roi:.2f}"
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.FONT_NUMBER if col > 3 else self.styles.FONT_NORMAL
                cell.alignment = self.styles.ALIGN_CENTER if col <= 3 else self.styles.ALIGN_RIGHT
                cell.border = self.styles.THIN_BORDER
                
                # 隔行换色
                cell.fill = self.styles.FILL_ALT1 if row % 2 == 0 else self.styles.FILL_ALT2
                
                # ROI 标记
                if col == 11 and roi < 1.5:
                    cell.fill = self.styles.FILL_NEGATIVE
                elif col == 11 and roi >= 2.0:
                    cell.fill = self.styles.FILL_POSITIVE
            
            row += 1
        
        # 调整列宽
        column_widths = [12, 15, 25, 15, 10, 15, 15, 12, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def create_anomaly_sheet(self, data: List[Dict], min_roi: float = 1.5, max_cpa: float = 100) -> None:
        """创建异常监控工作表"""
        ws = self.wb.create_sheet("异常监控")
        
        # 标题
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"异常广告计划 (ROI < {min_roi} 或 CPA > ¥{max_cpa})"
        title_cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        title_cell.alignment = self.styles.ALIGN_CENTER
        
        # 表头
        headers = ['平台', '计划 ID', '计划名称', '消耗', '转化', 'CPA', 'ROI', '异常类型']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.styles.FONT_HEADER
            cell.fill = self.styles.FILL_HEADER
            cell.alignment = self.styles.ALIGN_CENTER
            cell.border = self.styles.THIN_BORDER
        
        # 筛选异常计划
        platform_names = {'oceanengine': '抖音', 'kuaishou': '快手', 'tencent': '腾讯'}
        anomalies = []
        row = 3
        
        for item in data:
            cost = item.get('cost', 0)
            conversion = item.get('conversion', 0)
            
            if cost == 0:
                continue
            
            cpa = cost / conversion if conversion > 0 else float('inf')
            gmv = conversion * 200
            roi = gmv / cost
            
            anomaly_types = []
            if roi < min_roi and cost > 100:  # 消耗超过 100 元才标记
                anomaly_types.append(f'ROI 过低 ({roi:.2f})')
            if cpa > max_cpa and conversion > 0:
                anomaly_types.append(f'CPA 过高 (¥{cpa:.2f})')
            
            if anomaly_types:
                anomalies.append({
                    'platform': platform_names.get(item.get('platform', ''), ''),
                    'campaign_id': item.get('campaign_id', ''),
                    'campaign_name': item.get('campaign_name', ''),
                    'cost': cost,
                    'conversion': conversion,
                    'cpa': cpa,
                    'roi': roi,
                    'anomaly_types': ', '.join(anomaly_types)
                })
        
        # 写入数据
        for anomaly in anomalies:
            values = [
                anomaly['platform'],
                anomaly['campaign_id'],
                anomaly['campaign_name'],
                f"¥{anomaly['cost']:,.2f}",
                anomaly['conversion'],
                f"¥{anomaly['cpa']:.2f}",
                f"{anomaly['roi']:.2f}",
                anomaly['anomaly_types']
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles.FONT_NUMBER if col > 3 else self.styles.FONT_NORMAL
                cell.alignment = self.styles.ALIGN_CENTER if col in [1, 8] else self.styles.ALIGN_RIGHT
                cell.border = self.styles.THIN_BORDER
                cell.fill = self.styles.FILL_NEGATIVE
            
            row += 1
        
        # 调整列宽
        column_widths = [12, 15, 25, 15, 10, 12, 12, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 添加统计
        if row > 3:
            stat_row = row + 1
            ws[f'A{stat_row}'] = f"共发现 {len(anomalies)} 个异常计划"
            ws[f'A{stat_row}'].font = Font(size=10, bold=True, color='C0392B')
    
    def save(self, output_path: str) -> None:
        """保存 Excel 文件"""
        self.wb.save(output_path)


def load_data(file_path: str) -> List[Dict]:
    """加载 JSON 数据"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def main():
    parser = argparse.ArgumentParser(description='Excel 报表生成器')
    parser.add_argument('--input', '-i', type=str, required=True,
                        help='输入数据文件 (JSON 格式)')
    parser.add_argument('--output', '-o', type=str, required=True,
                        help='输出 Excel 文件路径')
    parser.add_argument('--date', '-d', type=str, default=None,
                        help='报表日期 (YYYY-MM-DD), 默认今天')
    parser.add_argument('--min-roi', type=float, default=1.5,
                        help='ROI 阈值 (低于此值标记为异常)')
    parser.add_argument('--max-cpa', type=float, default=100,
                        help='CPA 阈值 (高于此值标记为异常)')
    
    args = parser.parse_args()
    
    # 加载数据
    print(f"[DATA] 加载数据：{args.input}")
    data = load_data(args.input)
    print(f"   共 {len(data)} 条记录")
    
    # 确定日期
    date = args.date or datetime.now().strftime('%Y-%m-%d')
    
    # 创建报表生成器
    generator = ReportGenerator()
    
    # 生成各工作表
    print(f"\n[REPORT] 生成报表...")
    print(f"   - 创建汇总看板...")
    generator.create_summary_sheet(data, date)
    
    print(f"   - 创建分平台数据...")
    generator.create_platform_sheet(data)
    
    print(f"   - 创建计划明细...")
    generator.create_campaign_sheet(data)
    
    print(f"   - 创建异常监控...")
    generator.create_anomaly_sheet(data, min_roi=args.min_roi, max_cpa=args.max_cpa)
    
    # 保存文件
    print(f"\n[SAVE] 保存文件：{args.output}")
    generator.save(args.output)
    
    print(f"\n{'='*60}")
    print(f"[OK] Excel 报表生成完成!")
    print(f"   文件：{args.output}")
    print(f"   日期：{date}")
    print(f"   工作表：汇总看板、分平台数据、计划明细、异常监控")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
