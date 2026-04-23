# -*- coding: utf-8 -*-
import sys, os, traceback

log_path = r'C:\Users\Administrator\.openclaw\workspace\extract_run_log.txt'
log = open(log_path, 'w', encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import pymupdf, pdfplumber, re

    log.write('Imports OK\n')
    log.flush()

    PDF_FOLDER = r'C:\Users\Administrator\Desktop\车险保单'
    OUTPUT_PATH = r'C:\Users\Administrator\Desktop\车险保单提取结果_v4.xlsx'

    def safe_extract(text, patterns):
        for p in patterns:
            flags = 0
            if isinstance(p, tuple):
                pat, flags = p
            else:
                pat = p
            try:
                m = re.search(pat, text, flags)
                if m:
                    val = m.group(1).strip()
                    if val:
                        return val
            except:
                pass
        return ""

    def clean_data(data, text):
        for k, v in data.items():
            if isinstance(v, str) and v:
                for bad in ['/', '×', '✗', '=', '<', '>']:
                    if bad in v:
                        parts = re.split(f'[{re.escape(bad)}]', v)
                        v = parts[0].strip()
                v = re.sub(r'\t+', ' ', v)
                v = re.sub(r' +', ' ', v)
                v = v.strip(' .,;:\n\r\t')
                data[k] = v
            if data[k] is None:
                data[k] = ''
        return data

    def route_company(text):
        m = re.search(r"公司名称[：:]\s*(.{5,30}?)(?=公司地址|营业执照|注册地址|联系电话|地址|$)", text)
        if m:
            name = m.group(1).strip()
            for kw in ['太平洋', '中国人民', '亚太财产', '大地财产', '浙商财产', '华海']:
                if kw in name:
                    return kw
        return ''

    def route_type(text):
        if '交强险' in text or '机动车交通事故责任强制保险' in text:
            return '交强险'
        if '商业险' in text:
            return '商业险'
        return '未知'

    def parse_jiaoqiang(text, company=''):
        data = {
            '签单时间': '', '保险公司名称': company, '保单号': '',
            '保险起期': '', '车辆使用性质': '', '车架号': '',
            '车辆型号名称': '', '被保人姓名': '', '被保险人证件号': '',
            '被保险人手机号': '', '车牌号码': '', '险种名称原始': '',
            '实收保费': '', '车船税': ''
        }
        data['签单时间'] = safe_extract(text, [
            r"签单日期[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?)",
            r"签单时间[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?)",
        ])
        data['保单号'] = safe_extract(text, [r"保单号[：:]\s*([A-Z0-9]{10,})", r"保单号\s*([A-Z0-9]{10,})"])
        data['保险起期'] = safe_extract(text, [r"保险期间[：:]*\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?)"])
        data['车辆使用性质'] = safe_extract(text, [r"使用性质[：:]\s*([^\n]{2,10})"])
        data['车架号'] = safe_extract(text, [r"车架号[：:]\s*([A-Z0-9]{17})", r"车辆识别代码[：:]\s*([A-Z0-9]{17})"])
        data['车辆型号名称'] = safe_extract(text, [r"品牌型号[：:]\s*([^\n]{2,30})"])
        data['被保人姓名'] = safe_extract(text, [r"被保险人[：:]\s*([^\n]{2,15})"])
        data['被保险人证件号'] = safe_extract(text, [r"证件号码[：:]\s*([A-Z0-9]{15,18})"])
        data['被保险人手机号'] = safe_extract(text, [r"手机[：:]\s*(1[3-9]\d{9})"])
        data['车牌号码'] = safe_extract(text, [r"车牌号[码]?[：:]\s*([鲁豫京津冀晋蒙辽吉黑沪苏浙皖闽赣湘鄂粤桂琼渝川贵云藏陕甘青宁新]{1,2}[A-Z][A-HJ-NP-Z0-9]{5})"])
        data['险种名称原始'] = safe_extract(text, [r"险种名称[：:]\s*([^\n]{2,20})"])
        data['实收保费'] = safe_extract(text, [
            r"保险费合计（人民币大写）：[^¥]*¥[：:\s]*([0-9,]+\.?\d*)",
            r"保险费\s+大写：人民币[^小]*小写：CNY\s+([0-9,]+\.?\d*)",
            r"保险费合计[\s\S]*?[￥¥][：:\s\xa0]*([0-9,]+\.?\d*)",
            r"（￥：\s*([0-9,]+\.?\d*)元）",
            r"实收保费[：:\s]*[￥¥]?\s*([0-9,]+\.?\d*)",
            r"保险费合计[（(][^)]*)[）)]\s*[￥¥]?\s*([0-9,]+\.?\d*)",
            r"保险费合计（人民币大写）[\s\S]*?[￥¥][：:\s]*([0-9,]+\.?\d*)",
            r"保险费合计（大写）[\s\S]*?[￥¥][：:\s]*([0-9,]+\.?\d*)",
            r"保险费[\s\S]*?[￥¥][：:\s]*([0-9,]+\.?\d*)",
            r"保险费金额[\s\S]*?[￥¥][：:\s]*([0-9,]+\.?\d*)",
            r"含税总保险费[^0-9]*([0-9,]+\.?\d*)",
            r"总保险费[^\d]*([0-9,]+\.?\d*)",
            r"([0-9,]+\.00)\s*\n\s*柒佰柒拾元整",
            r"([0-9,]+\.00)\s*\n\s*叁佰陆拾元整",
            r"(\d+\.00)\n[^\n]*\n\u7396\u4f70\u4f0d\u62fe\u5143\u6574",
        ])
        data['车船税'] = safe_extract(text, [
            r"当年应缴\s*[\u00a5\uffe5\uff1а:]*\s*([0-9,]+\.?\d*)",
            r"车船税\s*[\u00a5\uffe5\uff1а:]*\s*([0-9,]+\.?\d*)",
            r"([0-9,]+\.00)\s*\n\s*叁佰陆拾元整",
            r"[一二三四五六七八九十零〇两三百千万亿万]+\s*\n\s*([0-9,]+\.?\d*)",
        ])
        if "华海" in text:
            data["保险公司名称"] = "华海财产保险股份有限公司"
        return data

    def parse_shangye(text, company=''):
        data = {
            '签单时间': '', '保险公司名称': company, '保单号': '',
            '保险起期': '', '车辆使用性质': '', '车架号': '',
            '车辆型号名称': '', '被保人姓名': '', '被保险人证件号': '',
            '被保险人手机号': '', '车牌号码': '', '险种名称原始': '',
            '实收保费': '', '车船税': ''
        }
        data['签单时间'] = safe_extract(text, [
            r"签单日期[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?)",
        ])
        data['保单号'] = safe_extract(text, [r"保单号[：:]\s*([A-Z0-9]{10,})", r"保单号\s*([A-Z0-9]{10,})"])
        data['保险起期'] = safe_extract(text, [r"保险期间[：:]*\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?)"])
        data['车辆使用性质'] = safe_extract(text, [r"使用性质[：:]\s*([^\n]{2,10})"])
        data['车架号'] = safe_extract(text, [r"车架号[：:]\s*([A-Z0-9]{17})", r"车辆识别代码[：:]\s*([A-Z0-9]{17})"])
        data['车辆型号名称'] = safe_extract(text, [r"品牌型号[：:]\s*([^\n]{2,30})"])
        data['被保人姓名'] = safe_extract(text, [r"被保险人[：:]\s*([^\n]{2,15})"])
        data['被保险人证件号'] = safe_extract(text, [r"证件号码[：:]\s*([A-Z0-9]{15,18})"])
        data['被保险人手机号'] = safe_extract(text, [r"手机[：:]\s*(1[3-9]\d{9})"])
        data['车牌号码'] = safe_extract(text, [r"车牌号[码]?[：:]\s*([鲁豫京津冀晋蒙辽吉黑沪苏浙皖闽赣湘鄂粤桂琼渝川贵云藏陕甘青宁新]{1,2}[A-Z][A-HJ-NP-Z0-9]{5})"])
        data['险种名称原始'] = safe_extract(text, [r"险种名称[：:]\s*([^\n]{2,20})"])
        data['实收保费'] = safe_extract(text, [
            r"([0-9,]+\.00)\s*\n\s*[一二三四五六七八九十零〇两整角分元角]+",
            r"保险费合计[\s\S]*?[￥¥][：:\s]*([0-9,]+\.?\d*)",
            r"保险费合计[（(][^)]*)[）)]\s*[￥¥]?\s*([0-9,]+\.?\d*)",
            r"含税总保险费[^0-9]*([0-9,]+\.?\d*)",
            r"总保险费[^\d]*([0-9,]+\.?\d*)",
            r"保险费合计[\s\S]*?[￥¥][：:\s\xa0]*([0-9,]+\.?\d*)",
        ])
        data['车船税'] = safe_extract(text, [r"车船税[：:]\s*([0-9,]+\.?\d*)"])
        if "华海" in text:
            data["保险公司名称"] = "华海财产保险股份有限公司"
        return data

    def parse_pdf(pdf_path):
        try:
            with pymupdf.open(pdf_path) as doc:
                pymupdf_text = ''
                for page in doc:
                    t = page.get_text()
                    if t:
                        pymupdf_text += t

            plumber_text = ''
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        plumber_text += t + '\n'

            def looks_valid(company_val):
                if not company_val or len(company_val) < 4:
                    return False
                bad = ("公司地址", "邮政编码", "服务电话", "签单日期", "保单号", "公司名称", "公司")
                return not company_val.startswith(bad)

            pymupdf_company = re.search(r"公司名称[：:]\s+([^\n]{2,30})(?=公司地址|营业执照|注册地址|联系电话|地址|$)", pymupdf_text)

            if not looks_valid(pymupdf_company.group(1) if pymupdf_company else ''):
                text = plumber_text
                company_check2 = ''
                for kw, full_name in [
                    ("太平洋", "中国太平洋财产保险股份有限公司"),
                    ("中国人民", "中国人民财产保险股份有限公司"),
                    ("亚太财产", "亚太财产保险有限公司"),
                    ("大地财产", "中国大地财产保险股份有限公司"),
                    ("华海", "华海财产保险股份有限公司"),
                ]:
                    if kw in plumber_text:
                        company_check2 = full_name
                        break
            else:
                text = pymupdf_text
                company_check2 = ''

            rt = route_type(text)
            if rt in ("交强险", "需人工判断"):
                return clean_data(parse_jiaoqiang(text, company_check2), text)
            elif rt == "商业险":
                return clean_data(parse_shangye(text, company_check2), text)
            else:
                return clean_data(parse_jiaoqiang(text, company_check2), text)
        except Exception as e:
            return {'error': str(e)}

    log.write('Functions defined, starting process\n')
    log.flush()

    headers = ['文件名', '保险公司名称', '保单号', '签单时间', '保险起期',
               '车辆使用性质', '车架号', '车辆型号名称', '被保人姓名',
               '被保险人证件号', '被保险人手机号', '车牌号码',
               '险种名称原始', '实收保费', '车船税']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '车险保单'
    ws.append(headers)
    log.write(f'Workbook created, output_path={OUTPUT_PATH}\n')
    log.flush()

    pdfs = sorted([f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')])
    log.write(f'Found {len(pdfs)} PDFs\n')
    log.flush()

    for fn in pdfs:
        pdf_path = os.path.join(PDF_FOLDER, fn)
        result = parse_pdf(pdf_path)
        row = [fn] + [result.get(h, '') for h in headers[1:]]
        ws.append(row)
        log.write(f'Done: {fn} -> {result.get("保险公司名称", "?")} / {result.get("实收保费", "?")}\n')
        log.flush()

    log.write(f'About to save to {OUTPUT_PATH}\n')
    log.flush()
    wb.save(OUTPUT_PATH)
    log.write(f'Saved OK, file exists={os.path.exists(OUTPUT_PATH)}\n')
    log.flush()

except Exception as e:
    log.write(f'ERROR: {e}\n')
    log.write(traceback.format_exc())
    log.flush()
finally:
    log.close()
    print('Done, check log file')