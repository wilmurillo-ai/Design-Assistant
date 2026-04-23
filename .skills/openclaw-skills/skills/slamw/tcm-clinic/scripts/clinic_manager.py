#!/usr/bin/env python3
"""中医诊所管理系统 - 统一命令行入口

使用方式:
    python clinic_manager.py <module> <action> [--data-dir DIR] [options...]

模块:
    patients     患者档案管理
    records      病历记录管理
    herbs        中药库存管理
    appointments 预约排班管理
    finance      财务收费管理
    init         初始化所有数据表

示例:
    python clinic_manager.py patients add --name "张三" --gender "男" --phone "13800138000"
    python clinic_manager.py patients search --name "张"
    python clinic_manager.py records add --patient-id "P20260402001" --complaint "头痛三日"
    python clinic_manager.py herbs alerts
    python clinic_manager.py finance summary --period month --month 2026-04
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("错误: 需要安装 openpyxl 库。请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DEFAULT_DATA_DIR = "clinic_data"

# ─── 表头定义 ───────────────────────────────────────────────

SCHEMAS = {
    "patients": {
        "filename": "patients.xlsx",
        "sheet": "patients",
        "headers": [
            "patient_id", "name", "gender", "birth_date", "age",
            "phone", "address", "constitution_type", "allergies",
            "chronic_diseases", "notes", "created_date", "last_visit_date"
        ],
    },
    "records": {
        "filename": "medical_records.xlsx",
        "sheet": "records",
        "headers": [
            "record_id", "patient_id", "patient_name", "visit_date",
            "chief_complaint", "tongue_condition", "pulse_condition",
            "observation", "listening_smelling", "inquiry",
            "diagnosis", "prescription", "advice", "visit_count", "notes"
        ],
    },
    "herbs": {
        "filename": "herbs_inventory.xlsx",
        "sheet": "herbs",
        "headers": [
            "herb_id", "name", "pinyin", "specification", "stock_quantity",
            "unit", "purchase_price", "retail_price", "supplier",
            "expiry_date", "entry_date", "minimum_stock", "category", "notes"
        ],
    },
    "appointments": {
        "filename": "appointments.xlsx",
        "sheet": "appointments",
        "headers": [
            "appointment_id", "patient_id", "patient_name",
            "appointment_date", "time_slot", "status",
            "purpose", "queue_number", "notes"
        ],
    },
    "finances": {
        "filename": "finances.xlsx",
        "sheet": "finances",
        "headers": [
            "finance_id", "record_id", "patient_id", "patient_name",
            "date", "type", "amount", "payment_method", "notes"
        ],
    },
}


# ─── 工具函数 ───────────────────────────────────────────────

def get_data_path(data_dir: str, module: str) -> str:
    schema = SCHEMAS[module]
    return os.path.join(data_dir, schema["filename"])


def ensure_dir(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)


def init_workbook(filepath: str, sheet_name: str, headers: list):
    if os.path.exists(filepath):
        return
    ensure_dir(filepath)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    wb.save(filepath)


def load_workbook_safe(filepath: str, sheet_name: str, headers: list):
    init_workbook(filepath, sheet_name, headers)
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]
    return wb, ws


def read_all_rows(data_dir: str, module: str) -> list:
    schema = SCHEMAS[module]
    filepath = get_data_path(data_dir, module)
    wb, ws = load_workbook_safe(filepath, schema["sheet"], schema["headers"])
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {h: (v if v is not None else "") for h, v in zip(schema["headers"], row)}
        rows.append(d)
    wb.close()
    return rows


def append_row(data_dir: str, module: str, row_data: dict):
    schema = SCHEMAS[module]
    filepath = get_data_path(data_dir, module)
    wb, ws = load_workbook_safe(filepath, schema["sheet"], schema["headers"])
    row_values = [row_data.get(h, "") for h in schema["headers"]]
    ws.append(row_values)
    wb.save(filepath)
    wb.close()


def find_patient_by_name(data_dir: str, name: str) -> list:
    rows = read_all_rows(data_dir, "patients")
    return [r for r in rows if name in (r.get("name") or "")]


# 模块名 → 对应 ID 字段名
_ID_FIELDS = {
    "patients": "patient_id",
    "records": "record_id",
    "herbs": "herb_id",
    "appointments": "appointment_id",
    "finances": "finance_id",
}


def generate_id(prefix: str, data_dir: str = None, module: str = None) -> str:
    """生成递增ID：前缀 + YYYYMMDD + 3位序号（从现有数据中取最大序号+1）"""
    today = datetime.now().strftime("%Y%m%d")
    next_seq = 1
    if data_dir and module:
        rows = read_all_rows(data_dir, module)
        today_prefix = f"{prefix}{today}"
        id_field = _ID_FIELDS.get(module)
        for row in rows:
            rid = (row.get(id_field) or "") if id_field else ""
            if rid and rid.startswith(today_prefix):
                seq_str = rid[len(today_prefix):]
                try:
                    seq = int(seq_str)
                    if seq >= next_seq:
                        next_seq = seq + 1
                except ValueError:
                    pass
    return f"{prefix}{today}{next_seq:03d}"


def get_today() -> str:
    return datetime.now().strftime("%Y-%m-%d")


# ─── 患者模块 ───────────────────────────────────────────────

def cmd_patients_add(args, data_dir):
    name = args.name
    if not name:
        print("错误: --name 参数为必填项", file=sys.stderr)
        return 1

    existing = find_patient_by_name(data_dir, name)
    if existing:
        print(f"提示: 已存在同名患者 '{name}'（ID: {existing[0]['patient_id']}），如需新增请确认。")

    row = {
        "patient_id": generate_id("P", data_dir, "patients"),
        "name": name,
        "gender": args.gender or "",
        "birth_date": args.birth_date or "",
        "age": args.age or 0,
        "phone": args.phone or "",
        "address": args.address or "",
        "constitution_type": args.constitution or "",
        "allergies": args.allergies or "",
        "chronic_diseases": args.chronic_diseases or "",
        "notes": args.notes or "",
        "created_date": get_today(),
        "last_visit_date": "",
    }
    append_row(data_dir, "patients", row)
    print(f"已登记患者: {row['patient_id']} - {name}")
    print(json.dumps(row, ensure_ascii=False, indent=2))
    return 0


def cmd_patients_search(args, data_dir):
    rows = read_all_rows(data_dir, "patients")
    if args.name:
        rows = [r for r in rows if args.name in (r.get("name") or "")]
    if args.phone:
        rows = [r for r in rows if args.phone in (r.get("phone") or "")]
    if args.patient_id:
        rows = [r for r in rows if r.get("patient_id") == args.patient_id]

    if not rows:
        print("未找到匹配的患者记录。")
        return 0

    print(f"共找到 {len(rows)} 条记录:")
    for r in rows:
        print(f"  [{r['patient_id']}] {r['name']} | {r.get('gender','')} | 电话: {r.get('phone','')} | 体质: {r.get('constitution_type','')} | 建档: {r.get('created_date','')}")
    return 0


def cmd_patients_list(args, data_dir):
    rows = read_all_rows(data_dir, "patients")
    if not rows:
        print("暂无患者记录。")
        return 0
    print(f"共 {len(rows)} 位患者:")
    for r in rows:
        print(f"  [{r['patient_id']}] {r['name']} | {r.get('gender','')} | 电话: {r.get('phone','')} | 体质: {r.get('constitution_type','')}")
    return 0


# ─── 病历模块 ───────────────────────────────────────────────

def cmd_records_add(args, data_dir):
    patient_id = args.patient_id
    if not patient_id:
        print("错误: --patient-id 参数为必填项", file=sys.stderr)
        return 1

    patients = read_all_rows(data_dir, "patients")
    patient = next((p for p in patients if p.get("patient_id") == patient_id), None)
    if not patient:
        print(f"错误: 未找到患者 ID '{patient_id}'，请先登记患者。", file=sys.stderr)
        return 1

    complaint = args.complaint or ""
    diagnosis = args.diagnosis or ""
    if not complaint and not diagnosis:
        print("错误: --complaint（主诉）和 --diagnosis（诊断）至少需要填写一项。", file=sys.stderr)
        return 1

    count = len([r for r in read_all_rows(data_dir, "records") if r.get("patient_id") == patient_id])

    row = {
        "record_id": generate_id("R", data_dir, "records"),
        "patient_id": patient_id,
        "patient_name": patient["name"],
        "visit_date": args.date or get_today(),
        "chief_complaint": complaint,
        "tongue_condition": args.tongue or "",
        "pulse_condition": args.pulse or "",
        "observation": args.observation or "",
        "listening_smelling": args.listening or "",
        "inquiry": args.inquiry or "",
        "diagnosis": diagnosis,
        "prescription": args.prescription or "",
        "advice": args.advice or "",
        "visit_count": count + 1,
        "notes": args.notes or "",
    }
    append_row(data_dir, "records", row)
    print(f"已添加病历: {row['record_id']}")
    print(json.dumps(row, ensure_ascii=False, indent=2))
    return 0


def cmd_records_search(args, data_dir):
    rows = read_all_rows(data_dir, "records")
    if args.patient_id:
        rows = [r for r in rows if r.get("patient_id") == args.patient_id]
    if args.patient_name:
        rows = [r for r in rows if args.patient_name in (r.get("patient_name") or "")]
    if args.date_from:
        rows = [r for r in rows if (r.get("visit_date") or "") >= args.date_from]
    if args.date_to:
        rows = [r for r in rows if (r.get("visit_date") or "") <= args.date_to]
    if args.diagnosis:
        rows = [r for r in rows if args.diagnosis in (r.get("diagnosis") or "")]

    if not rows:
        print("未找到匹配的病历记录。")
        return 0

    print(f"共找到 {len(rows)} 条病历记录:")
    for r in rows:
        print(f"  [{r['record_id']}] {r['visit_date']} | {r['patient_name']} | 主诉: {r.get('chief_complaint','')[:30]} | 诊断: {r.get('diagnosis','')}")
    return 0


def _register_cn_font():
    """注册中文字体，返回字体名称列表 (bold_font, normal_font)"""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    bold_font = None
    normal_font = None
    # macOS: 优先 STHeiti，备用 Arial Unicode
    candidates = [
        ("STHeitiBold", "/System/Library/Fonts/STHeiti Medium.ttc", 0),
        ("STHeiti", "/System/Library/Fonts/STHeiti Light.ttc", 0),
        ("ArialUnicode", "/Library/Fonts/Arial Unicode.ttf", None),
    ]
    for name, path, subidx in candidates:
        if not os.path.exists(path):
            continue
        try:
            kw = {} if subidx is None else {"subfontIndex": subidx}
            pdfmetrics.registerFont(TTFont(name, path, **kw))
            if bold_font is None:
                bold_font = name
            if normal_font is None and "Light" in name:
                normal_font = name
            elif normal_font is None and bold_font and name != bold_font:
                normal_font = name
        except Exception:
            continue
    # Linux 常见路径
    if bold_font is None:
        linux_fonts = [
            ("/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc", None),
            ("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", None),
            ("/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc", None),
            ("/usr/share/fonts/truetype/wqy/wqy-microhei.ttc", None),
        ]
        for path, subidx in linux_fonts:
            if not os.path.exists(path):
                continue
            try:
                kw = {} if subidx is None else {"subfontIndex": subidx}
                pdfmetrics.registerFont(TTFont("CjkFont", path, **kw))
                bold_font = "CjkFont"
                normal_font = "CjkFont"
                break
            except Exception:
                continue
    # Windows 常见路径
    if bold_font is None:
        win_candidates = [
            ("SimHei", "C:/Windows/Fonts/simhei.ttf"),
            ("MSYH", "C:/Windows/Fonts/msyh.ttc"),
        ]
        for name, path in win_candidates:
            if not os.path.exists(path):
                continue
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                bold_font = name
                normal_font = name
                break
            except Exception:
                continue
    if not normal_font:
        normal_font = bold_font
    return bold_font, normal_font


def cmd_records_export_pdf(args, data_dir):
    """生成患者电子病历 PDF 文件（用于微信发送）"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm, mm
        from reportlab.lib.colors import HexColor, black, white
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table,
            TableStyle, PageBreak, HRFlowable,
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        print("错误: 需要安装 reportlab 库。请运行: pip3 install reportlab", file=sys.stderr)
        return 1

    patient_id = args.patient_id
    if not patient_id:
        print("错误: --patient-id 参数为必填项", file=sys.stderr)
        return 1

    # 查找患者
    patients = read_all_rows(data_dir, "patients")
    patient = next((p for p in patients if p.get("patient_id") == patient_id), None)
    if not patient:
        print(f"错误: 未找到患者 ID '{patient_id}'", file=sys.stderr)
        return 1

    # 查找病历
    all_records = read_all_rows(data_dir, "records")
    patient_records = [r for r in all_records if r.get("patient_id") == patient_id]
    patient_records.sort(key=lambda r: r.get("visit_date", ""), reverse=True)

    if args.record_id:
        patient_records = [r for r in patient_records if r.get("record_id") == args.record_id]
        if not patient_records:
            print(f"错误: 未找到病历 ID '{args.record_id}'", file=sys.stderr)
            return 1

    if not patient_records:
        print(f"该患者暂无病历记录。")
        return 0

    # 注册中文字体
    bold_font, normal_font = _register_cn_font()
    if not bold_font:
        print("错误: 未找到可用的中文字体，无法生成 PDF", file=sys.stderr)
        return 1

    # 颜色定义
    COLOR_TITLE = HexColor("#2B579A")
    COLOR_SUBTITLE = HexColor("#4472C4")
    COLOR_SECTION_BG = HexColor("#D6E4F0")
    COLOR_LINE = HexColor("#4472C4")
    COLOR_LIGHT_BG = HexColor("#F2F7FC")

    # 输出文件
    today_str = get_today()
    safe_name = patient["name"].replace(" ", "_")
    if args.record_id:
        visit_date = patient_records[0].get("visit_date", today_str)
        pdf_filename = f"{safe_name}_{visit_date}_电子病历.pdf"
    else:
        pdf_filename = f"{safe_name}_病历汇总_{today_str}.pdf"
    pdf_path = os.path.join(data_dir, pdf_filename)

    # 页面参数
    PAGE_W, PAGE_H = A4
    MARGIN_LR = 2 * cm
    MARGIN_TB = 2 * cm
    CONTENT_W = PAGE_W - 2 * MARGIN_LR

    # 自定义页眉页脚
    class MedicalRecordDocTemplate(SimpleDocTemplate):
        def __init__(self, *args, **kwargs):
            self._page_count = 0
            super().__init__(*args, **kwargs)

        def afterPage(self):
            self._page_count += 1

    doc = MedicalRecordDocTemplate(
        pdf_path,
        pagesize=A4,
        leftMargin=MARGIN_LR,
        rightMargin=MARGIN_LR,
        topMargin=MARGIN_TB,
        bottomMargin=MARGIN_TB,
    )

    # 样式定义
    style_title = ParagraphStyle(
        "title", fontName=bold_font, fontSize=18, leading=24,
        textColor=COLOR_TITLE, alignment=TA_CENTER, spaceAfter=2*mm,
    )
    style_subtitle = ParagraphStyle(
        "subtitle", fontName=normal_font, fontSize=10, leading=14,
        textColor=HexColor("#666666"), alignment=TA_CENTER, spaceAfter=4*mm,
    )
    style_section = ParagraphStyle(
        "section", fontName=bold_font, fontSize=13, leading=18,
        textColor=COLOR_SUBTITLE, spaceBefore=6*mm, spaceAfter=3*mm,
    )
    style_subsection = ParagraphStyle(
        "subsection", fontName=bold_font, fontSize=11, leading=15,
        textColor=COLOR_TITLE, spaceBefore=3*mm, spaceAfter=2*mm,
    )
    style_label = ParagraphStyle(
        "label", fontName=bold_font, fontSize=10.5, leading=15,
        textColor=HexColor("#333333"),
    )
    style_text = ParagraphStyle(
        "text", fontName=normal_font, fontSize=10.5, leading=16,
        textColor=black,
    )
    style_table_header = ParagraphStyle(
        "th", fontName=bold_font, fontSize=10, leading=14,
        textColor=white, alignment=TA_CENTER,
    )
    style_table_cell = ParagraphStyle(
        "tc", fontName=normal_font, fontSize=10, leading=14,
        textColor=black,
    )

    # 辅助函数：创建分节线
    def section_line():
        return HRFlowable(width="100%", thickness=1, color=COLOR_LINE,
                          spaceBefore=1*mm, spaceAfter=3*mm)

    # 辅助函数：构建键值对表格
    def info_table(items, col_widths=None):
        """items: list of (label, value)"""
        if not col_widths:
            col_widths = [3.5*cm, CONTENT_W - 3.5*cm]
        rows = []
        row_data = []
        for i, (label, value) in enumerate(items):
            label_p = Paragraph(f"<b>{label}:</b>", style_label)
            value_str = str(value) if value else ""
            value_p = Paragraph(value_str.replace("\n", "<br/>"), style_text)
            row_data.append((label_p, value_p))
        # 2 columns per row
        table_data = []
        for i in range(0, len(row_data), 2):
            if i + 1 < len(row_data):
                table_data.append([row_data[i][0], row_data[i][1], row_data[i+1][0], row_data[i+1][1]])
            else:
                table_data.append([row_data[i][0], row_data[i][1], "", ""])
        w = [col_widths[0], col_widths[1], col_widths[0], col_widths[1]]
        t = Table(table_data, colWidths=w)
        t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 2*mm),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2*mm),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, HexColor("#E0E0E0")),
            ("BACKGROUND", (0, 0), (0, -1), COLOR_LIGHT_BG),
            ("BACKGROUND", (2, 0), (2, -1), COLOR_LIGHT_BG),
        ]))
        return t

    # 构建文档元素
    elements = []

    # 1. 标题区
    elements.append(Paragraph("电子病历", style_title))
    elements.append(Paragraph(f"生成日期: {today_str}", style_subtitle))
    elements.append(section_line())

    # 2. 患者基本信息
    elements.append(Paragraph("患者基本信息", style_section))
    patient_items = [
        ("姓名", patient.get("name", "")),
        ("性别", patient.get("gender", "")),
        ("年龄", patient.get("age", "") if patient.get("age") else ""),
        ("联系电话", patient.get("phone", "")),
        ("体质分型", patient.get("constitution_type", "")),
        ("过敏史", patient.get("allergies", "")),
        ("慢性病史", patient.get("chronic_diseases", "")),
    ]
    elements.append(info_table(patient_items))

    # 3. 就诊记录
    for rec_idx, rec in enumerate(patient_records):
        # 多条病历时换页
        if rec_idx > 0:
            elements.append(PageBreak())
            elements.append(Paragraph(f"电子病历（续）", style_title))
            elements.append(Paragraph(f"患者: {patient.get('name', '')}  |  {today_str}", style_subtitle))
            elements.append(section_line())

        visit_date = rec.get("visit_date", "未知日期")
        visit_count = rec.get("visit_count", rec_idx + 1)
        elements.append(Paragraph(
            f"第 {visit_count} 次就诊 — {visit_date}",
            style_section,
        ))
        elements.append(section_line())

        # 主诉
        complaint = rec.get("chief_complaint", "")
        if complaint:
            elements.append(Paragraph("<b>主诉</b>", style_subsection))
            elements.append(Paragraph(complaint.replace("\n", "<br/>"), style_text))

        # 四诊信息
        four_diag_items = [
            ("望诊", rec.get("observation", "")),
            ("闻诊", rec.get("listening_smelling", "")),
            ("问诊", rec.get("inquiry", "")),
            ("舌诊", rec.get("tongue_condition", "")),
            ("脉诊", rec.get("pulse_condition", "")),
        ]
        has_diag = any(v for _, v in four_diag_items)
        if has_diag:
            elements.append(Paragraph("<b>四诊合参</b>", style_subsection))
            diag_rows = []
            for label, value in four_diag_items:
                if value:
                    diag_rows.append([
                        Paragraph(f"<b>{label}</b>", style_label),
                        Paragraph(value.replace("\n", "<br/>"), style_text),
                    ])
            if diag_rows:
                w_diag = [2.5*cm, CONTENT_W - 2.5*cm]
                diag_table = Table(diag_rows, colWidths=w_diag)
                diag_table.setStyle(TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 1.5*mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5*mm),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.5, HexColor("#E0E0E0")),
                    ("BACKGROUND", (0, 0), (0, -1), COLOR_LIGHT_BG),
                ]))
                elements.append(diag_table)

        # 诊断
        diagnosis = rec.get("diagnosis", "")
        if diagnosis:
            elements.append(Paragraph("<b>诊断</b>", style_subsection))
            elements.append(Paragraph(diagnosis.replace("\n", "<br/>"), style_text))

        # 处方
        prescription = rec.get("prescription", "")
        if prescription:
            elements.append(Paragraph("<b>处方</b>", style_subsection))
            # 解析处方：按行分割，尝试解析为"药名 剂量"格式
            lines = [l.strip() for l in prescription.strip().split("\n") if l.strip()]
            if lines and ("、" in lines[0] or "g" in lines[0].lower() or "克" in lines[0]):
                # 表格形式展示处方
                rx_rows = [[
                    Paragraph("<b>序号</b>", style_table_header),
                    Paragraph("<b>药物组成</b>", style_table_header),
                ]]
                for i, line in enumerate(lines, 1):
                    rx_rows.append([
                        Paragraph(str(i), style_table_cell),
                        Paragraph(line, style_table_cell),
                    ])
                rx_table = Table(rx_rows, colWidths=[1.5*cm, CONTENT_W - 1.5*cm])
                rx_table.setStyle(TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 2*mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2*mm),
                    ("BACKGROUND", (0, 0), (-1, 0), COLOR_SUBTITLE),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, COLOR_LIGHT_BG]),
                    ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#D0D0D0")),
                ]))
                elements.append(rx_table)
            else:
                # 段落形式
                elements.append(Paragraph(prescription.replace("\n", "<br/>"), style_text))

        # 医嘱
        advice = rec.get("advice", "")
        if advice:
            elements.append(Paragraph("<b>医嘱建议</b>", style_subsection))
            elements.append(Paragraph(advice.replace("\n", "<br/>"), style_text))

        # 备注
        notes = rec.get("notes", "")
        if notes:
            elements.append(Spacer(1, 3*mm))
            elements.append(Paragraph(f"<i>备注: {notes}</i>", ParagraphStyle(
                "note", fontName=normal_font, fontSize=9, leading=13,
                textColor=HexColor("#888888"),
            )))

    # 页脚回调
    def add_page_number(canvas, doc):
        canvas.saveState()
        page_num = canvas.getPageNumber()
        # 页脚页码
        canvas.setFont(normal_font, 9)
        canvas.setFillColor(HexColor("#999999"))
        text = f"第 {page_num} 页"
        canvas.drawCentredString(PAGE_W / 2, 1.2*cm, text)
        # 底部分割线
        canvas.setStrokeColor(HexColor("#D0D0D0"))
        canvas.setLineWidth(0.5)
        canvas.line(MARGIN_LR, 1.5*cm, PAGE_W - MARGIN_LR, 1.5*cm)
        canvas.restoreState()

    # 生成 PDF
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    file_size = os.path.getsize(pdf_path)
    size_kb = file_size / 1024

    print(f"已生成电子病历 PDF: {pdf_path}")
    print(f"  患者: {patient.get('name', '')}")
    print(f"  病历数: {len(patient_records)} 条")
    print(f"  文件大小: {size_kb:.1f} KB")
    if size_kb > 500:
        print(f"  提示: 文件较大，建议拆分为单次就诊病历以优化微信传输")
    return 0


# ─── 中药库存模块 ───────────────────────────────────────────

def cmd_herbs_add(args, data_dir):
    name = args.name
    if not name:
        print("错误: --name 参数为必填项", file=sys.stderr)
        return 1

    row = {
        "herb_id": generate_id("H", data_dir, "herbs"),
        "name": name,
        "pinyin": args.pinyin or "",
        "specification": args.spec or "",
        "stock_quantity": args.quantity or 0,
        "unit": args.unit or "g",
        "purchase_price": args.purchase_price or 0,
        "retail_price": args.retail_price or 0,
        "supplier": args.supplier or "",
        "expiry_date": args.expiry_date or "",
        "entry_date": get_today(),
        "minimum_stock": args.min_stock or 0,
        "category": args.category or "",
        "notes": args.notes or "",
    }
    append_row(data_dir, "herbs", row)
    print(f"已入库药材: {row['herb_id']} - {name} ({row['stock_quantity']}{row['unit']})")
    return 0


def cmd_herbs_update(args, data_dir):
    herb_id = args.herb_id
    if not herb_id:
        print("错误: --herb-id 参数为必填项", file=sys.stderr)
        return 1

    schema = SCHEMAS["herbs"]
    filepath = get_data_path(data_dir, "herbs")
    wb, ws = load_workbook_safe(filepath, schema["sheet"], schema["headers"])

    updated = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == herb_id:
            if args.quantity is not None:
                current = row[4].value or 0
                row[4].value = current + args.quantity
            if args.purchase_price is not None:
                row[6].value = args.purchase_price
            if args.retail_price is not None:
                row[7].value = args.retail_price
            if args.expiry_date:
                row[9].value = args.expiry_date
            if args.min_stock is not None:
                row[11].value = args.min_stock
            updated = True
            print(f"已更新药材 {herb_id}")
            break

    if not updated:
        print(f"错误: 未找到药材 ID '{herb_id}'", file=sys.stderr)
        wb.close()
        return 1

    wb.save(filepath)
    wb.close()
    return 0


def cmd_herbs_search(args, data_dir):
    rows = read_all_rows(data_dir, "herbs")
    if args.name:
        rows = [r for r in rows if args.name in (r.get("name") or "") or args.name in (r.get("pinyin") or "")]
    if args.category:
        rows = [r for r in rows if r.get("category") == args.category]

    if not rows:
        print("未找到匹配的药材记录。")
        return 0

    print(f"共找到 {len(rows)} 种药材:")
    for r in rows:
        qty = r.get("stock_quantity", 0) or 0
        unit = r.get("unit", "g") or "g"
        print(f"  [{r['herb_id']}] {r['name']} | 库存: {qty}{unit} | 分类: {r.get('category','')} | 有效期至: {r.get('expiry_date','')}")
    return 0


def cmd_herbs_alerts(args, data_dir):
    rows = read_all_rows(data_dir, "herbs")
    today = datetime.now()

    # 保质期预警
    expiry_days = args.expiry_days or 30
    expiring = []
    for r in rows:
        exp_str = r.get("expiry_date", "")
        if exp_str:
            try:
                exp_date = datetime.strptime(exp_str, "%Y-%m-%d")
                diff = (exp_date - today).days
                if 0 <= diff <= expiry_days:
                    expiring.append((r, diff))
            except ValueError:
                pass

    # 库存预警
    low_stock = []
    for r in rows:
        qty = r.get("stock_quantity") or 0
        min_qty = r.get("minimum_stock") or 0
        if qty <= min_qty and min_qty > 0:
            low_stock.append(r)

    if expiring:
        print(f"⚠ 保质期预警（{expiry_days}天内过期，共 {len(expiring)} 项）:")
        for r, days in expiring:
            print(f"  {r['name']} | 剩余 {days} 天 | 库存: {r.get('stock_quantity',0)}{r.get('unit','g')}")

    if low_stock:
        print(f"\n⚠ 库存不足预警（共 {len(low_stock)} 项）:")
        for r in low_stock:
            qty = r.get("stock_quantity", 0) or 0
            min_q = r.get("minimum_stock", 0) or 0
            print(f"  {r['name']} | 当前: {qty}{r.get('unit','g')} | 最低: {min_q}{r.get('unit','g')}")

    if not expiring and not low_stock:
        print("所有药材状态正常，无需预警。")
    return 0


def cmd_herbs_list(args, data_dir):
    rows = read_all_rows(data_dir, "herbs")
    if not rows:
        print("暂无药材库存记录。")
        return 0
    print(f"共 {len(rows)} 种药材:")
    for r in rows:
        qty = r.get("stock_quantity", 0) or 0
        unit = r.get("unit", "g") or "g"
        print(f"  [{r['herb_id']}] {r['name']} | 库存: {qty}{unit} | 分类: {r.get('category','')} | 有效期: {r.get('expiry_date','')}")
    return 0


# ─── 预约模块 ───────────────────────────────────────────────

def cmd_appointments_add(args, data_dir):
    patient_id = args.patient_id
    if not patient_id:
        print("错误: --patient-id 参数为必填项", file=sys.stderr)
        return 1

    patients = read_all_rows(data_dir, "patients")
    patient = next((p for p in patients if p.get("patient_id") == patient_id), None)
    if not patient:
        print(f"错误: 未找到患者 ID '{patient_id}'", file=sys.stderr)
        return 1

    date = args.date or get_today()
    if not args.time_slot:
        print("错误: --time-slot 参数为必填项（如: 上午/下午/晚上）", file=sys.stderr)
        return 1

    row = {
        "appointment_id": generate_id("A", data_dir, "appointments"),
        "patient_id": patient_id,
        "patient_name": patient["name"],
        "appointment_date": date,
        "time_slot": args.time_slot,
        "status": "待诊",
        "purpose": args.purpose or "",
        "queue_number": "",
        "notes": args.notes or "",
    }
    append_row(data_dir, "appointments", row)
    print(f"已添加预约: {row['appointment_id']} | {date} {args.time_slot} | {patient['name']}")
    return 0


def cmd_appointments_search(args, data_dir):
    rows = read_all_rows(data_dir, "appointments")
    if args.date:
        rows = [r for r in rows if r.get("appointment_date") == args.date]
    if args.status:
        rows = [r for r in rows if r.get("status") == args.status]
    if args.patient_name:
        rows = [r for r in rows if args.patient_name in (r.get("patient_name") or "")]

    if not rows:
        print("未找到匹配的预约记录。")
        return 0

    print(f"共找到 {len(rows)} 条预约:")
    for r in rows:
        print(f"  [{r['appointment_id']}] {r['appointment_date']} {r.get('time_slot','')} | {r['patient_name']} | 状态: {r.get('status','')}")
    return 0


def cmd_appointments_today(args, data_dir):
    rows = read_all_rows(data_dir, "appointments")
    today = get_today()
    today_rows = [r for r in rows if r.get("appointment_date") == today]

    if not today_rows:
        print(f"今日（{today}）暂无预约。")
        return 0

    pending = [r for r in today_rows if r.get("status") == "待诊"]
    completed = [r for r in today_rows if r.get("status") == "已诊"]
    cancelled = [r for r in today_rows if r.get("status") == "取消"]
    noshow = [r for r in today_rows if r.get("status") == "未到"]

    print(f"今日预约总览（{today}）:")
    print(f"  待诊: {len(pending)} | 已诊: {len(completed)} | 取消: {len(cancelled)} | 未到: {len(noshow)}")
    if pending:
        print("  ── 待诊队列 ──")
        for r in pending:
            print(f"    [{r['appointment_id']}] {r.get('time_slot','')} | {r['patient_name']} | {r.get('purpose','')}")
    return 0


def cmd_appointments_list(args, data_dir):
    rows = read_all_rows(data_dir, "appointments")
    if not rows:
        print("暂无预约记录。")
        return 0
    print(f"共 {len(rows)} 条预约记录:")
    for r in rows:
        print(f"  [{r['appointment_id']}] {r['appointment_date']} {r.get('time_slot','')} | {r['patient_name']} | {r.get('status','')}")
    return 0


# ─── 财务模块 ───────────────────────────────────────────────

def cmd_finance_add(args, data_dir):
    patient_id = args.patient_id
    if not patient_id:
        print("错误: --patient-id 参数为必填项", file=sys.stderr)
        return 1

    patients = read_all_rows(data_dir, "patients")
    patient = next((p for p in patients if p.get("patient_id") == patient_id), None)
    patient_name = patient["name"] if patient else ""

    if not args.type:
        print("错误: --type 参数为必填项（挂号费/药费/针灸费/其他）", file=sys.stderr)
        return 1

    row = {
        "finance_id": generate_id("F", data_dir, "finances"),
        "record_id": args.record_id or "",
        "patient_id": patient_id,
        "patient_name": patient_name,
        "date": args.date or get_today(),
        "type": args.type,
        "amount": args.amount or 0,
        "payment_method": args.payment_method or "",
        "notes": args.notes or "",
    }
    append_row(data_dir, "finances", row)
    print(f"已添加收费记录: {row['finance_id']} | {args.type} | {row['amount']}元 | {patient_name}")
    return 0


def cmd_finance_summary(args, data_dir):
    rows = read_all_rows(data_dir, "finances")

    if args.period == "day":
        date = args.date or get_today()
        filtered = [r for r in rows if r.get("date") == date]
        title = f"日收入统计（{date}）"
    elif args.period == "month":
        month = args.month or datetime.now().strftime("%Y-%m")
        filtered = [r for r in rows if (r.get("date") or "")[:7] == month]
        title = f"月度收入统计（{month}）"
    elif args.period == "patient":
        if not args.patient_id:
            print("错误: --period patient 需要 --patient-id 参数", file=sys.stderr)
            return 1
        filtered = [r for r in rows if r.get("patient_id") == args.patient_id]
        title = f"患者费用汇总（{args.patient_id}）"
    else:
        filtered = rows
        title = "全部收入汇总"

    if not filtered:
        print(f"{title}: 无记录。")
        return 0

    total = 0
    type_totals = {}
    method_totals = {}
    for r in filtered:
        amt = r.get("amount") or 0
        total += amt
        t = r.get("type") or "未分类"
        type_totals[t] = type_totals.get(t, 0) + amt
        m = r.get("payment_method") or "未记录"
        method_totals[m] = method_totals.get(m, 0) + amt

    print(f"\n{title}")
    print(f"  总收入: {total:.2f} 元（{len(filtered)} 笔）")
    print(f"  ── 按费用类型 ──")
    for t, amt in sorted(type_totals.items(), key=lambda x: -x[1]):
        pct = (amt / total * 100) if total > 0 else 0
        print(f"    {t}: {amt:.2f} 元 ({pct:.1f}%)")
    print(f"  ── 按支付方式 ──")
    for m, amt in sorted(method_totals.items(), key=lambda x: -x[1]):
        pct = (amt / total * 100) if total > 0 else 0
        print(f"    {m}: {amt:.2f} 元 ({pct:.1f}%)")
    return 0


# ─── 报表生成 ───────────────────────────────────────────────

# 表头中文映射（用于报表显示）
_CN_HEADERS = {
    "patients": {
        "patient_id": "患者ID", "name": "姓名", "gender": "性别",
        "birth_date": "出生日期", "age": "年龄", "phone": "联系电话",
        "address": "地址", "constitution_type": "体质分型",
        "allergies": "过敏史", "chronic_diseases": "慢性病史",
        "notes": "备注", "created_date": "建档日期", "last_visit_date": "最近就诊",
    },
    "records": {
        "record_id": "病历ID", "patient_id": "患者ID", "patient_name": "患者姓名",
        "visit_date": "就诊日期", "chief_complaint": "主诉",
        "tongue_condition": "舌诊", "pulse_condition": "脉诊",
        "observation": "望诊", "listening_smelling": "闻诊",
        "inquiry": "问诊", "diagnosis": "诊断",
        "prescription": "处方", "advice": "医嘱",
        "visit_count": "就诊次数", "notes": "备注",
    },
    "herbs": {
        "herb_id": "药材ID", "name": "药材名称", "pinyin": "拼音",
        "specification": "规格", "stock_quantity": "库存量",
        "unit": "单位", "purchase_price": "进货价",
        "retail_price": "零售价", "supplier": "供应商",
        "expiry_date": "有效期至", "entry_date": "入库日期",
        "minimum_stock": "最低库存", "category": "分类", "notes": "备注",
    },
    "appointments": {
        "appointment_id": "预约ID", "patient_id": "患者ID",
        "patient_name": "患者姓名", "appointment_date": "预约日期",
        "time_slot": "时段", "status": "状态",
        "purpose": "就诊目的", "queue_number": "排队号", "notes": "备注",
    },
    "finances": {
        "finance_id": "财务ID", "record_id": "病历ID",
        "patient_id": "患者ID", "patient_name": "患者姓名",
        "date": "日期", "type": "费用类型",
        "amount": "金额(元)", "payment_method": "支付方式", "notes": "备注",
    },
}

# 报表中需要隐藏的内部字段
_HIDE_FIELDS = {
    "patients": ["patient_id", "notes"],
    "records": ["record_id", "patient_id", "notes"],
    "herbs": ["herb_id", "pinyin", "notes"],
    "appointments": ["appointment_id", "patient_id", "notes"],
    "finances": ["finance_id", "record_id", "patient_id", "notes"],
}


def _apply_title_style(ws, title_text, last_col):
    """在第一行写入合并标题并应用样式"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36


def _apply_header_style(ws, headers, row=2):
    """写入表头行并应用样式，返回列宽数据"""
    font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
    ws.row_dimensions[row].height = 24


def _write_data_rows(ws, rows, display_headers, data_headers, start_row=3):
    """批量写入数据行，返回实际写入行数"""
    money_fields = {"amount", "purchase_price", "retail_price", "stock_quantity"}
    qty_fields = {"stock_quantity", "minimum_stock", "age", "visit_count", "queue_number"}
    alert_font = Font(name="微软雅黑", size=10, color="CC0000", bold=True)
    normal_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    # 斑马纹填充
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for i, row in enumerate(rows):
        r = start_row + i
        fill = even_fill if i % 2 == 0 else odd_fill
        for col_idx, dh in enumerate(data_headers, 1):
            val = row.get(dh, "")
            cell = ws.cell(row=r, column=col_idx, value=val if val else "")
            cell.font = normal_font
            cell.fill = fill
            cell.border = thin_border
            # 对齐方式
            if dh in money_fields:
                cell.alignment = right_align
                if isinstance(val, (int, float)) and val != 0:
                    cell.number_format = '#,##0.00'
            elif dh in qty_fields:
                cell.alignment = right_align
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
            elif dh in ("phone",):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

            # 金额为0则显示空白
            if dh in money_fields and (val == 0 or val == "" or val is None):
                cell.value = ""

    return len(rows)


def _auto_column_width(ws, headers, data_rows, min_width=8, max_width=30):
    """根据表头和数据内容自动调整列宽"""
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row in data_rows:
            val = row.get(_CN_HEADERS.get(list(_CN_HEADERS.keys())[0], {}).get(
                [k for k, v in (_CN_HEADERS.get(list(_CN_HEADERS.keys())[0], {})).items() if v == h][0], ""
            ) if h in str(row.values()) else "")
        # 简化：基于表头和采样数据
        width = max(min_width, min(max_len * 2 + 2, max_width))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


def _setup_print(ws):
    """配置打印设置：A4、适合页宽、重复表头"""
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "2:2"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6


def _smart_column_width(ws, headers, data_dicts, header_map):
    """智能列宽：根据表头中文长度和数据采样计算"""
    for col_idx, h in enumerate(headers, 1):
        # 找到对应的英文字段名
        en_field = None
        for en, cn in header_map.items():
            if cn == h:
                en_field = en
                break
        if not en_field:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
            continue
        max_len = len(h) * 2  # 中文字符算2
        # 采样前20行数据
        for row in data_dicts[:20]:
            val = str(row.get(en_field, "") or "")
            val_len = sum(2 if ord(c) > 127 else 1 for c in val)
            if val_len > max_len:
                max_len = val_len
        width = max(8, min(max_len + 4, 35))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


def _build_sheet(ws, title, module, rows, data_dir):
    """通用工作表构建：标题 + 表头 + 数据 + 样式"""
    schema = SCHEMAS[module]
    hide = _HIDE_FIELDS.get(module, [])
    header_map = _CN_HEADERS[module]
    all_headers = schema["headers"]
    # 过滤隐藏字段
    show_headers = [h for h in all_headers if h not in hide]
    cn_headers = [header_map.get(h, h) for h in show_headers]

    last_col = len(cn_headers)
    _apply_title_style(ws, title, last_col)
    _apply_header_style(ws, cn_headers, row=2)
    count = _write_data_rows(ws, rows, cn_headers, show_headers, start_row=3)
    _smart_column_width(ws, cn_headers, rows, header_map)
    ws.freeze_panes = "A3"
    _setup_print(ws)

    # 低库存预警标记（仅 herbs 模块）
    if module == "herbs":
        alert_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for r_idx, row in enumerate(rows):
            qty = row.get("stock_quantity") or 0
            min_q = row.get("minimum_stock") or 0
            if min_q > 0 and qty <= min_q:
                for c in range(1, last_col + 1):
                    ws.cell(row=3 + r_idx, column=c).fill = alert_fill

    return count


def cmd_report(args, data_dir):
    """生成诊所汇总报表 Excel 文件"""
    today = get_today()
    report_file = os.path.join(data_dir, f"诊所汇总报表_{today}.xlsx")
    limit = args.limit or 30

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. 患者总览
    ws1 = wb.create_sheet("患者总览")
    patients = read_all_rows(data_dir, "patients")
    _build_sheet(ws1, f"患者总览（共 {len(patients)} 人）", "patients", patients, data_dir)

    # 2. 近期病历
    ws2 = wb.create_sheet("近期病历")
    records = read_all_rows(data_dir, "records")
    records_sorted = sorted(records, key=lambda r: r.get("visit_date", ""), reverse=True)[:limit]
    _build_sheet(ws2, f"近期病历（最近 {limit} 条）", "records", records_sorted, data_dir)

    # 3. 中药库存
    ws3 = wb.create_sheet("中药库存")
    herbs = read_all_rows(data_dir, "herbs")
    _build_sheet(ws3, f"中药库存（共 {len(herbs)} 种）", "herbs", herbs, data_dir)

    # 4. 预约记录
    ws4 = wb.create_sheet("预约记录")
    appointments = read_all_rows(data_dir, "appointments")
    appt_sorted = sorted(appointments, key=lambda r: r.get("appointment_date", ""), reverse=True)[:limit]
    _build_sheet(ws4, f"预约记录（最近 {limit} 条）", "appointments", appt_sorted, data_dir)

    # 5. 收入统计（按月汇总）
    ws5 = wb.create_sheet("收入统计")
    finances = read_all_rows(data_dir, "finances")
    # 按月聚合
    month_data = {}
    for f in finances:
        m = (f.get("date") or "")[:7]
        if not m:
            continue
        if m not in month_data:
            month_data[m] = {"total": 0, "types": {}, "methods": {}, "count": 0}
        amt = f.get("amount") or 0
        month_data[m]["total"] += amt
        month_data[m]["count"] += 1
        t = f.get("type") or "未分类"
        month_data[m]["types"][t] = month_data[m]["types"].get(t, 0) + amt
        p = f.get("payment_method") or "未记录"
        month_data[m]["methods"][p] = month_data[m]["methods"].get(p, 0) + amt

    # 收集所有出现过的费用类型和支付方式
    all_types = sorted(set(t for md in month_data.values() for t in md["types"]))
    all_methods = sorted(set(p for md in month_data.values() for p in md["methods"]))

    fin_headers = ["月份", "总收入(元)", "笔数"] + [f"{t}(元)" for t in all_types]
    last_col = len(fin_headers)
    _apply_title_style(ws5, "收入统计（按月汇总）", last_col)
    _apply_header_style(ws5, fin_headers, row=2)

    for i, m in enumerate(sorted(month_data.keys(), reverse=True)):
        md = month_data[m]
        row_vals = [m, md["total"], md["count"]] + [md["types"].get(t, 0) for t in all_types]
        r = 3 + i
        for c, v in enumerate(row_vals, 1):
            cell = ws5.cell(row=r, column=c, value=v)
            cell.font = Font(name="微软雅黑", size=10)
            cell.border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )
            if c >= 2:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00' if c != 3 else '#,##0'
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c == 3:
                cell.number_format = '#,##0'

    # 合计行
    total_row = 3 + len(month_data)
    ws5.cell(row=total_row, column=1, value="合计").font = Font(name="微软雅黑", size=10, bold=True)
    grand_total = sum(md["total"] for md in month_data.values())
    grand_count = sum(md["count"] for md in month_data.values())
    c2 = ws5.cell(row=total_row, column=2, value=grand_total)
    c2.font = Font(name="微软雅黑", size=10, bold=True)
    c2.number_format = '#,##0.00'
    c3 = ws5.cell(row=total_row, column=3, value=grand_count)
    c3.font = Font(name="微软雅黑", size=10, bold=True)
    c3.number_format = '#,##0'
    for ti, t in enumerate(all_types, 4):
        cv = sum(md["types"].get(t, 0) for md in month_data.values())
        cc = ws5.cell(row=total_row, column=ti, value=cv)
        cc.font = Font(name="微软雅黑", size=10, bold=True)
        cc.number_format = '#,##0.00'

    ws5.freeze_panes = "A3"
    _setup_print(ws5)
    for ci in range(1, last_col + 1):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

    # 6. 经营概况
    ws6 = wb.create_sheet("经营概况")
    _apply_title_style(ws6, "诊所经营概况", 2)
    ws6.column_dimensions["A"].width = 20
    ws6.column_dimensions["B"].width = 50

    this_month = today[:7]
    last_month = (datetime.now() - timedelta(days=30)).strftime("%Y-%m")
    month_records = [r for r in records if (r.get("visit_date") or "")[:7] == this_month]
    month_finances = [f for f in finances if (f.get("date") or "")[:7] == this_month]
    month_income = sum(f.get("amount") or 0 for f in month_finances)
    low_stock_count = sum(1 for h in herbs if (h.get("minimum_stock") or 0) > 0 and (h.get("stock_quantity") or 0) <= (h.get("minimum_stock") or 0))
    # 保质期预警
    from datetime import datetime as dt
    expiry_alerts = 0
    for h in herbs:
        exp = h.get("expiry_date", "")
        if exp:
            try:
                diff = (dt.strptime(exp, "%Y-%m-%d") - dt.now()).days
                if 0 <= diff <= 30:
                    expiry_alerts += 1
            except ValueError:
                pass
    today_appts = [a for a in appointments if a.get("appointment_date") == today]

    summary_items = [
        ("报表生成日期", today),
        ("", ""),
        ("── 患者统计 ──", ""),
        ("总患者数", f"{len(patients)} 人"),
        ("本月新增患者", f"{len([p for p in patients if (p.get('created_date') or '')[:7] == this_month])} 人"),
        ("", ""),
        ("── 本月就诊 ──", ""),
        ("本月就诊人次", f"{len(month_records)} 次"),
        ("本月收入", f"¥{month_income:,.2f}"),
        ("本月收费笔数", f"{len(month_finances)} 笔"),
        ("", ""),
        ("── 库存状态 ──", ""),
        ("药材品种数", f"{len(herbs)} 种"),
        ("低库存预警", f"{low_stock_count} 种" + (" ⚠️" if low_stock_count > 0 else "")),
        ("保质期预警（30天内）", f"{expiry_alerts} 种" + (" ⚠️" if expiry_alerts > 0 else "")),
        ("", ""),
        ("── 今日概况 ──", ""),
        ("今日预约数", f"{len(today_appts)} 人"),
        ("今日待诊", f"{len([a for a in today_appts if a.get('status') == '待诊'])} 人"),
    ]

    label_font = Font(name="微软雅黑", size=11, bold=True, color="2B579A")
    section_font = Font(name="微软雅黑", size=11, bold=True, color="4472C4")
    value_font = Font(name="微软雅黑", size=11)
    for i, (label, value) in enumerate(summary_items, 3):
        cl = ws6.cell(row=i, column=1, value=label)
        cv = ws6.cell(row=i, column=2, value=value)
        if label.startswith("──"):
            cl.font = section_font
            cv.font = section_font
        elif label:
            cl.font = label_font
            cv.font = value_font

    wb.save(report_file)
    print(f"✅ 诊所汇总报表已生成: {report_file}")
    print(f"  包含工作表: 患者总览、近期病历、中药库存、预约记录、收入统计、经营概况")
    return 0


# ─── 初始化 ────────────────────────────────────────────────

def cmd_init(args, data_dir):
    os.makedirs(data_dir, exist_ok=True)
    for module, schema in SCHEMAS.items():
        filepath = os.path.join(data_dir, schema["filename"])
        init_workbook(filepath, schema["sheet"], schema["headers"])
        print(f"  ✓ {schema['filename']}")
    print(f"\n诊所数据目录 '{data_dir}' 初始化完成。")
    return 0


# ─── 参数解析 ───────────────────────────────────────────────

def build_parser():
    parser = argparse.ArgumentParser(
        description="中医诊所管理系统",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--data-dir", default=DEFAULT_DATA_DIR, help="数据目录路径（默认: clinic_data）")

    sub = parser.add_subparsers(dest="module")

    # patients
    p_pat = sub.add_parser("patients", help="患者档案管理")
    pat_sub = p_pat.add_subparsers(dest="action")
    pat_add = pat_sub.add_parser("add", help="登记新患者")
    pat_add.add_argument("--name", required=True, help="患者姓名")
    pat_add.add_argument("--gender", help="性别")
    pat_add.add_argument("--birth-date", help="出生日期 (YYYY-MM-DD)")
    pat_add.add_argument("--age", type=int, help="年龄")
    pat_add.add_argument("--phone", help="联系电话")
    pat_add.add_argument("--address", help="地址")
    pat_add.add_argument("--constitution", help="体质分型")
    pat_add.add_argument("--allergies", help="过敏史")
    pat_add.add_argument("--chronic-diseases", help="慢性病史")
    pat_add.add_argument("--notes", help="备注")

    pat_search = pat_sub.add_parser("search", help="搜索患者")
    pat_search.add_argument("--name", help="姓名（模糊匹配）")
    pat_search.add_argument("--phone", help="电话（模糊匹配）")
    pat_search.add_argument("--patient-id", help="患者 ID")

    pat_sub.add_parser("list", help="列出所有患者")

    # records
    p_rec = sub.add_parser("records", help="病历记录管理")
    rec_sub = p_rec.add_subparsers(dest="action")
    rec_add = rec_sub.add_parser("add", help="添加病历记录")
    rec_add.add_argument("--patient-id", required=True, help="患者 ID")
    rec_add.add_argument("--date", help="就诊日期 (YYYY-MM-DD)")
    rec_add.add_argument("--complaint", help="主诉")
    rec_add.add_argument("--tongue", help="舌诊")
    rec_add.add_argument("--pulse", help="脉诊")
    rec_add.add_argument("--observation", help="望诊")
    rec_add.add_argument("--listening", help="闻诊")
    rec_add.add_argument("--inquiry", help="问诊")
    rec_add.add_argument("--diagnosis", help="诊断")
    rec_add.add_argument("--prescription", help="处方")
    rec_add.add_argument("--advice", help="医嘱")
    rec_add.add_argument("--notes", help="备注")

    rec_search = rec_sub.add_parser("search", help="搜索病历")
    rec_search.add_argument("--patient-id", help="患者 ID")
    rec_search.add_argument("--patient-name", help="患者姓名（模糊匹配）")
    rec_search.add_argument("--date-from", help="起始日期")
    rec_search.add_argument("--date-to", help="截止日期")
    rec_search.add_argument("--diagnosis", help="诊断（模糊匹配）")

    rec_export_pdf = rec_sub.add_parser("export-pdf", help="生成电子病历 PDF（用于微信发送）")
    rec_export_pdf.add_argument("--patient-id", required=True, help="患者 ID")
    rec_export_pdf.add_argument("--record-id", help="病历 ID（不指定则导出全部病历）")

    # herbs
    p_herbs = sub.add_parser("herbs", help="中药库存管理")
    herbs_sub = p_herbs.add_subparsers(dest="action")
    herbs_add = herbs_sub.add_parser("add", help="药材入库")
    herbs_add.add_argument("--name", required=True, help="药材名称")
    herbs_add.add_argument("--pinyin", help="拼音")
    herbs_add.add_argument("--spec", help="规格")
    herbs_add.add_argument("--quantity", type=float, help="库存量")
    herbs_add.add_argument("--unit", help="单位")
    herbs_add.add_argument("--purchase-price", type=float, help="进货价")
    herbs_add.add_argument("--retail-price", type=float, help="零售价")
    herbs_add.add_argument("--supplier", help="供应商")
    herbs_add.add_argument("--expiry-date", help="保质期 (YYYY-MM-DD)")
    herbs_add.add_argument("--min-stock", type=float, help="最低库存")
    herbs_add.add_argument("--category", help="分类")
    herbs_add.add_argument("--notes", help="备注")

    herbs_update = herbs_sub.add_parser("update", help="更新药材信息")
    herbs_update.add_argument("--herb-id", required=True, help="药材 ID")
    herbs_update.add_argument("--quantity", type=float, help="库存变动量（正数入库，负数出库）")
    herbs_update.add_argument("--purchase-price", type=float, help="进货价")
    herbs_update.add_argument("--retail-price", type=float, help="零售价")
    herbs_update.add_argument("--expiry-date", help="保质期")
    herbs_update.add_argument("--min-stock", type=float, help="最低库存")

    herbs_search = herbs_sub.add_parser("search", help="搜索药材")
    herbs_search.add_argument("--name", help="药材名称/拼音（模糊匹配）")
    herbs_search.add_argument("--category", help="分类")

    herbs_alerts = herbs_sub.add_parser("alerts", help="库存与保质期预警")
    herbs_alerts.add_argument("--expiry-days", type=int, default=30, help="保质期预警天数（默认30）")

    herbs_sub.add_parser("list", help="列出所有药材")

    # appointments
    p_appt = sub.add_parser("appointments", help="预约排班管理")
    appt_sub = p_appt.add_subparsers(dest="action")
    appt_add = appt_sub.add_parser("add", help="添加预约")
    appt_add.add_argument("--patient-id", required=True, help="患者 ID")
    appt_add.add_argument("--date", help="预约日期 (YYYY-MM-DD)")
    appt_add.add_argument("--time-slot", required=True, help="时段")
    appt_add.add_argument("--purpose", help="就诊目的")
    appt_add.add_argument("--notes", help="备注")

    appt_search = appt_sub.add_parser("search", help="搜索预约")
    appt_search.add_argument("--date", help="日期")
    appt_search.add_argument("--status", help="状态")
    appt_search.add_argument("--patient-name", help="患者姓名")

    appt_sub.add_parser("today", help="今日预约队列")
    appt_sub.add_parser("list", help="列出所有预约")

    # finance
    p_fin = sub.add_parser("finance", help="财务收费管理")
    fin_sub = p_fin.add_subparsers(dest="action")
    fin_add = fin_sub.add_parser("add", help="添加收费记录")
    fin_add.add_argument("--patient-id", required=True, help="患者 ID")
    fin_add.add_argument("--record-id", help="关联病历 ID")
    fin_add.add_argument("--date", help="日期 (YYYY-MM-DD)")
    fin_add.add_argument("--type", required=True, help="费用类型")
    fin_add.add_argument("--amount", type=float, help="金额")
    fin_add.add_argument("--payment-method", help="支付方式")
    fin_add.add_argument("--notes", help="备注")

    fin_summary = fin_sub.add_parser("summary", help="财务统计")
    fin_summary.add_argument("--period", choices=["day", "month", "patient", "all"], default="day", help="统计周期")
    fin_summary.add_argument("--date", help="日期（日统计用）")
    fin_summary.add_argument("--month", help="月份（月统计用，如 2026-04）")
    fin_summary.add_argument("--patient-id", help="患者 ID（患者统计用）")

    # report
    p_report = sub.add_parser("report", help="生成诊所汇总报表")
    p_report.add_argument("--limit", type=int, default=30, help="病历/预约记录显示条数（默认30）")

    # init
    sub.add_parser("init", help="初始化所有数据表")

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    if not args.module:
        parser.print_help()
        return 0

    data_dir = args.data_dir

    # 路由到对应命令
    handlers = {
        ("patients", "add"): cmd_patients_add,
        ("patients", "search"): cmd_patients_search,
        ("patients", "list"): cmd_patients_list,
        ("records", "add"): cmd_records_add,
        ("records", "search"): cmd_records_search,
        ("records", "export-pdf"): cmd_records_export_pdf,
        ("herbs", "add"): cmd_herbs_add,
        ("herbs", "update"): cmd_herbs_update,
        ("herbs", "search"): cmd_herbs_search,
        ("herbs", "alerts"): cmd_herbs_alerts,
        ("herbs", "list"): cmd_herbs_list,
        ("appointments", "add"): cmd_appointments_add,
        ("appointments", "search"): cmd_appointments_search,
        ("appointments", "today"): cmd_appointments_today,
        ("appointments", "list"): cmd_appointments_list,
        ("finance", "add"): cmd_finance_add,
        ("finance", "summary"): cmd_finance_summary,
    }

    if args.module == "init":
        return cmd_init(args, data_dir)

    if args.module == "report":
        return cmd_report(args, data_dir)

    key = (args.module, getattr(args, "action", None))
    handler = handlers.get(key)
    if handler:
        return handler(args, data_dir)
    else:
        parser.parse_args([args.module, "--help"])
        return 0


if __name__ == "__main__":
    sys.exit(main())
