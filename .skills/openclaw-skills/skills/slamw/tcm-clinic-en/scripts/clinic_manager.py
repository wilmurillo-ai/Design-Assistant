#!/usr/bin/env python3
"""TCM Clinic Management System - Unified CLI Entry Point

Usage:
    python clinic_manager.py <module> <action> [--data-dir DIR] [options...]

Modules:
    patients     Patient record management
    records      Medical chart management
    herbs        Herbal inventory management
    appointments Appointment scheduling management
    finance      Financial billing management
    init         Initialize all data tables

Examples:
    python clinic_manager.py patients add --name "John Smith" --gender "Male" --phone "555-0100"
    python clinic_manager.py patients search --name "John"
    python clinic_manager.py records add --patient-id "P20260402001" --complaint "Headache for 3 days"
    python clinic_manager.py herbs alerts
    python clinic_manager.py finance summary --period month --month 2026-04
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl library is required. Please run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DEFAULT_DATA_DIR = "clinic_data"

# ─── Schema Definitions ──────────────────────────────────────────

SCHEMAS = {
    "patients": {
        "filename": "patients.xlsx",
        "sheet": "patients",
        "headers": [
            "patient_id", "name", "gender", "birth_date", "age",
            "phone", "address", "constitution_type", "allergies",
            "chronic_diseases", "notes", "created_date", "last_visit_date"
        ],
    },
    "records": {
        "filename": "medical_records.xlsx",
        "sheet": "records",
        "headers": [
            "record_id", "patient_id", "patient_name", "visit_date",
            "chief_complaint", "tongue_condition", "pulse_condition",
            "observation", "listening_smelling", "inquiry",
            "diagnosis", "prescription", "advice", "visit_count", "notes"
        ],
    },
    "herbs": {
        "filename": "herbs_inventory.xlsx",
        "sheet": "herbs",
        "headers": [
            "herb_id", "name", "pinyin", "specification", "stock_quantity",
            "unit", "purchase_price", "retail_price", "supplier",
            "expiry_date", "entry_date", "minimum_stock", "category", "notes"
        ],
    },
    "appointments": {
        "filename": "appointments.xlsx",
        "sheet": "appointments",
        "headers": [
            "appointment_id", "patient_id", "patient_name",
            "appointment_date", "time_slot", "status",
            "purpose", "queue_number", "notes"
        ],
    },
    "finances": {
        "filename": "finances.xlsx",
        "sheet": "finances",
        "headers": [
            "finance_id", "record_id", "patient_id", "patient_name",
            "date", "type", "amount", "payment_method", "notes"
        ],
    },
}


# ─── Utility Functions ───────────────────────────────────────────

def get_data_path(data_dir: str, module: str) -> str:
    schema = SCHEMAS[module]
    return os.path.join(data_dir, schema["filename"])


def ensure_dir(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)


def init_workbook(filepath: str, sheet_name: str, headers: list):
    if os.path.exists(filepath):
        return
    ensure_dir(filepath)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    wb.save(filepath)


def load_workbook_safe(filepath: str, sheet_name: str, headers: list):
    init_workbook(filepath, sheet_name, headers)
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]
    return wb, ws


def read_all_rows(data_dir: str, module: str) -> list:
    schema = SCHEMAS[module]
    filepath = get_data_path(data_dir, module)
    wb, ws = load_workbook_safe(filepath, schema["sheet"], schema["headers"])
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {h: (v if v is not None else "") for h, v in zip(schema["headers"], row)}
        rows.append(d)
    wb.close()
    return rows


def append_row(data_dir: str, module: str, row_data: dict):
    schema = SCHEMAS[module]
    filepath = get_data_path(data_dir, module)
    wb, ws = load_workbook_safe(filepath, schema["sheet"], schema["headers"])
    row_values = [row_data.get(h, "") for h in schema["headers"]]
    ws.append(row_values)
    wb.save(filepath)
    wb.close()


def find_patient_by_name(data_dir: str, name: str) -> list:
    rows = read_all_rows(data_dir, "patients")
    return [r for r in rows if name in (r.get("name") or "")]


# Module name → corresponding ID field name
_ID_FIELDS = {
    "patients": "patient_id",
    "records": "record_id",
    "herbs": "herb_id",
    "appointments": "appointment_id",
    "finances": "finance_id",
}


def generate_id(prefix: str, data_dir: str = None, module: str = None) -> str:
    """Generate auto-incrementing ID: prefix + YYYYMMDD + 3-digit sequence"""
    today = datetime.now().strftime("%Y%m%d")
    next_seq = 1
    if data_dir and module:
        rows = read_all_rows(data_dir, module)
        today_prefix = f"{prefix}{today}"
        id_field = _ID_FIELDS.get(module)
        for row in rows:
            rid = (row.get(id_field) or "") if id_field else ""
            if rid and rid.startswith(today_prefix):
                seq_str = rid[len(today_prefix):]
                try:
                    seq = int(seq_str)
                    if seq >= next_seq:
                        next_seq = seq + 1
                except ValueError:
                    pass
    return f"{prefix}{today}{next_seq:03d}"


def get_today() -> str:
    return datetime.now().strftime("%Y-%m-%d")


# ─── Patient Module ──────────────────────────────────────────────

def cmd_patients_add(args, data_dir):
    name = args.name
    if not name:
        print("Error: --name is required", file=sys.stderr)
        return 1

    existing = find_patient_by_name(data_dir, name)
    if existing:
        print(f"Notice: Patient '{name}' already exists (ID: {existing[0]['patient_id']}). Confirm if you want to add a new record.")

    row = {
        "patient_id": generate_id("P", data_dir, "patients"),
        "name": name,
        "gender": args.gender or "",
        "birth_date": args.birth_date or "",
        "age": args.age or 0,
        "phone": args.phone or "",
        "address": args.address or "",
        "constitution_type": args.constitution or "",
        "allergies": args.allergies or "",
        "chronic_diseases": args.chronic_diseases or "",
        "notes": args.notes or "",
        "created_date": get_today(),
        "last_visit_date": "",
    }
    append_row(data_dir, "patients", row)
    print(f"Patient registered: {row['patient_id']} - {name}")
    print(json.dumps(row, ensure_ascii=False, indent=2))
    return 0


def cmd_patients_search(args, data_dir):
    rows = read_all_rows(data_dir, "patients")
    if args.name:
        rows = [r for r in rows if args.name in (r.get("name") or "")]
    if args.phone:
        rows = [r for r in rows if args.phone in (r.get("phone") or "")]
    if args.patient_id:
        rows = [r for r in rows if r.get("patient_id") == args.patient_id]

    if not rows:
        print("No matching patient records found.")
        return 0

    print(f"Found {len(rows)} record(s):")
    for r in rows:
        print(f"  [{r['patient_id']}] {r['name']} | {r.get('gender','')} | Phone: {r.get('phone','')} | Constitution: {r.get('constitution_type','')} | Created: {r.get('created_date','')}")
    return 0


def cmd_patients_list(args, data_dir):
    rows = read_all_rows(data_dir, "patients")
    if not rows:
        print("No patient records yet.")
        return 0
    print(f"Total {len(rows)} patient(s):")
    for r in rows:
        print(f"  [{r['patient_id']}] {r['name']} | {r.get('gender','')} | Phone: {r.get('phone','')} | Constitution: {r.get('constitution_type','')}")
    return 0


# ─── Medical Records Module ─────────────────────────────────────

def cmd_records_add(args, data_dir):
    patient_id = args.patient_id
    if not patient_id:
        print("Error: --patient-id is required", file=sys.stderr)
        return 1

    patients = read_all_rows(data_dir, "patients")
    patient = next((p for p in patients if p.get("patient_id") == patient_id), None)
    if not patient:
        print(f"Error: Patient ID '{patient_id}' not found. Please register the patient first.", file=sys.stderr)
        return 1

    complaint = args.complaint or ""
    diagnosis = args.diagnosis or ""
    if not complaint and not diagnosis:
        print("Error: At least one of --complaint or --diagnosis is required.", file=sys.stderr)
        return 1

    count = len([r for r in read_all_rows(data_dir, "records") if r.get("patient_id") == patient_id])

    row = {
        "record_id": generate_id("R", data_dir, "records"),
        "patient_id": patient_id,
        "patient_name": patient["name"],
        "visit_date": args.date or get_today(),
        "chief_complaint": complaint,
        "tongue_condition": args.tongue or "",
        "pulse_condition": args.pulse or "",
        "observation": args.observation or "",
        "listening_smelling": args.listening or "",
        "inquiry": args.inquiry or "",
        "diagnosis": diagnosis,
        "prescription": args.prescription or "",
        "advice": args.advice or "",
        "visit_count": count + 1,
        "notes": args.notes or "",
    }
    append_row(data_dir, "records", row)
    print(f"Medical record added: {row['record_id']}")
    print(json.dumps(row, ensure_ascii=False, indent=2))
    return 0


def cmd_records_search(args, data_dir):
    rows = read_all_rows(data_dir, "records")
    if args.patient_id:
        rows = [r for r in rows if r.get("patient_id") == args.patient_id]
    if args.patient_name:
        rows = [r for r in rows if args.patient_name in (r.get("patient_name") or "")]
    if args.date_from:
        rows = [r for r in rows if (r.get("visit_date") or "") >= args.date_from]
    if args.date_to:
        rows = [r for r in rows if (r.get("visit_date") or "") <= args.date_to]
    if args.diagnosis:
        rows = [r for r in rows if args.diagnosis in (r.get("diagnosis") or "")]

    if not rows:
        print("No matching medical records found.")
        return 0

    print(f"Found {len(rows)} medical record(s):")
    for r in rows:
        print(f"  [{r['record_id']}] {r['visit_date']} | {r['patient_name']} | Complaint: {r.get('chief_complaint','')[:30]} | Diagnosis: {r.get('diagnosis','')}")
    return 0


# ─── Herbal Inventory Module ─────────────────────────────────────

def cmd_herbs_add(args, data_dir):
    name = args.name
    if not name:
        print("Error: --name is required", file=sys.stderr)
        return 1

    row = {
        "herb_id": generate_id("H", data_dir, "herbs"),
        "name": name,
        "pinyin": args.pinyin or "",
        "specification": args.spec or "",
        "stock_quantity": args.quantity or 0,
        "unit": args.unit or "g",
        "purchase_price": args.purchase_price or 0,
        "retail_price": args.retail_price or 0,
        "supplier": args.supplier or "",
        "expiry_date": args.expiry_date or "",
        "entry_date": get_today(),
        "minimum_stock": args.min_stock or 0,
        "category": args.category or "",
        "notes": args.notes or "",
    }
    append_row(data_dir, "herbs", row)
    print(f"Herb stocked in: {row['herb_id']} - {name} ({row['stock_quantity']}{row['unit']})")
    return 0


def cmd_herbs_update(args, data_dir):
    herb_id = args.herb_id
    if not herb_id:
        print("Error: --herb-id is required", file=sys.stderr)
        return 1

    schema = SCHEMAS["herbs"]
    filepath = get_data_path(data_dir, "herbs")
    wb, ws = load_workbook_safe(filepath, schema["sheet"], schema["headers"])

    updated = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == herb_id:
            if args.quantity is not None:
                current = row[4].value or 0
                row[4].value = current + args.quantity
            if args.purchase_price is not None:
                row[6].value = args.purchase_price
            if args.retail_price is not None:
                row[7].value = args.retail_price
            if args.expiry_date:
                row[9].value = args.expiry_date
            if args.min_stock is not None:
                row[11].value = args.min_stock
            updated = True
            print(f"Herb {herb_id} updated")
            break

    if not updated:
        print(f"Error: Herb ID '{herb_id}' not found", file=sys.stderr)
        wb.close()
        return 1

    wb.save(filepath)
    wb.close()
    return 0


def cmd_herbs_search(args, data_dir):
    rows = read_all_rows(data_dir, "herbs")
    if args.name:
        rows = [r for r in rows if args.name in (r.get("name") or "") or args.name in (r.get("pinyin") or "")]
    if args.category:
        rows = [r for r in rows if r.get("category") == args.category]

    if not rows:
        print("No matching herb records found.")
        return 0

    print(f"Found {len(rows)} herb(s):")
    for r in rows:
        qty = r.get("stock_quantity", 0) or 0
        unit = r.get("unit", "g") or "g"
        print(f"  [{r['herb_id']}] {r['name']} | Stock: {qty}{unit} | Category: {r.get('category','')} | Expires: {r.get('expiry_date','')}")
    return 0


def cmd_herbs_alerts(args, data_dir):
    rows = read_all_rows(data_dir, "herbs")
    today = datetime.now()

    # Expiry alerts
    expiry_days = args.expiry_days or 30
    expiring = []
    for r in rows:
        exp_str = r.get("expiry_date", "")
        if exp_str:
            try:
                exp_date = datetime.strptime(exp_str, "%Y-%m-%d")
                diff = (exp_date - today).days
                if 0 <= diff <= expiry_days:
                    expiring.append((r, diff))
            except ValueError:
                pass

    # Low-stock alerts
    low_stock = []
    for r in rows:
        qty = r.get("stock_quantity") or 0
        min_qty = r.get("minimum_stock") or 0
        if qty <= min_qty and min_qty > 0:
            low_stock.append(r)

    if expiring:
        print(f"⚠ Expiry Alert (expiring within {expiry_days} days, {len(expiring)} item(s)):")
        for r, days in expiring:
            print(f"  {r['name']} | {days} day(s) remaining | Stock: {r.get('stock_quantity',0)}{r.get('unit','g')}")

    if low_stock:
        print(f"\n⚠ Low Stock Alert ({len(low_stock)} item(s)):")
        for r in low_stock:
            qty = r.get("stock_quantity", 0) or 0
            min_q = r.get("minimum_stock", 0) or 0
            print(f"  {r['name']} | Current: {qty}{r.get('unit','g')} | Minimum: {min_q}{r.get('unit','g')}")

    if not expiring and not low_stock:
        print("All herbs are in good condition. No alerts.")
    return 0


def cmd_herbs_list(args, data_dir):
    rows = read_all_rows(data_dir, "herbs")
    if not rows:
        print("No herb inventory records yet.")
        return 0
    print(f"Total {len(rows)} herb(s):")
    for r in rows:
        qty = r.get("stock_quantity", 0) or 0
        unit = r.get("unit", "g") or "g"
        print(f"  [{r['herb_id']}] {r['name']} | Stock: {qty}{unit} | Category: {r.get('category','')} | Expires: {r.get('expiry_date','')}")
    return 0


# ─── Appointment Module ─────────────────────────────────────────

def cmd_appointments_add(args, data_dir):
    patient_id = args.patient_id
    if not patient_id:
        print("Error: --patient-id is required", file=sys.stderr)
        return 1

    patients = read_all_rows(data_dir, "patients")
    patient = next((p for p in patients if p.get("patient_id") == patient_id), None)
    if not patient:
        print(f"Error: Patient ID '{patient_id}' not found", file=sys.stderr)
        return 1

    date = args.date or get_today()
    if not args.time_slot:
        print("Error: --time-slot is required (e.g., Morning/Afternoon/Evening)", file=sys.stderr)
        return 1

    row = {
        "appointment_id": generate_id("A", data_dir, "appointments"),
        "patient_id": patient_id,
        "patient_name": patient["name"],
        "appointment_date": date,
        "time_slot": args.time_slot,
        "status": "Pending",
        "purpose": args.purpose or "",
        "queue_number": "",
        "notes": args.notes or "",
    }
    append_row(data_dir, "appointments", row)
    print(f"Appointment added: {row['appointment_id']} | {date} {args.time_slot} | {patient['name']}")
    return 0


def cmd_appointments_search(args, data_dir):
    rows = read_all_rows(data_dir, "appointments")
    if args.date:
        rows = [r for r in rows if r.get("appointment_date") == args.date]
    if args.status:
        rows = [r for r in rows if r.get("status") == args.status]
    if args.patient_name:
        rows = [r for r in rows if args.patient_name in (r.get("patient_name") or "")]

    if not rows:
        print("No matching appointment records found.")
        return 0

    print(f"Found {len(rows)} appointment(s):")
    for r in rows:
        print(f"  [{r['appointment_id']}] {r['appointment_date']} {r.get('time_slot','')} | {r['patient_name']} | Status: {r.get('status','')}")
    return 0


def cmd_appointments_today(args, data_dir):
    rows = read_all_rows(data_dir, "appointments")
    today = get_today()
    today_rows = [r for r in rows if r.get("appointment_date") == today]

    if not today_rows:
        print(f"No appointments for today ({today}).")
        return 0

    pending = [r for r in today_rows if r.get("status") == "Pending"]
    completed = [r for r in today_rows if r.get("status") == "Completed"]
    cancelled = [r for r in today_rows if r.get("status") == "Cancelled"]
    noshow = [r for r in today_rows if r.get("status") == "No-show"]

    print(f"Today's Appointments Overview ({today}):")
    print(f"  Pending: {len(pending)} | Completed: {len(completed)} | Cancelled: {len(cancelled)} | No-show: {len(noshow)}")
    if pending:
        print("  -- Pending Queue --")
        for r in pending:
            print(f"    [{r['appointment_id']}] {r.get('time_slot','')} | {r['patient_name']} | {r.get('purpose','')}")
    return 0


def cmd_appointments_list(args, data_dir):
    rows = read_all_rows(data_dir, "appointments")
    if not rows:
        print("No appointment records yet.")
        return 0
    print(f"Total {len(rows)} appointment(s):")
    for r in rows:
        print(f"  [{r['appointment_id']}] {r['appointment_date']} {r.get('time_slot','')} | {r['patient_name']} | {r.get('status','')}")
    return 0


# ─── Finance Module ──────────────────────────────────────────────

def cmd_finance_add(args, data_dir):
    patient_id = args.patient_id
    if not patient_id:
        print("Error: --patient-id is required", file=sys.stderr)
        return 1

    patients = read_all_rows(data_dir, "patients")
    patient = next((p for p in patients if p.get("patient_id") == patient_id), None)
    patient_name = patient["name"] if patient else ""

    if not args.type:
        print("Error: --type is required (Consultation fee/Herbal fee/Acupuncture fee/Other)", file=sys.stderr)
        return 1

    row = {
        "finance_id": generate_id("F", data_dir, "finances"),
        "record_id": args.record_id or "",
        "patient_id": patient_id,
        "patient_name": patient_name,
        "date": args.date or get_today(),
        "type": args.type,
        "amount": args.amount or 0,
        "payment_method": args.payment_method or "",
        "notes": args.notes or "",
    }
    append_row(data_dir, "finances", row)
    print(f"Charge record added: {row['finance_id']} | {args.type} | {row['amount']} | {patient_name}")
    return 0


def cmd_finance_summary(args, data_dir):
    rows = read_all_rows(data_dir, "finances")

    if args.period == "day":
        date = args.date or get_today()
        filtered = [r for r in rows if r.get("date") == date]
        title = f"Daily Income Report ({date})"
    elif args.period == "month":
        month = args.month or datetime.now().strftime("%Y-%m")
        filtered = [r for r in rows if (r.get("date") or "")[:7] == month]
        title = f"Monthly Income Report ({month})"
    elif args.period == "patient":
        if not args.patient_id:
            print("Error: --period patient requires --patient-id", file=sys.stderr)
            return 1
        filtered = [r for r in rows if r.get("patient_id") == args.patient_id]
        title = f"Patient Charge Summary ({args.patient_id})"
    else:
        filtered = rows
        title = "All Income Summary"

    if not filtered:
        print(f"{title}: No records.")
        return 0

    total = 0
    type_totals = {}
    method_totals = {}
    for r in filtered:
        amt = r.get("amount") or 0
        total += amt
        t = r.get("type") or "Uncategorized"
        type_totals[t] = type_totals.get(t, 0) + amt
        m = r.get("payment_method") or "Not recorded"
        method_totals[m] = method_totals.get(m, 0) + amt

    print(f"\n{title}")
    print(f"  Total Income: {total:.2f} ({len(filtered)} transaction(s))")
    print(f"  -- By Fee Type --")
    for t, amt in sorted(type_totals.items(), key=lambda x: -x[1]):
        pct = (amt / total * 100) if total > 0 else 0
        print(f"    {t}: {amt:.2f} ({pct:.1f}%)")
    print(f"  -- By Payment Method --")
    for m, amt in sorted(method_totals.items(), key=lambda x: -x[1]):
        pct = (amt / total * 100) if total > 0 else 0
        print(f"    {m}: {amt:.2f} ({pct:.1f}%)")
    return 0


# ─── Initialize ──────────────────────────────────────────────────

def cmd_init(args, data_dir):
    os.makedirs(data_dir, exist_ok=True)
    for module, schema in SCHEMAS.items():
        filepath = os.path.join(data_dir, schema["filename"])
        init_workbook(filepath, schema["sheet"], schema["headers"])
        print(f"  ✓ {schema['filename']}")
    print(f"\nClinic data directory '{data_dir}' initialized.")
    return 0


# ─── Argument Parser ─────────────────────────────────────────────

def build_parser():
    parser = argparse.ArgumentParser(
        description="TCM Clinic Management System",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--data-dir", default=DEFAULT_DATA_DIR, help="Data directory path (default: clinic_data)")

    sub = parser.add_subparsers(dest="module")

    # patients
    p_pat = sub.add_parser("patients", help="Patient record management")
    pat_sub = p_pat.add_subparsers(dest="action")
    pat_add = pat_sub.add_parser("add", help="Register new patient")
    pat_add.add_argument("--name", required=True, help="Patient name")
    pat_add.add_argument("--gender", help="Gender")
    pat_add.add_argument("--birth-date", help="Date of birth (YYYY-MM-DD)")
    pat_add.add_argument("--age", type=int, help="Age")
    pat_add.add_argument("--phone", help="Contact phone")
    pat_add.add_argument("--address", help="Address")
    pat_add.add_argument("--constitution", help="Constitution type")
    pat_add.add_argument("--allergies", help="Allergy history")
    pat_add.add_argument("--chronic-diseases", help="Chronic disease history")
    pat_add.add_argument("--notes", help="Notes")

    pat_search = pat_sub.add_parser("search", help="Search patients")
    pat_search.add_argument("--name", help="Name (fuzzy match)")
    pat_search.add_argument("--phone", help="Phone (fuzzy match)")
    pat_search.add_argument("--patient-id", help="Patient ID")

    pat_sub.add_parser("list", help="List all patients")

    # records
    p_rec = sub.add_parser("records", help="Medical chart management")
    rec_sub = p_rec.add_subparsers(dest="action")
    rec_add = rec_sub.add_parser("add", help="Add medical record")
    rec_add.add_argument("--patient-id", required=True, help="Patient ID")
    rec_add.add_argument("--date", help="Visit date (YYYY-MM-DD)")
    rec_add.add_argument("--complaint", help="Chief complaint")
    rec_add.add_argument("--tongue", help="Tongue diagnosis")
    rec_add.add_argument("--pulse", help="Pulse diagnosis")
    rec_add.add_argument("--observation", help="Inspection findings")
    rec_add.add_argument("--listening", help="Auscultation/olfaction findings")
    rec_add.add_argument("--inquiry", help="Inquiry findings")
    rec_add.add_argument("--diagnosis", help="Diagnosis")
    rec_add.add_argument("--prescription", help="Prescription")
    rec_add.add_argument("--advice", help="Medical advice")
    rec_add.add_argument("--notes", help="Notes")

    rec_search = rec_sub.add_parser("search", help="Search medical records")
    rec_search.add_argument("--patient-id", help="Patient ID")
    rec_search.add_argument("--patient-name", help="Patient name (fuzzy match)")
    rec_search.add_argument("--date-from", help="Start date")
    rec_search.add_argument("--date-to", help="End date")
    rec_search.add_argument("--diagnosis", help="Diagnosis (fuzzy match)")

    # herbs
    p_herbs = sub.add_parser("herbs", help="Herbal inventory management")
    herbs_sub = p_herbs.add_subparsers(dest="action")
    herbs_add = herbs_sub.add_parser("add", help="Stock in herb")
    herbs_add.add_argument("--name", required=True, help="Herb name")
    herbs_add.add_argument("--pinyin", help="Pinyin name")
    herbs_add.add_argument("--spec", help="Specification")
    herbs_add.add_argument("--quantity", type=float, help="Stock quantity")
    herbs_add.add_argument("--unit", help="Unit")
    herbs_add.add_argument("--purchase-price", type=float, help="Purchase price")
    herbs_add.add_argument("--retail-price", type=float, help="Retail price")
    herbs_add.add_argument("--supplier", help="Supplier")
    herbs_add.add_argument("--expiry-date", help="Expiry date (YYYY-MM-DD)")
    herbs_add.add_argument("--min-stock", type=float, help="Minimum stock level")
    herbs_add.add_argument("--category", help="Category")
    herbs_add.add_argument("--notes", help="Notes")

    herbs_update = herbs_sub.add_parser("update", help="Update herb info")
    herbs_update.add_argument("--herb-id", required=True, help="Herb ID")
    herbs_update.add_argument("--quantity", type=float, help="Stock change (positive=stock in, negative=stock out)")
    herbs_update.add_argument("--purchase-price", type=float, help="Purchase price")
    herbs_update.add_argument("--retail-price", type=float, help="Retail price")
    herbs_update.add_argument("--expiry-date", help="Expiry date")
    herbs_update.add_argument("--min-stock", type=float, help="Minimum stock level")

    herbs_search = herbs_sub.add_parser("search", help="Search herbs")
    herbs_search.add_argument("--name", help="Herb name/pinyin (fuzzy match)")
    herbs_search.add_argument("--category", help="Category")

    herbs_alerts = herbs_sub.add_parser("alerts", help="Stock & expiry alerts")
    herbs_alerts.add_argument("--expiry-days", type=int, default=30, help="Expiry warning days (default 30)")

    herbs_sub.add_parser("list", help="List all herbs")

    # appointments
    p_appt = sub.add_parser("appointments", help="Appointment scheduling management")
    appt_sub = p_appt.add_subparsers(dest="action")
    appt_add = appt_sub.add_parser("add", help="Add appointment")
    appt_add.add_argument("--patient-id", required=True, help="Patient ID")
    appt_add.add_argument("--date", help="Appointment date (YYYY-MM-DD)")
    appt_add.add_argument("--time-slot", required=True, help="Time slot")
    appt_add.add_argument("--purpose", help="Visit purpose")
    appt_add.add_argument("--notes", help="Notes")

    appt_search = appt_sub.add_parser("search", help="Search appointments")
    appt_search.add_argument("--date", help="Date")
    appt_search.add_argument("--status", help="Status")
    appt_search.add_argument("--patient-name", help="Patient name")

    appt_sub.add_parser("today", help="Today's appointment queue")
    appt_sub.add_parser("list", help="List all appointments")

    # finance
    p_fin = sub.add_parser("finance", help="Financial billing management")
    fin_sub = p_fin.add_subparsers(dest="action")
    fin_add = fin_sub.add_parser("add", help="Add charge record")
    fin_add.add_argument("--patient-id", required=True, help="Patient ID")
    fin_add.add_argument("--record-id", help="Associated medical record ID")
    fin_add.add_argument("--date", help="Date (YYYY-MM-DD)")
    fin_add.add_argument("--type", required=True, help="Fee type")
    fin_add.add_argument("--amount", type=float, help="Amount")
    fin_add.add_argument("--payment-method", help="Payment method")
    fin_add.add_argument("--notes", help="Notes")

    fin_summary = fin_sub.add_parser("summary", help="Financial statistics")
    fin_summary.add_argument("--period", choices=["day", "month", "patient", "all"], default="day", help="Report period")
    fin_summary.add_argument("--date", help="Date (for daily report)")
    fin_summary.add_argument("--month", help="Month (for monthly report, e.g., 2026-04)")
    fin_summary.add_argument("--patient-id", help="Patient ID (for patient report)")

    # init
    sub.add_parser("init", help="Initialize all data tables")

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    if not args.module:
        parser.print_help()
        return 0

    data_dir = args.data_dir

    # Route to corresponding command handler
    handlers = {
        ("patients", "add"): cmd_patients_add,
        ("patients", "search"): cmd_patients_search,
        ("patients", "list"): cmd_patients_list,
        ("records", "add"): cmd_records_add,
        ("records", "search"): cmd_records_search,
        ("herbs", "add"): cmd_herbs_add,
        ("herbs", "update"): cmd_herbs_update,
        ("herbs", "search"): cmd_herbs_search,
        ("herbs", "alerts"): cmd_herbs_alerts,
        ("herbs", "list"): cmd_herbs_list,
        ("appointments", "add"): cmd_appointments_add,
        ("appointments", "search"): cmd_appointments_search,
        ("appointments", "today"): cmd_appointments_today,
        ("appointments", "list"): cmd_appointments_list,
        ("finance", "add"): cmd_finance_add,
        ("finance", "summary"): cmd_finance_summary,
    }

    if args.module == "init":
        return cmd_init(args, data_dir)

    key = (args.module, getattr(args, "action", None))
    handler = handlers.get(key)
    if handler:
        return handler(args, data_dir)
    else:
        parser.parse_args([args.module, "--help"])
        return 0


if __name__ == "__main__":
    sys.exit(main())
