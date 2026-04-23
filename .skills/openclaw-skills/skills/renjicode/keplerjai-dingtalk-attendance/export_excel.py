#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
考勤数据导出为 Excel（模板版）

规则：
1. 日单元格仅在出现迟到/请假/外出/旷工时填值，否则留空。
2. 迟到需要写入迟到分钟和时间。
3. 请假需要写入请假类型和时间段。
4. 交通补贴列不填（AV）。
5. 全勤奖 200 元：仅在无迟到、无请假、无旷工时发放。
6. 餐补 12 元/天：当日若请假或旷工则无餐补。
"""

import calendar
import io
import json
import os
import sys
from collections import Counter
from copy import copy
from datetime import datetime, timedelta

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    from openpyxl import load_workbook
except ImportError:
    print("正在安装必要依赖 openpyxl...")
    os.system('pip install openpyxl -q')
    from openpyxl import load_workbook


BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, 'data', 'attendance')
OUTPUT_DIR = os.path.join(BASE_DIR, 'data', 'excel')
TEMPLATE_DIR = os.path.join(BASE_DIR, 'template')
HOLIDAY_CALENDAR_FILE = os.path.join(BASE_DIR, 'holiday-calendar.json')

# 员工考勤表（模板第 1 个工作表）关键列
COL_DAY_START = 4   # D
COL_DAY_END = 34    # AH
COL_EXPECT = 35     # AI 应出勤
COL_ACTUAL = 36     # AJ 实出勤
COL_MEAL_DAYS = 37  # AK 享受餐补天数
COL_LATE_LE30 = 38  # AL 30分钟前迟到次数
COL_LATE_GT30 = 39  # AM 30分钟后迟到次数
COL_LEAVE_START = 40  # AN
COL_LEAVE_END = 46    # AT
COL_ABSENT = 47       # AU 旷工
COL_TRAFFIC = 48      # AV 交通补贴（不填）
COL_NOT_SIGNED = 49   # AW 未打卡次数
COL_OVERTIME = 50     # AX
COL_NIGHT_OVERTIME = 51  # AY
COL_OUTSIDE = 52      # AZ 外出
COL_FULL_AWARD = 56   # BD 全勤奖
COL_MEAL_AMOUNT = 57  # BE 餐补金额
COL_REMARK = 58       # BF 备注

WEEKDAY_CN = ['一', '二', '三', '四', '五', '六', '日']
WORK_AM_START = (9, 30)
WORK_AM_END = (12, 0)
WORK_PM_START = (13, 30)
WORK_PM_END = (18, 30)
LEAVE_COL_MAP = {
    '产检假': 40,  # AN
    '病假': 41,    # AO
    '婚假': 42,    # AP
    '事假': 43,    # AQ
    '年休': 44,    # AR
    '丧假': 45,    # AS
    '调休': 46,    # AT
}


def load_latest_data():
    if not os.path.exists(DATA_DIR):
        print("❌ 数据目录不存在")
        return None

    files = [
        f for f in os.listdir(DATA_DIR)
        if f.endswith('.json') and (
            f.startswith('attendance_') or f.endswith('考勤.json')
        )
    ]
    if not files:
        print("❌ 未找到考勤数据文件")
        return None

    latest_file = max(
        (os.path.join(DATA_DIR, f) for f in files),
        key=os.path.getmtime
    )
    print(f"📁 加载数据：{latest_file}")
    with open(latest_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if 'attendanceReports' not in data:
        print("❌ 数据格式不正确，缺少 attendanceReports 字段")
        return None
    return data


def find_template_file():
    if not os.path.exists(TEMPLATE_DIR):
        return None

    preferred = os.path.join(TEMPLATE_DIR, '考勤模板.xlsx')
    if os.path.exists(preferred):
        return preferred

    candidates = sorted(
        f for f in os.listdir(TEMPLATE_DIR)
        if f.lower().endswith('.xlsx') and not f.startswith('~$')
    )
    if not candidates:
        return None
    return os.path.join(TEMPLATE_DIR, candidates[0])


def parse_dt(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def get_month_context(data):
    dt = parse_dt(data.get('startDate')) or parse_dt(data.get('workDate')) or datetime.now()
    year, month = dt.year, dt.month
    days_in_month = calendar.monthrange(year, month)[1]

    period_start = parse_dt(data.get('startDate')) or parse_dt(data.get('workDate')) or datetime(year, month, 1)
    period_end = parse_dt(data.get('endDate')) or parse_dt(data.get('workDate')) or datetime(year, month, days_in_month)
    if period_start > period_end:
        period_start, period_end = period_end, period_start

    month_start = datetime(year, month, 1)
    month_end = datetime(year, month, days_in_month)
    period_start = max(period_start, month_start)
    period_end = min(period_end, month_end)

    calendar_cfg = load_holiday_calendar()
    holidays = set(calendar_cfg.get('holidays', []))
    makeup_workdays = set(calendar_cfg.get('workdays', []))

    workday_days = set()
    cur = period_start.date()
    end_date = period_end.date()
    while cur <= end_date:
        date_str = cur.strftime('%Y-%m-%d')
        if date_str in makeup_workdays:
            workday_days.add(cur.day)
        elif date_str in holidays:
            pass
        elif cur.weekday() < 5:
            workday_days.add(cur.day)
        cur += timedelta(days=1)

    return {
        'year': year,
        'month': month,
        'days_in_month': days_in_month,
        'expected_workdays': len(workday_days),
        'workday_days': workday_days,
        'period_start': period_start,
        'period_end': period_end,
    }


def load_holiday_calendar():
    if not os.path.exists(HOLIDAY_CALENDAR_FILE):
        return {'holidays': [], 'workdays': []}

    try:
        with open(HOLIDAY_CALENDAR_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return {
            'holidays': [str(x) for x in (data.get('holidays') or [])],
            'workdays': [str(x) for x in (data.get('workdays') or [])],
        }
    except Exception:
        print(f"⚠️ 节假日配置读取失败：{HOLIDAY_CALENDAR_FILE}，将按默认工作日规则计算")
        return {'holidays': [], 'workdays': []}


def iter_dates(start_dt, end_dt):
    cur = start_dt.date()
    end_date = end_dt.date()
    while cur <= end_date:
        yield cur
        cur += timedelta(days=1)


def leave_type_to_column(leave_type):
    if not leave_type:
        return None
    text = str(leave_type)
    for key, col in LEAVE_COL_MAP.items():
        if key in text:
            return col
    return None


def normalize_leave_type(leave_type):
    t = (leave_type or '').strip()
    return t if t else '请假'


def normalize_leave_display_type(leave_type):
    text = str(leave_type or '').strip()
    if '年休' in text or '年假' in text:
        return '年假'
    return text if text else '请假'


def is_tiaoxiu_leave(leave_type):
    return '调休' in str(leave_type or '')


def parse_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def is_outside_approval(approval):
    tag_name = str(approval.get('tag_name') or '').strip()
    biz_type = approval.get('biz_type')
    return ('外出' in tag_name) or (biz_type == 2)


def is_trip_approval(approval):
    tag_name = str(approval.get('tag_name') or '').strip()
    sub_type = str(approval.get('sub_type') or '').strip()
    return ('出差' in tag_name) or ('出差' in sub_type)


def is_resident_approval(approval):
    tag_name = str(approval.get('tag_name') or '').strip()
    sub_type = str(approval.get('sub_type') or '').strip()
    if ('常驻' in tag_name) or ('常驻' in sub_type):
        return True
    duration = parse_float(approval.get('duration'), 0.0)
    duration_unit = str(approval.get('duration_unit') or '').upper()
    return duration_unit == 'DAY' and duration >= 30


def iter_unique_approvals(report):
    """钉钉在部分场景会返回重复审批，按关键字段去重。"""
    seen = set()
    for approval in report.get('approveList', []):
        key = (
            approval.get('begin_time'),
            approval.get('end_time'),
            approval.get('sub_type'),
            approval.get('tag_name'),
            approval.get('biz_type'),
            approval.get('duration'),
            approval.get('duration_unit'),
            approval.get('procInst_id'),
        )
        if key in seen:
            continue
        seen.add(key)
        yield approval


def get_day_unit_label(begin_dt, end_dt, current_date, duration, duration_unit):
    """
    将请假/外出在日维度上折算成“半天”或“一天”。
    """
    unit = str(duration_unit or '').upper()
    if begin_dt.date() == end_dt.date():
        if unit == 'DAY':
            return '一天' if duration >= 1 else '半天'
        # 小于 8 小时按半天处理
        return '一天' if duration >= 8 else '半天'

    # 跨天：中间整天按一天，首尾按时间判断
    if current_date != begin_dt.date() and current_date != end_dt.date():
        return '一天'
    if current_date == begin_dt.date():
        return '一天' if (begin_dt.hour == 0 and begin_dt.minute == 0) else '半天'
    return '一天' if (end_dt.hour >= 23 and end_dt.minute >= 59) else '半天'


def combine_day_time(day_date, hm):
    return datetime(day_date.year, day_date.month, day_date.day, hm[0], hm[1], 0)


def overlap_seconds(start1, end1, start2, end2):
    latest_start = max(start1, start2)
    earliest_end = min(end1, end2)
    delta = (earliest_end - latest_start).total_seconds()
    return max(delta, 0)


def infer_day_unit_from_worktime(day_date, seg_start, seg_end):
    """按当日工作时段重叠时长折算为半天/一天。"""
    am_start = combine_day_time(day_date, WORK_AM_START)
    am_end = combine_day_time(day_date, WORK_AM_END)
    pm_start = combine_day_time(day_date, WORK_PM_START)
    pm_end = combine_day_time(day_date, WORK_PM_END)

    am_overlap = overlap_seconds(seg_start, seg_end, am_start, am_end)
    pm_overlap = overlap_seconds(seg_start, seg_end, pm_start, pm_end)
    total_work_overlap = am_overlap + pm_overlap

    # 覆盖一个完整工作日(7.5h)的大部分或上午+下午都有覆盖时，按一天。
    if total_work_overlap >= (7.5 * 3600 * 0.8):
        return '一天'
    if am_overlap >= (2.5 * 3600 * 0.8) and pm_overlap >= (5 * 3600 * 0.8):
        return '一天'
    return '半天'


def leave_unit_value(leave_unit):
    return '1' if leave_unit == '一天' else '0.5'


def blank_if_zero(value):
    if value is None:
        return None
    if isinstance(value, (int, float)) and abs(value) < 1e-9:
        return None
    return value


def build_user_stats(report, year, month, days_in_month, expected_workdays, workday_days):
    daily = {
        day: {
            'late_items': [],
            'leave_items': [],
            'outside_items': [],
            'not_signed_types': set(),
            'weekend_overtime_value': 0.0,
            'weekend_check_types': set(),
            'trip_type': None,  # 出差/常驻
            'absent': False,
        }
        for day in range(1, days_in_month + 1)
    }

    leave_day_type_counter = Counter()
    leave_days = set()
    penalizing_leave_days = set()  # 影响全勤奖的请假天（调休不计入）
    outside_days = set()
    weekend_overtime_days = set()
    weekend_overtime_total = 0.0
    trip_days = set()
    resident_days = set()
    absent_days = set()
    late_before_30 = 0
    late_after_30 = 0

    # 先预计算请假时段（上午/下午）用于豁免迟到和缺卡
    leave_slot_flags = {
        day: {'am_leave': False, 'pm_leave': False}
        for day in range(1, days_in_month + 1)
    }

    for approval in iter_unique_approvals(report):
        if is_outside_approval(approval):
            continue

        begin_dt = parse_dt(approval.get('begin_time'))
        end_dt = parse_dt(approval.get('end_time'))
        if not begin_dt or not end_dt:
            continue

        month_start = datetime(year, month, 1).date()
        month_end = datetime(year, month, days_in_month).date()
        overlap_start = max(begin_dt.date(), month_start)
        overlap_end = min(end_dt.date(), month_end)
        if overlap_start > overlap_end:
            continue

        for d in iter_dates(
            datetime.combine(overlap_start, datetime.min.time()),
            datetime.combine(overlap_end, datetime.min.time()),
        ):
            day = d.day
            day_start = datetime.combine(d, datetime.min.time())
            day_end = datetime.combine(d, datetime.max.time())
            seg_start = max(begin_dt, day_start)
            seg_end = min(end_dt, day_end)
            if seg_start >= seg_end:
                continue

            am_start = combine_day_time(d, WORK_AM_START)
            am_end = combine_day_time(d, WORK_AM_END)
            pm_start = combine_day_time(d, WORK_PM_START)
            pm_end = combine_day_time(d, WORK_PM_END)

            am_overlap = overlap_seconds(seg_start, seg_end, am_start, am_end)
            pm_overlap = overlap_seconds(seg_start, seg_end, pm_start, pm_end)

            # 上午 2.5h，下午 5h。覆盖 >= 50% 即视为该半天请假。
            if am_overlap >= (2.5 * 3600 * 0.5):
                leave_slot_flags[day]['am_leave'] = True
            if pm_overlap >= (5 * 3600 * 0.5):
                leave_slot_flags[day]['pm_leave'] = True

    for att in report.get('attendanceList', []):
        plan_dt = parse_dt(att.get('plan_check_time'))
        actual_dt = parse_dt(att.get('user_check_time'))
        dt = plan_dt or actual_dt
        if not dt or dt.year != year or dt.month != month:
            continue

        day = dt.day
        if day < 1 or day > days_in_month:
            continue

        # 周末打卡：双次打卡记加班1，单次打卡记加班0.5
        if dt.weekday() >= 5 and (att.get('user_check_time') or ''):
            check_type = (att.get('check_type') or '').strip()
            if check_type in ('OnDuty', 'OffDuty'):
                daily[day]['weekend_check_types'].add(check_type)
            else:
                daily[day]['weekend_check_types'].add('UNKNOWN')

        time_result = (att.get('time_result') or '').strip()
        check_type = (att.get('check_type') or '').strip()
        am_leave = leave_slot_flags[day]['am_leave']
        pm_leave = leave_slot_flags[day]['pm_leave']

        # 上午请假：上班迟到不计入
        if time_result == 'Late' and check_type == 'OnDuty' and am_leave:
            continue

        if time_result == 'Late':
            minutes = None
            if plan_dt and actual_dt:
                delta_minutes = int((actual_dt - plan_dt).total_seconds() // 60)
                minutes = max(delta_minutes, 0)
            daily[day]['late_items'].append({
                'minutes': minutes,
                'actual': actual_dt.strftime('%H:%M') if actual_dt else '',
            })
            if minutes is not None and minutes > 30:
                late_after_30 += 1
            else:
                late_before_30 += 1

        if time_result == 'NotSigned':
            # 上午请假：上班未打卡不计入；下午请假：下班未打卡不计入
            if check_type == 'OnDuty' and am_leave:
                continue
            if check_type == 'OffDuty' and pm_leave:
                continue

            if check_type == 'OnDuty':
                daily[day]['not_signed_types'].add('上午')
            elif check_type == 'OffDuty':
                daily[day]['not_signed_types'].add('下午')
            else:
                daily[day]['not_signed_types'].add('未标记')

        outside_remark = (att.get('outside_remark') or '').strip()
        if outside_remark:
            outside_days.add(day)
            daily[day]['outside_items'].append(outside_remark)

    for approval in iter_unique_approvals(report):
        begin_dt = parse_dt(approval.get('begin_time'))
        end_dt = parse_dt(approval.get('end_time'))
        if not begin_dt or not end_dt:
            continue

        is_outside = is_outside_approval(approval)
        is_resident = is_resident_approval(approval)
        is_trip = is_trip_approval(approval) and not is_resident
        leave_type = normalize_leave_type(approval.get('sub_type'))
        col = leave_type_to_column(leave_type)
        duration = parse_float(approval.get('duration'), 0.0)
        duration_unit = approval.get('duration_unit')

        month_start = datetime(year, month, 1).date()
        month_end = datetime(year, month, days_in_month).date()
        overlap_start = max(begin_dt.date(), month_start)
        overlap_end = min(end_dt.date(), month_end)
        if overlap_start > overlap_end:
            continue

        for d in iter_dates(
            datetime.combine(overlap_start, datetime.min.time()),
            datetime.combine(overlap_end, datetime.min.time()),
        ):
            day = d.day
            day_start = datetime.combine(d, datetime.min.time())
            day_end = datetime.combine(d, datetime.max.time())
            seg_start = max(begin_dt, day_start)
            seg_end = min(end_dt, day_end)
            if seg_start >= seg_end:
                continue

            if str(duration_unit or '').upper() == 'DAY' and duration >= 1:
                day_unit = '一天'
            else:
                day_unit = infer_day_unit_from_worktime(d, seg_start, seg_end)
            day_value = 0.5 if day_unit == '半天' else 1

            if is_resident:
                resident_days.add(day)
                daily[day]['trip_type'] = '常驻'
            elif is_trip:
                trip_days.add(day)
                if daily[day]['trip_type'] != '常驻':
                    daily[day]['trip_type'] = '出差'
            elif is_outside:
                outside_days.add(day)
                daily[day]['outside_items'].append('外出')
            else:
                leave_days.add(day)
                if not is_tiaoxiu_leave(leave_type):
                    penalizing_leave_days.add(day)
                if col:
                    leave_day_type_counter[col] += day_value
                daily[day]['leave_items'].append({'type': leave_type, 'unit': day_unit})

    not_signed_total = 0
    for day in range(1, days_in_month + 1):
        check_types = daily[day]['weekend_check_types']
        if check_types:
            if 'OnDuty' in check_types and 'OffDuty' in check_types:
                daily[day]['weekend_overtime_value'] = 1.0
            else:
                daily[day]['weekend_overtime_value'] = 0.5
            weekend_overtime_days.add(day)
            weekend_overtime_total += daily[day]['weekend_overtime_value']

        # 规则：请假当天不记录未打卡次数，也不显示未打卡文案
        if daily[day]['leave_items']:
            daily[day]['not_signed_types'] = set()

        signed_types = daily[day]['not_signed_types']
        not_signed_total += len(signed_types)
        if ('上午' in signed_types and '下午' in signed_types and day not in leave_days and day not in outside_days):
            daily[day]['absent'] = True
            absent_days.add(day)

    late_total = late_before_30 + late_after_30
    leave_days_on_workday = {d for d in leave_days if d in workday_days}
    absent_days_on_workday = {d for d in absent_days if d in workday_days}
    actual_workdays = max(expected_workdays - len(leave_days_on_workday) - len(absent_days_on_workday), 0)

    meal_days = 0
    meal_amount = 0
    for day in range(1, days_in_month + 1):
        if day not in workday_days:
            continue
        if daily[day]['leave_items'] or daily[day]['absent']:
            continue
        meal_days += 1
        if daily[day]['trip_type'] == '出差':
            meal_amount += 30
        else:
            # 常驻或普通出勤均按 12 元/天
            meal_amount += 12

    # 全勤奖规则：调休不影响全勤奖，其他请假/迟到/旷工会影响
    full_award = 200 if (late_total == 0 and not penalizing_leave_days and not absent_days) else 0

    return {
        'daily': daily,
        'expected_workdays': expected_workdays,
        'actual_workdays': actual_workdays,
        'meal_days': meal_days,
        'meal_amount': meal_amount,
        'late_before_30': late_before_30,
        'late_after_30': late_after_30,
        'leave_counter': leave_day_type_counter,
        'absent_days': absent_days,
        'outside_days': outside_days,
        'weekend_overtime_days': weekend_overtime_days,
        'weekend_overtime_total': weekend_overtime_total,
        'trip_days': trip_days,
        'resident_days': resident_days,
        'not_signed_total': not_signed_total,
        'full_award': full_award,
    }


def build_daily_cell_text(day_data):
    parts = []

    if day_data.get('trip_type') in ('出差', '常驻'):
        parts.append(day_data['trip_type'])

    overtime_value = day_data.get('weekend_overtime_value', 0.0)
    if overtime_value > 0:
        parts.append('加班1' if overtime_value >= 1 else '加班0.5')

    for late in day_data['late_items']:
        minutes = late.get('minutes')
        actual = late.get('actual')
        if minutes is None:
            parts.append(f"迟到({actual})" if actual else '迟到')
        else:
            parts.append(f"迟到{minutes}分({actual})" if actual else f"迟到{minutes}分")

    for leave in day_data['leave_items']:
        leave_type = normalize_leave_display_type(leave.get('type', '请假'))
        leave_unit = leave.get('unit', '半天')
        parts.append(f"{leave_type}{leave_unit_value(leave_unit)}")

    if day_data['outside_items']:
        parts.append('外出')

    if day_data['absent']:
        parts.append('旷工')
    else:
        # 只有单次未打卡时显示上班/下班未打卡；双次未打卡已归并为旷工
        if '上午' in day_data['not_signed_types']:
            parts.append('上班未打卡')
        if '下午' in day_data['not_signed_types']:
            parts.append('下班未打卡')

    return '；'.join(parts)


def clear_row_values(ws, row):
    for col in range(COL_DAY_START, COL_REMARK + 1):
        ws.cell(row=row, column=col, value=None)


def copy_row_style(ws, src_row, dst_row, max_col=COL_REMARK):
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)
        if src_cell.number_format:
            dst_cell.number_format = src_cell.number_format
        if src_cell.protection:
            dst_cell.protection = copy(src_cell.protection)
        if src_cell.alignment:
            dst_cell.alignment = copy(src_cell.alignment)

    if src_row in ws.row_dimensions and ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def update_header(ws, year, month, days_in_month):
    ws['A1'] = f"{year}{month}月在职员工出勤统计表"

    for day in range(1, 32):
        col = COL_DAY_START + day - 1
        if day <= days_in_month:
            ws.cell(row=2, column=col, value=day)
            weekday_idx = datetime(year, month, day).weekday()
            ws.cell(row=3, column=col, value=WEEKDAY_CN[weekday_idx])
        else:
            # 强制写空字符串，避免模板残留日期（例如 4 月出现 31 号）
            ws.cell(row=2, column=col, value='')
            ws.cell(row=3, column=col, value='')


def fill_employee_row(ws, row, stats, days_in_month):
    clear_row_values(ws, row)

    for day in range(1, 32):
        col = COL_DAY_START + day - 1
        if day <= days_in_month:
            text = build_daily_cell_text(stats['daily'][day])
            ws.cell(row=row, column=col, value=text if text else None)
        else:
            ws.cell(row=row, column=col, value=None)

    ws.cell(row=row, column=COL_EXPECT, value=stats['expected_workdays'])
    ws.cell(row=row, column=COL_ACTUAL, value=stats['actual_workdays'])
    ws.cell(row=row, column=COL_MEAL_DAYS, value=blank_if_zero(stats['meal_days']))
    ws.cell(row=row, column=COL_LATE_LE30, value=blank_if_zero(stats['late_before_30']))
    ws.cell(row=row, column=COL_LATE_GT30, value=blank_if_zero(stats['late_after_30']))

    for col in range(COL_LEAVE_START, COL_LEAVE_END + 1):
        ws.cell(row=row, column=col, value=blank_if_zero(stats['leave_counter'].get(col, 0)))

    ws.cell(row=row, column=COL_ABSENT, value=blank_if_zero(len(stats['absent_days'])))
    ws.cell(row=row, column=COL_TRAFFIC, value=None)  # 交通补贴按要求留空
    ws.cell(row=row, column=COL_NOT_SIGNED, value=blank_if_zero(stats['not_signed_total']))
    ws.cell(row=row, column=COL_OVERTIME, value=blank_if_zero(stats.get('weekend_overtime_total', 0)))
    ws.cell(row=row, column=COL_NIGHT_OVERTIME, value=None)
    ws.cell(row=row, column=COL_OUTSIDE, value=blank_if_zero(len(stats['outside_days'])))
    ws.cell(row=row, column=COL_FULL_AWARD, value=blank_if_zero(stats['full_award']))
    ws.cell(row=row, column=COL_MEAL_AMOUNT, value=blank_if_zero(stats['meal_amount']))
    ws.cell(row=row, column=COL_REMARK, value=None)


def create_excel(data, output_file):
    print("\n📊 正在按模板生成 Excel 文件...")

    template_file = find_template_file()
    if not template_file:
        print("❌ 未找到模板文件，请将 .xlsx 模板放到 template 目录")
        return None

    print(f"📄 使用模板：{template_file}")
    wb = load_workbook(template_file)
    ws = wb.worksheets[0]

    ctx = get_month_context(data)
    year = ctx['year']
    month = ctx['month']
    days_in_month = ctx['days_in_month']
    expected_workdays = ctx['expected_workdays']

    update_header(ws, year, month, days_in_month)

    reports = data.get('attendanceReports', [])
    report_by_name = {str(r.get('name', '')).strip(): r for r in reports if r.get('name')}

    # 优先使用 users 全量名单，确保无考勤记录员工也会导出
    all_names = []
    seen_names = set()
    for user in data.get('users', []):
        name = str(user.get('name', '')).strip()
        if name and name not in seen_names:
            seen_names.add(name)
            all_names.append(name)

    # 补齐 users 中未出现但在 attendanceReports 出现的员工
    for name in report_by_name.keys():
        if name and name not in seen_names:
            seen_names.add(name)
            all_names.append(name)

    existing_rows = {}
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row=row, column=3).value
        name = str(name).strip() if name else ''
        if name:
            existing_rows[name] = row

    start_row = min(existing_rows.values()) if existing_rows else 4
    style_row = start_row
    required_last_row = start_row + max(len(all_names), 1) - 1
    template_last_named_row = max(existing_rows.values()) if existing_rows else start_row
    clear_until = max(required_last_row, template_last_named_row)

    # 先清空将要使用的区域，避免模板残留影响
    for row in range(start_row, clear_until + 1):
        if row != style_row:
            copy_row_style(ws, style_row, row)
        ws.cell(row=row, column=1, value='=ROW()-3')
        ws.cell(row=row, column=2, value='')
        ws.cell(row=row, column=3, value='')
        clear_row_values(ws, row)

    # 顺序填充全部员工（即使没有任何考勤记录也会保留员工行）
    for idx, name in enumerate(all_names):
        row = start_row + idx
        ws.cell(row=row, column=1, value='=ROW()-3')
        ws.cell(row=row, column=2, value='')
        ws.cell(row=row, column=3, value=name)

        report = report_by_name.get(name, {'name': name, 'attendanceList': [], 'approveList': []})
        stats = build_user_stats(report, year, month, days_in_month, expected_workdays, ctx['workday_days'])
        fill_employee_row(ws, row, stats, days_in_month)

    print(f"👥 已导出员工总数：{len(all_names)}")

    print(f"💾 保存文件：{output_file}")
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
        except PermissionError:
            print("⚠️ 输出文件被占用，无法覆盖")
            return None

    wb.save(output_file)
    print("✅ Excel 文件生成成功！")
    return output_file


def build_output_name_base(data):
    start_dt = parse_dt(data.get('startDate')) or parse_dt(data.get('workDate'))
    end_dt = parse_dt(data.get('endDate')) or parse_dt(data.get('workDate')) or start_dt

    if not start_dt or not end_dt:
        now = datetime.now()
        return f"{now.year}.{now.month}月考勤"

    if start_dt.date() == end_dt.date():
        return f"{start_dt.strftime('%Y-%m-%d')}考勤"

    is_whole_month = (
        start_dt.year == end_dt.year and
        start_dt.month == end_dt.month and
        start_dt.day == 1 and
        end_dt.day == calendar.monthrange(end_dt.year, end_dt.month)[1]
    )

    if is_whole_month:
        return f"{start_dt.year}.{start_dt.month}月考勤"

    return f"{start_dt.strftime('%Y-%m-%d')}至{end_dt.strftime('%Y-%m-%d')}考勤"


def export_to_excel(input_file=None):
    print("=" * 60)
    print("📊 考勤数据导出 Excel（模板版）")
    print("=" * 60)

    if input_file:
        print(f"📁 加载数据：{input_file}")
        with open(input_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = load_latest_data()

    if not data:
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    output_file = os.path.join(OUTPUT_DIR, f'{build_output_name_base(data)}.xlsx')

    result = create_excel(data, output_file)
    if result:
        print("\n" + "=" * 60)
        print(f"📁 文件位置：{result}")
        print("=" * 60)


if __name__ == '__main__':
    input_file = sys.argv[1] if len(sys.argv) > 1 else None
    export_to_excel(input_file)
