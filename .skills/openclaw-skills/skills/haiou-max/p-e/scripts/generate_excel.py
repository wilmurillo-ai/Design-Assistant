#!/usr/bin/env python3
"""
Generate Excel file from extracted data.
从提取的数据生成 Excel 表格。

Usage:
    python generate_excel.py --json <data.json> [--output <path>]

Options:
    --json <file>      Input JSON file with extracted data
    --output <path>    Output Excel file path (default: product_list.xlsx)

JSON Format:
    {
        "fields": ["field1", "field2", "field3"],
        "data": [
            {"field1": "value1", "field2": "value2", "field3": "value3"},
            ...
        ]
    }
"""

import json
import sys
import os
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)


def generate_excel(fields, data, output_path="product_list.xlsx"):
    """Generate Excel workbook with extracted data."""

    if not data:
        print("Error: No data to process")
        return False

    if not fields:
        print("Error: No fields defined")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header
    for col_idx, col_name in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    row_height = 190
    for row_idx, row_data in enumerate(data, 2):
        ws.row_dimensions[row_idx].height = row_height

        for col_idx, col_name in enumerate(fields, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data.get(col_name, "") if isinstance(row_data, dict) else ""

            # Special handling for image column
            if col_name.lower() in ["image", "图片", "图"] and value:
                image_path = str(value)
                if os.path.exists(image_path):
                    try:
                        xl_image = XLImage(image_path)
                        xl_image.width = 320
                        xl_image.height = 240
                        ws.add_image(xl_image, cell.coordinate)
                    except Exception as e:
                        print(f"Warning: Failed to embed image {image_path}: {e}")
                        cell.value = "[Image failed]"
                else:
                    cell.value = "[File not found]"
            else:
                cell.value = value
                cell.border = thin_border

                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Adjust column widths
    for col_idx, col_name in enumerate(fields, 1):
        if col_name.lower() in ["image", "图片", "图"]:
            width = 55
        else:
            width = 15
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Save
    try:
        wb.save(output_path)
        print(f"✓ Excel generated: {output_path}")
        print(f"  Rows: {len(data)}")
        print(f"  Columns: {len(fields)}")
        return True
    except Exception as e:
        print(f"Error: Failed to save Excel: {e}")
        return False


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Excel from extracted data")
    parser.add_argument("--json", required=True, help="Input JSON file with extracted data")
    parser.add_argument("--output", default="product_list.xlsx", help="Output Excel file path")

    args = parser.parse_args()

    try:
        with open(args.json, 'r', encoding='utf-8') as f:
            input_data = json.load(f)

        if isinstance(input_data, dict):
            fields = input_data.get("fields", [])
            data = input_data.get("data", [])
        elif isinstance(input_data, list):
            if not input_data:
                print("Error: Empty data list")
                sys.exit(1)
            first_item = input_data[0]
            fields = list(first_item.keys()) if isinstance(first_item, dict) else []
            data = input_data
        else:
            print("Error: Invalid JSON format")
            sys.exit(1)

        if not fields:
            print("Error: No fields found in JSON")
            sys.exit(1)

        print(f"✓ Loaded {len(data)} rows with {len(fields)} fields")

        success = generate_excel(fields, data, args.output)
        sys.exit(0 if success else 1)

    except FileNotFoundError:
        print(f"Error: JSON file not found: {args.json}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
