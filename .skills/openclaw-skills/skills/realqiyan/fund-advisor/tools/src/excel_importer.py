"""
基金持仓管理系统 - Excel导入功能
"""
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook

from src.models import FundHolding, FundType
from src.database import Database


class ExcelImporter:
    """Excel导入器"""

    # Excel列名映射（与CSV保持一致）
    COLUMN_MAPPING = {
        "基金代码": "fund_code",
        "基金名称": "fund_name",
        "份额类别": "share_class",
        "基金管理人": "fund_manager",
        "基金账户": "fund_account",
        "销售机构": "sales_agency",
        "交易账户": "trade_account",
        "持有份额": "holding_shares",
        "份额日期": "share_date",
        "基金净值": "nav",
        "净值日期": "nav_date",
        "资产情况（结算币种）": "asset_value",
        "结算币种": "settlement_currency",
        "分红方式": "dividend_method"
    }

    # 列名别名映射（处理换行符、空格等变体）
    COLUMN_ALIASES = {
        "asset_value": ["资产情况（结算币种）", "资产情况\n（结算币种）", "资产情况 （结算币种）"]
    }

    def __init__(self, database: Database):
        self.database = database

    def _find_header_row(self, worksheet) -> int:
        """
        查找包含列标题的行号

        Args:
            worksheet: openpyxl worksheet对象

        Returns:
            列标题所在行号（1-based，与openpyxl一致），未找到返回-1
        """
        header_keywords = ["基金代码", "基金名称", "基金账户"]
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row is None:
                continue
            row_str = str(row)
            match_count = sum(1 for kw in header_keywords if kw in row_str)
            if match_count >= 2:
                return row_idx
        return -1

    def _is_data_row(self, row: dict) -> bool:
        """
        判断是否为有效数据行

        Args:
            row: Excel行数据（字典形式）

        Returns:
            是否为有效数据行
        """
        if not row:
            return False

        fund_code = str(row.get("基金代码", "")).strip()
        if not fund_code:
            return False

        if fund_code in ["序号", "总记录数", "数据来源", "说明", "数量单位", "免责声明"]:
            return False

        try:
            float(fund_code)
            return True
        except ValueError:
            return fund_code and not fund_code.startswith("总记录数")

    def import_from_excel(self, excel_path: str) -> Tuple[int, int, List[str]]:
        """
        从Excel文件导入基金持有信息

        Args:
            excel_path: Excel文件路径（支持.xlsx和.xls）

        Returns:
            (成功数量, 失败数量, 错误信息列表)
        """
        success_count = 0
        fail_count = 0
        errors = []

        excel_file = Path(excel_path)
        if not excel_file.exists():
            return 0, 0, [f"文件不存在: {excel_path}"]

        # 检查文件格式
        suffix = excel_file.suffix.lower()
        if suffix not in [".xlsx", ".xls"]:
            return 0, 0, [f"不支持的文件格式: {suffix}，仅支持.xlsx和.xls"]

        try:
            # 使用openpyxl读取Excel文件
            workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
            worksheet = workbook.active

            if worksheet is None:
                return 0, 0, ["Excel文件中没有工作表"]

            # 查找表头行
            header_row_idx = self._find_header_row(worksheet)
            if header_row_idx < 0:
                return 0, 0, ["无法找到有效的列标题行"]

            # 获取表头
            header_row = list(worksheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
            headers = [str(h).strip() if h else "" for h in header_row]

            # 标准化表头（处理换行符）
            normalized_headers = []
            for h in headers:
                normalized = h.replace("\n", "").replace("\r", "").strip()
                # 处理特殊情况："资产情况（结算币种）"中间可能有空格
                if "资产情况" in normalized and "结算币种" in normalized:
                    normalized = "资产情况（结算币种）"
                normalized_headers.append(normalized)

            # 读取数据行
            actual_row_num = header_row_idx + 1
             # 清空所有持仓记录
            self.database.clear_all_holdings()
            for row in worksheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if row is None or all(cell is None for cell in row):
                    actual_row_num += 1
                    continue

                # 构建行字典
                row_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(normalized_headers):
                        row_dict[normalized_headers[idx]] = str(cell).strip() if cell is not None else ""

                if not self._is_data_row(row_dict):
                    actual_row_num += 1
                    continue

                try:
                    holding = self._parse_row(row_dict, actual_row_num)
                    if holding:
                        self.database.upsert_fund_holding(holding)
                        success_count += 1
                except Exception as e:
                    fail_count += 1
                    errors.append(f"第{actual_row_num}行: {str(e)}")

                actual_row_num += 1

            workbook.close()

        except Exception as e:
            return 0, 0, [f"读取Excel文件失败: {str(e)}"]

        return success_count, fail_count, errors

    def _parse_row(self, row: dict, row_num: int) -> Optional[FundHolding]:
        """解析Excel行数据"""
        # 映射列名
        mapped_row = {}
        for csv_col, model_col in self.COLUMN_MAPPING.items():
            value = row.get(csv_col, "").strip()
            if not value:
                # 尝试其他可能的列名
                for key in row.keys():
                    if csv_col in key or key in csv_col:
                        value = row.get(key, "").strip()
                        break
            mapped_row[model_col] = value

        # 处理列名别名（如包含换行符的列名）
        for model_col, aliases in self.COLUMN_ALIASES.items():
            if not mapped_row.get(model_col):
                for alias in aliases:
                    # 直接匹配
                    if alias in row:
                        mapped_row[model_col] = row[alias].strip()
                        break
                    # 清理换行符后匹配
                    for key in row.keys():
                        cleaned_key = key.replace("\n", "").replace("\r", "")
                        cleaned_alias = alias.replace("\n", "").replace("\r", "")
                        if cleaned_key == cleaned_alias:
                            mapped_row[model_col] = row[key].strip()
                            break

        # 验证必填字段
        required_fields = ["fund_code", "fund_name", "fund_account", "trade_account",
                          "holding_shares", "share_date", "nav", "nav_date", "asset_value"]

        for field in required_fields:
            if not mapped_row.get(field):
                raise ValueError(f"缺少必填字段: {field}")

        # 解析数值字段
        try:
            holding_shares = float(str(mapped_row["holding_shares"]).replace(",", ""))
        except ValueError:
            raise ValueError(f"无效的持有份额: {mapped_row['holding_shares']}")

        try:
            nav = float(mapped_row["nav"])
        except ValueError:
            raise ValueError(f"无效的基金净值: {mapped_row['nav']}")

        try:
            asset_value = float(str(mapped_row["asset_value"]).replace(",", ""))
        except ValueError:
            raise ValueError(f"无效的资产情况: {mapped_row['asset_value']}")

        # 解析日期字段
        share_date = self._parse_date(mapped_row["share_date"])
        if not share_date:
            raise ValueError(f"无效的份额日期: {mapped_row['share_date']}")

        nav_date = self._parse_date(mapped_row["nav_date"])
        if not nav_date:
            raise ValueError(f"无效的净值日期: {mapped_row['nav_date']}")

        return FundHolding(
            fund_code=mapped_row["fund_code"],
            fund_name=mapped_row["fund_name"],
            share_class=mapped_row.get("share_class", ""),
            fund_manager=mapped_row.get("fund_manager", ""),
            fund_account=mapped_row["fund_account"],
            sales_agency=mapped_row.get("sales_agency", ""),
            trade_account=mapped_row["trade_account"],
            holding_shares=holding_shares,
            share_date=share_date,
            nav=nav,
            nav_date=nav_date,
            asset_value=asset_value,
            settlement_currency=mapped_row.get("settlement_currency", "人民币"),
            dividend_method=mapped_row.get("dividend_method", ""),
            fund_type=FundType.PUBLIC_FUND
        )

    def _parse_date(self, date_value) -> Optional[date]:
        """解析日期值（可能是字符串或datetime对象）"""
        if not date_value:
            return None

        # 如果是datetime或date对象
        if isinstance(date_value, datetime):
            return date_value.date()
        if isinstance(date_value, date):
            return date_value

        # 字符串处理
        date_str = str(date_value).strip()

        # 尝试多种日期格式
        formats = [
            "%Y/%m/%d",
            "%Y-%m-%d",
            "%Y.%m.%d",
            "%Y年%m月%d日"
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        return None

    def validate_excel(self, excel_path: str) -> Tuple[bool, List[str]]:
        """
        验证Excel文件格式

        Args:
            excel_path: Excel文件路径

        Returns:
            (是否有效, 错误信息列表)
        """
        errors = []
        excel_file = Path(excel_path)

        if not excel_file.exists():
            return False, [f"文件不存在: {excel_path}"]

        suffix = excel_file.suffix.lower()
        if suffix not in [".xlsx", ".xls"]:
            return False, [f"不是Excel文件: {excel_path}，仅支持.xlsx和.xls格式"]

        # 检查必需的列
        required_columns = ["基金代码", "基金名称", "基金账户", "交易账户",
                           "持有份额", "份额日期", "基金净值", "净值日期"]

        try:
            workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
            worksheet = workbook.active

            if worksheet is None:
                return False, ["Excel文件中没有工作表"]

            # 查找表头行
            header_row_idx = self._find_header_row(worksheet)
            if header_row_idx < 0:
                return False, ["无法找到有效的列标题行"]

            # 获取表头
            header_row = list(worksheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
            headers = [str(h).strip() if h else "" for h in header_row]

            # 标准化表头
            normalized_headers = []
            for h in headers:
                normalized = h.replace("\n", "").replace("\r", "").strip()
                if "资产情况" in normalized and "结算币种" in normalized:
                    normalized = "资产情况（结算币种）"
                normalized_headers.append(normalized)

            # 检查必需列
            missing_columns = []
            for col in required_columns:
                found = False
                for h in normalized_headers:
                    if col in h or h in col:
                        found = True
                        break
                if not found:
                    missing_columns.append(col)

            if missing_columns:
                errors.append(f"缺少必需列: {', '.join(missing_columns)}")
                return False, errors

            workbook.close()
            return True, []

        except Exception as e:
            errors.append(f"读取Excel文件失败: {str(e)}")
            return False, errors