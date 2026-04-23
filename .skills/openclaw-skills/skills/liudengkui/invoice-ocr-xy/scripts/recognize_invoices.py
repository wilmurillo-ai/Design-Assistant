#!/usr/bin/env python3
"""
发票 OCR 识别脚本
使用翔云 OCR API 识别发票并提取信息
支持混贴发票识别，自动识别 40+ 种发票类型
"""

import os
import sys
import io
import json
import base64
import urllib.request
import urllib.error
from pathlib import Path
from datetime import datetime

# Windows 控制台 GBK 编码不支持 emoji，强制 stdout/stderr 使用 UTF-8
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# 支持的文件格式
SUPPORTED_FORMATS = {'.pdf', '.ofd', '.jpg', '.jpeg', '.png', '.bmp'}

# 翔云混贴发票识别 typeId（固定值）
NETOCR_TYPEID = "20090"

# API 端点（官方拼写，保留 Invoive 拼写）
NETOCR_ENDPOINT = "https://netocr.com/api/v2/recogInvoiveBase64.do"


# ─────────────────────────── 配置管理 ───────────────────────────

def get_config_path() -> Path:
    return Path(__file__).parent.parent / "config.json"


def load_config() -> dict:
    """加载配置文件"""
    config_path = get_config_path()
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_config(config: dict):
    """保存配置文件"""
    config_path = get_config_path()
    config_path.parent.mkdir(parents=True, exist_ok=True)
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2, ensure_ascii=False)


# ─────────────────────────── API 调用 ───────────────────────────

def call_netocr_api(image_data: bytes, netocr_key: str, netocr_secret: str) -> dict:
    """
    调用翔云 OCR API（Base64 模式，标准 multipart/form-data）

    修复说明：
    - 原脚本使用字符串拼接构造 multipart body，CRLF 处理不规范，
      导致服务端解析失败（400 或识别错误）。
    - 现改为逐部分以字节流拼接，严格遵循 RFC 2046 规范。
    """
    img_b64 = base64.b64encode(image_data).decode('ascii')
    boundary = "----XYBoundary7MA4YWxkTrZu0gW"
    CRLF = b"\r\n"

    def make_part(name: str, value: str) -> bytes:
        return (
            f"--{boundary}\r\n"
            f"Content-Disposition: form-data; name=\"{name}\"\r\n"
            f"\r\n"
            f"{value}\r\n"
        ).encode('utf-8')

    body = b"".join([
        make_part("key",    netocr_key),
        make_part("secret", netocr_secret),
        make_part("typeId", NETOCR_TYPEID),
        make_part("format", "json"),
        make_part("img",    img_b64),
        f"--{boundary}--\r\n".encode('utf-8'),
    ])

    try:
        req = urllib.request.Request(NETOCR_ENDPOINT, data=body, method='POST')
        req.add_header('Content-Type', f'multipart/form-data; boundary={boundary}')
        req.add_header('Content-Length', str(len(body)))

        with urllib.request.urlopen(req, timeout=60) as response:
            raw = response.read().decode('utf-8', errors='replace')
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                return {"Code": "ParseError", "Message": "响应 JSON 解析失败", "raw": raw}

    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8', errors='replace')
        return {"Code": "HttpError", "Message": f"HTTP {e.code}: {error_body}"}
    except urllib.error.URLError as e:
        return {"Code": "NetworkError", "Message": f"网络错误: {e.reason}"}
    except Exception as e:
        return {"Code": "Error", "Message": str(e)}


# ─────────────────────────── 文件处理 ───────────────────────────

def collect_files(path: str) -> list:
    """
    收集待识别文件。
    修复说明：原脚本只支持文件夹，单文件路径会被当作目录处理而报错。
    现在同时支持单个文件和文件夹。
    """
    p = Path(path)
    if not p.exists():
        print(f"❌ 路径不存在: {path}")
        return []

    if p.is_file():
        if p.suffix.lower() not in SUPPORTED_FORMATS:
            print(f"❌ 不支持的文件格式: {p.suffix}")
            print(f"   支持的格式: {', '.join(sorted(SUPPORTED_FORMATS))}")
            return []
        return [str(p)]

    # 目录：递归扫描
    files = sorted([
        str(f) for f in p.rglob('*')
        if f.is_file() and f.suffix.lower() in SUPPORTED_FORMATS
    ])
    return files


# ─────────────────────────── 发票识别 ───────────────────────────

def parse_invoice_data(data) -> list:
    """
    解析翔云返回的 data 字段，标准化为发票信息列表。

    翔云实际响应结构示例：
    {
      "code": 0,
      "data": [
        {
          "invoice": { "invoiceType": 1, "invoiceCode": "...", ... },
          "dataMsg": "...",
          ...
        }
      ]
    }
    每个 data 列表项中，发票核心字段在 "invoice" 子字典内。
    本函数将其展开并合并外层字段（如 dataMsg、confidence 等）。
    """
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except json.JSONDecodeError:
            return []

    if isinstance(data, dict):
        # 兼容旧版直接返回发票字典 或 含 invoiceList 的格式
        if "invoiceList" in data:
            return parse_invoice_data(data["invoiceList"])
        # 含 invoice 子字典：展开
        if "invoice" in data:
            merged = {**data.get("invoice", {})}
            for k, v in data.items():
                if k != "invoice":
                    merged[f"_ext_{k}"] = v
            return [merged]
        return [data]

    if isinstance(data, list):
        results = []
        for item in data:
            results.extend(parse_invoice_data(item))
        return results

    return []


def recognize_invoice(file_path: str, netocr_key: str, netocr_secret: str) -> list:
    """识别单个发票文件，返回发票信息列表"""
    file_name = os.path.basename(file_path)
    print(f"  正在识别: {file_name} ...", end=" ", flush=True)

    try:
        with open(file_path, 'rb') as f:
            image_data = f.read()
    except IOError as e:
        print(f"❌ 读取文件失败: {e}")
        return [{"文件名": file_name, "状态": "读取失败", "错误信息": str(e)}]

    ocr_result = call_netocr_api(image_data, netocr_key, netocr_secret)

    # 检查 API 级别错误
    # 翔云实际返回小写 code（整数），0 = 成功；同时兼容旧版大写 Code 字符串格式
    api_code = ocr_result.get("code", ocr_result.get("Code", 0))
    api_msg  = ocr_result.get("msg",  ocr_result.get("Message", ""))
    # 非 0 且非 "Success" 均视为失败
    is_error = (isinstance(api_code, int) and api_code != 0) or \
               (isinstance(api_code, str) and api_code not in ("", "0", "Success"))
    if is_error:
        err_detail = ocr_result.get("raw", api_msg or "未知错误")
        print(f"❌ 识别失败 [code={api_code}]: {err_detail}")
        return [{"文件名": file_name, "状态": "识别失败", "错误码": str(api_code), "错误信息": err_detail}]

    # 解析 Data（翔云小写 data，兼容大写 Data）
    raw_data = ocr_result.get("data", ocr_result.get("Data", {}))
    invoices = parse_invoice_data(raw_data)

    if not invoices:
        print("⚠ 未识别到发票数据")
        return [{"文件名": file_name, "状态": "无数据", "原始响应": json.dumps(ocr_result, ensure_ascii=False)[:200]}]

    # 注入文件名方便追溯
    for inv in invoices:
        if isinstance(inv, dict):
            inv.setdefault("_来源文件", file_name)

    print(f"✓ 识别到 {len(invoices)} 张发票")
    return invoices


# ─────────────────────────── 结果输出 ───────────────────────────

def print_invoice(inv: dict, index: int):
    """格式化打印单张发票信息"""
    print(f"\n  ── 发票 #{index} ──────────────────────────────────")
    # 优先展示关键字段
    key_fields = [
        ("发票类型", "发票类型"), ("发票代码", "发票代码"), ("发票号码", "发票号码"),
        ("开票日期", "开票日期"), ("价税合计", "价税合计(小写)"),
        ("购买方名称", "购买方名称"), ("销售方名称", "销售方名称"),
    ]
    shown = set()
    for label, key in key_fields:
        val = inv.get(key) or inv.get(label)
        if val:
            print(f"  {label:<12}: {val}")
            shown.add(key)
    # 其余字段
    for k, v in inv.items():
        if k not in shown and not k.startswith("_") and v:
            print(f"  {k:<12}: {v}")


def export_json(results: list, output_path: str):
    """导出为 JSON 文件"""
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"✅ 已导出 JSON: {output_path}")


def export_excel(results: list, output_path: str):
    """
    导出为 Excel (.xlsx) 文件。
    依赖 openpyxl，未安装时自动提示安装。
    """
    if not results:
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("⚠️  导出 Excel 需要 openpyxl 库，正在自动安装...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # 收集所有字段（去掉内部字段 _ext_* 和 _来源文件）
    all_keys = []
    seen = set()
    for inv in results:
        if isinstance(inv, dict):
            for k in inv.keys():
                if k not in seen and not k.startswith("_"):
                    all_keys.append(k)
                    seen.add(k)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票识别结果"

    # ── 样式定义 ──
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── 写表头 ──
    for col_idx, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── 写数据行 ──
    for row_idx, inv in enumerate(results, 2):
        if not isinstance(inv, dict):
            continue
        for col_idx, key in enumerate(all_keys, 1):
            val = inv.get(key, "")
            # openpyxl 不支持 list/dict 类型，序列化为 JSON 字符串
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # ── 自动列宽 ──
    for col_idx, key in enumerate(all_keys, 1):
        # 表头宽度
        max_len = len(str(key))
        # 数据最大宽度（最多取前 50 行）
        for row_idx in range(2, min(len(results) + 2, 52)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                # 中文按 2 字符计算宽度
                cell_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
                max_len = max(max_len, cell_len)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # ── 冻结首行 ──
    ws.freeze_panes = "A2"

    # ── 自动筛选 ──
    if all_keys:
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(all_keys))}{len(results) + 1}"

    wb.save(output_path)
    print(f"✅ 已导出 Excel: {output_path}")


# ─────────────────────────── 主函数 ───────────────────────────

def cmd_config():
    """交互式配置翔云凭证"""
    print("请输入翔云配置信息（Ctrl+C 取消）:")
    try:
        netocr_key    = input("  netocr_key    : ").strip()
        netocr_secret = input("  netocr_secret : ").strip()
    except KeyboardInterrupt:
        print("\n已取消")
        return
    if not netocr_key or not netocr_secret:
        print("❌ 凭证不能为空")
        return
    save_config({"netocr_key": netocr_key, "netocr_secret": netocr_secret})
    print("✓ 配置已保存!")


def cmd_list_config():
    """查看当前配置"""
    config = load_config()
    if config:
        print("当前配置:")
        print(f"  netocr_key    : {config.get('netocr_key', '(未设置)')}")
        secret = config.get('netocr_secret', '')
        masked = ('*' * min(8, len(secret))) if secret else '(未设置)'
        print(f"  netocr_secret : {masked}")
    else:
        print("尚未配置，请运行: python recognize_invoices.py --config")


def main():
    """主函数"""
    # ── 帮助 / 无参数 ──
    if len(sys.argv) < 2 or sys.argv[1] in ("-h", "--help"):
        print(
            "发票 OCR 识别工具（翔云）\n"
            "\n用法:\n"
            "  python recognize_invoices.py <发票文件或文件夹路径>\n"
            "  python recognize_invoices.py <路径> --output <输出文件.xlsx|.json>\n"
            "  python recognize_invoices.py --config\n"
            "  python recognize_invoices.py --list-config\n"
            "\n支持格式: PDF, OFD, JPG, PNG, BMP\n"
            "支持类型: 增值税发票、火车票、出租车票、机票行程单、定额发票等 40+ 种\n"
            "导出格式: 默认导出 Excel (.xlsx)，也支持 JSON (.json)\n"
        )
        sys.exit(0 if "--help" in sys.argv or "-h" in sys.argv else 1)

    # ── 配置命令 ──
    if sys.argv[1] == "--config":
        cmd_config()
        return
    if sys.argv[1] == "--list-config":
        cmd_list_config()
        return

    # ── 加载凭证 ──
    config = load_config()
    netocr_key    = config.get("netocr_key")
    netocr_secret = config.get("netocr_secret")
    if not netocr_key or not netocr_secret:
        print("❌ 错误：未配置翔云凭证\n")
        print("请先运行以下命令配置凭证:")
        print("  python recognize_invoices.py --config")
        sys.exit(1)

    # ── 解析参数 ──
    input_path  = sys.argv[1]
    output_file = None
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            output_file = sys.argv[idx + 1]
        else:
            print("❌ --output 参数后需指定输出文件路径")
            sys.exit(1)

    # ── 收集文件 ──
    files = collect_files(input_path)
    if not files:
        print("❌ 未找到可识别的发票文件")
        sys.exit(1)
    print(f"📁 找到 {len(files)} 个发票文件\n")

    # ── 逐一识别 ──
    all_invoices = []
    success_count = 0
    fail_count    = 0

    for i, file_path in enumerate(files, 1):
        print(f"[{i}/{len(files)}]", end=" ")
        invoices = recognize_invoice(file_path, netocr_key, netocr_secret)

        for inv in invoices:
            if isinstance(inv, dict) and inv.get("状态") in ("识别失败", "读取失败", "无数据"):
                fail_count += 1
            else:
                success_count += 1
                if not output_file:      # 无导出时，打印详情
                    print_invoice(inv, success_count)

        all_invoices.extend(invoices)

    # ── 汇总 ──
    print(f"\n{'─'*50}")
    print(f"识别完成：成功 {success_count} 张，失败 {fail_count} 张，共 {len(all_invoices)} 条记录")

    # ── 导出结果 ──
    if output_file:
        suffix = Path(output_file).suffix.lower()
        if suffix == ".json":
            export_json(all_invoices, output_file)
        else:
            # 默认 Excel（含 .xlsx 及其他后缀）
            if not output_file.endswith(".xlsx"):
                output_file = Path(output_file).stem + ".xlsx"
            export_excel(all_invoices, output_file)
    else:
        # 未指定输出文件，自动生成带时间戳的 Excel
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        auto_output = str(Path(input_path).parent / f"invoice_results_{ts}.xlsx")
        export_excel(all_invoices, auto_output)


if __name__ == "__main__":
    main()
