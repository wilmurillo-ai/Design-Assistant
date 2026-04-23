#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
testcase-generator 技能 - 测试用例生成器（增强版）
根据需求文档自动生成标准格式的 Excel 测试用例
"""

import argparse
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_testcase_workbook():
    """创建测试用例工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    
    # 设置表头样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['用例编号', '用例标题', '用例类型', '优先级', '前置条件',
               '测试步骤', '预期结果', '测试数据', '关联需求', '备注']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 20
    
    return wb, ws


def parse_requirements(file_path):
    """解析需求文档，提取测试场景"""
    print(f"解析需求文档：{file_path}")
    
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    scenarios = []
    
    # 申购场景 - 增强版
    if '申购' in content or 'SUB' in content:
        sub_cases = [
            # 正常场景
            {'scene': '正常场景', 'desc': '绝对金额申购成功', 'priority': 'P0', 'target_type': '2 绝对金额', 'data': '申请金额=1000000', 'result': '申购成功，生成交易申请'},
            {'scene': '正常场景', 'desc': '大额申购成功', 'priority': 'P1', 'target_type': '2 绝对金额', 'data': '申请金额=10000000', 'result': '申购成功，触发大额预警'},
            # 异常场景 - 参数校验
            {'scene': '异常场景', 'desc': '绝对数量申购驳回', 'priority': 'P0', 'target_type': '1 绝对数量', 'data': '申请数量=10000', 'result': '驳回 - 申购不支持绝对数量'},
            {'scene': '异常场景', 'desc': '基金账号为空', 'priority': 'P0', 'target_type': '2 绝对金额', 'data': '基金账号=空', 'result': '驳回 - 基金账号不能为空'},
            {'scene': '异常场景', 'desc': '托管账号为空', 'priority': 'P0', 'target_type': '2 绝对金额', 'data': '托管账号=空', 'result': '驳回 - 托管账号不能为空'},
            {'scene': '异常场景', 'desc': '交易账号为空', 'priority': 'P0', 'target_type': '2 绝对金额', 'data': '交易账号=空', 'result': '驳回 - 交易账号不能为空'},
            {'scene': '异常场景', 'desc': '产品代码无映射', 'priority': 'P1', 'target_type': '2 绝对金额', 'data': '产品代码=未映射', 'result': '驳回 - 产品代码无映射关系'},
            {'scene': '异常场景', 'desc': '申请金额为 0', 'priority': 'P0', 'target_type': '2 绝对金额', 'data': '申请金额=0', 'result': '驳回 - 申请金额必须大于 0'},
            {'scene': '异常场景', 'desc': '申请金额为负数', 'priority': 'P0', 'target_type': '2 绝对金额', 'data': '申请金额=-1000', 'result': '驳回 - 申请金额必须大于 0'},
            {'scene': '异常场景', 'desc': '申请金额超限', 'priority': 'P1', 'target_type': '2 绝对金额', 'data': '申请金额=999999999', 'result': '驳回 - 超过单笔限额'},
            {'scene': '异常场景', 'desc': '申请金额精度超限', 'priority': 'P2', 'target_type': '2 绝对金额', 'data': '申请金额=1000.123', 'result': '驳回 - 金额精度最多 2 位小数'},
            # 边界场景
            {'scene': '边界场景', 'desc': '最小申购金额', 'priority': 'P2', 'target_type': '2 绝对金额', 'data': '申请金额=1', 'result': '申购成功或提示最小金额'},
            {'scene': '边界场景', 'desc': '最大申购金额', 'priority': 'P2', 'target_type': '2 绝对金额', 'data': '申请金额=99999999.99', 'result': '申购成功或提示超限'},
            # 业务规则
            {'scene': '业务规则', 'desc': '非交易时间申购', 'priority': 'P1', 'target_type': '2 绝对金额', 'data': '时间=20:00', 'result': '驳回 - 非交易时间'},
            {'scene': '业务规则', 'desc': '节假日申购', 'priority': 'P1', 'target_type': '2 绝对金额', 'data': '日期=周六', 'result': '驳回 - 非交易日'},
            {'scene': '业务规则', 'desc': '账户已冻结', 'priority': 'P0', 'target_type': '2 绝对金额', 'data': '账户状态=冻结', 'result': '驳回 - 账户已冻结'},
            {'scene': '业务规则', 'desc': '产品已暂停申购', 'priority': 'P0', 'target_type': '2 绝对金额', 'data': '产品状态=暂停申购', 'result': '驳回 - 产品暂停申购'},
        ]
        scenarios.append({'type': '申购', 'cases': sub_cases})
    
    # 赎回场景 - 增强版
    if '赎回' in content or 'RED' in content:
        red_cases = [
            # 正常场景
            {'scene': '正常场景', 'desc': '绝对数量赎回成功', 'priority': 'P0', 'target_type': '1 绝对数量', 'data': '申请份额=10000', 'result': '赎回成功，生成交易申请'},
            {'scene': '正常场景', 'desc': '是否全赎回=是', 'priority': 'P1', 'target_type': '1 绝对数量', 'data': '是否全赎回=是', 'result': '申请份额=剩余份额，确认日收益按 TA 配置'},
            {'scene': '正常场景', 'desc': '巨额赎回 - 剩余取消', 'priority': 'P1', 'target_type': '1 绝对数量', 'data': '巨额赎回=剩余取消', 'result': '放弃超额部分，确认剩余份额'},
            {'scene': '正常场景', 'desc': '巨额赎回 - 剩余延期', 'priority': 'P1', 'target_type': '1 绝对数量', 'data': '巨额赎回=剩余延期', 'result': '继续赎回，延期处理'},
            # 异常场景
            {'scene': '异常场景', 'desc': '绝对金额赎回驳回', 'priority': 'P0', 'target_type': '2 绝对金额', 'data': '指令目标类型=2', 'result': '驳回 - 赎回不支持绝对金额'},
            {'scene': '异常场景', 'desc': '基金账号为空', 'priority': 'P0', 'target_type': '1 绝对数量', 'data': '基金账号=空', 'result': '驳回 - 基金账号不能为空'},
            {'scene': '异常场景', 'desc': '申请份额为 0', 'priority': 'P0', 'target_type': '1 绝对数量', 'data': '申请份额=0', 'result': '驳回 - 申请份额必须大于 0'},
            {'scene': '异常场景', 'desc': '申请份额为负数', 'priority': 'P0', 'target_type': '1 绝对数量', 'data': '申请份额=-100', 'result': '驳回 - 申请份额必须大于 0'},
            {'scene': '异常场景', 'desc': '份额不足', 'priority': 'P0', 'target_type': '1 绝对数量', 'data': '申请份额=100000, 持有=50000', 'result': '驳回 - 可用份额不足'},
            {'scene': '异常场景', 'desc': '巨额赎回方式无效', 'priority': 'P1', 'target_type': '1 绝对数量', 'data': '巨额赎回=无效值', 'result': '驳回 - 巨额赎回方式无效'},
            # 边界场景
            {'scene': '边界场景', 'desc': '最小赎回份额', 'priority': 'P2', 'target_type': '1 绝对数量', 'data': '申请份额=0.01', 'result': '赎回成功或提示最小份额'},
            {'scene': '边界场景', 'desc': '最大赎回份额', 'priority': 'P2', 'target_type': '1 绝对数量', 'data': '申请份额=999999999', 'result': '赎回成功或提示超限'},
            {'scene': '边界场景', 'desc': '全部份额赎回', 'priority': 'P1', 'target_type': '1 绝对数量', 'data': '申请份额=持有份额', 'result': '赎回成功，账户份额清零'},
            {'scene': '边界场景', 'desc': '份额精度超限', 'priority': 'P2', 'target_type': '1 绝对数量', 'data': '申请份额=100.123', 'result': '驳回 - 份额精度最多 2 位小数'},
            # 业务规则
            {'scene': '业务规则', 'desc': '非交易时间赎回', 'priority': 'P1', 'target_type': '1 绝对数量', 'data': '时间=20:00', 'result': '驳回 - 非交易时间'},
            {'scene': '业务规则', 'desc': '节假日赎回', 'priority': 'P1', 'target_type': '1 绝对数量', 'data': '日期=周六', 'result': '驳回 - 非交易日'},
            {'scene': '业务规则', 'desc': '账户已冻结', 'priority': 'P0', 'target_type': '1 绝对数量', 'data': '账户状态=冻结', 'result': '驳回 - 账户已冻结'},
            {'scene': '业务规则', 'desc': '产品已暂停赎回', 'priority': 'P0', 'target_type': '1 绝对数量', 'data': '产品状态=暂停赎回', 'result': '驳回 - 产品暂停赎回'},
            {'scene': '业务规则', 'desc': '持有期不足', 'priority': 'P1', 'target_type': '1 绝对数量', 'data': '持有天数<最低持有期', 'result': '驳回 - 持有期不足'},
        ]
        scenarios.append({'type': '赎回', 'cases': red_cases})
    
    # 认购场景
    if '认购' in content or 'PSUB' in content:
        psub_cases = [
            {'scene': '正常场景', 'desc': '绝对金额认购成功', 'priority': 'P0', 'target_type': '2 绝对金额', 'data': '申请金额=1000000', 'result': '认购成功'},
            {'scene': '异常场景', 'desc': '绝对数量认购驳回', 'priority': 'P0', 'target_type': '1 绝对数量', 'data': '申请数量=10000', 'result': '驳回 - 认购不支持绝对数量'},
            {'scene': '异常场景', 'desc': '认购金额超限', 'priority': 'P1', 'target_type': '2 绝对金额', 'data': '申请金额=999999999', 'result': '驳回 - 超过认购限额'},
            {'scene': '业务规则', 'desc': '产品募集期外认购', 'priority': 'P1', 'target_type': '2 绝对金额', 'data': '日期=非募集期', 'result': '驳回 - 不在募集期内'},
        ]
        scenarios.append({'type': '认购', 'cases': psub_cases})
    
    # 交易撤销场景 - 增强版
    if '撤销' in content or '撤单' in content:
        cancel_cases = [
            # 正常场景
            {'scene': '正常场景', 'desc': '未确认交易撤销成功', 'priority': 'P0', 'data': '指令编号=ORDER001, 状态=待确认', 'result': '撤销成功，交易申请状态=已撤销'},
            {'scene': '正常场景', 'desc': '部分确认交易撤销', 'priority': 'P1', 'data': '指令编号=ORDER002, 状态=部分确认', 'result': '撤销未确认部分'},
            # 异常场景
            {'scene': '异常场景', 'desc': '指令编号不存在', 'priority': 'P1', 'data': '指令编号=NOT_EXIST', 'result': '撤销失败 - 指令编号不存在'},
            {'scene': '异常场景', 'desc': '已确认交易不可撤销', 'priority': 'P0', 'data': '指令编号=CONFIRMED, 状态=已确认', 'result': '撤销失败 - 交易已确认'},
            {'scene': '异常场景', 'desc': '已撤销交易重复撤销', 'priority': 'P1', 'data': '指令编号=CANCELLED, 状态=已撤销', 'result': '撤销失败 - 交易已撤销'},
            {'scene': '异常场景', 'desc': '指令编号为空', 'priority': 'P0', 'data': '指令编号=空', 'result': '撤销失败 - 指令编号不能为空'},
            {'scene': '异常场景', 'desc': '业务流水号不匹配', 'priority': 'P1', 'data': '业务流水号=不匹配', 'result': '撤销失败 - 业务流水号不匹配'},
            {'scene': '异常场景', 'desc': '交易类别错误', 'priority': 'P2', 'data': '交易类别=错误类型', 'result': '撤销失败 - 交易类别错误'},
            # 幂等性
            {'scene': '幂等场景', 'desc': '重复撤销请求', 'priority': 'P1', 'data': '指令编号=ORDER001(重复)', 'result': '返回之前结果，不重复处理'},
        ]
        scenarios.append({'type': '撤销', 'cases': cancel_cases})
    
    # 状态回传场景 - 增强版
    if '状态' in content and '回传' in content:
        status_cases = [
            # 正常场景
            {'scene': '正常场景', 'desc': '下单校验通过回传', 'priority': 'P0', 'data': '指令状态=成功', 'result': '回传成功，兴银接收状态'},
            {'scene': '正常场景', 'desc': '下单校验驳回回传', 'priority': 'P0', 'data': '指令状态=B00', 'result': '回传 B00-已驳回，包含驳回原因'},
            {'scene': '正常场景', 'desc': '交易确认回传', 'priority': 'P0', 'data': '指令状态=已确认', 'result': '回传确认结果，包含确认份额'},
            # 异常场景
            {'scene': '异常场景', 'desc': '指令编号不存在', 'priority': 'P1', 'data': '指令编号=NOT_EXIST', 'result': '回传失败 - 指令不存在'},
            {'scene': '异常场景', 'desc': '状态码无效', 'priority': 'P1', 'data': '指令状态=INVALID', 'result': '回传失败 - 状态码无效'},
            {'scene': '异常场景', 'desc': '交易类别错误', 'priority': 'P2', 'data': '交易类别=错误类型', 'result': '回传失败 - 交易类别错误'},
            # 接口异常
            {'scene': '接口异常', 'desc': 'IC04 接口超时', 'priority': 'P1', 'data': '超时=30 秒', 'result': '重试机制，记录日志'},
            {'scene': '接口异常', 'desc': 'IC04 接口失败', 'priority': 'P1', 'data': '返回=500', 'result': '记录失败，后续补传'},
        ]
        scenarios.append({'type': '状态回传', 'cases': status_cases})
    
    # 幂等性场景
    if '幂等' in content or '流水号' in content:
        idem_cases = [
            {'scene': '幂等场景', 'desc': '申购重复请求', 'priority': 'P0', 'data': '业务流水号=XY001(重复)', 'result': '返回之前结果，不重复生成申请'},
            {'scene': '幂等场景', 'desc': '赎回重复请求', 'priority': 'P0', 'data': '业务流水号=XY002(重复)', 'result': '返回之前结果，不重复扣减份额'},
            {'scene': '幂等场景', 'desc': '撤销重复请求', 'priority': 'P0', 'data': '业务流水号=XY003(重复)', 'result': '返回之前结果，不重复撤销'},
            {'scene': '幂等场景', 'desc': '不同流水号相同指令', 'priority': 'P1', 'data': '业务流水号=不同，指令编号=相同', 'result': '正常处理，视为不同请求'},
        ]
        scenarios.append({'type': '幂等性', 'cases': idem_cases})
    
    # 数据推送场景 - 增强版
    if '推送' in content:
        push_cases = [
            # 交易申请确认推送
            {'scene': '数据推送', 'desc': '交易申请确认推送', 'priority': 'P1', 'data': '确认份额=10000, 确认金额=1000000', 'result': '推送成功，兴银接收'},
            {'scene': '数据推送', 'desc': '确认推送 - 部分确认', 'priority': 'P1', 'data': '确认份额=5000, 原申请=10000', 'result': '推送部分确认结果'},
            {'scene': '数据推送', 'desc': '确认推送 - 确认失败', 'priority': 'P1', 'data': '确认结果=失败，失败原因=份额不足', 'result': '推送失败原因'},
            # 份额信息推送
            {'scene': '数据推送', 'desc': '份额信息推送', 'priority': 'P1', 'data': '基金账号=123456, 持有份额=100000', 'result': '推送成功'},
            {'scene': '数据推送', 'desc': '份额信息推送 - 多账户', 'priority': 'P2', 'data': '账户数=10', 'result': '批量推送成功'},
            {'scene': '数据推送', 'desc': '份额信息推送 - 数据异常', 'priority': 'P2', 'data': '份额=负数', 'result': '推送失败，记录异常'},
            # 推送异常
            {'scene': '推送异常', 'desc': '推送接口超时', 'priority': 'P1', 'data': '超时=30 秒', 'result': '重试机制，记录日志'},
            {'scene': '推送异常', 'desc': '推送数据格式错误', 'priority': 'P1', 'data': '数据=格式错误', 'result': '推送失败，记录错误'},
        ]
        scenarios.append({'type': '数据推送', 'cases': push_cases})
    
    # 性能场景
    if '性能' in content or '并发' in content or '响应' in content:
        perf_cases = [
            {'scene': '性能场景', 'desc': '接口响应时间', 'priority': 'P2', 'data': '并发数=10, 执行次数=100', 'result': '平均响应<2 秒，95%<3 秒'},
            {'scene': '性能场景', 'desc': '高并发申购', 'priority': 'P2', 'data': '并发数=100, 持续时间=1 分钟', 'result': '系统稳定，无崩溃'},
            {'scene': '性能场景', 'desc': '高并发赎回', 'priority': 'P2', 'data': '并发数=100, 持续时间=1 分钟', 'result': '系统稳定，份额扣减正确'},
            {'scene': '性能场景', 'desc': '大数据量推送', 'priority': 'P2', 'data': '推送记录=10000 条', 'result': '推送成功，耗时<5 分钟'},
            {'scene': '性能场景', 'desc': '内存使用', 'priority': 'P3', 'data': '持续运行=24 小时', 'result': '内存无泄漏，增长<10%'},
            {'scene': '性能场景', 'desc': '数据库连接池', 'priority': 'P2', 'data': '并发数=50', 'result': '连接池正常，无连接泄漏'},
        ]
        scenarios.append({'type': '性能', 'cases': perf_cases})
    
    # 安全场景
    if '安全' in content or '鉴权' in content or '权限' in content:
        sec_cases = [
            {'scene': '安全场景', 'desc': '接口鉴权失败', 'priority': 'P1', 'data': '鉴权信息=INVALID_TOKEN', 'result': '鉴权失败，拒绝访问'},
            {'scene': '安全场景', 'desc': '接口鉴权过期', 'priority': 'P1', 'data': '鉴权信息=EXPIRED_TOKEN', 'result': '鉴权过期，返回 401'},
            {'scene': '安全场景', 'desc': 'IP 白名单校验', 'priority': 'P1', 'data': '来源 IP=非白名单', 'result': '拒绝访问'},
            {'scene': '安全场景', 'desc': 'SQL 注入攻击', 'priority': 'P1', 'data': '参数=恶意 SQL', 'result': '拦截攻击，记录日志'},
            {'scene': '安全场景', 'desc': 'XSS 攻击', 'priority': 'P2', 'data': '参数=恶意脚本', 'result': '拦截攻击，转义输出'},
            {'scene': '安全场景', 'desc': '敏感数据加密', 'priority': 'P1', 'data': '数据=账号密码', 'result': '数据传输加密，日志脱敏'},
            {'scene': '安全场景', 'desc': '重放攻击', 'priority': 'P1', 'data': '重复请求=相同签名', 'result': '识别重放，拒绝请求'},
        ]
        scenarios.append({'type': '安全', 'cases': sec_cases})
    
    # 兼容性场景
    if '兼容' in content or '版本' in content:
        compat_cases = [
            {'scene': '兼容场景', 'desc': '接口版本兼容', 'priority': 'P2', 'data': '接口版本=v1/v2', 'result': '向后兼容，正常处理'},
            {'scene': '兼容场景', 'desc': '数据格式兼容', 'priority': 'P2', 'data': '数据格式=新旧格式', 'result': '兼容新旧格式'},
            {'scene': '兼容场景', 'desc': '字符集兼容', 'priority': 'P2', 'data': '字符集=GBK/UTF-8', 'result': '字符集转换正确'},
        ]
        scenarios.append({'type': '兼容性', 'cases': compat_cases})
    
    # 故障处理场景
    if '故障' in content or '异常' in content or '容灾' in content:
        fault_cases = [
            {'scene': '故障处理', 'desc': '数据库连接失败', 'priority': 'P1', 'data': '数据库=不可用', 'result': '切换备用库，记录告警'},
            {'scene': '故障处理', 'desc': '接口调用失败', 'priority': 'P1', 'data': '兴银接口=不可用', 'result': '重试机制，记录失败'},
            {'scene': '故障处理', 'desc': '网络中断恢复', 'priority': 'P1', 'data': '网络=中断后恢复', 'result': '自动重连，补传数据'},
            {'scene': '故障处理', 'desc': '系统重启恢复', 'priority': 'P1', 'data': '系统=重启', 'result': '数据不丢失，状态正确'},
        ]
        scenarios.append({'type': '故障处理', 'cases': fault_cases})
    
    print(f"提取到 {len(scenarios)} 个测试场景类别")
    total_cases = sum(len(s['cases']) for s in scenarios)
    print(f"共 {total_cases} 个测试用例")
    
    return scenarios


def generate_testcases(input_file, output_file, module_name=None):
    """生成测试用例主函数"""
    print(f"输入文件：{input_file}")
    print(f"输出文件：{output_file}")
    
    # 创建工作簿
    wb, ws = create_testcase_workbook()
    
    # 解析需求
    scenarios = parse_requirements(input_file)
    
    # 生成用例
    test_cases = []
    case_counter = {}
    
    type_name_map = {
        '申购': 'SUB', '赎回': 'RED', '认购': 'PSUB', '撤销': 'CANCEL',
        '状态回传': 'STATUS', '幂等性': 'IDEM', '数据推送': 'PUSH',
        '性能': 'PERF', '安全': 'SEC', '兼容性': 'COMPAT', '故障处理': 'FAULT'
    }
    
    for scenario in scenarios:
        stype = scenario['type']
        type_prefix = type_name_map.get(stype, 'TC')
        
        for case in scenario['cases']:
            case_counter[stype] = case_counter.get(stype, 0) + 1
            counter = case_counter[stype]
            
            case_id = f"TC_XY_{type_prefix}_{counter:03d}"
            title = f"兴银理财-{stype}-{case['scene']}-{case['desc']}"
            
            # 根据类型生成步骤和预期
            if stype in ['申购', '赎回', '认购']:
                steps = f"1.兴银系统发送 IC09 资管计划指令\n2.交易方向={stype}\n3.指令目标类型={case.get('target_type', '-')}\n4.{case.get('data', '')}\n5.销售系统接收并校验"
                expected = f"1.销售系统成功接收\n2.业务流水号、指令编号正确记录\n3.{case['result']}"
            elif stype == '撤销':
                steps = f"1.兴银系统发送 IC03 指令撤销\n2.{case.get('data', '')}\n3.销售系统接收并校验"
                expected = f"1.{case['result']}"
            elif stype == '状态回传':
                steps = f"1.销售系统接收 IC09 交易申请\n2.调用 IC04 指令状态回传\n3.{case.get('data', '')}"
                expected = f"1.正确调用 IC04 接口\n2.{case['result']}"
            elif stype == '幂等性':
                steps = f"1.兴银系统发送 IC09 指令\n2.{case.get('data', '')}\n3.销售系统校验幂等性"
                expected = f"1.识别重复请求\n2.{case['result']}"
            elif stype == '数据推送':
                steps = f"1.{case['desc']}\n2.销售系统推送数据\n3.{case.get('data', '')}"
                expected = f"1.正确推送数据\n2.{case['result']}"
            elif stype == '性能':
                steps = f"1.执行{case.get('data', '')}\n2.记录性能指标\n3.分析结果"
                expected = f"1.{case['result']}"
            elif stype == '安全':
                steps = f"1.发送请求\n2.{case.get('data', '')}\n3.销售系统校验"
                expected = f"1.{case['result']}"
            elif stype == '兼容性':
                steps = f"1.使用{case.get('data', '')}\n2.发送请求\n3.验证处理结果"
                expected = f"1.{case['result']}"
            elif stype == '故障处理':
                steps = f"1.模拟{case.get('data', '')}\n2.观察系统行为\n3.验证恢复机制"
                expected = f"1.{case['result']}"
            else:
                steps = f"1.执行测试\n2.{case.get('data', '')}"
                expected = case['result']
            
            test_cases.append({
                'case_id': case_id,
                'title': title,
                'case_type': f"{stype}测试",
                'priority': case['priority'],
                'precondition': '1.兴银理财系统已对接 2.相关配置已维护 3.测试环境准备就绪',
                'steps': steps,
                'expected_result': expected,
                'test_data': case.get('data', ''),
                'requirement': f'3-{stype}',
                'remark': case.get('remark', '')
            })
    
    # 写入用例
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(vertical="top", wrap_text=True)
    
    row = 2
    for case in test_cases:
        fields = ['case_id', 'title', 'case_type', 'priority', 'precondition',
                  'steps', 'expected_result', 'test_data', 'requirement', 'remark']
        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=row, column=col, value=case.get(field, ''))
            cell.border = thin_border
            cell.alignment = alignment
        row += 1
    
    # 保存文件
    wb.save(output_file)
    print(f"\n生成完成！共生成 {len(test_cases)} 条测试用例")
    print(f"输出文件：{os.path.abspath(output_file)}")
    
    # 统计信息
    case_types = {}
    priorities = {}
    for case in test_cases:
        ct = case.get('case_type', '其他')
        pr = case.get('priority', 'P3')
        case_types[ct] = case_types.get(ct, 0) + 1
        priorities[pr] = priorities.get(pr, 0) + 1
    
    print("\n用例类型分布:")
    for ct, count in sorted(case_types.items()):
        print(f"  {ct}: {count}条")
    
    print("\n优先级分布:")
    for pr, count in sorted(priorities.items()):
        print(f"  {pr}: {count}条")
    
    return True


def main():
    parser = argparse.ArgumentParser(description='testcase-generator - 测试用例生成器')
    parser.add_argument('--input', '-i', required=True, help='输入材料文件路径')
    parser.add_argument('--output', '-o', help='输出 Excel 文件路径')
    parser.add_argument('--module', '-m', help='模块名称，用于生成用例编号')
    parser.add_argument('--format', choices=['xlsx', 'csv'], default='xlsx', help='输出格式')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input):
        print(f"错误：输入文件不存在：{args.input}")
        sys.exit(1)
    
    if not args.output:
        base_name = os.path.splitext(args.input)[0]
        args.output = f"{base_name}_测试用例.{args.format}"
    
    success = generate_testcases(args.input, args.output, args.module)
    
    if not success:
        sys.exit(1)


if __name__ == '__main__':
    main()