#!/usr/bin/env python3
"""
Agent技能收集 - 根据Agent类型自动收集相关Skills
支持选择数据源：skillhub.tencent.com 或 cn.clawhub-mirror.com
"""

import subprocess
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import time
import sys
import urllib.request
import urllib.error
import re

# ========== Agent类型配置 ==========
AGENT_CONFIGS = {
    'hr': {
        'name': '人事Agent',
        'keywords': ['hr', '招聘', '人力', '考勤', '薪酬', '员工', '绩效', '简历', '面试', '入职']
    },
    'sales': {
        'name': '销售Agent',
        'keywords': ['sales', 'crm', '销售', '客户', '商机', '线索', '成交', '报价']
    },
    'customer_service': {
        'name': '客服Agent',
        'keywords': ['客服', '服务', '工单', 'support', 'chatbot', 'FAQ', '问答']
    },
    'data_analysis': {
        'name': '数据分析Agent',
        'keywords': ['data', 'analysis', 'analytics', 'bi', '数据', '分析', '报表', '可视化']
    },
    'marketing': {
        'name': '营销Agent',
        'keywords': ['marketing', '营销', '推广', 'SEO', '广告', '社媒', '文案', '运营']
    },
    'finance': {
        'name': '财务Agent',
        'keywords': ['finance', '财务', '金融', '发票', '报销', '预算', '会计']
    },
    'operations': {
        'name': '运营Agent',
        'keywords': ['运营', 'product', 'task', 'project', 'workflow', '流程', '日程', '会议']
    },
    'it': {
        'name': 'IT运维Agent',
        'keywords': ['IT', '运维', '监控', '安全', 'server', 'network', '系统', '日志']
    },
    'content': {
        'name': '内容创作Agent',
        'keywords': ['内容', '创作', '写作', '文案', 'blog', 'article', 'copywriting', '自媒体', '短视频']
    },
    'legal': {
        'name': '法务Agent',
        'keywords': ['法务', '法律', '合同', '律师', '合规', 'contract', 'legal']
    },
    'travel': {
        'name': '旅游Agent',
        'keywords': ['旅游', '旅行', '出行', '机票', '酒店', '签证', '景点', '攻略', '行程', '导游']
    }
}

# 数据源配置
DATA_SOURCES = {
    'clawhub': {
        'name': 'ClawHub (cn.clawhub-mirror.com)',
        'api': 'https://cn.clawhub-mirror.com/api/v1',
        'url_template': 'https://cn.clawhub-mirror.com/skill/{slug}',
        'install_template': '请先检查是否已安装 ClawHub 商店，若未安装，请根据 https://cn.clawhub-mirror.com/install 安装ClawHub商店，然后安装 {slug} 技能。\n\n若已安装，则直接安装 {slug} 技能。'
    },
    'skillhub': {
        'name': 'SkillHub (skillhub.tencent.com)',
        'api': None,
        'url_template': 'https://skillhub.tencent.com/skill/{slug}',
        'install_template': '请先检查是否已安装 SkillHub 商店，若未安装，请根据 https://skillhub.cn/install/skillhub.md 安装Skillhub商店（只安装CLI），然后安装 {slug} 技能。\n\n若已安装，则直接运行: skillhub install {slug}'
    }
}

# 类别关键词映射
CATEGORY_KEYWORDS = {
    '机票': ['flight', 'fly', '机票', '航班', '航空', 'airline', 'amadeus', 'fliggy', 'qunar', 'tuniu'],
    '酒店': ['hotel', '酒店', '住宿', 'booking', 'amadeus-hotels', 'ctrip-hotel', 'tuniu-hotel'],
    '地图': ['map', '地图', '导航', 'lbs', 'location', 'amap', 'baidu-nearby', 'google-maps', 'smart-map'],
    '天气': ['weather', '天气', '气象', 'forecast'],
    '行程': ['travel', 'trip', '行程', '规划', 'planner', 'itinerary', 'planit', 'event-planner'],
    '景点': ['景点', 'tourist', 'destination', 'guide', '旅游', '景区', '观光', 'guruwalk'],
    '签证': ['visa', '签证', '护照', 'ds160', '入境'],
    '铁路': ['train', '铁路', '火车', 'rail', 'oebb', 'bahn', 'mcporter', 'ns-trains', 'bvg'],
    '优惠': ['coupon', '优惠', '折扣', '省钱', 'coupon', 'trip-coupon', 'obtain-coupons'],
    '城市': ['new-york', 'hong-kong', 'dubai', 'amsterdam', 'sydney', 'berlin', 'seoul', 'vienna', 'london'],
    '语言': ['language', '语言', '翻译', 'translator', 'learn'],
    '工具': ['browser', '工具', 'utils', 'ocr', 'pdf', 'search', '天气'],
    '其他': []
}

TAG_KEYWORDS = [
    ('Python', 'python'), ('JavaScript', 'javascript'), ('API集成', 'api'),
    ('自动化', 'automation'), ('数据处理', 'data'), ('文档处理', 'doc'),
    ('OCR识别', 'ocr'), ('商业智能', 'bi'), ('机器学习', 'machine learning'),
    ('大语言模型', 'llm'), ('测试', 'test'), ('部署', 'deploy'),
    ('可视化', 'visualization'), ('实时通讯', 'chat'), ('监控', 'monitor'),
    ('安全', 'security'), ('搜索', 'search'), ('文件处理', 'file'),
]

def select_data_source_interactive():
    """交互式选择数据源 - 支持终端和参数"""
    # 默认使用方式2 (skillhub)
    print("\n" + "="*50)
    print("默认使用: skillhub.tencent.com (输出安装方式)")
    print("="*50)
    print("如需使用 cn.clawhub-mirror.com (输出网址)，请使用参数: --source 1")
    print("="*50)
    return 'skillhub'

def ask_clawhub_collect(output_path):
    """完成后询问是否需要用ClawHub模式再收集一次"""
    if not sys.stdin.isatty():
        return None
    
    print("\n" + "="*50)
    print("是否需要再用 ClawHub 模式收集一次？")
    print("="*50)
    print("  1. 是 (输出网址)")
    print("  2. 否")
    print("="*50)
    
    while True:
        try:
            choice = input("\n请输入选项 (1/2): ").strip()
            if choice == '1':
                # 生成ClawHub版本的文件名
                base, ext = output_path.rsplit('.', 1)
                clawhub_output = f"{base}_ClawHub.{ext}"
                return clawhub_output
            elif choice == '2' or choice == '':
                return None
            print("⚠️ 请输入 1 或 2")
        except EOFError:
            return None

def classify_skill(name, desc):
    """根据名称和描述自动分类"""
    text = (name + ' ' + desc).lower()
    
    for category, keywords in CATEGORY_KEYWORDS.items():
        if category == '其他':
            continue
        for kw in keywords:
            if kw.lower() in text:
                return category
    
    return '其他'

def search_skills(query, limit=15):
    """使用 skillhub CLI 搜索"""
    try:
        result = subprocess.run(
            ['skillhub', 'search', query, '--search-limit', str(limit), '--json'],
            capture_output=True, text=True, timeout=20
        )
        data = json.loads(result.stdout)
        results = data.get('results', [])
        
        skills = []
        for r in results:
            slug = r.get('slug', '')
            if not slug:
                continue
            
            # 提取中文描述
            desc = r.get('description', '') or r.get('summary', '')
            chinese_desc = ''
            for part in desc.split('\n'):
                if any('\u4e00' <= c <= '\u9fff' for c in part):
                    chinese_desc = part.strip()
                    break
            if not chinese_desc:
                chinese_desc = slug
            
            skills.append({
                'slug': slug,
                'name': r.get('name', slug),
                'description': chinese_desc[:200],
                'version': r.get('version', '')
            })
        return skills
    except Exception as e:
        print(f"   搜索出错: {e}")
        return []

def fetch_skill_stats(slug):
    """从 cn.clawhub-mirror.com 获取星星数和下载量"""
    try:
        url = f"https://cn.clawhub-mirror.com/api/v1/skills/{slug}"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=12) as response:
            data = json.loads(response.read().decode('utf-8'))
            skill = data.get('skill', {})
            stats = skill.get('stats', {})
            meta = data.get('metaContent', {})
            
            display_desc = meta.get('DisplayDescription', '')
            return {
                'stars': stats.get('stars', 0),
                'downloads': stats.get('downloads', 0),
                'description': display_desc
            }
    except urllib.error.HTTPError:
        return {'stars': 0, 'downloads': 0, 'description': ''}
    except Exception as e:
        return {'stars': 0, 'downloads': 0, 'description': ''}

def translate_to_chinese(text):
    """翻译英文到中文"""
    if not text:
        return ''
    
    if any('\u4e00' <= c <= '\u9fff' for c in text):
        return text
    
    try:
        import urllib.parse
        encoded_text = urllib.parse.quote(text[:500])
        url = f"https://api.mymemory.translated.net/get?q={encoded_text}&langpair=en|zh"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            result = json.loads(response.read().decode('utf-8'))
            if result.get('responseStatus') == 200:
                translated = result.get('responseData', {}).get('translatedText', '')
                if translated and translated != text:
                    return translated
    except:
        pass
    return text

def extract_tags(desc):
    """提取标签"""
    text = (desc or '').lower()
    matched = [cn for cn, en in TAG_KEYWORDS if en in text]
    return ', '.join(matched[:4]) if matched else '通用能力'

def calculate_relevance(name, desc, keywords):
    """计算关联度"""
    text = (name + ' ' + desc).lower()
    
    core_keywords = [kw.lower() for kw in keywords[:6]]
    for kw in core_keywords:
        if re.search(r'\b' + re.escape(kw) + r'\b', text):
            return '⭐⭐⭐ 核心'
    
    aux_keywords = [kw.lower() for kw in keywords[6:12]]
    for kw in aux_keywords:
        if re.search(r'\b' + re.escape(kw) + r'\b', text):
            return '⭐⭐ 辅助'
    
    return '⭐ 扩展'

def collect_agent_skills(agent_type, keywords, data_source, limit=100):
    """收集Agent相关Skills"""
    print(f"📡 开始搜索 {len(keywords)} 个关键词...")
    
    source_config = DATA_SOURCES[data_source]
    
    # 搜索
    all_skills = {}
    for kw in keywords:
        print(f"   搜索: {kw}")
        results = search_skills(kw, 15)
        
        for s in results:
            slug = s['slug']
            if slug and slug not in all_skills:
                all_skills[slug] = s
        time.sleep(0.4)
    
    print(f"   找到 {len(all_skills)} 个skills")
    
    # 获取详情
    final_results = []
    total = len(all_skills)
    
    for i, (slug, skill) in enumerate(all_skills.items()):
        print(f"   获取详情: {slug} ({i+1}/{total})")
        
        stars = 0
        downloads = 0
        desc = skill['description']
        
        # 获取星星/下载数
        stats = fetch_skill_stats(slug)
        stars = stats.get('stars', 0)
        downloads = stats.get('downloads', 0)
        
        # 强制中文描述
        if stats.get('description'):
            desc = stats['description']
        elif not any('\u4e00' <= c <= '\u9fff' for c in desc):
            desc = translate_to_chinese(desc)
        
        relevance = calculate_relevance(skill['name'], desc, keywords)
        tags = extract_tags(desc)
        
        # 自动分类
        category = classify_skill(skill['name'], desc)
        
        # 根据数据源决定网址列的内容
        if data_source == 'clawhub':
            url_content = source_config['url_template'].format(slug=slug)
        else:
            url_content = source_config['install_template'].format(slug=slug)
        
        final_results.append({
            'name': skill['name'],
            'slug': slug,
            'description': desc[:200] if desc else '',
            'url_or_install': url_content,
            'tags': tags,
            'category': category,
            'stars': stars,
            'downloads': downloads,
            'version': skill.get('version', ''),
            'relevance': relevance
        })
        
        time.sleep(0.8)
    
    # 排序
    def sort_key(x):
        order = {'⭐⭐⭐ 核心': 0, '⭐⭐ 辅助': 1, '⭐ 扩展': 2}
        return (order.get(x['relevance'], 3), -x['downloads'])
    
    final_results.sort(key=sort_key)
    return final_results[:limit]

def save_to_excel(data, output_path, data_source):
    """保存到Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agent技能库"
    
    # 格式：名称 | 关联度 | 类别 | 描述 | 网址/安装方式 | 标签 | 星星数 | 下载量 | 版本
    headers = ['名称', '关联度', '类别', '描述', '网址' if data_source == 'clawhub' else '安装方式', '标签', '星星数', '下载量', '版本']
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for row, item in enumerate(data, 2):
        ws.cell(row, 1, item['name'])
        ws.cell(row, 2, item['relevance'])
        ws.cell(row, 3, item['category'])
        ws.cell(row, 4, item['description'])
        ws.cell(row, 5, item['url_or_install'])
        ws.cell(row, 6, item['tags'])
        ws.cell(row, 7, item['stars'])
        ws.cell(row, 8, item['downloads'])
        ws.cell(row, 9, item['version'])
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    
    for row in range(2, len(data) + 2):
        ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row, 5).alignment = Alignment(wrap_text=True, vertical='top')
    
    wb.save(output_path)
    print(f"✅ 保存完成: {output_path}")

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Agent技能收集工具')
    parser.add_argument('agent_type', help='Agent类型 (如 travel, hr, content 等)')
    parser.add_argument('--source', '--data-source', '-s', 
                        choices=['1', '2', 'clawhub', 'skillhub'],
                        help='数据源: 1=cn.clawhub-mirror.com(网址), 2=skillhub.tencent.com(安装方式)')
    parser.add_argument('--output', '-o', '--output-path',
                        help='Excel输出路径')
    
    args = parser.parse_args()
    
    agent_type = args.agent_type.lower()
    
    if agent_type not in AGENT_CONFIGS:
        print(f"❌ 未知Agent类型: {agent_type}")
        print(f"支持的类型: {', '.join(AGENT_CONFIGS.keys())}")
        sys.exit(1)
    
    # 确定数据源
    if args.source:
        # 转换参数值
        if args.source in ['1', 'clawhub']:
            data_source = 'clawhub'
        else:
            data_source = 'skillhub'
    else:
        # 交互式选择
        data_source = select_data_source_interactive()
    
    # 确定输出路径
    output = args.output
    if not output:
        # 交互式输入
        print("\n" + "="*50)
        print("请输入Excel输出路径:")
        print("="*50)
        print("示例: /Users/xyqin/尚云数智/测试/travel_skills.xlsx")
        output = input("\n输出路径: ").strip()
        
        if not output:
            print("⚠️ 未输入路径，将使用默认路径")
            output = f"{agent_type}_agent_skills.xlsx"
    
    config = AGENT_CONFIGS[agent_type]
    agent_name = config['name']
    keywords = config['keywords']
    source_name = DATA_SOURCES[data_source]['name']
    
    print(f"\n🚀 正在收集[{agent_name}]相关Skills...")
    print(f"   数据源: {source_name}")
    print(f"   关键词: {', '.join(keywords[:5])}...")
    
    skills = collect_agent_skills(agent_type, keywords, data_source, 100)
    
    stats = {'⭐⭐⭐ 核心': 0, '⭐⭐ 辅助': 0, '⭐ 扩展': 0}
    for s in skills:
        stats[s['relevance']] = stats.get(s['relevance'], 0) + 1
    
    # 统计类别
    category_stats = {}
    for s in skills:
        cat = s['category']
        category_stats[cat] = category_stats.get(cat, 0) + 1
    
    print(f"\n📊 技能分布:")
    for k, v in stats.items():
        print(f"   {k}: {v}")
    
    print(f"\n📊 类别分布:")
    for cat, count in sorted(category_stats.items(), key=lambda x: -x[1]):
        print(f"   {cat}: {count}")
    
    save_to_excel(skills, output, data_source)
    print(f"\n✅ 完成! 共收集 {len(skills)} 条Skills")
    print(f"   文件: {output}")
    
    # 完成后询问是否需要用ClawHub模式再收集一次
    clawhub_output = ask_clawhub_collect(output)
    
    if clawhub_output:
        print(f"\n🔄 正在用 ClawHub 模式重新收集...")
        data_source = 'clawhub'
        source_name = DATA_SOURCES[data_source]['name']
        
        skills2 = collect_agent_skills(agent_type, keywords, data_source, 100)
        
        stats2 = {'⭐⭐⭐ 核心': 0, '⭐⭐ 辅助': 0, '⭐ 扩展': 0}
        for s in skills2:
            stats2[s['relevance']] = stats2.get(s['relevance'], 0) + 1
        
        print(f"\n📊 技能分布:")
        for k, v in stats2.items():
            print(f"   {k}: {v}")
        
        save_to_excel(skills2, clawhub_output, data_source)
        print(f"\n✅ ClawHub版本完成! 共收集 {len(skills2)} 条Skills")
        print(f"   文件: {clawhub_output}")

if __name__ == '__main__':
    main()