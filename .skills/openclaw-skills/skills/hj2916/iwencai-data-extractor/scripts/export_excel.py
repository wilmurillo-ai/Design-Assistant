#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
问财涨停数据 -> Excel 汇总报告导出
用法：cmd /c python export_excel.py
输出：OUT_PATH 指定路径的 .xlsx 文件（5个Sheet）
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from collections import Counter
import datetime

# ──────────────────────────────────────────────
#  ★ 用户配置区
# ──────────────────────────────────────────────
DB_PATH  = r"D:\workbuddyclaw\iwencai_zt.db"
OUT_PATH = r"D:\workbuddyclaw\iwencai_data\涨停数据汇总.xlsx"
# ──────────────────────────────────────────────

# 颜色
RED_DARK   = "C0392B"
RED_MID    = "E74C3C"
RED_LIGHT  = "FADBD8"
ORANGE     = "E67E22"
GRAY_DARK  = "2C3E50"
GRAY_MID   = "7F8C8D"
GRAY_LIGHT = "ECF0F1"
WHITE      = "FFFFFF"
GOLD       = "F39C12"

def hf(c): return PatternFill("solid", fgColor=c)
def tb():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, bg=GRAY_DARK, fg=WHITE, size=10):
    cell.font = Font(bold=True, color=fg, size=size, name="微软雅黑")
    cell.fill = hf(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = tb()

def dat(cell, align="center"):
    cell.font = Font(size=9, name="微软雅黑")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = tb()


# 读数据
conn = sqlite3.connect(DB_PATH)
df = pd.read_sql("SELECT * FROM zt_stocks ORDER BY trade_date, lb_count DESC, id", conn)
conn.close()

col_map = {
    "id":"序号","trade_date":"交易日期","stock_code":"股票代码","stock_name":"股票名称",
    "price":"收盘价","change_pct":"涨跌幅(%)","zt_time":"涨停时间","zt_status":"涨停状态",
    "volume":"成交量","amount":"成交额","first_zt_time":"首次涨停","lb_count":"连板数",
    "zt_type":"涨停类型","float_mv":"流通市值(亿)","vol_ratio":"量比","themes":"所属题材",
    "zt_tags":"涨停标签","total_mv":"总市值(亿)","created_at":"入库时间"
}
df.rename(columns=col_map, inplace=True)

daily = df.groupby("交易日期").agg(
    涨停总数=("股票代码","count"),
    首板数=("涨停类型", lambda x:(x=="首板涨停").sum()),
    连板数=("连板数", lambda x:(x>1).sum()),
    最高连板=("连板数","max"),
    平均涨幅=("涨跌幅(%)", lambda x:round(x.mean(),2))
).reset_index()
daily.columns = ["交易日期","涨停总数","首板数","连板数量","最高连板","平均涨幅(%)"]

lb_dist = df[df["连板数"]>0].groupby("连板数").size().reset_index(name="支数")
lb_dist.columns = ["连板数","支数"]

theme_counter = Counter()
for t in df["所属题材"].dropna():
    for item in str(t).split("+"):
        item = item.strip()
        if item: theme_counter[item] += 1
theme_df = pd.DataFrame(theme_counter.most_common(30), columns=["题材","涨停支数"])

wb = Workbook()

# ── Sheet1: 概览 ──
ws0 = wb.active
ws0.title = "概览"
ws0.sheet_view.showGridLines = False
ws0.merge_cells("B2:I3")
c = ws0["B2"]
c.value = "问财涨停数据抓取成果汇总报告"
c.font = Font(bold=True, size=20, color=WHITE, name="微软雅黑")
c.fill = hf(RED_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws0.merge_cells("B4:I4")
c = ws0["B4"]
c.value = f"数据区间：{daily['交易日期'].min()} ~ {daily['交易日期'].max()}  |  生成时间：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
c.font = Font(size=11, color=WHITE, name="微软雅黑")
c.fill = hf(RED_MID)
c.alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[2].height = 40
ws0.row_dimensions[3].height = 40
ws0.row_dimensions[4].height = 24

kpis = [("交易日数",f"{len(daily)} 天","B"),("涨停总支次",f"{len(df):,} 次","D"),
        ("最高连板",f"{df['连板数'].max()} 板","F"),
        ("涨停最多日",daily.sort_values('涨停总数',ascending=False).iloc[0]['交易日期'],"H")]
for label, val, col in kpis:
    for r, v, fsize, fcolor, bg in [(6,label,10,GRAY_MID,GRAY_LIGHT),(7,val,18,RED_DARK,WHITE)]:
        c = ws0.cell(row=r, column=ord(col)-64)
        c.value = v
        c.font = Font(bold=True, size=fsize, color=fcolor, name="微软雅黑")
        c.fill = hf(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = tb()
ws0.row_dimensions[6].height = 22
ws0.row_dimensions[7].height = 40

ws0["B9"].value = "每日涨停概况"
ws0["B9"].font = Font(bold=True, size=12, color=GRAY_DARK, name="微软雅黑")
ws0.row_dimensions[9].height = 28
for j, h in enumerate(["交易日期","涨停总数","首板数","连板数量","最高连板","平均涨幅(%)"], 2):
    hdr(ws0.cell(row=10, column=j))
    ws0.cell(row=10, column=j).value = h
for i, row in daily.iterrows():
    r = 11+i
    bg = RED_LIGHT if i%2==0 else WHITE
    for j, v in enumerate([row["交易日期"],row["涨停总数"],row["首板数"],row["连板数量"],row["最高连板"],row["平均涨幅(%)"]], 2):
        c = ws0.cell(row=r, column=j)
        c.value = v
        dat(c)
        c.fill = hf(bg)
        if j == 3: c.font = Font(bold=True, size=9, color=RED_DARK, name="微软雅黑")
    ws0.row_dimensions[r].height = 18
ws0.column_dimensions["A"].width = 2
for cl, w in zip(["B","C","D","E","F","G","H","I"],[14,10,8,10,10,8,14,5]):
    ws0.column_dimensions[cl].width = w

# ── Sheet2: 每日统计 ──
ws1 = wb.create_sheet("每日统计")
ws1.sheet_view.showGridLines = False
ws1.merge_cells("A1:F1")
c = ws1["A1"]; c.value = "每日涨停统计"
c.font = Font(bold=True, size=14, color=WHITE, name="微软雅黑")
c.fill = hf(RED_DARK); c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 32
for j, h in enumerate(["交易日期","涨停总数","首板数","连板数量","最高连板","平均涨幅(%)"],1):
    c = ws1.cell(row=2, column=j); c.value = h; hdr(c)
for i, row in daily.iterrows():
    r = 3+i; bg = RED_LIGHT if i%2==0 else WHITE
    for j, v in enumerate([row["交易日期"],row["涨停总数"],row["首板数"],row["连板数量"],row["最高连板"],row["平均涨幅(%)"]],1):
        c = ws1.cell(row=r, column=j); c.value = v; dat(c); c.fill = hf(bg)
sr = 3+len(daily)
for j, f in enumerate([("合计/平均",None),(f"=SUM(B3:B{sr-1})",None),(f"=SUM(C3:C{sr-1})",None),
                        (f"=SUM(D3:D{sr-1})",None),(f"=MAX(E3:E{sr-1})",None),(f"=AVERAGE(F3:F{sr-1})",None)],1):
    c = ws1.cell(row=sr, column=j); c.value = f[0]; c.font = Font(bold=True, size=10, color=WHITE, name="微软雅黑")
    c.fill = hf(ORANGE); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = tb()
ws1.conditional_formatting.add(f"B3:B{sr-1}",
    ColorScaleRule(start_type="min",start_color="FFFFFF",end_type="max",end_color="C0392B"))
for cl, w in zip(["A","B","C","D","E","F"],[13,10,8,10,10,12]): ws1.column_dimensions[cl].width = w
ws1.row_dimensions[2].height = 22
chart1 = BarChart(); chart1.type = "col"; chart1.title = "每日涨停支数"
chart1.height = 14; chart1.width = 22
d1 = Reference(ws1, min_col=2, min_row=2, max_row=2+len(daily))
c1 = Reference(ws1, min_col=1, min_row=3, max_row=2+len(daily))
chart1.add_data(d1, titles_from_data=True); chart1.set_categories(c1)
chart1.series[0].graphicalProperties.solidFill = RED_MID
ws1.add_chart(chart1, "H2")

# ── Sheet3: 连板分析 ──
ws2 = wb.create_sheet("连板分析")
ws2.sheet_view.showGridLines = False
ws2.merge_cells("A1:D1")
c = ws2["A1"]; c.value = "连板数分布统计"
c.font = Font(bold=True, size=14, color=WHITE, name="微软雅黑")
c.fill = hf(RED_DARK); c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32
for j, h in enumerate(["连板数","支数","占比(%)","说明"],1):
    c = ws2.cell(row=2, column=j); c.value = h; hdr(c)
total_lb = lb_dist["支数"].sum()
labels = {1:"首板",2:"二连板",3:"三连板",4:"四连板",5:"五连板"}
for i, row in lb_dist.iterrows():
    r = 3+i; bg = RED_LIGHT if i%2==0 else WHITE
    pct = round(row["支数"]/total_lb*100, 1)
    note = labels.get(int(row["连板数"]), f"{int(row['连板数'])}连板")
    for j, v in enumerate([row["连板数"],row["支数"],pct,note],1):
        c = ws2.cell(row=r, column=j); c.value = v; dat(c); c.fill = hf(bg)
sr2 = 3+len(lb_dist)
for j, v in enumerate(["合计",f"=SUM(B3:B{sr2-1})","100%",""],1):
    c = ws2.cell(row=sr2, column=j); c.value = v
    c.font = Font(bold=True, size=10, color=WHITE, name="微软雅黑")
    c.fill = hf(ORANGE); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = tb()

high_lb = df[df["连板数"]>=3][["交易日期","股票代码","股票名称","连板数","涨停类型","所属题材"]].sort_values(
    ["连板数","交易日期"],ascending=[False,True]).reset_index(drop=True)
start_r = sr2+2
ws2.cell(row=start_r, column=1).value = "高连板个股明细（>=3板）"
ws2.cell(row=start_r, column=1).font = Font(bold=True, size=12, color=GRAY_DARK, name="微软雅黑")
hd_r = start_r+1
for j, h in enumerate(["交易日期","股票代码","股票名称","连板数","涨停类型","所属题材"],1):
    c = ws2.cell(row=hd_r, column=j); c.value = h; hdr(c)
for i, row in high_lb.iterrows():
    r = hd_r+1+i; bg = RED_LIGHT if i%2==0 else WHITE
    for j, v in enumerate(row.values,1):
        c = ws2.cell(row=r, column=j); c.value = v
        dat(c, align="left" if j==6 else "center"); c.fill = hf(bg)
        if j==4: c.font = Font(bold=True, size=9, color=RED_DARK, name="微软雅黑")
for cl, w in zip(["A","B","C","D","E","F"],[12,10,10,8,12,35]): ws2.column_dimensions[cl].width = w
chart2 = BarChart(); chart2.type = "col"; chart2.title = "连板数分布"
chart2.height = 12; chart2.width = 16
d2 = Reference(ws2, min_col=2, min_row=2, max_row=2+len(lb_dist))
c2 = Reference(ws2, min_col=1, min_row=3, max_row=2+len(lb_dist))
chart2.add_data(d2, titles_from_data=True); chart2.set_categories(c2)
chart2.series[0].graphicalProperties.solidFill = GOLD
ws2.add_chart(chart2, "H2")

# ── Sheet4: 题材热度 ──
ws3 = wb.create_sheet("题材热度")
ws3.sheet_view.showGridLines = False
ws3.merge_cells("A1:C1")
c = ws3["A1"]; c.value = "涨停题材热度 TOP 30"
c.font = Font(bold=True, size=14, color=WHITE, name="微软雅黑")
c.fill = hf(RED_DARK); c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32
for j, h in enumerate(["排名","题材","涨停支数"],1):
    c = ws3.cell(row=2, column=j); c.value = h; hdr(c)
for i, row in theme_df.iterrows():
    r = 3+i; bg = RED_LIGHT if i%2==0 else WHITE
    for j, v in enumerate([i+1, row["题材"], row["涨停支数"]],1):
        c = ws3.cell(row=r, column=j); c.value = v
        dat(c, align="left" if j==2 else "center"); c.fill = hf(bg)
        if j==3: c.font = Font(bold=True, size=9, color=RED_DARK, name="微软雅黑")
ws3.conditional_formatting.add(f"C3:C{2+len(theme_df)}",
    ColorScaleRule(start_type="min",start_color="FFFFFF",end_type="max",end_color="C0392B"))
for cl, w in zip(["A","B","C"],[6,25,12]): ws3.column_dimensions[cl].width = w
chart3 = BarChart(); chart3.type = "bar"; chart3.title = "热门题材涨停支数 TOP 20"
chart3.height = 20; chart3.width = 20
d3 = Reference(ws3, min_col=3, min_row=2, max_row=22)
c3 = Reference(ws3, min_col=2, min_row=3, max_row=22)
chart3.add_data(d3, titles_from_data=True); chart3.set_categories(c3)
chart3.series[0].graphicalProperties.solidFill = RED_MID
ws3.add_chart(chart3, "E2")

# ── Sheet5: 全量明细 ──
ws4 = wb.create_sheet("全量明细")
ws4.sheet_view.showGridLines = False
ws4.merge_cells("A1:R1")
c = ws4["A1"]; c.value = f"涨停个股全量明细  共 {len(df)} 条记录"
c.font = Font(bold=True, size=13, color=WHITE, name="微软雅黑")
c.fill = hf(RED_DARK); c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 28
detail_cols = ["交易日期","股票代码","股票名称","收盘价","涨跌幅(%)","连板数",
               "涨停时间","首次涨停","涨停状态","涨停类型","成交额","成交量",
               "流通市值(亿)","总市值(亿)","量比","所属题材","涨停标签","入库时间"]
for j, h in enumerate(detail_cols,1):
    c = ws4.cell(row=2, column=j); c.value = h; hdr(c, size=9)
for i, row in df[detail_cols].iterrows():
    r = 3+i; bg = RED_LIGHT if i%2==0 else WHITE
    for j, col_name in enumerate(detail_cols,1):
        c = ws4.cell(row=r, column=j); c.value = row[col_name]
        dat(c, align="left" if col_name in ["所属题材","涨停标签"] else "center")
        c.fill = hf(bg)
        if col_name == "连板数" and isinstance(row[col_name],(int,float)) and row[col_name]>=3:
            c.font = Font(bold=True, size=9, color=RED_DARK, name="微软雅黑")
    ws4.row_dimensions[r].height = 16
for j, w in enumerate([12,10,10,8,10,8,10,10,10,10,12,12,14,12,8,35,30,18],1):
    ws4.column_dimensions[get_column_letter(j)].width = w
ws4.freeze_panes = "A3"
ws4.auto_filter.ref = f"A2:{get_column_letter(len(detail_cols))}2"

wb.save(OUT_PATH)
print(f"OK: {OUT_PATH}  ({len(df)} 条 / {len(daily)} 个交易日)")
