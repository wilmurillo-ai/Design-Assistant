#!/usr/bin/env python3
"""
室内设计自动预算生成器
Usage: python generate_budget.py <input.dxf> <output.xlsx> [total_area]
"""

import json
import sys
import ezdxf
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 加载标准项目库
DEFAULT_LIBRARY = "/Users/laobaobei/.openclaw/workspace/标准预算项目库.json"

def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    
    dxf_path = sys.argv[1]
    output_path = sys.argv[2]
    total_area = float(sys.argv[3]) if len(sys.argv) > 3 else None
    
    # 1. 读取DXF提取项目信息
    print(f"读取DXF文件: {dxf_path}")
    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()
    
    # 提取文字找项目名和层数
    project_name = "未知项目"
    floors_count = 1
    texts = []
    for entity in msp:
        if entity.dxftype() in ["TEXT", "MTEXT"]:
            if entity.dxftype() == "TEXT":
                t = entity.dxf.text.strip()
            else:
                t = entity.plain_text().strip()
            if len(t) > 1:
                texts.append(t)
                if "幢" in t and "单元" in t:
                    project_name = t
    
    # 估算层数
    if any("二层" in t for t in texts):
        floors_count = 2
    if any("三层" in t for t in texts):
        floors_count = 3
    if any("地下室" in t for t in texts):
        floors_count += 1
    
    # 如果没传面积，按平均估算
    if total_area is None:
        total_area = floors_count * 70  # 每层大概70平
    
    print(f"识别结果: 项目={project_name}, 层数={floors_count}, 估算面积={total_area}㎡")
    
    # 2. 加载标准项目库
    with open(DEFAULT_LIBRARY, "r", encoding="utf-8") as f:
        data = json.load(f)
    all_projects = data["projects"]
    print(f"加载标准项目: {len(all_projects)} 项")
    
    # 3. 分类分配项目（按空间）
    category_template = [
        ("一、新建项目", ["新砌", "新建", "新旧墙", "钢筋", "现浇", "抹灰", "纤维网"]),
        ("二、拆改工程", ["拆", "铲除", "拆除"]),
        ("三、玄关/客厅/过道", ["玄关", "客厅", "过道"]),
        ("四、餐厅", ["餐厅"]),
        ("五、厨房", ["厨房"]),
        ("六、卫生间", ["公卫", "主卫", "卫生间", "防水"]),
        ("七、阳台", ["阳台"]),
        ("八、卧室", ["主卧", "次卧", "男孩", "女孩"]),
        ("水电项目", ["水电", "开槽", "强电", "弱电", "管线", "水管", "底盒"]),
        ("其它项目", ["安装", "灯具", "开关", "插座", "五金", "卫浴", "地板", "门", "保洁", "运输", "管理", "垃圾"]),
    ]
    
    # 按层数展开
    final = []
    for f in range(floors_count):
        floor_name = f"第{f+1}层"
        for cat_name, keywords in category_template[:-2]:
            projects_in_cat = []
            for p in all_projects:
                if p["综合单价"] is None or p["综合单价"] <= 0:
                    continue
                matched = False
                for kw in keywords:
                    if kw in p["name"]:
                        matched = True
                        break
                if not matched and cat_name.lower() in p["name"].lower():
                    matched = True
                if matched:
                    qty = estimate_qty(p["name"], p["unit"], total_area)
                    p_new = p.copy()
                    p_new["qty"] = qty
                    p_new["amount"] = qty * p["combined_price"]
                    projects_in_cat.append(p_new)
            if len(projects_in_cat) > 0:
                final.append({
                    "floor": floor_name if floors_count > 1 else "",
                    "category": cat_name,
                    "projects": projects_in_cat
                })
    
    # 最后加水电和其他
    for cat_name, keywords in category_template[-2:]:
        projects_in_cat = []
        for p in all_projects:
            if p["综合单价"] is None or p["综合单价"] <= 0:
                continue
            matched = False
            for kw in keywords:
                if kw in p["name"]:
                    matched = True
                    break
            if not matched and cat_name.lower() in p["name"].lower():
                matched = True
            if matched:
                qty = estimate_qty(p["name"], p["unit"], total_area)
                p_new = p.copy()
                p_new["qty"] = qty
                p_new["amount"] = qty * p["combined_price"]
                projects_in_cat.append(p_new)
        if len(projects_in_cat) > 0:
            final.append({
                "floor": "",
                "category": cat_name,
                "projects": projects_in_cat
            })
    
    # 4. 生成Excel
    total_projects = sum(len(s["projects"]) for s in final)
    total_amount = sum(p["amount"] for s in final for p in s["projects"])
    print(f"整理完成: 总项目数={total_projects}, 估算总价={round(total_amount, 0)}元")
    
    generate_excel(final, project_name, total_area, output_path)
    print(f"\n✅ 生成完成! 输出: {output_path}")
    print(f"项目: {project_name}, 面积: {total_area}㎡, 项目数: {total_projects}, 总价: {round(total_amount, 0)}元")


def estimate_qty(name, unit, total_area=210):
    """根据项目类型估算工程量"""
    area = total_area
    if unit == "㎡":
        if "墙面" in name and "乳胶漆" in name:
            return round(area * 3.0, 2)
        elif "顶面" in name and "乳胶漆" in name:
            return round(area * 1.0, 2)
        elif "地砖" in name or "地面" in name:
            return round(area * 0.65, 2)
        elif "墙砖" in name:
            return round(area * 1.3, 2)
        elif "地板" in name:
            return round(area * 0.5, 2)
        elif "新砌墙" in name:
            return round(area * 0.4, 2)
        elif "吊顶" in name:
            return round(area * 0.35, 2)
        elif "防水" in name:
            return round(area * 0.18, 2) * 3  # 每个卫生间平均
        elif "找平" in name:
            return round(area * 0.8, 2)
        return round(area * 0.15, 2)
    elif unit == "m":
        if "包下水管" in name:
            return 3 * int((area / 70))  # 每层3根
        elif "门槛石" in name:
            return 2 * int((area / 70))
        elif "窗帘盒" in name:
            return 8 * int((area / 70))
        elif "管线" in name or "电" in name:
            return round(area * 4.5, 2)
        return round(area * 0.6, 2)
    elif unit == "项":
        return 1
    elif unit == "樘":
        if "门" in name:
            return 3 * int((area / 70))
        return 1
    else:
        return 1


def generate_excel(final_data, project_name, total_area, output_path):
    """生成Excel，和传统模板格式一致"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "预算表"

    # 样式
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题
    ws.merge_cells('A1:L1')
    ws['A1'] = "工程装饰（预）算书"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:L2')
    ws['A2'] = f"工程名称：{project_name}                       面积：{total_area}m²"
    ws['A2'].alignment = Alignment(horizontal='left')

    # 表头两行
    row = 3
    ws.cell(row=row, column=1, value="序号").font = header_font
    ws.cell(row=row, column=2, value="工程项目名称").font = header_font
    ws.cell(row=row, column=11, value="备    注").font = header_font
    for c in range(1, 12):
        ws.cell(row=row, column=c).border = thin_border

    row = 4
    ws.cell(row=row, column=3, value="单位").font = header_font
    ws.cell(row=row, column=4, value="工程量").font = header_font
    ws.cell(row=row, column=5, value="综合单价").font = header_font
    ws.cell(row=row, column=6, value="金额（元）").font = header_font
    ws.cell(row=row, column=7, value="人工").font = header_font
    ws.cell(row=row, column=8, value="辅材").font = header_font
    ws.cell(row=row, column=9, value="主材").font = header_font
    ws.cell(row=row, column=10, value="损耗").font = header_font
    for c in range(1, 12):
        ws.cell(row=row, column=c).border = thin_border

    row = 5
    seq_global = 1
    total_amount = 0

    for data in final_data:
        if data["floor"] != "":
            # 楼层标题
            ws.merge_cells(f'A{row}:L{row}')
            cell = ws.cell(row=row, column=1, value=data["floor"])
            cell.font = Font(bold=True)
            row += 1
        
        # 分类标题
        ws.merge_cells(f'A{row}:L{row}')
        cell = ws.cell(row=row, column=1, value=data["category"])
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        row += 1
        
        seq = 1
        for proj in data["projects"]:
            amount = proj["amount"]
            total_amount += amount
            
            ws.cell(row=row, column=1, value=seq).border = thin_border
            ws.cell(row=row, column=2, value=proj["name"]).border = thin_border
            ws.cell(row=row, column=3, value=proj["unit"]).border = thin_border
            ws.cell(row=row, column=4, value=round(proj["qty"], 2)).border = thin_border
            ws.cell(row=row, column=5, value=round(proj["combined_price"], 2)).border = thin_border
            ws.cell(row=row, column=6, value=round(amount, 2)).border = thin_border
            ws.cell(row=row, column=7, value=proj["labor_price"] if proj["labor_price"] is not None else "").border = thin_border
            ws.cell(row=row, column=8, value=proj["aux_price"] if proj["aux_price"] is not None else "").border = thin_border
            ws.cell(row=row, column=9, value=proj["material_price"] if proj["material_price"] is not None else "").border = thin_border
            ws.cell(row=row, column=10, value="").border = thin_border
            note = proj["note"].replace("\n", " ") if proj["note"] else ""
            ws.cell(row=row, column=11, value=note[:60]).border = thin_border
            
            row += 1
            seq += 1
            seq_global += 1

    # 总计
    row += 1
    ws.cell(row=row, column=1, value="总  计").font = Font(bold=True, size=14)
    ws.merge_cells(f'B{row}:E{row}')
    cell = ws.cell(row=row, column=6, value=round(total_amount, 2))
    cell.font = Font(bold=True, size=14)
    cell.border = thin_border

    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 6
    ws.column_dimensions['K'].width = 45

    wb.save(output_path)


if __name__ == "__main__":
    main()
