from __future__ import annotations

from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.news_item import NewsItem


class ExcelExporter:
    """
    Excel 导出器（增强显示版）。
    支持：
    - summary 概览 sheet
    - 主结果 sheet
    - 重复项 sheet
    - 无效项 sheet
    - 排序增强
    - 自动筛选 / 冻结首行 / 列宽调整 / 重要性着色 / 超链接样式
    """

    SUMMARY_SHEET_NAME = "summary"
    MAIN_SHEET_NAME = "news_table"
    DUP_SHEET_NAME = "duplicate_items"
    INVALID_SHEET_NAME = "invalid_items"

    def __init__(self, output_dir: str, filename_prefix: str = "AI资讯信息表") -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.filename_prefix = filename_prefix

    def export(
        self,
        items: List[NewsItem],
        run_date: str,
        filename: Optional[str] = None,
    ) -> Path:
        return self.export_full(
            valid_items=items,
            duplicate_items=[],
            invalid_items=[],
            run_date=run_date,
            filename=filename,
        )

    def export_full(
        self,
        valid_items: List[NewsItem],
        duplicate_items: List[NewsItem],
        invalid_items: List[NewsItem],
        run_date: str,
        filename: Optional[str] = None,
    ) -> Path:
        if filename:
            file_path = self.output_dir / filename
        else:
            safe_date = run_date.replace("-", "")
            file_path = self.output_dir / f"{self.filename_prefix}_{safe_date}.xlsx"

        main_df = self._build_dataframe(valid_items)
        dup_df = self._build_dataframe(duplicate_items)
        invalid_df = self._build_dataframe(invalid_items)

        main_df = self._sort_main_df(main_df)
        dup_df = self._sort_secondary_df(dup_df)
        invalid_df = self._sort_secondary_df(invalid_df)

        summary_df = self._build_summary_dataframe(
            valid_df=main_df,
            duplicate_df=dup_df,
            invalid_df=invalid_df,
            run_date=run_date,
        )

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, index=False, sheet_name=self.SUMMARY_SHEET_NAME)
            main_df.to_excel(writer, index=False, sheet_name=self.MAIN_SHEET_NAME)
            dup_df.to_excel(writer, index=False, sheet_name=self.DUP_SHEET_NAME)
            invalid_df.to_excel(writer, index=False, sheet_name=self.INVALID_SHEET_NAME)

            self._format_summary_sheet(writer.sheets[self.SUMMARY_SHEET_NAME], summary_df)
            self._format_sheet(writer.sheets[self.MAIN_SHEET_NAME], main_df)
            self._format_sheet(writer.sheets[self.DUP_SHEET_NAME], dup_df)
            self._format_sheet(writer.sheets[self.INVALID_SHEET_NAME], invalid_df)

        return file_path

    def _build_dataframe(self, items: List[NewsItem]) -> pd.DataFrame:
        rows = [item.to_excel_row() for item in items]
        df = pd.DataFrame(rows)

        if df.empty:
            df = pd.DataFrame(columns=self._default_columns())

        return df

    @staticmethod
    def _default_columns() -> List[str]:
        return [
            "资讯日期",
            "抓取日期",
            "时间窗口",
            "一级类别",
            "二级类别",
            "关联企业",
            "国家/地区",
            "标题",
            "摘要",
            "事件要点",
            "来源名称",
            "来源链接",
            "发布时间",
            "抓取时间",
            "是否官方来源",
            "重要性等级",
            "去重标记",
            "入表状态",
            "备注",
        ]

    def _sort_main_df(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df

        df = df.copy()

        importance_order = {"高": 0, "中": 1, "低": 2, "": 3}
        df["_importance_sort"] = df["重要性等级"].map(lambda x: importance_order.get(str(x), 3))
        df["_publish_sort"] = pd.to_datetime(df["发布时间"], errors="coerce")

        sort_cols = ["_importance_sort", "一级类别", "二级类别", "_publish_sort", "标题"]
        ascending = [True, True, True, False, True]

        df = df.sort_values(by=sort_cols, ascending=ascending, na_position="last")
        df = df.drop(columns=["_importance_sort", "_publish_sort"])
        df = df.reset_index(drop=True)
        return df

    def _sort_secondary_df(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df

        df = df.copy()
        if "发布时间" in df.columns:
            df["_publish_sort"] = pd.to_datetime(df["发布时间"], errors="coerce")
            df = df.sort_values(by=["_publish_sort", "标题"], ascending=[False, True], na_position="last")
            df = df.drop(columns=["_publish_sort"])
        else:
            df = df.sort_values(by=["标题"], ascending=[True])

        df = df.reset_index(drop=True)
        return df

    def _build_summary_dataframe(
        self,
        valid_df: pd.DataFrame,
        duplicate_df: pd.DataFrame,
        invalid_df: pd.DataFrame,
        run_date: str,
    ) -> pd.DataFrame:
        time_window = ""
        if not valid_df.empty and "时间窗口" in valid_df.columns:
            vals = [v for v in valid_df["时间窗口"].astype(str).tolist() if v and v != "nan"]
            if vals:
                time_window = vals[0]
        elif not invalid_df.empty and "时间窗口" in invalid_df.columns:
            vals = [v for v in invalid_df["时间窗口"].astype(str).tolist() if v and v != "nan"]
            if vals:
                time_window = vals[0]

        total_count = len(valid_df) + len(duplicate_df) + len(invalid_df)
        high_count = self._count_value(valid_df, "重要性等级", "高")
        medium_count = self._count_value(valid_df, "重要性等级", "中")
        low_count = self._count_value(valid_df, "重要性等级", "低")

        rows = [
            {"模块": "运行概览", "指标": "运行日期", "值": run_date},
            {"模块": "运行概览", "指标": "统计窗口", "值": time_window},
            {"模块": "运行概览", "指标": "总处理条数", "值": total_count},
            {"模块": "运行概览", "指标": "有效资讯条数", "值": len(valid_df)},
            {"模块": "运行概览", "指标": "重复资讯条数", "值": len(duplicate_df)},
            {"模块": "运行概览", "指标": "无效资讯条数", "值": len(invalid_df)},
            {"模块": "重要性分布", "指标": "高", "值": high_count},
            {"模块": "重要性分布", "指标": "中", "值": medium_count},
            {"模块": "重要性分布", "指标": "低", "值": low_count},
        ]

        if not valid_df.empty and "一级类别" in valid_df.columns:
            level1_counts = valid_df["一级类别"].fillna("").astype(str).value_counts()
            for k, v in level1_counts.items():
                if k:
                    rows.append({"模块": "一级分类分布", "指标": k, "值": int(v)})

        if not valid_df.empty and "二级类别" in valid_df.columns:
            level2_counts = valid_df["二级类别"].fillna("").astype(str).value_counts()
            for k, v in level2_counts.items():
                if k:
                    rows.append({"模块": "二级分类分布", "指标": k, "值": int(v)})

        return pd.DataFrame(rows)

    @staticmethod
    def _count_value(df: pd.DataFrame, col: str, value: str) -> int:
        if df.empty or col not in df.columns:
            return 0
        return int((df[col].fillna("").astype(str) == value).sum())

    def _format_sheet(self, ws, df: pd.DataFrame) -> None:
        self._freeze_header(ws)
        self._apply_auto_filter(ws)
        self._style_header(ws)
        self._adjust_column_width(ws, df)
        self._apply_wrap_text(ws)
        self._highlight_importance(ws)
        self._style_hyperlinks(ws)

    def _format_summary_sheet(self, ws, df: pd.DataFrame) -> None:
        self._freeze_header(ws)
        self._apply_auto_filter(ws)
        self._style_header(ws)
        self._adjust_column_width(ws, df)
        self._apply_wrap_text(ws)

        module_fill_map = {
            "运行概览": PatternFill(fill_type="solid", fgColor="D9EAF7"),
            "重要性分布": PatternFill(fill_type="solid", fgColor="FCE4D6"),
            "一级分类分布": PatternFill(fill_type="solid", fgColor="E2F0D9"),
            "二级分类分布": PatternFill(fill_type="solid", fgColor="FFF2CC"),
        }

        for row in range(2, ws.max_row + 1):
            module_value = str(ws.cell(row=row, column=1).value or "").strip()
            fill = module_fill_map.get(module_value)
            if fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill

    @staticmethod
    def _freeze_header(ws) -> None:
        ws.freeze_panes = "A2"

    @staticmethod
    def _apply_auto_filter(ws) -> None:
        ws.auto_filter.ref = ws.dimensions

    @staticmethod
    def _style_header(ws) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        header_font = Font(bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def _adjust_column_width(ws, df: pd.DataFrame) -> None:
        for idx, col in enumerate(df.columns, start=1):
            max_len = len(str(col))

            if not df.empty:
                series_max = df[col].astype(str).map(len).max()
                max_len = max(max_len, int(series_max))

            if col in {"标题"}:
                width = min(max(max_len + 2, 20), 50)
            elif col in {"摘要", "事件要点", "备注", "值"}:
                width = min(max(max_len + 2, 18), 60)
            elif col in {"来源链接"}:
                width = 50
            elif col in {"模块", "指标"}:
                width = min(max(max_len + 2, 12), 24)
            else:
                width = min(max_len + 2, 24)

            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = width

    @staticmethod
    def _apply_wrap_text(ws) -> None:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    @staticmethod
    def _highlight_importance(ws) -> None:
        header_map = {}
        for idx, cell in enumerate(ws[1], start=1):
            header_map[cell.value] = idx

        importance_col_idx = header_map.get("重要性等级")
        if not importance_col_idx:
            return

        high_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        medium_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        low_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=importance_col_idx)
            value = str(cell.value or "").strip()

            if value == "高":
                cell.fill = high_fill
            elif value == "中":
                cell.fill = medium_fill
            elif value == "低":
                cell.fill = low_fill

    @staticmethod
    def _style_hyperlinks(ws) -> None:
        header_map = {}
        for idx, cell in enumerate(ws[1], start=1):
            header_map[cell.value] = idx

        link_col_idx = header_map.get("来源链接")
        if not link_col_idx:
            return

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=link_col_idx)
            value = str(cell.value or "").strip()

            if value.startswith("http://") or value.startswith("https://"):
                cell.hyperlink = value
                cell.style = "Hyperlink"