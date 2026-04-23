from __future__ import annotations

from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.news_item import NewsItem


class SimpleExcelExporter:
    """
    简表版 Excel 导出器：
    - 仅导出有效数据
    - 仅保留一个 sheet
    - 仅保留指定表头
    - 按一级分类排序
    - 名单企业动态：浅蓝底
    - 全球AI产业动态：浅绿底
    - 增加完整边框线
    """

    SHEET_NAME = "valid_items"

    def __init__(self, output_dir: str, filename_prefix: str = "AI资讯有效数据简表") -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.filename_prefix = filename_prefix

    def export(
        self,
        valid_items: List[NewsItem],
        run_date: str,
        filename: Optional[str] = None,
    ) -> Path:
        if filename:
            file_path = self.output_dir / filename
        else:
            safe_date = run_date.replace("-", "")
            file_path = self.output_dir / f"{self.filename_prefix}_{safe_date}.xlsx"

        df = self._build_dataframe(valid_items)
        df = self._sort_dataframe(df)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=self.SHEET_NAME)
            ws = writer.sheets[self.SHEET_NAME]

            self._freeze_header(ws)
            self._style_header(ws)
            self._apply_wrap_text(ws)
            self._adjust_column_width(ws, df)
            self._style_hyperlinks(ws)
            self._fill_rows_by_category(ws)
            self._apply_borders(ws)

        return file_path

    def _build_dataframe(self, items: List[NewsItem]) -> pd.DataFrame:
        rows = []
        for item in items:
            rows.append({
                "资讯日期": item.event_date or "",
                "一级类别": item.category_level_1 or "",
                "关联企业": "；".join(item.related_companies) if item.related_companies else "",
                "标题": item.title_zh or item.title or "",
                "摘要": item.summary_zh or item.summary or "",
                "来源名称": item.source_name or "",
                "发布时间": item.format_dt(item.publish_time),
                "来源链接": item.source_url or "",
                "全球动态机构归属": item.global_region_label or "",
            })

        df = pd.DataFrame(rows)
        if df.empty:
            df = pd.DataFrame(columns=[
                "资讯日期", "一级类别", "关联企业", "标题", "摘要", "来源名称", "发布时间", "来源链接"
            ])
        return df

    @staticmethod
    def _sort_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df

        df = df.copy()

        def level1_sort(row):
            level1 = str(row.get("一级类别", "")).strip()
            region = str(row.get("全球动态机构归属", "")).strip()

            if level1 == "名单企业动态":
                return 0
            if level1 == "全球AI产业动态" and region == "国内":
                return 1
            if level1 == "全球AI产业动态" and region == "国际":
                return 2
            if level1 == "全球AI产业动态":
                return 3
            return 9

        df["_level1_sort"] = df.apply(level1_sort, axis=1)
        df["_publish_sort"] = pd.to_datetime(df["发布时间"], errors="coerce")

        df = df.sort_values(
            by=["_level1_sort", "_publish_sort", "标题"],
            ascending=[True, False, True],
            na_position="last",
        ).drop(columns=["_level1_sort", "_publish_sort", "全球动态机构归属"])

        df = df.reset_index(drop=True)
        return df

    @staticmethod
    def _freeze_header(ws) -> None:
        ws.freeze_panes = "A2"

    @staticmethod
    def _style_header(ws) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor="5B9BD5")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def _apply_wrap_text(ws) -> None:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    @staticmethod
    def _adjust_column_width(ws, df: pd.DataFrame) -> None:
        for idx, col in enumerate(df.columns, start=1):
            max_len = len(str(col))
            if not df.empty:
                series_max = df[col].astype(str).map(len).max()
                max_len = max(max_len, int(series_max))

            if col == "标题":
                width = min(max(max_len + 2, 18), 45)
            elif col == "摘要":
                width = min(max(max_len + 2, 20), 60)
            elif col == "来源链接":
                width = 50
            else:
                width = min(max_len + 2, 24)

            ws.column_dimensions[get_column_letter(idx)].width = width

    @staticmethod
    def _style_hyperlinks(ws) -> None:
        header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        link_col_idx = header_map.get("来源链接")
        if not link_col_idx:
            return

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=link_col_idx)
            value = str(cell.value or "").strip()
            if value.startswith("http://") or value.startswith("https://"):
                cell.hyperlink = value
                cell.style = "Hyperlink"

    @staticmethod
    def _fill_rows_by_category(ws) -> None:
        header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        level1_col_idx = header_map.get("一级类别")
        if not level1_col_idx:
            return

        company_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")  # 浅蓝
        global_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")   # 浅绿

        for row in range(2, ws.max_row + 1):
            level1_value = str(ws.cell(row=row, column=level1_col_idx).value or "").strip()

            if level1_value == "名单企业动态":
                fill = company_fill
            elif level1_value == "全球AI产业动态":
                fill = global_fill
            else:
                continue

            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    @staticmethod
    def _apply_borders(ws) -> None:
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border