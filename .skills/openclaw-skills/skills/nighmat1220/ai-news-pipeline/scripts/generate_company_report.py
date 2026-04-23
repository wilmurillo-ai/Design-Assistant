from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from zipfile import BadZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from brief_writer import load_latest_international_rows, write_daily_brief
from time_window import (
    existing_paths_from_stems,
    in_time_window,
    parse_time_window,
    target_date_strings_from_window,
)
from openpyxl import load_workbook


BASE_DIR = Path(os.getenv("AI_NEWS_WORKSPACE", os.getcwd())).resolve()
DATA_DIR = BASE_DIR / "data"
REPORT_DIR = BASE_DIR / "reports"
COMPANIES_PATH = BASE_DIR / "companies.txt"
STATE_DIR = BASE_DIR / "state"
SUMMARY_CACHE_PATH = STATE_DIR / "company_report_ai_cache.json"

ARK_API_BASE = os.getenv("ARK_API_BASE", "https://ark.cn-beijing.volces.com/api/v3")
ARK_MODEL = os.getenv("ARK_MODEL", "doubao-seed-2-0-mini-260215")
ARK_API_KEY = os.getenv("ARK_API_KEY", "").strip()
ARK_TIMEOUT_SECONDS = int(os.getenv("ARK_TIMEOUT_SECONDS", "60"))
MAX_CACHE_ITEMS = 5000


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate company news report from JSONL data files.")
    parser.add_argument(
        "--companies",
        type=Path,
        default=COMPANIES_PATH,
        help=f"Path to the company list file. Default: {COMPANIES_PATH}",
    )
    parser.add_argument(
        "--days",
        type=int,
        default=2,
        help="How many days of files to scan, counting today. Default: 2",
    )
    parser.add_argument(
        "--disable-ai",
        action="store_true",
        help="Skip AI summary/title generation even if ARK_API_KEY is set.",
    )
    parser.add_argument(
        "--time-window",
        type=str,
        default="",
        help="资讯时间窗口，例如：2026年3月15日9点到2026年3月15日18点",
    )
    return parser.parse_args()


def configure_output_encoding() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            reconfigure(encoding="utf-8", errors="replace")


def ensure_directories() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    COMPANIES_PATH.parent.mkdir(parents=True, exist_ok=True)


def dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value not in seen:
            result.append(value)
            seen.add(value)
    return result


def load_companies(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"Company list file not found: {path}")

    companies: list[dict[str, Any]] = []
    seen: set[str] = set()

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        parts = [part.strip() for part in line.split("|") if part.strip()]
        if not parts:
            continue

        canonical_name = parts[0]
        if canonical_name in seen:
            continue

        companies.append(
            {
                "name": canonical_name,
                "aliases": dedupe_preserve_order(parts),
            }
        )
        seen.add(canonical_name)

    return companies

def target_data_files(days: int) -> list[Path]:
    today = datetime.now().date()
    paths: list[Path] = []
    for offset in range(days):
        day = today - timedelta(days=offset)
        path = DATA_DIR / f"{day.isoformat()}.jsonl"
        if path.exists():
            paths.append(path)
    return paths


def target_data_files_by_window(time_window: str) -> tuple[list[Path], tuple[datetime, datetime]]:
    start, end = parse_time_window(time_window)
    stems = target_date_strings_from_window(start, end)
    paths = existing_paths_from_stems(DATA_DIR, stems)
    return paths, (start, end)


def read_records(paths: list[Path]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for path in paths:
        with path.open("r", encoding="utf-8") as file:
            for line_number, line in enumerate(file, start=1):
                cleaned = line.strip()
                if not cleaned:
                    continue
                try:
                    records.append(json.loads(cleaned))
                except json.JSONDecodeError as error:
                    print(
                        f"[WARN] 跳过无效 JSON: {path.name}:{line_number} {error}",
                        file=sys.stderr,
                    )
    return records


def record_time_value(record: dict[str, Any]) -> str:
    return str(record.get("published_at", "") or "")


def match_records(records: list[dict[str, Any]], companies: list[dict[str, Any]]) -> list[dict[str, str]]:
    matched_rows: list[dict[str, str]] = []
    for record in records:
        title = str(record.get("title", "") or "")
        content = str(record.get("content", "") or "")
        if not title and not content:
            continue

        combined_text = "\n".join(part for part in [title, content] if part)
        matched_names = [
            str(company["name"])
            for company in companies
            if any(alias in combined_text for alias in company.get("aliases", []))
        ]
        if not matched_names:
            continue

        for company_name in dedupe_preserve_order(matched_names):
            matched_rows.append(
                {
                    "企业名": company_name,
                    "资讯标题": title,
                    "资讯内容": content,
                    "资讯来源": str(record.get("source_name", "") or ""),
                    "资讯时间": record_time_value(record),
                    "资讯链接": str(record.get("link", "") or ""),
                    "AI标题": "",
                    "AI摘要": "",
                }
            )

    matched_rows.sort(
        key=lambda row: (row["企业名"], row["资讯时间"], row["资讯标题"]),
        reverse=True,
    )
    return matched_rows

def filter_records_by_time_window(
    records: list[dict[str, Any]],
    start: datetime,
    end: datetime,
) -> list[dict[str, Any]]:
    return [
        record
        for record in records
        if in_time_window(record_time_value(record), start, end)
    ]


def load_summary_cache() -> dict[str, dict[str, str]]:
    if not SUMMARY_CACHE_PATH.exists():
        return {}

    try:
        with SUMMARY_CACHE_PATH.open("r", encoding="utf-8") as file:
            data = json.load(file)
    except (OSError, json.JSONDecodeError):
        return {}

    if not isinstance(data, dict):
        return {}
    return {str(key): value for key, value in data.items() if isinstance(value, dict)}


def save_summary_cache(cache: dict[str, dict[str, str]]) -> None:
    items = list(cache.items())[-MAX_CACHE_ITEMS:]
    trimmed_cache = {key: value for key, value in items}
    with SUMMARY_CACHE_PATH.open("w", encoding="utf-8") as file:
        json.dump(trimmed_cache, file, ensure_ascii=False, indent=2)


def clear_summary_cache() -> None:
    try:
        if SUMMARY_CACHE_PATH.exists():
            SUMMARY_CACHE_PATH.write_text("{}\n", encoding="utf-8")
    except OSError:
        pass


def make_cache_key(row: dict[str, str]) -> str:
    raw = "||".join(
        [
            row["企业名"],
            row["资讯标题"],
            row["资讯内容"],
            row["资讯来源"],
            row["资讯时间"],
            row["资讯链接"],
        ]
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def build_ai_payload(row: dict[str, str]) -> dict[str, Any]:
    prompt = (
        "你是一名中文商业资讯编辑。请基于下面这条资讯生成 JSON。"
        "必须只返回 JSON 对象，格式为 {\"short_title\":\"...\",\"summary\":\"...\"}。"
        "要求："
        "1. short_title 目标为15个中文字符左右，通常不超过20个中文字符。英文字符、英文单词、缩写和数字不计入字数限制。"
        "2. short_title 必须把一句话说完整，不能为了压缩字数截断半句。"
        "3. short_title 可以保留必要的英文企业名、产品名或模型名。"
        "4. summary 目标为100个中文字符左右，通常不超过120个中文字符，信息准确、客观、完整。"
        "5. summary 必须完整表达，不能截断半句，并且必须以中文句号结尾。"
        "6. 不要编造原文没有的信息。"
        "7. 如果原文信息较少，也要尽量概括重点。"
        "\n\n"
        f"企业名：{row['企业名']}\n"
        f"原标题：{row['资讯标题']}\n"
        f"资讯内容：{row['资讯内容']}\n"
        f"资讯来源：{row['资讯来源']}\n"
        f"资讯时间：{row['资讯时间']}\n"
        f"资讯链接：{row['资讯链接']}\n"
    )

    return {
        "model": ARK_MODEL,
        "input": prompt,
    }

def extract_text_from_response(payload: dict[str, Any]) -> str:
    output_text = payload.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()

    output = payload.get("output")
    if isinstance(output, list):
        texts: list[str] = []
        for item in output:
            if not isinstance(item, dict):
                continue
            content = item.get("content")
            if not isinstance(content, list):
                continue
            for part in content:
                if not isinstance(part, dict):
                    continue
                text = part.get("text")
                if isinstance(text, str) and text.strip():
                    texts.append(text.strip())
        if texts:
            return "\n".join(texts)

    choices = payload.get("choices")
    if isinstance(choices, list):
        texts = []
        for choice in choices:
            if not isinstance(choice, dict):
                continue
            message = choice.get("message")
            if isinstance(message, dict):
                content = message.get("content")
                if isinstance(content, str) and content.strip():
                    texts.append(content.strip())
            text = choice.get("text")
            if isinstance(text, str) and text.strip():
                texts.append(text.strip())
        if texts:
            return "\n".join(texts)

    return ""


def ensure_cn_sentence_end(text: str) -> str:
    cleaned = " ".join(text.split()).strip()
    if not cleaned:
        return ""
    if cleaned[-1] not in "。！？":
        cleaned += "。"
    return cleaned


def call_ark_api(row: dict[str, str]) -> dict[str, str]:
    if not ARK_API_KEY:
        raise RuntimeError("未设置 ARK_API_KEY")

    payload = build_ai_payload(row)
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = Request(
        f"{ARK_API_BASE.rstrip('/')}/responses",
        data=body,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "Authorization": f"Bearer {ARK_API_KEY}",
        },
        method="POST",
    )

    with urlopen(request, timeout=ARK_TIMEOUT_SECONDS) as response:
        response_payload = json.loads(response.read().decode("utf-8"))

    text = extract_text_from_response(response_payload)
    if not text:
        raise RuntimeError("模型返回内容为空")

    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as error:
        raise RuntimeError(f"模型返回不是有效 JSON: {error}") from error

    short_title = str(parsed.get("short_title", "") or "").strip()
    summary = str(parsed.get("summary", "") or "").strip()
    if not short_title or not summary:
        raise RuntimeError("模型返回缺少 short_title 或 summary")

    summary = ensure_cn_sentence_end(summary)

    return {"AI标题": short_title, "AI摘要": summary}

def enrich_rows_with_ai(rows: list[dict[str, str]], disable_ai: bool) -> tuple[list[dict[str, str]], int]:
    if disable_ai:
        print("已按参数跳过 AI 摘要生成。")
        return rows, 0

    if not ARK_API_KEY:
        print("未设置 ARK_API_KEY，已跳过 AI 摘要生成。")
        return rows, 0

    if not rows:
        return rows, 0

    cache = load_summary_cache()
    generated_count = 0

    for index, row in enumerate(rows, start=1):
        cache_key = make_cache_key(row)
        cached = cache.get(cache_key)
        if cached:
            row["AI标题"] = str(cached.get("AI标题", "") or "")
            row["AI摘要"] = str(cached.get("AI摘要", "") or "")
            continue

        try:
            result = call_ark_api(row)
            row.update(result)
            cache[cache_key] = result
            generated_count += 1
            print(f"[AI] 已生成 {index}/{len(rows)}: {row['企业名']} | {row['资讯标题']}")
            time.sleep(0.4)
        except (HTTPError, URLError, TimeoutError, RuntimeError, OSError) as error:
            print(f"[AI WARN] {row['企业名']} 生成失败: {error}", file=sys.stderr)

    save_summary_cache(cache)
    return rows, generated_count


def autosize_worksheet(worksheet) -> None:
    widths = {
        "A": 18,
        "B": 42,
        "C": 80,
        "D": 20,
        "E": 28,
        "F": 48,
        "G": 26,
        "H": 56,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def load_existing_report_rows(path: Path, headers: list[str]) -> list[dict[str, str]]:
    if not path.exists():
        return []

    try:
        workbook = load_workbook(path, read_only=True)
        worksheet = workbook.active
        existing_headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        rows: list[dict[str, str]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            current = {str(existing_headers[index]): str(value or "") for index, value in enumerate(values)}
            rows.append({header: current.get(header, "") for header in headers})
        return rows
    except (BadZipFile, OSError, StopIteration, KeyError, ValueError):
        return []


def merge_rows(existing_rows: list[dict[str, str]], new_rows: list[dict[str, str]], headers: list[str]) -> list[dict[str, str]]:
    merged: dict[str, dict[str, str]] = {}
    order: list[str] = []

    def key_for(row: dict[str, str]) -> str:
        return "||".join(
            [
                row.get("企业名", ""),
                row.get("资讯标题", ""),
                row.get("资讯时间", ""),
                row.get("资讯链接", ""),
            ]
        )

    for row in existing_rows + new_rows:
        key = key_for(row)
        normalized = {header: row.get(header, "") for header in headers}
        if key not in merged:
            order.append(key)
        merged[key] = normalized

    return [merged[key] for key in order]


def resolve_writable_report_path(preferred_path: Path) -> Path:
    try:
        preferred_path.parent.mkdir(parents=True, exist_ok=True)
        if not preferred_path.exists():
            return preferred_path
        with preferred_path.open("ab"):
            pass
        return preferred_path
    except PermissionError:
        timestamp = datetime.now().strftime("%H%M%S")
        return preferred_path.with_name(f"{preferred_path.stem}_{timestamp}{preferred_path.suffix}")


def write_report(rows: list[dict[str, str]]) -> Path:
    preferred_path = REPORT_DIR / "company_mentions.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "企业资讯"

    headers = ["企业名", "资讯标题", "资讯内容", "资讯来源", "资讯时间", "资讯链接", "AI标题", "AI摘要"]
    worksheet.append(headers)

    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = wrap_alignment

    existing_rows = load_existing_report_rows(preferred_path, headers)
    merged_rows = merge_rows(existing_rows, rows, headers)

    for row in merged_rows:
        worksheet.append([row.get(header, "") for header in headers])

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    worksheet.freeze_panes = "A2"
    autosize_worksheet(worksheet)
    output_path = resolve_writable_report_path(preferred_path)
    workbook.save(output_path)
    return output_path


def main() -> int:
    configure_output_encoding()
    ensure_directories()
    args = parse_args()

    try:
        try:
            companies = load_companies(args.companies.resolve())
        except Exception as error:
            print(f"运行失败: {error}", file=sys.stderr)
            return 1

        if not companies:
            print("企业名单为空，请先填写 companies.txt。")
            return 0

        if args.time_window:
            paths, (start, end) = target_data_files_by_window(args.time_window)
        else:
            paths = target_data_files(max(args.days, 1))
            start = end = None
        if not paths:
            print("未找到可扫描的数据文件。")
            return 0

        records = read_records(paths)
        if start and end:
            records = filter_records_by_time_window(records, start, end)
        rows = match_records(records, companies)
        rows, generated_count = enrich_rows_with_ai(rows, args.disable_ai)
        output_path = write_report(rows)
        international_rows = load_latest_international_rows(
            REPORT_DIR,
            datetime.now().date().isoformat(),
            args.time_window,
        )
        word_output_path = write_daily_brief(REPORT_DIR, paths, rows, international_rows, args.time_window)

        print(f"已扫描 {len(paths)} 个数据文件，共 {len(records)} 条资讯。")
        print(f"命中 {len(rows)} 条企业资讯，AI 新生成 {generated_count} 条。")
        print(f"已写入 {output_path}")
        print(f"已写入 {word_output_path}")
        return 0
    finally:
        clear_summary_cache()


if __name__ == "__main__":
    raise SystemExit(main())
