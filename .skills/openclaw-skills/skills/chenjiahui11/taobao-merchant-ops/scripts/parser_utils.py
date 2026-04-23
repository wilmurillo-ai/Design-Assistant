#!/usr/bin/env python3
from __future__ import annotations

import ast
import operator
from pathlib import Path
from typing import Any

import openpyxl

from app_errors import ParseError


_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
}


def decode_header(cell: Any) -> str:
    if cell is None:
        return ""
    try:
        return cell.encode("latin1").decode("gbk")
    except Exception:
        return str(cell)


def safe_number(value: Any) -> float | str | None:
    if value is None or str(value).strip() == "" or str(value).upper() == "NULL":
        return None
    s = str(value).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return str(value)


def build_header_map(worksheet: Any) -> dict[str, int]:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ParseError("Excel header row is empty")
    return {
        decode_header(cell): idx
        for idx, cell in enumerate(header_row)
        if cell is not None and decode_header(cell)
    }


def load_first_data_row(worksheet: Any) -> tuple[Any, ...]:
    row = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
    if row is None:
        raise ParseError("Excel data row is empty")
    return row


def load_workbook_data(excel_path: str | Path) -> tuple[dict[str, int], tuple[Any, ...]]:
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.active
    return build_header_map(worksheet), load_first_data_row(worksheet)


def _eval_node(node: ast.AST, values: dict[str, float]) -> float:
    if isinstance(node, ast.Expression):
        return _eval_node(node.body, values)
    if isinstance(node, ast.Name):
        if node.id not in values:
            raise ParseError(f"Missing dependency: {node.id}")
        return values[node.id]
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return float(node.value)
    if isinstance(node, ast.Num):
        return float(node.n)
    if isinstance(node, ast.BinOp) and type(node.op) in _OPS:
        left = _eval_node(node.left, values)
        right = _eval_node(node.right, values)
        if isinstance(node.op, ast.Div) and right == 0:
            raise ZeroDivisionError
        return _OPS[type(node.op)](left, right)
    if isinstance(node, ast.UnaryOp) and type(node.op) in _OPS:
        return _OPS[type(node.op)](_eval_node(node.operand, values))
    raise ParseError(f"Unsupported expression node: {type(node).__name__}")


def compute_expression(expression: str, values: dict[str, Any]) -> float | None:
    numeric_values: dict[str, float] = {}
    for key, value in values.items():
        converted = safe_number(value)
        if converted is None:
            return None
        if isinstance(converted, str):
            raise ParseError(f"Non-numeric dependency for {key}: {converted}")
        numeric_values[key] = float(converted)
    try:
        tree = ast.parse(expression, mode="eval")
        result = _eval_node(tree, numeric_values)
    except ZeroDivisionError:
        return None
    except ParseError:
        raise
    except Exception as exc:
        raise ParseError(f"Expression evaluation failed: {expression} ({exc})") from exc
    if result != result or abs(result) == float("inf"):
        return None
    return round(result, 4)
