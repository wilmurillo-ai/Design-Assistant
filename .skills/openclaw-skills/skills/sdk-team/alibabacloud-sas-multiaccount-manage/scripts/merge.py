#!/usr/bin/env python3
"""merge.py — 公共 Excel 表格合并工具

用法（作为模块导入）:
    from merge import merge_excel
    merge_excel(
        export_name="output.xlsx",
        inputs=[
            {"filename": "a.xlsx", "account_id": "123456"},
            {"filename": "b.xlsx", "account_id": "789012"},
        ],
    )

用法（命令行）:
    uv run merge.py --output merged.xlsx --input a.xlsx:123 b.xlsx:456
"""

import argparse
import sys
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style",
    category=UserWarning,
)


def merge_excel(export_name, inputs):
    """将多个 Excel 文件合并为一个，并在末尾追加「资源管理账号」列。

    Args:
        export_name: 输出文件名（.xlsx）
        inputs: 列表，每项包含 filename(str) 和 account_id(str)

    Returns:
        输出文件的路径字符串
    """
    if not inputs:
        raise ValueError("inputs 不能为空")

    all_rows = []
    header = None
    valid_count = 0

    for item in inputs:
        filename = item["filename"]
        account_id = str(item["account_id"])
        if not Path(filename).exists():
            print("警告: 文件 {} 不存在，跳过".format(filename), file=sys.stderr)
            continue

        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb["Sheet0"] if "Sheet0" in wb.sheetnames else wb.active
        rows = list(ws.values)
        wb.close()

        if not rows:
            continue

        if header is None:
            header = list(rows[0]) + ["资源管理账号"]
            all_rows.append(header)

        for row in rows[1:]:
            all_rows.append(list(row) + [account_id])

        valid_count += 1

    if not all_rows or valid_count == 0:
        raise RuntimeError("没有可合并的有效文件")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Sheet0"
    for row in all_rows:
        out_ws.append(row)
    out_wb.save(export_name)

    data_rows = len(all_rows) - 1
    print("合并完成: {}（共 {} 行，{} 个账号）".format(export_name, data_rows, valid_count))
    return export_name


def main():
    parser = argparse.ArgumentParser(description="多账号 Excel 表格合并工具")
    parser.add_argument("--output", "-o", required=True, help="输出文件名")
    parser.add_argument(
        "--input",
        "-i",
        nargs="+",
        required=True,
        metavar="FILE:ACCOUNT_ID",
        help="输入文件，格式: 文件路径:账号ID，可指定多个",
    )
    args = parser.parse_args()

    inputs = []
    for item in args.input:
        parts = item.rsplit(":", 1)
        if len(parts) != 2:
            print(
                "错误: 输入格式不正确 '{}'，应为 '文件路径:账号ID'".format(item),
                file=sys.stderr,
            )
            sys.exit(1)
        inputs.append({"filename": parts[0], "account_id": parts[1]})

    merge_excel(args.output, inputs)


if __name__ == "__main__":
    main()
