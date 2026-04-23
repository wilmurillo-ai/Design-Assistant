#!/usr/bin/env python3
"""
深度分析版Markdown转Excel工具 - md2excel_deep.py

专为深度分析版测试用例格式设计的转换工具

支持解析：
- ### TC001: 标题
- **测试方法**: 正向测试
- **优先级**: 高
- **前置条件**:
```
1. 系统已部署
```
- **测试步骤**:
```
1. 进入模块
```
- **预期结果**:
```
1. 功能正常
```
"""

import sys
import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DeepMarkdownToExcelConverter:
    """深度分析版Markdown到Excel转换器"""

    def __init__(self, md_path):
        """
        初始化转换器

        Args:
            md_path: Markdown文件路径
        """
        self.md_path = Path(md_path)
        self.content = ""
        self.test_cases = []

    def load_markdown(self):
        """加载Markdown文件"""
        try:
            with open(self.md_path, 'r', encoding='utf-8') as f:
                self.content = f.read()
            logger.info(f"成功加载Markdown文件: {self.md_path}")
            return True
        except Exception as e:
            logger.error(f"Markdown文件加载失败: {e}")
            return False

    def extract_test_cases(self):
        """
        从深度分析版Markdown中提取测试用例

        Returns:
            List[Dict]: 测试用例列表
        """
        test_cases = []
        lines = self.content.split('\n')
        current_case = None
        in_code_block = False
        code_block_type = None
        code_content = []

        for i, line in enumerate(lines):
            line = line.rstrip()

            # 检查代码块开始/结束
            if line.startswith('```'):
                if in_code_block:
                    # 结束代码块
                    if code_block_type == "preconditions":
                        current_case["preconditions"] = '\n'.join(code_content)
                    elif code_block_type == "steps":
                        current_case["steps"] = '\n'.join(code_content)
                    elif code_block_type == "expected_result":
                        current_case["expected_result"] = '\n'.join(code_content)
                    in_code_block = False
                    code_block_type = None
                    code_content = []
                else:
                    # 开始代码块
                    in_code_block = True
                continue

            # 在代码块内时，收集内容
            if in_code_block:
                code_content.append(line)
                continue

            # 检查是否是新的测试用例标题
            case_match = re.match(r'^### TC(\d+): (.+)$', line)
            if case_match:
                # 如果有上一个用例，保存它
                if current_case:
                    test_cases.append(current_case)

                # 开始新的用例
                case_id = f"TC{case_match.group(1)}"
                title = case_match.group(2)
                current_case = {
                    "id": case_id,
                    "title": title,
                    "module": "",
                    "preconditions": "",
                    "steps": "",
                    "priority": "",
                    "expected_result": "",
                    "method": ""
                }
                continue

            # 检查字段
            if current_case:
                # 测试方法
                method_match = re.match(r'^\*\*测试方法\*\*: (.+)$', line)
                if method_match:
                    current_case["method"] = method_match.group(1)
                    continue

                # 优先级
                priority_match = re.match(r'^\*\*优先级\*\*: (.+)$', line)
                if priority_match:
                    current_case["priority"] = priority_match.group(1)
                    continue

                # 前置条件
                if line.startswith('**前置条件**:'):
                    code_block_type = "preconditions"
                    continue

                # 测试步骤
                if line.startswith('**测试步骤**:'):
                    code_block_type = "steps"
                    continue

                # 预期结果
                if line.startswith('**预期结果**:'):
                    code_block_type = "expected_result"
                    continue

                # 模块
                module_match = re.match(r'^## (.+)$', line)
                if module_match:
                    current_case["module"] = module_match.group(1)
                    continue

        # 添加最后一个用例
        if current_case:
            test_cases.append(current_case)

        logger.info(f"提取到 {len(test_cases)} 个测试用例")
        return test_cases

    def convert_to_excel(self, output_path=None):
        """
        转换为Excel文件

        Args:
            output_path: 输出文件路径，默认为Markdown同级目录

        Returns:
            str: 输出文件路径
        """
        if output_path is None:
            output_path = self.md_path.parent / f"{self.md_path.stem}.xlsx"

        # 准备数据
        headers = ['用例编号', '用例模块', '用例标题', '前置条件', '测试步骤', '优先级', '预期结果']
        data = [headers]

        for tc in self.test_cases:
            row = [
                tc.get('id', ''),
                tc.get('module', ''),
                tc.get('title', ''),
                tc.get('preconditions', ''),
                tc.get('steps', ''),
                tc.get('priority', ''),
                tc.get('expected_result', '')
            ]
            data.append(row)

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active

        # 写入数据
        for row_idx, row in enumerate(data, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 设置表头样式
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 应用表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        column_widths = {'A': 12, 'B': 12, 'C': 35, 'D': 30, 'E': 50, 'F': 10, 'G': 45}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 设置数据行样式
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = thin_border

        # 设置行高
        ws.row_dimensions[1].height = 30
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].height = 60

        # 保存
        wb.save(output_path)
        logger.info(f"Excel文件已生成: {output_path}")
        print(f"Excel文件已生成: {output_path}")
        print(f"总计测试用例: {len(self.test_cases)} 条")

        return str(output_path)


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(
        description="深度分析版Markdown转Excel工具 - 支持复杂测试用例格式",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python md2excel_deep.py 深度分析_完整测试用例集.md
  python md2excel_deep.py 深度分析_完整测试用例集.md --output 测试用例.xlsx
        """
    )

    parser.add_argument("md_file", help="Markdown文件路径")
    parser.add_argument("--output", help="输出Excel文件路径（可选）")

    args = parser.parse_args()

    # 创建转换器
    converter = DeepMarkdownToExcelConverter(args.md_file)

    # 加载并转换
    if converter.load_markdown():
        converter.test_cases = converter.extract_test_cases()
        excel_path = converter.convert_to_excel(args.output)
        print(f"\n转换完成！")
        print(f"Excel文件: {excel_path}")
    else:
        print("\n转换失败！")
        sys.exit(1)


if __name__ == "__main__":
    main()