#!/usr/bin/env python3
"""
Markdown转Excel工具 - md2excel.py

将测试用例的Markdown文件转换为Excel格式的测试用例表格
"""

import sys
import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any


class MarkdownToExcelConverter:
    """Markdown到Excel转换器"""

    def __init__(self, md_path):
        """
        初始化转换器

        Args:
            md_path: Markdown文件路径
        """
        self.md_path = Path(md_path)
        self.content = ""
        self.test_cases = []

    def load_markdown(self):
        """加载Markdown文件"""
        try:
            with open(self.md_path, 'r', encoding='utf-8') as f:
                self.content = f.read()
            print(f"成功加载Markdown文件: {self.md_path}")
            return True
        except Exception as e:
            print(f"Markdown文件加载失败: {e}")
            return False

    def extract_test_cases(self):
        """
        从Markdown中提取测试用例

        Returns:
            List[Dict]: 测试用例列表
        """
        test_cases = []

        # 分割文档为段落
        lines = self.content.split('\n')
        current_case = None
        in_code_block = False

        for line in lines:
            line = line.rstrip()

            # 检查代码块开始/结束
            if line.startswith('```'):
                in_code_block = not in_code_block
                continue

            # 跳过代码块中的内容
            if in_code_block:
                continue

            # 检查是否是新的测试用例标题
            case_match = re.match(r'^### (TC\d+): (.+)$', line)
            if case_match:
                # 如果有上一个用例，保存它
                if current_case:
                    test_cases.append(current_case)

                # 开始新的用例
                case_id = case_match.group(1)
                title = case_match.group(2)
                current_case = {
                    "id": case_id,
                    "title": title,
                    "module": "",
                    "preconditions": "",
                    "steps": "",
                    "priority": "中",
                    "expected_result": "",
                    "method": ""
                }
                continue

            # 检查字段
            if current_case:
                # 提取模块（从二级标题获取）
                module_match = re.match(r'^## (.+)$', line)
                if module_match and current_case["module"] == "":
                    current_case["module"] = module_match.group(1)

                # 检查优先级
                priority_match = re.match(r'^\*\*优先级\*\*: (.+)$', line)
                if priority_match:
                    current_case["priority"] = priority_match.group(1)
                    continue

                # 检查测试方法
                method_match = re.match(r'^\*\*测试方法\*\*: (.+)$', line)
                if method_match:
                    current_case["method"] = method_match.group(1)
                    continue

                # 检查前置条件
                if line.startswith('**前置条件**:'):  # 从行首开始
                    preconditions = line[len('**前置条件**:'):]  # 去掉前缀
                    # 检查是否是代码块
                    if line.endswith('```'):
                        # 下面的行是代码块内容
                        current_case["preconditions"] = ""
                        continue
                    else:
                        current_case["preconditions"] = preconditions.strip()
                    continue

                # 检查测试步骤
                if line.startswith('**测试步骤**:'):  # 从行首开始
                    steps = line[len('**测试步骤**:'):]  # 去掉前缀
                    current_case["steps"] = steps.strip()
                    continue

                # 检查预期结果
                if line.startswith('**预期结果**:'):  # 从行首开始
                    expected = line[len('**预期结果**:'):]  # 去掉前缀
                    current_case["expected_result"] = expected.strip()
                    continue

                # 检查代码块中的内容（如果前面有字段标记）
                if line.startswith('```') and current_case:
                    # 这是上一个字段的代码块内容
                    continue

                # 如果行不是空行且不在代码块中，尝试追加到当前字段
                if line.strip() and not line.startswith('### ') and not line.startswith('## ') and not line.startswith('---'):
                    # 追加到当前字段（如果有）
                    if current_case:
                        if not current_case["preconditions"] and line.startswith('1.'):
                            current_case["preconditions"] = line
                        elif not current_case["steps"] and line.startswith('1.'):
                            current_case["steps"] = line
                        elif not current_case["expected_result"] and line.startswith('1.'):
                            current_case["expected_result"] = line
                        elif current_case["preconditions"] and not current_case["steps"]:
                            # 追加到前置条件
                            current_case["preconditions"] += "\n" + line
                        elif current_case["steps"] and not current_case["expected_result"]:
                            # 追加到测试步骤
                            current_case["steps"] += "\n" + line
                        elif current_case["expected_result"]:
                            # 追加到预期结果
                            current_case["expected_result"] += "\n" + line

        # 添加最后一个用例
        if current_case:
            test_cases.append(current_case)

        print(f"提取到 {len(test_cases)} 个测试用例")
        return test_cases

    def convert_to_excel(self, output_path=None):
        """
        转换为Excel文件

        Args:
            output_path: 输出文件路径，默认为Markdown同级目录

        Returns:
            str: 输出文件路径
        """
        if output_path is None:
            output_path = self.md_path.parent / f"{self.md_path.stem}.xlsx"

        # 准备数据
        headers = ['用例编号', '用例模块', '用例标题', '前置条件', '测试步骤', '优先级', '预期结果']
        data = [headers]

        for tc in self.test_cases:
            row = [
                tc.get('id', ''),
                tc.get('module', ''),
                tc.get('title', ''),
                tc.get('preconditions', ''),
                tc.get('steps', ''),
                tc.get('priority', '中'),
                tc.get('expected_result', '')
            ]
            data.append(row)

        # 创建DataFrame
        df = pd.DataFrame(data[1:], columns=data[0])

        # 保存为Excel
        df.to_excel(output_path, index=False, engine='openpyxl')

        # 加载Excel进行格式化
        wb = Workbook()
        ws = wb.active

        # 写入数据
        for row_idx, row in enumerate(data, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 设置表头样式
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 应用表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        column_widths = {'A': 12, 'B': 12, 'C': 35, 'D': 30, 'E': 50, 'F': 10, 'G': 45}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 设置数据行样式
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = thin_border

        # 设置行高
        ws.row_dimensions[1].height = 30
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].height = 60

        # 保存
        wb.save(output_path)
        print(f"Excel文件已生成: {output_path}")
        print(f"总计测试用例: {len(self.test_cases)} 条")

        return str(output_path)


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(
        description="Markdown转Excel工具 - 将测试用例Markdown文件转换为Excel格式",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python md2excel.py testcases.md
  python md2excel.py testcases.md --output testcases.xlsx
        """
    )

    parser.add_argument("md_file", help="Markdown文件路径")
    parser.add_argument("--output", help="输出Excel文件路径（可选）")

    args = parser.parse_args()

    # 创建转换器
    converter = MarkdownToExcelConverter(args.md_file)

    # 加载并转换
    if converter.load_markdown():
        converter.test_cases = converter.extract_test_cases()
        excel_path = converter.convert_to_excel(args.output)
        print(f"\n转换完成！")
        print(f"Excel文件: {excel_path}")
    else:
        print("\n转换失败！")
        sys.exit(1)


if __name__ == "__main__":
    main()
