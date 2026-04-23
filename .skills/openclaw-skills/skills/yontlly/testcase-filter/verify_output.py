#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

# 验证输出文件
wb = openpyxl.load_workbook(r"C:\Users\yanghua1\.claude\skills\excel-automation\tmp\PC功能点用例_最终版.xlsx")
sheet = wb['登录']

print('输出文件 - 登录sheet的前10列表头:')
for col in range(1, min(11, sheet.max_column + 1)):
    val = sheet.cell(row=1, column=col).value
    if val:
        print(f'  第{col}列: {val}')
    else:
        print(f'  第{col}列: (空)')

print(f'\n总列数: {sheet.max_column}')

# 检查第2列是否已删除
if sheet.max_column == 15:  # 原来是16列，删除1列后应该是15列
    print('✓ 项目/产品列已成功删除')
else:
    print('✗ 项目/产品列删除失败')

wb.close()