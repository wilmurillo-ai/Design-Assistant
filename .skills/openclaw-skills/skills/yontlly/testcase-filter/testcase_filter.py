#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试用例筛选核心模块
本模块提供了筛选P0/P1测试用例、拆分合并单元格、重新编号等完整功能
"""

import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy


def find_priority_column(sheet):
    """查找优先级列"""
    for row in range(1, min(11, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and ('优先级' in str(cell_value) or 'Priority' in str(cell_value)):
                return col
    return None


def find_number_column(sheet):
    """查找编号列"""
    for row in range(1, min(11, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and ('编号' in str(cell_value) or 'ID' in str(cell_value) or
                           'No' in str(cell_value) or '序号' in str(cell_value)):
                return col
    return None


def find_product_project_column(sheet):
    """查找项目/产品相关列"""
    target_keywords = ['项目/产品', '产品/项目', '产品', '项目']
    for row in range(1, min(11, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value).strip()
                # 精确匹配关键词
                if cell_str in target_keywords:
                    return col
                # 也检查是否包含这些关键词（处理有前后空格或其他字符的情况）
                for keyword in target_keywords:
                    if keyword in cell_str:
                        return col
    return None


def delete_column(sheet, col_to_delete):
    """删除指定列并调整后续列"""
    # 使用openpyxl的delete_cols方法删除列
    sheet.delete_cols(col_to_delete)
    return True


def unmerge_and_fill_cells(sheet):
    """拆分合并单元格并填充内容"""
    if not sheet.merged_cells:
        return

    merged_ranges = list(sheet.merged_cells.ranges)

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
        top_left_value = sheet.cell(row=min_row, column=min_col).value

        sheet.unmerge_cells(str(merged_range))

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col, value=top_left_value)


def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def analyze_excel(file_path):
    """分析Excel文件结构"""
    print(f"正在分析文件: {file_path}")
    wb = openpyxl.load_workbook(file_path)

    print(f"\n文件包含 {len(wb.sheetnames)} 个sheet:")
    total_p0 = 0
    total_p1 = 0

    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        sheet = wb[sheet_name]
        print(f"\n{idx}. Sheet: {sheet_name}")
        print(f"   行数: {sheet.max_row}, 列数: {sheet.max_column}")

        priority_col = find_priority_column(sheet)
        if priority_col:
            print(f"   优先级列: 第{priority_col}列 ({get_column_letter(priority_col)})")
            p0_count, p1_count, empty_count = count_priority_levels(sheet, priority_col)
            print(f"   P0用例: {p0_count}, P1用例: {p1_count}, 空优先级: {empty_count}")
            total_p0 += p0_count
            total_p1 += p1_count
        else:
            print("   未找到优先级列")

        if sheet.merged_cells:
            merged_count = len(list(sheet.merged_cells.ranges))
            print(f"   合并单元格数量: {merged_count}")

    wb.close()

    print(f"\n总计:")
    print(f"  P0用例总数: {total_p0}")
    print(f"  P1用例总数: {total_p1}")
    print(f"  预计处理后用例数: {total_p0 + total_p1}")


def count_priority_levels(sheet, priority_col):
    """统计各优先级的数量"""
    p0_count = 0
    p1_count = 0
    empty_count = 0

    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row, column=priority_col).value
        if value is None or str(value).strip() == '':
            empty_count += 1
        elif str(value).strip() == 'P0':
            p0_count += 1
        elif str(value).strip() == 'P1':
            p1_count += 1

    return p0_count, p1_count, empty_count


def process_excel(input_file, output_file):
    """处理Excel文件 - 主函数"""
    print(f"\n开始处理文件: {input_file}")

    wb = openpyxl.load_workbook(input_file)
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)

    sheets_processed = 0
    sheets_skipped = 0

    for sheet_name in wb.sheetnames:
        print(f"\n处理 Sheet: {sheet_name}")
        original_sheet = wb[sheet_name]

        # 拆分合并单元格
        if original_sheet.merged_cells:
            merged_count = len(list(original_sheet.merged_cells.ranges))
            print(f"  拆分 {merged_count} 个合并单元格")
            unmerge_and_fill_cells(original_sheet)

        # 查找并删除"项目/产品"相关列
        product_project_col = find_product_project_column(original_sheet)
        if product_project_col:
            print(f"  删除列: 第{product_project_col}列 ({get_column_letter(product_project_col)}) - 项目/产品相关列")
            delete_column(original_sheet, product_project_col)
            print(f"  已删除项目/产品列")

        # 查找优先级列（在删除产品列之后）
        priority_col = find_priority_column(original_sheet)

        if not priority_col:
            print(f"  未找到优先级列，跳过此sheet")
            sheets_skipped += 1
            continue

        # 检查是否有P0或P1用例
        p0_count, p1_count, empty_count = count_priority_levels(original_sheet, priority_col)

        if p0_count == 0 and p1_count == 0:
            print(f"  没有P0或P1用例，跳过此sheet")
            sheets_skipped += 1
            continue

        print(f"  找到 {p0_count} 个P0用例, {p1_count} 个P1用例")

        # 创建新sheet
        new_sheet = new_wb.create_sheet(title=sheet_name)

        # 复制表头
        print(f"  复制表头...")
        for col in range(1, original_sheet.max_column + 1):
            source_cell = original_sheet.cell(row=1, column=col)
            target_cell = new_sheet.cell(row=1, column=col)
            target_cell.value = source_cell.value
            copy_cell_style(source_cell, target_cell)

        # 复制列宽
        for col in range(1, original_sheet.max_column + 1):
            col_letter = get_column_letter(col)
            if original_sheet.column_dimensions[col_letter].width:
                new_sheet.column_dimensions[col_letter].width = original_sheet.column_dimensions[col_letter].width

        # 复制行高
        new_sheet.row_dimensions[1].height = original_sheet.row_dimensions[1].height

        # 查找编号列（在删除产品列之后）
        number_col = find_number_column(original_sheet)
        if number_col:
            print(f"  编号列: 第{number_col}列 ({get_column_letter(number_col)})")

        # 复制P0和P1用例
        print(f"  复制P0和P1用例...")
        new_row = 2
        tc_counter = 1

        for row in range(2, original_sheet.max_row + 1):
            priority_value = original_sheet.cell(row=row, column=priority_col).value

            if priority_value and str(priority_value).strip() in ['P0', 'P1']:
                # 复制整行
                for col in range(1, original_sheet.max_column + 1):
                    source_cell = original_sheet.cell(row=row, column=col)
                    target_cell = new_sheet.cell(row=new_row, column=col)
                    target_cell.value = source_cell.value
                    copy_cell_style(source_cell, target_cell)

                # 重新编号
                if number_col:
                    new_number = f"TC{tc_counter:03d}"
                    new_sheet.cell(row=new_row, column=number_col, value=new_number)
                    tc_counter += 1

                # 复制行高
                if original_sheet.row_dimensions[row].height:
                    new_sheet.row_dimensions[new_row].height = original_sheet.row_dimensions[row].height

                new_row += 1

        print(f"  完成！共复制 {new_row - 2} 行用例")
        if number_col:
            print(f"  已重新编号 {tc_counter - 1} 个用例（TC001-TC{tc_counter - 1:03d}）")
        sheets_processed += 1

    # 保存新文件
    print(f"\n保存到: {output_file}")
    new_wb.save(output_file)

    wb.close()
    new_wb.close()

    print(f"\n处理完成！")
    print(f"处理的sheet数: {sheets_processed}")
    print(f"跳过的sheet数: {sheets_skipped}")

    return {
        'total_sheets': sheets_processed + sheets_skipped,
        'processed': sheets_processed,
        'skipped': sheets_skipped
    }


if __name__ == '__main__':
    import sys

    if len(sys.argv) >= 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2]

        # 先分析
        analyze_excel(input_file)

        # 再处理
        result = process_excel(input_file, output_file)

        print(f"\n✅ 处理完成！")
        print(f"输出文件: {output_file}")
    else:
        print("用法: python testcase_filter.py <输入文件> <输出文件>")
        print("示例: python testcase_filter.py input.xlsx output.xlsx")