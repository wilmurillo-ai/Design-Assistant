#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\yanghua1\.claude\skills\excel-automation\tmp\PC功能点用例.xlsx")
sheet = wb['登录']

print('表头内容:')
for col in range(1, sheet.max_column + 1):
    val = sheet.cell(row=1, column=col).value
    if val:
        print(f'  第{col}列: {val}')

wb.close()