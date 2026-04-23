from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Iterable
import xml.etree.ElementTree as ET

import requests
from openpyxl import Workbook

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
DEFAULT_PROJECT_ROOT = Path.cwd()
DEFAULT_ANALYSIS_DIR = DEFAULT_PROJECT_ROOT / 'analysis'
DEFAULT_SOQL_DIR = DEFAULT_ANALYSIS_DIR / 'soql'
DEFAULT_EXPORT_DIR = DEFAULT_ANALYSIS_DIR / 'excel_exports'

ANALYSIS_DIR = DEFAULT_ANALYSIS_DIR
SOQL_DIR = DEFAULT_SOQL_DIR
EXPORT_DIR = DEFAULT_EXPORT_DIR
NS = {
    'soapenv': 'http://schemas.xmlsoap.org/soap/envelope/',
    'urn': 'urn:enterprise.soap.sforce.com',
    'sf': 'urn:sobject.enterprise.soap.sforce.com',
}

OBJECTS = [
    'Lead',
    'Account',
    'Contact',
    'Opportunity',
    'Quote',
    'CRM_Contract__c',
    'CRM_Bidding_Project_Establishment__c',
    'CRM_Bidding_Review__c',
    'CRM_Order_Details__c',
    'CRM_Business_Trip_Application__c',
    'CRM_OrderList__c',
    'CRM_Technical_Review__c',
    'CRM_PaybackList__c',
    'CRM_InvoiceList__c',
]

EXTRA_LABELS = {
    'Id': '记录ID',
    'Name': '名称',
    'LastModifiedDate': '最后修改时间',
    'OwnerId': '所有人ID',
    'Owner.Name': '所有人名称',
    'Owner.CRM_Sales_Team__c': '所有人销售团队',
    'Owner.CRM_Sales_Group__c': '所有人销售组',
    'CRM_Opportunity__r.Name': '业务机会名称',
    'CRM_Opportunity__r.Owner.Name': '业务机会所有人名称',
    'CRM_Opportunity__r.Owner.CRM_Sales_Team__c': '业务机会所有人销售团队',
    'CRM_Opportunity__r.Owner.CRM_Sales_Group__c': '业务机会所有人销售组',
}


def local_name(tag: str) -> str:
    return tag.split('}', 1)[-1]


def load_labels(object_name: str) -> dict[str, str]:
    labels = {}
    describe_path = ANALYSIS_DIR / 'layout_extract' / f'{object_name}.describe.json'
    payload = json.loads(describe_path.read_text(encoding='utf-8'))
    for field in payload['result']['fields']:
        labels[field['name']] = field.get('label') or field['name']
    labels.update(EXTRA_LABELS)
    return labels


def parse_select_columns(soql_text: str) -> list[str]:
    columns: list[str] = []
    in_select = False
    in_typeof = False
    typeof_rel = ''
    for raw_line in soql_text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith('--'):
            continue
        if line == 'SELECT':
            in_select = True
            continue
        if not in_select:
            continue
        if line.startswith('FROM '):
            break
        if in_typeof:
            if line.startswith('WHEN '):
                _, _, field_part = line.partition(' THEN ')
                for field_name in [part.strip().rstrip(',') for part in field_part.split(',') if part.strip()]:
                    col = f'{typeof_rel}.{field_name}'
                    if col not in columns:
                        columns.append(col)
                continue
            if line.startswith('END'):
                in_typeof = False
                typeof_rel = ''
                continue
        if line.startswith('TYPEOF '):
            parts = line.split()
            typeof_rel = parts[1]
            in_typeof = True
            continue
        columns.append(line.rstrip(','))
    return columns


def sobject_to_dict(elem: ET.Element) -> dict:
    result: dict[str, object] = {}
    for child in elem:
        name = local_name(child.tag)
        if name == 'type':
            continue
        if list(child):
            result[name] = sobject_to_dict(child)
        else:
            result[name] = child.text or ''
    return result


def flatten_record(record: dict, prefix: str = '') -> dict[str, str]:
    flattened: dict[str, str] = {}
    for key, value in record.items():
        dotted = f'{prefix}.{key}' if prefix else key
        if isinstance(value, dict):
            flattened.update(flatten_record(value, dotted))
        else:
            flattened[dotted] = value
    return flattened


class SoapClient:
    def __init__(self, instance_url: str, access_token: str, api_version: str):
        self.endpoint = f"{instance_url}/services/Soap/c/{api_version}"
        self.session = requests.Session()
        self.headers = {
            'Content-Type': 'text/xml; charset=UTF-8',
            'SOAPAction': '""',
        }
        self.access_token = access_token

    def _post(self, body: str) -> ET.Element:
        resp = self.session.post(self.endpoint, data=body.encode('utf-8'), headers=self.headers, timeout=120)
        resp.raise_for_status()
        root = ET.fromstring(resp.text)
        fault = root.find('.//soapenv:Fault', NS)
        if fault is not None:
            code = fault.findtext('faultcode') or 'SOAP_FAULT'
            text = fault.findtext('faultstring') or 'Unknown SOAP fault'
            raise RuntimeError(f'{code}: {text}')
        return root

    def query(self, soql: str) -> tuple[list[dict], bool, str | None]:
        body = f'''<?xml version="1.0" encoding="utf-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:urn="urn:enterprise.soap.sforce.com">
  <soapenv:Header>
    <urn:SessionHeader>
      <urn:sessionId>{self.access_token}</urn:sessionId>
    </urn:SessionHeader>
  </soapenv:Header>
  <soapenv:Body>
    <urn:query>
      <urn:queryString><![CDATA[{soql}]]></urn:queryString>
    </urn:query>
  </soapenv:Body>
</soapenv:Envelope>'''
        root = self._post(body)
        return self._parse_query_result(root)

    def query_more(self, locator: str) -> tuple[list[dict], bool, str | None]:
        body = f'''<?xml version="1.0" encoding="utf-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:urn="urn:enterprise.soap.sforce.com">
  <soapenv:Header>
    <urn:SessionHeader>
      <urn:sessionId>{self.access_token}</urn:sessionId>
    </urn:SessionHeader>
  </soapenv:Header>
  <soapenv:Body>
    <urn:queryMore>
      <urn:queryLocator>{locator}</urn:queryLocator>
    </urn:queryMore>
  </soapenv:Body>
</soapenv:Envelope>'''
        root = self._post(body)
        return self._parse_query_result(root)

    def _parse_query_result(self, root: ET.Element) -> tuple[list[dict], bool, str | None]:
        result = root.find('.//urn:result', NS)
        if result is None:
            raise RuntimeError('SOAP response missing result node')
        records = [sobject_to_dict(rec) for rec in result.findall('urn:records', NS)]
        done = (result.findtext('urn:done', default='true', namespaces=NS) or 'true').lower() == 'true'
        locator = result.findtext('urn:queryLocator', default=None, namespaces=NS)
        return records, done, locator


def iter_query_records(client: SoapClient, soql: str) -> Iterable[dict]:
    records, done, locator = client.query(soql)
    for rec in records:
        yield rec
    while not done and locator:
        records, done, locator = client.query_more(locator)
        for rec in records:
            yield rec


def export_object(client: SoapClient, object_name: str) -> tuple[Path, int, list[str]]:
    soql_path = SOQL_DIR / f'{object_name}.soql'
    soql_text = soql_path.read_text(encoding='utf-8')
    columns = parse_select_columns(soql_text)
    labels = load_labels(object_name)
    output_path = EXPORT_DIR / f'{object_name}.xlsx'
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=object_name[:31])
    ws.append([labels.get(col, col) for col in columns])
    count = 0
    for record in iter_query_records(client, soql_text):
        flat = flatten_record(record)
        ws.append([flat.get(col, '') for col in columns])
        count += 1
    wb.save(output_path)
    return output_path, count, columns


def export_order_details(client: SoapClient, chunk_size: int) -> tuple[Path, int, list[str]]:
    order_soql = (SOQL_DIR / 'CRM_OrderList__c.soql').read_text(encoding='utf-8')
    order_ids: list[str] = []
    for record in iter_query_records(client, order_soql):
        rec_id = record.get('Id')
        if rec_id:
            order_ids.append(rec_id)
    columns = parse_select_columns((SOQL_DIR / 'CRM_Order_Details__c.soql').read_text(encoding='utf-8'))
    labels = load_labels('CRM_Order_Details__c')
    output_path = EXPORT_DIR / 'CRM_Order_Details__c.xlsx'
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title='CRM_Order_Details__c'[:31])
    ws.append([labels.get(col, col) for col in columns])
    count = 0
    for i in range(0, len(order_ids), chunk_size):
        chunk = order_ids[i:i + chunk_size]
        ids = ', '.join(f"'{v}'" for v in chunk)
        rendered_columns = ',\n  '.join(columns)
        soql = (
            "SELECT\n"
            f"  {rendered_columns}\n"
            "FROM CRM_Order_Details__c\n"
            "WHERE LastModifiedDate < 2026-03-22T16:00:00Z\n"
            f"AND CRM_Order__c IN ({ids})"
        )
        for record in iter_query_records(client, soql):
            flat = flatten_record(record)
            ws.append([flat.get(col, '') for col in columns])
            count += 1
    wb.save(output_path)
    return output_path, count, columns


def write_manifest(rows: list[dict[str, object]]) -> None:
    manifest_path = EXPORT_DIR / 'export_manifest.csv'
    with manifest_path.open('w', newline='', encoding='utf-8-sig') as fh:
        writer = csv.DictWriter(fh, fieldnames=['object', 'xlsx_file', 'row_count', 'column_count'])
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('--auth-json', required=True)
    parser.add_argument('--objects', nargs='*', default=OBJECTS)
    parser.add_argument('--analysis-dir', type=Path, default=DEFAULT_ANALYSIS_DIR)
    parser.add_argument('--soql-dir', type=Path, default=DEFAULT_SOQL_DIR)
    parser.add_argument('--export-dir', type=Path, default=DEFAULT_EXPORT_DIR)
    parser.add_argument('--detail-chunk-size', type=int, default=200)
    args = parser.parse_args()

    global ANALYSIS_DIR, SOQL_DIR, EXPORT_DIR
    ANALYSIS_DIR = args.analysis_dir.resolve()
    SOQL_DIR = args.soql_dir.resolve()
    EXPORT_DIR = args.export_dir.resolve()

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    auth = json.loads(Path(args.auth_json).read_text(encoding='utf-8'))['result']
    client = SoapClient(auth['instanceUrl'], auth['accessToken'], auth['apiVersion'])

    manifest_rows = []
    for object_name in args.objects:
        if object_name == 'CRM_Order_Details__c':
            output_path, row_count, columns = export_order_details(client, args.detail_chunk_size)
        else:
            output_path, row_count, columns = export_object(client, object_name)
        manifest_rows.append({
            'object': object_name,
            'xlsx_file': str(output_path),
            'row_count': row_count,
            'column_count': len(columns),
        })
        print(f'{object_name}: {row_count} rows -> {output_path}')
    write_manifest(manifest_rows)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
