#!/usr/bin/env python3
"""
起点经典网文推荐（万订/十万均订）
================================
从起点图(qidiantu.com)获取万订+十万均订徽章作品列表，
爬取详情后缓存到本地，支持随机推荐经典好书。

核心特性：
  - 预置数据：首次使用零等待，内置134本经典书库
  - 增量更新：每周仅抓列表页对比新增，秒级完成
  - 双层数据源：十万均订（34本顶级）+ 万订近百（最新100本）
  - IP 衍生品信息：电视剧/电影/动漫/手办/海外出圈自动标注
  - SBTI 人格筛选：复用三江推荐的 SBTI 机制
  - Excel 导入：支持从起点图领取的万订全量Excel导入
  - 本地缓存：7天有效，过期自动增量更新

数据加载策略：
  预置数据(134本) → 用户缓存 → 过期?增量更新(秒级) → 全量抓取(兜底)

数据源：
  - 十万均订：https://www.qidiantu.com/badge/shiwanjunding
  - 万订：https://www.qidiantu.com/badge/wanrenzhuipeng
  - 详情页：https://www.qidiantu.com/info/{book_id}

用法：
  python3 classic_picker.py                      # 默认推荐3本经典（秒加载）
  python3 classic_picker.py --count 5            # 推荐5本
  python3 classic_picker.py --sbti MALO          # 按 SBTI 人格推荐
  python3 classic_picker.py --tier 100k          # 只推荐十万均订
  python3 classic_picker.py --tier 10k           # 只推荐万订
  python3 classic_picker.py --check-update       # 增量检查新增书（秒级）
  python3 classic_picker.py --refresh            # 全量重新抓取
  python3 classic_picker.py --import-excel FILE  # 从Excel导入全量数据
  python3 classic_picker.py --dump-cache         # 输出完整缓存数据
  python3 classic_picker.py --setup              # 检测环境

依赖（自动安装）：beautifulsoup4 lxml
可选依赖（Excel导入用）：openpyxl
"""

import argparse
import hashlib
import json
import os
import random
import re
import subprocess
import sys
import time
from datetime import datetime, timedelta
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError


# ─── 环境检测与自动安装 ────────────────────────────────
REQUIRED_PACKAGES = {
    "bs4": "beautifulsoup4",
    "lxml": "lxml",
}


def check_and_install_deps(verbose=False):
    """检测并自动安装缺失的 Python 依赖。"""
    missing = []
    for import_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            __import__(import_name)
            if verbose:
                print(f"  ✅ {pip_name} 已安装", file=sys.stderr)
        except ImportError:
            missing.append((import_name, pip_name))
            if verbose:
                print(f"  ❌ {pip_name} 未安装", file=sys.stderr)

    if not missing:
        if verbose:
            print("[OK] 所有依赖已就绪！", file=sys.stderr)
        return True

    pip_packages = [pip_name for _, pip_name in missing]
    print(f"[INFO] 检测到缺失依赖: {', '.join(pip_packages)}，正在自动安装...", file=sys.stderr)

    pip_commands = [
        [sys.executable, "-m", "pip", "install"] + pip_packages,
        ["pip3", "install"] + pip_packages,
        ["pip", "install"] + pip_packages,
    ]

    for cmd in pip_commands:
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
            if result.returncode == 0:
                print(f"[OK] 依赖安装成功: {', '.join(pip_packages)}", file=sys.stderr)
                return True
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue
        except Exception:
            continue

    print(f"[ERROR] 自动安装失败，请手动: pip3 install {' '.join(pip_packages)}", file=sys.stderr)
    return False


# ─── 启动时自动检测依赖 ────────────────────────────────
try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    if check_and_install_deps():
        try:
            from bs4 import BeautifulSoup
            HAS_BS4 = True
        except ImportError:
            HAS_BS4 = False
    else:
        HAS_BS4 = False


# ─── 配置 ─────────────────────────────────────────────
BASE_URL = "https://www.qidiantu.com"
BADGE_100K_PATH = "/badge/shiwanjunding"   # 十万均订徽章页
BADGE_10K_PATH = "/badge/wanrenzhuipeng"   # 万订徽章页
BOOK_INFO_PATH = "/info/{book_id}"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}
REQUEST_TIMEOUT = 15
RETRY_COUNT = 2
RETRY_DELAY = 3

# 缓存路径
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(SCRIPT_DIR)  # 上级 = skill 根目录
CACHE_DIR = os.path.join(SCRIPT_DIR, ".cache")
CLASSIC_CACHE_FILE = os.path.join(CACHE_DIR, "classic_books.json")
PRESET_FILE = os.path.join(SKILL_DIR, "data", "preset_classics.json")
HISTORY_FILE = os.path.join(SCRIPT_DIR, ".classic_history.json")
CACHE_MAX_AGE_DAYS = 7
# 增量更新：仅抓取列表页，不逐本抓详情，超过此阈值才补详情
INCREMENTAL_DETAIL_BATCH = 20


# ─── IP 衍生品信息库 ──────────────────────────────────
# 有影视化/动漫化/手办/海外出圈的经典作品信息
# 格式：book_name -> {adaptations, overseas, forums}
IP_DATABASE = {
    "诡秘之主": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "动漫", "title": "诡秘之主 动画", "year": 2025, "platform": "腾讯视频/B站", "note": "IMDb开分9.3，首部评分人次过万的国漫"},
            {"type": "衍生品", "title": "卡牌/周边", "year": 2024, "platform": "阅文官方", "note": "大英图书馆联名周边，值夜者主题画展"},
        ],
        "overseas": {
            "status": "🌍 全球顶级IP",
            "highlights": [
                "入藏大英图书馆（2022年首批+2024年二批）",
                "海外读者评价比肩《指环王》《冰与火之歌》",
                "Reddit r/noveltranslations 最热门中文小说之一",
                "Fandom Wiki 拥有完整英文百科",
            ],
            "forums": [
                {"name": "诡秘之主官网", "url": "https://www.lomworld.com/"},
                {"name": "Lord of Mysteries Wiki (Fandom)", "url": "https://lordofthemysteries.fandom.com/"},
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/lord-of-the-mysteries_11022733006234505"},
                {"name": "Reddit讨论", "url": "https://www.reddit.com/r/LordofTheMysteries/"},
                {"name": "NovelUpdates", "url": "https://www.novelupdates.com/series/lord-of-the-mysteries/"},
            ],
        },
    },
    "宿命之环": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "动漫", "title": "宿命之环（预计）", "year": 2026, "platform": "待定", "note": "诡秘宇宙续作，IP联动开发中"},
        ],
        "overseas": {
            "status": "🌍 海外高热度续作",
            "highlights": [
                "诡秘之主续作，首日30字预告冲上微博热搜",
                "WebNovel同步连载英文版",
                "共享诡秘之主的全球粉丝群",
            ],
            "forums": [
                {"name": "诡秘之主官网", "url": "https://www.lomworld.com/"},
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/fate-unbound_26869085206424105"},
                {"name": "Reddit讨论", "url": "https://www.reddit.com/r/LordofTheMysteries/"},
            ],
        },
    },
    "大奉打更人": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "电视剧", "title": "大奉打更人", "year": 2025, "platform": "Netflix/腾讯视频", "note": "Netflix上线，海外热播"},
            {"type": "动漫", "title": "大奉打更人 动画", "year": 2024, "platform": "腾讯视频", "note": "改编动画"},
        ],
        "overseas": {
            "status": "🌍 Netflix热播",
            "highlights": [
                "Netflix全球上线",
                "WebNovel英文翻译连载",
            ],
            "forums": [
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/the-great-ruler_17921erta5050405"},
                {"name": "NovelUpdates", "url": "https://www.novelupdates.com/series/the-great-rulers-nightwatchman/"},
            ],
        },
    },
    "道诡异仙": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "动漫", "title": "道诡异仙（预计）", "year": 2026, "platform": "待定", "note": "暗黑诡异风格，动画化呼声极高"},
        ],
        "overseas": {
            "status": "🌏 海外高人气",
            "highlights": [
                "海外暗黑克苏鲁爱好者圈层热门",
                "WebNovel连载中",
            ],
            "forums": [
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/my-iyashikei-game_22636825706498505"},
                {"name": "NovelUpdates", "url": "https://www.novelupdates.com/series/my-house-of-horrors/"},
            ],
        },
    },
    "大王饶命": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "动漫", "title": "大王饶命 动画", "year": 2021, "platform": "腾讯视频/B站", "note": "改编动画播出"},
        ],
        "overseas": {
            "status": "🌏 海外经典",
            "highlights": [
                "会说话的肘子代表作",
                "WebNovel英文版 My House of Horrors 翻译完结",
            ],
            "forums": [
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/please-spare-me%2C-great-king!_19478027506188805"},
            ],
        },
    },
    "夜的命名术": {
        "tier_label": "🏆十万均订",
        "adaptations": [],
        "overseas": {
            "status": "🌏 海外连载",
            "highlights": [
                "肘子三本十万均订之一",
                "WebNovel同步连载",
            ],
            "forums": [
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/the-naming-of-the-night_22928163005120105"},
            ],
        },
    },
    "我师兄实在太稳健了": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "动漫", "title": "我师兄实在太稳健了 动画", "year": 2024, "platform": "腾讯视频/B站", "note": "苟道流鼻祖改编动画"},
        ],
        "overseas": {
            "status": "🌏 海外高人气",
            "highlights": [
                "苟道流开山之作",
                "WebNovel翻译连载",
            ],
            "forums": [
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/my-senior-brother-is-too-steady_17134937205939505"},
                {"name": "NovelUpdates", "url": "https://www.novelupdates.com/series/my-senior-brother-is-too-steady/"},
            ],
        },
    },
    "不科学御兽": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "动漫", "title": "不科学御兽 动画（预计）", "year": 2026, "platform": "待定", "note": "御兽流巅峰，动画化开发中"},
        ],
        "overseas": {
            "status": "🌏 海外连载",
            "highlights": ["WebNovel同步连载"],
            "forums": [
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/unscientific-beast-taming_22498703006009505"},
            ],
        },
    },
    "深海余烬": {
        "tier_label": "🏆十万均订",
        "adaptations": [],
        "overseas": {
            "status": "🌏 海外高口碑",
            "highlights": [
                "远瞳精品，克苏鲁+蒸汽朋克风格",
                "海外暗黑冒险爱好者关注",
            ],
            "forums": [
                {"name": "NovelUpdates", "url": "https://www.novelupdates.com/series/deep-sea-embers/"},
            ],
        },
    },
    "光阴之外": {
        "tier_label": "🏆十万均订",
        "adaptations": [],
        "overseas": {
            "status": "🌏 耳根海外粉丝群",
            "highlights": [
                "耳根回归之作",
                "WebNovel连载",
            ],
            "forums": [
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/beyond-the-timescape_23506481805498505"},
            ],
        },
    },
    "灵境行者": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "动漫", "title": "灵境行者 动画（预计）", "year": 2026, "platform": "待定", "note": "卖报小郎君第二本十万均订"},
        ],
        "overseas": {
            "status": "🌏 海外连载",
            "highlights": ["惊悚+玩梗风格"],
            "forums": [],
        },
    },
    "逼我重生是吧": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "短剧", "title": "逼我重生是吧 短剧", "year": 2025, "platform": "抖音/快手", "note": "重生电商文，短剧改编"},
        ],
        "overseas": {"status": "", "highlights": [], "forums": []},
    },
    "都重生了谁谈恋爱啊": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "影视", "title": "都重生了谁谈恋爱啊（开发中）", "year": 2026, "platform": "待定", "note": "重生搞钱流代表作，影视化开发中"},
        ],
        "overseas": {"status": "", "highlights": [], "forums": []},
    },
    "深空彼岸": {
        "tier_label": "🏆十万均订",
        "adaptations": [],
        "overseas": {
            "status": "🌏 辰东全球粉丝",
            "highlights": ["辰东回归之作", "WebNovel连载"],
            "forums": [
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/the-deep-beyond_23764325105684305"},
            ],
        },
    },
    "没钱修什么仙？": {
        "tier_label": "🏆十万均订",
        "adaptations": [],
        "overseas": {
            "status": "🔥 新晋十万均订",
            "highlights": ["2026年新晋十万均订", "赛博修仙/贷款修仙创意火爆"],
            "forums": [],
        },
    },
    "万族之劫": {
        "tier_label": "🏆十万均订",
        "adaptations": [],
        "overseas": {
            "status": "🌏 海外连载",
            "highlights": ["WebNovel翻译连载"],
            "forums": [
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/the-human-race's-calamity_23741125105459305"},
            ],
        },
    },
    "绍宋": {
        "tier_label": "🏆十万均订",
        "adaptations": [
            {"type": "电视剧", "title": "绍宋（开发中）", "year": 2026, "platform": "待定", "note": "历史文巅峰之作，影视化筹备中"},
        ],
        "overseas": {"status": "", "highlights": [], "forums": []},
    },
    "高武纪元": {
        "tier_label": "🏆十万均订",
        "adaptations": [],
        "overseas": {"status": "", "highlights": [], "forums": []},
    },
    "第一序列": {
        "tier_label": "🏆十万均订",
        "adaptations": [],
        "overseas": {
            "status": "🌏 海外完结经典",
            "highlights": ["肘子三部曲之一", "英文版完结"],
            "forums": [
                {"name": "WebNovel 英文版", "url": "https://www.webnovel.com/book/first-sequence_18670873405727305"},
            ],
        },
    },
    # ───── 以下为非十万均订但有 IP 衍生的万订作品 ─────
    "玄浑道章": {
        "adaptations": [],
        "overseas": {"status": "🌏 修仙文精品", "highlights": ["长篇修仙经典"], "forums": []},
    },
    "长夜君主": {
        "adaptations": [],
        "overseas": {"status": "", "highlights": [], "forums": []},
    },
}

# ─── SBTI 人格库（复用三江推荐的定义）──────────────────
SBTI_PERSONALITIES = [
    {"code": "SOLO", "name": "孤儿", "trait": "独行侠，莫挨老子",
     "keywords": ["武侠", "仙侠", "灵异", "悬疑", "末世"],
     "reason_templates": [
         "这本经典的主角就是SBTI里的{code}（{name}），一个人扛下所有——{trait}。能杀进十万/万订的狠角色。",
         "测出{code}（{name}）的朋友注意了！这本万订经典的主角比你还能独来独往。",
     ]},
    {"code": "MALO", "name": "吗喽", "trait": "躺平哲学，拒绝内卷",
     "keywords": ["都市", "轻小说", "日常", "生活"],
     "reason_templates": [
         "SBTI吗喽人狂喜！这本万订经典的主角摆得比你还平，后期躺赢得比你还香。",
         "这本经典就是{name}人格的圣经，主角把{trait}发挥到了艺术层面。万订实至名归。",
     ]},
    {"code": "FUCK", "name": "草者", "trait": "叛逆暴躁，人形野草不服管",
     "keywords": ["玄幻", "奇幻", "战争", "军事"],
     "reason_templates": [
         "SBTI草者（{code}）专属经典：主角脾气比你还暴，打脸比你还狠——难怪能杀进万订。",
         "测出{code}的朋友，这本经典的主角替你把想说的话都用拳头说了。",
     ]},
    {"code": "OJBK", "name": "无所谓人", "trait": "随波逐流，口头禅：随便",
     "keywords": ["历史", "二次元", "轻小说"],
     "reason_templates": [
         "你SBTI测出{code}？这本万订经典的主角比你还随便，结果随便着就封神了。",
         "无所谓人看了都得说一句'有所谓'的经典好书。",
     ]},
    {"code": "FAKE", "name": "伪人", "trait": "面具大师，隐藏真实自我",
     "keywords": ["悬疑", "灵异", "历史", "权谋"],
     "reason_templates": [
         "SBTI伪人（{code}）必入经典！主角面具大师段位爆表——万订不是盖的。",
         "这本经典就是{name}的操作手册，主角{trait}玩到飞起。",
     ]},
    {"code": "IMSB", "name": "自我攻击者", "trait": "内耗严重，又菜又爱想",
     "keywords": ["都市", "科幻", "现实"],
     "reason_templates": [
         "SBTI自我攻击者看这本经典能治病！主角逆袭之路让万千读者追到万订。",
         "{name}人格特供经典良药：原来内耗也是一种超能力。",
     ]},
    {"code": "JOKE-R", "name": "小丑", "trait": "用搞笑掩盖心碎，气氛组担当",
     "keywords": ["轻小说", "都市", "游戏", "日常"],
     "reason_templates": [
         "SBTI小丑（{code}）本丑推荐经典：笑着笑着就哭了，哭着哭着又追到万订。",
         "小丑人格专属经典治愈番：主角比你还能活跃气氛。",
     ]},
    {"code": "ZZZZ", "name": "装死者", "trait": "遇事就躲，逃避现实大师",
     "keywords": ["奇幻", "仙侠", "游戏", "异界"],
     "reason_templates": [
         "SBTI装死者找到经典组织！主角装死技能满级——装着装着就成绝世强者了。",
         "这本经典就是为{name}量身定做的：主角躺着躺着就赢了。",
     ]},
    {"code": "ATM", "name": "ATM-er", "trait": "人形提款机，总在为别人兜底",
     "keywords": ["都市", "历史", "现实"],
     "reason_templates": [
         "SBTI ATM-er看了想打人的经典好书：前期冤大头，后期让所有人跪着唱征服。",
         "ATM人格代入感拉满的万订经典：{trait}了三十章，剩下全在收利息。",
     ]},
    {"code": "DRUNK", "name": "酒鬼", "trait": "快乐全靠喝，一杯敬过往",
     "keywords": ["武侠", "历史", "仙侠"],
     "reason_templates": [
         "SBTI酒鬼看了能醉的经典好书！主角越喝越强——经典不愧是经典。",
         "给{name}人格的一个不喝酒也能快乐的方案——看这本万订经典。",
     ]},
    {"code": "WOC!", "name": "握草人", "trait": "易震惊体质，遇事只会卧槽",
     "keywords": ["玄幻", "科幻", "奇幻", "悬疑"],
     "reason_templates": [
         "SBTI握草人预警：这本万订经典每三章一个'卧槽'时刻。",
         "这本经典就是为{name}准备的精神过山车——万订读者的共同卧槽记忆。",
     ]},
    {"code": "MUM", "name": "妈妈", "trait": "天生母爱泛滥，到处照顾人",
     "keywords": ["轻小说", "都市", "日常", "游戏"],
     "reason_templates": [
         "SBTI妈妈人格看了想领养主角的经典好书。",
         "{name}人格专属经典：主角身边全是需要照顾的人。",
     ]},
    {"code": "SEXY", "name": "尤物", "trait": "自信撩人，自带魅力光环",
     "keywords": ["都市", "历史", "武侠", "现实"],
     "reason_templates": [
         "SBTI尤物本物推荐经典：主角魅力值拉满——万订认证。",
         "给{name}人格推荐的'如何征服世界'经典教科书。",
     ]},
    {"code": "CTRL", "name": "拿捏者", "trait": "控场大师，一切尽在掌握",
     "keywords": ["历史", "悬疑", "权谋", "科幻"],
     "reason_templates": [
         "SBTI拿捏者看了拍大腿的万订经典：主角控场能力让你说声'学到了'。",
         "{name}人格必看经典：这书主角的操盘手法连CTRL都得服。",
     ]},
    {"code": "SHIT", "name": "狗屎人", "trait": "愤世嫉俗，看啥都不顺眼",
     "keywords": ["现实", "末世", "科幻", "战争"],
     "reason_templates": [
         "SBTI狗屎人终于找到能看顺眼的经典：主角选择了掀桌子——过瘾。",
         "{name}人格特供万订经典：整个世界都是垃圾？主角一个人翻了个底朝天。",
     ]},
]

# SBTI 索引
_SBTI_INDEX = {}
for _p in SBTI_PERSONALITIES:
    _SBTI_INDEX[_p["code"].upper()] = _p
    _SBTI_INDEX[_p["name"]] = _p
    _clean_code = _p["code"].replace("!", "").replace("-", "").upper()
    _SBTI_INDEX[_clean_code] = _p


# ─── SBTI 工具函数 ────────────────────────────────────

def resolve_sbti_filter(query):
    """将用户输入解析为 SBTI 人格对象列表"""
    if not query:
        return []
    query_upper = query.strip().upper().replace("!", "").replace("-", "")
    query_stripped = query.strip()
    if query_upper in _SBTI_INDEX:
        return [_SBTI_INDEX[query_upper]]
    if query_stripped in _SBTI_INDEX:
        return [_SBTI_INDEX[query_stripped]]

    # 模糊匹配
    trait_map = {
        "躺平": "MALO", "摆烂": "MALO", "不想上班": "MALO", "咸鱼": "MALO",
        "暴躁": "FUCK", "叛逆": "FUCK", "不服": "FUCK", "莽": "FUCK",
        "社恐": "SOLO", "独来独往": "SOLO", "一个人": "SOLO", "孤独": "SOLO",
        "随便": "OJBK", "无所谓": "OJBK", "佛系": "OJBK", "都行": "OJBK",
        "内耗": "IMSB", "焦虑": "IMSB", "纠结": "IMSB",
        "搞笑": "JOKE-R", "逗比": "JOKE-R", "气氛组": "JOKE-R",
        "装死": "ZZZZ", "逃避": "ZZZZ", "鸵鸟": "ZZZZ",
        "老好人": "ATM", "讨好": "ATM", "心软": "ATM",
        "喝酒": "DRUNK", "干杯": "DRUNK",
        "卧槽": "WOC!", "震惊": "WOC!",
        "照顾": "MUM", "操心": "MUM",
        "自信": "SEXY", "魅力": "SEXY",
        "控制": "CTRL", "拿捏": "CTRL", "掌控": "CTRL",
        "愤世嫉俗": "SHIT", "看不惯": "SHIT",
        "假面": "FAKE", "伪装": "FAKE",
    }
    scores = []
    q_lower = query_stripped.lower()
    for p in SBTI_PERSONALITIES:
        score = 0
        for keyword, target_code in trait_map.items():
            if keyword in q_lower and p["code"] == target_code:
                score += 15
        for trait_word in p["trait"].replace("，", " ").replace("、", " ").split():
            if trait_word in q_lower or q_lower in trait_word:
                score += 5
        if p["name"] in query_stripped or query_stripped in p["name"]:
            score += 10
        if score > 0:
            scores.append((score, p))
    if not scores:
        return []
    scores.sort(key=lambda x: -x[0])
    return [s[1] for s in scores[:3]]


def assign_sbti_personality(book):
    """根据书籍分类分配 SBTI 人格"""
    category = book.get("category", "")
    book_id = book.get("book_id", "")
    matched = [p for p in SBTI_PERSONALITIES if any(kw in category for kw in p["keywords"])]
    if matched:
        idx = int(hashlib.md5(book_id.encode()).hexdigest(), 16) % len(matched)
        return matched[idx]
    idx = int(hashlib.md5(book_id.encode()).hexdigest(), 16) % len(SBTI_PERSONALITIES)
    return SBTI_PERSONALITIES[idx]


def generate_sbti_reason(book, personality):
    """生成 SBTI 必看理由"""
    templates = personality["reason_templates"]
    book_name = book.get("book_name", "")
    idx = int(hashlib.md5(book_name.encode()).hexdigest(), 16) % len(templates)
    return templates[idx].format(
        code=personality["code"],
        name=personality["name"],
        trait=personality["trait"],
    )


def label_book_sbti(book):
    """为单本书打上 SBTI 标签"""
    personality = assign_sbti_personality(book)
    book["sbti_code"] = personality["code"]
    book["sbti_name"] = personality["name"]
    book["sbti_trait"] = personality["trait"]
    book["sbti_reason"] = generate_sbti_reason(book, personality)
    return book


def dedupe_sbti_in_batch(books):
    """同批推荐里 SBTI 人格尽量不重复"""
    used_codes = set()
    for book in books:
        code = book.get("sbti_code", "")
        if code not in used_codes:
            used_codes.add(code)
            continue
        available = [p for p in SBTI_PERSONALITIES if p["code"] not in used_codes]
        if not available:
            used_codes.add(code)
            continue
        category = book.get("category", "")
        cat_matched = [p for p in available for kw in p["keywords"] if kw in category]
        pool = cat_matched or available
        idx = int(hashlib.md5(book.get("book_id", "").encode()).hexdigest(), 16) % len(pool)
        personality = pool[idx]
        used_codes.add(personality["code"])
        book["sbti_code"] = personality["code"]
        book["sbti_name"] = personality["name"]
        book["sbti_trait"] = personality["trait"]
        book["sbti_reason"] = generate_sbti_reason(book, personality)
    return books


# ─── 网络请求 ─────────────────────────────────────────

def http_get(url, retries=RETRY_COUNT):
    """带重试的 HTTP GET"""
    for attempt in range(retries + 1):
        try:
            req = Request(url, headers=HEADERS)
            with urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
                charset = resp.headers.get_content_charset() or "utf-8"
                return resp.read().decode(charset, errors="replace")
        except HTTPError as e:
            if e.code == 404:
                return None
            if attempt < retries:
                time.sleep(RETRY_DELAY)
            else:
                print(f"[ERROR] HTTP {e.code}: {url}", file=sys.stderr)
                return None
        except URLError as e:
            if attempt < retries:
                time.sleep(RETRY_DELAY)
            else:
                print(f"[ERROR] URL Error: {e}", file=sys.stderr)
                return None
        except Exception as e:
            print(f"[ERROR] Unexpected: {e}", file=sys.stderr)
            return None
    return None


# ─── 徽章页面解析 ─────────────────────────────────────

def parse_badge_page(html):
    """解析起点图徽章页面，提取书名和 book_id"""
    books = []
    if not html or not HAS_BS4:
        return books
    soup = BeautifulSoup(html, "lxml")

    # 策略1：找 /info/xxx 链接
    seen_ids = set()
    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        m = re.search(r"/info/(\d+)", href)
        if m:
            book_id = m.group(1)
            if book_id not in seen_ids:
                book_name = a_tag.get_text(strip=True)
                if book_name and len(book_name) > 1:
                    seen_ids.add(book_id)
                    books.append({
                        "book_id": book_id,
                        "book_name": book_name,
                    })
    return books


def fetch_badge_books(badge_path, tier_label):
    """从徽章页面获取书籍列表"""
    url = BASE_URL + badge_path
    print(f"[FETCH] 正在获取{tier_label}列表: {url}", file=sys.stderr)
    html = http_get(url)
    if not html:
        print(f"[ERROR] 无法获取{tier_label}页面", file=sys.stderr)
        return []
    books = parse_badge_page(html)
    for b in books:
        b["tier"] = tier_label
    print(f"[FETCH] {tier_label}: 解析到 {len(books)} 本", file=sys.stderr)
    return books


# ─── 详情解析 ─────────────────────────────────────────

def fetch_book_detail(book_id):
    """获取单本书详情"""
    url = BASE_URL + BOOK_INFO_PATH.format(book_id=book_id)
    html = http_get(url)
    if not html or not HAS_BS4:
        return {}

    soup = BeautifulSoup(html, "lxml")
    detail = {}
    text = soup.get_text()

    try:
        title_tag = soup.select_one("h1")
        if title_tag:
            name = title_tag.get_text(strip=True)
            name = re.sub(r"\(VIP\).*", "", name).strip()
            detail["book_name"] = name

        m = re.search(r"作者[：:]\s*([^\(（]+)", text)
        if m:
            detail["author"] = m.group(1).strip()

        m = re.search(r"标签[：:]\s*([^\s　]+)\s*([^\s　]*)", text)
        if m:
            cat = m.group(1).strip()
            sub = m.group(2).strip()
            detail["category"] = f"{cat} · {sub}" if sub else cat

        m = re.search(r"状态[：:]\s*(\S+)", text)
        if m:
            detail["status"] = m.group(1).strip()

        m = re.search(r"总字数[：:]\s*(\d+[\.?\d]*万?)", text)
        if m:
            detail["word_count"] = m.group(1)

        m = re.search(r"总收藏[：:]\s*(\d+[\.?\d]*万?)", text)
        if m:
            detail["collect"] = m.group(1)

        m = re.search(r"总推荐[：:]\s*(\d+[\.?\d]*万?k?)", text)
        if m:
            detail["recommend"] = m.group(1)

        # 简介
        intro = ""
        m = re.search(r"简介及书单.*?\n+(.*?)(?:展开更多简介|收起)", text, re.DOTALL)
        if m:
            intro = re.sub(r"\s+", " ", m.group(1).strip())
        if not intro:
            m = re.search(r"数据\s*\n+(.*?)(?:展开更多简介|收起|书单)", text, re.DOTALL)
            if m:
                intro = re.sub(r"\s+", " ", m.group(1).strip())
        if not intro:
            for p in soup.find_all(string=True):
                p_text = p.strip()
                if len(p_text) > 100 and "起点图" not in p_text and "鲁ICP" not in p_text:
                    intro = re.sub(r"\s+", " ", p_text)
                    break
        if intro:
            detail["intro"] = intro
    except Exception as e:
        print(f"[WARN] 详情解析失败 {book_id}: {e}", file=sys.stderr)

    return detail


# ─── IP 衍生品信息注入 ────────────────────────────────

def inject_ip_info(book):
    """注入 IP 衍生品和海外出圈信息"""
    name = book.get("book_name", "")
    ip_info = IP_DATABASE.get(name, {})

    book["has_ip"] = bool(ip_info)
    book["adaptations"] = ip_info.get("adaptations", [])

    overseas = ip_info.get("overseas", {})
    book["overseas_status"] = overseas.get("status", "")
    book["overseas_highlights"] = overseas.get("highlights", [])
    book["overseas_forums"] = overseas.get("forums", [])

    return book


# ─── 缓存管理 ─────────────────────────────────────────

def load_classic_cache(ignore_expiry=False):
    """加载经典书库缓存，缓存不存在时自动降级到预置数据"""
    # 1. 优先尝试用户缓存
    if os.path.exists(CLASSIC_CACHE_FILE):
        try:
            with open(CLASSIC_CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict) and "books" in data and data["books"]:
                # 检查过期
                if not ignore_expiry:
                    fetched = data.get("fetched_at", "")
                    if fetched:
                        try:
                            fetched_dt = datetime.strptime(fetched, "%Y-%m-%d %H:%M:%S")
                            age_days = (datetime.now() - fetched_dt).days
                            if age_days > CACHE_MAX_AGE_DAYS:
                                print(f"[CACHE] 缓存已过期({age_days}天)，需增量更新", file=sys.stderr)
                                data["_expired"] = True
                        except ValueError:
                            pass
                return data
        except (json.JSONDecodeError, IOError):
            pass

    # 2. 降级到预置数据
    if os.path.exists(PRESET_FILE):
        try:
            with open(PRESET_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict) and "books" in data and data["books"]:
                print(f"[PRESET] ✅ 加载预置经典书库: {len(data['books'])} 本 "
                      f"(版本: {data.get('preset_version', '未知')})", file=sys.stderr)
                data["_from_preset"] = True
                # 将预置数据写入缓存，下次直接读缓存
                save_classic_cache(data["books"])
                return data
        except (json.JSONDecodeError, IOError):
            pass

    return None


def save_classic_cache(books):
    """保存经典书库到缓存"""
    os.makedirs(CACHE_DIR, exist_ok=True)
    data = {
        "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "book_count": len(books),
        "tier_100k_count": len([b for b in books if b.get("tier") == "🏆十万均订"]),
        "tier_10k_count": len([b for b in books if b.get("tier") == "📈万订"]),
        "books": books,
    }
    try:
        with open(CLASSIC_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        fsize = os.path.getsize(CLASSIC_CACHE_FILE)
        print(f"[CACHE] 已缓存 {len(books)} 本经典到 {CLASSIC_CACHE_FILE} ({fsize/1024:.1f} KB)", file=sys.stderr)
    except IOError as e:
        print(f"[WARN] 缓存写入失败: {e}", file=sys.stderr)


# ─── 历史管理 ─────────────────────────────────────────

def load_history():
    if not os.path.exists(HISTORY_FILE):
        return {"last_book_ids": [], "last_date": "", "all_recommended": {}}
    try:
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {"last_book_ids": [], "last_date": "", "all_recommended": {}}


def save_history(book_ids, date_str):
    history = load_history()
    history["last_book_ids"] = book_ids
    history["last_date"] = date_str
    if "all_recommended" not in history:
        history["all_recommended"] = {}
    history["all_recommended"][date_str] = book_ids
    cutoff = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    history["all_recommended"] = {k: v for k, v in history["all_recommended"].items() if k >= cutoff}
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except IOError:
        pass


# ─── 全量抓取+缓存 ───────────────────────────────────

def fetch_and_cache_all():
    """全量抓取十万均订+万订数据，补充详情后缓存"""
    # 1. 抓取十万均订
    books_100k = fetch_badge_books(BADGE_100K_PATH, "🏆十万均订")

    # 2. 抓取万订（最新100本）
    books_10k = fetch_badge_books(BADGE_10K_PATH, "📈万订")

    # 去重（十万均订优先）
    seen_ids = set()
    all_books = []
    for b in books_100k:
        if b["book_id"] not in seen_ids:
            seen_ids.add(b["book_id"])
            all_books.append(b)
    for b in books_10k:
        if b["book_id"] not in seen_ids:
            seen_ids.add(b["book_id"])
            all_books.append(b)

    print(f"[FETCH] 合并去重后共 {len(all_books)} 本，开始获取详情...", file=sys.stderr)

    # 3. 逐本获取详情
    enriched = []
    for i, book in enumerate(all_books, 1):
        print(f"[FETCH] ({i}/{len(all_books)}) {book['book_name']}...", file=sys.stderr)
        detail = fetch_book_detail(book["book_id"])
        merged = {**book, **detail}
        merged.setdefault("book_name", f"未知({book['book_id']})")
        merged.setdefault("author", "未知作者")
        merged.setdefault("category", "未知分类")
        merged.setdefault("intro", "暂无简介")
        merged.setdefault("word_count", "")
        merged.setdefault("collect", "")
        merged.setdefault("recommend", "")
        merged["qidian_url"] = f"https://www.qidian.com/book/{book['book_id']}/?_trace=qidiandayrec_skill"

        # 注入 IP 信息
        merged = inject_ip_info(merged)

        # 打 SBTI 标签
        merged = label_book_sbti(merged)

        enriched.append(merged)
        if i < len(all_books):
            time.sleep(1.5)

    # 4. 缓存
    save_classic_cache(enriched)
    return enriched


def get_classic_books(force_refresh=False, check_update=False):
    """获取经典书库（预置 → 缓存 → 增量更新 → 全量抓取）

    数据加载优先级：
    1. 用户缓存 (scripts/.cache/classic_books.json)
    2. 预置数据 (data/preset_classics.json) — 首次使用零等待
    3. 缓存过期时 → 增量更新（仅抓列表页对比新增，秒级完成）
    4. 全量抓取（仅在无任何数据源时触发）

    Args:
        force_refresh: True=强制全量重新抓取
        check_update: True=仅执行增量检查，发现新书则补充入缓存
    """
    if force_refresh:
        print("[CACHE] 强制全量刷新...", file=sys.stderr)
        return fetch_and_cache_all()

    # 加载现有数据（缓存 or 预置）
    cached = load_classic_cache(ignore_expiry=True)

    if cached and cached.get("books"):
        books = cached["books"]
        from_preset = cached.get("_from_preset", False)
        is_expired = cached.get("_expired", False)
        source = "预置数据" if from_preset else "缓存"
        print(f"[CACHE] ✅ {source}命中: {len(books)} 本 "
              f"(🏆十万均订 {cached.get('tier_100k_count', '?')} + "
              f"📈万订 {cached.get('tier_10k_count', '?')}) "
              f"更新时间: {cached.get('fetched_at', '预置')}", file=sys.stderr)

        # 增量更新检查
        if check_update or is_expired or from_preset:
            new_books = incremental_update(books)
            if new_books is not None:
                books = new_books

        return books

    # 没有任何数据源，全量抓取
    print("[CACHE] 无缓存也无预置数据，首次全量抓取...", file=sys.stderr)
    return fetch_and_cache_all()


def incremental_update(existing_books):
    """增量更新：仅抓取列表页对比新增书目，秒级完成

    逻辑：
    1. 抓取十万均订 + 万订列表页（各1次请求，共2次）
    2. 与已有书目对比，找出新增 book_id
    3. 仅对新增书补详情 + IP + SBTI
    4. 合并到现有书库并更新缓存

    Returns:
        更新后的完整书目列表，无新增则返回 None
    """
    print("[UPDATE] 增量检查中（仅抓列表页对比新增）...", file=sys.stderr)
    existing_ids = {b.get("book_id", "") for b in existing_books}

    # 抓取列表页
    try:
        list_100k = fetch_badge_books(BADGE_100K_PATH, "🏆十万均订")
        time.sleep(1)
        list_10k = fetch_badge_books(BADGE_10K_PATH, "📈万订")
    except Exception as e:
        print(f"[UPDATE] 列表页抓取失败: {e}，跳过增量更新", file=sys.stderr)
        return None

    # 对比新增
    new_entries = []
    for b in list_100k + list_10k:
        if b["book_id"] not in existing_ids:
            new_entries.append(b)
            existing_ids.add(b["book_id"])  # 防重复

    # 检查是否有已有书需要升级 tier（万订 → 十万均订）
    upgrade_ids = {}
    for b in list_100k:
        bid = b["book_id"]
        for eb in existing_books:
            if eb.get("book_id") == bid and eb.get("tier") == "📈万订":
                upgrade_ids[bid] = "🏆十万均订"
                break

    if not new_entries and not upgrade_ids:
        print(f"[UPDATE] ✅ 无新增书目，书库保持 {len(existing_books)} 本", file=sys.stderr)
        # 刷新缓存时间戳
        save_classic_cache(existing_books)
        return None

    # 处理 tier 升级
    if upgrade_ids:
        for book in existing_books:
            bid = book.get("book_id", "")
            if bid in upgrade_ids:
                old_tier = book.get("tier", "")
                book["tier"] = upgrade_ids[bid]
                print(f"[UPDATE] ⬆️ {book.get('book_name', bid)}: {old_tier} → {upgrade_ids[bid]}", file=sys.stderr)

    # 补充新增书的详情
    if new_entries:
        print(f"[UPDATE] 🆕 发现 {len(new_entries)} 本新增书，补充详情中...", file=sys.stderr)
        enriched_new = []
        for i, book in enumerate(new_entries, 1):
            print(f"[UPDATE]   ({i}/{len(new_entries)}) {book.get('book_name', book['book_id'])}...", file=sys.stderr)
            detail = fetch_book_detail(book["book_id"])
            merged = {**book, **detail}
            merged.setdefault("book_name", f"未知({book['book_id']})")
            merged.setdefault("author", "未知作者")
            merged.setdefault("category", "未知分类")
            merged.setdefault("intro", "暂无简介")
            merged.setdefault("word_count", "")
            merged.setdefault("collect", "")
            merged.setdefault("recommend", "")
            merged["qidian_url"] = f"https://www.qidian.com/book/{book['book_id']}/?_trace=qidiandayrec_skill"
            merged = inject_ip_info(merged)
            merged = label_book_sbti(merged)
            enriched_new.append(merged)
            if i < len(new_entries):
                time.sleep(1.5)

        all_books = existing_books + enriched_new
        print(f"[UPDATE] ✅ 新增 {len(enriched_new)} 本，书库更新至 {len(all_books)} 本", file=sys.stderr)
    else:
        all_books = existing_books

    save_classic_cache(all_books)
    return all_books


# ─── Excel 导入 ───────────────────────────────────────

def import_from_excel(filepath):
    """从起点图万订全量Excel导入数据到缓存"""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 需要 openpyxl 库。安装: pip3 install openpyxl", file=sys.stderr)
        try:
            subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"],
                          capture_output=True, timeout=60)
            import openpyxl
        except Exception:
            print("[ERROR] openpyxl 安装失败", file=sys.stderr)
            return []

    print(f"[IMPORT] 正在从 Excel 导入: {filepath}", file=sys.stderr)
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # 自动检测列名
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {}
    for i, h in enumerate(headers):
        if "书名" in h or "小说" in h or "book" in h or "title" in h:
            col_map["book_name"] = i
        elif "作者" in h or "author" in h:
            col_map["author"] = i
        elif "id" in h or "编号" in h:
            col_map["book_id"] = i
        elif "分类" in h or "类型" in h or "category" in h:
            col_map["category"] = i

    if "book_name" not in col_map:
        print("[ERROR] Excel中未找到书名列，请确认表头包含'书名'", file=sys.stderr)
        return []

    imported = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name_idx = col_map["book_name"]
        if name_idx >= len(row) or not row[name_idx]:
            continue
        book_name = str(row[name_idx]).strip()
        book = {
            "book_name": book_name,
            "tier": "📈万订",  # Excel导入的默认为万订
        }
        if "author" in col_map and col_map["author"] < len(row) and row[col_map["author"]]:
            book["author"] = str(row[col_map["author"]]).strip()
        if "book_id" in col_map and col_map["book_id"] < len(row) and row[col_map["book_id"]]:
            book["book_id"] = str(row[col_map["book_id"]]).strip()
        if "category" in col_map and col_map["category"] < len(row) and row[col_map["category"]]:
            book["category"] = str(row[col_map["category"]]).strip()

        # 检查是否是十万均订作品（根据 IP_DATABASE 中的 tier_label）
        ip_info = IP_DATABASE.get(book_name, {})
        if ip_info.get("tier_label") == "🏆十万均订":
            book["tier"] = "🏆十万均订"

        imported.append(book)

    wb.close()
    print(f"[IMPORT] 从 Excel 导入 {len(imported)} 本", file=sys.stderr)

    # 合并到现有缓存（Excel数据优先级低于在线抓取）
    cached = load_classic_cache()
    if cached:
        existing_names = {b["book_name"] for b in cached["books"]}
        new_books = [b for b in imported if b["book_name"] not in existing_names]
        merged = cached["books"] + new_books
        print(f"[IMPORT] 与缓存合并：原有 {len(cached['books'])} + 新增 {len(new_books)} = {len(merged)}", file=sys.stderr)
    else:
        merged = imported

    # 补全缺失字段
    for b in merged:
        b.setdefault("book_id", "")
        b.setdefault("author", "未知作者")
        b.setdefault("category", "未知分类")
        b.setdefault("intro", "暂无简介")
        b.setdefault("word_count", "")
        b.setdefault("collect", "")
        b.setdefault("recommend", "")
        if b.get("book_id"):
            b.setdefault("qidian_url", f"https://www.qidian.com/book/{b['book_id']}/?_trace=qidiandayrec_skill")
        else:
            b.setdefault("qidian_url", "")
        b = inject_ip_info(b)
        if "sbti_code" not in b:
            b = label_book_sbti(b)

    save_classic_cache(merged)
    return merged


# ─── 推荐算法 ─────────────────────────────────────────

def pick_classic_books(books, count=3, exclude_ids=None, sbti_filter=None, tier_filter=None):
    """从经典书库中挑选推荐"""
    if exclude_ids is None:
        exclude_ids = set()
    else:
        exclude_ids = set(exclude_ids)

    candidates = list(books)

    # 层级筛选
    if tier_filter == "100k":
        candidates = [b for b in candidates if b.get("tier") == "🏆十万均订"]
        print(f"[FILTER] 仅十万均订 → {len(candidates)} 本", file=sys.stderr)
    elif tier_filter == "10k":
        candidates = [b for b in candidates if b.get("tier") == "📈万订"]
        print(f"[FILTER] 仅万订 → {len(candidates)} 本", file=sys.stderr)

    # SBTI 筛选
    if sbti_filter:
        target_codes = set(p["code"] for p in sbti_filter)
        sbti_matched = [b for b in candidates if b.get("sbti_code") in target_codes]
        if sbti_matched:
            print(f"[FILTER] SBTI → {len(sbti_matched)} 本", file=sys.stderr)
            candidates = sbti_matched

    # 去重
    deduped = [b for b in candidates if b.get("book_id", "") not in exclude_ids]
    if len(deduped) < count:
        deduped = list(candidates)

    # 随机打乱
    random.shuffle(deduped)

    # 优先选有 IP 衍生品的（加权）
    ip_books = [b for b in deduped if b.get("has_ip")]
    non_ip_books = [b for b in deduped if not b.get("has_ip")]

    selected = []
    used_categories = set()

    # 先从 IP 书中挑
    for book in ip_books:
        if len(selected) >= count:
            break
        cat = book.get("category", "")
        if cat not in used_categories or len(selected) < 1:
            selected.append(book)
            if cat:
                used_categories.add(cat)

    # 再从非 IP 书补充
    for book in non_ip_books:
        if len(selected) >= count:
            break
        cat = book.get("category", "")
        if cat not in used_categories:
            selected.append(book)
            if cat:
                used_categories.add(cat)

    # 兜底
    for book in deduped:
        if len(selected) >= count:
            break
        if book not in selected:
            selected.append(book)

    return selected[:count]


# ─── 输出格式化 ───────────────────────────────────────

def format_json_output(books):
    """JSON 输出"""
    return json.dumps(books, ensure_ascii=False, indent=2)


def format_markdown_output(books, sbti_query=None):
    """Markdown 输出——着重标注层级、IP衍生品、海外出圈"""
    sbti_note = ""
    if sbti_query:
        sbti_note = f"\n> 🎯 已按你的性格「{sbti_query}」筛选推荐\n"

    lines = [
        f"# 📚 经典网文推荐 | 起点万订精选",
        "",
        f"> 本期从起点万订/十万均订经典中精选 {len(books)} 本好书，附 SBTI 人格鉴定与 IP 衍生信息",
        sbti_note,
    ]

    for i, book in enumerate(books, 1):
        tier = book.get("tier", "")
        sbti_tag = ""
        if book.get("sbti_code"):
            sbti_tag = f" 🏷️ SBTI：{book['sbti_code']}（{book['sbti_name']}）"

        lines.append(f"## {i}. {tier} 《{book['book_name']}》{sbti_tag}")
        lines.append("")
        lines.append("| 项目 | 信息 |")
        lines.append("|------|------|")
        lines.append(f"| 作者 | {book.get('author', '未知')} |")
        lines.append(f"| 分类 | {book.get('category', '未知')} |")
        if book.get("word_count"):
            lines.append(f"| 字数 | {book['word_count']} |")
        if book.get("collect"):
            lines.append(f"| 收藏 | {book['collect']} |")
        if book.get("status"):
            lines.append(f"| 状态 | {book['status']} |")
        if book.get("qidian_url"):
            lines.append(f"| 起点链接 | [点击阅读]({book['qidian_url']}) |")
        lines.append("")

        # SBTI 必看理由
        if book.get("sbti_reason"):
            lines.append("**🎭 必看理由：**")
            lines.append(f"> {book['sbti_reason']}")
            lines.append("")

        # IP 衍生品信息
        if book.get("adaptations"):
            lines.append("**🎬 IP 衍生作品（中国IP希望之星）：**")
            lines.append("")
            for ad in book["adaptations"]:
                lines.append(f"- **{ad['type']}** | {ad['title']} ({ad['year']}) | {ad['platform']} — {ad['note']}")
            lines.append("")

        # 海外出圈信息
        if book.get("overseas_status"):
            lines.append(f"**{book['overseas_status']}**")
            lines.append("")
            if book.get("overseas_highlights"):
                for h in book["overseas_highlights"]:
                    lines.append(f"- {h}")
                lines.append("")
            if book.get("overseas_forums"):
                lines.append("**🔗 海外社区：**")
                for f_info in book["overseas_forums"]:
                    lines.append(f"- [{f_info['name']}]({f_info['url']})")
                lines.append("")

        # 简介
        intro = book.get("intro", "暂无简介")
        if len(intro) > 500:
            intro = intro[:500] + "…"
        lines.append("**📖 简介：**")
        lines.append(f"> {intro}")
        lines.append("")
        lines.append("---")
        lines.append("")

    lines.append(f"*数据来源：起点图 (qidiantu.com) | SBTI人格鉴定纯属娱乐 | 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}*")
    return "\n".join(lines)


# ─── 环境检测 ─────────────────────────────────────────

def setup_environment():
    """完整环境检测"""
    print("=" * 50, file=sys.stderr)
    print("🔍 起点经典推荐 - 环境检测", file=sys.stderr)
    print("=" * 50, file=sys.stderr)

    v = sys.version_info
    print(f"\n📌 Python: {v.major}.{v.minor}.{v.micro} ({sys.executable})", file=sys.stderr)

    print(f"\n📌 依赖检测:", file=sys.stderr)
    ok = check_and_install_deps(verbose=True)

    # openpyxl (可选)
    print(f"\n📌 可选依赖:", file=sys.stderr)
    try:
        import openpyxl
        print(f"  ✅ openpyxl 已安装（Excel导入可用）", file=sys.stderr)
    except ImportError:
        print(f"  ⚠️ openpyxl 未安装（Excel导入不可用，安装: pip3 install openpyxl）", file=sys.stderr)

    print(f"\n📌 数据状态:", file=sys.stderr)
    # 预置数据
    if os.path.exists(PRESET_FILE):
        fsize = os.path.getsize(PRESET_FILE)
        try:
            with open(PRESET_FILE, "r", encoding="utf-8") as f:
                pd = json.load(f)
            print(f"  📦 预置数据: {pd.get('book_count', '?')} 本 "
                  f"(v{pd.get('preset_version', '?')}) [{fsize/1024:.1f} KB]", file=sys.stderr)
        except Exception:
            print(f"  📦 预置数据: {PRESET_FILE} ({fsize/1024:.1f} KB)", file=sys.stderr)
    else:
        print(f"  📦 预置数据: 未找到（首次全量抓取可能较慢）", file=sys.stderr)
    # 用户缓存
    if os.path.exists(CLASSIC_CACHE_FILE):
        fsize = os.path.getsize(CLASSIC_CACHE_FILE)
        try:
            with open(CLASSIC_CACHE_FILE, "r", encoding="utf-8") as f:
                cd = json.load(f)
            age = ""
            fetched = cd.get("fetched_at", "")
            if fetched:
                try:
                    fetched_dt = datetime.strptime(fetched, "%Y-%m-%d %H:%M:%S")
                    days = (datetime.now() - fetched_dt).days
                    age = f"（{days}天前更新，{'⚠️ 已过期' if days > CACHE_MAX_AGE_DAYS else '✅ 有效'}）"
                except ValueError:
                    pass
            print(f"  📄 用户缓存: {cd.get('book_count', '?')} 本 "
                  f"[{fsize/1024:.1f} KB] {age}", file=sys.stderr)
        except Exception:
            print(f"  📄 用户缓存: {CLASSIC_CACHE_FILE} ({fsize/1024:.1f} KB)", file=sys.stderr)
    else:
        print(f"  📄 用户缓存: 尚未创建（首次运行自动从预置数据初始化）", file=sys.stderr)

    print(f"\n{'=' * 50}", file=sys.stderr)
    if ok:
        print("✅ 环境检测通过！", file=sys.stderr)
    else:
        print("❌ 有依赖未安装，请参考上方提示", file=sys.stderr)
    return ok


# ─── 主流程 ───────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="起点经典网文推荐（万订/十万均订）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  %(prog)s                          # 默认推荐3本经典（首次秒加载预置数据）
  %(prog)s --count 5                # 推荐5本
  %(prog)s --sbti MALO              # 按 SBTI 推荐
  %(prog)s --tier 100k              # 只推十万均订
  %(prog)s --tier 10k               # 只推万订
  %(prog)s --check-update           # 增量检查新增书（秒级，推荐日常使用）
  %(prog)s --refresh                # 全量重新抓取（耗时较长）
  %(prog)s --import-excel FILE      # 从Excel导入
  %(prog)s --dump-cache             # 输出缓存
        """
    )
    parser.add_argument("--count", type=int, default=3, help="推荐数量（默认3本）")
    parser.add_argument("--output", type=str, default="json", choices=["json", "markdown"],
                        help="输出格式（默认json）")
    parser.add_argument("--sbti", type=str, default=None, help="按 SBTI 人格筛选")
    parser.add_argument("--tier", type=str, default=None, choices=["100k", "10k"],
                        help="按层级筛选：100k=十万均订, 10k=万订")
    parser.add_argument("--refresh", action="store_true", help="强制全量刷新缓存（重新抓取所有数据）")
    parser.add_argument("--check-update", action="store_true",
                        help="增量检查更新（仅抓列表页对比新增，秒级完成）")
    parser.add_argument("--import-excel", type=str, default=None, metavar="FILE",
                        help="从Excel文件导入万订全量数据")
    parser.add_argument("--dump-cache", action="store_true", help="输出完整缓存")
    parser.add_argument("--no-save", action="store_true", help="不保存历史")
    parser.add_argument("--setup", action="store_true", help="环境检测")
    args = parser.parse_args()

    if args.setup:
        ok = setup_environment()
        sys.exit(0 if ok else 1)

    if not HAS_BS4:
        print("[ERROR] 核心依赖未安装: pip3 install beautifulsoup4 lxml", file=sys.stderr)
        sys.exit(1)

    # Excel 导入模式
    if args.import_excel:
        books = import_from_excel(args.import_excel)
        print(f"[OK] 导入完成，经典书库共 {len(books)} 本", file=sys.stderr)
        sys.exit(0)

    # 获取数据
    books = get_classic_books(force_refresh=args.refresh,
                              check_update=args.check_update)
    if not books:
        print("[ERROR] 无法获取经典书库数据", file=sys.stderr)
        sys.exit(1)
    print(f"[INFO] 经典书库: {len(books)} 本", file=sys.stderr)

    # dump-cache 模式
    if args.dump_cache:
        print(json.dumps(books, ensure_ascii=False, indent=2))
        sys.exit(0)

    # SBTI 筛选
    sbti_filter = None
    if args.sbti:
        sbti_filter = resolve_sbti_filter(args.sbti)
        if sbti_filter:
            names = [f"{p['code']}({p['name']})" for p in sbti_filter]
            print(f"[SBTI] 匹配: {', '.join(names)}", file=sys.stderr)

    # 历史去重
    history = load_history()
    last_ids = history.get("last_book_ids", [])

    # 挑选推荐
    selected = pick_classic_books(books, count=args.count,
                                  exclude_ids=last_ids,
                                  sbti_filter=sbti_filter,
                                  tier_filter=args.tier)
    selected = dedupe_sbti_in_batch(selected)
    print(f"[INFO] 精选 {len(selected)} 本推荐", file=sys.stderr)

    # 保存历史
    if not args.no_save:
        selected_ids = [b.get("book_id", "") for b in selected if b.get("book_id")]
        if selected_ids:
            save_history(selected_ids, datetime.now().strftime("%Y-%m-%d"))

    # 输出
    if args.output == "json":
        print(format_json_output(selected))
    else:
        print(format_markdown_output(selected, sbti_query=args.sbti))


if __name__ == "__main__":
    main()
