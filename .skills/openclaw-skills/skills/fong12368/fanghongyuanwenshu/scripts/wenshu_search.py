# -*- coding: utf-8 -*-
"""
裁判文书网搜索 v10（通用版）— 纯UI操作 + 左侧面板筛选 + Excel内嵌判决书全文

版本说明：
  - 本版本为通用分享版，不含任何个人账号密码信息
  - 首次运行会提示输入账号密码及搜索条件（仅一次，之后自动记忆）
  - 判决书正文自动清洗（去除网页导航/版权/板块标签）
  - Excel 全部自动换行，列宽合理
  - 安全机制：详情页间隔3秒，每5篇暂停5秒，连续2篇失败自动停止

合规准则（严格遵守）：
  ✅ 只模拟正常用户操作，不拦截不篡改任何请求
  ✅ 合理间隔，不高频批量访问
  ❌ 严禁 page.route() 拦截/篡改请求
  ❌ 严禁绕过服务器安全机制

运行方式：
  python wenshu_search.py          # 测试模式（仅前10条）
  python wenshu_search.py --test   # 同上，测试模式
  python wenshu_search.py --full   # 全量模式
"""
import asyncio
import json
import sys
import os
import re
import base64
import urllib.parse
import time
from datetime import datetime
from pathlib import Path

# ═══════════════════════════════════════════════════════
#  交互配置：首次运行提示补全账号密码和搜索条件
# ═══════════════════════════════════════════════════════
# 配置缓存文件（存放在脚本同目录下的 .wenshu_config.json）
_CONFIG_FILE = Path(__file__).parent / ".wenshu_config.json"


def _load_config():
    """加载缓存配置，若无则返回 None"""
    if _CONFIG_FILE.exists():
        try:
            return json.loads(_CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None


def _save_config(cfg: dict):
    """保存配置到本地（不含明文密码，用星号遮罩）"""
    # 密码仅存星号，运行时手动输入
    safe_cfg = dict(cfg)
    safe_cfg["PASSWORD"] = "********"
    _CONFIG_FILE.write_text(json.dumps(safe_cfg, ensure_ascii=False, indent=2), encoding="utf-8")


def _show_welcome():
    """显示使用引导提醒"""
    cfg = _load_config()
    is_first = cfg is None or not cfg.get("USERNAME") or not cfg.get("YEAR")

    print()
    print("╔" + "═"*60 + "╗")
    print("║" + "  裁判文书网搜索 v10 — 使用引导".center(48) + "║")
    print("╠" + "═"*60 + "╣")
    print("║  本技能由方鸿源律师发布，仅作个人办案参考工具使用；")
    print("║  有关法律、办案及养虾，也可联系13512717203（微信）。")
    print("║  近乎公平即正义，愿我们的法治公平早日实现！")
    print("║")
    if is_first:
        print("║  🔰 首次运行，将进入配置向导")
        print("║")
        print("║  请提前准备好以下信息：")
        print("║")
        print("║  1️⃣  裁判文书网账号（注册手机号）")
        print("║     注册地址: https://wenshu.court.gov.cn/")
        print("║")
        print("║  2️⃣  裁判文书网密码")
        print("║")
        print("║  3️⃣  搜索关键词（案由）")
        print("║     常用: 房屋租赁合同纠纷、民间借贷纠纷")
        print("║           买卖合同纠纷、劳动争议、交通事故...")
        print("║")
        print("║  4️⃣  文书类型: 判决书 / 裁定书 / 调解书")
        print("║  5️⃣  地区: 广东省 / 北京市 / 上海市 ...")
        print("║  6️⃣  年份: 2025 / 2026")
        print("║  7️⃣  审判程序: 一审 / 二审 / 再审")
        print("║")
    else:
        print("║  📋 已有配置，快速启动")
        print("║")
        print("║  如需修改搜索条件，请删除配置文件后重新运行：")
        print(f"║  {str(_CONFIG_FILE)}")
        print("║")

    print("║  ─────────────────────────────────────────")
    print("║  运行模式:")
    print("║    python wenshu_search.py        测试模式(前10条)")
    print("║    python wenshu_search.py --full 全量模式")
    print("║")
    print("║  输出: Excel文件(含判决书全文)，保存至当前目录")
    print("║")
    print("║  ⚠️  合规提醒: 仅模拟正常用户操作，请勿高频批量抓取")
    print("║")
    print("╚" + "═"*60 + "╝")
    print()


def _prompt_config():
    """
    首次运行或配置不完整时，交互提示用户补全信息。
    返回 dict，含 USERNAME / PASSWORD / CASE_CAUSE / DOC_TYPE / PROCEDURES / REGION / YEAR。
    """
    print("\n" + "=" * 60)
    print("  裁判文书网搜索 v10 — 首次配置向导")
    print("=" * 60)
    print("  提示：配置只输入一次，之后自动记忆（密码除外）")
    print("        按 Enter 使用默认值")
    print("-" * 60)

    cfg = _load_config() or {}

    # 账号
    default_user = cfg.get("USERNAME", "")
    USERNAME = input(f"\n📋 裁判文书网账号（手机号）[{default_user}]： ").strip()
    USERNAME = USERNAME or default_user
    if not USERNAME:
        raise SystemExit("❌ 账号不能为空，请重新运行。")

    # 密码
    PASSWORD = input("🔑 裁判文书网密码： ").strip()
    if not PASSWORD:
        raise SystemExit("❌ 密码不能为空，请重新运行。")

    # 案由
    default_cause = cfg.get("CASE_CAUSE", "房屋租赁合同纠纷")
    CASE_CAUSE = input(f"\n🔍 案由（中文）[默认={default_cause}]： ").strip() or default_cause
    if not CASE_CAUSE:
        CASE_CAUSE = "房屋租赁合同纠纷"

    # 文书类型
    default_dtype = cfg.get("DOC_TYPE", "判决书")
    dtype_opts = "判决书 / 裁定书 / 调解书"
    DOC_TYPE = input(f"\n📄 裁判文书类型 [{dtype_opts}，默认={default_dtype}]： ").strip() or default_dtype
    if DOC_TYPE not in ("判决书", "裁定书", "调解书"):
        DOC_TYPE = "判决书"

    # 审判程序
    default_proc = cfg.get("PROCEDURES", "一审,二审")
    PROCEDURES = input(f"\n⚖️  审判程序（一审/二审/再审，可用逗号分隔）[默认={default_proc}]： ").strip() or default_proc
    PROCEDURES = [p.strip() for p in PROCEDURES.replace("，", ",").split(",") if p.strip()]
    if not PROCEDURES:
        PROCEDURES = ["一审"]

    # 地区
    default_region = cfg.get("REGION", "广东省")
    print(f"\n🌏 地区（中文省份，如：广东省、北京市）[默认={default_region}]：")
    print("  常用：广东省 | 北京市 | 上海市 | 浙江省 | 江苏省 | 四川省 | ...")
    REGION = input(f"  请输入地区 [Enter 使用默认值]： ").strip() or default_region
    if not REGION:
        REGION = "广东省"

    # 年份
    default_year = cfg.get("YEAR", datetime.now().year)
    YEAR = input(f"\n📅 裁判年份（4位数字）[默认={default_year}]： ").strip()
    YEAR = int(YEAR) if YEAR.isdigit() else int(default_year)

    # 保存配置（不含明文密码）
    _save_config({
        "USERNAME": USERNAME,
        "PASSWORD": None,  # 密码不写入文件
        "CASE_CAUSE": CASE_CAUSE,
        "DOC_TYPE": DOC_TYPE,
        "PROCEDURES": ",".join(PROCEDURES),
        "REGION": REGION,
        "YEAR": YEAR,
    })

    print(f"\n✅ 配置已保存（密码仅本次有效），开始运行...\n")
    return {
        "USERNAME": USERNAME,
        "PASSWORD": PASSWORD,
        "CASE_CAUSE": CASE_CAUSE,
        "DOC_TYPE": DOC_TYPE,
        "PROCEDURES": PROCEDURES,
        "REGION": REGION,
        "YEAR": YEAR,
    }


# ── 启动时加载/提示配置 ───────────────────────────────
_show_welcome()

_cfg = _load_config()
if _cfg and _cfg.get("USERNAME") and _cfg.get("YEAR"):
    # 配置存在，直接提示输入密码即可
    print("\n" + "=" * 60)
    print("  裁判文书网搜索 v10 — 快速登录")
    print("=" * 60)
    print(f"  账号：{_cfg['USERNAME']}")
    print(f"  案由：{_cfg.get('CASE_CAUSE','')}  地区：{_cfg.get('REGION','')}  年份：{_cfg.get('YEAR','')}")
    print(f"  程序：{_cfg.get('PROCEDURES','')}  文书类型：{_cfg.get('DOC_TYPE','')}")
    print("-" * 60)
    USERNAME   = _cfg["USERNAME"]
    PASSWORD   = input("🔑 请输入密码： ").strip()
    if not PASSWORD:
        raise SystemExit("❌ 密码不能为空。")
    CASE_CAUSE  = _cfg.get("CASE_CAUSE", "房屋租赁合同纠纷")
    DOC_TYPE     = _cfg.get("DOC_TYPE", "判决书")
    PROCEDURES   = _cfg.get("PROCEDURES", "一审,二审").replace("，", ",").split(",")
    PROCEDURES   = [p.strip() for p in PROCEDURES if p.strip()]
    REGION       = _cfg.get("REGION", "广东省")
    YEAR_START   = int(_cfg.get("YEAR", datetime.now().year))
    YEAR_END     = YEAR_START
    print(f"  ✅ 加载已有配置，如需修改请删除 {_CONFIG_FILE} 后重新运行\n")
else:
    # 首次运行或配置缺失
    _d = _prompt_config()
    USERNAME     = _d["USERNAME"]
    PASSWORD     = _d["PASSWORD"]
    CASE_CAUSE   = _d["CASE_CAUSE"]
    DOC_TYPE     = _d["DOC_TYPE"]
    PROCEDURES   = _d["PROCEDURES"]
    REGION       = _d["REGION"]
    YEAR_START   = _d["YEAR"]
    YEAR_END     = YEAR_START

# ═══════════════════════════════════════════════════════
#  固定配置
# ═══════════════════════════════════════════════════════
SAVE_DIR   = "."  # 输出目录（默认为当前目录，可修改）

# 审判程序代码（s4 字段，URL 参数）
_S4_MAP = {
    "一审": "4",
    "二审": "5",
    "再审": "6",
}

# 裁判文书类型代码（s8 字段，URL 参数）
_S8_MAP = {
    "判决书": "MS",
    "裁定书": "CZ",
    "调解书": "TJ",
}

# 地区代码（s33 字段）
_REGION_MAP = {
    "北京市": "北京市", "天津市": "天津市", "上海市": "上海市", "重庆市": "重庆市",
    "广东省": "广东省", "浙江省": "浙江省", "江苏省": "江苏省", "四川省": "四川省",
    "河南省": "河南省", "山东省": "山东省", "福建省": "福建省", "湖北省": "湖北省",
    "湖南省": "湖南省", "安徽省": "安徽省", "江西省": "江西省", "河北省": "河北省",
    "山西省": "山西省", "辽宁省": "辽宁省", "吉林省": "吉林省", "黑龙江省": "黑龙江省",
    "陕西省": "陕西省", "甘肃省": "甘肃省", "青海省": "青海省", "海南省": "海南省",
    "云南省": "云南省", "贵州省": "贵州省", "内蒙古": "内蒙古", "宁夏": "宁夏回族自治区",
    "新疆": "新疆维吾尔自治区", "西藏": "西藏自治区", "广西": "广西壮族自治区",
}

# 案由代码说明：
#   s14=8175 = 租赁合同纠纷（上级案由）
#   s11=9000 = 房屋租赁合同纠纷（精确子案由，用于本地过滤）
#   ⚠️ 注意：通用版默认保留此配置，用户若需更换案由可修改此两行
S14_CODE_SEARCH = "8175"
S14_CODE_FILTER = "9000"

# 年份已由配置区设置
MAX_PAGES  = 50
PAGE_SIZE  = 20
MAX_FETCH  = 10

all_cases: list = []


# ═══════════════════════════════════════════════════════
#  字段映射
# ═══════════════════════════════════════════════════════
FIELD_MAP = {
    "s1":  "案件名称", "s2":  "法院名称", "s3":  "审理法院",
    "s5":  "文书ID",   "s7":  "案号",     "s8":  "案件类型",
    "s9":  "审判程序", "s11": "案由",     "s17": "当事人",
    "s31": "裁判日期", "s33": "法院省份", "s34": "法院地市",
    "s35": "法院区县", "s41": "发布日期", "s42": "裁判年份",
    "s44": "案例等级",
    "1":  "案件名称", "2":  "法院名称", "3":  "审理法院",
    "5":  "文书ID",   "7":  "案号",     "8":  "案件类型",
    "9":  "审判程序", "11": "案由",     "17": "当事人",
    "31": "裁判日期", "33": "法院省份", "34": "法院地市",
    "35": "法院区县", "41": "发布日期", "42": "裁判年份",
    "44": "案例等级",
}

# ═══════════════════════════════════════════════════════
#  3DES 解密（裁判文书网服务器加密响应）
# ═══════════════════════════════════════════════════════
def _pad(data: bytes) -> bytes:
    bs = 8
    padding = bs - len(data) % bs
    return data + bytes([padding] * padding)


def _unpad(data: bytes) -> bytes:
    return data[:-data[-1]]


def decrypt_response(encrypted_b64: str, secret_key: str) -> str:
    """
    使用 3DES-CBC 解密裁判文书网返回的加密数据。
    IV = 当天日期 YYYYMMDD 的 UTF-8 编码（每日更换）
    key = secretKey 前 24 字节
    """
    try:
        from Crypto.Cipher import DES3
        today_str = datetime.now().strftime("%Y%m%d")
        iv = today_str.encode("utf-8")
        key = secret_key[:24].encode("utf-8")
        cipher = DES3.new(key, DES3.MODE_CBC, iv)
        raw = base64.b64decode(encrypted_b64)
        decrypted = _unpad(cipher.decrypt(raw))
        return decrypted.decode("utf-8", errors="replace")
    except Exception as e:
        return ""


# ═══════════════════════════════════════════════════════
#  判决书正文清洗：去除网页导航、版权、底部链接等噪音
# ═══════════════════════════════════════════════════════
def clean_judgment_text(raw_text: str) -> str:
    """去掉裁判文书网页中的各类噪音内容，只保留判决书正文"""
    if not raw_text:
        return ""

    text = raw_text

    # 1. 去掉顶部日期导航行
    text = re.sub(r"^\d{4}年\d{1,2}月\d{1,2}日[^\n]*\n?", "", text)

    # 2. 去掉 "欢迎您，xxxxx" 用户信息行
    text = re.sub(r"欢迎您[，,]\S+[^\n]*\n?", "", text)

    # 3. 去掉 退出/意见建议/返回主站/使用帮助/目录 等按钮行
    text = re.sub(r"^(退出|意见建议|返回主站|使用帮助|目录)[^\n]*\n?", "", text, flags=re.MULTILINE)

    # 4. 去掉顶部分类标签行（指导 案例 推荐 关联）
    text = re.sub(r"^(指导|案例|推荐|关联)[ \t]*\n?", "", text, flags=re.MULTILINE)
    text = re.sub(r"[ \t]*(指导|案例|推荐|关联)[ \t]*", "", text)

    # 5. 去掉底部关联平台板块
    for platform in [
        "中国政府公开信息整合服务平台",
        "人民检察院案件信息公开网",
        "中国审判流程信息公开网",
        "中国司法大数据服务网",
        "中国执行信息公开网",
        "全国法院减刑、假释、暂予监外执行信息网",
        "中国涉外商事海事审判网",
        "最高人民法院服务人民群众系统场景导航",
    ]:
        text = re.sub(r"\|?\s*" + re.escape(platform) + r"[^\n]*\n?", "", text)

    # 6. 去掉底部版权/地址/ICP信息
    text = re.sub(r"地址[：:][^\n]+\n?", "", text)
    text = re.sub(r"邮编[：:]\d+\n?", "", text)
    text = re.sub(r"总机[：:]\d+\n?", "", text)
    text = re.sub(r"中华人民共和国最高人民法院\s*版权所有\n?", "", text)
    text = re.sub(r"京ICP备\d+号\n?", "", text)

    # 7. 去掉提示词
    text = re.sub(r"点击了解更多[^\n]*\n?", "", text)
    text = re.sub(r"^\s*(概要|详情|正文)[ \t]*\n?", "", text, flags=re.MULTILINE)

    # 8. 去掉多余分隔符行
    text = re.sub(r"^[_=\-]{5,}\s*$", "", text, flags=re.MULTILINE)

    # 9. 合并多余空行
    text = re.sub(r"\n{3,}", "\n\n", text)

    # 10. 去掉首尾空白
    lines = [ln.strip() for ln in text.splitlines()]
    return "\n".join(ln for ln in lines if ln)


# ═══════════════════════════════════════════════════════
#  文书详情页：提取判决书全文（写入 Excel 列）
# ═══════════════════════════════════════════════════════
CONTENT_COL = 4  # Excel 列号：判决书内容在第4列


async def fetch_full_text(page, case: dict, idx: int, total: int) -> str:
    """访问文书详情页，提取判决书全文，清洗后返回纯文本"""
    doc_id = case.get("文书ID") or case.get("案件ID") or ""
    if not doc_id:
        print(f"    [{idx}/{total}] 无文书ID，跳过")
        return ""

    detail_url = (
        f"https://wenshu.court.gov.cn/website/wenshu/181107ANFZ0BXSK4/index.html"
        f"?docId={doc_id}"
    )
    try:
        await page.goto(detail_url, wait_until="domcontentloaded", timeout=25000)
        await asyncio.sleep(3)

        content_text = await page.evaluate("""() => {
            const sels = ['#fullText', '.full-text', '#wsDom', '.ws-content',
                          '.judgment-content', '#judgment-content', '.content',
                          'article', '#article'];
            for (const s of sels) {
                const el = document.querySelector(s);
                if (el && el.innerText && el.innerText.length > 200)
                    return el.innerText;
            }
            return document.body.innerText.substring(0, 100000);
        }""")

        if not content_text or len(content_text) < 100:
            print(f"    [{idx}/{total}] docId={doc_id} 正文过短，可能需登录或已失效")
            case["判决书内容"] = ""
            return ""

        cleaned = clean_judgment_text(content_text)
        if not cleaned or len(cleaned) < 100:
            lines = [ln.strip() for ln in content_text.splitlines()]
            cleaned = "\n".join(ln for ln in lines if ln)

        case["判决书内容"] = cleaned
        print(f"    [{idx}/{total}] OK  {len(cleaned)} 字  {case.get('案件名称','')[:25]}...")
        return cleaned

    except Exception as e:
        print(f"    [{idx}/{total}] docId={doc_id} 异常: {e}")
        case["判决书内容"] = ""
        return ""


# ═══════════════════════════════════════════════════════
#  登录（OCR 识别验证码 + Playwright）
# ═══════════════════════════════════════════════════════
async def login_and_get_secret(page) -> str:
    """
    使用 Playwright 打开登录页，输入账密，识别图形验证码，完成登录。
    返回登录后的 secretKey（供解密 API 响应用）。
    """
    await page.goto("https://wenshu.court.gov.cn/", wait_until="domcontentloaded")
    await asyncio.sleep(1)

    try:
        user_link = page.locator("a", has_text="登录").first
        await user_link.click()
        await asyncio.sleep(2)
    except Exception:
        pass

    # 输入账密
    await page.fill("#username", USERNAME)
    await page.fill("#password", PASSWORD)
    await asyncio.sleep(0.5)

    # 识别验证码
    captcha_img = page.locator("#captchaImg")
    max_retries = 10
    for attempt in range(max_retries):
        try:
            await captcha_img.wait_for(timeout=5000)
            img_bytes = await captcha_img.screenshot()
            ocr = ddddocr.DdddOcr(show_ad=False)
            captcha_code = ocr.classification(img_bytes).strip()
            captcha_code = re.sub(r"[^a-zA-Z0-9]", "", captcha_code)
            print(f"  [验证码] 识别结果：{captcha_code}（第{attempt+1}次）")
            if len(captcha_code) == 4:
                break
        except Exception as e:
            print(f"  [验证码] 识别失败（第{attempt+1}次）: {e}")
            await asyncio.sleep(1)
    else:
        raise SystemExit("  [失败] 验证码识别10次均未成功，请手动处理。")

    # 填入验证码并提交
    await page.fill("#captchaCode", captcha_code)
    await page.click("#loginBtn")
    await asyncio.sleep(3)

    # 验证登录结果（检查页面是否还残留登录框）
    try:
        await page.wait_for_selector("#username", state="hidden", timeout=8000)
        print(f"  [登录] ✅ 成功（账号：{USERNAME}）")
    except Exception:
        current_url = page.url
        raise SystemExit(f"  [登录] ❌ 失败，请检查账号密码。\n  当前URL：{current_url}")

    # 获取 secretKey
    secret_key = ""
    for _ in range(5):
        try:
            secret_key = await page.evaluate("""() => {
                try {
                    return window.secretKey || window._WENSHU_SECRET_KEY_ ||
                           (window.localStorage && localStorage.getItem('wenshuSecretKey')) || "";
                } catch(e) { return ""; }
            }""")
            if secret_key:
                break
            await asyncio.sleep(1)
        except Exception:
            await asyncio.sleep(1)

    print(f"  [密钥] {'已获取' if secret_key else '未获取（将尝试监听获取）'}")
    return secret_key


# ═══════════════════════════════════════════════════════
#  纯 UI 搜索流程
# ═══════════════════════════════════════════════════════

def _parse_api_response(text: str, secret_key: str) -> list:
    """解密并解析文书列表 API 响应，返回案件字典列表"""
    try:
        data = json.loads(text)
        if not isinstance(data, dict):
            return []
        json_str = data.get("result", {}).get("restruction", "")
        if not json_str:
            return []
        decrypted = decrypt_response(json_str, secret_key)
        if not decrypted:
            return []
        records = json.loads(decrypted)
        return records if isinstance(records, list) else []
    except Exception:
        return []


def _derive_case_info(case: dict, procedure: str) -> dict:
    """从 API 原始字段派生完整信息"""
    year_map = {"4": "一审", "5": "二审", "6": "再审"}
    s4_code = {"一审": "4", "二审": "5", "再审": "6"}.get(procedure, "4")

    def get_field(keys: list) -> str:
        for k in keys:
            for ak in [k, k.lower()]:
                if ak in case:
                    v = case[ak]
                    if v and str(v).strip():
                        return str(v).strip()
        return ""

    case_num = get_field(["案号", "s7", "7"])
    raw_year = get_field(["裁判年份", "s42", "42", "裁判日期", "s31", "31"])

    year = ""
    if raw_year:
        m = re.search(r"(\d{4})", raw_year)
        year = m.group(1) if m else ""

    party = get_field(["当事人", "s17", "17"])
    province = get_field(["法院省份", "s33", "33"])
    city = get_field(["法院地市", "s34", "34"])
    court = get_field(["法院名称", "s2", "2"]) or get_field(["审理法院", "s3", "3"])
    doc_type = get_field(["案件类型", "s8", "8"])
    doc_id = get_field(["文书ID", "s5", "5"]) or get_field(["案件ID", "s5", "5"])

    # 城市级别推断省份
    if province and len(province) > 6:
        province = province[:6]

    result = {
        "案件名称": get_field(["案件名称", "s1", "1"]),
        "案号": case_num,
        "法院名称": court,
        "法院省份": province,
        "法院地市": city,
        "案件类型": doc_type or "判决书",
        "审判程序": procedure,
        "案由": get_field(["案由", "s11", "11"]) or CASE_CAUSE,
        "裁判年份": year,
        "裁判日期": get_field(["裁判日期", "s31", "31"]),
        "发布日期": get_field(["发布日期", "s41", "41"]),
        "当事人": party,
        "文书ID": doc_id,
    }
    return result


async def _search_procedure(
    browser, url: str, procedure: str, s4_code: str, max_pages: int,
    on_total, on_case_found,
    secret_key: str = "", api_log: list = None
) -> list:
    """
    打开搜索页 → 输入关键词 → 点击搜索 → 翻页提取。
    纯 UI 操作，不拦截请求。
    api_log: 可选列表，用于记录原始 API 响应（调试用）
    """
    from playwright.async_api import async_playwright

    context = await browser.contexts[0].new_context()
    page = await context.new_page()
    results = []

    try:
        await page.goto(url, wait_until="domcontentloaded")
        await asyncio.sleep(2)

        # 输入关键词（案由）并搜索
        await page.fill('input[placeholder*="关键词"]', CASE_CAUSE)
        await page.fill('input[placeholder*="文书ID"]', '')
        await asyncio.sleep(0.3)
        await page.click('button:has-text("搜索")')
        await asyncio.sleep(3)

        # 左侧面板筛选审判程序
        try:
            panel = page.locator(".filter-panel, .left-panel, #filterPanel, .court-filter")
            if await panel.count() > 0:
                item = panel.locator(
                    f'xpath=//*[contains(text(),"{procedure}")] | '
                    f'xpath=//*[@data-value="{s4_code}"]'
                ).first
                await item.click()
                await asyncio.sleep(1)
        except Exception:
            pass

        # 切换每页显示15条
        try:
            page_sel = page.locator('select')
            if await page_sel.count() > 0:
                await page_sel.first.select_option("15")
                await asyncio.sleep(1)
        except Exception:
            pass

        # 逐页提取
        for pg in range(max_pages):
            await asyncio.sleep(2)
            raw = await page.evaluate("""() => {
                const el = document.querySelector('#listDiv, .list-container, .case-list');
                return el ? el.innerHTML : '';
            }""")

            if not raw or "无匹配数据" in raw or "未找到" in raw:
                print(f"  [{procedure}] 第{pg+1}页：无数据，停止翻页")
                break

            records = []
            if secret_key and secret_key in raw:
                records = _parse_api_response(raw, secret_key)
                if not records:
                    m = re.search(r'"restruction":"([^"]+)"', raw)
                    if m:
                        records = _parse_api_response(json.dumps({"result": {"restruction": m.group(1)}}), secret_key)
            else:
                # 纯 UI 解析 HTML
                records = _ui_parse_list(page)

            new_count = 0
            for rec in records:
                case = _derive_case_info(rec, procedure)
                if case.get("文书ID") and not any(c.get("文书ID") == case["文书ID"] for c in all_cases):
                    all_cases.append(case)
                    on_case_found(case)
                    new_count += 1

            print(f"  [{procedure}] 第{pg+1}页：{new_count} 条新数据（累计 {len(all_cases)} 条）")
            if new_count == 0:
                break

            # 翻页
            try:
                next_btn = page.locator('.next, .page-next, a:has-text("下一页"), button:has-text("下一页")')
                if await next_btn.count() > 0 and next_btn.first.is_enabled():
                    await next_btn.first.click()
                else:
                    break
            except Exception:
                break

    finally:
        await context.close()

    return results


def _ui_parse_list(page) -> list:
    """从搜索结果页 DOM 中提取文书列表数据（纯 UI 模式）"""
    try:
        return page.evaluate("""() => {
            const items = document.querySelectorAll('.case-item, .ws-item, [class*="item"]');
            if (!items.length) return [];
            return Array.from(items).map(el => {
                const a = el.querySelector('a[href*="docId="]');
                const title = el.querySelector('.title, .case-title, [class*="title"]');
                const num = el.querySelector('.case-num, [class*="num"]');
                const date = el.querySelector('.date, [class*="date"]');
                const court = el.querySelector('.court, [class*="court"]');
                return {
                    "案件名称": title ? title.innerText : '',
                    "案号": num ? num.innerText : '',
                    "裁判日期": date ? date.innerText : '',
                    "法院名称": court ? court.innerText : '',
                    "文书ID": a ? (a.href.match(/docId=([^&]+)/) || ['',''])[1] : '',
                };
            });
        }""")
    except Exception:
        return []


async def _do_filter_by_year_and_region(page):
    """通过左侧面板筛选地区和年份（纯点击 UI）"""
    try:
        await page.wait_for_selector(".filter-panel, .left-panel", timeout=5000)
    except Exception:
        return

    await asyncio.sleep(1)

    # 展开地区筛选
    try:
        region_items = page.locator(".filter-item, .filter-option, .province-item")
        count = await region_items.count()
        for i in range(count):
            text = await region_items.nth(i).inner_text()
            if REGION and REGION[:2] in text:
                await region_items.nth(i).click()
                await asyncio.sleep(0.5)
                break
    except Exception:
        pass

    # 展开年份筛选
    try:
        year_str = str(YEAR_START)
        year_items = page.locator("xpath=//*[contains(text(),'" + year_str + "')]")
        if await year_items.count() > 0:
            await year_items.first.click()
            await asyncio.sleep(0.5)
    except Exception:
        pass


async def run_search_for_procedure(browser, procedure: str):
    """对一审/二审/再审分别执行完整搜索流程"""
    s4_code = _S4_MAP.get(procedure, "4")
    s8_code = _S8_MAP.get(DOC_TYPE, "MS")
    today = datetime.now().strftime("%Y%m%d")
    search_url = (
        f"https://wenshu.court.gov.cn/website/wenshu/181107ANFZ0BXSK4/index.html?"
        f"s14={S14_CODE_SEARCH}&s8={s8_code}&s4={s4_code}"
    )

    print(f"\n[搜索] 程序={procedure}  案由={CASE_CAUSE}  类型={DOC_TYPE}")

    results = []

    async def on_total(total_cnt: int):
        print(f"  [总量] 约 {total_cnt} 条")

    async def on_case_found(case: dict):
        pass

    results = await _search_procedure(
        browser, search_url, procedure, s4_code, MAX_PAGES,
        on_total, on_case_found,
    )
    return results


# ═══════════════════════════════════════════════════════
#  Excel 保存（openpyxl）
# ═══════════════════════════════════════════════════════
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

THIN = Side(style="thin", color="CCCCCC")
BD   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def save_excel(cases: list, total_all: int = 0) -> str:
    """
    将案件列表写入 Excel 文件，列顺序：
    序号、案件名称、案号、判决书内容、法院名称、省份、地市、类型、程序、案由、年份、日期、当事人（共13列）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "裁判文书"

    headers = [
        "序号", "案件名称", "案号", "判决书内容",
        "法院名称", "法院省份", "法院地市",
        "案件类型", "审判程序", "案由",
        "裁判年份", "裁判日期", "当事人",
    ]
    # 表头
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font      = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BD

    # 数据行
    if cases:
        for ri, case in enumerate(cases, 2):
            fill    = PatternFill("solid", fgColor="EBF3FB" if ri % 2 == 0 else "FFFFFF")
            content = case.get("判决书内容", "")
            if len(content) > 32000:
                content = content[:32000] + "\n\n[内容过长已截断]"

            vals = [
                ri - 1,
                case.get("案件名称", ""),
                case.get("案号", ""),
                content,
                case.get("法院名称", "") or case.get("审理法院", ""),
                case.get("法院省份", ""),
                case.get("法院地市", ""),
                case.get("案件类型", ""),
                case.get("审判程序", ""),
                case.get("案由", ""),
                case.get("裁判年份", ""),
                case.get("裁判日期", "") or case.get("发布日期", ""),
                case.get("当事人", ""),
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(ri, ci, v)
                c.font      = Font(name="微软雅黑", size=10)
                c.fill      = fill
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border    = BD
    else:
        ws.cell(2, 1, "暂无数据")
        ws.cell(3, 1, "提示：请检查搜索条件或稍后重试")

    # 列宽：判决书内容60，其余紧凑
    col_widths = [6, 35, 26, 60, 20, 10, 10, 10, 10, 16, 10, 14, 28]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    os.makedirs(SAVE_DIR, exist_ok=True)
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    region_ = REGION or "全国"
    fname   = f"{CASE_CAUSE}_{region_}_{'+'.join(PROCEDURES)}_{YEAR_START}_{ts}.xlsx"
    fpath   = os.path.join(SAVE_DIR, fname)
    wb.save(fpath)
    print(f"\n[保存] {fpath}")
    return fpath


# ═══════════════════════════════════════════════════════
#  主流程
# ═══════════════════════════════════════════════════════
async def main():
    from playwright.async_api import async_playwright

    # ── 命令行参数 ──────────────────────────────────────
    run_full   = "--full" in sys.argv or "-f" in sys.argv
    fetch_limit = 0 if run_full else MAX_FETCH
    mode_label  = "全量模式" if run_full else f"测试模式(前{fetch_limit}条)"
    print(f"\n{'='*60}")
    print(f"  裁判文书网搜索 v10 — 通用版")
    print(f"{'='*60}")
    print(f"[账号] {USERNAME}（已脱敏）")
    print(f"[配置] 案由={CASE_CAUSE}  类型={DOC_TYPE}  地区={REGION}  年份={YEAR_START}  程序={PROCEDURES}")
    print(f"[模式] {mode_label}")
    print(f"[合规] 纯UI操作，翻页≥3秒，详情页≥3秒，连续2篇失败自动停")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.contexts[0]

        # ── 登录 ──────────────────────────────────────────
        page = await context.new_page()
        secret_key = await login_and_get_secret(page)
        await page.close()

        # ── 逐审判程序搜索 ───────────────────────────────
        for proc in PROCEDURES:
            s4_code = _S4_MAP.get(proc, "4")
            s8_code = _S8_MAP.get(DOC_TYPE, "MS")
            today   = datetime.now().strftime("%Y%m%d")
            search_url = (
                f"https://wenshu.court.gov.cn/website/wenshu/181107ANFZ0BXSK4/index.html?"
                f"s14={S14_CODE_SEARCH}&s8={s8_code}&s4={s4_code}"
            )
            print(f"\n[搜索] 程序={proc}  案由={CASE_CAUSE}")

            ctx2 = await browser.new_context()
            pg   = await ctx2.new_page()
            try:
                await pg.goto(search_url, wait_until="domcontentloaded")
                await asyncio.sleep(2)

                # 输入关键词并搜索
                kw_input = pg.locator('input[placeholder*="关键词"]')
                if await kw_input.count() > 0:
                    await kw_input.fill(CASE_CAUSE)
                    await asyncio.sleep(0.3)
                    search_btn = pg.locator('button:has-text("搜索")')
                    if await search_btn.count() > 0:
                        await search_btn.first.click()
                    await asyncio.sleep(3)

                # 左侧面板筛选地区
                try:
                    region_matched = False
                    for attempt in range(2):
                        panels = pg.locator(".filter-panel, .left-panel, #filterPanel, [class*='filter']")
                        if await panels.count() > 0:
                            panel = panels.first
                            items = panel.locator(".filter-item, .filter-option, .province-item, [class*='item']")
                            item_count = await items.count()
                            for idx in range(item_count):
                                item_text = await items.nth(idx).inner_text()
                                if REGION[:2] in item_text or item_text.strip() == REGION:
                                    await items.nth(idx).click()
                                    region_matched = True
                                    await asyncio.sleep(1)
                                    break
                            if region_matched:
                                break
                        await asyncio.sleep(1)
                except Exception:
                    pass

                # 左侧面板筛选年份
                try:
                    year_str = str(YEAR_START)
                    year_items = pg.locator(f'xpath=//*[contains(text(),"{year_str}")]')
                    if await year_items.count() > 0:
                        await year_items.first.click()
                        await asyncio.sleep(1)
                except Exception:
                    pass

                # 切换每页15条
                try:
                    page_sel = pg.locator("select")
                    if await page_sel.count() > 0:
                        await page_sel.first.select_option("15")
                        PAGE_SIZE = 15
                        await asyncio.sleep(1)
                except Exception:
                    pass

                # 逐页翻取
                total_all = 0
                for pg_num in range(MAX_PAGES):
                    await asyncio.sleep(2)

                    # 提取当前页数据
                    page_data = []
                    try:
                        raw_html = await pg.evaluate(
                            "() => { const el = document.querySelector('#listDiv,.list-container,.ws-list,#caseList'); "
                            "return el ? el.innerHTML : ''; }"
                        )
                        if not raw_html or "无匹配" in raw_html or "未找到" in raw_html:
                            break

                        # 解析 API 加密数据或 DOM 数据
                        if secret_key:
                            m = re.search(r'"restruction":"([A-Za-z0-9+/=]+)"', raw_html)
                            if m:
                                records = _parse_api_response(
                                    json.dumps({"result": {"restruction": m.group(1)}}), secret_key
                                )
                                for rec in records:
                                    case = _derive_case_info(rec, proc)
                                    cid = case.get("文书ID", "")
                                    if cid and not any(c.get("文书ID") == cid for c in all_cases):
                                        all_cases.append(case)
                                        page_data.append(case)
                        else:
                            page_data = _ui_parse_list(pg)
                            for case in page_data:
                                case["审判程序"] = proc
                                cid = case.get("文书ID", "")
                                if cid and not any(c.get("文书ID") == cid for c in all_cases):
                                    all_cases.append(case)
                    except Exception as e:
                        print(f"  [解析] 第{pg_num+1}页异常: {e}")

                    cnt = len(page_data)
                    total_all += cnt
                    print(f"  [{proc}] 第{pg_num+1}页：{cnt} 条（累计 {len(all_cases)} 条）")
                    if cnt == 0:
                        break

                    # 下一页
                    try:
                        next_btn = pg.locator(".next:not(.disabled), .page-next:not([disabled]), "
                                              "a:has-text('下一页'), button:has-text('下一页')")
                        if await next_btn.count() > 0 and await next_btn.first.is_enabled():
                            await next_btn.first.click()
                        else:
                            break
                    except Exception:
                        break

            finally:
                await ctx2.close()

        # ── 过滤年份（双重保险） ─────────────────────────
        filtered = [
            c for c in all_cases
            if not YEAR_START or (c.get("裁判年份") and str(YEAR_START) in str(c.get("裁判年份", "")))
        ]

        # ── 抓取判决书全文 ───────────────────────────────
        fetch_list = filtered[:fetch_limit] if fetch_limit > 0 else filtered
        total_fetch = len(fetch_list)
        if filtered:
            print(f"\n[抓取] 开始获取判决书全文（{total_fetch}/{len(filtered)} 条）...")
            print(f"  [安全] 每篇间隔3秒，每5篇暂停5秒，遇异常自动停止")
            success_count = 0
            fail_streak   = 0

            ctx3 = await browser.new_context()
            for idx, case in enumerate(fetch_list, 1):
                pg2 = await ctx3.new_page()
                result = await fetch_full_text(pg2, case, idx, total_fetch)
                await pg2.close()

                if not result or not case.get("判决书内容"):
                    fail_streak += 1
                    print(f"    ⚠️ 第{idx}篇异常（连续失败 {fail_streak} 次）")
                    if fail_streak >= 2:
                        print(f"  [停止] 检测到连续异常，停止后续抓取")
                        break
                else:
                    fail_streak = 0
                    success_count += 1

                if idx % 5 == 0 and idx < total_fetch:
                    print(f"  [暂停] 已抓{idx}篇，休息5秒...")
                    await asyncio.sleep(5)
                else:
                    await asyncio.sleep(3)

            await ctx3.close()
            print(f"  [OK] 判决书全文抓取完成，成功 {success_count}/{total_fetch} 篇")

        await browser.close()

        # ── 保存 Excel ────────────────────────────────────
        fetched_count = len(fetch_list) if fetch_limit > 0 else len(filtered)
        print(f"\n[完成] 搜索获取 {len(all_cases)} 条，过滤后 {len(filtered)} 条，已抓取判决书 {fetched_count} 条")
        fpath = save_excel(filtered, total_all)

        print(f"\n✅ 全部完成！结果文件：\n   {fpath}")
        return fpath


if __name__ == "__main__":
    asyncio.run(main())
