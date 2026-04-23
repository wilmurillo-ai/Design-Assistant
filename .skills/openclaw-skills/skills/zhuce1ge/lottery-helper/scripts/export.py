"""
lottery-helper/scripts/export.py
将查询到的开奖数据导出为 txt / md / xlsx / sqlite
支持命令行调用：
  python export.py --codes 双色球,大乐透 --periods 50 --fmt xlsx --out C:/Users/vboxuser/Downloads
  python export.py --codes 全部 --periods 100 --fmt all --out C:/Users/vboxuser/Downloads
"""
import os
import sys
import argparse
from datetime import date
from pathlib import Path

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import sqlite3
    SQLITE3_OK = True
except ImportError:
    SQLITE3_OK = False


CODE_NAMES = {
    "ssq": "双色球", "dlt": "大乐透", "pl3": "排列三",
    "pl5": "排列五", "3d": "福彩3D", "kl8": "北京快乐8",
    "qlc": "七乐彩", "qxc": "七星彩",
}
NAME_TO_CODE = {v: k for k, v in CODE_NAMES.items()}

# 各彩种完整字段定义（已实测确认）
FIELD_DEFS = {
    "ssq": [
        "期号", "日期",
        "红1", "红2", "红3", "红4", "红5", "红6", "蓝球",
        "投注总额", "奖池金额",
        "一等奖注数", "一等奖金额",
        "二等奖注数", "二等奖金额",
        "三等奖注数", "三等奖金额",
        "四等奖注数", "四等奖金额",
        "五等奖注数", "五等奖金额",
        "六等奖注数", "六等奖金额",
        "福运奖注数", "福运奖金额",
    ],
    "dlt": [
        "期号", "日期",
        "前区1", "前区2", "前区3", "前区4", "前区5",
        "后区1", "后区2",
        "投注总额", "奖池金额",
        "一等奖注数", "一等奖金额",
        "二等奖注数", "二等奖金额",
        "三等奖注数", "三等奖金额",
        "四等奖注数", "四等奖金额",
        "五等奖注数", "五等奖金额",
        "六等奖注数", "六等奖金额",
        "七等奖注数", "七等奖金额",
        "八等奖注数", "八等奖金额",
        "九等奖注数", "九等奖金额",
        "追加一等奖注数", "追加一等奖金额",
        "追加二等奖注数", "追加二等奖金额",
    ],
    "pl3": [
        "期号", "日期",
        "百", "十", "个",
        "投注总额",
        "直选注数", "直选金额",
        "组三注数", "组三金额",
        "组六注数", "组六金额",
    ],
    "pl5": [
        "期号", "日期",
        "万", "千", "百", "十", "个",
        "投注总额",
        "直选注数", "直选金额",
    ],
    "3d": [
        "期号", "日期",
        "百", "十", "个",
        "投注总额",
        "直选注数", "直选金额",
    ],
    "kl8": [
        "期号", "日期",
        "码1","码2","码3","码4","码5","码6","码7","码8","码9","码10",
        "码11","码12","码13","码14","码15","码16","码17","码18","码19","码20",
        "投注总额", "奖池金额",
        # 选十
        "选十中十注数","选十中十奖金",
        "选十中九注数","选十中九奖金",
        "选十中八注数","选十中八奖金",
        "选十中七注数","选十中七奖金",
        "选十中六注数","选十中六奖金",
        "选十中五注数","选十中五奖金",
        "选十中零注数","选十中零奖金",
        # 选九
        "选九中九注数","选九中九奖金",
        "选九中八注数","选九中八奖金",
        "选九中七注数","选九中七奖金",
        "选九中六注数","选九中六奖金",
        "选九中五注数","选九中五奖金",
        "选九中四注数","选九中四奖金",
        "选九中零注数","选九中零奖金",
        # 选八
        "选八中八注数","选八中八奖金",
        "选八中七注数","选八中七奖金",
        "选八中六注数","选八中六奖金",
        "选八中五注数","选八中五奖金",
        "选八中四注数","选八中四奖金",
        "选八中三注数","选八中三奖金",
        "选八中零注数","选八中零奖金",
        # 选七
        "选七中七注数","选七中七奖金",
        "选七中六注数","选七中六奖金",
        "选七中五注数","选七中五奖金",
        "选七中四注数","选七中四奖金",
        "选七中三注数","选七中三奖金",
        "选七中二注数","选七中二奖金",
        "选七中零注数","选七中零奖金",
        # 选六
        "选六中六注数","选六中六奖金",
        "选六中五注数","选六中五奖金",
        "选六中四注数","选六中四奖金",
        "选六中三注数","选六中三奖金",
        "选六中二注数","选六中二奖金",
        "选六中一注数","选六中一奖金",
        # 选五
        "选五中五注数","选五中五奖金",
        "选五中四注数","选五中四奖金",
        "选五中三注数","选五中三奖金",
        "选五中二注数","选五中二奖金",
        "选五中一注数","选五中一奖金",
    ],
    "qlc": [
        "期号", "日期",
        "基本1","基本2","基本3","基本4","基本5","基本6","基本7",
        "特别码",
        "投注总额",
        "奖池金额",
        "一等奖注数","一等奖金额",
        "二等奖注数","二等奖金额",
        "三等奖注数","三等奖金额",
        "四等奖注数","四等奖金额",
        "五等奖注数","五等奖金额",
        "六等奖注数","六等奖金额",
        "七等奖注数","七等奖金额",
    ],
    "qxc": [
        "期号", "日期",
        "位1","位2","位3","位4","位5","位6","位7",
        "投注总额", "奖池金额",
        "一等奖注数","一等奖金额",
        "二等奖注数","二等奖金额",
        "三等奖注数","三等奖金额",
        "四等奖注数","四等奖金额",
        "五等奖注数","五等奖金额",
        "六等奖注数","六等奖金额",
    ],
}


def _pad_parts(parts, min_len):
    return parts + [""] * (min_len - len(parts))


def _dict_to_row(code, d):
    """
    将字段字典转为行列表。
    快乐8 的奖级字段已在 query.py 解析为独立 dict 键。
    """
    fields = FIELD_DEFS.get(code, [])
    result = []
    for f in fields:
        val = d.get(f, "")
        if isinstance(val, (list, tuple)):
            result.append(" ".join(str(v) for v in val))
        else:
            result.append(str(val) if val is not None else "")

    # 基本号码列拆分（快乐8除外，已在 query.py 解析为独立键）
    if code == "ssq" and "红球" in d:
        for i, v in enumerate(d["红球"][:6]):
            if i < len(result):
                result[2 + i] = str(v)
    if code == "dlt":
        if "前区" in d:
            for i, v in enumerate(d["前区"][:5]):
                result[2 + i] = str(v)
        if "后区" in d:
            for i, v in enumerate(d["后区"][:2]):
                result[7 + i] = str(v)
    if code == "qlc" and "基本号" in d:
        for i, v in enumerate(d["基本号"][:7]):
            result[2 + i] = str(v)
    if code == "qxc" and "号码" in d:
        for i, v in enumerate(d["号码"][:7]):
            result[2 + i] = str(v)
    if code in ("pl3", "3d") and "号码" in d:
        for i, v in enumerate(d["号码"][:3]):
            result[2 + i] = str(v)
    if code == "pl5" and "号码" in d:
        for i, v in enumerate(d["号码"][:5]):
            result[2 + i] = str(v)

    return result


def export_txt(results_by_code, output_path):
    """文本格式，每期一行基本信息"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("彩票开奖数据\n")
        f.write("=" * 60 + "\n\n")
        for name, records in results_by_code.items():
            f.write(f"【{name}】\n")
            for rec in records:
                f.write(f"  {rec['formatted']}\n")
            f.write("\n")
    print(f"[OK] 导出 TXT: {output_path}")


def export_md(results_by_code, output_path):
    """Markdown 格式，按彩种分节，每节用 FIELD_DEFS 列对齐表格"""
    lines = []
    today = date.today()
    lines.append("# 彩票开奖数据\n")
    lines.append(f"> 数据来源：17500.cn | 导出时间：{today}\n")

    for name, records in results_by_code.items():
        code = NAME_TO_CODE.get(name, name)
        fields = FIELD_DEFS.get(code, [])

        lines.append("---")
        lines.append(f"## {name}（近 {len(records)} 期）\n")

        if not records:
            lines.append("暂无数据\n")
            continue

        # 表头
        lines.append("| " + " | ".join(fields) + " |")
        lines.append("| " + " | ".join(["---"] * len(fields)) + " |")

        for rec in records:
            row = _dict_to_row(code, rec["parsed"])
            cells = [str(v) if v else "-" for v in row]
            lines.append("| " + " | ".join(cells) + " |")

        lines.append("")

    content = "\n".join(lines)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)
    size_kb = os.path.getsize(output_path) / 1024
    print(f"[OK] 导出 Markdown: {output_path} ({size_kb:.1f} KB)")


def export_xlsx(results_by_code, output_path):
    if not OPENPYXL_OK:
        print("[ERROR] openpyxl 未安装，无法导出 Excel")
        return
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, records in results_by_code.items():
        code = NAME_TO_CODE.get(name, name)
        ws = wb.create_sheet(name[:31])
        fields = FIELD_DEFS.get(code, [])
        ws.append(fields)
        for rec in records:
            ws.append(_dict_to_row(code, rec["parsed"]))
    wb.save(output_path)
    size_kb = os.path.getsize(output_path) / 1024
    print(f"[OK] 导出 Excel: {output_path} ({size_kb:.1f} KB)")


def export_sqlite(results_by_code, output_path):
    if not SQLITE3_OK:
        print("[ERROR] sqlite3 不可用，无法导出 SQLite")
        return
    conn = sqlite3.connect(output_path)
    cur = conn.cursor()
    for name, records in results_by_code.items():
        code = NAME_TO_CODE.get(name, name)
        fields = FIELD_DEFS.get(code, ["field" + str(i) for i in range(20)])
        col_defs = ", ".join([f'"{f}" TEXT' for f in fields])
        cur.execute(f"DROP TABLE IF EXISTS lottery_{code}")
        cur.execute(f"CREATE TABLE lottery_{code} ({col_defs})")
        for rec in records:
            row = _pad_parts(_dict_to_row(code, rec["parsed"]), len(fields))
            placeholders = ", ".join(["?"] * len(fields))
            cur.execute(f'INSERT INTO lottery_{code} VALUES ({placeholders})', row)
        conn.commit()
    conn.close()
    size_kb = os.path.getsize(output_path) / 1024
    print(f"[OK] 导出 SQLite: {output_path} ({size_kb:.1f} KB)")


def auto_export(results_by_code, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    today = date.today().strftime("%Y%m%d")
    base = os.path.join(output_dir, f"lottery_data_{today}")
    fmt_count = 0
    if OPENPYXL_OK:
        export_xlsx(results_by_code, base + ".xlsx")
        fmt_count += 1
    if SQLITE3_OK:
        export_sqlite(results_by_code, base + ".db")
        fmt_count += 1
    export_txt(results_by_code, base + ".txt")
    fmt_count += 1
    export_md(results_by_code, base + ".md")
    fmt_count += 1
    print(f"[OK] 全部导出完成（{fmt_count} 种格式）: {output_dir}")


# ── CLI ──────────────────────────────────────────────────────────────────────
def _build_parser():
    p = argparse.ArgumentParser(
        description="彩票开奖数据导出工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  python export.py --codes 双色球,大乐透 --periods 50 --fmt xlsx --out C:/Users/vboxuser/Downloads
  python export.py --codes 全部 --periods 100 --fmt all --out C:/Users/vboxuser/Downloads
  python export.py --codes 北京快乐8 --periods 20 --fmt md --out .
  python export.py --codes 双色球 --periods 10 --fmt txt --out .
        """,
    )
    p.add_argument(
        "--codes", default="全部",
        help="彩种名称，逗号分隔；写\"全部\"表示全部 8 种（默认：全部）",
    )
    p.add_argument(
        "--periods", type=int, default=50,
        help="每种最近多少期（默认：50）",
    )
    p.add_argument(
        "--fmt", default="all",
        choices=["xlsx", "md", "txt", "db", "all"],
        help="导出格式：xlsx / md / txt / db / all（默认：all）",
    )
    p.add_argument(
        "--out", default=".",
        help="输出目录（默认：当前目录）",
    )
    return p


if __name__ == "__main__":
    parser = _build_parser()
    args = parser.parse_args()

    # 解析彩种
    if args.codes.strip() == "全部":
        names = list(NAME_TO_CODE.keys())
    else:
        names = [n.strip() for n in args.codes.split(",")]
        unknown = [n for n in names if n not in NAME_TO_CODE]
        if unknown:
            print(f"[ERROR] 未知彩种: {', '.join(unknown)}")
            print(f"支持的彩种: {', '.join(NAME_TO_CODE.keys())}")
            sys.exit(1)

    # 查询
    sys.path.insert(0, str(Path(__file__).parent))
    import query

    queries = {name: args.periods for name in names}
    print(f"[INFO] 正在查询 {len(names)} 个彩种，每种最近 {args.periods} 期 ...")
    results = query.get_latest(queries)

    if not results:
        print("[ERROR] 未查询到任何数据")
        sys.exit(1)

    out_dir = os.path.abspath(args.out)
    os.makedirs(out_dir, exist_ok=True)
    today = date.today().strftime("%Y%m%d")

    # 每个彩种单独导出一个文件（避免 xlsx 单文件过大）
    if args.fmt == "all":
        fmts = ["xlsx", "md", "txt", "db"]
    else:
        fmts = [args.fmt]

    # xlsx：每个彩种单独一个文件
    if "xlsx" in fmts:
        for name, records in results.items():
            code = NAME_TO_CODE.get(name, name)
            safe = name.replace("/", "-")
            p = os.path.join(out_dir, f"{safe}_近{args.periods}期_{today}.xlsx")
            if OPENPYXL_OK:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
                ws = wb.create_sheet(name[:31])
                fields = FIELD_DEFS.get(code, [])
                ws.append(fields)
                for rec in records:
                    ws.append(_dict_to_row(code, rec["parsed"]))
                wb.save(p)
                size_kb = os.path.getsize(p) / 1024
                print(f"[OK] {name}: {p} ({size_kb:.1f} KB)")

    if "md" in fmts or "all" in fmts:
        lines = []
        lines.append("# 彩票开奖数据\n")
        lines.append(f"> 数据来源：17500.cn | 导出时间：{date.today()}\n")
        for name, records in results.items():
            code = NAME_TO_CODE.get(name, name)
            fields = FIELD_DEFS.get(code, [])
            lines.append("---")
            lines.append(f"## {name}（近 {len(records)} 期）\n")
            if not records:
                lines.append("暂无数据\n")
                continue
            lines.append("| " + " | ".join(fields) + " |")
            lines.append("| " + " | ".join(["---"] * len(fields)) + " |")
            for rec in records:
                row = _dict_to_row(code, rec["parsed"])
                cells = [str(v) if v else "-" for v in row]
                lines.append("| " + " | ".join(cells) + " |")
            lines.append("")
        p = os.path.join(out_dir, f"彩票数据_近{args.periods}期_{today}.md")
        with open(p, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        size_kb = os.path.getsize(p) / 1024
        print(f"[OK] Markdown: {p} ({size_kb:.1f} KB)")

    if "txt" in fmts:
        p = os.path.join(out_dir, f"彩票数据_近{args.periods}期_{today}.txt")
        export_txt(results, p)

    if "db" in fmts:
        p = os.path.join(out_dir, f"彩票数据_近{args.periods}期_{today}.db")
        export_sqlite(results, p)
