"""
Price Monitor & Daily Excel Report Bot
Argus AI Agent — OpenClaw Skill v1.0.0
"""

import os
import time
import smtplib
import logging
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path

import pandas as pd
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ── Logging ──────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("price-monitor")

# ── Config from env (OpenClaw secrets) ───────────────────────────────
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
RECIPIENT = os.getenv("REPORT_RECIPIENT", "")
URLS_FILE = Path("config/urls.txt")
DATA_DIR  = Path("data")
RAW_DIR   = DATA_DIR / "raw"
REPORT_DIR = DATA_DIR / "reports"

RAW_DIR.mkdir(parents=True, exist_ok=True)
REPORT_DIR.mkdir(parents=True, exist_ok=True)


def load_urls():
    if not URLS_FILE.exists():
        log.warning("urls.txt not found — using demo URLs")
        return ["https://www.amazon.com/dp/B08N5WRWNW"]
    return [u.strip() for u in URLS_FILE.read_text().splitlines() if u.strip()]


def scrape_price(driver, url: str) -> dict:
    """Scrape a single product URL and return price data."""
    try:
        driver.get(url)
        time.sleep(2)

        # Generic price selector — works for most e-commerce sites
        selectors = [
            '[data-automation="price"]',
            '.price', '#price', '.product-price',
            '[class*="price"]', '[id*="price"]',
        ]
        price_text = None
        for sel in selectors:
            elements = driver.find_elements(By.CSS_SELECTOR, sel)
            if elements:
                price_text = elements[0].text.strip()
                break

        title = driver.title[:80] if driver.title else url

        return {
            "url": url,
            "title": title,
            "price_raw": price_text or "N/A",
            "price": float(''.join(filter(lambda c: c.isdigit() or c == '.', price_text or '0')) or 0),
            "scraped_at": datetime.now().isoformat(),
            "status": "ok"
        }
    except Exception as e:
        log.error(f"Failed to scrape {url}: {e}")
        return {"url": url, "title": url, "price_raw": "ERROR", "price": 0, "status": "error"}


def run_scraper(urls: list) -> pd.DataFrame:
    """Run Selenium scraper on all URLs."""
    opts = Options()
    opts.add_argument("--headless")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0")

    driver = webdriver.Chrome(options=opts)
    results = []

    for i, url in enumerate(urls):
        log.info(f"[{i+1}/{len(urls)}] Scraping: {url}")
        results.append(scrape_price(driver, url))
        time.sleep(1.2)  # Polite delay

    driver.quit()
    return pd.DataFrame(results)


def compare_with_yesterday(today_df: pd.DataFrame) -> pd.DataFrame:
    """Load yesterday's data and compute price changes."""
    yesterday_files = sorted(REPORT_DIR.glob("raw_*.csv"))
    if not yesterday_files:
        today_df["prev_price"] = today_df["price"]
        today_df["change_pct"] = 0.0
        return today_df

    prev_df = pd.read_csv(yesterday_files[-1])[["url", "price"]].rename(columns={"price": "prev_price"})
    merged = today_df.merge(prev_df, on="url", how="left")
    merged["prev_price"] = merged["prev_price"].fillna(merged["price"])
    merged["change_pct"] = ((merged["price"] - merged["prev_price"]) / merged["prev_price"] * 100).round(2)
    return merged


def build_excel_report(df: pd.DataFrame, date_str: str) -> Path:
    """Create color-coded Excel report."""
    report_path = REPORT_DIR / f"price_report_{date_str}.xlsx"

    cols = ["title", "url", "prev_price", "price", "change_pct", "scraped_at"]
    available = [c for c in cols if c in df.columns]
    df[available].to_excel(report_path, index=False, sheet_name="Price Report")

    wb = load_workbook(report_path)
    ws = wb.active

    red_fill   = PatternFill("solid", fgColor="FFCCCC")   # price drop
    green_fill = PatternFill("solid", fgColor="CCFFCC")   # price rise

    change_col = available.index("change_pct") + 1 if "change_pct" in available else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if change_col:
            val = row[change_col - 1].value
            if isinstance(val, (int, float)):
                fill = red_fill if val < -5 else (green_fill if val > 5 else None)
                if fill:
                    for cell in row:
                        cell.fill = fill

    wb.save(report_path)
    log.info(f"Report saved: {report_path}")
    return report_path


def send_email(report_path: Path, date_str: str, summary: str):
    """Send Excel report via email."""
    if not all([SMTP_USER, SMTP_PASS, RECIPIENT]):
        log.warning("Email not configured — skipping send")
        return

    msg = MIMEMultipart()
    msg["From"] = SMTP_USER
    msg["To"] = RECIPIENT
    msg["Subject"] = f"📊 Price Report — {date_str}"
    msg.attach(MIMEText(f"Daily price monitoring report.\n\n{summary}", "plain"))

    with open(report_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={report_path.name}")
        msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, RECIPIENT, msg.as_string())
        log.info(f"Report emailed to {RECIPIENT}")
    except Exception as e:
        log.error(f"Email failed: {e} — report saved locally at {report_path}")


def main():
    date_str = datetime.now().strftime("%Y-%m-%d")
    log.info(f"=== Price Monitor started: {date_str} ===")

    urls = load_urls()
    log.info(f"Monitoring {len(urls)} URLs")

    today_df = run_scraper(urls)
    today_df.to_csv(RAW_DIR / f"raw_{date_str}.csv", index=False)

    result_df = compare_with_yesterday(today_df)
    report_path = build_excel_report(result_df, date_str)

    drops = result_df[result_df.get("change_pct", pd.Series()) < -5] if "change_pct" in result_df else pd.DataFrame()
    summary = (
        f"Total monitored: {len(result_df)}\n"
        f"Price drops (>5%): {len(drops)}\n"
        f"Errors: {len(result_df[result_df['status'] == 'error'])}\n"
    )
    print(summary)
    send_email(report_path, date_str, summary)
    log.info("=== Done ===")


if __name__ == "__main__":
    main()
