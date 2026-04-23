#!/usr/bin/env python3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "Templates"

THIN = Side(style="thin", color="A6A6A6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
META_FILL = PatternFill("solid", fgColor="D9EAF7")
REQ_FILL = PatternFill("solid", fgColor="FFF2CC")
OPT_FILL = PatternFill("solid", fgColor="EDEDED")
GUIDE_HEAD_FILL = PatternFill("solid", fgColor="B7D7E8")
INPUT_FONT = Font(name="Arial", size=11, color="0000FF")
HEAD_FONT = Font(name="Arial", size=11, bold=True)
BODY_FONT = Font(name="Arial", size=10)


def style_cell(cell, *, fill=None, font=None, alignment=None):
    cell.border = BORDER
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment


def apply_column_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_dropdown(ws, cell_range: str, options: list[str]):
    dv = DataValidation(type="list", formula1='"' + ','.join(options) + '"', allow_blank=True)
    dv.prompt = "请从下拉列表中选择"
    dv.error = "请输入模板允许的枚举值"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def build_guide_sheet(wb: Workbook, title: str, metadata_example: list[str], required_fields: list[str], optional_fields: list[str], notes: list[str]):
    ws = wb.create_sheet("填写说明")
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=14, bold=True)

    rows = [
        ["填写位置", "说明"],
        ["第 1 行", "元数据行，不是表头"],
        ["第 2 行", "字段表头，请不要改名"],
        ["第 3 行起", "正式数据"],
        ["元数据示例", " | ".join(metadata_example)],
        ["必填字段", "、".join(required_fields)],
        ["选填字段", "、".join(optional_fields) if optional_fields else "无"],
    ]
    for item in notes:
        rows.append(["注意事项", item])

    start_row = 3
    for r, values in enumerate(rows, start=start_row):
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            style_cell(
                cell,
                fill=GUIDE_HEAD_FILL if r == start_row else None,
                font=HEAD_FONT if r == start_row else BODY_FONT,
                alignment=Alignment(vertical="center", wrap_text=True),
            )
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 72


def build_data_sheet(wb: Workbook, sheet_name: str, metadata: list[str], headers: list[str], required_headers: set[str], dropdowns: dict[str, list[str]]):
    ws = wb.active
    ws.title = sheet_name

    for col_idx, value in enumerate(metadata, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        style_cell(cell, fill=META_FILL, font=HEAD_FONT, alignment=Alignment(horizontal="center"))

    for col_idx, header in enumerate(headers, start=1):
        fill = REQ_FILL if header in required_headers else OPT_FILL
        cell = ws.cell(row=2, column=col_idx, value=header)
        style_cell(cell, fill=fill, font=HEAD_FONT, alignment=Alignment(horizontal="center", vertical="center"))
        ws.cell(row=3, column=col_idx).font = INPUT_FONT
        ws.cell(row=3, column=col_idx).alignment = Alignment(vertical="center")
        ws.cell(row=3, column=col_idx).border = BORDER

    widths = []
    for header in headers:
        if header in {"WD Employee ID", "Employee Type", "Region", "BG", "Termination Category"}:
            widths.append(24)
        elif "Date" in header:
            widths.append(18)
        else:
            widths.append(22)
    apply_column_widths(ws, widths)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}3"

    for header, options in dropdowns.items():
        if header in headers:
            col_idx = headers.index(header) + 1
            add_dropdown(ws, f"{get_column_letter(col_idx)}3:{get_column_letter(col_idx)}500", options)


def create_active_template(path: Path):
    headers = [
        "WD Employee ID",
        "Employee Type",
        "Region",
        "Country/Territory",
        "Work Location",
        "BG",
        "Line",
        "Department",
        "Tencent Organization",
        "Hire Date",
    ]
    required = {"WD Employee ID", "Employee Type", "Region", "Country/Territory", "BG"}
    wb = Workbook()
    build_data_sheet(
        wb,
        "数据模板",
        ["dataset_type", "active", "snapshot_date", "2026-02-28"],
        headers,
        required,
        {
            "Employee Type": ["Regular", "Intern"],
            "Region": ["Americas", "APAC", "EMEA", "Greater China"],
        },
    )
    build_guide_sheet(
        wb,
        "在职明细上传模板说明",
        ["dataset_type", "active", "snapshot_date", "2026-02-28"],
        ["WD Employee ID", "Employee Type", "Region", "Country/Territory", "BG"],
        ["Work Location", "Line", "Department", "Tencent Organization", "Hire Date"],
        [
            "第 1 行必须保留为元数据行。",
            "第 2 行字段名请不要改动。",
            "第 3 行起填写明细，不要插入额外说明行。",
            "如上传多个在职文件，请保证每个文件的 snapshot_date 正确。",
        ],
    )
    wb.save(path)


def create_termination_template(path: Path):
    headers = [
        "WD Employee ID",
        "Employee Type",
        "Region",
        "Country/Territory",
        "Hire Date",
        "Termination Date",
        "Last Day of Work",
        "Termination Category",
        "BG",
    ]
    required = {"WD Employee ID", "Employee Type", "Region", "Country/Territory", "Termination Date", "Termination Category", "BG"}
    wb = Workbook()
    build_data_sheet(
        wb,
        "数据模板",
        ["dataset_type", "termination", "period_start", "2026-01-01", "period_end", "2026-02-28"],
        headers,
        required,
        {
            "Employee Type": ["Regular", "Intern"],
            "Region": ["Americas", "APAC", "EMEA", "Greater China"],
            "Termination Category": [
                "Terminate Employee > Voluntary",
                "Terminate Employee > Involuntary",
                "Terminate Employee > Others",
            ],
        },
    )
    build_guide_sheet(
        wb,
        "离职明细上传模板说明",
        ["dataset_type", "termination", "period_start", "2026-01-01", "period_end", "2026-02-28"],
        ["WD Employee ID", "Employee Type", "Region", "Country/Territory", "Termination Date", "Termination Category", "BG"],
        ["Hire Date", "Last Day of Work"],
        [
            "第 1 行请填写统计期间的起止日期。",
            "Termination Category 必须使用模板下拉中的标准值。",
            "第 2 行字段名请不要改动。",
            "第 3 行起填写离职人员明细。",
        ],
    )
    wb.save(path)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    create_active_template(OUT_DIR / "active_snapshot_template.xlsx")
    create_termination_template(OUT_DIR / "termination_detail_template.xlsx")
    print(OUT_DIR)


if __name__ == "__main__":
    main()
