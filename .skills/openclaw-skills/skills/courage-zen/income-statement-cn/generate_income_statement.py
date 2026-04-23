#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
盈盈利润表生成器 - 标准模板
根据明细表数据自动生成利润表
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
import os

def create_income_statement(source_file, output_file=None):
    """
    根据明细表生成利润表
    
    Args:
        source_file: 源 Excel 文件路径（含明细表）
        output_file: 输出利润表路径（可选，默认自动生成）
    
    Returns:
        生成的文件路径
    """
    # 自动提取月份
    filename = os.path.basename(source_file)
    month = ''
    for part in filename.split('_')[0]:
        if part.isdigit():
            month = part
            break
    
    if not output_file:
        output_file = f'/Users/zhengyong/.openclaw/workspace/利润表 {month}_明细生成.xlsx'
    
    wb_source = load_workbook(source_file)
    ws_detail = wb_source['明细']
    
    # 数据提取
    income_by_cat2_cat3 = {}
    expense_by_cat2_cat3 = {}
    
    for row in range(2, ws_detail.max_row + 1):
        cat1 = ws_detail.cell(row=row, column=3).value
        cat2 = ws_detail.cell(row=row, column=4).value
        cat3 = ws_detail.cell(row=row, column=5).value
        income_amt = ws_detail.cell(row=row, column=6).value
        expense_amt = ws_detail.cell(row=row, column=7).value
        amount = ws_detail.cell(row=row, column=8).value
        
        # 收入：优先发生金额
        if cat1 in ['收入', '实收']:
            amt = amount if amount else (income_amt if income_amt else 0)
            if amt and amt > 0:
                if cat2 not in income_by_cat2_cat3:
                    income_by_cat2_cat3[cat2] = {}
                if cat3 not in income_by_cat2_cat3[cat2]:
                    income_by_cat2_cat3[cat2][cat3] = 0
                income_by_cat2_cat3[cat2][cat3] += amt
        
        # 支出：优先发生金额
        if cat1 == '支出':
            amt = amount if amount else (expense_amt if expense_amt else 0)
            if amt and amt > 0:
                if cat2 not in expense_by_cat2_cat3:
                    expense_by_cat2_cat3[cat2] = {}
                if cat3 not in expense_by_cat2_cat3[cat2]:
                    expense_by_cat2_cat3[cat2][cat3] = 0
                expense_by_cat2_cat3[cat2][cat3] += amt
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '利润表'
    
    # 样式
    title_font = Font(name='微软雅黑', size=12, bold=True)
    bold_font = Font(name='微软雅黑', size=10, bold=True)
    
    # === 填写数据 ===
    # 标题行
    ws.merge_cells('B1:F1')
    ws['B1'] = f'利润表 {month}'
    ws['B1'].font = title_font
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 收入部分
    ws['B3'] = '收入'
    ws['B3'].font = bold_font
    
    # 动态填充收入明细，记录每个分类 2 的合计行号
    row = 4
    income_cat2_rows = []  # 记录分类 2 标题行号
    for cat2, cat3_dict in income_by_cat2_cat3.items():
        total = sum(cat3_dict.values())
        income_cat2_rows.append(row)  # 记录分类 2 行号
        ws.cell(row=row, column=3, value=cat2)
        ws.cell(row=row, column=5, value=total)
        row += 1
        for cat3, amount in cat3_dict.items():
            ws.cell(row=row, column=4, value=cat3)
            ws.cell(row=row, column=5, value=amount)
            row += 1
    
    income_end_row = row - 1
    
    # 设置收入合计公式（E3 = 所有分类 2 合计之和）
    income_formula = '=' + '+'.join([f'E{r}' for r in income_cat2_rows])
    ws.cell(row=3, column=5, value=income_formula)
    ws.cell(row=3, column=5).font = bold_font
    
    # 支出部分
    expense_header_row = row
    ws.cell(row=row, column=2, value='支出')
    ws.cell(row=row, column=2).font = bold_font
    
    # 记录支出分类 2 的行号
    row += 1
    expense_cat2_rows = []
    for cat2, cat3_dict in expense_by_cat2_cat3.items():
        total = sum(cat3_dict.values())
        expense_cat2_rows.append(row)  # 记录分类 2 行号
        ws.cell(row=row, column=3, value=cat2)
        ws.cell(row=row, column=5, value=total)
        row += 1
        for cat3, amount in cat3_dict.items():
            ws.cell(row=row, column=4, value=cat3)
            ws.cell(row=row, column=5, value=amount)
            row += 1
    
    expense_end_row = row - 1
    
    # 设置支出合计公式
    expense_formula = '=' + '+'.join([f'E{r}' for r in expense_cat2_rows])
    ws.cell(row=expense_header_row, column=5, value=expense_formula)
    ws.cell(row=expense_header_row, column=5).font = bold_font
    
    # 经营利润
    profit_row = expense_end_row + 1
    ws.cell(row=profit_row, column=2, value='经营利润')
    ws.cell(row=profit_row, column=2).font = bold_font
    ws.cell(row=profit_row, column=5, value=f'=E3-E{expense_header_row}')
    ws.cell(row=profit_row, column=5).font = bold_font
    
    # F 列留空
    for r in range(2, profit_row + 1):
        ws.cell(row=r, column=6, value='')
    
    # 列宽
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 40
    
    wb.save(output_file)
    print(f'✅ 利润表已生成：{output_file}')
    return output_file

if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        create_income_statement(sys.argv[1])
    else:
        print('用法：python3 generate_income_statement.py <财务报表路径>')
