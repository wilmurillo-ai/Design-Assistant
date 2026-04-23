#!/usr/bin/env python3
"""
交易记录表生成工具

根据银行源数据生成交易记录表 Excel 文件。

用法：
    python 交易记录表.py 202602 [--input-dir DIR] [--output-dir DIR] [--initial-balance AMOUNT]
"""

import argparse
import csv
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import xlrd
from openpyxl.styles import Font, Alignment


def find_source_file(pattern, input_dir, period):
    """查找匹配的源文件"""
    # 提取基础名称（如"招行活期"）和扩展名
    parts = pattern.split('YYYYMM')
    base_name = parts[0].strip()  # 去掉空格
    ext = parts[1] if len(parts) > 1 else ''
    
    # 查找包含基础名称和账期的文件
    for f in os.listdir(input_dir):
        if base_name in f and period in f and f.endswith(ext):
            return os.path.join(input_dir, f)
    return None


def clean_csv_file(filepath):
    """清理 CSV 文件：删除注释行和空行，统一逗号格式"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 替换中文逗号为英文逗号
    content = content.replace(',', ',')
    
    lines = content.split('\n')
    # 过滤注释行和空行
    clean_lines = [line for line in lines if line.strip() and not line.strip().startswith('#')]
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write('\n'.join(clean_lines))
    
    print(f"  ✓ 已清理 CSV 文件：{filepath}")


def read_checking_data(filepath):
    """读取招行活期数据"""
    data = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader, None)  # 跳过标题
        for row in reader:
            if len(row) >= 7:
                # 跳过标题行（如果有残留）
                if row[0] == '交易日期':
                    continue
                try:
                    data.append({
                        'date': row[0].strip(),
                        'income': float(row[2]) if row[2] else None,
                        'expense': float(row[3]) if row[3] else None,
                        'balance': float(row[4]) if row[4] else None,
                        'note': row[6].strip() if len(row) > 6 else ''
                    })
                except ValueError:
                    # 跳过无法解析的行
                    continue
    return data


def read_finance_data(filepath, period):
    """读取招行理财数据，过滤掉其他月份"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    year = int(period[:4])
    month = int(period[4:6])
    
    data = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None or row[0] in ['委托状态', '上期余额']:
            continue
        
        date_val = row[1] if len(row) > 1 else None
        if isinstance(date_val, datetime):
            if date_val.year != year or date_val.month != month:
                continue
        
        data.append({
            'date': date_val,
            'trans_type': row[4] if len(row) > 4 else '',
            'quantity': row[5] if len(row) > 5 else None,
            'amount': row[8] if len(row) > 8 else None,
            'raw': row[:14]
        })
    
    return data


def read_corp_data(filepath):
    """读取对公账户数据"""
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    
    data = []
    for i in range(ws.nrows):
        row = ws.row_values(i)
        if not row or all(c == '' or c is None for c in row):
            continue
        if row[0] in ['查询账号:', '借方交易笔数', '交易时间', '交易日期']:
            continue
        data.append(row)
    
    return data


def create_workbook(period, checking_data, finance_data, corp_data, initial_balance):
    """创建交易记录表工作簿"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # 1. 招行活期
    ws_checking = wb.create_sheet(title='招行活期')
    ws_checking.append(['交易日期', '收入', '支出', '余额', '交易备注'])
    for row in checking_data:
        ws_checking.append([row['date'], row['income'], row['expense'], row['balance'], row['note']])
    
    # 2. 招行理财
    ws_finance = wb.create_sheet(title='招行理财')
    if finance_data:
        # 添加标题行（从原始数据获取）
        headers = ['委托状态', '委托日期', '代码', '货币', '交易类型', '委托数量', 
                   '成交价格', '委托金额', '交易金额', '已提业绩报酬', '确认份额日期', 
                   '资金到账日期', '合同号', '交易详情']
        ws_finance.append(headers)
        
        for item in finance_data:
            ws_finance.append(list(item['raw']))
        
        # 添加上期余额行
        last_row = ws_finance.max_row + 1
        ws_finance.append(['上期余额', datetime(int(period[:4]), int(period[4:6]), 1), 
                          None, None, None, None, None, initial_balance, None])
    
    # 3. 对公账户
    ws_corp = wb.create_sheet(title='对公账户')
    # 添加标题行
    ws_corp.append(['交易日期', '收入', '支出', '余额', '交易备注'])
    for row in corp_data:
        date_val = row[0] if len(row) > 0 else ''
        # 转换日期格式：2026-01-08 13:30:35 → 20260108
        if isinstance(date_val, str) and '-' in date_val:
            date_str = date_val.split(' ')[0].replace('-', '')
        else:
            date_str = str(date_val) if date_val else ''
        
        if date_str:
            # 转换数值：字符串 '0.00' → 数字 0 → None（显示为空）
            def convert_value(v):
                if v is None or v == '':
                    return None
                if isinstance(v, str):
                    try:
                        num = float(v)
                        return None if num == 0 else num
                    except ValueError:
                        return v
                return None if v == 0 else v
            
            ws_corp.append([
                date_str,
                convert_value(row[2]) if len(row) > 2 else None,  # 收入（贷方）
                convert_value(row[1]) if len(row) > 1 else None,  # 支出（借方）
                convert_value(row[3]) if len(row) > 3 else None,  # 余额
                row[6] if len(row) > 6 else ''     # 账号/摘要
            ])
    
    # 4. 资金流水
    ws_print = wb.create_sheet(title='资金流水')
    
    # 构建数据
    all_rows = []
    
    # 总标题
    all_rows.append(('资金流水 ' + period, None, None, None, None))
    
    # 招行活期部分
    all_rows.append(('招行活期', None, None, None, None))
    all_rows.append(('交易日期', '收入', '支出', '余额', '交易备注'))
    for row in checking_data:
        all_rows.append([row['date'], row['income'], row['expense'], row['balance'], row['note']])
    
    # 空行
    all_rows.append((None, None, None, None, None))
    
    # 招行理财部分
    all_rows.append(('招行理财', None, None, None, None))
    all_rows.append(('交易日期', '收入', '支出', '余额', '交易备注'))
    
    # 分离份额分红和其他交易
    finance_transactions = []
    dividend_data = []
    
    for item in finance_data:
        if item['trans_type'] == '委托状态':
            continue
        
        date_str = item['date'].strftime('%Y%m%d') if isinstance(item['date'], datetime) else str(item['date'])
        
        if item['trans_type'] == '份额分红':
            dividend_data.append({'quantity': item['quantity']})
        else:
            finance_transactions.append({
                'date': date_str,
                'amount': item['amount'],
                'type': item['trans_type']
            })
    
    # 计算份额分红汇总
    dividend_total = sum(d['quantity'] for d in dividend_data if d['quantity'])
    dividend_count = len(dividend_data)
    
    # 生成理财数据行（过滤掉金额为 0 的行）
    start_row = len(all_rows) + 1
    
    # 先过滤掉金额为 0 的交易
    valid_transactions = [t for t in finance_transactions if t['amount'] and t['amount'] > 0]
    
    for idx, trans in enumerate(valid_transactions):
        row_num = start_row + idx
        
        if idx == 0:
            # 第一行：计算具体余额
            # 申购是收入（钱投入理财），赎回/快赎是支出（钱从理财取出）
            if trans['amount'] and trans['type'] == '申购':
                balance = initial_balance + trans['amount']
            elif trans['amount'] and trans['type'] in ['赎回', '快赎交易']:
                balance = initial_balance - trans['amount']
            else:
                balance = initial_balance
            
            all_rows.append([
                trans['date'],
                trans['amount'] if trans['type'] == '申购' else None,
                trans['amount'] if trans['type'] in ['赎回', '快赎交易'] else None,
                balance,
                trans['type']
            ])
        else:
            # 后续行：使用 Excel 公式
            prev_row = row_num - 1
            formula = f'=D{prev_row}+B{row_num}-C{row_num}'
            all_rows.append([
                trans['date'],
                trans['amount'] if trans['type'] == '申购' else None,
                trans['amount'] if trans['type'] in ['赎回', '快赎交易'] else None,
                formula,
                trans['type']
            ])
    
    # 添加份额分红汇总行
    dividend_row_num = start_row + len(valid_transactions)
    if valid_transactions:
        prev_row = dividend_row_num - 1
        formula = f'=D{prev_row}+B{dividend_row_num}-C{dividend_row_num}'
    else:
        formula = initial_balance
    
    # 月末日期
    year = int(period[:4])
    month = int(period[4:6])
    if month == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)
    last_day_str = last_day.strftime('%Y%m%d')
    
    all_rows.append([
        last_day_str,
        dividend_total,
        None,
        formula if isinstance(formula, str) else dividend_row_num,
        '利息'
    ])
    
    # 空行
    all_rows.append((None, None, None, None, None))
    
    # 对公账户部分
    all_rows.append(('对公账户', None, None, None, None))
    all_rows.append(('交易日期', '收入', '支出', '余额', '交易备注'))
    
    for row in corp_data:
        date_val = row[0] if len(row) > 0 else ''
        # 转换日期格式：2026-01-08 13:30:35 → 20260108
        if isinstance(date_val, str) and '-' in date_val:
            date_str = date_val.split(' ')[0].replace('-', '')
        else:
            date_str = str(date_val) if date_val else ''
        
        if date_str:
            # 转换数值：字符串 '0.00' → 数字 0，空字符串 → None
            def convert_value(v):
                if v is None or v == '':
                    return None
                if isinstance(v, str):
                    try:
                        num = float(v)
                        return num if num != 0 else None  # 0 转为 None
                    except ValueError:
                        return v
                return v if v != 0 else None  # 数字 0 转为 None
            
            all_rows.append([
                date_str,
                convert_value(row[2]) if len(row) > 2 else None,  # 收入（贷方）
                convert_value(row[1]) if len(row) > 1 else None,  # 支出（借方）
                convert_value(row[3]) if len(row) > 3 else None,  # 余额
                row[6] if len(row) > 6 else ''     # 摘要
            ])
    
    # 写入数据并设置格式
    for i, row in enumerate(all_rows, 1):
        if row is None or (row[0] is None and all(c is None for c in row)):
            continue
        
        if row[0] == '资金流水 ' + period:
            ws_print.merge_cells('A1:E1')
            cell = ws_print.cell(row=i, column=1, value=row[0])
            cell.font = Font(name='黑体', bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        elif row[0] in ['招行活期', '招行理财', '对公账户']:
            cell = ws_print.cell(row=i, column=1, value=row[0])
            cell.font = Font(name='黑体', bold=True, size=12)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        elif row[0] == '交易日期':
            for j, val in enumerate(row, 1):
                cell = ws_print.cell(row=i, column=j, value=val)
                cell.font = Font(name='黑体', bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        else:
            for j, val in enumerate(row, 1):
                # 把 0 改为 None（显示为空）
                if val == 0:
                    val = None
                cell = ws_print.cell(row=i, column=j, value=val)
                cell.font = Font(name='宋体', size=11)
                if j <= 3:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 设置列宽
    ws_print.column_dimensions['A'].width = 12
    ws_print.column_dimensions['B'].width = 10
    ws_print.column_dimensions['C'].width = 10
    ws_print.column_dimensions['D'].width = 12
    ws_print.column_dimensions['E'].width = 40
    
    return wb


def main():
    parser = argparse.ArgumentParser(description='生成交易记录表')
    parser.add_argument('period', help='账期 (YYYYMM 格式，如 202602)')
    parser.add_argument('--input-dir', default='.', help='源文件目录')
    parser.add_argument('--output-dir', default='.', help='输出目录')
    parser.add_argument('--initial-balance', type=float, default=71239.37, 
                        help='理财期初余额')
    
    args = parser.parse_args()
    
    period = args.period
    input_dir = args.input_dir
    output_dir = args.output_dir
    initial_balance = args.initial_balance
    
    print(f"生成交易记录表 {period}")
    print(f"  源文件目录：{input_dir}")
    print(f"  输出目录：{output_dir}")
    print(f"  理财期初余额：{initial_balance}")
    print()
    
    # 查找源文件
    checking_file = find_source_file('招行活期 YYYYMM.csv', input_dir, period)
    finance_file = find_source_file('招行理财 YYYYMM.xlsx', input_dir, period)
    corp_file = find_source_file('对公账户 YYYYMM.xls', input_dir, period)
    
    if not checking_file:
        print("错误：未找到招行活期 CSV 文件")
        sys.exit(1)
    if not finance_file:
        print("错误：未找到招行理财 Excel 文件")
        sys.exit(1)
    if not corp_file:
        print("错误：未找到对公账户 Excel 文件")
        sys.exit(1)
    
    print("找到源文件：")
    print(f"  招行活期：{checking_file}")
    print(f"  招行理财：{finance_file}")
    print(f"  对公账户：{corp_file}")
    print()
    
    # 清理 CSV 文件
    print("1. 清理 CSV 文件...")
    clean_csv_file(checking_file)
    
    # 读取数据
    print("\n2. 读取数据...")
    print("  读取招行活期...")
    checking_data = read_checking_data(checking_file)
    print(f"    {len(checking_data)} 条记录")
    
    print("  读取招行理财...")
    finance_data = read_finance_data(finance_file, period)
    print(f"    {len(finance_data)} 条记录")
    
    print("  读取对公账户...")
    corp_data = read_corp_data(corp_file)
    print(f"    {len(corp_data)} 条记录")
    
    # 创建工作簿
    print("\n3. 生成交易记录表...")
    wb = create_workbook(period, checking_data, finance_data, corp_data, initial_balance)
    
    # 保存
    output_file = os.path.join(output_dir, f'交易记录{period}.xlsx')
    wb.save(output_file)
    print(f"  ✓ 已保存：{output_file}")
    
    print("\n✅ 完成！")
    print("\nSheet 列表：")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}: {wb[sheet].max_row} 行")


if __name__ == '__main__':
    main()
