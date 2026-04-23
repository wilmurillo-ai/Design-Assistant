#!/usr/bin/env python3
"""
从交易记录自动生成账目 Excel 文件

用法：
    python 生成账目.py 202512 --input-dir /path/to/交易记录 --output-dir /path/to/输出 --accounting-date 20260110
"""

import argparse
import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment


# ============ 分类规则配置 ============

# 分类 2 关键词匹配规则（按优先级排序）
CATEGORY2_KEYWORDS = {
    '导师费': ['导师费', 'cw-', '成稳导师'],
    '员工开支': ['工资', '社保', '补贴', '分红'],
    '场地费': ['场地费', '旷济堂'],
    '工作坊': ['工作坊', '苏醒', 'sx', '成长班', '助人班', '存在主义', '呼吸', '禅修', '解梦', '服务课', '家族'],
    '个案': ['个案'],
    '其他': ['餐费', '水果', '零食', '矿泉水', '纸杯', '画纸', '垃圾袋', '办公耗材', '网银', '社保', '转账'],
}

# 转账关键词（优先匹配）
TRANSFER_KEYWORDS = ['理财转活期', '活期转理财', '对公账户转活期', '理财转交行', '交行转理财']

# 预收/实收关键词
PREPAY_KEYWORDS = ['存入', '报名', '预收', '学费']
CONFIRM_KEYWORDS = ['复训', '完成', '实收']

# 非人名关键词（用于排除）
NON_NAME_KEYWORDS = [
    '理财', '活期', '交行', '对公', '转账', ' cw', '成稳', '盈盈', '付付', 
    '餐费', '水果', '零食', '矿泉水', '纸杯', '画纸', '垃圾袋', '场地费',
    '社保', '工资', '导师费', '返点', '报销', '网银', '办公', '停车',
    '发票', '税', '利息', '分红', '借款', '还款', '贷款', '押金'
]

# ====== 2026-04-01 更新：分类规则优化 ======
# 根据实际账务处理经验，分类 1/2/3 列默认留空，由人工手动填写
# 原因：预收/实收判断需要人工确认，自动化容易出错
AUTO_FILL_CATEGORIES = False  # 设置为 False 时，分类列留空

# 常见人名姓氏（用于辅助判断）
COMMON_SURNAMES = [
    '李', '王', '张', '刘', '陈', '杨', '黄', '赵', '周', '吴', '徐', '孙',
    '马', '朱', '胡', '郭', '何', '高', '林', '罗', '郑', '梁', '谢', '宋',
    '唐', '许', '邓', '冯', '曾', '彭', '潘', '袁', '于', '董', '萧', '叶',
    '蒋', '田', '杜', '丁', '魏', '姚', '卢', '汪', '范', '金', '石', '廖',
    '贾', '夏', '韦', '付', '方', '白', '崔', '康', '毛', '邱', '秦', '江',
    '史', '顾', '侯', '邵', '孟', '龙', '万', '段', '漕', '钱', '汤', '尹',
    '黎', '易', '常', '武', '乔', '贺', '赖', '龚', '文', '成', '欧', '阳'
]


def extract_student_name(note):
    """从备注中提取学员姓名"""
    if not note:
        return None
    
    note_lower = note.lower()
    
    # 模式 1：姓名 - 课程名（最常见）
    match = re.match(r'^([^-]+?)-', note)
    if match:
        name = match.group(1).strip()
        # 排除非人名关键词
        for kw in NON_NAME_KEYWORDS:
            if name.lower() == kw.lower():
                return None
        # 排除纯数字
        if name.isdigit():
            return None
        # 排除包含数字的（可能是课程编号）
        if re.search(r'\d', name):
            return None
        # 检查是否以常见姓氏开头
        for surname in COMMON_SURNAMES:
            if name.startswith(surname):
                return name
        # 2-4 个字符，可能是人名
        if 2 <= len(name) <= 4:
            return name
    
    # 模式 2：备注末尾的人名
    match = re.search(r'[-\s]([A-Za-z\u4e00-\u9fa5]{2,4})$', note)
    if match:
        name = match.group(1).strip()
        for surname in COMMON_SURNAMES:
            if name.startswith(surname):
                return name
    
    return None


def extract_course_name(note):
    """从备注中提取课程名称"""
    if not note:
        return None
    
    # 排除转账类备注
    if any(kw in note for kw in TRANSFER_KEYWORDS):
        return None
    
    # 模式 1：完整的课程名（包含期数、日期）
    # 如：苏醒 87 届初阶 20251203-07、成长班 14 届、助人班 7 届
    patterns = [
        r'(苏醒\d+届 [^-\s,]+(?:20\d{6}(?:-\d{2,4})?)?)',
        r'(成长班\d+届)',
        r'(助人班\d+届)',
        r'(存在主义\d+期)',
        r'(家族治疗\d+届)',
        r'(呼吸\d+届)',
        r'(禅修\d+期)',
        r'(解梦\d+期)',
        r'(服务课\d*)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, note)
        if match:
            return match.group(1)
    
    # 模式 2：姓名 - 课程简写（如 sx87、czb14）
    match = re.search(r'-([A-Za-z]+)(\d+)', note)
    if match:
        abbr = match.group(1).lower()
        num = match.group(2)
        # 课程缩写映射
        course_map = {
            'sx': '苏醒',
            'czb': '成长班',
            'zrb': '助人班',
            'czy': '存在主义',
            'jz': '家族',
            'hx': '呼吸',
        }
        if abbr in course_map:
            return f"{course_map[abbr]}{num}届"
    
    return None


def determine_category(note, income, expense):
    """
    判断分类 1 和分类 2
    
    返回：(分类 1, 分类 2, 分类 3, is_prepay, is_confirmed)
    
    2026-04-01 更新：默认留空分类列，由人工手动填写
    原因：预收/实收判断需要人工确认，自动化容易出错
    """
    # 如果配置为不留空，才进行自动分类
    if not AUTO_FILL_CATEGORIES:
        return (None, None, None, False, False)
    
    if not note:
        return ('其他', None, None, False, False)
    
    # 转换数值类型（处理 Excel 公式字符串）
    income_val = None
    expense_val = None
    if income is not None and not isinstance(income, str):
        income_val = income
    if expense is not None and not isinstance(expense, str):
        expense_val = expense
    
    is_prepay = False  # 是否预收
    is_confirmed = False  # 是否已确认收入（实收）
    
    # 1. 先判断转账（最高优先级）
    for kw in TRANSFER_KEYWORDS:
        if kw in note:
            return ('转账', None, None, False, False)
    
    # 2. 判断支出类
    if expense_val and expense_val > 0:
        # 导师费
        if '导师费' in note or note.startswith('cw-'):
            return ('支出', '导师费', '成稳工作坊' if ('成稳' in note or note.startswith('cw-')) else '盈盈工作坊', False, False)
        
        # 员工开支
        for kw in CATEGORY2_KEYWORDS['员工开支']:
            if kw in note:
                subcat = '工资' if '工资' in note else ('社保' if '社保' in note else '补贴')
                return ('支出', '员工开支', subcat, False, False)
        
        # 场地费
        if '场地费' in note:
            return ('支出', '场地费', None, False, False)
        
        # 工作坊支出（返点、物料等）
        for kw in CATEGORY2_KEYWORDS['工作坊']:
            if kw in note.lower():
                course = extract_course_name(note)
                return ('支出', '工作坊', course or kw, False, False)
        
        # 其他支出
        return ('支出', '其他', None, False, False)
    
    # 3. 判断收入类
    if income_val and income_val > 0:
        # 个案收入
        if '个案' in note:
            return ('收入', '个案', extract_student_name(note) or '其他', False, False)
        
        # 工作坊收入（学费）
        for kw in CATEGORY2_KEYWORDS['工作坊']:
            if kw in note.lower():
                course = extract_course_name(note)
                
                # 判断是否预收/实收
                # 预收：包含"存入"、"报名"、"预收"等关键词
                for prepay_kw in PREPAY_KEYWORDS:
                    if prepay_kw in note:
                        is_prepay = True
                        break
                
                # 实收：课程已完成（根据日期判断或包含"完成"、"复训"等）
                # 简化处理：如果备注包含完整日期且日期已过，认为是实收
                date_match = re.search(r'20\d{6}(?:-\d{2,4})?', note)
                if date_match:
                    course_date = date_match.group()[:8]  # 提取 8 位日期
                    today = datetime.now().strftime('%Y%m%d')
                    if course_date < today:  # 课程日期已过
                        is_confirmed = True
                
                for confirm_kw in CONFIRM_KEYWORDS:
                    if confirm_kw in note:
                        is_confirmed = True
                        is_prepay = False  # 实收优先
                        break
                
                return ('收入', '工作坊', course or kw, is_prepay, is_confirmed)
        
        # 其他收入
        return ('收入', '其他', None, False, False)
    
    return ('其他', None, None, False, False)


def determine_accounting_date(transaction_date, manual_date=None):
    """
    确定做账日期
    
    如果提供了 manual_date，使用手动指定的日期
    否则使用交易日期所在月份的次月 10 日
    """
    if manual_date:
        return manual_date
    
    # 解析交易日期
    date_str = str(transaction_date)
    if len(date_str) == 8:
        year = int(date_str[:4])
        month = int(date_str[4:6])
        
        # 次月
        if month == 12:
            next_year = year + 1
            next_month = 1
        else:
            next_year = year
            next_month = month + 1
        
        # 次月 10 日
        return f"{next_year}{next_month:02d}10"
    
    return None


def process_transaction_sheet(ws, sheet_name, accounting_date=None):
    """
    处理交易记录 sheet，生成账目数据
    
    返回：账目数据列表
    """
    accounts_data = []
    
    # 跳过标题行，找到数据开始行
    start_row = 1
    for i in range(1, 10):
        cell = ws.cell(row=i, column=1).value
        if cell and isinstance(cell, str) and '交易日期' in cell:
            start_row = i + 1
            break
    
    # 读取数据行
    for row_idx in range(start_row, ws.max_row + 1):
        # 读取 5 列数据
        date = ws.cell(row=row_idx, column=1).value
        income = ws.cell(row=row_idx, column=2).value
        expense = ws.cell(row=row_idx, column=3).value
        balance = ws.cell(row=row_idx, column=4).value
        note = ws.cell(row=row_idx, column=5).value
        
        # 跳过空行
        if not date and not note:
            continue
        
        # 跳过标题行
        if date == '交易日期':
            continue
        
        # 转换日期格式
        if isinstance(date, int):
            date_str = str(date)
        elif isinstance(date, str):
            date_str = date.replace('-', '')
        else:
            date_str = str(date) if date else ''
        
        # 判断分类
        cat1, cat2, cat3, is_prepay, is_confirmed = determine_category(note, income, expense)
        
        # 提取学员姓名
        student = extract_student_name(note)
        
        # 确定会计月份
        accounting_month = date_str[:6] if len(date_str) >= 6 else None
        
        # 确定做账日期
        acc_date = determine_accounting_date(date_str, accounting_date)
        
        # 处理余额（Excel 公式或数值）
        if isinstance(balance, str) and balance.startswith('='):
            balance_formula = balance
        else:
            balance_formula = None
        
        # 构建账目行
        account_row = {
            'date': date_str,
            'cat1': cat1,
            'cat2': cat2,
            'cat3': cat3,
            'income': income,
            'expense': expense,
            'balance': balance if not balance_formula else None,
            'balance_formula': balance_formula,
            'note': note,
            'student': student,
            'accounting_month': accounting_month,
            'accounting_date': acc_date,
            'is_prepay': is_prepay,
            'is_confirmed': is_confirmed,
        }
        
        accounts_data.append(account_row)
    
    return accounts_data


def create_account_workbook(period, accounts_data_map, accounting_date=None):
    """
    创建账目 Excel 工作簿
    
    accounts_data_map: {sheet_name: [账目数据列表]}
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # 账户映射：交易记录 sheet 名 → 账目 sheet 名
    sheet_mapping = {
        '招行活期': '活期账户',
        '招行理财': '理财账户',  # 招行理财 → 理财账户
        '交行活期': '对公账户',  # 交行活期 → 对公账户
        '对公账户': '对公账户',  # 对公账户 → 对公账户
        '打印': '活期账户',  # 旧版交易记录用"打印"sheet
    }
    
    for source_sheet, data_list in accounts_data_map.items():
        # 确定目标 sheet 名
        target_sheet = sheet_mapping.get(source_sheet, source_sheet)
        
        # 创建或获取 sheet
        if target_sheet in wb.sheetnames:
            ws = wb[target_sheet]
        else:
            ws = wb.create_sheet(title=target_sheet)
        
        # 写入标题行
        headers = ['余额', '发生日期', '分类 1', '分类 2', '分类 3', '收入金额', '支出金额', 
                   '应付/应收/预收/未分配利润', '备注', '学员/单位', '会计月份', '做账日期']
        ws.append(headers)
        
        # 设置标题格式
        for col in range(1, 13):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据行
        for row_data in data_list:
            # 计算 H 列（应付/应收/预收/未分配利润）
            h_value = None
            cat1 = row_data['cat1']
            income = row_data['income']
            
            # 预收：负债增加，用负数表示
            if cat1 == '预收' or row_data.get('is_prepay'):
                h_value = -income if income else None
            # 实收：确认收入，冲减预收负债，用正数
            elif cat1 == '实收' or row_data.get('is_confirmed'):
                h_value = income if income else None
            # 应收：借款等，用负数表示
            elif cat1 == '应收':
                h_value = -income if income else None
            # 应付：贷款等，用正数表示
            elif cat1 == '应付':
                h_value = income if income else None
            # 未分配利润：分红等
            elif cat1 == '未分配利润':
                h_value = -income if income else None
            
            # 余额：使用 Excel 公式
            # 2026-04-01 更新：第一行也用公式，引用上月最后行
            # 公式统一为：=A(上月最后行)+F(本行)-G(本行)
            row_num = ws.max_row + 1
            if row_data['balance_formula']:
                balance_val = row_data['balance_formula']
            else:
                # 所有行都用公式：=上期 + 本行收入 - 本行支出
                balance_val = f'=A{row_num-1}+F{row_num}-G{row_num}'
            
            row = [
                balance_val,  # A 列：余额
                row_data['date'],  # B 列：发生日期
                row_data['cat1'],  # C 列：分类 1
                row_data['cat2'],  # D 列：分类 2
                row_data['cat3'],  # E 列：分类 3
                row_data['income'],  # F 列：收入金额
                row_data['expense'],  # G 列：支出金额
                h_value,  # H 列：应付/应收/预收
                row_data['note'],  # I 列：备注
                row_data['student'],  # J 列：学员/单位
                row_data['accounting_month'],  # K 列：会计月份
                row_data['accounting_date'],  # L 列：做账日期
            ]
            ws.append(row)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 35
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 12
    
    return wb


def main():
    parser = argparse.ArgumentParser(description='从交易记录生成账目 Excel')
    parser.add_argument('period', help='账期 (YYYYMM 格式，如 202512)')
    parser.add_argument('--input-dir', default='.', help='交易记录文件目录')
    parser.add_argument('--output-dir', default='.', help='输出目录')
    parser.add_argument('--accounting-date', help='做账日期 (YYYYMMDD 格式，如 20260110)')
    
    args = parser.parse_args()
    
    period = args.period
    input_dir = args.input_dir
    output_dir = args.output_dir
    accounting_date = args.accounting_date
    
    print(f"生成账目 {period}")
    print(f"  输入目录：{input_dir}")
    print(f"  输出目录：{output_dir}")
    print(f"  做账日期：{accounting_date or '自动计算'}")
    print()
    
    # 查找交易记录文件
    jl_file = None
    for f in os.listdir(input_dir):
        if f'交易记录{period}' in f and f.endswith('.xlsx'):
            jl_file = os.path.join(input_dir, f)
            break
    
    if not jl_file:
        print(f"错误：未找到交易记录{period}.xlsx")
        return 1
    
    print(f"找到交易记录文件：{jl_file}")
    print()
    
    # 读取交易记录
    wb = openpyxl.load_workbook(jl_file)
    print(f"Sheet 列表：{wb.sheetnames}")
    print()
    
    # 处理每个 sheet
    accounts_data_map = {}
    for sheet_name in wb.sheetnames:
        # 跳过非数据 sheet
        if sheet_name in ['Sheet1', 'Sheet2']:
            continue
        
        ws = wb[sheet_name]
        print(f"处理 {sheet_name}...")
        data = process_transaction_sheet(ws, sheet_name, accounting_date)
        print(f"  {len(data)} 条记录")
        accounts_data_map[sheet_name] = data
    
    wb.close()
    print()
    
    # 创建账目工作簿
    print("创建账目工作簿...")
    account_wb = create_account_workbook(period, accounts_data_map, accounting_date)
    
    # 保存
    output_file = os.path.join(output_dir, f'账目{period}.xlsx')
    account_wb.save(output_file)
    print(f"✓ 已保存：{output_file}")
    print()
    print(f"Sheet 列表：{account_wb.sheetnames}")
    for sn in account_wb.sheetnames:
        ws = account_wb[sn]
        print(f"  - {sn}: {ws.max_row - 1} 行")
    
    print("\n✅ 完成！")


if __name__ == '__main__':
    main()
