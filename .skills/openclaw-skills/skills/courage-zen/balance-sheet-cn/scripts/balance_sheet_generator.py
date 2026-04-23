#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
资产负债表生成器
根据财务报表源文件自动生成规范化的资产负债表
"""

import openpyxl
from copy import copy
from pathlib import Path


class BalanceSheetGenerator:
    """资产负债表生成器"""
    
    def __init__(self, src_path: str):
        """
        初始化生成器
        
        Args:
            src_path: 源文件路径（Excel）
        """
        self.src_path = Path(src_path)
        if not self.src_path.exists():
            raise FileNotFoundError(f"源文件不存在：{src_path}")
        
        # 加载源文件
        self.wb_src = openpyxl.load_workbook(self.src_path, data_only=False)
        self.bs_src = self.wb_src['资产负债表']
        self.detail_src = self.wb_src['明细']
        self.pl_values = openpyxl.load_workbook(self.src_path, data_only=True)['利润表']
        
        # 读取数据
        self._load_data()
    
    def _load_data(self):
        """从源文件读取所需数据"""
        # 银行余额（M 列求和）
        self.cmb_balance = 0  # 招商银行
        self.comm_balance = 0  # 交通银行
        for row in range(1, 200):
            m_cell = self.detail_src[f'M{row}']
            a_cell = self.detail_src[f'A{row}']
            m_val = m_cell.value if m_cell.value is not None else ''
            a_val = a_cell.value if a_cell.value is not None else 0
            if '招商银行' in str(m_val):
                self.cmb_balance += a_val
            elif '交通银行' in str(m_val):
                self.comm_balance += a_val
        self.cmb_balance = round(self.cmb_balance, 2)
        self.comm_balance = round(self.comm_balance, 2)
        
        # 分类统计（C 列和 H 列）
        self.receivable_total = 0  # 应收
        self.advance_total = 0     # 预收
        self.payable_total = 0     # 应付
        for row in range(2, 200):
            c_val = self.detail_src[f'C{row}'].value or ''
            h_val = self.detail_src[f'H{row}'].value or 0
            if c_val == '应收':
                self.receivable_total += h_val
            elif c_val == '预收':
                self.advance_total += h_val
            elif c_val == '应付':
                self.payable_total += h_val
        
        # 经营利润
        self.operating_profit = 0
        for row in range(1, 50):
            b_val = self.pl_values[f'B{row}'].value or ''
            if '经营利润' in str(b_val):
                self.operating_profit = self.pl_values[f'E{row}'].value or 0
                break
    
    def generate(self, output_path: str):
        """
        生成资产负债表
        
        Args:
            output_path: 输出文件路径
        """
        # 创建新工作簿
        wb_new = openpyxl.Workbook()
        wb_new.remove(wb_new.active)
        bs_new = wb_new.create_sheet('资产负债表')
        
        # 复制所有单元格（格式、公式、样式）
        self._copy_sheet(bs_new)
        
        # 按顺序执行规则
        self._apply_rules(bs_new)
        
        # 保存
        output = Path(output_path)
        wb_new.save(output)
        print(f'资产负债表已保存至：{output}')
        return output
    
    def _copy_sheet(self, bs_new):
        """复制源工作表的所有内容和格式"""
        for row in range(1, 50):
            for col_num in range(1, 20):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                src_cell = self.bs_src[f'{col_letter}{row}']
                dst_cell = bs_new[f'{col_letter}{row}']
                
                # 复制值
                if src_cell.value is not None:
                    dst_cell.value = src_cell.value
                
                # 复制样式
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = copy(src_cell.number_format)
                
                # 复制列宽
                if self.bs_src.column_dimensions[col_letter].width:
                    bs_new.column_dimensions[col_letter].width = self.bs_src.column_dimensions[col_letter].width
            
            # 复制行高
            if self.bs_src.row_dimensions[row].height:
                bs_new.row_dimensions[row].height = self.bs_src.row_dimensions[row].height
        
        # 复制合并单元格
        for merged_range in self.bs_src.merged_cells.ranges:
            bs_new.merge_cells(str(merged_range))
    
    def _apply_rules(self, bs_new):
        """按顺序执行 11 条规则"""
        # 规则 2：D 列数值 → B 列（在规则 4-5 修改之前）
        d7_original = self.bs_src['D7'].value
        d8_original = self.bs_src['D8'].value
        d16_original = self.bs_src['D16'].value
        
        bs_new['B7'].value = d7_original
        bs_new['B8'].value = d8_original
        bs_new['B16'].value = d16_original
        
        # 计算 D 列其他行的数值
        d_values_original = {}
        d_values_original[5] = self.bs_src['B5'].value
        d_values_original[6] = d7_original + d8_original
        d_values_original[9] = self.bs_src['B9'].value + self.bs_src['C9'].value
        d_values_original[4] = d_values_original[5] + d_values_original[6] + d_values_original[9]
        d_values_original[11] = self.bs_src['B11'].value + self.bs_src['C11'].value
        d_values_original[12] = 0
        d_values_original[10] = d_values_original[11] + d_values_original[12]
        d_values_original[14] = self.bs_src['B14'].value + self.bs_src['C14'].value
        d_values_original[15] = self.bs_src['B15'].value
        d_values_original[13] = d_values_original[14] + d_values_original[15]
        
        for row, val in d_values_original.items():
            if row not in [7, 8, 16]:
                bs_new[f'B{row}'].value = val
        
        # 规则 3：C5 = 0
        bs_new['C5'] = 0
        
        # 规则 4-5：修改 D7, D8
        bs_new['D7'] = self.cmb_balance
        bs_new['D8'] = self.comm_balance
        
        # 规则 6-10
        bs_new['C9'] = self.receivable_total
        bs_new['C11'] = self.advance_total
        bs_new['C12'] = self.payable_total
        bs_new['C14'] = self.operating_profit
        bs_new['C15'] = 0
        
        # 规则 1 和 11：保留公式
        bs_new['C4'].value = '=D4-B4'
        bs_new['C6'].value = '=C7+C8'
        bs_new['C7'].value = '=D7-B7'
        bs_new['C8'].value = '=D8-B8'
        bs_new['C10'].value = '=C11+C12'
        bs_new['C13'].value = '=C14+C15'


def main():
    """命令行入口"""
    import sys
    
    if len(sys.argv) < 2:
        print('用法：python3 generate_balance_sheet.py <源文件路径> [输出文件路径]')
        sys.exit(1)
    
    src_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else '资产负债表.xlsx'
    
    generator = BalanceSheetGenerator(src_path)
    generator.generate(output_path)


if __name__ == '__main__':
    main()
