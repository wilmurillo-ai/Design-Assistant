#!/usr/bin/env python3
"""
军工采购商机抓取工具 - 整合版

功能：
1. 军队采购网商机抓取
2. 全军武器装备采购信息网商机抓取
3. 国防科大采购信息网商机抓取
4. 整合输出Excel（Sheet1: 全军装备采购网, Sheet2: 军队采购网，Sheet3：国防科大信息采购网）

输入参数:
    keywords: 核心关键词列表，用于匹配项目名称
    exclude_keywords: 排除关键词列表，用于过滤无关项目
    regions: 地区字典，key=地区名, value=regionCode (仅军队采购网使用)
    date: 日期，格式 YYYY-MM-DD，默认今天

推荐等级规则:
    高: 包含模型/仿真/数据/决策/分析/智能/AI/软件/意向等高价值关键词
    中: 包含系统/信息等一般关键词
    空: 其他匹配项

特殊规则:
    - 包含"意向"关键词的项目优先推荐
    - 包含"中标公告"/"结果公示"的项目过滤掉（已结束的项目）
"""

import os
import requests
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional
import time
import re
import json
import subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 从配置模块加载默认参数
from .config import get_keywords, get_exclude_keywords, get_high_value_keywords, get_regions, get_proxies, get_output_dir


def _clean_description(text):
    """清洗项目描述：去除HTML/CSS乱码，提取20-50字简述"""
    if pd.isna(text) or str(text).strip() == '':
        return ''
    
    text = str(text)
    
    # 移除HTML/CSS
    text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL)
    text = re.sub(r'<style[^>]*>.*?</style>', '', text, flags=re.DOTALL)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'\{[^}]*\}', ' ', text)
    text = re.sub(r'\.[a-zA-Z][a-zA-Z0-9_]*', ' ', text)
    text = re.sub(r'&[a-zA-Z]+;', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()
    
    # 提取核心内容
    lines = text.split('。')
    result = []
    for line in lines[:3]:
        line = line.strip()
        if line and len(line) > 5:
            if any(kw in line for kw in ['项目名称', '项目编号', '预算', '采购', '服务', '物资', '系统', '软件', '数据', '分析', '智能']):
                result.append(line[:30])
    
    text = ' '.join(result) if result else ''
    
    if len(text) > 50:
        text = text[:47] + '...'
    
    return text if len(text) >= 5 else ''


def _generate_remarks(title: str, matched_kw: List[str]) -> str:
    """根据标题生成简明的分析备注"""
    if not title:
        return ""
    
    remarks = []
    
    # 意向项目优先标注
    if '意向' in title:
        remarks.append("采购意向前期可跟进")
    
    # 根据关键词匹配生成备注
    if '体系' in title:
        remarks.append("体系设计相关")
    if '模型' in title:
        remarks.append("模型开发可参与")
    if '仿真' in title:
        remarks.append("仿真项目可做")
    if '数据' in title:
        remarks.append("数据分析需求")
    if 'AI' in title or '智能' in title:
        remarks.append("AI智能项目")
    if '软件' in title:
        remarks.append("软件定制开发")
    if '系统' in title:
        remarks.append("系统集成类")
    if '训练' in title or '实验' in title:
        remarks.append("训练实验相关")
    if '决策' in title or '规划' in title:
        remarks.append("决策规划类")
    
    if not remarks:
        remarks.append("可关注")
    
    # 最多返回2条
    return "; ".join(remarks[:2])


# 默认配置（从 config.py 加载，保留向后兼容）
DEFAULT_REGIONS = get_regions()
DEFAULT_HIGH_VALUE_KW = get_high_value_keywords()
DEFAULT_CORE_KW = get_keywords()
DEFAULT_EXCLUDE_KW = get_exclude_keywords()

# 采购方式枚举
MANNER_MAP = {
    '1': '公开招标',
    '2': '邀请招标', 
    '3': '竞争性谈判',
    '4': '询价',
    '5': '单一来源',
    '6': '其他'
}

# 项目类别枚举
NATURE_MAP = {
    '1': '物资',
    '2': '工程',
    '3': '服务'
}

# 采购方式映射
WEAIN_PURCHASE_TYPE_MAP = {
    '公开招标': '公开招标',
    '邀请招标': '邀请招标',
    '竞争性谈判': '竞争性谈判',
    '询价': '询价'
}


# ============ 全军武器装备采购信息网获取函数 ============

def fetch_weain_bidding(
    keywords: Optional[List[str]] = None,
    date: Optional[str] = None,
    save_csv: bool = False
) -> pd.DataFrame:
    """
    获取全军武器装备采购信息网商机信息
    
    Args:
        keywords: 核心关键词列表
        date: 日期，默认今天，格式 YYYY-MM-DD
        save_csv: 是否保存CSV（此函数不保存CSV，统一在主函数输出Excel）
    
    Returns:
        DataFrame: 筛选后的项目列表
    """
    if keywords is None:
        keywords = DEFAULT_CORE_KW
    if date is None:
        date = datetime.now().strftime('%Y-%m-%d')
    
    # 转换日期格式为API所需的日期（去掉-）
    api_date = date.replace('-', '')
    
    # 获取当前时间戳
    try:
        result = subprocess.run(['date', '+%s%3N'], capture_output=True, text=True)
        _t = result.stdout.strip()
    except:
        _t = str(int(datetime.now().timestamp() * 1000))
    
    all_results = []
    
    print(f"[INFO] 开始抓取全军武器装备采购信息网 - 日期: {date}")
    print("-" * 60)
    
    # 采购类型
    purchase_types = [
        ("公开招标", "%E5%85%AC%E5%BC%80%E6%8B%9B%E6%A0%87"),
        ("邀请招标", "%E9%82%80%E8%AF%B7%E6%8B%9B%E6%A0%87"),
        ("竞争性谈判", "%E7%AB%9E%E4%BA%89%E6%80%A7%E8%B0%88%E5%88%A4"),
        ("询价", "%E8%AF%A2%E4%BB%B7")
    ]
    
    for ptype, encoded_ptype in purchase_types:
        url = f"https://www.weain.mil.cn/api/front/list/cggg/list?LMID=1149231276155707394&pageNo=1&purchaseType={encoded_ptype}&_t={_t}"
        
        try:
            resp = requests.get(url, timeout=10, proxies=get_proxies())
            data = resp.json()
            content_list = data.get('list', {}).get('contentList', [])
            
            for item in content_list:
                publish_time = item.get('publishTime', '')
                
                # 只取当日数据
                if date not in publish_time:
                    continue
                
                title = item.get('nonSecretTitle', '')
                
                # 检查关键词匹配
                matched_kw = [kw for kw in keywords if kw in title]
                
                if matched_kw:
                    # 获取字段
                    purchase_type = item.get('purchaseType', ptype)
                    secret_grade = item.get('secretGrade', '公开')
                    pc_url = item.get('pcUrl', '')
                    deadline = item.get('deadline', '')
                    
                    # 推荐等级判断 - 意向优先
                    recommendation = ""
                    if '意向' in title:
                        recommendation = "高"
                    elif any(kw in matched_kw for kw in DEFAULT_HIGH_VALUE_KW):
                        recommendation = "高"
                    elif '系统' in matched_kw:
                        recommendation = "中"
                    
                    # 生成备注
                    remarks = _generate_remarks(title, matched_kw)
                    
                    all_results.append({
                        '项目名称': title,
                        '公告类型': purchase_type,
                        '涉密等级': secret_grade,
                        '详情链接': f'https://www.weain.mil.cn{pc_url}' if pc_url else '',
                        '截止日期': deadline,
                        '发布日期': publish_time[:10] if publish_time else '',
                        '匹配关键词': ','.join(matched_kw),
                        '推荐等级': recommendation,
                        '备注': remarks
                    })
                    
        except Exception as e:
            print(f"[ERROR] 获取{ptype}数据失败: {e}")
            continue
    
    # 创建DataFrame
    df = pd.DataFrame(all_results)
    
    if len(df) > 0:
        df.index = range(1, len(df) + 1)
        df.index.name = '序号'
        # 添加序号列到第一列
        df.reset_index(inplace=True)
    
    print(f"[INFO] 全军武器装备采购信息网: 共 {len(df)} 条匹配项目")
    
    return df


# ============ 军队采购网获取函数（更新版） ============

def fetch_military_bidding(
    keywords: Optional[List[str]] = None,
    exclude_keywords: Optional[List[str]] = None,
    regions: Optional[Dict[str, str]] = None,
    date: Optional[str] = None,
    high_value_keywords: Optional[List[str]] = None,
    save_csv: bool = True,
    output_path: str = None
) -> pd.DataFrame:
    """
    获取并筛选军队采购网商机信息
    
    Args:
        keywords: 核心关键词列表，默认使用 DEFAULT_CORE_KW
        exclude_keywords: 排除关键词列表，默认使用 DEFAULT_EXCLUDE_KW
        regions: 地区字典，默认使用 DEFAULT_REGIONS
        date: 日期，默认今天，格式 YYYY-MM-DD
        high_value_keywords: 高价值关键词列表，用于评级
        save_csv: 是否保存CSV文件（此版本保留参数兼容，实际不使用）
        output_path: 输出文件路径（此版本保留参数兼容）
    
    Returns:
        DataFrame: 筛选后的项目列表
    """
    # 参数初始化
    if keywords is None:
        keywords = DEFAULT_CORE_KW
    if exclude_keywords is None:
        exclude_keywords = DEFAULT_EXCLUDE_KW
    if regions is None:
        regions = DEFAULT_REGIONS
    if high_value_keywords is None:
        high_value_keywords = DEFAULT_HIGH_VALUE_KW
    if date is None:
        date = datetime.now().strftime('%Y-%m-%d')
    
    # API基础参数
    base_url = 'https://www.plap.mil.cn/freecms/rest/v1/notice/selectInfoMoreChannel.do'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Referer': 'https://www.plap.mil.cn/freecms/site/juncai/cggg/index.html'
    }
    site_id = '404bb030-5be9-4070-85bd-c94b1473e8de'
    channel_id = 'c5bff13f-21ca-4dac-b158-cb40accd3035'
    
    all_results = []
    
    print(f"[INFO] 开始抓取军队采购网 - 日期: {date}")
    print(f"[INFO] 目标地区: {', '.join(regions.keys())}")
    print(f"[INFO] 关键词: {', '.join(keywords)}")
    print("-" * 60)
    
    for region_name, region_code in regions.items():
        page = 1
        region_items = []
        consecutive_empty = 0

        # 分页获取数据
        max_pages = 10
        while page <= max_pages:
            params = {
                'siteId': site_id,
                'channel': channel_id,
                'currPage': page,
                'pageSize': 20,
                'regionCode': region_code,
            }

            try:
                resp = requests.get(base_url, params=params, headers=headers, timeout=10, proxies=get_proxies())
                data = resp.json()
                items = data.get('data', [])

                if not items:
                    break

                # 筛选当日数据
                today_items = [item for item in items if date in (item.get('noticeTime') or '')]

                if today_items:
                    region_items.extend(today_items)
                    consecutive_empty = 0
                    page += 1
                else:
                    consecutive_empty += 1
                    if consecutive_empty >= 2:
                        break
                    page += 1

            except Exception as e:
                print(f"[ERROR] 地区 {region_name} 第{page}页获取失败: {e}")
                break
        
        print(f"[INFO] {region_name}: 获取 {len(region_items)} 条数据")
        
        # 筛选匹配项目
        for item in region_items:
            title = item.get('title', '')
            
            # ====== 新增规则：过滤中标公告/结果公示 ======
            if '中标公告' in title or '结果公示' in title:
                continue
            
            # 排除无关项目
            if any(ex in title for ex in exclude_keywords):
                continue
            
            # 检查关键词匹配
            matched_kw = [kw for kw in keywords if kw in title]
            
            if matched_kw:
                # 获取字段
                purchase_manner = str(item.get('purchaseManner', ''))
                purchase_nature = str(item.get('purchaseNature', ''))
                pageurl = item.get('pageurl', '')
                notice_time = item.get('noticeTime', '')[:10] if item.get('noticeTime') else ''
                description = item.get('description', '')[:200] if item.get('description') else ''
                
                # ====== 新增规则：意向优先推荐 ======
                recommendation = ""
                if '意向' in title:
                    recommendation = "高"
                elif any(kw in matched_kw for kw in high_value_keywords):
                    recommendation = "高"
                elif '系统' in matched_kw:
                    recommendation = "中"
                
                # 生成备注
                remarks = _generate_remarks(title, matched_kw)
                
                all_results.append({
                    '项目名称': title,
                    '项目描述': description,
                    '地区': region_name,
                    '采购方式': MANNER_MAP.get(purchase_manner, purchase_manner),
                    '项目类别': NATURE_MAP.get(purchase_nature, purchase_nature),
                    '发布时间': notice_time,
                    '详情链接': f'https://www.plap.mil.cn{pageurl}' if pageurl else '',
                    '推荐等级': recommendation,
                    '匹配关键词': ','.join(matched_kw),
                    '备注': remarks
                })
    
    # 创建DataFrame
    df = pd.DataFrame(all_results)
    
    if len(df) > 0:
        df.index = range(1, len(df) + 1)
        df.index.name = '序号'
        df.reset_index(inplace=True)
        # 调整列顺序: 序号,项目名称,项目描述,地区,采购方式,项目类别,详情链接,发布时间,匹配关键词,推荐等级,备注
        cols = ['序号', '项目名称', '项目描述', '地区', '采购方式', '项目类别', '详情链接', '发布时间', '匹配关键词', '推荐等级', '备注']
        df = df[cols]
        
        # 清洗项目描述（去除HTML/CSS乱码，20-50字简述）
        df['项目描述'] = df['项目描述'].apply(_clean_description)
    
    print("-" * 60)
    print(f"[INFO] 军队采购网: 共 {len(df)} 条匹配项目")
    
    return df


# ============ 国防科大采购信息网获取函数 ============

def _get_nudt_purchase_type(title: str) -> Optional[str]:
    """根据项目名称判断公告类型"""
    # 剔除结果公示/中标公告
    if '结果公示' in title or '中标公告' in title:
        return None
    
    # 判断类型（按优先级）
    if '流标' in title or '废标' in title:
        return '废标公告'
    elif '招标公告' in title or '招标变更' in title or '变更公告' in title:
        return '招标公告'
    elif '比价公告' in title:
        return '比价公告'
    elif '询价公告' in title:
        return '询价公告'
    elif '竞争性谈判' in title:
        return '竞争性谈判'
    elif '意向' in title:
        return '意向公开'
    else:
        return '其他'


def _get_nudt_recommend_level(matched_kw: List[str]) -> str:
    """根据匹配关键词返回推荐等级"""
    high_kw = ['体系', '模型', '仿真', '数据', '决策', '分析', '智能', 'AI', '软件']
    if any(k in matched_kw for k in high_kw):
        return "高"
    elif '系统' in matched_kw:
        return "中"
    return ""


def _get_nudt_remarks(title: str) -> str:
    """生成简明分析备注"""
    remarks = []
    if '体系' in title:
        remarks.append("体系相关")
    if '模型' in title:
        remarks.append("模型开发")
    if '仿真' in title:
        remarks.append("仿真项目")
    if '数据' in title:
        remarks.append("数据处理")
    if '软件' in title or '系统' in title:
        remarks.append("软件/系统类")
    if '智能' in title or 'AI' in title:
        remarks.append("智能AI")
    if not remarks:
        remarks.append("可关注")
    return "; ".join(remarks[:2])


def _fetch_nudt_page(page: int) -> List[Dict]:
    """获取国防科大采购信息网指定页的数据"""
    if page == 1:
        url = "https://www.nudt.edu.cn/cgxx/index.htm"
    else:
        url = f"https://www.nudt.edu.cn/cgxx/index{page-1}.htm"
    
    try:
        resp = requests.get(url, timeout=10, proxies=get_proxies())
        resp.encoding = 'utf-8'
        html = resp.text
        
        li_pattern = r'<li>\s*<a href="([^"]+)">\s*<div class="date"><span class="day[^"]*">(\d+)</span><span class="year[^"]*">(\d+-\d+)</span></div>\s*<div class="title[^"]*">(.*?)</div>'
        matches = re.findall(li_pattern, html)
        
        results = []
        for href, day, year_month, title in matches:
            year = "20" + year_month.split("-")[0]
            month = year_month.split("-")[1]
            full_date = f"{year}-{month}-{day}"
            
            results.append({
                'href': 'https://www.nudt.edu.cn/cgxx/' + href,
                'date': full_date,
                'title': title.strip()
            })
        
        return results
    except Exception as e:
        return []


def fetch_nudt_bidding(
    keywords: Optional[List[str]] = None,
    date: Optional[str] = None
) -> pd.DataFrame:
    """
    获取国防科大采购信息网商机信息
    
    Args:
        keywords: 核心关键词列表
        date: 日期，默认今天，格式 YYYY-MM-DD
    
    Returns:
        DataFrame: 筛选后的项目列表
    """
    if keywords is None:
        keywords = DEFAULT_CORE_KW
    if date is None:
        date = datetime.now().strftime('%Y-%m-%d')
    
    all_results = []
    
    print(f"[INFO] 开始抓取国防科大采购信息网 - 日期: {date}")
    print("-" * 60)
    
    # 获取多页直到不是目标日期
    page = 1
    max_pages = 5
    
    while page <= max_pages:
        results = _fetch_nudt_page(page)
        if not results:
            break
        
        # 检查是否有目标日期的数据
        has_target_date = any(r['date'] == date for r in results)
        if not has_target_date:
            break
        
        # 收集目标日期的数据
        for r in results:
            if r['date'] == date:
                all_results.append(r)
        
        page += 1
    
    # 处理数据
    processed_results = []
    for item in all_results:
        title = item['title']
        
        # 判断公告类型，剔除结果公示/中标公告
        purchase_type = _get_nudt_purchase_type(title)
        if purchase_type is None:
            continue
        
        # 匹配关键词
        matched_kw = [kw for kw in keywords if kw in title]
        
        # 只保留匹配的项目
        if matched_kw:
            recommend = _get_nudt_recommend_level(matched_kw)
            remarks = _get_nudt_remarks(title)
            
            processed_results.append({
                '项目名称': title,
                '公告类型': purchase_type,
                '详情链接': item['href'],
                '发布日期': item['date'],
                '匹配关键词': ','.join(matched_kw),
                '推荐等级': recommend,
                '备注': remarks
            })
    
    # 创建DataFrame
    df = pd.DataFrame(processed_results)
    
    if len(df) > 0:
        df.index = range(1, len(df) + 1)
        df.index.name = '序号'
        df.reset_index(inplace=True)
    
    print(f"[INFO] 国防科大采购信息网: 共 {len(df)} 条匹配项目")
    
    return df


# ============ 整合输出Excel函数 ============

def _get_weain_latest_date() -> str:
    """获取全军武器装备采购信息网的最新发布日期"""
    try:
        result = subprocess.run(['date', '+%s%3N'], capture_output=True, text=True)
        _t = result.stdout.strip()
        
        # 获取公开招标第一页，检查最新日期
        url = f"https://www.weain.mil.cn/api/front/list/cggg/list?LMID=1149231276155707394&pageNo=1&purchaseType=%E5%85%AC%E5%BC%80%E6%8B%9B%E6%A0%87&_t={_t}"
        resp = requests.get(url, timeout=10, proxies=get_proxies())
        data = resp.json()
        content_list = data.get('list', {}).get('contentList', [])
        
        if content_list:
            publish_time = content_list[0].get('publishTime', '')[:10]
            return publish_time
    except:
        pass
    return datetime.now().strftime('%Y-%m-%d')


def _get_nudt_latest_date() -> str:
    """获取国防科大采购信息网的最新发布日期"""
    try:
        url = "https://www.nudt.edu.cn/cgxx/index.htm"
        resp = requests.get(url, timeout=10, proxies=get_proxies())
        resp.encoding = 'utf-8'
        html = resp.text
        
        # 匹配所有日期记录
        li_pattern = r'<li>\s*<a href="[^"]+">\s*<div class="date"><span class="day[^"]*">(\d+)</span><span class="year[^"]*">(\d+-\d+)</span></div>'
        matches = re.findall(li_pattern, html)
        
        if matches:
            # 找到最新的日期（列表是按时间正序排列的，第一条是最旧的，最后一条是最新的）
            latest = matches[-1]
            day = latest[0]
            year_month = latest[1]
            year = "20" + year_month.split("-")[0]
            month = year_month.split("-")[1]
            return f"{year}-{month}-{day}"
    except Exception as e:
        print(f"获取国防科大最新日期失败: {e}")
        pass
    return datetime.now().strftime('%Y-%m-%d')


def fetch_all_bidding(
    keywords: Optional[List[str]] = None,
    exclude_keywords: Optional[List[str]] = None,
    regions: Optional[Dict[str, str]] = None,
    date: Optional[str] = None,
    high_value_keywords: Optional[List[str]] = None,
    output_path: str = None,
    auto_latest: bool = True
) -> tuple:
    """
    整合获取全军武器装备采购信息网和军队采购网的商机信息，输出Excel
    
    Args:
        keywords: 核心关键词列表
        exclude_keywords: 排除关键词列表（仅军队采购网使用）
        regions: 地区字典（仅军队采购网使用）
        date: 日期，默认今天，如果auto_latest为True则自动获取各渠道最新日期
        high_value_keywords: 高价值关键词列表
        output_path: 输出Excel路径
        auto_latest: 是否自动获取各渠道最新日期（默认True）
    
    Returns:
        tuple: (df_weain, df_military, df_nudt) 三个DataFrame
    """
    if keywords is None:
        keywords = DEFAULT_CORE_KW
    if exclude_keywords is None:
        exclude_keywords = DEFAULT_EXCLUDE_KW
    if regions is None:
        regions = DEFAULT_REGIONS
    if high_value_keywords is None:
        high_value_keywords = DEFAULT_HIGH_VALUE_KW
    
    today = datetime.now().strftime('%Y-%m-%d')

    # 确定各渠道日期：auto_latest 时并行预检，否则直接用指定日期
    if auto_latest and date is None:
        print("[INFO] 自动检测各渠道最新日期（并行）...")
        with ThreadPoolExecutor(max_workers=2) as ex:
            f_weain_date = ex.submit(_get_weain_latest_date)
            f_nudt_date = ex.submit(_get_nudt_latest_date)
            weain_date = f_weain_date.result()
            nudt_date = f_nudt_date.result()
        military_date = today
        file_date = today

        print(f"[INFO] 全军装备采购网最新: {weain_date}")
        print(f"[INFO] 军队采购网最新: {military_date}")
        print(f"[INFO] 国防科大采购网最新: {nudt_date}")
        print("=" * 60)
    else:
        if date is None:
            date = today
        weain_date = date
        military_date = date
        nudt_date = date
        file_date = date
        print("=" * 60)
        print(f"开始整合抓取 - 日期: {date}")
        print("=" * 60)

    if output_path is None:
        output_dir = os.path.expanduser(get_output_dir())
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f'军队采购商机汇总_{file_date}.xlsx')

    # 并行抓取三个数据源
    print("[INFO] 并行抓取三个数据源...")
    with ThreadPoolExecutor(max_workers=3) as ex:
        f_weain = ex.submit(fetch_weain_bidding, keywords, weain_date)
        f_military = ex.submit(fetch_military_bidding, keywords, exclude_keywords,
                               regions, military_date, high_value_keywords, False)
        f_nudt = ex.submit(fetch_nudt_bidding, keywords, nudt_date)
        df_weain = f_weain.result()
        df_military = f_military.result()
        df_nudt = f_nudt.result()

    weain_actual_date = weain_date if len(df_weain) > 0 else ""
    nudt_actual_date = nudt_date if len(df_nudt) > 0 else ""

    print()
    
    # 输出Excel
    _save_to_excel(df_weain, df_military, df_nudt, output_path)
    
    # 返回各渠道实际使用的日期
    return df_weain, df_military, df_nudt, {
        'weain': weain_actual_date,
        'military': military_date,
        'nudt': nudt_actual_date
    }


def _save_to_excel(df_weain: pd.DataFrame, df_military: pd.DataFrame, df_nudt: pd.DataFrame, output_path: str):
    """保存到Excel文件（三个sheet）"""
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 创建工作簿
    wb = Workbook()
    
    # ========== Sheet 1: 全军武器装备采购信息网 ==========
    ws1 = wb.active
    ws1.title = "全军武器装备采购信息网"
    
    # 写入标题行 - 使用自己的字段
    if len(df_weain) > 0:
        weain_headers = list(df_weain.columns)
        for col, header in enumerate(weain_headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        for row_idx, row_data in enumerate(df_weain.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws1.cell(row=row_idx, column=col_idx, value=value).border = thin_border
        
        # 调整列宽
        ws1.column_dimensions['A'].width = 8
        ws1.column_dimensions['B'].width = 60
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 10
        ws1.column_dimensions['E'].width = 70
        ws1.column_dimensions['F'].width = 12
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['H'].width = 20
        ws1.column_dimensions['I'].width = 10
        ws1.column_dimensions['J'].width = 30
    
    # ========== Sheet 2: 军队采购网 ==========
    ws2 = wb.create_sheet(title="军队采购网")
    
    # 写入标题行 - 使用自己的字段
    if len(df_military) > 0:
        military_headers = list(df_military.columns)
        for col, header in enumerate(military_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        for row_idx, row_data in enumerate(df_military.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws2.cell(row=row_idx, column=col_idx, value=value).border = thin_border
        
        # 调整列宽
        ws2.column_dimensions['A'].width = 8
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 35
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 12
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 60
        ws2.column_dimensions['H'].width = 12
        ws2.column_dimensions['I'].width = 20
        ws2.column_dimensions['J'].width = 10
        ws2.column_dimensions['K'].width = 30
    
    # ========== Sheet 3: 国防科大采购信息网 ==========
    ws3 = wb.create_sheet(title="国防科大采购信息网")
    
    # 写入标题行 - 使用自己的字段
    if len(df_nudt) > 0:
        nudt_headers = list(df_nudt.columns)
        for col, header in enumerate(nudt_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        for row_idx, row_data in enumerate(df_nudt.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws3.cell(row=row_idx, column=col_idx, value=value).border = thin_border
        
        # 调整列宽
        ws3.column_dimensions['A'].width = 8
        ws3.column_dimensions['B'].width = 60
        ws3.column_dimensions['C'].width = 12
        ws3.column_dimensions['D'].width = 70
        ws3.column_dimensions['E'].width = 12
        ws3.column_dimensions['F'].width = 20
        ws3.column_dimensions['G'].width = 10
        ws3.column_dimensions['H'].width = 30
    
    # 保存文件
    wb.save(output_path)
    print(f"[INFO] Excel已保存: {output_path}")


# ============ 兼容旧接口 ============

def fetch_with_custom_params(
    region_codes: Dict[str, str],
    target_date: str,
    custom_keywords: List[str],
    custom_exclude: List[str] = None
) -> pd.DataFrame:
    """
    自定义参数抓取接口
    
    Args:
        region_codes: 地区编码字典
        target_date: 目标日期 YYYY-MM-DD
        custom_keywords: 自定义关键词列表
        custom_exclude: 自定义排除词列表
    
    Returns:
        DataFrame: 筛选结果
    """
    return fetch_military_bidding(
        keywords=custom_keywords,
        exclude_keywords=custom_exclude,
        regions=region_codes,
        date=target_date,
        save_csv=True
    )


HELP_TEXT = """军工采购商机抓取工具

用法:
    python fetcher.py [选项]

选项:
    --today                 抓取今日数据
    --yesterday             抓取昨日数据（默认）
    --date DATE             抓取指定日期，格式 YYYY-MM-DD
    --keywords WORDS        核心关键词，逗号分隔，如: 模型,仿真,AI
    --exclude-keywords WORDS 排除关键词，逗号分隔，如: 服装,医疗
    --high-value-keywords WORDS 高价值关键词，逗号分隔
    --regions REGIONS       地区，逗号分隔，如: 北京,湖北,湖南
    --output PATH           输出文件路径
    --no-auto-latest       禁用自动获取最新日期
    --help                 显示此帮助信息

示例:
    python fetcher.py                     # 抓取昨日（默认）
    python fetcher.py --today             # 抓取今日
    python fetcher.py --date 2026-03-23  # 抓取指定日期
    python fetcher.py --keywords "模型,仿真" --exclude-keywords "服装,医疗"
"""


def main():
    import argparse
    import sys

    parser = argparse.ArgumentParser(
        description='军工采购商机抓取工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=HELP_TEXT
    )
    group = parser.add_mutually_exclusive_group()
    group.add_argument('--today', action='store_true', help='抓取今日数据')
    group.add_argument('--yesterday', action='store_true', help='抓取昨日数据（默认）')
    group.add_argument('--date', type=str, help='抓取指定日期，格式 YYYY-MM-DD')

    parser.add_argument('--keywords', type=str,
                       help='核心关键词，逗号分隔')
    parser.add_argument('--exclude-keywords', type=str,
                       help='排除关键词，逗号分隔')
    parser.add_argument('--high-value-keywords', type=str,
                       help='高价值关键词，逗号分隔')
    parser.add_argument('--regions', type=str,
                       help='地区，逗号分隔')
    parser.add_argument('--output', type=str,
                       help='输出文件路径')
    parser.add_argument('--no-auto-latest', action='store_true',
                       help='禁用自动获取最新日期')

    args = parser.parse_args()

    # 无参数或 --help 时显示帮助
    if len(sys.argv) == 1 or '--help' in sys.argv:
        print(HELP_TEXT)
        sys.exit(0)

    # 确定日期
    target_date = None
    auto_latest = not args.no_auto_latest

    if args.today:
        target_date = datetime.now().strftime('%Y-%m-%d')
        auto_latest = False
    elif args.yesterday:
        from datetime import timedelta
        target_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        auto_latest = False
    elif args.date:
        target_date = args.date

    # 解析关键词
    keywords = None
    if args.keywords:
        keywords = args.keywords.split(',')

    exclude_keywords = None
    if args.exclude_keywords:
        exclude_keywords = args.exclude_keywords.split(',')

    high_value_keywords = None
    if args.high_value_keywords:
        high_value_keywords = args.high_value_keywords.split(',')

    # 解析地区
    regions = None
    if args.regions:
        region_map = {
            '北京': '110000', '天津': '120000', '河北': '130000',
            '山西': '140000', '内蒙古': '150000', '辽宁': '210000',
            '吉林': '220000', '黑龙江': '230000', '上海': '310000',
            '江苏': '320000', '浙江': '330000', '安徽': '340000',
            '福建': '350000', '江西': '360000', '山东': '370000',
            '河南': '410000', '湖北': '420000', '湖南': '430000',
            '广东': '440000', '广西': '450000', '海南': '460000',
            '重庆': '500000', '四川': '510000', '贵州': '520000',
            '云南': '530000', '西藏': '540000', '陕西': '610000',
            '甘肃': '620000', '青海': '630000', '宁夏': '640000',
            '新疆': '650000'
        }
        region_names = args.regions.split(',')
        regions = {name: region_map.get(name, '') for name in region_names if name in region_map}

    # 执行抓取
    result = fetch_all_bidding(
        keywords=keywords,
        exclude_keywords=exclude_keywords,
        regions=regions,
        date=target_date,
        high_value_keywords=high_value_keywords,
        output_path=args.output,
        auto_latest=auto_latest
    )

    if len(result) == 4:
        df_weain, df_military, df_nudt, dates_info = result
    else:
        df_weain, df_military, df_nudt = result

    print("\n" + "=" * 60)
    print("[RESULT] 汇总结果:")
    print(f"  全军武器装备采购信息网: {len(df_weain)} 条")
    print(f"  军队采购网: {len(df_military)} 条")
    print(f"  国防科大采购信息网: {len(df_nudt)} 条")
    print("=" * 60)


# CLI入口
if __name__ == '__main__':
    main()
