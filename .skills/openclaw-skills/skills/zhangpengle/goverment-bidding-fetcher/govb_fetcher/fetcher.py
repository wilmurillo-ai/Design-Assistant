"""
govb-fetcher 核心抓取逻辑。

数据源：
  - 北京中建云智政府采购网（zbcg-bjzc.zhongcy.com）
"""

import argparse
import os
import re
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from govb_fetcher.config import (
    get_keywords, get_exclude_keywords, get_high_value_keywords,
    get_bjzc_bearer_token, get_bjzc_tbsession, get_bjzc_jsessionid, get_bjzc_alb_route,
    get_output_dir, save_to_env, get_proxies,
)

_DETAIL_WORKERS = 5       # 详情并发数，避免触发服务端限速
_PRINT_LOCK = threading.Lock()

BASE_URL = 'http://zbcg-bjzc.zhongcy.com/gt-jy-toubiao/api'
DETAIL_BASE = 'http://zbcg-bjzc.zhongcy.com/bjczj-jy-toubiao/index.html'

HNZC_LIST_URL   = 'http://www.ccgp-hunan.gov.cn/mvc/getNoticeList4Web.do'
HNZC_DETAIL_URL = 'http://www.ccgp-hunan.gov.cn/mvc/viewNoticeContent.do'
HNZC_PAGE_URL   = 'http://www.ccgp-hunan.gov.cn/page/notice/notice.jsp'


# ──────────────────────────────────────────────
# Session 管理
# ──────────────────────────────────────────────

def _build_session() -> requests.Session:
    session = requests.Session()
    proxies = get_proxies()
    if proxies:
        session.proxies.update(proxies)
    session.cookies.update({
        'YGCG_TBSESSION': get_bjzc_tbsession(),
        'JSESSIONID': get_bjzc_jsessionid(),
        'jcloud_alb_route': get_bjzc_alb_route(),
    })
    session.headers.update({
        'Accept': 'application/json, text/plain, */*',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Authorization': f'Bearer {get_bjzc_bearer_token()}',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Origin': 'http://zbcg-bjzc.zhongcy.com',
        'Pragma': 'no-cache',
        'Referer': 'http://zbcg-bjzc.zhongcy.com/bjczj-jy-toubiao/index.html',
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/146.0.0.0 Safari/537.36'
        ),
        'contentType': 'formType',
    })
    return session


def _refresh_session_cookie(session: requests.Session) -> None:
    """清除重复的 YGCG_TBSESSION，保留最新值，并同步写回 .env。"""
    values = [c.value for c in session.cookies if c.name == 'YGCG_TBSESSION']
    if len(values) > 1:
        latest = values[-1]
        others = [c for c in session.cookies if c.name != 'YGCG_TBSESSION']
        session.cookies.clear()
        for c in others:
            session.cookies.set(c.name, c.value)
        session.cookies.set('YGCG_TBSESSION', latest)
        # 自动写回 .env，保证下次启动时使用最新值
        try:
            save_to_env({'FETCHER_BJZC_TBSESSION': latest})
        except Exception:
            pass


def _t() -> int:
    return int(time.time())


# ──────────────────────────────────────────────
# 北京中建云智政府采购网
# ──────────────────────────────────────────────

def _fetch_bjzc_page(session: requests.Session, page: int, rows: int = 100) -> dict:
    url = f'{BASE_URL}/cggg/gonggao/queryZBGongGaoList.do'
    data = {
        'ggName': '', 'gcBH': '', 'gcName': '',
        'bdBH': '', 'bdName': '', 'xmStatus': '',
        'page': str(page), 'rows': str(rows),
    }
    resp = session.post(url, data=data, timeout=15)
    _refresh_session_cookie(session)
    return resp.json()


def _fetch_bjzc_gg_bd_list(session: requests.Session, gg_guid: str) -> list:
    """
    获取公告下的分包列表（queryGgBdList）。
    每条记录含 bdGuid、文件获取时间、开标时间等关键字段。
    """
    url = f'{BASE_URL}/cggg/gonggao/queryGgBdList.do'
    params = {'_t': _t(), 'ggGuid': gg_guid}
    try:
        resp = session.get(url, params=params, timeout=15)
        _refresh_session_cookie(session)
        data = resp.json()
        if isinstance(data, dict):
            # 响应结构：{"data": {"ggBdList": [...]}}
            inner = data.get('data', {})
            if isinstance(inner, dict):
                bd_list = inner.get('ggBdList', [])
                if isinstance(bd_list, list):
                    return bd_list
    except Exception as e:
        print(f'  [warn] 获取分包列表失败 ggGuid={gg_guid}: {e}')
    return []


def _fetch_bjzc_purchaser_info(session: requests.Session, gc_guid: str) -> dict:
    """
    获取采购人及代理机构信息（queryPurchaserInfo）。
    确认字段：zbRName、lianXiRenPhone、zbDLName、zbDLZBFuZeRenMobile/zbDLZBFuZeRenPhone
    """
    url = f'{BASE_URL}/cggg/gonggao/queryPurchaserInfo.do'
    params = {'_t': _t(), 'gcGuid': gc_guid}
    try:
        resp = session.get(url, params=params, timeout=15)
        _refresh_session_cookie(session)
        data = resp.json()
        if isinstance(data, dict) and isinstance(data.get('data'), dict):
            return data['data']
    except Exception as e:
        print(f'  [warn] 获取采购人信息失败 gcGuid={gc_guid}: {e}')
    return {}


def _fetch_bjzc_project_overview(session: requests.Session, gc_guid: str, gg_guid: str) -> str:
    """
    获取项目概况文本（projectOverview）。
    尝试字段：ggNeiRong、zbGkfw、zbRequire（按优先级取第一个非空值）
    """
    url = f'{BASE_URL}/cggg/gonggao/projectOverview.do'
    params = {'_t': _t(), 'gcGuid': gc_guid, 'ggGuid': gg_guid}
    try:
        resp = session.get(url, params=params, timeout=15)
        _refresh_session_cookie(session)
        data = resp.json()
        inner = data.get('data', {}) if isinstance(data, dict) else {}
        text = ''
        for field in ('ggNeiRong', 'zbGkfw', 'zbRequire', 'bgContent'):
            val = inner.get(field)
            if isinstance(val, str) and val.strip():
                text = val.strip()
                break
        text = re.sub(r'<[^>]+>', '', text)
        text = re.sub(r'&[a-zA-Z#0-9]+;', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        return text[:100] if text else ''
    except Exception as e:
        print(f'  [warn] 获取项目概况失败 gcGuid={gc_guid}: {e}')
    return ''


def _build_detail_url(bd_guid: str, gc_guid: str, gg_guid: str) -> str:
    return (
        f'{DETAIL_BASE}#/steps/noticepageyg'
        f'?bdGuid={bd_guid}&gcGuid={gc_guid}&ggGuid={gg_guid}'
    )


def _ts_to_date(ts_ms: Optional[int]) -> str:
    if not ts_ms:
        return ''
    try:
        return datetime.fromtimestamp(ts_ms / 1000).strftime('%Y-%m-%d')
    except Exception:
        return ''


def _ts_to_datetime(ts_ms) -> str:
    if not ts_ms:
        return ''
    try:
        return datetime.fromtimestamp(int(ts_ms) / 1000).strftime('%Y-%m-%d %H:%M')
    except Exception:
        return str(ts_ms)


def _extract_field(d: dict, *keys) -> str:
    """从字典中按优先顺序取第一个非空字符串值。"""
    for k in keys:
        v = d.get(k)
        if v is not None and str(v).strip() and str(v).strip() not in ('null', 'None', '0'):
            return str(v).strip()
    return ''


# ──────────────────────────────────────────────
# 湖南政府采购网（HNZC）
# ──────────────────────────────────────────────

def _build_hnzc_session() -> requests.Session:
    session = requests.Session()
    proxies = get_proxies()
    if proxies:
        session.proxies.update(proxies)
    session.headers.update({
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/146.0.0.0 Safari/537.36'
        ),
        'X-Requested-With': 'XMLHttpRequest',
    })
    return session


def _fetch_hnzc_page(session: requests.Session, target_date: str, page: int, page_size: int = 100) -> dict:
    data = {
        'nType': 'prcmNotices', 'pType': '', 'prcmPrjName': '',
        'prcmItemCode': '', 'prcmOrgName': '', 'startDate': target_date,
        'endDate': target_date, 'prcmPlanNo': '',
        'page': str(page), 'pageSize': str(page_size),
    }
    resp = session.post(HNZC_LIST_URL, data=data, timeout=15)
    return resp.json()


def _fetch_hnzc_detail(session: requests.Session, notice_id, notice_category_code: str) -> dict:
    """抓取湖南政采详情 HTML，用 regex 提取关键字段。"""
    try:
        resp = session.get(
            HNZC_DETAIL_URL,
            params={'noticeId': notice_id, 'area_id': '', 'isKJXY': 'null'},
            timeout=15,
        )
        resp.encoding = 'utf-8'
        html = resp.text

        # 去掉 script/style 块
        html = re.sub(r'<(script|style)[^>]*>.*?</\1>', '', html, flags=re.S)
        # 去掉所有 HTML 标签
        text = re.sub(r'<[^>]+>', '', html)
        # 合并连续空白（保留换行以便定位章节）
        text = re.sub(r'[ \t]+', ' ', text)

        def _rex(pattern, default=''):
            m = re.search(pattern, text, re.S)
            return m.group(1).strip() if m else default

        # 采购预算（两种写法均有）
        budget = _rex(r'(?:采购项目预算|采购预算)[：:]\s*([\d,]+)元')

        # 文件获取时间（第五章节 "有意参加投标者，于...至..."）
        file_start = _rex(r'于(\d{4}年\d{1,2}月\d{1,2}日)\s*至')
        file_end   = _rex(r'于.+?至(\d{4}年\d{1,2}月\d{1,2}日)')

        # 开标时间（第六章节 "3、开标时间："）
        open_bid = _rex(r'开标时间[：:]\s*(\d{4}年\d{1,2}月\d{1,2}日\s*\d{2}:\d{2})')

        # 采购人电话（十一章 采购人信息 (5)电话）
        purchaser_phone = _rex(r'1、采购人信息.+?（5）电\s+话[：:]\s*([^\s（\n]+)')

        # 代理机构名称
        agency_name = _rex(r'2、采购代理机构信息.+?（1）名\s+称[：:]\s*([^\n（]+)')

        # 代理机构电话
        agency_phone = _rex(r'2、采购代理机构信息.+?（5）电\s+话[：:]\s*([^\s（\n]+)')

        return {
            'budget': budget,
            'file_start': file_start,
            'file_end': file_end,
            'open_bid': open_bid,
            'purchaser_phone': purchaser_phone,
            'agency_name': agency_name.strip(),
            'agency_phone': agency_phone,
        }
    except Exception as e:
        print(f'  [warn] 获取湖南详情失败 noticeId={notice_id}: {e}')
        return {}


def fetch_hnzc_bidding(
    session: requests.Session,
    target_date: str,
    keywords: list,
    exclude_kw: list,
    high_value_kw: list,
    fetch_detail: bool = True,
) -> list:
    """抓取湖南政府采购网指定日期的公告，先过滤关键词，再对匹配项查详情。"""
    import math

    print(f'[hnzc] 抓取日期 {target_date}，逐页获取列表...')

    # 1. 翻页抓取
    raw_rows = []
    page = 1
    total_pages = 1
    while page <= total_pages:
        print(f'  第 {page} 页...', end=' ', flush=True)
        try:
            result = _fetch_hnzc_page(session, target_date, page)
        except Exception as e:
            print(f'请求失败: {e}，停止')
            break

        rows = result.get('rows') or []
        total = result.get('total') or 0
        total_pages = math.ceil(total / 100) if total > 0 else 1
        print(f'{len(rows)} 条 (共 {total} 条)')

        raw_rows.extend(rows)
        if len(rows) < 100 or page >= total_pages:
            break
        page += 1

    print(f'[hnzc] 当日原始记录 {len(raw_rows)} 条，开始关键词过滤...')

    # 2. 关键词过滤
    filtered = []
    for row in raw_rows:
        title = row.get('NOTICE_TITLE', '') or ''
        if any(ex in title for ex in exclude_kw):
            continue
        matched = [kw for kw in keywords if kw in title]
        if not matched:
            continue
        row['_matched_kw'] = matched
        filtered.append(row)

    print(f'[hnzc] 过滤后剩余 {len(filtered)} 条，{"开始补全详情..." if fetch_detail else "跳过详情补全"}')

    # 3. 构建基础记录
    def _build_base(row: dict) -> dict:
        notice_id    = row.get('NOTICE_ID', '')
        cat_code     = row.get('NOTICE_CATEGORY_CODE', '')
        title        = row.get('NOTICE_TITLE', '')
        matched_kw   = row.get('_matched_kw', [])
        detail_link  = f'{HNZC_PAGE_URL}?noticeId={notice_id}&noticeTypeCode={cat_code}'
        return {
            '项目名称':        title,
            '标段名称':        '',
            '招标方式':        row.get('PRCM_MODE_NAME', ''),
            '合同估价(元)':    '',
            '文件获取开始时间': '',
            '文件获取截止时间': '',
            '开标时间':        '',
            '采购人':          row.get('ORG_NAME', ''),
            '采购人电话':      '',
            '代理机构':        '',
            '详情链接':        detail_link,
            '发布日期':        row.get('NEWWORK_DATE', ''),
            '匹配关键词':      ','.join(matched_kw),
            '推荐等级':        _get_recommendation(title, matched_kw, high_value_kw),
            '备注':            _generate_remarks(title, matched_kw),
        }

    if not fetch_detail:
        return [_build_base(row) for row in filtered]

    # 并行补全详情
    def _fetch_one(args):
        i, row = args
        notice_id = row.get('NOTICE_ID', '')
        cat_code  = row.get('NOTICE_CATEGORY_CODE', '')
        title     = row.get('NOTICE_TITLE', '')
        record    = _build_base(row)
        d = _fetch_hnzc_detail(session, notice_id, cat_code)
        record.update({
            '合同估价(元)':    d.get('budget', ''),
            '文件获取开始时间': d.get('file_start', ''),
            '文件获取截止时间': d.get('file_end', ''),
            '开标时间':        d.get('open_bid', ''),
            '采购人电话':      d.get('purchaser_phone', ''),
            '代理机构':        d.get('agency_name', ''),
        })
        with _PRINT_LOCK:
            print(f'  [{i + 1}/{len(filtered)}] 完成: {title[:30]}...')
        return i, record

    ordered: list = [None] * len(filtered)
    with ThreadPoolExecutor(max_workers=_DETAIL_WORKERS) as executor:
        for i, record in executor.map(_fetch_one, enumerate(filtered)):
            ordered[i] = record

    return [r for r in ordered if r is not None]


# ──────────────────────────────────────────────
# 关键词过滤与评级
# ──────────────────────────────────────────────

def _filter_by_keywords(rows: list, keywords: list, exclude_kw: list) -> list:
    results = []
    for row in rows:
        title = row.get('ggName', '') or ''
        bd_name = row.get('bdNames', '') or ''
        combined = title + bd_name

        if any(ex in combined for ex in exclude_kw):
            continue
        matched = [kw for kw in keywords if kw in combined]
        if not matched:
            continue
        row['_matched_kw'] = matched
        results.append(row)
    return results


def _get_recommendation(title: str, matched_kw: list, high_value_kw: list) -> str:
    if '意向' in title:
        return '高'
    if any(kw in matched_kw for kw in high_value_kw):
        return '高'
    if '系统' in matched_kw:
        return '中'
    return '低'


def _generate_remarks(title: str, matched_kw: list) -> str:
    remarks = []
    if '意向' in title:
        remarks.append('采购意向前期可跟进')
    if '体系' in title:
        remarks.append('体系设计相关')
    if '模型' in title:
        remarks.append('模型开发可参与')
    if '仿真' in title:
        remarks.append('仿真项目可做')
    if '数据' in title:
        remarks.append('数据分析需求')
    if 'AI' in title or '智能' in title:
        remarks.append('AI智能项目')
    if '软件' in title:
        remarks.append('软件定制开发')
    if '系统' in title:
        remarks.append('系统集成类')
    if '算法' in title:
        remarks.append('算法研发类')
    if '平台' in title:
        remarks.append('平台建设类')
    if '决策' in title or '规划' in title:
        remarks.append('决策规划类')
    if not remarks:
        remarks.append('可关注')
    return '; '.join(remarks[:2])


# ──────────────────────────────────────────────
# 北京中建云智 主抓取流程
# ──────────────────────────────────────────────

def fetch_bjzc_bidding(
    session: requests.Session,
    target_date: str,
    keywords: list,
    exclude_kw: list,
    high_value_kw: list,
    fetch_detail: bool = True,
) -> list:
    """
    抓取北京中建云智政府采购网指定日期的公告，
    先过滤关键词，再对匹配项查询详情。
    """
    # 日期时间戳范围
    d = datetime.strptime(target_date, '%Y-%m-%d')
    day_start = int(datetime(d.year, d.month, d.day, 0, 0, 0).timestamp() * 1000)
    day_end   = int(datetime(d.year, d.month, d.day, 23, 59, 59).timestamp() * 1000)

    # 1. 抓取当日列表
    print(f'[bjzc] 抓取日期 {target_date}，逐页获取列表...')
    raw_rows = []
    page = 1
    while True:
        print(f'  第 {page} 页...', end=' ', flush=True)
        result = _fetch_bjzc_page(session, page)
        if not result.get('success') or not result.get('data', {}).get('rows'):
            print('无数据，停止')
            break
        rows = result['data']['rows']
        print(f'{len(rows)} 条')

        stop = False
        for row in rows:
            ts = row.get('ggStartTime')
            if ts is None:
                continue
            if ts < day_start:
                stop = True
                break
            if day_start <= ts <= day_end:
                raw_rows.append(row)

        if stop or len(rows) < 100:
            break
        page += 1

    print(f'[bjzc] 当日原始记录 {len(raw_rows)} 条，开始关键词过滤...')

    # 2. 关键词过滤
    filtered = _filter_by_keywords(raw_rows, keywords, exclude_kw)
    print(f'[bjzc] 过滤后剩余 {len(filtered)} 条，{"开始补全详情..." if fetch_detail else "跳过详情补全"}')

    # 3. 详情补全
    if not fetch_detail:
        results = []
        for row in filtered:
            title = row.get('ggName', '')
            matched_kw = row.get('_matched_kw', [])
            results.append({
                '项目名称': title,
                '标段名称': row.get('bdNames', ''),
                '招标方式': row.get('zbFangShiName', ''),
                '合同估价(元)': '',
                '文件获取开始时间': '',
                '文件获取截止时间': '',
                '开标时间': '',
                '采购人': '',
                '采购人电话': '',
                '代理机构': '',
                '详情链接': '',
                '发布日期': _ts_to_date(row.get('ggStartTime')),
                '匹配关键词': ','.join(matched_kw),
                '推荐等级': _get_recommendation(title, matched_kw, high_value_kw),
                '备注': _generate_remarks(title, matched_kw),
            })
        return results

    # 串行补全详情（bjzc 使用滚动式 session cookie，并发请求会导致旧 token 失效，字段为空）
    results = []
    for idx, row in enumerate(filtered, 1):
        gg_guid = row.get('ggGuid', '')
        gc_guid = row.get('gcGuid', '')
        title = row.get('ggName', '')
        matched_kw = row.get('_matched_kw', [])
        publish_date = _ts_to_date(row.get('ggStartTime'))

        base_record = {
            '项目名称': title,
            '标段名称': row.get('bdNames', ''),
            '招标方式': row.get('zbFangShiName', ''),
            '合同估价(元)': '',
            '文件获取开始时间': '',
            '文件获取截止时间': '',
            '开标时间': '',
            '采购人': '',
            '采购人电话': '',
            '代理机构': '',
            '详情链接': '',
            '发布日期': publish_date,
            '匹配关键词': ','.join(matched_kw),
            '推荐等级': _get_recommendation(title, matched_kw, high_value_kw),
            '备注': _generate_remarks(title, matched_kw),
        }

        print(f'  [{idx}/{len(filtered)}] 补全详情: {title[:30]}...')

        # Step1: queryGgBdList — 分包列表 + 文件获取时间 + 开标时间
        bd_list = _fetch_bjzc_gg_bd_list(session, gg_guid)
        if not bd_list:
            bd_list = [{}]

        # Step2: queryPurchaserInfo — 采购人 + 代理机构（每个公告只查一次）
        purchaser_info = _fetch_bjzc_purchaser_info(session, gc_guid) if gc_guid else {}
        purchaser_name  = purchaser_info.get('zbRName', '')
        purchaser_phone = purchaser_info.get('lianXiRenPhone', '')
        agency_name     = purchaser_info.get('zbDLName', '')
        agency_phone    = (purchaser_info.get('zbDLZBFuZeRenMobile')
                           or purchaser_info.get('zbDLZBFuZeRenPhone') or '')

        for bd in bd_list:
            bd_guid = bd.get('bdGuid', '')
            file_start = _ts_to_datetime(bd.get('zbWJHuoQuStartTime'))
            file_end   = _ts_to_datetime(bd.get('zbWJHuoQuEndTime'))
            open_bid   = _ts_to_datetime(bd.get('kbTime'))
            detail_url = _build_detail_url(bd_guid, gc_guid, gg_guid) if bd_guid else ''

            record = dict(base_record)
            contract_price = bd.get('bdHeTongGuJia')
            # 接口返回单位为分，转换为元并格式化为逗号分隔
            if contract_price:
                try:
                    yuan = int(contract_price) // 100
                    contract_price_fmt = f'{yuan:,}'
                except (ValueError, TypeError):
                    contract_price_fmt = str(contract_price)
            else:
                contract_price_fmt = ''
            record.update({
                '标段名称': bd.get('bdName') or base_record['标段名称'],
                '合同估价(元)': contract_price_fmt,
                '文件获取开始时间': file_start,
                '文件获取截止时间': file_end,
                '开标时间': open_bid,
                '采购人': purchaser_name,
                '采购人电话': purchaser_phone,
                '代理机构': agency_name,
                '详情链接': detail_url,
            })
            results.append(record)

    return results


# ──────────────────────────────────────────────
# 汇总入口
# ──────────────────────────────────────────────

def fetch_all_bidding(
    target_date: str,
    keywords: list = None,
    exclude_kw: list = None,
    high_value_kw: list = None,
    fetch_detail: bool = True,
) -> dict:
    keywords    = keywords    or get_keywords()
    exclude_kw  = exclude_kw  or get_exclude_keywords()
    high_value_kw = high_value_kw or get_high_value_keywords()

    def _run_bjzc():
        return fetch_bjzc_bidding(
            _build_session(), target_date, keywords, exclude_kw, high_value_kw, fetch_detail
        )

    def _run_hnzc():
        return fetch_hnzc_bidding(
            _build_hnzc_session(), target_date, keywords, exclude_kw, high_value_kw, fetch_detail
        )

    with ThreadPoolExecutor(max_workers=2) as executor:
        f_bjzc = executor.submit(_run_bjzc)
        f_hnzc = executor.submit(_run_hnzc)
        bjzc_results = f_bjzc.result()
        hnzc_results = f_hnzc.result()

    return {
        '北京政采': bjzc_results,
        '湖南政采': hnzc_results,
    }


# ──────────────────────────────────────────────
# Excel 输出
# ──────────────────────────────────────────────

COLUMNS = [
    ('序号',           8),
    ('项目名称',        55),
    ('标段名称',        40),
    ('招标方式',        12),
    ('合同估价(元)',    15),
    ('文件获取开始时间', 18),
    ('文件获取截止时间', 18),
    ('开标时间',        18),
    ('采购人',          25),
    ('采购人电话',      15),
    ('代理机构',        25),
    ('详情链接',        80),
    ('发布日期',        12),
    ('匹配关键词',      20),
    ('推荐等级',        10),
    ('备注',            30),
]

_HEADER_FONT  = Font(bold=True, color='FFFFFF')
_HEADER_FILL  = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_CELL_ALIGN   = Alignment(vertical='center', wrap_text=True)
_THIN_BORDER  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
_HIGH_FILL = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
_MID_FILL  = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
_LOW_FILL  = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')


def _active_columns(rows: list) -> list:
    """返回至少有一行有值的列（序号列始终保留）。用于 --no-detail 模式。"""
    result = []
    for col_name, col_width in COLUMNS:
        if col_name == '序号':
            result.append((col_name, col_width))
            continue
        if any(record.get(col_name) for record in rows):
            result.append((col_name, col_width))
    return result


def _write_sheet(ws, rows: list, columns: list = None) -> None:
    cols = columns if columns is not None else COLUMNS
    # 表头
    for col_idx, (col_name, col_width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = col_width
    ws.row_dimensions[1].height = 20

    # 数据行
    for row_idx, record in enumerate(rows, 2):
        recommendation = record.get('推荐等级', '')
        row_fill = (_HIGH_FILL if recommendation == '高' else
                    _MID_FILL  if recommendation == '中' else
                    _LOW_FILL  if recommendation == '低' else None)

        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        for col_idx, (col_name, _) in enumerate(cols[1:], 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ''))
            cell.alignment = _CELL_ALIGN
            cell.border = _THIN_BORDER
            if row_fill:
                cell.fill = row_fill

        # 序号列样式
        seq_cell = ws.cell(row=row_idx, column=1)
        seq_cell.alignment = Alignment(horizontal='center', vertical='center')
        seq_cell.border = _THIN_BORDER
        if row_fill:
            seq_cell.fill = row_fill

    ws.freeze_panes = 'A2'


def save_to_excel(all_results: dict, output_path: Path, no_detail: bool = False) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in all_results.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # sheet名最长31字符
        cols = _active_columns(rows) if no_detail else COLUMNS
        _write_sheet(ws, rows, cols)

    wb.save(output_path)
    print(f'\n[done] Excel 已保存: {output_path}')


# ──────────────────────────────────────────────
# 数据源凭证注册表
# 新增数据源时在此处添加一条记录即可
# ──────────────────────────────────────────────

# SOURCE_COOKIE_MAP[source_id] = {cookie_key_in_session_str: env_var_name}
SOURCE_COOKIE_MAP = {
    'bjzc': {
        '_bearer':        'FETCHER_BJZC_BEARER_TOKEN',
        'YGCG_TBSESSION': 'FETCHER_BJZC_TBSESSION',
        'JSESSIONID':     'FETCHER_BJZC_JSESSIONID',
        'jcloud_alb_route': 'FETCHER_BJZC_ALB_ROUTE',
    },
    # 后续新增示例：
    # 'ccgp': {
    #     '_bearer':   'FETCHER_CCGP_BEARER_TOKEN',
    #     'SESSION_ID': 'FETCHER_CCGP_SESSION',
    # },
}


# --set-cookie 子命令
def cmd_set_cookie(source: str, bearer: str, session_str: str) -> None:
    """解析并写入指定数据源的凭证到 .env 文件。"""
    source = source.lower().strip()
    if source not in SOURCE_COOKIE_MAP:
        known = ', '.join(SOURCE_COOKIE_MAP.keys())
        print(f'[error] 未知数据源 "{source}"，当前支持: {known}')
        sys.exit(1)

    cookie_map = SOURCE_COOKIE_MAP[source]
    updates = {}

    if bearer:
        token = bearer.strip()
        if token.lower().startswith('bearer '):
            token = token[7:].strip()
        if '_bearer' in cookie_map:
            updates[cookie_map['_bearer']] = token

    if session_str:
        for part in session_str.split(';'):
            part = part.strip()
            if '=' not in part:
                continue
            k, _, v = part.partition('=')
            k, v = k.strip(), v.strip()
            if k in cookie_map:
                updates[cookie_map[k]] = v

    if not updates:
        print('[error] 未能解析任何凭证，请检查 --bearer 和 --session 参数。')
        sys.exit(1)

    env_path = save_to_env(updates)
    print(f'[ok] [{source}] 凭证已写入: {env_path}')
    for k, v in updates.items():
        print(f'  {k} = {v[:12]}...' if len(v) > 12 else f'  {k} = {v}')


# ──────────────────────────────────────────────
# CLI 入口
# ──────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        prog='govb-fetcher',
        description='政府采购商机自动抓取工具',
    )

    # 凭证更新子命令
    parser.add_argument('--set-cookie', action='store_true',
                        help='更新 .env 中的凭证信息（不执行抓取）')
    parser.add_argument('--source', default='bjzc',
                        help=f'数据源标识，与 --set-cookie 配合使用，可选: {", ".join(SOURCE_COOKIE_MAP.keys())}（默认: bjzc）')
    parser.add_argument('--bearer', default='',
                        help='Bearer token，格式："Bearer xxx" 或 "xxx"')
    parser.add_argument('--session', default='',
                        help='Cookie 字符串，格式："YGCG_TBSESSION=xxx; JSESSIONID=xxx; ..."')

    # 日期参数（互斥）
    date_group = parser.add_mutually_exclusive_group()
    date_group.add_argument('--today', action='store_true', help='抓取今日数据')
    date_group.add_argument('--yesterday', action='store_true', help='抓取昨日数据（默认）')
    date_group.add_argument('--date', metavar='YYYY-MM-DD', help='抓取指定日期数据')

    # 过滤参数
    parser.add_argument('--keywords', help='核心关键词，逗号分隔')
    parser.add_argument('--exclude-keywords', dest='exclude_keywords',
                        help='排除关键词，逗号分隔')
    parser.add_argument('--high-value-keywords', dest='high_value_keywords',
                        help='高价值关键词，逗号分隔')

    # 输出参数
    parser.add_argument('--output', help='指定输出 Excel 路径')
    parser.add_argument('--no-detail', dest='no_detail', action='store_true',
                        help='跳过详情 API，仅保存列表字段（速度更快）')

    args = parser.parse_args()

    # 凭证更新模式
    if args.set_cookie:
        cmd_set_cookie(args.source, args.bearer, args.session)
        return

    # 确定目标日期（默认昨日）
    if args.today:
        target_date = datetime.now().strftime('%Y-%m-%d')
    elif args.date:
        target_date = args.date
    else:
        target_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')

    # 关键词
    keywords        = [k.strip() for k in args.keywords.split(',')] if args.keywords else None
    exclude_kw      = [k.strip() for k in args.exclude_keywords.split(',')] if args.exclude_keywords else None
    high_value_kw   = [k.strip() for k in args.high_value_keywords.split(',')] if args.high_value_keywords else None

    print(f'govb-fetcher  日期: {target_date}  详情补全: {"否" if args.no_detail else "是"}')
    print(f'关键词: {keywords or get_keywords()}')

    # 执行抓取
    all_results = fetch_all_bidding(
        target_date=target_date,
        keywords=keywords,
        exclude_kw=exclude_kw,
        high_value_kw=high_value_kw,
        fetch_detail=not args.no_detail,
    )

    # 统计
    print('\n[result]')
    total = 0
    for source, rows in all_results.items():
        print(f'  {source}: {len(rows)} 条')
        total += len(rows)
    print(f'  合计: {total} 条')

    if total == 0:
        print('[info] 无匹配数据，不生成 Excel。')
        return

    # 输出路径
    if args.output:
        output_path = Path(args.output)
    else:
        output_dir = get_output_dir()
        output_path = output_dir / f'政府采购商机汇总_{target_date}.xlsx'

    save_to_excel(all_results, output_path, no_detail=args.no_detail)


if __name__ == '__main__':
    main()
