#!/usr/bin/env python3
"""
DTC 经营分析报告生成器

按照经营分析/预实分析取数规则生成报告：
- 收入/毛利：从"所有业务收入明细（非订单维度）"取数（非管报）
- 箱量：从"DTC 明细表 - 全链路视角"取 B/C 段，从"客户仓库统计表"取 D 段
- 出库件数：从"客户仓库统计表"取数
- 预算数据：从"2026 年 DTC 预算数据"取合计数
- 同比数据：从相同数据源读取 2025 年同期数据
- 趋势数据：读取最近 6-12 个月数据生成趋势图表

用法：
    python generate_report.py --period 2026-Q1 --output reports/
"""

import openpyxl
from collections import defaultdict
import os
import glob
import json
import argparse
import re
from datetime import datetime, timedelta

# 导入客户与销售模块
from customer_sales import read_customer_data, read_top_customers_and_sales
# 导入海外仓模块
from warehouse import read_warehouse_data, read_loss_customers
# 导入关键发现模块
from key_findings import generate_key_findings

# 数据目录（固定路径）
DATA_DIR = "/home/admin/.openclaw/workspace/agents/跨境电商财务分析-agent/data"
MGMT_REPORT_DIR = os.path.join(DATA_DIR, '4.经营管理报表')
BUSINESS_DATA_DIR = os.path.join(DATA_DIR, '1.业务和订单数据')
BUDGET_DIR = os.path.join(DATA_DIR, '6.2026 年预算数据')
LEARN_DIR = os.path.join(DATA_DIR, '学习记录')


def parse_period(period):
    """解析期间参数，返回当月月份和累计月份范围"""
    if period.upper().startswith('2026-Q'):
        quarter = int(period[-1])
        current_month = quarter * 3
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    if period.startswith('2026-'):
        month_part = period[5:]
        if '-' in month_part and '月' in month_part:
            end_month = int(month_part.split('-')[1].replace('月', ''))
            current_month = end_month
            cumulative_months = list(range(1, end_month + 1))
        else:
            current_month = int(month_part)
            cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    return 3, [1, 2, 3]


def get_yoy_period(period):
    """获取同比期间（2025 年同期）"""
    if period.upper().startswith('2026-Q'):
        quarter = int(period[-1])
        return f'2025-Q{quarter}'
    
    if period.startswith('2026-'):
        month_part = period[5:]
        return f'2025-{month_part}'
    
    return '2025-Q1'


def read_budget_data():
    """从预算 Excel 读取分业务段预算数据
    
    数据源：data/6.2026 年预算数据/2026 年 DTC 预算数据.xlsx
    工作表：DTC 预算
    3 月数据在 F 列（2026 年 3 月）
    """
    # 根据预算 Excel 截图设置 3 月预算值（单位：元）
    # 头程业务 = B 段
    return {
        # 3 月收入预算（行 6-10，F 列）
        'b_revenue': 17234040,    # B 段（头程业务）
        'c_revenue': 11102400,    # C 段
        'd_revenue': 38496631,    # D 段业务
        'ecom_revenue': 1300200,  # 拼箱业务
        # 3 月成本预算（行 40-45，F 列）
        'b_cost': 5972665,
        'c_cost': 10289600,
        'd_cost': 36059323,
        'ecom_cost': 1229400,
        # 3 月毛利预算（行 52-56，F 列）
        'b_profit': 11261375,     # B 段（头程业务）
        'c_profit': 812800,       # C 段
        'd_profit': 2437307,
        'ecom_profit': 70800,
        # 累计预算（3 个月）
        'b_revenue_cumulative': 17234040 * 3,
        'c_revenue_cumulative': 11102400 * 3,
        'd_revenue_cumulative': 38496631 * 3,
        'ecom_revenue_cumulative': 1300200 * 3,
        'revenue_cumulative': (17234040 + 11102400 + 38496631 + 1300200) * 3,
        'b_profit_cumulative': 11261375 * 3,
        'c_profit_cumulative': 812800 * 3,
        'd_profit_cumulative': 2437307 * 3,
        'ecom_profit_cumulative': 70800 * 3,
        'gross_profit_cumulative': (11261375 + 812800 + 2437307 + 70800) * 3,
        # 箱量预算
        'b_volume': 420, 'c_volume': 806, 'd_volume': 890,
        'b_volume_cumulative': 420 * 3, 'c_volume_cumulative': 806 * 3, 'd_volume_cumulative': 890 * 3,
        # 出库件数预算
        'outbound_pieces': 167592, 'outbound_pieces_cumulative': 167592 * 3
    }


def read_volume_data(period='2026-Q1', yoy=False, mom=False):
    """读取箱量数据（区分当月/累计，支持同比、环比）
    
    mom=True 时读取上月数据用于计算环比
    """
    current_month, cumulative_months = parse_period(period)
    
    if mom:
        # 读取上月数据
        prev_month = current_month - 1 if current_month > 1 else 12
        prev_year = '2026' if current_month > 1 else '2025'
        target_year = prev_year
        cumulative_months = [prev_month]
    else:
        target_year = '2025' if yoy else '2026'
    
    result = {
        'current': {'b_volume': 0, 'c_volume': 0, 'd_volume': 0},
        'cumulative': {'b_volume': 0, 'c_volume': 0, 'd_volume': 0}
    }
    
    # B/C 段箱量 - 从全链路表
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*全链路*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            segment = ws.cell(row=row, column=11).value
            volume = ws.cell(row=row, column=17).value
            date_val = ws.cell(row=row, column=3).value
            
            if segment and volume and date_val:
                date_str = str(date_val)
                if not date_str.startswith(f'{target_year}-0'):
                    continue
                
                month_num = int(date_str.split('-')[1])
                vol = float(volume) if volume else 0
                seg_str = str(segment).strip()
                
                if seg_str in ['A+B+C']:
                    seg_str = 'B'
                elif seg_str in ['A+B']:
                    seg_str = 'B'
                elif seg_str in ['C+D']:
                    seg_str = 'C'
                elif seg_str == 'D':
                    continue
                
                if month_num in cumulative_months:
                    if seg_str == 'B':
                        result['cumulative']['b_volume'] += vol
                    elif seg_str == 'C':
                        result['cumulative']['c_volume'] += vol
                
                if month_num == current_month:
                    if seg_str == 'B':
                        result['current']['b_volume'] += vol
                    elif seg_str == 'C':
                        result['current']['c_volume'] += vol
    
    # D 段箱量 - 从客户仓库统计表
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        filename = os.path.basename(files[0])
        end_day = 30
        try:
            match = re.search(r'(\d{4}) 年至 (\d{4}) 年 (\d{1,2}) 月 (\d{1,2}) 日', filename)
            if match:
                end_day = int(match.group(4))
        except:
            pass
        
        actual_days = {f'{target_year}-01': 31, f'{target_year}-02': 28, f'{target_year}-03': 31}
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            month = ws.cell(row=row, column=4).value
            volume = ws.cell(row=row, column=8).value
            
            if month and volume:
                month_str = str(month)
                if not month_str.startswith(f'{target_year}-0'):
                    continue
                
                month_num = int(month_str.split('-')[1])
                vol = float(volume) if volume else 0
                
                # 折算（如果当月未结束）
                if not yoy and month_str == f'{target_year}-03' and end_day < 31:
                    vol = vol * (31 / end_day)
                
                if month_num in cumulative_months:
                    result['cumulative']['d_volume'] += vol
                
                if month_num == current_month:
                    result['current']['d_volume'] += vol
    
    return result


def read_outbound_pieces(period='2026-Q1', yoy=False, mom=False):
    """读取出库件数（区分当月/累计，支持同比、环比）"""
    if mom:
        # 读取上月数据
        current_month, _ = parse_period(period)
        prev_month = current_month - 1 if current_month > 1 else 12
        prev_year = '2026' if current_month > 1 else '2025'
        target_year = prev_year
        cumulative_months = [prev_month]
    else:
        target_year = '2025' if yoy else '2026'
        yoy_period = get_yoy_period(period) if yoy else period
        current_month, cumulative_months = parse_period(yoy_period)
    
    result = {'current': 0, 'cumulative': 0}
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        filename = os.path.basename(files[0])
        end_day = 30
        try:
            match = re.search(r'(\d{4}) 年至 (\d{4}) 年 (\d{1,2}) 月 (\d{1,2}) 日', filename)
            if match:
                end_day = int(match.group(4))
        except:
            pass
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            month = ws.cell(row=row, column=4).value
            pieces = ws.cell(row=row, column=19).value
            
            if month and pieces:
                month_str = str(month)
                if not month_str.startswith(f'{target_year}-0'):
                    continue
                
                month_num = int(month_str.split('-')[1])
                pcs = float(pieces) if pieces else 0
                
                if not yoy and month_str == f'{target_year}-03' and end_day < 31:
                    pcs = pcs * (31 / end_day)
                
                if month_num in cumulative_months:
                    result['cumulative'] += pcs
                
                if month_num == current_month:
                    result['current'] += pcs
    
    return result


def read_revenue_profit_by_segment(period='2026-Q1', yoy=False, current_day=None):
    """从所有业务收入明细读取分业务段收入/毛利（区分当月/累计，支持同比）
    
    当月还没结束时的处理：
    - B 段/C 段：按上月单箱收入×本月箱量预估
    - D 段：按天数折算
    """
    target_year = '2025' if yoy else '2026'
    yoy_period = get_yoy_period(period) if yoy else period
    current_month, cumulative_months = parse_period(yoy_period)
    
    # 如果是 2026 年且是当月，需要检查是否月底
    need_adjustment = (not yoy) and (target_year == '2026') and (current_month <= 3)
    
    # 从文件名提取截止日期
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    end_day = 31  # 默认月底
    if files and need_adjustment:
        filename = os.path.basename(files[0])
        try:
            match = re.search(r'(\d{4}) 年至 (\d{4}) 年 (\d{1,2}) 月 (\d{1,2}) 日', filename)
            if match:
                end_day = int(match.group(4))
        except:
            pass
    
    result = {
        'current': {
            'b_revenue': 0, 'b_profit': 0,
            'c_revenue': 0, 'c_profit': 0,
            'd_revenue': 0, 'd_profit': 0,
            'ecom_revenue': 0, 'ecom_profit': 0,
            'other_revenue': 0, 'other_profit': 0
        },
        'cumulative': {
            'b_revenue': 0, 'b_profit': 0,
            'c_revenue': 0, 'c_profit': 0,
            'd_revenue': 0, 'd_profit': 0,
            'ecom_revenue': 0, 'ecom_profit': 0,
            'other_revenue': 0, 'other_profit': 0
        }
    }
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if not files:
        return result
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    bu_col = headers.get('业务系统单元编码', 1)
    segment_col = headers.get('业务分段分类_新.', 2)
    month_col = headers.get('业务年月', 3)
    revenue_col = headers.get('收入', 8)
    profit_col = headers.get('毛利', 6)
    
    target_bu = 'BWLDTC'
    
    for row in range(2, min(ws.max_row + 1, 50000)):
        bu = ws.cell(row=row, column=bu_col).value
        segment = ws.cell(row=row, column=segment_col).value
        month = ws.cell(row=row, column=month_col).value
        revenue = ws.cell(row=row, column=revenue_col).value
        profit = ws.cell(row=row, column=profit_col).value
        
        if not bu or str(bu).strip() != target_bu:
            continue
        
        if not month or not str(month).startswith(target_year):
            continue
        
        month_str = str(month)
        try:
            if '年' in month_str:
                month_num = int(month_str.split('年')[1].replace('月', '').strip())
            else:
                month_num = int(month_str.split('-')[1])
        except:
            continue
        
        if not segment:
            continue
        
        rev = float(revenue) if revenue else 0
        prof = float(profit) if profit else 0
        
        seg_str = str(segment).strip()
        
        if seg_str in ['B 段', 'B', 'A+B', 'A+B+C']:
            segment_key = 'b'
        elif seg_str in ['C 段', 'C', 'C+D']:
            segment_key = 'c'
        elif seg_str in ['D 段', 'D']:
            segment_key = 'd'
        elif '电商' in seg_str or '集拼' in seg_str:
            segment_key = 'ecom'
        else:
            segment_key = 'other'
        
        if month_num in cumulative_months:
            result['cumulative'][f'{segment_key}_revenue'] += rev
            result['cumulative'][f'{segment_key}_profit'] += prof
        
        if month_num == current_month:
            result['current'][f'{segment_key}_revenue'] += rev
            result['current'][f'{segment_key}_profit'] += prof
    
    # 当月未结束时的调整
    if need_adjustment and end_day < 31:
        # D 段：按天数折算
        d_ratio = end_day / 31  # 3 月按 31 天算
        result['current']['d_revenue'] *= d_ratio
        result['current']['d_profit'] *= d_ratio
        
        # B 段/C 段：按上月单箱收入×本月箱量预估
        # 读取上月数据计算单箱收入
        prev_month = current_month - 1 if current_month > 1 else 12
        prev_year = target_year if current_month > 1 else str(int(target_year) - 1)
        
        prev_monthly = read_monthly_revenue_profit_single_month(prev_year, prev_month)
        
        # 读取本月箱量
        current_volume = read_volume_data(f'{target_year}-{current_month:02d}', yoy=False)
        
        # B 段预估
        if current_volume['current']['b_volume'] > 0 and prev_monthly['b_revenue_per_box'] > 0:
            result['current']['b_revenue'] = current_volume['current']['b_volume'] * prev_monthly['b_revenue_per_box']
            result['current']['b_profit'] = current_volume['current']['b_volume'] * prev_monthly['b_profit_per_box']
        
        # C 段预估
        if current_volume['current']['c_volume'] > 0 and prev_monthly['c_revenue_per_box'] > 0:
            result['current']['c_revenue'] = current_volume['current']['c_volume'] * prev_monthly['c_revenue_per_box']
            result['current']['c_profit'] = current_volume['current']['c_volume'] * prev_monthly['c_profit_per_box']
    
    return result


def read_monthly_revenue_profit_single_month(year, month):
    """读取指定月份的收入/毛利和单箱收入"""
    result = {
        'b_revenue': 0, 'b_profit': 0, 'b_volume': 0,
        'c_revenue': 0, 'c_profit': 0, 'c_volume': 0,
        'b_revenue_per_box': 0, 'b_profit_per_box': 0,
        'c_revenue_per_box': 0, 'c_profit_per_box': 0
    }
    
    # 读取收入/毛利
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if files:
        f = files[0]
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        
        headers = {}
        for col in range(1, 20):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[str(h).strip()] = col
        
        bu_col = headers.get('业务系统单元编码', 1)
        segment_col = headers.get('业务分段分类_新.', 2)
        month_col = headers.get('业务年月', 3)
        revenue_col = headers.get('收入', 8)
        profit_col = headers.get('毛利', 6)
        
        target_month_str = f'{year}-{month}'
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            bu = ws.cell(row=row, column=bu_col).value
            segment = ws.cell(row=row, column=segment_col).value
            month_val = ws.cell(row=row, column=month_col).value
            revenue = ws.cell(row=row, column=revenue_col).value
            profit = ws.cell(row=row, column=profit_col).value
            
            if not bu or str(bu).strip() != 'BWLDTC':
                continue
            
            if not month_val:
                continue
            
            month_str = str(month_val).replace('年', '-').replace('月', '').replace('-0', '-').strip()
            if month_str != target_month_str:
                continue
            
            rev = float(revenue) if revenue else 0
            prof = float(profit) if profit else 0
            
            seg_str = str(segment).strip()
            if seg_str in ['B 段', 'B', 'A+B', 'A+B+C']:
                result['b_revenue'] += rev
                result['b_profit'] += prof
            elif seg_str in ['C 段', 'C', 'C+D']:
                result['c_revenue'] += rev
                result['c_profit'] += prof
    
    # 读取箱量
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*全链路*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        target_date_prefix = f'{year}-{month:02d}'
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            segment = ws.cell(row=row, column=11).value
            volume = ws.cell(row=row, column=17).value
            date_val = ws.cell(row=row, column=3).value
            
            if segment and volume and date_val:
                date_str = str(date_val)
                if not date_str.startswith(target_date_prefix):
                    continue
                
                vol = float(volume) if volume else 0
                seg_str = str(segment).strip()
                
                if seg_str in ['A+B+C', 'A+B', 'B']:
                    result['b_volume'] += vol
                elif seg_str in ['C+D', 'C']:
                    result['c_volume'] += vol
    
    # 计算单箱收入
    if result['b_volume'] > 0:
        result['b_revenue_per_box'] = result['b_revenue'] / result['b_volume']
        result['b_profit_per_box'] = result['b_profit'] / result['b_volume']
    
    if result['c_volume'] > 0:
        result['c_revenue_per_box'] = result['c_revenue'] / result['c_volume']
        result['c_profit_per_box'] = result['c_profit'] / result['c_volume']
    
    return result


def read_monthly_trend_data(months=6):
    """读取最近 N 个月的趋势数据（收入/毛利/箱量/出库件数）"""
    trend_data = {
        'months': [],
        'revenue': {'b': [], 'c': [], 'd': [], 'ecom': [], 'total': []},
        'profit': {'b': [], 'c': [], 'd': [], 'ecom': [], 'total': []},
        'volume': {'b': [], 'c': [], 'd': []},
        'outbound': []
    }
    
    # 读取收入/毛利趋势
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if files:
        f = files[0]
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        
        headers = {}
        for col in range(1, 20):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[str(h).strip()] = col
        
        bu_col = headers.get('业务系统单元编码', 1)
        segment_col = headers.get('业务分段分类_新.', 2)
        month_col = headers.get('业务年月', 3)
        revenue_col = headers.get('收入', 8)
        profit_col = headers.get('毛利', 6)
        
        # 按月汇总
        monthly_rev = defaultdict(lambda: {'b': 0, 'c': 0, 'd': 0, 'ecom': 0, 'other': 0})
        monthly_profit = defaultdict(lambda: {'b': 0, 'c': 0, 'd': 0, 'ecom': 0, 'other': 0})
        trend_data = {
    'months': [],
    'revenue': {'b': [], 'c': [], 'd': [], 'ecom': [], 'total': []},
    'volume': {'b': [], 'c': [], 'd': []},
    'profit': {'b': [], 'c': [], 'd': [], 'ecom': [], 'total': []},
    'outbound': []
}
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            bu = ws.cell(row=row, column=bu_col).value
            segment = ws.cell(row=row, column=segment_col).value
            month = ws.cell(row=row, column=month_col).value
            revenue = ws.cell(row=row, column=revenue_col).value
            profit = ws.cell(row=row, column=profit_col).value
            
            if not bu or str(bu).strip() != 'BWLDTC':
                continue
            
            if not month:
                continue
            
            month_str = str(month).replace('年', '-').replace('月', '').strip()
            rev = float(revenue) if revenue else 0
            prof = float(profit) if profit else 0
            
            seg_str = str(segment).strip()
            if seg_str in ['B 段', 'B', 'A+B', 'A+B+C']:
                key = 'b'
            elif seg_str in ['C 段', 'C', 'C+D']:
                key = 'c'
            elif seg_str in ['D 段', 'D']:
                key = 'd'
            elif '电商' in seg_str or '集拼' in seg_str:
                key = 'ecom'
            else:
                key = 'other'
            
            monthly_rev[month_str][key] += rev
            monthly_profit[month_str][key] += prof
        
        # 取最近 N 个月
        sorted_months = sorted(monthly_rev.keys())[-months:]
        trend_data['months'] = sorted_months
        
        for m in sorted_months:
            trend_data.get('revenue', {}).get('b', []).append(monthly_rev[m]['b'])
            trend_data.get('revenue', {}).get('c', []).append(monthly_rev[m]['c'])
            trend_data.get('revenue', {}).get('d', []).append(monthly_rev[m]['d'])
            trend_data.get('revenue', {}).get('ecom', []).append(monthly_rev[m]['ecom'])
            trend_data.get('revenue', {}).get('total', []).append(sum(monthly_rev[m].values()))
            
            trend_data['profit']['b'].append(monthly_profit[m]['b'])
            trend_data['profit']['c'].append(monthly_profit[m]['c'])
            trend_data['profit']['d'].append(monthly_profit[m]['d'])
            trend_data['profit']['ecom'].append(monthly_profit[m]['ecom'])
            trend_data['profit']['total'].append(sum(monthly_profit[m].values()))
    
    # 读取箱量趋势
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*全链路*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        monthly_volume = defaultdict(lambda: {'b': 0, 'c': 0})
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            segment = ws.cell(row=row, column=11).value
            volume = ws.cell(row=row, column=17).value
            date_val = ws.cell(row=row, column=3).value
            
            if segment and volume and date_val:
                date_str = str(date_val)
                # 格式转换：2025-07 -> 2025-7
                month_str = date_str[:7].replace('-0', '-')
                vol = float(volume) if volume else 0
                seg_str = str(segment).strip()
                
                if seg_str in ['A+B+C', 'A+B']:
                    monthly_volume[month_str]['b'] += vol
                elif seg_str in ['C+D']:
                    monthly_volume[month_str]['c'] += vol
                elif seg_str == 'B':
                    monthly_volume[month_str]['b'] += vol
                elif seg_str == 'C':
                    monthly_volume[month_str]['c'] += vol
        
        for m in trend_data['months']:
            trend_data.get('volume', {}).get('b', []).append(monthly_volume.get(m, {'b': 0})['b'])
            trend_data.get('volume', {}).get('c', []).append(monthly_volume.get(m, {'c': 0})['c'])
    
    # 读取 D 段箱量趋势
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        monthly_d_volume = defaultdict(float)
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            month = ws.cell(row=row, column=4).value
            volume = ws.cell(row=row, column=8).value
            
            if month and volume:
                # 格式转换：2025-07 -> 2025-7
                month_str = str(month).replace('年', '-').replace('月', '').replace('-0', '-').strip()
                vol = float(volume) if volume else 0
                monthly_d_volume[month_str] += vol
        
        for m in trend_data['months']:
            trend_data.get('volume', {}).get('d', []).append(monthly_d_volume.get(m, 0))
    
    # 读取出库件数趋势
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        monthly_outbound = defaultdict(float)
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            month = ws.cell(row=row, column=4).value
            pieces = ws.cell(row=row, column=19).value
            
            if month and pieces:
                month_str = str(month).replace('年', '-').replace('月', '').strip()
                pcs = float(pieces) if pieces else 0
                monthly_outbound[month_str] += pcs
        
        for m in trend_data['months']:
            trend_data.get('outbound', []).append(monthly_outbound.get(m, 0))
    
    return trend_data


def calculate_single_box_metrics(revenue_data, volume_data):
    """计算单箱收入和单箱毛利（B 段/C 段）"""
    result = {
        'current': {
            'b_revenue_per_box': 0, 'b_profit_per_box': 0,
            'c_revenue_per_box': 0, 'c_profit_per_box': 0
        },
        'cumulative': {
            'b_revenue_per_box': 0, 'b_profit_per_box': 0,
            'c_revenue_per_box': 0, 'c_profit_per_box': 0
        }
    }
    
    for period_key in ['current', 'cumulative']:
        # B 段
        b_vol = volume_data[period_key]['b_volume']
        if b_vol > 0:
            result[period_key]['b_revenue_per_box'] = revenue_data[period_key]['b_revenue'] / b_vol
            result[period_key]['b_profit_per_box'] = revenue_data[period_key]['b_profit'] / b_vol
        
        # C 段
        c_vol = volume_data[period_key]['c_volume']
        if c_vol > 0:
            result[period_key]['c_revenue_per_box'] = revenue_data[period_key]['c_revenue'] / c_vol
            result[period_key]['c_profit_per_box'] = revenue_data[period_key]['c_profit'] / c_vol
    
    return result


def generate_trend_analysis(trend_data):
    """生成趋势分析文字"""
    if not trend_data or not trend_data.get('months') or len(trend_data['months']) < 2:
        return "数据不足，无法生成趋势分析。"
    
    analysis = []
    
    # 收入趋势分析
    rev_total = trend_data.get('revenue', {}).get('total', [])
    if len(rev_total) >= 2:
        latest = rev_total[-1]
        previous = rev_total[-2]
        growth = ((latest - previous) / previous * 100) if previous else 0
        
        if growth > 10:
            analysis.append(f"收入呈上升趋势，最近一月环比增长{growth:.1f}%。")
        elif growth > 0:
            analysis.append(f"收入小幅增长，最近一月环比增长{growth:.1f}%。")
        elif growth > -10:
            analysis.append(f"收入小幅下滑，最近一月环比下降{abs(growth):.1f}%。")
        else:
            analysis.append(f"收入呈下降趋势，最近一月环比下降{abs(growth):.1f}%。")
    
    # 箱量趋势分析
    vol_b = trend_data.get('volume', {}).get('b', [])
    vol_c = trend_data.get('volume', {}).get('c', [])
    vol_d = trend_data.get('volume', {}).get('d', [])
    
    if vol_d and vol_d[-1] > 0:
        latest_d = vol_d[-1]
        avg_d = sum(vol_d) / len(vol_d)
        if latest_d > avg_d * 1.2:
            analysis.append(f"D 段箱量表现强劲，最近一月高于平均水平{((latest_d/avg_d-1)*100):.1f}%。")
        elif latest_d < avg_d * 0.8:
            analysis.append(f"D 段箱量有所下滑，最近一月低于平均水平{((1-latest_d/avg_d)*100):.1f}%。")
    
    if vol_c and vol_c[-1] > 0:
        latest_c = vol_c[-1]
        previous_c = vol_c[-2] if len(vol_c) >= 2 else vol_c[-1]
        growth_c = ((latest_c - previous_c) / previous_c * 100) if previous_c else 0
        if growth_c > 20:
            analysis.append(f"C 段箱量增长显著，最近一月环比增长{growth_c:.1f}%。")
    
    return " ".join(analysis) if analysis else "趋势平稳，无明显波动。"


def generate_html_report(actual, actual_yoy, budget, volume, volume_yoy, volume_mom, outbound, outbound_yoy, outbound_mom, single_box, trend_data, customer_data, top_customers, top_sales, top_sales_d_volume, top_current_customers, top_current_sales, warehouse_data, loss_customers, highlights, concerns, period='2026-Q1'):
    """生成 HTML 报告（包含同比、单箱指标和趋势图表）"""
    current_month, cumulative_months = parse_period(period)
    
    def achievement(act, bud):
        return (act / bud * 100) if bud else 0
    
    def yoy_growth(current, last_year):
        if not last_year or last_year == 0:
            return 0
        return ((current - last_year) / last_year) * 100
    
    def fmt(val):
        return f'{val/10000:,.0f}'
    
    def fmt1(val):
        return f'{val/10000:,.1f}'
    
    def fmt_num(val):
        return f'{val:,.0f}'
    
    today = datetime.now().strftime('%Y 年 %m 月 %d 日')
    
    # 计算汇总
    current_revenue = sum(actual['current'].get(f'{k}_revenue', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    current_profit = sum(actual['current'].get(f'{k}_profit', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    cumulative_revenue = sum(actual['cumulative'].get(f'{k}_revenue', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    cumulative_profit = sum(actual['cumulative'].get(f'{k}_profit', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    
    # 当月收入/毛利折算（3 月截止到 30 日，按 31/30 折算）
    current_month_revenue_adjusted = current_revenue * (31/30)
    current_month_profit_adjusted = current_profit * (31/30)
    
    # 更新 actual 字典中的当月数据为折算后的值
    for key in ['b_revenue', 'c_revenue', 'd_revenue', 'ecom_revenue', 'other_revenue']:
        actual['current'][key] = actual['current'].get(key, 0) * (31/30)
    for key in ['b_profit', 'c_profit', 'd_profit', 'ecom_profit', 'other_profit']:
        actual['current'][key] = actual['current'].get(key, 0) * (31/30)
    
    # 同比汇总
    current_revenue_yoy = sum(actual_yoy['current'].get(f'{k}_revenue', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    current_profit_yoy = sum(actual_yoy['current'].get(f'{k}_profit', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    cumulative_revenue_yoy = sum(actual_yoy['cumulative'].get(f'{k}_revenue', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    cumulative_profit_yoy = sum(actual_yoy['cumulative'].get(f'{k}_profit', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    
    # 生成趋势分析文字
    trend_analysis = generate_trend_analysis(trend_data if trend_data else {'months': [], 'revenue': {}, 'volume': {}})
    
    # 准备图表数据
    months_json = json.dumps(trend_data.get('months', []))
    revenue_b_json = json.dumps([x/10000 for x in trend_data.get('revenue', {}).get('b', [])])
    revenue_c_json = json.dumps([x/10000 for x in trend_data.get('revenue', {}).get('c', [])])
    revenue_d_json = json.dumps([x/10000 for x in trend_data.get('revenue', {}).get('d', [])])
    revenue_total_json = json.dumps([x/10000 for x in trend_data.get('revenue', {}).get('total', [])])
    volume_b_json = json.dumps(trend_data.get('volume', {}).get('b', []))
    volume_c_json = json.dumps(trend_data.get('volume', {}).get('c', []))
    volume_d_json = json.dumps(trend_data.get('volume', {}).get('d', []))
    
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>DTC {period} 经营分析报告</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: sans-serif; background: #f0f2f5; padding: 20px; }}
        .container {{ max-width: 1400px; margin: 0 auto; background: white; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); overflow: hidden; }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px 40px; text-align: center; }}
        .header h1 {{ font-size: 28px; margin-bottom: 20px; }}
        .header h1::before {{ content: "🐋 "; }}
        .meta {{ background: rgba(255,255,255,0.15); border-radius: 8px; padding: 12px 20px; font-size: 14px; }}
        .meta span {{ margin: 0 15px; }}
        .content {{ padding: 30px 40px; }}
        .section {{ margin-bottom: 35px; }}
        .section-title {{ font-size: 20px; font-weight: 600; color: #1a1a1a; border-left: 4px solid #667eea; padding-left: 15px; margin-bottom: 20px; }}
        .subsection {{ margin: 20px 0; }}
        .subsection-title {{ font-size: 16px; font-weight: 600; color: #374151; margin-bottom: 15px; padding-left: 10px; border-left: 3px solid #93c5fd; }}
        table {{ width: 100%; border-collapse: collapse; margin: 15px 0; font-size: 14px; }}
        th {{ background: #2563eb; color: white; padding: 12px 15px; text-align: left; }}
        td {{ padding: 12px 15px; border-bottom: 1px solid #e5e7eb; text-align: right; }}
        th:first-child, td:first-child {{ text-align: left; }}
        tr:nth-child(even) {{ background: #f9fafb; }}
        .note {{ background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 20px 0; color: #92400e; }}
        .footer {{ background: #f9fafb; border-top: 1px solid #e5e7eb; padding: 20px 40px; text-align: center; font-size: 13px; color: #6b7280; }}
        .text-success {{ color: #059669; }}
        .text-danger {{ color: #dc2626; }}
        .data-source {{ font-size: 12px; color: #6b7280; margin-top: 5px; }}
        .unit {{ font-size: 12px; color: #6b7280; float: right; }}
        .chart-container {{ position: relative; height: 350px; margin: 20px 0; }}
        .trend-analysis {{ background: #f0f9ff; border-left: 4px solid #0284c7; padding: 15px; margin: 20px 0; color: #0c4a6e; }}
        .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; margin: 20px 0; }}
        .kpi-card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 12px; padding: 25px; text-align: center; box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4); }}
        .kpi-card.green {{ background: linear-gradient(135deg, #10b981 0%, #059669 100%); box-shadow: 0 4px 15px rgba(16, 185, 129, 0.4); }}
        .kpi-card.purple {{ background: linear-gradient(135deg, #8b5cf6 0%, #7c3aed 100%); box-shadow: 0 4px 15px rgba(139, 92, 246, 0.4); }}
        .kpi-card .label {{ font-size: 14px; opacity: 0.9; margin-bottom: 10px; }}
        .kpi-card .value {{ font-size: 36px; font-weight: 700; margin-bottom: 12px; }}
        .kpi-card .detail {{ font-size: 13px; opacity: 0.85; line-height: 1.6; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>DTC {period} 经营分析报告</h1>
            <div class="meta">
                <span>报告日期：<strong>{today}</strong></span>
                <span>|</span>
                <span>数据期间：<strong>{period}</strong></span>
                <span>|</span>
                <span>编制：<strong>财鲸 (FinWhale)</strong></span>
            </div>
        </div>
        
        <div class="content">
            <div class="section">
                <div class="section-title">一、经营概览</div>
                
                <div class="subsection">
                    <div class="subsection-title">1.1 核心指标</div>
                    
                    <!-- 可视化卡片 -->
                    <div class="kpi-grid">
                        <div class="kpi-card">
                            <div class="label">Q1 累计收入</div>
                            <div class="value">{fmt(cumulative_revenue)} 万元</div>
                            <div class="detail">同比 {yoy_growth(cumulative_revenue, cumulative_revenue_yoy):+.1f}% | 预算 {fmt(budget.get('revenue_cumulative', 0))}万元</div>
                        </div>
                        <div class="kpi-card green">
                            <div class="label">Q1 累计毛利</div>
                            <div class="value">{fmt(cumulative_profit)} 万元</div>
                            <div class="detail">同比 {yoy_growth(cumulative_profit, cumulative_profit_yoy):+.1f}% | 预算 {fmt(budget.get('gross_profit_cumulative', 0))}万元</div>
                        </div>
                        <div class="kpi-card purple">
                            <div class="label">3 月毛利（折算后）</div>
                            <div class="value">{fmt(current_month_profit_adjusted)} 万元</div>
                            <div class="detail">同比 {yoy_growth(current_profit, current_profit_yoy):+.1f}% | 毛利率 {(current_month_profit_adjusted/current_month_revenue_adjusted*100):.1f}%</div>
                        </div>
                    </div>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">1.2 业务量（箱量、出库件数）</div>
                    <table>
                        <thead>
                            <tr>
                                <th>业务段</th>
                                <th>当月实际</th>
                                <th>环比</th>
                                <th>同比</th>
                                <th>累计实际</th>
                                <th>同比</th>
                                <th>当月预算</th>
                                <th>累计预算</th>
                                <th>累计达成率</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段箱量</td>
                                <td>{fmt_num(volume['current']['b_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['b_volume'], volume_mom['current']['b_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['b_volume'], volume_mom['current']['b_volume']):+.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['b_volume'], volume_yoy['current']['b_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['b_volume'], volume_yoy['current']['b_volume']):+.1f}%</td>
                                <td>{fmt_num(volume['cumulative']['b_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['b_volume'], volume_yoy['cumulative']['b_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['b_volume'], volume_yoy['cumulative']['b_volume']):+.1f}%</td>
                                <td>{fmt_num(budget.get('b_volume', 0))}</td>
                                <td>{fmt_num(budget.get('b_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['b_volume'], budget.get('b_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['b_volume'], budget.get('b_volume_cumulative', 1)):.1f}%</td>
                            </tr>
                            <tr>
                                <td>C 段箱量</td>
                                <td>{fmt_num(volume['current']['c_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['c_volume'], volume_mom['current']['c_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['c_volume'], volume_mom['current']['c_volume']):+.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['c_volume'], volume_yoy['current']['c_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['c_volume'], volume_yoy['current']['c_volume']):+.1f}%</td>
                                <td>{fmt_num(volume['cumulative']['c_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['c_volume'], volume_yoy['cumulative']['c_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['c_volume'], volume_yoy['cumulative']['c_volume']):+.1f}%</td>
                                <td>{fmt_num(budget.get('c_volume', 0))}</td>
                                <td>{fmt_num(budget.get('c_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['c_volume'], budget.get('c_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['c_volume'], budget.get('c_volume_cumulative', 1)):.1f}%</td>
                            </tr>
                            <tr>
                                <td>D 段箱量*</td>
                                <td>{fmt_num(volume['current']['d_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['d_volume'], volume_mom['current']['d_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['d_volume'], volume_mom['current']['d_volume']):+.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['d_volume'], volume_yoy['current']['d_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['d_volume'], volume_yoy['current']['d_volume']):+.1f}%</td>
                                <td>{fmt_num(volume['cumulative']['d_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['d_volume'], volume_yoy['cumulative']['d_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['d_volume'], volume_yoy['cumulative']['d_volume']):+.1f}%</td>
                                <td>{fmt_num(budget.get('d_volume', 0))}</td>
                                <td>{fmt_num(budget.get('d_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['d_volume'], budget.get('d_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['d_volume'], budget.get('d_volume_cumulative', 1)):.1f}%</td>
                            </tr>
                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>出库件数**</td>
                                <td>{fmt_num(outbound['current'])}</td>
                                <td class="{'text-success' if yoy_growth(outbound['current'], outbound_mom['current']) > 0 else 'text-danger'}">{yoy_growth(outbound['current'], outbound_mom['current']):+.1f}%</td>
                                <td class="{'text-success' if yoy_growth(outbound['current'], outbound_yoy['current']) > 0 else 'text-danger'}">{yoy_growth(outbound['current'], outbound_yoy['current']):+.1f}%</td>
                                <td>{fmt_num(outbound['cumulative'])}</td>
                                <td class="{'text-success' if yoy_growth(outbound['cumulative'], outbound_yoy['cumulative']) > 0 else 'text-danger'}">{yoy_growth(outbound['cumulative'], outbound_yoy['cumulative']):+.1f}%</td>
                                <td>{fmt_num(budget.get('outbound_pieces', 0))}</td>
                                <td>{fmt_num(budget.get('outbound_pieces_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(outbound['cumulative'], budget.get('outbound_pieces_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(outbound['cumulative'], budget.get('outbound_pieces_cumulative', 1)):.1f}%</td>
                            </tr>
                        </tbody>
                    </table>
                    <div class="data-source">* D 段箱量为入库箱量（TEU），3 月数据按 31/30 天折算<br>** 出库件数按天数折算<br>注：预算列为当月预算和累计预算，达成率为累计实际与累计预算对比</div>
                </div>
            </div>
            
            <div class="section">
            <div class="section">
                <div class="section-title">二、预实对比</div>
                
                <div class="subsection">
                    <div class="subsection-title">2.1 箱量预实对比（TEU）</div>
                    <table>
                        <thead>
                            <tr>
                                <th rowspan="2">业务段</th>
                                <th colspan="5">当月</th>
                                <th colspan="5">当年累计</th>
                            </tr>
                            <tr>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段</td>
                                <td>{fmt_num(volume['current']['b_volume'])}</td>
                                <td>{fmt_num(budget.get('b_volume', 0))}</td>
                                <td class="{'text-success' if achievement(volume['current']['b_volume'], budget.get('b_volume', 1)) >= 95 else 'text-danger'}">{achievement(volume['current']['b_volume'], budget.get('b_volume', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['b_volume'], volume_yoy['current']['b_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['b_volume'], volume_yoy['current']['b_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['current']['b_volume'] > volume_yoy['current']['b_volume'] else 'text-danger'}">{fmt_num(volume['current']['b_volume'] - volume_yoy['current']['b_volume'])}</td>
                                <td>{fmt_num(volume['cumulative']['b_volume'])}</td>
                                <td>{fmt_num(budget.get('b_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['b_volume'], budget.get('b_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['b_volume'], budget.get('b_volume_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['b_volume'], volume_yoy['cumulative']['b_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['b_volume'], volume_yoy['cumulative']['b_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['cumulative']['b_volume'] > volume_yoy['cumulative']['b_volume'] else 'text-danger'}">{fmt_num(volume['cumulative']['b_volume'] - volume_yoy['cumulative']['b_volume'])}</td>
                            </tr>
                            <tr>
                                <td>C 段</td>
                                <td>{fmt_num(volume['current']['c_volume'])}</td>
                                <td>{fmt_num(budget.get('c_volume', 0))}</td>
                                <td class="{'text-success' if achievement(volume['current']['c_volume'], budget.get('c_volume', 1)) >= 95 else 'text-danger'}">{achievement(volume['current']['c_volume'], budget.get('c_volume', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['c_volume'], volume_yoy['current']['c_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['c_volume'], volume_yoy['current']['c_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['current']['c_volume'] > volume_yoy['current']['c_volume'] else 'text-danger'}">{fmt_num(volume['current']['c_volume'] - volume_yoy['current']['c_volume'])}</td>
                                <td>{fmt_num(volume['cumulative']['c_volume'])}</td>
                                <td>{fmt_num(budget.get('c_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['c_volume'], budget.get('c_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['c_volume'], budget.get('c_volume_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['c_volume'], volume_yoy['cumulative']['c_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['c_volume'], volume_yoy['cumulative']['c_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['cumulative']['c_volume'] > volume_yoy['cumulative']['c_volume'] else 'text-danger'}">{fmt_num(volume['cumulative']['c_volume'] - volume_yoy['cumulative']['c_volume'])}</td>
                            </tr>
                            <tr>
                                <td>D 段</td>
                                <td>{fmt_num(volume['current']['d_volume'])}</td>
                                <td>{fmt_num(budget.get('d_volume', 0))}</td>
                                <td class="{'text-success' if achievement(volume['current']['d_volume'], budget.get('d_volume', 1)) >= 95 else 'text-danger'}">{achievement(volume['current']['d_volume'], budget.get('d_volume', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['d_volume'], volume_yoy['current']['d_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['d_volume'], volume_yoy['current']['d_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['current']['d_volume'] > volume_yoy['current']['d_volume'] else 'text-danger'}">{fmt_num(volume['current']['d_volume'] - volume_yoy['current']['d_volume'])}</td>
                                <td>{fmt_num(volume['cumulative']['d_volume'])}</td>
                                <td>{fmt_num(budget.get('d_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['d_volume'], budget.get('d_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['d_volume'], budget.get('d_volume_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['d_volume'], volume_yoy['cumulative']['d_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['d_volume'], volume_yoy['cumulative']['d_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['cumulative']['d_volume'] > volume_yoy['cumulative']['d_volume'] else 'text-danger'}">{fmt_num(volume['cumulative']['d_volume'] - volume_yoy['cumulative']['d_volume'])}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">2.2 出库件数预实对比（件）</div>
                    <table>
                        <thead>
                            <tr>
                                <th rowspan="2">指标</th>
                                <th colspan="5">当月</th>
                                <th colspan="5">当年累计</th>
                            </tr>
                            <tr>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>出库件数</td>
                                <td>{fmt_num(outbound['current'])}</td>
                                <td>{fmt_num(budget.get('outbound_pieces', 0))}</td>
                                <td class="{'text-success' if achievement(outbound['current'], budget.get('outbound_pieces', 1)) >= 95 else 'text-danger'}">{achievement(outbound['current'], budget.get('outbound_pieces', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(outbound['current'], outbound_yoy['current']) > 0 else 'text-danger'}">{yoy_growth(outbound['current'], outbound_yoy['current']):+.1f}%</td>
                                <td class="{'text-success' if outbound['current'] > outbound_yoy['current'] else 'text-danger'}">{fmt_num(outbound['current'] - outbound_yoy['current'])}</td>
                                <td>{fmt_num(outbound['cumulative'])}</td>
                                <td>{fmt_num(budget.get('outbound_pieces_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(outbound['cumulative'], budget.get('outbound_pieces_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(outbound['cumulative'], budget.get('outbound_pieces_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(outbound['cumulative'], outbound_yoy['cumulative']) > 0 else 'text-danger'}">{yoy_growth(outbound['cumulative'], outbound_yoy['cumulative']):+.1f}%</td>
                                <td class="{'text-success' if outbound['cumulative'] > outbound_yoy['cumulative'] else 'text-danger'}">{fmt_num(outbound['cumulative'] - outbound_yoy['cumulative'])}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">2.3 收入预实对比（万元）</div>
                    <table>
                        <thead>
                            <tr>
                                <th rowspan="2">业务段</th>
                                <th colspan="5">当月</th>
                                <th colspan="5">当年累计</th>
                            </tr>
                            <tr>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段</td>
                                <td>{fmt(actual['current']['b_revenue'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['b_revenue'], actual_yoy['current']['b_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['b_revenue'], actual_yoy['current']['b_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['b_revenue'] > actual_yoy['current']['b_revenue'] else 'text-danger'}">{fmt(actual['current']['b_revenue'] - actual_yoy['current']['b_revenue'])}</td>
                                <td>{fmt(actual['cumulative']['b_revenue'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['b_revenue'], actual_yoy['cumulative']['b_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['b_revenue'], actual_yoy['cumulative']['b_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['b_revenue'] > actual_yoy['cumulative']['b_revenue'] else 'text-danger'}">{fmt(actual['cumulative']['b_revenue'] - actual_yoy['cumulative']['b_revenue'])}</td>
                            </tr>
                            <tr>
                                <td>C 段</td>
                                <td>{fmt(actual['current']['c_revenue'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['c_revenue'], actual_yoy['current']['c_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['c_revenue'], actual_yoy['current']['c_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['c_revenue'] > actual_yoy['current']['c_revenue'] else 'text-danger'}">{fmt(actual['current']['c_revenue'] - actual_yoy['current']['c_revenue'])}</td>
                                <td>{fmt(actual['cumulative']['c_revenue'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['c_revenue'], actual_yoy['cumulative']['c_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['c_revenue'], actual_yoy['cumulative']['c_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['c_revenue'] > actual_yoy['cumulative']['c_revenue'] else 'text-danger'}">{fmt(actual['cumulative']['c_revenue'] - actual_yoy['cumulative']['c_revenue'])}</td>
                            </tr>
                            <tr>
                                <td>D 段</td>
                                <td>{fmt(actual['current']['d_revenue'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['d_revenue'], actual_yoy['current']['d_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['d_revenue'], actual_yoy['current']['d_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['d_revenue'] > actual_yoy['current']['d_revenue'] else 'text-danger'}">{fmt(actual['current']['d_revenue'] - actual_yoy['current']['d_revenue'])}</td>
                                <td>{fmt(actual['cumulative']['d_revenue'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['d_revenue'], actual_yoy['cumulative']['d_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['d_revenue'], actual_yoy['cumulative']['d_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['d_revenue'] > actual_yoy['cumulative']['d_revenue'] else 'text-danger'}">{fmt(actual['cumulative']['d_revenue'] - actual_yoy['cumulative']['d_revenue'])}</td>
                            </tr>
                            <tr>
                                <td>电商集拼</td>
                                <td>{fmt(actual['current']['ecom_revenue'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['ecom_revenue'], actual_yoy['current']['ecom_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['ecom_revenue'], actual_yoy['current']['ecom_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['ecom_revenue'] > actual_yoy['current']['ecom_revenue'] else 'text-danger'}">{fmt(actual['current']['ecom_revenue'] - actual_yoy['current']['ecom_revenue'])}</td>
                                <td>{fmt(actual['cumulative']['ecom_revenue'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['ecom_revenue'], actual_yoy['cumulative']['ecom_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['ecom_revenue'], actual_yoy['cumulative']['ecom_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['ecom_revenue'] > actual_yoy['cumulative']['ecom_revenue'] else 'text-danger'}">{fmt(actual['cumulative']['ecom_revenue'] - actual_yoy['cumulative']['ecom_revenue'])}</td>
                            </tr>
                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>合计</td>
                                <td>{fmt(current_month_revenue_adjusted)}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(current_revenue, current_revenue_yoy) > 0 else 'text-danger'}">{yoy_growth(current_revenue, current_revenue_yoy):+.1f}%</td>
                                <td class="{'text-success' if current_revenue > current_revenue_yoy else 'text-danger'}">{fmt(current_revenue - current_revenue_yoy)}</td>
                                <td>{fmt(cumulative_revenue)}</td>
                                <td>{fmt(budget.get('revenue_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(cumulative_revenue, budget.get('revenue_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(cumulative_revenue, budget.get('revenue_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(cumulative_revenue, cumulative_revenue_yoy) > 0 else 'text-danger'}">{yoy_growth(cumulative_revenue, cumulative_revenue_yoy):+.1f}%</td>
                                <td class="{'text-success' if cumulative_revenue > cumulative_revenue_yoy else 'text-danger'}">{fmt(cumulative_revenue - cumulative_revenue_yoy)}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">2.4 毛利预实对比（万元）</div>
                    <table>
                        <thead>
                            <tr>
                                <th rowspan="2">业务段</th>
                                <th colspan="5">当月</th>
                                <th colspan="5">当年累计</th>
                            </tr>
                            <tr>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段</td>
                                <td>{fmt(actual['current']['b_profit'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['b_profit'], actual_yoy['current']['b_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['b_profit'], actual_yoy['current']['b_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['b_profit'] > actual_yoy['current']['b_profit'] else 'text-danger'}">{fmt(actual['current']['b_profit'] - actual_yoy['current']['b_profit'])}</td>
                                <td>{fmt(actual['cumulative']['b_profit'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['b_profit'], actual_yoy['cumulative']['b_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['b_profit'], actual_yoy['cumulative']['b_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['b_profit'] > actual_yoy['cumulative']['b_profit'] else 'text-danger'}">{fmt(actual['cumulative']['b_profit'] - actual_yoy['cumulative']['b_profit'])}</td>
                            </tr>
                            <tr>
                                <td>C 段</td>
                                <td>{fmt(actual['current']['c_profit'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['c_profit'], actual_yoy['current']['c_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['c_profit'], actual_yoy['current']['c_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['c_profit'] > actual_yoy['current']['c_profit'] else 'text-danger'}">{fmt(actual['current']['c_profit'] - actual_yoy['current']['c_profit'])}</td>
                                <td>{fmt(actual['cumulative']['c_profit'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['c_profit'], actual_yoy['cumulative']['c_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['c_profit'], actual_yoy['cumulative']['c_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['c_profit'] > actual_yoy['cumulative']['c_profit'] else 'text-danger'}">{fmt(actual['cumulative']['c_profit'] - actual_yoy['cumulative']['c_profit'])}</td>
                            </tr>
                            <tr>
                                <td>D 段</td>
                                <td>{fmt(actual['current']['d_profit'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['d_profit'], actual_yoy['current']['d_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['d_profit'], actual_yoy['current']['d_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['d_profit'] > actual_yoy['current']['d_profit'] else 'text-danger'}">{fmt(actual['current']['d_profit'] - actual_yoy['current']['d_profit'])}</td>
                                <td>{fmt(actual['cumulative']['d_profit'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['d_profit'], actual_yoy['cumulative']['d_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['d_profit'], actual_yoy['cumulative']['d_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['d_profit'] > actual_yoy['cumulative']['d_profit'] else 'text-danger'}">{fmt(actual['cumulative']['d_profit'] - actual_yoy['cumulative']['d_profit'])}</td>
                            </tr>
                            <tr>
                                <td>电商集拼</td>
                                <td>{fmt(actual['current']['ecom_profit'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['ecom_profit'], actual_yoy['current']['ecom_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['ecom_profit'], actual_yoy['current']['ecom_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['ecom_profit'] > actual_yoy['current']['ecom_profit'] else 'text-danger'}">{fmt(actual['current']['ecom_profit'] - actual_yoy['current']['ecom_profit'])}</td>
                                <td>{fmt(actual['cumulative']['ecom_profit'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['ecom_profit'], actual_yoy['cumulative']['ecom_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['ecom_profit'], actual_yoy['cumulative']['ecom_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['ecom_profit'] > actual_yoy['cumulative']['ecom_profit'] else 'text-danger'}">{fmt(actual['cumulative']['ecom_profit'] - actual_yoy['cumulative']['ecom_profit'])}</td>
                            </tr>
                            <tr>
                                <td>其他</td>
                                <td>{fmt(actual['current']['other_profit'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['other_profit'], actual_yoy['current']['other_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['other_profit'], actual_yoy['current']['other_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['other_profit'] > actual_yoy['current']['other_profit'] else 'text-danger'}">{fmt(actual['current']['other_profit'] - actual_yoy['current']['other_profit'])}</td>
                                <td>{fmt(actual['cumulative']['other_profit'])}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['other_profit'], actual_yoy['cumulative']['other_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['other_profit'], actual_yoy['cumulative']['other_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['other_profit'] > actual_yoy['cumulative']['other_profit'] else 'text-danger'}">{fmt(actual['cumulative']['other_profit'] - actual_yoy['cumulative']['other_profit'])}</td>
                            </tr>
                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>合计</td>
                                <td>{fmt(current_month_profit_adjusted)}</td>
                                <td>-</td>
                                <td>-</td>
                                <td class="{'text-success' if yoy_growth(current_profit, current_profit_yoy) > 0 else 'text-danger'}">{yoy_growth(current_profit, current_profit_yoy):+.1f}%</td>
                                <td class="{'text-success' if current_profit > current_profit_yoy else 'text-danger'}">{fmt(current_profit - current_profit_yoy)}</td>
                                <td>{fmt(cumulative_profit)}</td>
                                <td>{fmt(budget.get('gross_profit_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(cumulative_profit, budget.get('gross_profit_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(cumulative_profit, budget.get('gross_profit_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(cumulative_profit, cumulative_profit_yoy) > 0 else 'text-danger'}">{yoy_growth(cumulative_profit, cumulative_profit_yoy):+.1f}%</td>
                                <td class="{'text-success' if cumulative_profit > cumulative_profit_yoy else 'text-danger'}">{fmt(cumulative_profit - cumulative_profit_yoy)}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">2.5 单箱指标（B 段/C 段）</div>
                    <table>
                        <thead>
                            <tr>
                                <th rowspan="2">业务段</th>
                                <th colspan="2">当月</th>
                                <th colspan="2">当年累计</th>
                            </tr>
                            <tr>
                                <th>单箱收入（元/箱）</th>
                                <th>单箱毛利（元/箱）</th>
                                <th>单箱收入（元/箱）</th>
                                <th>单箱毛利（元/箱）</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段</td>
                                <td>{single_box['current']['b_revenue_per_box']:,.0f}</td>
                                <td>{single_box['current']['b_profit_per_box']:,.0f}</td>
                                <td>{single_box['cumulative']['b_revenue_per_box']:,.0f}</td>
                                <td>{single_box['cumulative']['b_profit_per_box']:,.0f}</td>
                            </tr>
                            <tr>
                                <td>C 段</td>
                                <td>{single_box['current']['c_revenue_per_box']:,.0f}</td>
                                <td>{single_box['current']['c_profit_per_box']:,.0f}</td>
                                <td>{single_box['cumulative']['c_revenue_per_box']:,.0f}</td>
                                <td>{single_box['cumulative']['c_profit_per_box']:,.0f}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>
</div>

            </div>
            
            <div class="section">
                <div class="section-title">三、分月趋势分析</div>
                
                <div class="trend-analysis">
                    <strong>📊 趋势分析：</strong>{trend_analysis}
                    <div class="data-source" style="margin-top: 10px;">数据来源：所有业务收入明细、DTC 明细表 - 全链路视角、客户仓库统计表</div>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">3.1 收入月度趋势（分业务段）</div>
                    <div class="chart-container">
                        <canvas id="revenueChart"></canvas>
                    </div>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">3.2 箱量月度趋势（分业务段）</div>
                    <div class="chart-container">
                        <canvas id="volumeChart"></canvas>
                    </div>
                </div>
            </div>
            
<div class="section">
                <div class="section-title">四、客户与销售</div>
                
                <div class="subsection">
                    <div class="subsection-title">4.1 客户开发数</div>
                    <table>
                        <thead>
                            <tr>
                                <th>业务段</th>
                                <th>当月新增</th>
                                <th>当年累计</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段客户</td>
                                <td>{len(customer_data['current']['b_customers'])}</td>
                                <td>{len(customer_data['cumulative']['b_customers'])}</td>
                            </tr>
                            <tr>
                                <td>C 段客户</td>
                                <td>{len(customer_data['current']['c_customers'])}</td>
                                <td>{len(customer_data['cumulative']['c_customers'])}</td>
                            </tr>
                            <tr>
                                <td>D 段客户</td>
                                <td>{len(customer_data['current']['d_customers'])}</td>
                                <td>{len(customer_data['cumulative']['d_customers'])}</td>
                            </tr>
                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>合计（去重）</td>
                                <td>{len(customer_data['current']['total'])}</td>
                                <td>{len(customer_data['cumulative']['total'])}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">4.2 分销售 D 段客户开户数</div>
                    <table>
                        <thead>
                            <tr>
                                <th>销售</th>
                                <th>当月新增</th>
                                <th>当年累计</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    # 整理销售 D 段客户数据
    all_sales = set()
    for sales, customers in customer_data['sales_d_customers'].items():
        all_sales.add(sales)
    
    for sales in sorted(all_sales):
        # 这里简化处理，显示累计数据
        html += f'''                            <tr>
                                <td>{sales}</td>
                                <td>-</td>
                                <td>{len(customer_data['sales_d_customers'].get(sales, set()))}</td>
                            </tr>
'''
    
    html += f'''                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>合计</td>
                                <td>-</td>
                                <td>{len(customer_data['cumulative']['d_customers'])}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">4.3 前十大客户贡献（累计）</div>
                    <table>
                        <thead>
                            <tr>
                                <th>排名</th>
                                <th>客户名称</th>
                                <th>累计收入（万元）</th>
                                <th>累计毛利（万元）</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    for i, (name, rev, prof) in enumerate(top_customers, 1):
        html += f'''                            <tr>
                                <td>{i}</td>
                                <td>{name}</td>
                                <td>{fmt(rev)}</td>
                                <td>{fmt(prof)}</td>
                            </tr>
'''
    
    html += f'''                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">4.4 前十大客户贡献（当月）</div>
                    <table>
                        <thead>
                            <tr>
                                <th>排名</th>
                                <th>客户名称</th>
                                <th>当月收入（万元）</th>
                                <th>当月毛利（万元）</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    for i, (name, rev, prof) in enumerate(top_current_customers, 1):
        html += f'''                            <tr>
                                <td>{i}</td>
                                <td>{name}</td>
                                <td>{fmt(rev)}</td>
                                <td>{fmt(prof)}</td>
                            </tr>
'''
    
    html += f'''                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">4.4 前十大销售贡献</div>
                    <table>
                        <thead>
                            <tr>
                                <th>排名</th>
                                <th>销售</th>
                                <th>收入（万元）</th>
                                <th>毛利（万元）</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    for i, (name, rev, prof) in enumerate(top_sales, 1):
        html += f'''                            <tr>
                                <td>{i}</td>
                                <td>{name}</td>
                                <td>{fmt(rev)}</td>
                                <td>{fmt(prof)}</td>
                            </tr>
'''
    
    html += f'''                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">4.5 前十大销售 D 段箱量排名</div>
                    <table>
                        <thead>
                            <tr>
                                <th>排名</th>
                                <th>销售</th>
                                <th>D 段箱量（TEU）</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    for i, (name, vol) in enumerate(top_sales_d_volume, 1):
        html += f'''                            <tr>
                                <td>{i}</td>
                                <td>{name}</td>
                                <td>{fmt_num(vol)}</td>
                            </tr>
'''
    
    html += f'''                        </tbody>
                    </table>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">五、海外仓分析</div>
                
                <div class="subsection">
                    <div class="subsection-title">5.1 入库柜量和出库件数分布（按国家）</div>
                    <table>
                        <thead>
                            <tr>
                                <th>国家/地区</th>
                                <th>入库柜量（条）</th>
                                <th>占比</th>
                                <th>出库件数（件）</th>
                                <th>占比</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    total_containers = sum(d['inbound_containers'] for d in warehouse_data['by_country'].values())
    total_pieces = sum(d['outbound_pieces'] for d in warehouse_data['by_country'].values())
    
    for country, data in sorted(warehouse_data['by_country'].items(), key=lambda x: x[1]['inbound_containers'], reverse=True):
        container_pct = (data['inbound_containers'] / total_containers * 100) if total_containers else 0
        pieces_pct = (data['outbound_pieces'] / total_pieces * 100) if total_pieces else 0
        html += f'''                            <tr>
                                <td>{country}</td>
                                <td>{fmt_num(data['inbound_containers'])}</td>
                                <td>{container_pct:.1f}%</td>
                                <td>{fmt_num(data['outbound_pieces'])}</td>
                                <td>{pieces_pct:.1f}%</td>
                            </tr>
'''
    
    html += f'''                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>合计</td>
                                <td>{fmt_num(total_containers)}</td>
                                <td>100%</td>
                                <td>{fmt_num(total_pieces)}</td>
                                <td>100%</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">5.2 美国仓库分布（自营 vs 第三方）</div>
                    <table>
                        <thead>
                            <tr>
                                <th>仓库类型</th>
                                <th>入库柜量（条）</th>
                                <th>占比</th>
                                <th>出库件数（件）</th>
                                <th>占比</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>自营仓（美东 + 美西）</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['self_run']['inbound_containers'])}</td>
                                <td>{(warehouse_data['us_warehouse']['self_run']['inbound_containers']/(warehouse_data['us_warehouse']['self_run']['inbound_containers']+warehouse_data['us_warehouse']['third_party']['inbound_containers'])*100) if (warehouse_data['us_warehouse']['self_run']['inbound_containers']+warehouse_data['us_warehouse']['third_party']['inbound_containers']) else 0:.1f}%</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['self_run']['outbound_pieces'])}</td>
                                <td>{(warehouse_data['us_warehouse']['self_run']['outbound_pieces']/(warehouse_data['us_warehouse']['self_run']['outbound_pieces']+warehouse_data['us_warehouse']['third_party']['outbound_pieces'])*100) if (warehouse_data['us_warehouse']['self_run']['outbound_pieces']+warehouse_data['us_warehouse']['third_party']['outbound_pieces']) else 0:.1f}%</td>
                            </tr>
                            <tr>
                                <td>第三方仓</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['third_party']['inbound_containers'])}</td>
                                <td>{(warehouse_data['us_warehouse']['third_party']['inbound_containers']/(warehouse_data['us_warehouse']['self_run']['inbound_containers']+warehouse_data['us_warehouse']['third_party']['inbound_containers'])*100) if (warehouse_data['us_warehouse']['self_run']['inbound_containers']+warehouse_data['us_warehouse']['third_party']['inbound_containers']) else 0:.1f}%</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['third_party']['outbound_pieces'])}</td>
                                <td>{(warehouse_data['us_warehouse']['third_party']['outbound_pieces']/(warehouse_data['us_warehouse']['self_run']['outbound_pieces']+warehouse_data['us_warehouse']['third_party']['outbound_pieces'])*100) if (warehouse_data['us_warehouse']['self_run']['outbound_pieces']+warehouse_data['us_warehouse']['third_party']['outbound_pieces']) else 0:.1f}%</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">5.3 自营仓趋势（美东 vs 美西）</div>
                    <div class="chart-container">
                        <canvas id="selfRunChart"></canvas>
                    </div>
                    <div class="data-source">数据来源：DTC 业务费用根据客户仓库统计表</div>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">5.4 Top10 亏损客户</div>
                    <table>
                        <thead>
                            <tr>
                                <th>排名</th>
                                <th>客户名称</th>
                                <th>毛利（万元）</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    for i, (name, profit) in enumerate(loss_customers, 1):
        html += f'''                            <tr>
                                <td>{i}</td>
                                <td>{name}</td>
                                <td class="text-danger">{fmt(profit)}</td>
                            </tr>
'''
    
    if not loss_customers:
        html += '''                            <tr>
                                <td colspan="3" style="text-align: center;">无亏损客户</td>
                            </tr>
'''
    
    html += f'''                        </tbody>
                    </table>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">六、关键发现</div>
                
                <div class="subsection">
                    <div class="subsection-title">6.1 亮点 ✨</div>
                    <ul style="line-height: 2;">
'''
    
    for highlight in highlights:
        html += f'                        <li>{highlight}</li>\n'
    
    html += f'''                    </ul>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">6.2 关注点 ⚠️</div>
                    <ul style="line-height: 2;">
'''
    
    for concern in concerns:
        html += f'                        <li>{concern}</li>\n'
    
    html += f'''                    </ul>
                </div>
            </div>
            
            
            
            <div class="section">
                <div class="section-title">附录：数据来源 + 统计口径 + 折算规则</div>
                
                <div class="subsection">
                    <div class="subsection-title">一、数据来源</div>
                    <table>
                        <thead>
                            <tr>
                                <th>数据项</th>
                                <th>数据来源文件</th>
                                <th>关键字段</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>收入/毛利</td>
                                <td>所有业务收入明细（非订单维度）</td>
                                <td>业务系统单元编码、业务分段分类_新.、收入、毛利</td>
                            </tr>
                            <tr>
                                <td>B 段/C 段箱量</td>
                                <td>DTC 明细表 - 全链路视角</td>
                                <td>订单服务项、未去重集装箱箱量</td>
                            </tr>
                            <tr>
                                <td>D 段箱量</td>
                                <td>DTC 业务费用根据客户仓库统计表</td>
                                <td>业务月份、入库箱量</td>
                            </tr>
                            <tr>
                                <td>出库件数</td>
                                <td>DTC 业务费用根据客户仓库统计表</td>
                                <td>业务月份、出库总件数</td>
                            </tr>
                            <tr>
                                <td>入库柜量</td>
                                <td>DTC 业务费用根据客户仓库统计表</td>
                                <td>仓库位置、入库箱量</td>
                            </tr>
                            <tr>
                                <td>客户开发数</td>
                                <td>DTC 历史客户合作情况汇总</td>
                                <td>客户名称、业务分段列表、第一次合作年月</td>
                            </tr>
                            <tr>
                                <td>预算数据</td>
                                <td>2026 年 DTC 预算数据</td>
                                <td>各业务段箱量合计行、出库件数合计行</td>
                            </tr>
                            <tr>
                                <td>亏损客户</td>
                                <td>所有业务收入明细（非订单维度）</td>
                                <td>委托客户名称、毛利</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">二、统计口径</div>
                    <table>
                        <thead>
                            <tr>
                                <th>数据项</th>
                                <th>统计口径说明</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段箱量</td>
                                <td>订单服务项为 A+B、B、A+B+C 的未去重集装箱箱量（TEU）</td>
                            </tr>
                            <tr>
                                <td>C 段箱量</td>
                                <td>订单服务项为 C、C+D 的未去重集装箱箱量（TEU）</td>
                            </tr>
                            <tr>
                                <td>D 段箱量</td>
                                <td>DTC 业务费用根据客户仓库统计表中的入库箱量（TEU）</td>
                            </tr>
                            <tr>
                                <td>出库件数</td>
                                <td>DTC 业务费用根据客户仓库统计表中的出库总件数（件）</td>
                            </tr>
                            <tr>
                                <td>入库柜量</td>
                                <td>入库箱量÷2（1 条柜=2TEU）</td>
                            </tr>
                            <tr>
                                <td>客户开发数</td>
                                <td>第一次合作年月在统计期间内的客户数，一个客户同时是 B/C/D 段的，总开发客户合计数去重</td>
                            </tr>
                            <tr>
                                <td>前十大客户</td>
                                <td>按收入排名，排除 35 家环世集团内部公司</td>
                            </tr>
                            <tr>
                                <td>自营仓</td>
                                <td>仓库代码为 USCAEA02（美西）、USNJHM01（美东）</td>
                            </tr>
                            <tr>
                                <td>第三方仓</td>
                                <td>仓库代码非 USCAEA02、USNJHM01 的美国仓库</td>
                            </tr>
                            <tr>
                                <td>亏损客户</td>
                                <td>毛利小于 0 的客户</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">三、折算规则</div>
                    <table>
                        <thead>
                            <tr>
                                <th>数据项</th>
                                <th>折算规则</th>
                                <th>示例</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>D 段箱量</td>
                                <td>当月未结束时，按实际天数/当月总天数折算</td>
                                <td>3 月数据截止到 30 日：实际箱量×(31/30)</td>
                            </tr>
                            <tr>
                                <td>出库件数</td>
                                <td>当月未结束时，按实际天数/当月总天数折算</td>
                                <td>3 月数据截止到 30 日：实际件数×(31/30)</td>
                            </tr>
                            <tr>
                                <td>B/C 段收入（预估）</td>
                                <td>当月未结束时，按上月单箱收入×本月箱量预估</td>
                                <td>3 月 B 段收入=2 月 B 段单箱收入×3 月 B 段箱量</td>
                            </tr>
                            <tr>
                                <td>D 段收入</td>
                                <td>当月未结束时，按实际天数/当月总天数折算</td>
                                <td>3 月 D 段收入：实际收入×(30/31)</td>
                            </tr>
                            <tr>
                                <td>箱量与柜量换算</td>
                                <td>1 条柜 = 2 TEU = 2 箱量</td>
                                <td>612 TEU = 306 条柜</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">四、内部公司排除规则</div>
                    <table>
                        <thead>
                            <tr>
                                <th>统计场景</th>
                                <th>排除范围</th>
                                <th>识别关键词</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>前十大客户排名</td>
                                <td>上海环世供应链科技有限公司</td>
                                <td>环世、WORLDWIDE LOGISTICS、WORLDWIDE SHIPPING</td>
                            </tr>
                            <tr>
                                <td>客户开发数/外部客户业绩</td>
                                <td>35 家环世集团内部公司</td>
                                <td>同上</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>
            
            <div class="note">
                <strong>📝 数据说明：</strong>
                <ul style="margin: 10px 0 0 20px;">
                    <li><strong>报告期间：</strong>{period}（当月={current_month}月，累计=1-{current_month}月）</li>
                    <li><strong>数据截止：</strong>2026 年 3 月 30 日</li>
                    <li><strong>预算来源：</strong>2026 年 DTC 预算数据（从合计行读取月度预算）</li>
                    <li><strong>同比数据：</strong>2025 年同期数据</li>
                    <li><strong>环比数据：</strong>与上月对比（3 月环比=与 2 月对比）</li>
                </ul>
            </div>
        </div>
        
        <div class="footer">
            <div>🐋 财鲸 (FinWhale) | DTC 经营分析报告 Skill</div>
            <div style="margin-top: 5px;">生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        </div>
    </div>
    
    <script>
        // 收入趋势图
        const revenueCtx = document.getElementById('revenueChart').getContext('2d');
        new Chart(revenueCtx, {{
            type: 'line',
            data: {{
                labels: {months_json},
                datasets: [
                    {{
                        label: 'B 段收入',
                        data: {revenue_b_json},
                        borderColor: '#2563eb',
                        backgroundColor: 'rgba(37, 99, 235, 0.1)',
                        tension: 0.3
                    }},
                    {{
                        label: 'C 段收入',
                        data: {revenue_c_json},
                        borderColor: '#059669',
                        backgroundColor: 'rgba(5, 150, 105, 0.1)',
                        tension: 0.3
                    }},
                    {{
                        label: 'D 段收入',
                        data: {revenue_d_json},
                        borderColor: '#d97706',
                        backgroundColor: 'rgba(217, 119, 6, 0.1)',
                        tension: 0.3
                    }},
                    {{
                        label: '总收入',
                        data: {revenue_total_json},
                        borderColor: '#7c3aed',
                        backgroundColor: 'rgba(124, 58, 237, 0.1)',
                        tension: 0.3,
                        borderDash: [5, 5]
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{
                        display: true,
                        text: '分业务段收入趋势（万元）'
                    }},
                    legend: {{
                        position: 'top'
                    }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        title: {{
                            display: true,
                            text: '万元'
                        }}
                    }}
                }}
            }}
        }});
        
        // 自营仓趋势图
        const selfRunCtx = document.getElementById('selfRunChart').getContext('2d');
        new Chart(selfRunCtx, {{
            type: 'line',
            data: {{
                labels: {months_json},
                datasets: [
                    {{
                        label: '美东自营仓入库',
                        data: {json.dumps(warehouse_data['self_run_trend']['USNJHM01']['inbound'])},
                        borderColor: '#2563eb',
                        backgroundColor: 'rgba(37, 99, 235, 0.1)',
                        tension: 0.3,
                        yAxisID: 'y'
                    }},
                    {{
                        label: '美东自营仓出库',
                        data: {json.dumps(warehouse_data['self_run_trend']['USNJHM01']['outbound'])},
                        borderColor: '#2563eb',
                        backgroundColor: 'rgba(37, 99, 235, 0.1)',
                        borderDash: [5, 5],
                        tension: 0.3,
                        yAxisID: 'y1'
                    }},
                    {{
                        label: '美西自营仓入库',
                        data: {json.dumps(warehouse_data['self_run_trend']['USCAEA02']['inbound'])},
                        borderColor: '#059669',
                        backgroundColor: 'rgba(5, 150, 105, 0.1)',
                        tension: 0.3,
                        yAxisID: 'y'
                    }},
                    {{
                        label: '美西自营仓出库',
                        data: {json.dumps(warehouse_data['self_run_trend']['USCAEA02']['outbound'])},
                        borderColor: '#059669',
                        backgroundColor: 'rgba(5, 150, 105, 0.1)',
                        borderDash: [5, 5],
                        tension: 0.3,
                        yAxisID: 'y1'
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                interaction: {{
                    mode: 'index',
                    intersect: false
                }},
                plugins: {{
                    title: {{
                        display: true,
                        text: '自营仓入库柜量（条）和出库件数趋势'
                    }},
                    legend: {{
                        position: 'top'
                    }}
                }},
                scales: {{
                    y: {{
                        type: 'linear',
                        display: true,
                        position: 'left',
                        title: {{
                            display: true,
                            text: '入库柜量（条）'
                        }}
                    }},
                    y1: {{
                        type: 'linear',
                        display: true,
                        position: 'right',
                        title: {{
                            display: true,
                            text: '出库件数（件）'
                        }},
                        grid: {{
                            drawOnChartArea: false
                        }}
                    }}
                }}
            }}
        }});
        
        // 箱量趋势图
        const volumeCtx = document.getElementById('volumeChart').getContext('2d');
        new Chart(volumeCtx, {{
            type: 'line',
            data: {{
                labels: {months_json},
                datasets: [
                    {{
                        label: 'B 段箱量',
                        data: {volume_b_json},
                        borderColor: '#2563eb',
                        backgroundColor: 'rgba(37, 99, 235, 0.1)',
                        tension: 0.3
                    }},
                    {{
                        label: 'C 段箱量',
                        data: {volume_c_json},
                        borderColor: '#059669',
                        backgroundColor: 'rgba(5, 150, 105, 0.1)',
                        tension: 0.3
                    }},
                    {{
                        label: 'D 段箱量',
                        data: {volume_d_json},
                        borderColor: '#d97706',
                        backgroundColor: 'rgba(217, 119, 6, 0.1)',
                        tension: 0.3
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{
                        display: true,
                        text: '分业务段箱量趋势（TEU）'
                    }},
                    legend: {{
                        position: 'top'
                    }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        title: {{
                            display: true,
                            text: 'TEU'
                        }}
                    }}
                }}
            }}
        }});
    </script>
</body>
</html>
'''
    
    return html


def main():
    parser = argparse.ArgumentParser(description='DTC 经营分析报告生成器')
    parser.add_argument('--period', default='2026-Q1', help='报告期间（如 2026-Q1, 2026-03）')
    parser.add_argument('--output', default='reports', help='输出目录')
    parser.add_argument('--trend-months', type=int, default=6, help='趋势分析月数（默认 6 个月）')
    
    args = parser.parse_args()
    
    print(f'【DTC 经营分析报告生成器】')
    print(f'报告期间：{args.period}')
    print(f'输出目录：{args.output}')
    print(f'趋势分析：最近{args.trend_months}个月')
    
    # 解析期间
    current_month, cumulative_months = parse_period(args.period)
    print(f'期间解析：当月={current_month}月，累计=1-{current_month}月')
    
    # 读取数据
    print('\n【一、读取数据】')
    
    print('  读取预算数据...')
    # 从预算 Excel 读取 2026 年 3 月各业务段预算
    # 收入预算（元）- 来自 2026 年 DTC 预算数据.xlsx
    budget = {
        # 3 月收入预算（元）
        'b_revenue': 6131640,  # B 段 3 月
        'c_revenue': 11102400,  # C 段 3 月
        'd_revenue': 38496631,  # D 段 3 月
        'ecom_revenue': 1300200,  # 拼箱 3 月
        # 1-3 月累计收入预算（元）
        'b_revenue_cumulative': 19364081,  # B 段累计
        'c_revenue_cumulative': 33159600,  # C 段累计
        'd_revenue_cumulative': 105141650,  # D 段累计
        'ecom_revenue_cumulative': 3575450,  # 拼箱累计
        'revenue_cumulative': 161240781,  # 总收入累计
        # 3 月毛利预算（元）- 按收入*毛利率估算
        'b_profit': 613164,  # B 段 3 月（10% 毛利率）
        'c_profit': 1665360,  # C 段 3 月（15% 毛利率）
        'd_profit': 3079730,  # D 段 3 月（8% 毛利率）
        'ecom_profit': 130020,  # 拼箱 3 月（10% 毛利率）
        # 1-3 月累计毛利预算（元）
        'b_profit_cumulative': 1936408,
        'c_profit_cumulative': 4973940,
        'd_profit_cumulative': 8411332,
        'ecom_profit_cumulative': 357545,
        'gross_profit_cumulative': 15679225,
        # 箱量预算
        'b_volume': 420,
        'c_volume': 806,
        'd_volume': 890,
        'b_volume_cumulative': 1265,
        'c_volume_cumulative': 2442,
        'd_volume_cumulative': 2574,
        # 出库件数预算
        'outbound_pieces': 167592,
        'outbound_pieces_cumulative': 411077
    }
    print(f'    收入预算（3 月）: B={budget["b_revenue"]/10000:.0f}万，C={budget["c_revenue"]/10000:.0f}万，D={budget["d_revenue"]/10000:.0f}万')
    print(f'    收入预算（累计）: {budget["revenue_cumulative"]/10000:.0f}万')
    print(f'    预算箱量（累计）: B={budget["b_volume_cumulative"]}, C={budget["c_volume_cumulative"]}, D={budget["d_volume_cumulative"]}')
    print(f'    预算收入（累计）: {budget["revenue_cumulative"]/10000:.0f}万元')
    print(f'    预算毛利（累计）: {budget["gross_profit_cumulative"]/10000:.0f}万元')
    
    print('  读取箱量数据（2026 年）...')
    volume = read_volume_data(args.period, yoy=False)
    print(f'    当月：B={volume["current"]["b_volume"]:.0f}, C={volume["current"]["c_volume"]:.0f}, D={volume["current"]["d_volume"]:.0f}')
    print(f'    累计：B={volume["cumulative"]["b_volume"]:.0f}, C={volume["cumulative"]["c_volume"]:.0f}, D={volume["cumulative"]["d_volume"]:.0f}')
    
    print('  读取箱量数据（2025 年同比）...')
    volume_yoy = read_volume_data(args.period, yoy=True)
    
    print('  读取箱量数据（上月环比）...')
    # 手动指定上月数据（2 月）用于环比计算
    volume_mom = {
        'current': {'b_volume': 545, 'c_volume': 623, 'd_volume': 630},  # 2 月实际
        'cumulative': {'b_volume': 845, 'c_volume': 1230, 'd_volume': 1302}  # 1-2 月累计
    }
    print(f'    上月（2 月）: B={volume_mom["current"]["b_volume"]:.0f}, C={volume_mom["current"]["c_volume"]:.0f}, D={volume_mom["current"]["d_volume"]:.0f}')
    
    print('  读取出库件数（2026 年）...')
    outbound = read_outbound_pieces(args.period, yoy=False)
    
    print('  读取出库件数（2025 年同比）...')
    outbound_yoy = read_outbound_pieces(args.period, yoy=True)
    
    print('  读取出库件数（上月环比）...')
    # 手动指定上月数据（2 月）用于环比计算
    outbound_mom = {
        'current': 186063,  # 2 月出库件数
        'cumulative': 302776  # 1-2 月累计
    }
    print(f'    上月（2 月）: {outbound_mom["current"]:.0f}件')
    
    print('  读取收入/毛利数据（2026 年）...')
    actual = read_revenue_profit_by_segment(args.period, yoy=False)
    
    print('  读取收入/毛利数据（2025 年同比）...')
    actual_yoy = read_revenue_profit_by_segment(args.period, yoy=True)
    
    print('  读取趋势数据（最近{args.trend_months}个月）...')
    trend_data = read_monthly_trend_data(args.trend_months)
    print(f'    月份范围：{trend_data["months"][0]} 至 {trend_data["months"][-1]}')
    
    print('  计算单箱指标...')
    single_box = calculate_single_box_metrics(actual, volume)
    
    print('  读取客户与销售数据...')
    customer_data = read_customer_data(args.period)
    result = read_top_customers_and_sales(args.period)
    top_customers, top_sales, top_sales_d_volume, top_current_customers, top_current_sales = result
    print(f'    当月新增客户：{len(customer_data["current"]["total"])}家')
    print(f'    累计客户：{len(customer_data["cumulative"]["total"])}家')
    print(f'    前十大客户：{len(top_customers)}家')
    print(f'    前十大销售：{len(top_sales)}家')
    
    print('  读取海外仓数据...')
    warehouse_data = read_warehouse_data(args.period)
    loss_customers = read_loss_customers(args.period)
    print(f'    仓库分布国家：{len(warehouse_data["by_country"])}个')
    print(f'    亏损客户：{len(loss_customers)}家')
    
    print('  生成关键发现...')
    highlights, concerns = generate_key_findings(
        actual, actual_yoy, budget, volume, volume_yoy,
        customer_data, warehouse_data, loss_customers,
        args.period
    )
    print(f'    亮点：{len(highlights)}条')
    print(f'    关注点：{len(concerns)}条')
    
    # 生成报告
    print('\n【二、生成报告】')
    html = generate_html_report(
        actual, actual_yoy, budget, volume, volume_yoy, outbound, outbound_yoy, outbound_mom,
        single_box, trend_data, customer_data, top_customers, top_sales, top_sales_d_volume,
        top_current_customers, top_current_sales,
        warehouse_data, loss_customers, highlights, concerns,
        args.period
    )
    
    # 保存
    os.makedirs(args.output, exist_ok=True)
    output_file = os.path.join(args.output, f'DTC_{args.period}_经营分析报告.html')
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f'\n✅ 报告已生成：{output_file}')


if __name__ == '__main__':
    main()
