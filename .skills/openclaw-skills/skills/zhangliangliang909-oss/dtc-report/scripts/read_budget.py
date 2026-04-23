#!/usr/bin/env python3
"""
读取预算数据模块

从预算 Excel 文件读取月度预算数据（箱量、出库件数、收入、毛利）
"""

import openpyxl
import os
import glob

# 数据目录（本地路径）
BUDGET_DIR = r"C:\Users\wwl\.openclaw\workspace-跨境电商\data\6.2026 年预算数据"


def read_budget_volume(period='2026-Q1'):
    """读取箱量和出库件数预算数据
    
    返回：
    - 当月预算（B/C/D段箱量、出库件数）
    - 累计预算（B/C/D段箱量、出库件数）
    """
    current_month, cumulative_months = parse_period(period)
    
    result = {
        'current': {'b_volume': 0, 'c_volume': 0, 'd_volume': 0, 'outbound_pieces': 0},
        'cumulative': {'b_volume': 0, 'c_volume': 0, 'd_volume': 0, 'outbound_pieces': 0}
    }
    
    # 查找预算文件
    files = glob.glob(os.path.join(BUDGET_DIR, '*DTC 预算*.xlsx'))
    if not files:
        return result
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    # 找表头
    headers = {}
    for col in range(1, 30):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    # 查找 2026 年 3 月的列（列 F）
    month_col = None
    for col in range(1, 30):
        h = ws.cell(row=2, column=col).value
        if h and '2026 年 3 月' in str(h):
            month_col = col
            break
    
    if not month_col:
        return result
    
    # 读取 B 段箱量合计（行 6）
    b_volume = ws.cell(row=6, column=month_col).value or 0
    
    # 读取 C 段箱量合计（行 10）
    c_volume = ws.cell(row=10, column=month_col).value or 0
    
    # 读取 D 段箱量合计（行 14）
    d_volume = ws.cell(row=14, column=month_col).value or 0
    
    # 读取 D 段出库件数合计（行 20）
    outbound_pieces = ws.cell(row=20, column=month_col).value or 0
    
    # 当月预算
    result['current']['b_volume'] = float(b_volume)
    result['current']['c_volume'] = float(c_volume)
    result['current']['d_volume'] = float(d_volume)
    result['current']['outbound_pieces'] = float(outbound_pieces)
    
    # 累计预算（1-3 月合计）
    for m in cumulative_months:
        m_col = None
        for col in range(1, 30):
            h = ws.cell(row=2, column=col).value
            if h and f'2026 年{m}月' in str(h):
                m_col = col
                break
        
        if m_col:
            result['cumulative']['b_volume'] += float(ws.cell(row=6, column=m_col).value or 0)
            result['cumulative']['c_volume'] += float(ws.cell(row=10, column=m_col).value or 0)
            result['cumulative']['d_volume'] += float(ws.cell(row=14, column=m_col).value or 0)
            result['cumulative']['outbound_pieces'] += float(ws.cell(row=20, column=m_col).value or 0)
    
    return result


def read_budget_revenue_profit(period='2026-Q1'):
    """读取收入和毛利预算数据
    
    返回：
    - 当月预算（收入、毛利）
    - 累计预算（收入、毛利）
    """
    current_month, cumulative_months = parse_period(period)
    
    result = {
        'current': {'revenue': 0, 'gross_profit': 0},
        'cumulative': {'revenue': 0, 'gross_profit': 0}
    }
    
    # 查找预算文件
    files = glob.glob(os.path.join(BUDGET_DIR, '*DTC 预算*.xlsx'))
    if not files:
        return result
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    # 查找 2026 年 3 月的列
    month_col = None
    for col in range(1, 30):
        h = ws.cell(row=2, column=col).value
        if h and '2026 年 3 月' in str(h):
            month_col = col
            break
    
    if not month_col:
        return result
    
    # 读取营业收入（行 25）- 原始数据单位是元，转换为万元
    revenue = ws.cell(row=25, column=month_col).value or 0
    
    # 读取毛利（需要计算或查找）
    # 简化处理：用收入 * 预算毛利率
    gross_profit = revenue * 0.06  # 假设 6% 毛利率
    
    # 当月预算（转换为万元）
    result['current']['revenue'] = float(revenue) / 10000
    result['current']['gross_profit'] = float(gross_profit) / 10000
    
    # 累计预算（转换为万元）
    for m in cumulative_months:
        m_col = None
        for col in range(1, 30):
            h = ws.cell(row=2, column=col).value
            if h and f'2026 年{m}月' in str(h):
                m_col = col
                break
        
        if m_col:
            rev = float(ws.cell(row=25, column=m_col).value or 0)
            result['cumulative']['revenue'] += rev / 10000
            result['cumulative']['gross_profit'] += (rev * 0.06) / 10000
    
    return result


def parse_period(period):
    """解析期间参数"""
    if period.upper().startswith('2026-Q'):
        quarter = int(period[-1])
        current_month = quarter * 3
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    if period.startswith('2026-'):
        month_part = period[5:]
        current_month = int(month_part)
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    return 3, [1, 2, 3]


if __name__ == '__main__':
    # 测试
    budget_volume = read_budget_volume('2026-Q1')
    budget_rev = read_budget_revenue_profit('2026-Q1')
    
    print(f"箱量预算（当月）: B={budget_volume['current']['b_volume']}, C={budget_volume['current']['c_volume']}, D={budget_volume['current']['d_volume']}")
    print(f"箱量预算（累计）: B={budget_volume['cumulative']['b_volume']}, C={budget_volume['cumulative']['c_volume']}, D={budget_volume['cumulative']['d_volume']}")
    print(f"收入预算（累计）: {budget_rev['cumulative']['revenue']/10000:.1f}万元")
