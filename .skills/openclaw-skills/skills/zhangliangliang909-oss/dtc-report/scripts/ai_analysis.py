#!/usr/bin/env python3
"""
AI 驱动的经营分析模块

功能：
- 异常波动识别
- 根因分析（业务/客户/销售）
- 行动建议生成
"""

import os
import sys
from pathlib import Path
from collections import defaultdict

# 数据目录（本地路径）
BUSINESS_DATA_DIR = r"C:\Users\wwl\.openclaw\workspace-跨境电商\data\1.业务和订单数据"


def parse_period(period):
    """解析期间参数"""
    if period.upper().startswith('2026-Q'):
        quarter = int(period[-1])
        current_month = quarter * 3
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    if period.startswith('2026-'):
        month_part = period[5:]
        current_month = int(month_part)
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    return 3, [1, 2, 3]


def analyze_business_performance(actual, budget, actual_yoy, volume, volume_yoy, period='2026-Q1'):
    """
    综合经营分析入口（增强版）
    
    Args:
        actual: 实际数据（收入/毛利）
        budget: 预算数据
        actual_yoy: 同比数据
        volume: 箱量数据（当期）
        volume_yoy: 箱量数据（同比）
        period: 报告期间
    
    Returns:
        dict: 分析结果
    """
    
    print('  进行异常波动识别...')
    anomalies = detect_anomalies(actual, budget, actual_yoy)
    
    print('  分析收入异常原因...')
    revenue_analysis = analyze_revenue_anomaly(anomalies, actual, actual_yoy, volume, volume_yoy, period)
    
    print('  生成行动建议...')
    recommendations = generate_recommendations(revenue_analysis, anomalies, period)
    
    return {
        'anomalies': anomalies,
        'revenue_analysis': revenue_analysis,
        'recommendations': recommendations,
        'success': True
    }


def detect_anomalies(actual, budget, actual_yoy, threshold_budget=0.7, threshold_yoy=-0.3):
    """
    识别异常波动点
    
    规则：
    - 预算达成率 < 70%
    - 同比下降 > 30%
    - 环比下降 > 40%
    """
    anomalies = []
    
    for segment in ['b', 'c', 'd', 'ecom']:
        seg_name = {'b': 'B 段', 'c': 'C 段', 'd': 'D 段', 'ecom': '电商集拼'}[segment]
        
        # 预算达成率异常
        actual_rev = actual.get('cumulative', {}).get(f'{segment}_revenue', 0)
        budget_rev = budget.get(f'{segment}_revenue_cumulative', 1)
        
        if budget_rev > 0:
            budget_achievement = actual_rev / budget_rev
            if budget_achievement < threshold_budget:
                anomalies.append({
                    'type': 'budget_miss',
                    'segment': segment,
                    'segment_name': seg_name,
                    'metric': '收入预算达成率',
                    'value': f'{budget_achievement:.1%}',
                    'actual': f'{actual_rev:,.0f}万元',
                    'budget': f'{budget_rev:,.0f}万元',
                    'severity': 'high' if budget_achievement < 0.5 else 'medium'
                })
        
        # 同比异常
        yoy_rev = actual_yoy.get('cumulative', {}).get(f'{segment}_revenue', 0)
        if actual_rev > 0 and yoy_rev > 0:
            yoy_growth = (actual_rev - yoy_rev) / yoy_rev
            if yoy_growth < threshold_yoy:
                anomalies.append({
                    'type': 'yoy_decline',
                    'segment': segment,
                    'segment_name': seg_name,
                    'metric': '收入同比增长',
                    'value': f'{yoy_growth:.1%}',
                    'severity': 'high' if yoy_growth < -0.5 else 'medium'
                })
    
    # 毛利异常
    actual_profit = actual.get('cumulative', {}).get('b_profit', 0) + \
                    actual.get('cumulative', {}).get('c_profit', 0) + \
                    actual.get('cumulative', {}).get('d_profit', 0)
    budget_profit = budget.get('gross_profit_cumulative', 1)
    
    if budget_profit > 0:
        profit_achievement = actual_profit / budget_profit
        if profit_achievement < 0.8:
            anomalies.append({
                'type': 'profit_miss',
                'segment': 'total',
                'segment_name': '整体',
                'metric': '毛利预算达成率',
                'value': f'{profit_achievement:.1%}',
                'severity': 'high' if profit_achievement < 0.6 else 'medium'
            })
    
    return anomalies


def analyze_revenue_anomaly(anomalies, actual, actual_yoy, volume, volume_yoy, period):
    """
    分析收入异常原因（深度分析）
    
    从以下维度分析：
    1. 箱量因素 vs 价格因素
    2. 客户集中度影响
    3. 销售团队表现
    """
    analysis = []
    
    for anomaly in anomalies:
        segment = anomaly['segment']
        seg_name = anomaly['segment_name']
        
        if anomaly['type'] == 'budget_miss':
            # 分析是箱量问题还是价格问题
            volume_analysis = analyze_volume_vs_price(segment, actual, actual_yoy, volume, volume_yoy)
            
            analysis.append({
                'segment': seg_name,
                'issue': f"{seg_name}{anomaly['metric']}仅{anomaly['value']}",
                'factors': volume_analysis,
                'impact': calculate_impact(actual, actual_yoy, segment)
            })
        
        elif anomaly['type'] == 'yoy_decline':
            # 分析同比下降原因（深度分析）
            yoy_analysis = analyze_yoy_decline(segment, actual, actual_yoy, volume, volume_yoy, period)
            
            analysis.append({
                'segment': seg_name,
                'issue': f"{seg_name}{anomaly['metric']}{anomaly['value']}",
                'factors': yoy_analysis,
                'impact': calculate_impact(actual, actual_yoy, segment)
            })
    
    return analysis


def analyze_volume_vs_price(segment, actual, actual_yoy, volume, volume_yoy):
    """
    分析箱量 vs 价格因素（深度分析）
    
    收入 = 箱量 × 单箱收入
    
    需要分析：
    1. 箱量变化贡献
    2. 单箱收入变化贡献
    """
    factors = []
    
    vol_key = f'{segment}_volume'
    rev_key = f'{segment}_revenue'
    
    # 当前数据
    current_vol = volume.get('cumulative', {}).get(vol_key, 0)
    current_rev = actual.get('cumulative', {}).get(rev_key, 0)
    
    # 去年同期数据
    last_year_vol = volume_yoy.get('cumulative', {}).get(vol_key, 0)
    last_year_rev = actual_yoy.get('cumulative', {}).get(rev_key, 0)
    
    # 计算单箱收入
    current_unit_rev = (current_rev / current_vol * 10000) if current_vol > 0 else 0  # 元/TEU
    last_year_unit_rev = (last_year_rev / last_year_vol * 10000) if last_year_vol > 0 else 0  # 元/TEU
    
    # 计算箱量变化和单箱收入变化对收入的影响
    if last_year_vol > 0 and last_year_unit_rev > 0:
        # 箱量变化影响 = (当前箱量 - 去年箱量) × 去年单箱收入
        vol_impact = (current_vol - last_year_vol) * last_year_unit_rev / 10000  # 万元
        
        # 单箱收入变化影响 = (当前单箱收入 - 去年单箱收入) × 当前箱量
        price_impact = (current_unit_rev - last_year_unit_rev) * current_vol / 10000  # 万元
        
        vol_change_pct = ((current_vol - last_year_vol) / last_year_vol * 100) if last_year_vol > 0 else 0
        unit_rev_change_pct = ((current_unit_rev - last_year_unit_rev) / last_year_unit_rev * 100) if last_year_unit_rev > 0 else 0
        
        factors.append({
            'type': 'volume_price_analysis',
            'current_vol': current_vol,
            'last_year_vol': last_year_vol,
            'vol_change': vol_change_pct,
            'current_unit_rev': current_unit_rev,
            'last_year_unit_rev': last_year_unit_rev,
            'unit_rev_change': unit_rev_change_pct,
            'vol_impact': vol_impact,
            'price_impact': price_impact
        })
        
        # 判断主因
        if abs(vol_impact) > abs(price_impact):
            factors.append(f"**主因是箱量变化**：箱量{current_vol:,.0f} TEU vs 去年{last_year_vol:,.0f} TEU（{vol_change_pct:+.1f}%），影响收入{vol_impact:,.0f}万元")
        else:
            factors.append(f"**主因是单箱收入变化**：单箱{current_unit_rev:,.0f}元/TEU vs 去年{last_year_unit_rev:,.0f}元/TEU（{unit_rev_change_pct:+.1f}%），影响收入{price_impact:,.0f}万元")
    
    return factors


def analyze_yoy_decline(segment, actual, actual_yoy, volume, volume_yoy, period):
    """
    分析同比下降原因（深度分析）
    
    需要分析：
    1. 箱量 vs 价格因素
    2. 客户维度分析（前十大客户变化）
    3. 销售维度分析
    """
    factors = []
    
    current_rev = actual.get('cumulative', {}).get(f'{segment}_revenue', 0)
    last_year_rev = actual_yoy.get('cumulative', {}).get(f'{segment}_revenue', 0)
    
    if last_year_rev > 0:
        decline_amount = last_year_rev - current_rev
        decline_rate = decline_amount / last_year_rev * 100
        
        factors.append(f"**收入同比下降{decline_rate:.1f}%**，减少{decline_amount:,.0f}万元")
        
        # 1. 箱量 vs 价格分析
        vol_price_analysis = analyze_volume_vs_price(segment, actual, actual_yoy, volume, volume_yoy)
        factors.extend(vol_price_analysis)
        
        # 2. 客户维度分析（读取客户数据）
        customer_analysis = analyze_customer_impact(segment, period)
        if customer_analysis:
            factors.append(customer_analysis)
        
        # 3. 建议深入分析方向
        factors.append("**建议深入分析**：")
        factors.append(f"- 对比去年同期{segment.upper()}段客户清单，识别流失客户和减少合作的客户")
        factors.append(f"- 分析{segment.upper()}段前 10 大客户同期收入变化")
        factors.append(f"- 分析{segment.upper()}段销售人员业绩变化")
    
    return factors


def analyze_customer_impact(segment, period):
    """
    分析客户维度影响
    
    需要读取客户数据，对比同期
    """
    try:
        import openpyxl
        import glob
        from datetime import datetime, timedelta
        
        # 读取所有业务收入明细
        files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
        if not files:
            return None
        
        f = files[0]
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        
        # 找表头
        headers = {}
        for col in range(1, 20):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[str(h).strip()] = col
        
        bu_col = headers.get('业务系统单元编码', 1)
        segment_col = headers.get('业务分段分类_新.', 2)
        month_col = headers.get('业务年月', 3)
        customer_col = headers.get('委托客户名称', 4)
        revenue_col = headers.get('收入', 8)
        
        # 解析期间
        current_month, cumulative_months = parse_period(period)
        
        # 汇总客户收入（当期 vs 去年同期）
        customer_revenue = {}
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            bu = ws.cell(row=row, column=bu_col).value
            seg = ws.cell(row=row, column=segment_col).value
            month = ws.cell(row=row, column=month_col).value
            customer = ws.cell(row=row, column=customer_col).value
            revenue = ws.cell(row=row, column=revenue_col).value
            
            if not bu or str(bu).strip() != 'BWLDTC':
                continue
            
            # 判断业务段
            seg_str = str(seg).strip() if seg else ''
            if segment == 'b' and seg_str not in ['B 段', 'B', 'A+B', 'A+B+C']:
                continue
            elif segment == 'c' and seg_str not in ['C 段', 'C', 'C+D']:
                continue
            elif segment == 'd' and seg_str not in ['D 段', 'D'] and '电商' not in seg_str:
                continue
            
            if not month or not customer or not revenue:
                continue
            
            # 处理月份
            if isinstance(month, (int, float)):
                base_date = datetime(1899, 12, 30)
                actual_date = base_date + timedelta(days=int(month))
                year = actual_date.year
                month_num = actual_date.month
            else:
                continue
            
            # 统计当期和去年同期
            if year == 2026 and month_num in cumulative_months:
                if customer not in customer_revenue:
                    customer_revenue[customer] = {'current': 0, 'last_year': 0}
                customer_revenue[customer]['current'] += float(revenue) if revenue else 0
            
            # 去年同期（2025 年同期）
            elif year == 2025 and month_num in cumulative_months:
                if customer not in customer_revenue:
                    customer_revenue[customer] = {'current': 0, 'last_year': 0}
                customer_revenue[customer]['last_year'] += float(revenue) if revenue else 0
        
        # 找出收入下降最多的客户
        declining_customers = []
        for customer, data in customer_revenue.items():
            if data['last_year'] > 0 and data['current'] < data['last_year']:
                decline = data['last_year'] - data['current']
                declining_customers.append((customer, data['last_year'], data['current'], decline))
        
        declining_customers.sort(key=lambda x: x[3], reverse=True)
        
        if declining_customers:
            analysis = f"**客户维度分析**：收入下降客户共{len(declining_customers)}家，其中：\n"
            for i, (cust, last, curr, decline) in enumerate(declining_customers[:5], 1):
                analysis += f"- Top{i} {cust}：{last/10000:.0f}万 → {curr/10000:.0f}万，减少{decline/10000:.0f}万元\n"
            return analysis
        else:
            return "**客户维度分析**：未发现明显收入下降客户，需分析新客户开发不足问题"
    
    except Exception as e:
        return f"**客户维度分析**：数据读取失败 ({e})"


def calculate_impact(actual, actual_yoy, segment):
    """计算影响程度"""
    current = actual.get('cumulative', {}).get(f'{segment}_revenue', 0)
    last_year = actual_yoy.get('cumulative', {}).get(f'{segment}_revenue', 0)
    
    if last_year > 0:
        change = current - last_year
        return f"影响整体收入{abs(change):,.0f}万元"
    return "影响待评估"


def generate_recommendations(revenue_analysis, anomalies, period):
    """生成行动建议"""
    recommendations = []
    
    # 根据异常类型生成建议
    for anomaly in anomalies:
        seg_name = anomaly['segment_name']
        
        if anomaly['type'] == 'budget_miss':
            if anomaly['severity'] == 'high':
                recommendations.append({
                    'priority': '高',
                    'issue': f"{seg_name}预算达成严重不足",
                    'actions': [
                        f"本周内召开{seg_name}专项分析会，明确差距原因",
                        "梳理{seg_name}前 10 大客户，逐一拜访了解需求",
                        "制定 Q2 追赶计划，明确责任人和时间节点"
                    ],
                    'owner': '业务负责人',
                    'timeline': '1 周内启动'
                })
            else:
                recommendations.append({
                    'priority': '中',
                    'issue': f"{seg_name}预算达成有待提升",
                    'actions': [
                        f"分析{seg_name}客户结构，识别增长机会",
                        "加强销售过程管理，提升转化率"
                    ],
                    'owner': '销售总监',
                    'timeline': '2 周内'
                })
        
        elif anomaly['type'] == 'yoy_decline':
            recommendations.append({
                'priority': '高' if anomaly['severity'] == 'high' else '中',
                'issue': f"{seg_name}收入同比下滑",
                'actions': [
                    f"对比去年同期{seg_name}客户清单，识别流失客户",
                    "分析市场竞争态势，评估价格策略",
                    "制定客户挽回计划和新增客户目标"
                ],
                'owner': '业务负责人',
                'timeline': '1 个月内'
            })
        
        elif anomaly['type'] == 'profit_miss':
            recommendations.append({
                'priority': '高',
                'issue': '整体毛利预算达成不足',
                'actions': [
                    '分析各业务段毛利率，识别低毛利业务',
                    '评估成本结构，寻找降本空间',
                    '优化客户结构，提升高毛利客户占比'
                ],
                'owner': '财务总监',
                'timeline': '2 周内'
            })
    
    # 添加通用建议
    if len(anomalies) > 3:
        recommendations.append({
            'priority': '高',
            'issue': '多项指标异常，需系统性复盘',
            'actions': [
                '召开 Q1 经营复盘会，全面分析问题',
                '制定 Q2 经营改善专项行动计划',
                '建立周度跟踪机制，确保改善措施落地'
            ],
            'owner': '总经理',
            'timeline': '1 周内'
        })
    
    return recommendations


def format_ai_analysis_html(analysis_result):
    """将分析结果格式化为 HTML"""
    if not analysis_result.get('success'):
        return '<div class="ai-analysis">AI 分析暂不可用</div>'
    
    html = '''
    <div class="section">
        <div class="section-title">🤖 AI 深度分析</div>
        
        <div class="subsection">
            <div class="subsection-title">异常波动识别</div>
            <table>
                <thead>
                    <tr>
                        <th>业务段</th>
                        <th>异常类型</th>
                        <th>数值</th>
                        <th>严重程度</th>
                    </tr>
                </thead>
                <tbody>
'''
    
    for anomaly in analysis_result.get('anomalies', []):
        severity_class = 'text-danger' if anomaly['severity'] == 'high' else 'text-warning'
        html += f'''
                    <tr>
                        <td>{anomaly['segment_name']}</td>
                        <td>{anomaly['metric']}</td>
                        <td>{anomaly['value']}</td>
                        <td class="{severity_class}">{"⚠️ 高" if anomaly['severity'] == 'high' else "◔ 中"}</td>
                    </tr>
'''
    
    html += '''
                </tbody>
            </table>
        </div>
        
        <div class="subsection">
            <div class="subsection-title">根因分析</div>
            <div class="ai-analysis-content">
'''
    
    for item in analysis_result.get('revenue_analysis', []):
        html += f'''
                <div class="analysis-item">
                    <strong>{item['segment']}：</strong>{item['issue']}<br>
                    <ul>
'''
        for factor in item['factors']:
            # 处理字典类型的因素（箱量/价格分析结果）
            if isinstance(factor, dict):
                if factor.get('type') == 'volume_price_analysis':
                    html += f'<li>箱量：{factor["current_vol"]:,.0f} TEU vs 去年{factor["last_year_vol"]:,.0f} TEU ({factor["vol_change"]:+.1f}%)</li>\n'
                    html += f'<li>单箱收入：{factor["current_unit_rev"]:,.0f}元/TEU vs 去年{factor["last_year_unit_rev"]:,.0f}元/TEU ({factor["unit_rev_change"]:+.1f}%)</li>\n'
                    html += f'<li>箱量影响：{factor["vol_impact"]:,.0f}万元，单箱收入影响：{factor["price_impact"]:,.0f}万元</li>\n'
            else:
                html += f'<li>{factor}</li>\n'
        html += f'''
                    </ul>
                    <div class="analysis-impact">影响：{item['impact']}</div>
                </div>
'''
    
    html += '''
            </div>
        </div>
        
        <div class="subsection">
            <div class="subsection-title">行动建议</div>
            <table>
                <thead>
                    <tr>
                        <th>优先级</th>
                        <th>问题</th>
                        <th>行动措施</th>
                        <th>责任人</th>
                        <th>时间</th>
                    </tr>
                </thead>
                <tbody>
'''
    
    for rec in analysis_result.get('recommendations', []):
        priority_class = 'text-danger' if rec['priority'] == '高' else 'text-warning'
        actions = '<br>'.join([f'• {a}' for a in rec['actions']])
        html += f'''
                    <tr>
                        <td class="{priority_class}">{rec['priority']}</td>
                        <td>{rec['issue']}</td>
                        <td>{actions}</td>
                        <td>{rec['owner']}</td>
                        <td>{rec['timeline']}</td>
                    </tr>
'''
    
    html += '''
                </tbody>
            </table>
        </div>
    </div>
'''
    
    return html


if __name__ == '__main__':
    # 测试代码
    print('AI 分析模块加载成功')
