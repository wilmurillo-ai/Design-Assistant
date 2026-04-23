import openpyxl
from collections import defaultdict
import os
import glob

# 数据目录（本地路径）
BUSINESS_DATA_DIR = r"C:\Users\wwl\.openclaw\workspace-跨境电商\data\1.业务和订单数据"

def parse_period(period):
    """解析期间参数"""
    if period.upper().startswith('2026-Q'):
        quarter = int(period[-1])
        current_month = quarter * 3
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    if period.startswith('2026-'):
        month_part = period[5:]
        current_month = int(month_part)
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    return 3, [1, 2, 3]


def read_customer_data(period='2026-Q1'):
    """读取客户开发数据
    
    返回：
    - 当月客户开发数（分 B/C/D 段，去重）
    - 累计客户开发数（分 B/C/D 段，去重）
    - 分销售 D 段客户开户数
    """
    current_month, cumulative_months = parse_period(period)
    
    result = {
        'current': {'b_customers': set(), 'c_customers': set(), 'd_customers': set(), 'total': set()},
        'cumulative': {'b_customers': set(), 'c_customers': set(), 'd_customers': set(), 'total': set()},
        'sales_d_customers': {},  # 销售 -> D 段客户集合（累计）
        'current_sales_d_customers': {}  # 销售 -> D 段客户集合（当月）
    }
    
    # 从历史客户合作情况汇总读取
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*历史客户合作情况汇总*.xlsx'))
    if not files:
        return result
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    # 找表头
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    customer_col = headers.get('客户名称', 1)
    segment_col = headers.get('业务分段列表', 2)
    year_col = headers.get('第一次合作年', 4)
    month_col = headers.get('第一次合作月', 5)
    sales_col = headers.get('销售', 7)
    
    for row in range(2, min(ws.max_row + 1, 5000)):
        customer = ws.cell(row=row, column=customer_col).value
        segment = ws.cell(row=row, column=segment_col).value
        year = ws.cell(row=row, column=year_col).value
        month = ws.cell(row=row, column=month_col).value
        sales = ws.cell(row=row, column=sales_col).value
        
        if not customer or not year or not month:
            continue
        
        # 判断是否在统计期间内
        if year == 2026 and month in cumulative_months:
            seg_str = str(segment).strip() if segment else ''
            
            # 判断业务段（电商集拼不属于 D 段）
            has_b = 'A' in seg_str or 'B' in seg_str
            has_c = 'C' in seg_str
            has_d = ('D' in seg_str or seg_str == 'D') and '电商' not in seg_str and '集拼' not in seg_str
            
            if has_b:
                result['cumulative']['b_customers'].add(customer)
            if has_c:
                result['cumulative']['c_customers'].add(customer)
            if has_d:
                result['cumulative']['d_customers'].add(customer)
                # 分销售统计
                if sales:
                    if sales not in result['sales_d_customers']:
                        result['sales_d_customers'][sales] = set()
                    result['sales_d_customers'][sales].add(customer)
            
            # 当月
            if month == current_month:
                if has_b:
                    result['current']['b_customers'].add(customer)
                if has_c:
                    result['current']['c_customers'].add(customer)
                if has_d:
                    result['current']['d_customers'].add(customer)
                    # 分销售统计（当月）
                    if sales:
                        if sales not in result['current_sales_d_customers']:
                            result['current_sales_d_customers'][sales] = set()
                        result['current_sales_d_customers'][sales].add(customer)
    
    # 计算总数（去重）
    result['cumulative']['total'] = (
        result['cumulative']['b_customers'] | 
        result['cumulative']['c_customers'] | 
        result['cumulative']['d_customers']
    )
    result['current']['total'] = (
        result['current']['b_customers'] | 
        result['current']['c_customers'] | 
        result['current']['d_customers']
    )
    
    return result


def read_top_customers_and_sales(period='2026-Q1', top_n=10):
    """读取前十大客户和销售贡献
    
    返回：
    - top_customers: [(客户名称，累计收入，累计毛利), ...]
    - top_current_customers: [(客户名称，当月收入，当月毛利), ...]
    - top_sales: [(销售，收入，毛利), ...]
    - top_sales_d_volume: [(销售，D 段箱量), ...]
    """
    current_month, cumulative_months = parse_period(period)
    
    # 从所有业务收入明细读取
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if not files:
        return [], [], []
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    bu_col = headers.get('业务系统单元编码', 1)
    customer_col = headers.get('委托客户名称', 4)
    sales_col = headers.get('销售', 5)
    revenue_col = headers.get('收入', 8)
    profit_col = headers.get('毛利', 6)
    month_col = headers.get('业务年月', 3)
    
    # 排除内部公司关键词
    internal_keywords = ['环世', 'WORLDWIDE LOGISTICS', 'WORLDWIDE SHIPPING']
    
    def is_internal(customer):
        if not customer:
            return False
        for kw in internal_keywords:
            if kw in str(customer):
                return True
        return False
    
    # 汇总客户数据（累计 + 当月）
    customer_data = defaultdict(lambda: {'revenue': 0, 'profit': 0})  # 累计
    customer_current_data = defaultdict(lambda: {'revenue': 0, 'profit': 0})  # 当月
    sales_data = defaultdict(lambda: {'revenue': 0, 'profit': 0})
    
    for row in range(2, min(ws.max_row + 1, 50000)):
        bu = ws.cell(row=row, column=bu_col).value
        month = ws.cell(row=row, column=month_col).value
        
        if not bu or str(bu).strip() != 'BWLDTC':
            continue
        
        if not month:
            continue
        
        # 处理月份：支持字符串和 Excel 日期序列号
        from datetime import datetime, timedelta
        if isinstance(month, (int, float)):
            # Excel 日期序列号
            base_date = datetime(1899, 12, 30)
            actual_date = base_date + timedelta(days=int(month))
            month_year = actual_date.year
            month_num = actual_date.month
        else:
            month_str = str(month)
            if not month_str.startswith('2026'):
                continue
            if '年' in month_str:
                month_year = int(month_str.split('年')[0].strip())
                month_num = int(month_str.split('年')[1].replace('月', '').strip())
            elif '-' in month_str:
                parts = month_str.split('-')
                month_year = int(parts[0].strip())
                month_num = int(parts[1].strip())
            else:
                continue
        
        if month_year != 2026 or month_num not in cumulative_months:
            continue
        
        customer = ws.cell(row=row, column=customer_col).value
        sales = ws.cell(row=row, column=sales_col).value
        revenue = ws.cell(row=row, column=revenue_col).value
        profit = ws.cell(row=row, column=profit_col).value
        
        # 排除内部公司（前十大客户）
        if customer and not is_internal(customer):
            customer_data[customer]['revenue'] += float(revenue) if revenue else 0
            customer_data[customer]['profit'] += float(profit) if profit else 0
            # 当月数据
            if month_num == current_month:
                customer_current_data[customer]['revenue'] += float(revenue) if revenue else 0
                customer_current_data[customer]['profit'] += float(profit) if profit else 0
        
        # 销售数据不排除
        if sales:
            sales_data[sales]['revenue'] += float(revenue) if revenue else 0
            sales_data[sales]['profit'] += float(profit) if profit else 0
    
    # 排序取 Top N（转换为万元）- 按累计收入排序
    top_customers = sorted(
        [(name, data['revenue']/10000, data['profit']/10000) for name, data in customer_data.items()],
        key=lambda x: x[1],
        reverse=True
    )[:top_n]
    
    # 当月 Top 客户（按当月收入排序）
    top_current_customers = sorted(
        [(name, data['revenue']/10000, data['profit']/10000) for name, data in customer_current_data.items()],
        key=lambda x: x[1],
        reverse=True
    )[:top_n]
    
    top_sales = sorted(
        [(name, data['revenue']/10000, data['profit']/10000) for name, data in sales_data.items()],
        key=lambda x: x[1],
        reverse=True
    )[:top_n]
    
    # 读取前十大销售的 D 段箱量（仅统计 2026 Q1）
    top_sales_names = [s[0] for s in top_sales]
    sales_d_volume = defaultdict(float)
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        headers = {}
        for col in range(1, 30):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[str(h).strip()] = col
        
        sales_col_idx = headers.get('销售', 2)
        month_col = headers.get('业务月份', 4)
        volume_col = headers.get('入库箱量', 8)
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            sales = ws.cell(row=row, column=sales_col_idx).value
            month = ws.cell(row=row, column=month_col).value
            volume = ws.cell(row=row, column=volume_col).value
            
            if not sales or not volume or not month:
                continue
            
            # 只统计 2026 Q1 数据
            month_str = str(month)
            if '-' in month_str:
                parts = month_str.split('-')
                year = parts[0].strip()
                month_num = int(parts[1].strip()) if len(parts) >= 2 else 0
                if year != '2026' or month_num not in [1, 2, 3]:
                    continue
            else:
                continue
            
            # 支持模糊匹配：夏胜利 CBE(夏胜利) 匹配 夏胜利
            sales_base = str(sales).split('CBE')[0].split('(')[0].strip()
            for top_sales_name in top_sales_names:
                if sales_base == top_sales_name or sales == top_sales_name:
                    sales_d_volume[top_sales_name] += float(volume)
                    break
    
    # D 段箱量转换为 TEU（原始数据是箱量，已经是 TEU 单位）
    top_sales_d_volume = [(name, sales_d_volume.get(name, 0)) for name in top_sales_names]
    top_sales_d_volume.sort(key=lambda x: x[1], reverse=True)
    
    # 读取当月数据用于排名
    current_month, _ = parse_period(period)
    current_customers = defaultdict(lambda: {'revenue': 0, 'profit': 0})
    current_sales = defaultdict(lambda: {'revenue': 0, 'profit': 0})
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if files:
        f = files[0]
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        
        headers = {}
        for col in range(1, 20):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[str(h).strip()] = col
        
        bu_col = headers.get('业务系统单元编码', 1)
        customer_col = headers.get('委托客户名称', 4)
        sales_col = headers.get('销售', 5)
        revenue_col = headers.get('收入', 8)
        profit_col = headers.get('毛利', 6)
        month_col = headers.get('业务年月', 3)
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            bu = ws.cell(row=row, column=bu_col).value
            month = ws.cell(row=row, column=month_col).value
            
            if not bu or str(bu).strip() != 'BWLDTC':
                continue
            
            if not month:
                continue
            
            # 处理月份：支持字符串和 Excel 日期序列号
            from datetime import datetime, timedelta
            if isinstance(month, (int, float)):
                base_date = datetime(1899, 12, 30)
                actual_date = base_date + timedelta(days=int(month))
                month_num = actual_date.month
            elif '年' in str(month):
                month_str = str(month).replace('年', '-').replace('月', '').strip()
                try:
                    month_num = int(month_str.split('-')[1])
                except:
                    continue
            elif '-' in str(month):
                month_str = str(month).replace('-0', '-').strip()
                try:
                    month_num = int(month_str.split('-')[1])
                except:
                    continue
            else:
                continue
            
            # 只统计当月数据
            if month_num != current_month:
                continue
            
            customer = ws.cell(row=row, column=customer_col).value
            sales = ws.cell(row=row, column=sales_col).value
            revenue = ws.cell(row=row, column=revenue_col).value
            profit = ws.cell(row=row, column=profit_col).value
            
            if customer and not is_internal(customer):
                current_customers[customer]['revenue'] += float(revenue) if revenue else 0
                current_customers[customer]['profit'] += float(profit) if profit else 0
            
            if sales:
                current_sales[sales]['revenue'] += float(revenue) if revenue else 0
                current_sales[sales]['profit'] += float(profit) if profit else 0
    
    # 当月排名前十大（转换为万元）
    top_current_customers = sorted(
        [(name, data['revenue']/10000, data['profit']/10000) for name, data in current_customers.items()],
        key=lambda x: x[1],
        reverse=True
    )[:10]
    
    top_current_sales = sorted(
        [(name, data['revenue']/10000, data['profit']/10000) for name, data in current_sales.items()],
        key=lambda x: x[1],
        reverse=True
    )[:10]
    
    return top_customers, top_sales, top_sales_d_volume, top_current_customers, top_current_sales
