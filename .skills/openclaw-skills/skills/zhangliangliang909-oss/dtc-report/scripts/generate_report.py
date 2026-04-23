#!/usr/bin/env python3
"""
DTC 经营分析报告生成器

按照经营分析/预实分析取数规则生成报告：
- 收入/毛利：从"所有业务收入明细（非订单维度）"取数（非管报）
- 箱量：从"DTC 明细表 - 全链路视角"取 B/C 段，从"客户仓库统计表"取 D 段
- 出库件数：从"客户仓库统计表"取数
- 预算数据：从"2026 年 DTC 预算数据"取合计数
- 同比数据：从相同数据源读取 2025 年同期数据
- 趋势数据：读取最近 6-12 个月数据生成趋势图表

用法：
    python generate_report.py --period 2026-Q1 --output reports/
"""

import openpyxl
from collections import defaultdict
import os
import glob
import json
import argparse
import re
from datetime import datetime, timedelta

# 导入客户与销售模块
from customer_sales import read_customer_data, read_top_customers_and_sales
# 导入海外仓模块
from warehouse import read_warehouse_data, read_loss_customers

# 导入 AI 分析模块
try:
    from ai_analysis import analyze_business_performance, format_ai_analysis_html
    AI_ANALYSIS_ENABLED = True
except ImportError:
    AI_ANALYSIS_ENABLED = False
    print('⚠️ AI 分析模块未加载，将使用规则分析')

# 导入客户年代分析模块
try:
    from customer_vintage_analysis import analyze_customer_vintage, format_customer_vintage_html
    CUSTOMER_VINTAGE_ENABLED = True
except ImportError:
    CUSTOMER_VINTAGE_ENABLED = False
    print('⚠️ 客户年代分析模块未加载')
# 导入关键发现模块
from key_findings import generate_key_findings

# 数据目录（本地路径）
DATA_DIR = r"C:\Users\wwl\.openclaw\workspace-跨境电商\data"
MGMT_REPORT_DIR = os.path.join(DATA_DIR, '4.经营管理报表')
BUSINESS_DATA_DIR = os.path.join(DATA_DIR, '1.业务和订单数据')
BUDGET_DIR = os.path.join(DATA_DIR, '6.2026 年预算数据')
LEARN_DIR = os.path.join(DATA_DIR, '学习记录')


def parse_period(period):
    """解析期间参数，返回当月月份和累计月份范围"""
    if period.upper().startswith('2026-Q'):
        quarter = int(period[-1])
        current_month = quarter * 3
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    if period.startswith('2026-'):
        month_part = period[5:]
        if '-' in month_part and '月' in month_part:
            end_month = int(month_part.split('-')[1].replace('月', ''))
            current_month = end_month
            cumulative_months = list(range(1, end_month + 1))
        else:
            current_month = int(month_part)
            cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    return 3, [1, 2, 3]


def get_yoy_period(period):
    """获取同比期间（2025 年同期）"""
    if period.upper().startswith('2026-Q'):
        quarter = int(period[-1])
        return f'2025-Q{quarter}'
    
    if period.startswith('2026-'):
        month_part = period[5:]
        return f'2025-{month_part}'
    
    return '2025-Q1'


def read_budget_data():
    """从 2026 年 DTC 预算 Excel 读取分月预算数据
    
    预算表结构：
    - 行 25-33：业务收入（B 段、C 段、D 段、电商集拼、其他）
    - 行 48-52：业务毛利（B 段、C 段、D 段、电商集拼、其他）
    - 列 D-F（4-6）：2026 年 1-3 月
    """
    # 使用 os.walk 查找预算文件（避免编码问题）
    budget_file = None
    for root, dirs, files in os.walk(DATA_DIR):
        for f in files:
            if '6.2026' in root and 'DTC' in f and '预算' in f and 'DTC1' not in f:
                budget_file = os.path.join(root, f)
                break
        if budget_file:
            break
    
    if not budget_file:
        print("  ⚠️ 未找到预算文件")
        return {}
    
    try:
        wb = openpyxl.load_workbook(budget_file, data_only=True)
        ws = wb.active
        
        monthly_rev = {'b': [], 'c': [], 'd': [], 'ecom': [], 'other': []}
        monthly_profit = {'b': [], 'c': [], 'd': [], 'ecom': [], 'other': []}
        
        for row in range(25, 55):
            a_col = ws.cell(row=row, column=1).value
            b_col = ws.cell(row=row, column=2).value
            if not a_col or not b_col:
                continue
            
            a_str = str(a_col).strip()
            b_str = str(b_col).strip().replace(' ', '')
            
            if '业务收入' in a_str:
                data_type = 'revenue'
            elif '业务毛利' in a_str:
                data_type = 'profit'
            else:
                continue
            
            if 'B' in b_str and '段' in b_str: segment = 'b'
            elif 'C' in b_str and '段' in b_str: segment = 'c'
            elif 'D' in b_str and '段' in b_str: segment = 'd'
            elif '电商集拼' in b_str or '集拼' in b_str: segment = 'ecom'
            elif b_str == '其他' or b_str == '其他业务': segment = 'other'
            else:
                continue
            
            month_data = []
            for col in range(4, 7):  # D-F 列
                val = ws.cell(row=row, column=col).value
                month_data.append(float(val) if val else 0)
            
            if data_type == 'revenue':
                monthly_rev[segment] = month_data
            else:
                monthly_profit[segment] = month_data
        
        # 打印调试信息
        # 转换为万元
        for seg in ['b', 'c', 'd', 'ecom', 'other']:
            monthly_rev[seg] = [v/10000 for v in monthly_rev[seg]]
            monthly_profit[seg] = [v/10000 for v in monthly_profit[seg]]
        
        print(f"    收入预算（3 月）: B={monthly_rev['b'][2]:.0f}万，C={monthly_rev['c'][2]:.0f}万，D={monthly_rev['d'][2]:.0f}万")
        print(f"    收入预算（Q1）: B={sum(monthly_rev['b']):.0f}万，C={sum(monthly_rev['c']):.0f}万，D={sum(monthly_rev['d']):.0f}万")
        print(f"    毛利预算（3 月）: B={monthly_profit['b'][2]:.1f}万，C={monthly_profit['c'][2]:.0f}万，D={monthly_profit['d'][2]:.0f}万")
        print(f"    毛利预算（Q1）: B={sum(monthly_profit['b']):.1f}万，C={sum(monthly_profit['c']):.0f}万，D={sum(monthly_profit['d']):.0f}万")
        
        return {
            # 3 月数据（万元）
            'b_revenue': monthly_rev['b'][2],
            'c_revenue': monthly_rev['c'][2],
            'd_revenue': monthly_rev['d'][2],
            'ecom_revenue': monthly_rev['ecom'][2],
            'other_revenue': monthly_rev['other'][2],
            'b_profit': monthly_profit['b'][2],
            'c_profit': monthly_profit['c'][2],
            'd_profit': monthly_profit['d'][2],
            'ecom_profit': monthly_profit['ecom'][2],
            'other_profit': monthly_profit['other'][2],
            # Q1 累计（逐月累加 1-3 月）
            'b_revenue_cumulative': sum(monthly_rev['b'][:3]),
            'c_revenue_cumulative': sum(monthly_rev['c'][:3]),
            'd_revenue_cumulative': sum(monthly_rev['d'][:3]),
            'ecom_revenue_cumulative': sum(monthly_rev['ecom'][:3]),
            'other_revenue_cumulative': sum(monthly_rev['other'][:3]),
            'revenue_cumulative': sum(sum(v[:3]) for v in monthly_rev.values()),
            'b_profit_cumulative': sum(monthly_profit['b'][:3]),
            'c_profit_cumulative': sum(monthly_profit['c'][:3]),
            'd_profit_cumulative': sum(monthly_profit['d'][:3]),
            'ecom_profit_cumulative': sum(monthly_profit['ecom'][:3]),
            'other_profit_cumulative': sum(monthly_profit['other'][:3]),
            'gross_profit_cumulative': sum(sum(v[:3]) for v in monthly_profit.values()),
            # 箱量预算
            'b_volume': 420, 'c_volume': 806, 'd_volume': 890,
            'b_volume_cumulative': 1265, 'c_volume_cumulative': 2442, 'd_volume_cumulative': 2574,
            # 出库件数预算
            'outbound_pieces': 167592, 'outbound_pieces_cumulative': 411077
        }
        
    except Exception as e:
        print(f"  读取预算失败：{e}")
        import traceback
        traceback.print_exc()
        return {}


def read_volume_data(period='2026-Q1', yoy=False, mom=False):
    """读取箱量数据（区分当月/累计，支持同比、环比）
    
    mom=True 时读取上月数据用于计算环比
    """
    current_month, cumulative_months = parse_period(period)
    
    if mom:
        # 读取上月数据
        prev_month = current_month - 1 if current_month > 1 else 12
        prev_year = '2026' if current_month > 1 else '2025'
        target_year = prev_year
        cumulative_months = [prev_month]
    else:
        target_year = '2025' if yoy else '2026'
    
    result = {
        'current': {'b_volume': 0, 'c_volume': 0, 'd_volume': 0},
        'cumulative': {'b_volume': 0, 'c_volume': 0, 'd_volume': 0}
    }
    
    # B/C 段箱量 - 从全链路表 (使用 read_only 模式节省内存)
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*全链路*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True, read_only=True)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # row 是元组：(列 1, 列 2, 列 3, ...)
            # 列 5=销售公司，列 11=订单服务项，列 3=业务日期，列 17=未去重集装箱箱量
            sales_company = row[4] if len(row) > 4 else None  # 列 5
            date_val = row[2] if len(row) > 2 else None  # 列 3
            segment = row[10] if len(row) > 10 else None  # 列 11
            volume = row[16] if len(row) > 16 else None  # 列 17
            
            # ⚠️ 2026-04-11 更新：仅统计销售公司=BWLDTC 的箱量
            if not sales_company or str(sales_company).strip() != 'BWLDTC':
                continue
            
            if segment and volume and date_val:
                date_str = str(date_val)
                if not date_str.startswith(f'{target_year}-0'):
                    continue
                
                month_num = int(date_str.split('-')[1])
                vol = float(volume) if volume else 0
                seg_str = str(segment).strip()
                
                if seg_str in ['A+B+C']:
                    seg_str = 'B'
                elif seg_str in ['A+B']:
                    seg_str = 'B'
                elif seg_str in ['C+D']:
                    seg_str = 'C'
                elif seg_str == 'D':
                    continue
                
                if month_num in cumulative_months:
                    if seg_str == 'B':
                        result['cumulative']['b_volume'] += vol
                    elif seg_str == 'C':
                        result['cumulative']['c_volume'] += vol
                
                if month_num == current_month:
                    if seg_str == 'B':
                        result['current']['b_volume'] += vol
                    elif seg_str == 'C':
                        result['current']['c_volume'] += vol
    
    # D 段箱量 - 从客户仓库统计表 (使用 read_only 模式节省内存)
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True, read_only=True)
        ws = wb.active
        
        filename = os.path.basename(files[0])
        end_day = 30
        try:
            match = re.search(r'(\d{4}) 年至 (\d{4}) 年 (\d{1,2}) 月 (\d{1,2}) 日', filename)
            if match:
                end_day = int(match.group(4))
        except:
            pass
        
        actual_days = {f'{target_year}-01': 31, f'{target_year}-02': 28, f'{target_year}-03': 31}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 列 4=业务月份，列 8=入库箱量
            month = row[3] if len(row) > 3 else None
            volume = row[7] if len(row) > 7 else None
            
            if month and volume:
                month_str = str(month)
                # 按年份过滤
                if not month_str.startswith(target_year):
                    continue
                
                # 标准化月份格式（处理 datetime 字符串如 "2026-03-01 00:00:00"）
                if len(month_str) >= 7:
                    month_num = int(month_str.split('-')[1])
                else:
                    continue
                
                vol = float(volume) if volume else 0
                
                # 累计和当月数据：都用原始数据，不做折算
                # 文件名显示数据截止月份，说明该月是完整月份
                if month_num in cumulative_months:
                    result['cumulative']['d_volume'] += vol
                
                if month_num == current_month:
                    result['current']['d_volume'] += vol
    
    return result


def read_outbound_pieces(period='2026-Q1', yoy=False, mom=False):
    """读取出库件数（区分当月/累计，支持同比、环比）"""
    from datetime import datetime
    
    if mom:
        # 读取上月数据
        current_month, _ = parse_period(period)
        prev_month = current_month - 1 if current_month > 1 else 12
        prev_year = '2026' if current_month > 1 else '2025'
        target_year = prev_year
        cumulative_months = [prev_month]
    else:
        target_year = '2025' if yoy else '2026'
        yoy_period = get_yoy_period(period) if yoy else period
        current_month, cumulative_months = parse_period(yoy_period)
    
    result = {'current': 0, 'cumulative': 0}
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        filename = os.path.basename(files[0])
        end_day = 30
        try:
            match = re.search(r'(\d{4}) 年至 (\d{4}) 年 (\d{1,2}) 月 (\d{1,2}) 日', filename)
            if match:
                end_day = int(match.group(4))
        except:
            pass
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            month = ws.cell(row=row, column=4).value
            pieces = ws.cell(row=row, column=19).value
            
            if not month or not pieces:
                continue
            
            # 处理 datetime 类型
            if isinstance(month, datetime):
                month_str = f"{month.year}-{month.month}"
                month_num = month.month
                year = month.year
            else:
                month_str = str(month)
                if '-' in month_str:
                    try:
                        parts = month_str.split('-')
                        if len(parts) >= 2:
                            year = int(parts[0])
                            month_num = int(parts[1])
                        else:
                            continue
                    except:
                        continue
                else:
                    continue
            
            # 只统计目标年份的数据
            if year != int(target_year):
                continue
            
            pcs = float(pieces) if pieces else 0
            
            if month_num in cumulative_months:
                result['cumulative'] += pcs
            
            if month_num == current_month:
                result['current'] += pcs
    
    return result


def read_revenue_profit_by_segment(period='2026-Q1', yoy=False, current_day=None):
    """从所有业务收入明细读取分业务段收入/毛利（区分当月/累计，支持同比）
    
    当月还没结束时的处理：
    - B 段/C 段：按上月单箱收入×本月箱量预估
    - D 段：按天数折算
    """
    target_year = '2025' if yoy else '2026'
    yoy_period = get_yoy_period(period) if yoy else period
    current_month, cumulative_months = parse_period(yoy_period)
    
    # 如果是 2026 年且是当月，需要检查是否月底
    need_adjustment = (not yoy) and (target_year == '2026') and (current_month <= 3)
    
    # 从文件名提取截止日期
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    end_day = 31  # 默认月底
    if files and need_adjustment:
        filename = os.path.basename(files[0])
        try:
            match = re.search(r'(\d{4}) 年至 (\d{4}) 年 (\d{1,2}) 月 (\d{1,2}) 日', filename)
            if match:
                end_day = int(match.group(4))
        except:
            pass
    
    result = {
        'current': {
            'b_revenue': 0, 'b_profit': 0,
            'c_revenue': 0, 'c_profit': 0,
            'd_revenue': 0, 'd_profit': 0,
            'ecom_revenue': 0, 'ecom_profit': 0,
            'other_revenue': 0, 'other_profit': 0
        },
        'cumulative': {
            'b_revenue': 0, 'b_profit': 0,
            'c_revenue': 0, 'c_profit': 0,
            'd_revenue': 0, 'd_profit': 0,
            'ecom_revenue': 0, 'ecom_profit': 0,
            'other_revenue': 0, 'other_profit': 0
        }
    }
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if not files:
        return result
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    bu_col = headers.get('业务系统单元编码', 1)
    # 优先使用"业务分段分类_新."字段（所有业务收入明细表），其次使用"订单服务项"
    segment_col = headers.get('业务分段分类_新.') or headers.get('订单服务项', 11)
    month_col = headers.get('业务年月', 3)
    revenue_col = headers.get('收入', 8)
    profit_col = headers.get('毛利', 6)
    
    target_bu = 'BWLDTC'
    
    for row in range(2, min(ws.max_row + 1, 50000)):
        bu = ws.cell(row=row, column=bu_col).value
        segment = ws.cell(row=row, column=segment_col).value
        month = ws.cell(row=row, column=month_col).value
        revenue = ws.cell(row=row, column=revenue_col).value
        profit = ws.cell(row=row, column=profit_col).value
        
        if not bu or str(bu).strip() != target_bu:
            continue
        
        if not month:
            continue
        
        # 处理月份：支持字符串和 Excel 日期序列号
        from datetime import datetime, timedelta
        if isinstance(month, (int, float)):
            # Excel 日期序列号（如 45658）
            base_date = datetime(1899, 12, 30)
            actual_date = base_date + timedelta(days=int(month))
            month_num = actual_date.month
            month_year = actual_date.year
        else:
            month_str = str(month)
            if '年' in month_str:
                month_year = int(month_str.split('年')[0].strip())
                month_num = int(month_str.split('年')[1].replace('月', '').strip())
            elif '-' in month_str:
                parts = month_str.split('-')
                month_year = int(parts[0].strip())
                month_num = int(parts[1].strip())
            else:
                continue
        
        # 检查年份
        if str(month_year) != target_year:
            continue
        
        if not segment:
            continue
        
        rev = float(revenue) if revenue else 0
        prof = float(profit) if profit else 0
        
        seg_str = str(segment).strip()
        
        if seg_str in ['B 段', 'B', 'A+B', 'A+B+C']:
            segment_key = 'b'
        elif seg_str in ['C 段', 'C', 'C+D']:
            segment_key = 'c'
        elif seg_str in ['D 段', 'D']:
            segment_key = 'd'
        elif '电商' in seg_str or '集拼' in seg_str:
            segment_key = 'ecom'
        else:
            segment_key = 'other'
        
        # 转换为万元（原始数据单位是元）
        rev_wan = rev / 10000
        prof_wan = prof / 10000
        
        if month_num in cumulative_months:
            result['cumulative'][f'{segment_key}_revenue'] += rev_wan
            result['cumulative'][f'{segment_key}_profit'] += prof_wan
        
        if month_num == current_month:
            result['current'][f'{segment_key}_revenue'] += rev_wan
            result['current'][f'{segment_key}_profit'] += prof_wan
    
    # 注意：文件名如"2025 年至 2026 年 3 月"表示完整月度数据，不再进行折算
    # 仅当文件名明确显示"XX 月 XX 日"且日期<28 时才考虑折算
    # 当前逻辑：不折算，直接使用原始数据
    # if need_adjustment and end_day < 28:
    #     # 仅当文件名显示日期<28 时才折算（如"3 月 20 日"）
    #     # D 段：按天数折算
    #     d_ratio = 31 / end_day  # 3 月按 31 天算
    #     result['current']['d_revenue'] *= d_ratio
    #     result['current']['d_profit'] *= d_ratio
    
    return result


def read_monthly_revenue_profit_single_month(year, month):
    """读取指定月份的收入/毛利和单箱收入"""
    result = {
        'b_revenue': 0, 'b_profit': 0, 'b_volume': 0,
        'c_revenue': 0, 'c_profit': 0, 'c_volume': 0,
        'b_revenue_per_box': 0, 'b_profit_per_box': 0,
        'c_revenue_per_box': 0, 'c_profit_per_box': 0
    }
    
    # 读取收入/毛利
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if files:
        f = files[0]
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        
        headers = {}
        for col in range(1, 20):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[str(h).strip()] = col
        
        bu_col = headers.get('业务系统单元编码', 1)
        segment_col = headers.get('订单服务项', 11)
        month_col = headers.get('业务年月', 3)
        revenue_col = headers.get('收入', 8)
        profit_col = headers.get('毛利', 6)
        
        target_month_str = f'{year}-{month}'
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            bu = ws.cell(row=row, column=bu_col).value
            segment = ws.cell(row=row, column=segment_col).value
            month_val = ws.cell(row=row, column=month_col).value
            revenue = ws.cell(row=row, column=revenue_col).value
            profit = ws.cell(row=row, column=profit_col).value
            
            if not bu or str(bu).strip() != 'BWLDTC':
                continue
            
            if not month_val:
                continue
            
            month_str = str(month_val).replace('年', '-').replace('月', '').replace('-0', '-').strip()
            if month_str != target_month_str:
                continue
            
            rev = float(revenue) if revenue else 0
            prof = float(profit) if profit else 0
            
            seg_str = str(segment).strip()
            if seg_str in ['B 段', 'B', 'A+B', 'A+B+C']:
                result['b_revenue'] += rev
                result['b_profit'] += prof
            elif seg_str in ['C 段', 'C', 'C+D']:
                result['c_revenue'] += rev
                result['c_profit'] += prof
    
    # 读取箱量
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*全链路*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        target_date_prefix = f'{year}-{month:02d}'
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            segment = ws.cell(row=row, column=11).value
            volume = ws.cell(row=row, column=17).value
            date_val = ws.cell(row=row, column=3).value
            
            if segment and volume and date_val:
                date_str = str(date_val)
                if not date_str.startswith(target_date_prefix):
                    continue
                
                vol = float(volume) if volume else 0
                seg_str = str(segment).strip()
                
                if seg_str in ['A+B+C', 'A+B', 'B']:
                    result['b_volume'] += vol
                elif seg_str in ['C+D', 'C']:
                    result['c_volume'] += vol
    
    # 计算单箱收入
    if result['b_volume'] > 0:
        result['b_revenue_per_box'] = result['b_revenue'] / result['b_volume']
        result['b_profit_per_box'] = result['b_profit'] / result['b_volume']
    
    if result['c_volume'] > 0:
        result['c_revenue_per_box'] = result['c_revenue'] / result['c_volume']
        result['c_profit_per_box'] = result['c_profit'] / result['c_volume']
    
    return result


def read_monthly_trend_data(months=6):
    """读取最近 N 个月的趋势数据（收入/毛利/箱量/出库件数）"""
    trend_data = {
        'months': [],
        'revenue': {'b': [], 'c': [], 'd': [], 'ecom': [], 'total': []},
        'profit': {'b': [], 'c': [], 'd': [], 'ecom': [], 'total': []},
        'volume': {'b': [], 'c': [], 'd': []},
        'outbound': []
    }
    
    # 读取收入/毛利趋势
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if files:
        f = files[0]
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        
        headers = {}
        for col in range(1, 20):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[str(h).strip()] = col
        
        bu_col = headers.get('业务系统单元编码', 1)
        # 优先使用"业务分段分类_新."字段（所有业务收入明细表），其次使用"订单服务项"
        segment_col = headers.get('业务分段分类_新.') or headers.get('订单服务项', 11)
        month_col = headers.get('业务年月', 3)
        revenue_col = headers.get('收入', 8)
        profit_col = headers.get('毛利', 6)
        
        # 按月汇总
        monthly_rev = defaultdict(lambda: {'b': 0, 'c': 0, 'd': 0, 'ecom': 0, 'other': 0})
        monthly_profit = defaultdict(lambda: {'b': 0, 'c': 0, 'd': 0, 'ecom': 0, 'other': 0})
        trend_data = {
    'months': [],
    'revenue': {'b': [], 'c': [], 'd': [], 'ecom': [], 'total': []},
    'volume': {'b': [], 'c': [], 'd': []},
    'profit': {'b': [], 'c': [], 'd': [], 'ecom': [], 'total': []},
    'outbound': []
}
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            bu = ws.cell(row=row, column=bu_col).value
            segment = ws.cell(row=row, column=segment_col).value
            month = ws.cell(row=row, column=month_col).value
            revenue = ws.cell(row=row, column=revenue_col).value
            profit = ws.cell(row=row, column=profit_col).value
            
            if not bu or str(bu).strip() != 'BWLDTC':
                continue
            
            if not month:
                continue
            
            # 处理月份：支持字符串和 Excel 日期序列号
            from datetime import datetime, timedelta
            if isinstance(month, (int, float)):
                # Excel 日期序列号（如 45658）
                base_date = datetime(1899, 12, 30)
                actual_date = base_date + timedelta(days=int(month))
                month_str = f"{actual_date.year}-{actual_date.month}"
            else:
                month_str = str(month).replace('年', '-').replace('月', '').strip()
            
            rev = float(revenue) if revenue else 0
            prof = float(profit) if profit else 0
            
            seg_str = str(segment).strip() if segment else ''
            # 使用订单服务项字段判断业务段
            if seg_str in ['B 段', 'B', 'A+B', 'A+B+C']:
                key = 'b'
            elif seg_str in ['C 段', 'C', 'C+D']:
                key = 'c'
            elif seg_str in ['D 段', 'D']:
                key = 'd'
            elif '电商集拼' in seg_str or '集拼' in seg_str:
                key = 'ecom'
            elif seg_str == '其他':
                key = 'other'
            else:
                continue  # 跳过不明确的分类'
            
            monthly_rev[month_str][key] += rev
            monthly_profit[month_str][key] += prof
        
        # 取最近 N 个月
        # 正确的月份排序
        def parse_month(m):
            try:
                parts = m.split('-')
                return (int(parts[0]), int(parts[1]))
            except:
                return (0, 0)
        
        # 过滤掉 4 月及以后的月份（Q1 报告只显示到 3 月）
        all_sorted = sorted(monthly_rev.keys(), key=parse_month)
        filtered_months = [m for m in all_sorted if not (m.startswith('2026-4') or m.startswith('2026-04'))]
        sorted_months = filtered_months[-months:] if len(filtered_months) >= months else filtered_months
        trend_data['months'] = sorted_months
        
        for m in sorted_months:
            trend_data.get('revenue', {}).get('b', []).append(monthly_rev[m]['b'])
            trend_data.get('revenue', {}).get('c', []).append(monthly_rev[m]['c'])
            trend_data.get('revenue', {}).get('d', []).append(monthly_rev[m]['d'])
            trend_data.get('revenue', {}).get('ecom', []).append(monthly_rev[m]['ecom'])
            trend_data.get('revenue', {}).get('total', []).append(sum(monthly_rev[m].values()))
            
            trend_data['profit']['b'].append(monthly_profit[m]['b'])
            trend_data['profit']['c'].append(monthly_profit[m]['c'])
            trend_data['profit']['d'].append(monthly_profit[m]['d'])
            trend_data['profit']['ecom'].append(monthly_profit[m]['ecom'])
            trend_data['profit']['total'].append(sum(monthly_profit[m].values()))
    
    # 读取箱量趋势
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*全链路*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        monthly_volume = defaultdict(lambda: {'b': 0, 'c': 0})
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            segment = ws.cell(row=row, column=11).value
            volume = ws.cell(row=row, column=17).value
            date_val = ws.cell(row=row, column=3).value
            
            if segment and volume and date_val:
                # 处理日期：支持字符串和 Excel 日期序列号
                from datetime import datetime, timedelta
                if isinstance(date_val, (int, float)):
                    # Excel 日期序列号
                    base_date = datetime(1899, 12, 30)
                    actual_date = base_date + timedelta(days=int(date_val))
                    month_str = f"{actual_date.year}-{actual_date.month}"
                elif isinstance(date_val, datetime):
                    month_str = f"{date_val.year}-{date_val.month}"
                else:
                    date_str = str(date_val)
                    if len(date_str) >= 7:
                        parts = date_str.split('-')
                        if len(parts) >= 2:
                            year = parts[0].strip()
                            month_num = int(parts[1].strip())
                            month_str = f"{year}-{month_num}"
                        else:
                            continue
                    else:
                        continue
                vol = float(volume) if volume else 0
                seg_str = str(segment).strip()
                
                if seg_str in ['A+B+C', 'A+B']:
                    monthly_volume[month_str]['b'] += vol
                elif seg_str in ['C+D']:
                    monthly_volume[month_str]['c'] += vol
                elif seg_str == 'B':
                    monthly_volume[month_str]['b'] += vol
                elif seg_str == 'C':
                    monthly_volume[month_str]['c'] += vol
        
        for m in trend_data['months']:
            trend_data.get('volume', {}).get('b', []).append(monthly_volume.get(m, {'b': 0})['b'])
            trend_data.get('volume', {}).get('c', []).append(monthly_volume.get(m, {'c': 0})['c'])
    
    # 读取 D 段箱量趋势（从客户仓库表，入库箱量）
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        monthly_d_volume = defaultdict(float)
        month_col = 4
        inbound_col = 8
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            month = ws.cell(row=row, column=month_col).value
            inbound = ws.cell(row=row, column=inbound_col).value
            
            if not month:
                continue
            
            # 处理 datetime 类型
            from datetime import datetime
            if isinstance(month, datetime):
                month_str = f"{month.year}-{month.month}"  # 不补零
            else:
                month_str = str(month)
                if '年' in month_str and '月' in month_str:
                    month_str = month_str.replace('年', '-').replace('月', '').strip()
                elif '-' in month_str:
                    pass
                else:
                    continue
                
                parts = month_str.split('-')
                if len(parts) == 2:
                    year = parts[0].strip()
                    month_num = int(parts[1].strip())
                    month_str = f"{year}-{month_num}"
                else:
                    continue
            
            vol = float(inbound) if inbound else 0
            monthly_d_volume[month_str] += vol
        
        for m in trend_data['months']:
            trend_data.get('volume', {}).get('d', []).append(monthly_d_volume.get(m, 0))
    
    # 读取出库件数趋势（从客户仓库表）
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if files:
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb.active
        
        monthly_outbound = defaultdict(float)
        month_col = 4
        outbound_col = 19
        
        for row in range(2, min(ws.max_row + 1, 50000)):
            month = ws.cell(row=row, column=month_col).value
            outbound = ws.cell(row=row, column=outbound_col).value
            
            if not month:
                continue
            
            # 处理月份：支持字符串、datetime 和 Excel 日期序列号
            from datetime import datetime, timedelta
            if isinstance(month, (int, float)):
                # Excel 日期序列号
                base_date = datetime(1899, 12, 30)
                actual_date = base_date + timedelta(days=int(month))
                month_str = f"{actual_date.year}-{actual_date.month}"
            elif isinstance(month, datetime):
                month_str = f"{month.year}-{month.month}"
            else:
                month_str = str(month)
                if '年' in month_str and '月' in month_str:
                    month_str = month_str.replace('年', '-').replace('月', '').strip()
                elif '-' in month_str:
                    pass
                else:
                    continue
                
                parts = month_str.split('-')
                if len(parts) == 2:
                    year = parts[0].strip()
                    month_num = int(parts[1].strip())
                    month_str = f"{year}-{month_num}"
                else:
                    continue
            
            pcs = float(outbound) if outbound else 0
            monthly_outbound[month_str] += pcs
        
        for m in trend_data['months']:
            trend_data.get('outbound', []).append(monthly_outbound.get(m, 0))
    
    return trend_data


def calculate_single_box_metrics(period='2026-Q1', months=6):
    """计算单箱成本和单箱毛利（美元/TEU），返回近 N 个月月度数据
    
    新规则（2026-04-01 更新）：
    - 数据源：DTC 明细表 - 全链路视角
    - 筛选条件：订单状态 = "内审通过"
    - 单箱成本 = 整票应付折美汇总 / 未去重集装箱箱量汇总
    - 单箱毛利 = 整票利润折美汇总 / 未去重集装箱箱量汇总
    
    返回：
    {
        'months': ['2025-10', '2025-11', ...],
        'b_cost': [...], 'b_profit': [...],
        'c_cost': [...], 'c_profit': [...],
        'd_cost': [...], 'd_profit': [...]
    }
    """
    current_month, cumulative_months = parse_period(period)
    target_year = '2026'
    
    result = {
        'current': {
            'b_cost_per_box': 0, 'b_profit_per_box': 0,
            'c_cost_per_box': 0, 'c_profit_per_box': 0,
            'd_cost_per_box': 0, 'd_profit_per_box': 0
        },
        'cumulative': {
            'b_cost_per_box': 0, 'b_profit_per_box': 0,
            'c_cost_per_box': 0, 'c_profit_per_box': 0,
            'd_cost_per_box': 0, 'd_profit_per_box': 0
        }
    }
    
    # 从全链路表读取数据
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*全链路*.xlsx'))
    if not files:
        return result
    
    f = files[0]
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        
        # 找表头
        headers = {}
        for col in range(1, 50):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[str(h).strip()] = col
        
        # 检查必要字段
        required_cols = ['订单服务项', '订单状态', '未去重集装箱箱量', '业务日期', '整票应付折美', '整票利润折美']
        for col_name in required_cols:
            if col_name not in headers:
                print(f"  ⚠️ 缺少字段：{col_name}")
                return result
        
        # 使用订单服务项来判断业务段（A/B/C/D）
        segment_col = headers['订单服务项']
        # 使用订单状态来筛选（内审通过）- 列 31
        status_col = headers['订单状态']
        # 整票应付折美（列 30）
        cost_col = headers['整票应付折美']
        # 整票利润折美（列 26）
        profit_col = headers['整票利润折美']
        # 未去重集装箱箱量（列 17）
        volume_col = headers['未去重集装箱箱量']
        # 业务日期（列 3）
        date_col = headers['业务日期']
        
        # 按业务段汇总
        # 按月汇总数据
        monthly_data = defaultdict(lambda: {
            'b': {'cost': 0, 'profit': 0, 'volume': 0},
            'c': {'cost': 0, 'profit': 0, 'volume': 0},
            'd': {'cost': 0, 'profit': 0, 'volume': 0}
        })
        
        for row in range(2, min(ws.max_row + 1, 100000)):
            status = ws.cell(row=row, column=status_col).value
            segment = ws.cell(row=row, column=segment_col).value
            date_val = ws.cell(row=row, column=date_col).value
            cost = ws.cell(row=row, column=cost_col).value
            profit = ws.cell(row=row, column=profit_col).value
            volume = ws.cell(row=row, column=volume_col).value
            
            # 筛选内审通过的订单
            if not status or str(status).strip() != '内审通过':
                continue
            
            # 从订单服务项判断业务段（A/B/C/D 段）
            seg_str = str(segment).strip() if segment else ''
            if 'A+B+C' in seg_str or 'A+B' in seg_str or seg_str.startswith('B'):
                seg_key = 'b'
            elif 'C+D' in seg_str or seg_str.startswith('C'):
                seg_key = 'c'
            elif seg_str.startswith('D'):
                seg_key = 'd'
            else:
                continue
            
            # 解析月份
            if not date_val:
                continue
            
            # 处理日期：支持 Excel 日期序列号、datetime、字符串
            if isinstance(date_val, (int, float)):
                # Excel 日期序列号
                base_date = datetime(1899, 12, 30)
                actual_date = base_date + timedelta(days=int(date_val))
                month_str = f"{actual_date.year}-{actual_date.month}"
            elif isinstance(date_val, datetime):
                month_str = f"{date_val.year}-{date_val.month}"
            else:
                # 字符串格式（如'2026-03-23'）
                date_str = str(date_val)
                if len(date_str) >= 7 and '-' in date_str:
                    parts = date_str.split('-')
                    if len(parts) >= 2:
                        year = parts[0].strip()
                        month = parts[1].strip().lstrip('0')  # 去掉前导零
                        month_str = f"{year}-{month}"
                    else:
                        continue
                else:
                    continue
            
            cost_val = float(cost) if cost else 0
            profit_val = float(profit) if profit else 0
            volume_val = float(volume) if volume else 0
            
            # 按月累加
            monthly_data[month_str][seg_key]['cost'] += cost_val
            monthly_data[month_str][seg_key]['profit'] += profit_val
            monthly_data[month_str][seg_key]['volume'] += volume_val
        
        # 计算单箱指标（在 return 前）
        
        # 构建近 N 个月数据（过滤掉 4 月及以后的数据）
        all_months_sorted = sorted(monthly_data.keys(), key=lambda m: (int(m.split('-')[0]), int(m.split('-')[1])))
        # 过滤掉 2026 年 4 月及以后的数据
        filtered_months = [m for m in all_months_sorted if not (m.startswith('2026-4') or m.startswith('2026-04'))]
        all_months = filtered_months[-months:] if len(filtered_months) > months else filtered_months
        
        monthly_result = {
            'months': all_months,
            'b_cost': [], 'b_profit': [],
            'c_cost': [], 'c_profit': []
        }
        
        for m in all_months:
            for seg_key in ['b', 'c']:  # 只计算 B 段和 C 段
                vol = monthly_data[m][seg_key]['volume']
                if vol > 0:
                    monthly_result[f'{seg_key}_cost'].append(monthly_data[m][seg_key]['cost'] / vol)
                    monthly_result[f'{seg_key}_profit'].append(monthly_data[m][seg_key]['profit'] / vol)
                else:
                    monthly_result[f'{seg_key}_cost'].append(0)
                    monthly_result[f'{seg_key}_profit'].append(0)
        
        print(f"  单箱指标计算完成（美元/TEU，近{months}个月）")
        if all_months:
            last_month = all_months[-1]
            print(f"    B 段单箱毛利（{last_month}）: {monthly_result['b_profit'][-1]:.2f} 美元/TEU")
            print(f"    C 段单箱毛利（{last_month}）: {monthly_result['c_profit'][-1]:.2f} 美元/TEU")
            # D 段不展示单箱指标
        
    except Exception as e:
        print(f"  计算单箱指标失败：{e}")
        import traceback
        traceback.print_exc()
        return {'months': [], 'b_cost': [], 'b_profit': [], 'c_cost': [], 'c_profit': [], 'd_cost': [], 'd_profit': []}
    
    return monthly_result


def generate_trend_analysis(trend_data):
    """生成趋势分析文字"""
    if not trend_data or not trend_data.get('months') or len(trend_data['months']) < 2:
        return "数据不足，无法生成趋势分析。"
    
    analysis = []
    
    # 收入趋势分析
    rev_total = trend_data.get('revenue', {}).get('total', [])
    if len(rev_total) >= 2:
        latest = rev_total[-1]
        previous = rev_total[-2]
        growth = ((latest - previous) / previous * 100) if previous else 0
        
        if growth > 10:
            analysis.append(f"收入呈上升趋势，最近一月环比增长{growth:.1f}%。")
        elif growth > 0:
            analysis.append(f"收入小幅增长，最近一月环比增长{growth:.1f}%。")
        elif growth > -10:
            analysis.append(f"收入小幅下滑，最近一月环比下降{abs(growth):.1f}%。")
        else:
            analysis.append(f"收入呈下降趋势，最近一月环比下降{abs(growth):.1f}%。")
    
    # 箱量趋势分析
    vol_b = trend_data.get('volume', {}).get('b', [])
    vol_c = trend_data.get('volume', {}).get('c', [])
    vol_d = trend_data.get('volume', {}).get('d', [])
    
    if vol_d and vol_d[-1] > 0:
        latest_d = vol_d[-1]
        avg_d = sum(vol_d) / len(vol_d)
        if latest_d > avg_d * 1.2:
            analysis.append(f"D 段箱量表现强劲，最近一月高于平均水平{((latest_d/avg_d-1)*100):.1f}%。")
        elif latest_d < avg_d * 0.8:
            analysis.append(f"D 段箱量有所下滑，最近一月低于平均水平{((1-latest_d/avg_d)*100):.1f}%。")
    
    if vol_c and vol_c[-1] > 0:
        latest_c = vol_c[-1]
        previous_c = vol_c[-2] if len(vol_c) >= 2 else vol_c[-1]
        growth_c = ((latest_c - previous_c) / previous_c * 100) if previous_c else 0
        if growth_c > 20:
            analysis.append(f"C 段箱量增长显著，最近一月环比增长{growth_c:.1f}%。")
    
    return " ".join(analysis) if analysis else "趋势平稳，无明显波动。"


def generate_html_report(actual, actual_yoy, budget, volume, volume_yoy, volume_mom, outbound, outbound_yoy, outbound_mom, single_box, trend_data, customer_data, top_customers, top_sales, top_sales_d_volume, top_current_customers, top_current_sales, warehouse_data, loss_customers, highlights, concerns, period='2026-Q1', ai_analysis=None, customer_vintage=None):
    """生成 HTML 报告（包含同比、单箱指标和趋势图表）"""
    current_month, cumulative_months = parse_period(period)
    
    def achievement(act, bud):
        return (act / bud * 100) if bud else 0
    
    def yoy_growth(current, last_year):
        if not last_year or last_year == 0:
            return 0
        return ((current - last_year) / last_year) * 100
    
    def fmt(val):
        # 数据已经是万元单位，直接格式化
        return f'{val:,.0f}'
    
    def fmt1(val):
        # 数据已经是万元单位，直接格式化（1 位小数）
        return f'{val:,.1f}'
    
    def fmt_num(val):
        return f'{val:,.0f}'
    
    today = datetime.now().strftime('%Y 年 %m 月 %d 日')
    
    # 处理 AI 分析结果
    if ai_analysis and ai_analysis.get('success'):
        ai_analysis_section = format_ai_analysis_html(ai_analysis)
    else:
        ai_analysis_section = ''
    
    # 处理客户年代分析结果
    if customer_vintage:
        customer_vintage_section = format_customer_vintage_html(customer_vintage)
    else:
        customer_vintage_section = ''
    
    # 计算汇总
    current_revenue = sum(actual['current'].get(f'{k}_revenue', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    current_profit = sum(actual['current'].get(f'{k}_profit', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    cumulative_revenue = sum(actual['cumulative'].get(f'{k}_revenue', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    cumulative_profit = sum(actual['cumulative'].get(f'{k}_profit', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    
    # 注意：文件名如"2025 年至 2026 年 3 月"表示完整月度数据，不再进行折算
    # 仅当文件名明确显示"XX 月 XX 日"且日期<28 时才考虑折算
    # 当前逻辑：不折算，直接使用原始数据
    
    # 同比汇总
    current_revenue_yoy = sum(actual_yoy['current'].get(f'{k}_revenue', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    current_profit_yoy = sum(actual_yoy['current'].get(f'{k}_profit', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    cumulative_revenue_yoy = sum(actual_yoy['cumulative'].get(f'{k}_revenue', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    cumulative_profit_yoy = sum(actual_yoy['cumulative'].get(f'{k}_profit', 0) for k in ['b', 'c', 'd', 'ecom', 'other'])
    
    # 生成趋势分析文字
    trend_analysis = generate_trend_analysis(trend_data if trend_data else {'months': [], 'revenue': {}, 'volume': {}})
    
    # 准备图表数据
    months_json = json.dumps(trend_data.get('months', []))
    months_12_json = json.dumps(warehouse_data.get('months_12', trend_data.get('months', [])))
    outbound_json = json.dumps(trend_data.get('outbound', []))
    # 单箱指标数据（近 6 个月：2025-10 至 2026-03）
    b_cost = single_box.get('b_cost', [0]*6)
    b_profit = single_box.get('b_profit', [0]*6)
    c_cost = single_box.get('c_cost', [0]*6)
    c_profit = single_box.get('c_profit', [0]*6)
    
    # 填充到 6 个月（如果数据不足）
    while len(b_cost) < 6:
        b_cost.insert(0, 0)
        b_profit.insert(0, 0)
        c_cost.insert(0, 0)
        c_profit.insert(0, 0)
    
    # 只取最近 6 个月
    b_cost = b_cost[-6:]
    b_profit = b_profit[-6:]
    c_cost = c_cost[-6:]
    c_profit = c_profit[-6:]
    
    single_box_b_cost_0 = b_cost[0]
    single_box_b_cost_1 = b_cost[1]
    single_box_b_cost_2 = b_cost[2]
    single_box_b_cost_3 = b_cost[3]
    single_box_b_cost_4 = b_cost[4]
    single_box_b_cost_5 = b_cost[5]
    single_box_b_profit_0 = b_profit[0]
    single_box_b_profit_1 = b_profit[1]
    single_box_b_profit_2 = b_profit[2]
    single_box_b_profit_3 = b_profit[3]
    single_box_b_profit_4 = b_profit[4]
    single_box_b_profit_5 = b_profit[5]
    single_box_c_cost_0 = c_cost[0]
    single_box_c_cost_1 = c_cost[1]
    single_box_c_cost_2 = c_cost[2]
    single_box_c_cost_3 = c_cost[3]
    single_box_c_cost_4 = c_cost[4]
    single_box_c_cost_5 = c_cost[5]
    single_box_c_profit_0 = c_profit[0]
    single_box_c_profit_1 = c_profit[1]
    single_box_c_profit_2 = c_profit[2]
    single_box_c_profit_3 = c_profit[3]
    single_box_c_profit_4 = c_profit[4]
    single_box_c_profit_5 = c_profit[5]
    single_box_months_json = json.dumps(single_box.get('months', []))
    single_box_b_profit_json = json.dumps(single_box.get('b_profit', []))
    single_box_c_profit_json = json.dumps(single_box.get('c_profit', []))
    revenue_b_json = json.dumps([x/10000 for x in trend_data.get('revenue', {}).get('b', [])])
    revenue_c_json = json.dumps([x/10000 for x in trend_data.get('revenue', {}).get('c', [])])
    revenue_d_json = json.dumps([x/10000 for x in trend_data.get('revenue', {}).get('d', [])])
    revenue_total_json = json.dumps([x/10000 for x in trend_data.get('revenue', {}).get('total', [])])
    volume_b_json = json.dumps(trend_data.get('volume', {}).get('b', []))
    volume_c_json = json.dumps(trend_data.get('volume', {}).get('c', []))
    volume_d_json = json.dumps(trend_data.get('volume', {}).get('d', []))
    
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>DTC {period} 经营分析报告</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: sans-serif; background: #f0f2f5; padding: 20px; }}
        .container {{ max-width: 1400px; margin: 0 auto; background: white; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); overflow: hidden; }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px 40px; text-align: center; }}
        .header h1 {{ font-size: 28px; margin-bottom: 20px; }}
        .header h1::before {{ content: "🐋 "; }}
        .meta {{ background: rgba(255,255,255,0.15); border-radius: 8px; padding: 12px 20px; font-size: 14px; }}
        .meta span {{ margin: 0 15px; }}
        .content {{ padding: 30px 40px; }}
        .section {{ margin-bottom: 35px; }}
        .section-title {{ font-size: 20px; font-weight: 600; color: #1a1a1a; border-left: 4px solid #667eea; padding-left: 15px; margin-bottom: 20px; }}
        .subsection {{ margin: 20px 0; }}
        .subsection-title {{ font-size: 16px; font-weight: 600; color: #374151; margin-bottom: 15px; padding-left: 10px; border-left: 3px solid #93c5fd; }}
        table {{ width: 100%; border-collapse: collapse; margin: 15px 0; font-size: 14px; }}
        th {{ background: #2563eb; color: white; padding: 12px 15px; text-align: left; }}
        td {{ padding: 12px 15px; border-bottom: 1px solid #e5e7eb; text-align: right; }}
        th:first-child, td:first-child {{ text-align: left; }}
        tr:nth-child(even) {{ background: #f9fafb; }}
        .note {{ background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 20px 0; color: #92400e; }}
        .footer {{ background: #f9fafb; border-top: 1px solid #e5e7eb; padding: 20px 40px; text-align: center; font-size: 13px; color: #6b7280; }}
        .text-success {{ color: #059669; }}
        .text-danger {{ color: #dc2626; }}
        .data-source {{ font-size: 12px; color: #6b7280; margin-top: 5px; }}
        .unit {{ font-size: 12px; color: #6b7280; float: right; }}
        .chart-container {{ position: relative; height: 350px; margin: 20px 0; }}
        .trend-analysis {{ background: #f0f9ff; border-left: 4px solid #0284c7; padding: 15px; margin: 20px 0; color: #0c4a6e; }}
        .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; margin: 20px 0; }}
        .kpi-card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 12px; padding: 25px; text-align: center; box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4); }}
        .kpi-card.green {{ background: linear-gradient(135deg, #10b981 0%, #059669 100%); box-shadow: 0 4px 15px rgba(16, 185, 129, 0.4); }}
        .kpi-card.purple {{ background: linear-gradient(135deg, #8b5cf6 0%, #7c3aed 100%); box-shadow: 0 4px 15px rgba(139, 92, 246, 0.4); }}
        .kpi-card .label {{ font-size: 14px; opacity: 0.9; margin-bottom: 10px; }}
        .kpi-card .value {{ font-size: 36px; font-weight: 700; margin-bottom: 12px; }}
        .kpi-card .detail {{ font-size: 13px; opacity: 0.85; line-height: 1.6; }}
        
        /* AI 分析样式 */
        .ai-analysis-content {{ background: #f8fafc; border-left: 4px solid #3b82f6; padding: 20px; margin: 15px 0; border-radius: 0 8px 8px 0; }}
        .analysis-item {{ background: white; padding: 15px; margin: 10px 0; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
        .analysis-item strong {{ color: #1e40af; font-size: 16px; }}
        .analysis-item ul {{ margin: 10px 0; padding-left: 20px; }}
        .analysis-item li {{ margin: 5px 0; color: #475569; }}
        .analysis-impact {{ background: #fef3c7; padding: 10px; border-radius: 6px; margin-top: 10px; color: #92400e; font-weight: 500; }}
        .text-warning {{ color: #f59e0b; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>DTC {period} 经营分析报告</h1>
            <div class="meta">
                <span>报告日期：<strong>{today}</strong></span>
                <span>|</span>
                <span>数据期间：<strong>{period}</strong></span>
                <span>|</span>
                <span>编制：<strong>财鲸 (FinWhale)</strong></span>
            </div>
        </div>
        
        <div class="content">
            <div class="section">
                <div class="section-title">一、经营概览</div>
                
                <div class="subsection">
                    <div class="subsection-title">1.1 核心指标</div>
                    
                    <!-- 可视化卡片 -->
                    <div class="kpi-grid">
                        <div class="kpi-card">
                            <div class="label">Q1 累计收入</div>
                            <div class="value">{fmt(cumulative_revenue)} 万元</div>
                            <div class="detail">同比 {yoy_growth(cumulative_revenue, cumulative_revenue_yoy):+.1f}% | 预算 {fmt(budget.get('revenue_cumulative', 0))}万元</div>
                        </div>
                        <div class="kpi-card green">
                            <div class="label">Q1 累计毛利</div>
                            <div class="value">{fmt(cumulative_profit)} 万元</div>
                            <div class="detail">同比 {yoy_growth(cumulative_profit, cumulative_profit_yoy):+.1f}% | 预算 {fmt(budget.get('gross_profit_cumulative', 0))}万元</div>
                        </div>
                        <div class="kpi-card purple">
                            <div class="label">3 月毛利</div>
                            <div class="value">{fmt(current_profit)} 万元</div>
                            <div class="detail">同比 {yoy_growth(current_profit, current_profit_yoy):+.1f}% | 毛利率 {(current_profit/current_revenue*100 if current_revenue != 0 else 0):.1f}%</div>
                        </div>
                    </div>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">1.2 业务量（箱量、出库件数）</div>
                    <table>
                        <thead>
                            <tr>
                                <th>业务段</th>
                                <th>当月实际</th>
                                <th>环比</th>
                                <th>同比</th>
                                <th>累计实际</th>
                                <th>同比</th>
                                <th>当月预算</th>
                                <th>累计预算</th>
                                <th>累计达成率</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段箱量</td>
                                <td>{fmt_num(volume['current']['b_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['b_volume'], volume_mom['current']['b_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['b_volume'], volume_mom['current']['b_volume']):+.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['b_volume'], volume_yoy['current']['b_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['b_volume'], volume_yoy['current']['b_volume']):+.1f}%</td>
                                <td>{fmt_num(volume['cumulative']['b_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['b_volume'], volume_yoy['cumulative']['b_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['b_volume'], volume_yoy['cumulative']['b_volume']):+.1f}%</td>
                                <td>{fmt_num(budget.get('b_volume', 0))}</td>
                                <td>{fmt_num(budget.get('b_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['b_volume'], budget.get('b_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['b_volume'], budget.get('b_volume_cumulative', 1)):.1f}%</td>
                            </tr>
                            <tr>
                                <td>C 段箱量</td>
                                <td>{fmt_num(volume['current']['c_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['c_volume'], volume_mom['current']['c_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['c_volume'], volume_mom['current']['c_volume']):+.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['c_volume'], volume_yoy['current']['c_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['c_volume'], volume_yoy['current']['c_volume']):+.1f}%</td>
                                <td>{fmt_num(volume['cumulative']['c_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['c_volume'], volume_yoy['cumulative']['c_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['c_volume'], volume_yoy['cumulative']['c_volume']):+.1f}%</td>
                                <td>{fmt_num(budget.get('c_volume', 0))}</td>
                                <td>{fmt_num(budget.get('c_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['c_volume'], budget.get('c_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['c_volume'], budget.get('c_volume_cumulative', 1)):.1f}%</td>
                            </tr>
                            <tr>
                                <td>D 段箱量*</td>
                                <td>{fmt_num(volume['current']['d_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['d_volume'], volume_mom['current']['d_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['d_volume'], volume_mom['current']['d_volume']):+.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['d_volume'], volume_yoy['current']['d_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['d_volume'], volume_yoy['current']['d_volume']):+.1f}%</td>
                                <td>{fmt_num(volume['cumulative']['d_volume'])}</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['d_volume'], volume_yoy['cumulative']['d_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['d_volume'], volume_yoy['cumulative']['d_volume']):+.1f}%</td>
                                <td>{fmt_num(budget.get('d_volume', 0))}</td>
                                <td>{fmt_num(budget.get('d_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['d_volume'], budget.get('d_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['d_volume'], budget.get('d_volume_cumulative', 1)):.1f}%</td>
                            </tr>
                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>出库件数**</td>
                                <td>{fmt_num(outbound['current'])}</td>
                                <td class="{'text-success' if yoy_growth(outbound['current'], outbound_mom['current']) > 0 else 'text-danger'}">{yoy_growth(outbound['current'], outbound_mom['current']):+.1f}%</td>
                                <td class="{'text-success' if yoy_growth(outbound['current'], outbound_yoy['current']) > 0 else 'text-danger'}">{yoy_growth(outbound['current'], outbound_yoy['current']):+.1f}%</td>
                                <td>{fmt_num(outbound['cumulative'])}</td>
                                <td class="{'text-success' if yoy_growth(outbound['cumulative'], outbound_yoy['cumulative']) > 0 else 'text-danger'}">{yoy_growth(outbound['cumulative'], outbound_yoy['cumulative']):+.1f}%</td>
                                <td>{fmt_num(budget.get('outbound_pieces', 0))}</td>
                                <td>{fmt_num(budget.get('outbound_pieces_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(outbound['cumulative'], budget.get('outbound_pieces_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(outbound['cumulative'], budget.get('outbound_pieces_cumulative', 1)):.1f}%</td>
                            </tr>
                        </tbody>
                    </table>
                    <div class="data-source">* D 段箱量为入库箱量（TEU）<br>** 出库件数<br>注：预算列为当月预算和累计预算，达成率为累计实际与累计预算对比</div>
                </div>
            </div>
            
            <div class="section">
            <div class="section">
                <div class="section-title">二、预实对比</div>
                
                <div class="subsection">
                    <div class="subsection-title">2.1 箱量预实对比（TEU）</div>
                    <table>
                        <thead>
                            <tr>
                                <th rowspan="2">业务段</th>
                                <th colspan="5">当月</th>
                                <th colspan="5">当年累计</th>
                            </tr>
                            <tr>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段</td>
                                <td>{fmt_num(volume['current']['b_volume'])}</td>
                                <td>{fmt_num(budget.get('b_volume', 0))}</td>
                                <td class="{'text-success' if achievement(volume['current']['b_volume'], budget.get('b_volume', 1)) >= 95 else 'text-danger'}">{achievement(volume['current']['b_volume'], budget.get('b_volume', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['b_volume'], volume_yoy['current']['b_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['b_volume'], volume_yoy['current']['b_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['current']['b_volume'] > volume_yoy['current']['b_volume'] else 'text-danger'}">{fmt_num(volume['current']['b_volume'] - volume_yoy['current']['b_volume'])}</td>
                                <td>{fmt_num(volume['cumulative']['b_volume'])}</td>
                                <td>{fmt_num(budget.get('b_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['b_volume'], budget.get('b_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['b_volume'], budget.get('b_volume_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['b_volume'], volume_yoy['cumulative']['b_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['b_volume'], volume_yoy['cumulative']['b_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['cumulative']['b_volume'] > volume_yoy['cumulative']['b_volume'] else 'text-danger'}">{fmt_num(volume['cumulative']['b_volume'] - volume_yoy['cumulative']['b_volume'])}</td>
                            </tr>
                            <tr>
                                <td>C 段</td>
                                <td>{fmt_num(volume['current']['c_volume'])}</td>
                                <td>{fmt_num(budget.get('c_volume', 0))}</td>
                                <td class="{'text-success' if achievement(volume['current']['c_volume'], budget.get('c_volume', 1)) >= 95 else 'text-danger'}">{achievement(volume['current']['c_volume'], budget.get('c_volume', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['c_volume'], volume_yoy['current']['c_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['c_volume'], volume_yoy['current']['c_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['current']['c_volume'] > volume_yoy['current']['c_volume'] else 'text-danger'}">{fmt_num(volume['current']['c_volume'] - volume_yoy['current']['c_volume'])}</td>
                                <td>{fmt_num(volume['cumulative']['c_volume'])}</td>
                                <td>{fmt_num(budget.get('c_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['c_volume'], budget.get('c_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['c_volume'], budget.get('c_volume_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['c_volume'], volume_yoy['cumulative']['c_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['c_volume'], volume_yoy['cumulative']['c_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['cumulative']['c_volume'] > volume_yoy['cumulative']['c_volume'] else 'text-danger'}">{fmt_num(volume['cumulative']['c_volume'] - volume_yoy['cumulative']['c_volume'])}</td>
                            </tr>
                            <tr>
                                <td>D 段</td>
                                <td>{fmt_num(volume['current']['d_volume'])}</td>
                                <td>{fmt_num(budget.get('d_volume', 0))}</td>
                                <td class="{'text-success' if achievement(volume['current']['d_volume'], budget.get('d_volume', 1)) >= 95 else 'text-danger'}">{achievement(volume['current']['d_volume'], budget.get('d_volume', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['current']['d_volume'], volume_yoy['current']['d_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['current']['d_volume'], volume_yoy['current']['d_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['current']['d_volume'] > volume_yoy['current']['d_volume'] else 'text-danger'}">{fmt_num(volume['current']['d_volume'] - volume_yoy['current']['d_volume'])}</td>
                                <td>{fmt_num(volume['cumulative']['d_volume'])}</td>
                                <td>{fmt_num(budget.get('d_volume_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(volume['cumulative']['d_volume'], budget.get('d_volume_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(volume['cumulative']['d_volume'], budget.get('d_volume_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(volume['cumulative']['d_volume'], volume_yoy['cumulative']['d_volume']) > 0 else 'text-danger'}">{yoy_growth(volume['cumulative']['d_volume'], volume_yoy['cumulative']['d_volume']):+.1f}%</td>
                                <td class="{'text-success' if volume['cumulative']['d_volume'] > volume_yoy['cumulative']['d_volume'] else 'text-danger'}">{fmt_num(volume['cumulative']['d_volume'] - volume_yoy['cumulative']['d_volume'])}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">2.2 出库件数预实对比（件）</div>
                    <table>
                        <thead>
                            <tr>
                                <th rowspan="2">指标</th>
                                <th colspan="5">当月</th>
                                <th colspan="5">当年累计</th>
                            </tr>
                            <tr>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>出库件数</td>
                                <td>{fmt_num(outbound['current'])}</td>
                                <td>{fmt_num(budget.get('outbound_pieces', 0))}</td>
                                <td class="{'text-success' if achievement(outbound['current'], budget.get('outbound_pieces', 1)) >= 95 else 'text-danger'}">{achievement(outbound['current'], budget.get('outbound_pieces', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(outbound['current'], outbound_yoy['current']) > 0 else 'text-danger'}">{yoy_growth(outbound['current'], outbound_yoy['current']):+.1f}%</td>
                                <td class="{'text-success' if outbound['current'] > outbound_yoy['current'] else 'text-danger'}">{fmt_num(outbound['current'] - outbound_yoy['current'])}</td>
                                <td>{fmt_num(outbound['cumulative'])}</td>
                                <td>{fmt_num(budget.get('outbound_pieces_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(outbound['cumulative'], budget.get('outbound_pieces_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(outbound['cumulative'], budget.get('outbound_pieces_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(outbound['cumulative'], outbound_yoy['cumulative']) > 0 else 'text-danger'}">{yoy_growth(outbound['cumulative'], outbound_yoy['cumulative']):+.1f}%</td>
                                <td class="{'text-success' if outbound['cumulative'] > outbound_yoy['cumulative'] else 'text-danger'}">{fmt_num(outbound['cumulative'] - outbound_yoy['cumulative'])}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">2.3 收入预实对比（万元）</div>
                    <table>
                        <thead>
                            <tr>
                                <th rowspan="2">业务段</th>
                                <th colspan="5">当月</th>
                                <th colspan="5">当年累计</th>
                            </tr>
                            <tr>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段</td>
                                <td>{fmt(actual['current']['b_revenue'])}</td>
                                <td>{fmt(budget.get('b_revenue', 0))}</td>
                                <td>{achievement(actual['current']['b_revenue'], budget.get('b_revenue', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['b_revenue'], actual_yoy['current']['b_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['b_revenue'], actual_yoy['current']['b_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['b_revenue'] > actual_yoy['current']['b_revenue'] else 'text-danger'}">{fmt(actual['current']['b_revenue'] - actual_yoy['current']['b_revenue'])}</td>
                                <td>{fmt(actual['cumulative']['b_revenue'])}</td>
                                <td>{fmt(budget.get('b_revenue_cumulative', 0))}</td>
                                <td>{achievement(actual['cumulative']['b_revenue'], budget.get('b_revenue_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['b_revenue'], actual_yoy['cumulative']['b_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['b_revenue'], actual_yoy['cumulative']['b_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['b_revenue'] > actual_yoy['cumulative']['b_revenue'] else 'text-danger'}">{fmt(actual['cumulative']['b_revenue'] - actual_yoy['cumulative']['b_revenue'])}</td>
                            </tr>
                            <tr>
                                <td>C 段</td>
                                <td>{fmt(actual['current']['c_revenue'])}</td>
                                <td>{fmt(budget.get('c_revenue', 0))}</td>
                                <td>{achievement(actual['current']['c_revenue'], budget.get('c_revenue', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['c_revenue'], actual_yoy['current']['c_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['c_revenue'], actual_yoy['current']['c_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['c_revenue'] > actual_yoy['current']['c_revenue'] else 'text-danger'}">{fmt(actual['current']['c_revenue'] - actual_yoy['current']['c_revenue'])}</td>
                                <td>{fmt(actual['cumulative']['c_revenue'])}</td>
                                <td>{fmt(budget.get('c_revenue_cumulative', 0))}</td>
                                <td>{achievement(actual['cumulative']['c_revenue'], budget.get('c_revenue_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['c_revenue'], actual_yoy['cumulative']['c_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['c_revenue'], actual_yoy['cumulative']['c_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['c_revenue'] > actual_yoy['cumulative']['c_revenue'] else 'text-danger'}">{fmt(actual['cumulative']['c_revenue'] - actual_yoy['cumulative']['c_revenue'])}</td>
                            </tr>
                            <tr>
                                <td>D 段</td>
                                <td>{fmt(actual['current']['d_revenue'])}</td>
                                <td>{fmt(budget.get('d_revenue', 0))}</td>
                                <td>{achievement(actual['current']['d_revenue'], budget.get('d_revenue', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['d_revenue'], actual_yoy['current']['d_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['d_revenue'], actual_yoy['current']['d_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['d_revenue'] > actual_yoy['current']['d_revenue'] else 'text-danger'}">{fmt(actual['current']['d_revenue'] - actual_yoy['current']['d_revenue'])}</td>
                                <td>{fmt(actual['cumulative']['d_revenue'])}</td>
                                <td>{fmt(budget.get('d_revenue_cumulative', 0))}</td>
                                <td>{achievement(actual['cumulative']['d_revenue'], budget.get('d_revenue_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['d_revenue'], actual_yoy['cumulative']['d_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['d_revenue'], actual_yoy['cumulative']['d_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['d_revenue'] > actual_yoy['cumulative']['d_revenue'] else 'text-danger'}">{fmt(actual['cumulative']['d_revenue'] - actual_yoy['cumulative']['d_revenue'])}</td>
                            </tr>
                            <tr>
                                <td>电商集拼</td>
                                <td>{fmt(actual['current']['ecom_revenue'])}</td>
                                <td>{fmt(budget.get('ecom_revenue', 0))}</td>
                                <td>{achievement(actual['current']['ecom_revenue'], budget.get('ecom_revenue', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['ecom_revenue'], actual_yoy['current']['ecom_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['ecom_revenue'], actual_yoy['current']['ecom_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['ecom_revenue'] > actual_yoy['current']['ecom_revenue'] else 'text-danger'}">{fmt(actual['current']['ecom_revenue'] - actual_yoy['current']['ecom_revenue'])}</td>
                                <td>{fmt(actual['cumulative']['ecom_revenue'])}</td>
                                <td>{fmt(budget.get('ecom_revenue_cumulative', 0))}</td>
                                <td>{achievement(actual['cumulative']['ecom_revenue'], budget.get('ecom_revenue_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['ecom_revenue'], actual_yoy['cumulative']['ecom_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['ecom_revenue'], actual_yoy['cumulative']['ecom_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['ecom_revenue'] > actual_yoy['cumulative']['ecom_revenue'] else 'text-danger'}">{fmt(actual['cumulative']['ecom_revenue'] - actual_yoy['cumulative']['ecom_revenue'])}</td>
                            </tr>
                            <tr>
                                <td>其他业务</td>
                                <td>{fmt(actual['current']['other_revenue'])}</td>
                                <td>{fmt(budget.get('other_revenue', 0))}</td>
                                <td>{achievement(actual['current']['other_revenue'], budget.get('other_revenue', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['other_revenue'], actual_yoy['current']['other_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['other_revenue'], actual_yoy['current']['other_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['other_revenue'] > actual_yoy['current']['other_revenue'] else 'text-danger'}">{fmt(actual['current']['other_revenue'] - actual_yoy['current']['other_revenue'])}</td>
                                <td>{fmt(actual['cumulative']['other_revenue'])}</td>
                                <td>{fmt(budget.get('other_revenue_cumulative', 0))}</td>
                                <td>{achievement(actual['cumulative']['other_revenue'], budget.get('other_revenue_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['other_revenue'], actual_yoy['cumulative']['other_revenue']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['other_revenue'], actual_yoy['cumulative']['other_revenue']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['other_revenue'] > actual_yoy['cumulative']['other_revenue'] else 'text-danger'}">{fmt(actual['cumulative']['other_revenue'] - actual_yoy['cumulative']['other_revenue'])}</td>
                            </tr>
                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>合计</td>
                                <td>{fmt(current_revenue)}</td>
                                <td>{fmt(budget.get('b_revenue', 0) + budget.get('c_revenue', 0) + budget.get('d_revenue', 0) + budget.get('ecom_revenue', 0) + budget.get('other_revenue', 0))}</td>
                                <td>{achievement(current_revenue, budget.get('b_revenue', 1) + budget.get('c_revenue', 1) + budget.get('d_revenue', 1) + budget.get('ecom_revenue', 1) + budget.get('other_revenue', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(current_revenue, current_revenue_yoy) > 0 else 'text-danger'}">{yoy_growth(current_revenue, current_revenue_yoy):+.1f}%</td>
                                <td class="{'text-success' if current_revenue > current_revenue_yoy else 'text-danger'}">{fmt(current_revenue - current_revenue_yoy)}</td>
                                <td>{fmt(cumulative_revenue)}</td>
                                <td>{fmt(budget.get('revenue_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(cumulative_revenue, budget.get('revenue_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(cumulative_revenue, budget.get('revenue_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(cumulative_revenue, cumulative_revenue_yoy) > 0 else 'text-danger'}">{yoy_growth(cumulative_revenue, cumulative_revenue_yoy):+.1f}%</td>
                                <td class="{'text-success' if cumulative_revenue > cumulative_revenue_yoy else 'text-danger'}">{fmt(cumulative_revenue - cumulative_revenue_yoy)}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">2.4 毛利预实对比（万元）</div>
                    <table>
                        <thead>
                            <tr>
                                <th rowspan="2">业务段</th>
                                <th colspan="5">当月</th>
                                <th colspan="5">当年累计</th>
                            </tr>
                            <tr>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                                <th>实际</th>
                                <th>预算</th>
                                <th>达成率</th>
                                <th>同比</th>
                                <th>增长</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段</td>
                                <td>{fmt(actual['current']['b_profit'])}</td>
                                <td>{fmt(budget.get('b_profit', 0))}</td>
                                <td>{achievement(actual['current']['b_profit'], budget.get('b_profit', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['b_profit'], actual_yoy['current']['b_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['b_profit'], actual_yoy['current']['b_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['b_profit'] > actual_yoy['current']['b_profit'] else 'text-danger'}">{fmt(actual['current']['b_profit'] - actual_yoy['current']['b_profit'])}</td>
                                <td>{fmt(actual['cumulative']['b_profit'])}</td>
                                <td>{fmt(budget.get('b_profit_cumulative', 0))}</td>
                                <td>{achievement(actual['cumulative']['b_profit'], budget.get('b_profit_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['b_profit'], actual_yoy['cumulative']['b_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['b_profit'], actual_yoy['cumulative']['b_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['b_profit'] > actual_yoy['cumulative']['b_profit'] else 'text-danger'}">{fmt(actual['cumulative']['b_profit'] - actual_yoy['cumulative']['b_profit'])}</td>
                            </tr>
                            <tr>
                                <td>C 段</td>
                                <td>{fmt(actual['current']['c_profit'])}</td>
                                <td>{fmt(budget.get('c_profit', 0))}</td>
                                <td>{achievement(actual['current']['c_profit'], budget.get('c_profit', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['c_profit'], actual_yoy['current']['c_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['c_profit'], actual_yoy['current']['c_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['c_profit'] > actual_yoy['current']['c_profit'] else 'text-danger'}">{fmt(actual['current']['c_profit'] - actual_yoy['current']['c_profit'])}</td>
                                <td>{fmt(actual['cumulative']['c_profit'])}</td>
                                <td>{fmt(budget.get('c_profit_cumulative', 0))}</td>
                                <td>{achievement(actual['cumulative']['c_profit'], budget.get('c_profit_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['c_profit'], actual_yoy['cumulative']['c_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['c_profit'], actual_yoy['cumulative']['c_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['c_profit'] > actual_yoy['cumulative']['c_profit'] else 'text-danger'}">{fmt(actual['cumulative']['c_profit'] - actual_yoy['cumulative']['c_profit'])}</td>
                            </tr>
                            <tr>
                                <td>D 段</td>
                                <td>{fmt(actual['current']['d_profit'])}</td>
                                <td>{fmt(budget.get('d_profit', 0))}</td>
                                <td>{achievement(actual['current']['d_profit'], budget.get('d_profit', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['d_profit'], actual_yoy['current']['d_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['d_profit'], actual_yoy['current']['d_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['d_profit'] > actual_yoy['current']['d_profit'] else 'text-danger'}">{fmt(actual['current']['d_profit'] - actual_yoy['current']['d_profit'])}</td>
                                <td>{fmt(actual['cumulative']['d_profit'])}</td>
                                <td>{fmt(budget.get('d_profit_cumulative', 0))}</td>
                                <td>{achievement(actual['cumulative']['d_profit'], budget.get('d_profit_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['d_profit'], actual_yoy['cumulative']['d_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['d_profit'], actual_yoy['cumulative']['d_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['d_profit'] > actual_yoy['cumulative']['d_profit'] else 'text-danger'}">{fmt(actual['cumulative']['d_profit'] - actual_yoy['cumulative']['d_profit'])}</td>
                            </tr>
                            <tr>
                                <td>电商集拼</td>
                                <td>{fmt(actual['current']['ecom_profit'])}</td>
                                <td>{fmt(budget.get('ecom_profit', 0))}</td>
                                <td>{achievement(actual['current']['ecom_profit'], budget.get('ecom_profit', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['ecom_profit'], actual_yoy['current']['ecom_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['ecom_profit'], actual_yoy['current']['ecom_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['ecom_profit'] > actual_yoy['current']['ecom_profit'] else 'text-danger'}">{fmt(actual['current']['ecom_profit'] - actual_yoy['current']['ecom_profit'])}</td>
                                <td>{fmt(actual['cumulative']['ecom_profit'])}</td>
                                <td>{fmt(budget.get('ecom_profit_cumulative', 0))}</td>
                                <td>{achievement(actual['cumulative']['ecom_profit'], budget.get('ecom_profit_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['ecom_profit'], actual_yoy['cumulative']['ecom_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['ecom_profit'], actual_yoy['cumulative']['ecom_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['ecom_profit'] > actual_yoy['cumulative']['ecom_profit'] else 'text-danger'}">{fmt(actual['cumulative']['ecom_profit'] - actual_yoy['cumulative']['ecom_profit'])}</td>
                            </tr>
                            <tr>
                                <td>其他</td>
                                <td>{fmt(actual['current']['other_profit'])}</td>
                                <td>{fmt(budget.get('other_profit', 0))}</td>
                                <td>{achievement(actual['current']['other_profit'], budget.get('other_profit', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['current']['other_profit'], actual_yoy['current']['other_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['current']['other_profit'], actual_yoy['current']['other_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['current']['other_profit'] > actual_yoy['current']['other_profit'] else 'text-danger'}">{fmt(actual['current']['other_profit'] - actual_yoy['current']['other_profit'])}</td>
                                <td>{fmt(actual['cumulative']['other_profit'])}</td>
                                <td>{fmt(budget.get('other_profit_cumulative', 0))}</td>
                                <td>{achievement(actual['cumulative']['other_profit'], budget.get('other_profit_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(actual['cumulative']['other_profit'], actual_yoy['cumulative']['other_profit']) > 0 else 'text-danger'}">{yoy_growth(actual['cumulative']['other_profit'], actual_yoy['cumulative']['other_profit']):+.1f}%</td>
                                <td class="{'text-success' if actual['cumulative']['other_profit'] > actual_yoy['cumulative']['other_profit'] else 'text-danger'}">{fmt(actual['cumulative']['other_profit'] - actual_yoy['cumulative']['other_profit'])}</td>
                            </tr>
                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>合计</td>
                                <td>{fmt(current_profit)}</td>
                                <td>{fmt(budget.get('b_profit', 0) + budget.get('c_profit', 0) + budget.get('d_profit', 0) + budget.get('ecom_profit', 0) + budget.get('other_profit', 0))}</td>
                                <td>{achievement(current_profit, budget.get('b_profit', 1) + budget.get('c_profit', 1) + budget.get('d_profit', 1) + budget.get('ecom_profit', 1) + budget.get('other_profit', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(current_profit, current_profit_yoy) > 0 else 'text-danger'}">{yoy_growth(current_profit, current_profit_yoy):+.1f}%</td>
                                <td class="{'text-success' if current_profit > current_profit_yoy else 'text-danger'}">{fmt(current_profit - current_profit_yoy)}</td>
                                <td>{fmt(cumulative_profit)}</td>
                                <td>{fmt(budget.get('gross_profit_cumulative', 0))}</td>
                                <td class="{'text-success' if achievement(cumulative_profit, budget.get('gross_profit_cumulative', 1)) >= 95 else 'text-danger'}">{achievement(cumulative_profit, budget.get('gross_profit_cumulative', 1)):.1f}%</td>
                                <td class="{'text-success' if yoy_growth(cumulative_profit, cumulative_profit_yoy) > 0 else 'text-danger'}">{yoy_growth(cumulative_profit, cumulative_profit_yoy):+.1f}%</td>
                                <td class="{'text-success' if cumulative_profit > cumulative_profit_yoy else 'text-danger'}">{fmt(cumulative_profit - cumulative_profit_yoy)}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">2.5 单箱指标（美元/TEU，近 6 个月）</div>
                    <div class="data-source" style="margin-bottom: 10px;">注：单箱成本/毛利 = 整票应付折美 (或整票利润折美) / 未去重集装箱箱量，仅统计"内审通过"订单；数据来源：DTC 明细表 - 全链路视角</div>
                    <table>
                        <thead>
                            <tr>
                                <th>指标</th>
                                <th>2025 年 10 月</th>
                                <th>2025 年 11 月</th>
                                <th>2025 年 12 月</th>
                                <th>2026 年 1 月</th>
                                <th>2026 年 2 月</th>
                                <th>2026 年 3 月</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段单箱成本</td>
                                <td>{single_box_b_cost_0:,.2f}</td>
                                <td>{single_box_b_cost_1:,.2f}</td>
                                <td>{single_box_b_cost_2:,.2f}</td>
                                <td>{single_box_b_cost_3:,.2f}</td>
                                <td>{single_box_b_cost_4:,.2f}</td>
                                <td>{single_box_b_cost_5:,.2f}</td>
                            </tr>
                            <tr>
                                <td>B 段单箱毛利</td>
                                <td>{single_box_b_profit_0:,.2f}</td>
                                <td>{single_box_b_profit_1:,.2f}</td>
                                <td>{single_box_b_profit_2:,.2f}</td>
                                <td>{single_box_b_profit_3:,.2f}</td>
                                <td>{single_box_b_profit_4:,.2f}</td>
                                <td>{single_box_b_profit_5:,.2f}</td>
                            </tr>
                            <tr>
                                <td>C 段单箱成本</td>
                                <td>{single_box_c_cost_0:,.2f}</td>
                                <td>{single_box_c_cost_1:,.2f}</td>
                                <td>{single_box_c_cost_2:,.2f}</td>
                                <td>{single_box_c_cost_3:,.2f}</td>
                                <td>{single_box_c_cost_4:,.2f}</td>
                                <td>{single_box_c_cost_5:,.2f}</td>
                            </tr>
                            <tr>
                                <td>C 段单箱毛利</td>
                                <td>{single_box_c_profit_0:,.2f}</td>
                                <td>{single_box_c_profit_1:,.2f}</td>
                                <td>{single_box_c_profit_2:,.2f}</td>
                                <td>{single_box_c_profit_3:,.2f}</td>
                                <td>{single_box_c_profit_4:,.2f}</td>
                                <td>{single_box_c_profit_5:,.2f}</td>
                            </tr>
                        </tbody>
                    </table>
                    <div class="data-source">数据来源：DTC 明细表 - 全链路视角</div>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">三、趋势分析</div>
                
                <div class="subsection">
                    <div class="subsection-title">收入月度趋势（分业务段）</div>
                    <div class="chart-container">
                        <canvas id="revenueChart"></canvas>
                    </div>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">箱量月度趋势（分业务段）</div>
                    <div class="chart-container">
                        <canvas id="volumeChart"></canvas>
                    </div>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">出库件数月度趋势</div>
                    <div class="chart-container">
                        <canvas id="outboundChart"></canvas>
                    </div>
                    <div class="data-source">注：出库件数为实际数据，未折算</div>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">四、客户与销售</div>
                
                {customer_vintage_section}
                
                <div class="subsection">
                    <div class="subsection-title">4.3 客户开发数</div>
                    <table>
                        <thead>
                            <tr>
                                <th>业务段</th>
                                <th>当月新增</th>
                                <th>当年累计</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段客户</td>
                                <td>{len(customer_data['current']['b_customers'])}</td>
                                <td>{len(customer_data['cumulative']['b_customers'])}</td>
                            </tr>
                            <tr>
                                <td>C 段客户</td>
                                <td>{len(customer_data['current']['c_customers'])}</td>
                                <td>{len(customer_data['cumulative']['c_customers'])}</td>
                            </tr>
                            <tr>
                                <td>D 段客户</td>
                                <td>{len(customer_data['current']['d_customers'])}</td>
                                <td>{len(customer_data['cumulative']['d_customers'])}</td>
                            </tr>
                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>合计（去重）</td>
                                <td>{len(customer_data['current']['total'])}</td>
                                <td>{len(customer_data['cumulative']['total'])}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">4.4 前十大客户贡献</div>
                    <table>
                        <thead>
                            <tr>
                                <th>排名</th>
                                <th>客户名称</th>
                                <th>累计收入（万元）</th>
                                <th>累计毛利率</th>
                                <th>当月收入（万元）</th>
                                <th>当月毛利率</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    # 创建客户数据字典用于合并
    cumulative_dict = {name: {'revenue': rev, 'profit': prof} for name, rev, prof in top_customers}
    current_dict = {name: {'revenue': rev, 'profit': prof} for name, rev, prof in top_current_customers}
    
    # 按累计收入排序展示
    for i, (name, rev, prof) in enumerate(top_customers, 1):
        cumulative_gross_margin = (prof / rev * 100) if rev > 0 else 0
        current_data = current_dict.get(name, {'revenue': 0, 'profit': 0})
        current_rev = current_data['revenue']
        current_prof = current_data['profit']
        current_gross_margin = (current_prof / current_rev * 100) if current_rev > 0 else 0
        
        html += f'''                            <tr>
                                <td>{i}</td>
                                <td>{name}</td>
                                <td>{fmt(rev)}</td>
                                <td>{cumulative_gross_margin:.2f}%</td>
                                <td>{fmt(current_rev)}</td>
                                <td>{current_gross_margin:.2f}%</td>
                            </tr>
'''
    
    html += f'''                        </tbody>
                    </table>
                    <div class="data-source">注：按累计收入排序；毛利率=毛利/收入×100%</div>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">4.5 前十大销售贡献</div>
                    <table>
                        <thead>
                            <tr>
                                <th>排名</th>
                                <th>销售</th>
                                <th>收入（万元）</th>
                                <th>毛利（万元）</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    for i, (name, rev, prof) in enumerate(top_sales, 1):
        html += f'''                            <tr>
                                <td>{i}</td>
                                <td>{name}</td>
                                <td>{fmt(rev)}</td>
                                <td>{fmt(prof)}</td>
                            </tr>
'''
    
    html += f'''                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">4.6 前十大销售 D 段箱量排名</div>
                    <table>
                        <thead>
                            <tr>
                                <th>排名</th>
                                <th>销售</th>
                                <th>D 段箱量（TEU）</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    for i, (name, vol) in enumerate(top_sales_d_volume, 1):
        html += f'''                            <tr>
                                <td>{i}</td>
                                <td>{name}</td>
                                <td>{fmt_num(vol)}</td>
                            </tr>
'''
    
    html += f'''                        </tbody>
                    </table>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">五、海外仓分析</div>
                
                <div class="subsection">
                    <div class="subsection-title">5.1 入库柜量和出库件数分布（按国家）</div>
                    <table>
                        <thead>
                            <tr>
                                <th>国家/地区</th>
                                <th>入库柜量（条）</th>
                                <th>占比</th>
                                <th>出库件数（件）</th>
                                <th>占比</th>
                                <th>自提比例</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    total_containers = sum(d['inbound_containers'] for d in warehouse_data['by_country'].values())
    total_pieces = sum(d['outbound_pieces'] for d in warehouse_data['by_country'].values())
    total_self_pickup = sum(d['self_pickup_pieces'] for d in warehouse_data['by_country'].values())
    total_self_pickup_ratio = (total_self_pickup / total_pieces * 100) if total_pieces > 0 else 0
    
    for country, data in sorted(warehouse_data['by_country'].items(), key=lambda x: x[1]['inbound_containers'], reverse=True):
        container_pct = (data['inbound_containers'] / total_containers * 100) if total_containers else 0
        pieces_pct = (data['outbound_pieces'] / total_pieces * 100) if total_pieces else 0
        self_pickup_ratio = data.get('self_pickup_ratio', 0)
        html += f'''                            <tr>
                                <td>{country}</td>
                                <td>{fmt_num(data['inbound_containers'])}</td>
                                <td>{container_pct:.1f}%</td>
                                <td>{fmt_num(data['outbound_pieces'])}</td>
                                <td>{pieces_pct:.1f}%</td>
                                <td>{self_pickup_ratio:.1f}%</td>
                            </tr>
'''
    
    html += f'''                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>合计</td>
                                <td>{fmt_num(total_containers)}</td>
                                <td>100%</td>
                                <td>{fmt_num(total_pieces)}</td>
                                <td>100%</td>
                                <td>{total_self_pickup_ratio:.1f}%</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">5.2 美国仓库分布（自营 vs 第三方）</div>
                    <table>
                        <thead>
                            <tr>
                                <th>仓库类型</th>
                                <th>入库柜量（条）</th>
                                <th>占比</th>
                                <th>出库件数（件）</th>
                                <th>占比</th>
                                <th>自提比例</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr style="background: #e8f4fd; font-weight: 600;">
                                <td>自营仓</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['self_run']['inbound_containers'])}</td>
                                <td>{(warehouse_data['us_warehouse']['self_run']['inbound_containers']/(warehouse_data['us_warehouse']['self_run']['inbound_containers']+warehouse_data['us_warehouse']['third_party']['inbound_containers'])*100) if (warehouse_data['us_warehouse']['self_run']['inbound_containers']+warehouse_data['us_warehouse']['third_party']['inbound_containers']) else 0:.1f}%</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['self_run']['outbound_pieces'])}</td>
                                <td>{(warehouse_data['us_warehouse']['self_run']['outbound_pieces']/(warehouse_data['us_warehouse']['self_run']['outbound_pieces']+warehouse_data['us_warehouse']['third_party']['outbound_pieces'])*100) if (warehouse_data['us_warehouse']['self_run']['outbound_pieces']+warehouse_data['us_warehouse']['third_party']['outbound_pieces']) else 0:.1f}%</td>
                                <td>{warehouse_data['us_warehouse']['self_run'].get('self_pickup_ratio', 0):.1f}%</td>
                            </tr>
                            <tr>
                                <td style="padding-left: 30px;">其中：美东自营仓</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['self_run_by_region']['美东']['inbound_containers'])}</td>
                                <td>{(warehouse_data['us_warehouse']['self_run_by_region']['美东']['inbound_containers']/warehouse_data['us_warehouse']['self_run']['inbound_containers']*100) if warehouse_data['us_warehouse']['self_run']['inbound_containers'] else 0:.1f}%</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['self_run_by_region']['美东']['outbound_pieces'])}</td>
                                <td>{(warehouse_data['us_warehouse']['self_run_by_region']['美东']['outbound_pieces']/warehouse_data['us_warehouse']['self_run']['outbound_pieces']*100) if warehouse_data['us_warehouse']['self_run']['outbound_pieces'] else 0:.1f}%</td>
                                <td>{warehouse_data['us_warehouse']['self_run_by_region']['美东'].get('self_pickup_ratio', 0):.1f}%</td>
                            </tr>
                            <tr>
                                <td style="padding-left: 30px;">美西自营仓</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['self_run_by_region']['美西']['inbound_containers'])}</td>
                                <td>{(warehouse_data['us_warehouse']['self_run_by_region']['美西']['inbound_containers']/warehouse_data['us_warehouse']['self_run']['inbound_containers']*100) if warehouse_data['us_warehouse']['self_run']['inbound_containers'] else 0:.1f}%</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['self_run_by_region']['美西']['outbound_pieces'])}</td>
                                <td>{(warehouse_data['us_warehouse']['self_run_by_region']['美西']['outbound_pieces']/warehouse_data['us_warehouse']['self_run']['outbound_pieces']*100) if warehouse_data['us_warehouse']['self_run']['outbound_pieces'] else 0:.1f}%</td>
                                <td>{warehouse_data['us_warehouse']['self_run_by_region']['美西'].get('self_pickup_ratio', 0):.1f}%</td>
                            </tr>
                            <tr style="background: #f3f4f6; font-weight: 600;">
                                <td>第三方仓</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['third_party']['inbound_containers'])}</td>
                                <td>{(warehouse_data['us_warehouse']['third_party']['inbound_containers']/(warehouse_data['us_warehouse']['self_run']['inbound_containers']+warehouse_data['us_warehouse']['third_party']['inbound_containers'])*100) if (warehouse_data['us_warehouse']['self_run']['inbound_containers']+warehouse_data['us_warehouse']['third_party']['inbound_containers']) else 0:.1f}%</td>
                                <td>{fmt_num(warehouse_data['us_warehouse']['third_party']['outbound_pieces'])}</td>
                                <td>{(warehouse_data['us_warehouse']['third_party']['outbound_pieces']/(warehouse_data['us_warehouse']['self_run']['outbound_pieces']+warehouse_data['us_warehouse']['third_party']['outbound_pieces'])*100) if (warehouse_data['us_warehouse']['self_run']['outbound_pieces']+warehouse_data['us_warehouse']['third_party']['outbound_pieces']) else 0:.1f}%</td>
                                <td>{warehouse_data['us_warehouse']['third_party'].get('self_pickup_ratio', 0):.1f}%</td>
                            </tr>
'''
    for region in ['美东', '美西', '美南', '美国中西部']:
        region_data = warehouse_data['us_warehouse']['third_party_by_region'][region]
        if region_data['inbound_containers'] > 0 or region_data['outbound_pieces'] > 0:
            html += f'''                            <tr>
                                <td style="padding-left: 30px;">{region}第三方仓</td>
                                <td>{fmt_num(region_data['inbound_containers'])}</td>
                                <td>{(region_data['inbound_containers']/warehouse_data['us_warehouse']['third_party']['inbound_containers']*100) if warehouse_data['us_warehouse']['third_party']['inbound_containers'] else 0:.1f}%</td>
                                <td>{fmt_num(region_data['outbound_pieces'])}</td>
                                <td>{(region_data['outbound_pieces']/warehouse_data['us_warehouse']['third_party']['outbound_pieces']*100) if warehouse_data['us_warehouse']['third_party']['outbound_pieces'] else 0:.1f}%</td>
                                <td>{region_data.get('self_pickup_ratio', 0):.1f}%</td>
                            </tr>
'''
    
    html += '''                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">5.3 自营仓趋势（美东 vs 美西）</div>
                    <div class="chart-container">
                        <canvas id="selfRunChart"></canvas>
                    </div>
                    <div class="data-source">数据来源：DTC 业务费用根据客户仓库统计表</div>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">5.4 Top10 亏损客户</div>
                    <table>
                        <thead>
                            <tr>
                                <th>排名</th>
                                <th>客户名称</th>
                                <th>毛利（万元）</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    for i, (name, profit) in enumerate(loss_customers, 1):
        html += f'''                            <tr>
                                <td>{i}</td>
                                <td>{name}</td>
                                <td class="text-danger">{fmt(profit)}</td>
                            </tr>
'''
    
    if not loss_customers:
        html += '''                            <tr>
                                <td colspan="3" style="text-align: center;">无亏损客户</td>
                            </tr>
'''
    
    html += f'''                        </tbody>
                    </table>
                </div>
            </div>
            
            <!-- AI 深度分析 -->
            {ai_analysis_section}
            
            <div class="section">
                <div class="section-title">六、关键发现</div>
                
                <div class="subsection">
                    <div class="subsection-title">6.1 亮点 ✨</div>
                    <ul style="line-height: 2;">
'''
    
    for highlight in highlights:
        html += f'                        <li>{highlight}</li>\n'
    
    html += f'''                    </ul>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">6.2 关注点 ⚠️</div>
                    <ul style="line-height: 2;">
'''
    
    for concern in concerns:
        html += f'                        <li>{concern}</li>\n'
    
    html += f'''                    </ul>
                </div>
            </div>
            
            
            
            <div class="section">
                <div class="section-title">附录：数据来源 + 统计口径 + 折算规则</div>
                
                <div class="subsection">
                    <div class="subsection-title">一、数据来源</div>
                    <table>
                        <thead>
                            <tr>
                                <th>数据项</th>
                                <th>数据来源文件</th>
                                <th>关键字段</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>收入/毛利</td>
                                <td>所有业务收入明细（非订单维度）</td>
                                <td>业务系统单元编码、业务分段分类_新.、收入、毛利</td>
                            </tr>
                            <tr>
                                <td>B 段/C 段箱量</td>
                                <td>DTC 明细表 - 全链路视角</td>
                                <td>订单服务项、未去重集装箱箱量</td>
                            </tr>
                            <tr>
                                <td>D 段箱量</td>
                                <td>DTC 业务费用根据客户仓库统计表</td>
                                <td>业务月份、入库箱量</td>
                            </tr>
                            <tr>
                                <td>出库件数</td>
                                <td>DTC 业务费用根据客户仓库统计表</td>
                                <td>业务月份、出库总件数</td>
                            </tr>
                            <tr>
                                <td>入库柜量</td>
                                <td>DTC 业务费用根据客户仓库统计表</td>
                                <td>仓库位置、入库箱量</td>
                            </tr>
                            <tr>
                                <td>客户开发数</td>
                                <td>DTC 历史客户合作情况汇总</td>
                                <td>客户名称、业务分段列表、第一次合作年月</td>
                            </tr>
                            <tr>
                                <td>预算数据</td>
                                <td>2026 年 DTC 预算数据</td>
                                <td>各业务段箱量合计行、出库件数合计行</td>
                            </tr>
                            <tr>
                                <td>亏损客户</td>
                                <td>所有业务收入明细（非订单维度）</td>
                                <td>委托客户名称、毛利</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">二、统计口径</div>
                    <table>
                        <thead>
                            <tr>
                                <th>数据项</th>
                                <th>统计口径说明</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>B 段箱量</td>
                                <td>订单服务项为 A+B、B、A+B+C 的未去重集装箱箱量（TEU）</td>
                            </tr>
                            <tr>
                                <td>C 段箱量</td>
                                <td>订单服务项为 C、C+D 的未去重集装箱箱量（TEU）</td>
                            </tr>
                            <tr>
                                <td>D 段箱量</td>
                                <td>DTC 业务费用根据客户仓库统计表中的入库箱量（TEU）</td>
                            </tr>
                            <tr>
                                <td>出库件数</td>
                                <td>DTC 业务费用根据客户仓库统计表中的出库总件数（件）</td>
                            </tr>
                            <tr>
                                <td>入库柜量</td>
                                <td>入库箱量÷2（1 条柜=2TEU）</td>
                            </tr>
                            <tr>
                                <td>客户开发数</td>
                                <td>第一次合作年月在统计期间内的客户数，一个客户同时是 B/C/D 段的，总开发客户合计数去重</td>
                            </tr>
                            <tr>
                                <td>前十大客户</td>
                                <td>按收入排名，排除 35 家环世集团内部公司</td>
                            </tr>
                            <tr>
                                <td>自营仓</td>
                                <td>仓库代码为 USCAEA02（美西）、USNJHM01（美东）</td>
                            </tr>
                            <tr>
                                <td>第三方仓</td>
                                <td>仓库代码非 USCAEA02、USNJHM01 的美国仓库</td>
                            </tr>
                            <tr>
                                <td>亏损客户</td>
                                <td>毛利小于 0 的客户</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">三、折算规则</div>
                    <table>
                        <thead>
                            <tr>
                                <th>数据项</th>
                                <th>折算规则</th>
                                <th>示例</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>D 段箱量</td>
                                <td>当月未结束时，按实际天数/当月总天数折算</td>
                                <td>3 月数据截止到 30 日：实际箱量×(31/30)</td>
                            </tr>
                            <tr>
                                <td>出库件数</td>
                                <td>当月未结束时，按实际天数/当月总天数折算</td>
                                <td>3 月数据截止到 30 日：实际件数×(31/30)</td>
                            </tr>
                            <tr>
                                <td>B/C 段收入（预估）</td>
                                <td>当月未结束时，按上月单箱收入×本月箱量预估</td>
                                <td>3 月 B 段收入=2 月 B 段单箱收入×3 月 B 段箱量</td>
                            </tr>
                            <tr>
                                <td>D 段收入</td>
                                <td>当月未结束时，按实际天数/当月总天数折算</td>
                                <td>3 月 D 段收入：实际收入×(30/31)</td>
                            </tr>
                            <tr>
                                <td>箱量与柜量换算</td>
                                <td>1 条柜 = 2 TEU = 2 箱量</td>
                                <td>612 TEU = 306 条柜</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                
                <div class="subsection">
                    <div class="subsection-title">四、内部公司排除规则</div>
                    <table>
                        <thead>
                            <tr>
                                <th>统计场景</th>
                                <th>排除范围</th>
                                <th>识别关键词</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>前十大客户排名</td>
                                <td>上海环世供应链科技有限公司</td>
                                <td>环世、WORLDWIDE LOGISTICS、WORLDWIDE SHIPPING</td>
                            </tr>
                            <tr>
                                <td>客户开发数/外部客户业绩</td>
                                <td>35 家环世集团内部公司</td>
                                <td>同上</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>
            
            <div class="note">
                <strong>📝 数据说明：</strong>
                <ul style="margin: 10px 0 0 20px;">
                    <li><strong>报告期间：</strong>{period}（当月={current_month}月，累计=1-{current_month}月）</li>
                    <li><strong>数据截止：</strong>2026 年 3 月 30 日</li>
                    <li><strong>预算来源：</strong>2026 年 DTC 预算数据（从合计行读取月度预算）</li>
                    <li><strong>同比数据：</strong>2025 年同期数据</li>
                    <li><strong>环比数据：</strong>与上月对比（3 月环比=与 2 月对比）</li>
                </ul>
            </div>
        </div>
        
        <div class="footer">
            <div>🐋 财鲸 (FinWhale) | DTC 经营分析报告 Skill</div>
            <div style="margin-top: 5px;">生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        </div>
    </div>
    
    <script>
        // 收入趋势图
        const revenueCtx = document.getElementById('revenueChart').getContext('2d');
        new Chart(revenueCtx, {{
            type: 'line',
            data: {{
                labels: {months_json},
                datasets: [
                    {{
                        label: 'B 段收入',
                        data: {revenue_b_json},
                        borderColor: '#2563eb',
                        backgroundColor: 'rgba(37, 99, 235, 0.1)',
                        tension: 0.3
                    }},
                    {{
                        label: 'C 段收入',
                        data: {revenue_c_json},
                        borderColor: '#059669',
                        backgroundColor: 'rgba(5, 150, 105, 0.1)',
                        tension: 0.3
                    }},
                    {{
                        label: 'D 段收入',
                        data: {revenue_d_json},
                        borderColor: '#d97706',
                        backgroundColor: 'rgba(217, 119, 6, 0.1)',
                        tension: 0.3
                    }},
                    {{
                        label: '总收入',
                        data: {revenue_total_json},
                        borderColor: '#7c3aed',
                        backgroundColor: 'rgba(124, 58, 237, 0.1)',
                        tension: 0.3,
                        borderDash: [5, 5]
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{
                        display: true,
                        text: '分业务段收入趋势（万元）'
                    }},
                    legend: {{
                        position: 'top'
                    }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        title: {{
                            display: true,
                            text: '万元'
                        }}
                    }}
                }}
            }}
        }});
        
        // 自营仓趋势图
        const selfRunCtx = document.getElementById('selfRunChart').getContext('2d');
        new Chart(selfRunCtx, {{
            type: 'line',
            data: {{
                labels: {months_12_json},
                datasets: [
                    {{
                        label: '美东自营仓入库',
                        data: {json.dumps(warehouse_data['self_run_trend']['USNJHM01']['inbound'])},
                        borderColor: '#2563eb',
                        backgroundColor: 'rgba(37, 99, 235, 0.1)',
                        tension: 0.3,
                        yAxisID: 'y'
                    }},
                    {{
                        label: '美东自营仓出库',
                        data: {json.dumps(warehouse_data['self_run_trend']['USNJHM01']['outbound'])},
                        borderColor: '#2563eb',
                        backgroundColor: 'rgba(37, 99, 235, 0.1)',
                        borderDash: [5, 5],
                        tension: 0.3,
                        yAxisID: 'y1'
                    }},
                    {{
                        label: '美西自营仓入库',
                        data: {json.dumps(warehouse_data['self_run_trend']['USCAEA02']['inbound'])},
                        borderColor: '#059669',
                        backgroundColor: 'rgba(5, 150, 105, 0.1)',
                        tension: 0.3,
                        yAxisID: 'y'
                    }},
                    {{
                        label: '美西自营仓出库',
                        data: {json.dumps(warehouse_data['self_run_trend']['USCAEA02']['outbound'])},
                        borderColor: '#059669',
                        backgroundColor: 'rgba(5, 150, 105, 0.1)',
                        borderDash: [5, 5],
                        tension: 0.3,
                        yAxisID: 'y1'
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                interaction: {{
                    mode: 'index',
                    intersect: false
                }},
                plugins: {{
                    title: {{
                        display: true,
                        text: '自营仓入库柜量（条）和出库件数趋势'
                    }},
                    legend: {{
                        position: 'top'
                    }}
                }},
                scales: {{
                    y: {{
                        type: 'linear',
                        display: true,
                        position: 'left',
                        title: {{
                            display: true,
                            text: '入库柜量（条）'
                        }}
                    }},
                    y1: {{
                        type: 'linear',
                        display: true,
                        position: 'right',
                        title: {{
                            display: true,
                            text: '出库件数（件）'
                        }},
                        grid: {{
                            drawOnChartArea: false
                        }}
                    }}
                }}
            }}
        }});
        
        // 箱量趋势图
        const volumeCtx = document.getElementById('volumeChart').getContext('2d');
        new Chart(volumeCtx, {{
            type: 'line',
            data: {{
                labels: {months_json},
                datasets: [
                    {{
                        label: 'B 段箱量',
                        data: {volume_b_json},
                        borderColor: '#2563eb',
                        backgroundColor: 'rgba(37, 99, 235, 0.1)',
                        tension: 0.3
                    }},
                    {{
                        label: 'C 段箱量',
                        data: {volume_c_json},
                        borderColor: '#059669',
                        backgroundColor: 'rgba(5, 150, 105, 0.1)',
                        tension: 0.3
                    }},
                    {{
                        label: 'D 段箱量',
                        data: {volume_d_json},
                        borderColor: '#d97706',
                        backgroundColor: 'rgba(217, 119, 6, 0.1)',
                        tension: 0.3
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{
                        display: true,
                        text: '分业务段箱量趋势（TEU）'
                    }},
                    legend: {{
                        position: 'top'
                    }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        title: {{
                            display: true,
                            text: 'TEU'
                        }}
                    }}
                }}
            }}
        }});
        
        // 出库件数趋势图
        const outboundCtx = document.getElementById('outboundChart').getContext('2d');
        new Chart(outboundCtx, {{
            type: 'line',
            data: {{
                labels: {months_json},
                datasets: [{{
                    label: '出库件数',
                    data: {outbound_json},
                    borderColor: '#8b5cf6',
                    backgroundColor: 'rgba(139, 92, 246, 0.1)',
                    tension: 0.4,
                    fill: true
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    title: {{
                        display: true,
                        text: '出库件数月度趋势（件）'
                    }},
                    legend: {{
                        position: 'top'
                    }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        title: {{
                            display: true,
                            text: '件'
                        }}
                    }}
                }}
            }}
        }});
        
        
    </script>
</body>
</html>
'''
    
    return html


def main():
    parser = argparse.ArgumentParser(description='DTC 经营分析报告生成器')
    parser.add_argument('--period', default='2026-Q1', help='报告期间（如 2026-Q1, 2026-03）')
    parser.add_argument('--output', default='reports', help='输出目录')
    parser.add_argument('--trend-months', type=int, default=6, help='趋势分析月数（默认 6 个月）')
    
    args = parser.parse_args()
    
    print(f'【DTC 经营分析报告生成器】')
    print(f'报告期间：{args.period}')
    print(f'输出目录：{args.output}')
    print(f'趋势分析：最近{args.trend_months}个月')
    
    # 解析期间
    current_month, cumulative_months = parse_period(args.period)
    print(f'期间解析：当月={current_month}月，累计=1-{current_month}月')
    
    # 读取数据
    print('\n【一、读取数据】')
    
    print('  读取预算数据...')
    # 从预算 Excel 读取 2026 年 Q1 各业务段预算（逐月累加）
    budget = read_budget_data()
    
    if budget and budget.get('revenue_cumulative'):
        print(f'    收入预算（Q1 逐月累加）: {budget["revenue_cumulative"]:.0f}万')
        print(f'      B 段：{budget["b_revenue_cumulative"]:.0f}万，C 段：{budget["c_revenue_cumulative"]:.0f}万，D 段：{budget["d_revenue_cumulative"]:.0f}万')
        print(f'    毛利预算（Q1 逐月累加）: {budget["gross_profit_cumulative"]:.0f}万')
        print(f'      B 段：{budget["b_profit_cumulative"]:.1f}万，C 段：{budget["c_profit_cumulative"]:.0f}万，D 段：{budget["d_profit_cumulative"]:.0f}万')
    else:
        print('    ⚠️ 预算数据读取失败')
    
    print('  读取箱量数据（2026 年）...')
    volume = read_volume_data(args.period, yoy=False)
    print(f'    当月：B={volume["current"]["b_volume"]:.0f}, C={volume["current"]["c_volume"]:.0f}, D={volume["current"]["d_volume"]:.0f}')
    print(f'    累计：B={volume["cumulative"]["b_volume"]:.0f}, C={volume["cumulative"]["c_volume"]:.0f}, D={volume["cumulative"]["d_volume"]:.0f}')
    
    print('  读取箱量数据（2025 年同比）...')
    volume_yoy = read_volume_data(args.period, yoy=True)
    
    print('  读取箱量数据（上月环比）...')
    # 手动指定上月数据（2 月）用于环比计算
    volume_mom = {
        'current': {'b_volume': 545, 'c_volume': 623, 'd_volume': 630},  # 2 月实际
        'cumulative': {'b_volume': 845, 'c_volume': 1230, 'd_volume': 1302}  # 1-2 月累计
    }
    print(f'    上月（2 月）: B={volume_mom["current"]["b_volume"]:.0f}, C={volume_mom["current"]["c_volume"]:.0f}, D={volume_mom["current"]["d_volume"]:.0f}')
    
    print('  读取出库件数（2026 年）...')
    outbound = read_outbound_pieces(args.period, yoy=False)
    
    print('  读取出库件数（2025 年同比）...')
    outbound_yoy = read_outbound_pieces(args.period, yoy=True)
    
    print('  读取出库件数（上月环比）...')
    # 手动指定上月数据（2 月）用于环比计算
    outbound_mom = {
        'current': 186063,  # 2 月出库件数
        'cumulative': 302776  # 1-2 月累计
    }
    print(f'    上月（2 月）: {outbound_mom["current"]:.0f}件')
    
    print('  读取收入/毛利数据（2026 年）...')
    actual = read_revenue_profit_by_segment(args.period, yoy=False)
    
    print('  读取收入/毛利数据（2025 年同比）...')
    actual_yoy = read_revenue_profit_by_segment(args.period, yoy=True)
    
    print(f'  读取趋势数据（最近{args.trend_months}个月）...')
    trend_data = read_monthly_trend_data(12)  # 近 12 个月
    if trend_data["months"]:
        print(f'    月份范围：{trend_data["months"][0]} 至 {trend_data["months"][-1]}')
    else:
        print('    ⚠️ 趋势数据为空')
    
    print('  计算单箱指标...')
    single_box = calculate_single_box_metrics(args.period, months=6)
    
    print('  读取客户与销售数据...')
    customer_data = read_customer_data(args.period)
    result = read_top_customers_and_sales(args.period)
    top_customers, top_sales, top_sales_d_volume, top_current_customers, top_current_sales = result
    print(f'    当月新增客户：{len(customer_data["current"]["total"])}家')
    print(f'    累计客户：{len(customer_data["cumulative"]["total"])}家')
    print(f'    前十大客户：{len(top_customers)}家')
    print(f'    前十大销售：{len(top_sales)}家')
    
    # 客户年代分析
    customer_vintage = None
    if CUSTOMER_VINTAGE_ENABLED:
        print('  进行客户年代分析...')
        try:
            customer_vintage = analyze_customer_vintage(args.period)
            print(f'    累计合作客户：{customer_vintage.get("total_unique_customers", 0)}家')
        except Exception as e:
            print(f'    ⚠️ 客户年代分析失败：{e}')
            customer_vintage = None
    
    print('  读取海外仓数据...')
    warehouse_data = read_warehouse_data(args.period)
    loss_customers = read_loss_customers(args.period)
    print(f'    仓库分布国家：{len(warehouse_data["by_country"])}个')
    print(f'    亏损客户：{len(loss_customers)}家')
    
    # AI 深度分析
    ai_analysis = None
    if AI_ANALYSIS_ENABLED:
        print('  调用 AI 分析...')
        try:
            # 读取同比箱量数据
            volume_yoy = read_volume_data(args.period, yoy=True)
            ai_analysis = analyze_business_performance(actual, budget, actual_yoy, volume, volume_yoy, args.period)
            print(f'    识别异常：{len(ai_analysis.get("anomalies", []))}个')
            print(f'    生成建议：{len(ai_analysis.get("recommendations", []))}条')
        except Exception as e:
            print(f'    ⚠️ AI 分析失败：{e}')
            import traceback
            traceback.print_exc()
            ai_analysis = None
    
    print('  生成关键发现...')
    highlights, concerns = generate_key_findings(
        actual, actual_yoy, budget, volume, volume_yoy,
        customer_data, warehouse_data, loss_customers,
        args.period
    )
    print(f'    亮点：{len(highlights)}条')
    print(f'    关注点：{len(concerns)}条')
    
    # 生成报告
    print('\n【二、生成报告】')
    html = generate_html_report(
        actual, actual_yoy, budget, volume, volume_yoy, volume_mom, outbound, outbound_yoy, outbound_mom,
        single_box, trend_data, customer_data, top_customers, top_sales, top_sales_d_volume,
        top_current_customers, top_current_sales,
        warehouse_data, loss_customers, highlights, concerns,
        args.period, ai_analysis, customer_vintage
    )
    
    # 保存
    os.makedirs(args.output, exist_ok=True)
    output_file = os.path.join(args.output, f'DTC_{args.period}_经营分析报告.html')
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f'\n✅ 报告已生成：{output_file}')


if __name__ == '__main__':
    main()
