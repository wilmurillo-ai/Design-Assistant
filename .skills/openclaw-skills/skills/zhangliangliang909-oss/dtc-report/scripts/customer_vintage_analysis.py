#!/usr/bin/env python3
"""
客户年代分析模块

功能：
1. 统计各年合作客户数（2022-2026）
2. 分析各年开发客户对本年收入的贡献
"""

import openpyxl
import glob
import os
from collections import defaultdict
from datetime import datetime, timedelta

# 数据目录（本地路径）
BUSINESS_DATA_DIR = r"C:\Users\wwl\.openclaw\workspace-跨境电商\data\1.业务和订单数据"


def analyze_customer_vintage(period='2026-Q1'):
    """
    客户年代分析
    
    返回：
    - 各年合作客户数统计
    - 各年开发客户对本年收入的贡献
    """
    print('  进行客户年代分析...')
    
    # 1. 读取历史客户合作数据，确定每个客户的首次合作年份
    customer_first_year = read_customer_first_cooperation_year()
    
    # 2. 读取 2022-2026 年各年合作客户数
    yearly_customers = read_yearly_cooperation_customers(customer_first_year)
    
    # 3. 读取 2026 Q1 收入数据，按客户汇总
    current_revenue_by_customer = read_current_revenue_by_customer(period)
    
    # 4. 分析各年开发客户对本年收入的贡献
    revenue_contribution_by_vintage = analyze_revenue_contribution_by_vintage(
        customer_first_year, 
        current_revenue_by_customer
    )
    
    return {
        'yearly_customers': yearly_customers,
        'revenue_contribution_by_vintage': revenue_contribution_by_vintage,
        'total_unique_customers': len(customer_first_year)
    }


def read_customer_first_cooperation_year():
    """
    读取每个客户的首次合作年份
    
    从 DTC 历史客户合作情况汇总.xlsx 读取
    """
    customer_first_year = {}
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*历史客户合作情况汇总*.xlsx'))
    if not files:
        print('    ⚠️ 未找到历史客户合作情况汇总文件')
        return customer_first_year
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    # 找表头
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    customer_col = headers.get('客户名称', 1)
    year_col = headers.get('第一次合作年', 4)
    segment_col = headers.get('业务分段列表', 2)
    
    for row in range(2, min(ws.max_row + 1, 10000)):
        customer = ws.cell(row=row, column=customer_col).value
        first_year = ws.cell(row=row, column=year_col).value
        segment = ws.cell(row=row, column=segment_col).value
        
        if not customer or not first_year:
            continue
        
        # 处理业务分段（按逗号分割后逐个判断）
        segments = []
        seg_str = str(segment).strip() if segment else ''
        seg_list = [s.strip() for s in seg_str.split(',')]
        
        for s in seg_list:
            if 'A' in s or 'B' in s:
                if 'b' not in segments:
                    segments.append('b')
            if 'C' in s:
                if 'c' not in segments:
                    segments.append('c')
            # D 段仅包含 D 段和 D，电商集拼/集拼/其他不属于 D 段
            if ('D' in s or s == 'D') and '电商' not in s and '集拼' not in s:
                if 'd' not in segments:
                    segments.append('d')
        
        # 记录客户首次合作年份（分业务段）
        if customer not in customer_first_year:
            customer_first_year[customer] = {'b': None, 'c': None, 'd': None}
        
        for seg in segments:
            try:
                year = int(first_year)
                if customer_first_year[customer][seg] is None or year < customer_first_year[customer][seg]:
                    customer_first_year[customer][seg] = year
            except:
                continue
    
    print(f'    读取历史客户：{len(customer_first_year)}家')
    return customer_first_year


def read_yearly_cooperation_customers(customer_first_year):
    """
    统计 2022-2026 年各年合作客户数
    
    新口径：按客户维度统计
    - 只要客户该年有 B 段收入（>0），就算 B 段客户
    - 只要客户该年有 C 段收入（>0），就算 C 段客户
    - 只要客户该年有 D 段收入（>0），就算 D 段客户
    - 同一客户可能同时出现在 B、C、D 段统计中
    - 合计客户数去重
    """
    # 按客户汇总各年各业务段收入
    customer_yearly_revenue = {}  # {customer: {year: {seg: revenue}}}
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if not files:
        return {}
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    # 找表头
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    bu_col = headers.get('业务系统单元编码', 1)
    segment_col = headers.get('业务分段分类_新.', 2)
    month_col = headers.get('业务年月', 3)
    customer_col = headers.get('委托客户名称', 4)
    revenue_col = headers.get('收入', 8)
    
    for row in range(2, min(ws.max_row + 1, 50000)):
        bu = ws.cell(row=row, column=bu_col).value
        segment = ws.cell(row=row, column=segment_col).value
        month = ws.cell(row=row, column=month_col).value
        customer = ws.cell(row=row, column=customer_col).value
        revenue = ws.cell(row=row, column=revenue_col).value
        
        if not bu or str(bu).strip() != 'BWLDTC':
            continue
        
        if not customer or not revenue or revenue <= 0:
            continue
        
        # 判断业务段
        seg_str = str(segment).strip() if segment else ''
        segments = []
        if 'B 段' in seg_str or 'B' in seg_str or 'A+B' in seg_str:
            segments.append('b')
        if 'C 段' in seg_str or 'C' in seg_str or 'C+D' in seg_str:
            segments.append('c')
        # D 段仅包含 D 段和 D，电商集拼/集拼/其他不属于 D 段
        if 'D 段' in seg_str or seg_str == 'D':
            segments.append('d')
        
        # 判断年份
        year = None
        if isinstance(month, (int, float)):
            base_date = datetime(1899, 12, 30)
            actual_date = base_date + timedelta(days=int(month))
            year = str(actual_date.year)
        elif isinstance(month, str):
            if '年' in month:
                year = month.split('年')[0].strip()
            elif '-' in month:
                year = month.split('-')[0].strip()
        
        if year not in ['2022', '2023', '2024', '2025', '2026']:
            continue
        
        # 按客户 + 年份 + 业务段汇总收入
        if customer not in customer_yearly_revenue:
            customer_yearly_revenue[customer] = {}
        if year not in customer_yearly_revenue[customer]:
            customer_yearly_revenue[customer][year] = {'b': 0, 'c': 0, 'd': 0}
        
        for seg in segments:
            customer_yearly_revenue[customer][year][seg] += float(revenue)
    
    # 统计各年合作客户数（只要该业务段收入>0，就算该业务段客户）
    yearly_customers = {}
    for year in ['2022', '2023', '2024', '2025', '2026']:
        yearly_customers[year] = {'b': set(), 'c': set(), 'd': set()}
        
        for customer, year_data in customer_yearly_revenue.items():
            if year not in year_data:
                continue
            
            # 只要该业务段有收入，就算该业务段客户
            if year_data[year]['b'] > 0:
                yearly_customers[year]['b'].add(customer)
            if year_data[year]['c'] > 0:
                yearly_customers[year]['c'].add(customer)
            if year_data[year]['d'] > 0:
                yearly_customers[year]['d'].add(customer)
    
    # 统计各年客户数
    result = {}
    for year in ['2022', '2023', '2024', '2025', '2026']:
        result[year] = {
            'b': len(yearly_customers[year]['b']),
            'c': len(yearly_customers[year]['c']),
            'd': len(yearly_customers[year]['d']),
            'total': len(yearly_customers[year]['b'] | yearly_customers[year]['c'] | yearly_customers[year]['d'])
        }
    
    return result


def read_current_revenue_by_customer(period='2026-Q1'):
    """
    读取 2026 Q1 各客户收入（分业务段）
    """
    current_month, cumulative_months = parse_period(period)
    
    revenue_by_customer = defaultdict(lambda: {'b': 0, 'c': 0, 'd': 0})
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if not files:
        return revenue_by_customer
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    # 找表头
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    bu_col = headers.get('业务系统单元编码', 1)
    segment_col = headers.get('业务分段分类_新.', 2)
    month_col = headers.get('业务年月', 3)
    customer_col = headers.get('委托客户名称', 4)
    revenue_col = headers.get('收入', 8)
    
    for row in range(2, min(ws.max_row + 1, 50000)):
        bu = ws.cell(row=row, column=bu_col).value
        segment = ws.cell(row=row, column=segment_col).value
        month = ws.cell(row=row, column=month_col).value
        customer = ws.cell(row=row, column=customer_col).value
        revenue = ws.cell(row=row, column=revenue_col).value
        
        if not bu or str(bu).strip() != 'BWLDTC':
            continue
        
        if not customer or not revenue:
            continue
        
        # 判断业务段
        seg_str = str(segment).strip() if segment else ''
        segments = []
        if 'B 段' in seg_str or 'B' in seg_str or 'A+B' in seg_str:
            segments.append('b')
        if 'C 段' in seg_str or 'C' in seg_str or 'C+D' in seg_str:
            segments.append('c')
        # D 段仅包含 D 段和 D，电商集拼/集拼/其他不属于 D 段
        if 'D 段' in seg_str or seg_str == 'D':
            segments.append('d')
        
        # 判断月份（需要同时检查年份和月份）
        month_num = None
        year = None
        if isinstance(month, (int, float)):
            base_date = datetime(1899, 12, 30)
            actual_date = base_date + timedelta(days=int(month))
            month_num = actual_date.month
            year = actual_date.year
        elif isinstance(month, str) and '-' in month:
            try:
                parts = month.split('-')
                year = int(parts[0])
                month_num = int(parts[1])
            except:
                continue
        
        # 检查年份（2026 年）和月份（1-3 月）
        target_year = 2026 if period.startswith('2026') else 2025
        if year != target_year or month_num not in cumulative_months:
            continue
        
        # 累加收入（转换为万元）
        for seg in segments:
            revenue_by_customer[customer][seg] += float(revenue) / 10000
    
    return revenue_by_customer


def analyze_revenue_contribution_by_vintage(customer_first_year, current_revenue_by_customer):
    """
    分析各年开发客户对本年收入的贡献
    
    按客户首次合作年份分组，统计本期收入贡献
    """
    contribution = {
        '2022': {'b': 0, 'c': 0, 'd': 0, 'customers': set()},
        '2023': {'b': 0, 'c': 0, 'd': 0, 'customers': set()},
        '2024': {'b': 0, 'c': 0, 'd': 0, 'customers': set()},
        '2025': {'b': 0, 'c': 0, 'd': 0, 'customers': set()},
        '2026': {'b': 0, 'c': 0, 'd': 0, 'customers': set()},
        'unknown': {'b': 0, 'c': 0, 'd': 0, 'customers': set()}
    }
    
    # 创建客户名称映射（标准化处理）
    def normalize_name(name):
        if not name:
            return ''
        # 去除空格和特殊字符
        return str(name).strip().replace(' ', '').replace('(', '（').replace(')', '）')
    
    # 构建标准化名称映射
    normalized_first_year = {}
    for customer, years in customer_first_year.items():
        normalized = normalize_name(customer)
        normalized_first_year[normalized] = years
    
    matched_count = 0
    unmatched_count = 0
    
    for customer, revenue in current_revenue_by_customer.items():
        # 获取客户首次合作年份（尝试标准化名称匹配）
        normalized_customer = normalize_name(customer)
        first_year = normalized_first_year.get(normalized_customer, customer_first_year.get(customer, {}))
        
        if not first_year and normalized_customer in normalized_first_year:
            first_year = normalized_first_year[normalized_customer]
        
        for seg in ['b', 'c', 'd']:
            if revenue[seg] > 0:
                # 确定年份
                year = first_year.get(seg, None) if first_year else None
                if year and str(year) in ['2022', '2023', '2024', '2025', '2026']:
                    vintage = str(year)
                    matched_count += 1
                else:
                    vintage = 'unknown'
                    if first_year:
                        matched_count += 1
                    else:
                        unmatched_count += 1
                
                contribution[vintage][seg] += revenue[seg]
                contribution[vintage]['customers'].add(customer)
    
    print(f'    客户匹配：{matched_count}家已匹配，{unmatched_count}家未匹配')
    
    # 格式化结果
    result = {}
    for vintage in ['2022', '2023', '2024', '2025', '2026', 'unknown']:
        result[vintage] = {
            'b_revenue': contribution[vintage]['b'],
            'c_revenue': contribution[vintage]['c'],
            'd_revenue': contribution[vintage]['d'],
            'total_revenue': contribution[vintage]['b'] + contribution[vintage]['c'] + contribution[vintage]['d'],
            'customer_count': len(contribution[vintage]['customers'])
        }
    
    return result


def parse_period(period):
    """解析期间参数"""
    if period.upper().startswith('2026-Q'):
        quarter = int(period[-1])
        current_month = quarter * 3
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    if period.startswith('2026-'):
        month_part = period[5:]
        current_month = int(month_part)
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    return 3, [1, 2, 3]


def format_customer_vintage_html(analysis_result):
    """将客户年代分析结果格式化为 HTML（只返回 subsection，不包含 section 标题）"""
    if not analysis_result:
        return ''
    
    html = '''
        <div class="subsection">
            <div class="subsection-title">4.1 各年合作客户数统计</div>
            <table style="width: auto; min-width: 500px;">
                <thead>
                    <tr>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: left;">业务分段</th>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: right;">2025 年</th>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: right;">2026 年</th>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: right;">合计</th>
                    </tr>
                </thead>
                <tbody>
'''
    
    yearly = analysis_result.get('yearly_customers', {})
    for seg in ['b', 'c', 'd']:
        seg_name = {'b': 'B 段', 'c': 'C 段', 'd': 'D 段'}[seg]
        y2025 = yearly.get('2025', {}).get(seg, 0)
        y2026 = yearly.get('2026', {}).get(seg, 0)
        row_total = y2025 + y2026
        html += f'''
                    <tr>
                        <td style="padding: 12px 15px; text-align: left;">{seg_name}</td>
                        <td style="padding: 12px 15px; text-align: right;">{y2025:,}</td>
                        <td style="padding: 12px 15px; text-align: right;">{y2026:,}</td>
                        <td style="padding: 12px 15px; text-align: right;">{row_total:,}</td>
                    </tr>
'''
    
    # 合计行
    total_2025 = yearly.get('2025', {}).get('total', 0)
    total_2026 = yearly.get('2026', {}).get('total', 0)
    grand_total = total_2025 + total_2026
    html += f'''
                    <tr style="background: #f3f4f6; font-weight: 600;">
                        <td style="padding: 12px 15px; text-align: left;">合计</td>
                        <td style="padding: 12px 15px; text-align: right;">{total_2025:,}</td>
                        <td style="padding: 12px 15px; text-align: right;">{total_2026:,}</td>
                        <td style="padding: 12px 15px; text-align: right;">{grand_total:,}</td>
                    </tr>
                </tbody>
            </table>
        </div>
        
        <div class="subsection">
            <div class="subsection-title">4.2 各年开发客户对本期收入的贡献</div>
            <table style="width: auto; min-width: 700px;">
                <thead>
                    <tr>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: left;">业务分段</th>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: right;">2022 年</th>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: right;">2023 年</th>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: right;">2024 年</th>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: right;">2025 年</th>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: right;">2026 年</th>
                        <th style="background: #2563eb; color: white; padding: 12px 15px; text-align: right;">合计</th>
                    </tr>
                </thead>
                <tbody>
'''
    
    vintage_data = analysis_result.get('revenue_contribution_by_vintage', {})
    
    for seg in ['b', 'c', 'd']:
        seg_name = {'b': 'B 段', 'c': 'C 段', 'd': 'D 段'}[seg]
        row_total = 0
        html += f'''
                    <tr>
                        <td style="padding: 12px 15px; text-align: left; font-weight: 600;">{seg_name}</td>
'''
        for year in ['2022', '2023', '2024', '2025', '2026']:
            rev = vintage_data.get(year, {}).get(f'{seg}_revenue', 0)
            row_total += rev
            html += f'''                        <td style="padding: 12px 15px; text-align: right;">{rev:,.1f}</td>
'''
        html += f'''                        <td style="padding: 12px 15px; text-align: right; font-weight: 600;">{row_total:,.1f}</td>
                    </tr>
'''
    
    # 合计行
    html += '''
                    <tr style="background: #f3f4f6; font-weight: 600;">
                        <td style="padding: 12px 15px; text-align: left;">合计</td>
'''
    grand_total = 0
    for year in ['2022', '2023', '2024', '2025', '2026']:
        total = vintage_data.get(year, {}).get('total_revenue', 0)
        grand_total += total
        html += f'''                        <td style="padding: 12px 15px; text-align: right;">{total:,.1f}</td>
'''
    html += f'''                        <td style="padding: 12px 15px; text-align: right; font-weight: 600;">{grand_total:,.1f}</td>
                    </tr>
                </tbody>
            </table>
            <div class="data-source" style="margin-top: 10px; font-size: 12px; color: #6b7280;">
                注：按客户首次合作年份分组，统计本期（2026 Q1）收入贡献（万元）；
                同一客户可能合作多个业务段，因此各业务段客户数之和可能大于总客户数
            </div>
        </div>
'''
    
    return html


if __name__ == '__main__':
    print('客户年代分析模块加载成功')
