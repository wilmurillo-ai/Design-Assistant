"""
PDF处理模块
支持PDF的格式转换、内容提取和批量处理
"""

import io
import logging
from typing import Dict, List, Optional, Any, Union
from enum import Enum

import pdfplumber
from PyPDF2 import PdfReader, PdfWriter
from docx import Document
from openpyxl import Workbook
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)


class ConversionFormat(str, Enum):
    """转换格式"""
    WORD = "word"
    EXCEL = "excel"
    PPT = "ppt"
    IMAGE = "image"


class ExtractionType(str, Enum):
    """提取类型"""
    TEXT = "text"
    TABLES = "tables"
    SUMMARY = "summary"
    KEY_POINTS = "key_points"


class ConversionRequest(BaseModel):
    """格式转换请求"""
    pdf_data: bytes = Field(..., description="PDF文件数据")
    target_format: ConversionFormat = Field(..., description="目标格式")
    preserve_formatting: bool = Field(True, description="是否保留格式")


class ExtractionRequest(BaseModel):
    """内容提取请求"""
    pdf_data: bytes = Field(..., description="PDF文件数据")
    extraction_type: ExtractionType = Field(..., description="提取类型")
    page_range: Optional[str] = Field(None, description="页码范围，如'1-5,8,10-12'")


class ExtractionResult(BaseModel):
    """提取结果"""
    content: str = Field(..., description="提取的内容")
    page_count: int = Field(..., description="总页数")
    word_count: int = Field(0, description="字数统计")
    metadata: Dict[str, Any] = Field(default_factory=dict, description="元数据")


class MergeRequest(BaseModel):
    """合并请求"""
    pdf_files: List[bytes] = Field(..., description="PDF文件列表")


class SplitRequest(BaseModel):
    """拆分请求"""
    pdf_data: bytes = Field(..., description="PDF文件数据")
    split_mode: str = Field("page", description="拆分模式: page/range")
    page_ranges: List[str] = Field(default_factory=list, description="页码范围列表")


class PDFProcessor:
    """PDF处理器"""
    
    def __init__(self):
        pass
        
    async def convert_format(
        self,
        request: ConversionRequest,
    ) -> bytes:
        """
        格式转换
        
        Args:
            request: 转换请求
            
        Returns:
            转换后的文件字节流
        """
        logger.info(f"开始PDF转换: {request.target_format.value}")
        
        return await self._local_convert(request)
        
    async def _local_convert(self, request: ConversionRequest) -> bytes:
        """本地转换"""
        if request.target_format == ConversionFormat.WORD:
            return await self._convert_to_word(request.pdf_data)
        elif request.target_format == ConversionFormat.EXCEL:
            return await self._convert_to_excel(request.pdf_data)
        else:
            raise ValueError(f"不支持的转换格式: {request.target_format.value}")
            
    async def _convert_to_word(self, pdf_data: bytes) -> bytes:
        """转换为Word"""
        doc = Document()
        
        with pdfplumber.open(io.BytesIO(pdf_data)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    doc.add_paragraph(text)
                    doc.add_page_break()
                    
        output = io.BytesIO()
        doc.save(output)
        output.seek(0)
        
        logger.info("PDF转Word完成")
        return output.getvalue()
        
    async def _convert_to_excel(self, pdf_data: bytes) -> bytes:
        """转换为Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "PDF内容"
        
        row_idx = 1
        
        with pdfplumber.open(io.BytesIO(pdf_data)) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()
                
                if tables:
                    for table in tables:
                        ws.cell(row=row_idx, column=1, value=f"第{page_num}页")
                        row_idx += 1
                        
                        for row in table:
                            for col_idx, cell in enumerate(row, 1):
                                ws.cell(row=row_idx, column=col_idx, value=cell)
                            row_idx += 1
                            
                        row_idx += 2
                else:
                    text = page.extract_text()
                    if text:
                        ws.cell(row=row_idx, column=1, value=f"第{page_num}页")
                        row_idx += 1
                        ws.cell(row=row_idx, column=1, value=text)
                        row_idx += 2
                        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("PDF转Excel完成")
        return output.getvalue()
        
    async def extract_content(
        self,
        request: ExtractionRequest,
    ) -> ExtractionResult:
        """
        内容提取
        
        Args:
            request: 提取请求
            
        Returns:
            提取结果
        """
        logger.info(f"开始内容提取: {request.extraction_type.value}")
        
        pdf_reader = PdfReader(io.BytesIO(request.pdf_data))
        page_count = len(pdf_reader.pages)
        
        pages_to_extract = self._parse_page_range(request.page_range, page_count)
        
        if request.extraction_type == ExtractionType.TEXT:
            content = await self._extract_text(request.pdf_data, pages_to_extract)
            
        elif request.extraction_type == ExtractionType.TABLES:
            content = await self._extract_tables(request.pdf_data, pages_to_extract)
            
        elif request.extraction_type == ExtractionType.SUMMARY:
            content = await self._extract_summary(request.pdf_data, pages_to_extract)
            
        elif request.extraction_type == ExtractionType.KEY_POINTS:
            content = await self._extract_key_points(request.pdf_data, pages_to_extract)
            
        else:
            content = ""
            
        word_count = len(content)
        
        logger.info(f"内容提取完成，共 {word_count} 字")
        
        return ExtractionResult(
            content=content,
            page_count=page_count,
            word_count=word_count,
            metadata={
                "extraction_type": request.extraction_type.value,
                "pages_extracted": len(pages_to_extract),
            }
        )
        
    def _parse_page_range(self, page_range: Optional[str], total_pages: int) -> List[int]:
        """解析页码范围"""
        if not page_range:
            return list(range(total_pages))
            
        pages = []
        parts = page_range.split(',')
        
        for part in parts:
            part = part.strip()
            if '-' in part:
                start, end = part.split('-')
                pages.extend(range(int(start) - 1, int(end)))
            else:
                pages.append(int(part) - 1)
                
        return [p for p in pages if 0 <= p < total_pages]
        
    async def _extract_text(self, pdf_data: bytes, pages: List[int]) -> str:
        """提取文本"""
        text_parts = []
        
        with pdfplumber.open(io.BytesIO(pdf_data)) as pdf:
            for page_num in pages:
                if page_num < len(pdf.pages):
                    text = pdf.pages[page_num].extract_text()
                    if text:
                        text_parts.append(f"--- 第{page_num + 1}页 ---\n{text}")
                        
        return "\n\n".join(text_parts)
        
    async def _extract_tables(self, pdf_data: bytes, pages: List[int]) -> str:
        """提取表格"""
        table_parts = []
        
        with pdfplumber.open(io.BytesIO(pdf_data)) as pdf:
            for page_num in pages:
                if page_num < len(pdf.pages):
                    tables = pdf.pages[page_num].extract_tables()
                    if tables:
                        for i, table in enumerate(tables, 1):
                            table_parts.append(f"--- 第{page_num + 1}页 表格{i} ---")
                            for row in table:
                                table_parts.append(" | ".join(str(cell) for cell in row))
                                
        return "\n".join(table_parts)
        
    async def _extract_summary(self, pdf_data: bytes, pages: List[int]) -> str:
        """提取摘要"""
        text = await self._extract_text(pdf_data, pages)
        
        if len(text) > 500:
            return text[:500] + "..."
        return text
        
    async def _extract_key_points(self, pdf_data: bytes, pages: List[int]) -> str:
        """提取关键要点"""
        text = await self._extract_text(pdf_data, pages)
        
        lines = text.split('\n')[:10]
        return "\n".join(f"- {line.strip()}" for line in lines if line.strip())
        
    async def merge_pdfs(
        self,
        request: MergeRequest,
    ) -> bytes:
        """
        合并PDF
        
        Args:
            request: 合并请求
            
        Returns:
            合并后的PDF字节流
        """
        logger.info(f"开始合并 {len(request.pdf_files)} 个PDF文件")
        
        writer = PdfWriter()
        
        for pdf_data in request.pdf_files:
            reader = PdfReader(io.BytesIO(pdf_data))
            for page in reader.pages:
                writer.add_page(page)
                
        output = io.BytesIO()
        writer.write(output)
        output.seek(0)
        
        logger.info("PDF合并完成")
        return output.getvalue()
        
    async def split_pdf(
        self,
        request: SplitRequest,
    ) -> List[bytes]:
        """
        拆分PDF
        
        Args:
            request: 拆分请求
            
        Returns:
            拆分后的PDF列表
        """
        logger.info("开始拆分PDF")
        
        reader = PdfReader(io.BytesIO(request.pdf_data))
        total_pages = len(reader.pages)
        results = []
        
        if request.split_mode == "page":
            for i in range(total_pages):
                writer = PdfWriter()
                writer.add_page(reader.pages[i])
                
                output = io.BytesIO()
                writer.write(output)
                output.seek(0)
                results.append(output.getvalue())
                
        elif request.split_mode == "range":
            for page_range in request.page_ranges:
                pages = self._parse_page_range(page_range, total_pages)
                
                writer = PdfWriter()
                for page_num in pages:
                    writer.add_page(reader.pages[page_num])
                    
                output = io.BytesIO()
                writer.write(output)
                output.seek(0)
                results.append(output.getvalue())
                
        logger.info(f"PDF拆分完成，共 {len(results)} 个文件")
        return results
        
    async def add_watermark(
        self,
        pdf_data: bytes,
        watermark_text: str,
        opacity: float = 0.3,
    ) -> bytes:
        """
        添加水印
        
        Args:
            pdf_data: PDF数据
            watermark_text: 水印文本
            opacity: 透明度
            
        Returns:
            添加水印后的PDF
        """
        logger.info(f"添加水印: {watermark_text}")
        
        reader = PdfReader(io.BytesIO(pdf_data))
        writer = PdfWriter()
        
        for page in reader.pages:
            page.merge_page(self._create_watermark_page(watermark_text, opacity))
            writer.add_page(page)
            
        output = io.BytesIO()
        writer.write(output)
        output.seek(0)
        
        logger.info("水印添加完成")
        return output.getvalue()
        
    def _create_watermark_page(self, text: str, opacity: float) -> Any:
        """创建水印页面"""
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        
        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=letter)
        
        can.setFillColorRGB(0.5, 0.5, 0.5, alpha=opacity)
        can.setFont("Helvetica", 40)
        
        can.save()
        packet.seek(0)
        
        from PyPDF2 import PdfReader as PR
        return PR(packet).pages[0]
