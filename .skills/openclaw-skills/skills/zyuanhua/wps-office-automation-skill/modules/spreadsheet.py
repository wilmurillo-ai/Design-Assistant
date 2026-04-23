"""
表格处理模块
支持Excel表格的数据清洗、智能分析和图表生成
"""

import io
import logging
from typing import Dict, List, Optional, Any, Union
from enum import Enum

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel, Field, ConfigDict

logger = logging.getLogger(__name__)


class ChartType(str, Enum):
    """图表类型"""
    BAR = "bar"
    LINE = "line"
    PIE = "pie"
    SCATTER = "scatter"


class AnalysisType(str, Enum):
    """分析类型"""
    SUM = "sum"
    AVERAGE = "average"
    COUNT = "count"
    MAX = "max"
    MIN = "min"
    FILTER = "filter"
    PIVOT = "pivot"


class DataCleaningRequest(BaseModel):
    """数据清洗请求"""
    model_config = ConfigDict(arbitrary_types_allowed=True)
    
    data: Any = Field(..., description="数据源")
    remove_duplicates: bool = Field(True, description="是否删除重复项")
    handle_missing: str = Field("mean", description="缺失值处理方式")
    remove_outliers: bool = Field(True, description="是否移除异常值")
    outlier_method: str = Field("iqr", description="异常值检测方法")


class DataCleaningReport(BaseModel):
    """数据清洗报告"""
    original_rows: int = Field(..., description="原始行数")
    cleaned_rows: int = Field(..., description="清洗后行数")
    duplicates_removed: int = Field(0, description="删除的重复项数量")
    missing_values_handled: int = Field(0, description="处理的缺失值数量")
    outliers_removed: int = Field(0, description="移除的异常值数量")
    cleaning_steps: List[str] = Field(default_factory=list, description="清洗步骤")


class AnalysisRequest(BaseModel):
    """智能分析请求"""
    model_config = ConfigDict(arbitrary_types_allowed=True)
    
    data: Any = Field(..., description="数据源")
    analysis_type: AnalysisType = Field(..., description="分析类型")
    columns: Optional[List[str]] = Field(None, description="目标列")
    group_by: Optional[str] = Field(None, description="分组列")
    filter_condition: Optional[str] = Field(None, description="筛选条件")


class ChartGenerationRequest(BaseModel):
    """图表生成请求"""
    model_config = ConfigDict(arbitrary_types_allowed=True)
    
    data: Any = Field(..., description="数据源")
    chart_type: ChartType = Field(..., description="图表类型")
    x_column: str = Field(..., description="X轴列")
    y_columns: List[str] = Field(..., description="Y轴列")
    title: Optional[str] = Field(None, description="图表标题")


class SpreadsheetProcessor:
    """表格处理器"""
    
    def __init__(self):
        pass
        
    def _load_data(self, data: Union[str, bytes, pd.DataFrame]) -> pd.DataFrame:
        """加载数据"""
        if isinstance(data, pd.DataFrame):
            return data.copy()
            
        if isinstance(data, bytes):
            return pd.read_excel(io.BytesIO(data))
            
        if isinstance(data, str):
            if data.endswith(('.xlsx', '.xls', '.et')):
                return pd.read_excel(data)
            elif data.endswith('.csv'):
                return pd.read_csv(data)
            else:
                raise ValueError(f"不支持的文件格式: {data}")
                
        raise ValueError("不支持的数据类型")
        
    async def clean_data(
        self,
        request: DataCleaningRequest,
    ) -> tuple[pd.DataFrame, DataCleaningReport]:
        """
        数据清洗
        
        Args:
            request: 数据清洗请求
            
        Returns:
            清洗后的数据和清洗报告
        """
        logger.info("开始数据清洗")
        
        df = self._load_data(request.data)
        report = DataCleaningReport(
            original_rows=len(df),
            cleaned_rows=len(df),
        )
        
        if request.remove_duplicates:
            before = len(df)
            df = df.drop_duplicates()
            report.duplicates_removed = before - len(df)
            report.cleaned_rows = len(df)
            report.cleaning_steps.append(f"删除重复项: {report.duplicates_removed} 行")
            
        if request.handle_missing != "none":
            missing_count = df.isnull().sum().sum()
            if missing_count > 0:
                if request.handle_missing == "mean":
                    numeric_cols = df.select_dtypes(include=['number']).columns
                    df[numeric_cols] = df[numeric_cols].fillna(df[numeric_cols].mean())
                elif request.handle_missing == "median":
                    numeric_cols = df.select_dtypes(include=['number']).columns
                    df[numeric_cols] = df[numeric_cols].fillna(df[numeric_cols].median())
                elif request.handle_missing == "drop":
                    df = df.dropna()
                    report.cleaned_rows = len(df)
                    
                report.missing_values_handled = missing_count
                report.cleaning_steps.append(f"处理缺失值: {missing_count} 个")
                
        if request.remove_outliers:
            numeric_cols = df.select_dtypes(include=['number']).columns
            before = len(df)
            
            for col in numeric_cols:
                if request.outlier_method == "iqr":
                    Q1 = df[col].quantile(0.25)
                    Q3 = df[col].quantile(0.75)
                    IQR = Q3 - Q1
                    lower_bound = Q1 - 1.5 * IQR
                    upper_bound = Q3 + 1.5 * IQR
                    df = df[(df[col] >= lower_bound) & (df[col] <= upper_bound)]
                    
            report.outliers_removed = before - len(df)
            report.cleaned_rows = len(df)
            if report.outliers_removed > 0:
                report.cleaning_steps.append(f"移除异常值: {report.outliers_removed} 行")
                
        logger.info(f"数据清洗完成: {report.original_rows} -> {report.cleaned_rows} 行")
        return df, report
        
    async def analyze_data(
        self,
        request: AnalysisRequest,
    ) -> Dict[str, Any]:
        """
        智能分析数据
        
        Args:
            request: 分析请求
            
        Returns:
            分析结果
        """
        logger.info(f"开始数据分析: {request.analysis_type.value}")
        
        df = self._load_data(request.data)
        result = {}
        
        if request.filter_condition:
            try:
                df = df.query(request.filter_condition)
                result["filtered_rows"] = len(df)
            except Exception as e:
                logger.warning(f"筛选条件执行失败: {str(e)}")
                
        columns = request.columns or df.select_dtypes(include=['number']).columns.tolist()
        
        if request.analysis_type == AnalysisType.SUM:
            if request.group_by:
                result["data"] = df.groupby(request.group_by)[columns].sum().to_dict()
            else:
                result["data"] = df[columns].sum().to_dict()
                
        elif request.analysis_type == AnalysisType.AVERAGE:
            if request.group_by:
                result["data"] = df.groupby(request.group_by)[columns].mean().to_dict()
            else:
                result["data"] = df[columns].mean().to_dict()
                
        elif request.analysis_type == AnalysisType.COUNT:
            if request.group_by:
                result["data"] = df.groupby(request.group_by).size().to_dict()
            else:
                result["data"] = {"total": len(df)}
                
        elif request.analysis_type == AnalysisType.MAX:
            if request.group_by:
                result["data"] = df.groupby(request.group_by)[columns].max().to_dict()
            else:
                result["data"] = df[columns].max().to_dict()
                
        elif request.analysis_type == AnalysisType.MIN:
            if request.group_by:
                result["data"] = df.groupby(request.group_by)[columns].min().to_dict()
            else:
                result["data"] = df[columns].min().to_dict()
                
        elif request.analysis_type == AnalysisType.PIVOT:
            if request.group_by and columns:
                pivot_table = df.pivot_table(
                    index=request.group_by,
                    values=columns[0],
                    aggfunc='sum'
                )
                result["data"] = pivot_table.to_dict()
            else:
                result["error"] = "透视分析需要指定group_by和columns"
                
        result["analysis_type"] = request.analysis_type.value
        logger.info("数据分析完成")
        return result
        
    async def generate_chart(
        self,
        request: ChartGenerationRequest,
    ) -> bytes:
        """
        生成图表
        
        Args:
            request: 图表生成请求
            
        Returns:
            Excel文件字节流（包含图表）
        """
        logger.info(f"开始生成图表: {request.chart_type.value}")
        
        df = self._load_data(request.data)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "数据"
        
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
                
        x_col_idx = df.columns.get_loc(request.x_column) + 1
        y_col_indices = [df.columns.get_loc(col) + 1 for col in request.y_columns]
        
        if request.chart_type == ChartType.BAR:
            chart = BarChart()
        elif request.chart_type == ChartType.LINE:
            chart = LineChart()
        elif request.chart_type == ChartType.PIE:
            chart = PieChart()
        else:
            chart = BarChart()
            
        data = Reference(ws, min_col=y_col_indices[0], min_row=1, max_row=len(df) + 1)
        cats = Reference(ws, min_col=x_col_idx, min_row=2, max_row=len(df) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        if request.title:
            chart.title = request.title
            
        chart.width = 15
        chart.height = 10
        
        ws.add_chart(chart, "E2")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("图表生成完成")
        return output.getvalue()
        
    async def smart_analysis(
        self,
        data: Union[str, bytes, pd.DataFrame],
        question: str,
    ) -> Dict[str, Any]:
        """
        自然语言智能分析
        
        Args:
            data: 数据源
            question: 自然语言问题
            
        Returns:
            分析结果
        """
        logger.info(f"智能分析问题: {question}")
        
        df = self._load_data(data)
        
        return {
            "error": "无法解析问题，请使用具体的分析指令",
            "suggestion": "例如: '计算销售额的总和' 或 '按地区分组计算平均值'",
        }
