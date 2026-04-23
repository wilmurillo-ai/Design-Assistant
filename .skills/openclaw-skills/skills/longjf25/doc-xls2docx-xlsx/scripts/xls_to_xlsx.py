#!/usr/bin/env python3
"""
xls_to_xlsx.py - 将 .xls 文件转换为 .xlsx 格式
用法:
    python xls_to_xlsx.py input.xls [output.xlsx]
    python xls_to_xlsx.py --batch 输入目录 [输出目录]
"""

import os
import sys
import xlrd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import argparse
from pathlib import Path
from datetime import datetime


def safe_str(s):
    """安全字符串转换，处理编码问题"""
    if isinstance(s, str):
        # 替换不常见字符
        return s.replace('\xa0', ' ').replace('\u2028', ' ')
    return str(s)


def xls_to_xlsx(xls_path, xlsx_path=None):
    """
    将单个 .xls 文件转换为 .xlsx 格式
    """
    if xlsx_path is None:
        xlsx_path = str(Path(xls_path).with_suffix('.xlsx'))

    # 打开 .xls 文件
    workbook = xlrd.open_workbook(xls_path)

    # 创建新的 .xlsx 工作簿
    new_workbook = openpyxl.Workbook()
    new_workbook.remove(new_workbook.active)  # 删除默认 sheet

    for sheet_idx in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_idx)
        sheet_name = safe_str(sheet.name or f'Sheet{sheet_idx + 1}')[:31]

        # 创建新 sheet
        new_sheet = new_workbook.create_sheet(title=sheet_name)

        # 设置默认列宽
        for col_idx in range(sheet.ncols):
            new_sheet.column_dimensions[get_column_letter(col_idx + 1)].width = 15

        # 写入单元格数据
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                new_cell = new_sheet.cell(row=row_idx + 1, column=col_idx + 1)

                # 复制值
                value = cell.value
                cell_type = cell.ctype

                if cell_type == xlrd.XL_CELL_DATE:
                    # 处理日期
                    try:
                        date_tuple = xlrd.xldate_as_tuple(value, workbook.datemode)
                        if date_tuple[0] == 0 and date_tuple[1] == 0:
                            # 时间
                            new_cell.value = datetime.time(*date_tuple[3:])
                            new_cell.number_format = 'HH:MM:SS'
                        else:
                            # 日期或日期时间
                            if date_tuple[3:] == (0, 0, 0):
                                new_cell.value = datetime.date(*date_tuple[:3])
                                new_cell.number_format = 'YYYY-MM-DD'
                            else:
                                new_cell.value = datetime(*date_tuple)
                                new_cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                    except:
                        new_cell.value = safe_str(value)
                elif cell_type == xlrd.XL_CELL_BOOLEAN:
                    new_cell.value = bool(value)
                elif cell_type == xlrd.XL_CELL_ERROR:
                    new_cell.value = xlrd.error_text.get(value, f'#{value}!')
                elif cell_type == xlrd.XL_CELL_BLANK:
                    pass  # 保持为空
                else:
                    new_cell.value = safe_str(value) if isinstance(value, str) else value

                # 复制基本格式
                xf_idx = cell.xf_index
                if xf_idx is not None and xf_idx < len(workbook.xf_list):
                    xf = workbook.xf_list[xf_idx]
                    font = workbook.font_list[xf.font_index]

                    # 字体
                    font_kwargs = {'name': 'Arial', 'size': 11}
                    if font.bold:
                        font_kwargs['bold'] = True
                    if font.height and font.height > 0:
                        font_kwargs['size'] = font.height / 20
                    if font.name:
                        font_kwargs['name'] = safe_str(font.name)

                    new_cell.font = Font(**font_kwargs)

                    # 对齐
                    alignment = xf.alignment
                    hor_align = alignment.hor_align
                    h_align = 'left'
                    if hor_align == 1:
                        h_align = 'left'
                    elif hor_align == 2:
                        h_align = 'center'
                    elif hor_align == 3:
                        h_align = 'right'
                    new_cell.alignment = Alignment(horizontal=h_align)

                    # 数字格式
                    if xf.format_key > 0:
                        try:
                            format_str = workbook.format_map.get(xf.format_key, None)
                            if format_str:
                                new_cell.number_format = safe_str(format_str.format_str)
                        except:
                            pass

    # 保存
    new_workbook.save(xlsx_path)
    try:
        print(f'[OK] {os.path.basename(xls_path)} -> {os.path.basename(xlsx_path)}')
    except:
        print(f'[OK] [file converted successfully]')
    return True


def batch_convert(input_dir, output_dir=None):
    """
    批量转换目录下的所有 .xls 文件
    """
    input_path = Path(input_dir)
    if output_dir is None:
        output_dir = input_path  # 保存到原目录
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

    xls_files = list(input_path.rglob('*.xls'))
    if not xls_files:
        print(f'在 {input_path} 中未找到 .xls 文件')
        return

    print(f'找到 {len(xls_files)} 个 .xls 文件\n')

    success = 0
    failed = []

    for xls_file in xls_files:
        rel_path = xls_file.relative_to(input_path)
        output_file = output_dir / rel_path.with_suffix('.xlsx')

        # 创建输出目录
        output_file.parent.mkdir(parents=True, exist_ok=True)

        try:
            xls_to_xlsx(str(xls_file), str(output_file))
            success += 1
        except Exception as e:
            try:
                print(f'[FAIL] {xls_file} - {e}')
            except:
                print(f'[FAIL] [conversion failed]')
            failed.append((str(xls_file), str(e)))

    print(f'\n{"="*50}')
    print(f'完成: 成功 {success}/{len(xls_files)}')
    if failed:
        print(f'失败 {len(failed)} 个')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='将 .xls 文件转换为 .xlsx 格式')
    parser.add_argument('input', help='输入 .xls 文件或包含 .xls 文件的目录')
    parser.add_argument('output', nargs='?', help='输出 .xlsx 文件或输出目录 (批量模式)')
    parser.add_argument('--batch', action='store_true', help='批量转换模式')

    args = parser.parse_args()

    input_path = Path(args.input)

    if args.batch or input_path.is_dir():
        batch_convert(args.input, args.output)
    else:
        if not input_path.exists():
            print(f'文件不存在: {input_path}')
            sys.exit(1)
        xls_to_xlsx(str(input_path), args.output)
