#!/usr/bin/env python3
"""
Content Plan Excel Generator
Supports multi-language: Chinese, English, German, French, Spanish, etc.
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

LANGUAGE_MAP = {
    "en": {
        "sheet_keywords": "Keywords",
        "sheet_competitors": "Competitors",
        "sheet_plan": "Content Plan",
        "title_keywords": "Industry Keywords Map",
        "title_competitors": "Competitor Analysis",
        "title_plan": "3-Month Content Plan",
        "headers_keywords": ["Category", "Keyword", "Search Intent", "Priority"],
        "headers_competitors": ["Competitor", "Main Topics", "Frequency", "Format", "Advantage"],
        "headers_plan": ["Month", "Week", "Theme", "Topic", "Content Type", "Target Keywords"],
        "month_prefix": "Month ",
        "month_suffix": "",
        "week_prefix": "Week ",
        "week_suffix": "",
    },
    "zh": {
        "sheet_keywords": "Keywords",
        "sheet_competitors": "Competitors",
        "sheet_plan": "Content Plan",
        "title_keywords": "Industry Keywords Map",
        "title_competitors": "Competitor Analysis",
        "title_plan": "3-Month Content Plan",
        "headers_keywords": ["Category", "Keyword", "Search Intent", "Priority"],
        "headers_competitors": ["Competitor", "Main Topics", "Frequency", "Format", "Advantage"],
        "headers_plan": ["Month", "Week", "Theme", "Topic", "Content Type", "Target Keywords"],
        "month_prefix": "Month ",
        "month_suffix": "",
        "week_prefix": "Week ",
        "week_suffix": "",
    },
    "de": {
        "sheet_keywords": "Keywords",
        "sheet_competitors": "Wettbewerber",
        "sheet_plan": "Inhaltsplan",
        "title_keywords": "Branchenschlüsselwörter",
        "title_competitors": "Wettbewerbsanalyse",
        "title_plan": "3-Monatiger Inhaltsplan",
        "headers_keywords": ["Kategorie", "Schlüsselwort", "Suchintention", "Priorität"],
        "headers_competitors": ["Wettbewerber", "Hauptthemen", "Häufigkeit", "Format", "Vorteil"],
        "headers_plan": ["Monat", "Woche", "Thema", "Titel", "Inhaltstyp", "Ziel-Keywords"],
        "month_prefix": "Monat ",
        "month_suffix": "",
        "week_prefix": "Woche ",
        "week_suffix": "",
    },
    "fr": {
        "sheet_keywords": "Mots-clés",
        "sheet_competitors": "Concurrents",
        "sheet_plan": "Plan de contenu",
        "title_keywords": "Carte des mots-clés",
        "title_competitors": "Analyse des concurrents",
        "title_plan": "Plan de contenu sur 3 mois",
        "headers_keywords": ["Catégorie", "Mot-clé", "Intention de recherche", "Priorité"],
        "headers_competitors": ["Concurrent", "Sujets principaux", "Fréquence", "Format", "Avantage"],
        "headers_plan": ["Mois", "Semaine", "Thème", "Sujet", "Type de contenu", "Mots-clés cibles"],
        "month_prefix": "Mois ",
        "month_suffix": "",
        "week_prefix": "Semaine ",
        "week_suffix": "",
    },
    "es": {
        "sheet_keywords": "Palabras clave",
        "sheet_competitors": "Competidores",
        "sheet_plan": "Plan de contenido",
        "title_keywords": "Mapa de palabras clave",
        "title_competitors": "Análisis de competidores",
        "title_plan": "Plan de contenido de 3 meses",
        "headers_keywords": ["Categoría", "Palabra clave", "Intención de búsqueda", "Prioridad"],
        "headers_competitors": ["Competidor", "Temas principales", "Frecuencia", "Formato", "Ventaja"],
        "headers_plan": ["Mes", "Semana", "Tema", "Título", "Tipo de contenido", "Palabras clave objetivo"],
        "month_prefix": "Mes ",
        "month_suffix": "",
        "week_prefix": "Semana ",
        "week_suffix": "",
    },
}

def get_lang_config(lang_code):
    """Get language configuration"""
    return LANGUAGE_MAP.get(lang_code, LANGUAGE_MAP["en"])

def format_month_week(month_num, week_num, config):
    """Format month and week numbers"""
    month = f"{config['month_prefix']}{month_num}{config['month_suffix']}"
    week = f"{config['week_prefix']}{week_num}{config['week_suffix']}"
    return month, week

def create_excel(analysis_data, output_file="content_plan.xlsx", lang="en"):
    """Create Excel content plan"""
    config = get_lang_config(lang)
    wb = Workbook()
    
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Keywords
    ws1 = wb.active
    ws1.title = config["sheet_keywords"]
    
    keywords = analysis_data.get("keywords", {})
    
    ws1["A1"] = config["title_keywords"]
    ws1["A1"].font = Font(bold=True, size=16)
    ws1.merge_cells("A1:D1")
    
    for col, header in enumerate(config["headers_keywords"], 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 4
    for category, words in keywords.items():
        for word in words:
            ws1.cell(row=row, column=1, value=category).border = thin_border
            ws1.cell(row=row, column=2, value=word.get("keyword", "")).border = thin_border
            ws1.cell(row=row, column=3, value=word.get("intent", "")).border = thin_border
            ws1.cell(row=row, column=4, value=word.get("priority", "")).border = thin_border
            row += 1
    
    for col in range(1, 5):
        ws1.column_dimensions[get_column_letter(col)].width = 22
    
    # Sheet 2: Competitors
    ws2 = wb.create_sheet(config["sheet_competitors"])
    
    ws2["A1"] = config["title_competitors"]
    ws2["A1"].font = Font(bold=True, size=16)
    ws2.merge_cells("A1:E1")
    
    competitors = analysis_data.get("competitors", [])
    
    for col, header in enumerate(config["headers_competitors"], 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 4
    for comp in competitors:
        ws2.cell(row=row, column=1, value=comp.get("name", "")).border = thin_border
        ws2.cell(row=row, column=2, value=comp.get("topics", "")).border = thin_border
        ws2.cell(row=row, column=3, value=comp.get("frequency", "")).border = thin_border
        ws2.cell(row=row, column=4, value=comp.get("format", "")).border = thin_border
        ws2.cell(row=row, column=5, value=comp.get("advantage", "")).border = thin_border
        row += 1
    
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 24
    
    # Sheet 3: Content Plan
    ws3 = wb.create_sheet(config["sheet_plan"])
    
    ws3["A1"] = config["title_plan"]
    ws3["A1"].font = Font(bold=True, size=16)
    ws3.merge_cells("A1:F1")
    
    plan = analysis_data.get("content_plan", [])
    
    for col, header in enumerate(config["headers_plan"], 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 4
    for item in plan:
        month_val = item.get("month", "")
        week_val = item.get("week", "")
        
        if isinstance(month_val, int):
            if isinstance(week_val, int):
                month_str, week_str = format_month_week(month_val, week_val, config)
            else:
                month_str = f"{config['month_prefix']}{month_val}{config['month_suffix']}"
                week_str = week_val
        else:
            month_str = month_val
            week_str = week_val
        
        ws3.cell(row=row, column=1, value=month_str).border = thin_border
        ws3.cell(row=row, column=2, value=week_str).border = thin_border
        ws3.cell(row=row, column=3, value=item.get("theme", "")).border = thin_border
        ws3.cell(row=row, column=4, value=item.get("topic", "")).border = thin_border
        ws3.cell(row=row, column=5, value=item.get("type", "")).border = thin_border
        ws3.cell(row=row, column=6, value=item.get("keywords", "")).border = thin_border
        row += 1
    
    for col in range(1, 7):
        ws3.column_dimensions[get_column_letter(col)].width = 25
    
    wb.save(output_file)
    print(f"Excel generated: {output_file}")

def main():
    input_file = "analysis_result.json"
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    
    output_file = "content_plan.xlsx"
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    lang = "en"
    if len(sys.argv) > 3:
        lang = sys.argv[3]
    
    try:
        with open(input_file, 'r', encoding='utf-8') as f:
            analysis_data = json.load(f)
    except FileNotFoundError:
        analysis_data = {
            "keywords": {
                "Core Keywords": [
                    {"keyword": "sample keyword 1", "intent": "purchase intent", "priority": "high"}
                ]
            },
            "competitors": [],
            "content_plan": []
        }
    
    create_excel(analysis_data, output_file, lang)

if __name__ == "__main__":
    main()
