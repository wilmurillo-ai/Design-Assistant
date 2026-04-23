#!/usr/bin/env python3
"""
Script de recalcul des formules Excel
Recalcule toutes les formules d'un fichier Excel via LibreOffice
Adapté pour OpenClawd
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

try:
    from office.soffice import get_soffice_env
except ImportError:
    # Fallback si le module office n'est pas disponible
    def get_soffice_env():
        env = os.environ.copy()
        env["SAL_USE_VCLPLUGIN"] = "svp"
        return env

try:
    from openpyxl import load_workbook
except ImportError:
    print("Erreur: openpyxl non installé. Exécuter: pip install openpyxl")
    sys.exit(1)

# Répertoire macro LibreOffice selon plateforme
MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "Module1.xba"

# Macro LibreOffice Basic pour recalcul
RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def has_gtimeout():
    """Vérifie si gtimeout est disponible sur macOS"""
    try:
        subprocess.run(
            ["gtimeout", "--version"], capture_output=True, timeout=1, check=False
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def setup_libreoffice_macro():
    """Configure la macro LibreOffice si pas déjà fait"""
    macro_dir = os.path.expanduser(
        MACRO_DIR_MACOS if platform.system() == "Darwin" else MACRO_DIR_LINUX
    )
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    # Vérifier si macro existe déjà
    if (
        os.path.exists(macro_file)
        and "RecalculateAndSave" in Path(macro_file).read_text()
    ):
        return True

    # Créer répertoire macro si nécessaire
    if not os.path.exists(macro_dir):
        try:
            subprocess.run(
                ["soffice", "--headless", "--terminate_after_init"],
                capture_output=True,
                timeout=10,
                env=get_soffice_env(),
            )
        except Exception:
            pass
        os.makedirs(macro_dir, exist_ok=True)

    # Écrire fichier macro
    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """
    Recalcule les formules d'un fichier Excel et rapporte les erreurs

    Args:
        filename: Chemin vers le fichier Excel
        timeout: Temps max d'attente pour le recalcul (secondes)

    Returns:
        dict avec emplacements et compteurs d'erreurs
    """
    if not Path(filename).exists():
        return {"error": f"Fichier {filename} inexistant"}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {"error": "Échec configuration macro LibreOffice"}

    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    # Encapsuler avec timeout si disponible
    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin" and has_gtimeout():
        cmd = ["gtimeout", str(timeout)] + cmd

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env(), timeout=timeout+10)
    except subprocess.TimeoutExpired:
        return {"error": "Timeout lors du recalcul"}

    if result.returncode != 0 and result.returncode != 124:  # 124 = code timeout
        error_msg = result.stderr or "Erreur inconnue lors du recalcul"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "Macro LibreOffice mal configurée"}
        return {"error": error_msg}

    # Vérifier erreurs Excel dans le fichier recalculé
    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        # Construire résumé
        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        # Ajouter catégories d'erreurs non vides
        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],  # Max 20 emplacements affichés
                }

        # Ajouter compte de formules
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count

        return result

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <fichier_excel> [timeout_secondes]")
        print("\nRecalcule toutes les formules d'un fichier Excel via LibreOffice")
        print("\nRetourne JSON avec détails d'erreurs:")
        print("  - status: 'success' ou 'errors_found'")
        print("  - total_errors: Nombre total d'erreurs Excel")
        print("  - total_formulas: Nombre de formules dans le fichier")
        print("  - error_summary: Détail par type avec emplacements")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
