"""
失信主体名单本地黑名单模块

数据来源：
1. 国家税务总局"重大税收违法失信案件信息公布栏"
   → https://www.chinatax.gov.cn/ 每月15日发布上月数据（ZIP含Excel）
2. 信用中国（扩展数据源，后续接入）

本地 SQLite 存储，报销时自动比对销售方税号/企业名称。
"""

import re
import json
import logging
import sqlite3
import zipfile
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Generator

logger = logging.getLogger(__name__)

BLACKLIST_TABLE = "blacklist"


def get_conn(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def init_blacklist_db(db_path: str):
    """初始化黑名单数据库表"""
    with get_conn(db_path) as conn:
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS {BLACKLIST_TABLE} (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                -- 统一社会信用代码 / 纳税人识别号
                taxpayer_id    TEXT,
                -- 企业名称（原始）
                company_name   TEXT,
                -- 失信类型：走逃失踪/虚开发票/偷税/骗税/其他
                violation_type TEXT,
                -- 案件性质描述
                case_desc      TEXT,
                -- 公布日期
                publish_date   TEXT,
                -- 数据来源：chinatax / creditchina
                source         TEXT,
                -- 本地收录时间
                fetched_at     TEXT,
                UNIQUE(taxpayer_id, source)
            )
        """)
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_taxpayer_id ON {BLACKLIST_TABLE}(taxpayer_id)")
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_company_name ON {BLACKLIST_TABLE}(company_name)")
        conn.commit()
    logger.info(f"黑名单库初始化完成: {db_path}")


def is_blacklisted(db_path: str, taxpayer_id: str = None, company_name: str = None) -> Optional[dict]:
    """
    查询企业是否在黑名单中（优先按税号精确匹配，再按名称模糊匹配）
    返回命中的第一条记录，无匹配返回 None
    """
    if not taxpayer_id and not company_name:
        return None

    taxpayer_id_norm = _norm_id(taxpayer_id) if taxpayer_id else None

    with get_conn(db_path) as conn:
        rows = conn.execute(f"SELECT * FROM {BLACKLIST_TABLE}").fetchall()

    # 1. 税号精确匹配（最严格）
    if taxpayer_id_norm:
        for row in rows:
            if _norm_id(row["taxpayer_id"]) == taxpayer_id_norm:
                return dict(row)

    # 2. 名称模糊匹配（容忍轻微差异）
    if company_name:
        name_norm = _norm_company(company_name)
        for row in rows:
            if _norm_company(row["company_name"]) == name_norm:
                return dict(row)

    return None


def check_invoice_seller(db_path: str, seller: str, seller_tax_id: str = None) -> dict:
    """
    报销时调用：对销售方做黑名单检查
    返回 {'hit': bool, 'record': dict or None}
    """
    hit = is_blacklisted(db_path, seller_tax_id, seller)
    return {"hit": hit is not None, "record": hit}


# ============================================================
# 税务局 Excel 数据获取
# ============================================================

def fetch_chinatax_blacklist(
    year: int = None,
    month: int = None,
    cache_dir: Path = None
) -> List[dict]:
    """
    下载并解析国家税务总局"重大税收违法失信案件信息公布栏"数据

    税务总局每月15日左右发布上月数据，格式为 ZIP 内含 Excel：
    https://www.chinatax.gov.cn/ 搜索"重大税收违法失信"

    参数:
        year/month: 指定年月，默认为上月
        cache_dir: 下载缓存目录，避免重复拉取

    返回: [{taxpayer_id, company_name, violation_type, case_desc, publish_date}]
    """
    import httpx
    import io

    if year is None or month is None:
        from datetime import date, timedelta
        d = date.today().replace(day=1) - timedelta(days=1)  # 上月最后一天
        year = year or d.year
        month = month or d.month

    records: List[dict] = []
    fetched_at = datetime.now().isoformat()

    # 尝试从多个来源获取
    sources = _chinatax_urls(year, month)
    for url, source_name in sources:
        try:
            logger.info(f"尝试下载黑名单: {url[:60]}")
            resp = httpx.get(url, timeout=60, follow_redirects=True)
            if resp.status_code != 200:
                continue

            content = resp.content
            if url.endswith(".zip") or "zip" in resp.headers.get("content-type", ""):
                records = _parse_zip(content, source_name, fetched_at)
            elif url.endswith(".xlsx") or url.endswith(".xls"):
                records = _parse_excel_bytes(content, source_name, fetched_at)
            else:
                records = _parse_html_list(resp.text, source_name, fetched_at)

            if records:
                logger.info(f"✅ 从 {source_name} 获取到 {len(records)} 条黑名单记录")
                break
        except Exception as e:
            logger.warning(f"下载失败 {url[:50]}: {e}")
            continue

    return records


def _chinatax_urls(year: int, month: int) -> List[tuple]:
    """
    生成当月可能的下载地址列表
    税务局文件名格式不固定，尝试多个路径
    """
    ym = f"{year}{month:02d}"
    ym_cn = f"{year}年{month}月"

    base = "https://www.chinatax.gov.cn"
    urls = [
        # 可能的路径（税务局结构有时会变）
        (f"{base}/chinatax/{ym}/soft/index_{ym}.zip", "chinatax_official"),
        (f"{base}/chinatax/{ym}/upload/{ym}_blacklist.xlsx", "chinatax"),
        # 通用搜索页（无法直接爬列表，只能手动维护URL）
    ]
    return urls


def _parse_zip(zip_bytes: bytes, source: str, fetched_at: str) -> List[dict]:
    """解析 ZIP 包内的 Excel 文件"""
    import io
    records = []
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            xlsx_files = [f for f in zf.namelist() if f.endswith((".xlsx", ".xls"))]
            for xlsx_name in xlsx_files:
                try:
                    with zf.open(xlsx_name) as f:
                        data = f.read()
                    records.extend(_parse_excel_bytes(data, source, fetched_at))
                except Exception as e:
                    logger.warning(f"解析 {xlsx_name} 失败: {e}")
    except Exception as e:
        logger.error(f"ZIP 解析失败: {e}")
    return records


def _parse_excel_bytes(data: bytes, source: str, fetched_at: str) -> List[dict]:
    """解析 Excel 内容为黑名单记录"""
    import io
    records = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # 尝试自动识别表头行
            header_map = _detect_header(rows[0])
            for row in rows[1:]:
                if not row or not any(row):
                    continue
                rec = _row_to_record(row, header_map, source, fetched_at)
                if rec:
                    records.append(rec)
    except Exception as e:
        logger.warning(f"Excel 解析失败: {e}")
    return records


def _detect_header(header_row: tuple) -> dict:
    """
    根据表头文字智能识别列索引
    税务局Excel表头可能有多种写法
    """
    if not header_row:
        return {}
    mapping = {}
    keywords = {
        "taxpayer_id": ["纳税人识别号", "税号", "统一社会信用代码", "信用代码", "纳税人编号"],
        "company_name": ["企业名称", "纳税人名称", "当事人", "公司名称", "名称"],
        "violation_type": ["失信类型", "违法行为", "案件性质", "违法类型", "类型"],
        "case_desc": ["案情描述", "主要违法事实", "处理情况", "详情"],
        "publish_date": ["公布日期", "发布日期", "日期", "公示日期"],
    }
    for col_idx, cell in enumerate(header_row):
        if not cell:
            continue
        cell_str = str(cell).strip()
        for field, aliases in keywords.items():
            if any(alias in cell_str for alias in aliases):
                mapping[field] = col_idx
    return mapping


def _row_to_record(row: tuple, header: dict, source: str, fetched_at: str) -> Optional[dict]:
    """将一行 Excel 数据转换为黑名单记录"""
    get = lambda f: row[header[f]] if header.get(f) is not None and header[f] < len(row) else None

    taxpayer_id = _clean(get("taxpayer_id"))
    company_name = _clean(get("company_name"))
    violation_type = _clean(get("violation_type")) or "未知"
    case_desc = _clean(get("case_desc"))
    publish_date = _normalize_date(str(get("publish_date") or ""))

    if not taxpayer_id and not company_name:
        return None

    return {
        "taxpayer_id": taxpayer_id,
        "company_name": company_name,
        "violation_type": violation_type,
        "case_desc": case_desc,
        "publish_date": publish_date,
        "source": source,
        "fetched_at": fetched_at,
    }


def _parse_html_list(html: str, source: str, fetched_at: str) -> List[dict]:
    """
    从 HTML 页面解析黑名单列表（备用方案）
    税务局网站的表格结构相对固定
    """
    records = []
    try:
        import re
        # 匹配表格行
        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.DOTALL | re.IGNORECASE)
        for row in rows[1:]:  # 跳过表头
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row, re.DOTALL | re.IGNORECASE)
            if len(cells) < 3:
                continue
            # 清理 HTML 标签
            clean_cells = [re.sub(r"<[^>]+>", "", c).strip() for c in cells]
            rec = {
                "taxpayer_id": _clean(clean_cells[0]) if len(clean_cells) > 0 else None,
                "company_name": _clean(clean_cells[1]) if len(clean_cells) > 1 else None,
                "violation_type": _clean(clean_cells[2]) if len(clean_cells) > 2 else "未知",
                "case_desc": _clean(clean_cells[3]) if len(clean_cells) > 3 else None,
                "publish_date": _normalize_date(clean_cells[4] if len(clean_cells) > 4 else ""),
                "source": source,
                "fetched_at": fetched_at,
            }
            if rec["taxpayer_id"] or rec["company_name"]:
                records.append(rec)
    except Exception as e:
        logger.warning(f"HTML 解析失败: {e}")
    return records


# ============================================================
# 同步到本地数据库
# ============================================================

def sync_blacklist(db_path: str, records: List[dict], replace: bool = False) -> dict:
    """
    将获取的黑名单记录写入本地数据库
    - replace=True: 清空旧数据后重新写入
    - replace=False: 仅新增/更新（upsert）
    """
    if not records:
        return {"added": 0, "updated": 0, "skipped": 0}

    if replace:
        with get_conn(db_path) as conn:
            conn.execute(f"DELETE FROM {BLACKLIST_TABLE}")
        logger.info("黑名单表已清空，将重新写入")

    added = updated = skipped = 0
    with get_conn(db_path) as conn:
        for rec in records:
            try:
                cur = conn.execute(f"""
                    INSERT OR REPLACE INTO {BLACKLIST_TABLE}
                    (taxpayer_id, company_name, violation_type, case_desc,
                     publish_date, source, fetched_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    rec.get("taxpayer_id"),
                    rec.get("company_name"),
                    rec.get("violation_type"),
                    rec.get("case_desc"),
                    rec.get("publish_date"),
                    rec.get("source", "unknown"),
                    rec.get("fetched_at"),
                ))
                rows_affected = cur.rowcount
                if rows_affected == 0:
                    skipped += 1
                else:
                    added += 1
            except Exception as e:
                skipped += 1
                logger.debug(f"写入跳过: {e}")

        conn.commit()

    logger.info(f"黑名单同步完成: 新增/更新={added}, 跳过={skipped}")
    return {"added": added, "updated": updated, "skipped": skipped}


def get_blacklist_stats(db_path: str) -> dict:
    """返回黑名单库统计"""
    with get_conn(db_path) as conn:
        total = conn.execute(f"SELECT COUNT(*) FROM {BLACKLIST_TABLE}").fetchone()[0]
        by_type = conn.execute(f"""
            SELECT violation_type, COUNT(*) as cnt
            FROM {BLACKLIST_TABLE}
            GROUP BY violation_type
            ORDER BY cnt DESC
        """).fetchall()
        latest = conn.execute(f"""
            SELECT publish_date, COUNT(*) as cnt
            FROM {BLACKLIST_TABLE}
            GROUP BY publish_date
            ORDER BY publish_date DESC LIMIT 5
        """).fetchall()
    return {
        "total": total,
        "by_violation_type": [dict(r) for r in by_type],
        "by_publish_date": [dict(r) for r in latest],
    }


# ============================================================
# 工具函数
# ============================================================

def _norm_id(tax_id: str) -> str:
    """标准化税号：统一社会信用代码（18位）或纳税人识别号（15-20位）"""
    if not tax_id:
        return ""
    # 去除空格和特殊字符
    s = re.sub(r"[\s_\-()（）【】]", "", str(tax_id)).upper()
    return s


def _norm_company(name: str) -> str:
    """标准化企业名称：去括号内容、括号、全角转半角"""
    if not name:
        return ""
    s = str(name).strip()
    # 去全角括号及其内容
    s = re.sub(r"[（(][^）)]*[）)]", "", s)
    # 去除"有限公司""股份有限公司"等常见后缀便于比对
    for suffix in ["有限公司", "股份有限公司", "有限责任公司",
                    "集团有限公司", "控股有限公司", "投资有限公司"]:
        if s.endswith(suffix):
            s = s[:-len(suffix)]
    # 全角转半角
    s = s.translate(str.maketrans("（）", "( )"))
    return re.sub(r"\s+", "", s).upper()


def _clean(val) -> Optional[str]:
    """清理单元格值"""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s not in ("无", "空", "-", "—", "NULL", "None", "") else None


def _normalize_date(val: str) -> str:
    """统一日期格式为 YYYY-MM-DD"""
    if not val:
        return ""
    # 尝试各种日期格式
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y.%m.%d",
                "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # 正则提取
    m = re.search(r"(\d{4})[年/\-.](%d{1,2})[月/\-.](%d{1,2})", val)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return val.strip()
