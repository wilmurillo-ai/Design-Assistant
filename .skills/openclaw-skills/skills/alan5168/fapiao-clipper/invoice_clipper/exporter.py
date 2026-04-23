"""
导出模块
- 生成 Excel 明细表（开票日期排序）
- 合并 PDF 报销包
- 复制文件到指定文件夹
"""
import logging
import shutil
from pathlib import Path
from typing import List, Optional
from datetime import datetime

logger = logging.getLogger(__name__)


def export_excel(invoices: List[dict], output_path: Path) -> Path:
    """
    生成报销 Excel 明细表
    列：序号、开票日期、发票号码、销售方名称、购买方名称、价税合计（小写）
    按开票日期升序排列，最后一行合计
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "发票明细"

    # 标题行
    headers = ["序号", "开票日期", "发票号码", "销售方名称", "购买方名称", "价税合计（小写）"]
    col_widths = [6, 14, 22, 35, 35, 18]

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    # 数据行（按日期排序）
    sorted_inv = sorted(invoices, key=lambda x: x.get("date") or "")
    even_fill = PatternFill("solid", fgColor="DCE6F1")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    total_amount = 0.0

    for row_idx, inv in enumerate(sorted_inv, 2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        amount = inv.get("amount_with_tax") or 0
        total_amount += amount

        values = [
            row_idx - 1,
            inv.get("date") or "",
            inv.get("invoice_number") or "",
            inv.get("seller") or "",
            inv.get("buyer") or "",
            amount,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 6:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row_idx].height = 18

    # 合计行
    total_row = len(sorted_inv) + 2
    total_fill = PatternFill("solid", fgColor="F2F2F2")
    total_font = Font(name="微软雅黑", bold=True, size=11)

    ws.merge_cells(f"A{total_row}:E{total_row}")
    total_label = ws.cell(row=total_row, column=1, value=f"合计（共 {len(sorted_inv)} 张）")
    total_label.font = total_font
    total_label.fill = total_fill
    total_label.alignment = Alignment(horizontal="center", vertical="center")
    total_label.border = border

    total_cell = ws.cell(row=total_row, column=6, value=total_amount)
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.number_format = "#,##0.00"
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.border = border
    ws.row_dimensions[total_row].height = 22

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(str(output_path))
    logger.info(f"Excel 导出完成: {output_path}，共 {len(sorted_inv)} 张，合计 {total_amount:.2f}")
    return output_path


def export_pdf_folder(invoices: List[dict], output_dir: Path) -> Path:
    """将发票文件复制到指定文件夹（不修改原文件）"""
    output_dir.mkdir(parents=True, exist_ok=True)
    count = 0
    for inv in sorted(invoices, key=lambda x: x.get("date") or ""):
        src = Path(inv.get("stored_path") or "")
        if not src.exists():
            logger.warning(f"文件不存在，跳过: {src}")
            continue
        # 目标文件名加序号避免重名
        dest = output_dir / src.name
        if dest.exists():
            stem, suffix = src.stem, src.suffix
            for i in range(1, 100):
                dest = output_dir / f"{stem}_{i}{suffix}"
                if not dest.exists():
                    break
        shutil.copy2(str(src), str(dest))
        count += 1
    logger.info(f"发票文件夹导出完成: {output_dir}，共 {count} 张")
    return output_dir


def export_merged_pdf(invoices: List[dict], output_path: Path) -> Optional[Path]:
    """合并发票为单一 PDF"""
    import fitz
    output_path.parent.mkdir(parents=True, exist_ok=True)

    merged = fitz.open()
    count = 0
    for inv in sorted(invoices, key=lambda x: x.get("date") or ""):
        src = Path(inv.get("stored_path") or "")
        if not src.exists():
            logger.warning(f"文件不存在，跳过: {src}")
            continue
        try:
            doc = fitz.open(str(src))
            merged.insert_pdf(doc)
            doc.close()
            count += 1
        except Exception as e:
            logger.warning(f"合并跳过 {src.name}: {e}")

    if count == 0:
        logger.warning("没有可合并的文件")
        return None

    merged.save(str(output_path))
    merged.close()
    logger.info(f"合并 PDF 完成: {output_path}，共 {count} 张")
    return output_path


def build_export_label(filters: dict) -> str:
    """根据筛选条件生成导出文件名标签"""
    parts = []
    if filters.get("date_from") and filters.get("date_to"):
        parts.append(f"{filters['date_from']}至{filters['date_to']}")
    elif filters.get("date_from"):
        parts.append(f"{filters['date_from']}起")
    elif filters.get("date_to"):
        parts.append(f"至{filters['date_to']}")
    if filters.get("buyer"):
        parts.append(f"购买方{filters['buyer']}")
    if filters.get("seller"):
        parts.append(f"销售方{filters['seller']}")
    if not parts:
        from datetime import datetime
        parts.append(datetime.now().strftime("%Y%m%d"))
    return "_".join(parts)


# ============================================================
# 问题发票独立导出
# ============================================================

def _parse_warnings(warnings_json: str) -> list:
    """安全解析 warnings JSON 字段"""
    import json
    try:
        return json.loads(warnings_json) if warnings_json else []
    except Exception:
        return []


def _warning_summary(warnings: list) -> str:
    """将警告列表汇总为一行文字"""
    if not warnings:
        return "正常"
    return " / ".join(w["code"] for w in warnings)


def export_problem_invoices(db_path: str, filters: dict, output_dir: Path) -> dict:
    """
    导出问题发票独立清单（单独Sheet，红色高亮）
    返回：{excel_path, problem_count, problem_amount}
    """
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from invoice_clipper.database import get_problem_invoices

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"问题发票清单_{timestamp}.xlsx"

    problems = get_problem_invoices(db_path, filters)

    wb = Workbook()

    # Sheet 1: 问题发票汇总
    ws_sum = wb.active
    ws_sum.title = "问题发票"

    headers = ["序号", "问题级别", "问题代码", "问题描述", "开票日期",
               "销售方", "购买方", "金额（含税）", "发票号码", "原始文件"]
    col_widths = [6, 10, 20, 50, 14, 30, 30, 16, 22, 30]

    # 红色标题
    header_fill = PatternFill("solid", fgColor="C00000")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws_sum.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws_sum.column_dimensions[get_column_letter(col)].width = w
    ws_sum.row_dimensions[1].height = 22

    # 按级别分颜色
    block_fill = PatternFill("solid", fgColor="FFE0E0")  # 红色：禁止
    warn_fill = PatternFill("solid", fgColor="FFFACD")   # 黄色：警告
    level_font = {"BLOCK": Font(color="C00000", bold=True), "WARN": Font(color="B8860B", bold=True)}

    total_problem_amount = 0.0
    for row_idx, inv in enumerate(sorted(problems, key=lambda x: x.get("date") or ""), 2):
        warnings = _parse_warnings(inv.get("warnings") or "[]")
        if not warnings:
            warnings = [{"level": "WARN", "code": "UNKNOWN", "message": "有风险标记"}]
        top_warning = warnings[0]
        fill = block_fill if top_warning["level"] == "BLOCK" else warn_fill
        f = level_font.get(top_warning["level"], Font())

        amount = inv.get("amount_with_tax") or 0
        total_problem_amount += amount

        row_data = [
            row_idx - 1,
            "🚫 禁止报销" if top_warning["level"] == "BLOCK" else "⚠️ 需确认",
            top_warning.get("code", ""),
            top_warning.get("message", ""),
            inv.get("date", ""),
            inv.get("seller", ""),
            inv.get("buyer", ""),
            amount,
            inv.get("invoice_number", ""),
            inv.get("original_filename", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 8:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
        # 级别列特殊字体
        ws_sum.cell(row=row_idx, column=2).font = f
        ws_sum.row_dimensions[row_idx].height = 30

    # 合计行
    total_row = len(problems) + 2
    total_fill = PatternFill("solid", fgColor="F2F2F2")
    total_font = Font(name="微软雅黑", bold=True, size=11)
    ws_sum.merge_cells(f"A{total_row}:G{total_row}")
    ws_sum.cell(row=total_row, column=1, value=f"合计（共 {len(problems)} 张问题发票，可报销金额 ¥{total_problem_amount:.2f}）").font = total_font
    ws_sum.cell(row=total_row, column=1).fill = total_fill
    ws_sum.cell(row=total_row, column=1).border = border
    ws_sum.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.cell(row=total_row, column=8, value=total_problem_amount)
    ws_sum.cell(row=total_row, column=8).number_format = "#,##0.00"
    ws_sum.cell(row=total_row, column=8).fill = total_fill
    ws_sum.cell(row=total_row, column=8).font = total_font
    ws_sum.cell(row=total_row, column=8).border = border
    ws_sum.cell(row=total_row, column=8).alignment = Alignment(horizontal="right", vertical="center")
    ws_sum.row_dimensions[total_row].height = 22

    ws_sum.freeze_panes = "A2"

    # Sheet 2: 所有警告详情
    ws_detail = wb.create_sheet("警告详情")
    detail_headers = ["发票ID", "问题级别", "问题代码", "问题描述", "开票日期", "销售方", "金额"]
    for col, h in enumerate(detail_headers, 1):
        c = ws_detail.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    row_d = 2
    for inv in problems:
        warnings = _parse_warnings(inv.get("warnings") or "[]")
        for w in warnings:
            fill = block_fill if w["level"] == "BLOCK" else warn_fill
            for col, val in enumerate([inv["id"], w["level"], w["code"], w["message"],
                                        inv.get("date", ""), inv.get("seller", ""),
                                        inv.get("amount_with_tax") or 0], 1):
                c = ws_detail.cell(row=row_d, column=col, value=val)
                c.fill = fill
                c.border = border
                if col == 7:
                    c.number_format = "#,##0.00"
            row_d += 1

    for col, w in enumerate([8, 10, 25, 55, 14, 30, 12], 1):
        ws_detail.column_dimensions[get_column_letter(col)].width = w

    wb.save(str(output_path))
    logger.info(f"问题发票清单导出: {output_path}，共 {len(problems)} 张")

    return {
        "path": output_path,
        "count": len(problems),
        "total_amount": total_problem_amount,
    }

