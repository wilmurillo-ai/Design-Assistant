import os
import sys
import subprocess

# --- 🛠️ 自动依赖检测与安装 ---
def install_and_import(package_name, import_name=None):
    """
    检测并安装依赖库。
    :param package_name: pip 安装时的包名 (如 openpyxl)
    :param import_name: import 时的模块名 (通常与包名相同，不同时需指定)
    """
    if import_name is None:
        import_name = package_name
        
    try:
        __import__(import_name)
    except ImportError:
        print(f"⚠️ 正在自动安装缺少的核心库: {package_name} ...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
            print(f"✅ {package_name} 安装成功！")
        except Exception as e:
            print(f"❌ 安装失败: {e}")
            sys.exit(1)

# 检测 openpyxl
install_and_import("openpyxl")

# --- 核心逻辑引入 ---
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils import get_column_letter

def create_excel_file(filename: str, sheet_name: str, data: list) -> dict:
    """
    创建 Excel 文件并写入数据（支持样式）。
    
    :param filename: 文件名 (如 'sales_report.xlsx')
    :param sheet_name: 工作表名称 (如 'Q1数据')
    :param data: 二维列表。每个元素代表一行。
                 行中的每个单元格可以是：
                 1. 基础值 (str, int, float) -> 直接写入
                 2. 字典 (dict) -> 支持样式配置
                    格式: {"value": "内容", "bold": True, "color": "FF0000", "bg_color": "FFFF00", "align": "center"}
    """
    try:
        # 1. 文件名处理
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # 2. 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name if sheet_name else "Sheet1"

        # 3. 遍历写入数据
        for row_idx, row_data in enumerate(data, 1): # Excel 行从 1 开始
            for col_idx, cell_data in enumerate(row_data, 1): # Excel 列从 1 开始
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # 情况 A: 简单数据 (字符串/数字)
                if not isinstance(cell_data, dict):
                    cell.value = cell_data
                
                # 情况 B: 复杂样式数据 (字典)
                else:
                    cell.value = cell_data.get("value", "")
                    
                    # --- 样式处理 ---
                    # 字体样式 (加粗、颜色)
                    is_bold = cell_data.get("bold", False)
                    font_color = cell_data.get("color", "000000") # 默认黑色
                    # 清洗颜色代码 (移除 #)
                    if font_color and font_color.startswith("#"): font_color = font_color[1:]
                    
                    cell.font = Font(bold=is_bold, color=font_color)

                    # 背景填充 (bg_color)
                    bg_color = cell_data.get("bg_color", None)
                    if bg_color:
                        if bg_color.startswith("#"): bg_color = bg_color[1:]
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

                    # 对齐方式 (align: left, center, right)
                    align_type = cell_data.get("align", "left").lower()
                    if align_type in ["center", "right", "left"]:
                        cell.alignment = Alignment(horizontal=align_type, vertical="center")

        # 4. 自动调整列宽 (简单的自适应逻辑)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # 获取列号字符 (A, B, C...)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 50) # 限制最大宽度

        # 5. 保存文件
        wb.save(filename)
        abs_path = os.path.abspath(filename)
        
        return {
            "status": "success",
            "message": "Excel 文件创建成功",
            "file_path": abs_path,
            "filename": filename
        }

    except Exception as e:
        return {
            "status": "error",
            "message": f"创建 Excel 失败: {str(e)}"
        }

# --- 🧪 本地测试入口 ---
if __name__ == "__main__":
    # 定义测试数据
    test_filename = "财务报表.xlsx"
    test_data = [
        # 第一行：表头 (加粗，白色文字，蓝色背景，居中)
        [
            {"value": "月份", "bold": True, "color": "FFFFFF", "bg_color": "4472C4", "align": "center"},
            {"value": "收入 (元)", "bold": True, "color": "FFFFFF", "bg_color": "4472C4", "align": "center"},
            {"value": "支出 (元)", "bold": True, "color": "FFFFFF", "bg_color": "4472C4", "align": "center"},
            {"value": "利润状态", "bold": True, "color": "FFFFFF", "bg_color": "4472C4", "align": "center"}
        ],
        # 第二行：数据
        ["1月", 50000, 30000, {"value": "盈利", "color": "008000", "bold": True}],
        # 第三行：数据
        ["2月", 20000, 25000, {"value": "亏损", "color": "FF0000", "bold": True}],
        # 第四行：普通数据
        ["3月", 60000, 40000, "盈利"]
    ]

    result = create_excel_file(test_filename, "2024第一季度", test_data)
    print(result)
