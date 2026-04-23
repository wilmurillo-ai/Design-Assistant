#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create an Excel spreadsheet (.xlsx)")
    parser.add_argument("--data", help="Path to JSON data file")
    parser.add_argument("--template", choices=["budget", "invoice"], help="Use a predefined template")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--force", action="store_true", help="Overwrite output file if it already exists")
    return parser.parse_args()


def ensure_output_path(output_path: Path, force: bool) -> None:
    if output_path.exists() and not force:
        raise FileExistsError(f"Output file already exists: {output_path}. Use --force to overwrite.")
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError("Output file must end with .xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise ValueError("Data JSON must be an object at the top level")
    return data


def style_header_row(worksheet, row_number: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=row_number, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(worksheet) -> None:
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(value) + 2)
    for col_idx, width in widths.items():
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 40)


def create_budget_template(worksheet) -> None:
    worksheet.title = "Monthly Budget"
    headers = ["Category", "Budgeted", "Actual", "Difference"]
    categories = ["Housing", "Food", "Transportation", "Utilities", "Insurance", "Entertainment", "Savings"]
    for col, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col, value=header)
    style_header_row(worksheet, 1, len(headers))
    for row, category in enumerate(categories, start=2):
        worksheet.cell(row=row, column=1, value=category)
        worksheet.cell(row=row, column=2, value=0)
        worksheet.cell(row=row, column=3, value=0)
        worksheet.cell(row=row, column=4, value=f"=B{row}-C{row}")
    total_row = len(categories) + 2
    worksheet.cell(row=total_row, column=1, value="Total")
    worksheet.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})")
    worksheet.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})")
    worksheet.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row - 1})")
    style_header_row(worksheet, total_row, 4)

    chart = BarChart()
    chart.title = "Budget vs Actual"
    data = Reference(worksheet, min_col=2, max_col=3, min_row=1, max_row=total_row - 1)
    categories_ref = Reference(worksheet, min_col=1, min_row=2, max_row=total_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories_ref)
    worksheet.add_chart(chart, "F2")
    autosize_columns(worksheet)


def create_invoice_template(worksheet) -> None:
    worksheet.title = "Invoice"
    worksheet["A1"] = "INVOICE"
    worksheet["A1"].font = Font(bold=True, size=18)
    worksheet["A6"] = "Item"
    worksheet["B6"] = "Quantity"
    worksheet["C6"] = "Unit Price"
    worksheet["D6"] = "Line Total"
    style_header_row(worksheet, 6, 4)
    for row in range(7, 12):
        worksheet.cell(row=row, column=4, value=f"=B{row}*C{row}")
    worksheet["C13"] = "Subtotal"
    worksheet["D13"] = "=SUM(D7:D11)"
    worksheet["C14"] = "GST"
    worksheet["D14"] = "=D13*0.10"
    worksheet["C15"] = "Total"
    worksheet["D15"] = "=D13+D14"
    autosize_columns(worksheet)


def populate_sheet_from_data(worksheet, sheet_data: dict[str, Any]) -> None:
    headers = sheet_data.get("headers", [])
    rows = sheet_data.get("rows", [])
    title = sheet_data.get("title")
    if title:
        worksheet.title = str(title)[:31]
    if headers:
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)
        style_header_row(worksheet, 1, len(headers))
    start_row = 2 if headers else 1
    for row_index, row_values in enumerate(rows, start=start_row):
        for col_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)
    autosize_columns(worksheet)


def build_from_json(workbook: Workbook, data: dict[str, Any]) -> None:
    sheets = data.get("sheets")
    if sheets:
        workbook.remove(workbook.active)
        for sheet_data in sheets:
            worksheet = workbook.create_sheet()
            populate_sheet_from_data(worksheet, sheet_data)
    else:
        populate_sheet_from_data(workbook.active, data)


def main() -> int:
    args = parse_args()
    output_path = Path(args.output).expanduser().resolve()
    try:
        ensure_output_path(output_path, args.force)
        workbook = Workbook()
        worksheet = workbook.active
        if args.template == "budget":
            create_budget_template(worksheet)
        elif args.template == "invoice":
            create_invoice_template(worksheet)
        elif args.data:
            data = load_json(Path(args.data).expanduser().resolve())
            build_from_json(workbook, data)
        else:
            worksheet["A1"] = "No template or data provided"
        workbook.save(str(output_path))
        print(f"Created Excel workbook: {output_path}")
        return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
