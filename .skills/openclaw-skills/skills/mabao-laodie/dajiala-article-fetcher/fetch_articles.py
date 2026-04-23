#!/usr/bin/env python3
"""
大咖啦文章获取脚本
从大咖啦 API 获取微信公众号文章列表
"""

import os
import sys
import json
import requests
import openpyxl
from datetime import datetime

# 配置 - 使用用户指定的路径
BASE_DIR = "/home/admin/每日茶动态"
ACCOUNTS_FILE = os.path.join(BASE_DIR, "公众号清单.xlsx")
DAJIALA_KEY = os.environ.get('DAJIALA_KEY', '')
API_URL = "https://www.dajiala.com/fbmain/monitor/v3/post_condition"


def load_accounts_from_excel(file_path):
    """从 Excel 文件读取公众号清单"""
    if not os.path.exists(file_path):
        print(f"错误：公众号清单文件不存在: {file_path}")
        return []
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    accounts = []
    # 跳过表头，从第二行开始
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 公众号ID
            accounts.append({
                'id': str(row[0]).strip(),
                'name': str(row[1]).strip() if len(row) > 1 and row[1] else ''
            })
    
    return accounts


def fetch_articles(account_id, account_name):
    """从大咖啦 API 获取单个公众号的文章"""
    headers = {
        "Content-Type": "application/json"
    }
    
    payload = {
        "biz": "",
        "url": "",
        "name": account_id,
        "key": DAJIALA_KEY,
        "verifycode": ""
    }
    
    try:
        response = requests.post(API_URL, json=payload, headers=headers, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        articles = []
        if data.get('code') == 0 and data.get('data'):
            for item in data.get('data', []):
                articles.append({
                    'account_name': account_name,
                    'title': item.get('title', ''),
                    'url': item.get('url', ''),
                    'publish_time': item.get('datetime', item.get('publish_time', ''))
                })
        elif data.get('code') == 0 and not data.get('data'):
            print(f"  {account_name}: 当天没有发文")
        
        return articles
    
    except Exception as e:
        print(f"获取公众号 {account_name} ({account_id}) 文章失败: {e}")
        return []


def save_to_excel(articles, output_file):
    """保存文章列表到 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文章列表"
    
    # 表头
    ws.append(["公众号名称", "文章标题", "文章链接", "发布时间"])
    
    # 数据
    for article in articles:
        ws.append([
            article.get('account_name', ''),
            article.get('title', ''),
            article.get('url', ''),
            article.get('publish_time', '')
        ])
    
    # 保存为 .xlsx 格式
    xlsx_file = output_file.replace('.xlsx', '.xlsx')
    wb.save(xlsx_file)
    print(f"已保存 {len(articles)} 篇文章到: {xlsx_file}")
    return xlsx_file


def main():
    if not DAJIALA_KEY:
        print("错误：请先配置 DAJIALA_KEY 环境变量")
        print("在 ~/.bashrc 中添加: export DAJIALA_KEY=你的KEY")
        sys.exit(1)
    
    today = datetime.now().strftime("%Y%m%d")
    output_dir = os.path.join(BASE_DIR, today)
    os.makedirs(output_dir, exist_ok=True)
    
    # 读取公众号清单
    accounts = load_accounts_from_excel(ACCOUNTS_FILE)
    if not accounts:
        print(f"错误：没有找到公众号清单，请创建 {ACCOUNTS_FILE}")
        sys.exit(1)
    
    print(f"开始获取 {len(accounts)} 个公众号的文章...")
    print(f"公众号清单: {ACCOUNTS_FILE}")
    print(f"输出目录: {output_dir}")
    
    all_articles = []
    for account in accounts:
        print(f"\n正在获取: {account['name']} ({account['id']})")
        articles = fetch_articles(account['id'], account['name'])
        all_articles.extend(articles)
        if articles:
            print(f"  获取到 {len(articles)} 篇文章")
    
    if all_articles:
        # 保存到 Excel
        output_file = os.path.join(output_dir, f"{today}文章列表.xlsx")
        save_to_excel(all_articles, output_file)
    else:
        print("\n没有获取到任何文章")
    
    print(f"\n完成！共获取 {len(all_articles)} 篇文章")


if __name__ == "__main__":
    main()
