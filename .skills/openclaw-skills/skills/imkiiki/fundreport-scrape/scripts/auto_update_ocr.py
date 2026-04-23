#!/usr/bin/env python3
"""
互认基金月度数据自动更新脚本 - OCR 增强版
支持文本提取 + OCR 双重识别，确保图表数据不遗漏
"""

import re
import json
import openpyxl
from pathlib import Path
import pdfplumber
import zipfile
import tempfile
import shutil
from pdf2image import convert_from_path
import pytesseract
import cv2
import numpy as np
from PIL import Image


class FundUpdater:
    """基金数据更新器（支持 OCR）"""
    
    def __init__(self, skill_path):
        """初始化"""
        template_path = Path(skill_path) / 'references/extraction_templates.json'
        with open(template_path, 'r', encoding='utf-8') as f:
            self.templates = json.load(f)
        
        self.pdf_to_sheet = {
            '华夏': '华夏', '南方东英': '南方东英', '博时': '博时', '弘收': '弘收',
            '高腾': '高腾亚洲', '摩根亚洲债券': '摩根亚洲', '摩根国际': '摩根国际',
            '易方达': '易方达', '汇丰亚洲债券': '汇丰亚洲债券', '汇丰亚洲高收益': '汇丰亚洲高收益',
            '汇丰亚洲高入息': '汇丰亚洲高入息', '汇丰亚洲多元资产': '汇丰亚洲多元资产高入息',
            '海通': '海通亚洲', '东亚联丰亚洲债券': '东亚联丰亚洲债券及货币',
            '东亚联丰亚洲策略': '东亚联丰亚洲策略', '中银香港全天候中国': '中银香港全天候中国高息',
            '中银香港全天候亚洲': '中银香港全天候亚洲',
        }
        
        # OCR 配置
        self.ocr_lang = 'chi_sim+eng'
        self.ocr_config = '--oem 3 --psm 6'
        
        # 行业分布关键词
        self.industry_keywords = [
            '基本原料', '通讯', '周期性消费品', '非周期性消费品', '能源',
            '房地产', '金融', '工业', '科技', '公用事业', '政府', '现金', '其他',
            '金融', '地產', '工業', '能源', '科技', '公用事業', '政府', '現金', '其他'
        ]
        
        # 地区分布关键词
        self.region_keywords = [
            '中国内地', '中国香港', '中国澳门', '美国', '新加坡', '英国', '印尼', '韩国',
            '日本', '台湾', '其他亚太', '其他欧洲', '其他地区', '澳門', '台灣', '其他'
        ]
        
        # 信用评级关键词
        self.rating_keywords = [
            'AAA', 'AA+', 'AA', 'AA-', 'A+', 'A', 'A-', 'BBB+', 'BBB', 'BBB-', 'BB', 'B', 'NR',
            '未評級', '无评级'
        ]
    
    def preprocess_image(self, img):
        """图像预处理，提高 OCR 识别率"""
        # 转灰度
        gray = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2GRAY)
        
        # 二值化（增强对比度）
        _, binary = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)
        
        # 去噪
        denoised = cv2.medianBlur(binary, 3)
        
        return Image.fromarray(denoised)
    
    def extract_with_ocr(self, pdf_path):
        """使用 OCR 提取 PDF 中的图表数据"""
        ocr_data = {
            'industry': {},
            'region': {},
            'credit': {},
            'duration': None,
            'ytm': None
        }
        
        try:
            # PDF 转图片（300 DPI）
            images = convert_from_path(pdf_path, dpi=200)
            
            for page_num, img in enumerate(images):
                # 图像预处理
                processed_img = self.preprocess_image(img)
                
                # OCR 识别
                text = pytesseract.image_to_string(
                    processed_img,
                    lang=self.ocr_lang,
                    config=self.ocr_config
                )
                
                # 提取行业分布
                for keyword in self.industry_keywords:
                    # 模式：关键词 + 数字+%
                    pattern = rf'{keyword}\s*(\d+\.?\d*)\s*%'
                    matches = re.findall(pattern, text)
                    for match in matches:
                        value = float(match) / 100
                        if 0 < value < 1:  # 合理范围
                            ocr_data['industry'][keyword] = value
                
                # 提取地区分布
                for keyword in self.region_keywords:
                    pattern = rf'{keyword}\s*(\d+\.?\d*)\s*%'
                    matches = re.findall(pattern, text)
                    for match in matches:
                        value = float(match) / 100
                        if 0 < value < 1:
                            ocr_data['region'][keyword] = value
                
                # 提取信用评级分布
                for keyword in self.rating_keywords:
                    # 转义特殊字符
                    keyword_escaped = re.escape(keyword)
                    pattern = rf'{keyword_escaped}\s*(\d+\.?\d*)\s*%'
                    matches = re.findall(pattern, text)
                    for match in matches:
                        value = float(match) / 100
                        if 0 < value < 1:
                            ocr_data['credit'][keyword] = value
                
                # 提取久期（从 OCR 文本）
                duration_match = re.search(r'久期 [：:]\s*([\d.]+)', text)
                if duration_match and not ocr_data['duration']:
                    val = float(duration_match.group(1))
                    if 0 < val < 20:
                        ocr_data['duration'] = val
                
                # 提取 YTM（从 OCR 文本）
                ytm_match = re.search(r'(?:收益率 |YTM|孳息率)[：:]\s*([\d.]+)\s*%', text)
                if ytm_match and not ocr_data['ytm']:
                    val = float(ytm_match.group(1)) / 100
                    if 0 < val < 1:
                        ocr_data['ytm'] = val
            
            print(f"  📷 OCR 识别完成：行业={len(ocr_data['industry'])}项，地区={len(ocr_data['region'])}项，评级={len(ocr_data['credit'])}项")
            
        except Exception as e:
            print(f"  ⚠️ OCR 识别失败：{e}")
        
        return ocr_data
    
    def extract_from_text(self, pdf_path):
        """从 PDF 文本提取数据（传统方法）"""
        data = {
            'duration': None,
            'ytm': None,
            'industry': {},
            'region': {},
            'credit': {}
        }
        
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text
            
            # 合并文本
            merged_text = full_text.replace('\n', ' ').replace('\r', ' ')
            merged_text = re.sub(r'\s+', ' ', merged_text)
            
            # 提取久期
            duration_match = re.search(r'([\d.]+)\s*Yea(?:rs?)?', merged_text, re.IGNORECASE)
            if duration_match:
                val = float(duration_match.group(1))
                if 0 < val < 30:
                    data['duration'] = val
            
            # 提取 YTM
            ytm_match = re.search(r'加權平均最差收益率\d*\s*([\d.]+)\s*%', merged_text)
            if ytm_match:
                data['ytm'] = float(ytm_match.group(1)) / 100
        
        return data
    
    def merge_data(self, text_data, ocr_data):
        """合并文本提取和 OCR 识别的结果"""
        merged = {
            'duration': text_data.get('duration') or ocr_data.get('duration'),
            'ytm': text_data.get('ytm') or ocr_data.get('ytm'),
            'industry': {**ocr_data.get('industry', {}), **text_data.get('industry', {})},
            'region': {**ocr_data.get('region', {}), **text_data.get('region', {})},
            'credit': {**ocr_data.get('credit', {}), **text_data.get('credit', {})}
        }
        
        # 优先使用文本提取的核心指标（更准确）
        if text_data.get('duration'):
            merged['duration'] = text_data['duration']
        if text_data.get('ytm'):
            merged['ytm'] = text_data['ytm']
        
        return merged
    
    def process_zip(self, zip_path, output_dir):
        """处理 ZIP 文件"""
        temp_dir = tempfile.mkdtemp()
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            excel_files = [f for f in Path(temp_dir).rglob('*.xlsx') if not f.name.startswith('._')]
            if not excel_files:
                return None, "未找到 Excel 文件"
            
            excel_path = excel_files[0]
            pdf_files = [f for f in Path(temp_dir).rglob('*.pdf') if not f.name.startswith('._')]
            if not pdf_files:
                return None, "未找到 PDF 文件"
            
            output_name = excel_path.stem + '_已更新.xlsx'
            output_path = Path(output_dir) / output_name
            
            count = self.update_excel(str(excel_path), str(temp_dir), str(output_path))
            
            return output_path, f"成功更新 {count} 个基金"
        
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
    
    def parse_excel_date(self, filename):
        """
        从 Excel 文件名解析日期，格式：互认基金月度更新_YYYYMMvsYYYYMM.xlsx
        返回：(last_month, current_month) 元组
        """
        match = re.search(r'_(\d{6})vs(\d{6})', filename)
        if match:
            return match.group(1), match.group(2)
        return None, None
    
    def parse_pdf_date(self, filename):
        """
        从 PDF 文件名解析日期，支持两种格式：
        - YYYYMM: 如 202601.pdf → 202601
        - YYMM: 如 2601.pdf → 202601（自动补全年份）
        
        返回：YYYYMM 格式字符串
        """
        # 先尝试匹配 YYYYMM 格式（6 位数字，年份以 20 开头）
        match_ym = re.search(r'(20\d{2})(\d{2})', filename)
        if match_ym:
            year, month = match_ym.groups()
            return f"{year}{int(month):02d}"
        
        # 再尝试匹配 YYMM 格式（4 位数字，后面跟.pdf 或_或空格或结尾）
        match_yy = re.search(r'(?<!\d)(\d{2})(\d{2})(?:\.pdf|_|\.|\s|$)', filename)
        if match_yy:
            yy, mm = match_yy.groups()
            yy_int = int(yy)
            if yy_int < 50:
                year = 2000 + yy_int
            else:
                year = 1900 + yy_int
            return f"{year}{int(mm):02d}"
        
        return None
    
    def update_excel(self, excel_path, pdf_dir, output_path, target_month=None):
        """
        更新 Excel（双重提取：文本+OCR）
        
        Args:
            excel_path: Excel 模板路径
            pdf_dir: PDF 文件夹路径
            output_path: 输出路径
            target_month: 目标月份（'202512' 或 '202601'），None 表示自动从文件名解析
        """
        wb = openpyxl.load_workbook(excel_path)
        pdf_dir = Path(pdf_dir)
        
        # 从 Excel 文件名解析日期
        last_month, current_month = self.parse_excel_date(Path(excel_path).name)
        if not last_month or not current_month:
            print(f"⚠️ 无法从文件名解析日期，使用默认值")
            last_month, current_month = "202512", "202601"
        
        print(f"  📅 Excel 日期：上月={last_month}, 本月={current_month}")
        
        # 确定目标月份（用于决定数据填入哪一列）
        if target_month is None:
            # 自动从 PDF 文件名判断
            target_month = current_month  # 默认处理本月
        
        pdf_content_to_sheet = [
            ('华夏精选固定收益', '华夏'), ('華夏精選固定收益', '华夏'),
            ('南方东英', '南方东英'), ('博時精選新興市場', '博时'),
            ('弘收', '弘收'), ('高騰亞洲收益', '高腾亚洲'),
            ('摩根亚洲债券', '摩根亚洲'), ('摩根国际', '摩根国际'),
            ('易方达', '易方达'), ('汇丰亚洲债券', '汇丰亚洲债券'),
            ('滙豐亞洲高收益', '汇丰亚洲高收益'), ('滙豐亞洲高入息', '汇丰亚洲高入息'),
            ('滙豐亞洲多元資產', '汇丰亚洲多元资产高入息'),
            ('海通亞洲', '海通亚洲'), ('东亚联丰亚洲债券', '东亚联丰亚洲债券及货币'),
            ('東亞聯豐亞洲策略', '东亚联丰亚洲策略'),
            ('中銀香港全天候中國', '中银香港全天候中国高息'),
            ('中銀香港全天候亞洲', '中银香港全天候亚洲'),
        ]
        
        seen = set()
        
        pdf_files = [f for f in pdf_dir.rglob('*.pdf') if not f.name.startswith('._')]
        
        for pdf_path in sorted(pdf_files):
            # 从 PDF 文件名解析日期
            pdf_date = self.parse_pdf_date(pdf_path.name)
            
            # 只处理目标月份的 PDF
            if pdf_date and target_month and pdf_date != target_month:
                continue
            
            # 从 PDF 内容识别基金名称
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text() if pdf.pages else ""
            
            # 匹配 Sheet
            sheet_name = None
            for keyword, name in pdf_content_to_sheet:
                if keyword in (text or ''):
                    sheet_name = name
                    break
            
            if not sheet_name or sheet_name in seen or sheet_name not in wb.sheetnames:
                continue
            
            seen.add(sheet_name)
            ws = wb[sheet_name]
            
            print(f"\n📄 处理：{pdf_path.name} → {sheet_name} (月份：{pdf_date})")
            
            # 更新日期字段（只在第一次处理时设置）
            if not hasattr(self, '_dates_updated'):
                self._update_dates(ws, last_month, current_month)
                self._dates_updated = True
            
            # 确定数据填入的列（4=上月，6=本月）
            if pdf_date == last_month:
                data_col = 4  # 上月列
                print(f"  📊 填入上月列（列 4）")
            else:
                data_col = 6  # 本月列
                print(f"  📊 填入本月列（列 6）")
            
            # === 双重提取：文本 + OCR ===
            # 方法 A：文本提取
            text_data = self.extract_from_text(pdf_path)
            print(f"  📝 文本提取：久期={text_data.get('duration')}, YTM={text_data.get('ytm')}")
            
            # 方法 B：OCR 识别（图表数据）
            ocr_data = self.extract_with_ocr(pdf_path)
            
            # 合并结果
            merged_data = self.merge_data(text_data, ocr_data)
            print(f"  📊 合并结果：久期={merged_data.get('duration')}, YTM={merged_data.get('ytm')}")
            print(f"     行业分布={len(merged_data['industry'])}项，地区分布={len(merged_data['region'])}项")
            
            # 填入数据到指定列
            if merged_data.get('duration'):
                ws.cell(row=12, column=data_col).value = merged_data['duration']
            
            if merged_data.get('ytm'):
                ws.cell(row=13, column=data_col).value = merged_data['ytm']
            
            # 填入行业分布（行 18-31）
            for row_idx in range(18, 32):
                excel_name = ws.cell(row=row_idx, column=5).value
                if excel_name and excel_name in merged_data['industry']:
                    ws.cell(row=row_idx, column=data_col).value = merged_data['industry'][excel_name]
            
            # 填入地区分布（行 36-49）
            for row_idx in range(36, 50):
                excel_name = ws.cell(row=row_idx, column=5).value
                if excel_name and excel_name in merged_data['region']:
                    ws.cell(row=row_idx, column=data_col).value = merged_data['region'][excel_name]
            
            # 填入信用评级分布（行 52-65）
            for row_idx in range(52, 66):
                excel_name = ws.cell(row=row_idx, column=5).value
                if excel_name and excel_name in merged_data['credit']:
                    ws.cell(row=row_idx, column=data_col).value = merged_data['credit'][excel_name]
        
        wb.save(output_path)
        wb.close()
        
        return len(seen)
    
    def _update_dates(self, ws, last_month, current_month):
        """更新日期字段"""
        ws.cell(row=7, column=3).value = last_month   # 上月
        ws.cell(row=8, column=3).value = current_month  # 本月
    
    def _left_shift_data(self, ws):
        """数据左移"""
        # 核心指标
        for row in [12, 13, 14]:
            val = ws.cell(row=row, column=6).value
            if val:
                ws.cell(row=row, column=4).value = val
        
        # 行业分布
        for row in range(18, 32):
            name = ws.cell(row=row, column=5).value
            val = ws.cell(row=row, column=6).value
            if name:
                ws.cell(row=row, column=3).value = name
            if val:
                ws.cell(row=row, column=4).value = val
        
        # 区域分布
        for row in range(36, 50):
            name = ws.cell(row=row, column=5).value
            val = ws.cell(row=row, column=6).value
            if name:
                ws.cell(row=row, column=3).value = name
            if val:
                ws.cell(row=row, column=4).value = val
        
        # 信用评级
        for row in range(52, 66):
            name = ws.cell(row=row, column=5).value
            val = ws.cell(row=row, column=6).value
            if name:
                ws.cell(row=row, column=3).value = name
            if val:
                ws.cell(row=row, column=4).value = val


def main():
    """主函数"""
    import sys
    
    if len(sys.argv) < 3:
        print("用法：python auto_update_ocr.py <zip_path> <output_dir>")
        print("   或：python auto_update_ocr.py <excel_path> <pdf_dir> <output_path>")
        sys.exit(1)
    
    skill_path = Path(__file__).parent.parent
    updater = FundUpdater(skill_path)
    
    if len(sys.argv) == 3:
        output_path, message = updater.process_zip(sys.argv[1], sys.argv[2])
        print(f"{message}")
        if output_path:
            print(f"文件：{output_path}")
    else:
        count = updater.update_excel(sys.argv[1], sys.argv[2], sys.argv[3])
        print(f"\n✅ 已更新 {count} 个基金")


if __name__ == '__main__':
    main()
