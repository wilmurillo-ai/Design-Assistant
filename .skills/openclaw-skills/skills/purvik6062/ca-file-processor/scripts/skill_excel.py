"""
OpenClaw Skill: Excel Processor (Open-Source Path)
===================================================
Reads .xlsx and .xls files using openpyxl.
Detects common CA document types (trial balance, P&L, ledger, etc.)
and structures the output for LLM consumption.

Dependencies:
    pip install openpyxl xlrd

Note: xlrd for legacy .xls support (pre-2007 Excel files).
      openpyxl handles all .xlsx files natively.

Usage in OpenClaw skill config:
    skill_name: excel_processor
    trigger: "*.xlsx uploaded" OR "*.xls uploaded"
    entry: process_excel(file_path)
"""

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


# ── CONFIG ────────────────────────────────────────────────────────────────────

MAX_ROWS_PER_SHEET = 2000            # Cap to avoid overloading LLM context
MAX_COLS = 50
PREVIEW_ROWS = 10                    # Rows to include in quick preview


# ── CA DOCUMENT TYPE DETECTION ────────────────────────────────────────────────

# Keywords that identify common CA Excel document types
CA_DOC_PATTERNS = {
    "trial_balance": [
        "trial balance", "debit", "credit", "closing balance", "opening balance"
    ],
    "profit_loss": [
        "profit", "loss", "revenue", "income", "expenditure", "expenses",
        "gross profit", "net profit", "ebitda"
    ],
    "balance_sheet": [
        "balance sheet", "assets", "liabilities", "equity", "capital",
        "fixed assets", "current assets", "current liabilities"
    ],
    "ledger": [
        "ledger", "dr", "cr", "narration", "voucher", "entry date"
    ],
    "bank_reconciliation": [
        "bank reconciliation", "brs", "passbook", "bank balance",
        "outstanding cheque", "deposit in transit"
    ],
    "payroll": [
        "salary", "payroll", "basic", "hra", "pf", "esic", "tds",
        "gross salary", "net salary", "ctc"
    ],
    "gst_workings": [
        "gstin", "gst", "igst", "cgst", "sgst", "taxable value",
        "hsn", "sac code"
    ],
}


# ── MAIN ENTRY POINT ─────────────────────────────────────────────────────────

def process_excel(file_path: str) -> dict:
    """
    Main entry point called by OpenClaw when an Excel file is received.

    Returns:
        {
          "file":       str,
          "sheets":     list[dict],     # one dict per sheet
          "doc_type":   str,            # detected CA document type
          "summary":    str,            # human-readable preview
          "row_counts": dict,           # sheet_name → row count
        }
    """
    path = Path(file_path)
    if not path.exists():
        return {"error": f"File not found: {file_path}"}

    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return _process_xlsx(path)
    elif suffix == ".xls":
        return _process_xls_legacy(path)
    else:
        return {"error": f"Unsupported extension: {suffix}. Expected .xlsx or .xls"}


# ── XLSX PROCESSING ───────────────────────────────────────────────────────────

def _process_xlsx(path: Path) -> dict:
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    sheets_data = []
    all_text_for_detection = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_result = _extract_sheet(ws, sheet_name)
        sheets_data.append(sheet_result)
        all_text_for_detection.extend(sheet_result["header_values"])

    wb.close()

    doc_type = _detect_ca_doc_type(all_text_for_detection)
    summary = _build_summary(sheets_data, doc_type)

    return {
        "file": path.name,
        "sheets": sheets_data,
        "doc_type": doc_type,
        "summary": summary,
        "row_counts": {s["name"]: s["row_count"] for s in sheets_data},
    }


def _extract_sheet(ws, sheet_name: str) -> dict:
    """Extract all rows from a single worksheet."""
    rows = []
    header_values = []
    row_count = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > MAX_ROWS_PER_SHEET:
            break

        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        cleaned_row = [_clean_cell(cell) for cell in row[:MAX_COLS]]
        rows.append(cleaned_row)
        row_count += 1

        # Collect header-area values for doc type detection
        if row_idx <= 10:
            header_values.extend([str(v).lower() for v in cleaned_row if v])

    return {
        "name": sheet_name,
        "rows": rows,
        "row_count": row_count,
        "header_values": header_values,
        "preview": rows[:PREVIEW_ROWS],      # First N rows as quick look
        "truncated": row_count >= MAX_ROWS_PER_SHEET,
    }


def _clean_cell(value) -> str:
    """Normalise cell value to string."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Avoid scientific notation for large financial numbers
        if value == int(value):
            return str(int(value))
        return f"{value:.2f}"
    return str(value).strip()


# ── LEGACY .XLS SUPPORT ───────────────────────────────────────────────────────

def _process_xls_legacy(path: Path) -> dict:
    """
    Fallback for old .xls files (pre-2007).
    Uses xlrd if available, otherwise returns a helpful error.
    """
    try:
        import xlrd
        wb = xlrd.open_workbook(str(path))
        sheets_data = []
        all_text = []

        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            rows = []
            header_values = []

            for row_idx in range(min(ws.nrows, MAX_ROWS_PER_SHEET)):
                row = [str(ws.cell_value(row_idx, col)).strip()
                       for col in range(min(ws.ncols, MAX_COLS))]
                if any(v for v in row):
                    rows.append(row)
                if row_idx < 10:
                    header_values.extend([v.lower() for v in row if v])

            sheets_data.append({
                "name": sheet_name,
                "rows": rows,
                "row_count": len(rows),
                "header_values": header_values,
                "preview": rows[:PREVIEW_ROWS],
                "truncated": ws.nrows >= MAX_ROWS_PER_SHEET,
            })
            all_text.extend(header_values)

        wb.release_resources()
        doc_type = _detect_ca_doc_type(all_text)
        return {
            "file": path.name,
            "sheets": sheets_data,
            "doc_type": doc_type,
            "summary": _build_summary(sheets_data, doc_type),
            "row_counts": {s["name"]: s["row_count"] for s in sheets_data},
        }

    except ImportError:
        return {
            "error": (
                ".xls file detected but xlrd is not installed. "
                "Run: pip install xlrd==1.2.0  "
                "(Note: xlrd 2.x dropped .xls support — use 1.2.0)"
            )
        }


# ── CA DOCUMENT TYPE DETECTION ────────────────────────────────────────────────

def _detect_ca_doc_type(header_values: list) -> str:
    """
    Scores each CA document type against extracted header keywords.
    Returns the best match or 'unknown'.
    """
    combined = " ".join(header_values).lower()
    scores = {}

    for doc_type, keywords in CA_DOC_PATTERNS.items():
        score = sum(1 for kw in keywords if kw in combined)
        if score > 0:
            scores[doc_type] = score

    if not scores:
        return "unknown"

    return max(scores, key=scores.get)


# ── SUMMARY BUILDER ───────────────────────────────────────────────────────────

def _build_summary(sheets_data: list, doc_type: str) -> str:
    """Build a human-readable summary for LLM context."""
    lines = [f"Document type: {doc_type.replace('_', ' ').title()}"]
    lines.append(f"Sheets: {len(sheets_data)}")

    for sheet in sheets_data:
        lines.append(f"\nSheet: {sheet['name']} ({sheet['row_count']} rows)")
        if sheet["preview"]:
            lines.append("First rows preview:")
            for row in sheet["preview"][:5]:
                non_empty = [v for v in row if v]
                if non_empty:
                    lines.append("  " + " | ".join(non_empty[:8]))

    return "\n".join(lines)


# ── OPENCLAW SKILL METADATA ───────────────────────────────────────────────────

SKILL_METADATA = {
    "name": "excel_processor",
    "version": "1.0.0",
    "description": "Reads Excel files and detects CA document type (trial balance, P&L, etc.)",
    "triggers": ["*.xlsx", "*.xls", "process excel", "read this spreadsheet"],
    "entry_function": "process_excel",
    "output_format": "dict → passed to LLM context",
    "ca_use_cases": [
        "Trial balance",
        "Profit & Loss statement",
        "Balance sheet workings",
        "Ledger extracts",
        "Bank reconciliation",
        "Payroll register",
        "GST workings / GSTR reconciliation",
    ]
}


# ── QUICK TEST ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python skill_excel.py <path_to_excel>")
        sys.exit(1)
    result = process_excel(sys.argv[1])
    print(json.dumps({k: v for k, v in result.items() if k != "sheets"}, indent=2))
