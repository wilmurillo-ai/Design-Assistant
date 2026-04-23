"""
Retail Trade Weekly Report Generator
Processes 12 Excel files (6 current week + 6 previous week) to generate consolidated weekly report
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import re
import os
from datetime import datetime
from collections import defaultdict


class RetailTradeReportGenerator:
    """Main class for generating retail trade weekly reports"""
    
    REGIONS = ['NCR', 'SLZ', 'NLZ', 'CLZ', 'EVIS', 'WVIS', 'MIN', 'Others']
    
    FILE_TYPES = {
        'DRP': 'DRP_Channel_Sales_Report_DRP',
        'TECNO': 'DRP_Special_SIM_Monitor_Report_Daily_TECNO',
        'LS': 'License_Store_Performance_Monitor_Report_LS',
        'DXS_Prepaid': 'DXS_Acquisition_Report_Mobile_Prepaid',
        'DXS_Postpaid': 'DXS_Acquisition_Report_Mobile_Postpaid',
        'DXS_FWA': 'DXS_Acquisition_Report_FWA'
    }
    
    def __init__(self, input_dir, mapping_csv_path):
        self.input_dir = input_dir
        self.mapping_csv_path = mapping_csv_path
        self.store_mapping = {}
        self.current_week_files = {}
        self.previous_week_files = {}
        self.current_week_data = {}
        self.previous_week_data = {}
        
    def load_store_mapping(self):
        """Load store to region mapping from CSV"""
        df = pd.read_csv(self.mapping_csv_path)
        
        for _, row in df.iterrows():
            store_name = str(row['Store Name']).strip()
            region = str(row['Region']).strip()
            aliases = str(row['Aliases']) if pd.notna(row['Aliases']) else ""
            
            # Add main name
            self.store_mapping[store_name.upper()] = region
            
            # Add aliases
            if aliases:
                for alias in aliases.split('|'):
                    alias = alias.strip()
                    if alias:
                        self.store_mapping[alias.upper()] = region
        
        print(f"Loaded {len(self.store_mapping)} store mappings")
    
    def map_store_to_region(self, store_name):
        """Map store name to region with fuzzy matching"""
        if pd.isna(store_name):
            return "Others"
        
        store_name = str(store_name).strip()
        store_upper = store_name.upper()
        
        # Exact match
        if store_upper in self.store_mapping:
            return self.store_mapping[store_upper]
        
        # Partial match
        for key, region in self.store_mapping.items():
            if key in store_upper or store_upper in key:
                return region
        
        # Default to Others
        return "Others"
    
    def extract_date_from_filename(self, filename):
        """Extract date range from filename"""
        pattern = r'_(\d+)_(\d+)-(\d+)_(\d+)'
        match = re.search(pattern, filename)
        if match:
            start_month, start_day, end_month, end_day = match.groups()
            # Use year 2026 for comparison
            start_date = datetime(2026, int(start_month), int(start_day))
            end_date = datetime(2026, int(end_month), int(end_day))
            return start_date, end_date, f"{start_month}/{start_day}-{end_month}/{end_day}"
        return None, None, None
    
    def identify_file_type(self, filename):
        """Identify file type from filename"""
        for file_type, pattern in self.FILE_TYPES.items():
            if pattern in filename:
                return file_type
        return None
    
    def identify_files(self):
        """Identify and group current week vs previous week files"""
        all_files = [f for f in os.listdir(self.input_dir) if f.endswith('.xlsx')]
        
        files_with_dates = []
        for filename in all_files:
            file_type = self.identify_file_type(filename)
            if file_type:
                start_date, end_date, date_str = self.extract_date_from_filename(filename)
                if start_date:
                    files_with_dates.append({
                        'filename': filename,
                        'filepath': os.path.join(self.input_dir, filename),
                        'type': file_type,
                        'start_date': start_date,
                        'end_date': end_date,
                        'date_str': date_str
                    })
        
        # Sort by date (most recent first)
        files_with_dates.sort(key=lambda x: x['end_date'], reverse=True)
        
        # Group by week (assuming 2 sets of 6 files)
        current_week_end = files_with_dates[0]['end_date']
        
        for file_info in files_with_dates:
            if file_info['end_date'] == current_week_end:
                self.current_week_files[file_info['type']] = file_info
            else:
                self.previous_week_files[file_info['type']] = file_info
        
        print(f"Current week files: {len(self.current_week_files)}")
        print(f"Previous week files: {len(self.previous_week_files)}")
        
        # Validate
        if len(self.current_week_files) != 6:
            raise ValueError(f"Expected 6 current week files, found {len(self.current_week_files)}")
        if len(self.previous_week_files) != 6:
            raise ValueError(f"Expected 6 previous week files, found {len(self.previous_week_files)}")
    
    def extract_drp_data(self, filepath):
        """Extract DRP channel sales data"""
        df = pd.read_excel(filepath, sheet_name='Sheet0', header=None)
        
        data = {}
        for idx in range(8, 16):  # Row 8-15 are data rows
            region = df.iloc[idx, 0]
            if pd.isna(region) or region == 'Total':
                continue
            
            data[region] = {
                'mobile_postpaid': self._safe_float(df.iloc[idx, 1]),
                'mobile_prepaid': self._safe_float(df.iloc[idx, 5]),
                'double_data': self._safe_float(df.iloc[idx, 6]),
                'fwa_4g': self._safe_float(df.iloc[idx, 9]),
                'fwa_5g': self._safe_float(df.iloc[idx, 10]) + self._safe_float(df.iloc[idx, 11])
            }
        
        return data
    
    def extract_tecno_data(self, filepath):
        """Extract TECNO activation data"""
        df = pd.read_excel(filepath, sheet_name='Sheet0', header=None)
        
        data = {}
        for idx in range(7, 15):  # Row 7-14 are data rows
            region = df.iloc[idx, 0]
            if pd.isna(region) or region == 'Total':
                continue
            
            data[region] = {
                'camon_40': self._safe_float(df.iloc[idx, 1]),
                'pova_7': self._safe_float(df.iloc[idx, 2]),
                'tecno_total': self._safe_float(df.iloc[idx, 3])
            }
        
        return data
    
    def extract_ls_data(self, filepath):
        """Extract License Store data"""
        df = pd.read_excel(filepath, sheet_name='Sheet0', header=None)
        
        stores_data = {
            'mobile_prepaid': {},
            'mobile_postpaid': {},
            'fwa_4g': {}
        }
        
        for idx in range(8, len(df)):
            store = df.iloc[idx, 0]
            if pd.isna(store):
                break
            
            stores_data['mobile_prepaid'][store] = self._safe_float(df.iloc[idx, 1])
            stores_data['mobile_postpaid'][store] = self._safe_float(df.iloc[idx, 3])
            stores_data['fwa_4g'][store] = self._safe_float(df.iloc[idx, 29])  # Column AD
        
        return stores_data
    
    def extract_dxs_prepaid_data(self, filepath):
        """Extract DXS Mobile Prepaid data"""
        df = pd.read_excel(filepath, sheet_name='Sheet1', header=None)
        
        stores_data = {}
        for idx in range(8, len(df)):
            store = df.iloc[idx, 0]
            if pd.isna(store) or store in ['Grand Total', '-']:
                continue
            
            value = self._safe_float(df.iloc[idx, 4])  # Column 4: Total
            if value > 0:
                stores_data[store] = value
        
        return stores_data
    
    def extract_dxs_postpaid_data(self, filepath):
        """Extract DXS Mobile Postpaid data"""
        df = pd.read_excel(filepath, sheet_name='Sheet1', header=None)
        
        stores_data = {}
        for idx in range(8, len(df)):
            store = df.iloc[idx, 0]
            if pd.isna(store) or store in ['Grand Total', '-']:
                continue
            
            value = self._safe_float(df.iloc[idx, 12])  # Column 12: Total
            if value > 0:
                stores_data[store] = value
        
        return stores_data
    
    def extract_dxs_fwa_data(self, filepath):
        """Extract DXS FWA data with 4G/5G split"""
        df = pd.read_excel(filepath, sheet_name='Sheet1', header=None)
        
        stores_data = {
            'fwa_4g': {},
            'fwa_5g': {}
        }
        
        for idx in range(8, len(df)):
            store = df.iloc[idx, 0]
            if pd.isna(store) or store in ['Grand Total', '-']:
                continue
            
            fwa_4g = self._safe_float(df.iloc[idx, 1])  # Column 1: 4G
            total = self._safe_float(df.iloc[idx, 18])  # Column 18: Total
            fwa_5g = total - fwa_4g
            
            if fwa_4g > 0:
                stores_data['fwa_4g'][store] = fwa_4g
            if fwa_5g > 0:
                stores_data['fwa_5g'][store] = fwa_5g
        
        return stores_data
    
    def aggregate_stores_by_region(self, stores_data):
        """Aggregate store-level data by region"""
        regional_data = {region: 0 for region in self.REGIONS}
        
        for store, value in stores_data.items():
            region = self.map_store_to_region(store)
            regional_data[region] += value
        
        return regional_data
    
    def process_all_files(self):
        """Process all files and aggregate data"""
        for week_label, file_dict in [('current', self.current_week_files), 
                                       ('previous', self.previous_week_files)]:
            week_data = {}
            
            # Process DRP
            if 'DRP' in file_dict:
                week_data['drp'] = self.extract_drp_data(file_dict['DRP']['filepath'])
            
            # Process TECNO
            if 'TECNO' in file_dict:
                week_data['tecno'] = self.extract_tecno_data(file_dict['TECNO']['filepath'])
            
            # Process License Store
            if 'LS' in file_dict:
                ls_raw = self.extract_ls_data(file_dict['LS']['filepath'])
                week_data['ls'] = {
                    'mobile_prepaid': self.aggregate_stores_by_region(ls_raw['mobile_prepaid']),
                    'mobile_postpaid': self.aggregate_stores_by_region(ls_raw['mobile_postpaid']),
                    'fwa_4g': self.aggregate_stores_by_region(ls_raw['fwa_4g'])
                }
            
            # Process DXS Prepaid
            if 'DXS_Prepaid' in file_dict:
                dxs_prepaid_raw = self.extract_dxs_prepaid_data(file_dict['DXS_Prepaid']['filepath'])
                week_data['dxs_prepaid'] = self.aggregate_stores_by_region(dxs_prepaid_raw)
            
            # Process DXS Postpaid
            if 'DXS_Postpaid' in file_dict:
                dxs_postpaid_raw = self.extract_dxs_postpaid_data(file_dict['DXS_Postpaid']['filepath'])
                week_data['dxs_postpaid'] = self.aggregate_stores_by_region(dxs_postpaid_raw)
            
            # Process DXS FWA
            if 'DXS_FWA' in file_dict:
                dxs_fwa_raw = self.extract_dxs_fwa_data(file_dict['DXS_FWA']['filepath'])
                week_data['dxs_fwa'] = {
                    'fwa_4g': self.aggregate_stores_by_region(dxs_fwa_raw['fwa_4g']),
                    'fwa_5g': self.aggregate_stores_by_region(dxs_fwa_raw['fwa_5g'])
                }
            
            if week_label == 'current':
                self.current_week_data = week_data
            else:
                self.previous_week_data = week_data
    
    def calculate_wow(self, current, previous):
        """Calculate week-over-week percentage"""
        if previous == 0 or pd.isna(previous):
            return "-"
        if current == 0 or pd.isna(current):
            return "-100%"
        
        wow = ((current - previous) / previous) * 100
        return f"{int(round(wow))}%"
    
    def _safe_float(self, value):
        """Safely convert value to float"""
        if pd.isna(value):
            return 0.0
        try:
            return float(value)
        except:
            return 0.0
    
    def generate_report(self, output_path):
        """Generate the final Excel report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Report"
        
        # Get date ranges
        current_date_str = self.current_week_files['DRP']['date_str']
        previous_date_str = self.previous_week_files['DRP']['date_str']
        
        # Write header
        row = 1
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'] = "Retail Trade Weekly Report"
        ws[f'A{row}'].font = Font(size=16, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'] = f"Last Week: {previous_date_str} | This Week: {current_date_str}"
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        row += 2
        
        # Mobile Prepaid Section
        row = self._write_product_section(ws, row, "Mobile Prepaid", 'mobile_prepaid')
        
        row += 2
        
        # DRP Prepaid Program Section
        row = self._write_drp_program_section(ws, row)
        
        row += 2
        
        # Mobile Postpaid Section
        row = self._write_product_section(ws, row, "Mobile Postpaid", 'mobile_postpaid')
        
        row += 2
        
        # FWA 4G Section
        row = self._write_product_section(ws, row, "FWA 4G", 'fwa_4g')
        
        row += 2
        
        # FWA 5G Section
        row = self._write_product_section(ws, row, "FWA 5G", 'fwa_5g')
        
        # Apply formatting
        self._apply_formatting(ws)
        
        # Adjust column widths
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save
        wb.save(output_path)
        print(f"Report saved to: {output_path}")
        return output_path
    
    def _write_product_section(self, ws, start_row, product_name, product_key):
        """Write a product section (Mobile Prepaid/Postpaid/FWA)"""
        row = start_row
        
        # Section header
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'] = product_name
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        row += 1
        
        # Column headers
        headers = ['Region', 'RT Total ADA', 'WoW', 'DXS ADA', 'WoW', 'LS ADA', 'WoW', 'DRP ADA', 'WoW']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Data rows
        total_current = {'rt': 0, 'dxs': 0, 'ls': 0, 'drp': 0}
        total_previous = {'rt': 0, 'dxs': 0, 'ls': 0, 'drp': 0}
        
        for region in self.REGIONS:
            # Get current week data
            drp_current = self.current_week_data['drp'].get(region, {}).get(product_key, 0)
            dxs_current = self._get_dxs_value(self.current_week_data, product_key, region)
            ls_current = self._get_ls_value(self.current_week_data, product_key, region)
            rt_current = drp_current + dxs_current + ls_current
            
            # Get previous week data
            drp_previous = self.previous_week_data['drp'].get(region, {}).get(product_key, 0)
            dxs_previous = self._get_dxs_value(self.previous_week_data, product_key, region)
            ls_previous = self._get_ls_value(self.previous_week_data, product_key, region)
            rt_previous = drp_previous + dxs_previous + ls_previous
            
            # Update totals
            total_current['rt'] += rt_current
            total_current['dxs'] += dxs_current
            total_current['ls'] += ls_current
            total_current['drp'] += drp_current
            total_previous['rt'] += rt_previous
            total_previous['dxs'] += dxs_previous
            total_previous['ls'] += ls_previous
            total_previous['drp'] += drp_previous
            
            # Write row
            ws.cell(row, 1, region).font = Font(bold=True)
            ws.cell(row, 2, self._format_number(rt_current))
            ws.cell(row, 3, self.calculate_wow(rt_current, rt_previous))
            ws.cell(row, 4, self._format_number(dxs_current))
            ws.cell(row, 5, self.calculate_wow(dxs_current, dxs_previous))
            ws.cell(row, 6, self._format_number(ls_current))
            ws.cell(row, 7, self.calculate_wow(ls_current, ls_previous))
            ws.cell(row, 8, self._format_number(drp_current))
            ws.cell(row, 9, self.calculate_wow(drp_current, drp_previous))
            
            # Apply WoW coloring
            self._color_wow_cell(ws.cell(row, 3))
            self._color_wow_cell(ws.cell(row, 5))
            self._color_wow_cell(ws.cell(row, 7))
            self._color_wow_cell(ws.cell(row, 9))
            
            row += 1
        
        # Total row
        ws.cell(row, 1, "Total").font = Font(bold=True)
        ws.cell(row, 2, self._format_number(total_current['rt']))
        ws.cell(row, 3, self.calculate_wow(total_current['rt'], total_previous['rt']))
        ws.cell(row, 4, self._format_number(total_current['dxs']))
        ws.cell(row, 5, self.calculate_wow(total_current['dxs'], total_previous['dxs']))
        ws.cell(row, 6, self._format_number(total_current['ls']))
        ws.cell(row, 7, self.calculate_wow(total_current['ls'], total_previous['ls']))
        ws.cell(row, 8, self._format_number(total_current['drp']))
        ws.cell(row, 9, self.calculate_wow(total_current['drp'], total_previous['drp']))
        
        # Color total row
        for col in range(1, 10):
            ws.cell(row, col).fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            ws.cell(row, col).font = Font(bold=True)
        
        self._color_wow_cell(ws.cell(row, 3))
        self._color_wow_cell(ws.cell(row, 5))
        self._color_wow_cell(ws.cell(row, 7))
        self._color_wow_cell(ws.cell(row, 9))
        
        return row + 1
    
    def _write_drp_program_section(self, ws, start_row):
        """Write DRP Prepaid Program section"""
        row = start_row
        
        # Section header
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'] = "DRP Prepaid Program"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        row += 1
        
        # Column headers
        headers = ['Region', 'Double Data ADA', 'WoW', 'TECNO ADA', 'WoW', 'CAMON 40', 'WoW', 'POVA 7', 'WoW']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Data rows
        total_current = {'double': 0, 'tecno': 0, 'camon': 0, 'pova': 0}
        total_previous = {'double': 0, 'tecno': 0, 'camon': 0, 'pova': 0}
        
        for region in self.REGIONS:
            # Current week
            double_current = self.current_week_data['drp'].get(region, {}).get('double_data', 0)
            camon_current = self.current_week_data['tecno'].get(region, {}).get('camon_40', 0)
            pova_current = self.current_week_data['tecno'].get(region, {}).get('pova_7', 0)
            tecno_current = camon_current + pova_current
            
            # Previous week
            double_previous = self.previous_week_data['drp'].get(region, {}).get('double_data', 0)
            camon_previous = self.previous_week_data['tecno'].get(region, {}).get('camon_40', 0)
            pova_previous = self.previous_week_data['tecno'].get(region, {}).get('pova_7', 0)
            tecno_previous = camon_previous + pova_previous
            
            # Update totals
            total_current['double'] += double_current
            total_current['tecno'] += tecno_current
            total_current['camon'] += camon_current
            total_current['pova'] += pova_current
            total_previous['double'] += double_previous
            total_previous['tecno'] += tecno_previous
            total_previous['camon'] += camon_previous
            total_previous['pova'] += pova_previous
            
            # Write row
            ws.cell(row, 1, region).font = Font(bold=True)
            ws.cell(row, 2, self._format_number(double_current))
            ws.cell(row, 3, self.calculate_wow(double_current, double_previous))
            ws.cell(row, 4, self._format_number(tecno_current))
            ws.cell(row, 5, self.calculate_wow(tecno_current, tecno_previous))
            ws.cell(row, 6, self._format_number(camon_current))
            ws.cell(row, 7, self.calculate_wow(camon_current, camon_previous))
            ws.cell(row, 8, self._format_number(pova_current))
            ws.cell(row, 9, self.calculate_wow(pova_current, pova_previous))
            
            # Apply coloring
            for col in [3, 5, 7, 9]:
                self._color_wow_cell(ws.cell(row, col))
            
            row += 1
        
        # Total row
        ws.cell(row, 1, "Total").font = Font(bold=True)
        ws.cell(row, 2, self._format_number(total_current['double']))
        ws.cell(row, 3, self.calculate_wow(total_current['double'], total_previous['double']))
        ws.cell(row, 4, self._format_number(total_current['tecno']))
        ws.cell(row, 5, self.calculate_wow(total_current['tecno'], total_previous['tecno']))
        ws.cell(row, 6, self._format_number(total_current['camon']))
        ws.cell(row, 7, self.calculate_wow(total_current['camon'], total_previous['camon']))
        ws.cell(row, 8, self._format_number(total_current['pova']))
        ws.cell(row, 9, self.calculate_wow(total_current['pova'], total_previous['pova']))
        
        # Color total row
        for col in range(1, 10):
            ws.cell(row, col).fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            ws.cell(row, col).font = Font(bold=True)
        
        for col in [3, 5, 7, 9]:
            self._color_wow_cell(ws.cell(row, col))
        
        return row + 1
    
    def _get_dxs_value(self, week_data, product_key, region):
        """Get DXS value for a product and region"""
        if product_key == 'mobile_prepaid':
            return week_data.get('dxs_prepaid', {}).get(region, 0)
        elif product_key == 'mobile_postpaid':
            return week_data.get('dxs_postpaid', {}).get(region, 0)
        elif product_key == 'fwa_4g':
            return week_data.get('dxs_fwa', {}).get('fwa_4g', {}).get(region, 0)
        elif product_key == 'fwa_5g':
            return week_data.get('dxs_fwa', {}).get('fwa_5g', {}).get(region, 0)
        return 0
    
    def _get_ls_value(self, week_data, product_key, region):
        """Get LS value for a product and region"""
        return week_data.get('ls', {}).get(product_key, {}).get(region, 0)
    
    def _format_number(self, value):
        """Format number for display"""
        if value < 10 and value > 0:
            return round(value, 1)
        return int(round(value))
    
    def _color_wow_cell(self, cell):
        """Apply color coding to WoW cell"""
        value = str(cell.value)
        if value == "-":
            cell.font = Font(color="808080")  # Gray
        elif "%" in value:
            try:
                pct = int(value.replace("%", ""))
                if pct > 0:
                    cell.font = Font(color="008000", bold=True)  # Green
                elif pct < 0:
                    cell.font = Font(color="FF0000", bold=True)  # Red
            except:
                pass
    
    def _apply_formatting(self, ws):
        """Apply borders to all cells"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border


def generate_weekly_report(input_dir, mapping_csv, output_path):
    """
    Main function to generate retail trade weekly report
    
    Args:
        input_dir: Directory containing input Excel files
        mapping_csv: Path to store mapping CSV file
        output_path: Output Excel file path
    
    Returns:
        Path to generated report
    """
    generator = RetailTradeReportGenerator(input_dir, mapping_csv)
    
    print("Step 1: Loading store mapping...")
    generator.load_store_mapping()
    
    print("Step 2: Identifying files...")
    generator.identify_files()
    
    print("Step 3: Processing all files...")
    generator.process_all_files()
    
    print("Step 4: Generating report...")
    return generator.generate_report(output_path)


if __name__ == "__main__":
    # Example usage
    input_dir = "/mnt/user-data/uploads/"
    mapping_csv = "/mnt/user-data/uploads/store_mapping.csv"
    output_path = "/mnt/user-data/outputs/Retail_Trade_Weekly_Report.xlsx"
    
    result = generate_weekly_report(input_dir, mapping_csv, output_path)
    print(f"✓ Report generated: {result}")
