import os
import re
import sys
import time
import random
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

# Accept topic URL from command line or prompt
if len(sys.argv) > 1:
    topic_url = sys.argv[1].rstrip('/')
else:
    topic_url = input("Enter the Lowyat forum topic URL: ").strip().rstrip('/')

# Extract topic ID from URL
topic_id = topic_url.split('/')[-1].split('/')[0]
output_file = f'{topic_id}.xlsx'
print(f"Scraping topic: {topic_id}")

# Fetch first page to auto-detect total pages
r = requests.get(topic_url)
soup = BeautifulSoup(r.content, 'html5lib')
last_offsets = re.findall(r'/topic/' + topic_id + r'/\+(\d+)', str(soup))
if last_offsets:
    max_offset = max(int(x) for x in last_offsets)
else:
    max_offset = 0
print(f"Detected {max_offset // 20 + 1} pages (max offset: {max_offset})")

# Build page offsets
offsets = [''] + [str(n) for n in range(20, max_offset + 1, 20)]

# Load or create workbook
if os.path.exists(output_file):
    wb = load_workbook(output_file)
    ws = wb.active
    print(f"Resuming with {ws.max_row - 1} existing rows in {output_file}")
else:
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Date', 'Comment'])

# Scrape all pages
for offset in tqdm(offsets, desc="Scraping", unit="page", dynamic_ncols=True):
    url = topic_url + '/+' + offset if offset else topic_url

    r = requests.get(url)
    soup = BeautifulSoup(r.content, 'html5lib')
    table = soup.find('div', attrs={'id': 'topic_content'})

    if table is None:
        tqdm.write(f"  Warning: no topic_content found, skipping {url}")
        continue

    # Get all post tables
    posts = table.findAll('table', attrs={'class': 'post_table'})

    for post in posts:
        # Username
        name_el = post.find('span', attrs={'class': 'normalname'})
        name = name_el.get_text(strip=True) if name_el else ''

        # Date
        date_el = post.find('span', attrs={'class': 'postdetails'})
        date = date_el.get_text(strip=True) if date_el else ''

        # Comment text
        text_el = post.find('div', attrs={'class': 'postcolor post_text'})
        comment = text_el.get_text(strip=True) if text_el else ''

        if comment:
            ws.append([name, date, comment])

    wb.save(output_file)
    time.sleep(random.uniform(0.5, 2.0))

print(f"\nDone! Output saved to {output_file}")
