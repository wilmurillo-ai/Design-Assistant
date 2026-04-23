#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel 表格转 JSON 工具
支持多种 JSON 结构输出
"""

import sys
import json
from pathlib import Path
from typing import Optional, List, Dict, Any

# 导入依赖
try:
    import openpyxl
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# 导入工具函数
sys.path.insert(0, str(Path(__file__).parent.parent))
from utils import print_success, print_error, print_info


def excel_to_json_records(excel_path: Path, sheet_name: Optional[str] = None,
                           header_row: int = 1) -> List[Dict[str, Any]]:
    """
    将 Excel 转换为记录列表（每行一个字典）
    
    Args:
        excel_path: Excel 文件路径
        sheet_name: 工作表名称（可选，默认第一个）
        header_row: 标题行号
    
    Returns:
        记录列表
    """
    if HAS_OPENPYXL:
        return _excel_to_json_openpyxl(excel_path, sheet_name, header_row)
    elif HAS_PANDAS:
        return _excel_to_json_pandas(excel_path, sheet_name, header_row)
    else:
        raise ImportError("Please install openpyxl or pandas: pip install openpyxl")


def _excel_to_json_openpyxl(excel_path: Path, sheet_name: Optional[str],
                             header_row: int) -> List[Dict[str, Any]]:
    """使用 openpyxl 转换"""
    wb = load_workbook(str(excel_path), data_only=True)
    
    # 选择工作表
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # 读取标题行
    headers = []
    for cell in ws[header_row]:
        headers.append(str(cell.value) if cell.value else f"Column_{cell.column}")
    
    # 读取数据行
    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        record = {}
        for col_idx, cell in enumerate(row):
            header = headers[col_idx] if col_idx < len(headers) else f"Column_{col_idx + 1}"
            record[header] = cell.value
        
        # 跳过空行
        if any(record.values()):
            records.append(record)
    
    wb.close()
    return records


def _excel_to_json_pandas(excel_path: Path, sheet_name: Optional[str],
                           header_row: int) -> List[Dict[str, Any]]:
    """使用 pandas 转换"""
    df = pd.read_excel(excel_path, sheet_name=sheet_name or 0, header=header_row - 1)
    return df.to_dict('records')


def excel_to_json_grouped(excel_path: Path, group_column: str,
                           sheet_name: Optional[str] = None) -> Dict[str, List[Dict]]:
    """
    将 Excel 转换为分组 JSON（按某列分组）
    
    Args:
        excel_path: Excel 文件路径
        group_column: 分组列名
        sheet_name: 工作表名称
    
    Returns:
        分组字典
    """
    records = excel_to_json_records(excel_path, sheet_name)
    
    grouped = {}
    for record in records:
        if group_column in record:
            key = str(record[group_column])
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(record)
    
    return grouped


def excel_to_json_nested(excel_path: Path, key_columns: List[str],
                          sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    将 Excel 转换为嵌套 JSON（按多列嵌套）
    
    Args:
        excel_path: Excel 文件路径
        key_columns: 键列名列表
        sheet_name: 工作表名称
    
    Returns:
        嵌套字典
    """
    records = excel_to_json_records(excel_path, sheet_name)
    
    nested = {}
    for record in records:
        current = nested
        for i, key_col in enumerate(key_columns[:-1]):
            key = str(record.get(key_col, 'Unknown'))
            if key not in current:
                current[key] = {}
            current = current[key]
        
        # 最后一层存储完整记录
        last_key = str(record.get(key_columns[-1], 'Unknown'))
        if last_key not in current:
            current[last_key] = []
        current[last_key].append(record)
    
    return nested


def excel_to_json_array(excel_path: Path, sheet_name: Optional[str] = None) -> List[List]:
    """
    将 Excel 转换为二维数组（包含标题）
    
    Args:
        excel_path: Excel 文件路径
        sheet_name: 工作表名称
    
    Returns:
        二维数组
    """
    if HAS_OPENPYXL:
        wb = load_workbook(str(excel_path), data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        array = []
        for row in ws.iter_rows(values_only=True):
            array.append(list(row))
        
        wb.close()
        return array
    
    elif HAS_PANDAS:
        df = pd.read_excel(excel_path, sheet_name=sheet_name or 0, header=None)
        return df.values.tolist()
    
    else:
        raise ImportError("Please install openpyxl or pandas")


def get_sheet_names(excel_path: Path) -> List[str]:
    """
    获取 Excel 文件中的所有工作表名称
    
    Args:
        excel_path: Excel 文件路径
    
    Returns:
        工作表名称列表
    """
    if HAS_OPENPYXL:
        wb = load_workbook(str(excel_path), read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    elif HAS_PANDAS:
        xl = pd.ExcelFile(excel_path)
        return xl.sheet_names
    else:
        return []


def convert_excel_to_json(excel_path: Path, output_path: Optional[Path] = None,
                           format_type: str = 'records',
                           sheet_name: Optional[str] = None,
                           group_column: Optional[str] = None,
                           key_columns: Optional[List[str]] = None,
                           indent: int = 2,
                           encoding: str = 'utf-8',
                           verbose: bool = True) -> bool:
    """
    将 Excel 转换为 JSON 文件
    
    Args:
        excel_path: Excel 文件路径
        output_path: 输出 JSON 路径（可选）
        format_type: JSON 格式类型 ('records', 'grouped', 'nested', 'array')
        sheet_name: 工作表名称
        group_column: 分组列名（仅用于 'grouped' 格式）
        key_columns: 键列名列表（仅用于 'nested' 格式）
        indent: JSON 缩进
        encoding: 文件编码
        verbose: 是否显示详细信息
    
    Returns:
        是否转换成功
    """
    if not excel_path.exists():
        if verbose:
            print_error(f"File not found: {excel_path}")
        return False
    
    try:
        if verbose:
            print_info(f"Reading: {excel_path.name}")
        
        # 根据格式类型转换
        if format_type == 'records':
            data = excel_to_json_records(excel_path, sheet_name)
        
        elif format_type == 'grouped':
            if not group_column:
                if verbose:
                    print_error("group_column is required for 'grouped' format")
                return False
            data = excel_to_json_grouped(excel_path, group_column, sheet_name)
        
        elif format_type == 'nested':
            if not key_columns:
                if verbose:
                    print_error("key_columns is required for 'nested' format")
                return False
            data = excel_to_json_nested(excel_path, key_columns, sheet_name)
        
        elif format_type == 'array':
            data = excel_to_json_array(excel_path, sheet_name)
        
        else:
            if verbose:
                print_error(f"Unknown format type: {format_type}")
            return False
        
        # 确定输出路径
        if output_path is None:
            output_path = excel_path.with_suffix('.json')
        
        # 写入 JSON 文件
        with open(output_path, 'w', encoding=encoding) as f:
            json.dump(data, f, indent=indent, ensure_ascii=False)
        
        if verbose:
            print_success(f"Created: {output_path}")
            print_info(f"Records: {len(data) if isinstance(data, list) else 'N/A'}")
        
        return True
    
    except Exception as e:
        if verbose:
            print_error(f"Conversion failed: {e}")
        return False


def batch_convert_excel_to_json(excel_files: List[Path], 
                                 format_type: str = 'records',
                                 verbose: bool = True) -> dict:
    """
    批量转换 Excel 文件为 JSON
    
    Args:
        excel_files: Excel 文件列表
        format_type: JSON 格式类型
        verbose: 是否显示详细信息
    
    Returns:
        转换结果统计
    """
    results = {'success': 0, 'failed': 0}
    
    for i, excel_file in enumerate(excel_files, 1):
        if verbose:
            print(f"\n[{i}/{len(excel_files)}] {excel_file.name}")
        
        success = convert_excel_to_json(
            excel_file,
            format_type=format_type,
            verbose=verbose
        )
        
        if success:
            results['success'] += 1
        else:
            results['failed'] += 1
    
    if verbose:
        print(f"\n{'='*50}")
        print(f"Conversion completed: {results['success']} success, {results['failed']} failed")
    
    return results
