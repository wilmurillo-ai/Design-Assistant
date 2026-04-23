#!/usr/bin/env python3
"""Convert an Excel contact sheet into the JSON payload required by Toingg."""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import List, Dict, Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required for xlsx parsing. Install with `pip install openpyxl`."
    ) from exc

REQUIRED_COLUMNS = {"name", "phone", "context"}


def normalize_header(header: str | None) -> str:
    return (header or "").strip().lower()


def sheet_to_contacts(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("Excel file is empty")

    headers = [normalize_header(str(cell)) for cell in rows[0]]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise SystemExit(
            f"Missing required columns: {', '.join(sorted(missing))}."
            " Ensure the first row has name, phone, context."
        )

    idx = {header: i for i, header in enumerate(headers)}
    contacts: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if row is None:
            continue
        name_value = row[idx["name"]] if idx["name"] < len(row) else None
        phone_value = row[idx["phone"]] if idx["phone"] < len(row) else None
        context_value = row[idx["context"]] if idx["context"] < len(row) else None

        name = str(name_value).strip() if name_value is not None else ""
        if isinstance(phone_value, (int, float)):
            phone = str(int(phone_value)).strip()
        else:
            phone = str(phone_value).strip() if phone_value is not None else ""
        context = str(context_value).strip() if context_value is not None else ""

        if not name or not phone:
            continue
        contacts.append(
            {
                "name": name,
                "phone": phone,
                "extraParams": {"context": context},
            }
        )
    if not contacts:
        raise SystemExit("No valid contacts found in spreadsheet")
    return contacts


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert xlsx contacts to JSON")
    parser.add_argument("xlsx", help="Path to Excel file with name/phone/context columns")
    parser.add_argument(
        "output",
        nargs="?",
        default="contacts.json",
        help="Where to write the JSON array (default: contacts.json)",
    )
    args = parser.parse_args()

    contacts = sheet_to_contacts(Path(args.xlsx))
    with open(args.output, "w", encoding="utf-8") as fh:
        json.dump(contacts, fh, indent=2)
        fh.write("\n")
    print(f"Wrote {len(contacts)} contacts to {args.output}")


if __name__ == "__main__":
    main()
