from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List


def generate_xlsx(spec: Dict[str, Any], out_path: str) -> Path:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError('Missing dependency: openpyxl. Install requirements first.') from exc

    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = 'Sheet1'

    sheets: List[Dict[str, Any]] = spec.get('contentSpec', {}).get('sheets', [])
    if sheets:
        workbook.remove(default_sheet)
        for sheet_spec in sheets:
            ws = workbook.create_sheet(title=sheet_spec.get('name', 'Sheet'))
            columns = sheet_spec.get('columns', [])
            rows = sheet_spec.get('rows', [])
            if columns:
                ws.append(columns)
                for idx, cell in enumerate(ws[1], start=1):
                    cell.font = Font(bold=True)
                    width = max(12, len(str(columns[idx - 1])) + 4)
                    ws.column_dimensions[get_column_letter(idx)].width = width
                ws.freeze_panes = 'A2'
                ws.auto_filter.ref = ws.dimensions
            for row in rows:
                ws.append(row)
    else:
        default_sheet['A1'] = spec.get('title', 'Workbook')

    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))
    return path
