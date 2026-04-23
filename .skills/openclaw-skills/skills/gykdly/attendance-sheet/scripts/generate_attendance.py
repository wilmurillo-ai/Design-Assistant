#!/usr/bin/env python3
"""
Attendance Sheet Generator
Generates professional xlsx attendance sheets from employee work information.
"""

import json
import sys
import argparse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_date(date_str):
    """Parse date string to datetime object."""
    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d",
        "%m-%d",
        "%Y%m%d"
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date: {date_str}")


def generate_date_range(start_date, end_date):
    """Generate list of dates between start and end."""
    start = parse_date(start_date)
    end = parse_date(end_date)
    dates = []
    current = start
    while current <= end:
        dates.append(current)
        current += timedelta(days=1)
    return dates


def create_attendance_sheet(
    employees,
    start_date,
    end_date,
    output_file="attendance_sheet.xlsx",
    attendance_types=None
):
    """
    Generate attendance sheet xlsx file.
    
    Args:
        employees: List of employee names
        start_date: Start date string (YYYY-MM-DD)
        end_date: End date string (YYYY-MM-DD)
        output_file: Output xlsx filename
        attendance_types: Dict mapping status to color hex codes
    """
    
    if attendance_types is None:
        attendance_types = {
            "正常出勤": "C6EFCE",  # Green
            "出勤": "C6EFCE",
            "迟到": "FFC7CE",     # Red
            "早退": "FFEB9C",     # Yellow
            "缺勤": "FFC7CE",     # Red
            "请假": "E2EFDA",     # Light green
            "加班": "DDEBF7",     # Blue
            "休息": "F2F2F2",     # Gray
        }
    
    dates = generate_date_range(start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤表"
    
    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Title
    ws.merge_cells(f'A1:{get_column_letter(len(dates) + 3)}1')
    ws['A1'] = f"{start_date} 至 {end_date} 考勤表"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    row = 3
    ws.cell(row=row, column=1, value="姓名").font = header_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=1).alignment = center_align
    
    for idx, date in enumerate(dates, start=2):
        cell = ws.cell(row=row, column=idx, value=f"{date.month}/{date.day}")
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
    
    ws.cell(row=row, column=len(dates) + 2, value="出勤天数").font = header_font
    ws.cell(row=row, column=len(dates) + 2).border = thin_border
    ws.cell(row=row, column=len(dates) + 2).alignment = center_align
    
    ws.cell(row=row, column=len(dates) + 3, value="备注").font = header_font
    ws.cell(row=row, column=len(dates) + 3).border = thin_border
    ws.cell(row=row, column=len(dates) + 3).alignment = center_align
    
    # Employee rows
    for emp_idx, emp_name in enumerate(employees, start=row + 1):
        ws.cell(row=emp_idx, column=1, value=emp_name).border = thin_border
        ws.cell(row=emp_idx, column=1).alignment = center_align
        
        # Initialize attendance count
        attendance_count = 0
        
        for date_idx, date in enumerate(dates, start=2):
            cell = ws.cell(row=emp_idx, column=date_idx, value="")
            cell.border = thin_border
            cell.alignment = center_align
        
        # Count formula
        ws.cell(row=emp_idx, column=len(dates) + 2, value=f"=COUNTA(B{emp_idx}:{get_column_letter(len(dates))}{emp_idx})")
        ws.cell(row=emp_idx, column=len(dates) + 2).border = thin_border
        ws.cell(row=emp_idx, column=len(dates) + 2).alignment = center_align
        
        ws.cell(row=emp_idx, column=len(dates) + 3, value="").border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    for idx in range(2, len(dates) + 2):
        ws.column_dimensions[get_column_letter(idx)].width = 6
    ws.column_dimensions[get_column_letter(len(dates) + 2)].width = 10
    ws.column_dimensions[get_column_letter(len(dates) + 3)].width = 15
    
    # Freeze panes (freeze first row and first column)
    ws.freeze_panes = 'B4'
    
    # Save
    wb.save(output_file)
    print(f"✅ 考勤表已生成: {output_file}")
    print(f"   员工数: {len(employees)}")
    print(f"   日期范围: {start_date} 至 {end_date} ({len(dates)}天)")
    
    return output_file


def main():
    parser = argparse.ArgumentParser(description="生成考勤表")
    parser.add_argument("--employees", "-e", required=True, help="员工列表，逗号分隔")
    parser.add_argument("--start", "-s", required=True, help="开始日期 YYYY-MM-DD")
    parser.add_argument("--end", "-d", required=True, help="结束日期 YYYY-MM-DD")
    parser.add_argument("--output", "-o", default="attendance_sheet.xlsx", help="输出文件名")
    parser.add_argument("--json", "-j", help="JSON 格式输入")
    
    args = parser.parse_args()
    
    if args.json:
        data = json.loads(args.json)
        employees = data.get("employees", [])
        start_date = data.get("start_date", args.start)
        end_date = data.get("end_date", args.end)
    else:
        employees = [e.strip() for e in args.employees.split(",")]
        start_date = args.start
        end_date = args.end
    
    create_attendance_sheet(employees, start_date, end_date, args.output)


if __name__ == "__main__":
    main()
