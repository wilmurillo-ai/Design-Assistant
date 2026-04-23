#!/usr/bin/env python3
"""
Excel/CSV Format Converter
Converts between Excel (.xlsx/.xls) and CSV formats, with support for
multiple sheets and encoding options.

Dependencies:
    pip install openpyxl xlrd

Usage:
    python excel_csv_converter.py <input_file> --format <target_format> [options]
    python excel_csv_converter.py --batch <input_dir> --format csv [options]
"""

import argparse
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    import csv
except ImportError:
    print("Error: csv module not available")
    sys.exit(1)


# Supported formats
SUPPORTED_FORMATS = {"csv", "xlsx", "xls"}


def detect_format(filepath: str) -> str:
    """Detect file format from extension."""
    ext = Path(filepath).suffix.lower().lstrip(".")
    
    if ext in ("xlsx", "xls", "csv"):
        return ext
    
    raise ValueError(
        f"Cannot determine format for '{filepath}'. "
        f"Supported formats: xlsx, xls, csv"
    )


def convert_csv_to_excel(
    input_path: str,
    output_path: str = None,
    delimiter: str = ",",
    encoding: str = "utf-8",
    sheet_name: str = "Sheet1",
    include_header: bool = True
) -> str:
    """Convert CSV file to Excel format."""
    
    input_file = Path(input_path)
    
    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    
    if output_path is None:
        output_path = input_file.with_suffix(".xlsx")
    else:
        output_path = Path(output_path)
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    print(f"Converting: {input_file.name} -> {output_path.name}")
    
    try:
        # Read CSV with proper encoding
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        row_count = 0
        
        with open(input_file, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                row_count += 1
        
        wb.save(str(output_path))
        
        size_kb = output_path.stat().st_size / 1024
        print(f"✓ Converted successfully ({row_count} rows, {size_kb:.1f}KB)")
        
        return str(output_path)
        
    except UnicodeDecodeError as e:
        print(f"✗ Encoding error: {e}. Try specifying a different encoding with --encoding.")
        raise
    except Exception as e:
        print(f"✗ Error converting {input_file.name}: {e}")
        raise


def convert_excel_to_csv(
    input_path: str,
    output_path: str = None,
    sheet_name: str = None,
    encoding: str = "utf-8",
    delimiter: str = ",",
    quotechar: str = '"',
    include_header: bool = True
) -> list[str]:
    """Convert Excel file to CSV (one file per sheet or specified sheet)."""
    
    input_file = Path(input_path)
    
    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    
    print(f"Reading: {input_file.name}")
    
    try:
        wb = load_workbook(input_path, read_only=True, data_only=True)
        
        results = []
        
        # Determine which sheets to process
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"
                )
            sheets_to_process = [sheet_name]
        else:
            sheets_to_process = wb.sheetnames
        
        for sname in sheets_to_process:
            ws = wb[sname]
            
            # Determine output path
            if output_path:
                out_path = Path(output_path)
                if len(sheets_to_process) > 1 and not sheet_name:
                    out_path = out_path.with_stem(f"{out_path.stem}_{sname}")
            else:
                if len(sheets_to_process) > 1:
                    out_path = input_file.with_name(f"{input_file.stem}_{sname}.csv")
                else:
                    out_path = input_file.with_suffix(".csv")
            
            out_path.parent.mkdir(parents=True, exist_ok=True)
            
            print(f"  Converting sheet '{sname}' -> {out_path.name}")
            
            with open(out_path, "w", newline="", encoding=encoding) as f:
                writer = csv.writer(f, delimiter=delimiter, quotechar=quotechar)
                
                row_count = 0
                header_written = False
                
                for row in ws.iter_rows(values_only=True):
                    # Convert values to strings
                    row_values = [
                        "" if v is None else (
                            str(v) if not isinstance(v, (int, float)) 
                            else v
                        ) 
                        for v in row
                    ]
                    
                    writer.writerow(row_values)
                    
                    if row_count == 0 and not include_header:
                        pass
                    
                    row_count += 1
            
            size_kb = out_path.stat().st_size / 1024
            print(f"  ✓ Saved ({row_count} rows, {size_kb:.1f}KB)")
            results.append(str(out_path))
        
        wb.close()
        return results
        
    except Exception as e:
        print(f"✗ Error converting {input_file.name}: {e}")
        raise


def get_excel_info(excel_path: str) -> dict:
    """Get detailed information about an Excel file."""
    input_file = Path(excel_path)
    
    if not input_file.exists():
        raise FileNotFoundError(f"File not found: {excel_path}")
    
    try:
        wb = load_workbook(input_file, read_only=True, data_only=True)
        
        info = {
            "file": input_file.name,
            "sheets": [],
            "total_sheets": len(wb.sheetnames),
            "file_size_bytes": input_file.stat().st_size,
            "file_size_kb": round(input_file.stat().st_size / 1024, 2),
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            info["sheets"].append({
                "name": sheet_name,
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            })
        
        wb.close()
        return info
        
    except Exception as e:
        return {"error": str(e)}


def batch_convert_files(
    input_dir: str,
    target_format: str,
    output_dir: str = None,
    **kwargs
) -> list[str]:
    """Batch convert all files of source type to target format."""
    
    input_path = Path(input_dir)
    
    if not input_path.is_dir():
        raise NotADirectoryError(f"Input directory not found: {input_dir}")
    
    target_format = target_format.lower().lstrip(".")
    
    if target_format not in SUPPORTED_FORMATS:
        raise ValueError(f"Unsupported target format: {target_format}")
    
    if output_dir is None:
        output_dir = input_dir
    
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    results = []
    errors = []
    
    # Determine source extensions based on target format
    if target_format == "csv":
        source_exts = [".xlsx", ".xls"]
    elif target_format in ("xlsx", "xls"):
        source_exts = [".csv"]
    else:
        raise ValueError(f"No conversion path defined for target format: {target_format}")
    
    # Find all matching files
    files_to_convert = []
    for ext in source_exts:
        files_to_convert.extend(input_path.glob(f"*{ext}"))
        files_to_convert.extend(input_path.glob(f"*{ext.upper()}"))
    
    if not files_to_convert:
        print(f"No convertible files found in: {input_dir}")
        return []
    
    print(f"Found {len(files_to_convert)} file(s) to convert")
    
    for src_file in sorted(files_to_convert):
        try:
            src_fmt = detect_format(str(src_file))
            
            if src_fmt == "csv" and target_format in ("xlsx", "xls"):
                result = convert_csv_to_excel(str(src_file), **kwargs)
                results.append(result)
                
            elif src_fmt in ("xlsx", "xls") and target_format == "csv":
                result_list = convert_excel_to_csv(str(src_file), **kwargs)
                results.extend(result_list)
            
            # Update kwargs with specific output path for batch mode
            if "output_path" not in kwargs:
                continue
                
        except Exception as e:
            errors.append((src_file.name, str(e)))
            continue
    
    summary = f"\nConverted {len(results)}/{len(files_to_convert)} operations successful"
    if errors:
        summary += f"\nErrors ({len(errors)}):"
        for fname, err in errors[:5]:
            summary += f"\n  - {fname}: {err}"
    print(summary)
    
    return results


def main():
    parser = argparse.ArgumentParser(
        description="Convert between Excel (.xlsx/.xls) and CSV formats",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Supported Formats:
  xlsx, xls, csv

Examples:
  python excel_csv_converter.py data.csv --format xlsx
  python excel_csv_converter.py data.xlsx --format csv
  python excel_csv_converter.py data.xlsx --format csv --sheet "Sales Data"
  python excel_csv_converter.py data.csv --format xlsx --delimiter ";" --encoding gbk
  python excel_csv_converter.py --batch ./data_folder --format csv
  python excel_csv_converter.py data.xlsx --info
        """
    )
    
    parser.add_argument("input", nargs="?", help="Input file or directory (with --batch)")
    parser.add_argument("--format", "-f", dest="target_format", help="Target format (csv/xlsx/xls)")
    parser.add_argument("--output", "-o", help="Output file/directory path")
    parser.add_argument("--batch", "-b", action="store_true", help="Batch mode")
    parser.add_argument("--output-dir", help="Output directory for batch mode")
    parser.add_argument("--sheet", "-s", help="Specific sheet name (Excel to CSV only)")
    parser.add_argument("--delimiter", "-d", default=",", help="CSV delimiter (default: comma)")
    parser.add_argument("--encoding", "-e", default="utf-8", help="File encoding (default: utf-8)")
    parser.add_argument("--no-header", action="store_true", dest="no_header", help="Skip header row")
    parser.add_argument("--info", "-i", action="store_true", help="Show file information")
    parser.add_argument("--sheet-name", default="Sheet1", help="Output sheet name (CSV to Excel)")
    
    args = parser.parse_args()
    
    try:
        if args.info:
            info = get_excel_info(args.input)
            print("\n=== File Information ===\n")
            
            if "error" in info:
                print(f"Error: {info['error']}")
                return
            
            print(f"File: {info['file']}")
            print(f"Size: {info['file_size_kb']} KB")
            print(f"Sheets: {info['total_sheets']}\n")
            
            for sheet in info["sheets"]:
                print(f"  Sheet: {sheet['name']}")
                print(f"    Dimensions: {sheet['dimensions']}")
                print(f"    Rows: {sheet['max_row']}, Columns: {sheet['max_column']}")
            return
        
        if not args.target_format and not args.info:
            parser.error("Target format required (--format). Use --info to view details.")
        
        kwargs = {
            "delimiter": args.delimiter,
            "encoding": args.encoding,
            "include_header": not args.no_header,
            "sheet_name": args.sheet_name,
        }
        
        if args.batch:
            batch_convert_files(args.input, args.target_format, args.output_dir or args.output, **kwargs)
        else:
            src_fmt = detect_format(args.input)
            
            if src_fmt == "csv" and args.target_format in ("xlsx", "xls"):
                convert_csv_to_excel(args.input, args.output, **kwargs)
            elif src_fmt in ("xlsx", "xls") and args.target_format == "csv":
                convert_excel_to_csv(
                    args.input, 
                    args.output, 
                    sheet_name=args.sheet,
                    **kwargs
                )
            else:
                print(f"Conversion from {src_fmt} to {args.target_format} is not supported directly.")
                sys.exit(1)
                
    except FileNotFoundError as e:
        print(f"Error: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
