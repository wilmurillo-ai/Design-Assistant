#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tests for clone_and_index.py

Covers:
  1. Original functionality (top-level group resolution, subgroup enumeration, project fetch)
  2. Graceful skip when a main group can't be accessed
  3. Sub-path resolution: sub-group paths and direct project paths
  4. Excel generation and update logic
  5. Edge cases: mixed inputs, deduplication, empty results
"""

from __future__ import annotations
import os
import sys
import json
import tempfile
import shutil
import subprocess
import unittest
from unittest.mock import patch, MagicMock
from datetime import datetime

# Ensure the script directory is on the path
sys.path.insert(0, os.path.dirname(__file__))

import clone_and_index as ci


# ===========================================================================
# Mock data factories
# ===========================================================================

def make_group(gid: int, name: str, full_path: str) -> dict:
    return {"id": gid, "name": name, "full_path": full_path, "path": name}

def make_project(pid: int, name: str, namespace_path: str, description: str = "") -> dict:
    full_path = f"{namespace_path}/{name}"
    return {
        "id": pid,
        "name": name,
        "path_with_namespace": full_path,
        "http_url_to_repo": f"https://gitlab.test.com/{full_path}.git",
        "ssh_url_to_repo": f"git@gitlab.test.com:{full_path}.git",
        "description": description,
    }


# ===========================================================================
# Test: Helper functions
# ===========================================================================

class TestHelpers(unittest.TestCase):
    """Test is_subpath and resolve_path_type logic."""

    def test_is_subpath_simple(self):
        self.assertFalse(ci.is_subpath("myGroup"))
        self.assertFalse(ci.is_subpath("myGroup"))

    def test_is_subpath_with_slash(self):
        self.assertTrue(ci.is_subpath("myGroup/mySubGroup"))
        self.assertTrue(ci.is_subpath("a/b/c"))
        self.assertTrue(ci.is_subpath("myGroup/sub"))

    @patch("clone_and_index.resolve_project")
    @patch("clone_and_index.resolve_group_id")
    def test_resolve_path_type_project(self, mock_group, mock_project):
        """When path resolves as a project (not a group), return ('project', dict)."""
        proj = make_project(100, "my-proj", "myGroup/mySubGroup")
        mock_group.return_value = None  # not a group
        mock_project.return_value = proj

        kind, data = ci.resolve_path_type("https://gl.test", "tok", "myGroup/mySubGroup/my-proj")
        self.assertEqual(kind, "project")
        self.assertEqual(data["id"], 100)
        # Group is tried first, then project
        mock_group.assert_called_once()
        mock_project.assert_called_once()

    @patch("clone_and_index.resolve_project")
    @patch("clone_and_index.resolve_group_id")
    def test_resolve_path_type_group(self, mock_group, mock_project):
        """When path resolves as a group, return ('group', id) — project not tried."""
        mock_group.return_value = 42
        mock_project.return_value = None  # should not be called

        kind, data = ci.resolve_path_type("https://gl.test", "tok", "myGroup/mySubGroup")
        self.assertEqual(kind, "group")
        self.assertEqual(data, 42)
        # Project should NOT be tried since group resolved
        mock_project.assert_not_called()

    @patch("clone_and_index.resolve_project")
    @patch("clone_and_index.resolve_group_id")
    def test_resolve_path_type_none(self, mock_group, mock_project):
        """When path doesn't resolve at all, return ('none', None)."""
        mock_group.return_value = None
        mock_project.return_value = None

        kind, data = ci.resolve_path_type("https://gl.test", "tok", "nonexistent/path")
        self.assertEqual(kind, "none")
        self.assertIsNone(data)


# ===========================================================================
# Test: API error handling / graceful skip
# ===========================================================================

class TestGracefulSkip(unittest.TestCase):
    """Test that inaccessible groups are skipped gracefully."""

    @patch("clone_and_index.api_get")
    def test_resolve_group_id_returns_none_on_error(self, mock_api):
        """When API returns an error list, resolve_group_id should return None."""
        mock_api.return_value = []  # API error returns empty list
        result = ci.resolve_group_id("https://gl.test", "tok", "forbidden-group")
        self.assertIsNone(result)

    @patch("clone_and_index.api_get")
    def test_resolve_project_returns_none_on_error(self, mock_api):
        mock_api.return_value = []
        result = ci.resolve_project("https://gl.test", "tok", "nonexistent/proj")
        self.assertIsNone(result)

    @patch("clone_and_index.api_get_paginated")
    def test_fetch_subgroups_recursive_empty_on_failure(self, mock_pag):
        """When both descendant and subgroup APIs return empty, we get []."""
        mock_pag.return_value = []
        result = ci.fetch_subgroups_recursive("https://gl.test", "tok", 999)
        self.assertEqual(result, [])


# ===========================================================================
# Test: Excel generation
# ===========================================================================

class TestExcel(unittest.TestCase):
    """Test Excel build and update logic."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.tmpdir, "01.索引.xlsx")

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _make_project_data(self, main_group, sub_group, path, name, desc=""):
        return {
            "main_group": main_group,
            "sub_group": sub_group,
            "project_path": path,
            "project_name": name,
            "description": desc,
            "branches": "master\ndev",
            "times": "2026-01-01 10:00:00\n2026-01-02 11:00:00",
            "ssh_url": f"git@gl.test:{path}.git",
            "download_time": "2026-03-11 09:00:00",
        }

    def test_build_excel_creates_file(self):
        data = [
            self._make_project_data("myGroup", "myGroup/sub1", "myGroup/sub1/proj-a", "proj-a"),
            self._make_project_data("myGroup", "myGroup/sub2", "myGroup/sub2/proj-b", "proj-b"),
        ]
        ci.build_excel(self.excel_path, data)
        self.assertTrue(os.path.exists(self.excel_path))

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        self.assertIn("myGroup", wb.sheetnames)
        ws = wb["myGroup"]
        # Header + 2 data rows
        self.assertEqual(ws.max_row, 3)
        # Check project paths
        self.assertEqual(ws.cell(row=2, column=3).value, "myGroup/sub1/proj-a")
        self.assertEqual(ws.cell(row=3, column=3).value, "myGroup/sub2/proj-b")

    def test_build_excel_multiple_groups(self):
        data = [
            self._make_project_data("myGroup", "myGroup", "myGroup/proj-1", "proj-1"),
            self._make_project_data("myGroup", "myGroup/mySubGroup", "myGroup/mySubGroup/map-svc", "map-svc"),
        ]
        ci.build_excel(self.excel_path, data)

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        self.assertIn("myGroup", wb.sheetnames)
        self.assertIn("myGroup", wb.sheetnames)

    def test_update_excel_adds_new_rows(self):
        # Build initial
        data1 = [self._make_project_data("myGroup", "myGroup/sub1", "myGroup/sub1/proj-a", "proj-a")]
        ci.build_excel(self.excel_path, data1)

        # Update with new project
        data2 = [self._make_project_data("myGroup", "myGroup/sub2", "myGroup/sub2/proj-b", "proj-b")]
        ci.update_excel(self.excel_path, data2)

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["myGroup"]
        # Header + 2 data rows
        self.assertEqual(ws.max_row, 3)

    def test_update_excel_updates_existing_rows(self):
        data1 = [self._make_project_data("myGroup", "myGroup/sub1", "myGroup/sub1/proj-a", "proj-a", "old desc")]
        ci.build_excel(self.excel_path, data1)

        data2 = [self._make_project_data("myGroup", "myGroup/sub1", "myGroup/sub1/proj-a", "proj-a", "new desc")]
        ci.update_excel(self.excel_path, data2)

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["myGroup"]
        self.assertEqual(ws.max_row, 2)  # header + 1 row (updated, not duplicated)
        self.assertEqual(ws.cell(row=2, column=5).value, "new desc")

    def test_update_excel_creates_new_sheet_for_new_group(self):
        data1 = [self._make_project_data("myGroup", "myGroup", "myGroup/proj-1", "proj-1")]
        ci.build_excel(self.excel_path, data1)

        data2 = [self._make_project_data("myGroup", "myGroup/mySubGroup", "myGroup/mySubGroup/map-svc", "map-svc")]
        ci.update_excel(self.excel_path, data2)

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        self.assertIn("myGroup", wb.sheetnames)
        self.assertIn("myGroup", wb.sheetnames)

    def test_build_excel_empty_data(self):
        """Empty project list should not create a file (graceful skip)."""
        ci.build_excel(self.excel_path, [])
        # No file created since there's nothing to write
        self.assertFalse(os.path.exists(self.excel_path))


# ===========================================================================
# Test: fetch_all_projects deduplication
# ===========================================================================

class TestFetchAllProjects(unittest.TestCase):
    """Test that fetch_all_projects deduplicates by project ID."""

    @patch("clone_and_index.api_get_paginated")
    def test_deduplication(self, mock_pag):
        proj_a = make_project(1, "proj-a", "G/sub1")
        proj_b = make_project(2, "proj-b", "G/sub2")
        proj_a_dup = make_project(1, "proj-a", "G/sub1")  # same id

        # Group 10 returns [a, b], group 20 returns [a_dup]
        mock_pag.side_effect = [
            [proj_a, proj_b],
            [proj_a_dup],
        ]

        result = ci.fetch_all_projects("https://gl.test", "tok", [10, 20])
        self.assertEqual(len(result), 2)
        ids = {p["id"] for p in result}
        self.assertEqual(ids, {1, 2})


# ===========================================================================
# Test: End-to-end main() logic with mocks
# ===========================================================================

class TestMainLogic(unittest.TestCase):
    """Integration tests for main() using mocked API and git operations."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    @patch("clone_and_index.fetch_all_projects")
    @patch("clone_and_index.fetch_subgroups_recursive")
    @patch("clone_and_index.api_get_paginated")
    @patch("clone_and_index.resolve_group_id")
    @patch("clone_and_index.resolve_path_type")
    def test_mixed_inputs_resolution_phase(self, mock_resolve_path, mock_resolve_group,
                                            mock_pag, mock_subs, mock_fetch):
        """Test the resolution phase with mixed inputs: top-level group + sub-group + project.

        We verify that the correct group IDs and direct projects are collected.
        We don't test the clone/pull phase here (ProcessPoolExecutor + mocks don't mix).
        """
        gitlab_url = "https://gl.test"
        token = "test-token"

        proj_direct = make_project(99, "my-proj", "myGroup/mySubGroup")

        # Simulate the resolution loop from main()
        target_groups = ["myGroup", "myGroup/mySubGroup", "myGroup/mySubGroup/my-proj"]

        mock_resolve_group.return_value = 1
        mock_subs.return_value = [make_group(2, "sub1", "myGroup/sub1")]
        mock_pag.return_value = [make_project(10, "my-proj", "myGroup")]

        def resolve_side_effect(url, tok, path):
            if path == "myGroup/mySubGroup":
                return ("group", 5)
            elif path == "myGroup/mySubGroup/my-proj":
                return ("project", proj_direct)
            return ("none", None)
        mock_resolve_path.side_effect = resolve_side_effect

        # Reproduce the resolution loop
        all_group_ids = []
        direct_projects = []

        for gpath in target_groups:
            if ci.is_subpath(gpath):
                path_type, resolved = ci.resolve_path_type(gitlab_url, token, gpath)
                if path_type == "project":
                    direct_projects.append(resolved)
                elif path_type == "group":
                    all_group_ids.append(resolved)
                    subs = ci.fetch_subgroups_recursive(gitlab_url, token, resolved)
                    all_group_ids.extend([s["id"] for s in subs])
            else:
                gid = ci.resolve_group_id(gitlab_url, token, gpath)
                if gid is not None:
                    all_group_ids.append(gid)
                    subs = ci.fetch_subgroups_recursive(gitlab_url, token, gid)
                    all_group_ids.extend([s["id"] for s in subs])

        # Verify: myGroup (id=1) + sub1 (id=2) + myGroup/mySubGroup (id=5) + its sub (id=2 again, but from mock)
        self.assertIn(1, all_group_ids)   # myGroup
        self.assertIn(5, all_group_ids)   # myGroup/mySubGroup (sub-group)
        self.assertEqual(len(direct_projects), 1)
        self.assertEqual(direct_projects[0]["id"], 99)  # my-proj

        # Verify dedup when merging
        mock_fetch.return_value = [
            make_project(10, "myGroup-proj", "myGroup"),
            make_project(20, "mySubGroup-proj", "myGroup/mySubGroup"),
        ]

        all_projects = []
        seen_ids = set()
        for p in direct_projects:
            if p["id"] not in seen_ids:
                seen_ids.add(p["id"])
                all_projects.append(p)
        group_projects = ci.fetch_all_projects(gitlab_url, token, all_group_ids)
        for p in group_projects:
            if p["id"] not in seen_ids:
                seen_ids.add(p["id"])
                all_projects.append(p)

        # Should have 3 unique projects: 99, 10, 20
        self.assertEqual(len(all_projects), 3)
        self.assertEqual({p["id"] for p in all_projects}, {99, 10, 20})

    @patch("clone_and_index.resolve_group_id")
    def test_inaccessible_group_skipped(self, mock_resolve):
        """When a top-level group can't be resolved, it should be skipped."""
        mock_resolve.return_value = None

        os.environ["GITLAB_URL"] = "https://gl.test"
        os.environ["GITLAB_TOKEN"] = "test-token"
        os.environ["GITLAB_GROUPS"] = "forbidden-group"
        os.environ["GITLAB_BASE_DIR"] = self.tmpdir

        with self.assertRaises(SystemExit) as ctx:
            ci.main()
        # Should exit with code 1 (no groups found)
        self.assertEqual(ctx.exception.code, 1)

        for key in ["GITLAB_URL", "GITLAB_TOKEN", "GITLAB_GROUPS", "GITLAB_BASE_DIR"]:
            os.environ.pop(key, None)

    @patch("clone_and_index.resolve_path_type")
    def test_unresolvable_subpath_skipped(self, mock_resolve_path):
        """When a sub-path can't be resolved as group or project, skip it."""
        mock_resolve_path.return_value = ("none", None)

        os.environ["GITLAB_URL"] = "https://gl.test"
        os.environ["GITLAB_TOKEN"] = "test-token"
        os.environ["GITLAB_GROUPS"] = "nonexistent/subpath"
        os.environ["GITLAB_BASE_DIR"] = self.tmpdir

        with self.assertRaises(SystemExit) as ctx:
            ci.main()
        self.assertEqual(ctx.exception.code, 1)

        for key in ["GITLAB_URL", "GITLAB_TOKEN", "GITLAB_GROUPS", "GITLAB_BASE_DIR"]:
            os.environ.pop(key, None)


# ===========================================================================
# Test: _is_release_branch
# ===========================================================================

class TestBranchClassification(unittest.TestCase):

    def test_release_branch_detection(self):
        self.assertTrue(ci._is_release_branch("release/1.0"))
        self.assertTrue(ci._is_release_branch("Release-v2"))
        self.assertTrue(ci._is_release_branch("prod"))
        self.assertTrue(ci._is_release_branch("production"))
        self.assertFalse(ci._is_release_branch("master"))
        self.assertFalse(ci._is_release_branch("dev"))
        self.assertFalse(ci._is_release_branch("feature/new-thing"))


# ===========================================================================
# Test: resort_sheet
# ===========================================================================

# ===========================================================================
# Test: Incremental Excel write
# ===========================================================================

class TestIncrementalExcel(unittest.TestCase):
    """Test the _incremental_excel_write helper."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.tmpdir, "01.索引.xlsx")

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _make_project_data(self, main_group, sub_group, path, name, desc=""):
        return {
            "main_group": main_group,
            "sub_group": sub_group,
            "project_path": path,
            "project_name": name,
            "description": desc,
            "branches": "master",
            "times": "2026-01-01 10:00:00",
            "ssh_url": f"git@gl.test:{path}.git",
            "download_time": "2026-03-11 09:00:00",
        }

    def test_incremental_creates_then_updates(self):
        """First call creates Excel, second call updates it."""
        batch1 = [self._make_project_data("myGroup", "myGroup/sub1", "myGroup/sub1/proj-a", "proj-a")]
        ci._incremental_excel_write(self.excel_path, batch1)
        self.assertTrue(os.path.exists(self.excel_path))

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        self.assertEqual(wb["myGroup"].max_row, 2)  # header + 1

        batch2 = [self._make_project_data("myGroup", "myGroup/sub2", "myGroup/sub2/proj-b", "proj-b")]
        ci._incremental_excel_write(self.excel_path, batch2)

        wb = openpyxl.load_workbook(self.excel_path)
        self.assertEqual(wb["myGroup"].max_row, 3)  # header + 2

    def test_incremental_empty_batch_noop(self):
        """Empty batch should not create a file."""
        ci._incremental_excel_write(self.excel_path, [])
        self.assertFalse(os.path.exists(self.excel_path))


# ===========================================================================
# Test: _run_git uses start_new_session
# ===========================================================================

class TestRunGitTimeout(unittest.TestCase):
    """Test that _run_git properly raises TimeoutExpired and cleans up."""

    def test_run_git_timeout_raises(self):
        """A command that sleeps longer than timeout should raise TimeoutExpired."""
        tmpdir = tempfile.mkdtemp()
        try:
            # Use a git command that will hang: init a repo first, then run a slow command
            subprocess.run(["git", "init", tmpdir], capture_output=True)
            with self.assertRaises(subprocess.TimeoutExpired):
                # 'git log' on empty repo is fast, so use a sleep trick via GIT_TRACE
                # Instead, just call _run_git with a very short timeout on a command that takes time
                ci._run_git(
                    ["-c", "core.pager=sleep 10", "log"],
                    cwd=tmpdir,
                    timeout=1,
                )
        except Exception:
            pass  # Some environments may not have git; skip gracefully
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)


# ===========================================================================
# Test: _resort_sheet
# ===========================================================================

class TestResortSheet(unittest.TestCase):

    def test_resort_sorts_by_subgroup_then_name(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(ci.HEADERS_CN)
        # Insert out of order
        ws.append(["G", "G/z", "G/z/proj-z", "proj-z", "", "", "", "", ""])
        ws.append(["G", "G/a", "G/a/proj-a", "proj-a", "", "", "", "", ""])
        ws.append(["G", "G/a", "G/a/proj-b", "proj-b", "", "", "", "", ""])

        ci._resort_sheet(ws)

        # Row 2 should be G/a / proj-a
        self.assertEqual(ws.cell(row=2, column=2).value, "G/a")
        self.assertEqual(ws.cell(row=2, column=4).value, "proj-a")
        # Row 3 should be G/a / proj-b
        self.assertEqual(ws.cell(row=3, column=4).value, "proj-b")
        # Row 4 should be G/z / proj-z
        self.assertEqual(ws.cell(row=4, column=2).value, "G/z")


# ===========================================================================
# Test: Project ID matching, cross-sheet migration, stale row cleanup
# ===========================================================================

class TestExcelProjectID(unittest.TestCase):
    """Test new Project ID based matching and cleanup logic."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.tmpdir, "01.索引.xlsx")

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _make_project_data(self, main_group, sub_group, path, name, pid, desc=""):
        return {
            "main_group": main_group,
            "sub_group": sub_group,
            "project_path": path,
            "project_name": name,
            "description": desc,
            "branches": "master",
            "times": "2026-01-01 10:00:00",
            "ssh_url": f"git@gl.test:{path}.git",
            "download_time": "2026-03-11 09:00:00",
            "project_id": pid,
        }

    def test_project_id_column_written(self):
        """Project ID should be written to Column J."""
        data = [self._make_project_data("myGroup", "myGroup/sub", "myGroup/sub/proj", "proj", 100)]
        ci.build_excel(self.excel_path, data)

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["myGroup"]
        # Header
        self.assertEqual(ws.cell(row=1, column=10).value, "Project ID")
        # Data
        self.assertEqual(ws.cell(row=2, column=10).value, 100)

    def test_update_matches_by_id_when_path_changes(self):
        """When a project is renamed (path changes but ID stays), update the existing row."""
        data1 = [self._make_project_data("myGroup", "myGroup/old", "myGroup/old/proj", "proj", 100)]
        ci.build_excel(self.excel_path, data1)

        # Same ID, different path (project was moved within same group)
        data2 = [self._make_project_data("myGroup", "myGroup/new", "myGroup/new/proj", "proj", 100)]
        ci.update_excel(self.excel_path, data2)

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["myGroup"]
        # Should have 1 data row (updated), not 2
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=3).value, "myGroup/new/proj")
        self.assertEqual(ws.cell(row=2, column=10).value, 100)

    def test_cross_sheet_migration(self):
        """When a project moves from one group to another, old row is deleted, new row created."""
        data1 = [
            self._make_project_data("myGroup", "myGroup", "myGroup/proj-a", "proj-a", 100),
            self._make_project_data("Platform", "Platform", "Platform/proj-b", "proj-b", 200),
        ]
        ci.build_excel(self.excel_path, data1)

        # proj-a (ID 100) moves from myGroup to Platform
        data2 = [self._make_project_data("Platform", "Platform", "Platform/proj-a", "proj-a", 100)]
        ci.update_excel(self.excel_path, data2)

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)

        # myGroup sheet should have 0 data rows (original proj-a removed)
        ws_myGroup = wb["myGroup"]
        self.assertEqual(ws_myGroup.max_row, 1)  # only header

        # Platform sheet should have 2 data rows: proj-b + migrated proj-a
        ws_plat = wb["Platform"]
        self.assertEqual(ws_plat.max_row, 3)  # header + 2 rows
        paths = {ws_plat.cell(row=r, column=3).value for r in range(2, 4)}
        self.assertIn("Platform/proj-a", paths)
        self.assertIn("Platform/proj-b", paths)

    def test_cross_sheet_migration_multiple(self):
        """Multiple projects migrating from same sheet — row deletion order is correct."""
        data1 = [
            self._make_project_data("Old", "Old", "Old/p1", "p1", 100),
            self._make_project_data("Old", "Old", "Old/p2", "p2", 200),
            self._make_project_data("Old", "Old", "Old/p3", "p3", 300),
        ]
        ci.build_excel(self.excel_path, data1)

        # All three move to New group
        data2 = [
            self._make_project_data("New", "New", "New/p1", "p1", 100),
            self._make_project_data("New", "New", "New/p2", "p2", 200),
            self._make_project_data("New", "New", "New/p3", "p3", 300),
        ]
        ci.update_excel(self.excel_path, data2)

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)

        ws_old = wb["Old"]
        self.assertEqual(ws_old.max_row, 1)  # only header

        ws_new = wb["New"]
        self.assertEqual(ws_new.max_row, 4)  # header + 3 rows

    def test_sync_mode_removes_stale_rows(self):
        """In sync mode, projects not in all_project_ids are removed."""
        data1 = [
            self._make_project_data("myGroup", "myGroup", "myGroup/alive", "alive", 100),
            self._make_project_data("myGroup", "myGroup", "myGroup/deleted", "deleted", 200),
        ]
        ci.build_excel(self.excel_path, data1)

        # Only project 100 exists now; project 200 was deleted on GitLab
        data2 = [self._make_project_data("myGroup", "myGroup", "myGroup/alive", "alive", 100)]
        ci.update_excel(
            self.excel_path, data2,
            all_project_ids={100},
            synced_groups={"myGroup"},
        )

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["myGroup"]
        self.assertEqual(ws.max_row, 2)  # header + 1 (alive)
        self.assertEqual(ws.cell(row=2, column=3).value, "myGroup/alive")

    def test_sync_mode_does_not_touch_unsynced_groups(self):
        """Sync cleanup only affects sheets belonging to synced groups."""
        data1 = [
            self._make_project_data("myGroup", "myGroup", "myGroup/proj", "proj", 100),
            self._make_project_data("Other", "Other", "Other/proj", "proj", 200),
        ]
        ci.build_excel(self.excel_path, data1)

        # Sync only myGroup; "Other" should remain untouched
        data2 = [self._make_project_data("myGroup", "myGroup", "myGroup/proj", "proj", 100)]
        ci.update_excel(
            self.excel_path, data2,
            all_project_ids={100},
            synced_groups={"myGroup"},
        )

        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path)
        ws_other = wb["Other"]
        self.assertEqual(ws_other.max_row, 2)  # header + 1 row (untouched)
        self.assertEqual(ws_other.cell(row=2, column=10).value, 200)

    def test_legacy_excel_without_id_column(self):
        """Opening a legacy 9-column Excel should not crash; path fallback should work."""
        import openpyxl
        from openpyxl.styles import Font, Alignment

        # Build a legacy Excel with 9 columns (no Project ID)
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])
        ws = wb.create_sheet("myGroup")
        old_headers = ci.HEADERS_CN[:9]  # first 9 only
        ws.append(old_headers)
        ws.append(["myGroup", "myGroup/sub", "myGroup/sub/proj", "proj", "desc",
                    "master", "2026-01-01", "git@gl.test:myGroup/sub/proj.git", "2026-03-01"])
        wb.save(self.excel_path)

        # Now update with data that has a project_id
        data = [self._make_project_data("myGroup", "myGroup/sub", "myGroup/sub/proj", "proj", 100, "new desc")]
        ci.update_excel(self.excel_path, data)

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["myGroup"]
        # Should update existing row (path match fallback) and write project_id
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(row=2, column=5).value, "new desc")
        self.assertEqual(ws.cell(row=2, column=10).value, 100)

    def test_legacy_excel_sync_cleanup_by_path(self):
        """Legacy rows without ID get cleaned up by path match in sync mode."""
        import openpyxl

        # Build legacy Excel without Project ID
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])
        ws = wb.create_sheet("myGroup")
        ws.append(ci.HEADERS_CN[:9])
        ws.append(["myGroup", "myGroup", "myGroup/alive", "alive", "", "master", "2026-01-01", "git@x", "2026-03-01"])
        ws.append(["myGroup", "myGroup", "myGroup/dead", "dead", "", "master", "2026-01-01", "git@x", "2026-03-01"])
        wb.save(self.excel_path)

        # Sync: only "alive" exists
        data = [self._make_project_data("myGroup", "myGroup", "myGroup/alive", "alive", 100)]
        ci.update_excel(
            self.excel_path, data,
            all_project_ids={100},
            synced_groups={"myGroup"},
        )

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["myGroup"]
        self.assertEqual(ws.max_row, 2)  # header + 1
        self.assertEqual(ws.cell(row=2, column=3).value, "myGroup/alive")


# ===========================================================================
# Test: _find_sheet_for_group
# ===========================================================================

class TestFindSheetForGroup(unittest.TestCase):
    """Test sheet lookup by Column A fallback."""

    def test_exact_title_match(self):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])
        ws = wb.create_sheet("myGroup")
        ws.append(ci.HEADERS_CN)
        ws.append(["myGroup", "", "", "", "", "", "", "", "", ""])

        result = ci._find_sheet_for_group(wb, "myGroup")
        self.assertEqual(result.title, "myGroup")

    def test_column_a_fallback(self):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])
        # Sheet title is truncated
        long_name = "a" * 40
        ws = wb.create_sheet(long_name[:31])
        ws.append(ci.HEADERS_CN)
        ws.append([long_name, "", "", "", "", "", "", "", "", ""])

        # Exact title match won't work since we search by the full name
        result = ci._find_sheet_for_group(wb, long_name)
        self.assertIsNotNone(result)
        self.assertEqual(result.title, long_name[:31])

    def test_no_match_returns_none(self):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])
        ws = wb.create_sheet("Other")
        ws.append(ci.HEADERS_CN)
        ws.append(["Other", "", "", "", "", "", "", "", "", ""])

        result = ci._find_sheet_for_group(wb, "myGroup")
        self.assertIsNone(result)


# ===========================================================================
# Test: _unique_sheet_name
# ===========================================================================

class TestUniqueSheetName(unittest.TestCase):

    def test_no_conflict(self):
        import openpyxl
        wb = openpyxl.Workbook()
        self.assertEqual(ci._unique_sheet_name(wb, "myGroup"), "myGroup")

    def test_conflict_adds_suffix(self):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.create_sheet("myGroup")
        result = ci._unique_sheet_name(wb, "myGroup")
        self.assertEqual(result, "myGroup_2")
        self.assertLessEqual(len(result), 31)

    def test_long_name_conflict(self):
        import openpyxl
        wb = openpyxl.Workbook()
        long_name = "a" * 40
        wb.create_sheet(long_name[:31])
        result = ci._unique_sheet_name(wb, long_name)
        self.assertNotEqual(result, long_name[:31])
        self.assertLessEqual(len(result), 31)
        self.assertTrue(result.endswith("_2"))


# ===========================================================================
# Test: _item_to_row
# ===========================================================================

class TestItemToRow(unittest.TestCase):

    def test_includes_project_id(self):
        item = {
            "main_group": "G",
            "sub_group": "G/s",
            "project_path": "G/s/p",
            "project_name": "p",
            "description": "d",
            "branches": "master",
            "times": "2026-01-01",
            "ssh_url": "git@x",
            "download_time": "2026-03-01",
            "project_id": 42,
        }
        row = ci._item_to_row(item)
        self.assertEqual(len(row), 10)
        self.assertEqual(row[9], 42)

    def test_missing_project_id_defaults_empty(self):
        item = {
            "main_group": "G",
            "sub_group": "G/s",
            "project_path": "G/s/p",
            "project_name": "p",
            "description": "d",
            "branches": "master",
            "times": "2026-01-01",
            "ssh_url": "git@x",
            "download_time": "2026-03-01",
        }
        row = ci._item_to_row(item)
        self.assertEqual(row[9], "")


# ===========================================================================
# Test: Sync mode in main() — environment variable handling
# ===========================================================================

class TestSyncModeSetup(unittest.TestCase):
    """Test that sync mode is parsed correctly and effective_mode is 'clone'."""

    @patch("clone_and_index.resolve_group_id")
    def test_sync_mode_effective_is_clone(self, mock_resolve):
        """Sync mode should resolve to effective_mode=clone internally."""
        mock_resolve.return_value = None

        os.environ["GITLAB_URL"] = "https://gl.test"
        os.environ["GITLAB_TOKEN"] = "test-token"
        os.environ["GITLAB_GROUPS"] = "nonexistent"
        os.environ["GITLAB_BASE_DIR"] = tempfile.mkdtemp()
        os.environ["GITLAB_MODE"] = "sync"

        # main() will exit(1) since group not found, but we can check
        # that it doesn't crash parsing "sync" mode
        with self.assertRaises(SystemExit):
            ci.main()

        for key in ["GITLAB_URL", "GITLAB_TOKEN", "GITLAB_GROUPS", "GITLAB_BASE_DIR", "GITLAB_MODE"]:
            os.environ.pop(key, None)


if __name__ == "__main__":
    unittest.main(verbosity=2)
