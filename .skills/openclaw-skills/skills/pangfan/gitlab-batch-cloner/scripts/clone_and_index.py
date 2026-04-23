#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GitLab Batch Clone & Index Generator

Clones all projects from GitLab group(s), checks out key branches,
and generates an Excel index with per-group sheets.

Supports:
  - Top-level groups: "myGroup" → enumerate all subgroups & projects under myGroup
  - Sub-group paths:  "myGroup/mySubGroup" → directly sync mySubGroup (and its
    descendants) without enumerating the entire myGroup tree
  - Direct project paths: "myGroup/mySubGroup/my-project" → sync only that project
  - Graceful skip when a top-level group's subgroups/projects can't be fetched

Environment variables:
    GITLAB_URL       - GitLab instance URL (required)
    GITLAB_TOKEN     - Personal access token (required)
    GITLAB_GROUPS    - Comma-separated group/sub-group/project paths (required)
    GITLAB_BASE_DIR  - Local storage path (default: ~/Desktop/Code)
    GITLAB_MODE      - "clone" (default), "update" (skip clone, only pull),
                       or "sync" (clone+pull+cleanup stale Excel rows)
    GITLAB_WORKERS   - Number of parallel workers (default: 4)
"""

from __future__ import annotations

import os
import sys
import json
import signal
import subprocess
import urllib.request
import urllib.parse
import urllib.error
import ssl
import time
import threading
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed

# ---------------------------------------------------------------------------
# Dependency bootstrap
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("[setup] Installing openpyxl ...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
CLONE_TIMEOUT = 300       # 5 minutes per project
FETCH_TIMEOUT = 300       # 5 minutes for fetch/pull
GIT_CMD_TIMEOUT = 60      # 1 minute for lightweight git commands
API_PER_PAGE = 100
INCREMENTAL_EXCEL_BATCH = 50  # Write Excel every N completed projects
DEFAULT_TOTAL_TIMEOUT = 0     # 0 = no global timeout; override via GITLAB_TOTAL_TIMEOUT

HEADERS_CN = [
    "主Group名称", "子Group路径", "Project路径", "Project名称",
    "Project描述", "已checkout分支", "分支最新提交时间", "SSH Git链接", "下载时间",
    "Project ID",
]

NUM_COLUMNS = len(HEADERS_CN)  # 10

COL_WIDTHS = {"A": 15, "B": 35, "C": 45, "D": 30, "E": 40, "F": 35, "G": 25, "H": 60, "I": 20, "J": 12}


# ===========================================================================
# GitLab API helpers (using urllib – no extra dependency)
# ===========================================================================

def _make_ssl_ctx():
    """Create a permissive SSL context (many internal GitLabs use self-signed certs)."""
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx

_SSL_CTX = _make_ssl_ctx()


def api_get(gitlab_url: str, token: str, path: str, params: dict | None = None) -> list | dict:
    """GET request to GitLab API. Returns parsed JSON."""
    url = f"{gitlab_url}/api/v4{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={"PRIVATE-TOKEN": token})
    try:
        with urllib.request.urlopen(req, context=_SSL_CTX, timeout=30) as resp:
            status = resp.status
            body = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        print(f"  [API error] {e.code} for {path}: {e.reason}")
        return []
    except Exception as e:
        print(f"  [API error] {path}: {e}")
        return []

    try:
        return json.loads(body)
    except json.JSONDecodeError:
        print(f"  [API error] Invalid JSON from {path}")
        return []


def api_get_paginated(gitlab_url: str, token: str, path: str, params: dict | None = None) -> list:
    """GET with automatic pagination. Returns flat list."""
    results = []
    page = 1
    base_params = dict(params or {})
    base_params["per_page"] = API_PER_PAGE
    while True:
        base_params["page"] = page
        data = api_get(gitlab_url, token, path, base_params)
        if not isinstance(data, list) or len(data) == 0:
            break
        results.extend(data)
        if len(data) < API_PER_PAGE:
            break
        page += 1
    return results


# ===========================================================================
# GitLab data fetching
# ===========================================================================

def fetch_subgroups_recursive(gitlab_url: str, token: str, group_id: int) -> list[dict]:
    """Recursively fetch all descendant subgroups of a group."""
    # Use the /groups/:id/descendant_groups endpoint if available (GitLab 13.5+)
    descendants = api_get_paginated(
        gitlab_url, token, f"/groups/{group_id}/descendant_groups",
        {"all_available": "false"},
    )
    if isinstance(descendants, list) and len(descendants) > 0:
        return descendants

    # Fallback: manual BFS via /groups/:id/subgroups
    all_subs = []
    queue = [group_id]
    visited = {group_id}
    while queue:
        gid = queue.pop(0)
        subs = api_get_paginated(gitlab_url, token, f"/groups/{gid}/subgroups")
        for s in subs:
            if s["id"] not in visited:
                visited.add(s["id"])
                all_subs.append(s)
                queue.append(s["id"])
    return all_subs


def resolve_group_id(gitlab_url: str, token: str, group_path: str) -> int | None:
    """Resolve a group path (e.g. 'myGroup') to its numeric ID."""
    encoded = urllib.parse.quote(group_path, safe="")
    data = api_get(gitlab_url, token, f"/groups/{encoded}")
    if isinstance(data, dict) and "id" in data:
        return data["id"]
    return None


def resolve_project(gitlab_url: str, token: str, project_path: str) -> dict | None:
    """Try to resolve a path as a project. Returns project dict or None."""
    encoded = urllib.parse.quote(project_path, safe="")
    data = api_get(gitlab_url, token, f"/projects/{encoded}")
    if isinstance(data, dict) and "id" in data:
        return data
    return None


def is_subpath(path: str) -> bool:
    """Check if a group input contains '/' — meaning the user specified a sub-path."""
    return "/" in path


def resolve_path_type(gitlab_url: str, token: str, path: str) -> tuple[str, dict | int | None]:
    """Resolve a user-provided path to either a project, group, or nothing.

    Returns:
        ("project", project_dict)  — path is a project
        ("group", group_id)        — path is a group/sub-group
        ("none", None)             — path could not be resolved

    Strategy: try group first (more common for paths with '/'), then project.
    This avoids a wasted 404 on /projects/ for sub-group paths like 'myGroup/mySubGroup'.
    """
    # Try as group first (sub-paths are more likely to be groups)
    gid = resolve_group_id(gitlab_url, token, path)
    if gid is not None:
        return ("group", gid)

    # Try as project
    proj = resolve_project(gitlab_url, token, path)
    if proj is not None:
        return ("project", proj)

    return ("none", None)


def fetch_all_projects(gitlab_url: str, token: str, group_ids: list[int]) -> list[dict]:
    """Fetch projects for a list of group IDs, deduplicated by project ID."""
    seen = set()
    projects = []
    for gid in group_ids:
        items = api_get_paginated(
            gitlab_url, token, f"/groups/{gid}/projects",
            {"include_subgroups": "false", "archived": "false"},
        )
        for p in items:
            if isinstance(p, dict) and "id" in p and p["id"] not in seen:
                seen.add(p["id"])
                projects.append(p)
    return projects


# ===========================================================================
# Git operations
# ===========================================================================

def _run_git(args: list[str], cwd: str, timeout: int = GIT_CMD_TIMEOUT) -> subprocess.CompletedProcess:
    """Run a git command with timeout.  Uses a new session so the entire
    process group can be killed on timeout (prevents orphan git-remote-https)."""
    proc = subprocess.Popen(
        ["git"] + args,
        cwd=cwd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        start_new_session=True,
    )
    try:
        stdout, stderr = proc.communicate(timeout=timeout)
        return subprocess.CompletedProcess(proc.args, proc.returncode, stdout, stderr)
    except subprocess.TimeoutExpired:
        # Kill entire process group to avoid orphan children
        try:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
        except OSError:
            proc.kill()
        proc.wait()
        raise


def clone_project(http_url: str, token: str, target_dir: str) -> None:
    """Clone a project via HTTPS with embedded token, then strip token from remote.
    Uses start_new_session + process-group kill to avoid orphan git processes."""
    url_with_token = http_url.replace("https://", f"https://oauth2:{token}@")
    proc = subprocess.Popen(
        ["git", "clone", "--quiet", url_with_token, target_dir],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        start_new_session=True,
    )
    try:
        proc.communicate(timeout=CLONE_TIMEOUT)
    except subprocess.TimeoutExpired:
        try:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
        except OSError:
            proc.kill()
        proc.wait()
        # Always try to strip token from remote (even if clone partially failed)
        _sanitize_remote(target_dir, http_url)
        raise
    # Always try to strip token from remote (even if clone partially failed)
    _sanitize_remote(target_dir, http_url)


def pull_project(target_dir: str) -> None:
    """Fetch all remotes and pull current branch."""
    _run_git(["fetch", "--all", "--quiet"], cwd=target_dir, timeout=FETCH_TIMEOUT)
    _run_git(["pull", "--quiet"], cwd=target_dir, timeout=FETCH_TIMEOUT)


def _sanitize_remote(target_dir: str, clean_url: str) -> None:
    """Reset origin URL to token-free version."""
    if os.path.isdir(os.path.join(target_dir, ".git")):
        _run_git(["remote", "set-url", "origin", clean_url], cwd=target_dir)


def get_default_branch(repo_path: str) -> str:
    """Detect the default branch from origin/HEAD or fall back heuristically."""
    # Try symbolic ref
    r = _run_git(["symbolic-ref", "refs/remotes/origin/HEAD", "--short"], cwd=repo_path)
    if r.returncode == 0 and r.stdout.strip():
        return r.stdout.strip().replace("origin/", "")

    # Fallback: current HEAD
    r = _run_git(["rev-parse", "--abbrev-ref", "HEAD"], cwd=repo_path)
    head = r.stdout.strip() if r.returncode == 0 else ""
    if head and head != "HEAD":
        return head

    # Last resort: check for master/main in remotes
    r = _run_git(["branch", "-r", "--list", "origin/master", "origin/main"], cwd=repo_path)
    for line in r.stdout.strip().splitlines():
        name = line.strip().replace("origin/", "")
        if name in ("master", "main"):
            return name
    return "master"


def get_remote_branches_sorted(repo_path: str) -> list[tuple[str, str]]:
    """Return remote branches sorted by committer date (newest first).
    Each item: (branch_name, 'YYYY-MM-DD HH:MM:SS')
    """
    r = _run_git(
        ["for-each-ref", "--sort=-committerdate", "refs/remotes/origin/",
         "--format=%(refname:short)\t%(committerdate:iso8601)"],
        cwd=repo_path,
    )
    if r.returncode != 0:
        return []

    result = []
    for line in r.stdout.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split("\t", 1)
        if len(parts) < 2:
            continue
        ref = parts[0]
        ts = parts[1]
        if "HEAD" in ref:
            continue
        branch = ref.replace("origin/", "", 1)
        # Normalize timestamp: strip timezone offset
        ts_clean = ts.split(" +")[0].split(" -")[0]
        result.append((branch, ts_clean))
    return result


def _is_release_branch(name: str) -> bool:
    low = name.lower()
    return "release" in low or "prod" in low


def checkout_branches(repo_path: str) -> None:
    """Checkout: default branch + latest active + latest release/prod."""
    default = get_default_branch(repo_path)
    remotes = get_remote_branches_sorted(repo_path)
    if not remotes:
        return

    to_checkout = [default]

    # Latest active branch (first non-default in time-sorted list)
    for name, _ in remotes:
        if name != default:
            to_checkout.append(name)
            break

    # Latest release/prod branch
    for name, _ in remotes:
        if name in to_checkout:
            continue
        if _is_release_branch(name):
            to_checkout.append(name)
            break

    for branch in to_checkout:
        # Already exists locally?
        r = _run_git(["branch", "--list", branch], cwd=repo_path)
        if r.stdout.strip():
            # Fast-forward existing local branch to match remote (no checkout needed)
            _run_git(["fetch", "origin", f"{branch}:{branch}"], cwd=repo_path)
        else:
            # Create local tracking branch from remote
            _run_git(["checkout", "-b", branch, f"origin/{branch}"], cwd=repo_path)

    # Return to default
    _run_git(["checkout", default], cwd=repo_path)


def get_local_branches_info(repo_path: str) -> list[tuple[str, str]]:
    """Return (branch_name, commit_time) for every local branch."""
    r = _run_git(
        ["for-each-ref", "--format=%(refname:short)\t%(committerdate:iso8601)", "refs/heads/"],
        cwd=repo_path,
    )
    if r.returncode != 0:
        return []
    result = []
    for line in r.stdout.strip().splitlines():
        parts = line.strip().split("\t", 1)
        if len(parts) < 2:
            continue
        name = parts[0]
        ts = parts[1].split(" +")[0].split(" -")[0]
        result.append((name, ts))
    return result


# ===========================================================================
# Per-project worker (runs in subprocess via ProcessPoolExecutor)
# ===========================================================================

def process_one_project(
    project_info: dict,
    base_dir: str,
    token: str,
    mode: str,
) -> dict:
    """Clone/pull a single project and return its metadata dict.

    Returns a dict with project data or an 'error' key on failure.
    """
    project_path = project_info["path_with_namespace"]
    http_url = project_info["http_url_to_repo"]
    ssh_url = project_info.get("ssh_url_to_repo", "")
    description = project_info.get("description", "") or ""
    target_dir = os.path.join(base_dir, project_path)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    is_existing = os.path.isdir(os.path.join(target_dir, ".git"))

    try:
        if is_existing:
            # Update mode or already cloned
            pull_project(target_dir)
        elif mode == "update":
            # Update-only mode but repo doesn't exist locally — skip
            return {"error": "skip", "project_path": project_path, "reason": "not cloned yet (update mode)"}
        else:
            os.makedirs(os.path.dirname(target_dir), exist_ok=True)
            clone_project(http_url, token, target_dir)

        if not os.path.isdir(os.path.join(target_dir, ".git")):
            return {"error": "clone_failed", "project_path": project_path, "reason": "no .git after clone"}

        # Ensure token is stripped
        _sanitize_remote(target_dir, http_url)

        # Checkout branches
        checkout_branches(target_dir)

        # Gather branch info
        branches = get_local_branches_info(target_dir)
        branches_str = "\n".join(b[0] for b in branches)
        times_str = "\n".join(b[1] for b in branches)

    except subprocess.TimeoutExpired:
        # Clean up token if partially cloned
        _sanitize_remote(target_dir, http_url)
        return {"error": "timeout", "project_path": project_path, "reason": "exceeded 5-min timeout"}
    except Exception as e:
        _sanitize_remote(target_dir, http_url)
        return {"error": "exception", "project_path": project_path, "reason": str(e)}

    # Compute group columns
    parts = project_path.split("/")
    main_group = parts[0] if parts else ""
    sub_group = "/".join(parts[:-1]) if len(parts) > 1 else main_group

    return {
        "main_group": main_group,
        "sub_group": sub_group,
        "project_path": project_path,
        "project_name": project_info["name"],
        "description": description,
        "branches": branches_str,
        "times": times_str,
        "ssh_url": ssh_url,
        "download_time": now_str,
        "project_id": project_info["id"],
    }


# ===========================================================================
# Excel generation
# ===========================================================================

def _item_to_row(item: dict) -> list:
    """Convert a project data dict to a row list matching HEADERS_CN."""
    return [
        item["main_group"],
        item["sub_group"],
        item["project_path"],
        item["project_name"],
        item["description"],
        item["branches"],
        item["times"],
        item["ssh_url"],
        item["download_time"],
        item.get("project_id", ""),
    ]


def build_excel(excel_path: str, projects_data: list[dict]) -> None:
    """Create or overwrite the Excel index file with per-group sheets."""
    if not projects_data:
        print("[excel] No project data to write, skipping Excel creation.")
        return

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Group projects by main_group
    groups: dict[str, list[dict]] = {}
    for p in projects_data:
        mg = p["main_group"]
        groups.setdefault(mg, []).append(p)

    for group_name in sorted(groups.keys()):
        items = groups[group_name]
        # Sort: sub_group asc, project_name asc
        items.sort(key=lambda x: (x["sub_group"], x["project_name"]))

        ws = wb.create_sheet(title=group_name[:31])  # sheet name max 31 chars
        ws.append(HEADERS_CN)

        for item in items:
            ws.append(_item_to_row(item))

        _format_sheet(ws)

    wb.save(excel_path)
    print(f"\n[excel] Saved: {excel_path}")


def _find_sheet_for_group(wb, group_name: str):
    """Find the sheet belonging to *group_name* by scanning Column A data.

    This avoids relying solely on sheet title (which may be truncated to 31 chars).
    Returns the worksheet or None.
    """
    # Fast path: exact title match
    truncated = group_name[:31]
    if truncated in wb.sheetnames:
        return wb[truncated]

    # Slow path: scan Column A (主Group名称) of every sheet
    for ws in wb.worksheets:
        if ws.max_row >= 2:
            val = ws.cell(row=2, column=1).value
            if val == group_name:
                return ws
    return None


def _build_global_id_lookup(wb) -> dict[int, tuple[str, int]]:
    """Build a global lookup: project_id (int) → (sheet_title, row_idx).

    Scans Column J (Project ID) of every sheet.  Rows without a valid ID are
    skipped — they will be matched by path as a fallback.
    """
    lookup: dict[int, tuple[str, int]] = {}
    for ws in wb.worksheets:
        for row_idx in range(2, ws.max_row + 1):
            raw = ws.cell(row=row_idx, column=10).value  # Column J
            if raw is not None and raw != "":
                try:
                    pid = int(raw)
                    lookup[pid] = (ws.title, row_idx)
                except (ValueError, TypeError):
                    pass
    return lookup


def _build_path_lookup(ws) -> dict[str, int]:
    """Build path → row_idx lookup for a single sheet (Column C)."""
    lookup: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        pp = ws.cell(row=row_idx, column=3).value
        if pp:
            lookup[pp] = row_idx
    return lookup


def update_excel(
    excel_path: str,
    projects_data: list[dict],
    *,
    all_project_ids: set[int] | None = None,
    synced_groups: set[str] | None = None,
) -> None:
    """Update an existing Excel file with project data.

    Matching priority:
      1. Project ID (Column J) — handles renames and path changes
      2. Project path (Column C) — fallback for legacy rows without ID

    When *all_project_ids* is provided (sync mode, final write), rows in sheets
    belonging to *synced_groups* whose Project ID is not in the set are deleted.
    This cleans up projects that were deleted/archived on GitLab.
    """
    if not os.path.exists(excel_path):
        build_excel(excel_path, projects_data)
        return

    wb = openpyxl.load_workbook(excel_path)

    # --- Phase 1: Build global ID lookup (for cross-sheet migration) ---
    global_id_lookup = _build_global_id_lookup(wb)

    # Collect rows to delete from other sheets (cross-group migration)
    # Key: sheet_title → list of row indices
    rows_to_delete: dict[str, list[int]] = {}

    # --- Phase 2: Upsert incoming project data ---
    groups: dict[str, list[dict]] = {}
    for p in projects_data:
        mg = p["main_group"]
        groups.setdefault(mg, []).append(p)

    for group_name, items in groups.items():
        ws = _find_sheet_for_group(wb, group_name)

        if ws is not None:
            path_lookup = _build_path_lookup(ws)

            for item in items:
                row_data = _item_to_row(item)
                pid = item.get("project_id")
                matched_row = None

                # Try ID match first (global — may be in a different sheet)
                if pid is not None and int(pid) in global_id_lookup:
                    old_sheet_title, old_row = global_id_lookup[int(pid)]
                    if old_sheet_title == ws.title:
                        # Same sheet — update in place
                        matched_row = old_row
                    else:
                        # Cross-sheet migration: mark old row for deletion
                        rows_to_delete.setdefault(old_sheet_title, []).append(old_row)
                        # Will append as new row below

                # Fallback: path match within current sheet (for legacy rows without ID)
                if matched_row is None and item["project_path"] in path_lookup:
                    matched_row = path_lookup[item["project_path"]]

                if matched_row is not None:
                    for col_i, val in enumerate(row_data, 1):
                        ws.cell(row=matched_row, column=col_i, value=val)
                else:
                    ws.append(row_data)

            _resort_sheet(ws)
            _format_sheet(ws)
        else:
            # New group — create sheet, but still check for cross-sheet migrations
            for item in items:
                pid = item.get("project_id")
                if pid is not None and int(pid) in global_id_lookup:
                    old_sheet_title, old_row = global_id_lookup[int(pid)]
                    rows_to_delete.setdefault(old_sheet_title, []).append(old_row)

            items.sort(key=lambda x: (x["sub_group"], x["project_name"]))
            sheet_name = _unique_sheet_name(wb, group_name)
            ws = wb.create_sheet(title=sheet_name)
            ws.append(HEADERS_CN)
            for item in items:
                ws.append(_item_to_row(item))
            _format_sheet(ws)

    # --- Phase 3: Delete cross-sheet migrated rows (from back to front) ---
    for sheet_title, row_indices in rows_to_delete.items():
        if sheet_title in wb.sheetnames:
            target_ws = wb[sheet_title]
            for ridx in sorted(row_indices, reverse=True):
                if ridx <= target_ws.max_row:
                    target_ws.delete_rows(ridx)
            _resort_sheet(target_ws)
            _format_sheet(target_ws)

    # --- Phase 4: Cleanup stale rows (sync mode only) ---
    if all_project_ids is not None and synced_groups is not None:
        for ws in wb.worksheets:
            # Only clean sheets belonging to synced groups
            group_in_sheet = ws.cell(row=2, column=1).value if ws.max_row >= 2 else None
            if group_in_sheet not in synced_groups:
                continue

            stale_rows = []
            for row_idx in range(2, ws.max_row + 1):
                raw_id = ws.cell(row=row_idx, column=10).value  # Column J
                if raw_id is not None and raw_id != "":
                    try:
                        if int(raw_id) not in all_project_ids:
                            stale_rows.append(row_idx)
                    except (ValueError, TypeError):
                        pass
                else:
                    # Legacy row without ID — check if path is still in current data
                    pp = ws.cell(row=row_idx, column=3).value
                    if pp:
                        path_still_exists = any(
                            p.get("project_path") == pp for p in projects_data
                        )
                        if not path_still_exists:
                            stale_rows.append(row_idx)

            for ridx in sorted(stale_rows, reverse=True):
                ws.delete_rows(ridx)

            if stale_rows:
                print(f"  [excel] Removed {len(stale_rows)} stale rows from sheet '{ws.title}'")
                _resort_sheet(ws)
                _format_sheet(ws)

    wb.save(excel_path)
    print(f"\n[excel] Updated: {excel_path}")


def _unique_sheet_name(wb, group_name: str) -> str:
    """Generate a unique sheet name (max 31 chars) for a group."""
    name = group_name[:31]
    if name not in wb.sheetnames:
        return name
    for i in range(2, 100):
        suffix = f"_{i}"
        candidate = group_name[:31 - len(suffix)] + suffix
        if candidate not in wb.sheetnames:
            return candidate
    return name  # fallback (should never happen)


def _resort_sheet(ws) -> None:
    """Re-sort data rows (2..max_row) by sub_group(B), project_name(D)."""
    if ws.max_row < 3:
        return
    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        data.append(list(row))

    data.sort(key=lambda r: (r[1] or "", r[3] or ""))

    # Determine column count from data (handles both 9-col legacy and 10-col new)
    max_cols = max((len(r) for r in data), default=NUM_COLUMNS)
    col_count = max(max_cols, NUM_COLUMNS)

    # Clear data rows
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, col_count + 1):
            ws.cell(row=row_idx, column=col_idx, value=None)

    # Write back
    for i, row_data in enumerate(data, 2):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=i, column=ci, value=val)


def _format_sheet(ws) -> None:
    """Apply formatting: widths, borders, alignment, freeze."""
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Header style
    hdr_font = Font(bold=True, size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.alignment = hdr_align

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=NUM_COLUMNS):
        for cell in row:
            cell.border = thin
            if cell.row == 1:
                continue
            if cell.column in (6, 7):  # F, G
                cell.alignment = left_wrap
            else:
                cell.alignment = center_wrap

    # Hide Project ID column (J) — internal use, not for end users
    ws.column_dimensions["J"].hidden = True

    ws.freeze_panes = "A2"


# ===========================================================================
# Main
# ===========================================================================

# ---------------------------------------------------------------------------
# Incremental Excel helper (thread-safe via lock)
# ---------------------------------------------------------------------------
_excel_lock = threading.Lock()


def _incremental_excel_write(
    excel_path: str,
    batch: list[dict],
    *,
    all_project_ids: set[int] | None = None,
    synced_groups: set[str] | None = None,
) -> None:
    """Thread-safe incremental write: merge *batch* into the Excel file.

    Pass *all_project_ids* and *synced_groups* only on the **final** write
    (sync mode) to trigger stale-row cleanup.  During incremental writes
    these should be None to avoid premature deletion.
    """
    has_cleanup = all_project_ids is not None
    if not batch and not has_cleanup:
        return
    with _excel_lock:
        if os.path.exists(excel_path):
            update_excel(
                excel_path, batch,
                all_project_ids=all_project_ids,
                synced_groups=synced_groups,
            )
        elif batch:
            build_excel(excel_path, batch)


# ---------------------------------------------------------------------------
# Signal-safe emergency Excel dump
# ---------------------------------------------------------------------------
_emergency_state: dict = {
    "excel_path": "",
    "pending": [],      # successes not yet written
}


def _emergency_excel_handler(signum, frame):
    """On SIGTERM / SIGINT, attempt to flush pending results to Excel."""
    pending = _emergency_state["pending"]
    excel_path = _emergency_state["excel_path"]
    if pending and excel_path:
        print(f"\n[signal] Caught signal {signum}, flushing {len(pending)} results to Excel ...")
        try:
            _incremental_excel_write(excel_path, list(pending))
            print(f"[signal] Emergency Excel write done: {excel_path}")
        except Exception as e:
            print(f"[signal] Emergency Excel write failed: {e}")
    sys.exit(1)


def main() -> None:
    gitlab_url = os.environ.get("GITLAB_URL", "").strip().rstrip("/")
    token = os.environ.get("GITLAB_TOKEN", "").strip()
    groups_input = os.environ.get("GITLAB_GROUPS", "").strip()
    base_dir = os.path.expanduser(os.environ.get("GITLAB_BASE_DIR", "~/Desktop/Code").strip())
    mode = os.environ.get("GITLAB_MODE", "clone").strip().lower()
    workers = int(os.environ.get("GITLAB_WORKERS", "4"))
    total_timeout = int(os.environ.get("GITLAB_TOTAL_TIMEOUT", str(DEFAULT_TOTAL_TIMEOUT)))

    if not all([gitlab_url, token, groups_input]):
        print("Error: GITLAB_URL, GITLAB_TOKEN, and GITLAB_GROUPS must be set.")
        sys.exit(1)

    target_groups = [g.strip() for g in groups_input.split(",") if g.strip()]
    os.makedirs(base_dir, exist_ok=True)

    excel_path = os.path.join(base_dir, "01.Index.xlsx")
    _emergency_state["excel_path"] = excel_path

    # Install signal handlers for emergency Excel dump
    signal.signal(signal.SIGTERM, _emergency_excel_handler)
    signal.signal(signal.SIGINT, _emergency_excel_handler)

    # -----------------------------------------------------------------------
    # 1. Resolve target groups/sub-groups/projects
    # -----------------------------------------------------------------------
    mode_label = {"update": "Update", "sync": "Sync", "clone": "Clone"}.get(mode, "Clone")
    is_sync = mode == "sync"
    # In sync mode, clone/pull behaves like "clone" mode, but final Excel
    # write also cleans up stale rows for synced groups.
    effective_mode = "clone" if is_sync else mode

    print(f"=== GitLab Batch {mode_label} ===")
    print(f"URL: {gitlab_url}")
    print(f"Groups: {', '.join(target_groups)}")
    print(f"Storage: {base_dir}")
    print(f"Workers: {workers}")
    if total_timeout > 0:
        print(f"Global timeout: {total_timeout}s")
    print()

    all_group_ids: list[int] = []
    direct_projects: list[dict] = []   # projects resolved directly by path

    for gpath in target_groups:
        print(f"[resolve] {gpath}")

        # --- Case A: User specified a sub-path (contains '/') ---
        if is_subpath(gpath):
            path_type, resolved = resolve_path_type(gitlab_url, token, gpath)

            if path_type == "project":
                # Direct project path — skip group enumeration entirely
                print(f"  → Resolved as PROJECT (id={resolved['id']})")
                direct_projects.append(resolved)
                continue

            elif path_type == "group":
                # Sub-group path — enumerate only this sub-group and its children
                gid = resolved
                print(f"  → Resolved as SUB-GROUP (id={gid}), fetching descendants ...")
                all_group_ids.append(gid)
                subs = fetch_subgroups_recursive(gitlab_url, token, gid)
                sub_ids = [s["id"] for s in subs]
                all_group_ids.extend(sub_ids)
                print(f"  Found {len(subs)} child subgroups ({len(sub_ids) + 1} total)")
                continue

            else:
                print(f"  ⚠ Could not resolve '{gpath}' as group or project, skipping")
                continue

        # --- Case B: Top-level group (no '/') ---
        gid = resolve_group_id(gitlab_url, token, gpath)
        if gid is None:
            print(f"  ⚠ Could not resolve group '{gpath}', skipping")
            continue

        all_group_ids.append(gid)
        print(f"  ID={gid}, fetching subgroups ...")
        subs = fetch_subgroups_recursive(gitlab_url, token, gid)
        if isinstance(subs, list):
            sub_ids = [s["id"] for s in subs]
        else:
            sub_ids = []

        # Check if we can actually access this group's data
        # Try fetching direct projects of this group as a canary
        canary = api_get_paginated(
            gitlab_url, token, f"/groups/{gid}/projects",
            {"include_subgroups": "false", "per_page": "1"},
        )
        if not isinstance(canary, list):
            # API returned non-list (error dict, etc.) — can't enumerate this group
            print(f"  ⚠ Cannot fetch subgroups/projects for '{gpath}' (access denied?), skipping")
            # Remove the group id we just added
            all_group_ids.pop()
            continue

        all_group_ids.extend(sub_ids)
        print(f"  Found {len(subs)} subgroups ({len(sub_ids) + 1} total groups)")

    if not all_group_ids and not direct_projects:
        print("No groups or projects found. Exiting.")
        sys.exit(1)

    # -----------------------------------------------------------------------
    # 2. Fetch all projects from groups + merge direct projects
    # -----------------------------------------------------------------------
    all_projects: list[dict] = []
    seen_ids: set[int] = set()

    # Add directly-resolved projects first
    for p in direct_projects:
        if p["id"] not in seen_ids:
            seen_ids.add(p["id"])
            all_projects.append(p)

    if all_group_ids:
        print(f"\n[projects] Fetching projects from {len(all_group_ids)} groups ...")
        group_projects = fetch_all_projects(gitlab_url, token, all_group_ids)
        for p in group_projects:
            if p["id"] not in seen_ids:
                seen_ids.add(p["id"])
                all_projects.append(p)

    print(f"[projects] Total unique projects: {len(all_projects)}")

    # Per-group project count breakdown
    group_count: dict[str, int] = {}
    for p in all_projects:
        top = p["path_with_namespace"].split("/")[0]
        group_count[top] = group_count.get(top, 0) + 1
    for gname, cnt in sorted(group_count.items()):
        print(f"  {gname}: {cnt} projects")

    if not all_projects:
        print("No projects found. Exiting.")
        sys.exit(0)

    # -----------------------------------------------------------------------
    # 3. Process projects (clone/pull + checkout + metadata)
    # -----------------------------------------------------------------------
    successes: list[dict] = []
    failures: list[dict] = []
    pending_batch: list[dict] = []   # buffer for incremental Excel writes
    total = len(all_projects)
    start_time = time.monotonic()

    print(f"\n[process] Starting with {workers} workers ...\n")

    timed_out_globally = False

    with ProcessPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(process_one_project, p, base_dir, token, effective_mode): p for p in all_projects}
        done_count = 0
        for future in as_completed(futures):
            done_count += 1
            proj = futures[future]
            elapsed = time.monotonic() - start_time
            try:
                result = future.result()
            except Exception as e:
                result = {"error": "worker_exception", "project_path": proj["path_with_namespace"], "reason": str(e)}

            tag = "✓" if "error" not in result else "✗"
            elapsed_str = f"{elapsed:.0f}s"
            print(f"  [{done_count}/{total}] {tag} {proj['path_with_namespace']}  ({elapsed_str})")

            if "error" in result:
                failures.append(result)
            else:
                successes.append(result)
                pending_batch.append(result)
                _emergency_state["pending"].append(result)

            # Incremental Excel write every INCREMENTAL_EXCEL_BATCH projects
            if len(pending_batch) >= INCREMENTAL_EXCEL_BATCH:
                print(f"  [excel] Incremental write ({len(successes)} projects so far) ...")
                _incremental_excel_write(excel_path, pending_batch)
                # Clear pending from both buffers (slice — O(n), no dict equality issues)
                _emergency_state["pending"] = _emergency_state["pending"][len(pending_batch):]
                pending_batch = []

            # Check global timeout
            if total_timeout > 0 and elapsed > total_timeout:
                print(f"\n⚠ Global timeout ({total_timeout}s) exceeded after {done_count}/{total} projects.")
                timed_out_globally = True
                # Cancel remaining futures
                for f in futures:
                    f.cancel()
                break

    # -----------------------------------------------------------------------
    # 4. Flush remaining results to Excel
    # -----------------------------------------------------------------------
    # In sync mode, the final write also cleans up stale rows.
    # Compute the full set of project IDs we fetched from GitLab (not just successes).
    sync_all_ids: set[int] | None = None
    sync_groups: set[str] | None = None
    if is_sync:
        sync_all_ids = {p["id"] for p in all_projects}
        sync_groups = {p["path_with_namespace"].split("/")[0] for p in all_projects}

    if pending_batch or is_sync:
        print(f"  [excel] Final write ({len(successes)} total projects) ...")
        _incremental_excel_write(
            excel_path, pending_batch,
            all_project_ids=sync_all_ids,
            synced_groups=sync_groups,
        )
        _emergency_state["pending"] = []
        pending_batch = []

    # -----------------------------------------------------------------------
    # 5. Summary
    # -----------------------------------------------------------------------
    total_elapsed = time.monotonic() - start_time
    elapsed_min = total_elapsed / 60

    print(f"\n{'=' * 60}")
    print(f"  Done!  Success: {len(successes)}  |  Failed/Timeout: {len(failures)}")
    if timed_out_globally:
        remaining = total - len(successes) - len(failures)
        print(f"  ⚠ Cancelled (global timeout): ~{remaining} projects")
    print(f"  Time: {elapsed_min:.1f} min")
    print(f"  Projects: {base_dir}")
    print(f"  Excel:    {excel_path}")
    print(f"{'=' * 60}")

    if failures:
        print(f"\n⚠ Failed / Timed-out projects ({len(failures)}):\n")
        print(f"  {'Project Path':<55} {'Reason'}")
        print(f"  {'-'*55} {'-'*30}")
        for f in failures:
            pp = f.get("project_path", "?")
            reason = f.get("reason", "unknown")
            print(f"  {pp:<55} {reason}")
        print()


if __name__ == "__main__":
    main()
