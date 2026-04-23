#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Xiaohongshu Data Export Tool

Usage:
    python export_data.py --format xlsx --output viral_data.xlsx
    python export_data.py --category Beauty --format csv --output beauty.csv
"""

import argparse
import json
import csv
from datetime import datetime
from typing import List, Dict


def get_mock_data() -> List[Dict]:
    """Retrieve mock data"""
    return [
        {
            "note_id": "note_0001",
            "title": "5-Minute No-Makeup Look for Early Risers | Foundation Hack My Coworkers Keep Asking About",
            "author_name": "BeautyGuru123",
            "followers": 3200,
            "category": "Beauty",
            "likes": 5200,
            "collections": 1800,
            "comments": 320,
            "shares": 150,
            "publish_time": "2024-01-15T08:30:00",
            "viral_score": 89
        },
        {
            "note_id": "note_0002",
            "title": "Fall/Winter Outfits for Petites | The Secret to Looking 10cm Taller",
            "author_name": "StyleLab",
            "followers": 1800,
            "category": "Fashion",
            "likes": 6800,
            "collections": 2400,
            "comments": 450,
            "shares": 280,
            "publish_time": "2024-01-14T19:00:00",
            "viral_score": 92
        },
        {
            "note_id": "note_0003",
            "title": "Solo Dining | 10-Minute Lazy Dinner Recipe",
            "author_name": "FoodieDiary",
            "followers": 4500,
            "category": "Food",
            "likes": 4100,
            "collections": 3200,
            "comments": 280,
            "shares": 520,
            "publish_time": "2024-01-13T12:00:00",
            "viral_score": 85
        },
        {
            "note_id": "note_0004",
            "title": "Weekend Getaway | Hidden Gem Photo Spots",
            "author_name": "WanderlustExplorer",
            "followers": 900,
            "category": "Travel",
            "likes": 3500,
            "collections": 2800,
            "comments": 180,
            "shares": 420,
            "publish_time": "2024-01-12T10:00:00",
            "viral_score": 78
        },
        {
            "note_id": "note_0005",
            "title": "Rental Makeover | 500 Yuan to Create an Instagram-Style Bedroom",
            "author_name": "HomeStylist",
            "followers": 2600,
            "category": "Home",
            "likes": 5900,
            "collections": 4100,
            "comments": 520,
            "shares": 380,
            "publish_time": "2024-01-11T15:30:00",
            "viral_score": 91
        }
    ]


def export_to_csv(data: List[Dict], filepath: str):
    """Export to CSV format"""
    if not data:
        print("No data to export")
        return
    
    fieldnames = [
        "note_id", "title", "author_name", "followers", "category",
        "likes", "collections", "comments", "shares", 
        "total_engagement", "engagement_rate", "publish_time", "viral_score"
    ]
    
    # Calculate derived fields
    for item in data:
        item["total_engagement"] = item["likes"] + item["collections"] + item["comments"] + item["shares"]
        item["engagement_rate"] = round(item["total_engagement"] / max(item["followers"], 1), 2)
    
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for item in data:
            row = {k: item.get(k, '') for k in fieldnames}
            writer.writerow(row)
    
    print(f"✅ CSV file exported: {filepath}")
    print(f"   Total {len(data)} records")


def export_to_json(data: List[Dict], filepath: str):
    """Export to JSON format"""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON file exported: {filepath}")
    print(f"   Total {len(data)} records")


def export_to_xlsx(data: List[Dict], filepath: str):
    """Export to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Viral Post Data"
        
        # Headers
        headers = [
            "Post ID", "Title", "Author", "Followers", "Category",
            "Likes", "Saves", "Comments", "Shares", "Total Engagement", "Engagement Rate",
            "Publish Time", "Viral Score"
        ]
        
        # Set header styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, item in enumerate(data, 2):
            total_engagement = item["likes"] + item["collections"] + item["comments"] + item["shares"]
            engagement_rate = round(total_engagement / max(item["followers"], 1), 2)
            
            ws.cell(row=row_idx, column=1, value=item["note_id"])
            ws.cell(row=row_idx, column=2, value=item["title"])
            ws.cell(row=row_idx, column=3, value=item["author_name"])
            ws.cell(row=row_idx, column=4, value=item["followers"])
            ws.cell(row=row_idx, column=5, value=item["category"])
            ws.cell(row=row_idx, column=6, value=item["likes"])
            ws.cell(row=row_idx, column=7, value=item["collections"])
            ws.cell(row=row_idx, column=8, value=item["comments"])
            ws.cell(row=row_idx, column=9, value=item["shares"])
            ws.cell(row=row_idx, column=10, value=total_engagement)
            ws.cell(row=row_idx, column=11, value=engagement_rate)
            ws.cell(row=row_idx, column=12, value=item["publish_time"])
            ws.cell(row=row_idx, column=13, value=item["viral_score"])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 10
        
        # Save file
        wb.save(filepath)
        print(f"✅ Excel file exported: {filepath}")
        print(f"   Total {len(data)} records")
        
    except ImportError:
        print("❌ openpyxl not installed. Please run: pip install openpyxl")
        print("   Exporting as CSV format instead...")
        csv_path = filepath.replace('.xlsx', '.csv')
        export_to_csv(data, csv_path)


def main():
    parser = argparse.ArgumentParser(description="Xiaohongshu Data Export Tool")
    parser.add_argument("--format", "-f", required=True,
                       choices=["csv", "json", "xlsx"],
                       help="Export format")
    parser.add_argument("--output", "-o", required=True,
                       help="Output file path")
    parser.add_argument("--category", "-c", help="Filter by category")
    parser.add_argument("--limit", "-l", type=int, default=100,
                       help="Export limit")
    
    args = parser.parse_args()
    
    # Retrieve data
    data = get_mock_data()
    
    # Filter by category
    if args.category:
        data = [d for d in data if d["category"] == args.category]
    
    # Limit quantity
    data = data[:args.limit]
    
    # Export
    if args.format == "csv":
        export_to_csv(data, args.output)
    elif args.format == "json":
        export_to_json(data, args.output)
    elif args.format == "xlsx":
        export_to_xlsx(data, args.output)


if __name__ == "__main__":
    main()